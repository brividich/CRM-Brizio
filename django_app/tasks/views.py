from __future__ import annotations

import json
import logging
from datetime import date, datetime, timedelta
from functools import wraps

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.views import redirect_to_login
from django.db import DatabaseError, connections, transaction
from django.db.models import Count, F, Max, Prefetch, Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from admin_portale.decorators import legacy_admin_required
from core.acl import user_can_modulo_action
from core.audit import log_action
from core.legacy_cache import get_cached_perm_map
from core.legacy_models import UtenteLegacy
from core.legacy_utils import get_legacy_user
from core.legacy_utils import is_legacy_admin
from core.legacy_utils import legacy_table_columns
from core.legacy_utils import sync_django_user_from_legacy
from core.models import AuditLog, Notifica, Profile
from core.module_branding import get_module_branding_context, handle_module_branding_post
from core.upload_mime import (
    UploadMimeValidationError,
    safe_filename,
    validate_extension_and_mime,
)
from attrezzature.models import (
    Attrezzatura,
    AttrezzaturaNota,
    AttrezzaturaStato,
    AttrezzaturaTask as GestioneAttrezzaturaTask,
    AttrezzaturaTaskStato as GestioneAttrezzaturaTaskStato,
    AttrezzaturaTaskTipo as GestioneAttrezzaturaTaskTipo,
    DisponibilitaStato,
)
from attrezzature.services import kickoff_integration as attrezzature_kickoff
from attrezzature.services import workflow as attrezzature_workflow

from .forms import (
    KickoffMeetingForm,
    ProjectCommentForm,
    ProjectKickoffForm,
    ProjectTaskGanttUpdateForm,
    SubTaskForm,
    SubTaskStatusForm,
    TaskAttachmentForm,
    TaskCommentForm,
    TaskDueDateForm,
    TaskFilterForm,
    TaskForm,
    TaskStatusForm,
    task_active_users_queryset,
)
from .models import (
    DependencyType,
    GanttBaseline,
    KickoffMeeting,
    MeetingIssue,
    MeetingIssueStatus,
    MeetingRoom,
    Project,
    ProjectComment,
    SubTask,
    TaskAccessLevel,
    Task,
    TaskDependency,
    TaskAttachment,
    TaskCategory,
    TaskCategoryField,
    TaskCategoryFieldType,
    TaskComment,
    TaskEvent,
    TaskEventType,
    TaskExtraRef,
    TaskImpostazioni,
    TaskPriority,
    TaskReminder,
    TaskRoleAssignment,
    TaskRoleAccessRule,
    TaskRoleDefinition,
    TaskRoleType,
    TaskStatus,
    TaskUserAccessRule,
    VRFDocStatus,
    VRFRiskAssessment,
)

TASK_MODULE_CODE = "tasks"
OPEN_STATUSES = {TaskStatus.TODO, TaskStatus.IN_PROGRESS}
KEY_EDIT_FIELDS = ("title", "priority", "due_date", "next_step_text", "next_step_due", "tags", "project_id")
User = get_user_model()
logger = logging.getLogger(__name__)

SYSTEM_TASK_ROLE_DEFINITIONS = (
    (TaskRoleType.PROJECT_MANAGER, "Project manager", "Ruolo collegato al campo Project manager del kickoff.", 10),
    (TaskRoleType.CAPO_COMMESSA, "Capocommessa", "Ruolo collegato al campo Capocommessa del kickoff.", 20),
    (TaskRoleType.PROGRAMMER, "Programmatore", "Ruolo collegato al campo Programmatore del kickoff.", 30),
)

MONTH_LABELS_IT = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre",
}

GANTT_WINDOW_OPTIONS = (
    (31, "1 mese"),
    (62, "2 mesi"),
    (93, "3 mesi"),
    (124, "4 mesi"),
    (0, "Auto"),
)
GANTT_CELL_OPTIONS = (
    ("s", "Compatta"),
    ("m", "Standard"),
    ("l", "Ampia"),
)
GANTT_NAME_WIDTH_OPTIONS = (
    (280, "Compatta"),
    (360, "Media"),
    (460, "Ampia"),
)
TASK_SETTINGS_TABS = ("config", "riepilogo", "record", "log", "ruoli", "accessi", "promemoria", "tipi")
TASK_ACCESS_LEVEL_ORDER = {
    TaskAccessLevel.NONE: 0,
    TaskAccessLevel.READ_ALL: 1,
    TaskAccessLevel.EDIT_ASSIGNED: 2,
    TaskAccessLevel.EDIT_ALL: 3,
}
ATTREZZATURA_KICKOFF_TASK_TYPES = {
    "creazione_attrezzo": GestioneAttrezzaturaTaskTipo.CREAZIONE_ATTREZZO,
    "verifica_disponibilita": GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
    "aggiorna_avanzamento": GestioneAttrezzaturaTaskTipo.AGGIORNA_AVANZAMENTO,
    "controllo_ritardo": GestioneAttrezzaturaTaskTipo.CONTROLLO_RITARDO,
    "conferma_pronta_produzione": GestioneAttrezzaturaTaskTipo.CONFERMA_PRONTA_PRODUZIONE,
}


def _normalize_tasks_settings_tab(raw_tab: str | None, *, default: str = "config") -> str:
    tab = str(raw_tab or "").strip().lower()
    if tab in TASK_SETTINGS_TABS:
        return tab
    return default


def _ensure_system_task_roles() -> None:
    for code, name, description, order_index in SYSTEM_TASK_ROLE_DEFINITIONS:
        role, created = TaskRoleDefinition.objects.get_or_create(
            code=code,
            defaults={
                "name": name,
                "description": description,
                "is_system": True,
                "is_active": True,
                "order_index": order_index,
            },
        )
        if not created and not role.is_system:
            role.is_system = True
            role.save(update_fields=["is_system", "updated_at"])


def _task_role_definitions(*, include_inactive: bool = False) -> list[TaskRoleDefinition]:
    _ensure_system_task_roles()
    qs = TaskRoleDefinition.objects.all()
    if not include_inactive:
        qs = qs.filter(is_active=True)
    return list(qs.order_by("order_index", "name", "id"))


def _task_role_label_map() -> dict[str, str]:
    return {role.code: role.name for role in _task_role_definitions(include_inactive=True)}


def _filter_task_user_rows(users, query: str):
    users = list(users)
    query = str(query or "").strip().casefold()
    if not query:
        return users

    legacy_ids = []
    for user in users:
        profile = getattr(user, "profile", None)
        legacy_user_id = getattr(profile, "legacy_user_id", None) if profile is not None else None
        if legacy_user_id:
            legacy_ids.append(legacy_user_id)

    legacy_by_id = {}
    if legacy_ids:
        try:
            legacy_by_id = {
                row["id"]: row
                for row in UtenteLegacy.objects.filter(id__in=legacy_ids).values("id", "nome", "email", "ruolo")
            }
        except DatabaseError:
            legacy_by_id = {}

    def _haystack(user) -> str:
        parts = [
            user.get_full_name(),
            getattr(user, "first_name", ""),
            getattr(user, "last_name", ""),
            getattr(user, "username", ""),
            getattr(user, "email", ""),
        ]
        profile = getattr(user, "profile", None)
        if profile is not None:
            legacy_user_id = getattr(profile, "legacy_user_id", None)
            parts.append(str(legacy_user_id or ""))
            legacy_row = legacy_by_id.get(legacy_user_id)
            if legacy_row:
                parts.extend([legacy_row.get("nome"), legacy_row.get("email"), legacy_row.get("ruolo")])
        return " ".join(str(part or "") for part in parts).casefold()

    return [user for user in users if query in _haystack(user)]


def _task_settings_users_queryset():
    """Utenti visibili nelle impostazioni task, sincronizzati dal legacy.

    La schermata admin utenti mostra `utenti` legacy; le assegnazioni task usano
    invece FK Django. Qui creiamo/aggiorniamo i corrispondenti auth_user mancanti
    solo per le pagine impostazioni, cosi' la matrice ruoli resta completa.
    """
    try:
        active_legacy_users = list(
            UtenteLegacy.objects.filter(attivo=True).only("id", "nome", "email", "ruolo", "ruolo_id")
        )
    except DatabaseError:
        return task_active_users_queryset()

    if not active_legacy_users:
        return task_active_users_queryset()

    active_legacy_ids = [int(user.id) for user in active_legacy_users]
    mapped_ids = set(
        Profile.objects.filter(legacy_user_id__in=active_legacy_ids).values_list("legacy_user_id", flat=True)
    )
    for legacy_user in active_legacy_users:
        if int(legacy_user.id) in mapped_ids:
            continue
        try:
            sync_django_user_from_legacy(legacy_user)
        except Exception:
            logger.exception(
                "Impossibile sincronizzare utente legacy task legacy_user_id=%s",
                getattr(legacy_user, "id", None),
            )

    return task_active_users_queryset()


def _coerce_positive_int(value, *, default: int, minimum: int = 1) -> int:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError, AttributeError):
        parsed = default
    return max(minimum, parsed)


def _build_tasks_settings_context(request, *, tab: str) -> dict:
    from django.core.paginator import Paginator

    today = timezone.localdate()
    total_tasks = Task.objects.count()
    total_projects = Project.objects.count()
    tasks_by_status_raw = dict(Task.objects.values_list("status").annotate(n=Count("id")).order_by())
    tasks_overdue = Task.objects.filter(
        due_date__lt=today,
        status__in=[TaskStatus.TODO, TaskStatus.IN_PROGRESS],
    ).count()

    context: dict[str, object] = {
        "tab": tab,
        "total_tasks": total_tasks,
        "total_projects": total_projects,
        "tasks_overdue": tasks_overdue,
        "todo_count": tasks_by_status_raw.get(TaskStatus.TODO, 0),
        "in_progress_count": tasks_by_status_raw.get(TaskStatus.IN_PROGRESS, 0),
        "done_count": tasks_by_status_raw.get(TaskStatus.DONE, 0),
    }

    if tab == "riepilogo":
        context["tasks_by_status"] = [
            (value, label, tasks_by_status_raw.get(value, 0))
            for value, label in TaskStatus.choices
        ]
        context["top_assignees"] = list(
            Task.objects.filter(assigned_to__isnull=False)
            .values("assigned_to__username")
            .annotate(n=Count("id"))
            .order_by("-n")[:10]
        )

    if tab == "record":
        q_task = request.GET.get("q_task", "").strip()
        q_proj = request.GET.get("q_proj", "").strip()
        filter_status = request.GET.get("filter_status", "").strip()

        tasks_qs = Task.objects.select_related("project", "assigned_to").order_by("-updated_at")
        if q_task:
            tasks_qs = tasks_qs.filter(title__icontains=q_task)
        if filter_status:
            tasks_qs = tasks_qs.filter(status=filter_status)

        projects_qs = Project.objects.select_related("project_manager").order_by("-updated_at")
        if q_proj:
            projects_qs = projects_qs.filter(Q(name__icontains=q_proj) | Q(client_name__icontains=q_proj))

        context.update(
            {
                "tasks_page": Paginator(tasks_qs, 50).get_page(request.GET.get("task_page")),
                "projects_page": Paginator(projects_qs, 50).get_page(request.GET.get("proj_page")),
                "q_task": q_task,
                "q_proj": q_proj,
                "filter_status": filter_status,
                "task_status_choices": TaskStatus.choices,
            }
        )

    if tab == "log":
        context["audit_entries"] = AuditLog.objects.filter(modulo="tasks").order_by("-created_at")[:100]

    if tab == "ruoli":
        role_definitions = _task_role_definitions(include_inactive=True)
        ruoli_filter_q = request.GET.get("q_user", "").strip()
        all_users = list(_task_settings_users_queryset())
        filtered_users = _filter_task_user_rows(all_users, ruoli_filter_q)
        assignments_raw = list(
            TaskRoleAssignment.objects.values_list("user_id", "role_type")
        )
        by_user: dict[int, set[str]] = {}
        for uid, rtype in assignments_raw:
            by_user.setdefault(uid, set()).add(rtype)

        roster = []
        for u in filtered_users:
            user_label = (u.get_full_name() or u.username)
            assigned_roles = by_user.get(u.id, set())
            roster.append({
                "id": u.id,
                "label": user_label,
                "email": u.email or "",
                "role_cells": [
                    {"code": role.code, "checked": role.code in assigned_roles}
                    for role in role_definitions
                ],
            })
        context.update({
            "ruoli_filter_q": ruoli_filter_q,
            "ruoli_roster": roster,
            "ruoli_role_definitions": role_definitions,
            "ruoli_colspan": 2 + len(role_definitions),
            "ruoli_stats": {
                "roles": len(role_definitions),
                "assignments": sum(len(v) for v in by_user.values()),
                "total_users": len(all_users),
                "filtered_users": len(roster),
            },
        })

    if tab == "accessi":
        role_definitions = _task_role_definitions(include_inactive=True)
        access_filter_q = request.GET.get("q_access_user", "").strip()
        all_users = list(_task_settings_users_queryset())
        filtered_users = _filter_task_user_rows(all_users, access_filter_q)
        role_rule_map = {
            role_type: access_level
            for role_type, access_level in TaskRoleAccessRule.objects.values_list("role_type", "access_level")
        }
        user_rule_map = {
            user_id: access_level
            for user_id, access_level in TaskUserAccessRule.objects.values_list("user_id", "access_level")
        }
        context.update(
            {
                "access_filter_q": access_filter_q,
                "access_role_rows": [
                    {
                        "code": role.code,
                        "label": role.name,
                        "help": role.description or (
                            "Valida sui task dei tipi associati a questo ruolo."
                        ),
                        "is_system": role.is_system,
                        "access_level": role_rule_map.get(role.code, TaskAccessLevel.NONE),
                    }
                    for role in role_definitions
                ],
                "access_user_rows": [
                    {
                        "id": user.id,
                        "label": user.get_full_name() or user.username,
                        "email": user.email or "",
                        "access_level": user_rule_map.get(user.id, ""),
                    }
                    for user in filtered_users
                ],
                "access_role_choices": [
                    (TaskAccessLevel.NONE, "Nessun accesso extra"),
                    (TaskAccessLevel.READ_ALL, "Vede tutti i task del kickoff"),
                    (TaskAccessLevel.EDIT_ASSIGNED, "Vede tutto + modifica solo i task assegnati"),
                    (TaskAccessLevel.EDIT_ALL, "Vede e modifica tutto il kickoff"),
                ],
                "access_user_choices": [
                    ("", "Eredita scope standard"),
                    (TaskAccessLevel.READ_ALL, "Vede tutto il modulo"),
                    (TaskAccessLevel.EDIT_ASSIGNED, "Vede tutto + modifica solo i task assegnati"),
                    (TaskAccessLevel.EDIT_ALL, "Vede e modifica tutto il modulo"),
                ],
                "access_stats": {
                    "role_rules": len(role_rule_map),
                    "user_overrides": len(user_rule_map),
                    "edit_all_roles": sum(1 for level in role_rule_map.values() if level == TaskAccessLevel.EDIT_ALL),
                    "edit_all_users": sum(1 for level in user_rule_map.values() if level == TaskAccessLevel.EDIT_ALL),
                    "total_users": len(all_users),
                    "filtered_users": len(filtered_users),
                },
            }
        )

    if tab == "promemoria":
        q_filter = (request.GET.get("filter_status") or "pending").strip().lower()
        qs = (
            TaskReminder.objects.select_related("task", "task__project", "task__assigned_to")
            .order_by("fire_at", "id")
        )
        if q_filter == "pending":
            qs = qs.filter(fired=False)
        elif q_filter == "fired":
            qs = qs.filter(fired=True)
        # else: all

        today = timezone.localdate()
        reminders_list = []
        for r in qs[:200]:
            reminders_list.append({
                "id": r.id,
                "task_id": r.task_id,
                "task_title": r.task.title if r.task else "(task eliminato)",
                "project_name": (r.task.project.name if r.task and r.task.project_id else ""),
                "assignee": (
                    r.task.assigned_to.get_full_name() or r.task.assigned_to.username
                ) if r.task and r.task.assigned_to_id else "",
                "legacy_user_id": r.legacy_user_id,
                "fire_at": r.fire_at,
                "due_date": r.task.due_date if r.task else None,
                "fired": r.fired,
                "fired_at": r.fired_at,
                "overdue": (not r.fired) and r.fire_at < today,
            })
        context.update({
            "promemoria_filter": q_filter,
            "promemoria_list": reminders_list,
            "promemoria_count_pending": TaskReminder.objects.filter(fired=False).count(),
            "promemoria_count_fired":   TaskReminder.objects.filter(fired=True).count(),
            "promemoria_count_overdue": TaskReminder.objects.filter(fired=False, fire_at__lt=today).count(),
            "promemoria_count_next7":   TaskReminder.objects.filter(
                fired=False, fire_at__gte=today, fire_at__lte=today + timedelta(days=7),
            ).count(),
        })

    if tab == "tipi":
        role_definitions = _task_role_definitions(include_inactive=True)
        categories = list(
            TaskCategory.objects.prefetch_related("fields").order_by("order_index", "name")
        )
        focus_id_raw = (request.GET.get("cat") or "").strip()
        focus_category = None
        try:
            focus_id = int(focus_id_raw)
        except (TypeError, ValueError):
            focus_id = 0
        if focus_id:
            for cat in categories:
                if cat.id == focus_id:
                    focus_category = cat
                    break
        if focus_category is None and categories:
            focus_category = categories[0]

        asset_type_choices: list[tuple[str, str]] = []
        asset_categories: list[dict] = []
        try:
            from assets.models import Asset, AssetCategory
            asset_type_choices = list(Asset.TYPE_CHOICES)
            asset_categories = [
                {"id": c.id, "label": getattr(c, "label", "") or c.code}
                for c in AssetCategory.objects.filter(is_active=True).order_by("sort_order", "label", "id")
            ]
        except Exception:
            pass
        context.update({
            "tipi_categories": categories,
            "tipi_focus_category": focus_category,
            "tipi_field_types": TaskCategoryFieldType.choices,
            "tipi_role_choices": [("", "Nessun ruolo dedicato")] + [
                (role.code, role.name) for role in role_definitions if role.is_active
            ],
            "tipi_role_label_map": _task_role_label_map(),
            "tipi_tasks_with_category": Task.objects.filter(category__isnull=False).count(),
            "tipi_asset_type_choices": asset_type_choices,
            "tipi_asset_categories": asset_categories,
        })

    return context


def _parse_vrf_excel(file_obj) -> dict:
    """Estrae i campi identificativi da MOD.073 VRF Rev.10 (Sheet 'VRF').
    Celle fisse: B3=P/N, I3=Descrizione, P3=Esp, O2=Preventivo n°, P2=Versione n°, B4=Cliente.
    Restituisce dict con chiavi stringa, mai None.
    """
    import io
    import openpyxl

    def _s(val):
        return str(val or "").strip()

    result = {
        "part_number": "",
        "vrf_description": "",
        "vrf_esp": "",
        "vrf_quote_number": "",
        "versione": "",
        "client_name": "",
    }
    try:
        data = file_obj.read() if hasattr(file_obj, "read") else file_obj
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        # Preferenza sheet "VRF", altrimenti primo sheet
        ws = wb["VRF"] if "VRF" in wb.sheetnames else wb.active
        # Leggi le 4 righe di intestazione (1-indexed)
        rows = {}
        for r in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            pass  # warming
        # iter_rows con read_only non supporta accesso per cella — rileggiamo
        wb.close()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb["VRF"] if "VRF" in wb.sheetnames else wb.active
        result["vrf_quote_number"] = _s(ws.cell(row=2, column=15).value)  # O2
        result["versione"]         = _s(ws.cell(row=2, column=16).value)  # P2
        result["part_number"]      = _s(ws.cell(row=3, column=2).value)   # B3
        result["vrf_description"]  = _s(ws.cell(row=3, column=9).value)   # I3
        result["vrf_esp"]          = _s(ws.cell(row=3, column=16).value)  # P3
        result["client_name"]      = _s(ws.cell(row=4, column=2).value)   # B4
        wb.close()
    except Exception:
        pass
    return result


def _vrf_status_detail(project, cfg: "TaskImpostazioni") -> dict:
    """Calcola lo stato del documento VRF per un kickoff.

    Restituisce dict con:
      status: 'ok' | 'na' | 'pending' | 'warning' | 'blocked'
      label: str (messaggio utente)
      is_blocked: bool
      days_pending: int | None
      show_upload_cta: bool
    """
    from django.utils import timezone as tz

    vrf_status = project.vrf_status

    if vrf_status == VRFDocStatus.UPLOADED:
        return {
            "status": "ok",
            "label": "Documento VRF caricato.",
            "is_blocked": False,
            "days_pending": None,
            "show_upload_cta": False,
        }

    if vrf_status == VRFDocStatus.NOT_REQUIRED:
        return {
            "status": "na",
            "label": "Documento VRF non richiesto per questo kickoff.",
            "is_blocked": False,
            "days_pending": None,
            "show_upload_cta": False,
        }

    # PENDING
    today = tz.localdate()
    created_date = project.created_at.date() if project.created_at else today
    days = (today - created_date).days

    if days >= cfg.vrf_blocking_days:
        return {
            "status": "blocked",
            "label": (
                f"Kickoff BLOCCATO: documento VRF non caricato da {days} giorni "
                f"(soglia blocco: {cfg.vrf_blocking_days} giorni). "
                "Non e possibile aggiungere o modificare attivita kickoff finche il documento non viene caricato."
            ),
            "is_blocked": True,
            "days_pending": days,
            "show_upload_cta": True,
        }

    if days >= cfg.vrf_reminder_days:
        return {
            "status": "warning",
            "label": (
                f"Attenzione: documento VRF non ancora caricato ({days} giorni dalla creazione del kickoff). "
                f"Il kickoff sara bloccato dopo {cfg.vrf_blocking_days} giorni."
            ),
            "is_blocked": False,
            "days_pending": days,
            "show_upload_cta": True,
        }

    return {
        "status": "pending",
        "label": f"Documento VRF non ancora caricato ({days} giorni dalla creazione del kickoff).",
        "is_blocked": False,
        "days_pending": days,
        "show_upload_cta": True,
    }


def _copy_project_vrf_file(project: Project, *, clear_part_number: bool = False) -> tuple[bytes | None, str]:
    if not project.vrf_file:
        return None, ""

    import io
    from pathlib import Path

    project.vrf_file.open("rb")
    try:
        file_bytes = project.vrf_file.read()
    finally:
        project.vrf_file.close()

    filename = project.vrf_original_name or Path(project.vrf_file.name).name

    if clear_part_number:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        try:
            worksheet = workbook["VRF"] if "VRF" in workbook.sheetnames else workbook.active
            worksheet["B3"] = ""
            output = io.BytesIO()
            workbook.save(output)
            file_bytes = output.getvalue()
        finally:
            workbook.close()

    return file_bytes, filename


def _duplicate_kickoff(source_project: Project, *, created_by, clear_part_number: bool = False) -> Project:
    from django.core.files.base import ContentFile

    kickoff = Project.objects.create(
        name="",
        description=source_project.description,
        client_name=source_project.client_name,
        project_manager=source_project.project_manager,
        capo_commessa=source_project.capo_commessa,
        programmer=source_project.programmer,
        control_method=source_project.control_method,
        part_number="" if clear_part_number else source_project.part_number,
        revisione=source_project.revisione,
        versione=source_project.versione,
        vrf_status=source_project.vrf_status,
        vrf_original_name=source_project.vrf_original_name,
        vrf_uploaded_at=source_project.vrf_uploaded_at,
        vrf_quote_number=source_project.vrf_quote_number,
        vrf_description=source_project.vrf_description,
        vrf_esp=source_project.vrf_esp,
        similar_project=source_project.similar_project,
        similar_work_note=source_project.similar_work_note,
        created_by=created_by,
    )

    file_bytes, filename = _copy_project_vrf_file(source_project, clear_part_number=clear_part_number)
    if file_bytes is not None:
        kickoff.vrf_file.save(filename, ContentFile(file_bytes), save=False)
        kickoff.vrf_original_name = filename
        kickoff.save(update_fields=["vrf_file", "vrf_original_name", "updated_at"])

    return kickoff


def _request_legacy_user(request):
    legacy_user = getattr(request, "legacy_user", None)
    if legacy_user is None:
        legacy_user = get_legacy_user(request.user)
        request.legacy_user = legacy_user
    return legacy_user


def _check_task_action_for_legacy(legacy_user, action_code: str) -> bool:
    """ACL action check locale app tasks (riuso cache permessi + override per-utente)."""
    if not legacy_user:
        return False
    if is_legacy_admin(legacy_user):
        return True

    role_id = getattr(legacy_user, "ruolo_id", None)
    if not role_id:
        return False

    action_norm = str(action_code or "").strip()
    if not action_norm:
        return False

    try:
        from core.models import UserPermissionOverride

        override = UserPermissionOverride.objects.filter(
            legacy_user_id=int(legacy_user.id),
            modulo__iexact=TASK_MODULE_CODE,
            azione__iexact=action_norm,
        ).first()
        if override is not None and override.can_view is not None:
            return bool(override.can_view)
    except Exception:
        pass

    perm_map = get_cached_perm_map(int(role_id))
    return bool(perm_map.get((TASK_MODULE_CODE, action_norm.lower()), False))


def _has_task_permission(request, action_code: str) -> bool:
    cache = getattr(request, "_task_perm_cache", None)
    if cache is None:
        cache = {}
        request._task_perm_cache = cache

    key = str(action_code or "").strip().lower()
    if key in cache:
        return bool(cache[key])

    if not request.user.is_authenticated:
        cache[key] = False
        return False
    if request.user.is_superuser:
        cache[key] = True
        return True

    legacy_user = _request_legacy_user(request)
    allowed = _check_task_action_for_legacy(legacy_user, action_code)
    cache[key] = bool(allowed)
    return bool(allowed)


def task_permissions_required(*action_codes: str):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect_to_login(request.get_full_path())
            for action_code in action_codes:
                if not _has_task_permission(request, action_code):
                    return render(
                        request,
                        "core/pages/forbidden.html",
                        {"page_title": "Accesso negato"},
                        status=403,
                    )
            return view_func(request, *args, **kwargs)

        return _wrapped

    return decorator


def _task_access_rank(level: str | None) -> int:
    return int(TASK_ACCESS_LEVEL_ORDER.get(str(level or "").strip(), 0))


def _task_access_allows_read(level: str | None) -> bool:
    return _task_access_rank(level) >= _task_access_rank(TaskAccessLevel.READ_ALL)


def _task_access_allows_edit_assigned(level: str | None) -> bool:
    return _task_access_rank(level) >= _task_access_rank(TaskAccessLevel.EDIT_ASSIGNED)


def _task_access_allows_edit_all(level: str | None) -> bool:
    return _task_access_rank(level) >= _task_access_rank(TaskAccessLevel.EDIT_ALL)


def _request_task_role_access_map(request) -> dict[str, str]:
    cached = getattr(request, "_task_role_access_map", None)
    if cached is not None:
        return cached
    cached = {
        role_type: access_level
        for role_type, access_level in TaskRoleAccessRule.objects.values_list("role_type", "access_level")
    }
    request._task_role_access_map = cached
    return cached


def _request_task_user_access_level(request) -> str:
    cached = getattr(request, "_task_user_access_level", None)
    if cached is not None:
        return cached
    if not getattr(request.user, "is_authenticated", False):
        cached = TaskAccessLevel.NONE
    else:
        cached = (
            TaskUserAccessRule.objects.filter(user=request.user)
            .values_list("access_level", flat=True)
            .first()
            or TaskAccessLevel.NONE
        )
    request._task_user_access_level = cached
    return cached


def _request_task_user_role_codes(request) -> set[str]:
    cached = getattr(request, "_task_user_role_codes", None)
    if cached is not None:
        return cached
    if not getattr(request.user, "is_authenticated", False):
        cached = set()
    else:
        cached = set(
            TaskRoleAssignment.objects.filter(user=request.user)
            .values_list("role_type", flat=True)
        )
    request._task_user_role_codes = cached
    return cached


def _project_role_access_level(request, project: Project | None) -> str:
    if project is None or not getattr(project, "pk", None):
        return TaskAccessLevel.NONE
    if not getattr(request.user, "is_authenticated", False):
        return TaskAccessLevel.NONE

    role_map = _request_task_role_access_map(request)
    levels: list[str] = []
    user_id = request.user.id
    if getattr(project, "project_manager_id", None) == user_id:
        levels.append(role_map.get(TaskRoleType.PROJECT_MANAGER, TaskAccessLevel.NONE))
    if getattr(project, "capo_commessa_id", None) == user_id:
        levels.append(role_map.get(TaskRoleType.CAPO_COMMESSA, TaskAccessLevel.NONE))
    if getattr(project, "programmer_id", None) == user_id:
        levels.append(role_map.get(TaskRoleType.PROGRAMMER, TaskAccessLevel.NONE))
    if not levels:
        return TaskAccessLevel.NONE
    return max(levels, key=_task_access_rank)


def _task_role_access_level(request, task: Task | None) -> str:
    if task is None:
        return TaskAccessLevel.NONE
    levels = [_project_role_access_level(request, getattr(task, "project", None))]
    role_type = getattr(task, "category", None)
    role_type = getattr(role_type, "role_type", "") or ""
    if role_type and role_type in _request_task_user_role_codes(request):
        levels.append(_request_task_role_access_map(request).get(role_type, TaskAccessLevel.NONE))
    return max(levels, key=_task_access_rank)


def _task_scope_filter_q(request) -> Q:
    user = request.user
    q = Q(created_by=user) | Q(assigned_to=user) | Q(subscribers=user)
    role_map = _request_task_role_access_map(request)
    if _task_access_allows_read(role_map.get(TaskRoleType.PROJECT_MANAGER)):
        q |= Q(project__project_manager=user)
    if _task_access_allows_read(role_map.get(TaskRoleType.CAPO_COMMESSA)):
        q |= Q(project__capo_commessa=user)
    if _task_access_allows_read(role_map.get(TaskRoleType.PROGRAMMER)):
        q |= Q(project__programmer=user)
    category_role_codes = [
        role_code
        for role_code in _request_task_user_role_codes(request)
        if _task_access_allows_read(role_map.get(role_code))
    ]
    if category_role_codes:
        q |= Q(category__role_type__in=category_role_codes)
    return q


def _project_scope_filter_q(request) -> Q:
    user = request.user
    q = (
        Q(created_by=user)
        | Q(tasks__created_by=user)
        | Q(tasks__assigned_to=user)
        | Q(tasks__subscribers=user)
    )
    role_map = _request_task_role_access_map(request)
    if _task_access_allows_read(role_map.get(TaskRoleType.PROJECT_MANAGER)):
        q |= Q(project_manager=user)
    if _task_access_allows_read(role_map.get(TaskRoleType.CAPO_COMMESSA)):
        q |= Q(capo_commessa=user)
    if _task_access_allows_read(role_map.get(TaskRoleType.PROGRAMMER)):
        q |= Q(programmer=user)
    category_role_codes = [
        role_code
        for role_code in _request_task_user_role_codes(request)
        if _task_access_allows_read(role_map.get(role_code))
    ]
    if category_role_codes:
        q |= Q(tasks__category__role_type__in=category_role_codes)
    return q


def _scoped_tasks_queryset(request):
    qs = Task.objects.select_related(
        "created_by",
        "assigned_to",
        "category",
        "project",
        "project__project_manager",
        "project__capo_commessa",
        "project__programmer",
        "project__similar_project",
    )
    if _has_task_permission(request, "tasks_admin"):
        return qs
    if _task_access_allows_read(_request_task_user_access_level(request)):
        return qs
    return qs.filter(_task_scope_filter_q(request)).distinct()


def _scoped_projects_queryset(request):
    qs = Project.objects.select_related(
        "created_by",
        "project_manager",
        "capo_commessa",
        "programmer",
        "similar_project",
    )
    if _has_task_permission(request, "tasks_admin"):
        return qs
    if _task_access_allows_read(_request_task_user_access_level(request)):
        return qs
    return qs.filter(_project_scope_filter_q(request)).distinct()


def _detail_queryset(request):
    return _scoped_tasks_queryset(request).prefetch_related(
        "subscribers",
        Prefetch("subtasks", queryset=SubTask.objects.select_related("assigned_to").order_by("order_index", "id")),
        Prefetch("comments", queryset=TaskComment.objects.select_related("author", "target_user").order_by("-created_at", "-id")),
        Prefetch("events", queryset=TaskEvent.objects.select_related("actor").order_by("-created_at", "-id")),
        Prefetch("attachments", queryset=TaskAttachment.objects.select_related("uploaded_by").order_by("-created_at", "-id")),
        Prefetch(
            "project__attachments",
            queryset=TaskAttachment.objects.select_related("uploaded_by").order_by("-created_at", "-id"),
        ),
    )


def _apply_default_ordering(qs):
    return qs.order_by(
        F("next_step_due").asc(nulls_last=True),
        F("due_date").asc(nulls_last=True),
        "-updated_at",
    )


def _tasks_shell_context(request, *, active: str, task: Task | None = None, project: Project | None = None) -> dict:
    current_project = project
    if current_project is None and task is not None and getattr(task, "project_id", None):
        current_project = task.project
    return {
        "tasks_shell_active": active,
        "tasks_shell_task": task,
        "tasks_shell_project": current_project,
        "tasks_shell_can_create": _has_task_permission(request, "tasks_create"),
        "tasks_shell_can_admin": user_can_modulo_action(request, "tasks", "admin_tasks"),
        "tasks_shell_is_scope_admin": _has_task_permission(request, "tasks_admin"),
    }


def _json_safe(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _log_event(task: Task, actor, event_type: str, payload: dict | None = None) -> None:
    TaskEvent.objects.create(
        task=task,
        actor=actor,
        type=event_type,
        payload=payload or {},
    )


def _task_snapshot(task: Task) -> dict:
    return {
        "status": task.status,
        "assigned_to_id": task.assigned_to_id,
        "title": task.title,
        "priority": task.priority,
        "due_date": task.due_date,
        "next_step_text": task.next_step_text,
        "next_step_due": task.next_step_due,
        "tags": task.tags,
        "project_id": task.project_id,
    }


def _task_notify_users_queryset(task: Task):
    user_ids: set[int] = {task.created_by_id}
    if task.assigned_to_id:
        user_ids.add(task.assigned_to_id)
    user_ids.update(task.subscribers.values_list("id", flat=True))
    user_ids.discard(None)
    return User.objects.filter(id__in=user_ids).order_by("first_name", "last_name", "username")


def _project_notify_users_queryset(project: Project):
    task_rows = project.tasks.values_list("created_by_id", "assigned_to_id")
    user_ids: set[int] = {project.created_by_id}
    for created_by_id, assigned_to_id in task_rows:
        if created_by_id:
            user_ids.add(created_by_id)
        if assigned_to_id:
            user_ids.add(assigned_to_id)
    user_ids.update(project.tasks.values_list("subscribers__id", flat=True))
    user_ids.discard(None)
    return User.objects.filter(id__in=user_ids).order_by("first_name", "last_name", "username")


def _legacy_user_id_for_user(user) -> int | None:
    profile = getattr(user, "profile", None)
    if profile and getattr(profile, "legacy_user_id", None):
        try:
            return int(profile.legacy_user_id)
        except Exception:
            return None
    return None


def _notify_user(target_user, *, message_text: str, action_url: str = "") -> None:
    target_legacy_user_id = _legacy_user_id_for_user(target_user)
    if not target_legacy_user_id:
        return
    Notifica.objects.create(
        legacy_user_id=target_legacy_user_id,
        tipo="generico",
        messaggio=str(message_text or "")[:500],
        url_azione=str(action_url or "")[:255],
    )


def _legacy_fetch_all_dict(sql: str, params: list | tuple | None = None) -> list[dict]:
    with connections["default"].cursor() as cursor:
        cursor.execute(sql, params or [])
        columns = [str(col[0]) for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _coerce_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        pass
    try:
        return date.fromisoformat(text[:10])
    except Exception:
        return None


def _absence_status_label(row: dict) -> str:
    raw_status = row.get("moderation_status")
    try:
        parsed = int(raw_status)
    except (TypeError, ValueError):
        parsed = None
    status_map = {
        0: "Approvato",
        1: "Rifiutato",
        2: "In attesa",
        3: "Bozza",
        4: "Programmato",
    }
    if parsed is not None:
        return status_map.get(parsed, "In attesa")
    consenso = str(row.get("consenso") or "").strip()
    if consenso:
        return consenso
    return "In attesa"


def _load_user_absences(user, *, date_start: date, date_end: date) -> list[dict]:
    if not user or date_end < date_start:
        return []

    columns = legacy_table_columns("assenze")
    if not columns:
        legacy_table_columns.cache_clear()
        columns = legacy_table_columns("assenze")
    if not columns or "data_inizio" not in columns or "data_fine" not in columns:
        return []

    full_name = str(user.get_full_name() or user.get_username() or "").strip()
    email = str(getattr(user, "email", "") or "").strip()

    identity_clauses: list[str] = []
    identity_params: list = []
    if "copia_nome" in columns and full_name:
        identity_clauses.append("UPPER(COALESCE(copia_nome,'')) = UPPER(%s)")
        identity_params.append(full_name)
    if "email_esterna" in columns and email:
        identity_clauses.append("UPPER(COALESCE(email_esterna,'')) = UPPER(%s)")
        identity_params.append(email)
    if not identity_clauses:
        return []

    range_start = datetime.combine(date_start, datetime.min.time())
    range_end = datetime.combine(date_end, datetime.max.time())

    status_clause = ""
    status_params: list = []
    if "moderation_status" in columns:
        status_clause = "AND COALESCE(moderation_status, 2) != %s"
        status_params.append(1)
    elif "consenso" in columns:
        status_clause = "AND UPPER(COALESCE(consenso,'')) NOT LIKE UPPER(%s)"
        status_params.append("%rifiut%")

    tipo_select = "tipo_assenza" if "tipo_assenza" in columns else "'' AS tipo_assenza"
    consenso_select = "consenso" if "consenso" in columns else "'' AS consenso"
    moderation_select = "moderation_status" if "moderation_status" in columns else "NULL AS moderation_status"

    sql = f"""
        SELECT
            id,
            data_inizio,
            data_fine,
            {tipo_select},
            {consenso_select},
            {moderation_select}
        FROM assenze
        WHERE ({' OR '.join(identity_clauses)})
          AND data_inizio IS NOT NULL
          AND data_fine IS NOT NULL
          AND data_fine >= %s
          AND data_inizio <= %s
          {status_clause}
        ORDER BY data_inizio ASC, id ASC
    """

    try:
        rows = _legacy_fetch_all_dict(sql, [*identity_params, range_start, range_end, *status_params])
    except DatabaseError:
        return []
    except Exception:
        return []

    normalized: list[dict] = []
    for row in rows:
        start_date = _coerce_date(row.get("data_inizio"))
        end_date = _coerce_date(row.get("data_fine"))
        if not start_date or not end_date:
            continue
        if end_date < start_date:
            start_date, end_date = end_date, start_date
        normalized.append(
            {
                "id": row.get("id"),
                "start_date": start_date,
                "end_date": end_date,
                "tipo": str(row.get("tipo_assenza") or "Assenza").strip() or "Assenza",
                "status": _absence_status_label(row),
            }
        )
    return normalized


def _task_date_absence_conflicts(task: Task) -> dict[str, list[dict]]:
    if not task.assigned_to_id:
        return {}

    target_dates: dict[str, date] = {}
    if task.due_date:
        target_dates["due_date"] = task.due_date
    if task.next_step_due:
        target_dates["next_step_due"] = task.next_step_due
    if not target_dates:
        return {}

    date_start = min(target_dates.values())
    date_end = max(target_dates.values())
    absences = _load_user_absences(task.assigned_to, date_start=date_start, date_end=date_end)
    if not absences:
        return {}

    conflicts: dict[str, list[dict]] = {}
    for field_name, target_date in target_dates.items():
        field_conflicts = [row for row in absences if row["start_date"] <= target_date <= row["end_date"]]
        if field_conflicts:
            conflicts[field_name] = field_conflicts
    return conflicts


def _add_task_absence_warnings(request, task: Task) -> None:
    conflicts = _task_date_absence_conflicts(task)
    if not conflicts or not task.assigned_to_id:
        return

    assignee_name = task.assigned_to.get_full_name() or task.assigned_to.get_username()
    field_labels = {
        "due_date": "data prevista conclusione",
        "next_step_due": "data obiettivo prossima azione",
    }
    for field_name in ("next_step_due", "due_date"):
        if field_name not in conflicts:
            continue
        target_date = getattr(task, field_name, None)
        if not target_date:
            continue
        unique_labels: list[str] = []
        for entry in conflicts[field_name]:
            label = f"{entry['tipo']} ({entry['status']})"
            if label not in unique_labels:
                unique_labels.append(label)
        label_text = ", ".join(unique_labels[:2])
        if len(unique_labels) > 2:
            label_text += ", ..."
        messages.warning(
            request,
            (
                f"Conflitto assenze: {field_labels[field_name]} del {target_date.strftime('%d/%m/%Y')} "
                f"coincide con assenza di {assignee_name} ({label_text})."
            ),
        )


def _sync_task_integrations(request, task: Task, form, *, action: str) -> None:
    """Dopo save: allinea evento Outlook + reminder portale secondo il form submit.

    `action` e' "create" o "edit" - controlla solo messages che vogliamo mostrare.
    Non blocca mai il flusso: errori Graph producono messages.warning ma il task resta salvato.
    """
    from .outlook_reminder import sync_task_outlook_event, sync_task_portal_reminder

    requested = bool(form.cleaned_data.get("add_to_outlook"))
    explicit_email = form.cleaned_data.get("outlook_target_email") or ""
    try:
        level, msg = sync_task_outlook_event(
            request=request, task=task, requested=requested, explicit_email=str(explicit_email),
        )
        if msg:
            if level == "success":
                messages.success(request, msg)
            elif level == "info":
                messages.info(request, msg)
            elif level == "warning":
                messages.warning(request, msg)
            elif level == "error":
                messages.error(request, msg)
    except Exception as exc:
        messages.warning(request, f"Integrazione Outlook non disponibile: {exc}")

    try:
        sync_task_portal_reminder(task=task)
    except Exception as exc:
        # Fire-and-forget: non bloccare il save del task
        logger.warning("Task reminder sync fallita (task=%s): %s", task.id, exc)


def _build_task_absence_day_map(tasks: list[Task], *, timeline_start: date, timeline_end: date) -> dict[int, dict[date, list[dict]]]:
    if timeline_end < timeline_start:
        return {}

    user_absence_cache: dict[int, list[dict]] = {}
    for task in tasks:
        if not task.assigned_to_id or task.assigned_to_id in user_absence_cache:
            continue
        user_absence_cache[task.assigned_to_id] = _load_user_absences(
            task.assigned_to,
            date_start=timeline_start,
            date_end=timeline_end,
        )

    per_task_day_map: dict[int, dict[date, list[dict]]] = {}
    for task in tasks:
        task_day_map: dict[date, list[dict]] = {}
        for absence in user_absence_cache.get(task.assigned_to_id, []):
            start_date = max(absence["start_date"], timeline_start)
            end_date = min(absence["end_date"], timeline_end)
            if end_date < start_date:
                continue
            current = start_date
            while current <= end_date:
                task_day_map.setdefault(current, []).append(absence)
                current += timedelta(days=1)
        per_task_day_map[task.id] = task_day_map
    return per_task_day_map


def _query_bool(query_data, key: str, *, default: bool = True) -> bool:
    raw = query_data.get(key)
    if raw is None:
        return bool(default)
    return str(raw).strip().lower() not in {"0", "false", "off", "no", ""}


def _parse_gantt_options(request):
    query_data = request.GET

    valid_windows = {row[0] for row in GANTT_WINDOW_OPTIONS}
    valid_cells = {row[0] for row in GANTT_CELL_OPTIONS}
    valid_name_widths = {row[0] for row in GANTT_NAME_WIDTH_OPTIONS}

    try:
        window_days = int(query_data.get("window_days", 31))
    except (TypeError, ValueError):
        window_days = 31
    if window_days not in valid_windows:
        window_days = 31

    cell_size = str(query_data.get("cell_size", "m") or "m").strip().lower()
    if cell_size not in valid_cells:
        cell_size = "m"

    try:
        name_width = int(query_data.get("name_width", 360))
    except (TypeError, ValueError):
        name_width = 360
    if name_width not in valid_name_widths:
        name_width = 360

    day_cell_px_map = {"s": 22, "m": 30, "l": 38}

    return {
        "window_days": window_days,
        "cell_size": cell_size,
        "name_width": name_width,
        "show_wbs": _query_bool(query_data, "show_wbs", default=True),
        "show_duration": _query_bool(query_data, "show_duration", default=True),
        "show_start": _query_bool(query_data, "show_start", default=True),
        "show_end": _query_bool(query_data, "show_end", default=True),
        "day_cell_px": day_cell_px_map[cell_size],
        "window_choices": GANTT_WINDOW_OPTIONS,
        "cell_choices": GANTT_CELL_OPTIONS,
        "name_width_choices": GANTT_NAME_WIDTH_OPTIONS,
        "return_qs": query_data.urlencode(),
    }


def _easter_date(year: int) -> date:
    """Calcola la domenica di Pasqua (algoritmo gregoriano anonimo)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _italian_holidays(year: int) -> set[date]:
    """Restituisce le festività nazionali italiane fisse + Pasqua/Lunedì dell'Angelo."""
    easter = _easter_date(year)
    return {
        date(year, 1, 1),   # Capodanno
        date(year, 1, 6),   # Epifania
        easter,              # Pasqua
        easter + timedelta(days=1),  # Lunedì dell'Angelo
        date(year, 4, 25),  # Liberazione
        date(year, 5, 1),   # Festa del lavoro
        date(year, 6, 2),   # Repubblica
        date(year, 8, 15),  # Ferragosto
        date(year, 11, 1),  # Ognissanti
        date(year, 12, 8),  # Immacolata Concezione
        date(year, 12, 25), # Natale
        date(year, 12, 26), # Santo Stefano
    }


def _build_holidays_set(start: date, end: date) -> set[date]:
    """Costruisce l'insieme di tutte le festività nell'arco temporale dato."""
    holidays: set[date] = set()
    for year in range(start.year, end.year + 1):
        holidays |= _italian_holidays(year)
    return holidays


def _is_non_working_day(d: date, holidays: set[date]) -> bool:
    return d.weekday() >= 5 or d in holidays


def _count_working_days(start: date, end: date, holidays: set[date]) -> int:
    return sum(1 for i in range((end - start).days + 1) if not _is_non_working_day(start + timedelta(days=i), holidays))


def _project_gantt_rows(tasks: list[Task], *, min_window_days: int = 31) -> dict:
    today = timezone.localdate()
    timeline_points: list[date] = [today]
    base_rows: list[dict] = []

    for index, task in enumerate(tasks, start=1):
        start = task.next_step_due or task.created_at.date()
        end = task.due_date or task.next_step_due or start
        active_start = min(start, end)
        active_end = max(start, end)
        invalid_range = bool(task.next_step_due and task.due_date and task.due_date <= task.next_step_due)

        timeline_points.extend([active_start, active_end])
        base_rows.append(
            {
                "task": task,
                "wbs": str(index),
                "start": start,
                "end": end,
                "active_start": active_start,
                "active_end": active_end,
                "invalid_range": invalid_range,
                "duration_days": max(1, (active_end - active_start).days + 1),
            }
        )

    if base_rows:
        timeline_start = min(timeline_points) - timedelta(days=1)
        timeline_end = max(timeline_points) + timedelta(days=1)
    else:
        timeline_start = today - timedelta(days=3)
        timeline_end = today + timedelta(days=10)

    if min_window_days and min_window_days > 0:
        current_days = max(1, (timeline_end - timeline_start).days + 1)
        if current_days < min_window_days:
            extra_days = min_window_days - current_days
            left_extra = extra_days // 2
            right_extra = extra_days - left_extra
            timeline_start -= timedelta(days=left_extra)
            timeline_end += timedelta(days=right_extra)

    total_days = max(1, (timeline_end - timeline_start).days + 1)
    days = [timeline_start + timedelta(days=offset) for offset in range(total_days)]
    holidays = _build_holidays_set(timeline_start, timeline_end)
    day_columns = [
        {
            "date": day_value,
            "is_weekend": day_value.weekday() >= 5,
            "is_holiday": day_value in holidays and day_value.weekday() < 5,
            "is_today": day_value == today,
        }
        for day_value in days
    ]

    month_spans = []
    if days:
        current_month = (days[0].year, days[0].month)
        span_start = 0
        for index, day_value in enumerate(days):
            key = (day_value.year, day_value.month)
            if key != current_month:
                month_spans.append(
                    {
                        "label": f"{MONTH_LABELS_IT.get(current_month[1], current_month[1])} {current_month[0]}",
                        "span": index - span_start,
                    }
                )
                current_month = key
                span_start = index
        month_spans.append(
            {
                "label": f"{MONTH_LABELS_IT.get(current_month[1], current_month[1])} {current_month[0]}",
                "span": len(days) - span_start,
            }
        )

    task_absence_days = _build_task_absence_day_map(tasks, timeline_start=timeline_start, timeline_end=timeline_end)

    rows: list[dict] = []
    for row in base_rows:
        start_index = max(0, (row["active_start"] - timeline_start).days)
        end_index = min(len(days) - 1, (row["active_end"] - timeline_start).days)
        row_absence_days = task_absence_days.get(row["task"].id, {})
        conflict_dates: list[date] = [
            day_value
            for day_value in sorted(row_absence_days.keys())
            if row["active_start"] <= day_value <= row["active_end"]
        ]
        cells = []
        for day_index, day_value in enumerate(days):
            classes = []
            marker = ""
            title = ""
            is_wknd = day_value.weekday() >= 5
            is_hol = day_value in holidays and not is_wknd
            if is_wknd:
                classes.append("is-weekend")
            if is_hol:
                classes.append("is-holiday")
            if day_value == today:
                classes.append("is-today")
            if start_index <= day_index <= end_index:
                classes.append("is-active")
                classes.append(f"status-{row['task'].status.lower()}")
                if row["invalid_range"]:
                    classes.append("is-invalid-range")
            absence_entries = row_absence_days.get(day_value, [])
            if absence_entries and start_index <= day_index <= end_index:
                classes.append("is-absence")
                marker = "X"
                labels = [f"{entry['tipo']} ({entry['status']})" for entry in absence_entries]
                title = "; ".join(dict.fromkeys(labels))
            cells.append(
                {
                    "classes": " ".join(classes),
                    "marker": marker,
                    "title": title,
                }
            )
        row["cells"] = cells
        row["start_index"] = start_index
        row["end_index"] = end_index
        row["absence_days"] = len(conflict_dates)
        row["absence_dates"] = conflict_dates[:6]
        row["has_absence_conflicts"] = bool(conflict_dates)
        row["duration_working_days"] = _count_working_days(row["active_start"], row["active_end"], holidays)
        rows.append(row)

    # Predecessore implicito (WBS order): ogni task conosce l'end_index del task precedente
    # per consentire il vincolo "task N non può iniziare prima della fine di task N-1" nel Gantt.
    for i, row in enumerate(rows):
        row["prev_end_index"] = rows[i - 1]["end_index"] if i > 0 else -1

    return {
        "rows": rows,
        "timeline_start": timeline_start,
        "timeline_end": timeline_end,
        "today": today,
        "days": day_columns,
        "month_spans": month_spans,
    }


def _can_edit_project_schedule(request, project: Project) -> bool:
    if _has_task_permission(request, "tasks_admin"):
        return True
    if project.created_by_id == request.user.id:
        return True
    global_level = _request_task_user_access_level(request)
    if _task_access_allows_edit_all(global_level):
        return True
    role_level = _project_role_access_level(request, project)
    if _task_access_allows_edit_all(role_level):
        return True
    if _task_access_allows_edit_assigned(global_level) and project.tasks.filter(assigned_to=request.user).exists():
        return True
    if _task_access_allows_edit_assigned(role_level) and project.tasks.filter(assigned_to=request.user).exists():
        return True
    return False


def _can_manage_project(request, project: Project) -> bool:
    if _has_task_permission(request, "tasks_admin"):
        return True
    if project.created_by_id == request.user.id:
        return True
    global_level = _request_task_user_access_level(request)
    if _task_access_allows_edit_all(global_level):
        return True
    role_level = _project_role_access_level(request, project)
    return _task_access_allows_edit_all(role_level)


def _is_project_lead_for_task(user, task: Task) -> bool:
    return bool(task.project_id and task.project and task.project.created_by_id == getattr(user, "id", None))


def _can_manage_task(request, task: Task) -> bool:
    if _has_task_permission(request, "tasks_admin"):
        return True
    if task.created_by_id == getattr(request.user, "id", None):
        return True
    if _is_project_lead_for_task(request.user, task):
        return True
    global_level = _request_task_user_access_level(request)
    if _task_access_allows_edit_all(global_level):
        return True
    role_level = _task_role_access_level(request, task)
    if _task_access_allows_edit_all(role_level):
        return True
    if task.assigned_to_id == getattr(request.user, "id", None):
        return _task_access_allows_edit_assigned(global_level) or _task_access_allows_edit_assigned(role_level)
    return False


def _can_update_task_due_date(request, task: Task) -> bool:
    if _has_task_permission(request, "tasks_admin"):
        return True
    if _can_manage_task(request, task):
        return True
    return bool(task.assigned_to_id and task.assigned_to_id == request.user.id)


def _compute_rollup_status(task: Task) -> str | None:
    subtask_statuses = list(task.subtasks.values_list("status", flat=True))
    if not subtask_statuses:
        return None
    if all(status == TaskStatus.DONE for status in subtask_statuses):
        return TaskStatus.DONE
    if all(status == TaskStatus.CANCELED for status in subtask_statuses):
        return TaskStatus.CANCELED
    if any(status == TaskStatus.IN_PROGRESS for status in subtask_statuses):
        return TaskStatus.IN_PROGRESS
    has_done = any(status == TaskStatus.DONE for status in subtask_statuses)
    has_todo = any(status == TaskStatus.TODO for status in subtask_statuses)
    if has_done and has_todo:
        return TaskStatus.IN_PROGRESS
    return TaskStatus.TODO


def _apply_subtask_rollup(task: Task, actor) -> None:
    next_status = _compute_rollup_status(task)
    if not next_status or next_status == task.status:
        return
    old_status = task.status
    task.status = next_status
    task.save(update_fields=["status", "updated_at"])
    _log_event(
        task,
        actor,
        TaskEventType.STATUS_CHANGE,
        {"from": old_status, "to": next_status, "source": "subtask_rollup"},
    )


def _log_task_update_events(task: Task, actor, before: dict) -> None:
    if before.get("status") != task.status:
        _log_event(
            task,
            actor,
            TaskEventType.STATUS_CHANGE,
            {"from": before.get("status"), "to": task.status},
        )

    if before.get("assigned_to_id") != task.assigned_to_id:
        _log_event(
            task,
            actor,
            TaskEventType.ASSIGNMENT_CHANGE,
            {"from_user_id": before.get("assigned_to_id"), "to_user_id": task.assigned_to_id},
        )

    key_changes = {}
    for field_name in KEY_EDIT_FIELDS:
        old_value = before.get(field_name)
        new_value = getattr(task, field_name)
        if old_value != new_value:
            key_changes[field_name] = {
                "from": _json_safe(old_value),
                "to": _json_safe(new_value),
            }
    if key_changes:
        _log_event(
            task,
            actor,
            TaskEventType.EDIT,
            {"changes": key_changes},
        )


@task_permissions_required("tasks_view")
def task_list(request):
    query_data = request.GET.copy()
    mine_raw = (query_data.get("mine") or "").strip().lower()
    mine_explicit_false = mine_raw in {"0", "false", "off", "no"}
    if mine_explicit_false:
        query_data.pop("mine", None)
    if not query_data and not mine_explicit_false:
        query_data["mine"] = "1"

    projects_qs = _scoped_projects_queryset(request).order_by("name", "id")
    filter_form = TaskFilterForm(query_data or None, user=request.user, project_queryset=projects_qs)
    scoped_base_qs = _scoped_tasks_queryset(request)
    tasks_qs = scoped_base_qs.prefetch_related("subscribers")

    if filter_form.is_valid():
        data = filter_form.cleaned_data
        mine_enabled = bool(data.get("mine")) and not mine_explicit_false
        if mine_enabled:
            user = request.user
            tasks_qs = tasks_qs.filter(Q(created_by=user) | Q(assigned_to=user) | Q(subscribers=user)).distinct()
        if data.get("status"):
            tasks_qs = tasks_qs.filter(status=data["status"])
        if data.get("priority"):
            tasks_qs = tasks_qs.filter(priority=data["priority"])
        if data.get("overdue"):
            tasks_qs = tasks_qs.filter(due_date__lt=timezone.localdate(), status__in=OPEN_STATUSES)
        if data.get("due_from"):
            tasks_qs = tasks_qs.filter(due_date__gte=data["due_from"])
        if data.get("due_to"):
            tasks_qs = tasks_qs.filter(due_date__lte=data["due_to"])
        if data.get("assigned_to"):
            tasks_qs = tasks_qs.filter(assigned_to=data["assigned_to"])
        if data.get("project"):
            tasks_qs = tasks_qs.filter(project=data["project"])
        if data.get("tag"):
            tasks_qs = tasks_qs.filter(tags__icontains=data["tag"].strip())
        if data.get("unassigned"):
            tasks_qs = tasks_qs.filter(assigned_to__isnull=True)
        if data.get("without_due_date"):
            tasks_qs = tasks_qs.filter(due_date__isnull=True)
        if data.get("without_project"):
            tasks_qs = tasks_qs.filter(project__isnull=True)

    tasks = _apply_default_ordering(tasks_qs)
    is_scope_admin = _has_task_permission(request, "tasks_admin")
    can_create = _has_task_permission(request, "tasks_create")
    can_edit = _has_task_permission(request, "tasks_edit")
    can_comment = _has_task_permission(request, "tasks_comment")

    active_project = filter_form.cleaned_data.get("project") if filter_form.is_valid() else None

    # Stats source: project-scoped when filtering by project, global otherwise
    stats_qs = tasks_qs.order_by() if active_project else scoped_base_qs.order_by()
    status_counter = {
        row["status"]: row["total"]
        for row in stats_qs.values("status").annotate(total=Count("id")).order_by()
    }
    total_count = stats_qs.count()
    done_count = int(status_counter.get(TaskStatus.DONE, 0))
    dashboard_stats = {
        "total": total_count,
        "todo": int(status_counter.get(TaskStatus.TODO, 0)),
        "in_progress": int(status_counter.get(TaskStatus.IN_PROGRESS, 0)),
        "done": done_count,
        "canceled": int(status_counter.get(TaskStatus.CANCELED, 0)),
        "overdue": stats_qs.filter(
            due_date__lt=timezone.localdate(),
            status__in=OPEN_STATUSES,
        ).count(),
    }
    done_pct = round(done_count / max(total_count, 1) * 100)

    admin_console = None
    admin_project_summary = []
    if is_scope_admin:
        today = timezone.localdate()
        now = timezone.now()
        admin_console = {
            "unassigned": stats_qs.filter(assigned_to__isnull=True).count(),
            "without_due_date": stats_qs.filter(due_date__isnull=True).count(),
            "without_project": stats_qs.filter(project__isnull=True).count(),
            "due_next_7d": stats_qs.filter(
                status__in=OPEN_STATUSES,
                due_date__gte=today,
                due_date__lte=today + timedelta(days=7),
            ).count(),
            "stale_in_progress": stats_qs.filter(
                status=TaskStatus.IN_PROGRESS,
                updated_at__lt=now - timedelta(days=7),
            ).count(),
        }
        admin_project_summary = list(
            Project.objects.filter(tasks__in=stats_qs)
            .order_by()
            .values("id", "name")
            .annotate(
                task_total=Count("tasks", distinct=True),
                open_total=Count("tasks", filter=Q(tasks__status__in=OPEN_STATUSES), distinct=True),
            )
            .order_by("-open_total", "name")[:6]
        )

    return render(
        request,
        "tasks/list.html",
        {
            **_tasks_shell_context(request, active="dashboard"),
            "page_title": "KICK-OFF",
            "tasks": tasks,
            "filter_form": filter_form,
            "can_create": can_create,
            "can_edit": can_edit,
            "can_comment": can_comment,
            "is_scope_admin": is_scope_admin,
            "dashboard_stats": dashboard_stats,
            "mine_explicit_false": mine_explicit_false,
            "showing_mine_default": (not is_scope_admin) or (not mine_explicit_false),
            "admin_console": admin_console,
            "admin_project_summary": admin_project_summary,
            "active_project": active_project,
            "done_pct": done_pct,
        },
    )


@task_permissions_required("tasks_view")
def task_detail(request, task_id: int):
    task = get_object_or_404(_detail_queryset(request), pk=task_id)
    can_manage = _can_manage_task(request, task)
    can_update_due_date = _can_update_task_due_date(request, task)
    comment_form = TaskCommentForm(
        user=request.user,
        notify_user_queryset=_task_notify_users_queryset(task),
    )

    return render(
        request,
        "tasks/detail.html",
        {
            **_tasks_shell_context(request, active="detail", task=task),
            "page_title": task.title,
            "task": task,
            "task_status_form": TaskStatusForm(instance=task),
            "task_due_date_form": TaskDueDateForm(instance=task),
            "comment_form": comment_form,
            "subtask_form": SubTaskForm(user=request.user),
            "attachment_form": TaskAttachmentForm(task=task),
            "can_edit": can_manage,
            "can_manage_task": can_manage,
            "can_update_due_date": can_update_due_date,
            "can_comment": _has_task_permission(request, "tasks_comment"),
            "task_status_choices": TaskStatus.choices,
            "subtask_status_choices": TaskStatus.choices,
            "task_extra_rows": _render_task_extra_data(task),
            "attrezzatura_embedded_context": _build_task_attrezzatura_context(task),
        },
    )


def _task_attrezzatura_part_number(task: Task) -> str:
    if task.project_id and task.project.part_number:
        return attrezzature_kickoff.normalize_part_number(task.project.part_number)
    data = task.extra_data or {}
    for key in ("part_number", "pn", "p_n", "particolare"):
        if data.get(key):
            return attrezzature_kickoff.normalize_part_number(data.get(key))
    return ""


def _task_attrezzatura_tipo(task: Task) -> str:
    data = task.extra_data or {}
    explicit = str(data.get("attrezzatura_task_type") or data.get("tipo_attrezzatura") or "").strip().lower()
    if explicit in ATTREZZATURA_KICKOFF_TASK_TYPES:
        return ATTREZZATURA_KICKOFF_TASK_TYPES[explicit]
    category = getattr(task, "category", None)
    if category is None:
        return ""
    for candidate in (category.slug, category.name):
        normalized = slugify(str(candidate or "").strip()).replace("-", "_")
        if normalized in ATTREZZATURA_KICKOFF_TASK_TYPES:
            return ATTREZZATURA_KICKOFF_TASK_TYPES[normalized]
    return ""


def _ensure_attrezzatura_task_link_for_kickoff_task(task: Task, user=None):
    part_number = _task_attrezzatura_part_number(task)
    tipo = _task_attrezzatura_tipo(task)
    if not part_number or not tipo:
        return None
    title = f"{task.title} - Gestione Attrezzatura"
    linked, _created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
        tipo=tipo,
        part_number=part_number,
        titolo=title,
        descrizione=task.description,
        kickoff_ref=task.project_id,
        kickoff_activity_ref=task.id,
        user=user,
    )
    return linked


def _handle_task_tooling_form_action(request, task: Task, form: TaskForm) -> None:
    mode = form.cleaned_data.get("tooling_mode") or TaskForm.TOOLING_NONE
    if mode == TaskForm.TOOLING_NONE:
        return
    part_number = attrezzature_kickoff.normalize_part_number(
        form.cleaned_data.get("tooling_part_number")
        or (task.project.part_number if task.project_id else "")
        or _task_attrezzatura_part_number(task)
    )
    if mode == TaskForm.TOOLING_LINK_EXISTING:
        tool = form.cleaned_data.get("tooling_existing_attrezzatura")
        if not tool:
            return
        linked, _created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
            tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
            part_number=part_number or tool.part_number,
            attrezzatura=tool,
            titolo=f"Gestire attrezzatura per {task.title}",
            descrizione=task.description,
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        attrezzature_kickoff.link_attrezzatura_to_kickoff_activity(
            tool,
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            task=linked,
        )
        log_action(request, "tooling_linked_from_task_form", "tasks", {"task_id": task.id, "attrezzatura_id": tool.id})
        messages.success(request, f"Attrezzatura {tool.codice or tool.pk} collegata all'attivita.")
    elif mode == TaskForm.TOOLING_REQUEST_NEW:
        tool = attrezzature_kickoff.create_draft_attrezzatura_from_kickoff(
            part_number=part_number,
            description=form.cleaned_data.get("tooling_description") or task.description,
            codice=form.cleaned_data.get("tooling_code") or "",
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        log_action(request, "tooling_requested_from_task_form", "tasks", {"task_id": task.id, "attrezzatura_id": tool.id})
        messages.success(request, f"Richiesta attrezzatura creata per P/N {tool.part_number}.")
    elif mode == TaskForm.TOOLING_VERIFICATION_REQUIRED:
        linked, created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
            tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
            part_number=part_number,
            titolo=f"Verificare disponibilita attrezzatura per {task.title}",
            descrizione=form.cleaned_data.get("tooling_description") or task.description,
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        log_action(request, "tooling_verification_requested_from_task_form", "tasks", {"task_id": task.id, "attrezzatura_task_id": linked.id})
        messages.success(request, "Verifica disponibilita attrezzatura creata." if created else "Verifica disponibilita gia collegata.")


def _build_task_attrezzatura_context(task: Task) -> dict | None:
    part_number = _task_attrezzatura_part_number(task)
    if not part_number:
        return None
    context = attrezzature_kickoff.build_kickoff_attrezzatura_context(
        part_number,
        kickoff_ref=task.project_id,
        kickoff_activity_ref=task.id,
    )
    context.update(
        {
            "action_url": reverse("tasks:attrezzature_action", kwargs={"task_id": task.id}),
            "source_task": task,
            "stato_choices": AttrezzaturaStato.choices,
            "disponibilita_choices": DisponibilitaStato.choices,
            "task_status_choices": GestioneAttrezzaturaTaskStato.choices,
        }
    )
    return context


@require_POST
@task_permissions_required("tasks_view")
def attrezzatura_action(request, task_id: int):
    task = get_object_or_404(_detail_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    part_number = _task_attrezzatura_part_number(task) or attrezzature_kickoff.normalize_part_number(
        request.POST.get("part_number", "")
    )
    if not part_number:
        messages.error(request, "P/N mancante: impossibile collegare Gestione Attrezzatura.")
        return redirect("tasks:detail", task_id=task.id)

    action = (request.POST.get("action") or "").strip()
    attrezzatura_id = request.POST.get("attrezzatura_id")
    attrezzatura = None
    if attrezzatura_id:
        attrezzatura = get_object_or_404(Attrezzatura, pk=attrezzatura_id)

    if action == "create_availability_task":
        linked, created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
            tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
            part_number=part_number,
            attrezzatura=attrezzatura,
            titolo=f"Verificare disponibilita attrezzatura per {task.title}",
            descrizione=task.description,
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        messages.success(request, "Task disponibilita creata." if created else f"Task disponibilita gia collegata: {linked.titolo}.")
    elif action == "create_new_tool_task":
        linked, created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
            tipo=GestioneAttrezzaturaTaskTipo.CREAZIONE_ATTREZZO,
            part_number=part_number,
            titolo=f"Creare nuovo attrezzo per P/N {part_number}",
            descrizione=task.description,
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        messages.success(request, "Task creazione attrezzo creata." if created else f"Task creazione gia collegata: {linked.titolo}.")
    elif action == "create_draft_tool":
        tool = attrezzature_kickoff.create_draft_attrezzatura_from_kickoff(
            part_number=part_number,
            description=request.POST.get("description", "") or task.description,
            codice=request.POST.get("codice", ""),
            kickoff_ref=task.project_id,
            kickoff_activity_ref=task.id,
            user=request.user,
        )
        messages.success(request, f"Bozza attrezzatura creata per P/N {tool.part_number}.")
    elif action == "link_tool":
        if attrezzatura is None:
            messages.error(request, "Seleziona un'attrezzatura da collegare.")
        else:
            linked = _ensure_attrezzatura_task_link_for_kickoff_task(task, user=request.user)
            if linked is None:
                linked, _created = attrezzature_kickoff.get_or_create_attrezzatura_task_for_kickoff_activity(
                    tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
                    part_number=part_number,
                    attrezzatura=attrezzatura,
                    titolo=f"Gestire attrezzatura per {task.title}",
                    kickoff_ref=task.project_id,
                    kickoff_activity_ref=task.id,
                    user=request.user,
                )
            attrezzature_kickoff.link_attrezzatura_to_kickoff_activity(
                attrezzatura,
                kickoff_ref=task.project_id,
                kickoff_activity_ref=task.id,
                task=linked,
            )
            messages.success(request, f"Attrezzatura {attrezzatura.codice or attrezzatura.pk} collegata.")
    elif action == "update_progress":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            raw_percent = (request.POST.get("percentuale") or "").strip()
            try:
                percent = int(raw_percent) if raw_percent else None
            except ValueError:
                messages.error(request, "Percentuale avanzamento non valida.")
                return redirect("tasks:detail", task_id=task.id)
            stato = request.POST.get("stato") or None
            attrezzature_kickoff.update_attrezzatura_progress_from_kickoff(
                attrezzatura,
                percentuale=percent,
                stato=stato,
                user=request.user,
                note=request.POST.get("note", ""),
                kickoff_ref=task.project_id,
                kickoff_activity_ref=task.id,
            )
            messages.success(request, "Avanzamento attrezzatura aggiornato.")
    elif action == "mark_blocked":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            reason = (request.POST.get("reason") or "").strip()
            if not reason:
                messages.error(request, "Indica il motivo del blocco.")
            else:
                attrezzature_workflow.mark_blocked(attrezzatura, user=request.user, reason=reason)
                attrezzature_kickoff.upsert_kickoff_link(
                    attrezzatura=attrezzatura,
                    kickoff_ref=task.project_id,
                    kickoff_activity_ref=task.id,
                    user=request.user,
                    notes=reason,
                )
                log_action(
                    request,
                    "tooling_blocked_from_kickoff",
                    "tasks",
                    {"task_id": task.id, "attrezzatura_id": attrezzatura.id, "reason": reason},
                )
                messages.success(request, "Attrezzatura bloccata.")
    elif action == "mark_to_create":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            attrezzature_workflow.mark_to_create(attrezzatura, user=request.user, note=request.POST.get("note", ""))
            attrezzature_kickoff.upsert_kickoff_link(
                attrezzatura=attrezzatura,
                kickoff_ref=task.project_id,
                kickoff_activity_ref=task.id,
                user=request.user,
            )
            log_action(
                request,
                "tooling_to_create_from_kickoff",
                "tasks",
                {"task_id": task.id, "attrezzatura_id": attrezzatura.id},
            )
            messages.success(request, "Attrezzatura segnata da creare.")
    elif action == "set_availability":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            attrezzature_workflow.set_availability_status(
                attrezzatura,
                request.POST.get("disponibilita_stato") or attrezzatura.disponibilita_stato,
                user=request.user,
                note=request.POST.get("note", ""),
            )
            attrezzature_kickoff.upsert_kickoff_link(
                attrezzatura=attrezzatura,
                kickoff_ref=task.project_id,
                kickoff_activity_ref=task.id,
                user=request.user,
            )
            messages.success(request, "Disponibilita attrezzatura aggiornata.")
    elif action == "add_note":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            text = (request.POST.get("note") or "").strip()
            if text:
                AttrezzaturaNota.objects.create(
                    attrezzatura=attrezzatura,
                    testo=text,
                    origine="kickoff_activity",
                    created_by=request.user,
                )
                messages.success(request, "Nota attrezzatura aggiunta.")
            else:
                messages.error(request, "Scrivi una nota prima di salvarla.")
    elif action == "confirm_ready":
        if attrezzatura is None:
            messages.error(request, "Attrezzatura non trovata.")
        else:
            attrezzature_kickoff.confirm_attrezzatura_ready_from_kickoff(
                attrezzatura,
                user=request.user,
                note=request.POST.get("note", ""),
                kickoff_ref=task.project_id,
                kickoff_activity_ref=task.id,
            )
            messages.success(request, "Attrezzatura confermata pronta produzione.")
    elif action in {"complete_attrezzatura_task", "block_attrezzatura_task"}:
        gestione_task_id = request.POST.get("attrezzatura_task_id")
        task_qs = GestioneAttrezzaturaTask.objects.filter(
            Q(part_number__iexact=part_number)
            | Q(external_kickoff_activity_id=str(task.id))
            | Q(external_kickoff_id=str(task.project_id or ""))
        )
        gestione_task = get_object_or_404(task_qs, pk=gestione_task_id)
        if action == "complete_attrezzatura_task":
            attrezzature_kickoff.complete_attrezzatura_task_from_kickoff(
                gestione_task,
                user=request.user,
                note=request.POST.get("note", ""),
            )
            messages.success(request, "Task Gestione Attrezzatura completata.")
        else:
            attrezzature_kickoff.block_attrezzatura_task_from_kickoff(
                gestione_task,
                user=request.user,
                reason=request.POST.get("reason", ""),
            )
            messages.success(request, "Task Gestione Attrezzatura bloccata.")
    else:
        messages.error(request, "Azione Gestione Attrezzatura non riconosciuta.")
    return redirect("tasks:detail", task_id=task.id)


def _suggest_task_start_date(project) -> date:
    """Return the automatic start date for a new task in ``project``.

    Day after the last existing task's end date (due_date or next_step_due),
    or today if the project has no tasks yet.
    """
    today = timezone.localdate()
    if project is None:
        return today
    last_task = (
        Task.objects.filter(project=project)
        .order_by("-id")
        .first()
    )
    if last_task:
        last_end = last_task.due_date or last_task.next_step_due
        if last_end:
            return max(last_end + timedelta(days=1), today)
    return today


@task_permissions_required("tasks_view")
def project_info_json(request, project_id: int):
    """Restituisce info progetto + lista task (ordine creazione) in formato JSON, per l'AJAX del form."""
    project = get_object_or_404(_scoped_projects_queryset(request).select_related(
        "project_manager", "capo_commessa", "programmer"
    ), pk=project_id)
    raw_tasks = list(
        Task.objects.filter(project=project)
        .order_by("id")
        .values("id", "title", "status", "next_step_due", "due_date",
                "assigned_to__first_name", "assigned_to__last_name", "assigned_to__username")
    )
    tasks_data = []
    for i, t in enumerate(raw_tasks, start=1):
        fn = t["assigned_to__first_name"] or ""
        ln = t["assigned_to__last_name"] or ""
        assignee = f"{fn} {ln}".strip() or t["assigned_to__username"] or ""
        tasks_data.append({
            "wbs": i,
            "id": t["id"],
            "title": t["title"],
            "status": t["status"],
            "next_step_due": t["next_step_due"].strftime("%d/%m/%Y") if t["next_step_due"] else "",
            "due_date": t["due_date"].strftime("%d/%m/%Y") if t["due_date"] else "",
            "assignee": assignee,
        })
    pm = project.project_manager
    cc = project.capo_commessa
    prog = project.programmer
    cfg = TaskImpostazioni.get_singleton()
    vrf_detail = _vrf_status_detail(project, cfg)
    return JsonResponse({
        "id": project.id,
        "name": project.name,
        "client_name": project.client_name or "",
        "part_number": project.part_number or "",
        "revisione": project.revisione or "",
        "versione": project.versione or "",
        "project_manager": (pm.get_full_name() or pm.username) if pm else "",
        "capo_commessa": (cc.get_full_name() or cc.username) if cc else "",
        "programmer": (prog.get_full_name() or prog.username) if prog else "",
        "task_total": len(tasks_data),
        "vrf_status": vrf_detail["status"],
        "vrf_label": vrf_detail["label"],
        "vrf_upload_url": reverse("tasks:project_vrf_upload", kwargs={"project_id": project.id}),
        "vrf_compile_url": reverse("tasks:project_vrf_compile", kwargs={"project_id": project.id}),
        "tasks": tasks_data,
        "suggested_start_date": _suggest_task_start_date(project).isoformat(),
    })


def _task_create_locked_project(request, projects_qs):
    project_param = (request.GET.get("project") or "").strip()
    if not project_param:
        return None
    try:
        project_id = int(project_param)
    except (TypeError, ValueError):
        return None
    return get_object_or_404(
        projects_qs.select_related("project_manager", "capo_commessa", "programmer"),
        pk=project_id,
    )


@task_permissions_required("tasks_view", "tasks_create")
def task_create(request):
    projects_qs = _scoped_projects_queryset(request).order_by("name", "id")
    locked_project = _task_create_locked_project(request, projects_qs)
    suggested_start_date = _suggest_task_start_date(locked_project)
    has_previous_task = (
        locked_project is not None
        and Task.objects.filter(project=locked_project).exists()
    )

    if request.method == "POST":
        form = TaskForm(
            request.POST,
            user=request.user,
            project_queryset=projects_qs,
            locked_project=locked_project,
        )
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.project = form.resolve_project(created_by=request.user)
            if task.project and not getattr(form, "new_project_created", False) and not _can_manage_project(request, task.project):
                messages.error(
                    request,
                    "Non hai i permessi per aggiungere nuove attivita a questo kickoff.",
                )
                return redirect("tasks:project_gantt", project_id=task.project.id)

            # Blocking guard: se il progetto esistente è bloccato per VRF mancante, blocca il salvataggio
            if task.project and not getattr(form, "new_project_created", False):
                cfg = TaskImpostazioni.get_singleton()
                vrf_detail = _vrf_status_detail(task.project, cfg)
                if vrf_detail["is_blocked"]:
                    messages.error(
                        request,
                        f"Kickoff bloccato: carica il documento VRF prima di procedere. "
                        f"({vrf_detail['days_pending']} giorni senza documento VRF)",
                    )
                    return redirect(
                        reverse("tasks:project_vrf_upload", kwargs={"project_id": task.project.id})
                        + f"?next={reverse('tasks:create')}"
                    )

            block_msgs = _check_blocking_asset_conflicts(task, request.POST, exclude_task_id=None)
            if block_msgs:
                for m in block_msgs:
                    messages.error(request, m)
                form.add_error(
                    None,
                    "Impossibile salvare: conflitto bloccante sugli asset selezionati.",
                )
                return render(
                    request,
                    "tasks/form.html",
                    {
                        **_tasks_shell_context(request, active="create"),
                        "page_title": "Nuova attivita kickoff",
                        "form": form,
                        "mode": "create",
                        "suggested_start_date": suggested_start_date,
                        "suggested_start_from_previous": has_previous_task,
                        "suggested_project_id": locked_project.id if locked_project is not None else None,
                        "locked_project": locked_project,
                        "locked_project_vrf_detail": _vrf_status_detail(locked_project, TaskImpostazioni.get_singleton()) if locked_project is not None else None,
                        "locked_project_task_total": Task.objects.filter(project=locked_project).count() if locked_project is not None else 0,
                        "task_extra_data_json": json.dumps(request.POST.dict()),
                    },
                )

            task.save()
            form.save_m2m()
            _persist_task_extra_data(task, request.POST)
            for mw_warn in _check_machine_work_overlaps(task, exclude_task_id=None):
                messages.warning(request, mw_warn)
            _ensure_attrezzatura_task_link_for_kickoff_task(task, user=request.user)
            _handle_task_tooling_form_action(request, task, form)
            if form.reused_existing_project is not None and task.project_id:
                if form.reused_existing_project_fields:
                    messages.info(
                        request,
                        (
                            f"Kickoff esistente riutilizzato per P/N {task.project.part_number or '-'}: "
                            "anagrafica kickoff aggiornata con i dati inseriti."
                        ),
                    )
                else:
                    messages.info(
                        request,
                        (
                            f"Kickoff esistente riutilizzato per P/N {task.project.part_number or '-'}: "
                            "l'attivita kickoff e stata agganciata senza creare duplicati."
                        ),
                    )
            if task.project_id:
                messages.success(request, f"Attivita kickoff creata nel kickoff '{task.project.name}'.")
            else:
                messages.success(request, "Attivita singola creata correttamente.")
            if form.assignment_conflicts and task.assigned_to_id:
                assignee_name = task.assigned_to.get_full_name() or task.assigned_to.get_username()
                messages.warning(
                    request,
                    (
                        f"Impegni sovrapposti: {assignee_name} ha gia {len(form.assignment_conflicts)} task attive "
                        f"nello stesso periodo ({form.assignment_conflict_summary(limit=3)})."
                    ),
                )
            if form.auto_raised_priority:
                messages.warning(request, "Priorita aggiornata automaticamente a High per conflitto impegni.")
            if task.is_overdue:
                messages.warning(request, "Attivita kickoff creata con scadenza gia oltre la data odierna.")
            _add_task_absence_warnings(request, task)

            # Applica preferenze reminder portale + integrazione Outlook
            task.reminder_portal_enabled = bool(form.cleaned_data.get("reminder_portal_enabled_field"))
            task.save(update_fields=["reminder_portal_enabled"])
            _sync_task_integrations(request, task, form, action="create")

            # Se è stato creato un nuovo progetto, redirect alla pagina upload VRF
            if getattr(form, "new_project_created", False) and task.project_id:
                next_url = reverse("tasks:detail", kwargs={"task_id": task.id})
                return redirect(
                    reverse("tasks:project_vrf_upload", kwargs={"project_id": task.project.id})
                    + f"?next={next_url}"
                )
            return redirect("tasks:detail", task_id=task.id)
    else:
        initial = {"next_step_due": suggested_start_date.isoformat()}
        if locked_project is not None:
            initial["project_choice"] = locked_project.id
            initial["task_scope"] = "project"
            initial["project_link_mode"] = TaskForm.PROJECT_LINK_EXISTING

        form = TaskForm(
            user=request.user,
            project_queryset=projects_qs,
            initial=initial,
            locked_project=locked_project,
        )

        return render(
            request,
            "tasks/form.html",
            {
                **_tasks_shell_context(request, active="create"),
                "page_title": "Nuova attivita kickoff",
                "form": form,
                "mode": "create",
                "suggested_start_date": suggested_start_date,
                "suggested_start_from_previous": has_previous_task,
                "suggested_project_id": locked_project.id if locked_project is not None else None,
                "locked_project": locked_project,
                "locked_project_vrf_detail": _vrf_status_detail(locked_project, TaskImpostazioni.get_singleton()) if locked_project is not None else None,
                "locked_project_task_total": Task.objects.filter(project=locked_project).count() if locked_project is not None else 0,
                "task_extra_data_json": json.dumps({}),
            },
        )

    return render(
        request,
        "tasks/form.html",
        {
            **_tasks_shell_context(request, active="create"),
            "page_title": "Nuova attivita kickoff",
            "form": form,
            "mode": "create",
            "suggested_start_date": suggested_start_date,
            "suggested_start_from_previous": has_previous_task,
            "suggested_project_id": locked_project.id if locked_project is not None else None,
            "locked_project": locked_project,
            "locked_project_vrf_detail": _vrf_status_detail(locked_project, TaskImpostazioni.get_singleton()) if locked_project is not None else None,
            "locked_project_task_total": Task.objects.filter(project=locked_project).count() if locked_project is not None else 0,
            "task_extra_data_json": json.dumps({}),
        },
    )


@task_permissions_required("tasks_view")
def task_edit(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request).prefetch_related("subscribers"), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    projects_qs = _scoped_projects_queryset(request).order_by("name", "id")
    before = _task_snapshot(task)

    if request.method == "POST":
        form = TaskForm(request.POST, instance=task, user=request.user, project_queryset=projects_qs)
        if form.is_valid():
            updated_task = form.save(commit=False)
            updated_task.project = form.resolve_project(created_by=request.user)
            if updated_task.project and updated_task.project_id != task.project_id and not _can_manage_project(request, updated_task.project):
                messages.error(
                    request,
                    "Non hai i permessi per spostare l'attivita su questo kickoff.",
                )
                return redirect("tasks:edit", task_id=task.id)

            # Blocking guard: progetto bloccato per VRF mancante
            if updated_task.project and not getattr(form, "new_project_created", False):
                cfg = TaskImpostazioni.get_singleton()
                vrf_detail = _vrf_status_detail(updated_task.project, cfg)
                if vrf_detail["is_blocked"]:
                    messages.error(
                        request,
                        "Kickoff bloccato: carica il documento VRF prima di modificare questa attivita kickoff.",
                    )
                    return redirect(
                        reverse("tasks:project_vrf_upload", kwargs={"project_id": updated_task.project.id})
                        + f"?next={reverse('tasks:edit', kwargs={'task_id': task.id})}"
                    )

            block_msgs = _check_blocking_asset_conflicts(updated_task, request.POST, exclude_task_id=task.id)
            if block_msgs:
                for m in block_msgs:
                    messages.error(request, m)
                form.add_error(
                    None,
                    "Impossibile salvare: conflitto bloccante sugli asset selezionati.",
                )
                return render(
                    request,
                    "tasks/form.html",
                    {
                        **_tasks_shell_context(request, active="edit", task=task),
                        "page_title": "Modifica attivita kickoff",
                        "form": form,
                        "task": task,
                        "mode": "edit",
                        "task_extra_data_json": json.dumps(request.POST.dict()),
                    },
                )

            updated_task.save()
            form.save_m2m()
            _persist_task_extra_data(updated_task, request.POST)
            for mw_warn in _check_machine_work_overlaps(updated_task, exclude_task_id=updated_task.pk):
                messages.warning(request, mw_warn)
            _ensure_attrezzatura_task_link_for_kickoff_task(updated_task, user=request.user)
            _handle_task_tooling_form_action(request, updated_task, form)
            _log_task_update_events(updated_task, request.user, before)
            if form.reused_existing_project is not None and updated_task.project_id:
                messages.info(
                    request,
                    (
                        f"Attivita kickoff riallineata al kickoff esistente '{updated_task.project.name}' "
                        f"(P/N {updated_task.project.part_number or '-'})."
                    ),
                )
            messages.success(request, "Attivita kickoff aggiornata.")
            if form.assignment_conflicts and updated_task.assigned_to_id:
                assignee_name = updated_task.assigned_to.get_full_name() or updated_task.assigned_to.get_username()
                messages.warning(
                    request,
                    (
                        f"Impegni sovrapposti: {assignee_name} ha gia {len(form.assignment_conflicts)} task attive "
                        f"nello stesso periodo ({form.assignment_conflict_summary(limit=3)})."
                    ),
                )
            if form.auto_raised_priority:
                messages.warning(request, "Priorita aggiornata automaticamente a High per conflitto impegni.")
            if updated_task.is_overdue:
                messages.warning(request, "Task in stato overdue.")
            _add_task_absence_warnings(request, updated_task)
            updated_task.reminder_portal_enabled = bool(form.cleaned_data.get("reminder_portal_enabled_field"))
            updated_task.save(update_fields=["reminder_portal_enabled"])
            _sync_task_integrations(request, updated_task, form, action="edit")
            return redirect("tasks:detail", task_id=updated_task.id)
    else:
        form = TaskForm(instance=task, user=request.user, project_queryset=projects_qs)

    return render(
        request,
        "tasks/form.html",
        {
            **_tasks_shell_context(request, active="edit", task=task),
            "page_title": "Modifica attivita kickoff",
            "form": form,
            "task": task,
            "mode": "edit",
            "task_extra_data_json": json.dumps(task.extra_data or {}),
        },
    )


@require_POST
@task_permissions_required("tasks_view")
def update_due_date(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_update_task_due_date(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )

    before = _task_snapshot(task)
    form = TaskDueDateForm(request.POST, instance=task)
    if form.is_valid():
        task = form.save()
        _log_task_update_events(task, request.user, before)
        messages.success(request, "Data prevista conclusione aggiornata.")
        if task.is_overdue:
            messages.warning(request, "Task in stato overdue.")
        _add_task_absence_warnings(request, task)
    else:
        messages.error(request, "Data prevista conclusione non valida.")

    return redirect("tasks:detail", task_id=task.id)


@require_POST
@task_permissions_required("tasks_view")
def change_status(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    old_status = task.status
    form = TaskStatusForm(request.POST, instance=task)
    if form.is_valid():
        task = form.save()
        if old_status != task.status:
            _log_event(
                task,
                request.user,
                TaskEventType.STATUS_CHANGE,
                {"from": old_status, "to": task.status},
            )
            messages.success(request, "Stato task aggiornato.")
    else:
        messages.error(request, "Stato non valido.")
    return redirect("tasks:detail", task_id=task_id)


@require_POST
@task_permissions_required("tasks_view", "tasks_comment")
def add_comment(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    notify_users_qs = _task_notify_users_queryset(task)
    form = TaskCommentForm(request.POST, user=request.user, notify_user_queryset=notify_users_qs)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.task = task
        comment.author = request.user
        comment.save()
        _log_event(
            task,
            request.user,
            TaskEventType.COMMENT_ADDED,
            {
                "comment_id": comment.id,
                "target_user_id": comment.target_user_id,
            },
        )
        if comment.target_user_id and comment.target_user_id != request.user.id:
            _notify_user(
                comment.target_user,
                message_text=f"Nuovo commento su task '{task.title}'.",
                action_url=reverse("tasks:detail", kwargs={"task_id": task.id}),
            )
        messages.success(request, "Commento aggiunto.")
    else:
        messages.error(request, "Commento non valido.")
    return redirect("tasks:detail", task_id=task_id)


@require_POST
@task_permissions_required("tasks_view")
def add_subtask(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    form = SubTaskForm(request.POST, user=request.user)
    if form.is_valid():
        subtask = form.save(commit=False)
        subtask.task = task
        subtask.save()
        _log_event(
            task,
            request.user,
            TaskEventType.SUBTASK_ADDED,
            {
                "subtask_id": subtask.id,
                "title": subtask.title,
                "status": subtask.status,
            },
        )
        _apply_subtask_rollup(task, request.user)
        messages.success(request, "Subtask aggiunta.")
    else:
        messages.error(request, "Subtask non valida.")
    return redirect("tasks:detail", task_id=task_id)


@require_POST
@task_permissions_required("tasks_view")
def edit_subtask_status(request, task_id: int, subtask_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    subtask = get_object_or_404(SubTask.objects.filter(task=task), pk=subtask_id)
    old_status = subtask.status

    form = SubTaskStatusForm(request.POST, instance=subtask)
    if form.is_valid():
        subtask = form.save()
        if old_status != subtask.status:
            _log_event(
                task,
                request.user,
                TaskEventType.SUBTASK_STATUS_CHANGE,
                {
                    "subtask_id": subtask.id,
                    "from": old_status,
                    "to": subtask.status,
                },
            )
            _apply_subtask_rollup(task, request.user)
            messages.success(request, "Stato subtask aggiornato.")
    else:
        messages.error(request, "Stato subtask non valido.")
    return redirect("tasks:detail", task_id=task_id)


@require_POST
@task_permissions_required("tasks_view")
def subtask_toggle(request, task_id: int, subtask_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return JsonResponse({"ok": False, "reason": "forbidden"}, status=403)
    subtask = get_object_or_404(SubTask.objects.filter(task=task), pk=subtask_id)
    old_status = subtask.status
    subtask.status = TaskStatus.TODO if subtask.status == TaskStatus.DONE else TaskStatus.DONE
    subtask.save(update_fields=["status"])
    if old_status != subtask.status:
        _log_event(
            task,
            request.user,
            TaskEventType.SUBTASK_STATUS_CHANGE,
            {"subtask_id": subtask.id, "from": old_status, "to": subtask.status},
        )
        _apply_subtask_rollup(task, request.user)
    task.refresh_from_db(fields=["status"])
    return JsonResponse(
        {
            "ok": True,
            "subtask_id": subtask.id,
            "status": subtask.status,
            "status_display": subtask.get_status_display(),
            "task_status": task.status,
            "task_status_display": task.get_status_display(),
        }
    )


@require_POST
@task_permissions_required("tasks_view")
def add_attachment(request, task_id: int):
    task = get_object_or_404(_scoped_tasks_queryset(request), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    form = TaskAttachmentForm(request.POST, request.FILES, task=task)

    if form.is_valid():
        attachment = form.save(commit=False)
        attachment.uploaded_by = request.user
        attachment.original_name = getattr(form.cleaned_data.get("file"), "name", "") or ""

        attach_to = form.cleaned_data.get("attach_to") or TaskAttachmentForm.TARGET_TASK
        if attach_to == TaskAttachmentForm.TARGET_PROJECT:
            if not task.project_id or not _can_manage_project(request, task.project):
                return render(
                    request,
                    "core/pages/forbidden.html",
                    {"page_title": "Accesso negato"},
                    status=403,
                )
            attachment.project = task.project
            attachment.task = None
            target = "project"
        else:
            attachment.task = task
            attachment.project = None
            target = "task"

        attachment.save()
        _log_event(
            task,
            request.user,
            TaskEventType.ATTACHMENT_ADDED,
            {
                "attachment_id": attachment.id,
                "target": target,
                "task_id": attachment.task_id,
                "project_id": attachment.project_id,
                "file_name": attachment.original_name,
            },
        )
        messages.success(request, "Allegato caricato.")
    else:
        messages.error(request, "Upload allegato non valido.")

    return redirect("tasks:detail", task_id=task.id)


@task_permissions_required("tasks_view", "tasks_create")
def project_create(request):
    """Crea un nuovo kickoff (Project) e avvia il workflow VRF.

    Flow: anagrafica kickoff -> salvataggio Project -> redirect a vrf_compile.
    Se l'utente compila P/N + revisione + versione gia' esistenti, il form
    intercetta il duplicato e reindirizza al kickoff gia' presente.
    """
    projects_qs = _scoped_projects_queryset(request)

    if request.method == "POST":
        form = ProjectKickoffForm(request.POST, project_queryset=projects_qs)
        if form.is_valid():
            if form.reused_existing_project is not None:
                reused = form.reused_existing_project
                messages.info(
                    request,
                    f"Kickoff gia' esistente per questo P/N: riutilizzo '{reused.name}'. "
                    "Procedi con la scheda VRF o aggiungi una nuova attivita.",
                )
                log_action(request, "kickoff_reused_on_create", "tasks", {
                    "project_id": reused.id,
                    "part_number": reused.part_number,
                    "message": f"Kickoff #{reused.id} riutilizzato da form Nuovo kickoff",
                })
                return redirect("tasks:project_vrf_compile", project_id=reused.id)

            project = form.save(commit=False)
            project.created_by = request.user
            project.save()
            log_action(request, "kickoff_created", "tasks", {
                "project_id": project.id,
                "part_number": project.part_number,
                "client_name": project.client_name,
                "message": f"Nuovo kickoff #{project.id} '{project.name}' creato",
            })
            messages.success(
                request,
                f"Kickoff '{project.name}' creato. Compila la scheda VRF per proseguire.",
            )
            return redirect("tasks:project_vrf_compile", project_id=project.id)
    else:
        form = ProjectKickoffForm(project_queryset=projects_qs)

    return render(
        request,
        "tasks/project_create.html",
        {
            **_tasks_shell_context(request, active="projects"),
            "page_title": "Nuovo kickoff",
            "form": form,
        },
    )


@task_permissions_required("tasks_view")
def project_list(request):
    today = timezone.localdate()
    projects_base_qs = _scoped_projects_queryset(request).order_by()

    # ── Filtri GET ──
    q_text   = (request.GET.get("q") or "").strip()
    q_client = (request.GET.get("client") or "").strip()
    q_vrf    = (request.GET.get("vrf_status") or "").strip()
    q_sort   = request.GET.get("sort") or "name"

    if q_text:
        projects_base_qs = projects_base_qs.filter(
            Q(name__icontains=q_text) | Q(part_number__icontains=q_text) | Q(client_name__icontains=q_text)
        )
    if q_client:
        projects_base_qs = projects_base_qs.filter(client_name__icontains=q_client)

    projects_qs = projects_base_qs.annotate(
        task_total=Count("tasks", distinct=True),
        task_open=Count("tasks", filter=Q(tasks__status__in=OPEN_STATUSES), distinct=True),
        task_done=Count("tasks", filter=Q(tasks__status=TaskStatus.DONE), distinct=True),
        task_overdue=Count(
            "tasks",
            filter=Q(tasks__status__in=OPEN_STATUSES, tasks__due_date__lt=today),
            distinct=True,
        ),
        task_progress_avg=Count("tasks", filter=Q(tasks__progress__gt=0), distinct=True),
        earliest_due=F("tasks__due_date"),
    )

    # Ordinamento
    if q_sort == "due":
        projects_qs = projects_qs.order_by(
            F("earliest_due").asc(nulls_last=True), "name", "id"
        )
    elif q_sort == "progress_desc":
        projects_qs = projects_qs.order_by(
            F("task_done").desc(nulls_last=True), "name", "id"
        )
    elif q_sort == "ral":
        projects_qs = projects_qs.order_by(
            F("task_overdue").desc(nulls_last=True), "name", "id"
        )
    else:
        projects_qs = projects_qs.order_by("name", "id")

    projects = list(projects_qs.distinct())
    cfg = TaskImpostazioni.get_singleton()

    # ── Calcola semaforo RAL e VRF detail per ciascun progetto ──
    for p in projects:
        p.vrf_detail = _vrf_status_detail(p, cfg)
        p.can_manage = _can_manage_project(request, p)
        overdue = getattr(p, "task_overdue", 0) or 0
        if overdue == 0:
            p.ral_status = "green"
        elif overdue <= 2:
            p.ral_status = "yellow"
        else:
            p.ral_status = "red"

    # Filtro VRF applicato lato Python (usa vrf_detail già calcolato)
    if q_vrf:
        projects = [p for p in projects if p.vrf_detail.get("status") == q_vrf]

    # Clienti distinti per il dropdown filtro
    client_choices = sorted(set(
        p.client_name for p in
        list(_scoped_projects_queryset(request).values_list("client_name", flat=True).distinct())
        if p
    ))

    return render(
        request,
        "tasks/projects.html",
        {
            **_tasks_shell_context(request, active="projects"),
            "page_title": "Portfolio kickoff",
            "projects": projects,
            "is_scope_admin": _has_task_permission(request, "tasks_admin"),
            "pf_filter_q": q_text,
            "pf_filter_client": q_client,
            "pf_filter_vrf": q_vrf,
            "pf_sort": q_sort,
            "pf_client_choices": client_choices,
        },
    )


@require_POST
@task_permissions_required("tasks_view", "tasks_create")
def copy_project_with_vrf(request, project_id: int):
    source_project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, source_project):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    kickoff = _duplicate_kickoff(source_project, created_by=request.user, clear_part_number=False)
    log_action(
        request,
        "copy_kickoff_vrf",
        "tasks",
        {
            "message": f"Copiato kickoff #{source_project.id} in kickoff #{kickoff.id} con duplicazione VRF.",
            "source_project_id": source_project.id,
            "copied_project_id": kickoff.id,
        },
    )
    messages.success(request, f"Kickoff duplicato correttamente: {kickoff.name}.")
    return redirect("tasks:project_gantt", project_id=kickoff.id)


@require_POST
@task_permissions_required("tasks_view", "tasks_create")
def copy_project_with_vrf_without_pn(request, project_id: int):
    source_project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, source_project):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    kickoff = _duplicate_kickoff(source_project, created_by=request.user, clear_part_number=True)
    log_action(
        request,
        "copy_kickoff_vrf_without_pn",
        "tasks",
        {
            "message": (
                f"Copiato kickoff #{source_project.id} in kickoff #{kickoff.id} "
                "svuotando il P/N e la cella B3 del file VRF."
            ),
            "source_project_id": source_project.id,
            "copied_project_id": kickoff.id,
            "without_part_number": True,
        },
    )
    messages.success(request, f"Kickoff duplicato senza P/N: {kickoff.name}.")
    return redirect("tasks:project_gantt", project_id=kickoff.id)


@task_permissions_required("tasks_view")
def project_gantt(request, project_id: int):
    project_qs = _scoped_projects_queryset(request).prefetch_related(
        Prefetch(
            "tasks",
            # Nel Gantt le task sono ordinate per ordine di creazione (id asc)
            # così WBS 1 = prima task creata, WBS 2 = seconda, ecc.
            # Questo mantiene la sequenza logica definita dall'utente
            # indipendentemente dalle date (che possono essere in disordine durante la pianificazione).
            queryset=Task.objects.select_related(
                "created_by",
                "assigned_to",
                "project",
                "project__project_manager",
                "project__capo_commessa",
                "project__programmer",
                "project__similar_project",
            ).prefetch_related("subscribers").order_by("id"),
        ),
        Prefetch(
            "comments",
            queryset=ProjectComment.objects.select_related("author", "target_user").order_by("-created_at", "-id"),
        ),
    )
    project = get_object_or_404(project_qs, pk=project_id)
    # Forziamo ordinamento lato Python (SQL Server può ignorare ORDER BY nei prefetch)
    tasks = sorted(project.tasks.all(), key=lambda t: t.id)
    can_edit_schedule = _can_edit_project_schedule(request, project)
    can_comment = _has_task_permission(request, "tasks_comment")
    gantt_options = _parse_gantt_options(request)
    gantt_meta = _project_gantt_rows(tasks, min_window_days=gantt_options["window_days"])
    comment_form = ProjectCommentForm(
        user=request.user,
        notify_user_queryset=_project_notify_users_queryset(project),
    )
    today = timezone.localdate()
    total_tasks = len(tasks)
    done_total = sum(1 for task in tasks if task.status == TaskStatus.DONE)
    open_total = sum(1 for task in tasks if task.status in OPEN_STATUSES)
    in_progress_total = sum(1 for task in tasks if task.status == TaskStatus.IN_PROGRESS)
    overdue_total = sum(1 for task in tasks if task.is_overdue)
    due_next_7d_total = sum(
        1
        for task in tasks
        if task.status in OPEN_STATUSES
        and task.due_date
        and today <= task.due_date <= today + timedelta(days=7)
    )
    unassigned_total = sum(1 for task in tasks if not task.assigned_to_id)
    high_priority_open_total = sum(
        1 for task in tasks if task.status in OPEN_STATUSES and task.priority == TaskPriority.HIGH
    )
    invalid_range_total = sum(1 for row in gantt_meta["rows"] if row.get("invalid_range"))
    absence_conflict_total = sum(1 for row in gantt_meta["rows"] if row.get("has_absence_conflicts"))
    progress_percent = int(round((done_total / total_tasks) * 100)) if total_tasks else 0
    project_summary = {
        "total_tasks": total_tasks,
        "open_tasks": open_total,
        "in_progress_tasks": in_progress_total,
        "done_tasks": done_total,
        "overdue_tasks": overdue_total,
        "due_next_7d_tasks": due_next_7d_total,
        "unassigned_tasks": unassigned_total,
        "high_priority_open_tasks": high_priority_open_total,
        "invalid_range_tasks": invalid_range_total,
        "absence_conflict_tasks": absence_conflict_total,
        "progress_percent": progress_percent,
        "planned_start": gantt_meta["timeline_start"],
        "planned_end": gantt_meta["timeline_end"],
    }

    task_update_forms = {}
    for row in gantt_meta["rows"]:
        row["can_edit"] = _can_manage_task(request, row["task"])
        if row["can_edit"]:
            task_update_forms[row["task"].id] = ProjectTaskGanttUpdateForm(
                instance=row["task"],
                prefix=f"task_{row['task'].id}",
            )
        row["update_form"] = task_update_forms.get(row["task"].id)

    can_edit_schedule = any(bool(row.get("can_edit")) for row in gantt_meta["rows"]) or _can_manage_project(request, project)

    cfg = TaskImpostazioni.get_singleton()
    vrf_detail = _vrf_status_detail(project, cfg)

    # ── Baseline Gantt ──
    baseline = GanttBaseline.objects.filter(project=project).first()
    for row in gantt_meta["rows"]:
        row["baseline_delta"] = baseline.get_task_delta_days(row["task"]) if baseline else {"start": None, "end": None}

    # ── Dipendenze task ──
    task_ids = [t.id for t in tasks]
    deps = list(
        TaskDependency.objects.filter(
            predecessor_id__in=task_ids,
            successor_id__in=task_ids,
        ).select_related("predecessor", "successor")
    )
    task_id_to_wbs = {row["task"].id: row["wbs"] for row in gantt_meta["rows"]}
    deps_display = [
        {
            "pred_id": d.predecessor_id,
            "succ_id": d.successor_id,
            "pred_wbs": task_id_to_wbs.get(d.predecessor_id, "?"),
            "succ_wbs": task_id_to_wbs.get(d.successor_id, "?"),
            "type": d.dependency_type,
            "lag_days": d.lag_days,
            "dep_id": d.pk,
        }
        for d in deps
    ]

    return render(
        request,
        "tasks/project_gantt.html",
        {
            **_tasks_shell_context(request, active="gantt", project=project),
            "page_title": f"Gantt kickoff - {project.name}",
            "project": project,
            "tasks": tasks,
            "gantt_rows": gantt_meta["rows"],
            "gantt_timeline_start": gantt_meta["timeline_start"],
            "gantt_timeline_end": gantt_meta["timeline_end"],
            "gantt_today": gantt_meta["today"],
            "gantt_days": gantt_meta["days"],
            "gantt_month_spans": gantt_meta["month_spans"],
            "can_edit_schedule": can_edit_schedule,
            "can_comment": can_comment,
            "task_update_forms": task_update_forms,
            "project_comment_form": comment_form,
            "project_summary": project_summary,
            "gantt_option_window_days": gantt_options["window_days"],
            "gantt_option_cell_size": gantt_options["cell_size"],
            "gantt_option_name_width": gantt_options["name_width"],
            "gantt_show_wbs": gantt_options["show_wbs"],
            "gantt_show_duration": gantt_options["show_duration"],
            "gantt_show_start": gantt_options["show_start"],
            "gantt_show_end": gantt_options["show_end"],
            "gantt_day_cell_px": gantt_options["day_cell_px"],
            "gantt_window_choices": gantt_options["window_choices"],
            "gantt_cell_choices": gantt_options["cell_choices"],
            "gantt_name_width_choices": gantt_options["name_width_choices"],
            "gantt_return_qs": gantt_options["return_qs"],
            "vrf_detail": vrf_detail,
            "gantt_baseline": baseline,
            "gantt_deps": deps_display,
            "dependency_type_choices": DependencyType.choices,
            "can_add_project_task": _has_task_permission(request, "tasks_create") and _can_manage_project(request, project),
        },
    )


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_fix_baseline(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, project):
        return render(request, "core/pages/forbidden.html", {"page_title": "Accesso negato"}, status=403)
    tasks = list(project.tasks.all())
    snapshot = {}
    for task in tasks:
        snapshot[str(task.pk)] = {
            "start": task.next_step_due.isoformat() if task.next_step_due else None,
            "end":   task.due_date.isoformat()      if task.due_date      else None,
        }
    GanttBaseline.objects.update_or_create(
        project=project,
        defaults={"snapshot": snapshot, "fixed_by": request.user},
    )
    log_action(request, "gantt_baseline_fixed", "tasks", {"project_id": project.pk, "task_count": len(tasks)})
    messages.success(request, "Baseline Gantt fissata.")
    return_qs = request.POST.get("return_qs", "")
    base_url = reverse("tasks:project_gantt", kwargs={"project_id": project_id})
    return redirect(f"{base_url}?{return_qs}" if return_qs else base_url)


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_clear_baseline(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, project):
        return render(request, "core/pages/forbidden.html", {"page_title": "Accesso negato"}, status=403)
    GanttBaseline.objects.filter(project=project).delete()
    log_action(request, "gantt_baseline_cleared", "tasks", {"project_id": project.pk})
    messages.success(request, "Baseline Gantt rimossa.")
    return_qs = request.POST.get("return_qs", "")
    base_url = reverse("tasks:project_gantt", kwargs={"project_id": project_id})
    return redirect(f"{base_url}?{return_qs}" if return_qs else base_url)


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_add_dependency(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, project):
        return JsonResponse({"ok": False, "error": "Accesso negato."}, status=403)
    pred_id  = request.POST.get("predecessor_id")
    succ_id  = request.POST.get("successor_id")
    dep_type = request.POST.get("dependency_type", DependencyType.FS)
    lag      = request.POST.get("lag_days", 0)
    if not pred_id or not succ_id:
        return JsonResponse({"ok": False, "error": "predecessor_id e successor_id obbligatori."}, status=400)
    if str(pred_id) == str(succ_id):
        return JsonResponse({"ok": False, "error": "Un task non può dipendere da se stesso."}, status=400)
    if dep_type not in dict(DependencyType.choices):
        dep_type = DependencyType.FS
    task_qs = Task.objects.filter(project=project)
    pred = get_object_or_404(task_qs, pk=pred_id)
    succ = get_object_or_404(task_qs, pk=succ_id)
    try:
        lag_int = int(lag)
    except (TypeError, ValueError):
        lag_int = 0
    dep, created = TaskDependency.objects.get_or_create(
        predecessor=pred,
        successor=succ,
        defaults={"dependency_type": dep_type, "lag_days": lag_int},
    )
    if not created:
        dep.dependency_type = dep_type
        dep.lag_days = lag_int
        dep.save(update_fields=["dependency_type", "lag_days"])
    log_action(request, "gantt_dependency_added", "tasks", {
        "project_id": project.pk,
        "predecessor_id": pred.pk,
        "successor_id": succ.pk,
        "type": dep_type,
        "lag_days": lag_int,
    })
    return JsonResponse({"ok": True, "dep_id": dep.pk, "created": created})


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_remove_dependency(request, project_id: int, dep_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, project):
        return JsonResponse({"ok": False, "error": "Accesso negato."}, status=403)
    task_ids = list(project.tasks.values_list("id", flat=True))
    dep = get_object_or_404(
        TaskDependency,
        pk=dep_id,
        predecessor_id__in=task_ids,
        successor_id__in=task_ids,
    )
    dep.delete()
    log_action(request, "gantt_dependency_removed", "tasks", {"project_id": project.pk, "dep_id": dep_id})
    return JsonResponse({"ok": True})


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_update_task(request, project_id: int, task_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    task = get_object_or_404(Task.objects.filter(project=project), pk=task_id)
    if not _can_manage_task(request, task):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )
    before = _task_snapshot(task)
    form = ProjectTaskGanttUpdateForm(request.POST, instance=task, prefix=f"task_{task.id}")
    if form.is_valid():
        task = form.save()
        _log_task_update_events(task, request.user, before)
        messages.success(request, f"Gantt aggiornato per task '{task.title}'.")
        _add_task_absence_warnings(request, task)
    else:
        errors = "; ".join(
            f"{form.fields[f].label or f}: {', '.join(errs)}"
            for f, errs in form.errors.items()
            if f != "__all__"
        )
        non_field = "; ".join(form.non_field_errors())
        detail = errors or non_field or "dati non validi"
        messages.error(request, f"Aggiornamento non salvato — {detail}")

    return_qs = str(request.POST.get("return_qs") or "").strip()
    target_url = reverse("tasks:project_gantt", kwargs={"project_id": project.id})
    if return_qs:
        target_url = f"{target_url}?{return_qs}"
    return redirect(target_url)


@require_POST
@task_permissions_required("tasks_view")
def project_gantt_shift_task(request, project_id: int, task_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    task = get_object_or_404(Task.objects.filter(project=project), pk=task_id)
    if not _can_manage_task(request, task):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    try:
        shift_days = int(str(request.POST.get("shift_days", "0")).strip())
    except (TypeError, ValueError):
        return JsonResponse({"ok": False, "error": "invalid_shift"}, status=400)

    if shift_days == 0:
        return JsonResponse(
            {
                "ok": True,
                "task_id": task.id,
                "shift_days": 0,
                "next_step_due": task.next_step_due.isoformat() if task.next_step_due else None,
                "due_date": task.due_date.isoformat() if task.due_date else None,
            }
        )

    # limite operativo anti-errori da drag involontario
    shift_days = max(-365, min(365, shift_days))

    # Vincolo predecessore implicito (WBS order): il task non può iniziare prima della fine
    # del task precedente. Il clamp garantisce la coerenza anche se il frontend è bypassato.
    predecessor = Task.objects.filter(project=project, id__lt=task.id).order_by("-id").first()
    if predecessor:
        pred_end = predecessor.due_date or predecessor.next_step_due
        if pred_end:
            orig_start = task.next_step_due or task.due_date
            if orig_start:
                min_shift = (pred_end + timedelta(days=1) - orig_start).days
                shift_days = max(shift_days, min_shift)

    # Salva la data originale di start PRIMA dello spostamento (per il cascade)
    original_active_start = task.next_step_due or task.due_date

    before = _task_snapshot(task)

    if task.next_step_due:
        task.next_step_due = task.next_step_due + timedelta(days=shift_days)
    if task.due_date:
        task.due_date = task.due_date + timedelta(days=shift_days)
    task.save(update_fields=["next_step_due", "due_date", "updated_at"])
    _log_task_update_events(task, request.user, before)

    # Cascade sequenziale: rispetta l'ordine WBS (creazione) e garantisce che
    # ogni task inizi DOPO la fine della precedente nella sequenza.
    cascade = str(request.POST.get("cascade", "1")).strip() != "0"
    cascade_count = 0
    if cascade:
        # Tutte le task del progetto in ordine WBS (id asc = ordine di creazione)
        all_project_tasks = list(Task.objects.filter(project=project).order_by("id"))

        # Sostituisce la task appena salvata con la versione aggiornata in memoria
        for i, t in enumerate(all_project_tasks):
            if t.id == task.id:
                all_project_tasks[i] = task
                dragged_idx = i
                break
        else:
            dragged_idx = -1

        if dragged_idx >= 0:
            for i in range(dragged_idx + 1, len(all_project_tasks)):
                ft = all_project_tasks[i]
                prev = all_project_tasks[i - 1]

                # Data di fine della task precedente
                prev_end = prev.due_date or prev.next_step_due
                if not prev_end:
                    continue

                # Data di inizio della task corrente
                ft_start = ft.next_step_due or ft.due_date
                if not ft_start:
                    continue

                if ft_start <= prev_end:
                    # La task corrente inizia prima o nello stesso giorno in cui finisce la precedente:
                    # la spostiamo al giorno successivo alla fine della precedente,
                    # mantenendo la durata originale.
                    new_start = prev_end + timedelta(days=1)
                    task_delta = (new_start - ft_start).days
                    if ft.next_step_due:
                        ft.next_step_due = ft.next_step_due + timedelta(days=task_delta)
                    if ft.due_date:
                        ft.due_date = ft.due_date + timedelta(days=task_delta)
                    ft.save(update_fields=["next_step_due", "due_date", "updated_at"])
                    cascade_count += 1

    conflict_messages: list[str] = []
    conflicts = _task_date_absence_conflicts(task)
    if conflicts:
        for field_name in ("next_step_due", "due_date"):
            if field_name not in conflicts:
                continue
            target_date = getattr(task, field_name)
            if not target_date:
                continue
            labels = [f"{entry['tipo']} ({entry['status']})" for entry in conflicts[field_name]]
            label_text = ", ".join(dict.fromkeys(labels))
            conflict_messages.append(
                f"{field_name}:{target_date.strftime('%d/%m/%Y')}:{label_text}"
            )

    return JsonResponse(
        {
            "ok": True,
            "task_id": task.id,
            "shift_days": shift_days,
            "cascade_count": cascade_count,
            "next_step_due": task.next_step_due.isoformat() if task.next_step_due else None,
            "due_date": task.due_date.isoformat() if task.due_date else None,
            "next_step_due_display": task.next_step_due.strftime("%d/%m/%Y") if task.next_step_due else "",
            "due_date_display": task.due_date.strftime("%d/%m/%Y") if task.due_date else "",
            "absence_conflicts": conflict_messages,
        }
    )


@require_POST
@task_permissions_required("tasks_view", "tasks_comment")
def add_project_comment(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    notify_users_qs = _project_notify_users_queryset(project)
    form = ProjectCommentForm(request.POST, user=request.user, notify_user_queryset=notify_users_qs)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.project = project
        comment.author = request.user
        comment.save()
        if comment.target_user_id and comment.target_user_id != request.user.id:
            _notify_user(
                comment.target_user,
                message_text=f"Nuovo commento nel kickoff '{project.name}'.",
                action_url=reverse("tasks:project_gantt", kwargs={"project_id": project.id}),
            )
        messages.success(request, "Commento kickoff aggiunto.")
    else:
        messages.error(request, "Commento kickoff non valido.")
    return_qs = str(request.POST.get("return_qs") or "").strip()
    target_url = reverse("tasks:project_gantt", kwargs={"project_id": project.id})
    if return_qs:
        target_url = f"{target_url}?{return_qs}"
    return redirect(target_url)


@legacy_admin_required
def gestione_admin(request):
    """Pagina di gestione interna Tasks — accesso solo admin."""
    from django.core.paginator import Paginator

    today = timezone.localdate()
    tab = request.GET.get("tab", "riepilogo")

    # --- Statistiche ---
    total_tasks = Task.objects.count()
    total_projects = Project.objects.count()
    _tasks_by_status_raw = dict(Task.objects.values_list("status").annotate(n=Count("id")).order_by())
    tasks_by_status = [(val, lbl, _tasks_by_status_raw.get(val, 0)) for val, lbl in TaskStatus.choices]
    tasks_overdue = Task.objects.filter(
        due_date__lt=today,
        status__in=[TaskStatus.TODO, TaskStatus.IN_PROGRESS],
    ).count()
    top_assignees = list(
        Task.objects.filter(assigned_to__isnull=False)
        .values("assigned_to__username")
        .annotate(n=Count("id"))
        .order_by("-n")[:10]
    )
    todo_count = _tasks_by_status_raw.get(TaskStatus.TODO, 0)
    in_progress_count = _tasks_by_status_raw.get(TaskStatus.IN_PROGRESS, 0)
    done_count = _tasks_by_status_raw.get(TaskStatus.DONE, 0)

    # --- Record ---
    q_task = request.GET.get("q_task", "").strip()
    q_proj = request.GET.get("q_proj", "").strip()
    filter_status = request.GET.get("filter_status", "").strip()

    tasks_qs = Task.objects.select_related("project", "assigned_to").order_by("-updated_at")
    if q_task:
        tasks_qs = tasks_qs.filter(Q(title__icontains=q_task))
    if filter_status:
        tasks_qs = tasks_qs.filter(status=filter_status)
    tasks_page = Paginator(tasks_qs, 50).get_page(request.GET.get("task_page"))

    projects_qs = Project.objects.select_related("project_manager").order_by("-updated_at")
    if q_proj:
        projects_qs = projects_qs.filter(Q(name__icontains=q_proj) | Q(client_name__icontains=q_proj))
    projects_page = Paginator(projects_qs, 50).get_page(request.GET.get("proj_page"))

    # --- Log ---
    audit_entries = AuditLog.objects.filter(modulo="tasks").order_by("-created_at")[:100]

    return render(
        request,
        "tasks/gestione_admin.html",
        {
            **_tasks_shell_context(request, active="admin"),
            "page_title": "Gestione KICK-OFF",
            "tab": tab,
            # stats
            "total_tasks": total_tasks,
            "total_projects": total_projects,
            "tasks_by_status": tasks_by_status,
            "tasks_overdue": tasks_overdue,
            "top_assignees": top_assignees,
            "task_status_choices": TaskStatus.choices,
            "todo_count": todo_count,
            "in_progress_count": in_progress_count,
            "done_count": done_count,
            # records
            "tasks_page": tasks_page,
            "projects_page": projects_page,
            "q_task": q_task,
            "q_proj": q_proj,
            "filter_status": filter_status,
            # log
            "audit_entries": audit_entries,
        },
    )


@legacy_admin_required
def impostazioni(request):
    """Impostazioni globali del modulo Task."""
    cfg = TaskImpostazioni.get_singleton()

    if request.method == "POST":
        branding_response = handle_module_branding_post(
            request,
            module_key="tasks",
            redirect_to="tasks:impostazioni",
            audit_module="tasks",
            fallback_label="KICK-OFF",
        )
        if branding_response is not None:
            return branding_response
        cfg.responsabile_email = request.POST.get("responsabile_email", "").strip()
        cfg.notifiche_scadenza_attive = bool(request.POST.get("notifiche_scadenza_attive"))
        cfg.giorni_preavviso = max(1, int(request.POST.get("giorni_preavviso") or 3))
        cfg.note_generali = request.POST.get("note_generali", "").strip()
        cfg.vrf_reminder_days = max(1, int(request.POST.get("vrf_reminder_days") or 7))
        cfg.vrf_blocking_days = max(1, int(request.POST.get("vrf_blocking_days") or 30))
        cfg.save()
        log_action(request, "modifica", "tasks", {"message": "Aggiornate impostazioni KICK-OFF"})
        messages.success(request, "Impostazioni salvate.")
        return redirect("tasks:impostazioni")

    return render(request, "tasks/impostazioni.html", {
        "cfg": cfg,
        "tasks_shell_active": "impostazioni",
        "tasks_shell_can_admin": True,
        **get_module_branding_context("tasks", fallback_label="KICK-OFF"),
    })


@legacy_admin_required
def gestione_admin(request):
    """Compat legacy: la vecchia gestione admin confluisce nelle impostazioni."""
    params = request.GET.copy()
    params["tab"] = _normalize_tasks_settings_tab(params.get("tab"), default="riepilogo")
    target_url = reverse("tasks:impostazioni")
    query_string = params.urlencode()
    if query_string:
        target_url = f"{target_url}?{query_string}"
    return redirect(target_url)


def _handle_tasks_roles_post(request):
    """POST tab ruoli: set completo delle assegnazioni ricalcolato dal form."""
    target_url = f"{reverse('tasks:impostazioni')}?tab=ruoli"
    q_user = request.POST.get("q_user", "").strip()
    if q_user:
        target_url = f"{target_url}&q_user={q_user}"

    role_admin_action = (request.POST.get("role_admin_action") or "").strip()
    if role_admin_action == "create_role":
        name = (request.POST.get("role_name") or "").strip()
        description = (request.POST.get("role_description") or "").strip()[:255]
        if not name:
            messages.error(request, "Indica il nome del nuovo ruolo.")
            return redirect(target_url)
        base_code = slugify(name).upper().replace("-", "_")[:32] or "RUOLO"
        code = base_code
        n = 2
        while TaskRoleDefinition.objects.filter(code=code).exists():
            suffix = f"_{n}"
            code = f"{base_code[:32 - len(suffix)]}{suffix}"
            n += 1
        max_order = TaskRoleDefinition.objects.aggregate(m=Max("order_index")).get("m") or 30
        role = TaskRoleDefinition.objects.create(
            code=code,
            name=name[:80],
            description=description,
            is_system=False,
            is_active=True,
            order_index=max_order + 10,
        )
        log_action(request, "ruolo_creato", "tasks", {"message": f"Creato ruolo operativo {role.name}", "role": role.code})
        messages.success(request, f"Ruolo '{role.name}' creato.")
        return redirect(target_url)

    if role_admin_action == "delete_role":
        role_code = (request.POST.get("role_code") or "").strip()
        try:
            role = TaskRoleDefinition.objects.get(code=role_code, is_system=False)
        except TaskRoleDefinition.DoesNotExist:
            messages.error(request, "Ruolo non eliminabile o non trovato.")
            return redirect(target_url)
        role_name = role.name
        with transaction.atomic():
            TaskRoleAssignment.objects.filter(role_type=role.code).delete()
            TaskRoleAccessRule.objects.filter(role_type=role.code).delete()
            TaskCategory.objects.filter(role_type=role.code).update(role_type="")
            role.delete()
        log_action(request, "ruolo_eliminato", "tasks", {"message": f"Eliminato ruolo operativo {role_name}", "role": role_code})
        messages.success(request, f"Ruolo '{role_name}' eliminato.")
        return redirect(target_url)

    valid_role_codes = {role.code for role in _task_role_definitions(include_inactive=False)}

    # raccoglie tutti i checkbox inviati: name="role__<ROLE_CODE>__<user_id>"
    desired: set[tuple[int, str]] = set()
    for key in request.POST:
        if not key.startswith("role__"):
            continue
        parts = key.split("__")
        if len(parts) != 3:
            continue
        _, role_code, uid_raw = parts
        if role_code not in valid_role_codes:
            continue
        try:
            uid = int(uid_raw)
        except (TypeError, ValueError):
            continue
        desired.add((uid, role_code))

    # valutiamo solo gli utenti visibili nel form (non tocchiamo utenti filtrati fuori)
    visible_uids = [int(x) for x in request.POST.getlist("visible_user_id")]
    if not visible_uids:
        messages.warning(request, "Nessun utente visibile: nessuna modifica ai ruoli.")
        return redirect(target_url)

    # delete + recreate (diff sul subset visible)
    existing = set(
        TaskRoleAssignment.objects.filter(user_id__in=visible_uids)
        .values_list("user_id", "role_type")
    )
    to_add = {pair for pair in desired if pair[0] in visible_uids} - existing
    to_remove = existing - {pair for pair in desired if pair[0] in visible_uids}

    with transaction.atomic():
        if to_remove:
            q = Q()
            for uid, rtype in to_remove:
                q |= Q(user_id=uid, role_type=rtype)
            TaskRoleAssignment.objects.filter(q).delete()
        TaskRoleAssignment.objects.bulk_create([
            TaskRoleAssignment(user_id=uid, role_type=rtype)
            for uid, rtype in to_add
        ])

    log_action(request, "ruoli_aggiornati", "tasks", {
        "message": f"Aggiornati ruoli operativi kickoff: +{len(to_add)} -{len(to_remove)}",
        "added": sorted(to_add),
        "removed": sorted(to_remove),
    })
    messages.success(
        request,
        f"Ruoli aggiornati: +{len(to_add)} aggiunti, {len(to_remove)} rimossi.",
    )
    return redirect(target_url)


def _handle_tasks_access_post(request):
    """POST tab accessi: regole ruolo operativo + override singolo utente."""
    from urllib.parse import urlencode

    params = {"tab": "accessi"}
    q_user = request.POST.get("q_access_user", "").strip()
    if q_user:
        params["q_access_user"] = q_user
    target_url = f"{reverse('tasks:impostazioni')}?{urlencode(params)}"

    valid_role_values = {
        TaskAccessLevel.NONE,
        TaskAccessLevel.READ_ALL,
        TaskAccessLevel.EDIT_ASSIGNED,
        TaskAccessLevel.EDIT_ALL,
    }
    valid_user_values = {
        TaskAccessLevel.READ_ALL,
        TaskAccessLevel.EDIT_ASSIGNED,
        TaskAccessLevel.EDIT_ALL,
    }

    role_updates: dict[str, str] = {}
    for role in _task_role_definitions(include_inactive=False):
        role_code = role.code
        raw_value = (request.POST.get(f"access_role__{role_code}") or TaskAccessLevel.NONE).strip()
        role_updates[role_code] = raw_value if raw_value in valid_role_values else TaskAccessLevel.NONE

    existing_role_map = {
        role_type: access_level
        for role_type, access_level in TaskRoleAccessRule.objects.values_list("role_type", "access_level")
    }
    role_changes = 0
    with transaction.atomic():
        for role_code, desired_level in role_updates.items():
            current_level = existing_role_map.get(role_code)
            if desired_level == TaskAccessLevel.NONE:
                if current_level is not None:
                    TaskRoleAccessRule.objects.filter(role_type=role_code).delete()
                    role_changes += 1
                continue
            if current_level != desired_level:
                TaskRoleAccessRule.objects.update_or_create(
                    role_type=role_code,
                    defaults={"access_level": desired_level},
                )
                role_changes += 1

        visible_uids: list[int] = []
        for raw_uid in request.POST.getlist("visible_access_user_id"):
            try:
                visible_uids.append(int(raw_uid))
            except (TypeError, ValueError):
                continue
        visible_uids = sorted(set(visible_uids))

        existing_user_map = {
            user_id: access_level
            for user_id, access_level in TaskUserAccessRule.objects.filter(user_id__in=visible_uids).values_list("user_id", "access_level")
        }
        desired_user_map: dict[int, str] = {}
        for user_id in visible_uids:
            raw_value = (request.POST.get(f"access_user__{user_id}") or "").strip()
            if raw_value in valid_user_values:
                desired_user_map[user_id] = raw_value

        user_changes = 0
        to_remove = sorted(set(existing_user_map) - set(desired_user_map))
        if to_remove:
            TaskUserAccessRule.objects.filter(user_id__in=to_remove).delete()
            user_changes += len(to_remove)

        for user_id, desired_level in desired_user_map.items():
            if existing_user_map.get(user_id) == desired_level:
                continue
            TaskUserAccessRule.objects.update_or_create(
                user_id=user_id,
                defaults={"access_level": desired_level},
            )
            user_changes += 1

    log_action(
        request,
        "accessi_aggiornati",
        "tasks",
        {
            "message": (
                f"Aggiornate regole accesso KICK-OFF: {role_changes} regole ruolo, "
                f"{user_changes} override utente."
            ),
            "role_updates": role_updates,
            "visible_users": visible_uids,
        },
    )
    messages.success(
        request,
        f"Regole accesso aggiornate: {role_changes} ruoli, {user_changes} override utente.",
    )
    return redirect(target_url)


def _handle_tasks_reminders_post(request):
    """POST tab promemoria: azioni su TaskReminder (delete/postpone/fire_now)."""
    target_url = f"{reverse('tasks:impostazioni')}?tab=promemoria"
    filter_status = (request.POST.get("filter_status") or "").strip()
    if filter_status:
        target_url = f"{target_url}&filter_status={filter_status}"

    action = (request.POST.get("reminder_action") or "").strip()
    ids_raw = request.POST.getlist("reminder_id")
    ids: list[int] = []
    for r in ids_raw:
        try:
            ids.append(int(r))
        except (TypeError, ValueError):
            pass

    if not ids:
        messages.warning(request, "Nessun promemoria selezionato.")
        return redirect(target_url)

    qs = TaskReminder.objects.filter(id__in=ids)

    if action == "delete":
        n = qs.count()
        qs.delete()
        log_action(request, "reminder_eliminati", "tasks", {
            "message": f"Eliminati {n} promemoria kickoff",
            "ids": ids,
        })
        messages.success(request, f"{n} promemoria eliminati.")
        return redirect(target_url)

    if action == "postpone":
        try:
            days = max(1, int(request.POST.get("postpone_days") or 0))
        except (TypeError, ValueError):
            days = 0
        if not days:
            messages.error(request, "Specifica il numero di giorni di rinvio.")
            return redirect(target_url)
        updated = 0
        for rem in qs.filter(fired=False):
            rem.fire_at = rem.fire_at + timedelta(days=days)
            rem.save(update_fields=["fire_at"])
            updated += 1
        log_action(request, "reminder_rimandati", "tasks", {
            "message": f"Rimandati {updated} promemoria kickoff di {days} giorni",
            "ids": ids,
            "days": days,
        })
        messages.success(request, f"{updated} promemoria rimandati di {days} giorni.")
        return redirect(target_url)

    if action == "fire_now":
        fired = 0
        skipped = 0
        for rem in qs.select_related("task", "task__project").filter(fired=False):
            task = rem.task
            if task is None:
                rem.fired = True
                rem.fired_at = timezone.now()
                rem.save(update_fields=["fired", "fired_at"])
                skipped += 1
                continue
            if task.status in {TaskStatus.DONE, TaskStatus.CANCELED}:
                rem.fired = True
                rem.fired_at = timezone.now()
                rem.save(update_fields=["fired", "fired_at"])
                skipped += 1
                continue
            project_label = f" [{task.project.name}]" if task.project_id and task.project else ""
            due_str = task.due_date.strftime("%d/%m/%Y") if task.due_date else "(senza data)"
            message_text = (
                f"Promemoria scadenza attivita kickoff: \"{task.title}\"{project_label} "
                f"in scadenza il {due_str}."
            )[:500]
            Notifica.objects.create(
                legacy_user_id=int(rem.legacy_user_id or 0),
                tipo="generico",
                messaggio=message_text,
                url_azione=reverse("tasks:detail", args=[task.id]),
            )
            rem.fired = True
            rem.fired_at = timezone.now()
            rem.save(update_fields=["fired", "fired_at"])
            fired += 1
        log_action(request, "reminder_forzati", "tasks", {
            "message": f"Forzato invio di {fired} promemoria kickoff ({skipped} saltati)",
            "ids": ids,
        })
        messages.success(request, f"Inviati {fired} promemoria ({skipped} saltati).")
        return redirect(target_url)

    messages.error(request, "Azione non riconosciuta.")
    return redirect(target_url)


def _render_task_extra_data(task) -> list[dict]:
    """Risolve i valori salvati in task.extra_data in righe display-ready
    per il template detail: [{'label', 'value', 'href'}]. Per i campi FK
    restituisce display name + link alla scheda.
    """
    if not task.category_id:
        return []
    rows: list[dict] = []
    data = task.extra_data or {}
    fields = list(task.category.fields.all().order_by("order_index", "id"))
    user_ids = {v for v in (data.get(f.code) for f in fields if f.field_type == TaskCategoryFieldType.USER) if v}
    asset_ids = {v for v in (data.get(f.code) for f in fields if f.field_type == TaskCategoryFieldType.ASSET) if v}
    users_map: dict[int, str] = {}
    assets_map: dict[int, tuple[str, str]] = {}
    if user_ids:
        users_map = {
            u.id: (u.get_full_name() or u.username)
            for u in User.objects.filter(pk__in=user_ids)
        }
    if asset_ids:
        try:
            from assets.models import Asset
            for a in Asset.objects.filter(pk__in=asset_ids):
                tag = getattr(a, "asset_tag", "") or getattr(a, "code", "")
                label = f"{tag} - {a.name}" if tag else a.name
                try:
                    href = reverse("assets:asset_view", kwargs={"id": a.id})
                except Exception:
                    href = ""
                assets_map[a.id] = (label, href)
        except Exception:
            pass

    for f in fields:
        raw = data.get(f.code)
        value: str = ""
        href: str = ""
        if f.field_type == TaskCategoryFieldType.CHECKBOX:
            value = "Si" if raw else "No"
        elif f.field_type == TaskCategoryFieldType.MULTISELECT:
            labels = {c["value"]: c["label"] for c in f.normalized_choices()}
            value = ", ".join(labels.get(v, str(v)) for v in (raw or []))
        elif f.field_type == TaskCategoryFieldType.SELECT:
            labels = {c["value"]: c["label"] for c in f.normalized_choices()}
            value = labels.get(raw, str(raw or ""))
        elif f.field_type == TaskCategoryFieldType.USER:
            if raw:
                value = users_map.get(int(raw), f"Utente #{raw}")
        elif f.field_type == TaskCategoryFieldType.ASSET:
            if raw:
                label, href = assets_map.get(int(raw), (f"Asset #{raw}", ""))
                value = label
        else:
            value = "" if raw in (None, "") else str(raw)
        if value == "":
            continue
        rows.append({"label": f.label, "value": value, "href": href, "code": f.code})
    return rows


def _serialize_category_field(field) -> dict:
    return {
        "id": field.id,
        "code": field.code,
        "label": field.label,
        "field_type": field.field_type,
        "required": field.required,
        "order": field.order_index,
        "help_text": field.help_text,
        "placeholder": field.placeholder,
        "choices": field.normalized_choices(),
        "depends_on_code": field.depends_on_code,
        "depends_on_value": field.depends_on_value,
        "asset_type_filter": field.asset_type_filter or "",
        "asset_category_filter": field.asset_category_filter or "",
    }


def _build_asset_options_for_field(field) -> list[dict]:
    """Opzioni asset per un TaskCategoryField di tipo asset.

    Applica i filtri configurati (asset_type_filter, asset_category_filter) e
    esclude in ogni caso gli asset con status RETIRED o IN_STOCK (non sono
    candidati ad attivita operative).
    """
    try:
        from assets.models import Asset
    except ImportError:
        return []
    qs = Asset.objects.exclude(status__in=[Asset.STATUS_RETIRED, Asset.STATUS_IN_STOCK])
    types = field.asset_type_filter_list()
    if types:
        qs = qs.filter(asset_type__in=types)
    cat_ids = field.asset_category_filter_ids()
    if cat_ids:
        qs = qs.filter(asset_category_id__in=cat_ids)
    out: list[dict] = []
    for a in qs.order_by("name")[:500]:
        tag = getattr(a, "asset_tag", "") or ""
        label = f"{tag} - {a.name}" if tag else a.name
        out.append({"value": a.id, "label": label})
    return out


@task_permissions_required("tasks_view")
def category_fields_json(request, category_id: int):
    """Espone la definizione dei campi di una categoria al form nuova attivita.

    Per i campi di tipo 'user' / 'asset' restituisce anche un piccolo dataset
    di opzioni (utenti attivi, asset base) cosi' il frontend puo' popolare
    la select senza una seconda chiamata.
    """
    try:
        category = TaskCategory.objects.prefetch_related("fields").get(pk=category_id, is_active=True)
    except TaskCategory.DoesNotExist:
        return JsonResponse({"ok": False, "reason": "not_found"}, status=404)

    raw_fields = list(category.fields.all().order_by("order_index", "id"))
    fields = [_serialize_category_field(f) for f in raw_fields]

    users_payload: list[dict] = []
    needs_users = any(f["field_type"] == TaskCategoryFieldType.USER for f in fields)

    # Popola le opzioni asset per-campo (ogni campo puo' avere filtri diversi).
    for raw, ser in zip(raw_fields, fields):
        if ser["field_type"] == TaskCategoryFieldType.ASSET:
            ser["choices"] = _build_asset_options_for_field(raw)

    if needs_users:
        users_payload = [
            {"value": u.id, "label": (u.get_full_name() or u.username)}
            for u in task_active_users_queryset()[:500]
        ]

    return JsonResponse({
        "ok": True,
        "category": {"id": category.id, "name": category.name, "slug": category.slug},
        "fields": fields,
        "users": users_payload,
        # retrocompat: il frontend accede a payload.assets; le choices asset
        # vivono ora per-campo in field.choices.
        "assets": [],
    })


def _parse_iso_date(raw: str):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _task_time_window(task: "Task"):
    """Finestra temporale di un task, normalizzata a (start, end).

    - Se next_step_due e due_date sono entrambi presenti, usa [next_step_due, due_date].
    - Se solo due_date, finestra = [due_date, due_date].
    - Se solo next_step_due, finestra = [next_step_due, next_step_due].
    - Se nessuno dei due, ritorna (None, None).
    """
    start = getattr(task, "next_step_due", None)
    end = getattr(task, "due_date", None)
    if start and end:
        if end < start:
            start, end = end, start
        return start, end
    if end:
        return end, end
    if start:
        return start, start
    return None, None


def _windows_overlap(a_start, a_end, b_start, b_end) -> bool:
    if not a_start or not a_end or not b_start or not b_end:
        return False
    return a_start <= b_end and b_start <= a_end


def _asset_availability_report(
    asset,
    start,
    end,
    exclude_task_id: int | None = None,
) -> dict:
    """Calcola il report di disponibilita per un asset nella finestra [start, end].

    Il risultato contiene: stato asset, OdL aperti, verifiche periodiche, ticket
    manutenzione aperti, task sovrapposti (via TaskExtraRef), severita
    complessiva.
    """
    from assets.models import Asset, WorkOrder, PeriodicVerification

    conflicts: list[dict] = []
    overall = "ok"

    def bump(level: str) -> None:
        nonlocal overall
        order = {"ok": 0, "warning": 1, "block": 2}
        if order[level] > order[overall]:
            overall = level

    # 1) Status asset
    if asset.status == Asset.STATUS_IN_REPAIR:
        conflicts.append({
            "type": "asset_status",
            "severity": "block",
            "title": f"Asset in riparazione: {asset.get_status_display()}",
            "detail": asset.notes or "",
            "when": "",
            "url": "",
        })
        bump("block")

    # 2) Ordini di lavoro aperti. Se la finestra e' nota, consideriamo solo
    #    quelli che si sovrappongono; altrimenti li riportiamo tutti (OPEN).
    wo_qs = WorkOrder.objects.filter(
        asset=asset,
        status=WorkOrder.STATUS_OPEN,
    ).order_by("opened_at")
    try:
        asset_view_url = reverse("assets:asset_view", kwargs={"id": asset.id})
    except Exception:
        asset_view_url = ""
    for wo in wo_qs[:50]:
        wo_start = wo.opened_at.date() if wo.opened_at else None
        wo_end = wo.closed_at.date() if wo.closed_at else None
        if start and end and wo_start:
            # consideriamo un OdL ancora aperto senza data di chiusura come
            # "in corso" fino a oggi incluso
            effective_end = wo_end or timezone.localdate()
            if not _windows_overlap(start, end, wo_start, effective_end):
                continue
        conflicts.append({
            "type": "workorder",
            "severity": "block",
            "title": f"OdL aperto #{wo.id}: {wo.title}",
            "detail": wo.get_kind_display(),
            "when": wo_start.isoformat() if wo_start else "",
            "url": asset_view_url,
        })
        bump("block")

    # 3) Verifiche periodiche in finestra
    pv_qs = PeriodicVerification.objects.filter(
        assets=asset,
        is_active=True,
        next_verification_date__isnull=False,
    )
    if start and end:
        pv_qs = pv_qs.filter(next_verification_date__gte=start, next_verification_date__lte=end)
    for pv in pv_qs.order_by("next_verification_date")[:20]:
        conflicts.append({
            "type": "verification",
            "severity": "warning",
            "title": f"Verifica periodica: {pv.name}",
            "detail": "",
            "when": pv.next_verification_date.isoformat() if pv.next_verification_date else "",
            "url": asset_view_url,
        })
        bump("warning")

    # 4) Ticket manutenzione aperti sull'asset
    try:
        from tickets.models import Ticket, StatoTicket
        open_states = [
            StatoTicket.APERTA,
            StatoTicket.IN_CARICO,
            StatoTicket.IN_ATTESA,
        ]
        for tk in Ticket.objects.filter(asset=asset, stato__in=open_states).order_by("-pk")[:20]:
            try:
                tk_url = reverse("tickets:detail", kwargs={"ticket_id": tk.id})
            except Exception:
                tk_url = ""
            conflicts.append({
                "type": "ticket",
                "severity": "warning",
                "title": f"Ticket {tk.numero_ticket}: {tk.titolo}",
                "detail": tk.get_stato_display(),
                "when": "",
                "url": tk_url,
            })
            bump("warning")
    except Exception:
        pass

    # 5) Task sovrapposti (stesso asset via TaskExtraRef, stato attivo)
    if start and end:
        tref_qs = (
            TaskExtraRef.objects.filter(asset=asset)
            .exclude(task__status__in=[TaskStatus.DONE, TaskStatus.CANCELED])
            .select_related("task")
        )
        if exclude_task_id:
            tref_qs = tref_qs.exclude(task_id=exclude_task_id)
        seen_task_ids: set[int] = set()
        for ref in tref_qs[:100]:
            if ref.task_id in seen_task_ids:
                continue
            other = ref.task
            o_start, o_end = _task_time_window(other)
            if not _windows_overlap(start, end, o_start, o_end):
                continue
            seen_task_ids.add(ref.task_id)
            try:
                t_url = reverse("tasks:detail", kwargs={"task_id": other.id})
            except Exception:
                t_url = ""
            when_label = o_start.isoformat() if o_start else ""
            if o_start and o_end and o_start != o_end:
                when_label = f"{o_start.isoformat()} -> {o_end.isoformat()}"
            conflicts.append({
                "type": "task",
                "severity": "block",
                "title": f"Attivita gia' assegnata: {other.title}",
                "detail": other.get_status_display(),
                "when": when_label,
                "url": t_url,
            })
            bump("block")

    return {
        "asset": {
            "id": asset.id,
            "tag": getattr(asset, "asset_tag", "") or "",
            "name": asset.name,
            "status": asset.status,
            "status_label": asset.get_status_display(),
        },
        "window": {
            "start": start.isoformat() if start else "",
            "end": end.isoformat() if end else "",
        },
        "conflicts": conflicts,
        "overall_severity": overall,
        "has_blocking_conflict": any(c["severity"] == "block" for c in conflicts),
    }


def _check_machine_work_overlaps(task: "Task", exclude_task_id: int | None) -> list[str]:
    """Ritorna avvisi (non bloccanti) se l'asset assegnato è già occupato da
    un altro task 'lavoro macchina' nella stessa finestra temporale."""
    if not task.category_id or not getattr(task.category, "is_machine_work", False):
        return []
    start = task.next_step_due
    end = task.due_date
    if not start and not end:
        return []
    asset_refs = list(
        task.extra_refs.filter(asset__isnull=False).select_related("asset")
    )
    if not asset_refs:
        return []
    warnings_out: list[str] = []
    for ref in asset_refs:
        qs = (
            TaskExtraRef.objects
            .filter(asset_id=ref.asset_id)
            .exclude(task_id=task.pk)
            .select_related("task__category", "task__assigned_to")
            .filter(
                task__category__is_machine_work=True,
                task__status__in=[TaskStatus.TODO, TaskStatus.IN_PROGRESS],
            )
        )
        if exclude_task_id:
            qs = qs.exclude(task_id=exclude_task_id)
        conflicts = []
        for other_ref in qs:
            t = other_ref.task
            t_start = t.next_step_due
            t_end = t.due_date
            overlaps = True
            if start and t_end and t_end < start:
                overlaps = False
            if end and t_start and t_start > end:
                overlaps = False
            if overlaps:
                conflicts.append(t)
        if conflicts:
            asset_label = f"{ref.asset.asset_tag} - {ref.asset.name}"
            titles = ", ".join(f"'{c.title}'" for c in conflicts[:3])
            suffix = f" (+{len(conflicts)-3} altri)" if len(conflicts) > 3 else ""
            warnings_out.append(
                f"Macchina {asset_label} già impegnata nello stesso periodo: {titles}{suffix}."
            )
    return warnings_out


def _check_blocking_asset_conflicts(task: "Task", post_data, exclude_task_id: int | None) -> list[str]:
    """Ritorna la lista di messaggi di errore per eventuali conflitti asset
    bloccanti, usando la stessa logica dell'endpoint availability. Se le
    impostazioni disabilitano il blocco, ritorna [].
    """
    cfg = TaskImpostazioni.get_singleton()
    if not cfg.asset_conflict_check_enabled or not cfg.asset_conflict_block:
        return []
    if not task.category_id:
        return []
    try:
        from assets.models import Asset
    except ImportError:
        return []
    fields = list(task.category.fields.all().order_by("order_index", "id"))
    asset_fields = [f for f in fields if f.field_type == TaskCategoryFieldType.ASSET]
    if not asset_fields:
        return []
    start, end = _task_time_window(task)
    errors: list[str] = []
    for f in asset_fields:
        raw = (post_data.get(f"extra__{f.code}") or "").strip()
        if not raw:
            continue
        try:
            aid = int(raw)
        except (TypeError, ValueError):
            continue
        try:
            asset = Asset.objects.get(pk=aid)
        except Asset.DoesNotExist:
            continue
        report = _asset_availability_report(asset, start, end, exclude_task_id)
        if report["has_blocking_conflict"]:
            tag = report["asset"].get("tag") or ""
            name = report["asset"].get("name") or ""
            asset_label = f"{tag} - {name}" if tag else name
            blocking = [c for c in report["conflicts"] if c["severity"] == "block"]
            details = "; ".join(c["title"] for c in blocking[:3])
            errors.append(
                f"Campo '{f.label}' ({asset_label}): conflitto bloccante - {details}"
            )
    return errors


@task_permissions_required("tasks_view")
def asset_availability_json(request, asset_id: int):
    """Ritorna il report di disponibilita per un asset nella finestra richiesta.

    Query string:
      start=YYYY-MM-DD (opzionale)
      end=YYYY-MM-DD (opzionale)
      exclude_task=<id> (opzionale, per l'edit di un task esistente)
    """
    try:
        from assets.models import Asset
        asset = Asset.objects.get(pk=asset_id)
    except Exception:
        return JsonResponse({"ok": False, "reason": "not_found"}, status=404)

    start = _parse_iso_date(request.GET.get("start") or "")
    end = _parse_iso_date(request.GET.get("end") or "")
    if start and end and end < start:
        start, end = end, start
    exclude_task_raw = request.GET.get("exclude_task") or ""
    exclude_task_id: int | None = None
    if exclude_task_raw:
        try:
            exclude_task_id = int(exclude_task_raw)
        except (TypeError, ValueError):
            exclude_task_id = None

    cfg = TaskImpostazioni.get_singleton()
    report = _asset_availability_report(asset, start, end, exclude_task_id)
    report["ok"] = True
    report["check_enabled"] = bool(cfg.asset_conflict_check_enabled)
    report["block_on_conflict"] = bool(cfg.asset_conflict_block)
    return JsonResponse(report)


def _persist_task_extra_data(task: "Task", post_data) -> None:
    """Legge i campi extra dal POST per la categoria di `task`, valida i
    valori secondo il catalogo TaskCategoryField e salva:
      - task.extra_data (JSON)
      - TaskExtraRef (righe normalizzate per FK user/asset)
    """
    task.extra_data = task.extra_data or {}

    if not task.category_id:
        if task.extra_data:
            task.extra_data = {}
            task.save(update_fields=["extra_data"])
        TaskExtraRef.objects.filter(task=task).delete()
        return

    fields = list(task.category.fields.all().order_by("order_index", "id"))
    parsed: dict[str, object] = {}
    refs_to_create: list[TaskExtraRef] = []

    for f in fields:
        name = f"extra__{f.code}"
        if f.field_type == TaskCategoryFieldType.CHECKBOX:
            parsed[f.code] = bool(post_data.get(name))
            continue
        if f.field_type == TaskCategoryFieldType.MULTISELECT:
            raw = post_data.getlist(name) if hasattr(post_data, "getlist") else post_data.get(name) or []
            if isinstance(raw, str):
                raw = [raw]
            allowed = {c["value"] for c in f.normalized_choices()}
            parsed[f.code] = [v for v in raw if v in allowed]
            continue
        raw_value = (post_data.get(name) or "").strip()
        if not raw_value:
            parsed[f.code] = ""
            continue
        if f.field_type == TaskCategoryFieldType.NUMBER:
            try:
                parsed[f.code] = float(raw_value) if "." in raw_value else int(raw_value)
            except (TypeError, ValueError):
                parsed[f.code] = raw_value
        elif f.field_type == TaskCategoryFieldType.SELECT:
            allowed = {c["value"] for c in f.normalized_choices()}
            parsed[f.code] = raw_value if raw_value in allowed else ""
        elif f.field_type == TaskCategoryFieldType.USER:
            try:
                uid = int(raw_value)
                if User.objects.filter(pk=uid).exists():
                    parsed[f.code] = uid
                    refs_to_create.append(TaskExtraRef(task=task, field_code=f.code, user_id=uid))
                else:
                    parsed[f.code] = ""
            except (TypeError, ValueError):
                parsed[f.code] = ""
        elif f.field_type == TaskCategoryFieldType.ASSET:
            try:
                aid = int(raw_value)
                from assets.models import Asset
                if Asset.objects.filter(pk=aid).exists():
                    parsed[f.code] = aid
                    refs_to_create.append(TaskExtraRef(task=task, field_code=f.code, asset_id=aid))
                else:
                    parsed[f.code] = ""
            except (TypeError, ValueError, ImportError):
                parsed[f.code] = ""
        else:
            parsed[f.code] = raw_value

    task.extra_data = parsed
    task.save(update_fields=["extra_data"])
    TaskExtraRef.objects.filter(task=task).delete()
    if refs_to_create:
        TaskExtraRef.objects.bulk_create(refs_to_create)


def _handle_tasks_categories_post(request):
    """POST tab tipi: CRUD TaskCategory + TaskCategoryField."""
    action = (request.POST.get("cat_action") or "").strip()
    base_url = f"{reverse('tasks:impostazioni')}?tab=tipi"

    def _url_for(category=None):
        if category is not None:
            return f"{base_url}&cat={category.id}"
        return base_url

    def _unique_slug(base: str, exclude_pk: int | None = None) -> str:
        candidate = slugify(base) or "tipo"
        candidate = candidate[:120]
        n = 2
        root = candidate
        while TaskCategory.objects.exclude(pk=exclude_pk or 0).filter(slug=candidate).exists():
            candidate = f"{root}-{n}"[:120]
            n += 1
        return candidate

    valid_field_types = {value for value, _ in TaskCategoryFieldType.choices}
    valid_role_codes = {role.code for role in _task_role_definitions(include_inactive=False)}

    def _clean_category_role() -> str:
        raw_role = (request.POST.get("role_type") or "").strip()
        return raw_role if raw_role in valid_role_codes else ""

    if action == "category_create":
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Indica un nome per il tipo attivita.")
            return redirect(_url_for())
        description = (request.POST.get("description") or "").strip()
        icon = (request.POST.get("icon") or "").strip()
        max_order = TaskCategory.objects.aggregate(m=Max("order_index")).get("m") or 0
        category = TaskCategory.objects.create(
            name=name[:120],
            slug=_unique_slug(name),
            description=description[:255],
            icon=icon[:80],
            role_type=_clean_category_role(),
            order_index=max_order + 10,
            is_machine_work=bool(request.POST.get("is_machine_work")),
        )
        log_action(request, "tipo_creato", "tasks", {
            "message": f"Creato tipo attivita '{category.name}'",
            "category_id": category.id,
        })
        messages.success(request, f"Tipo '{category.name}' creato.")
        return redirect(_url_for(category))

    if action == "category_update":
        try:
            category = TaskCategory.objects.get(pk=int(request.POST.get("category_id") or 0))
        except (TaskCategory.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Tipo attivita non trovato.")
            return redirect(_url_for())
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Il nome del tipo e' obbligatorio.")
            return redirect(_url_for(category))
        category.name = name[:120]
        category.description = (request.POST.get("description") or "").strip()[:255]
        category.icon = (request.POST.get("icon") or "").strip()[:80]
        category.role_type = _clean_category_role()
        category.is_active = bool(request.POST.get("is_active"))
        category.is_machine_work = bool(request.POST.get("is_machine_work"))
        try:
            category.order_index = max(0, int(request.POST.get("order_index") or 0))
        except (TypeError, ValueError):
            pass
        category.save()
        log_action(request, "tipo_modificato", "tasks", {
            "message": f"Modificato tipo attivita '{category.name}'",
            "category_id": category.id,
        })
        messages.success(request, "Tipo aggiornato.")
        return redirect(_url_for(category))

    if action == "category_delete":
        try:
            category = TaskCategory.objects.get(pk=int(request.POST.get("category_id") or 0))
        except (TaskCategory.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Tipo attivita non trovato.")
            return redirect(_url_for())
        name = category.name
        in_use = Task.objects.filter(category_id=category.id).count()
        category.delete()
        log_action(request, "tipo_eliminato", "tasks", {
            "message": f"Eliminato tipo attivita '{name}' (attivita collegate: {in_use})",
        })
        messages.success(request, f"Tipo '{name}' eliminato ({in_use} attivita scollegate).")
        return redirect(_url_for())

    if action in {"field_create", "field_update", "field_delete"}:
        try:
            category = TaskCategory.objects.get(pk=int(request.POST.get("category_id") or 0))
        except (TaskCategory.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Tipo attivita non trovato.")
            return redirect(_url_for())

        if action == "field_delete":
            try:
                field = TaskCategoryField.objects.get(pk=int(request.POST.get("field_id") or 0), category=category)
            except (TaskCategoryField.DoesNotExist, TypeError, ValueError):
                messages.error(request, "Campo non trovato.")
                return redirect(_url_for(category))
            label = field.label
            field.delete()
            log_action(request, "tipo_campo_eliminato", "tasks", {
                "message": f"Eliminato campo '{label}' dal tipo '{category.name}'",
                "category_id": category.id,
            })
            messages.success(request, f"Campo '{label}' eliminato.")
            return redirect(_url_for(category))

        label = (request.POST.get("label") or "").strip()
        code = (request.POST.get("code") or "").strip()
        field_type = (request.POST.get("field_type") or "").strip()
        if field_type not in valid_field_types:
            messages.error(request, "Tipo di campo non valido.")
            return redirect(_url_for(category))
        if not label:
            messages.error(request, "La label del campo e' obbligatoria.")
            return redirect(_url_for(category))
        if not code:
            code = slugify(label)[:60] or "campo"
        else:
            code = slugify(code)[:60] or "campo"
        required = bool(request.POST.get("required"))
        help_text = (request.POST.get("help_text") or "").strip()[:255]
        placeholder = (request.POST.get("placeholder") or "").strip()[:120]
        depends_on_code = slugify(request.POST.get("depends_on_code") or "")[:60]
        depends_on_value = (request.POST.get("depends_on_value") or "").strip()[:255]
        try:
            order_index = max(0, int(request.POST.get("order_index") or 0))
        except (TypeError, ValueError):
            order_index = 0

        asset_type_filter = ",".join(
            [v.strip() for v in request.POST.getlist("asset_type_filter_multi") if v.strip()]
        )[:255]
        asset_category_filter_raw = request.POST.getlist("asset_category_filter_multi")
        asset_category_filter_ids: list[str] = []
        for v in asset_category_filter_raw:
            v = (v or "").strip()
            if not v:
                continue
            try:
                asset_category_filter_ids.append(str(int(v)))
            except (TypeError, ValueError):
                continue
        asset_category_filter = ",".join(asset_category_filter_ids)[:255]

        choices_raw = (request.POST.get("choices_raw") or "").strip()
        choices_json: list[dict[str, str]] = []
        if field_type in {TaskCategoryFieldType.SELECT, TaskCategoryFieldType.MULTISELECT}:
            for line in choices_raw.splitlines():
                line = line.strip()
                if not line:
                    continue
                if "|" in line:
                    value, label_choice = line.split("|", 1)
                    value = value.strip()
                    label_choice = label_choice.strip() or value
                else:
                    value = line
                    label_choice = line
                if value:
                    choices_json.append({"value": value[:120], "label": label_choice[:180]})

        if action == "field_create":
            # unique code per category
            base_code = code
            n = 2
            while TaskCategoryField.objects.filter(category=category, code=code).exists():
                code = f"{base_code}-{n}"[:60]
                n += 1
            max_order = TaskCategoryField.objects.filter(category=category).aggregate(
                m=Max("order_index")
            ).get("m") or 0
            TaskCategoryField.objects.create(
                category=category,
                code=code,
                label=label[:180],
                field_type=field_type,
                required=required,
                order_index=order_index or (max_order + 10),
                help_text=help_text,
                placeholder=placeholder,
                choices_json=choices_json,
                depends_on_code=depends_on_code,
                depends_on_value=depends_on_value,
                asset_type_filter=asset_type_filter if field_type == TaskCategoryFieldType.ASSET else "",
                asset_category_filter=asset_category_filter if field_type == TaskCategoryFieldType.ASSET else "",
            )
            log_action(request, "tipo_campo_creato", "tasks", {
                "message": f"Aggiunto campo '{label}' al tipo '{category.name}'",
                "category_id": category.id,
            })
            messages.success(request, f"Campo '{label}' aggiunto.")
            return redirect(_url_for(category))

        # field_update
        try:
            field = TaskCategoryField.objects.get(pk=int(request.POST.get("field_id") or 0), category=category)
        except (TaskCategoryField.DoesNotExist, TypeError, ValueError):
            messages.error(request, "Campo non trovato.")
            return redirect(_url_for(category))
        # permetti cambio code solo se non confligge
        if code != field.code:
            if TaskCategoryField.objects.filter(category=category, code=code).exclude(pk=field.pk).exists():
                messages.error(request, f"Codice '{code}' gia in uso.")
                return redirect(_url_for(category))
            field.code = code
        field.label = label[:180]
        field.field_type = field_type
        field.required = required
        field.help_text = help_text
        field.placeholder = placeholder
        field.depends_on_code = depends_on_code
        field.depends_on_value = depends_on_value
        field.choices_json = choices_json
        if field_type == TaskCategoryFieldType.ASSET:
            field.asset_type_filter = asset_type_filter
            field.asset_category_filter = asset_category_filter
        else:
            field.asset_type_filter = ""
            field.asset_category_filter = ""
        if order_index:
            field.order_index = order_index
        field.save()
        log_action(request, "tipo_campo_modificato", "tasks", {
            "message": f"Modificato campo '{label}' del tipo '{category.name}'",
            "category_id": category.id,
        })
        messages.success(request, f"Campo '{label}' aggiornato.")
        return redirect(_url_for(category))

    messages.error(request, "Azione non riconosciuta.")
    return redirect(base_url)


@legacy_admin_required
def impostazioni(request):
    """Pagina canonica impostazioni/admin del modulo Task."""
    cfg = TaskImpostazioni.get_singleton()
    config_url = f"{reverse('tasks:impostazioni')}?tab=config"

    if request.method == "POST":
        posted_tab = (request.POST.get("tab") or "").strip().lower()
        if posted_tab == "ruoli":
            return _handle_tasks_roles_post(request)
        if posted_tab == "accessi":
            return _handle_tasks_access_post(request)
        if posted_tab == "promemoria":
            return _handle_tasks_reminders_post(request)
        if posted_tab == "tipi":
            return _handle_tasks_categories_post(request)

        branding_response = handle_module_branding_post(
            request,
            module_key="tasks",
            redirect_to=config_url,
            audit_module="tasks",
            fallback_label="KICK-OFF",
        )
        if branding_response is not None:
            return branding_response
        cfg.responsabile_email = request.POST.get("responsabile_email", "").strip()
        cfg.notifiche_scadenza_attive = bool(request.POST.get("notifiche_scadenza_attive"))
        cfg.giorni_preavviso = _coerce_positive_int(request.POST.get("giorni_preavviso"), default=3)
        cfg.note_generali = request.POST.get("note_generali", "").strip()
        cfg.vrf_reminder_days = _coerce_positive_int(request.POST.get("vrf_reminder_days"), default=7)
        cfg.vrf_blocking_days = _coerce_positive_int(request.POST.get("vrf_blocking_days"), default=30)
        cfg.asset_conflict_check_enabled = bool(request.POST.get("asset_conflict_check_enabled"))
        cfg.asset_conflict_block = bool(request.POST.get("asset_conflict_block"))
        cfg.save()
        log_action(request, "modifica", "tasks", {"message": "Aggiornate impostazioni KICK-OFF"})
        messages.success(request, "Impostazioni salvate.")
        return redirect(config_url)

    tab = _normalize_tasks_settings_tab(request.GET.get("tab"), default="config")
    return render(
        request,
        "tasks/impostazioni.html",
        {
            **_tasks_shell_context(request, active="impostazioni"),
            **_build_tasks_settings_context(request, tab=tab),
            "cfg": cfg,
            **get_module_branding_context("tasks", fallback_label="KICK-OFF"),
        },
    )


# ---------------------------------------------------------------------------
# VRF Document upload / management
# ---------------------------------------------------------------------------

@task_permissions_required("tasks_view")
def project_vrf_upload(request, project_id: int):
    """Upload documento VRF per un progetto: GET mostra form, POST gestisce upload/confirm/later/not_required."""
    project_qs = _scoped_projects_queryset(request)
    project = get_object_or_404(project_qs, pk=project_id)
    cfg = TaskImpostazioni.get_singleton()
    vrf_detail = _vrf_status_detail(project, cfg)

    next_url = request.GET.get("next") or request.POST.get("next") or reverse(
        "tasks:project_gantt", kwargs={"project_id": project_id}
    )

    # Sanitize next_url to only allow relative paths
    if next_url and not next_url.startswith("/"):
        next_url = reverse("tasks:project_gantt", kwargs={"project_id": project_id})

    session_key = f"tasks_vrf_preview_{project_id}"

    if request.method == "GET":
        request.session.pop(session_key, None)
        return render(request, "tasks/project_vrf_upload.html", {
            **_tasks_shell_context(request, active="projects", project=project),
            "page_title": f"Documento VRF - {project.name}",
            "project": project,
            "vrf_detail": vrf_detail,
            "preview": None,
            "next_url": next_url,
        })

    if not _can_manage_project(request, project):
        return render(
            request,
            "core/pages/forbidden.html",
            {"page_title": "Accesso negato"},
            status=403,
        )

    action = request.POST.get("action", "")

    # --- action: upload → parse e mostra preview ---
    if action == "upload":
        uploaded = request.FILES.get("vrf_file")
        if not uploaded:
            messages.error(request, "Nessun file selezionato.")
            return redirect(request.get_full_path())

        try:
            validate_extension_and_mime(
                uploaded,
                allowed_extensions={".xlsx", ".xls"},
                allowed_mimes={
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                    "application/octet-stream",
                    "application/zip",  # xlsx e' uno zip
                    "application/x-zip-compressed",
                },
                max_bytes=20 * 1024 * 1024,
                label=safe_filename(uploaded.name) or "Documento VRF",
                allow_empty=False,
            )
        except UploadMimeValidationError as exc:
            messages.error(request, str(exc))
            return redirect(request.get_full_path())

        try:
            file_bytes = uploaded.read()
            extracted = _parse_vrf_excel(file_bytes)
        except Exception:
            messages.error(request, "Impossibile leggere il file Excel. Verifica che sia un MOD.073 valido.")
            return redirect(request.get_full_path())

        request.session[session_key] = {
            "filename": uploaded.name,
            "file_bytes_b64": __import__("base64").b64encode(file_bytes).decode(),
            "extracted": extracted,
        }

        return render(request, "tasks/project_vrf_upload.html", {
            **_tasks_shell_context(request, active="projects", project=project),
            "page_title": f"Documento VRF - {project.name}",
            "project": project,
            "vrf_detail": vrf_detail,
            "preview": extracted,
            "filename": uploaded.name,
            "next_url": next_url,
        })

    # --- action: confirm → salva file e aggiorna progetto ---
    if action == "confirm":
        session_data = request.session.pop(session_key, None)
        if not session_data:
            messages.error(request, "Sessione scaduta. Ricarica il file.")
            return redirect(request.get_full_path())

        import base64, io
        from django.core.files.base import ContentFile
        from django.utils import timezone as tz

        extracted = session_data["extracted"]
        filename = session_data["filename"]
        file_bytes = base64.b64decode(session_data["file_bytes_b64"])

        project.vrf_file.save(filename, ContentFile(file_bytes), save=False)
        project.vrf_original_name = filename
        project.vrf_uploaded_at = tz.now()
        project.vrf_status = VRFDocStatus.UPLOADED

        if extracted.get("client_name"):
            project.client_name = extracted["client_name"]
        if extracted.get("part_number"):
            project.part_number = extracted["part_number"]
        if extracted.get("versione"):
            project.versione = extracted["versione"]
        if extracted.get("vrf_quote_number"):
            project.vrf_quote_number = extracted["vrf_quote_number"]
        if extracted.get("vrf_description"):
            project.vrf_description = extracted["vrf_description"]
        if extracted.get("vrf_esp"):
            project.vrf_esp = extracted["vrf_esp"]

        project.save()
        log_action(
            request,
            "upload_vrf",
            "tasks",
            {
                "message": f"Caricato documento VRF '{filename}' per kickoff #{project.id} — {project.name}",
                "project_id": project.id,
                "filename": filename,
            },
        )
        messages.success(request, f"Documento VRF caricato correttamente. Dati del kickoff aggiornati da '{filename}'.")
        return redirect(next_url)

    # --- action: discard → torna al form upload ---
    if action == "discard":
        request.session.pop(session_key, None)
        return redirect(request.get_full_path())

    # --- action: later → resta PENDING ---
    if action == "later":
        request.session.pop(session_key, None)
        log_action(
            request,
            "vrf_remind_skipped",
            "tasks",
            {
                "message": f"VRF rimandato per kickoff #{project.id} — {project.name}",
                "project_id": project.id,
            },
        )
        messages.info(request, "Documento VRF rimandato. Ricordati di caricarlo al piu presto.")
        return redirect(next_url)

    # --- action: not_required ---
    if action == "not_required":
        request.session.pop(session_key, None)
        project.vrf_status = VRFDocStatus.NOT_REQUIRED
        project.save(update_fields=["vrf_status"])
        log_action(
            request,
            "vrf_not_required",
            "tasks",
            {
                "message": f"VRF marcato come 'non richiesto' per kickoff #{project.id} — {project.name}",
                "project_id": project.id,
            },
        )
        messages.success(request, "Kickoff marcato come 'VRF non richiesto'.")
        return redirect(next_url)

    return redirect(request.get_full_path())


# ---------------------------------------------------------------------------
# VRF compilazione online (matrice rischi MOD.073 Rev.10)
# ---------------------------------------------------------------------------

@task_permissions_required("tasks_view")
def project_vrf_compile(request, project_id: int):
    """Compila la matrice rischi VRF online e genera l'xlsx a conferma.

    GET: mostra la matrice (con eventuale bozza salvata su VRFRiskAssessment).
    POST action=save_draft: aggiorna la valutazione senza generare file.
    POST action=confirm: genera xlsx, salva su project.vrf_file, marca UPLOADED.
    """
    from . import vrf_catalog
    from .vrf_generator import build_vrf_xlsx, parse_vrf_xlsx, vrf_filename_for

    project_qs = _scoped_projects_queryset(request)
    project = get_object_or_404(project_qs, pk=project_id)
    cfg = TaskImpostazioni.get_singleton()
    vrf_detail = _vrf_status_detail(project, cfg)

    if vrf_detail.get("is_blocked"):
        messages.error(request, "Compilazione VRF bloccata: periodo di blocco superato.")
        return redirect("tasks:project_gantt", project_id=project_id)

    assessment = getattr(project, "vrf_assessment", None)
    prefilled_from_xlsx = False
    if assessment is None:
        initial_data = vrf_catalog.default_scores()
        if request.method == "GET" and project.vrf_file and project.vrf_file.name:
            try:
                with project.vrf_file.open("rb") as fh:
                    initial_data = parse_vrf_xlsx(fh)
                prefilled_from_xlsx = True
            except Exception:
                initial_data = vrf_catalog.default_scores()
                prefilled_from_xlsx = False
        assessment = VRFRiskAssessment(project=project, data=initial_data)

    if request.method == "POST":
        if not _can_manage_project(request, project):
            return render(
                request,
                "core/pages/forbidden.html",
                {"page_title": "Accesso negato"},
                status=403,
            )
        action = request.POST.get("action", "")
        data = vrf_catalog.default_scores()

        for risk in vrf_catalog.RISKS:
            r_code = risk["code"]
            for ph in vrf_catalog.PHASES:
                k_raw = request.POST.get(f"k_{r_code}_{ph['key']}", "").strip()
                try:
                    k_val = int(k_raw)
                    k_val = max(vrf_catalog.K_RANGE[0], min(vrf_catalog.K_RANGE[1], k_val))
                except (TypeError, ValueError):
                    k_val = risk["k_default"][ph["key"]]
                data["risks"][r_code]["k"][ph["key"]] = k_val

                for sub in risk["sub_parameters"]:
                    raw = request.POST.get(f"score_{r_code}_{sub['code']}_{ph['key']}", "").strip()
                    if raw == "" or raw is None:
                        data["risks"][r_code]["subs"][sub["code"]][ph["key"]] = None
                    else:
                        try:
                            v = int(raw)
                            v = max(vrf_catalog.R_RANGE[0], min(vrf_catalog.R_RANGE[1], v))
                            data["risks"][r_code]["subs"][sub["code"]][ph["key"]] = v
                        except (TypeError, ValueError):
                            data["risks"][r_code]["subs"][sub["code"]][ph["key"]] = None

        assessment.data = data
        assessment.recompute_totals()
        assessment.compiled_by = request.user if getattr(request.user, "is_authenticated", False) else None
        assessment.compiled_at = timezone.now()
        assessment.save()

        if action == "save_draft":
            log_action(request, "vrf_compiled_draft", "tasks", {
                "project_id": project.id,
                "message": f"Bozza VRF online salvata per kickoff #{project.id}",
            })
            messages.success(request, "Bozza VRF salvata. I totali sono aggiornati.")
            return redirect("tasks:project_vrf_compile", project_id=project_id)

        if action == "skip_with_reminder":
            log_action(request, "vrf_skipped_with_reminder", "tasks", {
                "project_id": project.id,
                "message": f"VRF rimandato per kickoff #{project.id}, bozza salvata, stato PENDING",
            })
            messages.info(
                request,
                "Bozza VRF salvata. Il sistema ti ricordera' di completare la scheda VRF "
                "secondo le impostazioni del modulo (promemoria e blocco progressivo).",
            )
            return redirect("tasks:project_gantt", project_id=project_id)

        if action == "confirm":
            from django.core.files.base import ContentFile
            try:
                xlsx_bytes = build_vrf_xlsx(project, assessment)
            except FileNotFoundError:
                messages.error(request, "Template VRF non disponibile sul server. Contattare l'amministratore.")
                return redirect("tasks:project_vrf_compile", project_id=project_id)

            filename = vrf_filename_for(project)
            project.vrf_file.save(filename, ContentFile(xlsx_bytes), save=False)
            project.vrf_original_name = filename
            project.vrf_uploaded_at = timezone.now()
            project.vrf_status = VRFDocStatus.UPLOADED
            project.save()

            log_action(request, "vrf_compiled_inline", "tasks", {
                "project_id": project.id,
                "filename": filename,
                "total_p": assessment.total_p,
                "total_i": assessment.total_i,
                "total_c": assessment.total_c,
                "dig_triggered": assessment.dig_triggered,
                "message": f"VRF compilato online per kickoff #{project.id} - {project.name}",
            })
            if assessment.dig_triggered:
                messages.warning(request, "VRF salvato. Attenzione: TR >= 46 in almeno una fase - richiesto coinvolgimento DIG.")
            else:
                messages.success(request, "VRF compilato e salvato. Kickoff sbloccato.")
            return redirect("tasks:project_gantt", project_id=project_id)

        messages.error(request, "Azione non riconosciuta.")
        return redirect("tasks:project_vrf_compile", project_id=project_id)

    totals = vrf_catalog.compute_totals(assessment.data or {})
    risks_data = (assessment.data or {}).get("risks") or {}
    enriched_risks = []
    for risk in vrf_catalog.RISKS:
        r_data = risks_data.get(risk["code"]) or {}
        k_vals = r_data.get("k") or risk["k_default"]
        subs_data = r_data.get("subs") or {}
        enriched_subs = []
        for sub in risk["sub_parameters"]:
            sub_scores = subs_data.get(sub["code"]) or {}
            enriched_subs.append({
                "code": sub["code"],
                "row": sub["row"],
                "label": sub["label"],
                "score_p": "" if sub_scores.get("p") is None else str(sub_scores.get("p")),
                "score_i": "" if sub_scores.get("i") is None else str(sub_scores.get("i")),
                "score_c": "" if sub_scores.get("c") is None else str(sub_scores.get("c")),
            })
        enriched_risks.append({
            "code": risk["code"],
            "row": risk["row"],
            "title": risk["title"],
            "k_p": k_vals.get("p", risk["k_default"]["p"]),
            "k_i": k_vals.get("i", risk["k_default"]["i"]),
            "k_c": k_vals.get("c", risk["k_default"]["c"]),
            "sub_parameters": enriched_subs,
        })

    context = {
        **_tasks_shell_context(request, active="projects", project=project),
        "page_title": f"Compila VRF - {project.name}",
        "project": project,
        "vrf_detail": vrf_detail,
        "assessment": assessment,
        "prefilled_from_xlsx": prefilled_from_xlsx,
        "risks": enriched_risks,
        "phases": vrf_catalog.PHASES,
        "totals": totals,
        "dig_threshold": vrf_catalog.DIG_THRESHOLD,
        "k_min": vrf_catalog.K_RANGE[0],
        "k_max": vrf_catalog.K_RANGE[1],
        "r_min": vrf_catalog.R_RANGE[0],
        "r_max": vrf_catalog.R_RANGE[1],
    }
    return render(request, "tasks/project_vrf_compile.html", context)


@task_permissions_required("tasks_view")
def project_vrf_download(request, project_id: int):
    """Scarica l'xlsx corrente (da vrf_file) oppure rigenera dal template+assessment."""
    from django.http import FileResponse, HttpResponse
    from .vrf_generator import build_vrf_xlsx, vrf_filename_for

    project_qs = _scoped_projects_queryset(request)
    project = get_object_or_404(project_qs, pk=project_id)

    if project.vrf_file and project.vrf_file.name:
        try:
            return FileResponse(
                project.vrf_file.open("rb"),
                as_attachment=True,
                filename=project.vrf_original_name or vrf_filename_for(project),
            )
        except FileNotFoundError:
            pass

    assessment = getattr(project, "vrf_assessment", None)
    try:
        xlsx_bytes = build_vrf_xlsx(project, assessment)
    except FileNotFoundError:
        return HttpResponse("Template VRF non disponibile sul server.", status=500)
    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{vrf_filename_for(project)}"'
    return response


# ---------------------------------------------------------------------------
# Excel import / template download
# ---------------------------------------------------------------------------

_IMPORT_COLUMNS = [
    "Kickoff / Riferimento",
    "Cliente",
    "P/N",
    "Revisione",
    "Versione",
    "PM (username)",
    "Titolo attivita",
    "Descrizione attivita",
    "Stato",       # TODO / IN_PROGRESS / DONE / CANCELED
    "Priorita",    # LOW / MEDIUM / HIGH
    "Assegnato a (username)",
    "Data inizio (YYYY-MM-DD)",
    "Data fine (YYYY-MM-DD)",
    "Tag",
]


@task_permissions_required("tasks_create")
def download_excel_template(request):
    """Genera e scarica il template .xlsx per l'import attivita kickoff."""
    import io
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kickoff Import"

    # Header row
    for col_idx, col_name in enumerate(_IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Example row
    example = [
        "Rif. kickoff Alfa",
        "Cliente SRL",
        "PN-001",
        "A",
        "1.0",
        "",
        "Titolo attivita di esempio",
        "Descrizione operativa",
        "TODO",
        "MEDIUM",
        "",
        "",
        "",
        "",
    ]
    for col_idx, val in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Istruzioni foglio
    ws_info = wb.create_sheet("Istruzioni")
    ws_info["A1"] = "Istruzioni import attivita kickoff"
    ws_info["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    instructions = [
        ("Kickoff / Riferimento", "Riferimento libero per raggruppare piu righe nello stesso kickoff quando il P/N non e disponibile."),
        ("Cliente", "Ragione sociale cliente (testo libero)."),
        ("P/N", "Part Number (testo libero)."),
        ("Revisione", "Revisione del P/N (testo libero, es. A, B, 01)."),
        ("Versione", "Versione del P/N (testo libero, es. 1.0, 2.3)."),
        ("PM (username)", "Username del project manager. Deve esistere nel sistema."),
        ("Titolo attivita", "Titolo dell'attivita kickoff (obbligatorio)."),
        ("Descrizione attivita", "Descrizione operativa (opzionale)."),
        ("Stato", "TODO / IN_PROGRESS / DONE / CANCELED. Default: TODO."),
        ("Priorita", "LOW / MEDIUM / HIGH. Default: MEDIUM."),
        ("Assegnato a (username)", "Username dell'operatore. Deve esistere nel sistema."),
        ("Data inizio (YYYY-MM-DD)", "Data inizio (next_step_due). Formato: 2025-03-15."),
        ("Data fine (YYYY-MM-DD)", "Data fine prevista (due_date). Formato: 2025-04-30."),
        ("Tag", "Etichette separate da virgola."),
    ]
    ws_info["A3"] = "Campo"
    ws_info["B3"] = "Descrizione"
    ws_info["A3"].font = openpyxl.styles.Font(bold=True)
    ws_info["B3"].font = openpyxl.styles.Font(bold=True)
    for row_idx, (field, desc) in enumerate(instructions, start=4):
        ws_info.cell(row=row_idx, column=1, value=field)
        ws_info.cell(row=row_idx, column=2, value=desc)
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="template_import_kickoff.xlsx"'
    return response


@task_permissions_required("tasks_create")
def import_excel(request):
    """Upload + anteprima + commit import attivita kickoff da Excel."""
    import io
    import openpyxl

    if request.method == "GET":
        return render(request, "tasks/import.html", {
            **_tasks_shell_context(request, active="import"),
            "page_title": "Import attivita kickoff da Excel",
            "columns": _IMPORT_COLUMNS,
        })

    # --- STEP 1: upload file e anteprima ---
    if "file" in request.FILES and "confirm" not in request.POST:
        uploaded = request.FILES["file"]
        try:
            validate_extension_and_mime(
                uploaded,
                allowed_extensions={".xlsx", ".xls"},
                allowed_mimes={
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                    "application/octet-stream",
                    "application/zip",
                    "application/x-zip-compressed",
                },
                max_bytes=20 * 1024 * 1024,
                label=safe_filename(uploaded.name) or "Import Excel",
                allow_empty=False,
            )
        except UploadMimeValidationError as exc:
            messages.error(request, str(exc))
            return redirect("tasks:import_excel")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
        except Exception:
            messages.error(request, "File non valido. Carica un file .xlsx corretto.")
            return redirect("tasks:import_excel")

        ws = wb.active
        rows_raw = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        preview_rows = []
        errors = []
        for row_idx, row in enumerate(rows_raw, start=2):
            if not any(row):
                continue
            def _cell(n):
                try:
                    return str(row[n] or "").strip()
                except IndexError:
                    return ""

            kickoff_reference = _cell(0)
            part_number = _cell(2)
            revisione = _cell(3) if part_number else ""
            versione = _cell(4) if part_number else ""
            titolo = _cell(6)
            if not titolo:
                errors.append(f"Riga {row_idx}: campo Titolo attivita vuoto — riga ignorata.")
                continue

            stato_raw = _cell(8).upper() or "TODO"
            if stato_raw not in {s.value for s in TaskStatus}:
                stato_raw = "TODO"
            priorita_raw = _cell(9).upper() or "MEDIUM"
            if priorita_raw not in {p.value for p in TaskPriority}:
                priorita_raw = "MEDIUM"

            preview_rows.append({
                "progetto": kickoff_reference,
                "cliente": _cell(1),
                "part_number": part_number,
                "revisione": revisione,
                "versione": versione,
                "pm_username": _cell(5),
                "titolo": titolo,
                "descrizione": _cell(7),
                "stato": stato_raw,
                "priorita": priorita_raw,
                "assegnato_username": _cell(10),
                "data_inizio": _cell(11),
                "data_fine": _cell(12),
                "tag": _cell(13),
            })

        request.session["tasks_import_preview"] = preview_rows
        return render(request, "tasks/import.html", {
            **_tasks_shell_context(request, active="import"),
            "page_title": "Import attivita kickoff da Excel",
            "columns": _IMPORT_COLUMNS,
            "preview_rows": preview_rows,
            "errors": errors,
            "filename": uploaded.name,
        })

    # --- STEP 2: commit ---
    if "confirm" in request.POST:
        preview_rows = request.session.pop("tasks_import_preview", [])
        if not preview_rows:
            messages.error(request, "Sessione scaduta o nessun dato da importare. Ricarica il file.")
            return redirect("tasks:import_excel")

        created_projects = 0
        updated_projects = 0
        created_tasks = 0
        updated_project_ids: set[int] = set()
        cached_projects_by_identity: dict[tuple[str, str, str], Project] = {}
        cached_projects_by_reference: dict[str, Project] = {}

        try:
            with transaction.atomic():
                for row in preview_rows:
                    # Risolvi PM
                    pm_user = None
                    if row["pm_username"]:
                        pm_user = User.objects.filter(username__iexact=row["pm_username"]).first()

                    proj_defaults = {
                        "client_name": row["cliente"],
                        "project_manager": pm_user,
                    }
                    part_number = (row["part_number"] or "").strip()
                    revisione = (row["revisione"] or "").strip()
                    versione = (row["versione"] or "").strip()
                    kickoff_reference = (row["progetto"] or "").strip()

                    project = None
                    project_was_created = False
                    if part_number:
                        identity_key = (
                            part_number.lower(),
                            revisione.lower(),
                            versione.lower(),
                        )
                        project = cached_projects_by_identity.get(identity_key)
                        if project is None:
                            project = (
                                Project.objects.filter(
                                    part_number__iexact=part_number,
                                    revisione__iexact=revisione,
                                    versione__iexact=versione,
                                )
                                .order_by("id")
                                .first()
                            )
                            if project is None:
                                project = Project.objects.create(
                                    name="",
                                    client_name=row["cliente"],
                                    project_manager=pm_user,
                                    part_number=part_number,
                                    revisione=revisione,
                                    versione=versione,
                                    created_by=request.user,
                                )
                                project_was_created = True
                                created_projects += 1
                            cached_projects_by_identity[identity_key] = project
                    elif kickoff_reference:
                        reference_key = kickoff_reference.lower()
                        project = cached_projects_by_reference.get(reference_key)
                        if project is None:
                            project = Project.objects.create(
                                name="",
                                client_name=row["cliente"],
                                project_manager=pm_user,
                                created_by=request.user,
                            )
                            project_was_created = True
                            cached_projects_by_reference[reference_key] = project
                            created_projects += 1
                    else:
                        project = Project.objects.create(
                            name="",
                            client_name=row["cliente"],
                            project_manager=pm_user,
                            created_by=request.user,
                        )
                        project_was_created = True
                        created_projects += 1

                    updated_fields = []
                    for field_name, value in proj_defaults.items():
                        if getattr(project, field_name) != value:
                            setattr(project, field_name, value)
                            updated_fields.append(field_name)
                    if updated_fields:
                        project.save(update_fields=updated_fields + ["updated_at"])
                    if project.id not in updated_project_ids and updated_fields and not project_was_created:
                        updated_project_ids.add(project.id)
                        updated_projects += 1

                    # Risolvi assegnatario
                    assigned = None
                    if row["assegnato_username"]:
                        assigned = User.objects.filter(username__iexact=row["assegnato_username"]).first()

                    # Date
                    data_inizio = _coerce_date(row["data_inizio"])
                    data_fine = _coerce_date(row["data_fine"])
                    # Sanity: data_fine deve essere > data_inizio
                    if data_inizio and data_fine and data_fine <= data_inizio:
                        data_fine = None

                    Task.objects.create(
                        title=row["titolo"],
                        description=row["descrizione"],
                        status=row["stato"],
                        priority=row["priorita"],
                        project=project,
                        assigned_to=assigned,
                        next_step_due=data_inizio,
                        due_date=data_fine,
                        tags=row["tag"],
                        created_by=request.user,
                    )
                    created_tasks += 1

        except Exception as exc:
            messages.error(request, f"Errore durante l'import: {exc}")
            return redirect("tasks:import_excel")

        log_action(
            request,
            "import_excel",
            "tasks",
            {
                "message": (
                    f"Import Excel: {created_tasks} attivita kickoff create, "
                    f"{created_projects} kickoff nuovi, {updated_projects} kickoff aggiornati."
                ),
                "created_tasks": created_tasks,
                "created_projects": created_projects,
                "updated_projects": updated_projects,
            },
        )
        messages.success(
            request,
            (
                f"Import completato: {created_tasks} attivita kickoff create "
                f"({created_projects} kickoff nuovi, {updated_projects} aggiornati)."
            ),
        )
        return redirect("tasks:list")

    return redirect("tasks:import_excel")


# ---------------------------------------------------------------------------
# Incontri kickoff
# ---------------------------------------------------------------------------

def _sync_meeting_outlook(request, meeting) -> None:
    from .meeting_outlook import sync_meeting_outlook_event
    level, msg = sync_meeting_outlook_event(request=request, meeting=meeting)
    if level == "success":
        messages.success(request, msg)
    elif level in ("warning", "error"):
        messages.warning(request, msg)
    elif level == "info":
        messages.info(request, msg)


def _parse_optional_date(value: str):
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _issue_display_owner(issue: MeetingIssue) -> str:
    if issue.assigned_to_id:
        return issue.assigned_to.get_full_name() or issue.assigned_to.username
    return ""


def _meeting_issue_agenda_item(issue: MeetingIssue) -> dict:
    custom_fields = [{"label": "Stato", "value": issue.get_status_display()}]
    owner = _issue_display_owner(issue)
    if owner:
        custom_fields.append({"label": "Responsabile", "value": owner})
    if issue.due_date:
        custom_fields.append({"label": "Scadenza", "value": issue.due_date.strftime("%d/%m/%Y")})
    if issue.source_meeting_id:
        custom_fields.append({"label": "Origine", "value": f"Incontro {issue.source_meeting.numero}"})
    return {
        "id": f"issue-{issue.pk}",
        "titolo": f"Problema aperto: {issue.title}",
        "nota": issue.description,
        "task_id": issue.linked_task_id,
        "task_label": issue.linked_task.title if issue.linked_task_id else "",
        "issue_id": issue.pk,
        "source": "meeting_issue",
        "locked": True,
        "custom_fields": custom_fields,
        "done": issue.is_resolved,
    }


def _open_meeting_issues_for_project(project: Project):
    return (
        project.meeting_issues
        .filter(status=MeetingIssueStatus.OPEN)
        .select_related("source_meeting", "assigned_to", "linked_task")
        .order_by("due_date", "created_at", "id")
    )


def _meeting_issue_ids_from_agenda(meeting: KickoffMeeting) -> list[int]:
    ids: list[int] = []
    for item in meeting.agenda_items or []:
        if not isinstance(item, dict):
            continue
        try:
            issue_id = int(item.get("issue_id") or 0)
        except (TypeError, ValueError):
            issue_id = 0
        if issue_id:
            ids.append(issue_id)
    return ids


def _set_meeting_agenda_issue_done(meeting: KickoffMeeting, issue_id: int, done: bool) -> None:
    agenda_items = meeting.agenda_items or []
    if not isinstance(agenda_items, list):
        return
    changed = False
    for item in agenda_items:
        if not isinstance(item, dict):
            continue
        try:
            item_issue_id = int(item.get("issue_id") or 0)
        except (TypeError, ValueError):
            item_issue_id = 0
        if item_issue_id == issue_id and bool(item.get("done", False)) != done:
            item["done"] = done
            changed = True
    if changed:
        meeting.agenda_items = agenda_items
        meeting.save(update_fields=["agenda_items", "updated_at"])


def _meeting_issues_for_form(project: Project, meeting: KickoffMeeting | None = None):
    qs = project.meeting_issues.select_related(
        "source_meeting",
        "resolution_meeting",
        "assigned_to",
        "linked_task",
        "resolved_by",
    )
    if meeting is None:
        return qs.filter(status=MeetingIssueStatus.OPEN).order_by("due_date", "created_at", "id")
    return qs.filter(
        Q(status=MeetingIssueStatus.OPEN)
        | Q(source_meeting=meeting)
        | Q(resolution_meeting=meeting)
        | Q(pk__in=_meeting_issue_ids_from_agenda(meeting))
    ).distinct().order_by("status", "due_date", "created_at", "id")


def _sync_meeting_issues_from_post(request, project: Project, meeting: KickoffMeeting) -> None:
    displayed_ids: set[int] = set()
    for raw in request.POST.getlist("meeting_issue_ids"):
        try:
            displayed_ids.add(int(raw))
        except (TypeError, ValueError):
            continue

    resolved_ids: set[int] = set()
    for raw in request.POST.getlist("resolved_issue_ids"):
        try:
            resolved_ids.add(int(raw))
        except (TypeError, ValueError):
            continue

    if displayed_ids:
        for issue in MeetingIssue.objects.filter(project=project, pk__in=displayed_ids):
            note = request.POST.get(f"issue_resolution_{issue.pk}", "")
            if issue.pk in resolved_ids:
                issue.mark_resolved(meeting=meeting, user=request.user, note=note)
                issue.save(update_fields=[
                    "status",
                    "resolution_meeting",
                    "resolved_by",
                    "resolved_at",
                    "resolution_note",
                    "updated_at",
                ])
            elif issue.resolution_meeting_id == meeting.pk and issue.status == MeetingIssueStatus.RESOLVED:
                issue.reopen()
                issue.save(update_fields=[
                    "status",
                    "resolution_meeting",
                    "resolved_by",
                    "resolved_at",
                    "resolution_note",
                    "updated_at",
                ])

        agenda_changed = False
        agenda_items = meeting.agenda_items or []
        if isinstance(agenda_items, list):
            for item in agenda_items:
                if not isinstance(item, dict):
                    continue
                try:
                    item_issue_id = int(item.get("issue_id") or 0)
                except (TypeError, ValueError):
                    item_issue_id = 0
                if item_issue_id in displayed_ids:
                    next_done = item_issue_id in resolved_ids
                    if bool(item.get("done", False)) != next_done:
                        item["done"] = next_done
                        agenda_changed = True
        if agenda_changed:
            meeting.agenda_items = agenda_items
            meeting.save(update_fields=["agenda_items", "updated_at"])

    titles = request.POST.getlist("new_issue_title")
    descriptions = request.POST.getlist("new_issue_description")
    assigned_to_ids = request.POST.getlist("new_issue_assigned_to")
    due_dates = request.POST.getlist("new_issue_due_date")
    linked_task_ids = request.POST.getlist("new_issue_task")

    for idx, raw_title in enumerate(titles):
        title = (raw_title or "").strip()
        if not title:
            continue
        description = (descriptions[idx] if idx < len(descriptions) else "").strip()
        assigned_to = None
        assigned_to_id = (assigned_to_ids[idx] if idx < len(assigned_to_ids) else "").strip()
        if assigned_to_id:
            try:
                assigned_to = User.objects.get(pk=int(assigned_to_id))
            except (User.DoesNotExist, ValueError):
                assigned_to = None
        linked_task = None
        linked_task_id = (linked_task_ids[idx] if idx < len(linked_task_ids) else "").strip()
        if linked_task_id:
            try:
                linked_task = project.tasks.get(pk=int(linked_task_id))
            except (Task.DoesNotExist, ValueError):
                linked_task = None
        MeetingIssue.objects.create(
            project=project,
            source_meeting=meeting,
            title=title[:220],
            description=description,
            assigned_to=assigned_to,
            due_date=_parse_optional_date(due_dates[idx] if idx < len(due_dates) else ""),
            linked_task=linked_task,
            created_by=request.user if request.user.is_authenticated else None,
        )


@task_permissions_required("tasks_view")
def project_meetings(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meetings = project.meetings.select_related("created_by").prefetch_related("partecipanti_utenti").order_by("numero")
    can_manage = _can_manage_project(request, project)
    open_issue_count = project.meeting_issues.filter(status=MeetingIssueStatus.OPEN).count()
    return render(
        request,
        "tasks/project_meetings.html",
        {
            **_tasks_shell_context(request, active="meetings", project=project),
            "page_title": f"Incontri - {project.name}",
            "project": project,
            "meetings": meetings,
            "can_manage": can_manage,
            "open_issue_count": open_issue_count,
        },
    )


@task_permissions_required("tasks_create")
def project_meeting_create(request, project_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    if not _can_manage_project(request, project):
        messages.error(request, "Non hai i permessi per aggiungere incontri a questo kickoff.")
        return redirect("tasks:project_meetings", project_id=project_id)
    meeting_issues = list(_open_meeting_issues_for_project(project))
    if request.method == "POST":
        form = KickoffMeetingForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                meeting = form.save(commit=False)
                meeting.project = project
                meeting.created_by = request.user
                meeting.save()
                form.save_m2m()
                _sync_meeting_issues_from_post(request, project, meeting)
            _sync_meeting_outlook(request, meeting)
            log_action(request, "kickoff_meeting_create", "tasks", {"meeting_id": meeting.pk, "meeting_numero": meeting.numero, "project_id": project_id})
            messages.success(request, f"Incontro {meeting.numero} aggiunto.")
            return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting.pk)
    else:
        auto_agenda = [_meeting_issue_agenda_item(issue) for issue in meeting_issues]
        form = KickoffMeetingForm(initial={
            "data": timezone.localdate(),
            "agenda_items_raw": json.dumps(auto_agenda),
        })
    project_tasks = list(project.tasks.values("id", "title", "status").order_by("title")[:100])
    meeting_rooms = list(MeetingRoom.objects.values_list("nome", flat=True))
    active_users = task_active_users_queryset()
    return render(
        request,
        "tasks/project_meeting_form.html",
        {
            **_tasks_shell_context(request, active="meetings", project=project),
            "page_title": f"Nuovo incontro - {project.name}",
            "project": project,
            "form": form,
            "is_edit": False,
            "project_tasks_json": json.dumps(project_tasks),
            "meeting_rooms_json": json.dumps(meeting_rooms),
            "meeting_issues": meeting_issues,
            "active_users": active_users,
        },
    )


@task_permissions_required("tasks_view")
def project_meeting_detail(request, project_id: int, meeting_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(
        KickoffMeeting.objects.prefetch_related("partecipanti_utenti"),
        pk=meeting_id, project=project,
    )
    can_manage = _can_manage_project(request, project)
    # Righe next_steps non vuote (per CTA "Crea task da questo step")
    next_steps_lines = [l.strip() for l in (meeting.next_steps or "").splitlines() if l.strip()]
    agenda_issue_ids = _meeting_issue_ids_from_agenda(meeting)
    meeting_issues = list(
        project.meeting_issues
        .filter(
            Q(source_meeting=meeting)
            | Q(resolution_meeting=meeting)
            | Q(pk__in=agenda_issue_ids)
        )
        .select_related("source_meeting", "resolution_meeting", "assigned_to", "linked_task", "resolved_by")
        .distinct()
        .order_by("status", "due_date", "created_at", "id")
    )
    return render(
        request,
        "tasks/project_meeting_detail.html",
        {
            **_tasks_shell_context(request, active="meetings", project=project),
            "page_title": f"Incontro {meeting.numero} - {project.name}",
            "project": project,
            "meeting": meeting,
            "can_manage": can_manage,
            "meeting_issues": meeting_issues,
            "next_steps_lines": next_steps_lines,
            "active_users": task_active_users_queryset() if can_manage else [],
            "agenda_toggle_url_base": f"/tasks/projects/{project_id}/incontri/{meeting_id}/agenda-toggle/",
            "task_from_step_url": reverse("tasks:project_meeting_task_from_step", kwargs={"project_id": project_id, "meeting_id": meeting_id}),
        },
    )


@task_permissions_required("tasks_create")
def project_meeting_edit(request, project_id: int, meeting_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(KickoffMeeting, pk=meeting_id, project=project)
    if not _can_manage_project(request, project):
        messages.error(request, "Non hai i permessi per modificare questo incontro.")
        return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting_id)
    meeting_issues = list(_meeting_issues_for_form(project, meeting))
    if request.method == "POST":
        form = KickoffMeetingForm(request.POST, instance=meeting)
        if form.is_valid():
            with transaction.atomic():
                meeting = form.save()
                _sync_meeting_issues_from_post(request, project, meeting)
            _sync_meeting_outlook(request, meeting)
            log_action(request, "kickoff_meeting_edit", "tasks", {"meeting_id": meeting.pk, "meeting_numero": meeting.numero, "project_id": project_id})
            messages.success(request, "Incontro aggiornato.")
            return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting_id)
    else:
        form = KickoffMeetingForm(instance=meeting)
    project_tasks = list(project.tasks.values("id", "title", "status").order_by("title")[:100])
    meeting_rooms = list(MeetingRoom.objects.values_list("nome", flat=True))
    active_users = task_active_users_queryset()
    return render(
        request,
        "tasks/project_meeting_form.html",
        {
            **_tasks_shell_context(request, active="meetings", project=project),
            "page_title": f"Modifica incontro {meeting.numero} - {project.name}",
            "project": project,
            "meeting": meeting,
            "form": form,
            "is_edit": True,
            "project_tasks_json": json.dumps(project_tasks),
            "meeting_rooms_json": json.dumps(meeting_rooms),
            "meeting_issues": meeting_issues,
            "active_users": active_users,
        },
    )


@require_POST
@task_permissions_required("tasks_create")
def project_meeting_delete(request, project_id: int, meeting_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(KickoffMeeting, pk=meeting_id, project=project)
    if not _can_manage_project(request, project):
        messages.error(request, "Non hai i permessi per eliminare questo incontro.")
        return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting_id)
    numero = meeting.numero
    meeting.delete()
    log_action(request, "kickoff_meeting_delete", "tasks", {"meeting_numero": numero, "project_id": project_id})
    messages.success(request, f"Incontro {numero} eliminato.")
    return redirect("tasks:project_meetings", project_id=project_id)


@require_POST
@task_permissions_required("tasks_create")
def project_meeting_issue_status(request, project_id: int, meeting_id: int, issue_id: int):
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(KickoffMeeting, pk=meeting_id, project=project)
    issue = get_object_or_404(MeetingIssue, pk=issue_id, project=project)
    if not _can_manage_project(request, project):
        messages.error(request, "Non hai i permessi per aggiornare i problemi di questo kickoff.")
        return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting_id)

    action = (request.POST.get("action") or "").strip()
    if action == "resolve":
        issue.mark_resolved(
            meeting=meeting,
            user=request.user,
            note=request.POST.get("resolution_note", ""),
        )
        issue.save(update_fields=[
            "status",
            "resolution_meeting",
            "resolved_by",
            "resolved_at",
            "resolution_note",
            "updated_at",
        ])
        _set_meeting_agenda_issue_done(meeting, issue.pk, True)
        log_action(request, "kickoff_meeting_issue_resolve", "tasks", {"issue_id": issue.pk, "meeting_id": meeting.pk, "project_id": project.pk})
        messages.success(request, "Problema segnato come risolto.")
    elif action == "reopen":
        issue.reopen()
        issue.save(update_fields=[
            "status",
            "resolution_meeting",
            "resolved_by",
            "resolved_at",
            "resolution_note",
            "updated_at",
        ])
        _set_meeting_agenda_issue_done(meeting, issue.pk, False)
        log_action(request, "kickoff_meeting_issue_reopen", "tasks", {"issue_id": issue.pk, "meeting_id": meeting.pk, "project_id": project.pk})
        messages.info(request, "Problema riaperto e riportato nei prossimi incontri.")
    else:
        messages.error(request, "Azione non valida.")
    return redirect("tasks:project_meeting_detail", project_id=project_id, meeting_id=meeting_id)


@require_POST
@task_permissions_required("tasks_create")
def project_meeting_agenda_toggle(request, project_id: int, meeting_id: int, item_id: str):
    """Toggle done/undone su un singolo punto agenda dell'incontro."""
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(KickoffMeeting, pk=meeting_id, project=project)
    if not _can_manage_project(request, project):
        return JsonResponse({"ok": False, "reason": "forbidden"}, status=403)

    new_done = None
    items = meeting.agenda_items or []
    for item in items:
        if str(item.get("id", "")) == item_id:
            item["done"] = not bool(item.get("done", False))
            new_done = item["done"]
            break

    if new_done is None:
        return JsonResponse({"ok": False, "reason": "item_not_found"}, status=404)

    meeting.agenda_items = items
    meeting.save(update_fields=["agenda_items"])
    return JsonResponse({"ok": True, "done": new_done, "item_id": item_id})


@require_POST
@task_permissions_required("tasks_create")
def project_meeting_task_from_step(request, project_id: int, meeting_id: int):
    """Crea un task kickoff a partire da un next step dell'incontro."""
    project = get_object_or_404(_scoped_projects_queryset(request), pk=project_id)
    meeting = get_object_or_404(KickoffMeeting, pk=meeting_id, project=project)
    if not _can_manage_project(request, project):
        return JsonResponse({"ok": False, "reason": "forbidden"}, status=403)

    title = (request.POST.get("title") or "").strip()
    if not title:
        return JsonResponse({"ok": False, "reason": "title_required"}, status=400)

    # Assegnatario (opzionale)
    assigned_to = None
    assigned_to_id = (request.POST.get("assigned_to") or "").strip()
    if assigned_to_id:
        try:
            assigned_to = User.objects.get(pk=int(assigned_to_id))
        except (User.DoesNotExist, ValueError):
            pass

    # Scadenza (opzionale)
    due_date = None
    due_date_raw = (request.POST.get("due_date") or "").strip()
    if due_date_raw:
        try:
            due_date = date.fromisoformat(due_date_raw)
        except ValueError:
            pass

    # Priorità
    priority_raw = (request.POST.get("priority") or "MEDIUM").strip().upper()
    if priority_raw not in (TaskPriority.LOW, TaskPriority.MEDIUM, TaskPriority.HIGH):
        priority_raw = TaskPriority.MEDIUM

    task = Task(
        title=title,
        project=project,
        assigned_to=assigned_to,
        due_date=due_date,
        priority=priority_raw,
        status=TaskStatus.TODO,
        created_by=request.user,
        description=f"Dall'incontro #{meeting.numero} ({meeting.data:%d/%m/%Y})",
    )
    task.save()

    log_action(
        request, "kickoff_task_from_step", "tasks",
        {"task_id": task.pk, "title": title, "meeting_id": meeting.pk, "meeting_numero": meeting.numero, "project_id": project_id},
    )

    task_url = reverse("tasks:detail", kwargs={"task_id": task.pk})
    return JsonResponse({"ok": True, "task_id": task.pk, "task_url": task_url, "title": title})
