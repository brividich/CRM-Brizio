"""Generatore xlsx MOD.073 VRF Rev.10 dalla compilazione online.

Apre il template distribuito con il codice e scrive i valori compilati sulle
celle canoniche definite in vrf_catalog. Le formule originali del template
(medie, max, K x R, totali) restano intatte: Excel le ricalcola alla prima
apertura. La sorgente di verita' per i totali mostrati in portale resta
pero' VRFRiskAssessment.total_* (cache server-side).

API:
    build_vrf_xlsx(project, assessment) -> bytes
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.excel_export import write_cell_at

from . import vrf_catalog

_TEMPLATE_PATH = Path(__file__).parent / "vrf_template" / vrf_catalog.TEMPLATE_FILENAME


def _coerce_header_value(key: str, value: Any) -> Any:
    if value is None:
        return ""
    return value


def build_vrf_xlsx(project, assessment) -> bytes:
    """Ritorna i bytes di un .xlsx compilato per il progetto.

    project: tasks.models.Project (legge vrf_quote_number, part_number, ...)
    assessment: tasks.models.VRFRiskAssessment | None — se None scrive solo header.
    """
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template VRF non trovato: {_TEMPLATE_PATH}")

    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[vrf_catalog.TEMPLATE_SHEET] if vrf_catalog.TEMPLATE_SHEET in wb.sheetnames else wb.active

    header_values = {
        "vrf_quote_number": getattr(project, "vrf_quote_number", "") or "",
        "versione":         getattr(project, "versione", "") or "",
        "part_number":      getattr(project, "part_number", "") or "",
        "vrf_description":  getattr(project, "vrf_description", "") or "",
        "vrf_esp":          getattr(project, "vrf_esp", "") or "",
        "client_name":      getattr(project, "client_name", "") or "",
    }
    # Header: testo libero dal progetto (P/N, descrizione, cliente...). Passa da
    # core.excel_export.write_cell_at: `ws[cell] = "=..."` lo scriverebbe come
    # formula viva nel file scaricato (formula injection). I K/R sotto restano
    # int/None, quindi non sono iniettabili e conservano le formule del template.
    for key, cell in vrf_catalog.HEADER_CELLS.items():
        write_cell_at(ws, cell, _coerce_header_value(key, header_values.get(key)))

    data = (assessment.data if assessment else None) or {}
    risks_data = data.get("risks") or {}

    for risk in vrf_catalog.RISKS:
        r_code = risk["code"]
        r_info = risks_data.get(r_code) or {}
        k_vals = r_info.get("k") or risk["k_default"]
        subs = r_info.get("subs") or {}
        for ph in vrf_catalog.PHASES:
            ph_key = ph["key"]
            k_cell = f"{ph['k_col']}{risk['row']}"
            try:
                ws[k_cell] = int(k_vals.get(ph_key, risk["k_default"][ph_key]))
            except (TypeError, ValueError):
                ws[k_cell] = risk["k_default"][ph_key]
            for sub in risk["sub_parameters"]:
                v = (subs.get(sub["code"]) or {}).get(ph_key)
                cell = f"{ph['input_col']}{sub['row']}"
                if v is None or v == "":
                    ws[cell] = None
                else:
                    try:
                        ws[cell] = int(v)
                    except (TypeError, ValueError):
                        ws[cell] = None

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def parse_vrf_xlsx(file_obj) -> dict[str, Any]:
    """Legge un xlsx VRF (caricato o generato) e ritorna una struttura 'data'
    compatibile con VRFRiskAssessment.data / vrf_catalog.default_scores().

    Usata per pre-popolare la compilazione online a partire dal file gia'
    caricato dal kickoff. Valori mancanti / non convertibili restano None
    (default per i sub-parametri; il default del catalogo per i K)."""
    wb = load_workbook(file_obj, data_only=True)
    ws = wb[vrf_catalog.TEMPLATE_SHEET] if vrf_catalog.TEMPLATE_SHEET in wb.sheetnames else wb.active

    data = vrf_catalog.default_scores()
    for risk in vrf_catalog.RISKS:
        r_code = risk["code"]
        for ph in vrf_catalog.PHASES:
            ph_key = ph["key"]
            raw_k = ws[f"{ph['k_col']}{risk['row']}"].value
            try:
                k_val = int(raw_k)
                k_val = max(vrf_catalog.K_RANGE[0], min(vrf_catalog.K_RANGE[1], k_val))
                data["risks"][r_code]["k"][ph_key] = k_val
            except (TypeError, ValueError):
                pass
            for sub in risk["sub_parameters"]:
                raw = ws[f"{ph['input_col']}{sub['row']}"].value
                if raw is None or raw == "":
                    continue
                try:
                    v = int(raw)
                    v = max(vrf_catalog.R_RANGE[0], min(vrf_catalog.R_RANGE[1], v))
                    data["risks"][r_code]["subs"][sub["code"]][ph_key] = v
                except (TypeError, ValueError):
                    continue
    wb.close()
    return data


def vrf_filename_for(project) -> str:
    parts = []
    if getattr(project, "kickoff_number", None):
        parts.append(f"KICKOFF-{project.kickoff_number}")
    if getattr(project, "part_number", ""):
        parts.append(str(project.part_number).replace("/", "-").replace(" ", "_"))
    parts.append("VRF_MOD073_Rev10")
    return "_".join(parts) + ".xlsx"
