from __future__ import annotations

import ast
from datetime import datetime, timedelta
import io
from pathlib import Path
import shutil
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from core.legacy_cache import bump_legacy_cache_version
from core.legacy_models import Permesso
from core.models import Notifica, Profile, UserOnboarding
from attrezzature.models import (
    Attrezzatura,
    AttrezzaturaKickoffLink,
    AttrezzaturaStato,
    AttrezzaturaTask as GestioneAttrezzaturaTask,
    AttrezzaturaTaskOrigine as GestioneAttrezzaturaTaskOrigine,
    AttrezzaturaTaskStato as GestioneAttrezzaturaTaskStato,
    AttrezzaturaTaskTipo as GestioneAttrezzaturaTaskTipo,
)

from .models import (
    KickoffMeeting,
    MeetingIssue,
    MeetingIssueStatus,
    Project,
    ProjectComment,
    SubTask,
    TaskAccessLevel,
    Task,
    TaskAttachment,
    TaskCalendarEvent,
    TaskCategory,
    TaskComment,
    TaskEvent,
    TaskEventType,
    TaskImpostazioni,
    TaskPriority,
    TaskRoleAssignment,
    TaskRoleAccessRule,
    TaskRoleDefinition,
    TaskRoleType,
    TaskStatus,
    TaskUserAccessRule,
    VRFDocStatus,
)
from .forms import ProjectKickoffForm
from .views import _task_date_absence_conflicts

User = get_user_model()

TASK_ACTIONS = ("tasks_view", "tasks_create", "tasks_edit", "tasks_comment", "tasks_admin")


class TaskViewsImportHygieneTests(TestCase):
    def test_views_do_not_shadow_module_imports_with_local_imports(self):
        path = Path(__file__).with_name("views.py")
        tree = ast.parse(path.read_text(encoding="utf-8"))
        module_imports: set[str] = set()
        for node in tree.body:
            if isinstance(node, ast.Import):
                module_imports.update(alias.asname or alias.name.split(".")[0] for alias in node.names)
            elif isinstance(node, ast.ImportFrom):
                module_imports.update(alias.asname or alias.name for alias in node.names if alias.name != "*")

        offenders: list[str] = []
        for fn in (node for node in ast.walk(tree) if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))):
            for node in ast.walk(fn):
                imported_names: list[str] = []
                if isinstance(node, ast.Import):
                    imported_names = [alias.asname or alias.name.split(".")[0] for alias in node.names]
                elif isinstance(node, ast.ImportFrom):
                    imported_names = [alias.asname or alias.name for alias in node.names if alias.name != "*"]
                for name in imported_names:
                    if name in module_imports:
                        offenders.append(f"{fn.name}:{node.lineno}:{name}")

        self.assertEqual(offenders, [])


def _ensure_legacy_acl_tables() -> None:
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS ruoli (
                    id INTEGER PRIMARY KEY,
                    nome VARCHAR(100) NOT NULL
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS utenti (
                    id INTEGER PRIMARY KEY,
                    nome VARCHAR(200) NOT NULL,
                    email VARCHAR(200) NULL,
                    password VARCHAR(500) NOT NULL,
                    ruolo VARCHAR(100) NULL,
                    attivo INTEGER NOT NULL DEFAULT 1,
                    deve_cambiare_password INTEGER NOT NULL DEFAULT 0,
                    ruolo_id INTEGER NULL
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS permessi (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    modulo VARCHAR(100) NOT NULL,
                    azione VARCHAR(100) NOT NULL,
                    ruolo_id INTEGER NOT NULL,
                    consentito INTEGER NULL,
                    can_view INTEGER NULL,
                    can_edit INTEGER NULL,
                    can_delete INTEGER NULL,
                    can_approve INTEGER NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('ruoli', 'U') IS NULL
                CREATE TABLE ruoli (
                    id INT NOT NULL PRIMARY KEY,
                    nome NVARCHAR(100) NOT NULL
                )
                """
            )
            cursor.execute(
                """
                IF OBJECT_ID('utenti', 'U') IS NULL
                CREATE TABLE utenti (
                    id INT NOT NULL PRIMARY KEY,
                    nome NVARCHAR(200) NOT NULL,
                    email NVARCHAR(200) NULL,
                    password NVARCHAR(500) NOT NULL,
                    ruolo NVARCHAR(100) NULL,
                    attivo BIT NOT NULL DEFAULT 1,
                    deve_cambiare_password BIT NOT NULL DEFAULT 0,
                    ruolo_id INT NULL
                )
                """
            )
            cursor.execute(
                """
                IF OBJECT_ID('permessi', 'U') IS NULL
                CREATE TABLE permessi (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    modulo NVARCHAR(100) NOT NULL,
                    azione NVARCHAR(100) NOT NULL,
                    ruolo_id INT NOT NULL,
                    consentito INT NULL,
                    can_view INT NULL,
                    can_edit INT NULL,
                    can_delete INT NULL,
                    can_approve INT NULL
                )
                """
            )


def _clear_legacy_acl_tables() -> None:
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM permessi")
        cursor.execute("DELETE FROM utenti")
        cursor.execute("DELETE FROM ruoli")


def _legacy_table_has_identity(table_name: str) -> bool:
    if connection.vendor == "sqlite":
        return False
    with connection.cursor() as cursor:
        cursor.execute(
            f"SELECT COLUMNPROPERTY(OBJECT_ID('{table_name}'),'id','IsIdentity')"
        )
        row = cursor.fetchone()
    return bool(row and row[0])


def _legacy_upsert_by_id(table_name: str, record_id: int, values: dict[str, object]) -> None:
    assignments = ", ".join(f"{column} = %s" for column in values)
    insert_columns = ["id", *values.keys()]
    insert_placeholders = ", ".join(["%s"] * len(insert_columns))
    update_params = [*values.values(), record_id]
    insert_params = [record_id, *values.values()]

    with connection.cursor() as cursor:
        cursor.execute(f"UPDATE {table_name} SET {assignments} WHERE id = %s", update_params)
        if cursor.rowcount and cursor.rowcount > 0:
            return

        if _legacy_table_has_identity(table_name):
            cursor.execute(f"SET IDENTITY_INSERT {table_name} ON")
            try:
                cursor.execute(
                    f"INSERT INTO {table_name} ({', '.join(insert_columns)}) VALUES ({insert_placeholders})",
                    insert_params,
                )
            finally:
                cursor.execute(f"SET IDENTITY_INSERT {table_name} OFF")
            return

        cursor.execute(
            f"INSERT INTO {table_name} ({', '.join(insert_columns)}) VALUES ({insert_placeholders})",
            insert_params,
        )


def _ensure_assenze_table() -> None:
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS assenze (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    copia_nome VARCHAR(200) NULL,
                    email_esterna VARCHAR(200) NULL,
                    data_inizio DATETIME NULL,
                    data_fine DATETIME NULL,
                    tipo_assenza VARCHAR(100) NULL,
                    moderation_status INTEGER NULL,
                    consenso VARCHAR(100) NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('assenze', 'U') IS NULL
                CREATE TABLE assenze (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    copia_nome NVARCHAR(200) NULL,
                    email_esterna NVARCHAR(200) NULL,
                    data_inizio DATETIME NULL,
                    data_fine DATETIME NULL,
                    tipo_assenza NVARCHAR(100) NULL,
                    moderation_status INT NULL,
                    consenso NVARCHAR(100) NULL
                )
                """
            )


def _clear_assenze_table() -> None:
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM assenze")


def _ensure_role(role_id: int, name: str) -> None:
    _legacy_upsert_by_id("ruoli", role_id, {"nome": name})


def _grant_role_actions(role_id: int, actions: list[str]) -> None:
    for action in actions:
        Permesso.objects.update_or_create(
            ruolo_id=role_id,
            modulo="tasks",
            azione=action,
            defaults={
                "can_view": 1,
                "consentito": 1,
                "can_edit": 1,
                "can_delete": 1,
                "can_approve": 1,
            },
        )


def _create_user_with_legacy(*, username: str, legacy_user_id: int, role_id: int, role_name: str):
    user = User.objects.create_user(username=username, password="pass12345")
    UserOnboarding.objects.create(user=user, completed=True, completed_at=timezone.now())
    Profile.objects.create(
        user=user,
        legacy_user_id=legacy_user_id,
        legacy_ruolo_id=role_id,
        legacy_ruolo=role_name,
    )
    _legacy_upsert_by_id(
        "utenti",
        legacy_user_id,
        {
            "nome": username,
            "email": f"{username}@example.local",
            "password": "x",
            "ruolo": role_name,
            "attivo": True,
            "deve_cambiare_password": False,
            "ruolo_id": role_id,
        },
    )
    return user


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TasksBaseTestCase(TestCase):
    def setUp(self):
        super().setUp()
        _ensure_legacy_acl_tables()
        _clear_legacy_acl_tables()
        cache.clear()

    def _refresh_acl_cache(self):
        cache.clear()
        bump_legacy_cache_version()


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class MeetingIssueWorkflowTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "tasks")
        _grant_role_actions(2, ["tasks_view", "tasks_create"])
        self._refresh_acl_cache()
        self.user = _create_user_with_legacy(
            username="meeting-owner",
            legacy_user_id=201,
            role_id=2,
            role_name="tasks",
        )
        self.project = Project.objects.create(name="Kickoff test", created_by=self.user)

    def test_new_meeting_creates_managed_issue_and_carries_it_to_next_agenda(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:project_meeting_create", args=[self.project.id]),
            {
                "titolo": "Incontro 1",
                "data": "2026-04-29",
                "ora": "09:00",
                "luogo": "Sala test",
                "agenda_items_raw": "[]",
                "new_issue_title": ["Materiale mancante"],
                "new_issue_description": ["Serve conferma dal fornitore"],
                "new_issue_assigned_to": [str(self.user.id)],
                "new_issue_due_date": ["2026-05-02"],
                "new_issue_task": [""],
            },
        )

        self.assertEqual(response.status_code, 302)
        issue = MeetingIssue.objects.get(project=self.project)
        self.assertEqual(issue.status, MeetingIssueStatus.OPEN)
        self.assertEqual(issue.source_meeting.numero, 1)
        self.assertEqual(issue.assigned_to, self.user)

        response = self.client.get(reverse("tasks:project_meeting_create", args=[self.project.id]))
        self.assertContains(response, "Materiale mancante")
        self.assertContains(response, "Problema aperto")

    def test_meeting_form_can_resolve_existing_issue_and_status_endpoint_can_reopen(self):
        self.client.force_login(self.user)
        meeting_1 = KickoffMeeting.objects.create(
            project=self.project,
            titolo="Incontro 1",
            data=timezone.localdate(),
            created_by=self.user,
        )
        issue = MeetingIssue.objects.create(
            project=self.project,
            source_meeting=meeting_1,
            title="Quote non allineate",
            created_by=self.user,
        )
        meeting_2 = KickoffMeeting.objects.create(
            project=self.project,
            titolo="Incontro 2",
            data=timezone.localdate(),
            created_by=self.user,
        )

        response = self.client.post(
            reverse("tasks:project_meeting_edit", args=[self.project.id, meeting_2.id]),
            {
                "titolo": meeting_2.titolo,
                "data": meeting_2.data.isoformat(),
                "ora": "",
                "luogo": "",
                "agenda_items_raw": "[]",
                "meeting_issue_ids": [str(issue.id)],
                "resolved_issue_ids": [str(issue.id)],
                f"issue_resolution_{issue.id}": "Allineate in riunione",
            },
        )

        self.assertEqual(response.status_code, 302)
        issue.refresh_from_db()
        self.assertEqual(issue.status, MeetingIssueStatus.RESOLVED)
        self.assertEqual(issue.resolution_meeting, meeting_2)
        self.assertEqual(issue.resolution_note, "Allineate in riunione")

        detail_response = self.client.get(reverse("tasks:project_meeting_detail", args=[self.project.id, meeting_2.id]))
        self.assertContains(detail_response, "Quote non allineate")
        self.assertContains(detail_response, "Allineate in riunione")

        response = self.client.post(
            reverse("tasks:project_meeting_issue_status", args=[self.project.id, meeting_2.id, issue.id]),
            {"action": "reopen"},
        )

        self.assertEqual(response.status_code, 302)
        issue.refresh_from_db()
        self.assertEqual(issue.status, MeetingIssueStatus.OPEN)
        self.assertIsNone(issue.resolution_meeting)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskAdminSettingsTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(1, "admin")
        self._refresh_acl_cache()
        self.admin_user = _create_user_with_legacy(
            username="taskadmin",
            legacy_user_id=9001,
            role_id=1,
            role_name="admin",
        )

    def test_legacy_gestione_route_redirects_to_settings_tab(self):
        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("tasks:gestione_admin"))
        self.assertRedirects(response, f"{reverse('tasks:impostazioni')}?tab=riepilogo")

    def test_settings_page_shows_and_saves_all_editable_fields(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("tasks:impostazioni"), {"tab": "config"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tutte le impostazioni modificabili del modulo sono in questa tab")
        self.assertContains(response, 'name="responsabile_email"', html=False)
        self.assertContains(response, 'name="notifiche_scadenza_attive"', html=False)
        self.assertContains(response, 'name="giorni_preavviso"', html=False)
        self.assertContains(response, 'name="note_generali"', html=False)
        self.assertContains(response, 'name="vrf_reminder_days"', html=False)
        self.assertContains(response, 'name="vrf_blocking_days"', html=False)
        self.assertContains(response, 'name="branding_display_label"', html=False)

        post_response = self.client.post(
            f"{reverse('tasks:impostazioni')}?tab=config",
            {
                "responsabile_email": "kickoff@example.com",
                "notifiche_scadenza_attive": "on",
                "giorni_preavviso": "5",
                "note_generali": "Note test admin",
                "vrf_reminder_days": "11",
                "vrf_blocking_days": "40",
            },
        )
        self.assertRedirects(post_response, f"{reverse('tasks:impostazioni')}?tab=config")

        cfg = TaskImpostazioni.get_singleton()
        self.assertEqual(cfg.responsabile_email, "kickoff@example.com")
        self.assertTrue(cfg.notifiche_scadenza_attive)
        self.assertEqual(cfg.giorni_preavviso, 5)
        self.assertEqual(cfg.note_generali, "Note test admin")
        self.assertEqual(cfg.vrf_reminder_days, 11)
        self.assertEqual(cfg.vrf_blocking_days, 40)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskPermissionsScopeTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _ensure_role(3, "manager")
        _ensure_role(4, "ospite")

        _grant_role_actions(2, ["tasks_view"])
        _grant_role_actions(3, ["tasks_view", "tasks_admin"])
        self._refresh_acl_cache()

        self.owner = _create_user_with_legacy(username="owner", legacy_user_id=1001, role_id=2, role_name="utente")
        self.assignee = _create_user_with_legacy(
            username="assignee", legacy_user_id=1002, role_id=2, role_name="utente"
        )
        self.subscriber = _create_user_with_legacy(
            username="subscriber", legacy_user_id=1003, role_id=2, role_name="utente"
        )
        self.outsider = _create_user_with_legacy(
            username="outsider", legacy_user_id=1004, role_id=2, role_name="utente"
        )
        self.scope_admin = _create_user_with_legacy(
            username="scopeadmin", legacy_user_id=1005, role_id=3, role_name="manager"
        )
        self.blocked = _create_user_with_legacy(
            username="blocked", legacy_user_id=1006, role_id=4, role_name="ospite"
        )

        self.task_created = Task.objects.create(
            title="Created by owner",
            created_by=self.owner,
            assigned_to=self.outsider,
        )
        self.task_assigned = Task.objects.create(
            title="Assigned to owner",
            created_by=self.assignee,
            assigned_to=self.owner,
        )
        self.task_subscribed = Task.objects.create(
            title="Subscribed by owner",
            created_by=self.assignee,
            assigned_to=self.outsider,
        )
        self.task_subscribed.subscribers.add(self.owner)
        self.task_other = Task.objects.create(
            title="Other task",
            created_by=self.assignee,
            assigned_to=self.outsider,
        )

    def test_non_admin_scope_sees_only_related_tasks(self):
        self.client.force_login(self.owner)
        response = self.client.get(reverse("tasks:list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.task_created.title)
        self.assertContains(response, self.task_assigned.title)
        self.assertContains(response, self.task_subscribed.title)
        self.assertNotContains(response, self.task_other.title)

    def test_tasks_admin_scope_can_see_all(self):
        self.client.force_login(self.scope_admin)
        response = self.client.get(reverse("tasks:list"), {"mine": "0"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.task_created.title)
        self.assertContains(response, self.task_assigned.title)
        self.assertContains(response, self.task_subscribed.title)
        self.assertContains(response, self.task_other.title)

    def test_user_without_tasks_view_gets_forbidden(self):
        self.client.force_login(self.blocked)
        response = self.client.get(reverse("tasks:list"))
        self.assertEqual(response.status_code, 403)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskAntiIDORTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_edit", "tasks_comment"])
        self._refresh_acl_cache()

        self.user_a = _create_user_with_legacy(username="idora", legacy_user_id=2001, role_id=2, role_name="utente")
        self.user_b = _create_user_with_legacy(username="idorb", legacy_user_id=2002, role_id=2, role_name="utente")
        self.task_a = Task.objects.create(title="Task A", created_by=self.user_a)
        self.task_b = Task.objects.create(title="Task B", created_by=self.user_b)

    def test_detail_out_of_scope_returns_404(self):
        self.client.force_login(self.user_a)
        response = self.client.get(reverse("tasks:detail", args=[self.task_b.id]))
        self.assertEqual(response.status_code, 404)

    def test_edit_out_of_scope_returns_404(self):
        self.client.force_login(self.user_a)
        response = self.client.get(reverse("tasks:edit", args=[self.task_b.id]))
        self.assertEqual(response.status_code, 404)

    def test_status_change_out_of_scope_returns_404(self):
        self.client.force_login(self.user_a)
        response = self.client.post(
            reverse("tasks:change_status", args=[self.task_b.id]),
            {"status": TaskStatus.DONE},
        )
        self.assertEqual(response.status_code, 404)

    def test_attachment_upload_out_of_scope_returns_404(self):
        self.client.force_login(self.user_a)
        response = self.client.post(
            reverse("tasks:add_attachment", args=[self.task_b.id]),
            {"attach_to": "task"},
        )
        self.assertEqual(response.status_code, 404)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskAttrezzaturaEmbeddedPanelTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit", "tasks_comment"])
        self._refresh_acl_cache()
        self.user = _create_user_with_legacy(
            username="attpanel",
            legacy_user_id=2101,
            role_id=2,
            role_name="utente",
        )
        self.project = Project.objects.create(
            name="Kickoff panel",
            created_by=self.user,
            part_number=" pn-77 ",
        )
        self.task = Task.objects.create(
            title="Verifica attrezzatura",
            description="Controllare attrezzatura per produzione",
            created_by=self.user,
            assigned_to=self.user,
            project=self.project,
        )

    def test_detail_embeds_attrezzatura_panel_for_project_part_number(self):
        Attrezzatura.objects.create(
            codice="AT-77",
            part_number="PN-77",
            descrizione="Staffa prova",
            stato=AttrezzaturaStato.IN_CORSO,
            created_by=self.user,
        )
        GestioneAttrezzaturaTask.objects.create(
            part_number="PN-77",
            tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
            titolo="Verificare disponibilita PN-77",
            origine=GestioneAttrezzaturaTaskOrigine.KICKOFF,
            external_kickoff_id=str(self.project.id),
            external_kickoff_activity_id=str(self.task.id),
            created_by=self.user,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:detail", args=[self.task.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Attrezzatura collegata - P/N PN-77")
        self.assertContains(response, "AT-77")
        self.assertContains(response, "Verificare disponibilita PN-77")
        self.assertContains(response, reverse("tasks:attrezzature_action", args=[self.task.id]))

    def test_attrezzatura_action_creates_kickoff_availability_task(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("tasks:attrezzature_action", args=[self.task.id]),
            {"action": "create_availability_task"},
        )

        self.assertRedirects(response, reverse("tasks:detail", args=[self.task.id]))
        linked = GestioneAttrezzaturaTask.objects.get(
            tipo=GestioneAttrezzaturaTaskTipo.VERIFICA_DISPONIBILITA,
            external_kickoff_activity_id=str(self.task.id),
        )
        self.assertEqual(linked.part_number, "PN-77")
        self.assertEqual(linked.external_kickoff_id, str(self.project.id))
        self.assertEqual(linked.origine, GestioneAttrezzaturaTaskOrigine.KICKOFF)

    def test_create_task_can_request_new_tooling_workflow(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:create"),
            {
                "task_scope": "project",
                "project_link_mode": "existing",
                "project_choice": str(self.project.id),
                "title": "Creare attrezzo P/N 77",
                "description": "Serve attrezzo dedicato",
                "status": TaskStatus.TODO,
                "priority": TaskPriority.MEDIUM,
                "tooling_mode": "request_new",
                "tooling_code": "AT-REQ",
                "tooling_description": "Da creare per produzione",
                "reminder_portal_enabled_field": "on",
            },
        )
        task = Task.objects.get(title="Creare attrezzo P/N 77")
        self.assertRedirects(response, reverse("tasks:detail", args=[task.id]))
        tool = Attrezzatura.objects.get(codice="AT-REQ")
        self.assertEqual(tool.part_number, "PN-77")
        self.assertTrue(AttrezzaturaKickoffLink.objects.filter(attrezzatura=tool, task=task).exists())

    def test_create_task_can_link_existing_tooling(self):
        tool = Attrezzatura.objects.create(codice="AT-LINK", part_number="PN-77")
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:create"),
            {
                "task_scope": "project",
                "project_link_mode": "existing",
                "project_choice": str(self.project.id),
                "title": "Verificare attrezzo esistente",
                "status": TaskStatus.TODO,
                "priority": TaskPriority.MEDIUM,
                "tooling_mode": "link_existing",
                "tooling_existing_attrezzatura": str(tool.id),
                "reminder_portal_enabled_field": "on",
            },
        )
        task = Task.objects.get(title="Verificare attrezzo esistente")
        self.assertRedirects(response, reverse("tasks:detail", args=[task.id]))
        self.assertTrue(AttrezzaturaKickoffLink.objects.filter(attrezzatura=tool, task=task).exists())

    def test_attrezzatura_action_creates_draft_tool_from_kickoff_context(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("tasks:attrezzature_action", args=[self.task.id]),
            {
                "action": "create_draft_tool",
                "codice": "AT-DRAFT",
                "description": "Bozza da activity",
            },
        )

        self.assertRedirects(response, reverse("tasks:detail", args=[self.task.id]))
        tool = Attrezzatura.objects.get(codice="AT-DRAFT")
        self.assertEqual(tool.part_number, "PN-77")
        self.assertEqual(tool.descrizione, "Bozza da activity")
        self.assertEqual(tool.origine_import, "kickoff")

    def test_attrezzatura_action_updates_progress_and_completes_task(self):
        tool = Attrezzatura.objects.create(
            codice="AT-UPD",
            part_number="PN-77",
            stato=AttrezzaturaStato.IN_CORSO,
            created_by=self.user,
        )
        linked = GestioneAttrezzaturaTask.objects.create(
            attrezzatura=tool,
            part_number="PN-77",
            tipo=GestioneAttrezzaturaTaskTipo.AGGIORNA_AVANZAMENTO,
            titolo="Aggiornare AT-UPD",
            origine=GestioneAttrezzaturaTaskOrigine.KICKOFF,
            external_kickoff_id=str(self.project.id),
            external_kickoff_activity_id=str(self.task.id),
            created_by=self.user,
        )
        self.client.force_login(self.user)

        progress_response = self.client.post(
            reverse("tasks:attrezzature_action", args=[self.task.id]),
            {
                "action": "update_progress",
                "attrezzatura_id": str(tool.id),
                "percentuale": "80",
                "stato": AttrezzaturaStato.IN_ATTESA,
                "note": "Aggiornato da dettaglio activity",
            },
        )
        self.assertRedirects(progress_response, reverse("tasks:detail", args=[self.task.id]))
        tool.refresh_from_db()
        self.assertEqual(tool.avanzamento_percentuale, 80)
        self.assertEqual(tool.stato, AttrezzaturaStato.IN_ATTESA)
        self.assertEqual(tool.avanzamenti.count(), 1)

        complete_response = self.client.post(
            reverse("tasks:attrezzature_action", args=[self.task.id]),
            {
                "action": "complete_attrezzatura_task",
                "attrezzatura_task_id": str(linked.id),
                "note": "Completata da KICK-OFF",
            },
        )
        self.assertRedirects(complete_response, reverse("tasks:detail", args=[self.task.id]))
        linked.refresh_from_db()
        self.assertEqual(linked.stato, GestioneAttrezzaturaTaskStato.COMPLETATA)
        self.assertEqual(linked.completed_by, self.user)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskAuditTrailTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_edit", "tasks_comment"])
        self._refresh_acl_cache()

        self.user = _create_user_with_legacy(username="audituser", legacy_user_id=3001, role_id=2, role_name="utente")
        self.task = Task.objects.create(title="Audit task", created_by=self.user, status=TaskStatus.TODO)

    def test_status_change_creates_audit_event(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:change_status", args=[self.task.id]),
            {"status": TaskStatus.IN_PROGRESS},
        )
        self.assertEqual(response.status_code, 302)
        event = TaskEvent.objects.filter(task=self.task, type=TaskEventType.STATUS_CHANGE).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.payload.get("from"), TaskStatus.TODO)
        self.assertEqual(event.payload.get("to"), TaskStatus.IN_PROGRESS)

    def test_add_comment_creates_audit_event(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:add_comment", args=[self.task.id]),
            {"body": "Commento di test"},
        )
        self.assertEqual(response.status_code, 302)
        comment = TaskComment.objects.filter(task=self.task).first()
        self.assertIsNotNone(comment)
        event = TaskEvent.objects.filter(task=self.task, type=TaskEventType.COMMENT_ADDED).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.payload.get("comment_id"), comment.id)

    def test_subtask_events_are_created(self):
        self.client.force_login(self.user)
        response_add = self.client.post(
            reverse("tasks:add_subtask", args=[self.task.id]),
            {"title": "Sub 1", "order_index": 1},
        )
        self.assertEqual(response_add.status_code, 302)
        subtask = SubTask.objects.get(task=self.task, title="Sub 1")

        response_status = self.client.post(
            reverse("tasks:edit_subtask_status", args=[self.task.id, subtask.id]),
            {"status": TaskStatus.DONE},
        )
        self.assertEqual(response_status.status_code, 302)

        add_event = TaskEvent.objects.filter(task=self.task, type=TaskEventType.SUBTASK_ADDED).first()
        status_event = TaskEvent.objects.filter(task=self.task, type=TaskEventType.SUBTASK_STATUS_CHANGE).first()
        self.assertIsNotNone(add_event)
        self.assertIsNotNone(status_event)
        self.assertEqual(add_event.payload.get("subtask_id"), subtask.id)
        self.assertEqual(status_event.payload.get("subtask_id"), subtask.id)
        self.assertEqual(status_event.payload.get("from"), TaskStatus.TODO)
        self.assertEqual(status_event.payload.get("to"), TaskStatus.DONE)

    def test_subtask_rollup_updates_parent_task_status(self):
        self.client.force_login(self.user)
        self.client.post(
            reverse("tasks:add_subtask", args=[self.task.id]),
            {"title": "Sub rollup", "order_index": 1},
        )
        subtask = SubTask.objects.get(task=self.task, title="Sub rollup")
        response = self.client.post(
            reverse("tasks:edit_subtask_status", args=[self.task.id, subtask.id]),
            {"status": TaskStatus.DONE},
        )
        self.assertEqual(response.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.status, TaskStatus.DONE)
        rollup_events = TaskEvent.objects.filter(task=self.task, type=TaskEventType.STATUS_CHANGE)
        self.assertTrue(any((event.payload or {}).get("source") == "subtask_rollup" for event in rollup_events))


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskListFiltersTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _ensure_role(3, "manager")
        _grant_role_actions(2, ["tasks_view"])
        _grant_role_actions(3, ["tasks_view", "tasks_admin"])
        self._refresh_acl_cache()

        self.user = _create_user_with_legacy(username="filteruser", legacy_user_id=4001, role_id=2, role_name="utente")
        self.manager = _create_user_with_legacy(username="manager", legacy_user_id=4002, role_id=3, role_name="manager")
        self.other = _create_user_with_legacy(username="other", legacy_user_id=4003, role_id=2, role_name="utente")
        self.project_alpha = Project.objects.create(name="Project Alpha", created_by=self.manager)
        self.project_beta = Project.objects.create(name="Project Beta", created_by=self.manager)

        today = timezone.localdate()
        self.t_overdue = Task.objects.create(
            title="Overdue TODO",
            created_by=self.other,
            assigned_to=self.user,
            project=self.project_alpha,
            status=TaskStatus.TODO,
            priority=TaskPriority.HIGH,
            due_date=today - timedelta(days=2),
            tags="produzione, urgente",
        )
        self.t_done_past = Task.objects.create(
            title="Done old",
            created_by=self.other,
            assigned_to=self.user,
            status=TaskStatus.DONE,
            priority=TaskPriority.HIGH,
            due_date=today - timedelta(days=1),
        )
        self.t_future_medium = Task.objects.create(
            title="Future medium",
            created_by=self.other,
            assigned_to=self.other,
            project=self.project_beta,
            status=TaskStatus.IN_PROGRESS,
            priority=TaskPriority.MEDIUM,
            due_date=today + timedelta(days=4),
            tags="it, inventory",
        )
        self.t_future_low = Task.objects.create(
            title="Future low",
            created_by=self.other,
            assigned_to=self.user,
            status=TaskStatus.TODO,
            priority=TaskPriority.LOW,
            due_date=today + timedelta(days=10),
            tags="planning",
        )
        self.t_unassigned = Task.objects.create(
            title="Unassigned task",
            created_by=self.other,
            status=TaskStatus.TODO,
            priority=TaskPriority.MEDIUM,
            due_date=today + timedelta(days=2),
        )
        self.t_without_due = Task.objects.create(
            title="Task without due date",
            created_by=self.other,
            assigned_to=self.user,
            project=self.project_alpha,
            status=TaskStatus.TODO,
            priority=TaskPriority.MEDIUM,
            due_date=None,
        )

    def test_filter_overdue(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("tasks:list"), {"mine": "0", "overdue": "on"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_overdue.title)
        self.assertNotContains(response, self.t_done_past.title)
        self.assertNotContains(response, self.t_future_medium.title)

    def test_filter_status_priority_and_assigned_to(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "status": TaskStatus.TODO,
                "priority": TaskPriority.LOW,
                "assigned_to": str(self.user.id),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_future_low.title)
        self.assertNotContains(response, self.t_overdue.title)
        self.assertNotContains(response, self.t_future_medium.title)

    def test_filter_due_date_range(self):
        self.client.force_login(self.manager)
        today = timezone.localdate()
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "due_from": (today + timedelta(days=1)).isoformat(),
                "due_to": (today + timedelta(days=6)).isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_future_medium.title)
        self.assertNotContains(response, self.t_future_low.title)

    def test_filter_by_tag(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "tag": "urgente",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_overdue.title)
        self.assertNotContains(response, self.t_future_medium.title)

    def test_filter_by_project(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "project": str(self.project_alpha.id),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_overdue.title)
        self.assertNotContains(response, self.t_future_medium.title)

    def test_filter_unassigned(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "unassigned": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_unassigned.title)
        self.assertNotContains(response, self.t_overdue.title)

    def test_filter_without_due_date(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "without_due_date": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_without_due.title)
        self.assertNotContains(response, self.t_future_low.title)

    def test_filter_without_project(self):
        self.client.force_login(self.manager)
        response = self.client.get(
            reverse("tasks:list"),
            {
                "mine": "0",
                "without_project": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.t_future_low.title)
        self.assertContains(response, self.t_unassigned.title)
        self.assertNotContains(response, self.t_overdue.title)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskProjectsAndAttachmentsTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit"])
        self._refresh_acl_cache()

        self.user = _create_user_with_legacy(username="projectuser", legacy_user_id=5001, role_id=2, role_name="utente")

        tmp_root = Path(__file__).resolve().parents[1] / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        self._media_root = str(tmp_root / f"tasks_test_media_{uuid4().hex}")
        Path(self._media_root).mkdir(parents=True, exist_ok=True)
        self._media_override = override_settings(MEDIA_ROOT=self._media_root)
        self._media_override.enable()
        today = timezone.localdate()
        Path(self._media_root, "tasks_vrf", today.strftime("%Y"), today.strftime("%m")).mkdir(parents=True, exist_ok=True)
        Path(self._media_root, "tasks_attachments", today.strftime("%Y"), today.strftime("%m")).mkdir(parents=True, exist_ok=True)
        self.addCleanup(self._media_override.disable)
        self.addCleanup(shutil.rmtree, self._media_root, True)

    def _base_task_payload(self, title: str) -> dict:
        return {
            "title": title,
            "description": "Descrizione test",
            "status": TaskStatus.TODO,
            "priority": TaskPriority.MEDIUM,
            "task_scope": "single",
        }

    def _make_vrf_workbook_upload(
        self,
        *,
        part_number: str,
        description: str = "Descrizione VRF",
        version: str = "1.0",
        client_name: str = "Cliente Test",
        filename: str = "kickoff-vrf.xlsx",
    ) -> SimpleUploadedFile:
        import openpyxl

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "VRF"
        worksheet["O2"] = "Q-001"
        worksheet["P2"] = version
        worksheet["B3"] = part_number
        worksheet["I3"] = description
        worksheet["P3"] = "ESP"
        worksheet["B4"] = client_name

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return SimpleUploadedFile(
            filename,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _read_vrf_cell(self, project: Project, cell_ref: str) -> str:
        import openpyxl

        project.vrf_file.open("rb")
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(project.vrf_file.read()), data_only=True)
        finally:
            project.vrf_file.close()
        try:
            worksheet = workbook["VRF"] if "VRF" in workbook.sheetnames else workbook.active
            return str(worksheet[cell_ref].value or "")
        finally:
            workbook.close()

    def test_create_single_task_without_project(self):
        self.client.force_login(self.user)
        payload = self._base_task_payload("Task singola")
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)
        task = Task.objects.get(title="Task singola")
        self.assertIsNone(task.project_id)

    def test_create_task_with_new_project(self):
        self.client.force_login(self.user)
        payload = self._base_task_payload("Task con nuovo progetto")
        payload.update(
            {
                "task_scope": "project",
                "project_new_description": "Descrizione progetto A",
            }
        )
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)

        task = Task.objects.get(title="Task con nuovo progetto")
        self.assertIsNotNone(task.project_id)
        self.assertTrue(task.project.name.startswith("KICK-OFF "))
        self.assertIsNotNone(task.project.kickoff_number)
        self.assertEqual(task.project.created_by_id, self.user.id)

    def test_create_task_with_new_project_metadata(self):
        self.client.force_login(self.user)
        project_manager = User.objects.create_user(username="pm_user", password="pass12345")
        capo_commessa = User.objects.create_user(username="capo_user", password="pass12345")
        programmatore = User.objects.create_user(username="prog_user", password="pass12345")
        similar_project = Project.objects.create(name="Commessa simile", created_by=self.user)

        payload = self._base_task_payload("Task con metadati progetto")
        payload.update(
            {
                "task_scope": "project",
                "project_new_description": "Descrizione completa",
                "project_new_client": "Cliente Alfa",
                "project_new_manager": str(project_manager.id),
                "project_new_capo_commessa": str(capo_commessa.id),
                "project_new_programmer": str(programmatore.id),
                "project_new_control_method": "Checklist e test collaudo",
                "project_new_part_number": "PN-001",
                "project_similar_choice": str(similar_project.id),
            }
        )
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)

        task = Task.objects.get(title="Task con metadati progetto")
        self.assertIsNotNone(task.project_id)
        project = task.project
        self.assertEqual(project.client_name, "Cliente Alfa")
        self.assertEqual(project.project_manager_id, project_manager.id)
        self.assertEqual(project.capo_commessa_id, capo_commessa.id)
        self.assertEqual(project.programmer_id, programmatore.id)
        self.assertEqual(project.control_method, "Checklist e test collaudo")
        self.assertEqual(project.part_number, "PN-001")
        self.assertEqual(project.similar_project_id, similar_project.id)

    def test_create_task_reuses_existing_project_with_same_part_number_identity(self):
        self.client.force_login(self.user)
        project = Project.objects.create(
            name="Commessa storica",
            client_name="Cliente Legacy",
            part_number="PN-777",
            revisione="A",
            versione="1.0",
            created_by=self.user,
        )

        payload = self._base_task_payload("Task agganciata a progetto esistente per PN")
        payload.update(
            {
                "task_scope": "project",
                "project_new_client": "Cliente Nuovo",
                "project_new_part_number": "PN-777",
                "project_new_revisione": "A",
                "project_new_versione": "1.0",
            }
        )
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)

        task = Task.objects.get(title="Task agganciata a progetto esistente per PN")
        self.assertEqual(task.project_id, project.id)
        self.assertEqual(
            Project.objects.filter(part_number="PN-777", revisione="A", versione="1.0").count(),
            1,
        )
        project.refresh_from_db()
        self.assertEqual(project.client_name, "Cliente Nuovo")

    def test_new_project_requires_part_number_when_revision_or_version_is_set(self):
        # Regression: la validazione deve restare attiva sul ramo legacy
        # `project_link_mode=new` del TaskForm (il template ora usa il flow
        # dedicato /tasks/projects/new/ ma il backend mantiene compat).
        self.client.force_login(self.user)
        payload = self._base_task_payload("Task con revisione senza pn")
        payload.update(
            {
                "task_scope": "project",
                "project_link_mode": "new",
                "project_new_revisione": "B",
            }
        )
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Compila il P/N prima di usare revisione o versione.")
        self.assertFalse(Task.objects.filter(title="Task con revisione senza pn").exists())

    def test_create_task_with_existing_project(self):
        self.client.force_login(self.user)
        project = Project.objects.create(name="Progetto Esistente", created_by=self.user)
        payload = self._base_task_payload("Task su progetto esistente")
        payload.update(
            {
                "task_scope": "project",
                "project_choice": str(project.id),
            }
        )
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)

        task = Task.objects.get(title="Task su progetto esistente")
        self.assertEqual(task.project_id, project.id)
        self.assertEqual(Project.objects.count(), 1)

    def test_project_model_auto_generates_kickoff_name_and_number(self):
        first = Project.objects.create(name="", created_by=self.user)
        second = Project.objects.create(name="", created_by=self.user)

        self.assertTrue(first.name.startswith("KICK-OFF "))
        self.assertTrue(second.name.startswith("KICK-OFF "))
        self.assertIsNotNone(first.kickoff_number)
        self.assertIsNotNone(second.kickoff_number)
        self.assertNotEqual(first.kickoff_number, second.kickoff_number)

    def test_existing_kickoff_name_is_not_renamed(self):
        legacy = Project.objects.create(
            name="Legacy Kickoff",
            kickoff_number=77,
            client_name="Cliente Legacy",
            created_by=self.user,
        )
        legacy.description = "Aggiornato"
        legacy.save()
        legacy.refresh_from_db()

        self.assertEqual(legacy.name, "Legacy Kickoff")
        self.assertEqual(legacy.kickoff_number, 77)

    def test_create_form_no_longer_shows_removed_vrf_summary_section(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:create"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Scheda VRF")
        self.assertNotContains(response, "Contenuto operativo")
        self.assertContains(response, "Dettaglio attivita")

    def test_create_form_from_kickoff_context_hides_kickoff_selection(self):
        self.client.force_login(self.user)
        project = Project.objects.create(
            name="",
            client_name="Cliente Context",
            part_number="PN-CONTEXT-01",
            created_by=self.user,
        )

        response = self.client.get(f"{reverse('tasks:create')}?project={project.id}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Attivita gia collegata al kickoff corrente")
        self.assertContains(response, project.name)
        self.assertNotContains(response, "Come nasce questa attivita")
        self.assertNotContains(response, "Aggancio e anagrafica kickoff")

    def test_create_form_from_kickoff_context_forces_task_into_that_kickoff(self):
        self.client.force_login(self.user)
        project = Project.objects.create(
            name="",
            client_name="Cliente Locked",
            part_number="PN-LOCK-01",
            created_by=self.user,
        )

        payload = self._base_task_payload("Task dal kickoff contestuale")
        payload.update(
            {
                "task_scope": "single",
                "project_link_mode": "new",
                "project_new_client": "Cliente alternativo",
            }
        )

        response = self.client.post(f"{reverse('tasks:create')}?project={project.id}", payload)

        self.assertEqual(response.status_code, 302)
        task = Task.objects.get(title="Task dal kickoff contestuale")
        self.assertEqual(task.project_id, project.id)
        self.assertEqual(Project.objects.count(), 1)

    def test_project_list_shows_copy_buttons(self):
        self.client.force_login(self.user)
        Project.objects.create(name="", created_by=self.user)

        response = self.client.get(reverse("tasks:project_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Copia VRF")
        self.assertContains(response, "Copia senza P/N")

    def test_copy_project_with_vrf_duplicates_kickoff_metadata_and_file(self):
        self.client.force_login(self.user)
        project_manager = User.objects.create_user(username="copy_pm", password="pass12345")
        source_project = Project.objects.create(
            name="",
            description="Kickoff sorgente",
            client_name="Cliente Copy",
            project_manager=project_manager,
            control_method="Controllo finale",
            part_number="PN-COPY-001",
            revisione="A",
            versione="2.0",
            vrf_status=VRFDocStatus.UPLOADED,
            vrf_original_name="source-vrf.xlsx",
            vrf_quote_number="Q-500",
            vrf_description="Scheda sorgente",
            vrf_esp="ESP-1",
            created_by=self.user,
        )
        source_project.vrf_file.save(
            "source-vrf.xlsx",
            self._make_vrf_workbook_upload(
                part_number="PN-COPY-001",
                description="Scheda sorgente",
                version="2.0",
                client_name="Cliente Copy",
                filename="source-vrf.xlsx",
            ),
            save=True,
        )
        source_project.refresh_from_db()
        source_task = Task.objects.create(title="Task sorgente", created_by=self.user, project=source_project)
        ProjectComment.objects.create(project=source_project, author=self.user, body="Commento kickoff")
        TaskComment.objects.create(task=source_task, author=self.user, body="Commento task")
        TaskAttachment.objects.create(
            project=source_project,
            uploaded_by=self.user,
            file="tasks_attachments/mock.txt",
            original_name="mock.txt",
        )

        response = self.client.post(reverse("tasks:copy_project_with_vrf", args=[source_project.id]))
        self.assertEqual(response.status_code, 302)

        copied_project = Project.objects.exclude(id=source_project.id).get()
        self.assertTrue(copied_project.name.startswith("KICK-OFF "))
        self.assertNotEqual(copied_project.kickoff_number, source_project.kickoff_number)
        self.assertEqual(copied_project.description, source_project.description)
        self.assertEqual(copied_project.client_name, source_project.client_name)
        self.assertEqual(copied_project.project_manager_id, source_project.project_manager_id)
        self.assertEqual(copied_project.control_method, source_project.control_method)
        self.assertEqual(copied_project.part_number, "PN-COPY-001")
        self.assertEqual(copied_project.revisione, "A")
        self.assertEqual(copied_project.versione, "2.0")
        self.assertEqual(copied_project.vrf_status, VRFDocStatus.UPLOADED)
        self.assertEqual(copied_project.vrf_original_name, "source-vrf.xlsx")
        self.assertEqual(self._read_vrf_cell(source_project, "B3"), "PN-COPY-001")
        self.assertEqual(self._read_vrf_cell(copied_project, "B3"), "PN-COPY-001")
        self.assertEqual(Task.objects.filter(project=copied_project).count(), 0)
        self.assertEqual(ProjectComment.objects.filter(project=copied_project).count(), 0)
        self.assertEqual(TaskAttachment.objects.filter(project=copied_project).count(), 0)

    def test_copy_project_with_vrf_without_pn_clears_project_and_excel_part_number(self):
        self.client.force_login(self.user)
        source_project = Project.objects.create(
            name="",
            client_name="Cliente No PN",
            part_number="PN-COPY-002",
            revisione="B",
            versione="3.1",
            vrf_status=VRFDocStatus.UPLOADED,
            vrf_original_name="source-no-pn.xlsx",
            created_by=self.user,
        )
        source_project.vrf_file.save(
            "source-no-pn.xlsx",
            self._make_vrf_workbook_upload(
                part_number="PN-COPY-002",
                version="3.1",
                client_name="Cliente No PN",
                filename="source-no-pn.xlsx",
            ),
            save=True,
        )
        source_project.refresh_from_db()

        response = self.client.post(reverse("tasks:copy_project_with_vrf_without_pn", args=[source_project.id]))
        self.assertEqual(response.status_code, 302)

        copied_project = Project.objects.exclude(id=source_project.id).get()
        self.assertEqual(copied_project.part_number, "")
        self.assertEqual(copied_project.revisione, "B")
        self.assertEqual(copied_project.versione, "3.1")
        self.assertEqual(self._read_vrf_cell(source_project, "B3"), "PN-COPY-002")
        self.assertEqual(self._read_vrf_cell(copied_project, "B3"), "")

    def test_assignment_conflict_alert_on_create_with_keep_priority(self):
        self.client.force_login(self.user)
        assignee = User.objects.create_user(username="planner_keep", password="pass12345")
        today = timezone.localdate()
        Task.objects.create(
            title="Impegno esistente",
            created_by=self.user,
            assigned_to=assignee,
            status=TaskStatus.IN_PROGRESS,
            priority=TaskPriority.MEDIUM,
            next_step_due=today + timedelta(days=2),
            due_date=today + timedelta(days=5),
        )

        payload = self._base_task_payload("Task nuova con conflitto")
        payload.update(
            {
                "assigned_to": str(assignee.id),
                "priority": TaskPriority.LOW,
                "next_step_due": (today + timedelta(days=3)).isoformat(),
                "due_date": (today + timedelta(days=6)).isoformat(),
                "assignment_conflict_action": "keep_priority",
            }
        )
        response = self.client.post(reverse("tasks:create"), payload, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impegni sovrapposti")
        task = Task.objects.get(title="Task nuova con conflitto")
        self.assertEqual(task.priority, TaskPriority.LOW)

    def test_assignment_conflict_can_raise_priority_to_high(self):
        self.client.force_login(self.user)
        assignee = User.objects.create_user(username="planner_raise", password="pass12345")
        today = timezone.localdate()
        Task.objects.create(
            title="Impegno esistente raise",
            created_by=self.user,
            assigned_to=assignee,
            status=TaskStatus.TODO,
            priority=TaskPriority.MEDIUM,
            next_step_due=today + timedelta(days=1),
            due_date=today + timedelta(days=4),
        )

        payload = self._base_task_payload("Task nuova con priorita auto")
        payload.update(
            {
                "assigned_to": str(assignee.id),
                "priority": TaskPriority.LOW,
                "next_step_due": (today + timedelta(days=2)).isoformat(),
                "due_date": (today + timedelta(days=5)).isoformat(),
                "assignment_conflict_action": "raise_to_high",
            }
        )
        response = self.client.post(reverse("tasks:create"), payload, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Priorita aggiornata automaticamente a High")
        task = Task.objects.get(title="Task nuova con priorita auto")
        self.assertEqual(task.priority, TaskPriority.HIGH)

    def test_upload_attachment_to_task_creates_event(self):
        self.client.force_login(self.user)
        task = Task.objects.create(title="Task allegato", created_by=self.user)
        file_obj = SimpleUploadedFile("task-note.txt", b"contenuto allegato task", content_type="text/plain")
        response = self.client.post(
            reverse("tasks:add_attachment", args=[task.id]),
            {"attach_to": "task", "file": file_obj},
        )
        self.assertEqual(response.status_code, 302)

        attachment = TaskAttachment.objects.get(task=task)
        self.assertEqual(attachment.original_name, "task-note.txt")
        event = TaskEvent.objects.filter(task=task, type=TaskEventType.ATTACHMENT_ADDED).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.payload.get("target"), "task")
        self.assertEqual(event.payload.get("attachment_id"), attachment.id)

    def test_upload_attachment_to_project_creates_event(self):
        self.client.force_login(self.user)
        project = Project.objects.create(name="Project Attach", created_by=self.user)
        task = Task.objects.create(title="Task con progetto allegato", created_by=self.user, project=project)
        file_obj = SimpleUploadedFile("project-note.txt", b"contenuto allegato progetto", content_type="text/plain")
        response = self.client.post(
            reverse("tasks:add_attachment", args=[task.id]),
            {"attach_to": "project", "file": file_obj},
        )
        self.assertEqual(response.status_code, 302)

        attachment = TaskAttachment.objects.get(project=project, task__isnull=True)
        self.assertEqual(attachment.original_name, "project-note.txt")
        event = TaskEvent.objects.filter(task=task, type=TaskEventType.ATTACHMENT_ADDED).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.payload.get("target"), "project")
        self.assertEqual(event.payload.get("project_id"), project.id)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskProjectGanttAndNotificationsTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_edit", "tasks_comment"])
        self._refresh_acl_cache()

        self.owner = _create_user_with_legacy(username="g_owner", legacy_user_id=6001, role_id=2, role_name="utente")
        self.assignee = _create_user_with_legacy(
            username="g_assignee", legacy_user_id=6002, role_id=2, role_name="utente"
        )
        self.viewer = _create_user_with_legacy(username="g_viewer", legacy_user_id=6003, role_id=2, role_name="utente")
        self.outsider = _create_user_with_legacy(
            username="g_outsider", legacy_user_id=6004, role_id=2, role_name="utente"
        )

        self.project = Project.objects.create(
            name="Gantt Project",
            created_by=self.owner,
            programmer=self.assignee,
        )
        TaskRoleAccessRule.objects.create(
            role_type=TaskRoleType.PROGRAMMER,
            access_level=TaskAccessLevel.EDIT_ASSIGNED,
        )
        self.task = Task.objects.create(
            title="Task Gantt",
            created_by=self.owner,
            assigned_to=self.assignee,
            project=self.project,
            status=TaskStatus.TODO,
            next_step_due=timezone.localdate() + timedelta(days=2),
            due_date=timezone.localdate() + timedelta(days=7),
        )
        self.task.subscribers.add(self.viewer)

    def test_project_gantt_view_in_scope(self):
        self.client.force_login(self.viewer)
        response = self.client.get(reverse("tasks:project_gantt", args=[self.project.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.project.name)
        self.assertContains(response, self.task.title)

    def test_project_gantt_out_of_scope_returns_404(self):
        self.client.force_login(self.outsider)
        response = self.client.get(reverse("tasks:project_gantt", args=[self.project.id]))
        self.assertEqual(response.status_code, 404)

    def test_project_gantt_update_allowed_for_assignee(self):
        self.client.force_login(self.assignee)
        response = self.client.post(
            reverse("tasks:project_gantt_update_task", args=[self.project.id, self.task.id]),
            {
                "task_%s-next_step_due" % self.task.id: "2026-03-10",
                "task_%s-due_date" % self.task.id: "2026-03-15",
                "task_%s-status" % self.task.id: TaskStatus.IN_PROGRESS,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.status, TaskStatus.IN_PROGRESS)
        self.assertEqual(str(self.task.next_step_due), "2026-03-10")
        self.assertEqual(str(self.task.due_date), "2026-03-15")

    def test_project_gantt_update_rejects_equal_start_end(self):
        self.client.force_login(self.assignee)
        old_next = self.task.next_step_due
        old_due = self.task.due_date
        response = self.client.post(
            reverse("tasks:project_gantt_update_task", args=[self.project.id, self.task.id]),
            {
                "task_%s-next_step_due" % self.task.id: "2026-03-10",
                "task_%s-due_date" % self.task.id: "2026-03-10",
                "task_%s-status" % self.task.id: TaskStatus.IN_PROGRESS,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.next_step_due, old_next)
        self.assertEqual(self.task.due_date, old_due)
        self.assertEqual(self.task.status, TaskStatus.TODO)

    def test_project_gantt_update_denied_for_non_assignee(self):
        self.client.force_login(self.viewer)
        response = self.client.post(
            reverse("tasks:project_gantt_update_task", args=[self.project.id, self.task.id]),
            {
                "task_%s-next_step_due" % self.task.id: "2026-03-10",
                "task_%s-due_date" % self.task.id: "2026-03-12",
                "task_%s-status" % self.task.id: TaskStatus.IN_PROGRESS,
            },
        )
        self.assertEqual(response.status_code, 403)

    def test_project_gantt_shift_days_allowed_for_assignee_and_audited(self):
        self.client.force_login(self.assignee)
        old_next = self.task.next_step_due
        old_due = self.task.due_date
        response = self.client.post(
            reverse("tasks:project_gantt_shift_task", args=[self.project.id, self.task.id]),
            {"shift_days": "3"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        self.task.refresh_from_db()
        self.assertEqual(self.task.next_step_due, old_next + timedelta(days=3))
        self.assertEqual(self.task.due_date, old_due + timedelta(days=3))
        edit_events = TaskEvent.objects.filter(task=self.task, type=TaskEventType.EDIT)
        self.assertTrue(any("due_date" in (event.payload or {}).get("changes", {}) for event in edit_events))

    def test_project_gantt_shift_days_denied_for_non_assignee(self):
        self.client.force_login(self.viewer)
        response = self.client.post(
            reverse("tasks:project_gantt_shift_task", args=[self.project.id, self.task.id]),
            {"shift_days": "2"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 403)

    def test_project_gantt_shift_days_out_of_scope_returns_404(self):
        self.client.force_login(self.outsider)
        response = self.client.post(
            reverse("tasks:project_gantt_shift_task", args=[self.project.id, self.task.id]),
            {"shift_days": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 404)

    def test_task_comment_target_user_creates_notification(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("tasks:add_comment", args=[self.task.id]),
            {
                "body": "Controlla questa task",
                "target_user": str(self.assignee.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        comment = TaskComment.objects.filter(task=self.task).first()
        self.assertIsNotNone(comment)
        self.assertEqual(comment.target_user_id, self.assignee.id)
        notification = Notifica.objects.filter(
            legacy_user_id=6002,
            tipo="generico",
            messaggio__icontains="Task Gantt",
        ).first()
        self.assertIsNotNone(notification)
        self.assertIn("Task Gantt", notification.messaggio)

    def test_project_comment_target_user_creates_notification(self):
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("tasks:add_project_comment", args=[self.project.id]),
            {
                "body": "Aggiorna la timeline progetto",
                "target_user": str(self.assignee.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        comment = ProjectComment.objects.filter(project=self.project).first()
        self.assertIsNotNone(comment)
        self.assertEqual(comment.target_user_id, self.assignee.id)
        notification = Notifica.objects.filter(
            legacy_user_id=6002,
            tipo="generico",
            messaggio__icontains=self.project.name,
        ).first()
        self.assertIsNotNone(notification)
        self.assertIn(self.project.name, notification.messaggio)

    def test_project_gantt_marks_invalid_range_cells(self):
        self.client.force_login(self.owner)
        invalid_day = timezone.localdate() + timedelta(days=12)
        Task.objects.create(
            title="Task range invalido",
            created_by=self.owner,
            assigned_to=self.assignee,
            project=self.project,
            status=TaskStatus.IN_PROGRESS,
            next_step_due=invalid_day,
            due_date=invalid_day,
        )
        response = self.client.get(reverse("tasks:project_gantt", args=[self.project.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "is-invalid-range")
        self.assertContains(response, "Range non valido")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskAbsenceConflictTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create"])
        self._refresh_acl_cache()

        _ensure_assenze_table()
        _clear_assenze_table()

        self.owner = _create_user_with_legacy(username="absence_owner", legacy_user_id=6501, role_id=2, role_name="utente")
        self.assignee = _create_user_with_legacy(
            username="absence_assignee",
            legacy_user_id=6502,
            role_id=2,
            role_name="utente",
        )

    def _insert_absence(
        self,
        *,
        person_name: str,
        person_email: str,
        date_value,
        tipo: str = "Ferie",
        moderation_status: int = 0,
    ) -> None:
        start_dt = datetime.combine(date_value, datetime.min.time())
        end_dt = datetime.combine(date_value, datetime.max.time())
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO assenze (
                    copia_nome,
                    email_esterna,
                    data_inizio,
                    data_fine,
                    tipo_assenza,
                    moderation_status,
                    consenso
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                [person_name, person_email, start_dt, end_dt, tipo, moderation_status, "Approvato"],
            )

    def test_absence_conflicts_detected_on_task_dates(self):
        target_day = timezone.localdate() + timedelta(days=3)
        self._insert_absence(
            person_name=self.assignee.username,
            person_email=self.assignee.email,
            date_value=target_day,
        )
        task = Task.objects.create(
            title="Task con assenza",
            created_by=self.owner,
            assigned_to=self.assignee,
            status=TaskStatus.TODO,
            priority=TaskPriority.MEDIUM,
            due_date=target_day,
            next_step_due=target_day,
        )
        conflicts = _task_date_absence_conflicts(task)
        self.assertIn("due_date", conflicts)
        self.assertIn("next_step_due", conflicts)

    def test_project_gantt_marks_absence_cells(self):
        self.client.force_login(self.owner)
        target_day = timezone.localdate() + timedelta(days=2)
        self._insert_absence(
            person_name=self.assignee.username,
            person_email=self.assignee.email,
            date_value=target_day,
        )
        project = Project.objects.create(name="Project assenze", created_by=self.owner)
        Task.objects.create(
            title="Task conflitto gantt",
            created_by=self.owner,
            assigned_to=self.assignee,
            project=project,
            status=TaskStatus.IN_PROGRESS,
            next_step_due=target_day - timedelta(days=1),
            due_date=target_day + timedelta(days=1),
        )
        response = self.client.get(reverse("tasks:project_gantt", args=[project.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "is-absence")
        self.assertContains(response, "absence-x")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskEditAndDueDatePermissionsTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _ensure_role(3, "manager")
        _grant_role_actions(2, ["tasks_view"])
        _grant_role_actions(3, ["tasks_view", "tasks_admin"])
        self._refresh_acl_cache()

        self.project_lead = _create_user_with_legacy(
            username="lead_user", legacy_user_id=7001, role_id=2, role_name="utente"
        )
        self.assignee = _create_user_with_legacy(
            username="op_user", legacy_user_id=7002, role_id=2, role_name="utente"
        )
        self.viewer = _create_user_with_legacy(
            username="viewer_user", legacy_user_id=7003, role_id=2, role_name="utente"
        )
        self.scope_admin = _create_user_with_legacy(
            username="admin_user", legacy_user_id=7004, role_id=3, role_name="manager"
        )

        self.project = Project.objects.create(name="Project Lead Edit", created_by=self.project_lead)
        self.task = Task.objects.create(
            title="Task permessi edit",
            created_by=self.project_lead,
            assigned_to=self.assignee,
            project=self.project,
            due_date=timezone.localdate() + timedelta(days=5),
        )
        self.task.subscribers.add(self.viewer)

    def test_project_lead_can_open_and_submit_task_edit(self):
        self.client.force_login(self.project_lead)
        response_get = self.client.get(reverse("tasks:edit", args=[self.task.id]))
        self.assertEqual(response_get.status_code, 200)

        payload = {
            "title": "Task aggiornata da lead",
            "description": "Desc aggiornata",
            "tags": "lead-edit",
            "status": TaskStatus.TODO,
            "priority": TaskPriority.HIGH,
            "due_date": (timezone.localdate() + timedelta(days=9)).isoformat(),
            "next_step_text": "Nuovo step",
            "next_step_due": (timezone.localdate() + timedelta(days=7)).isoformat(),
            "assigned_to": str(self.assignee.id),
            "subscribers": [str(self.viewer.id)],
            "task_scope": "project",
            "project_choice": str(self.project.id),
            "project_new_name": "",
            "project_new_description": "",
        }
        response_post = self.client.post(reverse("tasks:edit", args=[self.task.id]), payload)
        self.assertEqual(response_post.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.title, "Task aggiornata da lead")
        self.assertEqual(self.task.priority, TaskPriority.HIGH)

    def test_assignee_without_manage_permission_cannot_open_task_edit(self):
        self.client.force_login(self.assignee)
        response = self.client.get(reverse("tasks:edit", args=[self.task.id]))
        self.assertEqual(response.status_code, 403)

    def test_assignee_can_update_due_date(self):
        self.client.force_login(self.assignee)
        target_due = timezone.localdate() + timedelta(days=14)
        response = self.client.post(
            reverse("tasks:update_due_date", args=[self.task.id]),
            {"due_date": target_due.isoformat()},
        )
        self.assertEqual(response.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.due_date, target_due)
        edit_events = TaskEvent.objects.filter(task=self.task, type=TaskEventType.EDIT)
        self.assertTrue(any("due_date" in (event.payload or {}).get("changes", {}) for event in edit_events))

    def test_viewer_in_scope_cannot_update_due_date(self):
        self.client.force_login(self.viewer)
        response = self.client.post(
            reverse("tasks:update_due_date", args=[self.task.id]),
            {"due_date": (timezone.localdate() + timedelta(days=20)).isoformat()},
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_can_update_due_date(self):
        self.client.force_login(self.scope_admin)
        target_due = timezone.localdate() + timedelta(days=30)
        response = self.client.post(
            reverse("tasks:update_due_date", args=[self.task.id]),
            {"due_date": target_due.isoformat()},
        )
        self.assertEqual(response.status_code, 302)
        self.task.refresh_from_db()
        self.assertEqual(self.task.due_date, target_due)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskRoleAssignmentTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(1, "admin")
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create"])
        self._refresh_acl_cache()
        self.admin = _create_user_with_legacy(
            username="roleadmin", legacy_user_id=9101, role_id=1, role_name="admin",
        )
        self.alice = _create_user_with_legacy(
            username="alice", legacy_user_id=9201, role_id=2, role_name="utente",
        )
        self.alice.first_name = "Alice"; self.alice.save()
        self.bob = _create_user_with_legacy(
            username="bob", legacy_user_id=9202, role_id=2, role_name="utente",
        )
        self.bob.first_name = "Bob"; self.bob.save()

    def test_no_assignment_means_all_users_visible_in_form(self):
        from .forms import _users_for_role
        qs = _users_for_role(TaskRoleType.PROJECT_MANAGER)
        # fallback: nessuna assegnazione -> tutti gli utenti
        self.assertIn(self.alice, list(qs))
        self.assertIn(self.bob, list(qs))

    def test_with_assignment_filter_applies(self):
        from .forms import _users_for_role
        TaskRoleAssignment.objects.create(user=self.alice, role_type=TaskRoleType.PROJECT_MANAGER)
        qs = list(_users_for_role(TaskRoleType.PROJECT_MANAGER))
        self.assertIn(self.alice, qs)
        self.assertNotIn(self.bob, qs)
        # altri ruoli: ancora fallback
        qs_cc = list(_users_for_role(TaskRoleType.CAPO_COMMESSA))
        self.assertIn(self.alice, qs_cc)
        self.assertIn(self.bob, qs_cc)

    def test_post_ruoli_tab_adds_and_removes_assignments(self):
        # alice parte con PM attivo
        TaskRoleAssignment.objects.create(user=self.alice, role_type=TaskRoleType.PROJECT_MANAGER)
        self.client.force_login(self.admin)
        # Post: alice non piu' PM ma CC, bob diventa PRG
        payload = {
            "tab": "ruoli",
            "visible_user_id": [str(self.alice.id), str(self.bob.id)],
            f"role__CC__{self.alice.id}": "on",
            f"role__PRG__{self.bob.id}": "on",
        }
        response = self.client.post(reverse("tasks:impostazioni"), payload)
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            TaskRoleAssignment.objects.filter(user=self.alice, role_type=TaskRoleType.PROJECT_MANAGER).exists()
        )
        self.assertTrue(
            TaskRoleAssignment.objects.filter(user=self.alice, role_type=TaskRoleType.CAPO_COMMESSA).exists()
        )
        self.assertTrue(
            TaskRoleAssignment.objects.filter(user=self.bob, role_type=TaskRoleType.PROGRAMMER).exists()
        )

    def test_ruoli_tab_search_matches_legacy_user_name(self):
        _legacy_upsert_by_id(
            "utenti",
            9201,
            {
                "nome": "Danesi Mario",
                "email": "m.danesi@example.local",
                "password": "x",
                "ruolo": "utente",
                "attivo": True,
                "deve_cambiare_password": False,
                "ruolo_id": 2,
            },
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:impostazioni"), {"tab": "ruoli", "q_user": "Danesi"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alice")
        self.assertNotContains(response, "Bob")

    def test_ruoli_tab_search_syncs_missing_active_legacy_user(self):
        _legacy_upsert_by_id(
            "utenti",
            9299,
            {
                "nome": "Danesi Simone",
                "email": "s.danesi@example.local",
                "password": "x",
                "ruolo": "utente",
                "attivo": True,
                "deve_cambiare_password": False,
                "ruolo_id": 2,
            },
        )
        self.assertFalse(Profile.objects.filter(legacy_user_id=9299).exists())

        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:impostazioni"), {"tab": "ruoli", "q_user": "Danesi"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Danesi")
        self.assertTrue(Profile.objects.filter(legacy_user_id=9299).exists())


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskReminderAdminTabTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(1, "admin")
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit"])
        self._refresh_acl_cache()
        self.admin = _create_user_with_legacy(
            username="remadmin", legacy_user_id=9401, role_id=1, role_name="admin",
        )
        self.user = _create_user_with_legacy(
            username="remuser", legacy_user_id=9501, role_id=2, role_name="utente",
        )
        self.project = Project.objects.create(name="KO rem", created_by=self.user)
        self.task = Task.objects.create(
            title="T rem", project=self.project, assigned_to=self.user,
            due_date=timezone.localdate() + timedelta(days=5),
            created_by=self.user, reminder_portal_enabled=True,
        )

    def test_postpone_shifts_fire_at(self):
        from .models import TaskReminder
        rem = TaskReminder.objects.create(
            task=self.task, legacy_user_id=9501,
            fire_at=timezone.localdate(),
        )
        original_fire = rem.fire_at
        self.client.force_login(self.admin)
        response = self.client.post(reverse("tasks:impostazioni"), {
            "tab": "promemoria",
            "reminder_action": "postpone",
            "reminder_id": [str(rem.id)],
            "postpone_days": "5",
        })
        self.assertEqual(response.status_code, 302)
        rem.refresh_from_db()
        self.assertEqual(rem.fire_at, original_fire + timedelta(days=5))

    def test_delete_removes_selected(self):
        from .models import TaskReminder
        rem = TaskReminder.objects.create(
            task=self.task, legacy_user_id=9501, fire_at=timezone.localdate(),
        )
        self.client.force_login(self.admin)
        response = self.client.post(reverse("tasks:impostazioni"), {
            "tab": "promemoria",
            "reminder_action": "delete",
            "reminder_id": [str(rem.id)],
        })
        self.assertEqual(response.status_code, 302)
        self.assertFalse(TaskReminder.objects.filter(id=rem.id).exists())

    def test_fire_now_creates_notifica_and_marks_fired(self):
        from .models import TaskReminder
        rem = TaskReminder.objects.create(
            task=self.task, legacy_user_id=9501,
            fire_at=timezone.localdate() + timedelta(days=3),
        )
        self.client.force_login(self.admin)
        response = self.client.post(reverse("tasks:impostazioni"), {
            "tab": "promemoria",
            "reminder_action": "fire_now",
            "reminder_id": [str(rem.id)],
        })
        self.assertEqual(response.status_code, 302)
        rem.refresh_from_db()
        self.assertTrue(rem.fired)
        self.assertEqual(Notifica.objects.filter(legacy_user_id=9501).count(), 1)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskScopedAccessRulesTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(1, "admin")
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit", "tasks_comment"])
        self._refresh_acl_cache()

        self.admin = _create_user_with_legacy(username="taskaccessadmin", legacy_user_id=9601, role_id=1, role_name="admin")
        self.owner = _create_user_with_legacy(username="ko_owner", legacy_user_id=9602, role_id=2, role_name="utente")
        self.capo = _create_user_with_legacy(username="ko_capo", legacy_user_id=9603, role_id=2, role_name="utente")
        self.programmer = _create_user_with_legacy(username="ko_prog", legacy_user_id=9604, role_id=2, role_name="utente")
        self.pm_user = _create_user_with_legacy(username="ko_pm", legacy_user_id=9605, role_id=2, role_name="utente")
        self.override_user = _create_user_with_legacy(username="ko_override", legacy_user_id=9606, role_id=2, role_name="utente")
        self.outsider = _create_user_with_legacy(username="ko_outsider", legacy_user_id=9607, role_id=2, role_name="utente")

        self.project = Project.objects.create(
            name="Kickoff accessi",
            created_by=self.owner,
            project_manager=self.pm_user,
            capo_commessa=self.capo,
            programmer=self.programmer,
        )
        self.task_prog = Task.objects.create(
            title="Task programmatore",
            created_by=self.owner,
            assigned_to=self.programmer,
            project=self.project,
            next_step_due=timezone.localdate() + timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=3),
        )
        self.task_other = Task.objects.create(
            title="Task altri",
            created_by=self.owner,
            assigned_to=self.owner,
            project=self.project,
            next_step_due=timezone.localdate() + timedelta(days=2),
            due_date=timezone.localdate() + timedelta(days=5),
        )

    def test_capo_commessa_edit_all_can_manage_entire_kickoff(self):
        TaskRoleAccessRule.objects.create(
            role_type=TaskRoleType.CAPO_COMMESSA,
            access_level=TaskAccessLevel.EDIT_ALL,
        )
        self.client.force_login(self.capo)

        response_projects = self.client.get(reverse("tasks:project_list"))
        self.assertEqual(response_projects.status_code, 200)
        self.assertContains(response_projects, self.project.name)

        response_edit = self.client.get(reverse("tasks:edit", args=[self.task_other.id]))
        self.assertEqual(response_edit.status_code, 200)

    def test_programmer_edit_assigned_can_only_edit_own_tasks(self):
        TaskRoleAccessRule.objects.create(
            role_type=TaskRoleType.PROGRAMMER,
            access_level=TaskAccessLevel.EDIT_ASSIGNED,
        )
        self.client.force_login(self.programmer)

        response_detail = self.client.get(reverse("tasks:detail", args=[self.task_other.id]))
        self.assertEqual(response_detail.status_code, 200)

        response_allowed = self.client.post(
            reverse("tasks:project_gantt_update_task", args=[self.project.id, self.task_prog.id]),
            {
                f"task_{self.task_prog.id}-next_step_due": "2026-03-10",
                f"task_{self.task_prog.id}-due_date": "2026-03-12",
                f"task_{self.task_prog.id}-status": TaskStatus.IN_PROGRESS,
            },
        )
        self.assertEqual(response_allowed.status_code, 302)

        response_denied = self.client.post(
            reverse("tasks:project_gantt_update_task", args=[self.project.id, self.task_other.id]),
            {
                f"task_{self.task_other.id}-next_step_due": "2026-03-10",
                f"task_{self.task_other.id}-due_date": "2026-03-12",
                f"task_{self.task_other.id}-status": TaskStatus.IN_PROGRESS,
            },
        )
        self.assertEqual(response_denied.status_code, 403)

    def test_project_manager_read_all_can_view_but_not_edit(self):
        TaskRoleAccessRule.objects.create(
            role_type=TaskRoleType.PROJECT_MANAGER,
            access_level=TaskAccessLevel.READ_ALL,
        )
        self.client.force_login(self.pm_user)

        response_detail = self.client.get(reverse("tasks:detail", args=[self.task_other.id]))
        self.assertEqual(response_detail.status_code, 200)

        response_edit = self.client.get(reverse("tasks:edit", args=[self.task_other.id]))
        self.assertEqual(response_edit.status_code, 403)

    def test_user_override_edit_all_grants_global_scope_and_manage(self):
        TaskUserAccessRule.objects.create(
            user=self.override_user,
            access_level=TaskAccessLevel.EDIT_ALL,
        )
        self.client.force_login(self.override_user)

        response_detail = self.client.get(reverse("tasks:detail", args=[self.task_other.id]))
        self.assertEqual(response_detail.status_code, 200)

        response_edit = self.client.get(reverse("tasks:edit", args=[self.task_other.id]))
        self.assertEqual(response_edit.status_code, 200)

    def test_settings_access_tab_persists_role_and_user_rules(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("tasks:impostazioni"),
            {
                "tab": "accessi",
                f"access_role__{TaskRoleType.CAPO_COMMESSA}": TaskAccessLevel.EDIT_ALL,
                f"access_role__{TaskRoleType.PROGRAMMER}": TaskAccessLevel.EDIT_ASSIGNED,
                f"access_role__{TaskRoleType.PROJECT_MANAGER}": TaskAccessLevel.READ_ALL,
                "visible_access_user_id": [str(self.override_user.id)],
                f"access_user__{self.override_user.id}": TaskAccessLevel.EDIT_ALL,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            TaskRoleAccessRule.objects.filter(
                role_type=TaskRoleType.CAPO_COMMESSA,
                access_level=TaskAccessLevel.EDIT_ALL,
            ).exists()
        )
        self.assertTrue(
            TaskRoleAccessRule.objects.filter(
                role_type=TaskRoleType.PROGRAMMER,
                access_level=TaskAccessLevel.EDIT_ASSIGNED,
            ).exists()
        )
        self.assertTrue(
            TaskUserAccessRule.objects.filter(
                user=self.override_user,
                access_level=TaskAccessLevel.EDIT_ALL,
            ).exists()
        )

    def test_custom_role_category_rule_grants_scope(self):
        role = TaskRoleDefinition.objects.create(
            code="COLLAUDATORE",
            name="Collaudatore",
            order_index=40,
        )
        category = TaskCategory.objects.create(
            name="Collaudo finale",
            slug="collaudo-finale",
            role_type=role.code,
        )
        task = Task.objects.create(
            title="Task collaudo",
            created_by=self.owner,
            assigned_to=self.owner,
            project=self.project,
            category=category,
            next_step_due=timezone.localdate() + timedelta(days=1),
            due_date=timezone.localdate() + timedelta(days=3),
        )
        TaskRoleAssignment.objects.create(user=self.outsider, role_type=role.code)
        TaskRoleAccessRule.objects.create(role_type=role.code, access_level=TaskAccessLevel.EDIT_ALL)
        self.client.force_login(self.outsider)

        response_detail = self.client.get(reverse("tasks:detail", args=[task.id]))
        self.assertEqual(response_detail.status_code, 200)

        response_edit = self.client.get(reverse("tasks:edit", args=[task.id]))
        self.assertEqual(response_edit.status_code, 200)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class TaskOutlookReminderTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit"])
        self._refresh_acl_cache()
        self.user = _create_user_with_legacy(
            username="outlookuser", legacy_user_id=8001, role_id=2, role_name="utente",
        )
        self.project = Project.objects.create(name="KO Reminder", created_by=self.user)
        self.impostazioni = TaskImpostazioni.get_singleton()
        self.impostazioni.notifiche_scadenza_attive = True
        self.impostazioni.giorni_preavviso = 3
        self.impostazioni.save()

    def test_reminder_created_on_task_save_when_enabled(self):
        from .models import TaskReminder

        target_due = timezone.localdate() + timedelta(days=10)
        self.client.force_login(self.user)
        payload = {
            "title": "Task con reminder",
            "description": "",
            "status": TaskStatus.TODO,
            "priority": TaskPriority.MEDIUM,
            "task_scope": "project",
            "project_link_mode": "existing",
            "project_choice": str(self.project.id),
            "due_date": target_due.isoformat(),
            "assigned_to": str(self.user.id),
            "reminder_portal_enabled_field": "on",
        }
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)
        task = Task.objects.get(title="Task con reminder")
        self.assertTrue(task.reminder_portal_enabled)
        reminders = TaskReminder.objects.filter(task=task)
        self.assertEqual(reminders.count(), 1)
        r = reminders.first()
        self.assertEqual(r.fire_at, target_due - timedelta(days=3))
        self.assertEqual(r.legacy_user_id, 8001)
        self.assertFalse(r.fired)

    def test_reminder_not_created_when_disabled(self):
        from .models import TaskReminder

        target_due = timezone.localdate() + timedelta(days=10)
        self.client.force_login(self.user)
        payload = {
            "title": "Task senza reminder",
            "description": "",
            "status": TaskStatus.TODO,
            "priority": TaskPriority.MEDIUM,
            "task_scope": "project",
            "project_link_mode": "existing",
            "project_choice": str(self.project.id),
            "due_date": target_due.isoformat(),
            "assigned_to": str(self.user.id),
            # reminder_portal_enabled_field NOT sent => False
        }
        response = self.client.post(reverse("tasks:create"), payload)
        self.assertEqual(response.status_code, 302)
        task = Task.objects.get(title="Task senza reminder")
        self.assertFalse(task.reminder_portal_enabled)
        self.assertEqual(TaskReminder.objects.filter(task=task).count(), 0)

    def test_send_task_reminders_command_creates_notifica(self):
        from .models import TaskReminder
        from django.core.management import call_command
        from io import StringIO

        task = Task.objects.create(
            title="T reminder fire",
            project=self.project,
            assigned_to=self.user,
            due_date=timezone.localdate() + timedelta(days=1),
            created_by=self.user,
            reminder_portal_enabled=True,
        )
        TaskReminder.objects.create(
            task=task, legacy_user_id=8001,
            fire_at=timezone.localdate(),
        )
        out = StringIO()
        call_command("send_task_reminders", stdout=out)
        r = TaskReminder.objects.get(task=task)
        self.assertTrue(r.fired)
        self.assertIsNotNone(r.fired_at)
        self.assertEqual(
            Notifica.objects.filter(legacy_user_id=8001).count(),
            1,
        )
        notif = Notifica.objects.get(legacy_user_id=8001)
        self.assertIn("T reminder fire", notif.messaggio)
        self.assertTrue(notif.url_azione)

    def test_send_task_reminders_skips_closed_task(self):
        from .models import TaskReminder
        from django.core.management import call_command
        from io import StringIO

        task = Task.objects.create(
            title="T chiuso",
            project=self.project, assigned_to=self.user,
            due_date=timezone.localdate() + timedelta(days=1),
            status=TaskStatus.DONE,
            created_by=self.user, reminder_portal_enabled=True,
        )
        TaskReminder.objects.create(
            task=task, legacy_user_id=8001, fire_at=timezone.localdate(),
        )
        call_command("send_task_reminders", stdout=StringIO())
        self.assertTrue(TaskReminder.objects.get(task=task).fired)
        self.assertEqual(Notifica.objects.filter(legacy_user_id=8001).count(), 0)

    def test_edit_form_prefills_existing_outlook_target(self):
        from .forms import TaskForm

        task = Task.objects.create(
            title="Task con evento Outlook",
            project=self.project,
            assigned_to=self.user,
            due_date=timezone.localdate() + timedelta(days=2),
            created_by=self.user,
            reminder_portal_enabled=True,
        )
        TaskCalendarEvent.objects.create(
            task=task,
            source_key=f"tasks.task:{task.id}:due",
            target_email="planner@example.com",
            target_display_name="Planner",
            due_date=task.due_date,
            subject="Task con evento Outlook",
            transaction_id="tx-001",
            graph_event_id="evt-001",
            graph_event_web_link="https://outlook.office.com/calendar/item/evt-001",
            created_by=self.user,
        )

        form = TaskForm(instance=task)

        self.assertTrue(form.initial["add_to_outlook"])
        self.assertEqual(form.initial["outlook_target_email"], "planner@example.com")

    def test_sync_task_outlook_event_access_denied_message_is_actionable(self):
        from unittest.mock import patch

        from .outlook_reminder import sync_task_outlook_event

        self.user.email = "outlook.user@cnovicrom.local"
        self.user.save(update_fields=["email"])
        task = Task.objects.create(
            title="Task access denied",
            project=self.project,
            assigned_to=self.user,
            due_date=timezone.localdate() + timedelta(days=4),
            created_by=self.user,
            reminder_portal_enabled=True,
        )

        with patch("tasks.outlook_reminder.graph_ready", return_value=True), patch(
            "tasks.outlook_reminder.create_event",
            side_effect=RuntimeError("Access is denied. Check credentials and try again."),
        ):
            level, message = sync_task_outlook_event(
                request=None,
                task=task,
                requested=True,
                explicit_email="",
            )

        self.assertEqual(level, "warning")
        self.assertIn("Calendars.ReadWrite", message)
        self.assertIn("Email calendario Outlook", message)
        self.assertIn(".local", message)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class ProjectCreateFlowTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit"])
        self._refresh_acl_cache()
        self.user = _create_user_with_legacy(
            username="kickoffuser", legacy_user_id=7001, role_id=2, role_name="utente",
        )

    def test_get_project_create_page(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:project_create"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nuovo kickoff")
        self.assertContains(response, 'name="part_number"')
        self.assertContains(response, 'name="client_name"')
        self.assertContains(response, 'name="safety_impact"')
        self.assertContains(response, "Impatto sulla sicurezza")

    def test_post_creates_project_and_redirects_to_vrf_compile(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:project_create"),
            {
                "client_name": "Cliente Test",
                "part_number": "PN-NEW-001",
                "revisione": "A",
                "versione": "1.0",
                "description": "Desc",
                "control_method": "",
                "vrf_quote_number": "Q-1",
                "vrf_description": "",
                "vrf_esp": "02",
                "project_manager": self.user.id,
            },
        )
        project = Project.objects.get(part_number="PN-NEW-001")
        self.assertRedirects(
            response,
            reverse("tasks:project_vrf_compile", args=[project.id]),
        )
        self.assertEqual(project.client_name, "Cliente Test")
        self.assertEqual(project.revisione, "A")
        self.assertEqual(project.created_by, self.user)
        self.assertFalse(project.safety_impact)

    def test_project_safety_impact_defaults_false(self):
        project = Project.objects.create(name="Kickoff safety default", created_by=self.user)

        self.assertFalse(project.safety_impact)

    def test_post_creates_project_with_safety_impact_true_without_changing_status_or_priority(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:project_create"),
            {
                "client_name": "Cliente Safety",
                "part_number": "PN-SAFE-001",
                "revisione": "A",
                "versione": "1.0",
                "description": "Desc",
                "control_method": "",
                "safety_impact": "on",
                "vrf_quote_number": "Q-SAFE",
                "vrf_description": "",
                "vrf_esp": "",
                "project_manager": self.user.id,
            },
        )

        project = Project.objects.get(part_number="PN-SAFE-001")
        self.assertRedirects(response, reverse("tasks:project_vrf_compile", args=[project.id]))
        self.assertTrue(project.safety_impact)
        self.assertEqual(project.vrf_status, VRFDocStatus.PENDING)

    def test_project_kickoff_form_edit_post_can_clear_safety_impact(self):
        project = Project.objects.create(
            client_name="Cliente Edit",
            part_number="PN-EDIT-SAFE",
            revisione="A",
            versione="1.0",
            safety_impact=True,
            created_by=self.user,
        )

        form = ProjectKickoffForm(
            {
                "client_name": "Cliente Edit",
                "part_number": "PN-EDIT-SAFE-2",
                "revisione": "A",
                "versione": "1.0",
                "description": project.description,
                "control_method": project.control_method,
                "vrf_quote_number": project.vrf_quote_number,
                "vrf_description": project.vrf_description,
                "vrf_esp": project.vrf_esp,
                "project_manager": self.user.id,
            },
            instance=project,
            project_queryset=Project.objects.exclude(pk=project.pk),
        )

        self.assertTrue(form.is_valid(), form.errors)
        updated = form.save()
        self.assertFalse(updated.safety_impact)

    def test_revisione_without_part_number_rejected(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:project_create"),
            {"client_name": "X", "part_number": "", "revisione": "A", "versione": "", "project_manager": self.user.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Revisione e versione sono valide solo con un P/N indicato.",
        )
        self.assertFalse(Project.objects.filter(client_name="X").exists())

    def test_duplicate_identity_reuses_existing_kickoff(self):
        existing = Project.objects.create(
            part_number="DUP-001", revisione="B", versione="2.0",
            client_name="Cliente Precedente", created_by=self.user,
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:project_create"),
            {
                "client_name": "Cliente Nuovo",
                "part_number": "DUP-001",
                "revisione": "B",
                "versione": "2.0",
                "project_manager": self.user.id,
            },
        )
        self.assertRedirects(
            response,
            reverse("tasks:project_vrf_compile", args=[existing.id]),
        )
        self.assertEqual(Project.objects.filter(part_number="DUP-001").count(), 1)
        existing.refresh_from_db()
        self.assertEqual(existing.client_name, "Cliente Precedente")

    def test_safety_impact_badge_visible_on_project_detail_and_list_only_when_true(self):
        safety_project = Project.objects.create(
            name="Kickoff safety",
            part_number="PN-SAFE-LIST",
            safety_impact=True,
            created_by=self.user,
        )
        Project.objects.create(
            name="Kickoff normal",
            part_number="PN-NORMAL-LIST",
            safety_impact=False,
            created_by=self.user,
        )
        self.client.force_login(self.user)

        list_response = self.client.get(reverse("tasks:project_list"))
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Impatto sicurezza", count=1)

        detail_response = self.client.get(reverse("tasks:project_gantt", args=[safety_project.id]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Impatto sicurezza")

        normal_response = self.client.get(reverse("tasks:project_gantt", args=[Project.objects.get(part_number="PN-NORMAL-LIST").id]))
        self.assertEqual(normal_response.status_code, 200)
        self.assertNotContains(normal_response, "Impatto sicurezza")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class VRFCompileOnlineTests(TasksBaseTestCase):
    def setUp(self):
        super().setUp()
        _ensure_role(2, "utente")
        _grant_role_actions(2, ["tasks_view", "tasks_create", "tasks_edit"])
        self._refresh_acl_cache()
        self.user = _create_user_with_legacy(
            username="vrfuser", legacy_user_id=6001, role_id=2, role_name="utente",
        )
        tmp_root = Path(__file__).resolve().parents[1] / ".tmp_tests"
        tmp_root.mkdir(parents=True, exist_ok=True)
        self._media_root = str(tmp_root / f"vrf_compile_{uuid4().hex}")
        Path(self._media_root).mkdir(parents=True, exist_ok=True)
        self._media_override = override_settings(MEDIA_ROOT=self._media_root)
        self._media_override.enable()
        today = timezone.localdate()
        Path(self._media_root, "tasks_vrf", today.strftime("%Y"), today.strftime("%m")).mkdir(parents=True, exist_ok=True)
        self.addCleanup(self._media_override.disable)
        self.addCleanup(shutil.rmtree, self._media_root, True)
        self.project = Project.objects.create(
            name="TBD",
            client_name="Test Client",
            part_number="PN-TEST-1",
            vrf_description="Particolare di test",
            vrf_esp="02",
            vrf_quote_number="Q-TEST",
            created_by=self.user,
        )

    def test_compile_page_renders_full_matrix(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("tasks:project_vrf_compile", args=[self.project.id])
        )
        self.assertEqual(response.status_code, 200)
        for r_code in ("R1", "R2", "R3", "R4", "R5", "R6", "R7", "R8", "R9"):
            self.assertContains(response, f'data-risk="{r_code}"')
        self.assertContains(response, 'name="score_R1_a_p"')
        self.assertContains(response, 'name="k_R2_c"')

    def test_save_draft_persists_scores_and_recomputes_totals(self):
        from .models import VRFRiskAssessment

        self.client.force_login(self.user)
        payload = {
            "action": "save_draft",
            "score_R1_a_p": "3", "score_R1_b_p": "2",
            "score_R2_a_p": "1", "score_R2_b_p": "2", "score_R2_c_p": "3",
            "k_R1_p": "3", "k_R1_i": "3", "k_R1_c": "3",
            "k_R2_p": "3", "k_R2_i": "3", "k_R2_c": "5",
        }
        response = self.client.post(
            reverse("tasks:project_vrf_compile", args=[self.project.id]), payload
        )
        self.assertEqual(response.status_code, 302)
        assessment = VRFRiskAssessment.objects.get(project=self.project)
        self.assertEqual(assessment.data["risks"]["R1"]["subs"]["a"]["p"], 3)
        self.assertEqual(assessment.data["risks"]["R2"]["subs"]["c"]["p"], 3)
        self.assertGreater(assessment.total_p, 0.0)
        self.assertEqual(assessment.total_i, 0.0)
        self.project.refresh_from_db()
        self.assertEqual(self.project.vrf_status, VRFDocStatus.PENDING)

    def test_confirm_generates_xlsx_and_marks_uploaded(self):
        self.client.force_login(self.user)
        payload = {
            "action": "confirm",
            "score_R1_a_p": "2", "score_R1_b_p": "3",
            "k_R1_p": "3",
        }
        response = self.client.post(
            reverse("tasks:project_vrf_compile", args=[self.project.id]), payload
        )
        self.assertEqual(response.status_code, 302)
        self.project.refresh_from_db()
        self.assertEqual(self.project.vrf_status, VRFDocStatus.UPLOADED)
        self.assertTrue(self.project.vrf_file.name)
        self.assertTrue(self.project.vrf_original_name.endswith(".xlsx"))
        import openpyxl
        self.project.vrf_file.open("rb")
        try:
            wb = openpyxl.load_workbook(io.BytesIO(self.project.vrf_file.read()))
        finally:
            self.project.vrf_file.close()
        ws = wb["VRF"]
        self.assertEqual(ws["B3"].value, "PN-TEST-1")
        self.assertEqual(ws["B4"].value, "Test Client")
        self.assertEqual(ws["U8"].value, 2)
        self.assertEqual(ws["U9"].value, 3)
        wb.close()

    def test_skip_with_reminder_saves_draft_and_redirects_to_gantt(self):
        from .models import VRFRiskAssessment

        self.client.force_login(self.user)
        payload = {
            "action": "skip_with_reminder",
            "score_R1_a_p": "2",
            "k_R1_p": "3",
        }
        response = self.client.post(
            reverse("tasks:project_vrf_compile", args=[self.project.id]), payload
        )
        self.assertRedirects(
            response,
            reverse("tasks:project_gantt", args=[self.project.id]),
        )
        # Stato rimane PENDING (non UPLOADED): il reminder progressivo resta attivo
        self.project.refresh_from_db()
        self.assertEqual(self.project.vrf_status, VRFDocStatus.PENDING)
        # La bozza e' comunque stata salvata
        assessment = VRFRiskAssessment.objects.get(project=self.project)
        self.assertEqual(assessment.data["risks"]["R1"]["subs"]["a"]["p"], 2)

    def test_dig_threshold_triggered_when_totals_high(self):
        from .models import VRFRiskAssessment
        from . import vrf_catalog

        self.client.force_login(self.user)
        payload = {"action": "save_draft"}
        for risk in vrf_catalog.RISKS:
            for sub in risk["sub_parameters"]:
                payload[f"score_{risk['code']}_{sub['code']}_p"] = "3"
            payload[f"k_{risk['code']}_p"] = "5"
            payload[f"k_{risk['code']}_i"] = str(risk["k_default"]["i"])
            payload[f"k_{risk['code']}_c"] = str(risk["k_default"]["c"])
        response = self.client.post(
            reverse("tasks:project_vrf_compile", args=[self.project.id]), payload
        )
        self.assertEqual(response.status_code, 302)
        assessment = VRFRiskAssessment.objects.get(project=self.project)
        self.assertGreaterEqual(assessment.total_p, vrf_catalog.DIG_THRESHOLD)
        self.assertTrue(assessment.dig_triggered)
