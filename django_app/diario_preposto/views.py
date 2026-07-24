from __future__ import annotations

import logging
import json
import mimetypes
import os
from datetime import timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import redirect_to_login
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from core.acl_v2 import request_has_permission_code
from core.module_branding import get_module_branding_context, handle_module_branding_post
from core.legacy_utils import get_legacy_user, is_legacy_admin
from core.models import ChecklistEsecuzione, ChecklistRisposta, ChecklistVoce
from core.upload_mime import (
    UploadMimeValidationError,
    safe_filename,
    validate_extension_and_mime,
)

from .acl_bootstrap import PERM_IMPOSTAZIONI_MANAGE
from .forms import SegnalazioneForm
from .models import DiarioPrepostoImpostazioni, SegnalazioneAllegato, SegnalazionePreposto

logger = logging.getLogger(__name__)

CHECKLIST_ISPEZIONE_PREPOSTO = "preposto_ispezione"


# ---------------------------------------------------------------------------
# Validazione allegati segnalazioni
# ---------------------------------------------------------------------------

DIARIO_ALLEGATO_MAX_BYTES = 20 * 1024 * 1024  # 20 MB
DIARIO_ALLEGATO_EXTENSIONS = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".txt",
    ".csv",
}
DIARIO_ALLEGATO_MIMES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/webp",
    "image/bmp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain",
    "text/csv",
    # Alcuni server identificano i CSV come text/plain, gia' coperto.
    "application/octet-stream",  # fallback per documenti firmati / formati legacy
}


def _validate_diario_allegato(uploaded_file):
    """Valida un allegato del Diario Preposto. Ritorna nome sicuro o solleva."""
    safe_name = safe_filename(getattr(uploaded_file, "name", ""))
    label = safe_name or "Allegato"
    validate_extension_and_mime(
        uploaded_file,
        allowed_extensions=DIARIO_ALLEGATO_EXTENSIONS,
        allowed_mimes=DIARIO_ALLEGATO_MIMES,
        max_bytes=DIARIO_ALLEGATO_MAX_BYTES,
        label=label,
        allow_empty=False,
    )
    return safe_name or "allegato"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _legacy_identity(request) -> tuple[str, str]:
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if legacy_user:
        name = (legacy_user.nome or "").strip() or request.user.get_full_name() or request.user.get_username()
        email = (legacy_user.email or "").strip().lower() or (request.user.email or "").strip().lower()
        return name, email
    name = request.user.get_full_name() or request.user.get_username()
    email = (request.user.email or "").strip().lower()
    return name, email


def _user_acl_identities(request) -> set[str]:
    """Raccoglie tutte le identità (lowercase) con cui l'utente può essere
    stato registrato in ``acl_scrittura``.

    L'elenco può contenere — a seconda di cosa salva il widget impostazioni —
    l'username Django (UPN), l'email di login aziendale o l'``aliasusername``
    legacy. Per evitare esclusioni da mismatch (es. ``a.astarita`` vs
    ``a.astarita@dominio``) il match considera tutte queste forme, incluso il
    prefisso dell'UPN. NON si considera l'email privata: l'identità valida è
    solo mail aziendale o username. Stesso pattern di ``tickets._user_acl_identities``.
    """
    identities: set[str] = set()

    def _add(value) -> None:
        text = str(value or "").strip().lower()
        if text:
            identities.add(text)
            if "@" in text:
                identities.add(text.split("@", 1)[0])  # prefisso UPN

    _add(request.user.get_username())
    _add(request.user.email)

    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    legacy_id = getattr(legacy_user, "id", None) if legacy_user else None
    if legacy_id:
        try:
            from core.legacy_models import AnagraficaDipendente
            ana = (
                AnagraficaDipendente.objects
                .filter(utente_id=legacy_id)
                .only("aliasusername", "email")
                .first()
            )
            if ana:
                _add(ana.aliasusername)   # username legacy
                _add(ana.email)           # mail aziendale / UPN
        except Exception:
            pass  # in caso di errore DB usiamo solo le identità Django
    return identities


def _acl_list_matches(acl: list, identities: set[str]) -> bool:
    """True se una qualsiasi voce della ACL list combacia con un'identità utente."""
    for raw in acl or []:
        value = str(raw or "").strip().lower()
        if not value:
            continue
        if value in identities:
            return True
        if "@" in value and value.split("@", 1)[0] in identities:
            return True
    return False


def _can_write(request) -> bool:
    """Controlla se l'utente puo creare/modificare/eliminare segnalazioni."""
    if not request.user.is_authenticated:
        return False
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if legacy_user and is_legacy_admin(legacy_user):
        return True
    cfg = DiarioPrepostoImpostazioni.objects.first()
    if not cfg:
        return True  # nessuna config = accesso aperto
    acl = cfg.acl_scrittura or []
    if not acl:
        return True  # elenco vuoto = accesso aperto a tutti gli autenticati
    return _acl_list_matches(acl, _user_acl_identities(request))


def _can_manage_settings(request) -> bool:
    if not request.user.is_authenticated:
        return False
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if getattr(request.user, "is_superuser", False) or (legacy_user and is_legacy_admin(legacy_user)):
        return True
    # ACL v2 canonico: concedibile per ruolo da /admin-portale/acl-canonico/.
    return request_has_permission_code(request, PERM_IMPOSTAZIONI_MANAGE)


def _can_view(request) -> bool:
    """Verifica l'accesso al modulo Diario Preposto per il download allegati.

    SEC-AUDIT-002: l'URL ``/diario-preposto/allegato/`` è registrato in
    ``_ACL_SHARED_PREFIXES`` (core/middleware.py) e quindi *bypassa* il
    middleware ACL. La vista di download deve perciò verificare da sé
    l'accesso al modulo, riusando la stessa decisione ACL che protegge la
    scheda dettaglio (``diario_preposto:lista``). Nessuna nuova policy:
    si applica la visibilità di modulo già esistente.
    """
    user = getattr(request, "user", None)
    if not user or not user.is_authenticated:
        return False
    # SEC-AUDIT-002 (fix): NON usare _can_write come scorciatoia di autorizzazione
    # alla lettura. _can_write è fail-open (ritorna True per qualsiasi utente
    # autenticato quando acl_scrittura è vuota = default), perciò — essendo questa
    # route esente dal gate ACL del middleware — rendeva il download degli allegati
    # accessibile a chiunque. La visibilità del download segue ORA solo l'ACL di
    # modulo (la stessa che protegge la scheda dettaglio), come da docstring.
    if getattr(user, "is_superuser", False):
        return True
    from core.legacy_utils import legacy_auth_enabled
    if not legacy_auth_enabled():
        return True
    from core.acl_v2 import resolve_acl_access
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(user)
    try:
        decision = resolve_acl_access(
            path=reverse("diario_preposto:lista"),
            legacy_user=legacy_user,
            django_user=user,
            request=request,
            include_legacy_diagnostic=False,
        )
    except Exception:
        return False
    return bool(decision.get("allowed", False))


def _write_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        if not _can_write(request):
            return render(request, "core/pages/forbidden.html", status=403)
        return view_func(request, *args, **kwargs)

    return _wrapped


def _json_err(msg: str, status: int = 400) -> JsonResponse:
    return JsonResponse({"ok": False, "error": msg}, status=status)


def _pdf_safe_text(value: object, *, preserve_breaks: bool = False) -> str:
    text = str(value or "-").strip() or "-"
    text = escape(text)
    if preserve_breaks:
        return text.replace("\n", "<br/>")
    return text


def _ensure_ispezione_voci() -> list[ChecklistVoce]:
    voci = list(
        ChecklistVoce.objects.filter(
            tipo_checklist=CHECKLIST_ISPEZIONE_PREPOSTO,
            is_active=True,
        ).order_by("categoria", "ordine", "id")
    )
    if voci:
        return voci
    defaults = [
        ("Area", "Area ordinata e percorsi liberi", "check", True, 10),
        ("Macchina", "Protezioni e ripari presenti", "check", True, 20),
        ("DPI", "DPI disponibili e correttamente utilizzati", "check", True, 30),
        ("Note", "Anomalie o azioni correttive", "testo", False, 40),
    ]
    created = [
        ChecklistVoce.objects.create(
            tipo_checklist=CHECKLIST_ISPEZIONE_PREPOSTO,
            categoria=categoria,
            label=label,
            tipo_campo=tipo_campo,
            obbligatorio=obbligatorio,
            ordine=ordine,
            is_active=True,
        )
        for categoria, label, tipo_campo, obbligatorio, ordine in defaults
    ]
    return created


def _ispezione_metadata(esecuzione: ChecklistEsecuzione) -> dict:
    try:
        data = json.loads(esecuzione.note or "{}")
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        return {}


def _checklist_value_from_post(voce: ChecklistVoce, post) -> str:
    key = f"voce_{voce.pk}"
    if voce.tipo_campo == "check":
        return "1" if post.get(key) else "0"
    return (post.get(key) or "").strip()


def _ispezione_actor(request) -> tuple[int, str]:
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    name, _email = _legacy_identity(request)
    actor_id = legacy_user.id if legacy_user else request.user.id
    return actor_id, name


def _prepare_ispezione_rows(esecuzioni) -> list[ChecklistEsecuzione]:
    rows = list(esecuzioni)
    for esecuzione in rows:
        esecuzione.meta = _ispezione_metadata(esecuzione)
        esecuzione.ok_count = sum(1 for risposta in esecuzione.risposte.all() if risposta.valore == "1")
        esecuzione.ko_count = sum(1 for risposta in esecuzione.risposte.all() if risposta.voce_tipo == "check" and risposta.valore != "1")
    return rows


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

@login_required
def lista(request):
    qs = SegnalazionePreposto.objects.all()

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(codice_identificativo__icontains=q)
            | Q(titolo__icontains=q)
            | Q(descrizione__icontains=q)
            | Q(chi_segnala__icontains=q)
            | Q(preposto__icontains=q)
        )

    filtro_preposto = request.GET.get("preposto", "").strip()
    if filtro_preposto:
        qs = qs.filter(preposto__icontains=filtro_preposto)

    preposti = (
        SegnalazionePreposto.objects.exclude(preposto="")
        .values_list("preposto", flat=True)
        .distinct()
        .order_by("preposto")
    )

    return render(
        request,
        "diario_preposto/pages/lista.html",
        {
            "segnalazioni": qs,
            "q": q,
            "filtro_preposto": filtro_preposto,
            "preposti": preposti,
            "can_write": _can_write(request),
            "can_manage_settings": _can_manage_settings(request),
            "totale": qs.count(),
        },
    )


@login_required
def ispezioni(request):
    voci = _ensure_ispezione_voci()
    cfg = DiarioPrepostoImpostazioni.get_singleton()
    esecuzioni = _prepare_ispezione_rows(
        ChecklistEsecuzione.objects.filter(tipo_checklist=CHECKLIST_ISPEZIONE_PREPOSTO)
        .prefetch_related("risposte")
        .order_by("-data_esecuzione")[:50]
    )
    ultima = esecuzioni[0] if esecuzioni else None
    prossima_scadenza = None
    ispezione_scaduta = False
    if ultima:
        prossima_scadenza = timezone.localtime(ultima.data_esecuzione).date() + timedelta(
            days=cfg.ispezione_frequenza_giorni
        )
        ispezione_scaduta = prossima_scadenza < timezone.localdate()

    return render(
        request,
        "diario_preposto/pages/ispezioni.html",
        {
            "voci": voci,
            "esecuzioni": esecuzioni,
            "ultima": ultima,
            "prossima_scadenza": prossima_scadenza,
            "ispezione_scaduta": ispezione_scaduta,
            "frequenza_giorni": cfg.ispezione_frequenza_giorni,
            "can_write": _can_write(request),
            "can_manage_settings": _can_manage_settings(request),
        },
    )


@_write_required
def ispezione_nuova(request):
    voci = _ensure_ispezione_voci()
    errors: list[str] = []
    form_data = {}
    for voce in voci:
        voce.submitted_value = ""

    if request.method == "POST":
        form_data = dict(request.POST)
        area = (request.POST.get("area") or "").strip()
        macchina = (request.POST.get("macchina") or "").strip()
        note = (request.POST.get("note") or "").strip()
        if not area:
            errors.append("Area obbligatoria.")

        risposte = []
        for voce in voci:
            value = _checklist_value_from_post(voce, request.POST)
            voce.submitted_value = value
            if voce.obbligatorio and voce.tipo_campo != "check" and not value:
                errors.append(f"Compilare la voce obbligatoria: {voce.label}.")
            if voce.tipo_campo == "select" and value and voce.scelte and value not in voce.scelte:
                errors.append(f"Valore non valido per: {voce.label}.")
            risposte.append((voce, value))

        if not errors:
            actor_id, actor_name = _ispezione_actor(request)
            metadata = {
                "area": area,
                "macchina": macchina,
                "note": note,
            }
            with transaction.atomic():
                esecuzione = ChecklistEsecuzione.objects.create(
                    legacy_user_id=actor_id,
                    utente_nome=area[:200],
                    tipo_checklist=CHECKLIST_ISPEZIONE_PREPOSTO,
                    eseguita_da_id=actor_id,
                    eseguita_da_nome=actor_name[:200],
                    note=json.dumps(metadata, ensure_ascii=False),
                    completata=True,
                )
                ChecklistRisposta.objects.bulk_create(
                    [
                        ChecklistRisposta(
                            esecuzione=esecuzione,
                            voce_id=voce.pk,
                            voce_label=voce.label,
                            voce_tipo=voce.tipo_campo,
                            valore=value,
                        )
                        for voce, value in risposte
                    ]
                )
            messages.success(request, "Ispezione periodica registrata.")
            return redirect("diario_preposto:ispezioni")

    return render(
        request,
        "diario_preposto/pages/ispezione_form.html",
        {
            "voci": voci,
            "errors": errors,
            "form_data": form_data,
        },
    )


@login_required
def dettaglio(request, pk):
    segnalazione = get_object_or_404(SegnalazionePreposto, pk=pk)
    allegati = segnalazione.allegati.all()
    return render(
        request,
        "diario_preposto/pages/dettaglio.html",
        {
            "segnalazione": segnalazione,
            "allegati": allegati,
            "can_write": _can_write(request),
        },
    )


@_write_required
def nuovo(request):
    nome, email = _legacy_identity(request)
    if request.method == "POST":
        form = SegnalazioneForm(request.POST)
        if form.is_valid():
            segnalazione = form.save(commit=False)
            segnalazione.creato_da = request.user
            segnalazione.preposto = nome
            segnalazione.chi_segnala = nome
            segnalazione.save()
            errori_allegati: list[str] = []
            for uploaded_file in request.FILES.getlist("allegati"):
                try:
                    safe_name = _validate_diario_allegato(uploaded_file)
                except UploadMimeValidationError as exc:
                    errori_allegati.append(str(exc))
                    continue
                SegnalazioneAllegato.objects.create(
                    segnalazione=segnalazione,
                    nome_file=safe_name,
                    file=uploaded_file,
                )
            if errori_allegati:
                for err in errori_allegati:
                    messages.warning(request, err)
            messages.success(request, "Segnalazione inserita con successo.")
            return redirect("diario_preposto:dettaglio", pk=segnalazione.pk)
    else:
        now_local = timezone.localtime(timezone.now())
        form = SegnalazioneForm(
            initial={
                "chi_segnala": nome,
                "data_segnalazione": now_local.strftime("%Y-%m-%dT%H:%M"),
            }
        )
    return render(
        request,
        "diario_preposto/pages/form.html",
        {
            "form": form,
            "action": "nuovo",
            "title": "Nuovo inserimento",
        },
    )


@_write_required
def modifica(request, pk):
    segnalazione = get_object_or_404(SegnalazionePreposto, pk=pk)
    if request.method == "POST":
        form = SegnalazioneForm(request.POST, instance=segnalazione)
        if form.is_valid():
            form.save()
            errori_allegati: list[str] = []
            for uploaded_file in request.FILES.getlist("allegati"):
                try:
                    safe_name = _validate_diario_allegato(uploaded_file)
                except UploadMimeValidationError as exc:
                    errori_allegati.append(str(exc))
                    continue
                SegnalazioneAllegato.objects.create(
                    segnalazione=segnalazione,
                    nome_file=safe_name,
                    file=uploaded_file,
                )
            if errori_allegati:
                for err in errori_allegati:
                    messages.warning(request, err)
            messages.success(request, "Segnalazione aggiornata.")
            return redirect("diario_preposto:dettaglio", pk=segnalazione.pk)
    else:
        initial = {}
        if segnalazione.data_segnalazione:
            local_dt = timezone.localtime(segnalazione.data_segnalazione)
            initial["data_segnalazione"] = local_dt.strftime("%Y-%m-%dT%H:%M")
        form = SegnalazioneForm(instance=segnalazione, initial=initial)
    return render(
        request,
        "diario_preposto/pages/form.html",
        {
            "form": form,
            "action": "modifica",
            "title": "Modifica segnalazione",
            "segnalazione": segnalazione,
            "allegati": segnalazione.allegati.all(),
        },
    )


@_write_required
@require_POST
def elimina(request, pk):
    segnalazione = get_object_or_404(SegnalazionePreposto, pk=pk)
    for allegato in segnalazione.allegati.all():
        try:
            if allegato.file and allegato.file.name:
                allegato.file.storage.delete(allegato.file.name)
        except Exception:
            pass
    segnalazione.delete()
    messages.success(request, "Segnalazione eliminata.")
    return redirect("diario_preposto:lista")


@login_required
def export_excel(request):
    """Esporta la lista delle segnalazioni in formato Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Titolo/descrizione/preposto sono testo libero: senza sanificazione
        # openpyxl li scriverebbe come formula viva (formula injection).
        from core.excel_export import write_cell

        qs = SegnalazionePreposto.objects.all()

        q = request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(codice_identificativo__icontains=q)
                | Q(titolo__icontains=q)
                | Q(descrizione__icontains=q)
                | Q(chi_segnala__icontains=q)
                | Q(preposto__icontains=q)
            )

        filtro_preposto = request.GET.get("preposto", "").strip()
        if filtro_preposto:
            qs = qs.filter(preposto__icontains=filtro_preposto)

        wb = Workbook()
        ws = wb.active
        ws.title = "Diario Preposto"

        headers = [
            "Codice identificativo",
            "Data segnalazione",
            "Titolo",
            "Descrizione",
            "Preposto",
            "Chi segnala",
            "Creato da",
            "Numero allegati",
            "created_at",
            "updated_at",
        ]

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="03787C", end_color="03787C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num, header in enumerate(headers, start=1):
            cell = write_cell(ws, 1, col_num, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, segnalazione in enumerate(qs, start=2):
            data_segnalazione = (
                timezone.localtime(segnalazione.data_segnalazione).strftime("%d-%m-%Y %H:%M")
                if segnalazione.data_segnalazione
                else ""
            )
            created_at = timezone.localtime(segnalazione.created_at).strftime("%d-%m-%Y %H:%M")
            updated_at = timezone.localtime(segnalazione.updated_at).strftime("%d-%m-%Y %H:%M")
            creato_da = segnalazione.creato_da.get_full_name() if segnalazione.creato_da else ""

            write_cell(ws, row_num, 1, segnalazione.codice_identificativo or "")
            write_cell(ws, row_num, 2, data_segnalazione)
            write_cell(ws, row_num, 3, segnalazione.titolo or "")
            write_cell(ws, row_num, 4, segnalazione.descrizione or "")
            write_cell(ws, row_num, 5, segnalazione.preposto or "")
            write_cell(ws, row_num, 6, segnalazione.chi_segnala or "")
            write_cell(ws, row_num, 7, creato_da)
            write_cell(ws, row_num, 8, segnalazione.allegati.count())
            write_cell(ws, row_num, 9, created_at)
            write_cell(ws, row_num, 10, updated_at)

            for col_num in range(1, 11):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"diario_preposto_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        return HttpResponse("openpyxl non disponibile", status=500)


@login_required
def export_pdf(request, pk):
    segnalazione = get_object_or_404(SegnalazionePreposto, pk=pk)
    allegati = list(segnalazione.allegati.all())

    try:
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, Spacer

        from core.pdf import (
            PdfTheme,
            build_styles,
            data_table,
            header_footer_callback,
            make_document,
            section_heading,
        )
    except ImportError:
        return HttpResponse("reportlab non disponibile", status=500)

    codice = segnalazione.codice_identificativo or f"PK-{segnalazione.pk}"
    data_segnalazione = (
        timezone.localtime(segnalazione.data_segnalazione).strftime("%d-%m-%Y %H:%M")
        if segnalazione.data_segnalazione
        else "-"
    )
    created_at = timezone.localtime(segnalazione.created_at).strftime("%d-%m-%Y %H:%M")
    updated_at = timezone.localtime(segnalazione.updated_at).strftime("%d-%m-%Y %H:%M")

    theme = PdfTheme.from_branding()
    styles = build_styles(theme)
    buf = BytesIO()
    doc = make_document(buf, title=f"Segnalazione {codice}")
    content_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    def meta_cell(label: str, value: object) -> list:
        return [
            Paragraph(_pdf_safe_text(label).upper(), styles["label"]),
            Paragraph(_pdf_safe_text(value), styles["value"]),
        ]

    story: list = [
        Paragraph(_pdf_safe_text(segnalazione.titolo), styles["title"]),
        Paragraph(
            "Report della segnalazione di sicurezza con riepilogo dati, descrizione completa e allegati collegati.",
            styles["body"],
        ),
        Spacer(1, 4 * mm),
    ]

    meta_table = data_table(
        [
            [meta_cell("ID segnalazione", codice), meta_cell("Data segnalazione", data_segnalazione)],
            [meta_cell("Preposto", segnalazione.preposto or "-"), meta_cell("Chi segnala", segnalazione.chi_segnala or "-")],
            [meta_cell("Creato il", created_at), meta_cell("Ultimo aggiornamento", updated_at)],
        ],
        theme,
        col_widths=[content_width / 2, content_width / 2],
        header=False,
        extra_style=[("VALIGN", (0, 0), (-1, -1), "TOP")],
    )
    story.extend([meta_table, Spacer(1, 6 * mm)])

    story.extend(section_heading("Descrizione della segnalazione", theme, styles))
    story.append(Paragraph(_pdf_safe_text(segnalazione.descrizione, preserve_breaks=True), styles["body"]))
    story.append(Spacer(1, 6 * mm))

    story.extend(section_heading("Allegati", theme, styles))
    if allegati:
        rows: list = [[
            Paragraph("#", styles["table_header"]),
            Paragraph("File", styles["table_header"]),
        ]]
        rows.extend(
            [Paragraph(str(index), styles["cell"]), Paragraph(_pdf_safe_text(allegato.nome_file), styles["cell"])]
            for index, allegato in enumerate(allegati, start=1)
        )
        story.append(
            data_table(rows, theme, col_widths=[14 * mm, content_width - 14 * mm], header=True)
        )
    else:
        story.append(Paragraph("Nessun allegato associato.", styles["body"]))

    draw = header_footer_callback(theme, title="DIARIO PREPOSTO", subtitle=f"Segnalazione {codice}")
    doc.build(story, onFirstPage=draw, onLaterPages=draw)
    buf.seek(0)
    filename = f"segnalazione_{codice.lower()}.pdf"
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# API allegati
# ---------------------------------------------------------------------------

@_write_required
@require_POST
def api_allegato_upload(request):
    pk = request.POST.get("segnalazione_id")
    if not pk:
        return _json_err("segnalazione_id obbligatorio")
    segnalazione = get_object_or_404(SegnalazionePreposto, pk=pk)
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return _json_err("Nessun file ricevuto")
    try:
        safe_name = _validate_diario_allegato(uploaded_file)
    except UploadMimeValidationError as exc:
        return _json_err(str(exc))
    allegato = SegnalazioneAllegato.objects.create(
        segnalazione=segnalazione,
        nome_file=safe_name,
        file=uploaded_file,
    )
    return JsonResponse(
        {
            "ok": True,
            "id": allegato.pk,
            "nome_file": allegato.nome_file,
            "url": reverse("diario_preposto:allegato_download", args=[allegato.pk]),
        }
    )


@_write_required
@require_POST
def api_allegato_delete(request):
    pk = request.POST.get("allegato_id")
    if not pk:
        return _json_err("allegato_id obbligatorio")
    allegato = get_object_or_404(SegnalazioneAllegato, pk=pk)
    try:
        if allegato.file and allegato.file.name:
            allegato.file.storage.delete(allegato.file.name)
    except Exception:
        pass
    allegato.delete()
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Download protetto allegati segnalazioni (sostituisce l'esposizione via /media/)
# ---------------------------------------------------------------------------

@login_required
def allegato_download(request, allegato_id: int):
    """Serve l'allegato di una segnalazione richiedendo l'autenticazione.

    L'autorizzazione segue la stessa visibilita' della scheda dettaglio:
    l'utente deve avere accesso al modulo Diario Preposto (vedi _can_view).
    Il controllo permessi precede il lookup dell'oggetto per non rivelare
    l'esistenza di un allegato a chi non è autorizzato al modulo.
    """
    from core.audit import log_action

    if not _can_view(request):
        log_action(
            request,
            "download_allegato",
            "diario_preposto",
            {
                "allegato_id": allegato_id,
                "esito": "denied",
                "motivo": "permission_denied",
            },
        )
        return render(request, "core/pages/forbidden.html", status=403)

    allegato = get_object_or_404(
        SegnalazioneAllegato.objects.select_related("segnalazione"),
        pk=allegato_id,
    )
    storage = allegato.file.storage
    if (
        not allegato.file
        or not allegato.file.name
        or not storage.exists(allegato.file.name)
    ):
        log_action(
            request,
            "download_allegato",
            "diario_preposto",
            {
                "allegato_id": allegato.id,
                "segnalazione_id": allegato.segnalazione_id,
                "esito": "not_found",
            },
        )
        return HttpResponse("Allegato non trovato.", status=404)

    filename = allegato.nome_file or Path(allegato.file.name).name
    content_type = (
        mimetypes.guess_type(filename)[0] or "application/octet-stream"
    )
    log_action(
        request,
        "download_allegato",
        "diario_preposto",
        {
            "allegato_id": allegato.id,
            "segnalazione_id": allegato.segnalazione_id,
            "filename": filename,
            "esito": "success",
        },
    )
    return FileResponse(
        storage.open(allegato.file.name, "rb"),
        as_attachment=True,
        filename=filename,
        content_type=content_type,
    )


@login_required
def api_cerca_utenti(request):
    """Ricerca dipendenti per il widget autorizzazioni — solo admin legacy.

    Stesso contratto di ``tickets.api_cerca_utenti``: ritorna nome, username
    (aliasusername legacy = fonte unica username) ed email di login aziendale.
    """
    if not _can_manage_settings(request):
        return _json_err("Non autorizzato", 403)

    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    results = []
    try:
        from core.legacy_models import AnagraficaDipendente

        qs = AnagraficaDipendente.objects.filter(
            Q(nome__icontains=q) | Q(cognome__icontains=q) |
            Q(email__icontains=q) | Q(aliasusername__icontains=q)
        ).order_by("cognome", "nome")[:20]

        for a in qs:
            nome_completo = f"{(a.nome or '').strip()} {(a.cognome or '').strip()}".strip()
            username = (a.aliasusername or "").strip()      # fonte unica username
            email_login = (a.email or "").strip()           # UPN / mail aziendale
            results.append({
                "nome": nome_completo,
                "username": username,
                "email": email_login,
                "label": nome_completo + (f" — {username}" if username else ""),
            })
    except Exception:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        qs = User.objects.filter(username__icontains=q).order_by("last_name", "first_name")[:20]
        for u in qs:
            nome = f"{u.first_name} {u.last_name}".strip() or u.username
            results.append({
                "nome": nome,
                "username": u.username,
                "email": u.email or "",
                "label": nome,
            })

    return JsonResponse({"results": results})


@login_required
def impostazioni(request):
    """Impostazioni modulo Diario Preposto — admin o grant canonico."""
    if not _can_manage_settings(request):
        return render(request, "core/pages/forbidden.html", status=403)

    cfg = DiarioPrepostoImpostazioni.objects.filter(pk=1).first()
    if cfg is None:
        cfg = DiarioPrepostoImpostazioni.objects.create(pk=1, acl_scrittura=[])

    if request.method == "POST":
        branding_response = handle_module_branding_post(
            request,
            module_key="diario_preposto",
            redirect_to="diario_preposto:impostazioni",
            audit_module="diario_preposto",
            fallback_label="Diario Preposto",
        )
        if branding_response is not None:
            return branding_response
        raw = request.POST.get("acl_scrittura", "").strip()
        entries = [e.strip() for e in raw.replace(",", "\n").splitlines() if e.strip()]
        raw_frequenza = request.POST.get("ispezione_frequenza_giorni", "").strip()
        try:
            frequenza = int(raw_frequenza)
        except (TypeError, ValueError):
            frequenza = cfg.ispezione_frequenza_giorni
        cfg.ispezione_frequenza_giorni = min(max(frequenza, 1), 365)
        cfg.acl_scrittura = entries
        cfg.save()
        from core.audit import log_action
        log_action(request, "modifica", "diario_preposto", "Aggiornate impostazioni Diario Preposto")
        messages.success(request, "Impostazioni salvate.")
        return redirect("diario_preposto:impostazioni")

    acl_text = "\n".join(cfg.acl_scrittura) if cfg.acl_scrittura else ""
    return render(request, "diario_preposto/pages/impostazioni.html", {
        "cfg": cfg,
        "acl_text": acl_text,
        **get_module_branding_context("diario_preposto", fallback_label="Diario Preposto"),
    })
