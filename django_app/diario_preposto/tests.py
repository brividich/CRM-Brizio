from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import ChecklistEsecuzione, ChecklistRisposta, ChecklistVoce, UserOnboarding

from .models import DiarioPrepostoImpostazioni, SegnalazionePreposto


def _aware_datetime(year: int, month: int, day: int, hour: int = 9, minute: int = 0):
    return timezone.make_aware(
        datetime(year, month, day, hour, minute),
        timezone.get_current_timezone(),
    )


class SegnalazionePrepostoCodiceTests(TestCase):
    def test_assigns_yearly_codes_starting_from_zero(self):
        first_2025 = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 1",
            data_segnalazione=_aware_datetime(2025, 1, 10),
            descrizione="Descrizione 1",
        )
        second_2025 = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 2",
            data_segnalazione=_aware_datetime(2025, 2, 11),
            descrizione="Descrizione 2",
        )
        first_2026 = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 3",
            data_segnalazione=_aware_datetime(2026, 1, 5),
            descrizione="Descrizione 3",
        )

        self.assertEqual(first_2025.codice_identificativo, "DP-2025-0000")
        self.assertEqual(second_2025.codice_identificativo, "DP-2025-0001")
        self.assertEqual(first_2026.codice_identificativo, "DP-2026-0000")

    def test_save_with_update_fields_restores_missing_code(self):
        first = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 1",
            data_segnalazione=_aware_datetime(2025, 1, 10),
            descrizione="Descrizione 1",
        )
        second = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 2",
            data_segnalazione=_aware_datetime(2025, 2, 11),
            descrizione="Descrizione 2",
        )

        SegnalazionePreposto.objects.filter(pk=second.pk).update(codice_identificativo="")
        second.refresh_from_db()
        second.descrizione = "Descrizione aggiornata"
        second.save(update_fields=["descrizione", "updated_at"])

        self.assertEqual(first.codice_identificativo, "DP-2025-0000")
        self.assertEqual(second.codice_identificativo, "DP-2025-0001")

    def test_existing_code_does_not_change_when_record_is_edited(self):
        segnalazione = SegnalazionePreposto.objects.create(
            titolo="Segnalazione 1",
            data_segnalazione=_aware_datetime(2025, 1, 10),
            descrizione="Descrizione 1",
        )

        segnalazione.data_segnalazione = _aware_datetime(2026, 1, 10)
        segnalazione.descrizione = "Descrizione aggiornata"
        segnalazione.save()

        self.assertEqual(segnalazione.codice_identificativo, "DP-2025-0000")

    def test_export_pdf_is_rendered_inline(self):
        user = get_user_model().objects.create_superuser(
            username="tester_pdf",
            email="tester_pdf@example.com",
            password="pwd12345",
        )
        segnalazione = SegnalazionePreposto.objects.create(
            titolo="Segnalazione PDF",
            data_segnalazione=_aware_datetime(2025, 3, 5),
            descrizione="Descrizione PDF",
        )

        self.client.force_login(user)
        response = self.client.get(reverse("diario_preposto:export_pdf", args=[segnalazione.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("inline;", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))


class SegnalazionePrepostoExcelExportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="tester_excel",
            email="tester_excel@example.com",
            password="pwd12345",
        )
        UserOnboarding.objects.create(user=self.user, completed=True, completed_at=timezone.now())
        self.client.force_login(self.user)
        self.matching = SegnalazionePreposto.objects.create(
            titolo="Segnalazione filtro",
            data_segnalazione=_aware_datetime(2026, 5, 4),
            descrizione="Descrizione inclusa",
            preposto="Mario Rossi",
            chi_segnala="Luca Verdi",
            creato_da=self.user,
        )
        SegnalazionePreposto.objects.create(
            titolo="Segnalazione esclusa",
            data_segnalazione=_aware_datetime(2026, 5, 5),
            descrizione="Descrizione esclusa",
            preposto="Anna Bianchi",
            chi_segnala="Giulia Neri",
            creato_da=self.user,
        )

    def test_export_excel_response_headers_workbook_columns_and_filters(self):
        response = self.client.get(reverse("diario_preposto:export_excel"), {"preposto": "Mario"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        expected_filename = f"diario_preposto_{timezone.now().strftime('%Y%m%d')}.xlsx"
        self.assertIn(expected_filename, response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        try:
            sheet = workbook["Diario Preposto"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                [str(header).lower() for header in headers],
                [
                    "codice identificativo",
                    "data segnalazione",
                    "titolo",
                    "descrizione",
                    "preposto",
                    "chi segnala",
                    "creato da",
                    "numero allegati",
                    "created_at",
                    "updated_at",
                ],
            )
            exported_titles = [row[2] for row in sheet.iter_rows(min_row=2, values_only=True)]
            self.assertEqual(exported_titles, [self.matching.titolo])
        finally:
            workbook.close()

    def test_export_excel_requires_login(self):
        self.client.logout()
        response = self.client.get(reverse("diario_preposto:export_excel"))

        self.assertEqual(response.status_code, 302)


class IspezioniPrepostoTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="preposto",
            email="preposto@example.com",
            password="pwd12345",
        )
        UserOnboarding.objects.create(user=self.user, completed=True, completed_at=timezone.now())
        self.client.force_login(self.user)

    def test_ispezioni_page_creates_default_template_voci(self):
        DiarioPrepostoImpostazioni.objects.create(
            pk=1,
            acl_scrittura=[],
            ispezione_frequenza_giorni=14,
        )

        response = self.client.get(reverse("diario_preposto:ispezioni"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ispezioni periodiche")
        self.assertEqual(
            ChecklistVoce.objects.filter(tipo_checklist="preposto_ispezione", is_active=True).count(),
            4,
        )
        self.assertContains(response, "14 gg")

    def test_ispezione_nuova_creates_checklist_execution_and_answers(self):
        self.client.get(reverse("diario_preposto:ispezioni"))
        voci = list(ChecklistVoce.objects.filter(tipo_checklist="preposto_ispezione"))
        data = {
            "area": "CNC",
            "macchina": "Tornio 1",
            "note": "Percorsi liberi.",
        }
        for voce in voci:
            if voce.tipo_campo == "check":
                data[f"voce_{voce.pk}"] = "1"
            elif voce.tipo_campo == "testo":
                data[f"voce_{voce.pk}"] = "Nessuna anomalia"

        response = self.client.post(reverse("diario_preposto:ispezione_nuova"), data)

        self.assertRedirects(response, reverse("diario_preposto:ispezioni"))
        esecuzione = ChecklistEsecuzione.objects.get(tipo_checklist="preposto_ispezione")
        metadata = json.loads(esecuzione.note)
        self.assertEqual(metadata["area"], "CNC")
        self.assertEqual(metadata["macchina"], "Tornio 1")
        self.assertEqual(ChecklistRisposta.objects.filter(esecuzione=esecuzione).count(), len(voci))


class SegnalazioneAllegatoDownloadTests(TestCase):
    """Verifica che gli allegati delle segnalazioni siano accessibili
    solo agli utenti autenticati (e non via URL pubblico /media/)."""

    def setUp(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from .models import SegnalazioneAllegato

        self.user = get_user_model().objects.create_user(
            username="dp_user",
            email="dp_user@example.com",
            password="pwd12345",
        )
        # Crea onboarding completato per evitare redirect
        from core.models import UserOnboarding, Profile
        UserOnboarding.objects.create(user=self.user, completed=True, skipped=True)
        # Crea legacy_user per ACL
        from core.legacy_models import UtenteLegacy, Ruolo
        role = Ruolo.objects.create(id=999, nome="Test Role")
        legacy_user = UtenteLegacy.objects.create(
            id=999,
            nome="Test User",
            email="dp_user@example.com",
            password="hash",
            ruolo=role.nome,
            ruolo_id=role.id,
            attivo=True,
        )
        # Crea Profile per collegare l'utente Django al legacy_user
        Profile.objects.create(user=self.user, legacy_user_id=legacy_user.id, legacy_ruolo_id=role.id, legacy_ruolo=role.nome)
        self.segnalazione = SegnalazionePreposto.objects.create(
            titolo="Segn allegato",
            data_segnalazione=_aware_datetime(2026, 1, 10),
            descrizione="Test allegato",
        )
        self.allegato = SegnalazioneAllegato.objects.create(
            segnalazione=self.segnalazione,
            nome_file="prova.txt",
            file=SimpleUploadedFile("prova.txt", b"contenuto riservato", content_type="text/plain"),
        )

    def tearDown(self):
        # Rimuove il file temporaneo creato dal test dallo storage privato.
        try:
            if self.allegato.file and self.allegato.file.name:
                self.allegato.file.storage.delete(self.allegato.file.name)
        except Exception:
            pass

    def test_allegato_download_requires_login(self):
        url = reverse("diario_preposto:allegato_download", args=[self.allegato.pk])
        response = self.client.get(url)
        # Utente anonimo: redirect verso login (non 200, non file).
        self.assertEqual(response.status_code, 302)
        self.assertNotIn("contenuto riservato", response.content.decode(errors="ignore"))

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": True})
    def test_allegato_download_authenticated_returns_file(self, _mock):
        # L'utente ha accesso al modulo (decisione ACL allow): scarica il file.
        # Dopo il fix SEC-AUDIT-002 il download segue l'ACL di modulo, non più il
        # fail-open di _can_write, quindi l'autorizzazione va resa esplicita.
        self.client.force_login(self.user)
        url = reverse("diario_preposto:allegato_download", args=[self.allegato.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        body = b"".join(response.streaming_content) if response.streaming else response.content
        self.assertEqual(body, b"contenuto riservato")

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": True})
    def test_allegato_download_authenticated_creates_audit_log(self, _mock):
        from core.models import AuditLog

        self.client.force_login(self.user)
        # Conta solo i log dell'azione specifica: il conteggio globale è fragile
        # (la richiesta può generare altri audit di piattaforma indipendenti).
        download_logs = lambda: AuditLog.objects.filter(
            azione="download_allegato", modulo="diario_preposto",
        )
        before = download_logs().count()
        url = reverse("diario_preposto:allegato_download", args=[self.allegato.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        created = download_logs().order_by("-id").first()
        self.assertIsNotNone(created)
        self.assertEqual(download_logs().count(), before + 1)
        payload = created.dettaglio or {}
        self.assertEqual(payload.get("esito"), "success")
        self.assertEqual(payload.get("allegato_id"), self.allegato.id)
        self.assertEqual(payload.get("segnalazione_id"), self.segnalazione.id)
        # Nessun contenuto file nel payload audit
        serialized = repr(payload)
        self.assertNotIn("contenuto riservato", serialized)


@override_settings(LEGACY_AUTH_ENABLED=True, SECURE_SSL_REDIRECT=False)
class AllegatoDownloadModuleAuthTests(TestCase):
    """SEC-AUDIT-002: l'URL /diario-preposto/allegato/ è registrato in
    _ACL_SHARED_PREFIXES (core/middleware.py) e quindi bypassa il middleware
    ACL. La vista allegato_download deve perciò verificare da sé l'accesso al
    modulo (helper _can_view), riusando la decisione ACL della scheda dettaglio.
    """

    def setUp(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        from core.models import UserOnboarding

        from .models import DiarioPrepostoImpostazioni, SegnalazioneAllegato

        User = get_user_model()
        # acl_scrittura non vuota che esclude gli utenti di test: _can_write False,
        # così il gate dipende solo dalla decisione ACL di modulo.
        DiarioPrepostoImpostazioni.objects.create(acl_scrittura=["altro-utente"])

        self.user = User.objects.create_user(
            username="dp_reader", email="dp_reader@example.com",
        )
        UserOnboarding.objects.create(user=self.user, completed=True, skipped=True)
        self.superuser = User.objects.create_superuser(
            username="dp_admin", email="dp_admin@example.com", password="pwd12345",
        )

        self.segnalazione = SegnalazionePreposto.objects.create(
            titolo="Segn riservata",
            data_segnalazione=_aware_datetime(2026, 1, 10),
            descrizione="Test SEC-AUDIT-002",
        )
        self.allegato = SegnalazioneAllegato.objects.create(
            segnalazione=self.segnalazione,
            nome_file="riservato.txt",
            file=SimpleUploadedFile(
                "riservato.txt", b"dati riservati", content_type="text/plain"
            ),
        )

    def tearDown(self):
        try:
            if self.allegato.file and self.allegato.file.name:
                self.allegato.file.storage.delete(self.allegato.file.name)
        except Exception:
            pass

    def _url(self, pk=None):
        return reverse("diario_preposto:allegato_download", args=[pk or self.allegato.pk])

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": True})
    def test_authorized_module_user_can_download(self, _mock):
        """1. Utente con accesso al modulo (decisione ACL allow) scarica l'allegato."""
        self.client.force_login(self.user)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 200)

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": False})
    def test_unauthorized_module_user_forbidden(self, _mock):
        """2. Utente senza accesso al modulo riceve 403, anche su id inesistente."""
        self.client.force_login(self.user)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 403)
        # 403 anche su id inesistente: il controllo permessi precede il lookup,
        # così non si rivela l'esistenza dell'allegato a chi non è autorizzato.
        resp_missing = self.client.get(self._url(pk=999999))
        self.assertEqual(resp_missing.status_code, 403)

    def test_superuser_can_download(self):
        """3. Il superuser accede sempre (nessuna dipendenza dalla decisione ACL)."""
        self.client.force_login(self.superuser)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 200)

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": True})
    def test_missing_attachment_returns_404_for_authorized_user(self, _mock):
        """4. Per un utente autorizzato, un allegato inesistente resta 404."""
        self.client.force_login(self.user)
        resp = self.client.get(self._url(pk=999999))
        self.assertEqual(resp.status_code, 404)

    @patch("core.acl_v2.resolve_acl_access", return_value={"allowed": False})
    def test_empty_write_acl_does_not_grant_download(self, _mock):
        """5. Regressione fail-open: con acl_scrittura VUOTA (default) _can_write
        è aperto a qualsiasi autenticato. Il download NON deve comunque essere
        concesso a chi non ha accesso al modulo (la vista segue solo l'ACL di
        modulo, non la scorciatoia di scrittura)."""
        DiarioPrepostoImpostazioni.objects.update(acl_scrittura=[])
        self.client.force_login(self.user)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 403)
