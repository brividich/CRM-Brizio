from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
from io import BytesIO
from datetime import datetime, timedelta, timezone as dt_timezone
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Case, When, Value, IntegerField, Count, OuterRef, Subquery
from django.contrib.auth.views import redirect_to_login
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.timezone import now as tz_now
from django.views.decorators.http import require_POST
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT

from core.pdf import PdfTheme, draw_canvas_footer, draw_canvas_header

from core.legacy_utils import get_legacy_user, is_legacy_admin
from core.audit import log_action
from core.upload_mime import UploadMimeValidationError, validate_extension_and_mime

from .models import (
    CATEGORIE_IT,
    CATEGORIE_MAN,
    CategoriaTicket,
    EsitoIntervento,
    PrioritaTicket,
    StatoTicket,
    Ticket,
    TicketAllegato,
    TicketCommento,
    TicketImpostazioni,
    TicketIntervento,
    TicketStatoLog,
    TipoFermo,
    TipoTicket,
    get_categorie,
)

TICKET_ALLOWED_UPLOAD_EXTENSIONS = {
    ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".bmp",
    ".doc", ".docx", ".xls", ".xlsx", ".odt", ".ods",
    ".txt", ".csv", ".msg", ".eml", ".zip", ".7z",
}
TICKET_ALLOWED_UPLOAD_MIMES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/gif",
    "image/bmp",
    "image/x-ms-bmp",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.oasis.opendocument.text",
    "application/vnd.oasis.opendocument.spreadsheet",
    "text/plain",
    "text/csv",
    "application/csv",
    "application/vnd.ms-outlook",
    "message/rfc822",
    "application/zip",
    "application/x-zip-compressed",
    "application/x-7z-compressed",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _legacy_identity(request) -> tuple[str, str, int | None]:
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if legacy_user:
        name  = (legacy_user.nome or "").strip() or request.user.get_full_name() or request.user.get_username()
        email = (legacy_user.email or "").strip().lower() or (request.user.email or "").strip().lower()
        return name, email, getattr(legacy_user, "id", None)
    name  = request.user.get_full_name() or request.user.get_username()
    email = (request.user.email or "").strip().lower()
    return name, email, None


def _build_ticket_activity_feed(ticket: Ticket) -> list[dict[str, object]]:
    """Costruisce una timeline operativa unificata per il dettaglio ticket."""
    feed: list[dict[str, object]] = []
    fallback_dt = datetime.min.replace(tzinfo=dt_timezone.utc)

    opened_meta = [ticket.label_tipo, ticket.label_categoria, ticket.label_priorita]
    if ticket.incide_sicurezza:
        opened_meta.append("Sicurezza")

    feed.append({
        "kind": "opened",
        "created_at": ticket.created_at,
        "actor": ticket.richiedente_nome or "Sistema",
        "title": "Ticket aperto",
        "summary": " | ".join(part for part in opened_meta if part),
        "body": "",
        "meta": "",
        "badge": "Apertura",
        "is_internal": False,
        "sort_at": ticket.created_at or fallback_dt,
    })

    for commento in ticket.commenti.all():
        feed.append({
            "kind": "comment",
            "created_at": commento.created_at,
            "actor": commento.autore_nome or "Utente",
            "title": "Nota interna" if commento.is_interno else "Commento",
            "summary": "",
            "body": commento.testo or "",
            "meta": "",
            "badge": "Interna" if commento.is_interno else "Commento",
            "is_internal": bool(commento.is_interno),
            "sort_at": commento.created_at or fallback_dt,
        })

    for log in ticket.stato_log.all():
        summary = f"{log.label_da} -> {log.label_a}" if log.da_stato else log.label_a
        feed.append({
            "kind": "state",
            "created_at": log.created_at,
            "actor": log.utente_nome or "Utente",
            "title": "Stato aggiornato",
            "summary": summary,
            "body": log.nota or "",
            "meta": "",
            "badge": log.label_a,
            "is_internal": False,
            "sort_at": log.created_at or fallback_dt,
        })

    for intervento in ticket.interventi.all():
        detail_lines: list[str] = []
        if intervento.componente_interessato:
            detail_lines.append(f"Componente: {intervento.componente_interessato}")
        if intervento.descrizione_lavoro:
            detail_lines.append(f"Lavoro: {intervento.descrizione_lavoro}")
        if intervento.azioni_svolte:
            detail_lines.append(f"Azioni: {intervento.azioni_svolte}")

        meta_parts = []
        if intervento.data_inizio:
            window = intervento.data_inizio.strftime("%d-%m-%Y %H:%M")
            if intervento.data_fine:
                window = f"{window} -> {intervento.data_fine.strftime('%d-%m-%Y %H:%M')}"
            meta_parts.append(window)
        if intervento.durata_ore is not None:
            meta_parts.append(f"{intervento.durata_ore:g}h")

        feed.append({
            "kind": "intervento",
            "created_at": intervento.data_inizio or intervento.created_at,
            "actor": intervento.tecnico_nome or "Tecnico",
            "title": "Intervento registrato",
            "summary": intervento.label_esito,
            "body": "\n".join(detail_lines),
            "meta": " | ".join(meta_parts),
            "badge": intervento.label_esito,
            "is_internal": False,
            "sort_at": intervento.data_inizio or intervento.created_at or fallback_dt,
        })

    feed.sort(key=lambda item: item.get("sort_at") or fallback_dt, reverse=True)
    for item in feed:
        item.pop("sort_at", None)
    return feed


def _user_acl_identities(request) -> set[str]:
    """Raccoglie tutte le identità (lowercase) con cui l'utente può essere
    stato registrato in una ACL list dei ticket.

    Le liste ``acl_apertura``/``acl_gestione`` possono contenere — a seconda di
    cosa salva il widget impostazioni — l'username Django (UPN), l'email di
    login aziendale o l'``aliasusername`` legacy. Per evitare 403 da mismatch
    (es. ``a.astarita`` vs ``a.astarita@dominio``) il match considera tutte
    queste forme, incluso il prefisso dell'UPN. NON si considera
    ``email_notifica`` (mail privata): l'identità valida è solo mail aziendale
    o username.
    """
    identities: set[str] = set()

    def _add(value) -> None:
        text = str(value or "").strip().lower()
        if text:
            identities.add(text)
            if "@" in text:
                identities.add(text.split("@", 1)[0])  # prefisso UPN

    _add(request.user.get_username())
    _add(request.user.email)

    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    legacy_id = getattr(legacy_user, "id", None) if legacy_user else None
    if legacy_id:
        try:
            from core.legacy_models import AnagraficaDipendente
            ana = (
                AnagraficaDipendente.objects
                .filter(utente_id=legacy_id)
                .only("aliasusername", "email")
                .first()
            )
            if ana:
                _add(ana.aliasusername)   # username legacy
                _add(ana.email)           # mail aziendale / UPN
        except Exception:
            pass  # in caso di errore DB usiamo solo le identità Django
    return identities


def _acl_list_matches(acl: list, identities: set[str]) -> bool:
    """True se una qualsiasi voce della ACL list combacia con un'identità utente."""
    for raw in acl or []:
        value = str(raw or "").strip().lower()
        if not value:
            continue
        if value in identities:
            return True
        if "@" in value and value.split("@", 1)[0] in identities:
            return True
    return False


def _can_open_tickets(request, tipo: str) -> bool:
    """Controlla se l'utente può aprire ticket del tipo dato."""
    if not request.user.is_authenticated:
        return False
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if legacy_user and is_legacy_admin(legacy_user):
        return True
    cfg = TicketImpostazioni.objects.filter(tipo=tipo).first()
    if not cfg:
        return True  # no config = open to all
    acl = cfg.acl_apertura or []
    if not acl:
        return True  # empty acl = open to all
    return _acl_list_matches(acl, _user_acl_identities(request))


def _can_manage_tickets(request, tipo: str | None = None) -> bool:
    """Controlla se l'utente è gestore ticket (IT o MAN o entrambi)."""
    if not request.user.is_authenticated:
        return False
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if legacy_user and is_legacy_admin(legacy_user):
        return True
    tipi = [tipo] if tipo else [TipoTicket.IT, TipoTicket.MAN]
    identities = _user_acl_identities(request)
    for t in tipi:
        cfg = TicketImpostazioni.objects.filter(tipo=t).first()
        if not cfg:
            continue
        if _acl_list_matches(cfg.acl_gestione or [], identities):
            return True
    return False


def _tickets_gestione_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        if not _can_manage_tickets(request):
            return render(request, "core/pages/forbidden.html", status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped


def _json_err(msg: str, status: int = 400) -> JsonResponse:
    return JsonResponse({"ok": False, "error": msg}, status=status)


def _int_or_none(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _get_user_reparto(legacy_id) -> str:
    """Ritorna il reparto dell'utente legacy (da UserExtraInfo o AnagraficaDipendente)."""
    if not legacy_id:
        return ""
    try:
        from core.models import UserExtraInfo
        extra = UserExtraInfo.objects.filter(legacy_user_id=legacy_id).only("reparto").first()
        if extra and extra.reparto:
            return extra.reparto.strip()
        from anagrafica.models import AnagraficaDipendente
        ana = AnagraficaDipendente.objects.filter(utente_id=legacy_id).only("reparto").first()
        if ana and ana.reparto:
            return ana.reparto.strip()
    except Exception:
        pass
    return ""


def _get_my_assets(legacy_id, name: str, reparto: str, tipo: str) -> tuple[list[dict], str]:
    """
    Ritorna (lista_asset, label) in base al tipo ticket.
    - IT:  asset assegnati all'utente; fallback → asset IT-type (PC/NB/Stampante/HW)
    - MAN: asset del reparto; fallback → tutte le macchine CNC/WORK_MACHINE attive
    Ritorna label vuota se la lista è vuota.
    """
    try:
        from assets.models import Asset as AssetModel

        def _serialize(qs):
            return [
                {
                    "id": a.id,
                    "name": a.name,
                    "asset_tag": a.asset_tag,
                    "asset_type_label": a.get_asset_type_display(),
                    "asset_category": a.asset_category.label if a.asset_category_id else "",
                    "model": a.model or "",
                }
                for a in qs.select_related("asset_category")
                          .filter(status__in=["IN_USE", "IN_STOCK"])
                          .order_by("name")[:50]
            ]

        if tipo == TipoTicket.IT:
            # 1° tentativo: asset assegnati all'utente
            f = Q()
            if legacy_id:
                f |= Q(assigned_legacy_user_id=legacy_id)
            if name:
                f |= Q(assignment_to__iexact=name)
            if f:
                items = _serialize(AssetModel.objects.filter(f))
                if items:
                    return items, "Miei asset"
            # Fallback: asset IT per tipo
            items = _serialize(AssetModel.objects.filter(
                asset_type__in=["PC", "NOTEBOOK", "STAMPANTE", "HW", "SERVER", "FIREWALL"]
            ))
            return items, "Asset IT" if items else ""

        elif tipo == TipoTicket.MAN:
            # 1° tentativo: asset del reparto utente
            if reparto:
                items = _serialize(AssetModel.objects.filter(
                    Q(reparto__iexact=reparto) | Q(assignment_reparto__iexact=reparto)
                ))
                if items:
                    return items, f"Asset reparto «{reparto}»"
            # Fallback: tutte le macchine/impianti
            items = _serialize(AssetModel.objects.filter(
                asset_type__in=["CNC", "WORK_MACHINE"]
            ))
            return items, "Macchine e impianti" if items else ""

    except Exception:
        pass
    return [], ""


def _get_assets_for_select() -> list[dict]:
    """Carica asset attivi per il datalist nel form."""
    try:
        from assets.models import Asset
        assets = (
            Asset.objects.filter(status__in=["IN_USE", "IN_STOCK"])
            .select_related("asset_category")
            .order_by("name")[:500]
        )
        return [
            {
                "id": asset.id,
                "name": asset.name,
                "asset_tag": asset.asset_tag,
                "asset_type": asset.asset_type,
                "asset_type_label": asset.get_asset_type_display(),
                "asset_category": asset.asset_category.label if asset.asset_category_id else "",
                "manufacturer": asset.manufacturer or "",
                "model": asset.model or "",
                "serial_number": asset.serial_number or "",
                "reparto": asset.reparto or "",
            }
            for asset in assets
        ]
    except Exception:
        return []


def _get_fornitori_for_select() -> list[dict]:
    """Carica fornitori attivi per la delega."""
    try:
        from anagrafica.models import Fornitore
        return list(
            Fornitore.objects.filter(is_active=True)
            .values("id", "ragione_sociale")
            .order_by("ragione_sociale")[:200]
        )
    except Exception:
        return []


def _ticket_form_context(tipo: str = "", error: str = "", form_data=None,
                         my_assets_list=None, my_assets_label: str = "") -> dict:
    return {
        "error": error,
        "tipo": tipo,
        "categorie_it": get_categorie(TipoTicket.IT),
        "categorie_man": get_categorie(TipoTicket.MAN),
        "assets_list": _get_assets_for_select(),
        "priorita_list": PrioritaTicket.choices,
        "tipi": TipoTicket.choices,
        "form_data": form_data,
        "my_assets_list": my_assets_list or [],
        "my_assets_label": my_assets_label,
        "include_in_maintenance_register": True if tipo == TipoTicket.MAN else False,
    }


def _ticket_access_flags(request, ticket: Ticket) -> dict:
    name, email, legacy_id = _legacy_identity(request)
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin = bool(legacy_user and is_legacy_admin(legacy_user))
    is_gestore = _can_manage_tickets(request, ticket.tipo)
    is_richiedente = (
        ticket.richiedente_user_id == request.user.id
        or (legacy_id and ticket.richiedente_legacy_user_id == legacy_id)
        or ticket.richiedente_nome == name
        or ticket.richiedente_email.lower() == email.lower()
    )
    return {
        "name": name,
        "email": email,
        "legacy_id": legacy_id,
        "is_admin": is_admin,
        "is_gestore": is_gestore,
        "is_richiedente": is_richiedente,
    }


def _api_require_ticket_access(request, ticket: Ticket):
    """Guard object-level per le API di gestione ticket.

    Il decoratore ``_tickets_gestione_required`` verifica solo che l'utente
    sia gestore di *almeno un* tipo di ticket; non garantisce l'accesso allo
    specifico ticket collegato all'oggetto manipolato. Questa funzione
    richiede che l'utente sia gestore del tipo del ticket (o admin legacy).

    Ritorna ``JsonResponse`` 403 se l'accesso è negato, altrimenti ``None``.
    """
    access = _ticket_access_flags(request, ticket)
    if access["is_gestore"] or access["is_admin"]:
        return None
    return _json_err("Accesso non consentito a questo ticket.", status=403)


# ── Costanti PDF ──────────────────────────────────────────────────────────────
_PDF_MARGIN   = 16 * mm
_PDF_COLOR_IT  = HexColor("#0369a1")
_PDF_COLOR_MAN = HexColor("#15803d")
_PDF_COLOR_DARK = HexColor("#0f172a")
_PDF_COLOR_GRAY = HexColor("#64748b")
_PDF_COLOR_LIGHT = HexColor("#f8fafc")
_PDF_COLOR_BORDER = HexColor("#e2e8f0")
_PDF_COLOR_URGENTE = HexColor("#dc2626")
_PDF_COLOR_ALTA    = HexColor("#d97706")
_PDF_FOOTER_H = 12 * mm


def _pdf_accent(ticket: Ticket) -> HexColor:
    return _PDF_COLOR_MAN if ticket.tipo == "MAN" else _PDF_COLOR_IT


def _pdf_wrap(text: str, max_chars: int) -> list[str]:
    """Spezza testo in righe di max_chars caratteri, rispettando le newline."""
    result = []
    for raw in (text or "-").splitlines():
        line = raw.strip() or " "
        while len(line) > max_chars:
            split_at = line.rfind(" ", 0, max_chars)
            if split_at < 10:
                split_at = max_chars
            result.append(line[:split_at].rstrip())
            line = line[split_at:].lstrip()
        result.append(line)
    return result or ["-"]


def _pdf_theme(pdf: canvas.Canvas) -> PdfTheme:
    """Tema PDF del portale, calcolato una volta e memorizzato sul canvas."""
    theme = getattr(pdf, "_hub_theme", None)
    if theme is None:
        theme = PdfTheme.from_branding()
        pdf._hub_theme = theme
    return theme


def _pdf_footer(pdf: canvas.Canvas, ticket: Ticket, page_num: int, page_width: float) -> None:
    """Footer standard del portale (nome portale + numero pagina)."""
    draw_canvas_footer(
        pdf,
        theme=_pdf_theme(pdf),
        page_width=page_width,
        left_margin=_PDF_MARGIN,
        right_margin=_PDF_MARGIN,
        baseline_y=_PDF_FOOTER_H - 4 * mm,
        page_number=page_num,
    )


def _pdf_new_page(pdf: canvas.Canvas, ticket: Ticket, page_num: int, page_width: float, page_height: float) -> tuple[float, int]:
    """Chiude la pagina corrente e ne apre una nuova, ritorna (y, nuovo_page_num)."""
    _pdf_footer(pdf, ticket, page_num, page_width)
    pdf.showPage()
    page_num += 1
    # Header di continuazione standard (logo/monogramma + numero ticket).
    content_y = draw_canvas_header(
        pdf,
        theme=_pdf_theme(pdf),
        page_width=page_width,
        page_height=page_height,
        left_margin=_PDF_MARGIN,
        right_margin=_PDF_MARGIN,
        title=ticket.numero_ticket,
        subtitle=(ticket.titolo or "")[:70],
    )
    return content_y, page_num


def _pdf_check_space(pdf, ticket, y, needed, page_num, page_width, page_height):
    if y - needed < _PDF_FOOTER_H + 8 * mm:
        y, page_num = _pdf_new_page(pdf, ticket, page_num, page_width, page_height)
    return y, page_num


def _pdf_section_title(pdf: canvas.Canvas, title: str, accent: HexColor, x: float, y: float, width: float) -> float:
    """Intestazione sezione con band colorata."""
    h = 7 * mm
    pdf.setFillColor(accent)
    pdf.rect(x, y - h, width, h, fill=1, stroke=0)
    pdf.setFillColor(white)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(x + 3 * mm, y - h + 2 * mm, title.upper())
    return y - h - 2 * mm


def _pdf_info_grid(pdf, items_left, items_right, x, y, width, accent) -> float:
    """Griglia 2-colonne per metadati."""
    col_w = (width - 4 * mm) / 2
    row_h = 7.5 * mm
    rows = max(len(items_left), len(items_right))
    for i in range(rows):
        ry = y - (i + 1) * row_h
        for col_idx, items in enumerate((items_left, items_right)):
            cx = x + col_idx * (col_w + 4 * mm)
            if i < len(items):
                lbl, val = items[i]
                pdf.setFillColor(_PDF_COLOR_LIGHT)
                pdf.setStrokeColor(_PDF_COLOR_BORDER)
                pdf.setLineWidth(0.4)
                pdf.rect(cx, ry, col_w, row_h - 1 * mm, fill=1, stroke=1)
                pdf.setFillColor(accent)
                pdf.setFont("Helvetica-Bold", 7)
                pdf.drawString(cx + 2 * mm, ry + row_h - 3.5 * mm, lbl.upper())
                pdf.setFillColor(_PDF_COLOR_DARK)
                pdf.setFont("Helvetica", 8.5)
                pdf.drawString(cx + 2 * mm, ry + 1.5 * mm, str(val)[:45])
    return y - rows * row_h - 3 * mm


def _pdf_text_block(pdf, ticket, text, x, y, width, page_num, page_width, page_height, font_size=9) -> tuple[float, int]:
    """Testo multiriga con gestione automatica di pagina."""
    max_chars = max(50, int(width / (font_size * 0.52)))
    lines = _pdf_wrap(text, max_chars)
    line_h = font_size * 0.45 * mm + 1.5 * mm
    pdf.setFont("Helvetica", font_size)
    pdf.setFillColor(_PDF_COLOR_DARK)
    for line in lines:
        y, page_num = _pdf_check_space(pdf, ticket, y, line_h + 2 * mm, page_num, page_width, page_height)
        pdf.drawString(x, y, line)
        y -= line_h
    return y, page_num


def _ticket_pdf_response(ticket: Ticket, *, commenti, allegati, include_internal: bool) -> HttpResponse:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    page_width, page_height = A4
    pdf.setTitle(f"Rapporto Intervento — {ticket.numero_ticket}")
    pdf.setAuthor("Portale Applicativo — Costruzioni Novicrom S.r.l.")
    pdf.setSubject(f"Ticket {ticket.numero_ticket}")

    accent   = _pdf_accent(ticket)
    mx       = _PDF_MARGIN
    bw       = page_width - 2 * mx
    page_num = 1
    tipo_label = "RAPPORTO DI INTERVENTO IT" if ticket.tipo == "IT" else "RAPPORTO DI MANUTENZIONE"

    # ── HEADER PAGINA 1 (template standard del portale) ───────────────────────
    y = draw_canvas_header(
        pdf,
        theme=_pdf_theme(pdf),
        page_width=page_width,
        page_height=page_height,
        left_margin=mx,
        right_margin=mx,
        title=tipo_label,
        subtitle="Costruzioni Novicrom S.r.l. · Portale Applicativo",
        right_title=ticket.numero_ticket,
        right_subtitle="Rapporto intervento",
    )
    y -= 3 * mm

    # Titolo ticket in grande (corpo)
    pdf.setFillColor(_PDF_COLOR_DARK)
    pdf.setFont("Helvetica-Bold", 14)
    # wrap titolo su più righe se lungo
    title_lines = _pdf_wrap(ticket.titolo, 80)
    for tl in title_lines[:2]:
        pdf.drawString(mx, y, tl)
        y -= 6 * mm

    # Badge sicurezza se attivo
    if ticket.incide_sicurezza:
        y -= 1 * mm
        pdf.setFillColor(HexColor("#fef2f2"))
        pdf.setStrokeColor(_PDF_COLOR_URGENTE)
        pdf.setLineWidth(0.8)
        pdf.roundRect(mx, y - 6 * mm, 70 * mm, 6 * mm, 2 * mm, fill=1, stroke=1)
        pdf.setFillColor(_PDF_COLOR_URGENTE)
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(mx + 3 * mm, y - 4.2 * mm, "⚠  PROBLEMA DI SICUREZZA SUL LAVORO — INTERVENTO PRIORITARIO")
        y -= 9 * mm

    y -= 4 * mm

    # ── GRIGLIA METADATI ──────────────────────────────────────────────────────
    y = _pdf_section_title(pdf, "Dati del ticket", accent, mx, y, bw)
    y -= 1 * mm

    asset_label = "-"
    if ticket.asset_id and ticket.asset:
        asset_label = f"{ticket.asset.name}  [{ticket.asset.asset_tag}]"
    elif ticket.asset_descrizione_libera:
        asset_label = ticket.asset_descrizione_libera

    prio_color_map = {"URGENTE": "🔴 Urgente", "ALTA": "🟠 Alta", "MEDIA": "🔵 Media", "BASSA": "🟢 Bassa"}
    prio_label = prio_color_map.get(ticket.priorita, ticket.label_priorita)

    items_left = [
        ("Tipo",         ticket.label_tipo),
        ("Categoria",    ticket.label_categoria),
        ("Priorità",     prio_label),
        ("Stato",        ticket.label_stato),
        ("Imp. sicurezza", "Sì" if ticket.incide_sicurezza else "No"),
    ]
    items_right = [
        ("Data apertura",  ticket.created_at.strftime("%d-%m-%Y %H:%M") if ticket.created_at else "-"),
        ("Data chiusura",  ticket.closed_at.strftime("%d-%m-%Y %H:%M") if ticket.closed_at else "-"),
        ("Richiedente",    ticket.richiedente_nome or "-"),
        ("Assegnato a",    ticket.assegnato_a or "-"),
        ("Asset / Macch.", asset_label[:45]),
    ]
    y = _pdf_info_grid(pdf, items_left, items_right, mx, y, bw, accent)
    y -= 5 * mm

    # ── DESCRIZIONE ───────────────────────────────────────────────────────────
    y, page_num = _pdf_check_space(pdf, ticket, y, 20 * mm, page_num, page_width, page_height)
    y = _pdf_section_title(pdf, "Descrizione del problema", accent, mx, y, bw)
    y -= 2 * mm
    # Box sfondo
    desc_lines = _pdf_wrap(ticket.descrizione, 105)
    box_h = len(desc_lines) * 4.8 * mm + 6 * mm
    pdf.setFillColor(_PDF_COLOR_LIGHT)
    pdf.setStrokeColor(_PDF_COLOR_BORDER)
    pdf.setLineWidth(0.4)
    pdf.rect(mx, y - box_h, bw, box_h, fill=1, stroke=1)
    txt_y = y - 4 * mm
    pdf.setFont("Helvetica", 9)
    pdf.setFillColor(_PDF_COLOR_DARK)
    for dl in desc_lines:
        txt_y, page_num = _pdf_check_space(pdf, ticket, txt_y, 5 * mm, page_num, page_width, page_height)
        pdf.drawString(mx + 3 * mm, txt_y, dl)
        txt_y -= 4.8 * mm
    y = txt_y - 4 * mm

    # ── ALLEGATI ──────────────────────────────────────────────────────────────
    if allegati:
        y, page_num = _pdf_check_space(pdf, ticket, y, 14 * mm, page_num, page_width, page_height)
        y = _pdf_section_title(pdf, f"Allegati ({len(allegati)})", accent, mx, y, bw)
        y -= 2 * mm
        pdf.setFont("Helvetica", 8.5)
        pdf.setFillColor(_PDF_COLOR_DARK)
        for a in allegati:
            y, page_num = _pdf_check_space(pdf, ticket, y, 5 * mm, page_num, page_width, page_height)
            pdf.drawString(mx + 3 * mm, y, f"•  {a.nome_originale}")
            y -= 4.5 * mm
        y -= 3 * mm

    # ── STORICO ATTIVITÀ ──────────────────────────────────────────────────────
    visible_comments = [c for c in commenti if include_internal or not c.is_interno]
    if visible_comments:
        y, page_num = _pdf_check_space(pdf, ticket, y, 14 * mm, page_num, page_width, page_height)
        y = _pdf_section_title(pdf, f"Storico attività ({len(visible_comments)} voci)", accent, mx, y, bw)
        y -= 3 * mm

        for c in visible_comments:
            is_int = include_internal and c.is_interno
            bg_col = HexColor("#fefce8") if is_int else HexColor("#f0f9ff")
            bord_col = HexColor("#fde68a") if is_int else HexColor("#bae6fd")
            c_lines = _pdf_wrap(c.testo, 95)
            c_h = len(c_lines) * 4.5 * mm + 9 * mm

            y, page_num = _pdf_check_space(pdf, ticket, y, c_h + 2 * mm, page_num, page_width, page_height)

            # Box commento
            pdf.setFillColor(bg_col)
            pdf.setStrokeColor(bord_col)
            pdf.setLineWidth(0.6)
            pdf.roundRect(mx, y - c_h, bw, c_h, 2 * mm, fill=1, stroke=1)

            # Header commento
            pdf.setFillColor(HexColor("#1e40af") if not is_int else HexColor("#92400e"))
            pdf.setFont("Helvetica-Bold", 8)
            date_str = c.created_at.strftime("%d-%m-%Y %H:%M") if c.created_at else ""
            label_int = "  [NOTA INTERNA]" if is_int else ""
            pdf.drawString(mx + 3 * mm, y - 4.5 * mm, f"{c.autore_nome}{label_int}")
            pdf.setFont("Helvetica", 7.5)
            pdf.setFillColor(_PDF_COLOR_GRAY)
            pdf.drawRightString(mx + bw - 3 * mm, y - 4.5 * mm, date_str)

            # Testo
            pdf.setFont("Helvetica", 8.5)
            pdf.setFillColor(_PDF_COLOR_DARK)
            txt_y2 = y - 9 * mm
            for cl in c_lines:
                pdf.drawString(mx + 3 * mm, txt_y2, cl)
                txt_y2 -= 4.5 * mm

            y -= c_h + 3 * mm
        y -= 3 * mm

    # ── FIRME (solo MAN o se ticket ha assegnato) ─────────────────────────────
    sig_needed = 38 * mm
    y, page_num = _pdf_check_space(pdf, ticket, y, sig_needed, page_num, page_width, page_height)
    y = _pdf_section_title(pdf, "Validazione e firme", accent, mx, y, bw)
    y -= 5 * mm
    col_w3 = (bw - 8 * mm) / 3
    for i, (lbl, name) in enumerate([
        ("Richiedente", ticket.richiedente_nome or ""),
        ("Tecnico / Assegnato", ticket.assegnato_a or ""),
        ("Responsabile", ""),
    ]):
        fx = mx + i * (col_w3 + 4 * mm)
        pdf.setFillColor(_PDF_COLOR_LIGHT)
        pdf.setStrokeColor(_PDF_COLOR_BORDER)
        pdf.setLineWidth(0.4)
        pdf.rect(fx, y - 28 * mm, col_w3, 28 * mm, fill=1, stroke=1)
        pdf.setFillColor(accent)
        pdf.setFont("Helvetica-Bold", 7)
        pdf.drawString(fx + 2 * mm, y - 4 * mm, lbl.upper())
        if name:
            pdf.setFillColor(_PDF_COLOR_DARK)
            pdf.setFont("Helvetica", 8)
            pdf.drawString(fx + 2 * mm, y - 9 * mm, name)
        # linea firma
        pdf.setStrokeColor(_PDF_COLOR_GRAY)
        pdf.setLineWidth(0.5)
        pdf.line(fx + 3 * mm, y - 22 * mm, fx + col_w3 - 3 * mm, y - 22 * mm)
        pdf.setFillColor(_PDF_COLOR_GRAY)
        pdf.setFont("Helvetica", 6.5)
        pdf.drawCentredString(fx + col_w3 / 2, y - 25.5 * mm, "Firma e data")

    # ── FOOTER ULTIMA PAGINA ──────────────────────────────────────────────────
    _pdf_footer(pdf, ticket, page_num, page_width)
    pdf.showPage()
    pdf.save()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{ticket.numero_ticket}.pdf"'
    response.write(buffer.getvalue())
    return response


# ---------------------------------------------------------------------------
# Dashboard utente (miei ticket)
# ---------------------------------------------------------------------------

@login_required
def ticket_dashboard(request):
    name, email, legacy_id = _legacy_identity(request)
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin    = bool(legacy_user and is_legacy_admin(legacy_user))
    is_gestore  = _can_manage_tickets(request)

    qs = Ticket.objects.select_related("asset")
    kpi_qs = Ticket.objects.all()

    # SEC: chi NON è gestore/admin vede solo i ticket di cui è richiedente.
    # Senza questo filtro la dashboard esponeva numero/titolo/categoria/asset e
    # flag "Sicurezza" di TUTTI i ticket a qualsiasi utente autenticato. Si usano
    # gli identificatori stabili (account Django, id legacy, email); il match
    # solo-per-nome è escluso di proposito perché ambiguo tra omonimi.
    if not (is_admin or is_gestore):
        from django.db.models import Q
        own = Q(richiedente_user=request.user)
        if legacy_id:
            own |= Q(richiedente_legacy_user_id=legacy_id)
        if email:
            own |= Q(richiedente_email__iexact=email)
        qs = qs.filter(own)
        kpi_qs = kpi_qs.filter(own)

    # Filtri GET
    tipo_f   = request.GET.get("tipo", "").strip().upper()
    stato_f  = request.GET.get("stato", "").strip().upper()
    prio_f   = request.GET.get("priorita", "").strip().upper()
    cerca_f  = request.GET.get("q", "").strip()

    if tipo_f  in (TipoTicket.IT, TipoTicket.MAN):
        qs = qs.filter(tipo=tipo_f)
    if stato_f in dict(StatoTicket.choices):
        qs = qs.filter(stato=stato_f)
    if prio_f  in dict(PrioritaTicket.choices):
        qs = qs.filter(priorita=prio_f)
    if cerca_f:
        qs = qs.filter(titolo__icontains=cerca_f) | qs.filter(numero_ticket__icontains=cerca_f)

    qs = qs.order_by("-created_at")

    can_open_it  = _can_open_tickets(request, TipoTicket.IT)
    can_open_man = _can_open_tickets(request, TipoTicket.MAN)

    ctx = {
        "tickets":       qs,
        "is_admin":      is_admin,
        "is_gestore":    is_gestore,
        "can_open_it":   can_open_it,
        "can_open_man":  can_open_man,
        "stati":         StatoTicket.choices,
        "priorita_list": PrioritaTicket.choices,
        "tipi":          TipoTicket.choices,
        "filtro_tipo":   tipo_f,
        "filtro_stato":  stato_f,
        "filtro_prio":   prio_f,
        "filtro_cerca":  cerca_f,
        # KPI
        "n_aperte":      kpi_qs.filter(stato=StatoTicket.APERTA).count(),
        "n_urgenti":     kpi_qs.filter(priorita=PrioritaTicket.URGENTE, stato__in=[StatoTicket.APERTA, StatoTicket.IN_CARICO]).count(),
        "n_in_carico":   kpi_qs.filter(stato=StatoTicket.IN_CARICO).count(),
    }
    return render(request, "tickets/pages/dashboard.html", ctx)


# ---------------------------------------------------------------------------
# Creazione ticket
# ---------------------------------------------------------------------------

@login_required
def ticket_nuovo(request):
    tipo = request.GET.get("tipo", "").strip().upper()
    if tipo not in (TipoTicket.IT, TipoTicket.MAN):
        tipo = ""

    if tipo and not _can_open_tickets(request, tipo):
        return render(request, "core/pages/forbidden.html", status=403)

    # Asset pre-selezionato via ?asset=<id> (proveniente dalla scheda asset)
    initial_asset_id = request.GET.get("asset", "").strip()
    initial_asset = None
    if initial_asset_id:
        try:
            from assets.models import Asset as AssetModel
            initial_asset = AssetModel.objects.filter(pk=int(initial_asset_id)).only(
                "pk", "name", "asset_tag"
            ).first()
        except (ValueError, TypeError):
            pass

    name, email, legacy_id = _legacy_identity(request)
    reparto = _get_user_reparto(legacy_id)

    def _ctx(tipo_eff=None, **kwargs):
        t = tipo_eff or tipo
        my, lbl = _get_my_assets(legacy_id, name, reparto, t)
        ctx = _ticket_form_context(tipo=t, my_assets_list=my, my_assets_label=lbl, **kwargs)
        ctx["initial_asset"] = initial_asset
        return ctx

    if request.method == "POST":
        tipo_post = (request.POST.get("tipo") or "").strip().upper()
        if tipo_post not in (TipoTicket.IT, TipoTicket.MAN):
            return render(request, "tickets/pages/nuovo.html",
                _ctx(tipo_eff=tipo_post, error="Tipo ticket non valido.", form_data=request.POST))

        if not _can_open_tickets(request, tipo_post):
            return render(request, "core/pages/forbidden.html", status=403)

        titolo      = (request.POST.get("titolo") or "").strip()[:300]
        descrizione = (request.POST.get("descrizione") or "").strip()
        categoria   = (request.POST.get("categoria") or "").strip()[:30]
        priorita    = (request.POST.get("priorita") or PrioritaTicket.MEDIA).strip()
        sicurezza_raw = (request.POST.get("incide_sicurezza") or "").strip()
        sicurezza   = sicurezza_raw == "1"
        asset_id    = (request.POST.get("asset_id") or "").strip()
        asset_libera= (request.POST.get("asset_descrizione_libera") or "").strip()[:300]
        include_maintenance = request.POST.get("include_in_maintenance_register") == "1"

        if sicurezza_raw not in {"0", "1"}:
            return render(request, "tickets/pages/nuovo.html",
                _ctx(tipo_eff=tipo_post,
                     error="Indica se il problema incide sulla sicurezza sul lavoro prima di proseguire.",
                     form_data=request.POST))

        if not titolo or not descrizione or not categoria:
            return render(request, "tickets/pages/nuovo.html",
                _ctx(tipo_eff=tipo_post,
                     error="Titolo, descrizione e categoria sono obbligatori.",
                     form_data=request.POST))

        if priorita not in dict(PrioritaTicket.choices):
            priorita = PrioritaTicket.MEDIA

        asset_obj = None
        if asset_id:
            try:
                from assets.models import Asset
                asset_obj = Asset.objects.filter(pk=int(asset_id)).first()
            except (ValueError, TypeError):
                pass

        ticket = Ticket(
            tipo=tipo_post,
            titolo=titolo,
            descrizione=descrizione,
            categoria=categoria,
            priorita=priorita,
            incide_sicurezza=sicurezza,
            asset=asset_obj,
            asset_descrizione_libera=asset_libera,
            richiedente_nome=name,
            richiedente_email=email,
            richiedente_legacy_user_id=legacy_id,
            richiedente_user=request.user,
            include_in_maintenance_register=include_maintenance if tipo_post == TipoTicket.MAN else False,
        )
        ticket.save()

        # Allegati
        for f in request.FILES.getlist("allegati"):
            TicketAllegato.objects.create(
                ticket=ticket,
                file=f,
                nome_originale=f.name[:255],
                tipo_mime=(f.content_type or "")[:100],
                uploaded_by_nome=name,
            )

        # Push SP (fire-and-forget, non blocca)
        try:
            _push_ticket_to_sharepoint(ticket)
        except Exception:
            pass

        return redirect("tickets:detail", pk=ticket.pk)

    return render(request, "tickets/pages/nuovo.html", _ctx())


# ---------------------------------------------------------------------------
# Dettaglio ticket (richiedente)
# ---------------------------------------------------------------------------

@login_required
def ticket_detail(request, pk: int):
    ticket = get_object_or_404(Ticket.objects.select_related("richiedente_user"), pk=pk)
    access = _ticket_access_flags(request, ticket)
    is_admin = access["is_admin"]
    is_gestore = access["is_gestore"]
    is_richiedente = access["is_richiedente"]
    if not (is_richiedente or is_gestore or is_admin):
        return render(request, "core/pages/forbidden.html", status=403)

    commenti = ticket.commenti.all()
    if not (is_gestore or is_admin):
        commenti = commenti.filter(is_interno=False)

    from core.legacy_models import AnagraficaDipendente
    richiedente_anagrafica_id = None
    if ticket.richiedente_user_id:
        richiedente_anagrafica_id = (
            AnagraficaDipendente.objects
            .filter(aliasusername=ticket.richiedente_user.username)
            .values_list("id", flat=True)
            .first()
        )

    ctx = {
        "ticket":        ticket,
        "commenti":      commenti,
        "allegati":      ticket.allegati.all(),
        "is_gestore":    is_gestore,
        "is_admin":      is_admin,
        "is_richiedente":is_richiedente,
        "richiedente_anagrafica_id": richiedente_anagrafica_id,
    }
    return render(request, "tickets/pages/detail.html", ctx)


@login_required
def ticket_download_allegato(request, allegato_id: int):
    allegato = get_object_or_404(TicketAllegato.objects.select_related("ticket"), pk=allegato_id)
    access = _ticket_access_flags(request, allegato.ticket)
    if not (access["is_richiedente"] or access["is_gestore"] or access["is_admin"]):
        log_action(
            request,
            "download_allegato",
            "tickets",
            {
                "allegato_id": allegato.id,
                "ticket_id": allegato.ticket_id,
                "esito": "denied",
                "motivo": "permission_denied",
            },
        )
        return render(request, "core/pages/forbidden.html", status=403)
    storage = allegato.file.storage
    if not allegato.file or not allegato.file.name or not storage.exists(allegato.file.name):
        log_action(
            request,
            "download_allegato",
            "tickets",
            {
                "allegato_id": allegato.id,
                "ticket_id": allegato.ticket_id,
                "esito": "not_found",
            },
        )
        return HttpResponse("Allegato non trovato.", status=404)
    content_type = allegato.tipo_mime or mimetypes.guess_type(allegato.nome_originale)[0] or "application/octet-stream"
    log_action(
        request,
        "download_allegato",
        "tickets",
        {
            "allegato_id": allegato.id,
            "ticket_id": allegato.ticket_id,
            "filename": allegato.nome_originale,
            "esito": "success",
        },
    )
    return FileResponse(
        storage.open(allegato.file.name, "rb"),
        as_attachment=True,
        filename=allegato.nome_originale,
        content_type=content_type,
    )


@login_required
def ticket_pdf(request, pk: int):
    ticket = get_object_or_404(Ticket, pk=pk)
    access = _ticket_access_flags(request, ticket)
    if not (access["is_richiedente"] or access["is_gestore"] or access["is_admin"]):
        return render(request, "core/pages/forbidden.html", status=403)

    include_internal = bool(access["is_gestore"] or access["is_admin"])
    commenti = ticket.commenti.all()
    if not include_internal:
        commenti = commenti.filter(is_interno=False)

    return _ticket_pdf_response(
        ticket,
        commenti=list(commenti),
        allegati=list(ticket.allegati.all()),
        include_internal=include_internal,
    )


# ---------------------------------------------------------------------------
# Gestione lista (team)
# ---------------------------------------------------------------------------

@_tickets_gestione_required
def ticket_gestione_list(request):
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin    = bool(legacy_user and is_legacy_admin(legacy_user))

    # ── Vista predefinita: tab IT/MAN/TUTTI ──
    # Priorità: parametro GET esplicito > sessione > auto-detect da ACL
    tipo_f = request.GET.get("tipo", "").strip().upper()
    if tipo_f in (TipoTicket.IT, TipoTicket.MAN, "TUTTI"):
        if tipo_f == "TUTTI":
            tipo_f = ""
        request.session["tkt_gestione_default_tipo"] = tipo_f
    else:
        # Recupera da sessione
        tipo_f = request.session.get("tkt_gestione_default_tipo", None)
        if tipo_f is None:
            # Auto-detect: se gestore solo di un tipo, predisponi quello
            can_it  = _can_manage_tickets(request, TipoTicket.IT)
            can_man = _can_manage_tickets(request, TipoTicket.MAN)
            if can_it and not can_man:
                tipo_f = TipoTicket.IT
            elif can_man and not can_it:
                tipo_f = TipoTicket.MAN
            else:
                tipo_f = ""
            request.session["tkt_gestione_default_tipo"] = tipo_f

    stato_f    = request.GET.get("stato", "").strip().upper()
    prio_f     = request.GET.get("priorita", "").strip().upper()
    cerca_f    = request.GET.get("q", "").strip()
    ass_f      = request.GET.get("assegnato", "").strip()
    asset_f    = request.GET.get("asset", "").strip()
    cat_f      = request.GET.get("categoria", "").strip()
    data_da_f  = request.GET.get("data_da", "").strip()
    data_a_f   = request.GET.get("data_a", "").strip()
    sicurezza_f  = request.GET.get("sicurezza", "").strip()
    ricorrente_f = request.GET.get("ricorrente", "").strip()
    ordine_f     = request.GET.get("ordine", "").strip()

    _priority_order = Case(
        When(priorita="URGENTE", then=Value(0)),
        When(priorita="ALTA",    then=Value(1)),
        When(priorita="MEDIA",   then=Value(2)),
        When(priorita="BASSA",   then=Value(3)),
        default=Value(4),
        output_field=IntegerField(),
    )
    _stato_order = Case(
        When(stato="APERTA",    then=Value(0)),
        When(stato="IN_CARICO", then=Value(1)),
        When(stato="RISOLTO",   then=Value(2)),
        When(stato="CHIUSO",    then=Value(3)),
        When(stato="ANNULLATO", then=Value(4)),
        default=Value(5),
        output_field=IntegerField(),
    )

    def _apply_filters(qs):
        if stato_f in dict(StatoTicket.choices):
            qs = qs.filter(stato=stato_f)
        if prio_f in dict(PrioritaTicket.choices):
            qs = qs.filter(priorita=prio_f)
        if cerca_f:
            qs = qs.filter(
                Q(titolo__icontains=cerca_f) |
                Q(numero_ticket__icontains=cerca_f) |
                Q(richiedente_nome__icontains=cerca_f)
            )
        if ass_f:
            qs = qs.filter(assegnato_a="") if ass_f == "__none__" else qs.filter(assegnato_a__icontains=ass_f)
        if asset_f:
            qs = qs.filter(
                Q(asset__name__icontains=asset_f) |
                Q(asset__asset_tag__icontains=asset_f) |
                Q(asset__internal_number__icontains=asset_f) |
                Q(asset_descrizione_libera__icontains=asset_f)
            )
        if cat_f:
            qs = qs.filter(categoria=cat_f)
        if sicurezza_f == "1":
            qs = qs.filter(incide_sicurezza=True)
        if ricorrente_f == "1":
            qs = qs.filter(ricorrente=True)
        if data_da_f:
            try:
                qs = qs.filter(created_at__date__gte=datetime.strptime(data_da_f, "%Y-%m-%d").date())
            except ValueError:
                pass
        if data_a_f:
            try:
                qs = qs.filter(created_at__date__lte=datetime.strptime(data_a_f, "%Y-%m-%d").date())
            except ValueError:
                pass
        # Ordinamento
        if ordine_f == "priorita":
            return qs.annotate(_po=_priority_order).order_by("_po", "-created_at")
        elif ordine_f == "stato":
            return qs.annotate(_so=_stato_order).order_by("_so", "-created_at")
        elif ordine_f == "assegnato":
            return qs.order_by("assegnato_a", "-created_at")
        elif ordine_f == "apertura_asc":
            return qs.order_by("created_at")
        else:
            return qs.order_by("-created_at")

    from core.legacy_models import AnagraficaDipendente
    base = Ticket.objects.select_related("asset", "richiedente_user").annotate(
        richiedente_anagrafica_id=Subquery(
            AnagraficaDipendente.objects.filter(
                aliasusername=OuterRef("richiedente_user__username")
            ).values("id")[:1]
        )
    )
    if tipo_f in (TipoTicket.IT, TipoTicket.MAN):
        tickets_it  = _apply_filters(base.filter(tipo=TipoTicket.IT))  if tipo_f == TipoTicket.IT  else None
        tickets_man = _apply_filters(base.filter(tipo=TipoTicket.MAN)) if tipo_f == TipoTicket.MAN else None
    else:
        tickets_it  = _apply_filters(base.filter(tipo=TipoTicket.IT))
        tickets_man = _apply_filters(base.filter(tipo=TipoTicket.MAN))

    ctx = {
        "tickets_it":    tickets_it,
        "tickets_man":   tickets_man,
        "tab_attivo":    tipo_f or "TUTTI",
        "is_admin":      is_admin,
        "stati":         StatoTicket.choices,
        "priorita_list": PrioritaTicket.choices,
        "filtro_stato":    stato_f,
        "filtro_prio":     prio_f,
        "filtro_cerca":    cerca_f,
        "filtro_ass":      ass_f,
        "filtro_asset":    asset_f,
        "filtro_cat":      cat_f,
        "filtro_data_da":  data_da_f,
        "filtro_data_a":   data_a_f,
        "filtro_sicurezza":  sicurezza_f,
        "filtro_ricorrente": ricorrente_f,
        "filtro_ordine":     ordine_f,
        "categorie_it":    get_categorie(TipoTicket.IT),
        "categorie_man":   get_categorie(TipoTicket.MAN),
        # KPI globali
        "n_aperte":    Ticket.objects.filter(stato=StatoTicket.APERTA).count(),
        "n_in_carico": Ticket.objects.filter(stato=StatoTicket.IN_CARICO).count(),
        "n_urgenti":   Ticket.objects.filter(priorita=PrioritaTicket.URGENTE, stato__in=[StatoTicket.APERTA, StatoTicket.IN_CARICO]).count(),
        "n_risolti":   Ticket.objects.filter(stato=StatoTicket.RISOLTO).count(),
    }
    return render(request, "tickets/pages/gestione_list.html", ctx)


# ---------------------------------------------------------------------------
# Gestione dettaglio (team)
# ---------------------------------------------------------------------------

@_tickets_gestione_required
def ticket_gestione_detail(request, pk: int):
    ticket = get_object_or_404(Ticket.objects.select_related("richiedente_user"), pk=pk)
    name, email, legacy_id = _legacy_identity(request)
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin    = bool(legacy_user and is_legacy_admin(legacy_user))

    # SEC: il decoratore garantisce solo "gestore di almeno un tipo". Il dettaglio
    # gestione (note interne, dati richiedente, costi) deve essere visibile solo al
    # gestore del TIPO specifico (o admin), non a un gestore di un altro tipo.
    if not (is_admin or _can_manage_tickets(request, ticket.tipo)):
        return render(request, "core/pages/forbidden.html", status=403)

    cfg         = TicketImpostazioni.get_or_create_for(ticket.tipo)
    fornitori   = _get_fornitori_for_select()
    commenti    = ticket.commenti.all()
    allegati    = ticket.allegati.all()
    interventi  = ticket.interventi.all()
    stato_log   = ticket.stato_log.all()

    reparto          = _get_user_reparto(legacy_id)
    my_assets, my_assets_label = _get_my_assets(legacy_id, name, reparto, ticket.tipo)

    # Workorder generati da questo ticket
    workorders_generati = ticket.workorders_generati.select_related("asset", "supplier").all()

    # Costi totali interventi
    costo_manodopera_tot = sum(
        float(i.costo_manodopera_eur) for i in interventi if i.costo_manodopera_eur
    )
    costo_materiali_tot = sum(
        float(i.materiali_costo_eur) for i in interventi if i.materiali_costo_eur
    )

    from core.legacy_models import AnagraficaDipendente
    richiedente_anagrafica_id = None
    if ticket.richiedente_user_id:
        richiedente_anagrafica_id = (
            AnagraficaDipendente.objects
            .filter(aliasusername=ticket.richiedente_user.username)
            .values_list("id", flat=True)
            .first()
        )

    ctx = {
        "ticket":           ticket,
        "commenti":         commenti,
        "allegati":         allegati,
        "cfg":              cfg,
        "stati":            StatoTicket.choices,
        "fornitori":        fornitori,
        "is_admin":         is_admin,
        "current_user_name":  name,
        "current_user_email": email,
        "assets_list":      _get_assets_for_select(),
        "my_assets_list":   my_assets,
        "my_assets_label":  my_assets_label,
        # Analytics
        "interventi":       interventi,
        "stato_log":        stato_log,
        "activity_feed":    _build_ticket_activity_feed(ticket),
        "tipo_fermo_choices": TipoFermo.choices,
        "esito_intervento_choices": EsitoIntervento.choices,
        # Manutenzione / costi
        "workorders_generati":   workorders_generati,
        "costo_manodopera_tot":  round(costo_manodopera_tot, 2) if costo_manodopera_tot else None,
        "costo_materiali_tot":   round(costo_materiali_tot, 2) if costo_materiali_tot else None,
        "costo_totale":          round(costo_manodopera_tot + costo_materiali_tot, 2) if (costo_manodopera_tot or costo_materiali_tot) else None,
        "richiedente_anagrafica_id": richiedente_anagrafica_id,
    }
    return render(request, "tickets/pages/gestione_detail.html", ctx)


# ---------------------------------------------------------------------------
# Impostazioni admin
# ---------------------------------------------------------------------------

@login_required
def ticket_impostazioni(request):
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if not (legacy_user and is_legacy_admin(legacy_user)):
        return render(request, "core/pages/forbidden.html", status=403)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        redirect_target = redirect("tickets:impostazioni")

        if action in {"add_categoria_ticket", "toggle_categoria_ticket", "delete_categoria_ticket"}:
            if action == "add_categoria_ticket":
                tipo = (request.POST.get("cat_tipo") or "").strip()
                codice = (request.POST.get("cat_codice") or "").strip().upper().replace(" ", "_")[:30]
                etichetta = (request.POST.get("cat_etichetta") or "").strip()[:100]
                ordine = _int_or_none(request.POST.get("cat_ordine")) or 0
                if tipo in (TipoTicket.IT, TipoTicket.MAN) and codice and etichetta:
                    _, created = CategoriaTicket.objects.get_or_create(
                        codice=codice,
                        defaults={
                            "tipo": tipo,
                            "etichetta": etichetta,
                            "ordine": ordine,
                            "attivo": True,
                        },
                    )
                    if created:
                        log_action(request, "add_categoria_ticket", "tickets", {"tipo": tipo, "codice": codice})
                        messages.success(request, f"Categoria '{etichetta}' aggiunta.")
                    else:
                        messages.warning(request, f"Codice '{codice}' gia esistente.")
                else:
                    messages.error(request, "Dati categoria mancanti o non validi.")
                return redirect_target

            cat_id = _int_or_none(request.POST.get("cat_id"))
            cat = CategoriaTicket.objects.filter(pk=cat_id).first() if cat_id else None
            if cat is None:
                messages.error(request, "Categoria ticket non trovata.")
                return redirect_target

            if action == "toggle_categoria_ticket":
                cat.attivo = not cat.attivo
                cat.save(update_fields=["attivo"])
                log_action(
                    request,
                    "toggle_categoria_ticket",
                    "tickets",
                    {"id": cat.pk, "attivo": cat.attivo},
                )
                messages.success(
                    request,
                    f"Categoria '{cat.etichetta}' {'attivata' if cat.attivo else 'disattivata'}.",
                )
                return redirect_target

            in_use = Ticket.objects.filter(categoria=cat.codice).exists()
            if in_use:
                messages.error(
                    request,
                    f"Impossibile eliminare '{cat.etichetta}': e usata da ticket esistenti. Disattivala.",
                )
                return redirect_target
            cat_label = cat.etichetta
            cat_id_value = cat.pk
            cat.delete()
            log_action(request, "delete_categoria_ticket", "tickets", {"id": cat_id_value})
            messages.success(request, f"Categoria '{cat_label}' eliminata.")
            return redirect_target

        if action == "save_escalation_config":
            from tickets.escalation_config import save_escalation_config
            attivo = (request.POST.get("esc_attivo") or "").strip().lower() in {"1", "on", "true", "si"}
            soglia = _int_or_none(request.POST.get("esc_soglia_ore")) or 0
            ora = _int_or_none(request.POST.get("esc_ora_invio"))
            ora = ora if ora is not None else 8
            ok = save_escalation_config(attivo=attivo, soglia_ore=soglia, ora_invio=ora)
            log_action(request, "save_ticket_escalation_config", "tickets",
                       {"attivo": attivo, "soglia_ore": soglia, "ora_invio": ora})
            if ok:
                messages.success(request, "Configurazione escalation ticket salvata.")
            else:
                messages.error(request, "Salvataggio escalation parzialmente fallito.")
            return redirect_target

    cfg_it  = TicketImpostazioni.get_or_create_for(TipoTicket.IT)
    cfg_man = TicketImpostazioni.get_or_create_for(TipoTicket.MAN)

    cutoff = tz_now() - timedelta(days=90)
    top_cause = list(
        Ticket.objects
        .filter(ricorrente=True, causa_radice__gt="", created_at__gte=cutoff)
        .values("causa_radice")
        .annotate(n=Count("id"))
        .order_by("-n")[:5]
    )

    from tickets.escalation_config import get_escalation_config

    ctx = {
        "cfg_it":  cfg_it,
        "cfg_man": cfg_man,
        "escalation_cfg": get_escalation_config(),
        "tipi":    TipoTicket.choices,
        "categorie_it":  get_categorie(TipoTicket.IT),
        "categorie_man": get_categorie(TipoTicket.MAN),
        "categorie_ticket_it": list(
            CategoriaTicket.objects.filter(tipo=TipoTicket.IT).order_by("ordine", "etichetta")
        ),
        "categorie_ticket_man": list(
            CategoriaTicket.objects.filter(tipo=TipoTicket.MAN).order_by("ordine", "etichetta")
        ),
        "top_cause": top_cause,
    }
    return render(request, "tickets/pages/impostazioni.html", ctx)


# ---------------------------------------------------------------------------
# SharePoint push (stub — da configurare con list IDs reali)
# ---------------------------------------------------------------------------

def _push_ticket_to_sharepoint(ticket: Ticket) -> None:
    """Push ticket a SP. Stub — da implementare quando disponibili list IDs."""
    try:
        from core.sharepoint_utils import get_sp_headers, get_sp_site_url
        import requests as req_lib

        cfg = TicketImpostazioni.objects.filter(tipo=ticket.tipo).first()
        if not cfg or not cfg.sharepoint_list_id:
            return

        headers = get_sp_headers()
        site    = get_sp_site_url()
        url     = f"{site}/_api/web/lists('{cfg.sharepoint_list_id}')/items"

        payload = {
            "__metadata": {"type": f"SP.Data.ListItem"},
            "Title":          ticket.titolo,
            "NumeroTicket":   ticket.numero_ticket,
            "Categoria":      ticket.categoria,
            "Priorita":       ticket.priorita,
            "IncideSicurezza": ticket.incide_sicurezza,
            "Stato":          ticket.stato,
            "Richiedente":    ticket.richiedente_nome,
            "Descrizione":    ticket.descrizione,
            "DataApertura":   ticket.created_at.isoformat() if ticket.created_at else "",
        }
        resp = req_lib.post(url, json=payload, headers=headers, timeout=10)
        if resp.ok:
            data = resp.json()
            sp_id = str(data.get("d", {}).get("ID") or data.get("ID") or "")
            if sp_id:
                Ticket.objects.filter(pk=ticket.pk).update(sharepoint_item_id=sp_id)
    except Exception:
        pass


def _update_ticket_sharepoint(ticket: Ticket) -> None:
    """Aggiorna item esistente su SP (stato, assegnazione)."""
    try:
        if not ticket.sharepoint_item_id:
            return
        from core.sharepoint_utils import get_sp_headers, get_sp_site_url
        import requests as req_lib

        cfg = TicketImpostazioni.objects.filter(tipo=ticket.tipo).first()
        if not cfg or not cfg.sharepoint_list_id:
            return

        headers = get_sp_headers()
        headers.update({"X-HTTP-Method": "MERGE", "IF-MATCH": "*"})
        site = get_sp_site_url()
        url  = f"{site}/_api/web/lists('{cfg.sharepoint_list_id}')/items({ticket.sharepoint_item_id})"

        payload = {
            "__metadata": {"type": "SP.Data.ListItem"},
            "Stato":       ticket.stato,
            "AssegnatoA":  ticket.assegnato_a,
            "DataChiusura": ticket.closed_at.isoformat() if ticket.closed_at else None,
        }
        req_lib.patch(url, json=payload, headers=headers, timeout=10)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# API: commento
# ---------------------------------------------------------------------------

@require_POST
@login_required
def api_commento(request):
    try:
        payload    = json.loads(request.body)
        ticket_id  = int(payload.get("ticket_id") or 0)
        testo      = (payload.get("testo") or "").strip()
        is_interno = bool(payload.get("is_interno"))
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    name, email, legacy_id = _legacy_identity(request)

    is_gestore = _can_manage_tickets(request, ticket.tipo)
    legacy_user= getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin   = bool(legacy_user and is_legacy_admin(legacy_user))

    # Solo gestori possono scrivere note interne
    if is_interno and not (is_gestore or is_admin):
        return _json_err("Non autorizzato", 403)

    # Verifica che l'utente possa commentare su questo ticket
    is_richiedente = (
        ticket.richiedente_user_id == request.user.id
        or (legacy_id and ticket.richiedente_legacy_user_id == legacy_id)
        or ticket.richiedente_nome == name
        or ticket.richiedente_email.lower() == email.lower()
    )
    if not (is_richiedente or is_gestore or is_admin):
        return _json_err("Non autorizzato", 403)

    if not testo:
        return _json_err("Testo vuoto")

    c = TicketCommento.objects.create(
        ticket=ticket,
        autore_nome=name,
        autore_email=email,
        testo=testo,
        is_interno=is_interno,
    )

    # Auto presa in carico: se il gestore commenta su un ticket ancora APERTA,
    # lo porta automaticamente in IN_CARICO (prima risposta operativa)
    auto_in_carico = False
    if ticket.stato == StatoTicket.APERTA and (is_gestore or is_admin):
        now = tz_now()
        ticket.stato = StatoTicket.IN_CARICO
        upd = ["stato", "updated_at"]
        if not ticket.data_presa_in_carico:
            ticket.data_presa_in_carico = now
            upd.append("data_presa_in_carico")
        ticket.save(update_fields=upd)
        TicketStatoLog.objects.create(
            ticket=ticket,
            da_stato=StatoTicket.APERTA,
            a_stato=StatoTicket.IN_CARICO,
            utente_nome=name,
            utente_email=email,
            nota="Preso in carico automaticamente alla prima risposta operativa",
        )
        TicketCommento.objects.create(
            ticket=ticket,
            autore_nome=name,
            autore_email=email,
            testo="Stato aggiornato: Aperta → In carico",
            is_interno=True,
        )
        auto_in_carico = True
        try:
            _update_ticket_sharepoint(ticket)
        except Exception:
            pass

    return JsonResponse({
        "ok": True,
        "commento_id": c.pk,
        "autore_nome": c.autore_nome,
        "testo": c.testo,
        "is_interno": c.is_interno,
        "created_at": c.created_at.strftime("%d-%m-%Y %H:%M"),
        "auto_in_carico": auto_in_carico,
    })


# ---------------------------------------------------------------------------
# API: allegato upload
# ---------------------------------------------------------------------------

@require_POST
@login_required
def api_allegato(request):
    ticket_id = int(request.POST.get("ticket_id") or 0)
    ticket    = get_object_or_404(Ticket, pk=ticket_id)
    name, email, legacy_id = _legacy_identity(request)

    is_gestore = _can_manage_tickets(request, ticket.tipo)
    legacy_user= getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    is_admin   = bool(legacy_user and is_legacy_admin(legacy_user))
    is_richiedente = (
        ticket.richiedente_user_id == request.user.id
        or (legacy_id and ticket.richiedente_legacy_user_id == legacy_id)
        or ticket.richiedente_nome == name
    )
    if not (is_richiedente or is_gestore or is_admin):
        return _json_err("Non autorizzato", 403)

    f = request.FILES.get("file")
    if not f:
        return _json_err("Nessun file")

    _MAX_SIZE = 20 * 1024 * 1024  # 20 MB
    try:
        detected_mime = validate_extension_and_mime(
            f,
            allowed_extensions=TICKET_ALLOWED_UPLOAD_EXTENSIONS,
            allowed_mimes=TICKET_ALLOWED_UPLOAD_MIMES,
            max_bytes=_MAX_SIZE,
        )
    except UploadMimeValidationError as exc:
        return _json_err(str(exc))

    allegato = TicketAllegato.objects.create(
        ticket=ticket,
        file=f,
        nome_originale=f.name[:255],
        tipo_mime=detected_mime[:100],
        uploaded_by_nome=name,
    )
    return JsonResponse({
        "ok": True,
        "allegato_id": allegato.pk,
        "nome": allegato.nome_originale,
        "url":  reverse("tickets:download_allegato", args=[allegato.pk]),
    })


# ---------------------------------------------------------------------------
# API: aggiorna stato (solo gestori)
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_stato(request):
    try:
        payload   = json.loads(request.body)
        ticket_id = int(payload.get("ticket_id") or 0)
        nuovo_stato = (payload.get("stato") or "").strip().upper()
        nota        = (payload.get("nota") or "").strip()
        data_prevista = (payload.get("data_prevista") or "").strip()
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    if nuovo_stato not in dict(StatoTicket.choices):
        return _json_err("Stato non valido")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)  # SEC: gestore del TIPO specifico
    if denied is not None:
        return denied
    vecchio = ticket.stato
    ticket.stato = nuovo_stato
    now = tz_now()

    if nuovo_stato in (StatoTicket.CHIUSO, StatoTicket.ANNULLATO, StatoTicket.RISOLTO):
        if not ticket.closed_at:
            ticket.closed_at = now
    else:
        ticket.closed_at = None

    # Analytics: registra data_presa_in_carico al primo passaggio IN_CARICO
    update_fields = ["stato", "closed_at", "updated_at"]
    name, email, _ = _legacy_identity(request)

    if nuovo_stato == StatoTicket.IN_CARICO and not ticket.data_presa_in_carico:
        ticket.data_presa_in_carico = now
        update_fields.append("data_presa_in_carico")

    # Data prevista risoluzione quando messo IN_ATTESA
    if nuovo_stato == StatoTicket.IN_ATTESA and data_prevista:
        try:
            ticket.data_prevista_risoluzione = datetime.strptime(data_prevista, "%Y-%m-%d").date()
            update_fields.append("data_prevista_risoluzione")
        except ValueError:
            pass

    # Analytics: chi ha risolto
    if nuovo_stato == StatoTicket.RISOLTO and not ticket.risolto_da_nome:
        ticket.risolto_da_nome = name
        update_fields.append("risolto_da_nome")

    ticket.save(update_fields=update_fields)

    # Log strutturato cambio stato
    TicketStatoLog.objects.create(
        ticket=ticket,
        da_stato=vecchio,
        a_stato=nuovo_stato,
        utente_nome=name,
        utente_email=email,
        nota=nota,
    )

    # Commento automatico cambio stato
    label_stato = dict(StatoTicket.choices).get(nuovo_stato, nuovo_stato)
    testo_auto  = f"Stato aggiornato: {dict(StatoTicket.choices).get(vecchio, vecchio)} → {label_stato}"
    if nota:
        testo_auto += f"\n{nota}"
    TicketCommento.objects.create(
        ticket=ticket,
        autore_nome=name,
        autore_email=email,
        testo=testo_auto,
        is_interno=True,
    )

    try:
        _update_ticket_sharepoint(ticket)
    except Exception:
        pass

    return JsonResponse({"ok": True, "stato": nuovo_stato, "label": label_stato})


# ---------------------------------------------------------------------------
# API: assegna tecnico (solo gestori)
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_assegna(request):
    try:
        payload       = json.loads(request.body)
        ticket_id     = int(payload.get("ticket_id") or 0)
        assegnato_a   = (payload.get("assegnato_a") or "").strip()[:200]
        assegnato_email = (payload.get("assegnato_email") or "").strip()[:200]
        fornitore_id  = payload.get("fornitore_id")
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)  # SEC: gestore del TIPO specifico
    if denied is not None:
        return denied
    ticket.assegnato_a    = assegnato_a
    ticket.assegnato_email= assegnato_email

    if fornitore_id:
        try:
            from anagrafica.models import Fornitore
            ticket.delegato_fornitore = Fornitore.objects.filter(pk=int(fornitore_id)).first()
        except (ValueError, TypeError):
            ticket.delegato_fornitore = None
    else:
        ticket.delegato_fornitore = None

    if ticket.stato == StatoTicket.APERTA and assegnato_a:
        ticket.stato = StatoTicket.IN_CARICO

    ticket.save(update_fields=["assegnato_a", "assegnato_email", "delegato_fornitore", "stato", "updated_at"])

    name, email, _ = _legacy_identity(request)
    desc = assegnato_a or (ticket.delegato_fornitore.ragione_sociale if ticket.delegato_fornitore else "—")
    TicketCommento.objects.create(
        ticket=ticket,
        autore_nome=name,
        autore_email=email,
        testo=f"Ticket assegnato a: {desc}",
        is_interno=True,
    )

    try:
        _update_ticket_sharepoint(ticket)
    except Exception:
        pass

    return JsonResponse({
        "ok": True,
        "assegnato_a": ticket.assegnato_a,
        "stato": ticket.stato,
    })


# ---------------------------------------------------------------------------
# API: Copilota triage AI (solo gestori) - proposta read-only, l'umano firma
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_copilota_triage(request):
    """Propone (NON salva) categoria/priorita/assegnatario/bozza dal testo ticket.

    Read-only: il gestore rivede la proposta e applica con i pulsanti esistenti
    (assegna/stato). Audit solo-metadati, nessun prompt/testo salvato.
    """
    try:
        payload = json.loads(request.body)
        ticket_id = payload.get("ticket_id")
        tipo = (payload.get("tipo") or "").strip().upper()
        titolo = (payload.get("titolo") or "").strip()
        descrizione = (payload.get("descrizione") or "").strip()
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = None
    tid = _int_or_none(ticket_id)
    if tid:
        ticket = Ticket.objects.filter(pk=tid).first()
        if ticket is not None:
            tipo = ticket.tipo
            titolo = titolo or ticket.titolo
            descrizione = descrizione or ticket.descrizione

    if tipo not in (TipoTicket.IT, TipoTicket.MAN):
        return _json_err("Tipo ticket non valido")
    if not (titolo or descrizione):
        return _json_err("Testo del ticket mancante")

    categorie = list(get_categorie(tipo))
    cfg = TicketImpostazioni.objects.filter(tipo=tipo).first()
    gestori = (cfg.team_gestori if cfg else []) or []

    from .ai_copilota import proponi_triage
    proposta = proponi_triage(
        titolo=titolo, descrizione=descrizione, tipo=tipo,
        categorie=categorie, gestori=gestori,
    )

    # Audit metadata-only: niente prompt, testo o contenuti del modello.
    log_action(request, "ticket_copilota_triage", "tickets", {
        "ticket_id": ticket.pk if ticket else None,
        "tipo": tipo,
        "ai_disponibile": proposta.get("ai_disponibile"),
        "ha_categoria": bool(proposta.get("categoria")),
        "ha_assegnatario": bool(proposta.get("assegnatario")),
    })

    return JsonResponse({"ok": True, "proposta": proposta})


# ---------------------------------------------------------------------------
# API: aggiorna asset collegato al ticket (gestori e admin)
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_asset(request):
    try:
        payload   = json.loads(request.body)
        ticket_id = int(payload.get("ticket_id") or 0)
        asset_id  = payload.get("asset_id")           # int o null
        asset_lib = (payload.get("asset_descrizione_libera") or "").strip()[:300]
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)  # SEC: gestore del TIPO specifico
    if denied is not None:
        return denied

    asset_obj = None
    if asset_id:
        try:
            from assets.models import Asset as AssetModel
            asset_obj = AssetModel.objects.filter(pk=int(asset_id)).first()
        except (ValueError, TypeError):
            pass

    ticket.asset = asset_obj
    ticket.asset_descrizione_libera = asset_lib if not asset_obj else ""
    ticket.save(update_fields=["asset", "asset_descrizione_libera", "updated_at"])

    name, email, _ = _legacy_identity(request)
    label = f"{asset_obj.name} [{asset_obj.asset_tag}]" if asset_obj else (asset_lib or "—")
    TicketCommento.objects.create(
        ticket=ticket,
        autore_nome=name,
        autore_email=email,
        testo=f"Asset collegato: {label}",
        is_interno=True,
    )

    return JsonResponse({
        "ok":         True,
        "asset_id":   asset_obj.pk if asset_obj else None,
        "asset_name": asset_obj.name if asset_obj else "",
        "asset_tag":  asset_obj.asset_tag if asset_obj else "",
        "asset_url":  f"/assets/{asset_obj.pk}/" if asset_obj else "",
        "asset_lib":  ticket.asset_descrizione_libera,
    })


# ---------------------------------------------------------------------------
# API: autocomplete asset per filtro gestione lista
# ---------------------------------------------------------------------------

@login_required
def api_assets_autocomplete(request):
    q = request.GET.get("q", "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})
    try:
        from assets.models import Asset as AssetModel
        assets = (
            AssetModel.objects.filter(status__in=["IN_USE", "IN_STOCK"])
            .filter(Q(name__icontains=q) | Q(asset_tag__icontains=q) | Q(internal_number__icontains=q))
            .order_by("name")[:20]
        )
        results = [
            {"name": a.name, "tag": a.asset_tag or "", "internal_number": a.internal_number or "", "tipo": a.get_asset_type_display() if hasattr(a, 'get_asset_type_display') else ""}
            for a in assets
        ]
    except Exception:
        results = []
    return JsonResponse({"results": results})


# ---------------------------------------------------------------------------
# API: salva impostazioni (solo admin)
# ---------------------------------------------------------------------------

@require_POST
@login_required
def api_impostazioni(request):
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if not (legacy_user and is_legacy_admin(legacy_user)):
        return _json_err("Non autorizzato", 403)

    try:
        payload = json.loads(request.body)
        tipo    = (payload.get("tipo") or "").strip().upper()
    except json.JSONDecodeError:
        return _json_err("JSON non valido")

    if tipo not in (TipoTicket.IT, TipoTicket.MAN):
        return _json_err("Tipo non valido")

    cfg = TicketImpostazioni.get_or_create_for(tipo)
    cfg.sharepoint_list_id = (payload.get("sharepoint_list_id") or "").strip()[:100]

    # team_gestori: lista di {nome, email}
    raw_team = payload.get("team_gestori")
    if isinstance(raw_team, list):
        cfg.team_gestori = [
            {"nome": (m.get("nome") or "").strip(), "email": (m.get("email") or "").strip()}
            for m in raw_team if isinstance(m, dict)
        ]

    # acl_apertura / acl_gestione: lista di stringhe username/email
    for field in ("acl_apertura", "acl_gestione"):
        raw = payload.get(field)
        if isinstance(raw, list):
            setattr(cfg, field, [str(v).strip() for v in raw if str(v).strip()])

    cfg.save()
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# API: ricerca utenti (per autocomplete impostazioni)
# ---------------------------------------------------------------------------

@login_required
def api_cerca_utenti(request):
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if not (legacy_user and is_legacy_admin(legacy_user)):
        return _json_err("Non autorizzato", 403)

    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    results = []
    try:
        from core.legacy_models import AnagraficaDipendente, UtenteLegacy
        from django.db.models import Q

        # Cerca in anagrafica_dipendenti (nome + cognome) e utenti (email/UPN)
        qs = AnagraficaDipendente.objects.filter(
            Q(nome__icontains=q) | Q(cognome__icontains=q) |
            Q(email__icontains=q) | Q(aliasusername__icontains=q) |
            Q(email_notifica__icontains=q)
        ).select_related("utente").order_by("cognome", "nome")[:20]

        for a in qs:
            nome_completo = f"{(a.nome or '').strip()} {(a.cognome or '').strip()}".strip()
            username      = (a.aliasusername or "").strip()
            email_login   = (a.email or "").strip()          # UPN (login)
            email_notifica= (a.email_notifica or "").strip() # email reale
            results.append({
                "nome":     nome_completo,
                "username": username,
                "email":    email_login,
                "email_notifica": email_notifica,
                "label":    f"{nome_completo}" + (f" — {username}" if username else ""),
            })
    except Exception:
        # Fallback su Django auth users
        from django.contrib.auth import get_user_model
        User = get_user_model()
        qs = User.objects.filter(
            username__icontains=q
        ).order_by("last_name", "first_name")[:20]
        for u in qs:
            nome = f"{u.first_name} {u.last_name}".strip() or u.username
            results.append({
                "nome":     nome,
                "username": u.username,
                "email":    u.email or "",
                "email_notifica": u.email or "",
                "label":    f"{nome} — {u.username}",
            })

    return JsonResponse({"results": results})


# ---------------------------------------------------------------------------
# API: test connessione SharePoint (verifica list ID)
# ---------------------------------------------------------------------------

@require_POST
@login_required
def api_test_sp(request):
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if not (legacy_user and is_legacy_admin(legacy_user)):
        return _json_err("Non autorizzato", 403)

    try:
        payload = json.loads(request.body)
        list_id = (payload.get("sharepoint_list_id") or "").strip()
    except json.JSONDecodeError:
        return _json_err("JSON non valido")

    if not list_id:
        return _json_err("List ID vuoto")

    try:
        import os
        from core.graph_utils import acquire_graph_token, is_placeholder_value
        from config.env_config import get_first_env_value

        tenant_id = get_first_env_value("GRAPH_TENANT_ID", "AZURE_TENANT_ID")
        client_id = get_first_env_value("GRAPH_CLIENT_ID", "AZURE_CLIENT_ID")
        client_secret = get_first_env_value("GRAPH_CLIENT_SECRET", "AZURE_CLIENT_SECRET")
        site_id = get_first_env_value("GRAPH_SITE_ID")

        if any(is_placeholder_value(v) or not v for v in [tenant_id, client_id, client_secret, site_id]):
            return _json_err("Configurazione Graph incompleta in .env (tenant_id/client_id/client_secret/site_id)")

        import requests as req_lib
        token   = acquire_graph_token(tenant_id, client_id, client_secret)
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

        # GET lista — titolo
        url_list = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}"
        r = req_lib.get(url_list, headers=headers, timeout=10)
        if r.status_code == 404:
            return _json_err("Lista non trovata (404). Verifica il List ID.")
        if r.status_code == 403:
            return _json_err("Permessi insufficienti (403). Verifica che l'app Azure abbia accesso alla lista.")
        if not r.ok:
            return _json_err(f"Errore Graph {r.status_code}: {r.text[:200]}")

        list_title = r.json().get("displayName") or r.json().get("name") or list_id

        # Conta item (prima pagina + @odata.count)
        r2 = req_lib.get(
            f"{url_list}/items?$top=1&$count=true",
            headers={**headers, "ConsistencyLevel": "eventual"},
            timeout=10,
        )
        item_count = r2.json().get("@odata.count", "—") if r2.ok else "—"

        return JsonResponse({"ok": True, "list_title": list_title, "item_count": item_count})

    except Exception as e:
        return _json_err(str(e))


# ---------------------------------------------------------------------------
# API: import CSV/Excel ticket (solo admin)
# ---------------------------------------------------------------------------

def _xlsx_rows_to_dicts(file_obj) -> list[dict]:
    """Parses the first non-Mappa sheet of an xlsx file into a list of row dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = next(
        (wb[s] for s in wb.sheetnames if "mappa" not in s.lower()),
        wb.active,
    )
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(rows_iter)]
    result = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        d: dict[str, str] = {}
        for h, v in zip(headers, row):
            if v is None:
                d[h] = ""
            elif isinstance(v, bool):
                d[h] = "1" if v else "0"
            elif hasattr(v, "isoformat"):
                d[h] = v.isoformat(sep=" ")[:19]
            else:
                d[h] = str(v)
        result.append(d)
    return result


def _resolve_asset_by_name(name: str, all_assets: list) -> "object | None":
    """3-level asset matching: exact → forward icontains → reverse icontains.

    Reverse: ticket name 'HH - HERMLE C1200U' contains asset name 'HERMLE C1200U'.
    Only returns a result when the match is unambiguous.
    """
    if not name:
        return None
    try:
        from assets.models import Asset as _Asset
        # 1. Exact (case-insensitive)
        exact = list(_Asset.objects.filter(name__iexact=name))
        if len(exact) == 1:
            return exact[0]
        if len(exact) > 1:
            return None
        # 2. Forward: DB asset name contains the ticket machine name
        forward = list(_Asset.objects.filter(name__icontains=name))
        if len(forward) == 1:
            return forward[0]
        if len(forward) > 1:
            return None
        # 3. Reverse: ticket machine name contains the DB asset name (Python filter)
        name_lower = name.lower()
        reverse = [a for a in all_assets if len(a.name) > 4 and a.name.lower() in name_lower]
        if not reverse:
            return None
        reverse.sort(key=lambda a: len(a.name), reverse=True)
        if len(reverse) == 1:
            return reverse[0]
        # Multiple candidates: accept only if top is clearly more specific than second
        if len(reverse[0].name) > len(reverse[1].name) + 3:
            return reverse[0]
    except Exception:
        pass
    return None


@require_POST
@login_required
def api_import_csv(request):
    from django.utils.timezone import make_aware

    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if not (legacy_user and is_legacy_admin(legacy_user)):
        return _json_err("Non autorizzato", 403)

    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        return _json_err("Nessun file caricato")

    filename = (csv_file.name or "").lower()
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xlsm")

    if is_excel:
        try:
            rows = _xlsx_rows_to_dicts(csv_file)
        except Exception as exc:
            return _json_err(f"Impossibile leggere il file Excel: {exc}")
        reader = iter(rows)
    else:
        try:
            raw = csv_file.read()
            try:
                content = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                content = raw.decode("latin-1")
        except Exception:
            return _json_err("Impossibile leggere il file. Usa codifica UTF-8.")
        reader = csv.DictReader(io.StringIO(content))

    # Preload all assets once for reverse-matching (avoids N+1)
    try:
        from assets.models import Asset as _AssetModel
        _all_assets = list(_AssetModel.objects.only("id", "name"))
    except Exception:
        _all_assets = []
    update_existing = request.POST.get("update_existing") == "1"
    tipo_override   = (request.POST.get("tipo_override") or "").strip().upper()

    VALID_TIPI     = {TipoTicket.IT, TipoTicket.MAN}
    VALID_STATI    = set(dict(StatoTicket.choices).keys())
    VALID_PRIORITA = set(dict(PrioritaTicket.choices).keys())

    # Normalizza priorità italiana (SharePoint) → valore interno
    _PRIO_MAP = {
        "alta": "ALTA", "media": "MEDIA", "bassa": "BASSA",
        "urgente": "URGENTE", "critica": "URGENTE",
    }

    def _get(row, *keys):
        """Cerca il primo campo non-vuoto tra più alias di colonna."""
        for k in keys:
            v = row.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""

    created = 0
    updated = 0
    skipped = 0
    linked  = 0
    errors  = []

    def _parse_dt(raw: str):
        for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return make_aware(datetime.strptime(raw.strip(), fmt))
            except (ValueError, OverflowError):
                continue
        return None

    def _build_fields(row, tipo, priorita, stato, incide_sicurezza) -> dict:
        return dict(
            tipo=tipo,
            titolo=_get(row, "titolo", "Titolo", "Problema")[:300],
            descrizione=_get(row, "descrizione", "Descrizione del problema"),
            categoria=_get(row, "categoria", "Tipologia problema", "Tipologia")[:30],
            priorita=priorita,
            stato=stato,
            incide_sicurezza=incide_sicurezza,
            richiedente_nome=_get(row, "richiedente_nome", "Problema registrato da")[:200],
            richiedente_email=_get(row, "richiedente_email")[:200],
            assegnato_a=_get(row, "assegnato_a", "Assegnato a:", "Assegnato a")[:200],
            assegnato_email=_get(row, "assegnato_email")[:200],
            note_interne=_get(row, "note_interne", "Note"),
            asset_descrizione_libera=_get(row, "asset_descrizione_libera", "Macchinario")[:300],
            sharepoint_item_id=_get(row, "sharepoint_item_id", "ID")[:100],
        )

    for row_num, row in enumerate(reader, start=2):
        # tipo: da override POST oppure da colonna CSV
        if tipo_override in VALID_TIPI:
            tipo = tipo_override
        else:
            tipo = (row.get("tipo") or "").strip().upper()
            if tipo not in VALID_TIPI:
                errors.append(f"Riga {row_num}: tipo '{tipo}' non valido (usa IT o MAN)")
                continue

        titolo           = _get(row, "titolo", "Titolo", "Problema")
        descrizione      = _get(row, "descrizione", "Descrizione del problema")
        categoria        = _get(row, "categoria", "Tipologia problema", "Tipologia")
        richiedente_nome = _get(row, "richiedente_nome", "Problema registrato da")

        if not titolo or not descrizione or not categoria or not richiedente_nome:
            errors.append(
                f"Riga {row_num}: campi obbligatori mancanti "
                "(titolo/Problema, descrizione/Descrizione del problema, "
                "categoria/Tipologia, richiedente_nome/Problema registrato da)"
            )
            continue

        # Priorità: normalizza sia formato interno che italiano SharePoint
        priorita_raw = _get(row, "priorita", "Priorità").upper()
        priorita = _PRIO_MAP.get(priorita_raw.lower(), priorita_raw)
        if priorita not in VALID_PRIORITA:
            priorita = PrioritaTicket.MEDIA

        # Stato: da colonna "stato" oppure derivato da "Chiuso" (formato SharePoint)
        stato_raw = (row.get("stato") or "").strip().upper()
        if stato_raw in VALID_STATI:
            stato = stato_raw
        else:
            chiuso_raw = _get(row, "Chiuso").lower()
            stato = StatoTicket.CHIUSO if chiuso_raw in ("1", "true", "si", "sì", "yes") else StatoTicket.APERTA

        incide_raw       = _get(row, "incide_sicurezza", "Impatto su sicurezza").lower()
        incide_sicurezza = incide_raw in ("1", "true", "si", "sì", "yes")
        if incide_sicurezza:
            priorita = PrioritaTicket.URGENTE

        numero_ticket = (row.get("numero_ticket") or "").strip()
        sp_item_id    = _get(row, "sharepoint_item_id", "ID")
        fields        = _build_fields(row, tipo, priorita, stato, incide_sicurezza)

        # ── Deduplicazione: prima per numero_ticket, poi per sharepoint_item_id ──
        existing = None
        if numero_ticket:
            existing = Ticket.objects.filter(numero_ticket=numero_ticket).first()
        if not existing and sp_item_id:
            existing = Ticket.objects.filter(sharepoint_item_id=sp_item_id, tipo=tipo).first()

        if existing:
            if not update_existing:
                skipped += 1
                continue
            for attr, val in fields.items():
                setattr(existing, attr, val)
            existing.save()
            ticket = existing
            updated += 1
        else:
            ticket = Ticket(**fields)
            if numero_ticket:
                ticket.numero_ticket = numero_ticket
            else:
                # Usa l'anno da created_at del CSV per il numero ticket (storico)
                from .models import _next_ticket_number
                created_at_raw_pre = _get(row, "created_at", "Data/ora creazione", "Data segnalazione")
                pre_year = None
                if created_at_raw_pre:
                    dt_pre = _parse_dt(created_at_raw_pre)
                    if dt_pre:
                        pre_year = dt_pre.year
                ticket.numero_ticket = _next_ticket_number(tipo, pre_year)
            ticket.save()
            created += 1

        # Campi auto_now_add / nullable: aggiornati con .update()
        dt_updates = {}
        created_at_raw = _get(row, "created_at", "Data/ora creazione", "Data segnalazione")
        if created_at_raw:
            dt_val = _parse_dt(created_at_raw)
            if dt_val:
                dt_updates["created_at"] = dt_val

        closed_at_raw = _get(row, "closed_at", "Data chiusura ticket", "Data fine lavori")
        if closed_at_raw:
            dt_val = _parse_dt(closed_at_raw)
            if dt_val:
                dt_updates["closed_at"] = dt_val
        elif not closed_at_raw and update_existing:
            dt_updates["closed_at"] = None

        if dt_updates:
            Ticket.objects.filter(pk=ticket.pk).update(**dt_updates)

        # ── Auto-link asset: se asset_descrizione_libera è valorizzato e FK è vuota ──
        if not ticket.asset_id and ticket.asset_descrizione_libera:
            matched = _resolve_asset_by_name(ticket.asset_descrizione_libera, _all_assets)
            if matched:
                Ticket.objects.filter(pk=ticket.pk).update(asset=matched)
                linked += 1

    return JsonResponse({
        "ok":      True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "linked":  linked,
        "errors":  errors[:30],
    })


# ---------------------------------------------------------------------------
# API: azioni bulk su più ticket (solo gestori)
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_bulk(request):
    try:
        payload = json.loads(request.body)
        ids     = [int(x) for x in (payload.get("ids") or [])]
        azione  = (payload.get("azione") or "").strip()
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    if not ids:
        return _json_err("Nessun ticket selezionato")
    if len(ids) > 200:
        return _json_err("Troppi ticket (max 200)")

    tickets = list(Ticket.objects.filter(pk__in=ids))
    if not tickets:
        return _json_err("Ticket non trovati")

    # SEC: opera solo sui ticket di cui l'utente è gestore del TIPO specifico (o
    # admin). _tickets_gestione_required garantisce solo "gestore di almeno un tipo".
    tickets = [t for t in tickets if _api_require_ticket_access(request, t) is None]
    if not tickets:
        return _json_err("Accesso non consentito ai ticket selezionati", status=403)

    name, email, _ = _legacy_identity(request)
    aggiornati = 0

    # ── Cambia stato ──────────────────────────────────────────────────────────
    if azione == "stato":
        nuovo_stato = (payload.get("stato") or "").strip().upper()
        nota        = (payload.get("nota") or "").strip()
        if nuovo_stato not in dict(StatoTicket.choices):
            return _json_err("Stato non valido")

        label_nuovo = dict(StatoTicket.choices).get(nuovo_stato, nuovo_stato)
        for t in tickets:
            vecchio = t.stato
            t.stato = nuovo_stato
            if nuovo_stato in (StatoTicket.CHIUSO, StatoTicket.ANNULLATO, StatoTicket.RISOLTO):
                if not t.closed_at:
                    t.closed_at = tz_now()
            else:
                t.closed_at = None
            t.save(update_fields=["stato", "closed_at", "updated_at"])
            testo_auto = f"Stato aggiornato (bulk): {dict(StatoTicket.choices).get(vecchio, vecchio)} → {label_nuovo}"
            if nota:
                testo_auto += f"\n{nota}"
            TicketCommento.objects.create(
                ticket=t, autore_nome=name, autore_email=email,
                testo=testo_auto, is_interno=True,
            )
            try:
                _update_ticket_sharepoint(t)
            except Exception:
                pass
            aggiornati += 1

    # ── Assegna tecnico ───────────────────────────────────────────────────────
    elif azione == "assegna":
        assegnato_a     = (payload.get("assegnato_a") or "").strip()[:200]
        assegnato_email = (payload.get("assegnato_email") or "").strip()[:200]
        for t in tickets:
            t.assegnato_a     = assegnato_a
            t.assegnato_email = assegnato_email
            if t.stato == StatoTicket.APERTA and assegnato_a:
                t.stato = StatoTicket.IN_CARICO
            t.save(update_fields=["assegnato_a", "assegnato_email", "stato", "updated_at"])
            TicketCommento.objects.create(
                ticket=t, autore_nome=name, autore_email=email,
                testo=f"Ticket assegnato (bulk) a: {assegnato_a or '—'}",
                is_interno=True,
            )
            try:
                _update_ticket_sharepoint(t)
            except Exception:
                pass
            aggiornati += 1

    else:
        return _json_err(f"Azione non riconosciuta: {azione}")

    return JsonResponse({"ok": True, "aggiornati": aggiornati})


# ---------------------------------------------------------------------------
# API: aggiornamento campi analitici ticket
# ---------------------------------------------------------------------------

@require_POST
@_tickets_gestione_required
def api_ticket_analytics(request):
    """Salva i campi analitici (componente, causa_radice, tipo_fermo, ore_fermo,
    ricorrente, ticket_origine) sul ticket indicato."""
    try:
        payload   = json.loads(request.body)
        ticket_id = int(payload.get("ticket_id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)  # SEC: gestore del TIPO specifico
    if denied is not None:
        return denied
    update_fields = ["updated_at"]

    if "componente" in payload:
        ticket.componente = (payload["componente"] or "").strip()[:200]
        update_fields.append("componente")
    if "causa_radice" in payload:
        ticket.causa_radice = (payload["causa_radice"] or "").strip()[:500]
        update_fields.append("causa_radice")
    if "tipo_fermo" in payload:
        valore = (payload["tipo_fermo"] or "").strip().upper()
        if valore not in dict(TipoFermo.choices):
            return _json_err("tipo_fermo non valido")
        ticket.tipo_fermo = valore
        update_fields.append("tipo_fermo")
    if "ore_fermo_macchina" in payload:
        try:
            ore = payload["ore_fermo_macchina"]
            ticket.ore_fermo_macchina = float(ore) if ore not in (None, "") else None
        except (TypeError, ValueError):
            return _json_err("ore_fermo_macchina non valido")
        update_fields.append("ore_fermo_macchina")
    if "ricorrente" in payload:
        ticket.ricorrente = bool(payload["ricorrente"])
        update_fields.append("ricorrente")
    if "ticket_origine_id" in payload:
        orig_id = payload["ticket_origine_id"]
        if orig_id:
            try:
                orig_id = int(orig_id)
                if orig_id == ticket.pk:
                    return _json_err("Il ticket origine non può essere se stesso")
                if not Ticket.objects.filter(pk=orig_id).exists():
                    return _json_err("Ticket origine non trovato")
                ticket.ticket_origine_id = orig_id
            except (TypeError, ValueError):
                return _json_err("ticket_origine_id non valido")
        else:
            ticket.ticket_origine_id = None
        update_fields.append("ticket_origine_id")

    ticket.save(update_fields=update_fields)
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# API: CRUD interventi (sessioni di lavoro)
# ---------------------------------------------------------------------------

@_tickets_gestione_required
def api_intervento(request):
    if request.method == "POST":
        return _api_intervento_create(request)
    if request.method == "PATCH":
        return _api_intervento_update(request)
    if request.method == "DELETE":
        return _api_intervento_delete(request)
    return _json_err("Metodo non supportato", status=405)


def _api_intervento_create(request):
    try:
        payload   = json.loads(request.body)
        ticket_id = int(payload.get("ticket_id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)
    if denied is not None:
        return denied
    name, email, _ = _legacy_identity(request)

    try:
        data_inizio = datetime.fromisoformat(payload["data_inizio"])
    except (KeyError, ValueError, TypeError):
        return _json_err("data_inizio non valida (formato ISO 8601)")

    data_fine = None
    if payload.get("data_fine"):
        try:
            data_fine = datetime.fromisoformat(payload["data_fine"])
        except (ValueError, TypeError):
            return _json_err("data_fine non valida (formato ISO 8601)")

    ore_lavorate = None
    if payload.get("ore_lavorate") not in (None, ""):
        try:
            ore_lavorate = float(payload["ore_lavorate"])
            if ore_lavorate < 0:
                return _json_err("ore_lavorate non può essere negativo")
        except (TypeError, ValueError):
            return _json_err("ore_lavorate non valido")

    esito = (payload.get("esito") or EsitoIntervento.IN_CORSO).strip().upper()
    if esito not in dict(EsitoIntervento.choices):
        return _json_err("esito non valido")

    tecnico_nome  = (payload.get("tecnico_nome") or name).strip()[:200]
    tecnico_email = (payload.get("tecnico_email") or email).strip()[:200]

    def _decimal_or_none(val):
        if val in (None, ""):
            return None
        try:
            v = float(val)
            return v if v >= 0 else None
        except (TypeError, ValueError):
            return None

    interv = TicketIntervento.objects.create(
        ticket=ticket,
        tecnico_nome=tecnico_nome,
        tecnico_email=tecnico_email,
        data_inizio=data_inizio,
        data_fine=data_fine,
        ore_lavorate=ore_lavorate,
        descrizione_lavoro=(payload.get("descrizione_lavoro") or "").strip(),
        componente_interessato=(payload.get("componente_interessato") or "").strip()[:200],
        azioni_svolte=(payload.get("azioni_svolte") or "").strip(),
        esito=esito,
        note=(payload.get("note") or "").strip(),
        costo_manodopera_eur=_decimal_or_none(payload.get("costo_manodopera_eur")),
        materiali_costo_eur=_decimal_or_none(payload.get("materiali_costo_eur")),
        numero_rapportino=(payload.get("numero_rapportino") or "").strip()[:50],
    )

    # Aggiorna data_primo_intervento sul ticket se è il primo
    if not ticket.data_primo_intervento:
        ticket.data_primo_intervento = interv.created_at
        ticket.save(update_fields=["data_primo_intervento", "updated_at"])

    return JsonResponse({
        "ok": True,
        "id": interv.pk,
        "tecnico_nome": interv.tecnico_nome,
        "data_inizio": interv.data_inizio.strftime("%d-%m-%Y %H:%M"),
        "esito": interv.esito,
        "label_esito": interv.label_esito,
        "durata_ore": interv.durata_ore,
    })


def _api_intervento_update(request):
    try:
        payload      = json.loads(request.body)
        intervento_id = int(payload.get("id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    interv = get_object_or_404(TicketIntervento.objects.select_related("ticket"), pk=intervento_id)
    denied = _api_require_ticket_access(request, interv.ticket)
    if denied is not None:
        return denied
    update_fields = ["updated_at"]

    if "data_fine" in payload:
        if payload["data_fine"]:
            try:
                interv.data_fine = datetime.fromisoformat(payload["data_fine"])
                update_fields.append("data_fine")
            except (ValueError, TypeError):
                return _json_err("data_fine non valida")
        else:
            interv.data_fine = None
            update_fields.append("data_fine")

    if "ore_lavorate" in payload:
        ore = payload["ore_lavorate"]
        interv.ore_lavorate = float(ore) if ore not in (None, "") else None
        update_fields.append("ore_lavorate")

    if "esito" in payload:
        esito = (payload["esito"] or "").strip().upper()
        if esito not in dict(EsitoIntervento.choices):
            return _json_err("esito non valido")
        interv.esito = esito
        update_fields.append("esito")

    for campo in ("descrizione_lavoro", "componente_interessato", "azioni_svolte", "note", "numero_rapportino"):
        if campo in payload:
            setattr(interv, campo, (payload[campo] or "").strip())
            update_fields.append(campo)

    for campo in ("costo_manodopera_eur", "materiali_costo_eur"):
        if campo in payload:
            val = payload[campo]
            try:
                setattr(interv, campo, float(val) if val not in (None, "") else None)
            except (TypeError, ValueError):
                pass
            else:
                update_fields.append(campo)

    interv.save(update_fields=update_fields)
    return JsonResponse({"ok": True, "durata_ore": interv.durata_ore, "label_esito": interv.label_esito})


def _api_intervento_delete(request):
    try:
        payload      = json.loads(request.body)
        intervento_id = int(payload.get("id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    interv = get_object_or_404(TicketIntervento.objects.select_related("ticket"), pk=intervento_id)
    ticket = interv.ticket
    denied = _api_require_ticket_access(request, ticket)
    if denied is not None:
        return denied
    interv.delete()

    # Ricalcola data_primo_intervento
    primo = ticket.interventi.order_by("data_inizio").first()
    ticket.data_primo_intervento = primo.created_at if primo else None
    ticket.save(update_fields=["data_primo_intervento", "updated_at"])

    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# API: CRUD componenti sostituiti (ricambi per sessione di intervento)
# ---------------------------------------------------------------------------

@_tickets_gestione_required
def api_componente(request):
    if request.method == "POST":
        return _api_componente_create(request)
    if request.method == "DELETE":
        return _api_componente_delete(request)
    return _json_err("Metodo non supportato", status=405)


def _api_componente_create(request):
    from .models import TicketComponenteSostituito
    try:
        payload       = json.loads(request.body)
        intervento_id = int(payload.get("intervento_id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    interv = get_object_or_404(TicketIntervento.objects.select_related("ticket"), pk=intervento_id)
    denied = _api_require_ticket_access(request, interv.ticket)
    if denied is not None:
        return denied
    nome   = (payload.get("nome") or "").strip()[:200]
    if not nome:
        return _json_err("Nome componente obbligatorio")

    def _dec(val):
        if val in (None, ""):
            return None
        try:
            v = float(val)
            return v if v >= 0 else None
        except (TypeError, ValueError):
            return None

    comp = TicketComponenteSostituito.objects.create(
        intervento=interv,
        nome=nome,
        codice_parte=(payload.get("codice_parte") or "").strip()[:100],
        quantita=max(1, int(payload.get("quantita") or 1)),
        costo_unitario_eur=_dec(payload.get("costo_unitario_eur")),
        note=(payload.get("note") or "").strip()[:300],
    )
    return JsonResponse({
        "ok": True,
        "id": comp.pk,
        "nome": comp.nome,
        "codice_parte": comp.codice_parte,
        "quantita": comp.quantita,
        "costo_unitario_eur": float(comp.costo_unitario_eur) if comp.costo_unitario_eur is not None else None,
        "costo_totale_eur": comp.costo_totale_eur,
        "note": comp.note,
    })


def _api_componente_delete(request):
    from .models import TicketComponenteSostituito
    try:
        payload = json.loads(request.body)
        comp_id = int(payload.get("id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    comp = get_object_or_404(
        TicketComponenteSostituito.objects.select_related("intervento__ticket"), pk=comp_id
    )
    denied = _api_require_ticket_access(request, comp.intervento.ticket)
    if denied is not None:
        return denied
    comp.delete()
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# API: Crea WorkOrder (manutenzione straordinaria) da ticket
# ---------------------------------------------------------------------------

@_tickets_gestione_required
def api_crea_workorder_da_ticket(request):
    if request.method != "POST":
        return _json_err("Metodo non supportato", status=405)

    try:
        payload   = json.loads(request.body)
        ticket_id = int(payload.get("ticket_id") or 0)
    except (json.JSONDecodeError, ValueError):
        return _json_err("Dati non validi")

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    denied = _api_require_ticket_access(request, ticket)
    if denied is not None:
        return denied

    if not ticket.asset_id:
        return _json_err("Il ticket non ha un asset collegato. Collegare prima l'asset dalla tab 'Asset e KPI'.")

    from assets.models import WorkOrder

    title       = (payload.get("title") or ticket.titolo).strip()[:255]
    description = (payload.get("description") or ticket.descrizione).strip()

    wo = WorkOrder.objects.create(
        asset=ticket.asset,
        ticket=ticket,
        supplier=ticket.delegato_fornitore,
        kind=WorkOrder.KIND_CORRECTIVE,
        title=title,
        description=description,
    )

    wo_url = reverse("assets:wo_view", args=[wo.pk])
    return JsonResponse({
        "ok": True,
        "workorder_id": wo.pk,
        "workorder_url": wo_url,
        "numero": wo.pk,
        "title": wo.title,
        "asset_name": ticket.asset.name,
        "asset_tag": ticket.asset.asset_tag,
    })
