"""Import assenze da file Excel (.xlsx) verso la tabella legacy ``assenze``.

Pensato per l'export della lista SharePoint "Calendario assenze", con header in
prima riga e colonne ``Data inizio | Data fine | Nome Cognome | Tipoassenza |
Stato approvazione`` (le colonne accessorie dell'export - Nome, Ricercanome,
Tipo di elemento, Percorso - vengono ignorate).

Comportamento idempotente:

* riga assente in DB            -> INSERT
* riga presente e diversa       -> UPDATE dei soli campi cambiati
* riga presente e identica      -> nessuna scrittura (conteggiata "invariate")

La corrispondenza usa la chiave naturale ``(nominativo, giorno inizio, giorno
fine)``: fra i candidati dello stesso giorno vince prima la riga con orari
identici, poi quella con lo stesso tipo di assenza, infine la piu' vecchia.
Ogni riga di DB puo' essere abbinata a una sola riga del file.

NOTA SharePoint: il comando scrive solo sul DB locale. Le righe inserite non
hanno ``sharepoint_item_id`` e verrebbero quindi CREATE su SharePoint dal push
di sincronizzazione (``/assenze/api/sync/push``), generando doppioni nella
lista. Se la lista SharePoint e' ancora attiva, la via corretta e'
``manage.py sync_assenze_sharepoint --force``.

Esempi:
    python manage.py import_assenze_xlsx assenze.xlsx --dry-run
    python manage.py import_assenze_xlsx assenze.xlsx
    python manage.py import_assenze_xlsx assenze.xlsx --sheet "query (3)" --limit 20 --verbose
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import connections, transaction
from django.utils import timezone

from core.legacy_utils import legacy_table_columns

from assenze.views import (
    _CONSENSO_TO_MOD,
    _as_int,
    _fetch_all_dict,
    _insert_row_and_return_id,
    _norm_consenso,
    _quote_identifier,
    _quoted_columns,
    _tipo_for_storage,
)


# ---------------------------------------------------------------------------
# Header mapping
# ---------------------------------------------------------------------------

def _norm_header(raw) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(raw or "").lower())


HEADER_ALIAS: dict[str, str] = {
    "datainizio": "data_inizio",
    "datafine": "data_fine",
    "nomecognome": "copia_nome",
    "nominativo": "copia_nome",
    "dipendente": "copia_nome",
    "tipoassenza": "tipo_assenza",
    "tipodiassenza": "tipo_assenza",
    "statoapprovazione": "consenso",
    "consenso": "consenso",
    "motivazionerichiesta": "motivazione_richiesta",
    "motivazione": "motivazione_richiesta",
    "emailesterna": "email_esterna",
    # L'export SharePoint ripete il nominativo in "Nome": usato solo se manca
    # la colonna "Nome Cognome".
    "nome": "copia_nome_alt",
    # Colonne dell'export senza corrispondenza in DB
    "ricercanome": "_skip",
    "tipodielemento": "_skip",
    "percorso": "_skip",
    "id": "_skip",
}

# Campi confrontati (e aggiornati) quando la riga esiste gia'.
BASE_COMPARE_FIELDS = ("tipo_assenza", "consenso", "moderation_status", "data_inizio", "data_fine")
OPTIONAL_COMPARE_FIELDS = ("motivazione_richiesta", "email_esterna")


def _build_header_index(header_row) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, raw in enumerate(header_row):
        if raw is None:
            continue
        mapped = HEADER_ALIAS.get(_norm_header(raw))
        if mapped and mapped != "_skip" and mapped not in out:
            out[mapped] = i
    return out


# ---------------------------------------------------------------------------
# Parsing valori
# ---------------------------------------------------------------------------

_DT_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
)


def _coerce_dt(value) -> datetime | None:
    """Converte un valore Excel/DB in datetime naive (ora locale)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        dt = value
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt, timezone.get_current_timezone()).replace(tzinfo=None)
        return dt.replace(microsecond=0)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    raw = str(value).strip()
    if not raw:
        return None
    parsed = None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        parsed = None
    if parsed is None:
        for fmt in _DT_FORMATS:
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
    if parsed is None:
        return None
    if timezone.is_aware(parsed):
        parsed = timezone.localtime(parsed, timezone.get_current_timezone()).replace(tzinfo=None)
    return parsed.replace(microsecond=0)


def _dt_key(value) -> datetime | None:
    """Datetime al minuto: i secondi non sono significativi per il confronto."""
    dt = _coerce_dt(value)
    return dt.replace(second=0, microsecond=0) if dt else None


def _name_key(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _norm_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _compare_value(field: str, value):
    """Valore normalizzato per il confronto file <-> DB."""
    if field in {"data_inizio", "data_fine"}:
        return _dt_key(value)
    if field == "tipo_assenza":
        return _tipo_for_storage(value)
    if field == "consenso":
        return _norm_consenso(value)
    if field == "moderation_status":
        return _as_int(value)
    if field == "email_esterna":
        return _norm_text(value).lower()
    return _norm_text(value)


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class _DryRunAbort(Exception):
    pass


class Command(BaseCommand):
    help = (
        "Importa le assenze da un file Excel (.xlsx): inserisce le nuove, "
        "aggiorna le modificate, salta le invariate."
    )

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Percorso al file .xlsx")
        parser.add_argument("--sheet", type=str, default="", help="Nome foglio (default: primo)")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla, mostra solo il report")
        parser.add_argument("--limit", type=int, default=0, help="Limita a N righe (0 = tutte)")
        parser.add_argument("--verbose", action="store_true", help="Mostra ogni riga processata")

    # -- lettura file -------------------------------------------------------

    def _read_rows(self, opts) -> tuple[list[dict], dict[str, int], dict[str, int]]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - dipendenza dichiarata in requirements
            raise CommandError("openpyxl non installato.") from exc

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        # read_only tiene aperto il file finche' non si chiude il workbook.
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            return self._parse_workbook(wb, opts)
        finally:
            wb.close()

    def _parse_workbook(self, wb, opts) -> tuple[list[dict], dict[str, int], dict[str, int]]:
        sheet_name = opts.get("sheet") or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' non trovato. Disponibili: {wb.sheetnames}")
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise CommandError("Foglio vuoto.") from exc

        idx = _build_header_index(header_row)
        if "copia_nome" not in idx and "copia_nome_alt" in idx:
            idx["copia_nome"] = idx["copia_nome_alt"]
        missing = [k for k in ("copia_nome", "data_inizio", "data_fine", "tipo_assenza") if k not in idx]
        if missing:
            raise CommandError(
                f"Colonne obbligatorie mancanti: {missing}\nHeader letto: {list(header_row)}"
            )

        limit = int(opts.get("limit") or 0)
        parsed: list[dict] = []
        skipped: dict[str, int] = {"dati_mancanti": 0, "date_non_valide": 0}

        def cell(row, key):
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        for excel_row, raw_row in enumerate(rows_iter, start=2):
            if limit and len(parsed) >= limit:
                break
            if raw_row is None or all(c in (None, "") for c in raw_row):
                continue

            nome = _norm_text(cell(raw_row, "copia_nome"))
            if not nome:
                skipped["dati_mancanti"] += 1
                continue

            dt_inizio = _coerce_dt(cell(raw_row, "data_inizio"))
            dt_fine = _coerce_dt(cell(raw_row, "data_fine"))
            if dt_inizio is None or dt_fine is None:
                skipped["date_non_valide"] += 1
                continue

            consenso = _norm_consenso(cell(raw_row, "consenso"))
            record = {
                "_excel_row": excel_row,
                "copia_nome": nome,
                "tipo_assenza": _tipo_for_storage(cell(raw_row, "tipo_assenza")),
                "data_inizio": dt_inizio,
                "data_fine": dt_fine,
                "consenso": consenso,
                "moderation_status": _CONSENSO_TO_MOD.get(consenso, 2),
            }
            if "motivazione_richiesta" in idx:
                record["motivazione_richiesta"] = _norm_text(cell(raw_row, "motivazione_richiesta"))
            if "email_esterna" in idx:
                record["email_esterna"] = _norm_text(cell(raw_row, "email_esterna")).lower()
            parsed.append(record)

        return parsed, idx, skipped

    # -- indice righe gia' presenti -----------------------------------------

    def _load_existing(self, records: list[dict], cols: set[str]) -> dict[tuple, list[dict]]:
        wanted = [
            "id",
            "copia_nome",
            "data_inizio",
            "data_fine",
            "tipo_assenza",
            "consenso",
            "moderation_status",
            "motivazione_richiesta",
            "email_esterna",
            "sharepoint_item_id",
        ]
        selected = [c for c in wanted if c in cols]
        sql = f"SELECT {_quoted_columns(selected)} FROM assenze"
        params: list = []
        # Riduce la finestra letta: un match richiede lo stesso giorno di inizio.
        starts = [r["data_inizio"] for r in records]
        if starts:
            lo = min(starts).replace(hour=0, minute=0, second=0, microsecond=0)
            hi = max(starts).replace(hour=23, minute=59, second=59, microsecond=0)
            sql += " WHERE data_inizio >= %s AND data_inizio <= %s"
            params = [lo, hi]

        index: dict[tuple, list[dict]] = {}
        for row in _fetch_all_dict(sql, params):
            dt_i = _coerce_dt(row.get("data_inizio"))
            dt_f = _coerce_dt(row.get("data_fine"))
            if dt_i is None or dt_f is None:
                continue
            key = (_name_key(row.get("copia_nome")), dt_i.date(), dt_f.date())
            index.setdefault(key, []).append(row)
        for rows in index.values():
            rows.sort(key=lambda r: _as_int(r.get("id")) or 0)
        return index

    @staticmethod
    def _pick_candidate(candidates: list[dict], record: dict) -> dict | None:
        if not candidates:
            return None
        for row in candidates:
            if (
                _dt_key(row.get("data_inizio")) == _dt_key(record["data_inizio"])
                and _dt_key(row.get("data_fine")) == _dt_key(record["data_fine"])
            ):
                return row
        for row in candidates:
            if _tipo_for_storage(row.get("tipo_assenza")) == record["tipo_assenza"]:
                return row
        return candidates[0]

    # -- handle -------------------------------------------------------------

    def handle(self, *args, **opts):
        cols = legacy_table_columns("assenze")
        if not cols:
            raise CommandError("Tabella 'assenze' non disponibile sul database configurato.")
        for required in ("copia_nome", "data_inizio", "data_fine"):
            if required not in cols:
                raise CommandError(f"Tabella 'assenze' priva della colonna '{required}'.")

        records, idx, skipped = self._read_rows(opts)
        dry_run = bool(opts.get("dry_run"))
        verbose = bool(opts.get("verbose"))

        compare_fields = [f for f in BASE_COMPARE_FIELDS if f in cols]
        compare_fields += [f for f in OPTIONAL_COMPARE_FIELDS if f in cols and f in idx]

        self.stdout.write(
            self.style.NOTICE(
                f"Import assenze da {Path(opts['file']).name} "
                f"[sheet={opts.get('sheet') or 'primo'}, righe={len(records)}, dry_run={dry_run}]"
            )
        )

        stats = {
            "lette": len(records),
            "create": 0,
            "aggiornate": 0,
            "invariate": 0,
            "saltate_dati_mancanti": skipped["dati_mancanti"],
            "saltate_date_non_valide": skipped["date_non_valide"],
            "errori": 0,
        }
        errors: list[str] = []
        create_senza_sp = 0

        try:
            with transaction.atomic():
                existing = self._load_existing(records, cols)
                used_ids: set[int] = set()

                with connections["default"].cursor() as cursor:
                    for record in records:
                        excel_row = record.pop("_excel_row")
                        key = (
                            _name_key(record["copia_nome"]),
                            record["data_inizio"].date(),
                            record["data_fine"].date(),
                        )
                        candidates = [
                            row for row in existing.get(key, [])
                            if (_as_int(row.get("id")) or 0) not in used_ids
                        ]
                        match = self._pick_candidate(candidates, record)

                        try:
                            if match is None:
                                payload = {k: v for k, v in record.items() if k in cols}
                                insert_cols = list(payload.keys())
                                new_id = _insert_row_and_return_id(
                                    cursor, "assenze", insert_cols, [payload[c] for c in insert_cols]
                                )
                                stats["create"] += 1
                                create_senza_sp += 1
                                if verbose:
                                    nome = record["copia_nome"]
                                    self.stdout.write(
                                        f"  riga {excel_row}: CREATA id={new_id} {nome} {record['tipo_assenza']}"
                                    )
                                continue

                            row_id = _as_int(match.get("id"))
                            used_ids.add(row_id or 0)
                            updates = {
                                field: record[field]
                                for field in compare_fields
                                if field in record
                                and _compare_value(field, match.get(field))
                                != _compare_value(field, record[field])
                            }
                            if not updates:
                                stats["invariate"] += 1
                                if verbose:
                                    self.stdout.write(f"  riga {excel_row}: invariata id={row_id}")
                                continue

                            sets = ", ".join(f"{_quote_identifier(k)} = %s" for k in updates)
                            cursor.execute(
                                f"UPDATE assenze SET {sets} WHERE id = %s",
                                [*updates.values(), row_id],
                            )
                            stats["aggiornate"] += 1
                            if verbose:
                                self.stdout.write(
                                    f"  riga {excel_row}: AGGIORNATA id={row_id} campi={sorted(updates)}"
                                )
                        except Exception as exc:  # pragma: no cover - difensivo
                            stats["errori"] += 1
                            errors.append(f"riga {excel_row} ({record['copia_nome']}): {exc}")

                if dry_run:
                    raise _DryRunAbort()
        except _DryRunAbort:
            pass

        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("Riepilogo"))
        for label, value in stats.items():
            self.stdout.write(f"  {label:<26} {value}")

        if errors:
            self.stdout.write("")
            self.stdout.write(self.style.ERROR(f"Errori ({len(errors)}):"))
            for line in errors[:20]:
                self.stdout.write(f"  - {line}")

        if create_senza_sp and not dry_run:
            self.stdout.write("")
            self.stdout.write(
                self.style.WARNING(
                    f"{create_senza_sp} righe inserite senza sharepoint_item_id: un push di "
                    "sincronizzazione le creerebbe come nuovi elementi nella lista SharePoint."
                )
            )

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna modifica salvata."))
        else:
            self.stdout.write(self.style.SUCCESS("Import completato."))
