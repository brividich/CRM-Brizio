"""Import di un MOD.174 SGI (Excel compilato) nel registro OFI centralizzato.

Legge il foglio col layout canonico del MOD.174 (intestazioni a riga 4, dati da
riga 5, colonne A→T) e mappa ogni riga su ``RegistroOFI``. La **fase PDCA** è
derivata dalle X cumulative nelle colonne P/Q/R/S (1→PLAN … 4→ACT; CHIUSO se
``DATA CLOSED`` è valorizzata), coerente con la colonna TOT del foglio (che è una
formula e quindi NON viene letta). L'import è **idempotente**: upsert sul numero
di registro (colonna REF), spazio numeri condiviso con l'OFI legacy.

Separato dal management command per essere testabile senza I/O su disco.
"""
from __future__ import annotations

from datetime import date, datetime

# Posizioni colonna (1-based) del MOD.174 — layout fisso del modulo SGI.
COL = {
    "ref": 1, "data": 2, "tipo": 3, "normative": 4, "rif_norma": 5,
    "processo": 6, "opportunita": 7, "plan": 8, "allegato": 9, "do": 10,
    "check": 11, "act": 12, "data_required": 13, "data_closed": 14, "owner": 15,
    "p": 16, "d": 17, "c": 18, "a": 19,  # tot=20 è una formula: ignorata
}

# Colonne che indicano contenuto reale (per saltare le righe pre-numerate vuote).
_CONTENT_COLS = (2, 3, 6, 7, 8, 10, 11, 12, 14, 15)


def _txt(value) -> str:
    return "" if value is None else str(value).strip()


def _is_x(value) -> bool:
    return _txt(value).upper() == "X"


def _to_int(value):
    if value is None or _txt(value) == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_date(value):
    """Converte una cella in ``date`` (datetime/date nativo o stringa gg/mm/aaaa)."""
    if value is None or _txt(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _txt(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_header_row(ws) -> int:
    """Riga delle intestazioni: la prima con colonna A == 'REF' (fallback: 4)."""
    for r in range(1, min(ws.max_row or 1, 30) + 1):
        if _txt(ws.cell(row=r, column=1).value).upper() == "REF":
            return r
    raise ValueError("Il file non sembra un MOD.174: intestazione 'REF' non trovata.")


def _fase_da_marks(marks: int, data_chiusura) -> str:
    from .models import RegistroOFI as R
    if data_chiusura is not None:
        return R.FASE_CHIUSO
    return {1: R.FASE_PLAN, 2: R.FASE_DO, 3: R.FASE_CHECK, 4: R.FASE_ACT}.get(marks, R.FASE_PLAN)


def leggi_righe_mod174(fileobj, *, sheet: str | None = None) -> list[dict]:
    """Legge il MOD.174 e ritorna la lista dei dict-riga (una per OFI compilata).

    ``fileobj`` è un path o un file-like (bytes). Le righe pre-numerate ma prive
    di contenuto reale vengono saltate. Solleva ``ValueError`` se il file non ha
    l'intestazione del MOD.174.
    """
    import openpyxl

    wb = openpyxl.load_workbook(fileobj, data_only=False, read_only=False)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    header = _find_header_row(ws)

    righe: list[dict] = []
    for r in range(header + 1, (ws.max_row or header) + 1):
        # salta le righe senza contenuto reale (le REF pre-stampate 1..N vuote)
        if not any(_txt(ws.cell(row=r, column=c).value) for c in _CONTENT_COLS):
            continue
        numero = _to_int(ws.cell(row=r, column=COL["ref"]).value)
        if not numero:
            continue
        normative = _txt(ws.cell(row=r, column=COL["normative"]).value)
        tipo_raw = _txt(ws.cell(row=r, column=COL["tipo"]).value).upper()
        data_chiusura = _to_date(ws.cell(row=r, column=COL["data_closed"]).value)
        marks = sum(1 for k in ("p", "d", "c", "a") if _is_x(ws.cell(row=r, column=COL[k]).value))
        righe.append({
            "numero": numero,
            "data_apertura": _to_date(ws.cell(row=r, column=COL["data"]).value),
            "tipo": "NC" if "NC" in tipo_raw else "OFI",
            "norma_iso27001": "27001" in normative,
            "norma_iso45001": "45001" in normative,
            "norma_en9100": "9100" in normative,
            "rif_norma": _txt(ws.cell(row=r, column=COL["rif_norma"]).value)[:200],
            "processo": _txt(ws.cell(row=r, column=COL["processo"]).value)[:200],
            "opportunita": _txt(ws.cell(row=r, column=COL["opportunita"]).value),
            "plan": _txt(ws.cell(row=r, column=COL["plan"]).value),
            "allegato_link": _txt(ws.cell(row=r, column=COL["allegato"]).value)[:500],
            "do": _txt(ws.cell(row=r, column=COL["do"]).value),
            "verifica": _txt(ws.cell(row=r, column=COL["check"]).value),
            "act": _txt(ws.cell(row=r, column=COL["act"]).value),
            "data_richiesta": _to_date(ws.cell(row=r, column=COL["data_required"]).value),
            "data_chiusura": data_chiusura,
            "owner_processo": _txt(ws.cell(row=r, column=COL["owner"]).value)[:150],
            "fase": _fase_da_marks(marks, data_chiusura),
        })
    return righe


# Campi del dict-riga scrivibili direttamente su RegistroOFI.
_CAMPI = (
    "data_apertura", "tipo", "norma_iso27001", "norma_iso45001", "norma_en9100",
    "rif_norma", "processo", "opportunita", "plan", "allegato_link", "do",
    "verifica", "act", "data_richiesta", "data_chiusura", "owner_processo", "fase",
)


def importa_voci(righe: list[dict], *, modulo_origine: str = "", dry_run: bool = True) -> dict:
    """Upsert delle righe nel registro OFI (idempotente sul numero).

    ``dry_run`` (default) NON scrive: conta soltanto creati/aggiornati. Ritorna
    ``{creati, aggiornati, saltati, dettagli}``.
    """
    from django.utils import timezone

    from .models import RegistroOFI

    out = {"creati": 0, "aggiornati": 0, "saltati": 0, "dettagli": []}
    for r in righe:
        numero = r.get("numero")
        if not numero:
            out["saltati"] += 1
            continue
        defaults = {k: r.get(k) for k in _CAMPI}
        defaults["data_apertura"] = r.get("data_apertura") or timezone.localdate()
        defaults["modulo_origine"] = modulo_origine

        esiste = RegistroOFI.objects.filter(numero=numero).exists()
        if not dry_run:
            RegistroOFI.objects.update_or_create(numero=numero, defaults=defaults)
        chiave = "aggiornati" if esiste else "creati"
        out[chiave] += 1
        out["dettagli"].append({"numero": numero, "tipo": defaults["tipo"],
                                "processo": defaults["processo"], "fase": defaults["fase"],
                                "azione": chiave})
    return out
