"""Converte i listoni Excel del gestionale -> CSV template `import_specifiche_storico`.

Adattatore di intake (vedi `gestione_specifiche/intake_export.py`):
- `--spte FILE.xls`   : Specifiche Tecniche (gestionale) -> `<out>/specifiche_spte.csv`
                        (+ `<out>/allegati_spte.csv` con i percorsi UNC dei PDF).
- `--cliente FILE.xlsx`: Registro Specifiche Cliente -> `<out>/specifiche_cliente.csv`.

NON scrive su DB: produce solo i CSV, da validare poi con `import_specifiche_storico`
(dry-run) e applicare con `--apply`. Tieni i file e i CSV FUORI dal repo (dati reali).

Esempi::

    python manage.py converti_export_gestionale \\
        --spte "C:/import/OK SPTE.xls" \\
        --cliente "C:/import/MOD.097.xlsx" \\
        --out C:/import/csv
"""
from __future__ import annotations

import csv
import os
from collections import Counter

from django.core.management.base import BaseCommand, CommandError

from gestione_specifiche import intake_export as ix


def _norm(s: str) -> str:
    return " ".join(str(s or "").split()).strip().lower()


class Command(BaseCommand):
    help = "Converte i listoni Excel (SPTE .xls / Specifiche Cliente .xlsx) nei CSV template F8."

    def add_arguments(self, parser):
        parser.add_argument("--spte", help="File .xls del gestionale (Specifiche Tecniche).")
        parser.add_argument("--cliente", help="File .xlsx Registro Specifiche Cliente.")
        parser.add_argument("--out", required=True, help="Cartella di output dei CSV.")
        parser.add_argument("--sheet", default="Registro", help="Foglio del file cliente (default 'Registro').")
        parser.add_argument("--includi-fvali", action="store_true",
                            help="NON escludere le righe SPTE con fvali valorizzato (default: escluse).")

    def handle(self, *args, **opts):
        if not opts.get("spte") and not opts.get("cliente"):
            raise CommandError("Specificare almeno --spte e/o --cliente.")
        out = opts["out"]
        os.makedirs(out, exist_ok=True)

        if opts.get("spte"):
            self._spte(opts["spte"], out, escludi_fvali=not opts["includi_fvali"])
        if opts.get("cliente"):
            self._cliente(opts["cliente"], out, sheet=opts["sheet"])

    # ------------------------------------------------------------------ helpers
    def _scrivi(self, path, righe):
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=ix.EXPECTED_COLUMNS, delimiter=";")
            w.writeheader()
            for r in righe:
                w.writerow({c: r.get(c, "") for c in ix.EXPECTED_COLUMNS})

    # --------------------------------------------------------------------- SPTE
    def _spte(self, path, out, *, escludi_fvali):
        try:
            import xlrd
        except ImportError as e:
            raise CommandError(f"xlrd non installato (serve per leggere .xls): {e}")
        if not os.path.exists(path):
            raise CommandError(f"File SPTE non trovato: {path}")
        wb = xlrd.open_workbook(path)
        sh = wb.sheets()[0]
        H = {self._h(sh.cell_value(0, c)): c for c in range(sh.ncols)}
        cnt = Counter()
        seen = set()
        righe, allegati = [], []
        for r in range(1, sh.nrows):
            cnt["totali"] += 1
            rec = {name: sh.cell_value(r, idx) for name, idx in H.items()}
            res = ix.mappa_spte(rec, escludi_fvali=escludi_fvali)
            if res["riga"] is None:
                cnt["scartati"] += 1
                cnt["scarto:" + res["scarto"]] += 1
                continue
            key = (res["riga"]["codice"].upper(), res["riga"]["revisione"].upper())
            if key in seen:
                cnt["dedup"] += 1
                continue
            seen.add(key)
            righe.append(res["riga"])
            if res["path"]:
                allegati.append({"codice": res["riga"]["codice"],
                                 "revisione": res["riga"]["revisione"], "path": res["path"]})
        csv_path = os.path.join(out, "specifiche_spte.csv")
        self._scrivi(csv_path, righe)
        all_path = os.path.join(out, "allegati_spte.csv")
        with open(all_path, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=["codice", "revisione", "path"], delimiter=";")
            w.writeheader()
            w.writerows(allegati)
        self.stdout.write(self.style.MIGRATE_HEADING("SPTE -> specifiche_spte.csv"))
        self.stdout.write(f"  righe valide: {len(righe)} | con path UNC: {len(allegati)} | {dict(cnt)}")
        self.stdout.write(f"  scritti: {csv_path}\n           {all_path}")

    @staticmethod
    def _h(v):
        return str(v).strip()

    # ------------------------------------------------------------------ CLIENTE
    def _cliente(self, path, out, *, sheet):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(f"openpyxl non installato: {e}")
        if not os.path.exists(path):
            raise CommandError(f"File cliente non trovato: {path}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            raise CommandError(f"Foglio '{sheet}' assente. Fogli: {wb.sheetnames}")
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 5:
            raise CommandError("Foglio cliente troppo corto / struttura inattesa.")
        # intestazione a riga 4 (indice 3); mappa per etichetta, fallback a indici noti.
        hdr = {_norm(h): i for i, h in enumerate(rows[3]) if h}
        col = {
            "cliente": hdr.get("cliente", 0),
            "codice": hdr.get("n° documento", hdr.get("n documento", 1)),
            "revisione": hdr.get("revisione", 2),
            "dataric": hdr.get("data ricevuto", 4),
            "sup": hdr.get("sospese/ superate", hdr.get("sospese/superate", 14)),
            "sosp": hdr.get("sosp.", 16),
        }
        cnt = Counter()
        seen = set()
        righe = []
        for row in rows[4:]:
            if not row:
                continue
            rec = {k: (row[i] if i < len(row) else "") for k, i in col.items()}
            cnt["totali"] += 1
            res = ix.mappa_cliente(rec)
            if res["riga"] is None:
                cnt["scartati"] += 1
                cnt["scarto:" + res["scarto"]] += 1
                continue
            key = (res["riga"]["codice"].upper(), res["riga"]["revisione"].upper())
            if key in seen:
                cnt["dedup"] += 1
                continue
            seen.add(key)
            righe.append(res["riga"])
        csv_path = os.path.join(out, "specifiche_cliente.csv")
        self._scrivi(csv_path, righe)
        clienti = Counter(r["cliente"] for r in righe)
        self.stdout.write(self.style.MIGRATE_HEADING("CLIENTE -> specifiche_cliente.csv"))
        self.stdout.write(f"  righe valide: {len(righe)} | {dict(cnt)}")
        self.stdout.write("  clienti: " + ", ".join(f"{k}x{v}" for k, v in clienti.most_common(8)))
        self.stdout.write(f"  scritto: {csv_path}")
