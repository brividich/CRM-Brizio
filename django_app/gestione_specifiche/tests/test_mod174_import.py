"""Test import MOD.174 SGI (Excel compilato) → registro OFI.

Costruisce in memoria un foglio con lo stesso layout del MOD.174 (header a riga 4,
dati da riga 5, colonne A→T) e verifica parsing, dry-run, apply e idempotenza.
Nessun dato reale.
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
from django.core.management import call_command
from django.test import TestCase

from gestione_specifiche.mod174_import import importa_voci, leggi_righe_mod174
from gestione_specifiche.models import RegistroOFI

_HEADERS = ["REF", "DATA", "OFI\nNC", "Normative (27001,45001, 9100...)", "REF NORMA",
            "PROCESSO", "OPPORTUNITY", "PLAN", "Allegato/Link", "DO ", "CHECK", "ACT",
            "DATA REQUIRED", "DATA CLOSED", "OWNER", "P", "D", "C", "A", "TOT"]


def _fake_mod174() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDCA 27001"
    ws["A1"] = "Costruzioni Novicrom"
    ws["T1"] = "MOD.174 - SGI Registro OFI_NC Rev.0"
    for i, h in enumerate(_HEADERS, start=1):
        ws.cell(row=4, column=i, value=h)
    # Riga 5 — OFI n.1, fase DO (2 X), normative 27001+9100
    ws.cell(row=5, column=1, value=1)
    ws.cell(row=5, column=2, value=date(2026, 1, 15))
    ws.cell(row=5, column=3, value="OFI")
    ws.cell(row=5, column=4, value="27001, 9100")
    ws.cell(row=5, column=5, value="8.5.1")
    ws.cell(row=5, column=6, value="Trattamenti")
    ws.cell(row=5, column=7, value="Migliorare la resa")
    ws.cell(row=5, column=8, value="Piano X")
    ws.cell(row=5, column=10, value="Attuazione X")
    ws.cell(row=5, column=15, value="Rossi")
    ws.cell(row=5, column=16, value="X")  # P
    ws.cell(row=5, column=17, value="X")  # D
    # Riga 6 — NC n.2, chiusa (DATA CLOSED valorizzata) + 4 X
    ws.cell(row=6, column=1, value=2)
    ws.cell(row=6, column=2, value=date(2026, 2, 1))
    ws.cell(row=6, column=3, value="NC")
    ws.cell(row=6, column=6, value="Saldatura")
    ws.cell(row=6, column=7, value="Rivedere WPS")
    ws.cell(row=6, column=14, value=date(2026, 3, 1))  # DATA CLOSED
    for col in (16, 17, 18, 19):
        ws.cell(row=6, column=col, value="X")
    # Riga 7 — pre-numerata ma VUOTA (nessun contenuto) → da saltare
    ws.cell(row=7, column=1, value=3)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


class ParseMod174Tests(TestCase):
    def test_parsing_righe(self):
        righe = leggi_righe_mod174(_fake_mod174())
        self.assertEqual(len(righe), 2)  # la riga 7 (vuota) è saltata
        r1, r2 = righe
        self.assertEqual(r1["numero"], 1)
        self.assertEqual(r1["tipo"], "OFI")
        self.assertTrue(r1["norma_iso27001"])
        self.assertTrue(r1["norma_en9100"])
        self.assertFalse(r1["norma_iso45001"])
        self.assertEqual(r1["processo"], "Trattamenti")
        self.assertEqual(r1["data_apertura"], date(2026, 1, 15))
        self.assertEqual(r1["fase"], RegistroOFI.FASE_DO)  # 2 X → DO
        self.assertEqual(r2["numero"], 2)
        self.assertEqual(r2["tipo"], "NC")
        self.assertEqual(r2["data_chiusura"], date(2026, 3, 1))
        self.assertEqual(r2["fase"], RegistroOFI.FASE_CHIUSO)  # chiusa

    def test_file_non_mod174_solleva(self):
        wb = openpyxl.Workbook()
        wb.active["A1"] = "qualcosa d'altro"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        with self.assertRaises(ValueError):
            leggi_righe_mod174(bio)


class ImportaVociTests(TestCase):
    def test_dry_run_non_scrive(self):
        res = importa_voci(leggi_righe_mod174(_fake_mod174()), dry_run=True)
        self.assertEqual(res["creati"], 2)
        self.assertEqual(RegistroOFI.objects.count(), 0)

    def test_apply_crea_voci(self):
        res = importa_voci(leggi_righe_mod174(_fake_mod174()),
                           dry_run=False, modulo_origine="sgi_27001")
        self.assertEqual(res["creati"], 2)
        self.assertEqual(RegistroOFI.objects.count(), 2)
        v = RegistroOFI.objects.get(numero=1)
        self.assertEqual(v.processo, "Trattamenti")
        self.assertEqual(v.fase, RegistroOFI.FASE_DO)
        self.assertEqual(v.modulo_origine, "sgi_27001")
        self.assertTrue(v.pdca_marks["d"])

    def test_idempotente(self):
        importa_voci(leggi_righe_mod174(_fake_mod174()), dry_run=False)
        res2 = importa_voci(leggi_righe_mod174(_fake_mod174()), dry_run=False)
        self.assertEqual(res2["aggiornati"], 2)
        self.assertEqual(res2["creati"], 0)
        self.assertEqual(RegistroOFI.objects.count(), 2)


class CommandTests(TestCase):
    def _scrivi_file(self, tmp_path):
        with open(tmp_path, "wb") as f:
            f.write(_fake_mod174().read())

    def test_command_apply(self):
        import tempfile
        import os
        d = tempfile.mkdtemp()
        p = os.path.join(d, "mod174.xlsx")
        self._scrivi_file(p)
        call_command("import_mod174", p, "--apply", "--modulo", "sgi_27001")
        self.assertEqual(RegistroOFI.objects.count(), 2)

    def test_command_dry_run_default(self):
        import tempfile
        import os
        d = tempfile.mkdtemp()
        p = os.path.join(d, "mod174.xlsx")
        self._scrivi_file(p)
        call_command("import_mod174", p)  # niente --apply
        self.assertEqual(RegistroOFI.objects.count(), 0)
