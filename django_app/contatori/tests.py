import datetime as dt
from unittest import mock
from django.test import TestCase
from django.urls import reverse
from django.core.management import call_command
from contatori import services
from contatori.models import Macchina, LetturaContatori, ImpostazioniSNMP
from contatori.snmp import SNMPError


class _AuthedClientMixin:
    """Nell'HUB ogni vista è dietro ACLMiddleware: autentichiamo un superuser
    (che bypassa l'ACL) prima di ogni richiesta del client."""

    def setUp(self):
        super().setUp()
        from django.contrib.auth import get_user_model
        U = get_user_model()
        u, _ = U.objects.get_or_create(
            username="contatori_test_admin",
            defaults={"is_staff": True, "is_superuser": True},
        )
        u.is_staff = True
        u.is_superuser = True
        u.save()
        self.client.force_login(u)


class RiconciliazioneTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_nessuna_anomalia_su_tutti_i_trimestri(self):
        for trim in ["2025-Q4", "2026-Q1", "2026-Q2"]:
            righe, riep = services.riconcilia(trim)
            self.assertEqual(riep["contatori"], 16, trim)
            self.assertEqual(riep["anomalie"], 0, f"{trim}: attese 0 anomalie")
            self.assertTrue(riep["ok"], trim)

    def test_pool_2020_124_somma_cqf_cqi(self):
        # A4 BN pool Q2: 155251 (CQF) + 56037 (CQI) = 211288, fornitore 210798, scarto +490
        righe, _ = services.riconcilia("2026-Q2")
        r = next(x for x in righe if x["contratto"] == "2020/124" and x["contatore"] == "A4 BN")
        self.assertTrue(r["pool"])
        self.assertEqual(r["ns"], 211288)
        self.assertEqual(r["scarto"], 490)

    def test_monotonia_rileva_refuso_logistica_a3col(self):
        problemi = services.controllo_monotonia()
        self.assertEqual(len(problemi), 1)
        p = problemi[0]
        self.assertEqual(p["matricola"], "2YM19974")
        self.assertEqual(p["contatore"], "A3 COL")
        self.assertEqual(p["da_val"], 539)
        self.assertEqual(p["a_val"], 366)

    def test_scarto_negativo_genera_anomalia(self):
        # abbasso una lettura interna sotto la fattura -> deve scattare l'allarme
        log = Macchina.objects.get(matricola="2YM19974")
        l = LetturaContatori.objects.get(macchina=log, trimestre="2026-Q2")
        l.a4_bn = 100000  # < fornitore 187917
        l.save()
        _, riep = services.riconcilia("2026-Q2")
        self.assertGreaterEqual(riep["anomalie"], 1)
        self.assertFalse(riep["ok"])


class UltimeRilevazioniTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_ultima_aggiornata(self):
        # ultima lettura Q2 è del 2026-07-06; a 4 giorni è "aggiornata"
        ur = services.ultime_rilevazioni(oggi=dt.date(2026, 7, 10))
        log = Macchina.objects.get(matricola="2YM19974")
        r = ur[log.id]
        self.assertEqual(r["lettura"].trimestre, "2026-Q2")
        self.assertEqual(r["giorni_fa"], 4)
        self.assertEqual(r["stato"], "aggiornata")

    def test_ultima_da_aggiornare(self):
        ur = services.ultime_rilevazioni(oggi=dt.date(2027, 1, 1))
        log = Macchina.objects.get(matricola="2YM19974")
        self.assertEqual(ur[log.id]["stato"], "da_aggiornare")

    def test_mai_letta(self):
        m = Macchina.objects.create(reparto="NUOVO", matricola="NUOVA1")
        ur = services.ultime_rilevazioni()
        self.assertEqual(ur[m.id]["stato"], "mai")


class AnalisiTest(_AuthedClientMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_andamento_totale_q2(self):
        a = services.andamento_trimestri()
        self.assertEqual([x["trimestre"] for x in a], ["2025-Q4", "2026-Q1", "2026-Q2"])
        self.assertEqual(a[-1]["totale"], 726853)
        self.assertEqual(a[-1]["bn"], 559338)
        self.assertEqual(a[-1]["col"], 167515)

    def test_consumo_delta_q2_e_anomalia(self):
        consumo, anomalie = services.consumo_per_trimestre()
        self.assertEqual([c["trimestre"] for c in consumo], ["2026-Q1", "2026-Q2"])
        q2 = consumo[-1]
        self.assertEqual(q2["totale"], 44101)
        self.assertEqual(q2["bn"], 29682)
        self.assertEqual(q2["col"], 14419)
        # il calo a3_col di LOGISTICA (539->366) è l'unica anomalia
        self.assertEqual(anomalie, 1)

    def test_classifica_ordinata(self):
        c = services.classifica_reparti()
        self.assertEqual(len(c), 5)
        self.assertEqual(c[0]["reparto"], "LOGISTICA")
        self.assertEqual(c[0]["totale"], 32251)
        totali = [r["totale"] for r in c]
        self.assertEqual(totali, sorted(totali, reverse=True))

    def test_ripartizione_percentuali(self):
        r = services.ripartizione()
        self.assertEqual(r["trimestre"], "2026-Q2")
        self.assertEqual(r["totale"], 726853)
        self.assertEqual(r["bn_pct"], 77)
        self.assertEqual(r["col_pct"], 23)
        self.assertEqual(r["a4_pct"], 97)
        self.assertEqual(r["a3_pct"], 3)

    def test_vista_analisi_ok(self):
        r = self.client.get(reverse("contatori:analisi"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Classifica reparti")
        self.assertContains(r, "LOGISTICA")


class ConsumabiliTest(_AuthedClientMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_parsing_livelli(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.212"; m.save()
        raw = [("1.1", "Black Toner", 47, 100),
               ("1.5", "Waste Toner", -3, 100),   # residuo non quantificato
               ("1.6", "Drum", -2, 100)]          # sconosciuto
        with mock.patch("contatori.snmp._consumabili_raw", return_value=raw):
            from contatori.snmp import leggi_consumabili
            out = leggi_consumabili(m)
        self.assertEqual(out[0], {"nome": "Black Toner", "pct": 47, "nota": ""})
        self.assertEqual(out[1]["pct"], None)
        self.assertEqual(out[1]["nota"], "presente")
        self.assertEqual(out[2]["pct"], None)

    def test_vista_consumabili_ok(self):
        m = Macchina.objects.first()
        dati = [{"nome": "Black Toner", "pct": 47, "nota": ""},
                {"nome": "Fuser Unit", "pct": 10, "nota": ""}]
        with mock.patch("contatori.snmp.leggi_consumabili", return_value=dati):
            r = self.client.post(reverse("contatori:macchina_consumabili", args=[m.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Black Toner")
        self.assertContains(r, "47%")

    def test_vista_consumabili_errore(self):
        m = Macchina.objects.first()
        with mock.patch("contatori.snmp.leggi_consumabili",
                        side_effect=SNMPError("host irraggiungibile")):
            r = self.client.post(reverse("contatori:macchina_consumabili", args=[m.pk]))
        self.assertContains(r, "host irraggiungibile")

    def test_consumabili_solo_post(self):
        m = Macchina.objects.first()
        r = self.client.get(reverse("contatori:macchina_consumabili", args=[m.pk]))
        self.assertEqual(r.status_code, 405)

    def test_flotta_solo_macchine_con_host(self):
        m = Macchina.objects.get(matricola="2YM19974")
        m.host = "10.0.0.212"; m.save()
        r = self.client.get(reverse("contatori:consumabili"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "10.0.0.212")
        self.assertContains(r, "LOGISTICA")
        # le macchine senza host non compaiono
        self.assertNotContains(r, "PRESETTING")

    def test_riepilogo_critici(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.99"; m.save()
        dati = [{"nome": "Black Toner", "pct": 8, "nota": ""},
                {"nome": "Cyan Toner", "pct": 60, "nota": ""}]
        with mock.patch("contatori.snmp.leggi_consumabili", return_value=dati):
            r = self.client.get(reverse("contatori:consumabili_riepilogo", args=[m.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "critico")
        self.assertContains(r, "Black Toner")


class ExportExcelTest(_AuthedClientMixin, TestCase):
    XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_export_riconciliazione(self):
        r = self.client.get(reverse("contatori:export_riconciliazione", args=["2026-Q2"]))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], self.XLSX)
        self.assertIn("riconciliazione_2026-Q2.xlsx", r["Content-Disposition"])
        self.assertTrue(len(r.content) > 500)

    def test_export_analisi_fogli(self):
        import io
        from openpyxl import load_workbook
        r = self.client.get(reverse("contatori:export_analisi"))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], self.XLSX)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.sheetnames,
                         ["Andamento", "Consumo", "Classifica reparti", "Ripartizione"])
        # verifica un valore noto: totale andamento Q2
        ws = wb["Andamento"]
        ultima = list(ws.iter_rows(values_only=True))[-1]
        self.assertEqual(ultima[0], "2026-Q2")
        self.assertEqual(ultima[-1], 726853)


class ViewsTest(_AuthedClientMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_dashboard_ok(self):
        r = self.client.get(reverse("contatori:dashboard"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Dashboard contatori")

    def test_riconciliazione_ok(self):
        r = self.client.get(reverse("contatori:riconciliazione"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Fattura corretta")

    def test_macchina_detail_ok(self):
        m = Macchina.objects.first()
        r = self.client.get(reverse("contatori:macchina", args=[m.pk]))
        self.assertEqual(r.status_code, 200)


class GestioneStampantiTest(_AuthedClientMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_lista_ok(self):
        r = self.client.get(reverse("contatori:macchine"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Stampanti")
        self.assertContains(r, "AMMINISTRAZIONE")

    def test_crea_macchina(self):
        prima = Macchina.objects.count()
        r = self.client.post(reverse("contatori:macchina_nuova"), {
            "reparto": "UFFICIO TEST", "matricola": "TEST0001",
            "modello": "iR-ADV C5535i", "contratto": "9999/1",
            "fornitore": "BASE", "host": "192.168.1.50", "attiva": "on",
        })
        self.assertRedirects(r, reverse("contatori:macchine"))
        self.assertEqual(Macchina.objects.count(), prima + 1)
        self.assertTrue(Macchina.objects.filter(matricola="TEST0001").exists())

    def test_modifica_ip(self):
        m = Macchina.objects.get(matricola="2YM19959")
        r = self.client.post(reverse("contatori:macchina_edit", args=[m.pk]), {
            "reparto": m.reparto, "matricola": m.matricola, "modello": m.modello,
            "contratto": m.contratto, "fornitore": m.fornitore,
            "host": "10.0.0.7", "attiva": "on",
        })
        self.assertRedirects(r, reverse("contatori:macchine"))
        m.refresh_from_db()
        self.assertEqual(m.host, "10.0.0.7")

    def test_test_snmp_successo(self):
        m = Macchina.objects.first()
        valori = {"a4_bn": 111, "a3_bn": 22, "a4_col": 33, "a3_col": 4}
        with mock.patch("contatori.snmp.leggi_macchina", return_value=valori):
            r = self.client.post(reverse("contatori:macchina_test_snmp", args=[m.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "risponde")
        self.assertContains(r, "111")
        # banner di conferma out-of-band
        self.assertContains(r, "Lettura SNMP riuscita")
        self.assertContains(r, 'hx-swap-oob="true"')

    def test_test_snmp_errore(self):
        m = Macchina.objects.first()
        with mock.patch("contatori.snmp.leggi_macchina",
                        side_effect=SNMPError("host non impostato")):
            r = self.client.post(reverse("contatori:macchina_test_snmp", args=[m.pk]))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "host non impostato")

    def test_test_snmp_solo_post(self):
        m = Macchina.objects.first()
        r = self.client.get(reverse("contatori:macchina_test_snmp", args=[m.pk]))
        self.assertEqual(r.status_code, 405)

    def test_impostazioni_default(self):
        cfg = ImpostazioniSNMP.get_solo()
        self.assertEqual(cfg.community, "novicromprinter")
        self.assertEqual(cfg.version, "v1")
        self.assertEqual(cfg.port, 161)

    def test_salva_impostazioni_snmp(self):
        r = self.client.post(reverse("contatori:macchine"), {
            "community": "altrogruppo", "port": "1610",
            "timeout": "5", "version": "v2c",
        })
        self.assertRedirects(r, reverse("contatori:macchine"))
        cfg = ImpostazioniSNMP.get_solo()
        self.assertEqual(cfg.community, "altrogruppo")
        self.assertEqual(cfg.version, "v2c")
        self.assertEqual(cfg.port, 1610)

    def test_test_snmp_usa_config_salvata(self):
        ImpostazioniSNMP.objects.update_or_create(
            pk=1, defaults={"community": "xyz", "port": 161, "timeout": 2, "version": "v2c"})
        m = Macchina.objects.first()
        with mock.patch("contatori.snmp.leggi_macchina", return_value={
                "a4_bn": 1, "a3_bn": 2, "a4_col": 3, "a3_col": 4}) as fake:
            self.client.post(reverse("contatori:macchina_test_snmp", args=[m.pk]))
        _, kwargs = fake.call_args
        self.assertEqual(kwargs["community"], "xyz")
        self.assertEqual(kwargs["version"], "v2c")
        self.assertEqual(kwargs["timeout"], 2)


class CollegamentoAssetTest(_AuthedClientMixin, TestCase):
    """Sotto-progetto D1: collegamento Macchina↔Asset dell'HUB."""

    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def test_fk_asset_nullable_default(self):
        m = Macchina.objects.first()
        self.assertIsNone(m.asset)

    def test_collega_asset_command_dryrun_non_fallisce(self):
        from io import StringIO
        out = StringIO()
        call_command("collega_asset", stdout=out)
        self.assertIn("match trovati", out.getvalue())

    def test_collega_asset_per_matricola(self):
        from assets.models import Asset
        m = Macchina.objects.first()
        Asset.objects.create(asset_tag="SOC-TEST-1", name="MFC Test", serial_number=m.matricola)
        from io import StringIO
        out = StringIO()
        call_command("collega_asset", "--apply", stdout=out)
        m.refresh_from_db()
        self.assertIsNotNone(m.asset)
        self.assertEqual(m.asset.serial_number, m.matricola)

    def test_scheda_macchina_mostra_asset_collegato(self):
        from assets.models import Asset
        m = Macchina.objects.first()
        m.asset = Asset.objects.create(asset_tag="SOC-TEST-2", name="MFC Collegata", serial_number=m.matricola)
        m.save(update_fields=["asset"])
        r = self.client.get(reverse("contatori:macchina", args=[m.pk]))
        self.assertContains(r, "Asset collegato")
        self.assertContains(r, "MFC Collegata")


class ModelloValidatoEAssetNelFormTest(_AuthedClientMixin, TestCase):
    """Fix bug «nessuna counter_map per modello 'CANON'» + aggancio asset dal form.

    Causa radice: `modello` era testo libero, quindi dalla UI si poteva scrivere la
    marca ('CANON') e la lookup esatta su COUNTER_MAP falliva.
    """

    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    def _payload(self, m, **over):
        data = {
            "reparto": m.reparto,
            "matricola": m.matricola,
            "modello": m.modello,
            "contratto": m.contratto,
            "fornitore": m.fornitore,
            "host": "",
            "attiva": "on",
        }
        data.update(over)
        return data

    def test_modello_marca_generica_rifiutato(self):
        from contatori.forms import MacchinaForm
        m = Macchina.objects.first()
        form = MacchinaForm(self._payload(m, modello="CANON"), instance=m)
        self.assertFalse(form.is_valid())
        self.assertIn("modello", form.errors)

    def test_modello_mappato_accettato(self):
        from contatori.forms import MacchinaForm
        m = Macchina.objects.first()
        form = MacchinaForm(self._payload(m, modello="iR-ADV DX C5840i"), instance=m)
        self.assertTrue(form.is_valid(), form.errors)

    def test_ogni_modello_selezionabile_ha_una_counter_map(self):
        """Guardia: le choices e COUNTER_MAP non devono divergere."""
        from contatori.snmp import COUNTER_MAP
        for valore, _label in Macchina.Modello.choices:
            self.assertIn(valore, COUNTER_MAP, f"modello selezionabile senza counter_map: {valore}")

    def test_form_espone_asset_e_permette_aggancio(self):
        from assets.models import Asset
        from contatori.forms import MacchinaForm
        self.assertIn("asset", MacchinaForm().fields)
        m = Macchina.objects.first()
        a = Asset.objects.create(asset_tag="SOC-FORM-1", name="MFC dal form", serial_number=m.matricola)
        form = MacchinaForm(self._payload(m, asset=a.pk), instance=m)
        self.assertTrue(form.is_valid(), form.errors)
        form.save()
        m.refresh_from_db()
        self.assertEqual(m.asset_id, a.pk)

    def test_errore_modello_non_mappato_e_esplicativo(self):
        from contatori.snmp import leggi_macchina
        m = Macchina.objects.first()
        m.modello = "CANON"
        m.host = "10.0.0.1"
        with self.assertRaises(SNMPError) as ctx:
            leggi_macchina(m)
        msg = str(ctx.exception)
        self.assertIn("CANON", msg)
        self.assertIn("iR-ADV", msg)  # deve elencare i modelli mappati


class DiscoverySNMPTest(_AuthedClientMixin, TestCase):
    """Discovery di rete dal portale: trova le stampanti e scopre gli IP sbagliati.

    L'abbinamento e' sulla MATRICOLA (seriale letto dal dispositivo), non sull'IP:
    e' proprio l'IP quello che puo' essere sbagliato.
    """

    @classmethod
    def setUpTestData(cls):
        call_command("seed_demo")

    # -- validazioni che non toccano la rete -------------------------------
    def test_rete_non_valida_solleva_errore(self):
        from contatori.snmp import scansiona_rete
        with self.assertRaises(SNMPError):
            scansiona_rete("non-una-rete")

    def test_range_troppo_ampio_rifiutato(self):
        from contatori.snmp import scansiona_rete
        with self.assertRaises(SNMPError) as ctx:
            scansiona_rete("10.0.0.0/16")
        self.assertIn("troppo ampio", str(ctx.exception))

    # -- logica di abbinamento (pura) --------------------------------------
    def test_abbina_rileva_ip_diverso(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.155"
        m.save(update_fields=["host"])
        righe = services.abbina_discovery(
            [{"host": "10.0.0.77", "descr": "Canon iR-ADV", "nome": "mfc", "matricola": m.matricola}]
        )
        self.assertEqual(righe[0]["stato"], "ip_diverso")
        self.assertEqual(righe[0]["macchina"], m)

    def test_abbina_ip_corretto_e_ok(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.77"
        m.save(update_fields=["host"])
        righe = services.abbina_discovery(
            [{"host": "10.0.0.77", "descr": "Canon", "nome": "", "matricola": m.matricola}]
        )
        self.assertEqual(righe[0]["stato"], "ok")

    def test_abbina_dispositivo_sconosciuto_e_nuovo(self):
        righe = services.abbina_discovery(
            [{"host": "10.0.0.99", "descr": "Canon", "nome": "", "matricola": "SERIALE-IGNOTO"}]
        )
        self.assertEqual(righe[0]["stato"], "nuova")
        self.assertIsNone(righe[0]["macchina"])

    # -- pagina ------------------------------------------------------------
    def test_pagina_discovery_si_apre(self):
        r = self.client.get(reverse("contatori:discovery"))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Discovery SNMP")

    def test_scansione_mostra_ip_diverso_in_pagina(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.155"
        m.save(update_fields=["host"])
        finti = [{"host": "10.0.0.77", "descr": "Canon iR-ADV DX C5840i",
                  "nome": "mfc-amm", "matricola": m.matricola}]
        with mock.patch("contatori.snmp.scansiona_rete", return_value=finti):
            r = self.client.post(reverse("contatori:discovery"),
                                 {"rete": "10.0.0.0/24", "community": "x",
                                  "version": "v1", "timeout": "2"})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "IP diverso")
        self.assertContains(r, "10.0.0.77")

    def test_applica_ip_corregge_la_macchina(self):
        m = Macchina.objects.first()
        m.host = "10.0.0.155"
        m.save(update_fields=["host"])
        r = self.client.post(reverse("contatori:discovery_applica_ip", args=[m.pk]),
                             {"host": "10.0.0.77"})
        self.assertEqual(r.status_code, 302)
        m.refresh_from_db()
        self.assertEqual(str(m.host), "10.0.0.77")


class ExportXlsxInjectionTest(_AuthedClientMixin, TestCase):
    """Formula injection: descrizioni/reparti dei contatori sono testo libero e
    finiscono nell'.xlsx scaricabile. Devono restare TESTO, mai formule vive.
    Dati sintetici (services mockati: nessun dato reale).
    """

    RIGA_EVIL = {
        "contratto": '=HYPERLINK("http://evil.test","clicca")',
        "descrizione": "@SUM(A1:A9)",
        "pool": False,
        "contatore": "A4 BN",
        "fornitore": 100,
        "ns": 100,
        "scarto": 0,
        "esito": "OK",
        "livello": "ok",
    }
    RIEP = {"contatori": 1, "anomalie": 0, "ok": True}

    def _assert_no_formula(self, content):
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content))
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    self.assertNotEqual(
                        cell.data_type, "f",
                        f"{ws.title}!{cell.coordinate} = {cell.value!r} scritta come formula",
                    )
        return wb

    def test_riconciliazione_non_produce_formule(self):
        with mock.patch.object(services, "riconcilia",
                               return_value=([self.RIGA_EVIL], self.RIEP)):
            r = self.client.get(reverse("contatori:export_riconciliazione",
                                        args=["2026-Q2"]))
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        wb = self._assert_no_formula(r.content)
        ws = wb.worksheets[0]
        self.assertEqual(ws.cell(row=2, column=1).value, self.RIGA_EVIL["contratto"])
        self.assertTrue(ws.cell(row=2, column=1).quotePrefix)
        self.assertTrue(ws.cell(row=2, column=2).quotePrefix)
        # non-regressione: i numeri restano numeri
        self.assertEqual(ws.cell(row=2, column=6).value, 100)
        self.assertEqual(ws.cell(row=2, column=6).data_type, "n")

    def test_analisi_non_produce_formule(self):
        evil = '=cmd|\' /c calc\'!A1'
        with mock.patch.object(services, "andamento_trimestri", return_value=[]), \
             mock.patch.object(services, "consumo_per_trimestre", return_value=([], 0)), \
             mock.patch.object(services, "classifica_reparti", return_value=[
                 {"reparto": evil, "matricola": "SN-TEST", "bn": 10, "col": 5,
                  "totale": 15, "quota_col": 33.3}]), \
             mock.patch.object(services, "ripartizione", return_value={
                 "trimestre": "2026-Q2", "totale": 15, "bn_pct": 66.7,
                 "col_pct": 33.3, "a4_pct": 90.0, "a3_pct": 10.0}):
            r = self.client.get(reverse("contatori:export_analisi"))
        self.assertEqual(r.status_code, 200)
        wb = self._assert_no_formula(r.content)
        ws = wb["Classifica reparti"]
        self.assertEqual(ws.cell(row=2, column=2).value, evil)
        self.assertTrue(ws.cell(row=2, column=2).quotePrefix)
        self.assertEqual(ws.cell(row=2, column=4).value, 10)  # numeri intatti
