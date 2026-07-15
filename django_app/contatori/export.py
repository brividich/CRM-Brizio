"""Export Excel (.xlsx) di riconciliazione e analisi. Usa openpyxl.

Le celle passano da `core.excel_export.append_row` (sede unica della
sanificazione anti formula injection): `ws.append` scriverebbe come formula
viva qualunque stringa che inizia con "=" (descrizioni/reparti dal DB).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.excel_export import append_row

from . import services

_HEADER = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="12395F")
_DANGER = Font(color="C0392B", bold=True)


def _intesta(ws, colonne):
    append_row(ws, colonne)
    for cell in ws[1]:
        cell.font = _HEADER
        cell.fill = _HEADER_FILL


def _autolarghezza(ws):
    for col in ws.columns:
        larg = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(larg + 3, 48)


def riconciliazione_xlsx(trimestre):
    """Workbook con la riconciliazione di un trimestre."""
    righe, riep = services.riconcilia(trimestre)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Riconciliazione {trimestre}"[:31]
    _intesta(ws, ["Contratto", "Unità", "Pool", "Contatore",
                  "Fornitore", "Interno", "Scarto", "Esito"])
    for r in righe:
        append_row(ws, [r["contratto"], r["descrizione"], "sì" if r["pool"] else "",
                        r["contatore"], r["fornitore"], r["ns"], r["scarto"], r["esito"]])
        if r["livello"] == "danger":
            for cell in ws[ws.max_row]:
                cell.font = _DANGER
    append_row(ws)
    append_row(ws, ["Trimestre", trimestre, "", "Contatori", riep["contatori"],
                    "Anomalie", riep["anomalie"], "OK" if riep["ok"] else "DA CONTROLLARE"])
    _autolarghezza(ws)
    return wb


def analisi_xlsx():
    """Workbook multi-foglio con andamento, consumo, classifica, ripartizione."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Andamento"
    _intesta(ws, ["Trimestre", "A4 BN", "A3 BN", "A4 COL", "A3 COL",
                  "BN", "Colore", "Totale"])
    for a in services.andamento_trimestri():
        append_row(ws, [a["trimestre"], a["a4_bn"], a["a3_bn"], a["a4_col"], a["a3_col"],
                        a["bn"], a["col"], a["totale"]])
    _autolarghezza(ws)

    consumo, anomalie = services.consumo_per_trimestre()
    ws = wb.create_sheet("Consumo")
    _intesta(ws, ["Trimestre", "BN", "Colore", "Totale prodotto"])
    for c in consumo:
        append_row(ws, [c["trimestre"], c["bn"], c["col"], c["totale"]])
    append_row(ws)
    append_row(ws, ["Cali ignorati (refusi monotonìa)", anomalie])
    _autolarghezza(ws)

    ws = wb.create_sheet("Classifica reparti")
    _intesta(ws, ["#", "Reparto", "Matricola", "BN", "Colore", "Totale", "Quota colore %"])
    for i, r in enumerate(services.classifica_reparti(), 1):
        append_row(ws, [i, r["reparto"], r["matricola"], r["bn"], r["col"],
                        r["totale"], r["quota_col"]])
    _autolarghezza(ws)

    rip = services.ripartizione()
    ws = wb.create_sheet("Ripartizione")
    _intesta(ws, ["Metrica", "Valore"])
    append_row(ws, ["Trimestre", rip.get("trimestre", "")])
    append_row(ws, ["Totale copie", rip.get("totale", 0)])
    append_row(ws, ["% Bianco/Nero", rip.get("bn_pct", 0)])
    append_row(ws, ["% Colore", rip.get("col_pct", 0)])
    append_row(ws, ["% A4", rip.get("a4_pct", 0)])
    append_row(ws, ["% A3", rip.get("a3_pct", 0)])
    _autolarghezza(ws)

    return wb
