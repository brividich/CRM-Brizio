from __future__ import annotations

import hashlib
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from assets.models import Asset, WorkMachine


TRUE_VALUES = {"1", "S", "SI", "YES", "Y", "TRUE", "ON", "OK", "PRESENTE", "X", "CHECK", "CHECKED", "V"}
FALSE_VALUES = {"0", "N", "NO", "FALSE", "OFF", "NONE", "NULL", "N/A", "NA", "ND", "N D", "-", ""}

HEADER_MAP = {
    "REPARTO": "reparto",
    "NAME": "name",
    "X MM": "x_mm",
    "Y MM": "y_mm",
    "Z MM": "z_mm",
    "DIAMETER MM": "diameter_mm",
    "SPINDLE MM": "spindle_mm",
    "YEAR": "year",
    "TMC": "tmc",
    "TCR": "tcr_enabled",
    "PRESSURE BAR": "pressure_bar",
    "CNC": "cnc_controlled",
    "5 AXES": "five_axes",
    "ACCURACY FROM": "accuracy_from",
}

KNOWN_MANUFACTURERS = [
    ("DMG MORI", "DMG Mori"),
    ("DMG", "DMG"),
    ("HERMLE", "HERMLE"),
    ("MIKRON", "MIKRON"),
    ("ZEISS", "ZEISS"),
    ("HEXAGON", "HEXAGON"),
    ("MAZAK", "MAZAK"),
    ("OKUMA", "OKUMA"),
    ("FANUC", "FANUC"),
]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    row = str(value).strip().upper()
    if not row:
        return ""
    row = row.replace("Ø", " DIAMETER ")
    row = row.replace("O/", " DIAMETER ")
    row = row.replace("DIAM.", " DIAMETER ")
    row = unicodedata.normalize("NFKD", row)
    row = "".join(ch for ch in row if not unicodedata.combining(ch))
    row = re.sub(r"[^A-Z0-9]+", " ", row)
    return re.sub(r"\s+", " ", row).strip()


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("%f" % value).rstrip("0").rstrip(".")
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    return _clean_str(value) == ""


def _nullish(value: Any) -> bool:
    token = _clean_str(value).strip().upper()
    return token in FALSE_VALUES


def _parse_int(value: Any) -> int | None:
    if _is_blank(value) or _nullish(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    token = _clean_str(value).replace(",", ".")
    try:
        if "." in token:
            parsed = float(token)
            return int(parsed) if parsed.is_integer() else None
        return int(token)
    except ValueError:
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    if _is_blank(value) or _nullish(value):
        return None
    token = _clean_str(value).replace(",", ".")
    try:
        return Decimal(token)
    except (InvalidOperation, ValueError):
        return None


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    token = _clean_str(value).strip()
    if token in {"✓", "✔", "X", "x"}:
        return True
    normalized = _normalize_header(token)
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    return bool(token)


def _split_machine_name(name: str) -> tuple[str, str]:
    normalized = name.strip()
    upper_name = normalized.upper()
    for manufacturer_match, manufacturer_label in KNOWN_MANUFACTURERS:
        if upper_name.startswith(manufacturer_match + " "):
            model = normalized[len(manufacturer_match) :].strip()
            return manufacturer_label, model or normalized
    parts = normalized.split(" ", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", normalized


def _build_source_key(sheet_name: str, reparto: str, name: str, year: int | None, x_mm: int | None, y_mm: int | None, z_mm: int | None) -> str:
    parts = [sheet_name.strip().upper(), reparto.strip().upper(), name.strip().upper()]
    if year is not None:
        parts.append(f"YEAR={year}")
    else:
        parts.extend([f"X={x_mm or ''}", f"Y={y_mm or ''}", f"Z={z_mm or ''}"])
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()


class Command(BaseCommand):
    help = "Importa macchine di lavoro da Excel in Asset + WorkMachine."

    def add_arguments(self, parser):
        parser.add_argument("--file", default="Macchine di lavoro.xlsx", help="Path del file Excel.")
        parser.add_argument("--sheet", default="", help="Nome foglio da importare. Se omesso usa il primo.")
        parser.add_argument("--dry-run", action="store_true", help="Simula import senza scrivere su DB.")
        parser.add_argument(
            "--update",
            dest="update",
            action="store_true",
            default=True,
            help="Aggiorna record esistenti (default).",
        )
        parser.add_argument(
            "--no-update",
            dest="update",
            action="store_false",
            help="Non aggiornare record esistenti.",
        )

    def handle(self, *args, **options):
        file_path = Path(str(options.get("file") or "")).expanduser()
        if not file_path.exists():
            raise CommandError(f"File non trovato: {file_path}")

        dry_run = bool(options.get("dry_run"))
        update_existing = bool(options.get("update", True))

        try:
            workbook = load_workbook(filename=str(file_path), data_only=True)
        except Exception as exc:
            raise CommandError(f"Impossibile aprire workbook: {exc}") from exc

        requested_sheet = str(options.get("sheet") or "").strip()
        if requested_sheet:
            if requested_sheet not in workbook.sheetnames:
                raise CommandError(f"Foglio non trovato: {requested_sheet}")
            ws = workbook[requested_sheet]
        else:
            ws = workbook[workbook.sheetnames[0]]

        headers_by_col: dict[int, str] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            raw_header = ws.cell(row=1, column=col_idx).value
            normalized = _normalize_header(raw_header)
            if normalized in HEADER_MAP:
                headers_by_col[col_idx] = HEADER_MAP[normalized]

        if "name" not in headers_by_col.values():
            raise CommandError("Colonna 'Name' non trovata nel foglio.")

        stats = {
            "rows_seen": 0,
            "rows_skipped": 0,
            "assets_created": 0,
            "assets_updated": 0,
            "machines_created": 0,
            "machines_updated": 0,
        }

        self.stdout.write(f"Workbook: {file_path}")
        self.stdout.write(f"Foglio selezionato: {ws.title}")
        self.stdout.write(f"Modalita: {'DRY-RUN' if dry_run else 'IMPORT REALE'} | update={update_existing}")

        for row_idx in range(2, (ws.max_row or 0) + 1):
            stats["rows_seen"] += 1
            row_data: dict[str, Any] = {}
            for col_idx, field_name in headers_by_col.items():
                row_data[field_name] = ws.cell(row=row_idx, column=col_idx).value

            if all(_is_blank(value) for value in row_data.values()):
                continue

            name = _clean_str(row_data.get("name"))
            reparto = _clean_str(row_data.get("reparto"))
            if not name:
                stats["rows_skipped"] += 1
                continue

            x_mm = _parse_int(row_data.get("x_mm"))
            y_mm = _parse_int(row_data.get("y_mm"))
            z_mm = _parse_int(row_data.get("z_mm"))
            year = _parse_int(row_data.get("year"))
            source_key = _build_source_key(ws.title, reparto, name, year, x_mm, y_mm, z_mm)
            manufacturer, model = _split_machine_name(name)

            asset = Asset.objects.filter(source_key=source_key).first()
            asset_created = False
            if asset is None:
                asset = Asset(
                    source_key=source_key,
                    name=name[:255],
                    asset_type=Asset.TYPE_WORK_MACHINE,
                    reparto=reparto[:120],
                    manufacturer=manufacturer[:120] or None,
                    model=model[:120] or None,
                    status=Asset.STATUS_IN_USE,
                )
                if not dry_run:
                    asset.save()
                stats["assets_created"] += 1
                asset_created = True
            elif update_existing:
                changed_fields: list[str] = []

                def update_asset_field(field_name: str, next_value: Any) -> None:
                    if next_value is None:
                        return
                    if getattr(asset, field_name) != next_value:
                        setattr(asset, field_name, next_value)
                        changed_fields.append(field_name)

                update_asset_field("name", name[:255])
                update_asset_field("asset_type", Asset.TYPE_WORK_MACHINE)
                update_asset_field("reparto", reparto[:120])
                if manufacturer:
                    update_asset_field("manufacturer", manufacturer[:120])
                if model:
                    update_asset_field("model", model[:120])

                if changed_fields:
                    if not dry_run:
                        asset.save(update_fields=sorted(set(changed_fields + ["updated_at"])))
                    stats["assets_updated"] += 1
            else:
                continue

            machine = WorkMachine.objects.filter(asset=asset).first() if asset.pk else None
            machine_payload = {
                "source_key": source_key,
                "x_mm": x_mm,
                "y_mm": y_mm,
                "z_mm": z_mm,
                "diameter_mm": _parse_int(row_data.get("diameter_mm")),
                "spindle_mm": _parse_int(row_data.get("spindle_mm")),
                "year": year,
                "tmc": _parse_int(row_data.get("tmc")),
                "tcr_enabled": _parse_bool(row_data.get("tcr_enabled")),
                "pressure_bar": _parse_decimal(row_data.get("pressure_bar")),
                "cnc_controlled": _parse_bool(row_data.get("cnc_controlled")),
                "five_axes": _parse_bool(row_data.get("five_axes")),
                "accuracy_from": _clean_str(row_data.get("accuracy_from"))[:120],
            }

            if machine is None:
                if dry_run:
                    stats["machines_created"] += 1
                    continue
                WorkMachine.objects.create(asset=asset, **machine_payload)
                stats["machines_created"] += 1
                continue

            if not update_existing:
                continue

            machine_changed: list[str] = []
            for field_name, next_value in machine_payload.items():
                if getattr(machine, field_name) != next_value:
                    setattr(machine, field_name, next_value)
                    machine_changed.append(field_name)

            if machine_changed:
                if not dry_run:
                    machine.save(update_fields=sorted(set(machine_changed + ["updated_at"])))
                stats["machines_updated"] += 1
            elif asset_created and dry_run:
                stats["machines_created"] += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Import completato | "
                f"righe lette={stats['rows_seen']}, "
                f"saltate={stats['rows_skipped']}, "
                f"asset creati={stats['assets_created']}, "
                f"asset aggiornati={stats['assets_updated']}, "
                f"macchine create={stats['machines_created']}, "
                f"macchine aggiornate={stats['machines_updated']}"
            )
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna modifica e stata salvata nel database."))
