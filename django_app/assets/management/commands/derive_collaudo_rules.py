from __future__ import annotations

import datetime as dt
import statistics
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from assets.maintenance import upsert_asset_maintenance_rule_state
from assets.models import (
    Asset,
    AssetMaintenanceRuleState,
    MaintenanceInterventionTemplate,
    MaintenanceRule,
)

EXCLUDED_CAUSALI = {"A10", "A11", "A12", "A13"}
SKIP_TOKENS = {"NUOVO", "SALTA", "SKIP", "ORFANO", ""}

# Frequenze canoniche (giorni) su cui "snappare" la mediana dei gap osservati.
CANON_DAYS = [15, 30, 60, 91, 122, 182, 274, 365, 547, 730, 1096, 1461, 1826, 2557, 3652, 5478, 7305]
CANON_LABELS = {
    15: "15 giorni", 30: "mensile", 60: "bimestrale", 91: "trimestrale", 122: "quadrimestrale",
    182: "semestrale", 274: "9 mesi", 365: "annuale", 547: "18 mesi", 730: "biennale",
    1096: "triennale", 1461: "quadriennale", 1826: "quinquennale", 2557: "7 anni",
    3652: "decennale", 5478: "15 anni", 7305: "ventennale",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return " ".join(_clean(value).lower().split())


def _as_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _template_code(causale: str) -> str:
    return f"caus-{slugify(causale)}"


def _snap(days: float) -> int:
    return min(CANON_DAYS, key=lambda c: abs(c - days))


def _warning_days(threshold: int) -> int:
    return max(15, min(180, round(threshold * 0.08)))


class Command(BaseCommand):
    help = (
        "Deriva MaintenanceRule (categoria x causale) dalle frequenze osservate nello storico collaudo "
        "e seed di AssetMaintenanceRuleState con l'ultima esecuzione, per le macchine confermate nel foglio "
        "di riconciliazione. Idempotente, dry-run di default. Cosi lo scadenzario parte gia accurato."
    )

    def add_arguments(self, parser):
        parser.add_argument("--storico", default="storico collaudo.xlsx")
        parser.add_argument("--mapping", default="docs/assets/riconciliazione_collaudo_macchine_PROD.xlsx")
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument("--dry-run", action="store_true")
        mode.add_argument("--commit", action="store_true")

    def _load_mapping(self, path: Path) -> dict[tuple[str, str], str]:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [_norm_header(c) for c in rows[0]]

        def col(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        i_centro, i_mac, i_conf = col("centro"), col("macc.ut", "macc ut"), col("conferma")
        if None in (i_centro, i_mac, i_conf):
            raise CommandError(f"Colonne mancanti nel foglio. Header: {header}")
        mapping: dict[tuple[str, str], str] = {}
        for r in rows[1:]:
            centro, mac, conf = _clean(r[i_centro]).upper(), _clean(r[i_mac]), _clean(r[i_conf])
            if centro and mac and conf.upper() not in SKIP_TOKENS:
                mapping[(centro, mac)] = conf
        return mapping

    def handle(self, *args, **options):
        storico_path = Path(str(options.get("storico") or "")).expanduser()
        mapping_path = Path(str(options.get("mapping") or "")).expanduser()
        if not storico_path.exists():
            raise CommandError(f"Storico non trovato: {storico_path}")
        if not mapping_path.exists():
            raise CommandError(f"Foglio di riconciliazione non trovato: {mapping_path}")
        dry_run = bool(options.get("dry_run"))

        mapping = self._load_mapping(mapping_path)
        assets_by_tag = {a.asset_tag: a for a in Asset.objects.select_related("asset_category").filter(asset_tag__in=set(mapping.values()))}
        templates_by_code = {t.code: t for t in MaintenanceInterventionTemplate.objects.all()}

        wb = load_workbook(filename=str(storico_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [_norm_header(c) for c in rows[0]]

        def col(name):
            for i, h in enumerate(header):
                if name in h:
                    return i
            return None

        i_dpian, i_dult = col("data pianif"), col("dt.ult")
        i_cod, i_clav, i_mac = col("cod. test"), col("c.di lav"), col("macc.ut")

        # raccolta: gap per (categoria, template); ultima esecuzione per (asset, categoria, template)
        gaps: dict[tuple[int, int], list[int]] = {}
        last_exec: dict[tuple[int, int, int], dt.date] = {}
        cat_label: dict[int, str] = {}
        tpl_label: dict[int, str] = {}
        skipped_no_cat = set()

        for r in rows[1:]:
            tag = mapping.get((_clean(r[i_clav]).upper(), _clean(r[i_mac])))
            if not tag:
                continue
            cod = _clean(r[i_cod]).upper()
            if cod in EXCLUDED_CAUSALI:
                continue
            asset = assets_by_tag.get(tag)
            template = templates_by_code.get(_template_code(cod))
            if asset is None or template is None:
                continue
            if not asset.asset_category_id:
                skipped_no_cat.add(tag)
                continue
            ck = (asset.asset_category_id, template.id)
            cat_label[asset.asset_category_id] = getattr(asset.asset_category, "label", str(asset.asset_category_id))
            tpl_label[template.id] = template.label

            dp, du = _as_date(r[i_dpian]), _as_date(r[i_dult])
            if dp and du:
                delta = (dp - du).days
                if delta > 0:
                    gaps.setdefault(ck, []).append(delta)
            if du:
                ak = (asset.id, asset.asset_category_id, template.id)
                if ak not in last_exec or du > last_exec[ak]:
                    last_exec[ak] = du

        # deriva soglie
        derived: dict[tuple[int, int], dict[str, Any]] = {}
        for ck, gl in gaps.items():
            median = int(statistics.median(gl))
            snapped = _snap(median)
            derived[ck] = {
                "median": median, "snapped": snapped, "n": len(gl),
                "label": CANON_LABELS.get(snapped, f"{snapped} gg"),
                "warning": _warning_days(snapped),
            }

        rules_created = rules_updated = states_upserted = 0

        def _apply():
            nonlocal rules_created, rules_updated, states_upserted
            rule_by_ck: dict[tuple[int, int], MaintenanceRule] = {}
            for (cat_id, tpl_id), info in derived.items():
                rule, created = MaintenanceRule.objects.get_or_create(
                    asset_category_id=cat_id,
                    intervention_template_id=tpl_id,
                    defaults={
                        "threshold_type": MaintenanceRule.THRESHOLD_DAYS,
                        "threshold_value": info["snapped"],
                        "warning_days": info["warning"],
                        "is_active": True,
                        "notes": f"Frequenza derivata da storico collaudo (mediana {info['median']}gg, n={info['n']}).",
                    },
                )
                if created:
                    rules_created += 1
                else:
                    changed = []
                    if rule.threshold_value != info["snapped"]:
                        rule.threshold_value = info["snapped"]; changed.append("threshold_value")
                    if rule.warning_days != info["warning"]:
                        rule.warning_days = info["warning"]; changed.append("warning_days")
                    if changed:
                        rule.save(update_fields=[*changed, "updated_at"]); rules_updated += 1
                rule_by_ck[(cat_id, tpl_id)] = rule

            for (asset_id, cat_id, tpl_id), executed in last_exec.items():
                rule = rule_by_ck.get((cat_id, tpl_id))
                if rule is None:
                    continue
                upsert_asset_maintenance_rule_state(
                    asset=assets_by_tag_by_id[asset_id], base_rule=rule, executed_on=executed,
                )
                states_upserted += 1

        assets_by_tag_by_id = {a.id: a for a in assets_by_tag.values()}

        if not dry_run:
            with transaction.atomic():
                _apply()

        # report
        self.stdout.write(f"Storico: {storico_path}")
        self.stdout.write(f"Mapping: macchine confermate={len(mapping)}")
        self.stdout.write(f"Modalita: {'DRY-RUN' if dry_run else 'COMMIT'}")
        self.stdout.write(f"Regole (categoria x causale) derivate: {len(derived)} | stati asset da seed: {len(last_exec)}")
        if skipped_no_cat:
            self.stdout.write(self.style.WARNING(f"Asset senza categoria (saltati per le regole): {len(skipped_no_cat)} -> {sorted(skipped_no_cat)[:5]}..."))
        self.stdout.write("")
        self.stdout.write(f"{'CATEGORIA':16s} {'CAUSALE':34s} {'n':>3s} {'mediana':>8s} {'->':>3s} {'FREQUENZA':>12s}")
        for (cat_id, tpl_id), info in sorted(derived.items(), key=lambda x: (cat_label.get(x[0][0], ''), -x[1]['n'])):
            self.stdout.write(
                f"{cat_label.get(cat_id,'?')[:16]:16s} {tpl_label.get(tpl_id,'?')[:34]:34s} "
                f"{info['n']:>3d} {info['median']:>6d}gg -> {info['snapped']:>5d}gg {info['label']}"
            )
        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna regola/stato salvato."))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"Regole create={rules_created}, aggiornate={rules_updated}, stati asset={states_upserted}. "
                "Lo scadenzario ora parte dall'ultima esecuzione reale."
            ))
