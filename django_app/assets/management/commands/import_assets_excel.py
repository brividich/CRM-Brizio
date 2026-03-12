from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from assets.models import Asset, AssetCustomField, AssetEndpoint, AssetITDetails, WorkOrder

DEFAULT_PRIMARY_SHEETS = [
    "LAN A 203.0.113.x",
    "LAN B 198.51.100.x",
    "LAN C 192.0.2.x",
]

OPTIONAL_SHEETS = [
    "CCTV 198.51.100.X",
    "GUEST-LAN 203.0.113.X",
    "MASS-STORAGE",
    "Telefonia",
    "SIM Telefonica",
]

TRUE_VALUES = {
    "1",
    "S",
    "SI",
    "YES",
    "Y",
    "TRUE",
    "ON",
    "OK",
    "ATTIVO",
    "ABILITATO",
    "PRESENTE",
}
FALSE_VALUES = {
    "0",
    "N",
    "NO",
    "FALSE",
    "OFF",
    "NONE",
    "NULL",
    "N/A",
    "NA",
    "ND",
    "N D",
    "-",
}

SENSITIVE_KEYWORDS = {
    "PASSWORD",
    "PASS",
    "PWD",
    "PSW",
    "PIN",
    "PUK",
    "SECRET",
    "CREDENZIAL",
    "CREDENTIAL",
}

HEADER_SCAN_MAX_ROWS = 30
MAX_COLS_PER_ROW = 250


HEADER_ALIASES = {
    "reparto": ["REPARTO", "DEPARTMENT", "REP"],
    "name": [
        "NOME PC",
        "NOME",
        "HOSTNAME",
        "NOME MACCHINA",
        "MACCHINA",
        "ASSET NAME",
        "DEVICE NAME",
        "APPARATO",
    ],
    "type": ["TIPO", "TIPOLOGIA", "CATEGORY", "CATEGORIA"],
    "manufacturer": ["MANUFACTURER", "MARCA", "VENDOR", "PRODUTTORE"],
    "model": ["MODELLO", "MODEL"],
    "serial": ["ID", "SERIALE", "SERIAL", "SERIAL NUMBER", "SN", "S N", "MATRICOLA", "IMEI", "ICCID"],
    "vlan": ["VLAN"],
    "ip": ["IP", "IP ADDRESS", "INDIRIZZO IP"],
    "switch_name": ["SWITCH", "SWITCH NAME", "NOME SWITCH"],
    "switch_port": ["PORTA SWITCH", "PORTA SW", "SWITCH PORT", "PORT SWITCH"],
    "punto": ["PUNTO", "PRESA", "PATCH", "LOCATION POINT"],
    "os": ["OS", "OPERATING SYSTEM", "SISTEMA OPERATIVO"],
    "cpu": ["CPU", "PROCESSOR", "PROCESSORE"],
    "ram": ["RAM", "MEMORY", "MEMORIA", "MEMORIA RAM"],
    "disco": ["DISCO", "DISK", "STORAGE", "HDD", "SSD"],
    "domain": ["DOMAIN", "DOMINIO"],
    "edpr": ["EDPR", "EDR"],
    "ad360": ["AD360"],
    "office_2fa": ["2FA OFFICE", "OFFICE 2FA", "MFA OFFICE", "2FA"],
    "bios_pwd": ["PSW BIOS", "BIOS PASSWORD", "PASSWORD BIOS", "PWD BIOS"],
    "ultima_mtz": ["ULTIMA MTZ", "ULTIMA MANUTENZIONE", "LAST MAINTENANCE", "DATA MANUTENZIONE"],
    "status": ["STATO", "STATUS"],
    "notes": ["NOTE", "NOTES", "COMMENTI"],
    "assignment_to": ["ASSIGNMENT TO", "ASSEGNATO A", "DIPENDENTE", "UTENTE", "USER"],
    "assignment_reparto": ["ASSIGNMENT REPARTO", "REPARTO ASSEGNAZIONE", "REPARTO ASSEGNATO"],
    "assignment_location": ["ASSIGNMENT LOCATION", "LOCATION", "SEDE", "UBICAZIONE", "POSIZIONE"],
}

KNOWN_CANONICAL_KEYS = set(HEADER_ALIASES.keys())


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    row = str(value).strip()
    if not row:
        return ""
    row = unicodedata.normalize("NFKD", row)
    row = "".join(ch for ch in row if not unicodedata.combining(ch))
    row = row.replace("\n", " ").replace("\r", " ")
    row = re.sub(r"[^A-Za-z0-9]+", " ", row)
    row = re.sub(r"\s+", " ", row).strip().upper()
    return row


def _compact(value: Any) -> str:
    return _normalize_text(value).replace(" ", "")


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("%f" % value).rstrip("0").rstrip(".")
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    return _clean_str(value) == ""


def _sheet_kind(sheet_name: str) -> str:
    normalized = _normalize_text(sheet_name)
    if "TVCC" in normalized or "CCTV" in normalized:
        return "CCTV"
    if "SIM" in normalized or "TELEFONIA" in normalized:
        return "TELEFONIA"
    if "TRUST" in normalized or "SEGMENTATION" in normalized or "MANAGEMENT" in normalized:
        return "IT"
    return "GENERIC"


def _map_asset_type(raw_value: Any, sheet_name: str) -> str:
    token = _normalize_text(raw_value)
    compact = token.replace(" ", "")

    if compact in {"PC", "DESKTOP"}:
        return Asset.TYPE_PC
    if compact in {"NOTEBOOK", "LAPTOP", "NB"}:
        return Asset.TYPE_NOTEBOOK
    if compact in {"SERVER", "SRV"}:
        return Asset.TYPE_SERVER
    if compact in {"VM", "VIRTUALMACHINE", "VIRTUAL"}:
        return Asset.TYPE_VM
    if compact in {"FIREWALL", "FW", "ROUTER", "SWITCH", "NETWORK", "NETWORKING", "ACCESSPOINT", "AP"}:
        return Asset.TYPE_FIREWALL
    if compact in {"STAMPANTE", "PRINTER", "MULTIFUNZIONE"}:
        return Asset.TYPE_STAMPANTE
    if compact in {"HW", "HARDWARE", "TELEFONIA", "PHONE", "TABLET", "SIM"}:
        return Asset.TYPE_HW
    if compact in {"CCTV", "TVCC", "CAMERA", "VIDEOSORVEGLIANZA"}:
        return Asset.TYPE_CCTV

    if token:
        cnc_markers = [
            "CNC",
            "TORNIO",
            "FRESA",
            "PRESSA",
            "TRAPANO",
            "LASER",
            "PIEGA",
            "OFFICINA",
            "MACCHINA",
            "MACCHINARIO",
            "ATTREZZATURA",
            "CENTRO LAVORO",
        ]
        for marker in cnc_markers:
            if marker in token:
                return Asset.TYPE_CNC

    sheet_kind = _sheet_kind(sheet_name)
    if sheet_kind == "CCTV":
        return Asset.TYPE_CCTV
    if sheet_kind == "TELEFONIA":
        return Asset.TYPE_HW
    return Asset.TYPE_OTHER


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
        return None
    token = _normalize_text(value)
    if not token:
        return None
    if token in TRUE_VALUES:
        return True
    if token in FALSE_VALUES:
        return False
    return None


def _to_bool(value: Any, default: bool = False) -> bool:
    parsed = _parse_bool(value)
    if parsed is None:
        if _is_blank(value):
            return default
        token = _normalize_text(value)
        return token not in FALSE_VALUES
    return parsed


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    token = _clean_str(value).replace(",", ".")
    if not token:
        return None
    try:
        if "." in token:
            number = float(token)
            if number.is_integer():
                return int(number)
            return None
        return int(token)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        if 20000 <= float(value) <= 80000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()
        return None

    token = _clean_str(value)
    if not token:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _to_aware_datetime(row_date: date) -> datetime:
    naive = datetime.combine(row_date, time(hour=8, minute=0, second=0))
    tz = timezone.get_current_timezone()
    return timezone.make_aware(naive, tz)


def _looks_sensitive(header_label: str) -> bool:
    normalized = _normalize_text(header_label)
    if not normalized:
        return False
    return any(keyword in normalized for keyword in SENSITIVE_KEYWORDS)


def _canonical_from_header(header_label: str, alias_map: dict[str, str]) -> str | None:
    normalized = _normalize_text(header_label)
    compact = normalized.replace(" ", "")
    if not normalized:
        return None

    if compact in alias_map:
        return alias_map[compact]

    if "PORTA" in normalized and "SW" in normalized:
        return "switch_port"
    if "PORTA" in normalized and "SWITCH" in normalized:
        return "switch_port"
    if "SWITCH" in normalized and "PORT" in normalized:
        return "switch_port"
    if normalized == "SWITCH" or normalized.endswith(" SWITCH"):
        return "switch_name"
    if "VLAN" in normalized:
        return "vlan"
    if normalized in {"IP", "INDIRIZZO IP", "IP ADDRESS"}:
        return "ip"
    if "PUNTO" in normalized or "PRESA" in normalized:
        return "punto"
    if "REPARTO" in normalized:
        return "reparto"
    if "MODEL" in normalized:
        return "model"
    if "SERIAL" in normalized or "SERIALE" in normalized or normalized == "ID":
        return "serial"
    if normalized == "TIPO" or "TIPOLOGIA" in normalized:
        return "type"
    if normalized.startswith("NOME") or normalized.endswith(" NAME"):
        return "name"
    return None


def _build_alias_map() -> dict[str, str]:
    alias_map: dict[str, str] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = _compact(alias)
            if key:
                alias_map[key] = canonical
    return alias_map


def _resolve_requested_sheets(
    workbook_sheet_names: list[str],
    requested_csv: str,
    include_optional: bool,
    all_sheets: bool,
) -> tuple[list[str], list[str]]:
    if all_sheets:
        return workbook_sheet_names, []

    requested_names = [row.strip() for row in requested_csv.split(",") if row.strip()]
    if not requested_names:
        requested_names = list(DEFAULT_PRIMARY_SHEETS)
    if include_optional:
        requested_names.extend(OPTIONAL_SHEETS)

    normalized_map = {_compact(name): name for name in workbook_sheet_names}
    selected: list[str] = []
    missing: list[str] = []

    for requested in requested_names:
        normalized_requested = _compact(requested)
        matched = normalized_map.get(normalized_requested)
        if not matched:
            for wb_name in workbook_sheet_names:
                wb_norm = _compact(wb_name)
                if normalized_requested and (normalized_requested in wb_norm or wb_norm in normalized_requested):
                    matched = wb_name
                    break
        if matched:
            if matched not in selected:
                selected.append(matched)
        else:
            missing.append(requested)

    if not selected:
        return workbook_sheet_names, missing

    return selected, missing


def _detect_header_row(ws, alias_map: dict[str, str]) -> int:
    max_row = min(ws.max_row or 0, HEADER_SCAN_MAX_ROWS)
    best_row = 5
    best_score = -1

    for row_idx in range(1, max_row + 1):
        score = 0
        non_empty = 0
        for col_idx in range(1, min((ws.max_column or 0), MAX_COLS_PER_ROW) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if _is_blank(value):
                continue
            non_empty += 1
            canonical = _canonical_from_header(_clean_str(value), alias_map)
            if canonical in KNOWN_CANONICAL_KEYS:
                score += 1
        if score > best_score and non_empty > 0:
            best_score = score
            best_row = row_idx

    return best_row


def _iter_header_cells(ws, header_row: int) -> dict[int, str]:
    headers: dict[int, str] = {}
    max_col = min(ws.max_column or 0, MAX_COLS_PER_ROW)
    for col_idx in range(1, max_col + 1):
        label = _clean_str(ws.cell(row=header_row, column=col_idx).value)
        if label:
            headers[col_idx] = label
    return headers


def _source_key(sheet_name: str, reparto: str, tipo: str, nome: str, serial: str, vlan: str, ip: str) -> str:
    raw = f"{sheet_name}|{reparto}|{tipo}|{nome}|{serial}|{vlan}|{ip}"
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()


class Command(BaseCommand):
    help = "Import massivo asset da Excel con supporto fogli e colonne dinamiche (upsert su source_key)."

    def add_arguments(self, parser):
        parser.add_argument("--file", default="CN - Asset Inventory (1).xlsx", help="Path file Excel da importare.")
        parser.add_argument("--sheets", default=",".join(DEFAULT_PRIMARY_SHEETS), help="Lista fogli separata da virgola.")
        parser.add_argument(
            "--include-optional",
            action="store_true",
            help="Include automaticamente i fogli opzionali (TVCC/Telefonia/SIM/etc).",
        )
        parser.add_argument("--all-sheets", action="store_true", help="Importa tutti i fogli presenti nel workbook.")
        parser.add_argument("--dry-run", action="store_true", help="Simula import senza scrivere su DB.")
        parser.add_argument(
            "--update",
            dest="update",
            action="store_true",
            default=True,
            help="Aggiorna record esistenti (default).",
        )
        parser.add_argument(
            "--no-update",
            dest="update",
            action="store_false",
            help="Non aggiornare record esistenti (solo insert).",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"File non trovato: {file_path}")

        dry_run = bool(options.get("dry_run"))
        update_existing = bool(options.get("update", True))
        include_optional = bool(options.get("include_optional"))
        all_sheets = bool(options.get("all_sheets"))
        requested_sheets = str(options.get("sheets") or "")

        try:
            workbook = load_workbook(filename=str(file_path), data_only=True)
        except Exception as exc:
            raise CommandError(f"Impossibile aprire workbook: {exc}") from exc

        workbook_sheets = list(workbook.sheetnames)
        selected_sheets, missing_sheets = _resolve_requested_sheets(
            workbook_sheet_names=workbook_sheets,
            requested_csv=requested_sheets,
            include_optional=include_optional,
            all_sheets=all_sheets,
        )

        if missing_sheets:
            self.stdout.write(
                self.style.WARNING(
                    "Fogli richiesti non trovati (proseguo con quelli disponibili): " + ", ".join(missing_sheets)
                )
            )

        if not selected_sheets:
            raise CommandError("Nessun foglio disponibile da importare.")

        self.stdout.write(f"Workbook: {file_path}")
        self.stdout.write(f"Fogli selezionati: {', '.join(selected_sheets)}")
        self.stdout.write(f"Modalita: {'DRY-RUN' if dry_run else 'IMPORT REALE'} | update={update_existing}")

        alias_map = _build_alias_map()

        field_by_label_key = {_normalize_text(field.label): field for field in AssetCustomField.objects.all()}
        field_by_code = {field.code: field for field in AssetCustomField.objects.all()}

        stats = {
            "rows_seen": 0,
            "rows_skipped": 0,
            "assets_created": 0,
            "assets_updated": 0,
            "assets_unchanged": 0,
            "assets_skipped_update": 0,
            "endpoints_created": 0,
            "endpoints_updated": 0,
            "it_details_updated": 0,
            "workorders_created": 0,
            "custom_fields_created": 0,
            "errors": 0,
        }

        def ensure_custom_field(label: str, field_type: str, sensitive: bool = False) -> AssetCustomField | None:
            effective_label = f"{label} (presente)" if sensitive else label
            key = _normalize_text(effective_label)
            if key in field_by_label_key:
                return field_by_label_key[key]
            if dry_run:
                return None

            base_code = slugify(effective_label)[:75] or f"extra-{hashlib.sha1(effective_label.encode('utf-8')).hexdigest()[:8]}"
            candidate = base_code
            suffix = 2
            while candidate in field_by_code:
                candidate = f"{base_code[: max(1, 75 - len(str(suffix)) - 1)]}-{suffix}"
                suffix += 1

            field = AssetCustomField.objects.create(
                code=candidate,
                label=effective_label[:120],
                field_type=field_type,
                sort_order=900,
                is_active=True,
            )
            field_by_label_key[key] = field
            field_by_code[field.code] = field
            stats["custom_fields_created"] += 1
            return field

        for sheet_name in selected_sheets:
            ws = workbook[sheet_name]
            header_row = _detect_header_row(ws, alias_map)
            header_cells = _iter_header_cells(ws, header_row)
            if not header_cells:
                self.stdout.write(self.style.WARNING(f"[{sheet_name}] Nessun header rilevato, foglio ignorato."))
                continue

            header_to_canonical: dict[str, str | None] = {}
            for label in header_cells.values():
                header_to_canonical[label] = _canonical_from_header(label, alias_map)

            first_data_row = header_row + 1
            max_row = ws.max_row or first_data_row
            self.stdout.write(f"[{sheet_name}] header row={header_row}, righe={max(0, max_row - header_row)}")

            for row_idx in range(first_data_row, max_row + 1):
                stats["rows_seen"] += 1
                row_cells: dict[str, Any] = {}
                has_any_value = False

                for col_idx, header_label in header_cells.items():
                    raw_value = ws.cell(row=row_idx, column=col_idx).value
                    row_cells[header_label] = raw_value
                    if not _is_blank(raw_value):
                        has_any_value = True

                if not has_any_value:
                    continue

                canonical_values: dict[str, Any] = {}
                extra_values: dict[str, Any] = {}
                for header_label, raw_value in row_cells.items():
                    canonical_key = header_to_canonical.get(header_label)
                    if canonical_key:
                        if canonical_key not in canonical_values or _is_blank(canonical_values.get(canonical_key)):
                            canonical_values[canonical_key] = raw_value
                    else:
                        extra_values[header_label] = raw_value

                name_value = _clean_str(canonical_values.get("name"))
                endpoint_name_value = name_value
                serial_value = _clean_str(canonical_values.get("serial"))
                ip_value = _clean_str(canonical_values.get("ip"))
                reparto_value = _clean_str(canonical_values.get("reparto"))
                type_raw_value = _clean_str(canonical_values.get("type"))
                vlan_int = _parse_int(canonical_values.get("vlan"))
                vlan_value = "" if vlan_int is None else str(vlan_int)

                if not name_value:
                    name_value = endpoint_name_value or serial_value or ip_value
                if not name_value:
                    name_value = f"Asset {sheet_name} #{row_idx}"

                if not any([name_value, serial_value, ip_value, reparto_value, type_raw_value]):
                    stats["rows_skipped"] += 1
                    continue

                asset_type_value = _map_asset_type(type_raw_value, sheet_name)
                source_key = _source_key(
                    sheet_name=sheet_name,
                    reparto=reparto_value,
                    tipo=type_raw_value,
                    nome=name_value,
                    serial=serial_value,
                    vlan=vlan_value,
                    ip=ip_value,
                )

                try:
                    asset = Asset.objects.filter(source_key=source_key).first()
                    if asset is None:
                        asset = Asset(
                            source_key=source_key,
                            name=name_value[:255],
                            asset_type=asset_type_value,
                            reparto=reparto_value[:120],
                            manufacturer=_clean_str(canonical_values.get("manufacturer"))[:120] or None,
                            model=_clean_str(canonical_values.get("model"))[:120] or None,
                            serial_number=serial_value[:120] or None,
                            status=Asset.STATUS_IN_USE,
                            notes=_clean_str(canonical_values.get("notes")),
                            assignment_to=_clean_str(canonical_values.get("assignment_to"))[:200],
                            assignment_reparto=_clean_str(canonical_values.get("assignment_reparto"))[:120],
                            assignment_location=_clean_str(canonical_values.get("assignment_location"))[:200],
                        )

                        status_text = _normalize_text(canonical_values.get("status"))
                        if status_text:
                            if "REPAIR" in status_text or "RIPAR" in status_text:
                                asset.status = Asset.STATUS_IN_REPAIR
                            elif "RETIR" in status_text or "DISMES" in status_text:
                                asset.status = Asset.STATUS_RETIRED
                            elif "STOCK" in status_text or "MAGAZ" in status_text:
                                asset.status = Asset.STATUS_IN_STOCK

                        if not dry_run:
                            asset.save()
                        stats["assets_created"] += 1
                    else:
                        if not update_existing:
                            stats["assets_skipped_update"] += 1
                            continue

                        changed_fields: list[str] = []

                        def update_if_value(field_name: str, value: Any, max_len: int | None = None):
                            if value is None:
                                return
                            raw = _clean_str(value)
                            if raw == "":
                                return
                            next_value: Any = raw
                            if max_len is not None:
                                next_value = raw[:max_len]
                            current = getattr(asset, field_name)
                            if current != next_value:
                                setattr(asset, field_name, next_value)
                                changed_fields.append(field_name)

                        update_if_value("name", name_value, 255)
                        if asset_type_value != Asset.TYPE_OTHER and asset.asset_type != asset_type_value:
                            asset.asset_type = asset_type_value
                            changed_fields.append("asset_type")
                        update_if_value("reparto", reparto_value, 120)

                        manufacturer = _clean_str(canonical_values.get("manufacturer"))
                        if manufacturer and asset.manufacturer != manufacturer[:120]:
                            asset.manufacturer = manufacturer[:120]
                            changed_fields.append("manufacturer")

                        model = _clean_str(canonical_values.get("model"))
                        if model and asset.model != model[:120]:
                            asset.model = model[:120]
                            changed_fields.append("model")

                        if serial_value and asset.serial_number != serial_value[:120]:
                            asset.serial_number = serial_value[:120]
                            changed_fields.append("serial_number")

                        notes = _clean_str(canonical_values.get("notes"))
                        if notes and asset.notes != notes:
                            asset.notes = notes
                            changed_fields.append("notes")

                        assign_to = _clean_str(canonical_values.get("assignment_to"))
                        if assign_to and asset.assignment_to != assign_to[:200]:
                            asset.assignment_to = assign_to[:200]
                            changed_fields.append("assignment_to")

                        assign_rep = _clean_str(canonical_values.get("assignment_reparto"))
                        if assign_rep and asset.assignment_reparto != assign_rep[:120]:
                            asset.assignment_reparto = assign_rep[:120]
                            changed_fields.append("assignment_reparto")

                        assign_loc = _clean_str(canonical_values.get("assignment_location"))
                        if assign_loc and asset.assignment_location != assign_loc[:200]:
                            asset.assignment_location = assign_loc[:200]
                            changed_fields.append("assignment_location")

                        status_text = _normalize_text(canonical_values.get("status"))
                        if status_text:
                            mapped_status = asset.status
                            if "REPAIR" in status_text or "RIPAR" in status_text:
                                mapped_status = Asset.STATUS_IN_REPAIR
                            elif "RETIR" in status_text or "DISMES" in status_text:
                                mapped_status = Asset.STATUS_RETIRED
                            elif "STOCK" in status_text or "MAGAZ" in status_text:
                                mapped_status = Asset.STATUS_IN_STOCK
                            elif "USE" in status_text or "USO" in status_text or "ATTIV" in status_text:
                                mapped_status = Asset.STATUS_IN_USE
                            if mapped_status != asset.status:
                                asset.status = mapped_status
                                changed_fields.append("status")

                        if changed_fields:
                            if not dry_run:
                                asset.save(update_fields=sorted(set(changed_fields + ["updated_at"])))
                            stats["assets_updated"] += 1
                        else:
                            stats["assets_unchanged"] += 1

                    merged_extra = dict(asset.extra_columns or {}) if hasattr(asset, "extra_columns") else {}
                    extra_changed = False
                    for header_label, raw_value in extra_values.items():
                        if _is_blank(raw_value):
                            continue

                        is_sensitive = _looks_sensitive(header_label)
                        if is_sensitive:
                            field = ensure_custom_field(header_label, AssetCustomField.TYPE_BOOL, sensitive=True)
                            if field is None:
                                continue
                            value_to_store = not _is_blank(raw_value)
                        else:
                            parsed_bool = _parse_bool(raw_value)
                            parsed_date = _parse_date(raw_value)
                            parsed_int = _parse_int(raw_value)
                            field_type = AssetCustomField.TYPE_TEXT
                            if parsed_bool is not None:
                                field_type = AssetCustomField.TYPE_BOOL
                            elif parsed_date is not None:
                                field_type = AssetCustomField.TYPE_DATE
                            elif parsed_int is not None:
                                field_type = AssetCustomField.TYPE_NUMBER

                            field = ensure_custom_field(header_label, field_type, sensitive=False)
                            if field is None:
                                continue

                            if field.field_type == AssetCustomField.TYPE_BOOL:
                                value_to_store = _to_bool(raw_value, default=False)
                            elif field.field_type == AssetCustomField.TYPE_DATE:
                                date_val = _parse_date(raw_value)
                                value_to_store = date_val.isoformat() if date_val else _clean_str(raw_value)
                            elif field.field_type == AssetCustomField.TYPE_NUMBER:
                                number_val = _parse_int(raw_value)
                                value_to_store = number_val if number_val is not None else _clean_str(raw_value)
                            else:
                                value_to_store = _clean_str(raw_value)

                        if merged_extra.get(field.code) != value_to_store:
                            merged_extra[field.code] = value_to_store
                            extra_changed = True

                    if extra_changed and not dry_run:
                        asset.extra_columns = merged_extra
                        asset.save(update_fields=["extra_columns", "updated_at"])

                    has_endpoint_payload = any(
                        [
                            endpoint_name_value,
                            vlan_int is not None,
                            ip_value,
                            _clean_str(canonical_values.get("switch_name")),
                            _clean_str(canonical_values.get("switch_port")),
                            _clean_str(canonical_values.get("punto")),
                        ]
                    )

                    if has_endpoint_payload and not dry_run:
                        endpoint, endpoint_created = AssetEndpoint.objects.get_or_create(
                            asset=asset,
                            endpoint_name=endpoint_name_value[:255],
                            vlan=vlan_int,
                            ip=ip_value[:80] or None,
                            defaults={
                                "switch_name": _clean_str(canonical_values.get("switch_name"))[:120],
                                "switch_port": _clean_str(canonical_values.get("switch_port"))[:120],
                                "punto": _clean_str(canonical_values.get("punto"))[:120],
                            },
                        )
                        if endpoint_created:
                            stats["endpoints_created"] += 1
                        else:
                            endpoint_changed = False
                            switch_name = _clean_str(canonical_values.get("switch_name"))[:120]
                            switch_port = _clean_str(canonical_values.get("switch_port"))[:120]
                            punto = _clean_str(canonical_values.get("punto"))[:120]
                            if switch_name and endpoint.switch_name != switch_name:
                                endpoint.switch_name = switch_name
                                endpoint_changed = True
                            if switch_port and endpoint.switch_port != switch_port:
                                endpoint.switch_port = switch_port
                                endpoint_changed = True
                            if punto and endpoint.punto != punto:
                                endpoint.punto = punto
                                endpoint_changed = True
                            if endpoint_changed:
                                endpoint.save(update_fields=["switch_name", "switch_port", "punto"])
                                stats["endpoints_updated"] += 1

                    has_it_payload = any(
                        [
                            not _is_blank(canonical_values.get("os")),
                            not _is_blank(canonical_values.get("cpu")),
                            not _is_blank(canonical_values.get("ram")),
                            not _is_blank(canonical_values.get("disco")),
                            not _is_blank(canonical_values.get("domain")),
                            not _is_blank(canonical_values.get("edpr")),
                            not _is_blank(canonical_values.get("ad360")),
                            not _is_blank(canonical_values.get("office_2fa")),
                            not _is_blank(canonical_values.get("bios_pwd")),
                        ]
                    )

                    if has_it_payload and not dry_run:
                        details, _ = AssetITDetails.objects.get_or_create(asset=asset)
                        details_changed = False

                        for field_name in ["os", "cpu", "ram", "disco"]:
                            value = _clean_str(canonical_values.get(field_name))
                            if value and getattr(details, field_name) != value[:120]:
                                setattr(details, field_name, value[:120])
                                details_changed = True

                        bool_map = {
                            "domain": "domain_joined",
                            "edpr": "edr_enabled",
                            "ad360": "ad360_managed",
                            "office_2fa": "office_2fa_enabled",
                        }
                        for source_key, target_field in bool_map.items():
                            source_value = canonical_values.get(source_key)
                            if _is_blank(source_value):
                                continue
                            parsed = _to_bool(source_value, default=False)
                            if getattr(details, target_field) != parsed:
                                setattr(details, target_field, parsed)
                                details_changed = True

                        if not _is_blank(canonical_values.get("bios_pwd")):
                            bios_set = not _is_blank(canonical_values.get("bios_pwd"))
                            if details.bios_pwd_set != bios_set:
                                details.bios_pwd_set = bios_set
                                details_changed = True

                        if details_changed:
                            details.save()
                            stats["it_details_updated"] += 1

                    mtz_date = _parse_date(canonical_values.get("ultima_mtz"))
                    if mtz_date and not dry_run:
                        mtz_dt = _to_aware_datetime(mtz_date)
                        exists_wo = WorkOrder.objects.filter(
                            asset=asset,
                            kind=WorkOrder.KIND_PREVENTIVE,
                            status=WorkOrder.STATUS_DONE,
                            title="Manutenzione importata",
                            closed_at__date=mtz_date,
                        ).exists()
                        if not exists_wo:
                            WorkOrder.objects.create(
                                asset=asset,
                                kind=WorkOrder.KIND_PREVENTIVE,
                                status=WorkOrder.STATUS_DONE,
                                opened_at=mtz_dt,
                                closed_at=mtz_dt,
                                title="Manutenzione importata",
                                description=f"Importata da Excel ({sheet_name})",
                                resolution="Record storico importato",
                                downtime_minutes=0,
                            )
                            stats["workorders_created"] += 1

                except Exception as exc:
                    stats["errors"] += 1
                    self.stdout.write(self.style.ERROR(f"[{sheet_name}] Riga {row_idx}: {exc}"))

        self.stdout.write(
            self.style.SUCCESS(
                "Import completato | "
                f"righe lette={stats['rows_seen']}, "
                f"saltate={stats['rows_skipped']}, "
                f"asset creati={stats['assets_created']}, "
                f"asset aggiornati={stats['assets_updated']}, "
                f"asset invariati={stats['assets_unchanged']}, "
                f"asset skip-update={stats['assets_skipped_update']}, "
                f"endpoint creati={stats['endpoints_created']}, "
                f"endpoint aggiornati={stats['endpoints_updated']}, "
                f"it-details aggiornati={stats['it_details_updated']}, "
                f"WO creati={stats['workorders_created']}, "
                f"campi custom creati={stats['custom_fields_created']}, "
                f"errori={stats['errors']}"
            )
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna modifica e stata salvata nel database."))

