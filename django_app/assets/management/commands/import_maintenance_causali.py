from __future__ import annotations

from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from assets.models import MaintenanceInterventionTemplate


# Causali escluse da assets: corsi / formazione / sanitario / qualifiche.
# Vanno nel modulo Formazione (piano "Salute e Sicurezza"), dove sono gia
# gestite come TrainingCourse. Tenerle fuori da assets per coerenza + GDPR.
#   A10 Cliente | A11 Corso Esterno | A12 Visita Medica | A13 Ente Esterno
EXCLUDED_CAUSALI = {"A10", "A11", "A12", "A13"}

# Prefisso namespace per i code dei template (SlugField unico globale),
# evita collisioni con altri seed (es. "prova-antincendio").
CODE_PREFIX = "caus"

HEADER_CODE = "codice test"
HEADER_DESC = "descrizione test di collaudo"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return " ".join(_clean(value).lower().split())


def _template_code(causale: str) -> str:
    return f"{CODE_PREFIX}-{slugify(causale)}"


class Command(BaseCommand):
    help = (
        "Importa il catalogo causali di manutenzione (Causali manutenzione.xlsx) "
        "come MaintenanceInterventionTemplate idempotenti. Esclude le causali di "
        "corsi/formazione/sanitario (A10-A13), gestite nel modulo Formazione."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="Causali manutenzione.xlsx",
            help="Path del file Excel con le causali (Codice Test + Descrizione).",
        )
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument("--dry-run", action="store_true", help="Simula senza scrivere su DB.")
        mode.add_argument("--commit", action="store_true", help="Esegue in transaction.atomic().")

    def handle(self, *args, **options):
        file_path = Path(str(options.get("file") or "")).expanduser()
        if not file_path.exists():
            raise CommandError(f"File non trovato: {file_path}")

        dry_run = bool(options.get("dry_run"))

        try:
            workbook = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        except Exception as exc:  # pragma: no cover - dipende dal file
            raise CommandError(f"Impossibile aprire workbook: {exc}") from exc

        ws = workbook[workbook.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            raise CommandError("Foglio vuoto: nessuna causale da importare.")

        header = [_norm_header(c) for c in rows[0]]
        try:
            i_code = header.index(HEADER_CODE)
            i_desc = header.index(HEADER_DESC)
        except ValueError as exc:
            raise CommandError(
                f"Header non riconosciuto. Attese colonne '{HEADER_CODE}' e '{HEADER_DESC}', "
                f"trovate: {header}"
            ) from exc

        parsed: list[tuple[str, str]] = []  # (causale, descrizione)
        seen_codes: set[str] = set()
        skipped_excluded = 0
        skipped_empty = 0
        skipped_dup = 0

        for raw in rows[1:]:
            causale = _clean(raw[i_code]).upper()
            descrizione = " ".join(_clean(raw[i_desc]).split())
            if not causale or not descrizione:
                skipped_empty += 1
                continue
            if causale in EXCLUDED_CAUSALI:
                skipped_excluded += 1
                continue
            if causale in seen_codes:
                skipped_dup += 1
                continue
            seen_codes.add(causale)
            parsed.append((causale, descrizione))

        created = 0
        updated = 0
        unchanged = 0
        report_lines: list[str] = []

        def _apply() -> None:
            nonlocal created, updated, unchanged
            for sort_index, (causale, descrizione) in enumerate(parsed, start=1):
                code = _template_code(causale)
                sort_order = sort_index * 10
                ref_note = f"Causale collaudo {causale}"
                existing = MaintenanceInterventionTemplate.objects.filter(code=code).first()
                if existing is None:
                    if not dry_run:
                        MaintenanceInterventionTemplate.objects.create(
                            code=code,
                            label=descrizione[:120],
                            description=ref_note,
                            asset_category=None,
                            sort_order=sort_order,
                            is_active=True,
                        )
                    created += 1
                    report_lines.append(f"  + {causale:5s} -> {code} | {descrizione}")
                    continue

                changed: list[str] = []
                if existing.label != descrizione[:120]:
                    existing.label = descrizione[:120]
                    changed.append("label")
                if not (existing.description or "").strip():
                    existing.description = ref_note
                    changed.append("description")
                if existing.sort_order != sort_order:
                    existing.sort_order = sort_order
                    changed.append("sort_order")
                if changed:
                    if not dry_run:
                        existing.save(update_fields=sorted(set(changed + ["updated_at"])))
                    updated += 1
                    report_lines.append(f"  ~ {causale:5s} -> {code} | {', '.join(changed)}")
                else:
                    unchanged += 1

        if dry_run:
            _apply()
        else:
            with transaction.atomic():
                _apply()

        self.stdout.write(f"File: {file_path}")
        self.stdout.write(f"Modalita: {'DRY-RUN' if dry_run else 'COMMIT'}")
        self.stdout.write(
            "Conteggi | "
            f"causali valide={len(parsed)}, "
            f"create={created}, aggiornate={updated}, invariate={unchanged}, "
            f"escluse formazione/sanitario={skipped_excluded}, "
            f"righe vuote={skipped_empty}, duplicati={skipped_dup}"
        )
        for line in report_lines:
            self.stdout.write(line)

        if skipped_excluded:
            self.stdout.write(
                self.style.WARNING(
                    f"Escluse {skipped_excluded} causali corsi/sanitario "
                    f"({', '.join(sorted(EXCLUDED_CAUSALI))}) -> modulo Formazione."
                )
            )
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna modifica salvata nel database."))
        else:
            self.stdout.write(self.style.SUCCESS("Import causali completato."))
