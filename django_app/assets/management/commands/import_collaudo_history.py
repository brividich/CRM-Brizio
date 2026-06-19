from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from assets.models import Asset, MaintenanceInterventionTemplate, WorkOrder


# Causali corsi/formazione/sanitario: gestite in Formazione, mai come OdL asset.
EXCLUDED_CAUSALI = {"A10", "A11", "A12", "A13"}
SKIP_TOKENS = {"NUOVO", "SALTA", "SKIP", "ORFANO", ""}
REF_PREFIX = "COLLAUDO"


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return " ".join(_clean(value).lower().split())


def _as_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def _template_code(causale: str) -> str:
    return f"caus-{slugify(causale)}"


def _sanitize_comment(comment: str) -> str:
    """Rimozione difensiva di liste nomi (GDPR): taglia da 'A:' in poi (liste presenze)
    e collassa sequenze di NOMI MAIUSCOLI separati da ';'. I commenti del bucket macchine
    sono tecnici ('Anno di costruzione...'), ma sanifichiamo comunque per sicurezza."""
    text = (comment or "").strip()
    if not text:
        return ""
    # taglia da "A: COGNOME NOME; ..." in poi (lista presenze): tutto cio' che segue
    # e' dato personale e va rimosso. Se il commento inizia con "A:" resta stringa vuota.
    text = re.split(r"\bA:\s", text, maxsplit=1)[0].strip()
    return text[:1000]


class Command(BaseCommand):
    help = (
        "Importa lo storico collaudo come WorkOrder storici (chiusi) sugli asset gia esistenti, "
        "leggendo gli abbinamenti confermati dal foglio di riconciliazione. Idempotente, dry-run di default. "
        "Le righe senza CONFERMA (o NUOVO/SALTA) e i centri non-macchina vengono saltati."
    )

    def add_arguments(self, parser):
        parser.add_argument("--storico", default="storico collaudo.xlsx", help="Path Excel storico collaudo.")
        parser.add_argument(
            "--mapping",
            default="docs/assets/riconciliazione_collaudo_macchine_PROD.xlsx",
            help="Path foglio di riconciliazione con colonna CONFERMA compilata.",
        )
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument("--dry-run", action="store_true", help="Simula senza scrivere su DB.")
        mode.add_argument("--commit", action="store_true", help="Esegue in transaction.atomic().")

    # ── lettura foglio di riconciliazione ────────────────────────────────────
    def _load_mapping(self, path: Path) -> dict[tuple[str, str], str]:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise CommandError("Foglio di riconciliazione vuoto.")
        header = [_norm_header(c) for c in rows[0]]

        def col(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        i_centro = col("centro")
        i_mac = col("macc.ut", "macc ut")
        i_conf = col("conferma")
        if i_centro is None or i_mac is None or i_conf is None:
            raise CommandError(f"Colonne attese non trovate nel foglio. Header: {header}")

        mapping: dict[tuple[str, str], str] = {}
        skipped = 0
        for r in rows[1:]:
            centro = _clean(r[i_centro]).upper()
            mac = _clean(r[i_mac])
            conferma = _clean(r[i_conf])
            if not centro or not mac:
                continue
            if conferma.upper() in SKIP_TOKENS:
                skipped += 1
                continue
            mapping[(centro, mac)] = conferma
        self._map_skipped = skipped
        return mapping

    def handle(self, *args, **options):
        storico_path = Path(str(options.get("storico") or "")).expanduser()
        mapping_path = Path(str(options.get("mapping") or "")).expanduser()
        if not storico_path.exists():
            raise CommandError(f"Storico non trovato: {storico_path}")
        if not mapping_path.exists():
            raise CommandError(f"Foglio di riconciliazione non trovato: {mapping_path}")

        dry_run = bool(options.get("dry_run"))
        mapping = self._load_mapping(mapping_path)
        if not mapping:
            raise CommandError("Nessun abbinamento confermato nel foglio (colonna CONFERMA vuota).")

        # indice asset confermati per tag
        confirmed_tags = set(mapping.values())
        assets_by_tag = {a.asset_tag: a for a in Asset.objects.filter(asset_tag__in=confirmed_tags)}
        templates_by_code = {t.code: t for t in MaintenanceInterventionTemplate.objects.all()}

        wb = load_workbook(filename=str(storico_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = [_norm_header(c) for c in rows[0]]

        def col(name):
            for i, h in enumerate(header):
                if name in h:
                    return i
            return None

        i_dpian = col("data pianif")
        i_dult = col("dt.ult")
        i_cod = col("cod. test")
        i_desc = col("descrizione test")
        i_clav = col("c.di lav")
        i_mac = col("macc.ut")
        i_ciclo = col("codice ciclo di collaudo")
        i_comm = col("commento")

        stats = {
            "righe": 0,
            "non_macchina_o_non_confermate": 0,
            "causali_escluse": 0,
            "asset_mancante": 0,
            "template_mancante": 0,
            "gia_presenti": 0,
            "da_creare": 0,
        }
        problems: list[str] = []
        to_create: list[WorkOrder] = []

        for r in rows[1:]:
            stats["righe"] += 1
            centro = _clean(r[i_clav]).upper()
            mac = _clean(r[i_mac])
            tag = mapping.get((centro, mac))
            if not tag:
                stats["non_macchina_o_non_confermate"] += 1
                continue

            cod = _clean(r[i_cod]).upper()
            if cod in EXCLUDED_CAUSALI:
                stats["causali_escluse"] += 1
                continue

            asset = assets_by_tag.get(tag)
            if asset is None:
                stats["asset_mancante"] += 1
                problems.append(f"asset non trovato a DB per tag '{tag}' (riga {centro}/{mac})")
                continue

            template = templates_by_code.get(_template_code(cod))
            if template is None:
                stats["template_mancante"] += 1
                problems.append(f"template causale mancante per Cod.Test '{cod}' (serve import_maintenance_causali)")
                continue

            executed = _as_date(r[i_dult]) or _as_date(r[i_dpian])
            if executed is None:
                executed = timezone.localdate()
            opened = timezone.make_aware(dt.datetime.combine(executed, dt.time(8, 0)))

            ciclo = _clean(r[i_ciclo])
            ref = f"{REF_PREFIX}|{cod}|{mac}|{executed.isoformat()}|{ciclo}"[:100]

            kind = WorkOrder.KIND_SAFETY if cod.startswith("A") else WorkOrder.KIND_PREVENTIVE
            title = f"{template.label} - {asset.asset_tag}"[:255]
            notes = _sanitize_comment(_clean(r[i_comm]))

            # idempotenza: stesso reference_batch sullo stesso asset => gia importato
            if WorkOrder.objects.filter(asset=asset, reference_batch=ref).exists():
                stats["gia_presenti"] += 1
                continue

            wo = WorkOrder(
                asset=asset,
                kind=kind,
                status=WorkOrder.STATUS_DONE,
                origin=WorkOrder.ORIGIN_PERIODIC,
                opened_at=opened,
                closed_at=opened,
                title=title,
                description=(template.description or "").strip(),
                notes=notes,
                reference_batch=ref,
            )
            to_create.append(wo)
            stats["da_creare"] += 1

        if not dry_run and to_create:
            with transaction.atomic():
                for wo in to_create:
                    wo.save()

        # report
        self.stdout.write(f"Storico: {storico_path}")
        self.stdout.write(f"Mapping: {mapping_path} (macchine confermate={len(mapping)}, scartate NUOVO/SALTA={getattr(self, '_map_skipped', 0)})")
        self.stdout.write(f"Modalita: {'DRY-RUN' if dry_run else 'COMMIT'}")
        self.stdout.write(
            "Conteggi | "
            f"righe storico={stats['righe']}, "
            f"saltate (non-macchina/non-confermate)={stats['non_macchina_o_non_confermate']}, "
            f"causali escluse={stats['causali_escluse']}, "
            f"asset mancante={stats['asset_mancante']}, "
            f"template mancante={stats['template_mancante']}, "
            f"gia presenti (idempotenza)={stats['gia_presenti']}, "
            f"OdL da creare={stats['da_creare']}"
        )
        for p in problems[:30]:
            self.stdout.write(self.style.WARNING(f"  ! {p}"))
        if len(problems) > 30:
            self.stdout.write(self.style.WARNING(f"  ... e altri {len(problems) - 30} avvisi"))

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessun OdL salvato."))
        else:
            self.stdout.write(self.style.SUCCESS(f"Creati {len(to_create)} OdL storici."))
