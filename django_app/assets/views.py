from __future__ import annotations

import io
import json
import logging
import mimetypes
import os
import re
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from html import escape
from pathlib import Path, PurePosixPath
from types import SimpleNamespace
from urllib.parse import parse_qs, quote, urlencode, urlsplit, urlunsplit
from uuid import NAMESPACE_URL, uuid4, uuid5

import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.paginator import Paginator
from django.db import DatabaseError, IntegrityError, connections, transaction
from django.db.models import Avg, Case, Count, IntegerField, Max, Q, When
from django.http import FileResponse, Http404, HttpRequest, HttpResponse, HttpResponseForbidden, JsonResponse
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404, redirect, render
from django.templatetags.static import static
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.text import slugify
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph

from admin_portale.decorators import legacy_admin_required
from config.env_config import get_first_env_value, load_env_file_values, resolve_env_value, update_env_file_values
from core.acl import user_can_modulo_action
from core.public_headers import risposta_pubblica
from core.audit import log_action, storico_oggetto

# Etichetta con cui le voci di audit si agganciano all'asset: e' la stessa che
# `core.audit.storico_oggetto` usa per ripescarle sulla scheda. Costante e non
# stringa sparsa perche' un refuso qui non romperebbe nulla — semplicemente la
# voce non comparirebbe mai nello storico, e non se ne accorgerebbe nessuno.
AUDIT_OGGETTO_ASSET = "assets.asset"
from core.graph_utils import acquire_graph_token, is_placeholder_value
from core.legacy_models import AnagraficaDipendente, UtenteLegacy
from core.legacy_utils import get_legacy_user, is_legacy_admin
from core.models import AuditLog, SiteConfig, UserDashboardLayout, UserExtraInfo
from core.module_branding import (
    get_module_branding_context,
    handle_module_branding_post,
    resolve_module_logo,
)
from core.module_registry import resolve_module_label
from core.pdf import (
    PdfTheme,
    draw_canvas_footer,
    draw_canvas_header,
)
from core.table_pdf import render_table_pdf as _report_table_pdf
from core.upload_mime import UploadMimeValidationError, safe_filename, validate_extension_and_mime
from .forms import (
    AssistanceContractForm,
    AssetAdministrativeDeadlineForm,
    AssetAssignmentForm,
    AssetComponentForm,
    AssetFilterForm,
    AssetForm,
    AssetLabelTemplateForm,
    ChemicalAssetForm,
    MaintenanceChecklistStepFormSet,
    MaintenanceInterventionTemplateForm,
    MaintenanceRuleForm,
    MaintenanceRuleAssetOverrideForm,
    PeriodicVerificationForm,
    PlantLayoutForm,
    SoftwareLicenseForm,
    WorkMachineAssetForm,
    DeviceFilterForm,
    IT_DEVICE_TYPES,
    PRODUCTION_ASSET_TYPES,
    WorkMachineFilterForm,
    WorkOrderCloseForm,
    WorkOrderForm,
)
from .models import _add_months
from .models import (
    Asset,
    AssetActionButton,
    AssetAdministrativeDeadline,
    AssetAdministrativeDeadlineCompletion,
    AssetAdministrativeDeadlineCompletionAttachment,
    AssetCategory,
    AssetCategoryDocumentFolder,
    AssetCategoryField,
    AssetComponent,
    AssetCustomField,
    AssetDashboardConfig,
    AssetDetailField,
    AssetDetailSectionLayout,
    AssetDocument,
    AssetHeaderTool,
    AssetLabelTemplate,
    AssetListLayout,
    AssetListOption,
    AssetMaintenanceBudget,
    AssetMaintenanceRuleState,
    AssetMeter,
    AssetMeterHistory,
    AssetCalendarEvent,
    AssistanceContract,
    AssetReportDefinition,
    AssetReportTemplate,
    AssetSidebarButton,
    AssetTimelineEntry,
    AssetTimelineHiddenEvent,
    SoftwareLicense,
    MaintenanceInterventionTemplate,
    MaintenanceRule,
    MaintenanceRuleAssetOverride,
    PeriodicVerification,
    PlantLayout,
    PlantLayoutArea,
    PlantLayoutMarker,
    WorkMachine,
    WorkOrder,
    WorkOrderAttachment,
    WorkOrderChecklist,
    WorkOrderExecutionDay,
    WorkOrderLog,
)
from .maintenance import (
    build_workorder_prefill_payload,
    build_day_based_maintenance_schedule_rows,
    contract_state_payload,
    copy_template_checklist_to_workorder,
    get_applicable_assistance_contracts,
    get_primary_assistance_contract,
    get_workorder_overdue_days,
    normalize_workorder_source,
    preview_maintenance_rule_impact,
    resolve_asset_maintenance_rules,
    sync_workorder_maintenance_state,
    upsert_asset_maintenance_rule_state,
)
from .notifications import notify_workorder_assigned, notify_workorder_reassigned, notify_workorder_taken_over
from .services.dashboard_kpi import (
    get_asset_maintenance_costs,
    get_downtime_by_family,
    get_families_distribution,
    get_family_dashboard_kpis,
    get_fire_safety_kpis,
    get_maintenance_by_family,
    get_maintenance_kpis_for_types,
    get_maintenance_performance_kpis,
)
from .services.maintenance_kpi import build_maintenance_report_kpis
from .services.sidebar_categories import (
    category_sidebar_active_match as _service_category_sidebar_active_match,
    category_sidebar_target as _service_category_sidebar_target,
)

logger = logging.getLogger(__name__)

DEFAULT_IMPORT_SHEETS = ",".join(
    [
        "LAN A 203.0.113.x",
        "LAN B 198.51.100.x",
        "LAN C 192.0.2.x",
    ]
)
ASSET_DOCUMENT_ALLOWED_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".csv",
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".msg",
}
ASSET_DOCUMENT_ALLOWED_MIMES = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "text/csv",
    "application/csv",
    "text/plain",
    "image/jpeg",
    "image/png",
    "image/webp",
    # File messaggio Outlook (.msg): contenitore OLE Compound Document.
    # libmagic puo riportarlo con diverse etichette a seconda della versione.
    "application/vnd.ms-outlook",
    "application/x-ole-storage",
    "application/cdfv2",
}
# File di sistema da ignorare silenziosamente negli upload di intere cartelle.
ASSET_DOCUMENT_IGNORED_FILENAMES = {"thumbs.db", "desktop.ini", ".ds_store"}
ASSET_DOCUMENT_MAX_BYTES = 50 * 1024 * 1024
ASSET_DOCUMENT_UPLOAD_FIELDS = {
    AssetDocument.CATEGORY_SPECIFICHE: "upload_specs_files",
    AssetDocument.CATEGORY_MANUALI: "upload_manuals_files",
    AssetDocument.CATEGORY_INTERVENTI: "upload_interventions_files",
}
ASSET_DOCUMENT_CATEGORY_LABELS = dict(AssetDocument.CATEGORY_CHOICES)
# Prefisso del campo upload per le cartelle documento extra (oltre alle 3 di base).
ASSET_DOCUMENT_CUSTOM_FIELD_PREFIX = "upload_cat_"
REPORT_TEMPLATE_ALLOWED_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".ppt",
    ".pptx",
    ".html",
    ".htm",
}
DEFAULT_REPORT_DEFINITIONS = [
    {
        "code": AssetReportTemplate.REPORT_ASSET_DETAIL,
        "label": "Scheda asset PDF",
        "description": "Report PDF del singolo asset con riepilogo, documenti e storico.",
        "sort_order": 10,
    },
    {
        "code": AssetReportTemplate.REPORT_WORK_MACHINE_MAINTENANCE,
        "label": "Manutenzioni macchine mese",
        "description": "Report mensile delle manutenzioni pianificate per le macchine di lavoro.",
        "sort_order": 20,
    },
]

LIST_ACTIONS = {
    "create_list_option",
    "update_list_option",
    "delete_list_option",
}

BUTTON_ACTIONS = {
    "create_action_button",
    "update_action_button",
    "delete_action_button",
}

DETAIL_FIELD_ACTIONS = {
    "seed_detail_fields",
    "create_detail_field",
    "update_detail_field",
    "update_detail_field_bulk",
    "delete_detail_field",
}

DETAIL_LAYOUT_ACTIONS = {
    "update_detail_section_layout",
    "update_detail_section_layout_bulk",
    "move_detail_section_layout",
}

LIST_LAYOUT_ACTIONS = {
    "update_asset_list_layout",
    "reset_asset_list_layout",
}

CATEGORY_ACTIONS = {
    "create_asset_category",
    "update_asset_category",
    "delete_asset_category",
    "create_asset_category_field",
    "update_asset_category_field",
    "delete_asset_category_field",
    "create_sidebar_button_for_category",
}

HEADER_TOOL_ACTIONS = {"update_header_tool"}

SIDEBAR_ACTIONS = {
    "seed_sidebar_buttons",
    "create_sidebar_button",
    "update_sidebar_button",
    "delete_sidebar_button",
    "reset_sidebar_buttons",
    "clear_sidebar_buttons",
}

UI_LABEL_TRANSLATIONS = {
    "Dashboard": "Cruscotto",
    "Hardware": "Dispositivi",
    "Servers": "Server",
    "Workstations": "Postazioni di lavoro",
    "Networking": "Rete",
    "Software Licenses": "Licenze software",
    "Lifecycle Tracking": "Tracciamento ciclo di vita",
    "Compliance Reports": "Report conformita",
    "Main Navigation": "Navigazione principale",
    "Analytics & Risk": "Analisi e rischio",
    "Operations": "Operativita",
    "Print Label": "Stampa etichetta",
    "Edit Details": "Modifica dettagli",
    "Reassign": "Riassegna",
    "Log Repair": "Registra intervento",
    "Refresh Data": "Aggiorna dati",
    "Retire Asset": "Dismetti bene",
    "Manufacturer": "Produttore",
    "Model": "Modello",
    "Assignment to": "Assegnato a",
    "Assignment reparto": "Reparto assegnazione",
    "Assignment location": "Posizione assegnazione",
    "Header": "Intestazione",
    "Quick Actions": "Azioni rapide",
    "Link": "Collegamento",
    "Print": "Stampa",
    "Refresh": "Aggiorna",
    "Default": "Predefinito",
    "Primary": "Primario",
    "Secondary": "Secondario",
    "Danger": "Pericolo",
}

ASSET_LIST_BASE_COLUMN_CHOICES = [
    ("name", "Nome & Tag"),
    ("status", "Stato"),
    ("category", "Categoria"),
    ("assigned", "Assegnato a"),
    ("assignment_location", "Collocazione"),
    ("manufacturer", "Produttore"),
    ("model", "Modello"),
    ("serial_number", "Seriale"),
    ("last_seen", "Ultimo aggiornamento"),
    ("reparto", "Reparto"),
    ("vlan", "VLAN"),
    ("ip", "IP"),
]
ASSET_LIST_COMMON_COLUMNS = [
    "name",
    "status",
    "category",
    "assigned",
    "assignment_location",
    "manufacturer",
    "model",
    "serial_number",
    "last_seen",
]
ITALIAN_MONTH_NAMES = [
    "gennaio",
    "febbraio",
    "marzo",
    "aprile",
    "maggio",
    "giugno",
    "luglio",
    "agosto",
    "settembre",
    "ottobre",
    "novembre",
    "dicembre",
]
OUTLOOK_CALENDAR_TIMEZONE = "W. Europe Standard Time"
OUTLOOK_CALENDAR_START_HOUR = 8
OUTLOOK_CALENDAR_DURATION_MINUTES = 60


def _clean_string(value: str | None) -> str:
    return (value or "").strip()


def _ui_label(value: str | None) -> str:
    label = _clean_string(value)
    if not label:
        return ""
    return UI_LABEL_TRANSLATIONS.get(label, label)


def _ui_choices(raw_choices) -> list[tuple[str, str]]:
    return [(code, _ui_label(label)) for code, label in raw_choices]


def _coalesce_str(*values) -> str:
    for value in values:
        row = _clean_string(str(value)) if value is not None else ""
        if row:
            return row
    return ""


def _format_filesize(num_bytes: int | None) -> str:
    try:
        size = int(num_bytes or 0)
    except (TypeError, ValueError):
        return ""
    if size <= 0:
        return ""
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            if unit == "B":
                return f"{int(value)} {unit}"
            return f"{value:.1f} {unit}"
        value /= 1024.0
    return ""


def _default_asset_report_definition_objects() -> list[SimpleNamespace]:
    return [
        SimpleNamespace(
            id=0,
            code=str(item["code"]),
            label=str(item["label"]),
            description=str(item["description"]),
            sort_order=int(item["sort_order"]),
            is_active=True,
        )
        for item in DEFAULT_REPORT_DEFINITIONS
    ]


def _model_table_exists(model_class) -> bool:
    connection = connections[model_class.objects.db]
    try:
        with connection.cursor() as cursor:
            table_names = connection.introspection.table_names(cursor)
    except DatabaseError:
        return False
    return model_class._meta.db_table in table_names


def _active_asset_report_template(report_code: str) -> AssetReportTemplate | None:
    report_value = _clean_string(report_code)
    if not report_value or not _model_table_exists(AssetReportTemplate):
        return None
    try:
        return (
            AssetReportTemplate.objects.filter(report_code=report_value, is_active=True)
            .select_related("uploaded_by")
            .order_by("-updated_at", "-id")
            .first()
        )
    except DatabaseError:
        return None


def _ensure_default_asset_report_definitions() -> list[AssetReportDefinition | SimpleNamespace]:
    if not _model_table_exists(AssetReportDefinition):
        return _default_asset_report_definition_objects()
    try:
        existing = {row.code: row for row in AssetReportDefinition.objects.all()}
        for item in DEFAULT_REPORT_DEFINITIONS:
            code = str(item["code"])
            row = existing.get(code)
            if row is None:
                row = AssetReportDefinition.objects.create(
                    code=code,
                    label=str(item["label"]),
                    description=str(item["description"]),
                    sort_order=int(item["sort_order"]),
                    is_active=True,
                )
                existing[code] = row
                continue
            changed = False
            if not row.label:
                row.label = str(item["label"])
                changed = True
            if not row.description:
                row.description = str(item["description"])
                changed = True
            if row.sort_order != int(item["sort_order"]):
                row.sort_order = int(item["sort_order"])
                changed = True
            if changed:
                row.save(update_fields=["label", "description", "sort_order", "updated_at"])
        return list(AssetReportDefinition.objects.order_by("sort_order", "label", "id"))
    except DatabaseError:
        return _default_asset_report_definition_objects()


def _asset_report_definition_map() -> dict[str, AssetReportDefinition | SimpleNamespace]:
    return {row.code: row for row in _ensure_default_asset_report_definitions()}


def _report_templates_grouped() -> list[dict[str, object]]:
    grouped: list[dict[str, object]] = []
    templates_table_exists = _model_table_exists(AssetReportTemplate)
    for definition in _ensure_default_asset_report_definitions():
        report_code = definition.code
        if not templates_table_exists:
            rows = []
        else:
            try:
                rows = list(
                    AssetReportTemplate.objects.filter(report_code=report_code)
                    .select_related("uploaded_by")
                    .order_by("-is_active", "-updated_at", "-id")
                )
            except DatabaseError:
                rows = []
        grouped.append(
            {
                "code": report_code,
                "label": definition.label,
                "description": definition.description,
                "definition": definition,
                "active": next((row for row in rows if row.is_active), None),
                "rows": rows,
            }
        )
    return grouped

LABEL_TEMPLATE_DEFAULT_CODE = "default"
HEX_COLOR_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")
LABEL_TEMPLATE_COPY_FIELDS = [
    "page_width_mm",
    "page_height_mm",
    "qr_size_mm",
    "qr_position",
    "show_logo",
    "logo_height_mm",
    "logo_alignment",
    "title_font_size_pt",
    "body_font_size_pt",
    "show_border",
    "border_radius_mm",
    "show_field_labels",
    "show_target_label",
    "show_help_text",
    "show_target_url",
    "background_color",
    "border_color",
    "text_color",
    "accent_color",
    "title_primary_field",
    "title_secondary_field",
]


def _normalize_label_hex(value: str | None, fallback: str) -> str:
    row = _clean_string(value).upper()
    if HEX_COLOR_RE.match(row):
        return row
    return fallback


def _asset_label_field_catalog() -> list[dict[str, str]]:
    catalog = [
        {"key": "asset_tag", "label": "Tag asset", "group": "Asset"},
        {"key": "name", "label": "Nome bene", "group": "Asset"},
        {"key": "asset_category", "label": "Categoria asset", "group": "Asset"},
        {"key": "asset_type", "label": "Tipo bene", "group": "Asset"},
        {"key": "status", "label": "Stato", "group": "Asset"},
        {"key": "reparto", "label": "Reparto", "group": "Asset"},
        {"key": "manufacturer", "label": "Produttore", "group": "Asset"},
        {"key": "model", "label": "Modello", "group": "Asset"},
        {"key": "serial_number", "label": "Numero seriale", "group": "Asset"},
        {"key": "purchase_date", "label": "Data acquisto", "group": "Asset"},
        {"key": "production_date", "label": "Data produzione", "group": "Asset"},
        {"key": "assignment_to", "label": "Assegnato a", "group": "Assegnazione"},
        {"key": "assignment_reparto", "label": "Reparto assegnazione", "group": "Assegnazione"},
        {"key": "assignment_location", "label": "Posizione assegnazione", "group": "Assegnazione"},
        {"key": "year", "label": "Anno macchina", "group": "Macchina"},
        {"key": "x_mm", "label": "Corsa X", "group": "Macchina"},
        {"key": "y_mm", "label": "Corsa Y", "group": "Macchina"},
        {"key": "z_mm", "label": "Corsa Z", "group": "Macchina"},
        {"key": "diameter_mm", "label": "Diametro", "group": "Macchina"},
        {"key": "spindle_mm", "label": "Mandrino", "group": "Macchina"},
        {"key": "tmc", "label": "TMC", "group": "Macchina"},
        {"key": "tcr_enabled", "label": "TCR", "group": "Macchina"},
        {"key": "pressure_bar", "label": "Pressione", "group": "Macchina"},
        {"key": "cnc_controlled", "label": "Controllo CNC", "group": "Macchina"},
        {"key": "five_axes", "label": "5 assi", "group": "Macchina"},
        {"key": "accuracy_from", "label": "Accuracy from", "group": "Macchina"},
        {"key": "next_maintenance_date", "label": "Prossima manutenzione", "group": "Macchina"},
        {"key": "maintenance_reminder_days", "label": "Soglia reminder", "group": "Macchina"},
    ]
    for field_def in AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"):
        catalog.append(
            {
                "key": f"extra__{field_def.code}",
                "label": field_def.label,
                "group": "Campi personalizzati",
            }
        )
    return catalog


def _asset_label_field_catalog_map() -> dict[str, dict[str, str]]:
    return {row["key"]: row for row in _asset_label_field_catalog()}


def _asset_label_field_choices() -> list[tuple[str, str]]:
    return [(row["key"], f"{row['label']} [{row['group']}]") for row in _asset_label_field_catalog()]


def _asset_type_label(asset_type: str | None) -> str:
    return dict(Asset.TYPE_CHOICES).get(_clean_string(asset_type), _clean_string(asset_type) or "Generale")


def _get_default_asset_label_template() -> AssetLabelTemplate:
    template, _created = AssetLabelTemplate.objects.get_or_create(
        code=LABEL_TEMPLATE_DEFAULT_CODE,
        defaults={"scope": AssetLabelTemplate.SCOPE_DEFAULT, "asset_type": "", "name": "Etichetta predefinita"},
    )
    changed = False
    if template.scope != AssetLabelTemplate.SCOPE_DEFAULT:
        template.scope = AssetLabelTemplate.SCOPE_DEFAULT
        changed = True
    if template.asset_type:
        template.asset_type = ""
        changed = True
    if template.asset_id is not None:
        template.asset = None
        changed = True
    if changed:
        template.save()
    return template


def _default_asset_label_logo_path() -> Path:
    return Path(__file__).resolve().parents[1] / "core" / "static" / "core" / "img" / "logo_novicrom.png"


def _default_asset_label_logo_url() -> str:
    return static("core/img/logo_novicrom.png")


def _asset_label_logo_source_path(template: AssetLabelTemplate) -> str:
    if template.logo_file and getattr(template.logo_file, "path", ""):
        logo_path = Path(template.logo_file.path)
        if logo_path.exists():
            return str(logo_path)
    default_logo = _default_asset_label_logo_path()
    if default_logo.exists():
        return str(default_logo)
    return ""


def _asset_label_logo_preview_url(template: AssetLabelTemplate) -> str:
    if template.logo_file and getattr(template.logo_file, "name", ""):
        try:
            return template.logo_file.url
        except Exception:
            return ""
    return _default_asset_label_logo_url()


def _asset_label_logo_meta(template: AssetLabelTemplate) -> dict[str, str]:
    has_custom = bool(template.logo_file and getattr(template.logo_file, "name", ""))
    return {
        "url": _asset_label_logo_preview_url(template),
        "default_url": _default_asset_label_logo_url(),
        "source": "custom" if has_custom else "default",
        "name": Path(template.logo_file.name).name if has_custom else "logo_novicrom.png",
    }


def _find_asset_type_label_template(asset_type: str | None) -> AssetLabelTemplate | None:
    asset_type_value = _clean_string(asset_type)
    if not asset_type_value:
        return None
    return (
        AssetLabelTemplate.objects.filter(
            scope=AssetLabelTemplate.SCOPE_ASSET_TYPE,
            asset_type=asset_type_value,
            asset__isnull=True,
        )
        .order_by("id")
        .first()
    )


def _find_asset_override_label_template(asset: Asset | None) -> AssetLabelTemplate | None:
    if asset is None or not getattr(asset, "pk", None):
        return None
    return (
        AssetLabelTemplate.objects.filter(
            scope=AssetLabelTemplate.SCOPE_ASSET,
            asset=asset,
        )
        .order_by("id")
        .first()
    )


def _resolve_asset_label_template(asset: Asset | None = None) -> AssetLabelTemplate:
    if asset is not None:
        template = _find_asset_override_label_template(asset)
        if template is not None:
            return template
        template = _find_asset_type_label_template(asset.asset_type)
        if template is not None:
            return template
    return _get_default_asset_label_template()


def _clone_asset_label_template(
    source: AssetLabelTemplate,
    *,
    asset: Asset | None = None,
    asset_type: str = "",
) -> AssetLabelTemplate:
    template = AssetLabelTemplate()
    for field_name in LABEL_TEMPLATE_COPY_FIELDS:
        value = getattr(source, field_name)
        if field_name == "body_fields":
            value = list(value or [])
        setattr(template, field_name, value)
    template.body_fields = list(getattr(source, "body_fields", []) or [])
    if source.logo_file and getattr(source.logo_file, "name", ""):
        template.logo_file = source.logo_file.name
    template.asset = asset
    template.asset_type = _clean_string(asset_type)
    base_name = _clean_string(getattr(source, "name", "")) or "Etichetta"
    if asset is not None:
        scope_name = asset.asset_tag or asset.name or f"Asset {asset.pk}"
        template.name = f"{base_name} - {scope_name}"[:120]
    elif template.asset_type:
        template.name = f"{base_name} - {_asset_type_label(template.asset_type)}"[:120]
    else:
        template.name = base_name[:120]
    return template


def _get_asset_label_template_for_scope(
    *,
    scope: str,
    asset: Asset | None = None,
    asset_type: str = "",
) -> tuple[AssetLabelTemplate, bool]:
    if scope == AssetLabelTemplate.SCOPE_ASSET and asset is not None:
        template = _find_asset_override_label_template(asset)
        if template is not None:
            return template, True
        base_template = _find_asset_type_label_template(asset.asset_type) or _get_default_asset_label_template()
        return _clone_asset_label_template(base_template, asset=asset), False

    if scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and _clean_string(asset_type):
        template = _find_asset_type_label_template(asset_type)
        if template is not None:
            return template, True
        return _clone_asset_label_template(_get_default_asset_label_template(), asset_type=asset_type), False

    return _get_default_asset_label_template(), True


def _normalize_asset_label_fields(keys: list[str] | tuple[str, ...] | None, catalog_map: dict[str, dict[str, str]]) -> list[str]:
    cleaned: list[str] = []
    for row in keys or []:
        key = _clean_string(row)
        if key and key in catalog_map and key not in cleaned:
            cleaned.append(key)
    return cleaned


def _format_label_number(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    text = str(value)
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text.replace(".", ",")


def _format_asset_label_value(asset: Asset | None, field_key: str, catalog_map: dict[str, dict[str, str]] | None = None) -> str:
    if asset is None:
        return ""
    work_machine = getattr(asset, "work_machine", None)
    extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}

    if field_key.startswith("extra__"):
        raw_value = extra.get(field_key.replace("extra__", ""), "")
    elif field_key == "asset_category":
        raw_value = asset.category_label
    elif field_key == "asset_type":
        raw_value = asset.get_asset_type_display()
    elif field_key == "status":
        raw_value = asset.get_status_display()
    elif field_key in {
        "year",
        "x_mm",
        "y_mm",
        "z_mm",
        "diameter_mm",
        "spindle_mm",
        "tmc",
        "tcr_enabled",
        "pressure_bar",
        "cnc_controlled",
        "five_axes",
        "accuracy_from",
        "next_maintenance_date",
        "maintenance_reminder_days",
    }:
        raw_value = getattr(work_machine, field_key, "") if work_machine is not None else ""
    else:
        raw_value = getattr(asset, field_key, "")

    if raw_value in (None, ""):
        return ""
    if isinstance(raw_value, bool):
        return "Si" if raw_value else "No"
    if isinstance(raw_value, datetime):
        raw_value = raw_value.date()
    if isinstance(raw_value, date):
        return raw_value.strftime("%d-%m-%Y")

    if field_key in {"x_mm", "y_mm", "z_mm", "diameter_mm", "spindle_mm"}:
        return f"{_format_label_number(raw_value)} mm"
    if field_key == "pressure_bar":
        return f"{_format_label_number(raw_value)} bar"
    if field_key == "maintenance_reminder_days":
        return f"{_format_label_number(raw_value)} gg"
    return _clean_string(_format_label_number(raw_value) if field_key in {"year", "tmc"} else str(raw_value))


def _default_asset_label_preview_values() -> dict[str, str]:
    return {
        "asset_tag": "ML-000001",
        "name": "Centro di lavoro 5 assi",
        "asset_type": "Macchina di lavoro",
        "status": "In uso",
        "reparto": "CN5",
        "manufacturer": "DMG Mori",
        "model": "DMC 85",
        "serial_number": "DMG-550",
        "assignment_to": "Officina",
        "assignment_reparto": "CN5",
        "assignment_location": "Corsia A",
        "year": "2022",
        "x_mm": "850 mm",
        "y_mm": "700 mm",
        "z_mm": "500 mm",
        "diameter_mm": "120 mm",
        "spindle_mm": "180 mm",
        "tmc": "48",
        "tcr_enabled": "Si",
        "pressure_bar": "6,5 bar",
        "cnc_controlled": "Si",
        "five_axes": "Si",
        "accuracy_from": "0.010",
        "next_maintenance_date": (timezone.localdate() + timedelta(days=30)).strftime("%d-%m-%Y"),
        "maintenance_reminder_days": "15 gg",
    }


def _build_asset_label_preview_context(
    request: HttpRequest,
    *,
    template: AssetLabelTemplate,
    asset: Asset | None,
    target: str = "detail",
) -> dict[str, object]:
    catalog_map = _asset_label_field_catalog_map()
    field_values = _default_asset_label_preview_values()
    preview_asset_name = "Anteprima generica"
    preview_asset_tag = "Nessun asset selezionato"
    target_url = _portal_absolute_uri(request, reverse("assets:asset_list"))
    target_label = "Elenco asset"
    if asset is not None:
        preview_asset_name = asset.name or "Asset"
        preview_asset_tag = asset.asset_tag or "Asset"
        target_url, target_label = _asset_qr_target_url(request, asset, target=target)
        for key in catalog_map:
            field_values[key] = _format_asset_label_value(asset, key, catalog_map)

    selected_body_fields = _normalize_asset_label_fields(template.body_fields, catalog_map)
    title_primary_key = template.title_primary_field if template.title_primary_field in catalog_map else "asset_tag"
    title_secondary_key = template.title_secondary_field if template.title_secondary_field in catalog_map else "name"

    return {
        "catalog": [
            {
                "key": row["key"],
                "label": row["label"],
                "group": row["group"],
                "value": field_values.get(row["key"], ""),
            }
            for row in _asset_label_field_catalog()
        ],
        "catalog_map": catalog_map,
        "field_values": field_values,
        "selected_body_fields": selected_body_fields,
        "title_primary_key": title_primary_key,
        "title_secondary_key": title_secondary_key,
        "preview_asset_name": preview_asset_name,
        "preview_asset_tag": preview_asset_tag,
        "target_url": target_url,
        "target_label": target_label,
    }


def _truncate_pdf_text(pdf: canvas.Canvas, text: str, *, font_name: str, font_size: float, max_width: float) -> str:
    row = _clean_string(text)
    if not row:
        return ""
    if pdf.stringWidth(row, font_name, font_size) <= max_width:
        return row
    suffix = "..."
    while row and pdf.stringWidth(f"{row}{suffix}", font_name, font_size) > max_width:
        row = row[:-1]
    return f"{row.rstrip()}{suffix}" if row else suffix


def _month_start_from_value(raw_value: str | None, *, today: date | None = None) -> date:
    reference = today or timezone.localdate()
    month_value = _clean_string(raw_value)
    if month_value:
        try:
            return datetime.strptime(month_value, "%Y-%m").date().replace(day=1)
        except ValueError:
            pass
    return reference.replace(day=1)


def _month_end(month_start: date) -> date:
    if month_start.month == 12:
        return date(month_start.year + 1, 1, 1) - timedelta(days=1)
    return date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)


def _month_label(month_start: date) -> str:
    return f"{ITALIAN_MONTH_NAMES[month_start.month - 1]} {month_start.year}"


def _work_machine_maintenance_month_pdf_url(*, month_code: str, reparto_filter: str = "") -> str:
    params = [f"month={quote(month_code)}"]
    reparto_value = _clean_string(reparto_filter)
    if reparto_value:
        params.append(f"reparto={quote(reparto_value)}")
    return f'{reverse("assets:work_machine_maintenance_month_pdf")}?{"&".join(params)}'


PERIODIC_EXECUTION_WINDOW_CHOICES: tuple[tuple[str, str], ...] = (
    ("12", "Ultimi 12 mesi"),
    ("24", "Ultimi 24 mesi"),
    ("all", "Tutto lo storico"),
)
PERIODIC_EXECUTION_WINDOW_DEFAULT = "12"


def _normalize_periodic_execution_window(value: str | None) -> str:
    candidate = _clean_string(value).lower()
    valid = {key for key, _ in PERIODIC_EXECUTION_WINDOW_CHOICES}
    if candidate in valid:
        return candidate
    return PERIODIC_EXECUTION_WINDOW_DEFAULT


def _periodic_verifications_page_url(
    *,
    asset_id: int = 0,
    edit_id: int = 0,
    scope: str = "",
    window: str = "",
    view: str = "",
    q: str = "",
    create: bool = False,
) -> str:
    params: list[str] = []
    scope_value = _normalize_reports_scope(scope) if scope else ""
    if scope_value:
        params.append(f"scope={quote(scope_value)}")
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    if edit_id:
        params.append(f"edit={int(edit_id)}")
    if window:
        normalized_window = _normalize_periodic_execution_window(window)
        if normalized_window != PERIODIC_EXECUTION_WINDOW_DEFAULT:
            params.append(f"window={quote(normalized_window)}")
    view_value = _clean_string(view).lower()
    if view_value in {"active", "attention", "planned", "archive"} and view_value != "active":
        params.append(f"view={quote(view_value)}")
    q_value = _clean_string(q)
    if q_value:
        params.append(f"q={quote(q_value)}")
    if create:
        params.append("create=1")
    base_url = reverse("assets:periodic_verifications")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _asset_component_page_url(*, asset_id: int = 0) -> str:
    if asset_id:
        return reverse("assets:asset_component_list_for_asset", kwargs={"asset_id": int(asset_id)})
    return reverse("assets:asset_component_list")


def _asset_component_create_page_url(*, asset_id: int = 0) -> str:
    base_url = reverse("assets:asset_component_create")
    if asset_id:
        return f"{base_url}?asset={int(asset_id)}"
    return base_url


def _asset_administrative_deadline_page_url(
    *,
    asset_id: int = 0,
    component_id: int = 0,
    deadline_type: str = "",
    status: str = "",
    q: str = "",
) -> str:
    params: list[str] = []
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    if component_id:
        params.append(f"component={int(component_id)}")
    deadline_type_value = _clean_string(deadline_type).upper()
    if deadline_type_value:
        params.append(f"deadline_type={quote(deadline_type_value)}")
    status_value = _clean_string(status).lower()
    if status_value:
        params.append(f"status={quote(status_value)}")
    q_value = _clean_string(q)
    if q_value:
        params.append(f"q={quote(q_value)}")
    base_url = reverse("assets:asset_administrative_deadline_list")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _asset_administrative_deadline_create_page_url(*, asset_id: int = 0, component_id: int = 0) -> str:
    params: list[str] = []
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    if component_id:
        params.append(f"component={int(component_id)}")
    base_url = reverse("assets:asset_administrative_deadline_create")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _maintenance_settings_page_url(request: HttpRequest | None = None, *, tab: str = "catalogo") -> str:
    """URL del governo manutenzione, preservando i filtri principali."""
    tab_value = tab if tab in {"catalogo", "piani", "copertura"} else "catalogo"
    params: list[str] = [f"tab={tab_value}"]
    if request is not None:
        category_id = _as_int(request.GET.get("category"), default=0)
        if category_id:
            params.append(f"category={category_id}")
        active_value = _clean_string(request.GET.get("active")).lower()
        if active_value in {"active", "inactive", "all"}:
            params.append(f"active={quote(active_value)}")
        q = _clean_string(request.GET.get("q"))
        if q:
            params.append(f"q={quote(q)}")
    return f"{reverse('assets:maintenance_impostazioni')}?{'&'.join(params)}"


def _maintenance_template_list_page_url(*, category_id: int = 0, active: str = "") -> str:
    params: list[str] = []
    if category_id:
        params.append(f"category={int(category_id)}")
    active_value = _clean_string(active).lower()
    if active_value:
        params.append(f"active={quote(active_value)}")
    base_url = reverse("assets:maintenance_template_list")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _maintenance_rule_list_page_url(
    *,
    category_id: int = 0,
    template_id: int = 0,
    threshold_type: str = "",
    active: str = "",
) -> str:
    params: list[str] = []
    if category_id:
        params.append(f"category={int(category_id)}")
    if template_id:
        params.append(f"template={int(template_id)}")
    threshold_value = _clean_string(threshold_type).upper()
    if threshold_value:
        params.append(f"threshold_type={quote(threshold_value)}")
    active_value = _clean_string(active).lower()
    if active_value:
        params.append(f"active={quote(active_value)}")
    base_url = reverse("assets:maintenance_rule_list")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _maintenance_rule_form_state(form: MaintenanceRuleForm, *, is_edit: bool) -> dict[str, object]:
    selected_category_id = _as_int(form["asset_category"].value(), default=0)
    selected_category = None
    if selected_category_id:
        selected_category = AssetCategory.objects.filter(pk=selected_category_id).only("id", "label").first()

    category_option_count = form.fields["asset_category"].queryset.count()
    available_template_count = form.fields["intervention_template"].queryset.count()
    active_template_count = MaintenanceInterventionTemplate.objects.filter(is_active=True).count()
    general_template_count = MaintenanceInterventionTemplate.objects.filter(
        is_active=True,
        asset_category__isnull=True,
    ).count()

    no_categories_available = category_option_count == 0
    no_templates_available = active_template_count == 0
    no_templates_for_selected_category = selected_category is not None and available_template_count == 0

    alert_level = ""
    alert_message = ""
    if no_categories_available:
        alert_level = "warning"
        alert_message = "Non ci sono categorie asset attive. Prima crea almeno una categoria asset."
    elif no_templates_available:
        alert_level = "warning"
        alert_message = "Non ci sono template manutenzione attivi. Crea prima almeno un template generale o associato a una categoria."
    elif no_templates_for_selected_category:
        alert_message = (
            f'Nessun template compatibile con la categoria "{selected_category.label}". '
            "Puoi creare un template generale oppure dedicato a questa categoria."
        )

    template_list_base_url = reverse("assets:maintenance_template_list")
    template_create_base_url = reverse("assets:maintenance_template_create")
    template_list_url = _maintenance_template_list_page_url(category_id=selected_category_id)
    template_create_url = template_create_base_url
    if selected_category_id:
        template_create_url = f"{template_create_base_url}?category={selected_category_id}"

    return {
        "category_option_count": category_option_count,
        "active_template_count": active_template_count,
        "general_template_count": general_template_count,
        "available_template_count": available_template_count,
        "available_template_scope_label": (
            f'Compatibili con "{selected_category.label}"' if selected_category is not None else "Visibili nel form"
        ),
        "selected_category": selected_category,
        "alert_level": alert_level,
        "alert_message": alert_message,
        "disable_submit": not is_edit and (no_categories_available or available_template_count == 0),
        "template_list_base_url": template_list_base_url,
        "template_create_base_url": template_create_base_url,
        "template_list_url": template_list_url,
        "template_create_url": template_create_url,
        "category_admin_url": f"{reverse('assets:gestione_admin')}?tab=categorie",
    }


def _asset_maintenance_rule_list_page_url(asset_id: int, *, focus_rule_id: int = 0) -> str:
    base_url = reverse("assets:asset_maintenance_rule_list", kwargs={"asset_id": int(asset_id)})
    if focus_rule_id:
        return f"{base_url}?focus_rule={int(focus_rule_id)}#rule-{int(focus_rule_id)}"
    return base_url


def _maintenance_schedule_page_url(
    *,
    asset_id: int = 0,
    status: str = "",
    category_id: int = 0,
    reparto: str = "",
    coverage: str = "",
    q: str = "",
) -> str:
    params = []
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    status_value = _clean_string(status)
    if status_value:
        params.append(f"status={quote(status_value)}")
    if category_id:
        params.append(f"category={int(category_id)}")
    reparto_value = _clean_string(reparto)
    if reparto_value:
        params.append(f"reparto={quote(reparto_value)}")
    coverage_value = _clean_string(coverage)
    if coverage_value:
        params.append(f"coverage={quote(coverage_value)}")
    q_value = _clean_string(q)
    if q_value:
        params.append(f"q={quote(q_value)}")
    base_url = reverse("assets:maintenance_schedule")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _workorder_create_page_url(
    *,
    asset_id: int,
    rule_id: int = 0,
    periodic_verification_id: int = 0,
    source: str = "",
) -> str:
    params: list[str] = []
    if rule_id:
        params.append(f"rule={int(rule_id)}")
    if periodic_verification_id:
        params.append(f"periodic={int(periodic_verification_id)}")
    normalized_source = normalize_workorder_source(source)
    if normalized_source and normalized_source != "manual":
        params.append(f"source={quote(normalized_source)}")
    base_url = reverse("assets:wo_create", kwargs={"id": int(asset_id)})
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _assistance_contracts_page_url(
    *,
    asset_id: int = 0,
    edit_id: int = 0,
    supplier_filter: int = 0,
    state: str = "",
    scope: str = "",
    q: str = "",
) -> str:
    params = []
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    if edit_id:
        params.append(f"edit={int(edit_id)}")
    if supplier_filter:
        params.append(f"supplier={int(supplier_filter)}")
    state_value = _clean_string(state)
    if state_value:
        params.append(f"state={quote(state_value)}")
    scope_value = _clean_string(scope)
    if scope_value:
        params.append(f"scope={quote(scope_value)}")
    q_value = _clean_string(q)
    if q_value:
        params.append(f"q={quote(q_value)}")
    base_url = reverse("assets:assistance_contract_list")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _software_licenses_page_url(
    *,
    asset_id: int = 0,
    anagrafica_id: int = 0,
    edit_id: int = 0,
    category: str = "",
    status: str = "",
    assignee: str = "",
    q: str = "",
) -> str:
    params: list[str] = []
    if asset_id:
        params.append(f"asset={int(asset_id)}")
    if anagrafica_id:
        params.append(f"anagrafica={int(anagrafica_id)}")
    if edit_id:
        params.append(f"edit={int(edit_id)}")
    category_value = _clean_string(category)
    if category_value:
        params.append(f"category={quote(category_value)}")
    status_value = _clean_string(status)
    if status_value:
        params.append(f"status={quote(status_value)}")
    assignee_value = _clean_string(assignee)
    if assignee_value:
        params.append(f"assignee={quote(assignee_value)}")
    q_value = _clean_string(q)
    if q_value:
        params.append(f"q={quote(q_value)}")
    base_url = reverse("assets:software_license_list")
    return f"{base_url}?{'&'.join(params)}" if params else base_url


def _asset_maintenance_rule_override_create_page_url(*, asset_id: int, rule_id: int) -> str:
    return reverse(
        "assets:asset_maintenance_rule_override_create",
        kwargs={"asset_id": int(asset_id), "rule_id": int(rule_id)},
    )


def _asset_maintenance_rule_override_edit_page_url(*, asset_id: int, override_id: int) -> str:
    return reverse(
        "assets:asset_maintenance_rule_override_edit",
        kwargs={"asset_id": int(asset_id), "id": int(override_id)},
    )


def _asset_maintenance_rule_override_reset_page_url(*, asset_id: int, override_id: int) -> str:
    return reverse(
        "assets:asset_maintenance_rule_override_reset",
        kwargs={"asset_id": int(asset_id), "id": int(override_id)},
    )


def _asset_report_pdf_url(asset_id: int) -> str:
    return reverse("assets:asset_report_pdf", kwargs={"id": int(asset_id)})


def _asset_administrative_deadline_state(
    deadline: AssetAdministrativeDeadline,
    *,
    today: date | None = None,
) -> dict[str, object]:
    current_day = today or timezone.localdate()
    days_until_due = deadline.days_until_due(reference_date=current_day)
    if not deadline.is_active:
        return {
            "status": "inactive",
            "label": "Disattiva",
            "badge_class": "muted",
            "days_until_due": days_until_due,
            "days_label": "Monitoraggio disattivato",
        }
    if days_until_due is None:
        return {
            "status": "upcoming",
            "label": "Programmato",
            "badge_class": "ok",
            "days_until_due": None,
            "days_label": "Data non disponibile",
        }
    warning_days = max(0, int(deadline.warning_days or 0))
    if days_until_due < 0:
        overdue_days = abs(days_until_due)
        return {
            "status": "overdue",
            "label": "Scaduta",
            "badge_class": "danger",
            "days_until_due": days_until_due,
            "days_label": f"Scaduta da {overdue_days} gg",
        }
    if days_until_due <= warning_days:
        if days_until_due == 0:
            day_label = "Scade oggi"
        else:
            day_label = f"Scade tra {days_until_due} gg"
        return {
            "status": "warning",
            "label": "In scadenza",
            "badge_class": "warn",
            "days_until_due": days_until_due,
            "days_label": day_label,
        }
    return {
        "status": "upcoming",
        "label": "Programmato",
        "badge_class": "ok",
        "days_until_due": days_until_due,
        "days_label": f"Scade tra {days_until_due} gg",
    }


def _asset_status_payload(asset: Asset) -> dict[str, str]:
    if asset.status == Asset.STATUS_IN_USE:
        badge_class = "ok"
    elif asset.status == Asset.STATUS_IN_REPAIR:
        badge_class = "danger"
    else:
        badge_class = "muted"
    return {
        "label": asset.get_status_display(),
        "badge_class": badge_class,
    }


def _workorder_status_payload(status: str) -> dict[str, str]:
    if status == WorkOrder.STATUS_DONE:
        return {"label": "Chiuso", "badge_class": "ok"}
    if status == WorkOrder.STATUS_CANCELED:
        return {"label": "Annullato", "badge_class": "muted"}
    return {"label": "Aperto", "badge_class": "info"}


def _workorder_kind_payload(kind: str) -> dict[str, str]:
    mapping = {
        WorkOrder.KIND_PREVENTIVE: "ok",
        WorkOrder.KIND_CORRECTIVE: "warn",
        WorkOrder.KIND_SAFETY: "danger",
        WorkOrder.KIND_CALIBRATION: "info",
        WorkOrder.KIND_OTHER: "muted",
    }
    return {
        "label": dict(WorkOrder.KIND_CHOICES).get(kind, kind),
        "badge_class": mapping.get(kind, "muted"),
    }


def _coverage_status_payload(*, is_covered: bool, contract: AssistanceContract | None = None) -> dict[str, str]:
    if is_covered or contract is not None:
        label = "Coperto da contratto"
        if contract is not None:
            label = "Contratto collegato"
        return {"label": label, "badge_class": "info"}
    return {"label": "Copertura non disponibile", "badge_class": "muted"}


def _contextual_maintenance_suggestions(
    *,
    asset: Asset,
    schedule_row: dict[str, object] | None,
    contract: AssistanceContract | None,
    source: str,
) -> list[dict[str, str]]:
    suggestions: list[dict[str, str]] = []
    if schedule_row is not None:
        base_rule = schedule_row.get("base_rule")
        base_rule_id = getattr(base_rule, "id", 0)
        schedule_status = str(schedule_row.get("schedule_status") or "")
        if schedule_status == "overdue" and base_rule_id:
            suggestions.append(
                {
                    "label": "Crea intervento",
                    "href": _workorder_create_page_url(asset_id=asset.id, rule_id=base_rule_id, source=source),
                    "description": "La manutenzione risulta scaduta e puo essere aperta gia collegata alla regola.",
                    "style": "primary",
                }
            )
        if schedule_status == "missing" and base_rule_id:
            suggestions.append(
                {
                    "label": "Imposta prima esecuzione",
                    "href": _asset_maintenance_rule_list_page_url(asset_id=asset.id, focus_rule_id=base_rule_id),
                    "description": "Manca una baseline: registra la prima esecuzione per calcolare la prossima scadenza.",
                    "style": "default",
                }
            )
    if contract is None:
        suggestions.append(
            {
                "label": "Verifica copertura",
                "href": _assistance_contracts_page_url(asset_id=asset.id),
                "description": "L'asset non ha un contratto attivo collegato.",
                "style": "secondary",
            }
        )
    return suggestions[:3]


def _maintenance_row_primary_action(
    *,
    asset: Asset,
    base_rule: MaintenanceRule | None,
    schedule_status: str,
    source: str,
) -> dict[str, str]:
    base_rule_id = getattr(base_rule, "id", 0)
    if schedule_status == "missing" and base_rule_id:
        return {
            "label": "Imposta prima esecuzione",
            "url": _asset_maintenance_rule_list_page_url(asset_id=asset.id, focus_rule_id=base_rule_id),
        }
    return {
        "label": "Crea intervento",
        "url": _workorder_create_page_url(asset_id=asset.id, rule_id=base_rule_id, source=source),
    }


def _contract_scope_payload(contract: AssistanceContract) -> dict[str, str]:
    if contract.asset_id and contract.asset:
        return {
            "label": "Asset",
            "detail": f"{contract.asset.asset_tag} - {contract.asset.name}",
            "badge_class": "info",
        }
    if contract.asset_category_id and contract.asset_category:
        return {
            "label": "Categoria",
            "detail": f"Categoria {contract.asset_category.label}",
            "badge_class": "info",
        }
    return {
        "label": "Generale",
        "detail": "Copertura trasversale",
        "badge_class": "muted",
    }


def _software_license_state_payload(license_row: SoftwareLicense, *, today: date | None = None) -> dict[str, object]:
    current_day = today or timezone.localdate()
    expiry_date = license_row.expiry_date
    if not license_row.is_active:
        return {"status": "inactive", "label": "Disattiva", "badge_class": "muted", "days_label": "Monitoraggio spento"}
    if isinstance(expiry_date, date):
        delta_days = (expiry_date - current_day).days
        if delta_days < 0:
            return {
                "status": "expired",
                "label": "Scaduta",
                "badge_class": "danger",
                "days_label": f"Scaduta da {abs(delta_days)} gg",
            }
        if delta_days <= 30:
            return {
                "status": "expiring",
                "label": "In scadenza",
                "badge_class": "warn",
                "days_label": f"Scade tra {delta_days} gg" if delta_days else "Scade oggi",
            }
    return {"status": "active", "label": "Attiva", "badge_class": "ok", "days_label": "In uso"}


def _anagrafica_employee_options() -> tuple[list[tuple[str, str]], dict[str, dict[str, str]]]:
    try:
        rows = list(
            AnagraficaDipendente.objects.all()
            .values(
                "id",
                "nome",
                "cognome",
                "aliasusername",
                "reparto",
                "email",
                "email_notifica",
                "utente_id",
            )
        )
    except DatabaseError:
        return [], {}

    options: list[tuple[str, str]] = []
    details: dict[str, dict[str, str]] = {}
    for row in rows:
        anagrafica_id = int(row.get("id") or 0)
        if anagrafica_id <= 0:
            continue
        nome = _clean_string(row.get("nome"))
        cognome = _clean_string(row.get("cognome"))
        alias = _clean_string(row.get("aliasusername"))
        display_name = " ".join([value for value in [cognome, nome] if value]).strip() or alias
        email = _clean_string(row.get("email"))
        notification_email = _clean_string(row.get("email_notifica"))
        label_email = notification_email or email
        label = f"{display_name} - {label_email}" if label_email else display_name or f"Dipendente #{anagrafica_id}"
        options.append((str(anagrafica_id), label))
        details[str(anagrafica_id)] = {
            "display_name": display_name or f"Dipendente #{anagrafica_id}",
            "email": email,
            "notification_email": notification_email,
            "reparto": _clean_string(row.get("reparto")),
            "legacy_user_id": str(row.get("utente_id") or "").strip(),
        }
    options.sort(key=lambda item: item[1].casefold())
    return options, details


def _build_work_machine_maintenance_month_dataset(
    *,
    month_value: str | None = None,
    reparto_filter: str = "",
    today: date | None = None,
) -> dict[str, object]:
    current_day = today or timezone.localdate()
    month_start = _month_start_from_value(month_value, today=current_day)
    month_end = _month_end(month_start)
    reparto_value = _clean_string(reparto_filter)

    queryset = (
        Asset.objects.filter(
            asset_type__in=PRODUCTION_ASSET_TYPES,
            work_machine__next_maintenance_date__gte=month_start,
            work_machine__next_maintenance_date__lte=month_end,
        )
        .select_related("work_machine")
        .order_by("work_machine__next_maintenance_date", "reparto", "name", "asset_tag")
    )
    if reparto_value:
        queryset = queryset.filter(reparto=reparto_value)

    rows: list[dict[str, object]] = []
    status_counts = {"overdue": 0, "warning": 0, "ok": 0}
    for asset in queryset:
        machine = getattr(asset, "work_machine", None)
        if not isinstance(machine, WorkMachine):
            continue
        state = _work_machine_maintenance_state(machine, current_day)
        rows.append({"asset": asset, "machine": machine, "state": state})
        if state["status"] in status_counts:
            status_counts[state["status"]] += 1

    return {
        "month_start": month_start,
        "month_end": month_end,
        "month_code": month_start.strftime("%Y-%m"),
        "month_label": _month_label(month_start),
        "period_label": f'{month_start.strftime("%d-%m-%Y")} - {month_end.strftime("%d-%m-%Y")}',
        "reparto_filter": reparto_value,
        "rows": rows,
        "total_count": len(rows),
        "overdue_count": status_counts["overdue"],
        "warning_count": status_counts["warning"],
        "ok_count": status_counts["ok"],
    }


def _draw_work_machine_maintenance_month_report_header(
    pdf: canvas.Canvas,
    *,
    theme: PdfTheme,
    page_width: float,
    page_height: float,
    dataset: dict[str, object],
    generated_at: datetime,
) -> float:
    margin_x = 14 * mm
    reparto_filter = _clean_string(str(dataset.get("reparto_filter") or ""))
    subtitle = f'Periodo: {dataset["month_label"]} ({dataset["period_label"]})'
    if reparto_filter:
        subtitle = f"{subtitle} | Reparto: {reparto_filter}"
    content_y = draw_canvas_header(
        pdf,
        theme=theme,
        page_width=page_width,
        page_height=page_height,
        left_margin=margin_x,
        right_margin=margin_x,
        title="Report manutenzioni macchine",
        subtitle=subtitle,
        right_subtitle=f'Generato il {generated_at.strftime("%d-%m-%Y %H:%M")}',
    )

    cards_top = content_y
    cards_height = 16 * mm
    cards_gap = 4 * mm
    cards_width = (page_width - (2 * margin_x) - (3 * cards_gap)) / 4
    cards = [
        ("Totale mese", int(dataset.get("total_count") or 0), theme.primary),
        ("Scadute", int(dataset.get("overdue_count") or 0), "#dc2626"),
        ("In soglia", int(dataset.get("warning_count") or 0), "#d97706"),
        ("Pianificate", int(dataset.get("ok_count") or 0), "#15803d"),
    ]
    for idx, (label, value, accent) in enumerate(cards):
        left = margin_x + idx * (cards_width + cards_gap)
        pdf.setFillColor(HexColor("#ffffff"))
        pdf.setStrokeColor(theme.c_border())
        pdf.roundRect(left, cards_top - cards_height, cards_width, cards_height, 6, fill=1, stroke=1)
        pdf.setFillColor(HexColor(accent))
        pdf.rect(left, cards_top - 4, cards_width, 4, fill=1, stroke=0)
        pdf.setFillColor(theme.c_muted())
        pdf.setFont("Helvetica", 8)
        pdf.drawString(left + 10, cards_top - 15, label.upper())
        pdf.setFillColor(theme.c_text())
        pdf.setFont("Helvetica-Bold", 19)
        pdf.drawString(left + 10, cards_top - 31, str(value))

    return cards_top - cards_height - (8 * mm)


def _draw_work_machine_maintenance_month_report_footer(
    pdf: canvas.Canvas,
    *,
    theme: PdfTheme,
    page_width: float,
    margin_x: float,
    bottom_limit: float,
    page_number: int,
) -> None:
    draw_canvas_footer(
        pdf,
        theme=theme,
        page_width=page_width,
        left_margin=margin_x,
        right_margin=margin_x,
        baseline_y=bottom_limit - 10,
        page_number=page_number,
    )


def _draw_work_machine_maintenance_month_table_header(
    pdf: canvas.Canvas,
    *,
    theme: PdfTheme,
    start_x: float,
    table_y: float,
    column_defs: list[tuple[str, float]],
) -> float:
    total_width = sum(width for _, width in column_defs)
    row_height = 10 * mm
    pdf.setFillColor(theme.c_primary())
    pdf.setStrokeColor(theme.c_primary())
    pdf.rect(start_x, table_y - row_height, total_width, row_height, fill=1, stroke=1)
    pdf.setFillColor(HexColor("#ffffff"))
    pdf.setFont("Helvetica-Bold", 8)

    current_x = start_x
    for label, width in column_defs:
        pdf.drawString(current_x + 6, table_y - 18, label.upper())
        current_x += width
    return table_y - row_height


def _draw_work_machine_maintenance_month_pdf(
    pdf: canvas.Canvas,
    *,
    theme: PdfTheme,
    dataset: dict[str, object],
    generated_at: datetime,
) -> None:
    page_width, page_height = landscape(A4)
    margin_x = 14 * mm
    bottom_limit = 16 * mm
    row_height = 9 * mm
    column_defs = [
        ("Data", 24 * mm),
        ("Tag", 28 * mm),
        ("Macchina", 82 * mm),
        ("Reparto", 32 * mm),
        ("Stato", 62 * mm),
        ("Soglia", 20 * mm),
    ]
    total_width = sum(width for _, width in column_defs)
    status_colors = {
        "overdue": HexColor("#b91c1c"),
        "warning": HexColor("#a16207"),
        "ok": HexColor("#15803d"),
    }

    page_number = 1
    table_y = _draw_work_machine_maintenance_month_report_header(
        pdf,
        theme=theme,
        page_width=page_width,
        page_height=page_height,
        dataset=dataset,
        generated_at=generated_at,
    )
    table_y = _draw_work_machine_maintenance_month_table_header(
        pdf,
        theme=theme,
        start_x=margin_x,
        table_y=table_y,
        column_defs=column_defs,
    )
    current_y = table_y

    rows = list(dataset.get("rows") or [])
    if not rows:
        pdf.setFillColor(theme.c_muted())
        pdf.setFont("Helvetica", 11)
        pdf.drawString(
            margin_x,
            current_y - 22,
            "Nessuna macchina con manutenzione pianificata nel periodo selezionato.",
        )
        _draw_work_machine_maintenance_month_report_footer(
            pdf,
            theme=theme,
            page_width=page_width,
            margin_x=margin_x,
            bottom_limit=bottom_limit,
            page_number=page_number,
        )
        return

    for index, row in enumerate(rows):
        if current_y - row_height < bottom_limit:
            _draw_work_machine_maintenance_month_report_footer(
                pdf,
                theme=theme,
                page_width=page_width,
                margin_x=margin_x,
                bottom_limit=bottom_limit,
                page_number=page_number,
            )
            pdf.showPage()
            page_number += 1
            table_y = _draw_work_machine_maintenance_month_report_header(
                pdf,
                theme=theme,
                page_width=page_width,
                page_height=page_height,
                dataset=dataset,
                generated_at=generated_at,
            )
            table_y = _draw_work_machine_maintenance_month_table_header(
                pdf,
                theme=theme,
                start_x=margin_x,
                table_y=table_y,
                column_defs=column_defs,
            )
            current_y = table_y

        asset = row["asset"]
        machine = row["machine"]
        state = row["state"]
        pdf.setFillColor(HexColor("#ffffff" if index % 2 == 0 else "#fbfdff"))
        pdf.setStrokeColor(theme.c_border())
        pdf.rect(margin_x, current_y - row_height, total_width, row_height, fill=1, stroke=1)

        values = [
            machine.next_maintenance_date.strftime("%d-%m-%Y") if machine.next_maintenance_date else "-",
            _coalesce_str(asset.asset_tag, "-"),
            _coalesce_str(asset.name, "-"),
            _coalesce_str(asset.reparto, "-"),
            _coalesce_str(str(state.get("label") or ""), "-"),
            f"{int(machine.maintenance_reminder_days or 0)} gg",
        ]
        font_name = "Helvetica"
        font_size = 8.5
        text_y = current_y - 17
        current_x = margin_x
        for value_idx, value in enumerate(values):
            width = column_defs[value_idx][1]
            pdf.setFillColor(theme.c_text())
            if value_idx == 4:
                pdf.setFillColor(status_colors.get(str(state.get("status") or ""), theme.c_text()))
            pdf.setFont(font_name, font_size)
            pdf.drawString(
                current_x + 6,
                text_y,
                _truncate_pdf_text(
                    pdf,
                    value,
                    font_name=font_name,
                    font_size=font_size,
                    max_width=width - 12,
                ),
            )
            current_x += width

        current_y -= row_height

    _draw_work_machine_maintenance_month_report_footer(
        pdf,
        theme=theme,
        page_width=page_width,
        margin_x=margin_x,
        bottom_limit=bottom_limit,
        page_number=page_number,
    )


def _draw_asset_label_pdf(
    pdf: canvas.Canvas,
    *,
    asset: Asset,
    template: AssetLabelTemplate,
    target_url: str,
    target_label: str,
) -> None:
    catalog_map = _asset_label_field_catalog_map()
    body_field_keys = _normalize_asset_label_fields(template.body_fields, catalog_map)
    title_primary_key = template.title_primary_field if template.title_primary_field in catalog_map else "asset_tag"
    title_secondary_key = template.title_secondary_field if template.title_secondary_field in catalog_map else "name"

    width = float(template.page_width_mm or 100) * mm
    height = float(template.page_height_mm or 62) * mm
    margin = 6 * mm
    qr_size = float(template.qr_size_mm or 24) * mm
    qr_gap = 5 * mm
    qr_on_left = template.qr_position == AssetLabelTemplate.QR_POSITION_LEFT
    qr_x = margin if qr_on_left else max(margin, width - margin - qr_size)
    qr_y = max(margin, (height - qr_size) / 2)
    text_x = qr_x + qr_size + qr_gap if qr_on_left else margin
    text_width = width - (margin * 2) - qr_size - qr_gap

    background_color = HexColor(_normalize_label_hex(template.background_color, "#FFFFFF"))
    border_color = HexColor(_normalize_label_hex(template.border_color, "#111827"))
    text_color = HexColor(_normalize_label_hex(template.text_color, "#0F172A"))
    accent_color = HexColor(_normalize_label_hex(template.accent_color, "#1D4ED8"))

    pdf.setFillColor(background_color)
    pdf.setStrokeColor(border_color if template.show_border else background_color)
    pdf.roundRect(
        3 * mm,
        3 * mm,
        width - (6 * mm),
        height - (6 * mm),
        float(template.border_radius_mm or 0) * mm,
        stroke=1 if template.show_border else 0,
        fill=1,
    )
    _draw_pdf_qr(pdf, target_url, x=qr_x, y=qr_y, size=qr_size)

    cursor_y = height - margin - 2 * mm
    title_size = float(template.title_font_size_pt or 16)
    body_size = float(template.body_font_size_pt or 8)
    secondary_size = max(10.0, title_size - 4.0)
    line_gap = body_size + 3

    primary_text = _format_asset_label_value(asset, title_primary_key, catalog_map) or asset.asset_tag or "Asset"
    secondary_text = _format_asset_label_value(asset, title_secondary_key, catalog_map) if title_secondary_key else ""

    if template.show_logo:
        logo_path = _asset_label_logo_source_path(template)
        if logo_path:
            try:
                logo_reader = ImageReader(logo_path)
                logo_width_px, logo_height_px = logo_reader.getSize()
                target_logo_height = max(6 * mm, float(template.logo_height_mm or 10) * mm)
                target_logo_width = target_logo_height * (float(logo_width_px) / max(1.0, float(logo_height_px)))
                if target_logo_width > text_width:
                    scale_ratio = text_width / target_logo_width
                    target_logo_width = text_width
                    target_logo_height = target_logo_height * scale_ratio
                if template.logo_alignment == AssetLabelTemplate.LOGO_ALIGNMENT_CENTER:
                    logo_x = text_x + max(0, (text_width - target_logo_width) / 2)
                elif template.logo_alignment == AssetLabelTemplate.LOGO_ALIGNMENT_RIGHT:
                    logo_x = text_x + max(0, text_width - target_logo_width)
                else:
                    logo_x = text_x
                logo_y = cursor_y - target_logo_height
                pdf.drawImage(
                    logo_reader,
                    logo_x,
                    logo_y,
                    width=target_logo_width,
                    height=target_logo_height,
                    preserveAspectRatio=True,
                    mask="auto",
                )
                cursor_y = logo_y - 3 * mm
            except Exception:
                pass

    pdf.setFillColor(accent_color)
    pdf.setFont("Helvetica-Bold", title_size)
    pdf.drawString(
        text_x,
        cursor_y,
        _truncate_pdf_text(pdf, primary_text, font_name="Helvetica-Bold", font_size=title_size, max_width=text_width),
    )
    cursor_y -= title_size + 2

    if secondary_text:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", secondary_size)
        pdf.drawString(
            text_x,
            cursor_y,
            _truncate_pdf_text(pdf, secondary_text, font_name="Helvetica-Bold", font_size=secondary_size, max_width=text_width),
        )
        cursor_y -= secondary_size + 2

    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica", body_size)
    reserved_lines = 0
    if template.show_target_label:
        reserved_lines += 1
    if template.show_help_text:
        reserved_lines += 1
    if template.show_target_url:
        reserved_lines += 1
    min_bottom = margin + (reserved_lines * line_gap)

    for field_key in body_field_keys:
        if cursor_y <= min_bottom:
            break
        meta = catalog_map.get(field_key)
        value = _format_asset_label_value(asset, field_key, catalog_map)
        if not meta or not value:
            continue
        row_text = f"{meta['label']}: {value}" if template.show_field_labels else value
        pdf.drawString(
            text_x,
            cursor_y,
            _truncate_pdf_text(pdf, row_text, font_name="Helvetica", font_size=body_size, max_width=text_width),
        )
        cursor_y -= line_gap

    if template.show_target_label and cursor_y > margin:
        row_text = f"Target QR: {target_label}"
        pdf.drawString(
            text_x,
            cursor_y,
            _truncate_pdf_text(pdf, row_text, font_name="Helvetica", font_size=body_size, max_width=text_width),
        )
        cursor_y -= line_gap
    if template.show_help_text and cursor_y > margin:
        help_text = "Scansiona per aprire la scheda o la cartella."
        pdf.drawString(
            text_x,
            cursor_y,
            _truncate_pdf_text(pdf, help_text, font_name="Helvetica", font_size=body_size, max_width=text_width),
        )
        cursor_y -= line_gap
    if template.show_target_url and cursor_y > (margin - 1):
        url_size = max(6.5, body_size - 1.0)
        pdf.setFont("Helvetica", url_size)
        pdf.drawString(
            text_x,
            max(margin, cursor_y),
            _truncate_pdf_text(pdf, target_url, font_name="Helvetica", font_size=url_size, max_width=text_width),
        )


def _build_asset_report_snapshot(asset: Asset) -> dict[str, object]:
    extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}
    it_details = getattr(asset, "it_details", None)
    work_machine = getattr(asset, "work_machine", None)
    category_field_values = extra.get("_category_fields") if isinstance(extra.get("_category_fields"), dict) else {}

    summary_rows = [
        ("Tag asset", _coalesce_str(asset.asset_tag, "-")),
        ("Nome", _coalesce_str(asset.name, "-")),
        ("Categoria", _coalesce_str(asset.category_label, "-")),
        ("Tipologia", _coalesce_str(asset.get_asset_type_display(), "-")),
        ("Stato", _coalesce_str(asset.get_status_display(), "-")),
        ("PART 145", "Sì" if getattr(asset, "part_145", False) else "No"),
        ("Reparto", _coalesce_str(asset.reparto, "-")),
        ("Produttore", _coalesce_str(asset.manufacturer, "-")),
        ("Modello", _coalesce_str(asset.model, "-")),
        ("Seriale", _coalesce_str(asset.serial_number, "-")),
        ("Assegnato a", _coalesce_str(asset.assignment_to, "Non assegnato")),
        ("Posizione", _coalesce_str(asset.assignment_location, "-")),
    ]

    technical_rows: list[tuple[str, str]] = []
    if isinstance(work_machine, WorkMachine):
        technical_rows.extend(
            [
                ("Anno macchina", _coalesce_str(work_machine.year, "-")),
                ("Corsa X", _format_asset_detail_value(work_machine.x_mm, AssetDetailField.FORMAT_MM)),
                ("Corsa Y", _format_asset_detail_value(work_machine.y_mm, AssetDetailField.FORMAT_MM)),
                ("Corsa Z", _format_asset_detail_value(work_machine.z_mm, AssetDetailField.FORMAT_MM)),
                ("Diametro", _format_asset_detail_value(work_machine.diameter_mm, AssetDetailField.FORMAT_MM)),
                ("Mandrino", _format_asset_detail_value(work_machine.spindle_mm, AssetDetailField.FORMAT_MM)),
                ("TMC", _coalesce_str(work_machine.tmc, "-")),
                ("TCR", _format_asset_detail_value(work_machine.tcr_enabled, AssetDetailField.FORMAT_BOOL)),
                ("Pressione", _format_asset_detail_value(work_machine.pressure_bar, AssetDetailField.FORMAT_BAR)),
                ("CNC", _format_asset_detail_value(work_machine.cnc_controlled, AssetDetailField.FORMAT_BOOL)),
                ("5 assi", _format_asset_detail_value(work_machine.five_axes, AssetDetailField.FORMAT_BOOL)),
                ("Accuracy from", _coalesce_str(work_machine.accuracy_from, "-")),
                ("Prossima manutenzione", _format_asset_detail_value(work_machine.next_maintenance_date, AssetDetailField.FORMAT_DATE)),
            ]
        )
    elif it_details is not None:
        technical_rows.extend(
            [
                ("CPU", _coalesce_str(it_details.cpu, "-")),
                ("RAM", _coalesce_str(it_details.ram, "-")),
                ("Sistema operativo", _coalesce_str(it_details.os, "-")),
                ("Disco", _coalesce_str(it_details.disco, "-")),
            ]
        )

    category_rows: list[tuple[str, str]] = []
    if asset.asset_category_id:
        for field_def in asset.asset_category.category_fields.filter(is_active=True).order_by("sort_order", "label", "id"):
            value = _coalesce_str(category_field_values.get(field_def.code), "-")
            category_rows.append((field_def.label, value))

    custom_rows: list[tuple[str, str]] = []
    for field_def in AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"):
        value = extra.get(field_def.code, extra.get(field_def.label, ""))
        if value in ("", None, [], {}):
            continue
        custom_rows.append((field_def.label, _coalesce_str(str(value), "-")))

    workorder_rows = [
        (
            wo.opened_at.strftime("%d-%m-%Y") if wo.opened_at else "-",
            _coalesce_str(wo.get_kind_display(), "-"),
            _coalesce_str(wo.title, "-"),
            _coalesce_str(wo.get_status_display(), "-"),
        )
        for wo in asset.workorders.all().order_by("-opened_at", "-id")[:8]
    ]
    ticket_rows = [
        (
            _coalesce_str(ticket.numero_ticket, "-"),
            _coalesce_str(ticket.label_tipo, "-"),
            _coalesce_str(ticket.titolo, "-"),
            _coalesce_str(ticket.label_stato, "-"),
        )
        for ticket in asset.tickets.all().order_by("-created_at", "-id")[:8]
    ]
    document_rows = [
        (
            _coalesce_str(doc.get_category_display(), "-"),
            _coalesce_str(doc.original_name or Path(doc.file.name).name, "-"),
            doc.created_at.strftime("%d-%m-%Y") if doc.created_at else "-",
            _format_filesize(getattr(doc.file, "size", 0)),
        )
        for doc in asset.documents.all().order_by("category", "-created_at", "-id")[:12]
    ]
    periodic_rows = [
        (
            _coalesce_str(verification.name, "-"),
            _coalesce_str(
                verification.next_verification_date.strftime("%d-%m-%Y") if verification.next_verification_date else "",
                "-",
            ),
            _coalesce_str(getattr(getattr(verification, "supplier", None), "ragione_sociale", ""), "-"),
        )
        for verification in asset.periodic_verifications.all().order_by("name", "id")
    ]
    return {
        "summary_rows": summary_rows,
        "technical_rows": technical_rows,
        "category_rows": category_rows,
        "custom_rows": custom_rows,
        "workorder_rows": workorder_rows,
        "ticket_rows": ticket_rows,
        "document_rows": document_rows,
        "periodic_rows": periodic_rows,
    }


def render_asset_report_pdf(asset: Asset, snapshot: dict[str, object], *, template_name: str = "") -> bytes:
    """Report PDF della scheda asset col template standard del portale (``core.pdf``).

    Allinea la veste a tutti gli altri report (header/footer branded via
    ``header_footer_callback``, tabelle zebra col colore primario) e mette in
    evidenza il flag PART 145.
    """
    import io as _io
    from html import escape as _escape

    from reportlab.lib import colors as _colors
    from reportlab.lib.styles import ParagraphStyle as _ParagraphStyle
    from reportlab.lib.units import mm as _mm
    from reportlab.platypus import Paragraph as _Paragraph, Spacer as _Spacer, Table as _Table, TableStyle as _TableStyle

    from core.pdf import build_styles, data_table, header_footer_callback, make_document, section_heading

    theme = PdfTheme.from_branding()
    styles = build_styles(theme)
    buf = _io.BytesIO()
    doc = make_document(buf, title=f"Report asset {asset.asset_tag or ''}".strip())
    page_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin
    story: list = []

    # Evidenza PART 145 (rosso) in testa quando applicabile.
    if getattr(asset, "part_145", False):
        badge_style = _ParagraphStyle(
            "p145Badge", fontName="Helvetica-Bold", fontSize=11, textColor=_colors.white, alignment=1,
        )
        note_style = _ParagraphStyle(
            "p145Note", fontName="Helvetica-Bold", fontSize=9.5, textColor=_colors.HexColor("#7f1d1d"),
        )
        banner = _Table(
            [[
                _Paragraph("PART 145", badge_style),
                _Paragraph("Asset soggetto al regolamento aeronautico PART 145", note_style),
            ]],
            colWidths=[30 * _mm, page_w - 30 * _mm],
        )
        banner.setStyle(_TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), _colors.HexColor("#dc2626")),
            ("BACKGROUND", (1, 0), (1, 0), _colors.HexColor("#fef2f2")),
            ("BOX", (0, 0), (-1, -1), 0.8, _colors.HexColor("#dc2626")),
            ("INNERGRID", (0, 0), (-1, -1), 0.8, _colors.HexColor("#dc2626")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(banner)
        story.append(_Spacer(1, 5 * _mm))

    def kv_section(title: str, rows) -> None:
        rows = list(rows or [])
        if not rows:
            return
        story.extend(section_heading(title, theme, styles))
        table_rows = [
            [
                _Paragraph(_escape(str(key)), styles["label"]),
                _Paragraph(_escape(str(value)).replace("\n", "<br/>"), styles["value"]),
            ]
            for key, value in rows
        ]
        story.append(data_table(table_rows, theme, header=False, col_widths=[54 * _mm, page_w - 54 * _mm]))
        story.append(_Spacer(1, 3 * _mm))

    def list_section(title: str, headers, rows) -> None:
        rows = list(rows or [])
        if not rows:
            return
        story.extend(section_heading(title, theme, styles))
        col_w = page_w / max(1, len(headers))
        table_rows = [[_Paragraph(_escape(str(h)), styles["table_header"]) for h in headers]]
        for row in rows:
            table_rows.append([_Paragraph(_escape(str(cell)), styles["cell"]) for cell in row])
        story.append(data_table(table_rows, theme, col_widths=[col_w] * len(headers), repeat_rows=1))
        story.append(_Spacer(1, 3 * _mm))

    kv_section("Riepilogo asset", snapshot.get("summary_rows"))
    kv_section("Dati tecnici", snapshot.get("technical_rows"))
    kv_section("Campi categoria", snapshot.get("category_rows"))
    kv_section("Campi personalizzati", snapshot.get("custom_rows"))
    list_section("Documenti", ["Categoria", "File", "Data", "Peso"], snapshot.get("document_rows"))
    list_section("Work order recenti", ["Data", "Tipo", "Titolo", "Stato"], snapshot.get("workorder_rows"))
    list_section("Ticket collegati", ["Ticket", "Tipo", "Titolo", "Stato"], snapshot.get("ticket_rows"))
    list_section("Manutenzione periodica", ["Piano", "Prossima data", "Fornitore"], snapshot.get("periodic_rows"))

    if not story:
        story.append(_Paragraph("Nessun dato disponibile per questo asset.", styles["body"]))

    subtitle = f"{asset.asset_tag or '-'} — {asset.name or '-'}"
    if template_name:
        subtitle = f"{subtitle} | Template: {template_name}"
    draw = header_footer_callback(theme, title="REPORT ASSET", subtitle=subtitle)
    doc.build(story, onFirstPage=draw, onLaterPages=draw)
    return buf.getvalue()


def _sanitize_document_segment(value: str | None) -> str:
    row = re.sub(r'[\\/:*?"<>|#%{}~&]+', "-", _clean_string(value))
    row = re.sub(r"\s+", " ", row)
    return row.strip(" .")


def _sanitize_document_filename(value: str | None, *, fallback: str = "documento") -> str:
    filename = safe_filename(value, max_length=180)
    if not filename:
        filename = fallback
    filename = re.sub(r'["*:<>?|#%{}~&]', "-", filename)
    filename = re.sub(r"\s+", " ", filename).strip(" .")
    return filename or fallback


def _asset_document_upload_basename(value: str | None) -> str:
    normalized = _clean_string(value).replace("\\", "/")
    return PurePosixPath(normalized).name if normalized else ""


def _asset_document_relative_path_field(field_name: str) -> str:
    return f"{field_name}_relative_path"


def _asset_document_custom_field_name(slug: str) -> str:
    """Nome del campo upload HTML per una cartella documento extra."""
    return f"{ASSET_DOCUMENT_CUSTOM_FIELD_PREFIX}{slug}_files"


def _asset_document_folder_specs(asset: "Asset | None" = None) -> list[dict]:
    """Cartelle documento disponibili: le 3 di base piu le extra della AssetCategory.

    Ogni spec contiene: ``code`` (chiave salvata in ``AssetDocument.category``),
    ``label`` visibile,
    ``field`` nome del campo upload, ``removable`` (True solo per le extra) e
    ``folder_id`` (pk della ``AssetCategoryDocumentFolder`` o ``None`` per le base).
    """
    specs: list[dict] = []
    for code, label in AssetDocument.CATEGORY_CHOICES:
        specs.append(
            {
                "code": code,
                "label": label,
                "field": ASSET_DOCUMENT_UPLOAD_FIELDS[code],
                "removable": False,
                "folder_id": None,
            }
        )
    category = getattr(asset, "asset_category", None)
    if category is not None:
        for folder in category.document_folders.filter(is_active=True):
            specs.append(
                {
                    "code": folder.slug,
                    "label": folder.name,
                    "field": _asset_document_custom_field_name(folder.slug),
                    "removable": True,
                    "folder_id": folder.id,
                }
            )
    return specs


def _asset_document_category_label(specs: list[dict], code: str) -> str:
    """Etichetta visibile di una cartella documento dato il suo codice."""
    for spec in specs:
        if spec["code"] == code:
            return spec["label"]
    return ASSET_DOCUMENT_CATEGORY_LABELS.get(code, code)


def _sanitize_document_relative_path(value: str | None, *, fallback_filename: str = "") -> str:
    normalized = _normalize_document_path(value)
    if not normalized:
        return ""
    parts = [part for part in normalized.split("/") if part and part not in {".", ".."}]
    if not parts:
        return ""
    folders = [_sanitize_document_segment(part) for part in parts[:-1]]
    folders = [part for part in folders if part]
    filename = _sanitize_document_filename(parts[-1], fallback=fallback_filename or "documento")
    if not folders:
        return filename
    return "/".join([*folders, filename])


def _normalize_document_path(value: str | None) -> str:
    text = _clean_string(value).replace("\\", "/")
    while "//" in text:
        text = text.replace("//", "/")
    return text.strip("/")


def _asset_qr_target_url(request: HttpRequest, asset: Asset, *, target: str = "detail") -> tuple[str, str]:
    desired = _clean_string(target).lower()
    public_token = _clean_string(getattr(asset, "public_qr_token", ""))
    public_qr_enabled = bool(public_token) and getattr(asset, "public_qr_enabled", True)
    # "sharepoint" e' un valore legacy (link e template etichetta salvati prima
    # della rimozione dell'archivio SharePoint): vale come "landing".
    if desired in {"sharepoint", "landing"}:
        # Landing QR: pubblica (token opaco) quando disponibile, così il QR resta
        # leggibile da tecnici/ispettori esterni senza login.
        if public_qr_enabled:
            public_landing = reverse("assets:asset_qr_public_landing", kwargs={"public_qr_token": public_token})
            return _portal_absolute_uri(request, public_landing), "Landing QR pubblica"
        landing_url = reverse("assets:asset_qr_landing", kwargs={"asset_tag": asset.asset_tag})
        return _portal_absolute_uri(request, landing_url), "Landing mobile QR"
    detail_url = reverse("assets:asset_view", kwargs={"id": asset.id})
    return _portal_absolute_uri(request, detail_url), "Scheda asset"


def _portal_absolute_uri(request: HttpRequest, path: str) -> str:
    site_url = _clean_string(getattr(settings, "SITE_URL", "")).rstrip("/")
    if site_url:
        normalized_path = _clean_string(path)
        if not normalized_path.startswith("/"):
            normalized_path = f"/{normalized_path}"
        return f"{site_url}{normalized_path}"
    return request.build_absolute_uri(path)


def _draw_pdf_qr(pdf: canvas.Canvas, value: str, *, x: float, y: float, size: float) -> None:
    widget = qr.QrCodeWidget(value)
    bounds = widget.getBounds()
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    drawing = Drawing(size, size, transform=[size / width, 0, 0, size / height, 0, 0])
    drawing.add(widget)
    renderPDF.draw(drawing, pdf, x, y)


def _shorten_text(value: str, limit: int = 56) -> str:
    text = _clean_string(value)
    if len(text) <= limit:
        return text
    return f"{text[: max(0, limit - 1)].rstrip()}…"


def _build_asset_documents_by_category(asset: Asset) -> tuple[dict[str, str], dict[str, list[dict]]]:
    """Costruisce i documenti asset raggruppati per categoria e cartella.

    Per ogni categoria ritorna una lista di gruppi ``{"folder": ..., "documents": [...]}``;
    il gruppo con ``folder`` vuoto raccoglie i file singoli (loose), gli altri
    raggruppano i file caricati con "Carica cartella" mantenendo la cartella visibile.
    """
    flat: dict[str, list[dict]] = defaultdict(list)
    extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}
    raw_docs = extra.get("documents")
    if isinstance(raw_docs, list):
        for row in raw_docs:
            if not isinstance(row, dict):
                continue
            name = _coalesce_str(row.get("name"), row.get("filename"))
            if not name:
                continue
            category = _coalesce_str(row.get("category"), AssetDocument.CATEGORY_SPECIFICHE).upper()
            if category not in ASSET_DOCUMENT_CATEGORY_LABELS:
                category = AssetDocument.CATEGORY_SPECIFICHE
            flat[category].append(
                {
                    "name": name,
                    "size": _coalesce_str(row.get("size"), ""),
                    "date": _coalesce_str(row.get("date"), ""),
                    "url": _coalesce_str(row.get("url"), ""),
                    "kind": "external",
                    "meta": "",
                    "folder": "",
                }
            )

    for uploaded in asset.documents.all():
        size_text = ""
        try:
            size_text = _format_filesize(uploaded.file.size)
        except Exception:
            size_text = ""
        meta_parts = []
        if _clean_string(uploaded.notes):
            meta_parts.append(_clean_string(uploaded.notes))
        flat[uploaded.category].append(
            {
                "id": uploaded.id,
                "name": uploaded.original_name or Path(uploaded.file.name).name,
                "size": size_text,
                "date": uploaded.document_date.strftime("%d-%m-%Y") if uploaded.document_date else uploaded.created_at.strftime("%d-%m-%Y"),
                "url": reverse("assets:asset_document_download", args=[uploaded.id]),
                "kind": "uploaded",
                "meta": " | ".join(meta_parts),
                "folder": _normalize_document_path(uploaded.relative_folder),
            }
        )

    # Categorie da mostrare: le 3 di base + le extra della AssetCategory, piu
    # eventuali codici "orfani" presenti solo su documenti gia esistenti.
    specs = _asset_document_folder_specs(asset)
    category_labels: dict[str, str] = {spec["code"]: spec["label"] for spec in specs}
    for code in flat:
        category_labels.setdefault(code, ASSET_DOCUMENT_CATEGORY_LABELS.get(code, code))

    documents_by_category: dict[str, list[dict]] = {}
    for category in category_labels:
        groups_map: dict[str, list[dict]] = {}
        for doc in flat.get(category, []):
            groups_map.setdefault(doc.get("folder", ""), []).append(doc)
        groups: list[dict] = []
        if groups_map.get(""):
            groups.append({"folder": "", "documents": groups_map.pop("")})
        else:
            groups_map.pop("", None)
        for folder in sorted(groups_map.keys(), key=str.lower):
            groups.append({"folder": folder, "documents": groups_map[folder]})
        documents_by_category[category] = groups
    return category_labels, documents_by_category


def _build_uploaded_documents_context(asset: Asset | None) -> dict[str, list[AssetDocument]]:
    grouped: dict[str, list[AssetDocument]] = {
        spec["code"]: [] for spec in _asset_document_folder_specs(asset)
    }
    if not asset or not asset.pk:
        return grouped
    for document in asset.documents.all():
        grouped.setdefault(document.category, []).append(document)
    return grouped


# Content-type per cui l'apertura inline e' sicura: nessuno di questi puo'
# eseguire script nell'origine del portale. Tutto il resto (html, svg, ...) va
# servito come allegato, altrimenti un upload diventa una XSS same-origin.
_INLINE_SAFE_CONTENT_TYPES = frozenset(
    {
        "application/pdf",
        "image/png",
        "image/jpeg",
        "image/gif",
        "image/webp",
        "image/bmp",
        "text/plain",
    }
)


def _document_file_response(storage, file_name: str, filename: str) -> FileResponse:
    """FileResponse con Content-Disposition inline solo per i tipi sicuri."""
    content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    inline = content_type in _INLINE_SAFE_CONTENT_TYPES
    response = FileResponse(
        storage.open(file_name, "rb"),
        as_attachment=not inline,
        filename=filename,
        content_type=content_type,
    )
    response["X-Content-Type-Options"] = "nosniff"
    return response


def _can_download_asset_document(request: HttpRequest) -> bool:
    """Chi puo' scaricare un documento asset da sessione autenticata.

    Policy: **qualunque utente autenticato**. I documenti asset (manuali,
    specifiche, procedure di intervento) sono per scelta di prodotto leggibili
    da chiunque abbia accesso fisico alla macchina, via il token QR stampato
    sopra (``asset_document_qr_download``, senza login). Il QR e' quindi il
    *pavimento* dell'accessibilita', non il tetto: un utente autenticato non puo'
    vedere meno di un anonimo con il QR in mano.

    Cio' che rendeva pericoloso il vecchio ``@login_required`` nudo non era
    questa apertura, ma il fatto che gli stessi file fossero **anche** serviti da
    IIS su ``/media/assets_documents/`` in anonimo, senza alcun controllo ne'
    traccia. Chiuso quello (deny IIS) e passato tutto da view con audit, l'accesso
    autenticato universale e' una decisione consapevole, non un IDOR.

    Il tetto resta comunque l'ACL di modulo: questa rotta sta sotto ``/assets/``,
    quindi ``ACLMiddleware`` la gatea con il binding canonico del modulo — chi non
    ha accesso ad assets non arriva nemmeno qui. Restringere la policy significa
    toccare solo questo predicato (il diniego e' gia' cablato e auditato).
    """
    return bool(getattr(request.user, "is_authenticated", False))


@login_required
def asset_document_download(request, document_id: int):
    document = get_object_or_404(
        AssetDocument.objects.select_related("asset", "uploaded_by"),
        pk=document_id,
    )
    if not _can_download_asset_document(request):
        log_action(
            request,
            "download_asset_document",
            "assets",
            {
                "document_id": document.id,
                "asset_id": document.asset_id,
                "category": document.category,
                "esito": "denied",
                "motivo": "permission_denied",
            },
        )
        return render(request, "core/pages/forbidden.html", status=403)
    storage = document.file.storage if document.file else None
    file_name = document.file.name if document.file else ""
    if not storage or not file_name or not storage.exists(file_name):
        log_action(
            request,
            "download_asset_document",
            "assets",
            {
                "document_id": document.id,
                "asset_id": document.asset_id,
                "category": document.category,
                "esito": "not_found",
            },
        )
        return HttpResponse("Documento non trovato.", status=404)
    filename = document.original_name or Path(file_name).name
    log_action(
        request,
        "download_asset_document",
        "assets",
        {
            "document_id": document.id,
            "asset_id": document.asset_id,
            "category": document.category,
            "filename": filename,
            "esito": "success",
        },
    )
    # Stessa resa del QR: anteprima nel browser per i tipi che non possono
    # eseguire script (era il comportamento con l'URL /media/ diretto), allegato
    # per tutti gli altri.
    return _document_file_response(storage, file_name, filename)


@risposta_pubblica
def asset_document_qr_download(request: HttpRequest, public_qr_token: str, document_id: int):
    """Download di un documento asset tramite il token QR stampato sulla macchina.

    Il token e' la chiave d'accesso: i file NON sono raggiungibili via /media/
    (deny IIS su media/assets_documents). Nessun filtro di categoria — dal QR si
    vedono tutti i documenti dell'asset — ma il documento deve appartenere
    esattamente all'asset di quel token: nessun accesso cross-asset.
    """
    token = _clean_string(public_qr_token)
    if not token:
        raise Http404("Link non disponibile.")
    asset = Asset.objects.filter(public_qr_token=token, public_qr_enabled=True).first()
    if asset is None:
        raise Http404("Link non disponibile.")

    document = AssetDocument.objects.filter(pk=document_id, asset_id=asset.id).first()
    if document is None:
        # Documento inesistente o di un altro asset: stessa risposta, nessun oracolo.
        log_action(
            request,
            "download_asset_document_qr",
            "assets",
            {
                "document_id": document_id,
                "asset_id": asset.id,
                "asset_tag": asset.asset_tag,
                "qr_token_prefix": token[:8],
                "esito": "denied",
                "motivo": "document_not_in_asset",
            },
        )
        raise Http404("Documento non disponibile.")

    storage = document.file.storage if document.file else None
    file_name = document.file.name if document.file else ""
    if not storage or not file_name or not storage.exists(file_name):
        log_action(
            request,
            "download_asset_document_qr",
            "assets",
            {
                "document_id": document.id,
                "asset_id": asset.id,
                "asset_tag": asset.asset_tag,
                "qr_token_prefix": token[:8],
                "esito": "not_found",
            },
        )
        raise Http404("Documento non disponibile.")

    filename = document.original_name or Path(file_name).name
    log_action(
        request,
        "download_asset_document_qr",
        "assets",
        {
            "document_id": document.id,
            "asset_id": asset.id,
            "asset_tag": asset.asset_tag,
            "category": document.category,
            "filename": filename,
            "qr_token_prefix": token[:8],
            "esito": "success",
        },
    )
    return _document_file_response(storage, file_name, filename)


@login_required
def workorder_attachment_download(request: HttpRequest, attachment_id: int):
    """Allegati OdL: serviti solo da qui (deny IIS su media/assets_workorders)."""
    attachment = get_object_or_404(
        WorkOrderAttachment.objects.select_related("work_order__asset"),
        pk=attachment_id,
    )
    storage = attachment.file.storage if attachment.file else None
    file_name = attachment.file.name if attachment.file else ""
    if not storage or not file_name or not storage.exists(file_name):
        log_action(
            request,
            "download_workorder_attachment",
            "assets",
            {
                "attachment_id": attachment.id,
                "work_order_id": attachment.work_order_id,
                "asset_id": attachment.work_order.asset_id,
                "esito": "not_found",
            },
        )
        return HttpResponse("Allegato non trovato.", status=404)
    filename = attachment.original_name or Path(file_name).name
    log_action(
        request,
        "download_workorder_attachment",
        "assets",
        {
            "attachment_id": attachment.id,
            "work_order_id": attachment.work_order_id,
            "asset_id": attachment.work_order.asset_id,
            "filename": filename,
            "esito": "success",
        },
    )
    return _document_file_response(storage, file_name, filename)


def _validate_asset_document_uploads(
    request: HttpRequest, asset: Asset | None = None
) -> tuple[dict[str, list], list[str]]:
    uploads: dict[str, list] = {}
    errors: list[str] = []
    for spec in _asset_document_folder_specs(asset):
        category = spec["code"]
        field_name = spec["field"]
        valid_files = []
        relative_paths = request.POST.getlist(_asset_document_relative_path_field(field_name))
        for index, upload in enumerate(request.FILES.getlist(field_name)):
            filename = getattr(upload, "name", "") or ""
            basename = _asset_document_upload_basename(filename)
            if not basename:
                continue
            # Upload di intere cartelle: ignora i file di sistema senza errore.
            if basename.lower() in ASSET_DOCUMENT_IGNORED_FILENAMES:
                continue
            relative_path = relative_paths[index] if index < len(relative_paths) else ""
            try:
                validate_extension_and_mime(
                    upload,
                    allowed_extensions=ASSET_DOCUMENT_ALLOWED_EXTENSIONS,
                    allowed_mimes=ASSET_DOCUMENT_ALLOWED_MIMES,
                    max_bytes=ASSET_DOCUMENT_MAX_BYTES,
                    label=relative_path or basename,
                    allow_empty=False,
                )
            except UploadMimeValidationError as exc:
                errors.append(str(exc))
                continue
            upload.name = _sanitize_document_filename(basename)
            upload._document_relative_path = _sanitize_document_relative_path(
                relative_path,
                fallback_filename=upload.name,
            )
            valid_files.append(upload)
        uploads[category] = valid_files
    return uploads, errors


def _asset_document_upload_count(uploads: dict[str, list]) -> int:
    return sum(len(files) for files in uploads.values())


def _apply_asset_document_changes(
    asset: Asset,
    *,
    uploads: dict[str, list],
    remove_ids: set[int],
    actor,
) -> None:
    """Applica rimozioni e nuovi upload sull'archivio documenti locale dell'asset."""
    if remove_ids:
        for document in asset.documents.filter(id__in=remove_ids):
            document.delete()

    for category, files in uploads.items():
        for upload in files:
            relative_path = getattr(upload, "_document_relative_path", "")
            relative_folder = "/".join(relative_path.split("/")[:-1]) if "/" in relative_path else ""
            AssetDocument.objects.create(
                asset=asset,
                category=category,
                file=upload,
                original_name=_sanitize_document_filename(getattr(upload, "name", ""))[:255],
                relative_folder=relative_folder[:400],
                uploaded_by=actor if getattr(actor, "is_authenticated", False) else None,
            )


def _work_machine_maintenance_state(machine: WorkMachine, today) -> dict[str, object]:
    next_date = getattr(machine, "next_maintenance_date", None)
    reminder_days = int(getattr(machine, "maintenance_reminder_days", 30) or 0)
    if not next_date:
        return {"status": "missing", "label": "Da pianificare", "days": None, "date": None}
    delta_days = (next_date - today).days
    if delta_days < 0:
        return {"status": "overdue", "label": f"Scaduta da {abs(delta_days)} gg", "days": delta_days, "date": next_date}
    if delta_days <= reminder_days:
        return {"status": "warning", "label": f"In soglia ({delta_days} gg)", "days": delta_days, "date": next_date}
    return {"status": "ok", "label": f"Pianificata tra {delta_days} gg", "days": delta_days, "date": next_date}


def _periodic_verification_state(verification: PeriodicVerification, today=None) -> dict[str, object]:
    current_day = today or timezone.localdate()
    if not verification.is_active:
        return {"status": "inactive", "label": "Disattivata", "date": verification.next_verification_date}
    next_date = verification.next_verification_date
    if not next_date:
        return {"status": "missing", "label": "Da pianificare", "date": None}
    delta_days = (next_date - current_day).days
    if delta_days < 0:
        return {"status": "overdue", "label": f"Scaduta da {abs(delta_days)} gg", "date": next_date}
    if delta_days <= 30:
        return {"status": "warning", "label": f"In scadenza ({delta_days} gg)", "date": next_date}
    return {"status": "ok", "label": f"Pianificata tra {delta_days} gg", "date": next_date}


def _periodic_execution_window_cutoff(window: str, today=None) -> "date | None":
    current_day = today or timezone.localdate()
    normalized = _normalize_periodic_execution_window(window)
    if normalized == "all":
        return None
    months = int(normalized)
    cutoff = current_day - timedelta(days=months * 31)
    return cutoff


def _parse_execution_cost_input(raw_value: str | None) -> "Decimal | None":
    cleaned = _clean_string(raw_value)
    if not cleaned:
        return None
    try:
        value = Decimal(cleaned.replace(",", "."))
    except (ArithmeticError, ValueError) as exc:
        raise ValueError("invalid cost") from exc
    if value < 0:
        raise ValueError("negative cost")
    return value


def _build_execution_workorder(
    *,
    asset: Asset,
    title: str,
    description: str,
    executed_on: date,
    duration_minutes: int,
    cost_value: "Decimal | None",
    resolution_text: str,
    periodic_verification: PeriodicVerification | None = None,
    maintenance_rule: MaintenanceRule | None = None,
    supplier=None,
) -> WorkOrder:
    executed_at_dt = timezone.make_aware(
        datetime.combine(executed_on, datetime.min.time().replace(hour=12)),
        timezone.get_current_timezone(),
    )
    workorder = WorkOrder.objects.create(
        asset=asset,
        periodic_verification=periodic_verification,
        maintenance_rule=maintenance_rule,
        supplier=supplier,
        kind=WorkOrder.KIND_PREVENTIVE,
        status=WorkOrder.STATUS_OPEN,
        opened_at=executed_at_dt,
        title=title[:255],
        description=description,
    )
    workorder.closed_at = executed_at_dt
    workorder.status = WorkOrder.STATUS_DONE
    if resolution_text:
        workorder.resolution = resolution_text
    workorder.intervention_duration_minutes = max(0, int(duration_minutes or 0))
    if cost_value is not None:
        workorder.cost_eur = cost_value
    workorder.full_clean()
    workorder.save()
    return workorder


def _serialize_execution_workorder(workorder: WorkOrder) -> dict[str, object]:
    return {
        "workorder": workorder,
        "asset": workorder.asset,
        "asset_label": f"{workorder.asset.asset_tag} - {workorder.asset.name}",
        "supplier_label": str(workorder.supplier) if workorder.supplier else "",
        "executed_at": workorder.closed_at,
        "duration_minutes": workorder.intervention_duration_minutes,
        "downtime_minutes": workorder.downtime_minutes,
        "cost_eur": workorder.resolved_total_cost_eur,
        "resolution": workorder.resolution,
        "title": workorder.title,
        "url": reverse("assets:wo_view", kwargs={"id": workorder.id}),
    }


def _periodic_execution_rows_for_verification(
    *,
    verification_id: int,
    asset_id: int = 0,
    cutoff_date: "date | None" = None,
    limit: int = 25,
) -> list[dict[str, object]]:
    qs = (
        WorkOrder.objects.select_related("asset", "supplier")
        .filter(periodic_verification_id=verification_id, status=WorkOrder.STATUS_DONE)
    )
    if asset_id:
        qs = qs.filter(asset_id=int(asset_id))
    if cutoff_date is not None:
        qs = qs.filter(closed_at__date__gte=cutoff_date)
    qs = qs.order_by("-closed_at", "-id")[: max(1, int(limit))]
    return [_serialize_execution_workorder(workorder) for workorder in qs]


def _maintenance_rule_execution_rows(
    *,
    asset_id: int,
    base_rule_id: int,
    cutoff_date: "date | None" = None,
    limit: int = 10,
) -> list[dict[str, object]]:
    qs = (
        WorkOrder.objects.select_related("asset", "supplier")
        .filter(asset_id=int(asset_id), maintenance_rule_id=int(base_rule_id), status=WorkOrder.STATUS_DONE)
    )
    if cutoff_date is not None:
        qs = qs.filter(closed_at__date__gte=cutoff_date)
    qs = qs.order_by("-closed_at", "-id")[: max(1, int(limit))]
    return [_serialize_execution_workorder(workorder) for workorder in qs]


def _deadline_completion_rows(
    *,
    deadline_id: int,
    cutoff_date: "date | None" = None,
    limit: int = 10,
) -> list[dict[str, object]]:
    qs = (
        AssetAdministrativeDeadlineCompletion.objects.select_related("completed_by")
        .prefetch_related("attachments")
        .filter(deadline_id=int(deadline_id))
    )
    if cutoff_date is not None:
        qs = qs.filter(completed_on__gte=cutoff_date)
    qs = qs.order_by("-completed_on", "-id")[: max(1, int(limit))]
    rows: list[dict[str, object]] = []
    for completion in qs:
        attachment_rows: list[dict[str, object]] = []
        for attachment in completion.attachments.all():
            attachment_rows.append(
                {
                    "id": attachment.id,
                    "name": attachment.original_name or Path(attachment.file.name).name,
                    "url": reverse("assets:admin_deadline_attachment_download", args=[attachment.id]),
                }
            )
        rows.append(
            {
                "completion": completion,
                "completed_on": completion.completed_on,
                "completed_by": completion.completed_by,
                "completed_by_label": (
                    completion.completed_by.get_full_name() or completion.completed_by.username
                    if completion.completed_by_id
                    else ""
                ),
                "duration_minutes": completion.duration_minutes,
                "cost_eur": completion.cost_eur,
                "notes": completion.notes,
                "next_due_date": completion.next_due_date,
                "attachment_rows": attachment_rows,
            }
        )
    return rows


@login_required
def admin_deadline_attachment_download(request, attachment_id: int):
    attachment = get_object_or_404(
        AssetAdministrativeDeadlineCompletionAttachment.objects.select_related("completion__deadline__asset"),
        pk=attachment_id,
    )
    if not _is_assets_admin(request):
        log_action(
            request,
            "download_admin_deadline_attachment",
            "assets",
            {
                "attachment_id": attachment.id,
                "completion_id": attachment.completion_id,
                "deadline_id": attachment.completion.deadline_id,
                "asset_id": attachment.completion.deadline.asset_id,
                "esito": "denied",
                "motivo": "permission_denied",
            },
        )
        return render(request, "core/pages/forbidden.html", status=403)
    storage = attachment.file.storage
    if not attachment.file or not attachment.file.name or not storage.exists(attachment.file.name):
        log_action(
            request,
            "download_admin_deadline_attachment",
            "assets",
            {
                "attachment_id": attachment.id,
                "completion_id": attachment.completion_id,
                "deadline_id": attachment.completion.deadline_id,
                "asset_id": attachment.completion.deadline.asset_id,
                "esito": "not_found",
            },
        )
        return HttpResponse("Allegato non trovato.", status=404)
    filename = attachment.original_name or Path(attachment.file.name).name
    content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    log_action(
        request,
        "download_admin_deadline_attachment",
        "assets",
        {
            "attachment_id": attachment.id,
            "completion_id": attachment.completion_id,
            "deadline_id": attachment.completion.deadline_id,
            "asset_id": attachment.completion.deadline.asset_id,
            "filename": filename,
            "esito": "success",
        },
    )
    return FileResponse(
        storage.open(attachment.file.name, "rb"),
        as_attachment=True,
        filename=filename,
        content_type=content_type,
    )


def _compute_ticket_kpi_for_asset(asset: Asset) -> dict:
    """Calcola KPI ticket per il singolo asset.

    Restituisce un dizionario pronto per il template. Non genera N+1:
    carica tutti i ticket dell'asset in una query e tutti gli interventi
    in una seconda query, poi elabora in Python.
    """
    from tickets.models import StatoTicket, TicketIntervento

    tickets_qs = list(
        asset.tickets
        .only(
            "pk", "stato", "tipo", "categoria", "componente", "causa_radice",
            "tipo_fermo", "ore_fermo_macchina",
            "created_at", "closed_at", "data_presa_in_carico",
            "risolto_da_nome",
        )
    )
    if not tickets_qs:
        return {"has_tickets": False}

    stati_aperti = {StatoTicket.APERTA, StatoTicket.IN_CARICO, StatoTicket.IN_ATTESA}
    stati_chiusi = {StatoTicket.RISOLTO, StatoTicket.CHIUSO}

    totale = len(tickets_qs)
    aperti = sum(1 for t in tickets_qs if t.stato in stati_aperti)
    chiusi = sum(1 for t in tickets_qs if t.stato in stati_chiusi)

    # MTTR — media ore dalla creazione alla chiusura per ticket risolti/chiusi con closed_at
    durate_risoluzione = []
    for t in tickets_qs:
        if t.stato in stati_chiusi and t.closed_at and t.created_at:
            delta_h = (t.closed_at - t.created_at).total_seconds() / 3600
            durate_risoluzione.append(delta_h)
    mttr_ore = round(sum(durate_risoluzione) / len(durate_risoluzione), 1) if durate_risoluzione else None

    # Ore fermo macchina totali
    ore_fermo_totali = sum(
        float(t.ore_fermo_macchina)
        for t in tickets_qs
        if t.ore_fermo_macchina is not None
    )

    # Top 3 componenti per frequenza
    from collections import Counter
    componenti_counter = Counter(
        t.componente for t in tickets_qs if t.componente
    )
    top_componenti = componenti_counter.most_common(3)

    # Top 3 cause radice per frequenza
    cause_counter = Counter(
        t.causa_radice for t in tickets_qs if t.causa_radice
    )
    top_cause = cause_counter.most_common(3)

    # Interventi — ore per tecnico
    ticket_ids = [t.pk for t in tickets_qs]
    interventi_qs = list(
        TicketIntervento.objects
        .filter(ticket_id__in=ticket_ids)
        .only("tecnico_nome", "ore_lavorate", "data_inizio", "data_fine")
    )
    ore_per_tecnico: dict[str, float] = {}
    for interv in interventi_qs:
        ore = interv.durata_ore
        if ore:
            ore_per_tecnico[interv.tecnico_nome] = ore_per_tecnico.get(interv.tecnico_nome, 0.0) + ore
    top_tecnici = sorted(ore_per_tecnico.items(), key=lambda x: -x[1])[:3]
    ore_intervento_totali = round(sum(ore_per_tecnico.values()), 2)

    return {
        "has_tickets": True,
        "totale": totale,
        "aperti": aperti,
        "chiusi": chiusi,
        "mttr_ore": mttr_ore,
        "ore_fermo_totali": round(ore_fermo_totali, 2) if ore_fermo_totali else 0,
        "ore_intervento_totali": ore_intervento_totali,
        "top_componenti": top_componenti,
        "top_cause": top_cause,
        "top_tecnici": top_tecnici,
    }


def _build_asset_status_band(
    *,
    primary_contract: AssistanceContract | None,
    primary_contract_state: dict[str, str] | None,
    next_deadline_row: dict[str, object] | None,
    asset_assistance_contracts_url: str,
    asset_administrative_deadline_list_url: str,
) -> dict[str, dict[str, str]]:
    """Banda di stato del dettaglio asset: copertura assistenza + prossima scadenza
    amministrativa unite in un'unica card. Lo stato asset resta nella pill header e
    la prossima manutenzione viene mostrata nel Registro manutenzione."""
    contract_value = str(primary_contract.supplier) if primary_contract is not None else "Copertura non disponibile"
    contract_meta = "Copertura non disponibile"
    if primary_contract is not None:
        contract_meta = primary_contract.coverage_summary or primary_contract.target_label
    coverage = {
        "label": "Copertura assistenza",
        "value": contract_value,
        "meta": contract_meta or "Nessun contratto attivo collegato.",
        "badge_label": (
            str(primary_contract_state.get("label"))
            if primary_contract_state is not None
            else "Senza contratto"
        ),
        "badge_class": (
            str(primary_contract_state.get("badge_class"))
            if primary_contract_state is not None
            else "muted"
        ),
        "link_label": "Apri contratti",
        "link_url": asset_assistance_contracts_url,
    }
    deadline_value = "Nessuna scadenza"
    deadline_meta = "Nessuna scadenza amministrativa registrata."
    deadline_badge_class = "muted"
    deadline_badge_label = "Nessuna"
    if next_deadline_row is not None:
        deadline = next_deadline_row["deadline"]
        deadline_state = next_deadline_row["state"]
        deadline_value = str(deadline.title)
        deadline_meta = str(deadline_state.get("days_label") or "")
        deadline_badge_class = str(deadline_state.get("badge_class") or "muted")
        deadline_badge_label = str(deadline_state.get("label") or "Programmato")
    deadline_card = {
        "label": "Scadenze amministrative",
        "value": deadline_value,
        "meta": deadline_meta,
        "badge_label": deadline_badge_label,
        "badge_class": deadline_badge_class,
        "link_label": "Apri scadenze",
        "link_url": asset_administrative_deadline_list_url,
    }
    return {"coverage": coverage, "deadline": deadline_card}


def _build_asset_primary_kpis(
    *,
    asset: Asset,
    asset_status: dict[str, str],
    assignment_rows: list[dict[str, str]],
    primary_contract: AssistanceContract | None,
    primary_contract_state: dict[str, str] | None,
    next_maintenance_row: dict[str, object] | None,
    next_deadline_row: dict[str, object] | None,
    preferred_detail_metrics: list[dict[str, object]],
    detail_metrics: list[dict[str, object]],
    asset_assign_url: str,
    asset_assistance_contracts_url: str,
    asset_maintenance_schedule_url: str,
    asset_administrative_deadline_list_url: str,
) -> list[dict[str, str]]:
    # I KPI di default (stato/copertura/prossima manutenzione/scadenze) sono stati
    # ridistribuiti: stato nella pill header, copertura+scadenze nella status band,
    # prossima manutenzione nel Registro manutenzione. Qui restano solo le eventuali
    # metriche "preferite" configurate via AssetDetailField.
    default_cards: list[dict[str, str]] = []

    if not any(card["meta"] for card in default_cards) and detail_metrics:
        for index, metric in enumerate(detail_metrics[:4]):
            if index >= len(default_cards):
                break
            default_cards[index]["meta"] = str(metric.get("value") or "")

    cards: list[dict[str, str]] = []
    if preferred_detail_metrics:
        for metric in preferred_detail_metrics:
            label = _coalesce_str(metric.get("label"), "")
            if not label:
                continue
            cards.append(
                {
                    "label": label,
                    "value": _coalesce_str(metric.get("value"), "N/D"),
                    "meta": "",
                    "badge_label": "",
                    "badge_class": "info",
                    "link_label": "",
                    "link_url": "",
                    "size_class": str(metric.get("size_class") or "af-span-third"),
                }
            )
            if len(cards) >= 4:
                return cards[:4]

    used_labels = {str(card.get("label")) for card in cards}
    for card in default_cards:
        if str(card.get("label")) in used_labels:
            continue
        cards.append(card)
        if len(cards) >= 4:
            break

    return cards[:4]


def _is_assets_admin(request: HttpRequest) -> bool:
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    return bool(request.user.is_superuser or (legacy_user and is_legacy_admin(legacy_user)))


def _can_manage_asset_detail_layout(request: HttpRequest) -> bool:
    if _is_assets_admin(request):
        return True
    return bool(
        user_can_modulo_action(request, "assets", "admin_assets")
        or user_can_modulo_action(request, "assets", "asset_detail_layout")
    )


def _can_manage_asset_document_folders(request: HttpRequest) -> bool:
    """Solo admin/gestori asset possono aggiungere o disattivare cartelle documento."""
    if _is_assets_admin(request):
        return True
    return bool(user_can_modulo_action(request, "assets", "admin_assets"))


def _can_manage_asset_timeline_entries(request: HttpRequest) -> bool:
    """Chi puo' aggiungere voci manuali alla timeline di vita di un asset.

    Di serie gli amministratori/gestori asset. La platea e' pero' "opzionabile":
    dal pannello Accessi si concede l'azione ``assets/asset_timeline_entry`` (per
    ruolo o come override utente) ai manutentori, senza toccare il codice. Senza
    grant il pulsante non compare e la POST viene rifiutata.
    """
    if _is_assets_admin(request):
        return True
    return bool(
        user_can_modulo_action(request, "assets", "admin_assets")
        or user_can_modulo_action(request, "assets", "asset_timeline_entry")
    )


def _asset_timeline_entry_payload(request: HttpRequest) -> tuple[dict | None, str]:
    """Legge dalla POST i campi di una voce manuale di timeline.

    Stessa lettura per inserimento e modifica: ritorna ``(payload, errore)``,
    con payload ``None`` quando manca un dato obbligatorio.
    """
    entry_title = _clean_string(request.POST.get("timeline_title"))[:160]
    raw_event_date = _clean_string(request.POST.get("timeline_date"))
    try:
        entry_date = date.fromisoformat(raw_event_date) if raw_event_date else None
    except ValueError:
        entry_date = None
    entry_color = _clean_string(request.POST.get("timeline_color")).lower()
    if entry_color not in {code for code, _ in AssetTimelineEntry.COLOR_CHOICES}:
        entry_color = AssetTimelineEntry.COLOR_BLUE
    if not entry_title:
        return None, "Inserisci il titolo dell'evento da registrare in timeline."
    if entry_date is None:
        return None, "Inserisci una data evento valida."
    return (
        {
            "event_date": entry_date,
            "title": entry_title,
            "tag": _clean_string(request.POST.get("timeline_tag"))[:40].upper(),
            "description": _clean_string(request.POST.get("timeline_description"))[:2000],
            "meta": _clean_string(request.POST.get("timeline_meta"))[:120],
            "color": entry_color,
        },
        "",
    )


def _asset_timeline_hidden_keys(asset: Asset) -> set[str]:
    """Chiavi degli eventi automatici che l'utente ha rimosso dalla timeline."""
    return set(asset.timeline_hidden_events.values_list("event_key", flat=True))


def _asset_timeline_manual_enabled(asset: Asset) -> bool:
    """L'inserimento manuale e' attivabile/disattivabile per categoria asset."""
    category = getattr(asset, "asset_category", None)
    if category is None:
        return True
    return bool(getattr(category, "detail_timeline_manual_enabled", True))


def _can_manage_asset_list_layout(request: HttpRequest) -> bool:
    if _is_assets_admin(request):
        return True
    return bool(
        user_can_modulo_action(request, "assets", "admin_assets")
        or user_can_modulo_action(request, "assets", "asset_list_layout")
    )


def _legacy_employee_options() -> tuple[list[tuple[str, str]], dict[str, dict[str, str]]]:
    try:
        users = list(
            UtenteLegacy.objects.filter(attivo=True)
            .exclude(id__isnull=True)
            .order_by("nome", "email", "id")
        )
    except DatabaseError:
        return [], {}

    user_ids = [int(u.id) for u in users if getattr(u, "id", None)]
    anagrafica_map: dict[int, AnagraficaDipendente] = {}
    extra_map: dict[int, UserExtraInfo] = {}

    if user_ids:
        try:
            for row in AnagraficaDipendente.objects.filter(utente_id__in=user_ids):
                if row.utente_id is not None:
                    anagrafica_map[int(row.utente_id)] = row
        except DatabaseError:
            anagrafica_map = {}
        try:
            for row in UserExtraInfo.objects.filter(legacy_user_id__in=user_ids):
                extra_map[int(row.legacy_user_id)] = row
        except Exception:
            extra_map = {}

    options: list[tuple[str, str]] = []
    details: dict[str, dict[str, str]] = {}
    for user in users:
        uid = int(user.id)
        anagrafica = anagrafica_map.get(uid)
        extra = extra_map.get(uid)
        nome = _clean_string(user.nome)
        email = _clean_string(user.email)
        notification_email = _clean_string(getattr(anagrafica, "email_notifica", ""))
        calendar_user_id = email or notification_email
        if not nome and anagrafica:
            nome = " ".join(
                [
                    _clean_string(getattr(anagrafica, "nome", "")),
                    _clean_string(getattr(anagrafica, "cognome", "")),
                ]
            ).strip()
        display_name = nome or email or f"Utente #{uid}"
        reparto = _clean_string(getattr(extra, "reparto", "")) or _clean_string(getattr(anagrafica, "reparto", ""))
        label_email = calendar_user_id
        label = f"{display_name} - {label_email}" if label_email else display_name
        options.append((str(uid), label))
        details[str(uid)] = {
            "display_name": display_name,
            "email": email,
            "notification_email": notification_email,
            "calendar_user_id": calendar_user_id,
            "reparto": reparto,
        }
    return options, details


def _anagrafica_assignment_options() -> tuple[list[tuple[str, str]], dict[str, dict[str, str]], list[str]]:
    try:
        rows = list(
            AnagraficaDipendente.objects.all()
            .order_by("cognome", "nome", "id")[:1000]
        )
    except DatabaseError:
        return [], {}, []

    choices: list[tuple[str, str]] = []
    details: dict[str, dict[str, str]] = {}
    departments: set[str] = set()
    for row in rows:
        display_name = " ".join(
            [
                _clean_string(getattr(row, "cognome", "")),
                _clean_string(getattr(row, "nome", "")),
            ]
        ).strip() or _clean_string(getattr(row, "aliasusername", "")) or f"Dipendente #{row.id}"
        reparto = _clean_string(getattr(row, "reparto", ""))
        mansione = _clean_string(getattr(row, "mansione", ""))
        label_bits = [display_name]
        if reparto:
            label_bits.append(reparto)
            departments.add(reparto)
        if mansione:
            label_bits.append(mansione)
        label = " - ".join(label_bits)
        key = str(row.id)
        choices.append((key, label))
        details[key] = {
            "display_name": display_name,
            "reparto": reparto,
            "mansione": mansione,
            "legacy_user_id": str(getattr(row, "utente_id", "") or ""),
        }

    for value in (
        Asset.objects.exclude(reparto="")
        .values_list("reparto", flat=True)
        .distinct()[:300]
    ):
        cleaned = _clean_string(value)
        if cleaned:
            departments.add(cleaned)
    return choices, details, sorted(departments, key=lambda row: row.lower())


def _assignment_form_kwargs(asset: Asset | None = None) -> dict[str, object]:
    employee_choices, employee_details, department_choices = _anagrafica_assignment_options()
    include_in_layout = bool(asset and PlantLayoutMarker.objects.filter(asset=asset).exists())
    return {
        "assignment_employee_choices": employee_choices,
        "assignment_employee_details": employee_details,
        "assignment_department_choices": department_choices,
        "default_include_in_plant_layout": include_in_layout,
    }


def _resolve_asset_plant_layout(asset: Asset) -> tuple["PlantLayout | None", bool]:
    """Sceglie la planimetria attiva più coerente con il reparto dell'asset.

    Ritorna ``(layout, is_specific)``:
      - (a) layout attivo la cui ``category`` coincide col reparto dell'asset;
      - (b) layout attivo con un'area il cui ``reparto_code`` coincide col reparto;
      - (c) fallback: primo layout attivo per (category, name, id).
    ``is_specific`` è True solo per (a)/(b). ``layout`` è ``None`` se non esiste
    alcuna planimetria attiva.
    """
    active = PlantLayout.objects.filter(is_active=True)
    reparto = _clean_string(getattr(asset, "reparto", ""))
    if reparto:
        specific = active.filter(category__iexact=reparto).order_by("category", "name", "id").first()
        if specific is not None:
            return specific, True
        by_area = (
            active.filter(areas__reparto_code__iexact=reparto)
            .order_by("category", "name", "id")
            .first()
        )
        if by_area is not None:
            return by_area, True
    return active.order_by("category", "name", "id").first(), False


def _ensure_asset_plant_layout_marker(asset: Asset) -> str:
    layout, is_specific = _resolve_asset_plant_layout(asset)
    if layout is None:
        return "Asset salvato, ma non inserito in piantina: nessuna planimetria attiva disponibile."

    # Se la risoluzione è specifica per reparto e l'asset ha già un marker su un
    # layout diverso, spostalo su quello corretto (rispettando il vincolo unico
    # layout+asset).
    if is_specific:
        strays = list(PlantLayoutMarker.objects.filter(asset=asset).exclude(layout=layout))
        if strays:
            target_exists = PlantLayoutMarker.objects.filter(layout=layout, asset=asset).exists()
            if not target_exists:
                moved = strays.pop(0)
                moved.layout = layout
                moved.is_visible = True
                moved.save(update_fields=["layout", "is_visible", "updated_at"])
            for stray in strays:
                stray.delete()

    marker, created = PlantLayoutMarker.objects.get_or_create(
        layout=layout,
        asset=asset,
        defaults={
            "label": asset.name[:120],
            "x_percent": 50,
            "y_percent": 50,
            "is_visible": True,
            "sort_order": 100,
        },
    )
    if not created and not marker.is_visible:
        marker.is_visible = True
        marker.save(update_fields=["is_visible", "updated_at"])
    return ""


def _outlook_calendar_graph_settings() -> dict[str, str]:
    return {
        "tenant_id": _clean_string(get_first_env_value("GRAPH_TENANT_ID", "AZURE_TENANT_ID")),
        "client_id": _clean_string(get_first_env_value("GRAPH_CLIENT_ID", "AZURE_CLIENT_ID")),
        "client_secret": _clean_string(get_first_env_value("GRAPH_CLIENT_SECRET", "AZURE_CLIENT_SECRET")),
    }


def _outlook_calendar_graph_ready() -> bool:
    config = _outlook_calendar_graph_settings()
    return all(config.values()) and not any(is_placeholder_value(value) for value in config.values())


def _outlook_calendar_graph_token() -> str:
    config = _outlook_calendar_graph_settings()
    if not _outlook_calendar_graph_ready():
        raise RuntimeError("Configurazione Graph incompleta: tenant, client o secret mancanti.")
    return acquire_graph_token(config["tenant_id"], config["client_id"], config["client_secret"])


def _outlook_calendar_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {_outlook_calendar_graph_token()}",
        "Content-Type": "application/json",
    }


def _asset_calendar_target_email(target_details: dict[str, str]) -> str:
    return _clean_string(target_details.get("calendar_user_id") or target_details.get("email"))


def _asset_calendar_default_user_id(
    asset: Asset | None,
    calendar_user_details: dict[str, dict[str, str]],
) -> str:
    suggested_user_id = str(getattr(asset, "assigned_legacy_user_id", "") or "")
    if suggested_user_id and suggested_user_id not in calendar_user_details:
        return ""
    return suggested_user_id


def _asset_calendar_source_key(
    *,
    event_kind: str,
    asset_id: int,
    source_id: int,
    due_date: date,
    target_legacy_user_id: int,
) -> str:
    return (
        f"{_clean_string(event_kind)}:{int(asset_id or 0)}:{int(source_id or 0)}:"
        f"{due_date.isoformat()}:{int(target_legacy_user_id or 0)}"
    )


def _asset_calendar_transaction_id(*, source_key: str) -> str:
    return str(uuid5(NAMESPACE_URL, f"https://novicrom.local/assets/calendar/{source_key}"))


def _outlook_event_time_window(due_date: date) -> tuple[datetime, datetime]:
    start_dt = datetime.combine(due_date, datetime.min.time()).replace(hour=OUTLOOK_CALENDAR_START_HOUR)
    end_dt = start_dt + timedelta(minutes=OUTLOOK_CALENDAR_DURATION_MINUTES)
    return start_dt, end_dt


def _build_outlook_event_payload(
    *,
    subject: str,
    body_html: str,
    location_label: str,
    due_date: date,
    transaction_id: str,
) -> dict[str, object]:
    start_dt, end_dt = _outlook_event_time_window(due_date)
    return {
        "subject": _clean_string(subject)[:255],
        "body": {
            "contentType": "HTML",
            "content": body_html,
        },
        "start": {
            "dateTime": start_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "timeZone": OUTLOOK_CALENDAR_TIMEZONE,
        },
        "end": {
            "dateTime": end_dt.strftime("%Y-%m-%dT%H:%M:%S"),
            "timeZone": OUTLOOK_CALENDAR_TIMEZONE,
        },
        "location": {
            "displayName": _clean_string(location_label) or "Promemoria assets",
        },
        "showAs": "busy",
        "responseRequested": False,
        "transactionId": transaction_id,
    }


def _maintenance_schedule_redirect_from_request(request: HttpRequest) -> str:
    return _maintenance_schedule_page_url(
        asset_id=_as_int(request.POST.get("filter_asset"), default=0),
        status=_clean_string(request.POST.get("filter_status")),
        category_id=_as_int(request.POST.get("filter_category"), default=0),
        reparto=_clean_string(request.POST.get("filter_reparto")),
        coverage=_clean_string(request.POST.get("filter_coverage")),
        q=_clean_string(request.POST.get("filter_q")),
    )


def _deadline_list_redirect_from_request(request: HttpRequest) -> str:
    return _asset_administrative_deadline_page_url(
        asset_id=_as_int(request.POST.get("filter_asset"), default=0),
        component_id=_as_int(request.POST.get("filter_component"), default=0),
        deadline_type=_clean_string(request.POST.get("filter_deadline_type")),
        status=_clean_string(request.POST.get("filter_status")),
        q=_clean_string(request.POST.get("filter_q")),
    )


def _periodic_verification_redirect_from_request(request: HttpRequest) -> str:
    return _periodic_verifications_page_url(
        asset_id=_as_int(request.POST.get("filter_asset"), default=0),
        edit_id=_as_int(request.POST.get("filter_edit"), default=0),
        scope=_clean_string(request.POST.get("filter_scope") or request.POST.get("scope")),
        window=_clean_string(request.POST.get("filter_window")),
        view=_clean_string(request.POST.get("filter_view")),
        q=_clean_string(request.POST.get("filter_q")),
    )


def _assistance_contract_redirect_from_request(request: HttpRequest) -> str:
    return _assistance_contracts_page_url(
        asset_id=_as_int(request.POST.get("filter_asset"), default=0),
        supplier_filter=_as_int(request.POST.get("filter_supplier"), default=0),
        state=_clean_string(request.POST.get("filter_state")),
        scope=_clean_string(request.POST.get("filter_scope")),
        q=_clean_string(request.POST.get("filter_q")),
    )


def _software_license_redirect_from_request(request: HttpRequest) -> str:
    return _software_licenses_page_url(
        asset_id=_as_int(request.POST.get("filter_asset"), default=0),
        anagrafica_id=_as_int(request.POST.get("filter_anagrafica"), default=0),
        edit_id=_as_int(request.POST.get("filter_edit"), default=0),
        category=_clean_string(request.POST.get("filter_category")),
        status=_clean_string(request.POST.get("filter_status")),
        assignee=_clean_string(request.POST.get("filter_assignee")),
        q=_clean_string(request.POST.get("filter_q")),
    )


def _maintenance_schedule_row_for_asset_rule(*, asset: Asset, base_rule_id: int) -> dict[str, object] | None:
    rows = build_day_based_maintenance_schedule_rows(
        asset_queryset=Asset.objects.filter(pk=asset.id).select_related("asset_category"),
    )
    for row in rows:
        if getattr(row.get("base_rule"), "id", 0) == int(base_rule_id or 0):
            return row
    return None


def _build_maintenance_outlook_event_payload(
    *,
    request: HttpRequest,
    asset: Asset,
    schedule_row: dict[str, object],
    target_display_name: str,
    transaction_id: str,
) -> dict[str, object]:
    due_date = schedule_row.get("due_date")
    if not isinstance(due_date, date):
        raise RuntimeError("La regola selezionata non ha una scadenza calendarizzabile.")

    template = schedule_row["effective_intervention_template"]
    template_label = _clean_string(getattr(template, "label", "Intervento manutentivo")) or "Intervento manutentivo"
    threshold_label = _clean_string(schedule_row.get("effective_threshold_label"))
    threshold_value = schedule_row.get("effective_threshold_value")
    warning_days = int(schedule_row.get("effective_warning_days") or 0)
    contract = get_primary_assistance_contract(asset)
    asset_url = request.build_absolute_uri(reverse("assets:asset_view", kwargs={"id": asset.id}))
    workorder_url = request.build_absolute_uri(
        _workorder_create_page_url(asset_id=asset.id, rule_id=schedule_row["base_rule"].id, source="maintenance_schedule")
    )
    due_date_label = due_date.strftime("%d-%m-%Y")
    subject = f"Manutenzione {asset.asset_tag} - {template_label}"[:255]

    start_dt = datetime.combine(due_date, datetime.min.time()).replace(hour=OUTLOOK_CALENDAR_START_HOUR)
    end_dt = start_dt + timedelta(minutes=OUTLOOK_CALENDAR_DURATION_MINUTES)
    contract_label = (
        escape(_clean_string(getattr(contract, "title", "")) or str(getattr(contract, "supplier", "")))
        if contract is not None
        else "Nessuna copertura contrattuale attiva"
    )
    body_html = (
        "<p>Promemoria manutenzione generato dal modulo Assets.</p>"
        f"<p><strong>Destinatario:</strong> {escape(target_display_name or 'Utente selezionato')}<br>"
        f"<strong>Asset:</strong> {escape(asset.asset_tag)} - {escape(asset.name)}<br>"
        f"<strong>Reparto:</strong> {escape(_clean_string(asset.reparto) or '-')}<br>"
        f"<strong>Intervento:</strong> {escape(template_label)}<br>"
        f"<strong>Scadenza:</strong> {escape(due_date_label)}<br>"
        f"<strong>Periodicita:</strong> {escape(str(threshold_value or 0))} {escape(threshold_label.lower() or 'giorni')}<br>"
        f"<strong>Warning:</strong> {warning_days} gg<br>"
        f"<strong>Contratto:</strong> {contract_label}</p>"
        f"<p><a href=\"{escape(asset_url)}\">Apri scheda asset</a><br>"
        f"<a href=\"{escape(workorder_url)}\">Apri creazione intervento</a></p>"
    )
    return _build_outlook_event_payload(
        subject=subject,
        body_html=body_html,
        location_label=_clean_string(asset.reparto) or "Manutenzione asset",
        due_date=due_date,
        transaction_id=transaction_id,
    )


def _build_deadline_outlook_event_payload(
    *,
    request: HttpRequest,
    deadline: AssetAdministrativeDeadline,
    target_display_name: str,
    transaction_id: str,
) -> dict[str, object]:
    asset = deadline.asset
    component_label = (
        f"{deadline.component.name} ({deadline.component.code})"
        if deadline.component_id and deadline.component and deadline.component.code
        else getattr(deadline.component, "name", "")
    )
    asset_url = request.build_absolute_uri(reverse("assets:asset_view", kwargs={"id": asset.id}))
    edit_url = request.build_absolute_uri(reverse("assets:asset_administrative_deadline_edit", kwargs={"id": deadline.id}))
    body_html = (
        "<p>Promemoria scadenza amministrativa/tecnica generato dal modulo Assets.</p>"
        f"<p><strong>Destinatario:</strong> {escape(target_display_name or 'Utente selezionato')}<br>"
        f"<strong>Asset:</strong> {escape(asset.asset_tag)} - {escape(asset.name)}<br>"
        f"<strong>Scadenza:</strong> {escape(deadline.title)}<br>"
        f"<strong>Tipo:</strong> {escape(deadline.get_deadline_type_display())}<br>"
        f"<strong>Data:</strong> {deadline.due_date:%d-%m-%Y}<br>"
        f"<strong>Preavviso:</strong> {int(deadline.warning_days or 0)} gg<br>"
        f"<strong>Riferimento:</strong> {escape(_clean_string(deadline.reference_code) or '-')}<br>"
        f"<strong>Ente/Rilasciato da:</strong> {escape(_clean_string(deadline.issuer) or '-')}<br>"
        f"<strong>Componente:</strong> {escape(_clean_string(component_label) or '-')}</p>"
        f"<p><a href=\"{escape(asset_url)}\">Apri scheda asset</a><br>"
        f"<a href=\"{escape(edit_url)}\">Apri scadenza amministrativa</a></p>"
    )
    subject = f"Scadenza {asset.asset_tag} - {deadline.title}"[:255]
    return _build_outlook_event_payload(
        subject=subject,
        body_html=body_html,
        location_label=_clean_string(asset.reparto) or "Scadenza asset",
        due_date=deadline.due_date,
        transaction_id=transaction_id,
    )


def _build_periodic_verification_outlook_event_payload(
    *,
    request: HttpRequest,
    asset: Asset,
    verification: PeriodicVerification,
    target_display_name: str,
    transaction_id: str,
) -> dict[str, object]:
    next_date = verification.next_verification_date
    if not isinstance(next_date, date):
        raise RuntimeError("La manutenzione periodica non ha una prossima data calendarizzabile.")
    asset_url = request.build_absolute_uri(reverse("assets:asset_view", kwargs={"id": asset.id}))
    verification_url = request.build_absolute_uri(_periodic_verifications_page_url(asset_id=asset.id, edit_id=verification.id))
    supplier_label = _clean_string(str(getattr(verification, "supplier", "") or "")) or "Fornitore non impostato"
    body_html = (
        "<p>Promemoria manutenzione periodica generato dal modulo Assets.</p>"
        f"<p><strong>Destinatario:</strong> {escape(target_display_name or 'Utente selezionato')}<br>"
        f"<strong>Asset:</strong> {escape(asset.asset_tag)} - {escape(asset.name)}<br>"
        f"<strong>Manutenzione periodica:</strong> {escape(verification.name)}<br>"
        f"<strong>Prossima data:</strong> {next_date:%d-%m-%Y}<br>"
        f"<strong>Cadenza:</strong> {int(verification.frequency_months or 0)} mesi<br>"
        f"<strong>Fornitore:</strong> {escape(supplier_label)}</p>"
        f"<p><a href=\"{escape(asset_url)}\">Apri scheda asset</a><br>"
        f"<a href=\"{escape(verification_url)}\">Apri manutenzione periodica</a></p>"
    )
    subject = f"Manutenzione periodica {asset.asset_tag} - {verification.name}"[:255]
    return _build_outlook_event_payload(
        subject=subject,
        body_html=body_html,
        location_label=_clean_string(asset.reparto) or "Manutenzione periodica asset",
        due_date=next_date,
        transaction_id=transaction_id,
    )


def _build_assistance_contract_outlook_event_payload(
    *,
    request: HttpRequest,
    asset: Asset,
    contract: AssistanceContract,
    target_display_name: str,
    transaction_id: str,
) -> dict[str, object]:
    if not isinstance(contract.end_date, date):
        raise RuntimeError("Il contratto selezionato non ha una data scadenza calendarizzabile.")
    supplier_label = _clean_string(str(getattr(contract, "supplier", "") or "")) or "Fornitore non impostato"
    asset_url = request.build_absolute_uri(reverse("assets:asset_view", kwargs={"id": asset.id}))
    contracts_url = request.build_absolute_uri(_assistance_contracts_page_url(asset_id=asset.id, edit_id=contract.id))
    scope_payload = _contract_scope_payload(contract)
    body_html = (
        "<p>Promemoria scadenza contratto assistenza generato dal modulo Assets.</p>"
        f"<p><strong>Destinatario:</strong> {escape(target_display_name or 'Utente selezionato')}<br>"
        f"<strong>Asset di contesto:</strong> {escape(asset.asset_tag)} - {escape(asset.name)}<br>"
        f"<strong>Contratto:</strong> {escape(contract.title)}<br>"
        f"<strong>Fornitore:</strong> {escape(supplier_label)}<br>"
        f"<strong>Codice:</strong> {escape(_clean_string(contract.code) or '-')}<br>"
        f"<strong>Scadenza:</strong> {contract.end_date:%d-%m-%Y}<br>"
        f"<strong>Ambito:</strong> {escape(scope_payload.get('label') or '-')}<br>"
        f"<strong>Dettaglio:</strong> {escape(scope_payload.get('detail') or '-')}</p>"
        f"<p><a href=\"{escape(asset_url)}\">Apri scheda asset</a><br>"
        f"<a href=\"{escape(contracts_url)}\">Apri contratti assistenza</a></p>"
    )
    subject = f"Scadenza contratto {asset.asset_tag} - {contract.title}"[:255]
    return _build_outlook_event_payload(
        subject=subject,
        body_html=body_html,
        location_label=_clean_string(asset.reparto) or "Contratto assistenza",
        due_date=contract.end_date,
        transaction_id=transaction_id,
    )


def _outlook_calendar_create_event(*, target_email: str, payload: dict[str, object]) -> dict[str, object]:
    if not _outlook_calendar_graph_ready():
        raise RuntimeError("Integrazione Outlook non configurata: mancano le credenziali Graph valide.")
    target_value = _clean_string(target_email)
    if not target_value:
        raise RuntimeError("Utente Outlook non valido.")
    url = f"https://graph.microsoft.com/v1.0/users/{quote(target_value, safe='')}/calendar/events"
    response = requests.post(url, headers=_outlook_calendar_headers(), json=payload, timeout=20)
    if response.status_code in {200, 201}:
        return response.json()
    try:
        payload = response.json()
    except ValueError:
        payload = {}
    graph_error = payload.get("error") if isinstance(payload, dict) else None
    message = ""
    if isinstance(graph_error, dict):
        message = _clean_string(graph_error.get("message"))
    raise RuntimeError(message or response.text or f"Errore Outlook Calendar {response.status_code}")


def _create_asset_calendar_event_record(
    *,
    asset: Asset,
    event_kind: str,
    source_key: str,
    due_date: date,
    target_legacy_user_id: int,
    target_details: dict[str, str],
    payload: dict[str, object],
    request_user,
    maintenance_rule: MaintenanceRule | None = None,
    administrative_deadline: AssetAdministrativeDeadline | None = None,
    periodic_verification: PeriodicVerification | None = None,
    assistance_contract: AssistanceContract | None = None,
) -> tuple[AssetCalendarEvent, bool]:
    target_email = _asset_calendar_target_email(target_details)
    if not target_email:
        raise RuntimeError("L'utente selezionato non ha un identificatore Outlook configurato.")
    existing = AssetCalendarEvent.objects.filter(source_key=source_key).order_by("id").first()
    if existing is not None:
        return existing, False

    target_display_name = _clean_string(target_details.get("display_name")) or target_email
    graph_event = _outlook_calendar_create_event(target_email=target_email, payload=payload)
    try:
        entry = AssetCalendarEvent.objects.create(
            asset=asset,
            event_kind=event_kind,
            source_key=source_key,
            maintenance_rule=maintenance_rule,
            administrative_deadline=administrative_deadline,
            periodic_verification=periodic_verification,
            assistance_contract=assistance_contract,
            due_date=due_date,
            target_legacy_user_id=int(target_legacy_user_id),
            target_display_name=target_display_name,
            target_email=target_email,
            subject=_clean_string(payload.get("subject")),
            transaction_id=_clean_string(payload.get("transactionId")),
            graph_event_id=_clean_string(graph_event.get("id")),
            graph_event_web_link=_clean_string(graph_event.get("webLink")),
            created_by=request_user if getattr(request_user, "is_authenticated", False) else None,
        )
    except IntegrityError:
        entry = AssetCalendarEvent.objects.get(source_key=source_key)
        return entry, False
    return entry, True


def _create_asset_maintenance_calendar_event(
    *,
    request: HttpRequest,
    asset: Asset,
    schedule_row: dict[str, object],
    target_legacy_user_id: int,
    target_details: dict[str, str],
) -> tuple[AssetCalendarEvent, bool]:
    due_date = schedule_row.get("due_date")
    if not isinstance(due_date, date):
        raise RuntimeError("La manutenzione selezionata non ha una scadenza impostata.")
    source_key = _asset_calendar_source_key(
        event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
        asset_id=asset.id,
        source_id=schedule_row["base_rule"].id,
        due_date=due_date,
        target_legacy_user_id=target_legacy_user_id,
    )
    payload = _build_maintenance_outlook_event_payload(
        request=request,
        asset=asset,
        schedule_row=schedule_row,
        target_display_name=_clean_string(target_details.get("display_name")),
        transaction_id=_asset_calendar_transaction_id(source_key=source_key),
    )
    return _create_asset_calendar_event_record(
        asset=asset,
        event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
        source_key=source_key,
        due_date=due_date,
        target_legacy_user_id=target_legacy_user_id,
        target_details=target_details,
        payload=payload,
        request_user=request.user,
        maintenance_rule=schedule_row["base_rule"],
    )


def _create_asset_deadline_calendar_event(
    *,
    request: HttpRequest,
    deadline: AssetAdministrativeDeadline,
    target_legacy_user_id: int,
    target_details: dict[str, str],
) -> tuple[AssetCalendarEvent, bool]:
    due_date = deadline.due_date
    source_key = _asset_calendar_source_key(
        event_kind=AssetCalendarEvent.KIND_ADMINISTRATIVE_DEADLINE,
        asset_id=deadline.asset_id,
        source_id=deadline.id,
        due_date=due_date,
        target_legacy_user_id=target_legacy_user_id,
    )
    payload = _build_deadline_outlook_event_payload(
        request=request,
        deadline=deadline,
        target_display_name=_clean_string(target_details.get("display_name")),
        transaction_id=_asset_calendar_transaction_id(source_key=source_key),
    )
    return _create_asset_calendar_event_record(
        asset=deadline.asset,
        event_kind=AssetCalendarEvent.KIND_ADMINISTRATIVE_DEADLINE,
        source_key=source_key,
        due_date=due_date,
        target_legacy_user_id=target_legacy_user_id,
        target_details=target_details,
        payload=payload,
        request_user=request.user,
        administrative_deadline=deadline,
    )


def _create_asset_periodic_verification_calendar_event(
    *,
    request: HttpRequest,
    asset: Asset,
    verification: PeriodicVerification,
    target_legacy_user_id: int,
    target_details: dict[str, str],
) -> tuple[AssetCalendarEvent, bool]:
    next_date = verification.next_verification_date
    if not isinstance(next_date, date):
        raise RuntimeError("La manutenzione periodica selezionata non ha una prossima data impostata.")
    source_key = _asset_calendar_source_key(
        event_kind=AssetCalendarEvent.KIND_PERIODIC_VERIFICATION,
        asset_id=asset.id,
        source_id=verification.id,
        due_date=next_date,
        target_legacy_user_id=target_legacy_user_id,
    )
    payload = _build_periodic_verification_outlook_event_payload(
        request=request,
        asset=asset,
        verification=verification,
        target_display_name=_clean_string(target_details.get("display_name")),
        transaction_id=_asset_calendar_transaction_id(source_key=source_key),
    )
    return _create_asset_calendar_event_record(
        asset=asset,
        event_kind=AssetCalendarEvent.KIND_PERIODIC_VERIFICATION,
        source_key=source_key,
        due_date=next_date,
        target_legacy_user_id=target_legacy_user_id,
        target_details=target_details,
        payload=payload,
        request_user=request.user,
        periodic_verification=verification,
    )


def _create_asset_assistance_contract_calendar_event(
    *,
    request: HttpRequest,
    asset: Asset,
    contract: AssistanceContract,
    target_legacy_user_id: int,
    target_details: dict[str, str],
) -> tuple[AssetCalendarEvent, bool]:
    if not isinstance(contract.end_date, date):
        raise RuntimeError("Il contratto selezionato non ha una data scadenza impostata.")
    source_key = _asset_calendar_source_key(
        event_kind=AssetCalendarEvent.KIND_ASSISTANCE_CONTRACT,
        asset_id=asset.id,
        source_id=contract.id,
        due_date=contract.end_date,
        target_legacy_user_id=target_legacy_user_id,
    )
    payload = _build_assistance_contract_outlook_event_payload(
        request=request,
        asset=asset,
        contract=contract,
        target_display_name=_clean_string(target_details.get("display_name")),
        transaction_id=_asset_calendar_transaction_id(source_key=source_key),
    )
    return _create_asset_calendar_event_record(
        asset=asset,
        event_kind=AssetCalendarEvent.KIND_ASSISTANCE_CONTRACT,
        source_key=source_key,
        due_date=contract.end_date,
        target_legacy_user_id=target_legacy_user_id,
        target_details=target_details,
        payload=payload,
        request_user=request.user,
        assistance_contract=contract,
    )


def _as_int(value, default: int = 100) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _unique_custom_field_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"campo-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetCustomField.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


def _update_custom_field_values_after_delete(field_code: str) -> int:
    updated_count = 0
    for asset in Asset.objects.only("id", "extra_columns").iterator():
        if not isinstance(asset.extra_columns, dict):
            continue
        if field_code not in asset.extra_columns:
            continue
        next_extra = dict(asset.extra_columns)
        next_extra.pop(field_code, None)
        asset.extra_columns = next_extra
        asset.save(update_fields=["extra_columns"])
        updated_count += 1
    return updated_count


def _update_asset_category_values_after_delete(field_code: str) -> int:
    updated_count = 0
    for asset in Asset.objects.only("id", "extra_columns").iterator():
        if not isinstance(asset.extra_columns, dict):
            continue
        category_values = asset.extra_columns.get("_category_fields", {})
        if not isinstance(category_values, dict) or field_code not in category_values:
            continue
        next_extra = dict(asset.extra_columns)
        next_category_values = dict(category_values)
        next_category_values.pop(field_code, None)
        if next_category_values:
            next_extra["_category_fields"] = next_category_values
        else:
            next_extra.pop("_category_fields", None)
        asset.extra_columns = next_extra
        asset.save(update_fields=["extra_columns"])
        updated_count += 1
    return updated_count


def _build_asset_list_suggestions(
    employee_details: dict[str, dict[str, str]] | None = None,
) -> dict[str, list[str]]:
    field_keys = [key for key, _ in AssetListOption.FIELD_CHOICES]
    merged: dict[str, set[str]] = {key: set() for key in field_keys}

    option_qs = AssetListOption.objects.filter(is_active=True).order_by("field_key", "sort_order", "value")
    for option in option_qs:
        cleaned_value = _clean_string(option.value)
        if cleaned_value:
            merged.setdefault(option.field_key, set()).add(cleaned_value)

    for field_key in field_keys:
        if not hasattr(Asset, field_key):
            continue
        db_values = (
            Asset.objects.exclude(**{f"{field_key}__isnull": True})
            .exclude(**{field_key: ""})
            .values_list(field_key, flat=True)
            .distinct()[:300]
        )
        for value in db_values:
            cleaned_value = _clean_string(str(value))
            if cleaned_value:
                merged.setdefault(field_key, set()).add(cleaned_value)

    if employee_details:
        for details in employee_details.values():
            display_name = _clean_string(details.get("display_name"))
            reparto = _clean_string(details.get("reparto"))
            if display_name:
                merged.setdefault(AssetListOption.FIELD_ASSIGNMENT_TO, set()).add(display_name)
            if reparto:
                merged.setdefault(AssetListOption.FIELD_ASSIGNMENT_REPARTO, set()).add(reparto)

    normalized: dict[str, list[str]] = {}
    for key, values in merged.items():
        if values:
            normalized[key] = sorted(values, key=lambda row: row.lower())
    return normalized


def _handle_list_option_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_field_keys = {key for key, _ in AssetListOption.FIELD_CHOICES}

    if action == "create_list_option":
        field_key = _clean_string(request.POST.get("field_key"))
        value = _clean_string(request.POST.get("value"))
        if field_key not in valid_field_keys:
            return False, "Campo lista non valido."
        if not value:
            return False, "Inserisci il valore lista."
        sort_order = _as_int(request.POST.get("sort_order"), default=100)
        is_active = bool(request.POST.get("is_active"))
        option, created = AssetListOption.objects.get_or_create(
            field_key=field_key,
            value=value,
            defaults={"sort_order": sort_order, "is_active": is_active},
        )
        if created:
            return True, f"Valore \"{value}\" aggiunto."
        option.sort_order = sort_order
        option.is_active = is_active
        option.save(update_fields=["sort_order", "is_active", "updated_at"])
        return True, f"Valore \"{value}\" aggiornato."

    if action == "update_list_option":
        option_id = _as_int(request.POST.get("option_id"), default=0)
        option = AssetListOption.objects.filter(pk=option_id).first()
        if not option:
            return False, "Valore lista non trovato."
        field_key = _clean_string(request.POST.get("field_key")) or option.field_key
        value = _clean_string(request.POST.get("value"))
        if field_key not in valid_field_keys:
            return False, "Campo lista non valido."
        if not value:
            return False, "Il valore lista non puo essere vuoto."
        sort_order = _as_int(request.POST.get("sort_order"), default=option.sort_order)
        is_active = bool(request.POST.get("is_active"))
        duplicate_qs = AssetListOption.objects.filter(field_key=field_key, value=value).exclude(pk=option.pk)
        if duplicate_qs.exists():
            return False, "Valore gia presente per questo campo."
        option.field_key = field_key
        option.value = value
        option.sort_order = sort_order
        option.is_active = is_active
        option.save(update_fields=["field_key", "value", "sort_order", "is_active", "updated_at"])
        return True, f"Valore \"{value}\" salvato."

    if action == "delete_list_option":
        option_id = _as_int(request.POST.get("option_id"), default=0)
        option = AssetListOption.objects.filter(pk=option_id).first()
        if not option:
            return False, "Valore lista non trovato."
        label = option.value
        option.delete()
        return True, f"Valore \"{label}\" eliminato."

    return False, "Azione lista non valida."


def _unique_action_button_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"button-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetActionButton.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


def _unique_sidebar_button_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"menu-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetSidebarButton.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


def _unique_detail_field_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"dettaglio-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetDetailField.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


def _unique_asset_category_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"categoria-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetCategory.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


def _unique_asset_category_field_code(label: str, requested_code: str | None = None) -> str:
    seed = _clean_string(requested_code) or _clean_string(label)
    base = slugify(seed)[:70]
    if not base:
        base = f"categoria-campo-{uuid4().hex[:8]}"
    candidate = base
    index = 2
    while AssetCategoryField.objects.filter(code=candidate).exists():
        suffix = f"-{index}"
        candidate = f"{base[: max(1, 80 - len(suffix))]}{suffix}"
        index += 1
    return candidate


ASSET_DETAIL_SOURCE_PRESETS: list[tuple[str, str]] = [
    ("computed:travel_xyz", "Calcolato · Corse XYZ"),
    ("computed:machine_configuration", "Calcolato · Configurazione macchina"),
    ("computed:battery_health", "Calcolato · Salute batteria"),
    ("computed:cpu_load", "Calcolato · Carico medio CPU"),
    ("computed:storage_free", "Calcolato · Spazio libero"),
    ("computed:purchase_date", "Calcolato · Data acquisto"),
    ("computed:production_date", "Calcolato · Data produzione"),
    ("computed:sync_text", "Calcolato · Ultimo sync"),
    ("asset:asset_tag", "Asset · Tag asset"),
    ("asset:name", "Asset · Nome"),
    ("asset:asset_category", "Asset · Categoria asset"),
    ("asset:asset_type", "Asset · Tipo asset"),
    ("asset:reparto", "Asset · Reparto"),
    ("asset:manufacturer", "Asset · Produttore"),
    ("asset:model", "Asset · Modello"),
    ("asset:serial_number", "Asset · Numero seriale"),
    ("asset:status", "Asset · Stato"),
    ("asset:assignment_to", "Asset · Assegnato a"),
    ("asset:assignment_reparto", "Asset · Reparto assegnazione"),
    ("asset:assignment_location", "Asset · Posizione assegnazione"),
    ("asset:updated_at", "Asset · Ultimo aggiornamento"),
    ("it:cpu", "IT · Processore"),
    ("it:ram", "IT · Memoria RAM"),
    ("it:os", "IT · Sistema operativo"),
    ("it:disco", "IT · Archiviazione"),
    ("work_machine:x_mm", "Macchina · Corsa X"),
    ("work_machine:y_mm", "Macchina · Corsa Y"),
    ("work_machine:z_mm", "Macchina · Corsa Z"),
    ("work_machine:diameter_mm", "Macchina · Diametro"),
    ("work_machine:spindle_mm", "Macchina · Mandrino"),
    ("work_machine:year", "Macchina · Anno"),
    ("work_machine:tmc", "Macchina · TMC"),
    ("work_machine:tcr_enabled", "Macchina · TCR"),
    ("work_machine:pressure_bar", "Macchina · Pressione"),
    ("work_machine:cnc_controlled", "Macchina · CNC"),
    ("work_machine:five_axes", "Macchina · 5 assi"),
    ("work_machine:accuracy_from", "Macchina · Accuracy from"),
    ("work_machine:next_maintenance_date", "Macchina · Prossima manutenzione"),
    ("work_machine:maintenance_reminder_days", "Macchina · Soglia reminder"),
    ("extra:graphics", "Extra · Grafica"),
    ("extra:display", "Extra · Schermo"),
    ("extra:po_ref", "Extra · Riferimento ordine"),
    ("extra:owner_dept", "Extra · Reparto owner"),
    ("extra:purchase_date", "Extra · Data acquisto"),
]


def _asset_detail_source_choices() -> list[tuple[str, str]]:
    choices = list(ASSET_DETAIL_SOURCE_PRESETS)
    for field in AssetCustomField.objects.order_by("sort_order", "label", "id"):
        choices.append((f"custom:{field.code}", f"Campo custom · {field.label}"))
    for field in AssetCategoryField.objects.select_related("category").order_by(
        "category__sort_order",
        "category__label",
        "sort_order",
        "label",
        "id",
    ):
        choices.append((f"category:{field.code}", f"Campo categoria · {field.category.label} · {field.label}"))
    return choices


def _asset_detail_source_refs() -> set[str]:
    return {value for value, _label in _asset_detail_source_choices()}


def _default_asset_detail_field_seed_rows() -> list[dict[str, object]]:
    return [
        {
            "label": "Corse XYZ",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "computed:travel_xyz",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Anno macchina",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:year",
            "value_format": AssetDetailField.FORMAT_AUTO,
            "sort_order": 20,
        },
        {
            "label": "Configurazione",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "computed:machine_configuration",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Salute batteria",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "computed:battery_health",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Carico medio CPU",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "computed:cpu_load",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "Spazio libero",
            "section": AssetDetailField.SECTION_METRICS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "computed:storage_free",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Produttore",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "asset:manufacturer",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Modello",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "asset:model",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "Numero seriale",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "asset:serial_number",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Reparto",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "asset:reparto",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 40,
        },
        {
            "label": "Corsa X",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:x_mm",
            "value_format": AssetDetailField.FORMAT_MM,
            "sort_order": 50,
        },
        {
            "label": "Corsa Y",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:y_mm",
            "value_format": AssetDetailField.FORMAT_MM,
            "sort_order": 60,
        },
        {
            "label": "Corsa Z",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:z_mm",
            "value_format": AssetDetailField.FORMAT_MM,
            "sort_order": 70,
        },
        {
            "label": "Diametro",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:diameter_mm",
            "value_format": AssetDetailField.FORMAT_MM,
            "sort_order": 80,
        },
        {
            "label": "Mandrino",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:spindle_mm",
            "value_format": AssetDetailField.FORMAT_MM,
            "sort_order": 90,
        },
        {
            "label": "Anno",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:year",
            "value_format": AssetDetailField.FORMAT_AUTO,
            "sort_order": 100,
        },
        {
            "label": "TMC",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:tmc",
            "value_format": AssetDetailField.FORMAT_AUTO,
            "sort_order": 110,
        },
        {
            "label": "TCR",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:tcr_enabled",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 120,
        },
        {
            "label": "Pressione",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:pressure_bar",
            "value_format": AssetDetailField.FORMAT_BAR,
            "sort_order": 130,
        },
        {
            "label": "CNC",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:cnc_controlled",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 140,
        },
        {
            "label": "5 assi",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:five_axes",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 150,
        },
        {
            "label": "Accuracy from",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:accuracy_from",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 160,
        },
        {
            "label": "Prossima manutenzione",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:next_maintenance_date",
            "value_format": AssetDetailField.FORMAT_DATE,
            "sort_order": 170,
        },
        {
            "label": "Soglia reminder",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:maintenance_reminder_days",
            "value_format": AssetDetailField.FORMAT_AUTO,
            "sort_order": 180,
        },
        {
            "label": "Processore",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "it:cpu",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Numero seriale",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "asset:serial_number",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "Memoria",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "it:ram",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Sistema operativo",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "it:os",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 40,
        },
        {
            "label": "Archiviazione",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "it:disco",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 50,
        },
        {
            "label": "Grafica",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "extra:graphics",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 60,
        },
        {
            "label": "Schermo",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "extra:display",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 70,
        },
        {
            "label": "Data acquisto",
            "section": AssetDetailField.SECTION_SPECS,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "computed:purchase_date",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 80,
        },
        {
            "label": "Tag asset",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_ALL,
            "source_ref": "asset:asset_tag",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Reparto",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "asset:reparto",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "TCR",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:tcr_enabled",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 30,
        },
        {
            "label": "CNC",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:cnc_controlled",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 40,
        },
        {
            "label": "5 assi",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:five_axes",
            "value_format": AssetDetailField.FORMAT_BOOL,
            "sort_order": 50,
        },
        {
            "label": "Prossima manutenzione",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:next_maintenance_date",
            "value_format": AssetDetailField.FORMAT_DATE,
            "sort_order": 60,
        },
        {
            "label": "Soglia reminder",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:maintenance_reminder_days",
            "value_format": AssetDetailField.FORMAT_AUTO,
            "sort_order": 70,
        },
        {
            "label": "Accuracy from",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_WORK_MACHINE,
            "source_ref": "work_machine:accuracy_from",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 80,
        },
        {
            "label": "Produttore",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "asset:manufacturer",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "Modello",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "asset:model",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Ultimo sync",
            "section": AssetDetailField.SECTION_PROFILE,
            "asset_scope": AssetDetailField.SCOPE_STANDARD,
            "source_ref": "computed:sync_text",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 40,
        },
        {
            "label": "Reparto",
            "section": AssetDetailField.SECTION_ASSIGNMENT,
            "asset_scope": AssetDetailField.SCOPE_ALL,
            "source_ref": "asset:assignment_reparto",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 10,
        },
        {
            "label": "Posizione",
            "section": AssetDetailField.SECTION_ASSIGNMENT,
            "asset_scope": AssetDetailField.SCOPE_ALL,
            "source_ref": "asset:assignment_location",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 20,
        },
        {
            "label": "Assegnato a",
            "section": AssetDetailField.SECTION_ASSIGNMENT,
            "asset_scope": AssetDetailField.SCOPE_ALL,
            "source_ref": "asset:assignment_to",
            "value_format": AssetDetailField.FORMAT_TEXT,
            "sort_order": 30,
        },
        {
            "label": "Ultimo aggiornamento",
            "section": AssetDetailField.SECTION_ASSIGNMENT,
            "asset_scope": AssetDetailField.SCOPE_ALL,
            "source_ref": "asset:updated_at",
            "value_format": AssetDetailField.FORMAT_DATE,
            "sort_order": 40,
        },
    ]


def _seed_default_asset_detail_fields(*, create_only_if_empty: bool = True) -> int:
    if create_only_if_empty and AssetDetailField.objects.exists():
        return 0
    created = 0
    for row in _default_asset_detail_field_seed_rows():
        label = _clean_string(row.get("label"))
        code_seed = row.get("code") or f"{row.get('section')}-{row.get('asset_scope')}-{row.get('source_ref')}"
        defaults = {
            "section": row["section"],
            "asset_scope": row["asset_scope"],
            "source_ref": row["source_ref"],
            "value_format": row["value_format"],
            "sort_order": row["sort_order"],
            "show_if_empty": bool(row.get("show_if_empty", True)),
            "is_active": bool(row.get("is_active", True)),
        }
        code = _unique_detail_field_code(label, code_seed)
        _obj, created_flag = AssetDetailField.objects.get_or_create(
            code=code,
            defaults={"label": label[:120], **defaults},
        )
        if created_flag:
            created += 1
    return created


def _default_asset_detail_section_layout_rows() -> list[dict[str, object]]:
    return [
        {
            "code": AssetDetailSectionLayout.SECTION_SPECS,
            "grid_size": AssetDetailSectionLayout.SIZE_WIDE,
            "sort_order": 100,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_TIMELINE,
            "grid_size": AssetDetailSectionLayout.SIZE_WIDE,
            "sort_order": 110,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_MAINTENANCE,
            "grid_size": AssetDetailSectionLayout.SIZE_FULL,
            "sort_order": 120,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_TICKETS,
            "grid_size": AssetDetailSectionLayout.SIZE_WIDE,
            "sort_order": 130,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_PROFILE,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 200,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_ASSIGNMENT,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 210,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_LICENSES,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 215,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_PERIODIC,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 220,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_DOCUMENTS,
            "grid_size": AssetDetailSectionLayout.SIZE_WIDE,
            "sort_order": 230,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_QR,
            "grid_size": AssetDetailSectionLayout.SIZE_THIRD,
            "sort_order": 280,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_QUICK_ACTIONS,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 250,
            "is_visible": True,
        },
        {
            "code": AssetDetailSectionLayout.SECTION_MAP,
            "grid_size": AssetDetailSectionLayout.SIZE_HALF,
            "sort_order": 270,
            "is_visible": True,
        },
    ]


def _ensure_default_asset_detail_section_layouts() -> list[AssetDetailSectionLayout]:
    rows = _default_asset_detail_section_layout_rows()
    existing = {
        row.code: row
        for row in AssetDetailSectionLayout.objects.all()
    }
    for item in rows:
        code = str(item["code"])
        if code in existing:
            continue
        existing[code] = AssetDetailSectionLayout.objects.create(
            code=code,
            grid_size=str(item["grid_size"]),
            sort_order=int(item["sort_order"]),
            is_visible=bool(item["is_visible"]),
        )
    return list(AssetDetailSectionLayout.objects.order_by("sort_order", "id"))


def _detail_field_matches_asset_scope(detail_field: AssetDetailField, work_machine: WorkMachine | None) -> bool:
    if detail_field.asset_scope == AssetDetailField.SCOPE_ALL:
        return True
    if detail_field.asset_scope == AssetDetailField.SCOPE_WORK_MACHINE:
        return isinstance(work_machine, WorkMachine)
    if detail_field.asset_scope == AssetDetailField.SCOPE_STANDARD:
        return not isinstance(work_machine, WorkMachine)
    return False


def _resolve_asset_detail_source_value(
    *,
    source_ref: str,
    asset: Asset,
    it_details: AssetITDetails | None,
    work_machine: WorkMachine | None,
    extra: dict[str, object],
    custom_fields_by_code: dict[str, AssetCustomField],
    sync_text: str,
) -> object:
    source_kind, _, source_key = _clean_string(source_ref).partition(":")
    source_kind = source_kind.lower()
    source_key = _clean_string(source_key)
    category_values = extra.get("_category_fields", {}) if isinstance(extra.get("_category_fields"), dict) else {}
    if not source_kind or not source_key:
        return ""

    if source_kind == "asset":
        if source_key == "asset_category":
            return asset.category_label
        value = getattr(asset, source_key, "")
        if source_key == "asset_type" and value:
            return asset.get_asset_type_display()
        if source_key == "status" and value:
            return asset.get_status_display()
        return value
    if source_kind == "it":
        return getattr(it_details, source_key, "") if it_details is not None else ""
    if source_kind == "work_machine":
        return getattr(work_machine, source_key, "") if isinstance(work_machine, WorkMachine) else ""
    if source_kind == "extra":
        return extra.get(source_key, "")
    if source_kind == "custom":
        field = custom_fields_by_code.get(source_key)
        if field and field.label and field.label in extra:
            return extra.get(source_key, extra.get(field.label, ""))
        return extra.get(source_key, "")
    if source_kind == "category":
        return category_values.get(source_key, "")
    if source_kind != "computed":
        return ""

    if source_key == "travel_xyz":
        if not isinstance(work_machine, WorkMachine):
            return ""
        travel_parts = [str(value) for value in [work_machine.x_mm, work_machine.y_mm, work_machine.z_mm] if value is not None]
        return " x ".join(travel_parts) + " mm" if travel_parts else ""
    if source_key == "machine_configuration":
        if not isinstance(work_machine, WorkMachine):
            return ""
        machine_flags: list[str] = []
        if work_machine.tcr_enabled:
            machine_flags.append("TCR")
        if work_machine.cnc_controlled:
            machine_flags.append("CNC")
        if work_machine.five_axes:
            machine_flags.append("5 assi")
        return ", ".join(machine_flags) if machine_flags else "Standard"
    if source_key == "battery_health":
        return _coalesce_str(extra.get("battery_health"), extra.get("batteria"), "")
    if source_key == "cpu_load":
        return _coalesce_str(extra.get("avg_cpu_load"), extra.get("cpu_load"), "")
    if source_key == "storage_free":
        return _coalesce_str(extra.get("storage_free"), extra.get("free_storage"), getattr(it_details, "disco", ""), "")
    if source_key == "purchase_date":
        return _coalesce_str(
            asset.purchase_date.strftime("%d-%m-%Y") if asset.purchase_date else "",
            extra.get("purchase_date"),
            asset.created_at.strftime("%d-%m-%Y") if asset.created_at else "",
            "",
        )
    if source_key == "production_date":
        return _coalesce_str(
            asset.production_date.strftime("%d-%m-%Y") if asset.production_date else "",
            extra.get("production_date"),
            "",
        )
    if source_key == "sync_text":
        return sync_text
    return ""


def _format_asset_detail_value(value, value_format: str) -> str:
    if value_format == AssetDetailField.FORMAT_BOOL:
        if value in (None, ""):
            return "N/D"
        return "Si" if bool(value) else "No"
    if value_format == AssetDetailField.FORMAT_DATE:
        if isinstance(value, datetime):
            return timezone.localtime(value).strftime("%d-%m-%Y")
        if isinstance(value, date):
            return value.strftime("%d-%m-%Y")
        cleaned = _clean_string(value)
        return cleaned or "N/D"
    if value_format == AssetDetailField.FORMAT_MM:
        if value in (None, ""):
            return "N/D"
        return f"{value} mm"
    if value_format == AssetDetailField.FORMAT_BAR:
        if value in (None, ""):
            return "N/D"
        return f"{value} bar"

    if isinstance(value, bool):
        return "Si" if value else "No"
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    cleaned = _clean_string(str(value) if value not in (None, "") else "")
    return cleaned or "N/D"


def _is_empty_asset_detail_value(value) -> bool:
    if isinstance(value, bool):
        return False
    if value is None:
        return True
    if isinstance(value, (list, tuple, set, dict)):
        return not value
    cleaned = _clean_string(str(value))
    return cleaned.casefold() in {"", "n/d", "-"}


def _should_skip_asset_detail_row(section: str, formatted_value, show_if_empty: bool) -> bool:
    is_empty = _is_empty_asset_detail_value(formatted_value)
    if section == AssetDetailField.SECTION_SPECS:
        return is_empty
    return is_empty and not show_if_empty


def _build_configured_asset_detail_sections(
    *,
    asset: Asset,
    it_details: AssetITDetails | None,
    work_machine: WorkMachine | None,
    extra: dict[str, object],
    custom_fields_by_code: dict[str, AssetCustomField],
    sync_text: str,
) -> tuple[dict[str, list[dict[str, str]]], bool]:
    configured = list(AssetDetailField.objects.filter(is_active=True).order_by("section", "asset_scope", "sort_order", "id"))
    sections: dict[str, list[dict[str, str]]] = defaultdict(list)
    has_matching_config = False
    for detail_field in configured:
        if not _detail_field_matches_asset_scope(detail_field, work_machine):
            continue
        has_matching_config = True
        raw_value = _resolve_asset_detail_source_value(
            source_ref=detail_field.source_ref,
            asset=asset,
            it_details=it_details,
            work_machine=work_machine,
            extra=extra,
            custom_fields_by_code=custom_fields_by_code,
            sync_text=sync_text,
        )
        formatted_value = _format_asset_detail_value(raw_value, detail_field.value_format)
        if _should_skip_asset_detail_row(detail_field.section, formatted_value, detail_field.show_if_empty):
            continue
        sections[detail_field.section].append(
            {
                "label": detail_field.label,
                "value": formatted_value,
                "size": detail_field.card_size,
                "source_ref": detail_field.source_ref,
            }
        )
    return sections, has_matching_config


def _build_asset_category_detail_sections(asset: Asset, extra: dict[str, object]) -> dict[str, list[dict[str, str]]]:
    category = getattr(asset, "asset_category", None)
    if category is None:
        return {}
    category_values = extra.get("_category_fields", {})
    if not isinstance(category_values, dict):
        category_values = {}

    sections: dict[str, list[dict[str, str]]] = defaultdict(list)
    field_qs = category.category_fields.filter(is_active=True, show_in_detail=True).order_by("sort_order", "label", "id")
    for field_def in field_qs:
        raw_value = category_values.get(field_def.code, "")
        formatted_value = _format_asset_detail_value(raw_value, field_def.detail_value_format)
        if _should_skip_asset_detail_row(field_def.detail_section, formatted_value, field_def.show_if_empty):
            continue
        sections[field_def.detail_section].append(
            {
                "label": field_def.label,
                "value": formatted_value,
                "size": field_def.detail_card_size,
                "source_ref": f"category:{field_def.code}",
            }
        )
    return sections


def _preferred_asset_detail_metrics(detail_metrics: list[dict[str, object]]) -> list[dict[str, object]]:
    default_metric_signatures = {
        (_clean_string(row.get("label")), _clean_string(row.get("source_ref")))
        for row in _default_asset_detail_field_seed_rows()
        if row.get("section") == AssetDetailField.SECTION_METRICS
    }
    preferred_metrics: list[dict[str, object]] = []
    for metric in detail_metrics:
        label = _clean_string(metric.get("label"))
        source_ref = _clean_string(metric.get("source_ref"))
        size = _clean_string(metric.get("size")).upper()
        if not label:
            continue
        if source_ref.startswith("custom:") or source_ref.startswith("category:"):
            preferred_metrics.append(metric)
            continue
        if size in {AssetDetailField.CARD_HALF, AssetDetailField.CARD_WIDE, AssetDetailField.CARD_FULL}:
            preferred_metrics.append(metric)
            continue
        if (label, source_ref) not in default_metric_signatures:
            preferred_metrics.append(metric)
    return preferred_metrics


def _detail_grid_size_class(size_code: str) -> str:
    normalized = _clean_string(size_code).upper()
    if normalized == AssetDetailField.CARD_HALF:
        return "af-span-half"
    if normalized == AssetDetailField.CARD_WIDE:
        return "af-span-wide"
    if normalized == AssetDetailField.CARD_FULL:
        return "af-span-full"
    return "af-span-third"


def _resolve_sidebar_url(raw_url: str, rows: int = 25) -> str:
    target = _clean_string(raw_url).replace("{rows}", str(rows))
    if not target:
        return reverse("assets:asset_list")
    if target.startswith("django:"):
        route_expr = target.split("django:", 1)[1]
        route_name, _, query = route_expr.partition("?")
        try:
            base_url = reverse(route_name)
            return f"{base_url}?{query}" if query else base_url
        except NoReverseMatch:
            return "#"
    return target


def _is_sidebar_button_active(request: HttpRequest, button: AssetSidebarButton, resolved_url: str) -> bool:
    active_match = _clean_string(button.active_match)
    full_path = request.get_full_path()
    if active_match:
        if "=" in active_match and not active_match.startswith("/"):
            active_qs = parse_qs(active_match.lstrip("?"), keep_blank_values=True)
            if active_qs:
                return all(request.GET.get(key, "") in values for key, values in active_qs.items())
        return active_match in full_path

    parsed = urlsplit(resolved_url)
    if parsed.path and parsed.path != request.path:
        return False

    target_qs = parse_qs(parsed.query, keep_blank_values=True)
    if not target_qs:
        return parsed.path == request.path

    # Per la lista asset evitiamo che "Dashboard" risulti attivo insieme ai filtri tipo.
    try:
        asset_list_path = reverse("assets:asset_list")
    except NoReverseMatch:
        asset_list_path = "/assets/"
    if parsed.path == asset_list_path and "asset_type" not in target_qs and _clean_string(request.GET.get("asset_type")):
        return False
    if parsed.path == asset_list_path and "asset_category" not in target_qs and _clean_string(request.GET.get("asset_category")):
        return False
    # "Asset produzione" (?group=production) e' una lista a se': un target
    # asset_list senza group non deve risultare attivo quando il gruppo e' attivo.
    if parsed.path == asset_list_path and "group" not in target_qs and _clean_string(request.GET.get("group")):
        return False

    for key, values in target_qs.items():
        current_value = request.GET.get(key, "")
        if current_value not in values:
            return False
    return True


def _normalize_reports_scope(value: str | None) -> str:
    normalized = _clean_string(value).casefold()
    if normalized in {"it", "device", "devices", "dispositivi", "dispositivi_it"}:
        return "it"
    return "production"


def _reports_scope_context(scope: str) -> dict[str, object]:
    if scope == "it":
        return {
            "scope": "it",
            "title": "Report dispositivi IT",
            "subtitle": "KPI e interventi filtrati su PC, server, rete, fonia, TVCC e dispositivi IT.",
            "asset_types": IT_DEVICE_TYPES,
            "maintenance_month_enabled": False,
            "empty_month_message": "Il piano mese e il PDF mensile sono disponibili nei report asset produzione.",
        }
    return {
        "scope": "production",
        "title": "Report asset produzione",
        "subtitle": "KPI e interventi filtrati su CNC, macchine utensili e carroponti.",
        "asset_types": PRODUCTION_ASSET_TYPES,
        "maintenance_month_enabled": True,
        "empty_month_message": "",
    }


def _periodic_scope_context(scope: str) -> dict[str, object]:
    if scope == "it":
        return {
            "scope": "it",
            "title": "Manutenzione periodica dispositivi IT",
            "subtitle": "Piani ricorrenti filtrati su PC, server, rete, fonia, TVCC e dispositivi IT.",
            "asset_types": IT_DEVICE_TYPES,
        }
    return {
        "scope": "production",
        "title": "Manutenzione periodica asset produzione",
        "subtitle": "Piani ricorrenti filtrati su CNC, macchine utensili e carroponti.",
        "asset_types": PRODUCTION_ASSET_TYPES,
    }


def _default_sidebar_buttons(request: HttpRequest, rows: int = 25) -> list[dict]:
    base_list = reverse("assets:asset_list")
    asset_dashboard = reverse("assets:asset_dashboard")
    asset_components = reverse("assets:asset_component_list")
    asset_deadlines = reverse("assets:asset_administrative_deadline_list")
    maintenance_hub = reverse("assets:maintenance_hub")
    maintenance_templates = reverse("assets:maintenance_template_list")
    maintenance_rules = reverse("assets:maintenance_rule_list")
    maintenance_schedule = reverse("assets:maintenance_schedule")
    assistance_contracts = reverse("assets:assistance_contract_list")
    software_licenses = reverse("assets:software_license_list")
    production_assets_url = f"{reverse('assets:asset_list')}?group=production"
    work_machine_dashboard = reverse("assets:work_machine_dashboard")
    it_periodic_verifications = f"{reverse('assets:periodic_verifications')}?scope=it"
    production_periodic_verifications = f"{reverse('assets:periodic_verifications')}?scope=production"
    plant_layout_map = reverse("assets:plant_layout_map")
    reports = reverse("assets:reports")
    it_reports = f"{reports}?scope=it"
    production_reports = f"{reports}?scope=production"
    wo_list = reverse("assets:wo_list")
    asset_quick_report = reverse("assets:asset_quick_report")
    maintenance_todo = reverse("assets:maintenance_todo")
    il_mio_turno_url = reverse("assets:il_mio_turno")
    current_type = _clean_string(request.GET.get("asset_type"))
    current_group = _clean_string(request.GET.get("group")).lower()
    current_route = getattr(getattr(request, "resolver_match", None), "url_name", "")
    current_report_scope = _normalize_reports_scope(request.GET.get("scope")) if current_route == "reports" else ""
    current_periodic_scope = _normalize_reports_scope(request.GET.get("scope")) if current_route == "periodic_verifications" else ""
    return [
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Dashboard",
            "url": asset_dashboard,
            "is_subitem": False,
            "active": current_route == "asset_dashboard",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Il mio turno",
            "url": il_mio_turno_url,
            "is_subitem": False,
            "active": current_route == "il_mio_turno",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Cruscotto",
            "url": f"{base_list}?rows={rows}",
            "is_subitem": False,
            "active": not current_type and current_route == "asset_list",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Dispositivi",
            "url": f"{base_list}?asset_type=HW&rows={rows}",
            "is_subitem": False,
            "active": current_type == Asset.TYPE_HW and current_route == "asset_list",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Server",
            "url": f"{base_list}?asset_type={Asset.TYPE_SERVER}&rows={rows}",
            "is_subitem": True,
            "active": current_type == Asset.TYPE_SERVER and current_route == "asset_list",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Postazioni di lavoro",
            "url": f"{base_list}?asset_type={Asset.TYPE_PC}&rows={rows}",
            "is_subitem": True,
            "active": current_type in {Asset.TYPE_PC, Asset.TYPE_NOTEBOOK} and current_route == "asset_list",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Rete",
            "url": f"{base_list}?asset_type={Asset.TYPE_FIREWALL}&rows={rows}",
            "is_subitem": True,
            "active": current_type == Asset.TYPE_FIREWALL and current_route == "asset_list",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Scadenze amministrative",
            "url": asset_deadlines,
            "is_subitem": False,
            "active": current_route in {"asset_administrative_deadline_list", "asset_administrative_deadline_create", "asset_administrative_deadline_edit"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Componenti",
            "url": asset_components,
            "is_subitem": False,
            "active": current_route in {"asset_component_list", "asset_component_list_for_asset", "asset_component_create", "asset_component_edit"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Manutenzione",
            "url": maintenance_hub,
            "is_subitem": False,
            "active": current_route in {
                "maintenance_hub",
                "maintenance_template_list", "maintenance_template_create", "maintenance_template_edit",
                "maintenance_rule_list", "maintenance_rule_create", "maintenance_rule_edit",
                "maintenance_schedule", "maintenance_todo", "maintenance_history",
                "wo_list", "wo_view", "wo_create", "wo_close",
                "assistance_contract_list",
            },
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Pannello manutenzione",
            "url": maintenance_hub,
            "is_subitem": True,
            "active": current_route == "maintenance_hub",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Centro operativo",
            "url": maintenance_todo,
            "is_subitem": True,
            "active": current_route == "maintenance_todo",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Scadenzario",
            "url": maintenance_schedule,
            "is_subitem": True,
            "active": current_route == "maintenance_schedule",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Catalogo attivita",
            "url": maintenance_templates,
            "is_subitem": True,
            "active": current_route in {"maintenance_template_list", "maintenance_template_create", "maintenance_template_edit"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Piani ordinari",
            "url": maintenance_rules,
            "is_subitem": True,
            "active": current_route in {"maintenance_rule_list", "maintenance_rule_create", "maintenance_rule_edit"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Archivio piani IT",
            "url": it_periodic_verifications,
            "is_subitem": True,
            "active": current_route == "periodic_verifications" and current_periodic_scope == "it",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Report dispositivi IT",
            "url": it_reports,
            "is_subitem": True,
            "active": current_route == "reports" and current_report_scope == "it",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Macchine di lavoro",
            "url": production_assets_url,
            "is_subitem": False,
            "active": (current_route == "asset_list" and current_group == "production")
            or current_route in {"work_machine_list", "work_machine_create", "work_machine_edit"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Dashboard officina",
            "url": work_machine_dashboard,
            "is_subitem": True,
            "active": current_route == "work_machine_dashboard",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Mappa officina",
            "url": plant_layout_map,
            "is_subitem": True,
            "active": current_route in {"plant_layout_map", "plant_layout_editor"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Archivio piani produzione",
            "url": production_periodic_verifications,
            "is_subitem": True,
            "active": current_route == "periodic_verifications" and current_periodic_scope == "production",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Report asset produzione",
            "url": production_reports,
            "is_subitem": True,
            "active": current_route == "reports" and current_report_scope == "production",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Interventi",
            "url": wo_list,
            "is_subitem": True,
            "active": current_route in {"wo_list", "wo_view", "wo_create", "wo_close"},
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Segnala un problema",
            "url": asset_quick_report,
            "is_subitem": True,
            "active": current_route == "asset_quick_report",
        },
        {
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Licenze software",
            "url": software_licenses,
            "is_subitem": False,
            "active": current_route == "software_license_list",
        },
        {
            "section": AssetSidebarButton.SECTION_OPERATIONS,
            "label": "Prossime manutenzioni",
            "url": maintenance_schedule,
            "is_subitem": False,
            "active": current_route == "maintenance_schedule",
        },
        {
            "section": AssetSidebarButton.SECTION_OPERATIONS,
            "label": "Contratti assistenza",
            "url": assistance_contracts,
            "is_subitem": False,
            "active": current_route == "assistance_contract_list",
        },
        {
            "section": AssetSidebarButton.SECTION_ANALYTICS,
            "label": "Scadenzario operativo",
            "url": maintenance_schedule,
            "is_subitem": False,
            "active": current_route == "maintenance_schedule",
        },
    ]


def _default_sidebar_seed_rows() -> list[dict]:
    return [
        {
            "code": "dashboard",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Cruscotto",
            "target_url": "django:assets:asset_list?rows={rows}",
            "active_match": "",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 10,
            "is_visible": True,
        },
        {
            "code": "hardware",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Dispositivi",
            "target_url": "django:assets:asset_list?asset_type=HW&rows={rows}",
            "active_match": "asset_type=HW",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 20,
            "is_visible": True,
        },
        {
            "code": "servers",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Server",
            "target_url": "django:assets:asset_list?asset_type=SERVER&rows={rows}",
            "active_match": "asset_type=SERVER",
            "is_subitem": True,
            "parent_code": "hardware",
            "sort_order": 30,
            "is_visible": True,
        },
        {
            "code": "workstations",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Postazioni di lavoro",
            "target_url": "django:assets:asset_list?asset_type=PC&rows={rows}",
            "active_match": "asset_type=PC",
            "is_subitem": True,
            "parent_code": "hardware",
            "sort_order": 40,
            "is_visible": True,
        },
        {
            "code": "networking",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Rete",
            "target_url": "django:assets:asset_list?asset_type=FIREWALL&rows={rows}",
            "active_match": "asset_type=FIREWALL",
            "is_subitem": True,
            "parent_code": "hardware",
            "sort_order": 50,
            "is_visible": True,
        },
        {
            "code": "work_machines",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Macchine di lavoro",
            # Confluita nell'inventario unico: stessa pagina/lista, filtrata al
            # gruppo produzione (?group=production).
            "target_url": "django:assets:asset_list?group=production&rows={rows}",
            "active_match": "group=production",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 60,
            "is_visible": True,
        },
        {
            "code": "work_machines_dashboard",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Dashboard officina",
            "target_url": "django:assets:work_machine_dashboard",
            "active_match": "/assets/work-machines/dashboard/",
            "is_subitem": True,
            "parent_code": "work_machines",
            "sort_order": 61,
            "is_visible": True,
        },
        {
            "code": "periodic_verifications",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Manutenzione periodica produzione",
            "target_url": "django:assets:periodic_verifications?scope=production",
            "active_match": "",
            "is_subitem": True,
            "parent_code": "manutenzione_hub",
            "sort_order": 57,
            "is_visible": True,
        },
        {
            "code": "asset_deadlines",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Scadenze amministrative",
            "target_url": "django:assets:asset_administrative_deadline_list",
            "active_match": "/assets/scadenze/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 52,
            "is_visible": True,
        },
        {
            "code": "asset_components",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Componenti",
            "target_url": "django:assets:asset_component_list",
            "active_match": "/componenti/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 53,
            "is_visible": True,
        },
        {
            "code": "manutenzione_hub",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Manutenzione",
            "target_url": "django:assets:maintenance_hub",
            "active_match": "/assets/manutenzione/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 54,
            "is_visible": True,
        },
        {
            "code": "maintenance_templates",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Catalogo attivita",
            "target_url": "django:assets:maintenance_template_list",
            "active_match": "/assets/manutenzione/templates/",
            "is_subitem": True,
            "parent_code": "manutenzione_hub",
            "sort_order": 55,
            "is_visible": True,
        },
        {
            "code": "maintenance_rules",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Piani ordinari",
            "target_url": "django:assets:maintenance_rule_list",
            "active_match": "/assets/manutenzione/regole/",
            "is_subitem": True,
            "parent_code": "manutenzione_hub",
            "sort_order": 56,
            "is_visible": True,
        },
        {
            "code": "plant_layout_map",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Mappa officina",
            "target_url": "django:assets:plant_layout_map",
            "active_match": "/assets/work-machines/map/",
            "is_subitem": True,
            "parent_code": "work_machines",
            "sort_order": 61,
            "is_visible": True,
        },
        {
            "code": "software_licenses",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Licenze software",
            "target_url": "django:assets:software_license_list",
            "active_match": "/assets/licenze/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 63,
            "is_visible": True,
        },
        {
            "code": "maintenance_schedule",
            "section": AssetSidebarButton.SECTION_OPERATIONS,
            "label": "Prossime manutenzioni",
            "target_url": "django:assets:maintenance_schedule",
            "active_match": "/assets/manutenzione/prossime/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 66,
            "is_visible": True,
        },
        {
            "code": "assistance_contracts",
            "section": AssetSidebarButton.SECTION_OPERATIONS,
            "label": "Contratti assistenza",
            "target_url": "django:assets:assistance_contract_list",
            "active_match": "/assets/manutenzione/contratti/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 67,
            "is_visible": True,
        },
        {
            "code": "lifecycle_tracking",
            "section": AssetSidebarButton.SECTION_MAIN,
            "label": "Report asset produzione",
            "target_url": "django:assets:reports?scope=production",
            "active_match": "",
            "is_subitem": True,
            "parent_code": "manutenzione_hub",
            "sort_order": 58,
            "is_visible": True,
        },
        {
            "code": "compliance_reports",
            "section": AssetSidebarButton.SECTION_ANALYTICS,
            "label": "Scadenzario operativo",
            "target_url": "django:assets:maintenance_schedule",
            "active_match": "/assets/manutenzione/prossime/",
            "is_subitem": False,
            "parent_code": "",
            "sort_order": 80,
            "is_visible": True,
        },
    ]


def _create_default_sidebar_buttons() -> int:
    payload = _default_sidebar_seed_rows()
    created = 0
    created_by_code: dict[str, AssetSidebarButton] = {}
    for row in payload:
        button, _created = AssetSidebarButton.objects.get_or_create(
            code=row["code"],
            defaults={
                "section": row["section"],
                "label": row["label"],
                "target_url": row["target_url"],
                "active_match": row["active_match"],
                "is_subitem": row["is_subitem"],
                "sort_order": row["sort_order"],
                "is_visible": row["is_visible"],
            },
        )
        created_by_code[row["code"]] = button
        created += 1
    for row in payload:
        parent_code = _clean_string(row.get("parent_code"))
        if not parent_code:
            continue
        button = created_by_code.get(row["code"])
        parent = created_by_code.get(parent_code)
        if button is None or parent is None or button.parent_id == parent.id:
            continue
        button.parent = parent
        button.is_subitem = True
        button.section = parent.section
        button.save(update_fields=["parent", "is_subitem", "section", "updated_at"])
    return created


def _sidebar_input_suggestions() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    target_suggestions: list[dict[str, str]] = []
    active_match_suggestions: list[dict[str, str]] = []
    seen_targets: set[str] = set()
    seen_active_matches: set[str] = set()

    def add_target(value: str, label: str = "") -> None:
        normalized = _clean_string(value)
        if not normalized or normalized in seen_targets:
            return
        seen_targets.add(normalized)
        target_suggestions.append({"value": normalized, "label": label})

    def add_active_match(value: str, label: str = "") -> None:
        normalized = _clean_string(value)
        if not normalized or normalized in seen_active_matches:
            return
        seen_active_matches.add(normalized)
        active_match_suggestions.append({"value": normalized, "label": label})

    route_targets = [
        ("django:assets:asset_list?rows={rows}", "Cruscotto inventario"),
        ("django:assets:asset_component_list", "Componenti"),
        ("django:assets:asset_administrative_deadline_list", "Scadenze amministrative"),
        ("django:assets:maintenance_template_list", "Template manutenzione"),
        ("django:assets:maintenance_rule_list", "Regole manutenzione"),
        ("django:assets:work_machine_list", "Macchine di lavoro"),
        ("django:assets:work_machine_dashboard", "Dashboard officina"),
        ("django:assets:periodic_verifications?scope=it", "Manutenzione periodica dispositivi IT"),
        ("django:assets:periodic_verifications?scope=production", "Manutenzione periodica asset produzione"),
        ("django:assets:plant_layout_map", "Mappa officina"),
        ("django:assets:wo_list", "Interventi / Work order"),
        ("django:assets:software_license_list", "Licenze software"),
        ("django:assets:maintenance_schedule", "Prossime manutenzioni"),
        ("django:assets:assistance_contract_list", "Contratti assistenza"),
        ("django:assets:reports", "Report asset"),
        ("django:assets:reports?scope=it", "Report dispositivi IT"),
        ("django:assets:reports?scope=production", "Report asset produzione"),
        ("django:assets:gestione_admin", "Impostazioni assets"),
    ]
    for value, label in route_targets:
        add_target(value, label)

    for asset_type_code, asset_type_label in Asset.TYPE_CHOICES:
        add_target(
            f"django:assets:asset_list?asset_type={asset_type_code}&rows={{rows}}",
            f"Lista asset: {asset_type_label}",
        )
        add_active_match(f"asset_type={asset_type_code}", f"Filtro asset_type: {asset_type_label}")

    for category in AssetCategory.objects.filter(is_active=True).only("id", "label").order_by("sort_order", "label", "id"):
        add_target(
            f"django:assets:asset_list?asset_category={category.id}&rows={{rows}}",
            f"Lista categoria: {category.label}",
        )
        add_active_match(f"asset_category={category.id}", f"Filtro categoria: {category.label}")

    route_active_matches = [
        ("assets:asset_list", "Lista inventario"),
        ("assets:asset_component_list", "Componenti"),
        ("assets:asset_administrative_deadline_list", "Scadenze amministrative"),
        ("assets:maintenance_template_list", "Template manutenzione"),
        ("assets:maintenance_rule_list", "Regole manutenzione"),
        ("assets:work_machine_list", "Macchine di lavoro"),
        ("assets:work_machine_dashboard", "Dashboard officina"),
        ("assets:periodic_verifications", "Manutenzione periodica"),
        ("assets:plant_layout_map", "Mappa officina"),
        ("assets:wo_list", "Interventi / Work order"),
        ("assets:software_license_list", "Licenze software"),
        ("assets:maintenance_schedule", "Prossime manutenzioni"),
        ("assets:assistance_contract_list", "Contratti assistenza"),
        ("assets:reports", "Report asset"),
        ("assets:gestione_admin", "Impostazioni assets"),
    ]
    for route_name, label in route_active_matches:
        try:
            add_active_match(reverse(route_name), label)
        except NoReverseMatch:
            continue

    for value, label in [
        ("q=", "Filtro ricerca libera"),
        ("reparto=", "Filtro reparto"),
        ("vlan=", "Filtro VLAN"),
        ("ip=", "Filtro IP"),
        ("scope=it", "Report dispositivi IT"),
        ("scope=production", "Report asset produzione"),
        ("rows={rows}", "Placeholder righe per pagina"),
    ]:
        add_active_match(value, label)

    for button in AssetSidebarButton.objects.exclude(target_url="").only("target_url"):
        add_target(button.target_url, "Gia configurato")
    for button in AssetSidebarButton.objects.exclude(active_match="").only("active_match"):
        add_active_match(button.active_match, "Gia configurato")

    return target_suggestions, active_match_suggestions


def _dashboard_open_workorder_alert_rows(limit: int = 4) -> list[WorkOrder]:
    try:
        return list(
            WorkOrder.objects.select_related("asset")
            .only(
                "id",
                "asset_id",
                "status",
                "opened_at",
                "title",
                "description",
                "asset__id",
                "asset__asset_tag",
                "asset__name",
            )
            .filter(status=WorkOrder.STATUS_OPEN)
            .order_by("-opened_at")[:limit]
        )
    except DatabaseError:
        logger.warning("Impossibile caricare gli alert work order per la dashboard assets: schema WorkOrder non allineato.", exc_info=True)
        return []


def _sidebar_button_payload(
    request: HttpRequest,
    button: AssetSidebarButton,
    *,
    rows: int = 25,
    force_subitem: bool | None = None,
) -> dict[str, object]:
    url = _resolve_sidebar_url(button.target_url, rows=rows)
    return {
        "id": button.id,
        "label": _ui_label(button.label),
        "url": url,
        "is_subitem": button.is_subitem if force_subitem is None else force_subitem,
        "active": _is_sidebar_button_active(request, button, url),
    }


def _nest_sidebar_items(flat_items: list[dict]) -> list[dict]:
    nested: list[dict] = []
    current_parent: dict | None = None

    for item in flat_items:
        payload = dict(item)
        payload["children"] = []
        payload["has_children"] = False
        payload["expanded"] = bool(payload.get("active"))

        if payload.get("is_subitem") and current_parent is not None:
            current_parent["children"].append(payload)
            current_parent["has_children"] = True
            if payload.get("active"):
                current_parent["expanded"] = True
            continue

        nested.append(payload)
        current_parent = payload if not payload.get("is_subitem") else None

    return nested


def _build_sidebar_groups(request: HttpRequest, rows: int = 25) -> list[dict]:
    section_label = dict(AssetSidebarButton.SECTION_CHOICES)
    section_order = {
        AssetSidebarButton.SECTION_MAIN: 0,
        AssetSidebarButton.SECTION_ANALYTICS: 1,
        AssetSidebarButton.SECTION_OPERATIONS: 2,
    }
    configured = list(
        AssetSidebarButton.objects.filter(is_visible=True)
        .select_related("parent")
        .order_by("section", "sort_order", "id")
    )
    grouped: dict[str, list[dict]] = defaultdict(list)

    if configured:
        visible_by_id = {button.id: button for button in configured}
        roots_by_section: dict[str, list[AssetSidebarButton]] = defaultdict(list)
        children_by_parent: dict[int, list[AssetSidebarButton]] = defaultdict(list)

        for button in configured:
            if button.parent_id and button.parent_id in visible_by_id:
                children_by_parent[button.parent_id].append(button)
            else:
                roots_by_section[button.section].append(button)

        for section, root_buttons in roots_by_section.items():
            for button in root_buttons:
                root_payload = _sidebar_button_payload(request, button, rows=rows, force_subitem=button.is_subitem)
                root_payload["children"] = []
                root_payload["has_children"] = False
                root_payload["expanded"] = bool(root_payload.get("active"))
                grouped[section].append(root_payload)
                for child in children_by_parent.get(button.id, []):
                    child_payload = _sidebar_button_payload(request, child, rows=rows, force_subitem=True)
                    child_payload["children"] = []
                    child_payload["has_children"] = False
                    child_payload["expanded"] = bool(child_payload.get("active"))
                    root_payload["children"].append(child_payload)
                    root_payload["has_children"] = True
                    if child_payload.get("active"):
                        root_payload["expanded"] = True
    else:
        for payload in _default_sidebar_buttons(request, rows=rows):
            grouped[payload["section"]].append(payload)
        for section, items in list(grouped.items()):
            grouped[section] = _nest_sidebar_items(items)

    output = []
    for section, items in sorted(grouped.items(), key=lambda row: section_order.get(row[0], 99)):
        output.append({"section": section, "label": section_label.get(section, section), "items": items})
    return output


def _sidebar_parent_choices() -> list[AssetSidebarButton]:
    return list(
        AssetSidebarButton.objects.filter(parent__isnull=True)
        .order_by("section", "sort_order", "label", "id")
    )


def _category_sidebar_target(category_id: int) -> str:
    return _service_category_sidebar_target(category_id)


def _category_sidebar_active_match(category_id: int) -> str:
    return _service_category_sidebar_active_match(category_id)


def _category_inventory_url(category_id: int, rows: int = 25) -> str:
    return f"{reverse('assets:asset_list')}?asset_category={int(category_id)}&rows={int(rows)}"


def _category_subtree_ids(category) -> list[int]:
    """Ritorna gli id della categoria e di tutte le sue discendenti.

    L'albero categorie e' su due livelli (radice -> sotto-categorie); il terzo
    livello e' gestito per sicurezza nel caso venga introdotto in futuro.
    """
    ids = [category.id]
    child_ids = list(
        AssetCategory.objects.filter(parent_id=category.id).values_list("id", flat=True)
    )
    ids.extend(child_ids)
    if child_ids:
        ids.extend(
            AssetCategory.objects.filter(parent_id__in=child_ids).values_list("id", flat=True)
        )
    return ids


def _build_asset_category_admin_rows(categories: list[AssetCategory]) -> list[dict[str, object]]:
    category_ids = [category.id for category in categories]
    if not category_ids:
        return []

    asset_stats = {
        row["asset_category_id"]: row
        for row in (
            Asset.objects.filter(asset_category_id__in=category_ids)
            .values("asset_category_id")
            .annotate(
                total=Count("id"),
                in_use=Count("id", filter=Q(status=Asset.STATUS_IN_USE)),
                in_repair=Count("id", filter=Q(status=Asset.STATUS_IN_REPAIR)),
                last_updated=Max("updated_at"),
            )
            .order_by()
        )
    }
    field_stats = {
        row["category_id"]: row
        for row in (
            AssetCategoryField.objects.filter(category_id__in=category_ids)
            .values("category_id")
            .annotate(
                total=Count("id"),
                active=Count("id", filter=Q(is_active=True)),
            )
            .order_by()
        )
    }
    child_counts = {
        row["parent_id"]: row["total"]
        for row in AssetCategory.objects.filter(parent_id__in=category_ids).values("parent_id").annotate(total=Count("id")).order_by()
    }
    open_workorders = {
        row["asset__asset_category_id"]: row["total"]
        for row in (
            WorkOrder.objects.filter(asset__asset_category_id__in=category_ids, status=WorkOrder.STATUS_OPEN)
            .values("asset__asset_category_id")
            .annotate(total=Count("id"))
            .order_by()
        )
    }
    sample_assets_by_category: dict[int, list[Asset]] = defaultdict(list)
    for asset in (
        Asset.objects.filter(asset_category_id__in=category_ids)
        .select_related("asset_category")
        .order_by("asset_category_id", "-updated_at", "name", "asset_tag")
    ):
        bucket = sample_assets_by_category[asset.asset_category_id]
        if len(bucket) < 6:
            bucket.append(asset)

    sidebar_targets = {_category_sidebar_target(category_id) for category_id in category_ids}
    sidebar_matches = {_category_sidebar_active_match(category_id) for category_id in category_ids}
    sidebar_by_target = {
        row.target_url: row
        for row in AssetSidebarButton.objects.filter(target_url__in=sidebar_targets)
    }
    sidebar_by_match = {
        row.active_match: row
        for row in AssetSidebarButton.objects.filter(active_match__in=sidebar_matches)
    }
    type_labels = dict(Asset.TYPE_CHOICES)
    rows: list[dict[str, object]] = []
    for category in categories:
        stats = asset_stats.get(category.id, {})
        fields = field_stats.get(category.id, {})
        sidebar_target = _category_sidebar_target(category.id)
        sidebar_match = _category_sidebar_active_match(category.id)
        rows.append(
            {
                "category": category,
                "base_asset_type_label": type_labels.get(category.base_asset_type, category.base_asset_type),
                "asset_count": int(stats.get("total") or 0),
                "in_use_count": int(stats.get("in_use") or 0),
                "in_repair_count": int(stats.get("in_repair") or 0),
                "open_workorder_count": int(open_workorders.get(category.id) or 0),
                "field_count": int(fields.get("total") or 0),
                "active_field_count": int(fields.get("active") or 0),
                "child_count": int(child_counts.get(category.id) or 0),
                "last_updated": stats.get("last_updated"),
                "sample_assets": sample_assets_by_category.get(category.id, []),
                "inventory_url": _category_inventory_url(category.id),
                "sidebar_target_url": sidebar_target,
                "sidebar_active_match": sidebar_match,
                "sidebar_button": sidebar_by_target.get(sidebar_target) or sidebar_by_match.get(sidebar_match),
            }
        )
    return rows


def _header_tool_visibility(is_admin: bool) -> dict[str, bool]:
    """Restituisce can_hdr_<code> per ogni strumento header in base ai settings DB."""
    tools = {t.code: t for t in AssetHeaderTool.objects.all()}

    def _visible(code: str) -> bool:
        t = tools.get(code)
        if t is None:
            return True  # default: visibile se non ancora in DB
        if not t.is_active:
            return False
        if t.admin_only and not is_admin:
            return False
        return True

    return {
        "can_hdr_avvisi": _visible(AssetHeaderTool.TOOL_AVVISI),
        "can_hdr_widget": _visible(AssetHeaderTool.TOOL_WIDGET),
        "can_hdr_sync": _visible(AssetHeaderTool.TOOL_SYNC),
    }


def _handle_header_tool_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    if action == "update_header_tool":
        tool_id = _as_int(request.POST.get("tool_id"), default=0)
        tool = AssetHeaderTool.objects.filter(pk=tool_id).first()
        if not tool:
            return False, "Strumento non trovato."
        tool.is_active = request.POST.get("is_active") == "1"
        tool.admin_only = request.POST.get("admin_only") == "1"
        tool.save(update_fields=["is_active", "admin_only"])
        return True, f"Strumento «{tool.label}» aggiornato."
    return False, "Azione non riconosciuta."


def _handle_sidebar_button_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_sections = {key for key, _ in AssetSidebarButton.SECTION_CHOICES}

    if action == "seed_sidebar_buttons":
        if AssetSidebarButton.objects.exists():
            return False, "Menu sidebar già configurato."
        payload = _default_sidebar_seed_rows()
        created = 0
        created_by_code: dict[str, AssetSidebarButton] = {}
        for row in payload:
            button, _created = AssetSidebarButton.objects.get_or_create(
                code=row["code"],
                defaults={
                    "section": row["section"],
                    "label": row["label"],
                    "target_url": row["target_url"],
                    "active_match": row["active_match"],
                    "is_subitem": row["is_subitem"],
                    "sort_order": row["sort_order"],
                    "is_visible": row["is_visible"],
                },
            )
            created_by_code[row["code"]] = button
            created += 1
        for row in payload:
            parent_code = _clean_string(row.get("parent_code"))
            if not parent_code:
                continue
            button = created_by_code.get(row["code"])
            parent = created_by_code.get(parent_code)
            if button is None or parent is None or button.parent_id == parent.id:
                continue
            button.parent = parent
            button.is_subitem = True
            button.section = parent.section
            button.save(update_fields=["parent", "is_subitem", "section", "updated_at"])
        return True, f"Menu sidebar inizializzato ({created} voci)."

    if action == "create_sidebar_button":
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci etichetta menu."
        parent_id = _as_int(
            request.POST.get("parent_sidebar_button_id") or request.POST.get("parent_id"),
            default=0,
        )
        parent_button = AssetSidebarButton.objects.filter(pk=parent_id).first() if parent_id else None
        if parent_button and parent_button.parent_id:
            return False, "La voce padre deve essere di primo livello."
        section = _clean_string(request.POST.get("section")) or AssetSidebarButton.SECTION_MAIN
        if section not in valid_sections:
            section = AssetSidebarButton.SECTION_MAIN
        if parent_button is not None:
            section = parent_button.section
        code = _unique_sidebar_button_code(label, request.POST.get("code"))
        AssetSidebarButton.objects.create(
            code=code,
            section=section,
            parent=parent_button,
            label=label[:120],
            target_url=_clean_string(request.POST.get("target_url")),
            active_match=_clean_string(request.POST.get("active_match")),
            is_subitem=True if parent_button is not None else bool(request.POST.get("is_subitem")),
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            is_visible=bool(request.POST.get("is_visible")),
        )
        return True, f"Voce menu \"{label}\" creata."

    if action == "update_sidebar_button":
        button_id = _as_int(request.POST.get("sidebar_button_id"), default=0)
        button = AssetSidebarButton.objects.filter(pk=button_id).first()
        if not button:
            return False, "Voce menu non trovata."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Etichetta menu obbligatoria."
        parent_id = _as_int(
            request.POST.get("parent_sidebar_button_id") or request.POST.get("parent_id"),
            default=0,
        )
        parent_button = AssetSidebarButton.objects.filter(pk=parent_id).first() if parent_id else None
        if parent_button and parent_button.id == button.id:
            return False, "Una voce non puo essere padre di se stessa."
        if parent_button and parent_button.parent_id:
            return False, "La voce padre deve essere di primo livello."
        section = _clean_string(request.POST.get("section")) or button.section
        if section not in valid_sections:
            section = button.section
        if parent_button is not None:
            section = parent_button.section
        button.section = section
        button.parent = parent_button
        button.label = label[:120]
        button.target_url = _clean_string(request.POST.get("target_url"))
        button.active_match = _clean_string(request.POST.get("active_match"))
        button.is_subitem = True if parent_button is not None else bool(request.POST.get("is_subitem"))
        button.sort_order = _as_int(request.POST.get("sort_order"), default=button.sort_order)
        button.is_visible = bool(request.POST.get("is_visible"))
        button.save(
            update_fields=[
                "section",
                "parent",
                "label",
                "target_url",
                "active_match",
                "is_subitem",
                "sort_order",
                "is_visible",
                "updated_at",
            ]
        )
        return True, f"Voce menu \"{button.label}\" aggiornata."

    if action == "delete_sidebar_button":
        button_id = _as_int(request.POST.get("sidebar_button_id"), default=0)
        button = AssetSidebarButton.objects.filter(pk=button_id).first()
        if not button:
            return False, "Voce menu non trovata."
        label = button.label
        button.delete()
        return True, f"Voce menu \"{label}\" eliminata."

    if action == "reset_sidebar_buttons":
        deleted, _ = AssetSidebarButton.objects.all().delete()
        payload = _default_sidebar_seed_rows()
        created = 0
        created_by_code: dict[str, AssetSidebarButton] = {}
        for row in payload:
            button, _ = AssetSidebarButton.objects.get_or_create(
                code=row["code"],
                defaults={
                    "section": row["section"],
                    "label": row["label"],
                    "target_url": row["target_url"],
                    "active_match": row["active_match"],
                    "is_subitem": row["is_subitem"],
                    "sort_order": row["sort_order"],
                    "is_visible": row["is_visible"],
                },
            )
            created_by_code[row["code"]] = button
            created += 1
        for row in payload:
            parent_code = _clean_string(row.get("parent_code"))
            if not parent_code:
                continue
            btn = created_by_code.get(row["code"])
            parent = created_by_code.get(parent_code)
            if btn is None or parent is None or btn.parent_id == parent.id:
                continue
            btn.parent = parent
            btn.is_subitem = True
            btn.section = parent.section
            btn.save(update_fields=["parent", "is_subitem", "section", "updated_at"])
        return True, f"Menu sidebar reimpostato ai valori predefiniti ({created} voci, {deleted} eliminate)."

    if action == "clear_sidebar_buttons":
        deleted, _ = AssetSidebarButton.objects.all().delete()
        return True, f"Menu sidebar svuotato ({deleted} voci eliminate). Il portale usa ora il menu predefinito aggiornato."

    return False, "Azione menu non valida."


def _assets_shell_context(
    request: HttpRequest,
    *,
    rows: int = 25,
    search_action: str | None = None,
    new_url: str | None = None,
    new_label: str | None = None,
    search_placeholder: str | None = None,
) -> dict[str, object]:
    logo_url = resolve_module_logo("assets", legacy_logo_keys=("assets_logo_image",)) or None
    display_label = resolve_module_label("assets", fallback="Assets", surface="display")
    return {
        "assets_sidebar_groups": _build_sidebar_groups(request, rows=rows),
        "assets_section_nav": _assets_section_nav(request),
        "assets_shell_search_action": search_action or reverse("assets:asset_list"),
        "assets_shell_new_url": new_url or reverse("assets:asset_create"),
        "assets_shell_new_label": new_label or "+ Nuovo asset",
        "assets_shell_search_placeholder": search_placeholder or "Ricerca rapida per asset, seriali o utenti (Ctrl + K)",
        "can_gestione_admin": user_can_modulo_action(request, "assets", "admin_assets"),
        "assets_logo_url": logo_url,
        "assets_brand_label": display_label,
    }


def _assets_section_nav(request: HttpRequest) -> dict[str, object] | None:
    current_route = _clean_string(getattr(getattr(request, "resolver_match", None), "url_name", ""))
    if not current_route:
        return None

    route_to_item = {
        "maintenance_hub": "todo",
        "maintenance_todo": "todo",
        "maintenance_schedule": "schedule",
        "maintenance_scadenzario": "schedule",
        "maintenance_history": "history",
        "wo_list": "workorders",
        "wo_view": "workorders",
        "wo_create": "workorders",
        "wo_close": "workorders",
        "reports": "reports",
        "report_template_admin": "report_templates",
        "maintenance_impostazioni": "settings",
        "maintenance_suppliers": "suppliers",
        "maintenance_template_list": "settings",
        "maintenance_template_create": "settings",
        "maintenance_template_edit": "settings",
        "maintenance_rule_list": "settings",
        "maintenance_rule_create": "settings",
        "maintenance_rule_edit": "settings",
        "asset_maintenance_rule_list": "settings",
        "asset_maintenance_rule_override_create": "settings",
        "asset_maintenance_rule_override_edit": "settings",
        "asset_maintenance_rule_override_reset": "settings",
        "periodic_verifications": "settings",
        "assistance_contract_list": "settings",
    }
    active_item = route_to_item.get(current_route)
    if current_route == "wo_list" and (
        _clean_string(request.GET.get("view")).lower() == "closed"
        or _clean_string(request.GET.get("status")).upper() in {WorkOrder.STATUS_DONE, WorkOrder.STATUS_CANCELED}
    ):
        active_item = "history"
    if active_item == "report_templates":
        active_item = "reports"
    if active_item is None:
        return None

    report_scope = _normalize_reports_scope(request.GET.get("scope"))
    workorders_url = reverse("assets:wo_list")
    settings_url = reverse("assets:maintenance_impostazioni")
    items = [
        {
            "key": "todo",
            "label": "Oggi",
            "url": reverse("assets:maintenance_hub"),
        },
        {
            "key": "schedule",
            "label": "Scadenzario",
            "url": reverse("assets:maintenance_schedule"),
        },
        {
            "key": "workorders",
            "label": "Interventi",
            "url": workorders_url,
        },
        {
            "key": "history",
            "label": "Storico",
            "url": reverse("assets:maintenance_history"),
        },
        {
            "key": "settings",
            "label": "Catalogo e piani",
            "url": settings_url,
        },
        {
            "key": "reports",
            "label": "Report",
            "url": f"{reverse('assets:reports')}?scope={report_scope}",
        },
        {
            "key": "suppliers",
            "label": "Fornitori",
            "url": reverse("assets:maintenance_suppliers"),
        },
    ]
    for item in items:
        item["active"] = item["key"] == active_item

    active_label = next((item["label"] for item in items if item["active"]), "Manutenzione")
    return {
        "label": "Manutenzione",
        "active_label": active_label,
        "items": items,
        "breadcrumbs": [
            {"label": "Assets", "url": reverse("assets:asset_dashboard")},
            {"label": "Manutenzione", "url": reverse("assets:maintenance_hub")},
            {"label": active_label, "url": ""},
        ],
        "actions": [
            {
                "key": "new-workorder",
                "label": "+ Nuovo intervento",
                "url": f"{workorders_url}?create=1",
                "kind": "primary",
            },
            {
                "key": "export-workorders",
                "label": "Esporta OdL",
                "url": f"{workorders_url}?export=1",
                "kind": "secondary",
            },
            {
                "key": "new-plan",
                "label": "+ Nuovo piano",
                "url": reverse("assets:maintenance_rule_create"),
                "kind": "secondary",
            },
        ],
    }


def _safe_editor_json_rows(raw_value) -> list[dict[str, object]]:
    if not raw_value:
        return []
    try:
        payload = json.loads(raw_value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    if not isinstance(payload, list):
        return []
    return [row for row in payload if isinstance(row, dict)]


def _plant_layout_editor_area_rows(layout: PlantLayout | None) -> list[dict[str, object]]:
    if layout is None:
        return []
    return [
        {
            "id": area.id,
            "name": area.name,
            "reparto_code": area.reparto_code,
            "color": area.color,
            "notes": area.notes,
            "x_percent": float(area.x_percent),
            "y_percent": float(area.y_percent),
            "width_percent": float(area.width_percent),
            "height_percent": float(area.height_percent),
            "sort_order": area.sort_order,
        }
        for area in layout.areas.all().order_by("sort_order", "id")
    ]


def _plant_layout_editor_marker_rows(layout: PlantLayout | None) -> list[dict[str, object]]:
    if layout is None:
        return []
    return [
        {
            "id": marker.id,
            "asset_id": marker.asset_id,
            "label": marker.label,
            "x_percent": float(marker.x_percent),
            "y_percent": float(marker.y_percent),
            "sort_order": marker.sort_order,
        }
        for marker in layout.markers.select_related("asset").all().order_by("sort_order", "id")
    ]


# Parent code delle AssetCategory considerate IT/ufficio: gli asset sotto questi
# rami (o con asset_type IT) NON sono posizionabili in planimetria. Tutto il resto
# (gru, impianti a pressione, macchine NO CE, Kardex, attrezzature varie...) e'
# inseribile. Filtro per esclusione cosi nuove categorie di produzione/officina
# restano incluse automaticamente.
_PLANT_LAYOUT_EXCLUDED_CATEGORY_PARENTS = ("information-technology", "physical-security")


def _plant_layout_machine_queryset():
    excluded_category_ids = list(
        AssetCategory.objects.filter(
            Q(code__in=_PLANT_LAYOUT_EXCLUDED_CATEGORY_PARENTS)
            | Q(code__startswith="information-technology")
            | Q(code__startswith="physical-security")
        ).values_list("id", flat=True)
    )
    return (
        Asset.objects.exclude(asset_type__in=IT_DEVICE_TYPES)
        .exclude(asset_category_id__in=excluded_category_ids)
        .select_related("work_machine")
        .order_by("reparto", "name", "asset_tag", "id")
    )


def _plant_layout_machine_catalog() -> list[dict[str, object]]:
    machines = _plant_layout_machine_queryset()
    catalog: list[dict[str, object]] = []
    for asset in machines:
        machine = getattr(asset, "work_machine", None)
        catalog.append(
            {
                "id": asset.id,
                "asset_tag": asset.asset_tag,
                "internal_number": _clean_string(asset.internal_number),
                "name": asset.name,
                "reparto": _clean_string(asset.reparto),
                "status": asset.get_status_display(),
                "status_code": asset.status,
                "location": _clean_string(asset.assignment_location),
                "manufacturer": _clean_string(asset.manufacturer),
                "model": _clean_string(asset.model),
                "detail_url": reverse("assets:asset_view", kwargs={"id": asset.id}),
                "next_maintenance_date": (
                    machine.next_maintenance_date.strftime("%d-%m-%Y")
                    if isinstance(machine, WorkMachine) and machine.next_maintenance_date
                    else ""
                ),
                "cnc_controlled": bool(getattr(machine, "cnc_controlled", False)),
                "five_axes": bool(getattr(machine, "five_axes", False)),
            }
        )
    return catalog


def _plant_layout_queryset():
    return PlantLayout.objects.prefetch_related("areas", "markers", "markers__asset", "markers__asset__work_machine")


def _preferred_plant_layout_category(
    active_layouts: list[PlantLayout],
    *,
    requested_category: str = "",
    fallback_category: str = PlantLayout.DEFAULT_CATEGORY,
) -> str:
    requested = _clean_string(requested_category)
    if requested:
        for layout in active_layouts:
            if _clean_string(layout.category).casefold() == requested.casefold():
                return layout.category
    for layout in active_layouts:
        if _clean_string(layout.category).casefold() == _clean_string(fallback_category).casefold():
            return layout.category
    return active_layouts[0].category if active_layouts else ""


def _plant_layout_category_switches(
    *,
    active_layouts: list[PlantLayout],
    selected_category: str,
    focus_asset_id: int = 0,
) -> list[dict[str, object]]:
    try:
        base_url = reverse("assets:plant_layout_map")
    except NoReverseMatch:
        base_url = "/assets/work-machines/map/"
    switches: list[dict[str, object]] = []
    seen_categories: set[str] = set()
    for layout in active_layouts:
        category_key = _clean_string(layout.category).casefold()
        if category_key in seen_categories:
            continue
        seen_categories.add(category_key)
        params = [f"category={quote(layout.category)}"]
        if focus_asset_id:
            params.append(f"asset={focus_asset_id}")
        switches.append(
            {
                "category": layout.category,
                "active": _clean_string(layout.category).casefold() == _clean_string(selected_category).casefold(),
                "url": f"{base_url}?{'&'.join(params)}",
                "layout_name": layout.name,
            }
        )
    return switches


def _hex_to_rgba(hex_color: str, alpha: float) -> str:
    """Convert #RRGGBB to rgba(r,g,b,alpha) for CSS compatibility without color-mix()."""
    h = str(hex_color or "").strip().lstrip("#")
    try:
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except (ValueError, IndexError):
        r, g, b = 37, 99, 235  # fallback #2563EB
    return f"rgba({r},{g},{b},{alpha})"


def _open_tickets_by_asset(asset_ids) -> dict[int, list[dict[str, object]]]:
    """Mappa asset_id -> lista ticket aperti (una sola query batch, difensiva).

    Usata per l'overlay ticket sulla planimetria (marker in rosso quando l'asset
    ha ticket aperti). Stati aperti = APERTA/IN_CARICO/IN_ATTESA.
    """
    result: dict[int, list[dict[str, object]]] = defaultdict(list)
    ids = [i for i in (asset_ids or []) if i]
    if not ids:
        return result
    try:
        from tickets.models import StatoTicket, Ticket

        open_states = (StatoTicket.APERTA, StatoTicket.IN_CARICO, StatoTicket.IN_ATTESA)
        rows = (
            Ticket.objects.filter(asset_id__in=ids, stato__in=open_states)
            .order_by("-created_at", "-id")
            .values("id", "asset_id", "numero_ticket", "titolo", "stato", "priorita")
        )
        label_map = dict(StatoTicket.choices)
        for row in rows:
            try:
                url = reverse("tickets:detail", args=[row["id"]])
            except Exception:
                url = ""
            result[row["asset_id"]].append({
                "id": row["id"],
                "numero": row["numero_ticket"] or f"T-{row['id']}",
                "titolo": row["titolo"] or "Ticket",
                "stato": label_map.get(row["stato"], row["stato"]),
                "priorita": row["priorita"] or "",
                "url": url,
            })
    except Exception:
        return result
    return result


def _plant_layout_public_payload(layout: PlantLayout | None) -> dict[str, object]:
    if layout is None:
        return {"layout": None, "areas": [], "markers": [], "machine_catalog": [], "markers_with_tickets": 0}

    machine_catalog = _plant_layout_machine_catalog()
    machines_by_id = {row["id"]: row for row in machine_catalog}
    reparto_machine_rows: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in machine_catalog:
        reparto_machine_rows[_clean_string(str(row.get("reparto") or ""))].append(dict(row))

    reparto_area_ids: dict[str, list[int]] = defaultdict(list)
    area_payload: list[dict[str, object]] = []
    for area in layout.areas.all().order_by("sort_order", "id"):
        reparto_code = _clean_string(area.reparto_code)
        reparto_area_ids[reparto_code].append(area.id)
        area_payload.append(
            {
                "id": area.id,
                "name": area.name,
                "reparto_code": reparto_code,
                "color": area.color,
                "bg_color": _hex_to_rgba(area.color, 0.22),
                "active_bg_color": _hex_to_rgba(area.color, 0.40),
                "notes": area.notes,
                "x_percent": float(area.x_percent),
                "y_percent": float(area.y_percent),
                "width_percent": float(area.width_percent),
                "height_percent": float(area.height_percent),
                "machine_count": len(reparto_machine_rows.get(reparto_code, [])),
                "machines": list(reparto_machine_rows.get(reparto_code, [])),
            }
        )

    marker_payload: list[dict[str, object]] = []
    for marker in layout.markers.select_related("asset", "asset__work_machine").all().order_by("sort_order", "id"):
        asset_payload = dict(machines_by_id.get(marker.asset_id) or {})
        asset_payload["marker_id"] = marker.id
        marker_payload.append(
            {
                "id": marker.id,
                "asset_id": marker.asset_id,
                "label": marker.label or asset_payload.get("asset_tag") or asset_payload.get("name") or f"Marker {marker.id}",
                "x_percent": float(marker.x_percent),
                "y_percent": float(marker.y_percent),
                "area_ids": list(reparto_area_ids.get(_clean_string(str(asset_payload.get("reparto") or "")), [])),
                "machine": asset_payload,
            }
        )

    # Overlay ticket aperti (#5): arricchisce ogni marker con i ticket aperti
    # dell'asset (una sola query batch) per evidenziarlo in rosso sulla mappa.
    ticket_map = _open_tickets_by_asset({m["asset_id"] for m in marker_payload if m.get("asset_id")})
    markers_with_tickets = 0
    for m in marker_payload:
        tickets = ticket_map.get(m["asset_id"], [])
        m["open_tickets"] = len(tickets)
        m["tickets"] = tickets
        if tickets:
            markers_with_tickets += 1

    return {
        "layout": {
            "id": layout.id,
            "name": layout.name,
            "description": layout.description,
            "image_url": layout.image.url if layout.image else "",
            "is_active": layout.is_active,
        },
        "areas": area_payload,
        "markers": marker_payload,
        "machine_catalog": machine_catalog,
        "markers_with_tickets": markers_with_tickets,
    }


def _asset_edit_route_name(asset: Asset) -> str:
    if asset.asset_type == Asset.TYPE_WORK_MACHINE:
        return "assets:work_machine_edit"
    return "assets:asset_edit"


def _resolve_button_target(button: AssetActionButton, asset: Asset) -> str:
    target = _clean_string(button.target)
    replacements = {
        "{asset_id}": str(asset.id),
        "{asset_tag}": asset.asset_tag or "",
        "{asset_name}": asset.name or "",
        "{asset_type}": asset.asset_type or "",
        "{assigned_user_id}": str(asset.assigned_legacy_user_id or ""),
    }
    for key, value in replacements.items():
        target = target.replace(key, value)
    return target


def _default_action_buttons(asset: Asset, *, create_workorder_url: str = "") -> dict[str, list[dict]]:
    try:
        edit_url = reverse(_asset_edit_route_name(asset), kwargs={"id": asset.id})
    except NoReverseMatch:
        edit_url = ""
    try:
        assign_url = reverse("assets:asset_assign", kwargs={"id": asset.id})
    except NoReverseMatch:
        assign_url = ""
    try:
        wo_url = create_workorder_url or reverse("assets:wo_create", kwargs={"id": asset.id})
    except NoReverseMatch:
        wo_url = create_workorder_url
    try:
        refresh_url = reverse("assets:asset_view", kwargs={"id": asset.id})
    except NoReverseMatch:
        refresh_url = ""
    try:
        qr_url = reverse("assets:asset_qr_label", kwargs={"id": asset.id})
    except NoReverseMatch:
        qr_url = ""
    try:
        qr_landing_url = reverse("assets:asset_qr_landing", kwargs={"asset_tag": asset.asset_tag})
    except NoReverseMatch:
        qr_landing_url = ""

    return {
        AssetActionButton.ZONE_HEADER: [
            {"label": "Etichetta QR", "style": AssetActionButton.STYLE_DEFAULT, "href": qr_url, "data_action": "", "new_tab": True},
            {"label": "Modifica dettagli", "style": AssetActionButton.STYLE_PRIMARY, "href": edit_url, "data_action": "", "new_tab": False},
        ],
        AssetActionButton.ZONE_QUICK: [
            {"label": "Riassegna", "style": AssetActionButton.STYLE_DEFAULT, "href": assign_url, "data_action": "", "new_tab": False},
            {"label": "Crea intervento", "style": AssetActionButton.STYLE_DEFAULT, "href": wo_url, "data_action": "", "new_tab": False},
            {"label": "Vista QR mobile", "style": AssetActionButton.STYLE_SECONDARY, "href": qr_landing_url, "data_action": "", "new_tab": False},
            {"label": "Aggiorna dati", "style": AssetActionButton.STYLE_SECONDARY, "href": refresh_url, "data_action": "", "new_tab": False},
            {"label": "Dismetti bene", "style": AssetActionButton.STYLE_DANGER, "href": edit_url, "data_action": "", "new_tab": False},
        ],
    }


def _system_action_buttons_for_asset(asset: Asset) -> dict[str, list[dict]]:
    buttons = {
        AssetActionButton.ZONE_HEADER: [],
        AssetActionButton.ZONE_QUICK: [],
    }
    try:
        buttons[AssetActionButton.ZONE_HEADER].append(
            {
                "label": "Etichetta QR",
                "style": AssetActionButton.STYLE_SECONDARY,
                "href": reverse("assets:asset_qr_label", kwargs={"id": asset.id}),
                "data_action": "",
                "new_tab": True,
            }
        )
    except NoReverseMatch:
        pass
    return buttons


def _append_unique_action_buttons(base: dict[str, list[dict]], extra: dict[str, list[dict]]) -> dict[str, list[dict]]:
    for zone, buttons in extra.items():
        seen = {
            (
                _clean_string(button.get("label")).lower(),
                _clean_string(button.get("href")),
                _clean_string(button.get("data_action")).lower(),
            )
            for button in base.get(zone, [])
        }
        for button in buttons:
            signature = (
                _clean_string(button.get("label")).lower(),
                _clean_string(button.get("href")),
                _clean_string(button.get("data_action")).lower(),
            )
            if signature in seen:
                continue
            base.setdefault(zone, []).append(button)
            seen.add(signature)
    return base


def _build_action_buttons_for_asset(asset: Asset, *, create_workorder_url: str = "") -> dict[str, list[dict]]:
    configured = list(AssetActionButton.objects.filter(is_active=True).order_by("zone", "sort_order", "id"))
    defaults = _default_action_buttons(asset, create_workorder_url=create_workorder_url)
    output: dict[str, list[dict]] = {
        AssetActionButton.ZONE_HEADER: [],
        AssetActionButton.ZONE_QUICK: [],
    }
    if not configured:
        return _append_unique_action_buttons(defaults, _system_action_buttons_for_asset(asset))

    for button in configured:
        if button.zone not in output:
            continue
        payload = {
            "label": _ui_label(button.label),
            "style": button.style,
            "href": "",
            "data_action": "",
            "new_tab": bool(button.open_in_new_tab),
        }
        if button.action_type == AssetActionButton.TYPE_PRINT:
            payload["data_action"] = "print"
        elif button.action_type == AssetActionButton.TYPE_REFRESH:
            payload["data_action"] = "refresh"
        else:
            payload["href"] = _resolve_button_target(button, asset)
        output[button.zone].append(payload)

    if not output[AssetActionButton.ZONE_HEADER]:
        output[AssetActionButton.ZONE_HEADER] = defaults[AssetActionButton.ZONE_HEADER]
    if not output[AssetActionButton.ZONE_QUICK]:
        output[AssetActionButton.ZONE_QUICK] = defaults[AssetActionButton.ZONE_QUICK]
    return _append_unique_action_buttons(output, _system_action_buttons_for_asset(asset))


def _match_action_button(button: dict, *, href: str = "", label: str = "") -> bool:
    href_value = _clean_string(href)
    label_value = _clean_string(label).casefold()
    button_href = _clean_string(button.get("href"))
    button_label = _clean_string(button.get("label")).casefold()
    if href_value and button_href == href_value:
        return True
    if label_value and button_label == label_value:
        return True
    return False


def _promote_asset_detail_actions(
    buttons_by_zone: dict[str, list[dict]],
    *,
    assign_url: str,
    qr_url: str,
) -> dict[str, list[dict]]:
    header_buttons = list(buttons_by_zone.get(AssetActionButton.ZONE_HEADER, []))
    quick_buttons = list(buttons_by_zone.get(AssetActionButton.ZONE_QUICK, []))

    if qr_url:
        header_buttons = [button for button in header_buttons if not _match_action_button(button, href=qr_url)]
        quick_buttons = [button for button in quick_buttons if not _match_action_button(button, href=qr_url)]

    if assign_url:
        quick_buttons = [button for button in quick_buttons if not _match_action_button(button, href=assign_url)]
        if not any(_match_action_button(button, href=assign_url) for button in header_buttons):
            header_buttons.insert(
                0,
                {
                    "label": "Riassegna",
                    "style": AssetActionButton.STYLE_SECONDARY,
                    "href": assign_url,
                    "data_action": "",
                    "new_tab": False,
                },
            )

    buttons_by_zone[AssetActionButton.ZONE_HEADER] = header_buttons
    buttons_by_zone[AssetActionButton.ZONE_QUICK] = quick_buttons
    return buttons_by_zone


def _handle_action_button_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_zones = {key for key, _ in AssetActionButton.ZONE_CHOICES}
    valid_action_types = {key for key, _ in AssetActionButton.ACTION_CHOICES}
    valid_styles = {key for key, _ in AssetActionButton.STYLE_CHOICES}

    if action == "create_action_button":
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci etichetta pulsante."
        zone = _clean_string(request.POST.get("zone")) or AssetActionButton.ZONE_QUICK
        if zone not in valid_zones:
            zone = AssetActionButton.ZONE_QUICK
        action_type = _clean_string(request.POST.get("action_type")) or AssetActionButton.TYPE_LINK
        if action_type not in valid_action_types:
            action_type = AssetActionButton.TYPE_LINK
        style = _clean_string(request.POST.get("style")) or AssetActionButton.STYLE_DEFAULT
        if style not in valid_styles:
            style = AssetActionButton.STYLE_DEFAULT
        target = _clean_string(request.POST.get("target"))
        if action_type == AssetActionButton.TYPE_LINK and not target:
            return False, "Per i pulsanti LINK devi inserire un target."
        code = _unique_action_button_code(label, request.POST.get("code"))
        AssetActionButton.objects.create(
            code=code,
            zone=zone,
            label=label[:120],
            action_type=action_type,
            target=target,
            style=style,
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            open_in_new_tab=bool(request.POST.get("open_in_new_tab")),
            is_active=bool(request.POST.get("is_active")),
        )
        return True, f"Pulsante \"{label}\" creato."

    if action == "update_action_button":
        button_id = _as_int(request.POST.get("button_id"), default=0)
        button = AssetActionButton.objects.filter(pk=button_id).first()
        if not button:
            return False, "Pulsante non trovato."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Etichetta pulsante obbligatoria."
        zone = _clean_string(request.POST.get("zone")) or button.zone
        if zone not in valid_zones:
            zone = button.zone
        action_type = _clean_string(request.POST.get("action_type")) or button.action_type
        if action_type not in valid_action_types:
            action_type = button.action_type
        style = _clean_string(request.POST.get("style")) or button.style
        if style not in valid_styles:
            style = button.style
        target = _clean_string(request.POST.get("target"))
        if action_type == AssetActionButton.TYPE_LINK and not target:
            return False, "Per i pulsanti LINK devi inserire un target."
        button.zone = zone
        button.label = label[:120]
        button.action_type = action_type
        button.target = target
        button.style = style
        button.sort_order = _as_int(request.POST.get("sort_order"), default=button.sort_order)
        button.open_in_new_tab = bool(request.POST.get("open_in_new_tab"))
        button.is_active = bool(request.POST.get("is_active"))
        button.save(
            update_fields=[
                "zone",
                "label",
                "action_type",
                "target",
                "style",
                "sort_order",
                "open_in_new_tab",
                "is_active",
                "updated_at",
            ]
        )
        return True, f"Pulsante \"{button.label}\" aggiornato."

    if action == "delete_action_button":
        button_id = _as_int(request.POST.get("button_id"), default=0)
        button = AssetActionButton.objects.filter(pk=button_id).first()
        if not button:
            return False, "Pulsante non trovato."
        label = button.label
        button.delete()
        return True, f"Pulsante \"{label}\" eliminato."

    return False, "Azione pulsante non valida."


def _handle_detail_field_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_sections = {key for key, _ in AssetDetailField.SECTION_CHOICES}
    valid_scopes = {key for key, _ in AssetDetailField.SCOPE_CHOICES}
    valid_formats = {key for key, _ in AssetDetailField.FORMAT_CHOICES}
    valid_card_sizes = {key for key, _ in AssetDetailField.CARD_SIZE_CHOICES}

    if action == "seed_detail_fields":
        created = _seed_default_asset_detail_fields(create_only_if_empty=True)
        if created <= 0:
            return False, "Campi dettaglio gia configurati."
        return True, f"Schema dettaglio asset inizializzato ({created} campi)."

    if action == "create_detail_field":
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci etichetta campo dettaglio."
        section = _clean_string(request.POST.get("section")) or AssetDetailField.SECTION_SPECS
        if section not in valid_sections:
            section = AssetDetailField.SECTION_SPECS
        asset_scope = _clean_string(request.POST.get("asset_scope")) or AssetDetailField.SCOPE_ALL
        if asset_scope not in valid_scopes:
            asset_scope = AssetDetailField.SCOPE_ALL
        value_format = _clean_string(request.POST.get("value_format")) or AssetDetailField.FORMAT_AUTO
        if value_format not in valid_formats:
            value_format = AssetDetailField.FORMAT_AUTO
        card_size = _clean_string(request.POST.get("card_size")) or AssetDetailField.CARD_THIRD
        if card_size not in valid_card_sizes:
            card_size = AssetDetailField.CARD_THIRD
        source_ref = _clean_string(request.POST.get("source_ref"))
        if not source_ref:
            return False, "Seleziona il dato da mostrare."
        AssetDetailField.objects.create(
            code=_unique_detail_field_code(label, request.POST.get("code")),
            label=label[:120],
            section=section,
            asset_scope=asset_scope,
            source_ref=source_ref[:120],
            value_format=value_format,
            card_size=card_size,
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            show_if_empty=bool(request.POST.get("show_if_empty")),
            is_active=bool(request.POST.get("is_active")),
        )
        return True, f"Campo dettaglio \"{label}\" creato."

    if action == "update_detail_field_bulk":
        bulk_field = _clean_string(request.POST.get("bulk_field"))
        apply_scope = _clean_string(request.POST.get("apply_scope")) or "selected"
        selected_ids = [
            row
            for row in {
                _as_int(value, default=0)
                for value in request.POST.getlist("selected_detail_field_ids")
            }
            if row > 0
        ]

        if apply_scope == "all":
            detail_fields = list(
                AssetDetailField.objects.order_by("section", "asset_scope", "sort_order", "label", "id")
            )
        else:
            detail_fields = list(
                AssetDetailField.objects.filter(pk__in=selected_ids).order_by(
                    "section", "asset_scope", "sort_order", "label", "id"
                )
            )

        if not detail_fields:
            return False, "Seleziona almeno un campo oppure usa l'opzione per applicare a tutti."

        if bulk_field == "section":
            bulk_value = _clean_string(request.POST.get("bulk_section"))
            if bulk_value not in valid_sections:
                return False, "Sezione bulk non valida."
            for detail_field in detail_fields:
                detail_field.section = bulk_value
                detail_field.save(update_fields=["section", "updated_at"])
            return True, f"Sezione aggiornata per {len(detail_fields)} campi."

        if bulk_field == "asset_scope":
            bulk_value = _clean_string(request.POST.get("bulk_asset_scope"))
            if bulk_value not in valid_scopes:
                return False, "Ambito bulk non valido."
            for detail_field in detail_fields:
                detail_field.asset_scope = bulk_value
                detail_field.save(update_fields=["asset_scope", "updated_at"])
            return True, f"Ambito aggiornato per {len(detail_fields)} campi."

        if bulk_field == "value_format":
            bulk_value = _clean_string(request.POST.get("bulk_value_format"))
            if bulk_value not in valid_formats:
                return False, "Formato bulk non valido."
            for detail_field in detail_fields:
                detail_field.value_format = bulk_value
                detail_field.save(update_fields=["value_format", "updated_at"])
            return True, f"Formato aggiornato per {len(detail_fields)} campi."

        if bulk_field == "card_size":
            bulk_value = _clean_string(request.POST.get("bulk_card_size"))
            if bulk_value not in valid_card_sizes:
                return False, "Dimensione bulk non valida."
            for detail_field in detail_fields:
                detail_field.card_size = bulk_value
                detail_field.save(update_fields=["card_size", "updated_at"])
            return True, f"Dimensione aggiornata per {len(detail_fields)} campi."

        if bulk_field == "show_if_empty":
            bulk_value = _clean_string(request.POST.get("bulk_show_if_empty"))
            if bulk_value not in {"show", "hide"}:
                return False, "Valore bulk per campi vuoti non valido."
            show_if_empty = bulk_value == "show"
            for detail_field in detail_fields:
                detail_field.show_if_empty = show_if_empty
                detail_field.save(update_fields=["show_if_empty", "updated_at"])
            return True, f"Visibilita campi vuoti aggiornata per {len(detail_fields)} campi."

        if bulk_field == "is_active":
            bulk_value = _clean_string(request.POST.get("bulk_is_active"))
            if bulk_value not in {"active", "inactive"}:
                return False, "Stato attivo bulk non valido."
            is_active = bulk_value == "active"
            for detail_field in detail_fields:
                detail_field.is_active = is_active
                detail_field.save(update_fields=["is_active", "updated_at"])
            return True, f"Stato attivo aggiornato per {len(detail_fields)} campi."

        return False, "Parametro bulk non valido."

    if action == "update_detail_field":
        detail_field_id = _as_int(request.POST.get("detail_field_id"), default=0)
        detail_field = AssetDetailField.objects.filter(pk=detail_field_id).first()
        if not detail_field:
            return False, "Campo dettaglio non trovato."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Etichetta campo dettaglio obbligatoria."
        section = _clean_string(request.POST.get("section")) or detail_field.section
        if section not in valid_sections:
            section = detail_field.section
        asset_scope = _clean_string(request.POST.get("asset_scope")) or detail_field.asset_scope
        if asset_scope not in valid_scopes:
            asset_scope = detail_field.asset_scope
        value_format = _clean_string(request.POST.get("value_format")) or detail_field.value_format
        if value_format not in valid_formats:
            value_format = detail_field.value_format
        card_size = _clean_string(request.POST.get("card_size")) or detail_field.card_size
        if card_size not in valid_card_sizes:
            card_size = detail_field.card_size
        source_ref = _clean_string(request.POST.get("source_ref"))
        if not source_ref:
            return False, "Seleziona il dato da mostrare."
        detail_field.label = label[:120]
        detail_field.section = section
        detail_field.asset_scope = asset_scope
        detail_field.source_ref = source_ref[:120]
        detail_field.value_format = value_format
        detail_field.card_size = card_size
        detail_field.sort_order = _as_int(request.POST.get("sort_order"), default=detail_field.sort_order)
        detail_field.show_if_empty = bool(request.POST.get("show_if_empty"))
        detail_field.is_active = bool(request.POST.get("is_active"))
        detail_field.save(
            update_fields=[
                "label",
                "section",
                "asset_scope",
                "source_ref",
                "value_format",
                "card_size",
                "sort_order",
                "show_if_empty",
                "is_active",
                "updated_at",
            ]
        )
        return True, f"Campo dettaglio \"{detail_field.label}\" aggiornato."

    if action == "delete_detail_field":
        detail_field_id = _as_int(request.POST.get("detail_field_id"), default=0)
        detail_field = AssetDetailField.objects.filter(pk=detail_field_id).first()
        if not detail_field:
            return False, "Campo dettaglio non trovato."
        label = detail_field.label
        detail_field.delete()
        return True, f"Campo dettaglio \"{label}\" eliminato."

    return False, "Azione dettaglio asset non valida."


def _handle_detail_section_layout_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_sizes = {key for key, _ in AssetDetailSectionLayout.SIZE_CHOICES}
    valid_codes = {key for key, _ in AssetDetailSectionLayout.SECTION_CHOICES}

    if action == "update_detail_section_layout_bulk":
        bulk_field = _clean_string(request.POST.get("bulk_field"))
        apply_scope = _clean_string(request.POST.get("apply_scope")) or "selected"
        selected_ids = [
            row
            for row in {
                _as_int(value, default=0)
                for value in request.POST.getlist("selected_layout_ids")
            }
            if row > 0
        ]

        if apply_scope == "all":
            layouts = list(AssetDetailSectionLayout.objects.order_by("sort_order", "id"))
        else:
            layouts = list(AssetDetailSectionLayout.objects.filter(pk__in=selected_ids).order_by("sort_order", "id"))

        if not layouts:
            return False, "Seleziona almeno un riquadro oppure usa l'opzione per applicare a tutti."

        if bulk_field == "grid_size":
            bulk_value = _clean_string(request.POST.get("bulk_grid_size"))
            if bulk_value not in valid_sizes:
                return False, "Dimensione bulk non valida."
            for layout in layouts:
                layout.grid_size = bulk_value
                layout.save(update_fields=["grid_size", "updated_at"])
            return True, f"Dimensione aggiornata per {len(layouts)} riquadri."

        if bulk_field == "is_visible":
            bulk_value = _clean_string(request.POST.get("bulk_is_visible"))
            if bulk_value not in {"visible", "hidden"}:
                return False, "Stato visibilita bulk non valido."
            is_visible = bulk_value == "visible"
            for layout in layouts:
                layout.is_visible = is_visible
                layout.save(update_fields=["is_visible", "updated_at"])
            return True, f"Visibilita aggiornata per {len(layouts)} riquadri."

        return False, "Parametro bulk non valido."

    if action == "move_detail_section_layout":
        layout_id = _as_int(request.POST.get("layout_id"), default=0)
        move_direction = _clean_string(request.POST.get("direction")).lower()
        if move_direction not in {"up", "down"}:
            return False, "Direzione di spostamento non valida."
        ordered_layouts = list(AssetDetailSectionLayout.objects.order_by("sort_order", "id"))
        current_index = next(
            (index for index, row in enumerate(ordered_layouts) if row.id == layout_id),
            -1,
        )
        if current_index < 0:
            return False, "Riquadro dettaglio non trovato."
        target_index = current_index - 1 if move_direction == "up" else current_index + 1
        if target_index < 0 or target_index >= len(ordered_layouts):
            return True, "Posizione riquadro invariata."
        moved_layout = ordered_layouts.pop(current_index)
        ordered_layouts.insert(target_index, moved_layout)
        changed_layouts: list[AssetDetailSectionLayout] = []
        for index, row in enumerate(ordered_layouts, start=1):
            expected_order = index * 10
            if row.sort_order != expected_order:
                row.sort_order = expected_order
                changed_layouts.append(row)
        if changed_layouts:
            AssetDetailSectionLayout.objects.bulk_update(changed_layouts, ["sort_order"])
        return True, f'Riquadro "{moved_layout.get_code_display()}" spostato.'

    if action != "update_detail_section_layout":
        return False, "Azione layout dettaglio non valida."

    layout_id = _as_int(request.POST.get("layout_id"), default=0)
    layout = AssetDetailSectionLayout.objects.filter(pk=layout_id).first()
    if not layout:
        return False, "Riquadro dettaglio non trovato."

    code = _clean_string(request.POST.get("code")) or layout.code
    if code not in valid_codes:
        code = layout.code
    grid_size = _clean_string(request.POST.get("grid_size")) or layout.grid_size
    if grid_size not in valid_sizes:
        grid_size = layout.grid_size

    layout.code = code
    layout.grid_size = grid_size
    layout.sort_order = _as_int(request.POST.get("sort_order"), default=layout.sort_order)
    layout.is_visible = bool(request.POST.get("is_visible"))
    layout.save(update_fields=["code", "grid_size", "sort_order", "is_visible", "updated_at"])
    return True, f"Riquadro \"{layout.get_code_display()}\" aggiornato."


def _handle_asset_list_layout_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    if action not in LIST_LAYOUT_ACTIONS:
        return False, "Azione layout lista non valida."

    try:
        layout_id = _as_int(request.POST.get("layout_id"), default=0)
        layout = AssetListLayout.objects.filter(pk=layout_id).first()
        if not layout:
            return False, "Preset lista non trovato."

        definition = _asset_list_context_definition_map().get(layout.context_key, {})
        default_columns = list(definition.get("visible_columns", []) or definition.get("default_columns", []) or [])
        custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
        valid_keys = _asset_list_valid_column_keys(custom_fields)

        if action == "reset_asset_list_layout":
            layout.visible_columns = list(default_columns)
            layout.is_customized = False
            layout.save(update_fields=["visible_columns", "is_customized", "updated_at"])
            return True, f"Vista \"{layout.get_context_key_display()}\" ripristinata."

        selected_columns = _sanitize_asset_list_visible_columns(
            request.POST.getlist("visible_columns"),
            valid_keys,
            fallback=[],
        )
        if not selected_columns:
            return False, "Seleziona almeno una colonna per il preset centrale."

        layout.visible_columns = selected_columns
        layout.is_customized = True
        layout.save(update_fields=["visible_columns", "is_customized", "updated_at"])
        return True, f"Vista \"{layout.get_context_key_display()}\" aggiornata."
    except DatabaseError:
        return False, "Preset lista non disponibile finche il database non e' allineato con le migration."


def _handle_asset_category_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    valid_asset_types = {key for key, _ in Asset.TYPE_CHOICES}
    valid_field_types = {key for key, _ in AssetCategoryField.TYPE_CHOICES}
    valid_sections = {key for key, _ in AssetDetailField.SECTION_CHOICES}
    valid_formats = {key for key, _ in AssetDetailField.FORMAT_CHOICES}
    valid_card_sizes = {key for key, _ in AssetDetailField.CARD_SIZE_CHOICES}
    valid_sidebar_sections = {key for key, _ in AssetSidebarButton.SECTION_CHOICES}

    if action == "create_asset_category":
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci il nome categoria."
        base_asset_type = _clean_string(request.POST.get("base_asset_type")) or Asset.TYPE_OTHER
        if base_asset_type not in valid_asset_types:
            base_asset_type = Asset.TYPE_OTHER
        parent_id = _as_int(request.POST.get("parent_id"), default=0)
        parent = AssetCategory.objects.filter(pk=parent_id).first() if parent_id else None
        AssetCategory.objects.create(
            code=_unique_asset_category_code(label, request.POST.get("code")),
            label=label[:120],
            parent=parent,
            base_asset_type=base_asset_type,
            description=_clean_string(request.POST.get("description")),
            detail_specs_title=_clean_string(request.POST.get("detail_specs_title"))[:120],
            detail_profile_title=_clean_string(request.POST.get("detail_profile_title"))[:120],
            detail_assignment_title=_clean_string(request.POST.get("detail_assignment_title"))[:120],
            detail_timeline_title=_clean_string(request.POST.get("detail_timeline_title"))[:120],
            detail_maintenance_title=_clean_string(request.POST.get("detail_maintenance_title"))[:120],
            detail_timeline_manual_enabled=bool(request.POST.get("detail_timeline_manual_enabled")),
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            is_active=bool(request.POST.get("is_active")),
        )
        return True, f"Categoria asset \"{label}\" creata."

    if action == "update_asset_category":
        category_id = _as_int(request.POST.get("category_id"), default=0)
        category = AssetCategory.objects.filter(pk=category_id).first()
        if not category:
            return False, "Categoria asset non trovata."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Il nome categoria e obbligatorio."
        base_asset_type = _clean_string(request.POST.get("base_asset_type")) or category.base_asset_type
        if base_asset_type not in valid_asset_types:
            base_asset_type = category.base_asset_type
        parent_id = _as_int(request.POST.get("parent_id"), default=0)
        parent = AssetCategory.objects.filter(pk=parent_id).first() if parent_id else None
        if parent and parent.id == category.id:
            return False, "Una categoria non puo essere padre di se stessa."
        category.label = label[:120]
        category.parent = parent
        category.base_asset_type = base_asset_type
        category.description = _clean_string(request.POST.get("description"))
        category.detail_specs_title = _clean_string(request.POST.get("detail_specs_title"))[:120]
        category.detail_profile_title = _clean_string(request.POST.get("detail_profile_title"))[:120]
        category.detail_assignment_title = _clean_string(request.POST.get("detail_assignment_title"))[:120]
        category.detail_timeline_title = _clean_string(request.POST.get("detail_timeline_title"))[:120]
        category.detail_maintenance_title = _clean_string(request.POST.get("detail_maintenance_title"))[:120]
        category.detail_timeline_manual_enabled = bool(request.POST.get("detail_timeline_manual_enabled"))
        category.sort_order = _as_int(request.POST.get("sort_order"), default=category.sort_order)
        category.is_active = bool(request.POST.get("is_active"))
        category.save(
            update_fields=[
                "label",
                "parent",
                "base_asset_type",
                "description",
                "detail_specs_title",
                "detail_profile_title",
                "detail_assignment_title",
                "detail_timeline_title",
                "detail_maintenance_title",
                "detail_timeline_manual_enabled",
                "sort_order",
                "is_active",
                "updated_at",
            ]
        )
        return True, f"Categoria asset \"{category.label}\" aggiornata."

    if action == "create_sidebar_button_for_category":
        category_id = _as_int(request.POST.get("category_id"), default=0)
        category = AssetCategory.objects.filter(pk=category_id).first()
        if not category:
            return False, "Categoria asset non trovata."
        if not AssetSidebarButton.objects.exists():
            _create_default_sidebar_buttons()
        target_url = _category_sidebar_target(category.id)
        active_match = _category_sidebar_active_match(category.id)
        existing = AssetSidebarButton.objects.filter(Q(target_url=target_url) | Q(active_match=active_match)).first()
        if existing:
            return False, f"La categoria \"{category.label}\" ha gia una voce menu: {existing.label}."

        parent_id = _as_int(request.POST.get("parent_sidebar_button_id") or request.POST.get("parent_id"), default=0)
        parent_button = AssetSidebarButton.objects.filter(pk=parent_id).first() if parent_id else None
        if parent_button and parent_button.parent_id:
            return False, "La voce padre deve essere di primo livello."
        section = _clean_string(request.POST.get("section")) or AssetSidebarButton.SECTION_MAIN
        if section not in valid_sidebar_sections:
            section = AssetSidebarButton.SECTION_MAIN
        if parent_button is not None:
            section = parent_button.section

        label = _clean_string(request.POST.get("sidebar_label")) or category.label
        AssetSidebarButton.objects.create(
            code=_unique_sidebar_button_code(f"categoria-{category.code}", request.POST.get("code")),
            section=section,
            parent=parent_button,
            label=label[:120],
            target_url=target_url,
            active_match=active_match,
            is_subitem=True if parent_button is not None else bool(request.POST.get("is_subitem")),
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            is_visible=True,
        )
        return True, f"Voce menu per categoria \"{category.label}\" creata."

    if action == "delete_asset_category":
        category_id = _as_int(request.POST.get("category_id"), default=0)
        category = AssetCategory.objects.filter(pk=category_id).first()
        if not category:
            return False, "Categoria asset non trovata."
        linked_assets = category.assets.count()
        if linked_assets:
            return False, f"La categoria \"{category.label}\" e assegnata a {linked_assets} asset: rimuovila prima dagli asset collegati."
        label = category.label
        category.delete()
        return True, f"Categoria asset \"{label}\" eliminata."

    if action == "create_asset_category_field":
        category_id = _as_int(request.POST.get("category_id"), default=0)
        category = AssetCategory.objects.filter(pk=category_id).first()
        if not category:
            return False, "Seleziona una categoria valida."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci il nome campo categoria."
        field_type = _clean_string(request.POST.get("field_type")) or AssetCategoryField.TYPE_TEXT
        if field_type not in valid_field_types:
            field_type = AssetCategoryField.TYPE_TEXT
        detail_section = _clean_string(request.POST.get("detail_section")) or AssetDetailField.SECTION_SPECS
        if detail_section not in valid_sections:
            detail_section = AssetDetailField.SECTION_SPECS
        detail_value_format = _clean_string(request.POST.get("detail_value_format")) or AssetDetailField.FORMAT_AUTO
        if detail_value_format not in valid_formats:
            detail_value_format = AssetDetailField.FORMAT_AUTO
        detail_card_size = _clean_string(request.POST.get("detail_card_size")) or AssetDetailField.CARD_THIRD
        if detail_card_size not in valid_card_sizes:
            detail_card_size = AssetDetailField.CARD_THIRD
        AssetCategoryField.objects.create(
            category=category,
            code=_unique_asset_category_field_code(label, request.POST.get("code")),
            label=label[:120],
            field_type=field_type,
            detail_section=detail_section,
            detail_value_format=detail_value_format,
            detail_card_size=detail_card_size,
            placeholder=_clean_string(request.POST.get("placeholder"))[:160],
            help_text=_clean_string(request.POST.get("help_text"))[:255],
            sort_order=_as_int(request.POST.get("sort_order"), default=100),
            is_required=bool(request.POST.get("is_required")),
            show_in_form=bool(request.POST.get("show_in_form", "1")),
            show_in_detail=bool(request.POST.get("show_in_detail", "1")),
            show_if_empty=bool(request.POST.get("show_if_empty")),
            is_active=bool(request.POST.get("is_active", "1")),
        )
        return True, f"Campo categoria \"{label}\" creato."

    if action == "update_asset_category_field":
        field_id = _as_int(request.POST.get("category_field_id"), default=0)
        field = AssetCategoryField.objects.select_related("category").filter(pk=field_id).first()
        if not field:
            return False, "Campo categoria non trovato."
        category_id = _as_int(request.POST.get("category_id"), default=field.category_id)
        category = AssetCategory.objects.filter(pk=category_id).first()
        if not category:
            return False, "Categoria asset non valida."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Il nome campo categoria e obbligatorio."
        field_type = _clean_string(request.POST.get("field_type")) or field.field_type
        if field_type not in valid_field_types:
            field_type = field.field_type
        detail_section = _clean_string(request.POST.get("detail_section")) or field.detail_section
        if detail_section not in valid_sections:
            detail_section = field.detail_section
        detail_value_format = _clean_string(request.POST.get("detail_value_format")) or field.detail_value_format
        if detail_value_format not in valid_formats:
            detail_value_format = field.detail_value_format
        detail_card_size = _clean_string(request.POST.get("detail_card_size")) or field.detail_card_size
        if detail_card_size not in valid_card_sizes:
            detail_card_size = field.detail_card_size
        field.category = category
        field.label = label[:120]
        field.field_type = field_type
        field.detail_section = detail_section
        field.detail_value_format = detail_value_format
        field.detail_card_size = detail_card_size
        field.placeholder = _clean_string(request.POST.get("placeholder"))[:160]
        field.help_text = _clean_string(request.POST.get("help_text"))[:255]
        field.sort_order = _as_int(request.POST.get("sort_order"), default=field.sort_order)
        field.is_required = bool(request.POST.get("is_required"))
        field.show_in_form = bool(request.POST.get("show_in_form"))
        field.show_in_detail = bool(request.POST.get("show_in_detail"))
        field.show_if_empty = bool(request.POST.get("show_if_empty"))
        field.is_active = bool(request.POST.get("is_active"))
        field.save(
            update_fields=[
                "category",
                "label",
                "field_type",
                "detail_section",
                "detail_value_format",
                "detail_card_size",
                "placeholder",
                "help_text",
                "sort_order",
                "is_required",
                "show_in_form",
                "show_in_detail",
                "show_if_empty",
                "is_active",
                "updated_at",
            ]
        )
        return True, f"Campo categoria \"{field.label}\" aggiornato."

    if action == "delete_asset_category_field":
        field_id = _as_int(request.POST.get("category_field_id"), default=0)
        field = AssetCategoryField.objects.filter(pk=field_id).first()
        if not field:
            return False, "Campo categoria non trovato."
        label = field.label
        code = field.code
        field.delete()
        touched = _update_asset_category_values_after_delete(code)
        return True, f"Campo categoria \"{label}\" eliminato ({touched} asset ripuliti)."

    return False, "Azione categoria asset non valida."


def _query_url(request: HttpRequest, **overrides) -> str:
    params = request.GET.copy()
    for key, value in overrides.items():
        if value is None:
            params.pop(key, None)
            continue
        params[key] = str(value)
    query = params.urlencode()
    return f"?{query}" if query else ""


def _asset_list_context_definitions() -> list[dict[str, object]]:
    default_columns = list(ASSET_LIST_COMMON_COLUMNS)
    return [
        {
            "key": AssetListLayout.CONTEXT_ALL,
            "label": "Inventario completo",
            "asset_type": "",
            "default_columns": default_columns,
            "sort_order": 100,
        },
        {
            "key": AssetListLayout.CONTEXT_DEVICES,
            "label": "Dispositivi",
            "asset_type": Asset.TYPE_HW,
            "default_columns": default_columns,
            "sort_order": 110,
        },
        {
            "key": AssetListLayout.CONTEXT_SERVERS,
            "label": "Server",
            "asset_type": Asset.TYPE_SERVER,
            "default_columns": default_columns,
            "sort_order": 120,
        },
        {
            "key": AssetListLayout.CONTEXT_WORKSTATIONS,
            "label": "Postazioni di lavoro",
            "asset_type": Asset.TYPE_PC,
            "default_columns": default_columns,
            "sort_order": 130,
        },
        {
            "key": AssetListLayout.CONTEXT_NETWORK,
            "label": "Rete",
            "asset_type": Asset.TYPE_FIREWALL,
            "default_columns": default_columns,
            "sort_order": 140,
        },
        {
            "key": AssetListLayout.CONTEXT_VIRTUAL_MACHINES,
            "label": "Macchine virtuali",
            "asset_type": Asset.TYPE_VM,
            "default_columns": default_columns,
            "sort_order": 150,
        },
        {
            "key": AssetListLayout.CONTEXT_CCTV,
            "label": "Videosorveglianza",
            "asset_type": Asset.TYPE_CCTV,
            "default_columns": default_columns,
            "sort_order": 160,
        },
    ]


def _asset_list_context_definition_map() -> dict[str, dict[str, object]]:
    return {
        str(row["key"]): row
        for row in _asset_list_context_definitions()
    }


def _asset_list_context(asset_type: str) -> tuple[str, str]:
    normalized = _clean_string(asset_type).upper()
    if normalized in {Asset.TYPE_PC, Asset.TYPE_NOTEBOOK}:
        return AssetListLayout.CONTEXT_WORKSTATIONS, "Postazioni di lavoro"
    if normalized == Asset.TYPE_SERVER:
        return AssetListLayout.CONTEXT_SERVERS, "Server"
    if normalized == Asset.TYPE_FIREWALL:
        return AssetListLayout.CONTEXT_NETWORK, "Rete"
    if normalized == Asset.TYPE_HW:
        return AssetListLayout.CONTEXT_DEVICES, "Dispositivi"
    if normalized == Asset.TYPE_VM:
        return AssetListLayout.CONTEXT_VIRTUAL_MACHINES, "Macchine virtuali"
    if normalized == Asset.TYPE_CCTV:
        return AssetListLayout.CONTEXT_CCTV, "Videosorveglianza"
    return AssetListLayout.CONTEXT_ALL, "Inventario completo"


def _asset_list_default_columns(asset_type: str) -> list[str]:
    return list(ASSET_LIST_COMMON_COLUMNS)


def _default_asset_list_layout_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for definition in _asset_list_context_definitions():
        asset_type = _clean_string(definition.get("asset_type"))
        rows.append(
            {
                "context_key": str(definition["key"]),
                "sort_order": int(definition["sort_order"]),
                "visible_columns": _asset_list_default_columns(asset_type),
                "is_customized": False,
            }
        )
    return rows


def _default_asset_list_layout_instances() -> list[AssetListLayout]:
    return [
        AssetListLayout(
            context_key=str(item["context_key"]),
            sort_order=int(item["sort_order"]),
            visible_columns=list(item["visible_columns"]),
            is_customized=bool(item["is_customized"]),
        )
        for item in _default_asset_list_layout_rows()
    ]


def _ensure_default_asset_list_layouts() -> list[AssetListLayout]:
    defaults = _default_asset_list_layout_rows()
    try:
        existing = {
            row.context_key: row
            for row in AssetListLayout.objects.all()
        }
        for item in defaults:
            context_key = str(item["context_key"])
            if context_key in existing:
                row = existing[context_key]
                if not isinstance(row.visible_columns, list) or not row.visible_columns:
                    row.visible_columns = list(item["visible_columns"])
                    row.save(update_fields=["visible_columns", "updated_at"])
                elif not row.is_customized and row.visible_columns != list(item["visible_columns"]):
                    row.visible_columns = list(item["visible_columns"])
                    row.save(update_fields=["visible_columns", "updated_at"])
                continue
            existing[context_key] = AssetListLayout.objects.create(
                context_key=context_key,
                sort_order=int(item["sort_order"]),
                visible_columns=list(item["visible_columns"]),
                is_customized=bool(item["is_customized"]),
            )
        return list(AssetListLayout.objects.order_by("sort_order", "id"))
    except DatabaseError:
        return _default_asset_list_layout_instances()


def _asset_list_valid_column_keys(custom_fields: list[AssetCustomField]) -> set[str]:
    return set(ASSET_LIST_COMMON_COLUMNS)


def _sanitize_asset_list_visible_columns(columns: object, valid_keys: set[str], fallback: list[str] | None = None) -> list[str]:
    cleaned: list[str] = []
    for value in columns if isinstance(columns, list) else []:
        key = _clean_string(value)
        if key and key in valid_keys and key not in cleaned:
            cleaned.append(key)
    if cleaned:
        return cleaned
    return list(fallback or [])


def _asset_list_layout_revision(layout: AssetListLayout | None) -> str:
    if layout is None or layout.updated_at is None:
        return "default"
    return layout.updated_at.strftime("%Y%m%d%H%M%S")


def _asset_list_layout_manage_url(request: HttpRequest, context_key: str) -> str:
    if not _can_manage_asset_list_layout(request):
        return ""
    try:
        return f"{reverse('assets:asset_list_layout_admin')}?context={quote(context_key)}"
    except NoReverseMatch:
        return ""


def _asset_table_layout_storage_user_id(request: HttpRequest) -> int | None:
    if not getattr(request.user, "is_authenticated", False):
        return None
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    try:
        if legacy_user is not None and getattr(legacy_user, "id", None) is not None:
            return int(legacy_user.id)
        user_id = getattr(request.user, "pk", None)
        return -int(user_id) if user_id is not None else None
    except (TypeError, ValueError):
        return None


def _load_user_dashboard_layout_payload(storage_user_id: int | None) -> dict[str, object]:
    if storage_user_id is None:
        return {}
    try:
        row = UserDashboardLayout.objects.filter(legacy_user_id=storage_user_id).first()
    except DatabaseError:
        return {}
    payload = getattr(row, "layout", {})
    return payload if isinstance(payload, dict) else {}


def _sanitize_asset_table_column_order(value: object, valid_keys: set[str]) -> list[str]:
    cleaned: list[str] = []
    for item in value if isinstance(value, list) else []:
        key = _clean_string(item)
        if key and key in valid_keys and key not in cleaned:
            cleaned.append(key)
    return cleaned


def _sanitize_asset_table_column_widths(value: object, valid_keys: set[str]) -> dict[str, int]:
    cleaned: dict[str, int] = {}
    if not isinstance(value, dict):
        return cleaned
    for raw_key, raw_width in value.items():
        key = _clean_string(raw_key)
        if not key or key not in valid_keys or key in cleaned:
            continue
        try:
            width = int(raw_width)
        except (TypeError, ValueError):
            continue
        if 90 <= width <= 1600:
            cleaned[key] = width
    return cleaned


def _sanitize_asset_table_layout(value: object, valid_keys: set[str]) -> dict[str, object]:
    if not isinstance(value, dict):
        return {}
    return {
        "visible_columns": _sanitize_asset_list_visible_columns(value.get("visible_columns"), valid_keys, fallback=[]),
        "column_order": _sanitize_asset_table_column_order(value.get("column_order"), valid_keys),
        "column_widths": _sanitize_asset_table_column_widths(value.get("column_widths"), valid_keys),
    }


def _load_user_asset_table_layout(request: HttpRequest, context_key: str, valid_keys: set[str]) -> dict[str, object]:
    payload = _load_user_dashboard_layout_payload(_asset_table_layout_storage_user_id(request))
    contexts = payload.get("assets_table")
    if not isinstance(contexts, dict):
        return {}
    return _sanitize_asset_table_layout(contexts.get(context_key), valid_keys)


def _persist_user_asset_table_layout(
    request: HttpRequest,
    context_key: str,
    payload: dict[str, object],
    valid_keys: set[str],
) -> dict[str, object] | None:
    storage_user_id = _asset_table_layout_storage_user_id(request)
    if storage_user_id is None:
        return None
    sanitized = _sanitize_asset_table_layout(payload, valid_keys)
    try:
        current = _load_user_dashboard_layout_payload(storage_user_id)
        updated = dict(current)
        contexts = updated.get("assets_table")
        if not isinstance(contexts, dict):
            contexts = {}

        has_payload = bool(
            sanitized.get("visible_columns") or sanitized.get("column_order") or sanitized.get("column_widths")
        )
        if has_payload:
            contexts[context_key] = sanitized
        else:
            contexts.pop(context_key, None)

        if contexts:
            updated["assets_table"] = contexts
        else:
            updated.pop("assets_table", None)

        if updated:
            UserDashboardLayout.objects.update_or_create(
                legacy_user_id=storage_user_id,
                defaults={"layout": updated},
            )
        else:
            UserDashboardLayout.objects.filter(legacy_user_id=storage_user_id).delete()
        return sanitized
    except DatabaseError:
        return None


def _handle_asset_table_layout_request(request: HttpRequest, payload: dict[str, object]) -> JsonResponse:
    context_key = _clean_string(payload.get("context_key"))
    if context_key not in _asset_list_context_definition_map():
        return JsonResponse({"ok": False, "error": "Contesto layout non valido."}, status=400)

    custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
    valid_keys = _asset_list_valid_column_keys(custom_fields)
    saved_layout = _persist_user_asset_table_layout(
        request,
        context_key,
        {
            "visible_columns": payload.get("visible_columns"),
            "column_order": payload.get("column_order"),
            "column_widths": payload.get("column_widths"),
        },
        valid_keys,
    )
    if saved_layout is None:
        return JsonResponse({"ok": False, "error": "Impossibile salvare le preferenze tabella."}, status=400)
    return JsonResponse({"ok": True, "layout": saved_layout})


def _asset_list_preview_url(context_key: str, rows: int = 25) -> str:
    definition = _asset_list_context_definition_map().get(context_key, {})
    base_url = reverse("assets:asset_list")
    asset_type = _clean_string(definition.get("asset_type"))
    if asset_type:
        return f"{base_url}?asset_type={quote(asset_type)}&rows={rows}"
    return f"{base_url}?rows={rows}"


def _asset_extra_has_custom_value(extra: object, field: AssetCustomField) -> bool:
    if not isinstance(extra, dict) or field is None:
        return False
    sentinel = object()
    value = extra.get(field.code, sentinel)
    if value is sentinel:
        value = extra.get(field.label, sentinel)
    if value is sentinel:
        return False
    if value in ("", None, [], {}):
        return False
    return True


def _asset_list_relevant_custom_columns(assets_qs, custom_fields: list[AssetCustomField], sample_size: int = 250) -> list[str]:
    if not custom_fields:
        return []
    relevant_codes: list[str] = []
    remaining = {field.code: field for field in custom_fields}
    for asset in assets_qs[:sample_size]:
        extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}
        for code, field in list(remaining.items()):
            if _asset_extra_has_custom_value(extra, field):
                relevant_codes.append(code)
                remaining.pop(code, None)
        if not remaining:
            break
    return relevant_codes


def _asset_endpoint_column_summary(asset: Asset) -> dict[str, str]:
    endpoints = list(asset.endpoints.all())

    def _join_unique(values: list[str]) -> str:
        seen: list[str] = []
        for value in values:
            normalized = _clean_string(value)
            if normalized and normalized not in seen:
                seen.append(normalized)
        return ", ".join(seen) if seen else "-"

    vlan_values = []
    ip_values = []
    for endpoint in endpoints:
        if endpoint.vlan is not None:
            vlan_values.append(str(endpoint.vlan))
        if endpoint.ip:
            ip_values.append(str(endpoint.ip))
    return {
        "vlan": _join_unique(vlan_values),
        "ip": _join_unique(ip_values),
    }


def _handle_custom_field_request(request: HttpRequest) -> tuple[bool, str]:
    action = _clean_string(request.POST.get("action"))
    allowed_types = {choice[0] for choice in AssetCustomField.TYPE_CHOICES}

    if action == "create_custom_field":
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Inserisci il nome del campo."
        field_type = _clean_string(request.POST.get("field_type")) or AssetCustomField.TYPE_TEXT
        if field_type not in allowed_types:
            field_type = AssetCustomField.TYPE_TEXT
        sort_order = _as_int(request.POST.get("sort_order"), default=100)
        code = _unique_custom_field_code(label, requested_code=request.POST.get("code"))
        is_active = bool(request.POST.get("is_active"))
        AssetCustomField.objects.create(
            code=code,
            label=label[:120],
            field_type=field_type,
            sort_order=sort_order,
            is_active=is_active,
        )
        return True, f"Campo \"{label}\" creato."

    if action == "update_custom_field":
        field_id = _as_int(request.POST.get("field_id"), default=0)
        field = AssetCustomField.objects.filter(pk=field_id).first()
        if not field:
            return False, "Campo non trovato."
        label = _clean_string(request.POST.get("label"))
        if not label:
            return False, "Il nome campo non puo essere vuoto."
        field_type = _clean_string(request.POST.get("field_type")) or field.field_type
        if field_type not in allowed_types:
            field_type = field.field_type
        field.label = label[:120]
        field.field_type = field_type
        field.sort_order = _as_int(request.POST.get("sort_order"), default=field.sort_order)
        field.is_active = bool(request.POST.get("is_active"))
        field.save(update_fields=["label", "field_type", "sort_order", "is_active", "updated_at"])
        return True, f"Campo \"{field.label}\" aggiornato."

    if action == "delete_custom_field":
        field_id = _as_int(request.POST.get("field_id"), default=0)
        field = AssetCustomField.objects.filter(pk=field_id).first()
        if not field:
            return False, "Campo non trovato."
        field_code = field.code
        field_label = field.label
        field.delete()
        touched = _update_custom_field_values_after_delete(field_code)
        return True, f"Campo \"{field_label}\" eliminato ({touched} asset aggiornati)."

    return False, "Azione non valida."


def _handle_excel_import_request(request: HttpRequest) -> tuple[bool, str]:
    uploaded_file = request.FILES.get("excel_file")
    if not uploaded_file:
        return False, "Seleziona un file Excel prima di avviare l'import."

    file_name = (uploaded_file.name or "").lower()
    if not file_name.endswith((".xlsx", ".xlsm")):
        return False, "Formato non supportato. Usa file .xlsx oppure .xlsm."

    sheets_csv = _clean_string(request.POST.get("import_sheets")) or DEFAULT_IMPORT_SHEETS
    dry_run = bool(request.POST.get("dry_run"))
    include_optional = bool(request.POST.get("include_optional"))
    all_sheets = bool(request.POST.get("all_sheets"))
    update_existing = bool(request.POST.get("update_existing", "1") == "1")

    tmp_path = ""
    output = io.StringIO()
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
            tmp_path = temp_file.name

        call_command(
            "import_assets_excel",
            file=tmp_path,
            sheets=sheets_csv,
            dry_run=dry_run,
            include_optional=include_optional,
            all_sheets=all_sheets,
            update=update_existing,
            stdout=output,
            stderr=output,
        )
        command_output = output.getvalue().strip()
        mode = "DRY-RUN" if dry_run else "IMPORT REALE"
        if command_output:
            return True, f"{mode} completato. {command_output.splitlines()[-1]}"
        return True, f"{mode} completato con successo."
    except Exception as exc:
        command_output = output.getvalue().strip()
        if command_output:
            return False, f"{exc} | {command_output}"
        return False, str(exc)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def _build_assets_admin_snapshot() -> dict:
    return {
        "generated_at": timezone.now().isoformat(),
        "asset_categories": list(
            AssetCategory.objects.order_by("sort_order", "label", "id").values(
                "code",
                "label",
                "base_asset_type",
                "description",
                "detail_specs_title",
                "detail_profile_title",
                "detail_assignment_title",
                "detail_timeline_title",
                "detail_maintenance_title",
                "detail_timeline_manual_enabled",
                "sort_order",
                "is_active",
            )
        ),
        "asset_category_fields": list(
            AssetCategoryField.objects.select_related("category")
            .order_by("category__sort_order", "category__label", "sort_order", "label", "id")
            .values(
                "code",
                "category__code",
                "category__label",
                "label",
                "field_type",
                "detail_section",
                "detail_value_format",
                "detail_card_size",
                "placeholder",
                "help_text",
                "sort_order",
                "is_required",
                "show_in_form",
                "show_in_detail",
                "show_if_empty",
                "is_active",
            )
        ),
        "custom_fields": list(
            AssetCustomField.objects.order_by("sort_order", "id").values(
                "code",
                "label",
                "field_type",
                "sort_order",
                "is_active",
            )
        ),
        "list_options": list(
            AssetListOption.objects.order_by("field_key", "sort_order", "value", "id").values(
                "field_key",
                "value",
                "sort_order",
                "is_active",
            )
        ),
        "action_buttons": list(
            AssetActionButton.objects.order_by("zone", "sort_order", "label", "id").values(
                "code",
                "label",
                "zone",
                "action_type",
                "target",
                "style",
                "sort_order",
                "open_in_new_tab",
                "is_active",
            )
        ),
        "detail_fields": list(
            AssetDetailField.objects.order_by("section", "asset_scope", "sort_order", "label", "id").values(
                "code",
                "label",
                "section",
                "asset_scope",
                "source_ref",
                "value_format",
                "card_size",
                "sort_order",
                "show_if_empty",
                "is_active",
            )
        ),
        "detail_section_layouts": list(
            AssetDetailSectionLayout.objects.order_by("sort_order", "id").values(
                "code",
                "grid_size",
                "sort_order",
                "is_visible",
            )
        ),
        "sidebar_buttons": list(
            AssetSidebarButton.objects.order_by("section", "sort_order", "label", "id").values(
                "code",
                "label",
                "section",
                "parent_id",
                "parent__code",
                "target_url",
                "active_match",
                "is_subitem",
                "sort_order",
                "is_visible",
            )
        ),
    }


@login_required
def asset_part_145_list(request: HttpRequest) -> HttpResponse:
    """Sezione dedicata: elenco degli asset che rientrano in PART 145."""
    part_145_assets = (
        Asset.objects.filter(part_145=True)
        .select_related("asset_category")
        .order_by("name", "asset_tag", "id")
    )
    context = {
        "page_title": "Asset PART 145",
        "part_145_assets": part_145_assets,
        "part_145_total": part_145_assets.count(),
        # Shell context: senza questo la sidebar del modulo asset resta vuota.
        **_assets_shell_context(request, search_action=reverse("assets:asset_list")),
    }
    return render(request, "assets/pages/part_145_list.html", context)


def _part_145_export_dataset() -> tuple[list[str], list[list[str]]]:
    """Header + righe (stringhe) per l'export PART 145; stesso ordine della lista."""
    assets = (
        Asset.objects.filter(part_145=True)
        .select_related("asset_category")
        .order_by("name", "asset_tag", "id")
    )
    headers = ["Asset", "Tag", "Stato", "Categoria", "Reparto", "Produttore", "Modello", "Seriale"]
    rows = [
        [
            a.name or "",
            a.asset_tag or "",
            a.get_status_display(),
            a.category_label or "",
            a.reparto or "",
            a.manufacturer or "",
            a.model or "",
            a.serial_number or "",
        ]
        for a in assets
    ]
    return headers, rows


@login_required
def asset_part_145_export_pdf(request: HttpRequest) -> HttpResponse:
    """Esporta l'elenco PART 145 in PDF (report tabellare standard)."""
    headers, rows = _part_145_export_dataset()
    today = timezone.localdate().strftime("%Y%m%d")
    pdf_bytes = _report_table_pdf(title="Asset PART 145", headers=headers, rows=rows)
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="asset_part145_{today}.pdf"'
    return resp


@login_required
def asset_part_145_export_excel(request: HttpRequest) -> HttpResponse:
    """Esporta l'elenco PART 145 in Excel (.xlsx)."""
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    from core.excel_export import write_cell, write_row

    headers, rows = _part_145_export_dataset()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset PART 145"

    header_fill = PatternFill(fill_type="solid", fgColor="DC2626")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for col_idx, header in enumerate(headers, 1):
        cell = write_cell(ws, 1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for row_idx, values in enumerate(rows, 2):
        write_row(ws, row_idx, values)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today_str = timezone.localdate().strftime("%Y%m%d")
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="asset_part145_{today_str}.xlsx"'
    return resp


@login_required
def asset_list(request: HttpRequest) -> HttpResponse:
    can_manage_custom_fields = _is_assets_admin(request)

    if request.method == "GET" and request.GET.get("category") and not request.GET.get("asset_category"):
        next_query = request.GET.copy()
        next_query["asset_category"] = request.GET.get("category")
        next_query.pop("category", None)
        query_string = next_query.urlencode()
        target_url = reverse("assets:asset_list")
        return redirect(f"{target_url}?{query_string}" if query_string else target_url)

    if request.method == "POST":
        json_payload: dict[str, object] = {}
        if "application/json" in str(getattr(request, "content_type", "") or "").lower():
            try:
                decoded = json.loads((request.body or b"{}").decode("utf-8"))
            except (json.JSONDecodeError, UnicodeDecodeError):
                decoded = {}
            if isinstance(decoded, dict):
                json_payload = decoded
        action = _clean_string(json_payload.get("action")) if json_payload else _clean_string(request.POST.get("action"))
        if action == "save_asset_table_layout":
            return _handle_asset_table_layout_request(request, json_payload)
        if action == "import_excel":
            # L'import crea e (update_existing default=1) SOVRASCRIVE asset in blocco:
            # e' una scrittura di massa, non un'azione operativa. Solo admin asset,
            # come tutte le altre azioni di questo dispatcher.
            if not can_manage_custom_fields:
                log_action(
                    request,
                    "import_assets_excel",
                    "assets",
                    {"esito": "denied", "motivo": "permission_denied"},
                )
                return render(request, "core/pages/forbidden.html", status=403)
            ok, text = _handle_excel_import_request(request)
            log_action(
                request,
                "import_assets_excel",
                "assets",
                {"esito": "success" if ok else "error", "messaggio": text[:500]},
            )
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, f"Import Excel fallito: {text}")
            return redirect("assets:asset_list")
        if action in {"create_custom_field", "update_custom_field", "delete_custom_field"}:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare i campi personalizzati.")
                return redirect("assets:asset_list")
            ok, text = _handle_custom_field_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in LIST_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare le liste.")
                return redirect("assets:asset_list")
            ok, text = _handle_list_option_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in BUTTON_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare i pulsanti.")
                return redirect("assets:asset_list")
            ok, text = _handle_action_button_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in DETAIL_FIELD_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare il dettaglio asset.")
                return redirect("assets:asset_list")
            ok, text = _handle_detail_field_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in CATEGORY_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare categorie e campi asset.")
                return redirect("assets:asset_list")
            ok, text = _handle_asset_category_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in SIDEBAR_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare il menu sidebar.")
                return redirect("assets:asset_list")
            ok, text = _handle_sidebar_button_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action in HEADER_TOOL_ACTIONS:
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo modificare gli strumenti header.")
                return redirect("assets:asset_list")
            ok, text = _handle_header_tool_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return redirect("assets:asset_list")
        if action == "export_admin_snapshot":
            if not can_manage_custom_fields:
                messages.error(request, "Solo admin puo esportare la configurazione.")
                return redirect("assets:asset_list")
            payload = _build_assets_admin_snapshot()
            content = json.dumps(payload, ensure_ascii=False, indent=2)
            response = HttpResponse(content, content_type="application/json; charset=utf-8")
            response["Content-Disposition"] = (
                f'attachment; filename=\"assets_admin_snapshot_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json\"'
            )
            return response

    form = AssetFilterForm(request.GET or None)
    assets = Asset.objects.select_related("asset_category").all().prefetch_related("endpoints")

    # Contesto "Asset produzione": la vecchia pagina dedicata (work_machine_list)
    # e' confluita qui. ?group=production restringe l'inventario ai soli asset
    # di produzione (CNC / macchine di lavoro / carroponte), riusando la stessa
    # lista, colonne e ricerca degli altri tipi.
    # IMPORTANTE: il filtro di gruppo vale solo per la VISTA di navigazione
    # (nessuna ricerca). Appena c'e' una query testuale la ricerca deve spaziare
    # su TUTTO l'inventario: l'utente cerca un asset ovunque sia classificato,
    # anche se il suo tipo non e' di produzione (es. una macchina registrata con
    # tipo generico). Confinare la ricerca al gruppo nasconderebbe risultati.
    production_group = _clean_string(request.GET.get("group")).lower() == "production"
    search_active = bool(_clean_string(request.GET.get("q")))
    if production_group:
        assets = assets.select_related("work_machine")
        if not search_active:
            assets = assets.filter(asset_type__in=PRODUCTION_ASSET_TYPES)

    if form.is_valid():
        q = _clean_string(form.cleaned_data.get("q"))
        asset_type = _clean_string(form.cleaned_data.get("asset_type"))
        asset_category = form.cleaned_data.get("asset_category")
        reparto = _clean_string(form.cleaned_data.get("reparto"))
        vlan = form.cleaned_data.get("vlan")
        ip = _clean_string(form.cleaned_data.get("ip"))
        cnc_only = bool(form.cleaned_data.get("cnc_only"))
        five_axes_only = bool(form.cleaned_data.get("five_axes_only"))
        tcr_only = bool(form.cleaned_data.get("tcr_only"))

        if q:
            assets = assets.filter(
                Q(asset_tag__icontains=q)
                | Q(internal_number__icontains=q)
                | Q(name__icontains=q)
                | Q(serial_number__icontains=q)
                | Q(manufacturer__icontains=q)
                | Q(model__icontains=q)
                | Q(endpoints__endpoint_name__icontains=q)
                | Q(endpoints__ip__icontains=q)
            )
        if asset_type:
            assets = assets.filter(asset_type=asset_type)
        if asset_category:
            # Filtro per sottoalbero: la categoria selezionata e tutte le sue
            # discendenti (clic su una radice mostra tutti i suoi asset).
            assets = assets.filter(asset_category_id__in=_category_subtree_ids(asset_category))
        if reparto:
            assets = assets.filter(reparto__icontains=reparto)
        if vlan is not None:
            assets = assets.filter(endpoints__vlan=vlan)
        if ip:
            assets = assets.filter(endpoints__ip__icontains=ip)
        # Filtri di produzione (capability del profilo WorkMachine): attivi solo
        # quando l'utente li spunta, indipendentemente dal gruppo selezionato.
        if cnc_only:
            assets = assets.filter(work_machine__cnc_controlled=True)
        if five_axes_only:
            assets = assets.filter(work_machine__five_axes=True)
        if tcr_only:
            assets = assets.filter(work_machine__tcr_enabled=True)

    assets_filtered = assets.distinct().order_by("name", "asset_tag")

    allowed_rows = [10, 25, 50, 100]
    rows = _as_int(request.GET.get("rows"), default=25)
    if rows not in allowed_rows:
        rows = 25
    paginator = Paginator(assets_filtered, rows)
    page_number = _as_int(request.GET.get("page"), default=1)
    page_obj = paginator.get_page(page_number)
    assets = page_obj.object_list
    visible_count = assets_filtered.count()
    page_start = ((page_obj.number - 1) * rows + 1) if visible_count else 0
    page_end = (page_start + len(assets) - 1) if visible_count else 0

    rows_options = [
        {
            "value": value,
            "active": value == rows,
            "url": _query_url(request, rows=value, page=1),
        }
        for value in allowed_rows
    ]

    page_links = []
    if paginator.num_pages > 0:
        start_page = max(1, page_obj.number - 2)
        end_page = min(paginator.num_pages, page_obj.number + 2)
        for number in range(start_page, end_page + 1):
            page_links.append(
                {
                    "number": number,
                    "active": number == page_obj.number,
                    "url": _query_url(request, page=number, rows=rows),
                }
            )

    prev_page_url = _query_url(request, page=page_obj.previous_page_number(), rows=rows) if page_obj.has_previous() else ""
    next_page_url = _query_url(request, page=page_obj.next_page_number(), rows=rows) if page_obj.has_next() else ""
    current_asset_type = _clean_string(request.GET.get("asset_type"))
    selected_asset_category = form.cleaned_data.get("asset_category") if form.is_valid() else None
    asset_list_context_key, asset_list_context_label = _asset_list_context(current_asset_type)
    if production_group:
        asset_list_context_label = "Asset produzione"
    custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
    all_custom_fields = list(AssetCustomField.objects.order_by("sort_order", "id"))
    list_layouts_by_context = {
        row.context_key: row
        for row in _ensure_default_asset_list_layouts()
    }
    current_list_layout = list_layouts_by_context.get(asset_list_context_key)
    valid_list_column_keys = _asset_list_valid_column_keys(custom_fields)
    user_asset_table_layout = _load_user_asset_table_layout(request, asset_list_context_key, valid_list_column_keys)
    suggested_visible_columns = _asset_list_default_columns(current_asset_type)
    suggested_visible_columns = list(dict.fromkeys(suggested_visible_columns))
    if current_list_layout and current_list_layout.is_customized:
        asset_list_default_visible_columns = _sanitize_asset_list_visible_columns(
            current_list_layout.visible_columns,
            valid_list_column_keys,
            fallback=suggested_visible_columns,
        )
    else:
        asset_list_default_visible_columns = _sanitize_asset_list_visible_columns(
            suggested_visible_columns,
            valid_list_column_keys,
            fallback=suggested_visible_columns,
        )
    if user_asset_table_layout.get("visible_columns"):
        asset_list_default_visible_columns = _sanitize_asset_list_visible_columns(
            user_asset_table_layout.get("visible_columns"),
            valid_list_column_keys,
            fallback=asset_list_default_visible_columns,
        )
    asset_list_layout_revision = _asset_list_layout_revision(current_list_layout)
    list_options = list(AssetListOption.objects.order_by("field_key", "sort_order", "value", "id"))
    action_buttons = list(AssetActionButton.objects.order_by("zone", "sort_order", "label", "id"))
    detail_fields = list(AssetDetailField.objects.order_by("section", "asset_scope", "sort_order", "label", "id"))
    asset_categories = list(AssetCategory.objects.order_by("sort_order", "label", "id"))
    asset_category_fields = list(
        AssetCategoryField.objects.select_related("category").order_by(
            "category__sort_order",
            "category__label",
            "sort_order",
            "label",
            "id",
        )
    )
    sidebar_buttons = list(AssetSidebarButton.objects.select_related("parent").order_by("section", "sort_order", "label", "id"))
    sidebar_parent_choices = _sidebar_parent_choices()
    sidebar_target_suggestions, sidebar_active_match_suggestions = _sidebar_input_suggestions()
    for button in action_buttons:
        button.label = _ui_label(button.label)
    for detail_item in detail_fields:
        detail_item.label = _ui_label(detail_item.label)
    for sidebar_item in sidebar_buttons:
        sidebar_item.label = _ui_label(sidebar_item.label)
    for parent_item in sidebar_parent_choices:
        parent_item.label = _ui_label(parent_item.label)
    admin_metrics = {
        "custom_fields_total": len(all_custom_fields),
        "custom_fields_active": sum(1 for row in all_custom_fields if row.is_active),
        "list_options_total": len(list_options),
        "list_options_active": sum(1 for row in list_options if row.is_active),
        "action_buttons_total": len(action_buttons),
        "action_buttons_active": sum(1 for row in action_buttons if row.is_active),
        "detail_fields_total": len(detail_fields),
        "detail_fields_active": sum(1 for row in detail_fields if row.is_active),
        "asset_categories_total": len(asset_categories),
        "asset_categories_active": sum(1 for row in asset_categories if row.is_active),
        "asset_category_fields_total": len(asset_category_fields),
        "asset_category_fields_active": sum(1 for row in asset_category_fields if row.is_active),
        "sidebar_total": len(sidebar_buttons),
        "sidebar_visible": sum(1 for row in sidebar_buttons if row.is_visible),
        "sidebar_hidden": sum(1 for row in sidebar_buttons if not row.is_visible),
    }
    admin_checks = []
    if can_manage_custom_fields:
        if admin_metrics["sidebar_total"] == 0:
            admin_checks.append("Menu laterale personalizzato non configurato: usa il caricamento iniziale.")
        if admin_metrics["action_buttons_active"] == 0:
            admin_checks.append("Nessun pulsante azione attivo sul dettaglio asset.")
        if admin_metrics["detail_fields_active"] == 0:
            admin_checks.append("Nessun campo dettaglio attivo: la scheda asset usera il fallback predefinito.")
        if admin_metrics["asset_categories_active"] == 0:
            admin_checks.append("Nessuna categoria asset attiva: il modulo usa solo le tipologie tecniche standard.")
        if admin_metrics["custom_fields_active"] == 0:
            admin_checks.append("Nessun campo personalizzato attivo: verifica se e voluto.")
        if not admin_checks:
            admin_checks.append("Configurazione amministratore completa e coerente.")
    # KPI: scope by structural filters (category + type) to reflect the current view
    kpi_qs = Asset.objects.all()
    if form.is_valid():
        _kf_type = _clean_string(form.cleaned_data.get("asset_type"))
        _kf_cat = form.cleaned_data.get("asset_category")
        if _kf_type:
            kpi_qs = kpi_qs.filter(asset_type=_kf_type)
        if _kf_cat:
            kpi_qs = kpi_qs.filter(asset_category_id__in=_category_subtree_ids(_kf_cat))

    total_assets = kpi_qs.count()
    in_use_count = kpi_qs.filter(status=Asset.STATUS_IN_USE).count()
    in_repair_count = kpi_qs.filter(status=Asset.STATUS_IN_REPAIR).count()
    assigned_count = kpi_qs.exclude(assignment_to__isnull=True).exclude(assignment_to="").count()
    _kpi_wo_qs = WorkOrder.objects.filter(asset__in=kpi_qs)
    open_wo_count = _kpi_wo_qs.filter(status=WorkOrder.STATUS_OPEN).count()
    maintenance_due_count = _kpi_wo_qs.filter(
        status=WorkOrder.STATUS_OPEN,
        opened_at__lt=timezone.now() - timedelta(days=get_workorder_overdue_days()),
    ).count()
    work_machine_total = Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES).count()

    health_percent = 0.0
    in_use_percent = 0.0
    if total_assets > 0:
        health_percent = max(0.0, round((1 - (in_repair_count / total_assets)) * 100, 1))
        in_use_percent = max(0.0, min(100.0, round((in_use_count / total_assets) * 100, 1)))
    risk_count = in_repair_count + open_wo_count

    phase_1_count = Asset.objects.filter(
        asset_type__in=[
            Asset.TYPE_PC,
            Asset.TYPE_NOTEBOOK,
            Asset.TYPE_STAMPANTE,
            Asset.TYPE_HW,
        ]
    ).count()
    phase_2_count = Asset.objects.filter(
        asset_type__in=[
            Asset.TYPE_SERVER,
            Asset.TYPE_VM,
            Asset.TYPE_FIREWALL,
            Asset.TYPE_CCTV,
        ]
    ).count()
    phase_3_count = Asset.objects.filter(
        asset_type__in=[Asset.TYPE_CNC, Asset.TYPE_WORK_MACHINE, Asset.TYPE_OTHER],
    ).count()
    lifecycle_total = phase_1_count + phase_2_count + phase_3_count
    if lifecycle_total <= 0:
        lifecycle_total = 1
    lifecycle_phase_1 = int(round((phase_1_count / lifecycle_total) * 100))
    lifecycle_phase_2 = int(round((phase_2_count / lifecycle_total) * 100))
    lifecycle_phase_3 = int(round((phase_3_count / lifecycle_total) * 100))

    for asset in assets:
        endpoint_summary = _asset_endpoint_column_summary(asset)
        asset.endpoint_vlan_display = endpoint_summary["vlan"]
        asset.endpoint_ip_display = endpoint_summary["ip"]

    recent_alerts: list[dict[str, str]] = []
    open_wo_alerts = _dashboard_open_workorder_alert_rows(limit=4)
    for workorder in open_wo_alerts:
        is_critical = workorder.opened_at < timezone.now() - timedelta(days=14)
        recent_alerts.append(
            {
                "title": workorder.title or f"Intervento su {workorder.asset.asset_tag}",
                "message": _coalesce_str(workorder.description, f"Asset {workorder.asset.asset_tag} richiede attenzione."),
                "time": workorder.opened_at.strftime("%d-%m-%Y %H:%M"),
                "level": "critical" if is_critical else "warning",
            }
        )
    if len(recent_alerts) < 5:
        repair_assets = list(
            Asset.objects.filter(status=Asset.STATUS_IN_REPAIR)
            .order_by("-updated_at")
            .values("asset_tag", "name", "updated_at")[: (5 - len(recent_alerts))]
        )
        for row in repair_assets:
            recent_alerts.append(
                {
                    "title": f"Asset in riparazione: {row.get('asset_tag')}",
                    "message": _coalesce_str(row.get("name"), "Asset segnalato in riparazione."),
                    "time": row["updated_at"].strftime("%d-%m-%Y %H:%M") if row.get("updated_at") else "-",
                    "level": "warning",
                }
            )
    if not recent_alerts:
        recent_alerts.append(
            {
                "title": "Nessun alert critico",
                "message": "La situazione asset e stabile.",
                "time": timezone.now().strftime("%d-%m-%Y %H:%M"),
                "level": "ok",
            }
        )

    return render(
        request,
        "assets/pages/asset_list.html",
        {
            "page_title": "Asset produzione" if production_group else "Inventario asset",
            "production_group": production_group,
            "filters_form": form,
            "assets": assets,
            "total_assets": total_assets,
            "in_use_count": in_use_count,
            "in_repair_count": in_repair_count,
            "open_wo_count": open_wo_count,
            "visible_count": visible_count,
            "rows": rows,
            "rows_options": rows_options,
            "page_obj": page_obj,
            "page_links": page_links,
            "prev_page_url": prev_page_url,
            "next_page_url": next_page_url,
            "page_start": page_start,
            "page_end": page_end,
            "custom_fields": custom_fields,
            "all_custom_fields": all_custom_fields,
            "asset_list_context_key": asset_list_context_key,
            "asset_list_context_label": asset_list_context_label,
            "selected_asset_category": selected_asset_category,
            "asset_list_default_visible_columns": asset_list_default_visible_columns,
            "asset_list_default_visible_columns_json": json.dumps(asset_list_default_visible_columns),
            "asset_table_saved_layout_json": json.dumps(user_asset_table_layout),
            "asset_table_layout_can_persist": bool(_asset_table_layout_storage_user_id(request) is not None),
            "asset_list_layout_revision": asset_list_layout_revision,
            "asset_list_layout_manage_url": _asset_list_layout_manage_url(request, asset_list_context_key),
            "asset_list_layout_is_customized": bool(current_list_layout and current_list_layout.is_customized),
            "custom_type_choices": _ui_choices(AssetCustomField.TYPE_CHOICES),
            "list_options": list_options,
            "list_option_choices": _ui_choices(AssetListOption.FIELD_CHOICES),
            "action_buttons": action_buttons,
            "button_zone_choices": _ui_choices(AssetActionButton.ZONE_CHOICES),
            "button_action_choices": _ui_choices(AssetActionButton.ACTION_CHOICES),
            "button_style_choices": _ui_choices(AssetActionButton.STYLE_CHOICES),
            "detail_fields": detail_fields,
            "asset_categories": asset_categories,
            "asset_category_fields": asset_category_fields,
            "asset_category_type_choices": _ui_choices(Asset.TYPE_CHOICES),
            "asset_category_field_type_choices": _ui_choices(AssetCategoryField.TYPE_CHOICES),
            "detail_section_choices": _ui_choices(AssetDetailField.SECTION_CHOICES),
            "detail_scope_choices": _ui_choices(AssetDetailField.SCOPE_CHOICES),
            "detail_format_choices": _ui_choices(AssetDetailField.FORMAT_CHOICES),
            "detail_source_choices": _asset_detail_source_choices(),
            "sidebar_buttons": sidebar_buttons,
            "sidebar_parent_choices": sidebar_parent_choices,
            "sidebar_section_choices": _ui_choices(AssetSidebarButton.SECTION_CHOICES),
            "sidebar_target_suggestions": sidebar_target_suggestions,
            "sidebar_active_match_suggestions": sidebar_active_match_suggestions,
            "can_manage_custom_fields": can_manage_custom_fields,
            "can_gestione_admin": user_can_modulo_action(request, "assets", "admin_assets"),
            "header_tools": list(AssetHeaderTool.objects.order_by("sort_order", "code")),
            **_header_tool_visibility(can_manage_custom_fields),
            "admin_metrics": admin_metrics,
            "admin_checks": admin_checks,
            "table_colspan": 7 + len(custom_fields),
            "health_percent": health_percent,
            "in_use_percent": in_use_percent,
            "risk_count": risk_count,
            "maintenance_due_count": maintenance_due_count,
            "assigned_count": assigned_count,
            "work_machine_total": work_machine_total,
            "phase_1_count": phase_1_count,
            "phase_2_count": phase_2_count,
            "phase_3_count": phase_3_count,
            "lifecycle_phase_1": lifecycle_phase_1,
            "lifecycle_phase_2": lifecycle_phase_2,
            "lifecycle_phase_3": lifecycle_phase_3,
            "recent_alerts": recent_alerts,
            **_assets_shell_context(request, rows=rows),
        },
    )


@login_required
def asset_detail_layout_admin(request: HttpRequest) -> HttpResponse:
    if not _can_manage_asset_detail_layout(request):
        messages.error(request, "Non hai i permessi per configurare il layout del dettaglio asset.")
        return redirect("assets:asset_list")

    preview_asset_id = _as_int(request.POST.get("asset_id") if request.method == "POST" else request.GET.get("asset"), default=0)
    preview_asset = (
        Asset.objects.select_related("asset_category")
        .filter(pk=preview_asset_id)
        .first()
    )

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action in DETAIL_LAYOUT_ACTIONS:
            ok, text = _handle_detail_section_layout_request(request)
        elif action in DETAIL_FIELD_ACTIONS:
            ok, text = _handle_detail_field_request(request)
        elif action == "update_asset_category_field":
            ok, text = _handle_asset_category_request(request)
        else:
            ok, text = False, "Azione layout dettaglio non riconosciuta."
        if ok:
            messages.success(request, text)
        else:
            messages.error(request, text)
        redirect_url = reverse("assets:asset_detail_layout_admin")
        if preview_asset_id:
            redirect_url = f"{redirect_url}?asset={preview_asset_id}"
        return redirect(redirect_url)

    section_layouts = _ensure_default_asset_detail_section_layouts()
    detail_fields = list(AssetDetailField.objects.order_by("section", "asset_scope", "sort_order", "label", "id"))
    asset_categories = list(
        AssetCategory.objects.prefetch_related("category_fields").order_by("sort_order", "label", "id")
    )
    detail_source_help = [
        "asset:manufacturer",
        "asset:model",
        "asset:serial_number",
        "it:cpu",
        "work_machine:x_mm",
        "custom:centro_costo",
        "computed:travel_xyz",
    ]

    return render(
        request,
        "assets/pages/asset_detail_layout_admin.html",
        {
            "page_title": "Configura dettaglio asset",
            "section_layouts": section_layouts,
            "detail_fields": detail_fields,
            "asset_categories": asset_categories,
            "detail_section_choices": _ui_choices(AssetDetailField.SECTION_CHOICES),
            "detail_scope_choices": _ui_choices(AssetDetailField.SCOPE_CHOICES),
            "detail_format_choices": _ui_choices(AssetDetailField.FORMAT_CHOICES),
            "detail_card_size_choices": _ui_choices(AssetDetailField.CARD_SIZE_CHOICES),
            "detail_section_layout_choices": _ui_choices(AssetDetailSectionLayout.SECTION_CHOICES),
            "detail_source_help": detail_source_help,
            "preview_asset": preview_asset,
            "preview_asset_url": reverse("assets:asset_view", kwargs={"id": preview_asset.id}) if preview_asset else "",
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


def _build_asset_detail_section_cards(
    *,
    detail_specs_title: str,
    detail_timeline_title: str,
    detail_maintenance_title: str,
    detail_tickets_title: str,
    profile_card_title: str,
    detail_assignment_title: str,
    quick_action_buttons: list[dict],
    map_marker,
    doc_category_labels: dict,
    spec_pairs: list[tuple[str, str]],
    profile_rows: list[dict[str, str]],
    license_rows: list[dict[str, object]],
    ticket_rows: list[dict[str, object]],
) -> list[dict[str, str]]:
    cards_by_code = {
        AssetDetailSectionLayout.SECTION_SPECS: {
            "code": AssetDetailSectionLayout.SECTION_SPECS,
            "title": detail_specs_title,
            # Fusa nella card "Anagrafica e assegnazione" come gruppo Specifiche tecniche.
            "render": False,
        },
        AssetDetailSectionLayout.SECTION_TIMELINE: {
            "code": AssetDetailSectionLayout.SECTION_TIMELINE,
            "title": detail_timeline_title,
            # Fusa nella card "Anagrafica e assegnazione" (gruppo storico ciclo di vita).
            "render": False,
        },
        AssetDetailSectionLayout.SECTION_MAINTENANCE: {
            "code": AssetDetailSectionLayout.SECTION_MAINTENANCE,
            "title": detail_maintenance_title,
            "render": True,
        },
        AssetDetailSectionLayout.SECTION_TICKETS: {
            "code": AssetDetailSectionLayout.SECTION_TICKETS,
            "title": detail_tickets_title,
            # Il "Registro manutenzione" include SOLO i ticket MAN con
            # include_in_maintenance_register=True: i ticket IT (e i MAN non
            # marcati) non comparirebbero da nessuna parte. La card resta quindi
            # la vista completa dei ticket collegati all'asset.
            "render": bool(ticket_rows),
        },
        AssetDetailSectionLayout.SECTION_PROFILE: {
            "code": AssetDetailSectionLayout.SECTION_PROFILE,
            "title": profile_card_title,
            "render": bool(profile_rows),
        },
        AssetDetailSectionLayout.SECTION_LICENSES: {
            "code": AssetDetailSectionLayout.SECTION_LICENSES,
            "title": "Licenze software",
            "render": bool(license_rows),
        },
        AssetDetailSectionLayout.SECTION_PERIODIC: {
            "code": AssetDetailSectionLayout.SECTION_PERIODIC,
            "title": "Manutenzione periodica",
            # Fusa nel "Registro manutenzione" come accordion dedicato.
            "render": False,
        },
        AssetDetailSectionLayout.SECTION_QR: {
            "code": AssetDetailSectionLayout.SECTION_QR,
            "title": "QR asset",
            "render": True,
        },
        AssetDetailSectionLayout.SECTION_QUICK_ACTIONS: {
            "code": AssetDetailSectionLayout.SECTION_QUICK_ACTIONS,
            "title": "Azioni rapide",
            "render": bool(quick_action_buttons) and len(quick_action_buttons) <= 2,
        },
        AssetDetailSectionLayout.SECTION_ASSIGNMENT: {
            "code": AssetDetailSectionLayout.SECTION_ASSIGNMENT,
            "title": detail_assignment_title,
            # Fusa nella card PROFILE ("Anagrafica e assegnazione"): non renderizzata
            # come card a sé per evitare duplicazione (reparto/posizione/assegnatario).
            "render": False,
        },
        AssetDetailSectionLayout.SECTION_MAP: {
            "code": AssetDetailSectionLayout.SECTION_MAP,
            "title": "Posizione in officina",
            "render": bool(map_marker),
        },
        AssetDetailSectionLayout.SECTION_DOCUMENTS: {
            "code": AssetDetailSectionLayout.SECTION_DOCUMENTS,
            "title": "Documenti",
            "render": bool(doc_category_labels),
        },
    }

    cards: list[dict[str, str]] = []
    for layout in _ensure_default_asset_detail_section_layouts():
        payload = cards_by_code.get(layout.code)
        if payload is None or not layout.is_visible or not payload["render"]:
            continue
        cards.append(
            {
                **payload,
                "size_class": _detail_grid_size_class(layout.grid_size),
            }
        )
    return cards


def _can_manage_ticket_type_for_asset_view(request: HttpRequest, ticket_type: str) -> bool:
    from tickets.models import TicketImpostazioni
    if not request.user.is_authenticated:
        return False
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    if request.user.is_superuser or is_legacy_admin(legacy_user):
        return True
    cfg = TicketImpostazioni.objects.filter(tipo=ticket_type).first()
    if not cfg:
        return False
    acl = cfg.acl_gestione or []
    username = request.user.get_username().strip().lower()
    email = (request.user.email or "").strip().lower()
    return any(str(value).strip().lower() in {username, email} for value in acl if value)


def _ticket_belongs_to_request_user(request: HttpRequest, ticket: Ticket) -> bool:
    legacy_user = getattr(request, "legacy_user", None) or get_legacy_user(request.user)
    request_name = (
        (getattr(legacy_user, "nome", "") or "").strip()
        or request.user.get_full_name()
        or request.user.get_username()
    )
    request_email = (
        (getattr(legacy_user, "email", "") or "").strip().lower()
        or (request.user.email or "").strip().lower()
    )
    request_legacy_id = getattr(legacy_user, "id", None)
    ticket_email = (ticket.richiedente_email or "").strip().lower()
    return bool(
        (request_legacy_id and ticket.richiedente_legacy_user_id == request_legacy_id)
        or (request_name and ticket.richiedente_nome == request_name)
        or (request_email and ticket_email == request_email)
    )


def _asset_ticket_detail_url(
    request: HttpRequest,
    ticket: Ticket,
    manageable_ticket_types: set[str],
) -> str:
    if ticket.tipo in manageable_ticket_types:
        return reverse("tickets:gestione_detail", kwargs={"pk": ticket.pk})
    if _ticket_belongs_to_request_user(request, ticket):
        return reverse("tickets:detail", kwargs={"pk": ticket.pk})
    return ""


@login_required
def asset_detail(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(
        Asset.objects.select_related("asset_category", "it_details", "work_machine", "prodotto_chimico", "prodotto_chimico__reparto").prefetch_related(
            "components",
            "administrative_deadlines",
            "administrative_deadlines__component",
            "endpoints",
            "workorders",
            "documents",
            "tickets",
            "periodic_verifications",
            "periodic_verifications__supplier",
            "software_licenses",
        ),
        pk=id,
    )
    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action == "prepare_planned_maintenance_report":
            base_rule_id = _as_int(request.POST.get("base_rule_id"), default=0)
            schedule_row = _maintenance_schedule_row_for_asset_rule(
                asset=asset,
                base_rule_id=base_rule_id,
            )
            if schedule_row is None:
                messages.error(request, "La manutenzione pianificata selezionata non è disponibile per questo asset.")
                return redirect("assets:asset_view", id=asset.id)

            workorder = None
            created = False
            with transaction.atomic():
                # Serializza il fallback manuale sulla regola: due click ravvicinati
                # devono aprire lo stesso rapporto, non due OdL concorrenti.
                MaintenanceRule.objects.select_for_update().get(pk=base_rule_id)
                workorder = (
                    WorkOrder.objects.filter(
                        asset=asset,
                        maintenance_rule_id=base_rule_id,
                        status=WorkOrder.STATUS_OPEN,
                    )
                    .order_by("opened_at", "id")
                    .first()
                )
                if workorder is None:
                    rule = schedule_row["base_rule"]
                    template = schedule_row["effective_intervention_template"]
                    workorder = WorkOrder(
                        asset=asset,
                        maintenance_rule=rule,
                        origin=WorkOrder.ORIGIN_PERIODIC,
                        kind=getattr(template, "workorder_kind", WorkOrder.KIND_PREVENTIVE),
                        status=WorkOrder.STATUS_OPEN,
                        title=f"{template.label} — {asset.asset_tag}",
                        description=getattr(template, "description", "") or "",
                        assigned_to=rule.assigned_to,
                        supplier=rule.supplier,
                        reference_batch=f"asset-{timezone.localdate():%Y%m%d}",
                    )
                    workorder.full_clean()
                    workorder.save()
                    copied_steps = copy_template_checklist_to_workorder(
                        workorder,
                        template_id=getattr(template, "pk", None),
                    )
                    log_note = "Rapporto di manutenzione pianificata generato dalla scheda asset."
                    if copied_steps:
                        log_note = f"{log_note} Checklist precompilata: {copied_steps} attività."
                    WorkOrderLog.objects.create(
                        work_order=workorder,
                        note=log_note,
                        author=request.user if request.user.is_authenticated else None,
                    )
                    created = True

            if created:
                notify_workorder_assigned(
                    workorder,
                    actor=request.user if request.user.is_authenticated else None,
                )
                messages.success(request, f"Rapporto #{workorder.id} generato: completa i dati e chiudilo.")
            else:
                messages.info(request, f"Il rapporto #{workorder.id} era già aperto: completa i dati e chiudilo.")
            return redirect("assets:wo_close", id=workorder.id)
        if action == "add_asset_timeline_entry":
            if not _can_manage_asset_timeline_entries(request):
                messages.error(request, "Permessi insufficienti per inserire voci nella timeline di vita.")
                return redirect("assets:asset_view", id=asset.id)
            if not _asset_timeline_manual_enabled(asset):
                messages.error(request, "L'inserimento manuale in timeline e' disattivato per questa categoria di asset.")
                return redirect("assets:asset_view", id=asset.id)
            entry_payload, payload_error = _asset_timeline_entry_payload(request)
            if entry_payload is None:
                messages.error(request, payload_error)
            else:
                entry = AssetTimelineEntry.objects.create(
                    asset=asset,
                    created_by=request.user if request.user.is_authenticated else None,
                    **entry_payload,
                )
                log_action(
                    request,
                    "add_asset_timeline_entry",
                    "assets",
                    {
                        "asset_id": asset.id,
                        "entry_id": entry.id,
                        "event_date": entry.event_date.isoformat(),
                        "title": entry.title,
                    },
                    oggetto_tipo=AUDIT_OGGETTO_ASSET,
                    oggetto_id=asset.id,
                )
                messages.success(request, f"Voce \"{entry.title}\" aggiunta alla timeline di vita.")
            return redirect("assets:asset_view", id=asset.id)
        if action in {"update_asset_timeline_entry", "delete_asset_timeline_entry"}:
            # Correzione e rimozione non dipendono dal flag di categoria: se
            # l'inserimento manuale viene spento, le voci gia' registrate devono
            # restare sistemabili da chi ha il permesso.
            if not _can_manage_asset_timeline_entries(request):
                messages.error(request, "Permessi insufficienti per gestire le voci della timeline di vita.")
                return redirect("assets:asset_view", id=asset.id)
            entry_id = _as_int(request.POST.get("entry_id"), default=0)
            entry = asset.timeline_entries.filter(pk=entry_id).first() if entry_id > 0 else None
            if entry is None:
                messages.error(request, "Voce di timeline non trovata.")
                return redirect("assets:asset_view", id=asset.id)
            if action == "delete_asset_timeline_entry":
                entry_label = entry.title
                entry.delete()
                log_action(
                    request,
                    "delete_asset_timeline_entry",
                    "assets",
                    {"asset_id": asset.id, "entry_id": entry_id, "title": entry_label},
                    oggetto_tipo=AUDIT_OGGETTO_ASSET,
                    oggetto_id=asset.id,
                )
                messages.success(request, f"Voce \"{entry_label}\" rimossa dalla timeline di vita.")
                return redirect("assets:asset_view", id=asset.id)
            entry_payload, payload_error = _asset_timeline_entry_payload(request)
            if entry_payload is None:
                messages.error(request, payload_error)
            else:
                for field_name, field_value in entry_payload.items():
                    setattr(entry, field_name, field_value)
                entry.save(update_fields=[*entry_payload.keys(), "updated_at"])
                log_action(
                    request,
                    "update_asset_timeline_entry",
                    "assets",
                    {
                        "asset_id": asset.id,
                        "entry_id": entry.id,
                        "event_date": entry.event_date.isoformat(),
                        "title": entry.title,
                    },
                    oggetto_tipo=AUDIT_OGGETTO_ASSET,
                    oggetto_id=asset.id,
                )
                messages.success(request, f"Voce \"{entry.title}\" aggiornata.")
            return redirect("assets:asset_view", id=asset.id)
        if action == "hide_asset_timeline_event":
            # Gli eventi automatici non sono righe di database: "eliminarli"
            # significa registrare che su questo asset non vanno piu' mostrati.
            if not _can_manage_asset_timeline_entries(request):
                messages.error(request, "Permessi insufficienti per gestire le voci della timeline di vita.")
                return redirect("assets:asset_view", id=asset.id)
            event_key = _clean_string(request.POST.get("event_key"))[:40]
            event_labels = dict(AssetTimelineHiddenEvent.KEY_CHOICES)
            if event_key not in event_labels:
                messages.error(request, "Voce di timeline non trovata.")
                return redirect("assets:asset_view", id=asset.id)
            _, hidden_created = AssetTimelineHiddenEvent.objects.get_or_create(
                asset=asset,
                event_key=event_key,
                defaults={"hidden_by": request.user if request.user.is_authenticated else None},
            )
            if hidden_created:
                log_action(
                    request,
                    "hide_asset_timeline_event",
                    "assets",
                    {"asset_id": asset.id, "event_key": event_key},
                    oggetto_tipo=AUDIT_OGGETTO_ASSET,
                    oggetto_id=asset.id,
                )
            messages.success(request, f"Voce \"{event_labels[event_key]}\" rimossa dalla timeline di vita.")
            return redirect("assets:asset_view", id=asset.id)
        if action == "upload_asset_documents":
            uploads, upload_errors = _validate_asset_document_uploads(request, asset)
            upload_count = _asset_document_upload_count(uploads)
            for error in upload_errors:
                messages.error(request, error)
            if not upload_errors and upload_count <= 0:
                messages.error(request, "Seleziona almeno un file da caricare.")
            if not upload_errors and upload_count > 0:
                _apply_asset_document_changes(
                    asset,
                    uploads=uploads,
                    remove_ids=set(),
                    actor=request.user,
                )
                messages.success(request, f"{upload_count} file caricato/i in archivio.")
            return redirect("assets:asset_view", id=asset.id)
        if action == "delete_asset_document":
            document_id = _as_int(request.POST.get("document_id"), default=0)
            document = asset.documents.filter(id=document_id).first() if document_id > 0 else None
            if document is None:
                messages.error(request, "Documento non trovato.")
            else:
                doc_label = document.original_name or Path(document.file.name).name
                doc_category = document.category
                _apply_asset_document_changes(
                    asset,
                    uploads={},
                    remove_ids={document.id},
                    actor=request.user,
                )
                log_action(
                    request,
                    "delete_asset_document",
                    "assets",
                    {
                        "document_id": document_id,
                        "asset_id": asset.id,
                        "category": doc_category,
                        "filename": doc_label,
                    },
                )
                messages.success(request, f"Documento \"{doc_label}\" eliminato.")
            return redirect("assets:asset_view", id=asset.id)
        if action == "add_asset_document_folder":
            if not _can_manage_asset_document_folders(request):
                messages.error(request, "Permessi insufficienti per gestire le cartelle documento.")
                return redirect("assets:asset_view", id=asset.id)
            category = asset.asset_category
            if category is None:
                messages.error(request, "Assegna prima una categoria all'asset per aggiungere cartelle documento.")
                return redirect("assets:asset_view", id=asset.id)
            raw_name = _clean_string(request.POST.get("folder_name"))[:120]
            slug = slugify(raw_name)[:60]
            base_codes = {code.lower() for code, _ in AssetDocument.CATEGORY_CHOICES}
            if not raw_name or not slug:
                messages.error(request, "Nome cartella non valido.")
            elif slug in base_codes:
                messages.error(request, "Esiste gia una cartella di base con questo nome.")
            elif category.document_folders.filter(slug=slug).exists():
                messages.error(request, "Esiste gia una cartella documento con questo nome per la categoria.")
            else:
                next_order = (category.document_folders.aggregate(m=Max("order"))["m"] or 0) + 1
                AssetCategoryDocumentFolder.objects.create(
                    category=category, name=raw_name, slug=slug, order=next_order
                )
                log_action(
                    request,
                    "add_asset_document_folder",
                    "assets",
                    {"asset_id": asset.id, "category_id": category.id, "slug": slug, "name": raw_name},
                )
                messages.success(
                    request,
                    f"Cartella documento \"{raw_name}\" aggiunta alla categoria {category.label}.",
                )
            return redirect("assets:asset_view", id=asset.id)
        if action == "deactivate_asset_document_folder":
            if not _can_manage_asset_document_folders(request):
                messages.error(request, "Permessi insufficienti per gestire le cartelle documento.")
                return redirect("assets:asset_view", id=asset.id)
            folder_id = _as_int(request.POST.get("folder_id"), default=0)
            folder = None
            if folder_id > 0 and asset.asset_category is not None:
                folder = asset.asset_category.document_folders.filter(id=folder_id, is_active=True).first()
            if folder is None:
                messages.error(request, "Cartella documento non trovata.")
            elif AssetDocument.objects.filter(
                asset__asset_category=folder.category, category=folder.slug
            ).exists():
                messages.error(request, "Impossibile disattivare: la cartella contiene ancora documenti.")
            else:
                folder.is_active = False
                folder.save(update_fields=["is_active"])
                log_action(
                    request,
                    "deactivate_asset_document_folder",
                    "assets",
                    {"asset_id": asset.id, "folder_id": folder.id, "slug": folder.slug},
                )
                messages.success(request, f"Cartella documento \"{folder.name}\" disattivata.")
            return redirect("assets:asset_view", id=asset.id)

    recent_workorders = asset.workorders.select_related("asset").all()[:10]
    component_rows = list(asset.components.all())
    administrative_deadlines = list(
        asset.administrative_deadlines.all().order_by("due_date", "title", "id")
    )
    custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
    custom_fields_by_code = {field.code: field for field in custom_fields}
    extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}
    mapped_keys = {field.code for field in custom_fields} | {field.label for field in custom_fields}
    unmapped_extra = [(k, v) for k, v in extra.items() if k not in mapped_keys]

    now = timezone.now()
    today = timezone.localdate()
    age = max(0, int((now - asset.updated_at).total_seconds() // 60)) if asset.updated_at else 0
    if age < 1:
        sync_text = "Ultimo sync: adesso"
    elif age < 60:
        sync_text = f"Ultimo sync: {age} minuti fa"
    else:
        hours = age // 60
        sync_text = f"Ultimo sync: {hours} ore fa"

    it_details = getattr(asset, "it_details", None)
    work_machine = getattr(asset, "work_machine", None)
    travel_xyz = "N/D"
    machine_flags: list[str] = []
    if isinstance(work_machine, WorkMachine):
        travel_parts = [str(value) for value in [work_machine.x_mm, work_machine.y_mm, work_machine.z_mm] if value is not None]
        travel_xyz = " x ".join(travel_parts) + " mm" if travel_parts else "N/D"
        if work_machine.tcr_enabled:
            machine_flags.append("TCR")
        if work_machine.cnc_controlled:
            machine_flags.append("CNC")
        if work_machine.five_axes:
            machine_flags.append("5 assi")

    if isinstance(work_machine, WorkMachine):
        default_detail_metrics = [
            {"label": "Corse XYZ", "value": travel_xyz, "size": AssetDetailField.CARD_THIRD},
            {"label": "Anno macchina", "value": _coalesce_str(work_machine.year, "N/D"), "size": AssetDetailField.CARD_THIRD},
            {
                "label": "Configurazione",
                "value": ", ".join(machine_flags) if machine_flags else "Standard",
                "size": AssetDetailField.CARD_THIRD,
            },
        ]
        default_spec_pairs = [
            # Identità (produttore/modello/seriale/reparto) spostata nella card
            # "Anagrafica e assegnazione"; qui restano solo le caratteristiche tecniche.
            ("Corsa X", _format_asset_detail_value(work_machine.x_mm, AssetDetailField.FORMAT_MM)),
            ("Corsa Y", _format_asset_detail_value(work_machine.y_mm, AssetDetailField.FORMAT_MM)),
            ("Corsa Z", _format_asset_detail_value(work_machine.z_mm, AssetDetailField.FORMAT_MM)),
            ("Diametro", _format_asset_detail_value(work_machine.diameter_mm, AssetDetailField.FORMAT_MM)),
            ("Mandrino", _format_asset_detail_value(work_machine.spindle_mm, AssetDetailField.FORMAT_MM)),
            ("Anno", _coalesce_str(work_machine.year, "N/D")),
            ("TMC", _coalesce_str(work_machine.tmc, "N/D")),
            ("TCR", _format_asset_detail_value(work_machine.tcr_enabled, AssetDetailField.FORMAT_BOOL)),
            ("Pressione", _format_asset_detail_value(work_machine.pressure_bar, AssetDetailField.FORMAT_BAR)),
            ("CNC", _format_asset_detail_value(work_machine.cnc_controlled, AssetDetailField.FORMAT_BOOL)),
            ("5 assi", _format_asset_detail_value(work_machine.five_axes, AssetDetailField.FORMAT_BOOL)),
            ("Accuracy from", _coalesce_str(work_machine.accuracy_from, "N/D")),
            ("Prossima manutenzione", _format_asset_detail_value(work_machine.next_maintenance_date, AssetDetailField.FORMAT_DATE)),
            ("Soglia reminder", f"{work_machine.maintenance_reminder_days} gg"),
        ]
        # Solo identità: i campi tecnici (TCR/CNC/5 assi/manutenzione/accuracy)
        # restano nella card "Caratteristiche tecniche" per evitare duplicazione.
        default_profile_rows = [
            {"label": "Tag asset", "value": asset.asset_tag},
            {"label": "Produttore", "value": _coalesce_str(asset.manufacturer, "-")},
            {"label": "Modello", "value": _coalesce_str(asset.model, asset.name, "-")},
            {"label": "Numero seriale", "value": _coalesce_str(asset.serial_number, "-")},
            {"label": "Reparto macchina", "value": _coalesce_str(asset.reparto, "-")},
        ]
        profile_card_title = "Anagrafica e assegnazione"
    else:
        metric_battery = _coalesce_str(extra.get("battery_health"), extra.get("batteria"), "N/D")
        metric_cpu = _coalesce_str(extra.get("avg_cpu_load"), extra.get("cpu_load"), "N/D")
        metric_storage = _coalesce_str(extra.get("storage_free"), extra.get("free_storage"), it_details.disco if it_details else "", "N/D")
        default_detail_metrics = [
            {"label": "Salute batteria", "value": metric_battery, "size": AssetDetailField.CARD_THIRD},
            {"label": "Carico medio CPU", "value": metric_cpu, "size": AssetDetailField.CARD_THIRD},
            {"label": "Spazio libero", "value": metric_storage, "size": AssetDetailField.CARD_THIRD},
        ]
        default_spec_pairs = [
            ("Processore", _coalesce_str(it_details.cpu if it_details else "", asset.model, "N/D")),
            ("Numero seriale", _coalesce_str(asset.serial_number, "N/D")),
            ("Memoria", _coalesce_str(it_details.ram if it_details else "", "N/D")),
            ("Sistema operativo", _coalesce_str(it_details.os if it_details else "", "N/D")),
            ("Archiviazione", _coalesce_str(it_details.disco if it_details else "", "N/D")),
            ("Grafica", _coalesce_str(extra.get("graphics"), "N/D")),
            ("Schermo", _coalesce_str(extra.get("display"), "N/D")),
            ("Data acquisto", _coalesce_str(asset.purchase_date.strftime("%d-%m-%Y") if asset.purchase_date else "", extra.get("purchase_date"), asset.created_at.strftime("%d-%m-%Y") if asset.created_at else "", "N/D")),
            ("Data produzione", _coalesce_str(asset.production_date.strftime("%d-%m-%Y") if asset.production_date else "", extra.get("production_date"), "N/D")),
        ]
        default_profile_rows = [
            {"label": "Tag asset", "value": asset.asset_tag},
            {"label": "Produttore", "value": _coalesce_str(asset.manufacturer, "-")},
            {"label": "Modello", "value": _coalesce_str(asset.model, "-")},
            {"label": "Ultimo sync", "value": sync_text},
        ]
        profile_card_title = "Anagrafica e assegnazione"

    default_assignment_rows = [
        {"label": "Reparto", "value": _coalesce_str(asset.assignment_reparto, "-")},
        {"label": "Posizione", "value": _coalesce_str(asset.assignment_location, "-")},
        {"label": "Assegnato a", "value": _coalesce_str(asset.assignment_to, "Non assegnato")},
        {"label": "Assegnato dal", "value": asset.updated_at.strftime("%d-%m-%Y") if asset.updated_at else "-"},
    ]
    category_detail_sections = _build_asset_category_detail_sections(asset, extra)

    configured_sections, has_matching_detail_layout = _build_configured_asset_detail_sections(
        asset=asset,
        it_details=it_details,
        work_machine=work_machine,
        extra=extra,
        custom_fields_by_code=custom_fields_by_code,
        sync_text=sync_text,
    )
    if has_matching_detail_layout:
        detail_metrics = configured_sections.get(AssetDetailField.SECTION_METRICS, [])
        spec_rows = configured_sections.get(AssetDetailField.SECTION_SPECS, [])
        profile_rows = configured_sections.get(AssetDetailField.SECTION_PROFILE, [])
        assignment_rows = configured_sections.get(AssetDetailField.SECTION_ASSIGNMENT, [])
    else:
        detail_metrics = default_detail_metrics
        spec_rows = [
            {"label": key, "value": value}
            for key, value in default_spec_pairs
            if not _is_empty_asset_detail_value(value)
        ]
        profile_rows = default_profile_rows
        assignment_rows = default_assignment_rows
    detail_metrics = [*detail_metrics, *category_detail_sections.get(AssetDetailField.SECTION_METRICS, [])]
    spec_rows = [*spec_rows, *category_detail_sections.get(AssetDetailField.SECTION_SPECS, [])]
    profile_rows = [*profile_rows, *category_detail_sections.get(AssetDetailField.SECTION_PROFILE, [])]
    assignment_rows = [*assignment_rows, *category_detail_sections.get(AssetDetailField.SECTION_ASSIGNMENT, [])]
    for metric in detail_metrics:
        metric["size_class"] = _detail_grid_size_class(str(metric.get("size") or ""))
    spec_pairs = [(row["label"], row["value"]) for row in spec_rows]
    detail_specs_title = _coalesce_str(getattr(asset.asset_category, "detail_specs_title", ""), "Specifiche tecniche")
    profile_card_title = _coalesce_str(getattr(asset.asset_category, "detail_profile_title", ""), profile_card_title)
    detail_assignment_title = _coalesce_str(getattr(asset.asset_category, "detail_assignment_title", ""), "Responsabile attuale")
    detail_timeline_title = _coalesce_str(getattr(asset.asset_category, "detail_timeline_title", ""), "Timeline ciclo di vita")
    detail_maintenance_title = _coalesce_str(
        getattr(asset.asset_category, "detail_maintenance_title", ""),
        "Registro manutenzione",
    )
    detail_tickets_title = "Ticket collegati"

    timeline_events: list[dict] = []
    if asset.assignment_to:
        timeline_events.append(
            {
                "title": f"Assegnato a {asset.assignment_to}",
                "tag": "ASSEGNAZIONE",
                "auto_key": AssetTimelineHiddenEvent.KEY_ASSIGNMENT,
                "description": _coalesce_str(asset.assignment_location, "Asset in uso"),
                "date": asset.updated_at,
                "meta": _coalesce_str(asset.assignment_reparto, "Inventario"),
                "color": "green",
            }
        )
    timeline_events.append(
        {
            "title": "Registrazione inventario",
            "tag": "AMMINISTRAZIONE",
            "auto_key": AssetTimelineHiddenEvent.KEY_INVENTORY,
            "description": _coalesce_str(asset.source_key, "Asset aggiunto al sistema."),
            "date": asset.created_at,
            "meta": "Sistema",
            "color": "blue",
        }
    )
    if asset.created_at:
        timeline_events.append(
            {
                "title": "Acquisto / Provisioning",
                "tag": "APPROVVIGIONAMENTO",
                "auto_key": AssetTimelineHiddenEvent.KEY_PROVISIONING,
                "description": _coalesce_str(extra.get("po_ref"), "Asset provisionato"),
                "date": asset.created_at - timedelta(days=3),
                "meta": _coalesce_str(extra.get("owner_dept"), "Approvvigionamenti"),
                "color": "amber",
            }
        )
    if isinstance(work_machine, WorkMachine) and work_machine.year:
        try:
            machine_start = timezone.make_aware(datetime(int(work_machine.year), 1, 1, 8, 0, 0), timezone.get_current_timezone())
        except (TypeError, ValueError):
            machine_start = None
        if machine_start is not None:
            timeline_events.append(
                {
                    "title": "Messa in servizio macchina",
                    "tag": "OFFICINA",
                    "auto_key": AssetTimelineHiddenEvent.KEY_MACHINE_START,
                    "description": _coalesce_str(asset.reparto, "Macchina operativa"),
                    "date": machine_start,
                    "meta": _coalesce_str(asset.manufacturer, "Produzione"),
                    "color": "blue",
                }
            )
    # Voci inserite a mano: eventi reali che non passano dal portale (fermi,
    # traslochi, collaudi del fornitore) e che qui si mescolano agli automatici.
    for entry in asset.timeline_entries.all()[:200]:
        try:
            entry_at = timezone.make_aware(
                datetime(entry.event_date.year, entry.event_date.month, entry.event_date.day, 12, 0),
                timezone.get_current_timezone(),
            )
        except (TypeError, ValueError, OverflowError):
            entry_at = None
        timeline_events.append(
            {
                "title": entry.title,
                "tag": entry.tag or "MANUALE",
                "description": entry.description,
                "date": entry_at,
                "meta": _coalesce_str(entry.meta, "Inserimento manuale"),
                "color": entry.color,
                "manual": True,
                "entry_id": entry.id,
                "entry_date_iso": entry.event_date.isoformat(),
                "entry_tag": entry.tag,
                "entry_meta": entry.meta,
                "entry_description": entry.description,
            }
        )
    hidden_auto_keys = _asset_timeline_hidden_keys(asset)
    if hidden_auto_keys:
        # Gli eventi automatici rimossi restano fuori dalla timeline: la riga di
        # AssetTimelineHiddenEvent e' l'unica traccia, i dati dell'asset no.
        timeline_events = [event for event in timeline_events if event.get("auto_key") not in hidden_auto_keys]
    timeline_events.sort(key=lambda item: item.get("date") or now, reverse=True)

    # Aggiungi ticket MAN inclusi nel registro manutenzione
    from assets.services.maintenance_register import collect_asset_maintenance_register
    maintenance_register = collect_asset_maintenance_register(asset, include_tickets=True)
    # Converti le righe del registro in formato compatibile con il template
    maintenance_rows_from_register = []
    for row in maintenance_register:
        if row["source"] == "TICKET":
            maintenance_rows_from_register.append({
                "id": row["ticket"].id,
                "title": row["title"],
                "description": row["description"],
                "status": row["status"],
                "executed_at": row["executed_at"],
                "created_at": row["created_at"],
                "source": "TICKET",
                "ticket_number": row["ticket_number"],
                "url": row["url"],
                "technician": row["technician"],
                "priority": row["priority"],
                "category": row["category"],
            })
        elif row["source"] == "WORKORDER":
            maintenance_rows_from_register.append({
                "id": row["workorder"].id,
                "title": row["title"],
                "description": row["description"],
                "status": row["status"],
                "executed_at": row["executed_at"],
                "created_at": row["created_at"],
                "source": "WORKORDER",
                "url": row["url"],
                "technician": row["technician"],
                "get_status_display": row["workorder"].get_status_display,
                "get_kind_display": row["workorder"].get_kind_display,
                "supplier": row["workorder"].supplier,
            })
    # Ordina per data decrescente
    maintenance_rows_from_register.sort(key=lambda x: (x["executed_at"] or x["created_at"] or timezone.now()), reverse=True)
    # Limita a 10 righe
    maintenance_rows = maintenance_rows_from_register[:10]

    # Timeline P1.2: completamenti scadenze amministrative per questo asset
    deadline_completion_history = list(
        AssetAdministrativeDeadlineCompletion.objects
        .filter(deadline__asset=asset)
        .select_related("deadline", "completed_by")
        .order_by("-completed_on", "-id")[:20]
    )

    # P3.4: analisi costi manutenzione per asset
    try:
        asset_maintenance_costs = get_asset_maintenance_costs(asset.id, today=today)
    except Exception:
        asset_maintenance_costs = {"has_data": False}

    # AS3: budget manutenzione per categoria asset (anno corrente e precedente)
    asset_budget_rows: list[dict] = []
    if asset.asset_category_id:
        from decimal import Decimal as _Dec
        _zero = _Dec("0")
        _year = today.year
        for _y in (_year, _year - 1):
            try:
                _budget_obj = AssetMaintenanceBudget.objects.filter(
                    asset_category_id=asset.asset_category_id, year=_y
                ).first()
                if _budget_obj is None:
                    continue
                _budget = _budget_obj.budget_eur
                _spent = asset_maintenance_costs.get("cost_year" if _y == _year else "cost_prev_year", _zero) or _zero
                _pct = round(float(_spent / _budget * 100), 1) if _budget else 0.0
                _residual = _budget - _spent
                asset_budget_rows.append({
                    "year": _y,
                    "budget": _budget,
                    "spent": _spent,
                    "residual": _residual,
                    "pct": min(_pct, 100.0),
                    "over_budget": _spent > _budget,
                })
            except Exception:
                pass

    # AS4: timeline storico tecnico (OdL + verifiche + incidenti)
    timeline_tech_rows: list[dict] = []
    try:
        from rilevazione_incidenti.models import RilevazioneIncidente as _RI
        _incidents = list(
            _RI.objects.filter(asset_id=asset.id)
            .order_by("-data_segnalazione")[:50]
        )
        for _inc in _incidents:
            timeline_tech_rows.append({
                "tipo": "incidente",
                "tipo_label": _inc.tipologia_scheda,
                "data": _inc.data_segnalazione,
                "titolo": _inc.tipologia_scheda,
                "dettaglio": _inc.descrizione_avvenimento or _inc.nominativo,
                "id": _inc.id,
            })
    except Exception:
        pass
    try:
        _wo_all = list(
            asset.workorders.order_by("-opened_at")
            .values("id", "title", "kind", "status", "opened_at", "closed_at")[:50]
        )
        _kind_label = {
            WorkOrder.KIND_PREVENTIVE: "Preventiva",
            WorkOrder.KIND_CORRECTIVE: "Correttiva",
            WorkOrder.KIND_SAFETY: "Sicurezza",
            WorkOrder.KIND_CALIBRATION: "Taratura",
            WorkOrder.KIND_OTHER: "Altro",
        }
        for _wo in _wo_all:
            timeline_tech_rows.append({
                "tipo": "workorder",
                "tipo_label": f"OdL — {_kind_label.get(_wo['kind'], _wo['kind'])}",
                "data": _wo["closed_at"] or _wo["opened_at"],
                "titolo": _wo["title"],
                "dettaglio": f"#{_wo['id']} · {WorkOrder.STATUS_CHOICES and dict(WorkOrder.STATUS_CHOICES).get(_wo['status'], _wo['status'])}",
                "id": _wo["id"],
            })
    except Exception:
        pass
    try:
        for _pv in asset.periodic_verifications.all().order_by("name")[:20]:
            _pv_wos = list(
                _pv.workorders.filter(asset_id=asset.id).order_by("-opened_at")
                .values("id", "title", "opened_at", "closed_at", "status")[:5]
            )
            for _wo in _pv_wos:
                timeline_tech_rows.append({
                    "tipo": "verifica",
                    "tipo_label": f"Verifica — {_pv.name}",
                    "data": _wo["closed_at"] or _wo["opened_at"],
                    "titolo": _pv.name,
                    "dettaglio": f"#{_wo['id']} · {_wo['status']}",
                    "id": _wo["id"],
                })
    except Exception:
        pass
    timeline_tech_rows.sort(key=lambda r: r["data"] if r["data"] else timezone.datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    timeline_tech_rows = timeline_tech_rows[:40]

    from tickets.models import TipoTicket
    manageable_ticket_types = {
        ticket_type
        for ticket_type in (TipoTicket.IT, TipoTicket.MAN)
        if _can_manage_ticket_type_for_asset_view(request, ticket_type)
    }
    ticket_rows = [
        {
            "numero_ticket": ticket.numero_ticket,
            "tipo_label": ticket.label_tipo,
            "titolo": ticket.titolo,
            "stato_label": ticket.label_stato,
            "priorita_label": ticket.label_priorita,
            "componente": ticket.componente,
            "created_at": ticket.created_at,
            "closed_at": ticket.closed_at,
            "detail_url": _asset_ticket_detail_url(request, ticket, manageable_ticket_types),
        }
        for ticket in asset.tickets.all()
    ]
    ticket_kpi = _compute_ticket_kpi_for_asset(asset)
    doc_category_labels, documents_by_category = _build_asset_documents_by_category(asset)
    doc_category_specs = _asset_document_folder_specs(asset)
    doc_upload_field_map = {spec["code"]: spec["field"] for spec in doc_category_specs}
    can_manage_doc_folders = _can_manage_asset_document_folders(request)
    can_manage_asset_timeline = _can_manage_asset_timeline_entries(request)
    asset_timeline_manual_enabled = _asset_timeline_manual_enabled(asset)
    periodic_verification_rows = []
    asset_execution_cutoff = _periodic_execution_window_cutoff(PERIODIC_EXECUTION_WINDOW_DEFAULT, today=today)
    for verification in asset.periodic_verifications.all().order_by("name", "id"):
        execution_rows = _periodic_execution_rows_for_verification(
            verification_id=verification.id,
            asset_id=asset.id,
            cutoff_date=asset_execution_cutoff,
            limit=10,
        )
        periodic_verification_rows.append(
            {
                "verification": verification,
                "state": _periodic_verification_state(verification, today=today),
                "execution_rows": execution_rows,
                "execution_count": len(execution_rows),
            }
        )
    active_component_count = sum(1 for component in component_rows if component.is_active)
    active_deadline_rows = [deadline for deadline in administrative_deadlines if deadline.is_active]
    deadline_state_rows = [
        {"deadline": deadline, "state": _asset_administrative_deadline_state(deadline, today=today)}
        for deadline in administrative_deadlines
    ]
    overdue_deadline_count = sum(1 for row in deadline_state_rows if row["state"]["status"] == "overdue")
    warning_deadline_count = sum(1 for row in deadline_state_rows if row["state"]["status"] == "warning")
    next_deadline_row = next(
        (row for row in deadline_state_rows if row["deadline"].is_active),
        None,
    )
    resolved_maintenance_rule_rows = resolve_asset_maintenance_rules(asset)
    asset_schedule_rows = build_day_based_maintenance_schedule_rows(
        asset_queryset=Asset.objects.filter(pk=asset.id).select_related("asset_category"),
        today=today,
    )
    open_workorders_by_rule: dict[int, WorkOrder] = {}
    rule_ids = [row["base_rule"].id for row in asset_schedule_rows]
    for workorder in (
        WorkOrder.objects.filter(
            asset=asset,
            maintenance_rule_id__in=rule_ids,
            status=WorkOrder.STATUS_OPEN,
        )
        .select_related("maintenance_rule")
        .order_by("opened_at", "id")
    ):
        open_workorders_by_rule.setdefault(workorder.maintenance_rule_id, workorder)
    for row in asset_schedule_rows:
        open_workorder = open_workorders_by_rule.get(row["base_rule"].id)
        row["open_workorder"] = open_workorder
        row["workorder_close_url"] = (
            reverse("assets:wo_close", kwargs={"id": open_workorder.id})
            if open_workorder is not None
            else ""
        )
    next_maintenance_row = next(
        (
            row
            for row in asset_schedule_rows
            if row["schedule_status"] in {"overdue", "warning", "upcoming", "missing"}
        ),
        None,
    )
    primary_contract = get_primary_assistance_contract(asset, today=today)
    primary_contract_state = contract_state_payload(primary_contract, today=today) if primary_contract is not None else None
    asset_license_rows = [
        {
            "license": license_row,
            "state": _software_license_state_payload(license_row, today=today),
            "asset_url": reverse("assets:asset_view", kwargs={"id": asset.id}),
            "employee_url": (
                reverse("anagrafica:dipendente_detail", args=[license_row.assigned_anagrafica_id])
                if license_row.assigned_anagrafica_id
                else ""
            ),
        }
        for license_row in asset.software_licenses.all().order_by("category", "vendor", "product_name", "id")
    ]
    asset_create_workorder_url = _workorder_create_page_url(
        asset_id=asset.id,
        rule_id=getattr(next_maintenance_row.get("base_rule"), "id", 0) if next_maintenance_row else 0,
        source="asset_detail",
    )
    maintenance_suggestions = _contextual_maintenance_suggestions(
        asset=asset,
        schedule_row=next_maintenance_row,
        contract=primary_contract,
        source="asset_detail",
    )
    asset_status = _asset_status_payload(asset)
    can_manage_asset_maintenance_rules = _is_assets_admin(request)
    # "Metti fuori uso"/"Riattiva": riusa l'endpoint gated asset_bulk_update (solo admin).
    can_retire_asset = _is_assets_admin(request)
    asset_assign_url = reverse("assets:asset_assign", kwargs={"id": asset.id})

    buttons_by_zone = _build_action_buttons_for_asset(asset, create_workorder_url=asset_create_workorder_url)
    buttons_by_zone = _promote_asset_detail_actions(
        buttons_by_zone,
        assign_url=asset_assign_url,
        qr_url=reverse("assets:asset_qr_label", kwargs={"id": asset.id}),
    )
    quick_action_buttons = [
        button
        for button in buttons_by_zone.get(AssetActionButton.ZONE_QUICK, [])
        if _clean_string(button.get("label")) not in {"Riassegna", "Crea intervento", "Etichetta QR"}
    ]

    assigned_user_admin_url = ""
    if asset.assigned_legacy_user_id:
        try:
            anag = AnagraficaDipendente.objects.filter(
                utente_id=int(asset.assigned_legacy_user_id)
            ).values_list("id", flat=True).first()
            if anag:
                assigned_user_admin_url = reverse(
                    "anagrafica:dipendente_detail",
                    kwargs={"legacy_id": anag},
                )
        except (NoReverseMatch, ValueError, TypeError):
            assigned_user_admin_url = ""
    collection_url = reverse("assets:work_machine_list") if asset.asset_type == Asset.TYPE_WORK_MACHINE else reverse("assets:asset_list")
    collection_label = "Macchine di lavoro" if asset.asset_type == Asset.TYPE_WORK_MACHINE else "Inventario"

    # Planimetria della scheda: mostra quella COERENTE COL REPARTO dell'asset,
    # non un marker qualunque. Gli asset con marker su una planimetria non
    # coerente (storici / creati prima del riallineamento al reparto) vengono
    # auto-riparati qui, spostando il marker sul layout corretto (idempotente).
    resolved_layout, layout_is_specific = _resolve_asset_plant_layout(asset)
    if resolved_layout is not None and layout_is_specific:
        has_marker = PlantLayoutMarker.objects.filter(asset=asset, layout__is_active=True).exists()
        on_resolved = PlantLayoutMarker.objects.filter(asset=asset, layout=resolved_layout).exists()
        if has_marker and not on_resolved:
            _ensure_asset_plant_layout_marker(asset)
    map_marker = None
    if resolved_layout is not None:
        map_marker = (
            PlantLayoutMarker.objects.filter(asset=asset, layout=resolved_layout, layout__is_active=True)
            .select_related("layout")
            .first()
        )
    if map_marker is None:
        map_marker = (
            PlantLayoutMarker.objects.filter(asset=asset, layout__is_active=True)
            .select_related("layout")
            .order_by("layout__category", "layout__name", "id")
            .first()
        )
    if map_marker:
        map_url = reverse("assets:plant_layout_map") + f"?asset={asset.id}&category={quote(map_marker.layout.category)}"
    else:
        map_url = ""

    shell_kwargs = {}
    if asset.asset_type == Asset.TYPE_WORK_MACHINE:
        shell_kwargs = {
            "search_action": reverse("assets:work_machine_list"),
            "new_url": reverse("assets:work_machine_create"),
            "new_label": "+ Nuova macchina",
            "search_placeholder": "Ricerca rapida per macchina, tag, reparto o seriale",
        }

    linked_task_rows: list[dict] = []
    upcoming_task_rows: list[dict] = []
    try:
        from tasks.models import TaskExtraRef, TaskStatus
        refs = (
            TaskExtraRef.objects.filter(asset=asset)
            .select_related("task", "task__category", "task__project", "task__assigned_to")
            .order_by("-task__updated_at", "-task_id")[:50]
        )
        today = timezone.localdate()
        seen_task_ids = set()
        for ref in refs:
            if not ref.task_id or ref.task_id in seen_task_ids:
                continue
            seen_task_ids.add(ref.task_id)
            t = ref.task
            cat_field_label = ""
            if t.category_id:
                cat_field = next((f for f in t.category.fields.all() if f.code == ref.field_code), None)
                if cat_field:
                    cat_field_label = cat_field.label
            base_row = {
                "id": t.id,
                "title": t.title,
                "status": t.get_status_display(),
                "status_code": t.status,
                "due_date": t.due_date,
                "start_date": getattr(t, "next_step_due", None),
                "assignee": (t.assigned_to.get_full_name() or t.assigned_to.username) if t.assigned_to_id else "",
                "category_name": t.category.name if t.category_id else "",
                "field_label": cat_field_label or ref.field_code,
                "project_name": t.project.name if t.project_id else "",
                "url": reverse("tasks:detail", args=[t.id]),
            }
            linked_task_rows.append(base_row)
            # "AdL previste": solo task attive con finestra temporale rilevante (in corso oppure futura)
            if t.status in {TaskStatus.TODO, TaskStatus.IN_PROGRESS}:
                start = base_row["start_date"]
                end = base_row["due_date"]
                effective_end = end or start
                if effective_end and effective_end >= today:
                    upcoming_task_rows.append(base_row)
        upcoming_task_rows.sort(
            key=lambda r: (
                r.get("start_date") or r.get("due_date") or today,
                r.get("due_date") or today,
                r.get("id") or 0,
            )
        )
    except Exception:
        linked_task_rows = []
        upcoming_task_rows = []
    detail_section_cards = _build_asset_detail_section_cards(
        detail_specs_title=detail_specs_title,
        detail_timeline_title=detail_timeline_title,
        detail_maintenance_title=detail_maintenance_title,
        detail_tickets_title=detail_tickets_title,
        profile_card_title=profile_card_title,
        detail_assignment_title=detail_assignment_title,
        quick_action_buttons=quick_action_buttons,
        map_marker=map_marker,
        doc_category_labels=doc_category_labels,
        spec_pairs=spec_pairs,
        profile_rows=profile_rows,
        license_rows=asset_license_rows,
        ticket_rows=ticket_rows,
    )
    asset_primary_kpis = _build_asset_primary_kpis(
        asset=asset,
        asset_status=asset_status,
        assignment_rows=assignment_rows,
        primary_contract=primary_contract,
        primary_contract_state=primary_contract_state,
        next_maintenance_row=next_maintenance_row,
        next_deadline_row=next_deadline_row,
        preferred_detail_metrics=_preferred_asset_detail_metrics(detail_metrics),
        detail_metrics=detail_metrics,
        asset_assign_url=asset_assign_url,
        asset_assistance_contracts_url=_assistance_contracts_page_url(asset_id=asset.id),
        asset_maintenance_schedule_url=_maintenance_schedule_page_url(asset_id=asset.id),
        asset_administrative_deadline_list_url=_asset_administrative_deadline_page_url(asset_id=asset.id),
    )
    asset_status_band = _build_asset_status_band(
        primary_contract=primary_contract,
        primary_contract_state=primary_contract_state,
        next_deadline_row=next_deadline_row,
        asset_assistance_contracts_url=_assistance_contracts_page_url(asset_id=asset.id),
        asset_administrative_deadline_list_url=_asset_administrative_deadline_page_url(asset_id=asset.id),
    )
    # Import in-funzione: assets non deve dipendere staticamente da
    # schede_sicurezza, che resta la fonte unica dei dati chimici.
    from schede_sicurezza import pittogrammi as ghs

    scheda_chimica = asset.prodotto_chimico.scheda_corrente() if asset.prodotto_chimico_id else None
    return render(
        request,
        "assets/pages/asset_detail.html",
        {
            "page_title": f"Dettaglio asset {asset.asset_tag}",
            "asset": asset,
            "prodotto_chimico": asset.prodotto_chimico,
            "scheda_corrente": scheda_chimica,
            # Senza SDS restano i pittogrammi dichiarati sul prodotto.
            "pittogrammi_clp": (
                ghs.dettaglio(asset.prodotto_chimico.pittogrammi_effettivi(scheda_chimica))
                if asset.prodotto_chimico_id else []
            ),
            "asset_completeness": asset.completeness(),
            "asset_status": asset_status,
            # Chi ha toccato QUESTO asset, e quando. Vedi core.audit.storico_oggetto:
            # compaiono le voci agganciate all'asset, non tutte quelle del modulo.
            "storico_audit": storico_oggetto(
                oggetto_tipo=AUDIT_OGGETTO_ASSET, oggetto_id=asset.id, limit=25,
            ),
            "recent_workorders": recent_workorders,
            "custom_fields": custom_fields,
            "unmapped_extra": unmapped_extra,
            "assigned_user_admin_url": assigned_user_admin_url,
            "sync_text": sync_text,
            "detail_metrics": detail_metrics,
            "asset_primary_kpis": asset_primary_kpis,
            "asset_status_band": asset_status_band,
            "detail_specs_title": detail_specs_title,
            "spec_pairs": spec_pairs,
            "profile_rows": profile_rows,
            "profile_card_title": profile_card_title,
            "assignment_rows": assignment_rows,
            "detail_assignment_title": detail_assignment_title,
            "timeline_events": timeline_events,
            "detail_timeline_title": detail_timeline_title,
            "can_manage_asset_timeline": can_manage_asset_timeline,
            "asset_timeline_manual_enabled": asset_timeline_manual_enabled,
            "asset_timeline_color_choices": _ui_choices(AssetTimelineEntry.COLOR_CHOICES),
            "maintenance_rows": maintenance_rows,
            "deadline_completion_history": deadline_completion_history,
            "asset_maintenance_costs": asset_maintenance_costs,
            "asset_budget_rows": asset_budget_rows,
            "timeline_tech_rows": timeline_tech_rows,
            "detail_maintenance_title": detail_maintenance_title,
            "ticket_rows": ticket_rows,
            "ticket_kpi": ticket_kpi,
            "detail_tickets_title": detail_tickets_title,
            "work_machine": work_machine,
            "collection_url": collection_url,
            "collection_label": collection_label,
            "asset_license_rows": asset_license_rows,
            "asset_license_manage_url": _software_licenses_page_url(asset_id=asset.id),
            "doc_category_labels": doc_category_labels,
            "doc_category_specs": doc_category_specs,
            "documents_by_category": dict(documents_by_category),
            "document_upload_field_map": doc_upload_field_map,
            "can_manage_doc_folders": can_manage_doc_folders,
            "asset_report_pdf_url": _asset_report_pdf_url(asset.id),
            "asset_qr_url": reverse("assets:asset_qr_label", kwargs={"id": asset.id}),
            "asset_label_designer_url": (
                reverse("assets:asset_label_designer") + f"?scope=asset&asset_id={asset.id}"
                if _is_assets_admin(request)
                else ""
            ),
            "periodic_verification_rows": periodic_verification_rows,
            "periodic_verification_manage_url": _periodic_verifications_page_url(asset_id=asset.id),
            "asset_component_count": len(component_rows),
            "asset_active_component_count": active_component_count,
            "asset_component_list_url": _asset_component_page_url(asset_id=asset.id),
            "asset_component_create_url": _asset_component_create_page_url(asset_id=asset.id),
            "asset_administrative_deadline_count": len(administrative_deadlines),
            "asset_active_administrative_deadline_count": len(active_deadline_rows),
            "asset_overdue_administrative_deadline_count": overdue_deadline_count,
            "asset_warning_administrative_deadline_count": warning_deadline_count,
            "asset_next_administrative_deadline": next_deadline_row["deadline"] if next_deadline_row else None,
            "asset_next_administrative_deadline_state": next_deadline_row["state"] if next_deadline_row else None,
            "asset_administrative_deadline_list_url": _asset_administrative_deadline_page_url(asset_id=asset.id),
            "asset_administrative_deadline_create_url": _asset_administrative_deadline_create_page_url(asset_id=asset.id),
            "asset_maintenance_rule_count": len(resolved_maintenance_rule_rows),
            "asset_overridden_maintenance_rule_count": sum(
                1 for row in resolved_maintenance_rule_rows if row["status"] == "overridden"
            ),
            "asset_disabled_maintenance_rule_count": sum(
                1 for row in resolved_maintenance_rule_rows if row["status"] == "disabled"
            ),
            "asset_maintenance_rule_list_url": _asset_maintenance_rule_list_page_url(asset_id=asset.id),
            "can_manage_asset_maintenance_rules": can_manage_asset_maintenance_rules,
            "asset_linked_task_rows": linked_task_rows,
            "asset_upcoming_task_rows": upcoming_task_rows,
            "asset_maintenance_schedule_url": _maintenance_schedule_page_url(asset_id=asset.id),
            "asset_assistance_contracts_url": _assistance_contracts_page_url(asset_id=asset.id),
            "asset_primary_contract": primary_contract,
            "asset_primary_contract_state": primary_contract_state,
            "asset_next_maintenance_row": next_maintenance_row,
            "asset_schedule_rows": asset_schedule_rows,
            "asset_create_workorder_url": asset_create_workorder_url,
            "maintenance_suggestions": maintenance_suggestions,
            "can_manage_licenses": _is_assets_admin(request),
            "asset_assign_url": asset_assign_url,
            "header_action_buttons": buttons_by_zone.get(AssetActionButton.ZONE_HEADER, []),
            "quick_action_buttons": quick_action_buttons,
            "detail_section_cards": detail_section_cards,
            "layout_manage_url": (
                reverse("assets:asset_detail_layout_admin") + f"?asset={asset.id}"
                if _can_manage_asset_detail_layout(request)
                else ""
            ),
            "map_marker": map_marker,
            "map_url": map_url,
            "can_retire_asset": can_retire_asset,
            "asset_is_retired": asset.status == Asset.STATUS_RETIRED,
            "asset_bulk_update_url": reverse("assets:asset_bulk_update"),
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25), **shell_kwargs),
        },
    )


@login_required
def asset_label_designer(request: HttpRequest) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo personalizzare l'etichetta QR.")
        return redirect("assets:asset_list")

    field_choices = _asset_label_field_choices()
    scope_raw = request.POST.get("scope") if request.method == "POST" else request.GET.get("scope")
    scope = _clean_string(scope_raw).upper() or AssetLabelTemplate.SCOPE_DEFAULT
    if scope not in {AssetLabelTemplate.SCOPE_DEFAULT, AssetLabelTemplate.SCOPE_ASSET_TYPE, AssetLabelTemplate.SCOPE_ASSET}:
        scope = AssetLabelTemplate.SCOPE_DEFAULT

    raw_scope_asset_id = request.POST.get("scope_asset_id") if request.method == "POST" else request.GET.get("asset_id")
    scope_asset_id = _as_int(raw_scope_asset_id, default=0)
    scope_asset = None
    if scope_asset_id:
        scope_asset = Asset.objects.filter(pk=scope_asset_id).select_related("work_machine").first()
    if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is None:
        messages.error(request, "Seleziona un asset valido per l'override etichetta.")
        return redirect("assets:gestione_admin")

    raw_scope_asset_type = request.POST.get("scope_asset_type") if request.method == "POST" else request.GET.get("asset_type")
    scope_asset_type = _clean_string(raw_scope_asset_type).upper()
    valid_asset_types = {code for code, _label in Asset.TYPE_CHOICES}
    if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is not None:
        scope_asset_type = scope_asset.asset_type
    elif scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and scope_asset_type not in valid_asset_types:
        messages.error(request, "Seleziona una tipologia asset valida per il template.")
        return redirect(f"{reverse('assets:gestione_admin')}?tab=config")
    elif scope == AssetLabelTemplate.SCOPE_DEFAULT:
        scope_asset_type = ""

    template, template_exists = _get_asset_label_template_for_scope(
        scope=scope,
        asset=scope_asset,
        asset_type=scope_asset_type,
    )

    raw_preview_asset_id = request.POST.get("preview_asset_id") if request.method == "POST" else request.GET.get("preview_asset_id")
    preview_asset_id = _as_int(raw_preview_asset_id, default=0)
    preview_asset = None
    if preview_asset_id:
        preview_asset = Asset.objects.filter(pk=preview_asset_id).select_related("work_machine").first()
    if preview_asset is not None and scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and preview_asset.asset_type != scope_asset_type:
        preview_asset = None
    if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is not None:
        preview_asset = scope_asset
    elif preview_asset is None and scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and scope_asset_type:
        preview_asset = (
            Asset.objects.filter(asset_type=scope_asset_type)
            .select_related("work_machine")
            .order_by("name", "asset_tag")
            .first()
        )
    if preview_asset is None:
        preview_asset = (
            Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES)
            .select_related("work_machine")
            .order_by("name", "asset_tag")
            .first()
        )
    if preview_asset is None:
        preview_asset = Asset.objects.select_related("work_machine").order_by("name", "asset_tag").first()

    if request.method == "POST":
        form = AssetLabelTemplateForm(request.POST, request.FILES, instance=template, field_choices=field_choices)
        if form.is_valid():
            template = form.save()
            if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is not None:
                messages.success(request, f"Override etichetta salvato per {scope_asset.asset_tag}.")
            elif scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and scope_asset_type:
                messages.success(request, f"Template etichetta salvato per {_asset_type_label(scope_asset_type)}.")
            else:
                messages.success(request, "Template etichetta generale aggiornato.")
            next_url = reverse("assets:asset_label_designer")
            query_parts = [f"scope={scope}"]
            if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is not None:
                query_parts.append(f"asset_id={scope_asset.id}")
            elif scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and scope_asset_type:
                query_parts.append(f"asset_type={scope_asset_type}")
            if preview_asset is not None:
                query_parts.append(f"preview_asset_id={preview_asset.id}")
            if query_parts:
                next_url = f"{next_url}?{'&'.join(query_parts)}"
            return redirect(next_url)
    else:
        form = AssetLabelTemplateForm(instance=template, field_choices=field_choices)

    preview_context = _build_asset_label_preview_context(request, template=template, asset=preview_asset)
    preview_asset_qr_url = ""
    logo_meta = _asset_label_logo_meta(template)
    if preview_asset is not None:
        preview_asset_qr_url = reverse("assets:asset_qr_label", kwargs={"id": preview_asset.id})

    if scope == AssetLabelTemplate.SCOPE_ASSET and scope_asset is not None:
        scope_title = f"Override asset - {scope_asset.asset_tag}"
        scope_description = "Template personale applicato solo a questo asset."
        back_url = reverse("assets:asset_view", kwargs={"id": scope_asset.id})
    elif scope == AssetLabelTemplate.SCOPE_ASSET_TYPE and scope_asset_type:
        scope_title = f"Template tipologia - {_asset_type_label(scope_asset_type)}"
        scope_description = "Template generale applicato a tutti gli asset di questa tipologia, salvo override del singolo asset."
        back_url = f"{reverse('assets:gestione_admin')}?tab=config"
    else:
        scope_title = "Template generale"
        scope_description = "Fallback usato quando non esiste un template dedicato per tipologia o per asset."
        back_url = f"{reverse('assets:gestione_admin')}?tab=config"

    return render(
        request,
        "assets/pages/asset_label_designer.html",
        {
            "page_title": "Designer etichetta QR",
            "form": form,
            "scope": scope,
            "scope_title": scope_title,
            "scope_description": scope_description,
            "scope_asset": scope_asset,
            "scope_asset_type": scope_asset_type,
            "scope_asset_type_label": _asset_type_label(scope_asset_type),
            "template_exists": template_exists,
            "back_url": back_url,
            "preview_asset": preview_asset,
            "preview_asset_qr_url": preview_asset_qr_url,
            "preview_catalog": preview_context["catalog"],
            "preview_selected_body_fields": preview_context["selected_body_fields"],
            "preview_field_values": preview_context["field_values"],
            "preview_title_primary_key": preview_context["title_primary_key"],
            "preview_title_secondary_key": preview_context["title_secondary_key"],
            "preview_target_url": preview_context["target_url"],
            "preview_target_label": preview_context["target_label"],
            "preview_target_meta": {
                "label": preview_context["target_label"],
                "url": preview_context["target_url"],
                "fallbackPrimary": preview_context["preview_asset_tag"],
                "fallbackSecondary": preview_context["preview_asset_name"],
            },
            "preview_logo_meta": logo_meta,
            "preview_asset_name": preview_context["preview_asset_name"],
            "preview_asset_tag": preview_context["preview_asset_tag"],
            "preview_asset_id": preview_asset.id if preview_asset is not None else "",
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def asset_qr_landing(request: HttpRequest, asset_tag: str) -> HttpResponse:
    """P3.3 — Landing mobile-first raggiungibile via QR code fisico sull'asset.
    Mostra: nome asset, stato, OdL aperti, ultima manutenzione, link segnalazione.
    """
    asset = Asset.objects.filter(asset_tag=asset_tag).select_related("asset_category").first()
    if asset is None:
        return render(
            request,
            "assets/pages/asset_qr_landing.html",
            {
                "page_title": "Asset non trovato",
                "asset": None,
                "asset_tag": asset_tag,
                "base_template": "core/base.html",
                **_assets_shell_context(request, rows=25),
            },
        )
    return _render_asset_qr_landing(request, asset, public=False)


@risposta_pubblica
def asset_qr_public_landing(request: HttpRequest, public_qr_token: str) -> HttpResponse:
    """Landing QR pubblica (token opaco): consultabile senza login da tecnici/ispettori esterni.

    Sola lettura: le azioni (segnalazione, registrazione intervento, scheda interna)
    restano dietro login tramite i decoratori delle rispettive view.
    """
    token = _clean_string(public_qr_token)
    if not token:
        raise Http404("Link non disponibile.")
    asset = (
        Asset.objects.filter(public_qr_token=token, public_qr_enabled=True)
        .select_related("asset_category")
        .first()
    )
    if asset is None:
        raise Http404("Link non disponibile.")
    return _render_asset_qr_landing(request, asset, public=True)


def _render_asset_qr_landing(request: HttpRequest, asset: Asset, *, public: bool) -> HttpResponse:
    asset_tag = asset.asset_tag
    today = timezone.localdate()

    # OdL aperti per questo asset (max 5)
    open_workorders = list(
        WorkOrder.objects.filter(asset=asset, status=WorkOrder.STATUS_OPEN)
        .select_related("maintenance_rule__intervention_template")
        .order_by("opened_at")[:5]
    )

    # Ultimo intervento chiuso
    last_wo = (
        WorkOrder.objects.filter(asset=asset, status=WorkOrder.STATUS_DONE)
        .order_by("-closed_at")
        .values("id", "title", "closed_at", "kind")
        .first()
    )

    # Prossima scadenza amministrativa (campo corretto: due_date; niente flag 'completed').
    next_deadline = (
        AssetAdministrativeDeadline.objects.filter(asset=asset, is_active=True)
        .order_by("due_date")
        .values("id", "title", "due_date")
        .first()
    )

    # Stato manutenzione: giorni dall'ultimo intervento
    days_since_last = None
    if last_wo and last_wo["closed_at"]:
        closed_date = last_wo["closed_at"].date() if hasattr(last_wo["closed_at"], "date") else last_wo["closed_at"]
        days_since_last = (today - closed_date).days

    # Manutenzioni programmate in scadenza per questa macchina (dalle regole) — solo azionabili.
    from .maintenance import build_maintenance_schedule_rows

    _status_order = {"overdue": 0, "warning": 1, "upcoming": 2, "missing": 3}
    qr_schedule_rows: list[dict[str, object]] = []
    for row in build_maintenance_schedule_rows(
        asset_queryset=Asset.objects.filter(pk=asset.id).select_related("asset_category")
    ):
        status = str(row.get("schedule_status") or "")
        if status not in _status_order:
            continue
        base_rule = row["base_rule"]
        qr_schedule_rows.append(
            {
                "label": getattr(row.get("effective_intervention_template"), "label", "") or "Manutenzione",
                "due_date": row.get("due_date"),
                "status": status,
                "schedule_label": row.get("schedule_label"),
                "is_external": base_rule.is_external,
                "supplier": str(base_rule.supplier) if (base_rule.is_external and base_rule.supplier_id) else "",
                "record_url": _workorder_create_page_url(
                    asset_id=asset.id, rule_id=base_rule.id, source="qr_landing"
                ),
            }
        )
    qr_schedule_rows.sort(key=lambda r: (_status_order.get(r["status"], 9), r["due_date"] or today))
    qr_schedule_rows = qr_schedule_rows[:12]
    qr_overdue_count = sum(1 for r in qr_schedule_rows if r["status"] == "overdue")

    # Documenti dell'asset (manuali, specifiche, interventi).
    # I file non sono mai raggiungibili via /media/ (deny IIS): il link passa dalla
    # view a token per il QR pubblico, dalla view autenticata per l'utente loggato.
    asset_documents = []
    qr_token = _clean_string(asset.public_qr_token) if asset.public_qr_enabled else ""
    for doc in asset.documents.all()[:15]:
        if public:
            # Il visitatore anonimo ha in mano solo il token: download a token dei
            # file locali, nessun altro percorso esposto.
            open_url = (
                reverse(
                    "assets:asset_document_qr_download",
                    kwargs={"public_qr_token": qr_token, "document_id": doc.id},
                )
                if (qr_token and doc.file)
                else ""
            )
        else:
            open_url = (
                reverse("assets:asset_document_download", kwargs={"document_id": doc.id})
                if doc.file
                else ""
            )
        asset_documents.append(
            {
                "name": doc.original_name or "Documento",
                "category_label": ASSET_DOCUMENT_CATEGORY_LABELS.get(doc.category, doc.category),
                "document_date": doc.document_date,
                "open_url": open_url,
            }
        )
    # URL azioni
    detail_url = reverse("assets:asset_view", kwargs={"id": asset.id})
    report_url = f"{reverse('assets:asset_quick_report')}?asset={asset.id}"
    wo_list_url = f"{reverse('assets:wo_list')}?asset={asset.id}"
    schedule_url = _maintenance_schedule_page_url(asset_id=asset.id)

    context: dict[str, object] = {
        "page_title": f"{asset.asset_tag} — {asset.name}",
        "asset": asset,
        "asset_tag": asset_tag,
        "open_workorders": open_workorders,
        "last_wo": last_wo,
        "next_deadline": next_deadline,
        "days_since_last": days_since_last,
        "today": today,
        "qr_schedule_rows": qr_schedule_rows,
        "qr_overdue_count": qr_overdue_count,
        "asset_documents": asset_documents,
        "detail_url": detail_url,
        "report_url": report_url,
        "wo_list_url": wo_list_url,
        "schedule_url": schedule_url,
        "qr_public": public,
    }
    if public:
        # Nessuna shell applicativa (sidebar/nav/ACL) per i visitatori non autenticati.
        context["base_template"] = "core/base_public.html"
    else:
        context["base_template"] = "core/base.html"
        context.update(_assets_shell_context(request, rows=25))
    return render(request, "assets/pages/asset_qr_landing.html", context)


@login_required
def asset_qr_label(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(Asset, pk=id)
    target_value = _clean_string(request.GET.get("target")).lower() or "landing"
    target_url, target_label = _asset_qr_target_url(request, asset, target=target_value)
    template = _resolve_asset_label_template(asset)
    width = float(template.page_width_mm or 100) * mm
    height = float(template.page_height_mm or 62) * mm
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{asset.asset_tag or "asset"}-qr-label.pdf"'

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(width, height))
    pdf.setTitle(f"QR {asset.asset_tag}")
    pdf.setAuthor("Portale Applicativo")

    _draw_asset_label_pdf(
        pdf,
        asset=asset,
        template=template,
        target_url=target_url,
        target_label=target_label,
    )

    pdf.showPage()
    pdf.save()
    response.write(buffer.getvalue())
    return response


#: Convenzione storica del numero interno asset in produzione: "Int.NNN".
INTERNAL_NUMBER_PREFIX = "Int"
INTERNAL_NUMBER_SEP = "."
INTERNAL_NUMBER_PAD = 3


@login_required
def asset_internal_number_next(request: HttpRequest) -> JsonResponse:
    """Prossimo numero interno progressivo (opt-in): il form lo richiede solo su
    click del bottone "Assegna progressivo". Non è una chiave univoca.

    Segue la convenzione storica "Int.NNN" (es. Int.263 = max Int.262 + 1),
    assorbendo nella stessa sequenza gli eventuali valori nudi legacy (es. 196/197)."""
    from core.numbering import next_prefixed

    value = next_prefixed(
        Asset.objects.values_list("internal_number", flat=True),
        prefix=INTERNAL_NUMBER_PREFIX,
        sep=INTERNAL_NUMBER_SEP,
        pad=INTERNAL_NUMBER_PAD,
    )
    return JsonResponse({"next": value})


@login_required
def chemical_asset_create(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ChemicalAssetForm(request.POST)
        if form.is_valid():
            asset = form.save()
            messages.success(request, "Asset prodotto chimico creato.")
            return redirect("assets:asset_view", id=asset.id)
    else:
        form = ChemicalAssetForm()
    return render(
        request,
        "assets/pages/chemical_asset_form.html",
        {
            "page_title": "Nuovo prodotto chimico",
            "form": form,
            "is_edit": False,
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def chemical_asset_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(Asset, pk=id)
    if request.method == "POST":
        form = ChemicalAssetForm(request.POST, instance=asset)
        if form.is_valid():
            asset = form.save()
            messages.success(request, "Asset prodotto chimico aggiornato.")
            return redirect("assets:asset_view", id=asset.id)
    else:
        initial = {}
        if asset.prodotto_chimico_id:
            initial = {"prodotto_mode": "existing", "prodotto_chimico": asset.prodotto_chimico_id}
        form = ChemicalAssetForm(instance=asset, initial=initial)
    return render(
        request,
        "assets/pages/chemical_asset_form.html",
        {
            "page_title": f"Modifica {asset.asset_tag}",
            "form": form,
            "asset": asset,
            "is_edit": True,
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def asset_create(request: HttpRequest) -> HttpResponse:
    if _clean_string(request.GET.get("asset_type")) == Asset.TYPE_WORK_MACHINE:
        return redirect("assets:work_machine_create")
    if _clean_string(request.GET.get("asset_type")) == Asset.TYPE_CHEMICAL:
        return redirect("assets:chemical_create")
    custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
    list_suggestions = _build_asset_list_suggestions()
    assignment_kwargs = _assignment_form_kwargs()
    if request.method == "POST":
        form = AssetForm(request.POST, request.FILES, custom_fields=custom_fields, list_suggestions=list_suggestions, **assignment_kwargs)
        if form.is_valid():
            asset = form.save()
            if form.cleaned_data.get("include_in_plant_layout"):
                marker_warning = _ensure_asset_plant_layout_marker(asset)
                if marker_warning:
                    messages.warning(request, marker_warning)

            messages.success(request, "Asset creato correttamente.")
            return redirect("assets:asset_view", id=asset.id)
    else:
        form = AssetForm(custom_fields=custom_fields, list_suggestions=list_suggestions, **assignment_kwargs)
    return render(
        request,
        "assets/pages/asset_form.html",
        {
            "page_title": "Nuovo asset",
            "form": form,
            "is_edit": False,
            "base_field_names": form.base_field_names,
            "category_field_groups": form.category_field_groups,
            "category_dynamic_field_names": form.category_dynamic_field_names,
            "dynamic_field_names": form.dynamic_field_names,
            "verification_field_names": form.verification_field_names,
            "list_suggestions": list_suggestions,
            "assignment_department_choices": assignment_kwargs["assignment_department_choices"],
            "assignment_employee_details": assignment_kwargs["assignment_employee_details"],

            "plant_layout_field_names": ["include_in_plant_layout"],
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def asset_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(Asset, pk=id)
    if asset.asset_type == Asset.TYPE_WORK_MACHINE:
        return redirect("assets:work_machine_edit", id=asset.id)
    if asset.asset_type == Asset.TYPE_CHEMICAL:
        return redirect("assets:chemical_edit", id=asset.id)
    custom_fields = list(AssetCustomField.objects.filter(is_active=True).order_by("sort_order", "id"))
    list_suggestions = _build_asset_list_suggestions()
    assignment_kwargs = _assignment_form_kwargs(asset)
    if request.method == "POST":
        form = AssetForm(
            request.POST,
            request.FILES,
            instance=asset,
            custom_fields=custom_fields,
            list_suggestions=list_suggestions,
            **assignment_kwargs,
        )
        if form.is_valid():
            asset = form.save()
            if form.cleaned_data.get("include_in_plant_layout"):
                marker_warning = _ensure_asset_plant_layout_marker(asset)
                if marker_warning:
                    messages.warning(request, marker_warning)

            messages.success(request, "Asset aggiornato.")
            return redirect("assets:asset_view", id=asset.id)
    else:
        form = AssetForm(instance=asset, custom_fields=custom_fields, list_suggestions=list_suggestions, **assignment_kwargs)
    return render(
        request,
        "assets/pages/asset_form.html",
        {
            "page_title": f"Modifica {asset.asset_tag}",
            "form": form,
            "asset": asset,
            "is_edit": True,
            "base_field_names": form.base_field_names,
            "category_field_groups": form.category_field_groups,
            "category_dynamic_field_names": form.category_dynamic_field_names,
            "dynamic_field_names": form.dynamic_field_names,
            "verification_field_names": form.verification_field_names,
            "list_suggestions": list_suggestions,
            "assignment_department_choices": assignment_kwargs["assignment_department_choices"],
            "assignment_employee_details": assignment_kwargs["assignment_employee_details"],

            "plant_layout_field_names": ["include_in_plant_layout"],
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def asset_component_list(request: HttpRequest, asset_id: int | None = None) -> HttpResponse:
    selected_asset = None
    selected_asset_id = asset_id if asset_id is not None else _as_int(request.GET.get("asset"), default=0)
    if asset_id is not None:
        selected_asset = get_object_or_404(
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type"),
            pk=asset_id,
        )
    elif selected_asset_id:
        selected_asset = (
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type")
            .filter(pk=selected_asset_id)
            .first()
        )

    q = _clean_string(request.GET.get("q"))
    active_filter = _clean_string(request.GET.get("active")).lower() or "active"
    if active_filter not in {"active", "inactive", "all"}:
        active_filter = "active"

    component_qs = (
        AssetComponent.objects.select_related("asset")
        .prefetch_related("administrative_deadlines")
        .order_by("-is_active", "asset__name", "name", "code", "id")
    )
    if selected_asset is not None:
        component_qs = component_qs.filter(asset_id=selected_asset.id)
    if q:
        component_qs = component_qs.filter(
            Q(code__icontains=q)
            | Q(name__icontains=q)
            | Q(component_type__icontains=q)
            | Q(serial_number__icontains=q)
            | Q(manufacturer__icontains=q)
            | Q(model__icontains=q)
            | Q(notes__icontains=q)
            | Q(asset__asset_tag__icontains=q)
            | Q(asset__name__icontains=q)
        )
    if active_filter == "active":
        component_qs = component_qs.filter(is_active=True)
    elif active_filter == "inactive":
        component_qs = component_qs.filter(is_active=False)

    component_rows: list[dict[str, object]] = []
    for component in component_qs:
        deadlines = list(component.administrative_deadlines.all())
        active_deadline_count = sum(1 for deadline in deadlines if deadline.is_active)
        component_rows.append(
            {
                "component": component,
                "deadline_total": len(deadlines),
                "active_deadline_count": active_deadline_count,
                "edit_url": reverse("assets:asset_component_edit", kwargs={"id": component.id}),
                "deadline_list_url": _asset_administrative_deadline_page_url(
                    asset_id=component.asset_id,
                    component_id=component.id,
                ),
            }
        )

    asset_options = list(
        Asset.objects.only("id", "asset_tag", "name").order_by("name", "asset_tag", "id")
    )
    selected_asset_component_count = sum(1 for row in component_rows if row["component"].is_active)
    create_url = _asset_component_create_page_url(asset_id=selected_asset.id if selected_asset else 0)

    return render(
        request,
        "assets/pages/asset_component_list.html",
        {
            "page_title": "Componenti asset",
            "component_rows": component_rows,
            "component_total": len(component_rows),
            "active_component_count": sum(1 for row in component_rows if row["component"].is_active),
            "inactive_component_count": sum(1 for row in component_rows if not row["component"].is_active),
            "selected_asset": selected_asset,
            "selected_asset_component_count": selected_asset_component_count,
            "asset_options": asset_options,
            "active_filter": active_filter,
            "q": q,
            "create_url": create_url,
            "clear_filters_url": _asset_component_page_url(asset_id=selected_asset.id if asset_id is not None and selected_asset else 0),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_component_page_url(asset_id=selected_asset.id if asset_id is not None and selected_asset else 0),
                new_url=create_url,
                new_label="+ Nuovo componente",
                search_placeholder="Ricerca rapida per componente, codice, seriale o asset",
            ),
        },
    )


@login_required
def asset_component_create(request: HttpRequest, asset_id: int | None = None) -> HttpResponse:
    selected_asset = None
    selected_asset_id = asset_id if asset_id is not None else _as_int(request.GET.get("asset"), default=0)
    if asset_id is not None:
        selected_asset = get_object_or_404(
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type"),
            pk=asset_id,
        )
    elif selected_asset_id:
        selected_asset = (
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type")
            .filter(pk=selected_asset_id)
            .first()
        )

    if request.method == "POST":
        form = AssetComponentForm(request.POST, locked_asset=selected_asset)
        if form.is_valid():
            component = form.save()
            log_action(
                request,
                "create_asset_component",
                "assets",
                {"component_id": component.id, "asset_id": component.asset_id, "code": component.code, "name": component.name},
            )
            messages.success(request, "Componente salvato correttamente.")
            return redirect(_asset_component_page_url(asset_id=component.asset_id))
    else:
        form = AssetComponentForm(
            locked_asset=selected_asset,
            initial={"asset": selected_asset.id} if selected_asset else None,
        )

    return render(
        request,
        "assets/pages/asset_component_form.html",
        {
            "page_title": "Nuovo componente",
            "form": form,
            "selected_asset": selected_asset,
            "is_edit": False,
            "back_url": _asset_component_page_url(asset_id=selected_asset.id if selected_asset else 0),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_component_page_url(asset_id=selected_asset.id if selected_asset else 0),
                new_url=_asset_component_create_page_url(asset_id=selected_asset.id if selected_asset else 0),
                new_label="+ Nuovo componente",
                search_placeholder="Ricerca rapida per componente, codice, seriale o asset",
            ),
        },
    )


@login_required
def asset_component_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_component_list")
    component = get_object_or_404(
        AssetComponent.objects.select_related("asset"),
        pk=id,
    )
    selected_asset = component.asset

    if request.method == "POST":
        form = AssetComponentForm(request.POST, instance=component, locked_asset=selected_asset)
        if form.is_valid():
            component = form.save()
            log_action(
                request,
                "update_asset_component",
                "assets",
                {"component_id": component.id, "asset_id": component.asset_id, "code": component.code, "name": component.name},
            )
            messages.success(request, "Componente aggiornato.")
            return redirect(_asset_component_page_url(asset_id=component.asset_id))
    else:
        form = AssetComponentForm(instance=component, locked_asset=selected_asset)

    return render(
        request,
        "assets/pages/asset_component_form.html",
        {
            "page_title": f"Modifica componente {component.name}",
            "form": form,
            "component": component,
            "selected_asset": selected_asset,
            "is_edit": True,
            "back_url": _asset_component_page_url(asset_id=selected_asset.id),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_component_page_url(asset_id=selected_asset.id),
                new_url=_asset_component_create_page_url(asset_id=selected_asset.id),
                new_label="+ Nuovo componente",
                search_placeholder="Ricerca rapida per componente, codice, seriale o asset",
            ),
        },
    )


@login_required
def asset_administrative_deadline_list(request: HttpRequest) -> HttpResponse:
    can_manage_outlook_calendar = _is_assets_admin(request)
    calendar_user_choices: list[tuple[str, str]] = []
    calendar_user_details: dict[str, dict[str, str]] = {}
    if can_manage_outlook_calendar:
        calendar_user_choices, calendar_user_details = _legacy_employee_options()

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action == "create_outlook_calendar_event":
            redirect_url = _deadline_list_redirect_from_request(request)
            if not can_manage_outlook_calendar:
                messages.error(request, "Solo gli admin assets possono creare eventi Outlook su calendari utente.")
                return redirect(redirect_url)

            deadline_id = _as_int(request.POST.get("deadline_id"), default=0)
            target_legacy_user_id = _as_int(request.POST.get("target_legacy_user_id"), default=0)
            deadline = (
                AssetAdministrativeDeadline.objects.select_related("asset", "component")
                .filter(pk=deadline_id)
                .first()
            )
            target_details = calendar_user_details.get(str(target_legacy_user_id))
            if deadline is None:
                messages.error(request, "Scadenza amministrativa non trovata.")
                return redirect(redirect_url)
            if target_details is None:
                messages.error(request, "Seleziona un utente valido per il calendario Outlook.")
                return redirect(redirect_url)

            try:
                entry, created = _create_asset_deadline_calendar_event(
                    request=request,
                    deadline=deadline,
                    target_legacy_user_id=target_legacy_user_id,
                    target_details=target_details,
                )
                target_label = entry.target_display_name or entry.target_email or f"utente {entry.target_legacy_user_id}"
                if created:
                    log_action(
                        request,
                        "create_deadline_calendar_event",
                        "assets",
                        {
                            "deadline_id": deadline.id,
                            "asset_id": deadline.asset_id,
                            "component_id": deadline.component_id,
                            "due_date": str(deadline.due_date),
                            "target_legacy_user_id": entry.target_legacy_user_id,
                            "target_email": entry.target_email,
                            "graph_event_id": entry.graph_event_id,
                        },
                        oggetto_tipo=AUDIT_OGGETTO_ASSET,
                        oggetto_id=deadline.asset_id,
                    )
                    messages.success(
                        request,
                        f"Evento Outlook creato per {target_label} sulla scadenza del {entry.due_date:%d-%m-%Y}.",
                    )
                else:
                    messages.info(
                        request,
                        f"Esiste gia un evento Outlook per {target_label} su questa stessa scadenza.",
                    )
            except Exception as exc:
                messages.error(request, f"Creazione evento Outlook fallita: {exc}")
            return redirect(redirect_url)

        if action == "complete_administrative_deadline":
            redirect_url = _deadline_list_redirect_from_request(request)
            if not _is_assets_admin(request):
                messages.error(request, "Solo admin puo completare scadenze amministrative.")
                return redirect(redirect_url)
            deadline_id = _as_int(request.POST.get("deadline_id"), default=0)
            deadline = (
                AssetAdministrativeDeadline.objects.select_related("asset", "component")
                .filter(pk=deadline_id)
                .first()
            )
            if deadline is None:
                messages.error(request, "Scadenza amministrativa non trovata.")
                return redirect(redirect_url)
            today_local = timezone.localdate()
            executed_on_raw = _clean_string(request.POST.get("execution_date"))
            if executed_on_raw:
                try:
                    completed_on = date.fromisoformat(executed_on_raw)
                except ValueError:
                    messages.error(request, "Data di completamento non valida.")
                    return redirect(redirect_url)
            else:
                completed_on = today_local
            if completed_on > today_local:
                messages.error(request, "La data di completamento non puo essere futura.")
                return redirect(redirect_url)
            next_due_raw = _clean_string(request.POST.get("execution_next_due"))
            next_due_date = None
            if next_due_raw:
                try:
                    next_due_date = date.fromisoformat(next_due_raw)
                except ValueError:
                    messages.error(request, "Prossima scadenza non valida.")
                    return redirect(redirect_url)
                if next_due_date < completed_on:
                    messages.error(request, "La prossima scadenza non puo precedere la data di completamento.")
                    return redirect(redirect_url)
            duration_minutes = max(0, _as_int(request.POST.get("execution_duration_minutes"), default=0))
            try:
                cost_value = _parse_execution_cost_input(request.POST.get("execution_cost_eur"))
            except ValueError:
                messages.error(request, "Costo non valido: usa un numero non negativo (es. 120.50).")
                return redirect(redirect_url)
            notes_text = _clean_string(request.POST.get("execution_notes"))
            uploads, upload_errors = _validate_workorder_attachment_uploads(
                request, field_name="completion_files"
            )
            if upload_errors:
                for error in upload_errors:
                    messages.error(request, error)
                return redirect(redirect_url)
            attachments_total = 0
            try:
                with transaction.atomic():
                    completion = AssetAdministrativeDeadlineCompletion.objects.create(
                        deadline=deadline,
                        completed_on=completed_on,
                        completed_by=request.user if request.user.is_authenticated else None,
                        cost_eur=cost_value,
                        duration_minutes=duration_minutes,
                        notes=notes_text,
                        next_due_date=next_due_date,
                    )
                    for upload in uploads:
                        AssetAdministrativeDeadlineCompletionAttachment.objects.create(
                            completion=completion,
                            file=upload,
                            original_name=Path(getattr(upload, "name", "") or "").name[:255],
                            uploaded_by=request.user if request.user.is_authenticated else None,
                        )
                        attachments_total += 1
                    if next_due_date is not None:
                        deadline.due_date = next_due_date
                        deadline.is_active = True
                        deadline.save(update_fields=["due_date", "is_active", "updated_at"])
                    else:
                        deadline.is_active = False
                        deadline.save(update_fields=["is_active", "updated_at"])
            except ValidationError as exc:
                messages.error(request, f"Completamento non registrabile: {exc}")
                return redirect(redirect_url)
            except Exception as exc:
                messages.error(request, f"Errore registrazione completamento: {exc}")
                return redirect(redirect_url)
            log_action(
                request,
                "complete_administrative_deadline",
                "assets",
                {
                    "deadline_id": deadline.id,
                    "asset_id": deadline.asset_id,
                    "completion_id": completion.id,
                    "completed_on": str(completed_on),
                    "next_due_date": str(next_due_date) if next_due_date else None,
                    "cost_eur": str(cost_value) if cost_value is not None else None,
                    "duration_minutes": duration_minutes,
                    "attachments": attachments_total,
                },
                oggetto_tipo=AUDIT_OGGETTO_ASSET,
                oggetto_id=deadline.asset_id,
            )
            attach_suffix = f" ({attachments_total} allegati)" if attachments_total else ""
            if next_due_date is not None:
                messages.success(
                    request,
                    f"Completamento registrato. Nuova scadenza fissata al {next_due_date:%d-%m-%Y}{attach_suffix}.",
                )
            else:
                messages.success(
                    request,
                    f"Completamento registrato il {completed_on:%d-%m-%Y}. Scadenza chiusa{attach_suffix}.",
                )
            return redirect(redirect_url)

    today = timezone.localdate()
    selected_asset_id = _as_int(request.GET.get("asset"), default=0)
    selected_component_id = _as_int(request.GET.get("component"), default=0)
    selected_asset = None
    selected_component = None

    if selected_component_id:
        selected_component = (
            AssetComponent.objects.select_related("asset")
            .filter(pk=selected_component_id)
            .first()
        )
        if selected_component is not None and not selected_asset_id:
            selected_asset_id = selected_component.asset_id
    if selected_asset_id:
        selected_asset = (
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type")
            .filter(pk=selected_asset_id)
            .first()
        )
    if selected_component is not None and selected_asset is None:
        selected_asset = selected_component.asset

    q = _clean_string(request.GET.get("q"))
    deadline_type = _clean_string(request.GET.get("deadline_type")).upper()
    valid_deadline_types = {code for code, _label in AssetAdministrativeDeadline.TYPE_CHOICES}
    if deadline_type not in valid_deadline_types:
        deadline_type = ""
    status_filter = _clean_string(request.GET.get("status")).lower() or "all"
    if status_filter not in {"all", "overdue", "warning", "upcoming", "inactive"}:
        status_filter = "all"

    _DEADLINE_FAMILY_TYPES: dict[str, list[str]] = {
        "it":               ["PC", "NOTEBOOK", "SERVER", "VM", "FIREWALL", "STAMPANTE", "HW", "FONIA"],
        "produzione":       ["CNC", "WORK_MACHINE", "CARROPONTE"],
        "videosorveglianza": ["CCTV"],
        "altro":            ["OTHER"],
    }
    _DEADLINE_FAMILY_LABELS: dict[str, str] = {
        "it": "Asset IT", "produzione": "Produzione",
        "videosorveglianza": "Videosorveglianza", "altro": "Altro",
    }
    family_filter = _clean_string(request.GET.get("family")).lower()
    if family_filter not in _DEADLINE_FAMILY_TYPES:
        family_filter = ""
    family_label = _DEADLINE_FAMILY_LABELS.get(family_filter, "")

    deadline_qs = (
        AssetAdministrativeDeadline.objects.select_related("asset", "component")
        .order_by("due_date", "asset__name", "title", "id")
    )
    if selected_asset is not None:
        deadline_qs = deadline_qs.filter(asset_id=selected_asset.id)
    if selected_component is not None:
        deadline_qs = deadline_qs.filter(component_id=selected_component.id)
    if deadline_type:
        deadline_qs = deadline_qs.filter(deadline_type=deadline_type)
    if family_filter:
        deadline_qs = deadline_qs.filter(asset__asset_type__in=_DEADLINE_FAMILY_TYPES[family_filter])
    if q:
        deadline_qs = deadline_qs.filter(
            Q(title__icontains=q)
            | Q(reference_code__icontains=q)
            | Q(issuer__icontains=q)
            | Q(notes__icontains=q)
            | Q(asset__asset_tag__icontains=q)
            | Q(asset__name__icontains=q)
            | Q(component__name__icontains=q)
            | Q(component__code__icontains=q)
        )

    deadline_rows_all: list[dict[str, object]] = []
    for deadline in deadline_qs:
        state = _asset_administrative_deadline_state(deadline, today=today)
        deadline_rows_all.append(
            {
                "deadline": deadline,
                "state": state,
                "edit_url": reverse("assets:asset_administrative_deadline_edit", kwargs={"id": deadline.id}),
                "asset_url": reverse("assets:asset_view", kwargs={"id": deadline.asset_id}),
                "component_url": _asset_component_page_url(asset_id=deadline.asset_id),
            }
        )

    calendar_event_map: dict[int, list[AssetCalendarEvent]] = defaultdict(list)
    deadline_ids = [row["deadline"].id for row in deadline_rows_all]
    if deadline_ids:
        for entry in (
            AssetCalendarEvent.objects.select_related("administrative_deadline")
            .filter(
                event_kind=AssetCalendarEvent.KIND_ADMINISTRATIVE_DEADLINE,
                administrative_deadline_id__in=deadline_ids,
            )
            .order_by("target_display_name", "target_email", "created_at", "id")
        ):
            if entry.administrative_deadline_id:
                calendar_event_map[entry.administrative_deadline_id].append(entry)

    if status_filter == "all":
        deadline_rows = deadline_rows_all
    else:
        deadline_rows = [row for row in deadline_rows_all if row["state"]["status"] == status_filter]

    deadline_completion_cutoff = _periodic_execution_window_cutoff(
        PERIODIC_EXECUTION_WINDOW_DEFAULT, today=today
    )
    for row in deadline_rows:
        deadline = row["deadline"]
        row["calendar_event_rows"] = list(calendar_event_map.get(deadline.id, []))
        row["default_calendar_user_id"] = _asset_calendar_default_user_id(deadline.asset, calendar_user_details)
        row["completion_rows"] = _deadline_completion_rows(
            deadline_id=deadline.id,
            cutoff_date=deadline_completion_cutoff,
            limit=10,
        )
        row["completion_count"] = len(row["completion_rows"])
        row["last_completion"] = row["completion_rows"][0] if row["completion_rows"] else None

    component_options = []
    if selected_asset is not None:
        component_options = list(
            AssetComponent.objects.filter(asset_id=selected_asset.id)
            .order_by("-is_active", "name", "code", "id")
        )
    elif selected_component is not None:
        component_options = [selected_component]

    create_url = _asset_administrative_deadline_create_page_url(
        asset_id=selected_asset.id if selected_asset else 0,
        component_id=selected_component.id if selected_component else 0,
    )

    return render(
        request,
        "assets/pages/asset_admin_deadline_list.html",
        {
            "page_title": "Scadenze amministrative",
            "deadline_rows": deadline_rows,
            "deadline_total": len(deadline_rows),
            "deadline_total_all": len(deadline_rows_all),
            "active_deadline_count": sum(1 for row in deadline_rows_all if row["deadline"].is_active),
            "overdue_deadline_count": sum(1 for row in deadline_rows_all if row["state"]["status"] == "overdue"),
            "warning_deadline_count": sum(1 for row in deadline_rows_all if row["state"]["status"] == "warning"),
            "selected_asset": selected_asset,
            "selected_component": selected_component,
            "asset_options": list(Asset.objects.only("id", "asset_tag", "name").order_by("name", "asset_tag", "id")),
            "component_options": component_options,
            "deadline_type_choices": AssetAdministrativeDeadline.TYPE_CHOICES,
            "selected_deadline_type": deadline_type,
            "status_filter": status_filter,
            "q": q,
            "create_url": create_url,
            "clear_filters_url": _asset_administrative_deadline_page_url(),
            "periodic_verification_url": (
                _periodic_verifications_page_url(asset_id=selected_asset.id if selected_asset else 0)
                if selected_asset
                else reverse("assets:periodic_verifications")
            ),
            "can_manage_outlook_calendar": can_manage_outlook_calendar,
            "can_manage_assets": _is_assets_admin(request),
            "outlook_calendar_ready": _outlook_calendar_graph_ready() if can_manage_outlook_calendar else False,
            "calendar_user_choices": calendar_user_choices,
            "today_iso": today.isoformat(),
            "family_filter": family_filter,
            "family_label": family_label,
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:asset_administrative_deadline_list"),
                new_url=create_url,
                new_label="+ Nuova scadenza",
                search_placeholder="Ricerca rapida per scadenza, riferimento, asset o componente",
            ),
        },
    )


@login_required
def asset_administrative_deadline_create(request: HttpRequest) -> HttpResponse:
    selected_asset_id = _as_int(request.GET.get("asset"), default=0)
    selected_component_id = _as_int(request.GET.get("component"), default=0)
    selected_component = None
    selected_asset = None

    if selected_component_id:
        selected_component = (
            AssetComponent.objects.select_related("asset")
            .filter(pk=selected_component_id)
            .first()
        )
        if selected_component is not None and not selected_asset_id:
            selected_asset_id = selected_component.asset_id
    if selected_asset_id:
        selected_asset = (
            Asset.objects.only("id", "asset_tag", "name", "reparto", "asset_type")
            .filter(pk=selected_asset_id)
            .first()
        )
    if selected_component is not None and selected_asset is None:
        selected_asset = selected_component.asset

    if request.method == "POST":
        form = AssetAdministrativeDeadlineForm(
            request.POST,
            locked_asset=selected_asset,
            preselected_component=selected_component,
        )
        if form.is_valid():
            deadline = form.save()
            log_action(
                request,
                "create_asset_deadline",
                "assets",
                {
                    "deadline_id": deadline.id,
                    "asset_id": deadline.asset_id,
                    "component_id": deadline.component_id,
                    "deadline_type": deadline.deadline_type,
                    "due_date": deadline.due_date.isoformat() if deadline.due_date else "",
                },
                oggetto_tipo=AUDIT_OGGETTO_ASSET,
                oggetto_id=deadline.asset_id,
            )
            messages.success(request, "Scadenza salvata correttamente.")
            return redirect(
                _asset_administrative_deadline_page_url(
                    asset_id=deadline.asset_id,
                    component_id=deadline.component_id or 0,
                )
            )
    else:
        initial = {}
        if selected_asset is not None:
            initial["asset"] = selected_asset.id
        if selected_component is not None:
            initial["component"] = selected_component.id
        form = AssetAdministrativeDeadlineForm(
            initial=initial,
            locked_asset=selected_asset,
            preselected_component=selected_component,
        )

    return render(
        request,
        "assets/pages/asset_admin_deadline_form.html",
        {
            "page_title": "Nuova scadenza amministrativa",
            "form": form,
            "selected_asset": selected_asset,
            "selected_component": selected_component,
            "is_edit": False,
            "back_url": _asset_administrative_deadline_page_url(
                asset_id=selected_asset.id if selected_asset else 0,
                component_id=selected_component.id if selected_component else 0,
            ),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:asset_administrative_deadline_list"),
                new_url=_asset_administrative_deadline_create_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    component_id=selected_component.id if selected_component else 0,
                ),
                new_label="+ Nuova scadenza",
                search_placeholder="Ricerca rapida per scadenza, riferimento, asset o componente",
            ),
        },
    )


@login_required
def asset_administrative_deadline_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_administrative_deadline_list")
    deadline = get_object_or_404(
        AssetAdministrativeDeadline.objects.select_related("asset", "component"),
        pk=id,
    )
    selected_asset = deadline.asset
    selected_component = deadline.component

    if request.method == "POST":
        form = AssetAdministrativeDeadlineForm(
            request.POST,
            instance=deadline,
            locked_asset=selected_asset,
        )
        if form.is_valid():
            deadline = form.save()
            log_action(
                request,
                "update_asset_deadline",
                "assets",
                {
                    "deadline_id": deadline.id,
                    "asset_id": deadline.asset_id,
                    "component_id": deadline.component_id,
                    "deadline_type": deadline.deadline_type,
                    "due_date": deadline.due_date.isoformat() if deadline.due_date else "",
                },
                oggetto_tipo=AUDIT_OGGETTO_ASSET,
                oggetto_id=deadline.asset_id,
            )
            messages.success(request, "Scadenza aggiornata.")
            return redirect(
                _asset_administrative_deadline_page_url(
                    asset_id=deadline.asset_id,
                    component_id=deadline.component_id or 0,
                )
            )
    else:
        form = AssetAdministrativeDeadlineForm(instance=deadline, locked_asset=selected_asset)

    return render(
        request,
        "assets/pages/asset_admin_deadline_form.html",
        {
            "page_title": f"Modifica scadenza {deadline.title}",
            "form": form,
            "deadline": deadline,
            "selected_asset": selected_asset,
            "selected_component": selected_component,
            "is_edit": True,
            "back_url": _asset_administrative_deadline_page_url(
                asset_id=selected_asset.id,
                component_id=selected_component.id if selected_component else 0,
            ),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:asset_administrative_deadline_list"),
                new_url=_asset_administrative_deadline_create_page_url(asset_id=selected_asset.id),
                new_label="+ Nuova scadenza",
                search_placeholder="Ricerca rapida per scadenza, riferimento, asset o componente",
            ),
        },
    )


@login_required
def maintenance_template_list(request: HttpRequest) -> HttpResponse:
    """Deprecata: template e regole sono ora un'unica vista in Impostazioni manutenzione.
    Redirige alla tab 'Interventi & Regole' preservando i filtri (categoria/attivo/ricerca)."""
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire i template manutenzione.")
        return redirect("assets:asset_list")
    return redirect(_maintenance_settings_page_url(request, tab="catalogo"), permanent=True)


def _save_template_checklist_formset(formset, template) -> None:
    """Salva il formset degli step checklist e normalizza la numerazione a 10, 20, 30...
    Gli step lasciati senza numero finiscono in coda."""
    from .models import MaintenanceChecklistStep

    formset.instance = template
    saved = formset.save(commit=False)
    for obj in formset.deleted_objects:
        obj.delete()
    for offset, obj in enumerate(saved):
        if not obj.step_number:
            obj.step_number = 9000 + offset
        obj.save()
    for idx, step in enumerate(
        MaintenanceChecklistStep.objects.filter(intervention_template=template).order_by("step_number", "id"),
        start=1,
    ):
        target = idx * 10
        if step.step_number != target:
            step.step_number = target
            step.save(update_fields=["step_number"])


def _template_next_step_rule_url(template) -> str:
    """URL di creazione regola precompilato con il template appena salvato (CTA continuità)."""
    next_url = f"{reverse('assets:maintenance_rule_create')}?template={template.id}"
    if template.asset_category_id:
        next_url += f"&category={template.asset_category_id}"
    return next_url


@login_required
def maintenance_template_create(request: HttpRequest) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire i template manutenzione.")
        return redirect("assets:asset_list")

    selected_category_id = _as_int(request.GET.get("category"), default=0)
    initial = {"asset_category": selected_category_id} if selected_category_id else None

    if request.method == "POST":
        form = MaintenanceInterventionTemplateForm(request.POST)
        checklist_formset = MaintenanceChecklistStepFormSet(request.POST, prefix="checklist")
        if form.is_valid() and checklist_formset.is_valid():
            template = form.save()
            _save_template_checklist_formset(checklist_formset, template)
            log_action(
                request,
                "create_maintenance_template",
                "assets",
                {"template_id": template.id, "code": template.code, "asset_category_id": template.asset_category_id},
            )
            messages.success(
                request,
                "Tipo di attivita creato. Definisci ora il piano che lo applica agli asset.",
            )
            return redirect(_template_next_step_rule_url(template))
    else:
        form = MaintenanceInterventionTemplateForm(initial=initial)
        checklist_formset = MaintenanceChecklistStepFormSet(prefix="checklist")

    return render(
        request,
        "assets/pages/maintenance_template_form.html",
        {
            "page_title": "Nuovo tipo di attivita",
            "form": form,
            "checklist_formset": checklist_formset,
            "is_edit": False,
            "back_url": _maintenance_template_list_page_url(category_id=selected_category_id),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:maintenance_template_list"),
                new_url=reverse("assets:maintenance_template_create"),
                new_label="+ Nuovo tipo",
                search_placeholder="Ricerca rapida per attivita, codice o categoria",
            ),
        },
    )


@login_required
def maintenance_template_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire i template manutenzione.")
        return redirect("assets:asset_list")
    if id is None:
        return redirect("assets:maintenance_template_list")

    template = get_object_or_404(MaintenanceInterventionTemplate.objects.select_related("asset_category"), pk=id)
    if request.method == "POST":
        form = MaintenanceInterventionTemplateForm(request.POST, instance=template)
        checklist_formset = MaintenanceChecklistStepFormSet(request.POST, instance=template, prefix="checklist")
        if form.is_valid() and checklist_formset.is_valid():
            template = form.save()
            _save_template_checklist_formset(checklist_formset, template)
            log_action(
                request,
                "update_maintenance_template",
                "assets",
                {"template_id": template.id, "code": template.code, "asset_category_id": template.asset_category_id},
            )
            messages.success(request, "Tipo di attivita aggiornato.")
            return redirect(_maintenance_template_list_page_url(category_id=template.asset_category_id or 0))
    else:
        form = MaintenanceInterventionTemplateForm(instance=template)
        checklist_formset = MaintenanceChecklistStepFormSet(instance=template, prefix="checklist")

    return render(
        request,
        "assets/pages/maintenance_template_form.html",
        {
            "page_title": f"Modifica attivita {template.label}",
            "form": form,
            "checklist_formset": checklist_formset,
            "template": template,
            "is_edit": True,
            "rule_create_url": _template_next_step_rule_url(template),
            "back_url": _maintenance_template_list_page_url(category_id=template.asset_category_id or 0),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:maintenance_template_list"),
                new_url=reverse("assets:maintenance_template_create"),
                new_label="+ Nuovo tipo",
                search_placeholder="Ricerca rapida per attivita, codice o categoria",
            ),
        },
    )


@login_required
def maintenance_rule_list(request: HttpRequest) -> HttpResponse:
    """Deprecata: template e regole sono ora un'unica vista in Impostazioni manutenzione.
    Redirige alla tab 'Interventi & Regole' preservando i filtri (categoria/attivo/ricerca)."""
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire le regole manutenzione.")
        return redirect("assets:asset_list")
    return redirect(_maintenance_settings_page_url(request, tab="piani"), permanent=True)


@login_required
def maintenance_rule_create(request: HttpRequest) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire le regole manutenzione.")
        return redirect("assets:asset_list")

    selected_category_id = _as_int(request.GET.get("category"), default=0)
    selected_template_id = _as_int(request.GET.get("template"), default=0)
    initial = {}
    if selected_category_id:
        initial["asset_category"] = selected_category_id
    if selected_template_id:
        initial["intervention_template"] = selected_template_id

    if request.method == "POST":
        form = MaintenanceRuleForm(request.POST)
        if form.is_valid():
            rule = form.save()
            log_action(
                request,
                "create_maintenance_rule",
                "assets",
                {
                    "rule_id": rule.id,
                    "asset_category_id": rule.asset_category_id,
                    "template_id": rule.intervention_template_id,
                    "threshold_type": rule.threshold_type,
                    "threshold_value": rule.threshold_value,
                },
            )
            messages.success(request, "Piano di manutenzione creato.")
            return redirect(
                _maintenance_rule_list_page_url(
                    category_id=rule.asset_category_id,
                    template_id=rule.intervention_template_id,
                )
            )
    else:
        form = MaintenanceRuleForm(initial=initial)

    form_state = _maintenance_rule_form_state(form, is_edit=False)

    return render(
        request,
        "assets/pages/maintenance_rule_form.html",
        {
            "page_title": "Nuovo piano manutenzione",
            "form": form,
            "maintenance_rule_form_state": form_state,
            "is_edit": False,
            "back_url": _maintenance_rule_list_page_url(
                category_id=selected_category_id,
                template_id=selected_template_id,
            ),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:maintenance_rule_list"),
                new_url=reverse("assets:maintenance_rule_create"),
                new_label="+ Nuovo piano",
                search_placeholder="Ricerca rapida per piano, attivita o categoria",
            ),
        },
    )


@login_required
def maintenance_rule_impact_preview(request: HttpRequest) -> HttpResponse:
    """Endpoint HTMX-only: anteprima non persistita di quanti asset copre una regola e delle
    prime scadenze, mentre l'utente compila il form. Non salva nulla."""
    if not _is_assets_admin(request):
        return HttpResponseForbidden("Permesso negato.")

    rule_pk = _as_int(request.POST.get("rule_id"), default=0) or None
    scope_type = request.POST.get("scope_type") or MaintenanceRule.SCOPE_CATEGORY
    asset_ids = [_as_int(v) for v in request.POST.getlist("assets") if _as_int(v)]
    first_due_date = parse_date(request.POST.get("first_due_date") or "") if request.POST.get("first_due_date") else None

    impact = preview_maintenance_rule_impact(
        asset_category_id=_as_int(request.POST.get("asset_category"), default=0) or None,
        scope_type=scope_type,
        asset_ids=asset_ids,
        threshold_type=request.POST.get("threshold_type") or MaintenanceRule.THRESHOLD_DAYS,
        threshold_value=_as_int(request.POST.get("threshold_value"), default=0),
        warning_days=_as_int(request.POST.get("warning_days"), default=0),
        first_due_date=first_due_date,
        rule_pk=rule_pk,
    )
    return render(
        request,
        "assets/components/_maintenance_rule_impact.html",
        {"impact": impact},
    )


@login_required
def maintenance_rule_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire le regole manutenzione.")
        return redirect("assets:asset_list")
    if id is None:
        return redirect("assets:maintenance_rule_list")

    rule = get_object_or_404(
        MaintenanceRule.objects.select_related("asset_category", "intervention_template"),
        pk=id,
    )
    if request.method == "POST":
        form = MaintenanceRuleForm(request.POST, instance=rule)
        if form.is_valid():
            rule = form.save()
            log_action(
                request,
                "update_maintenance_rule",
                "assets",
                {
                    "rule_id": rule.id,
                    "asset_category_id": rule.asset_category_id,
                    "template_id": rule.intervention_template_id,
                    "threshold_type": rule.threshold_type,
                    "threshold_value": rule.threshold_value,
                },
            )
            messages.success(request, "Piano di manutenzione aggiornato.")
            return redirect(
                _maintenance_rule_list_page_url(
                    category_id=rule.asset_category_id,
                    template_id=rule.intervention_template_id,
                )
            )
    else:
        form = MaintenanceRuleForm(instance=rule)

    form_state = _maintenance_rule_form_state(form, is_edit=True)

    return render(
        request,
        "assets/pages/maintenance_rule_form.html",
        {
            "page_title": f"Modifica piano {rule.intervention_template.label}",
            "form": form,
            "maintenance_rule_form_state": form_state,
            "rule": rule,
            "is_edit": True,
            "back_url": _maintenance_rule_list_page_url(
                category_id=rule.asset_category_id,
                template_id=rule.intervention_template_id,
            ),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:maintenance_rule_list"),
                new_url=reverse("assets:maintenance_rule_create"),
                new_label="+ Nuovo piano",
                search_placeholder="Ricerca rapida per piano, attivita o categoria",
            ),
        },
    )


def _asset_maintenance_status_payload(status: str) -> dict[str, str]:
    if status == "disabled":
        return {"label": "Disabilitata", "badge_class": "muted"}
    if status == "overridden":
        return {"label": "Personalizzata", "badge_class": "warn"}
    return {"label": "Ereditata", "badge_class": "ok"}


@login_required
def asset_maintenance_rule_list(request: HttpRequest, asset_id: int) -> HttpResponse:
    asset = get_object_or_404(Asset.objects.select_related("asset_category"), pk=asset_id)
    focus_rule_id = _as_int(request.GET.get("focus_rule"), default=0)
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire le regole manutenzione per asset.")
        return redirect("assets:asset_view", id=asset.id)

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        base_rule = get_object_or_404(
            MaintenanceRule.objects.select_related("asset_category", "intervention_template"),
            pk=_as_int(request.POST.get("base_rule_id"), default=0),
            asset_category_id=asset.asset_category_id,
        )
        if action == "update_rule_execution":
            raw_execution_date = _clean_string(request.POST.get("last_execution_date"))
            execution_notes = _clean_string(request.POST.get("last_execution_notes"))
            try:
                execution_date = datetime.strptime(raw_execution_date, "%Y-%m-%d").date()
            except ValueError:
                execution_date = None
            if execution_date is None:
                messages.error(request, "Inserisci una data valida per l'ultima esecuzione.")
            else:
                upsert_asset_maintenance_rule_state(
                    asset=asset,
                    base_rule=base_rule,
                    executed_on=execution_date,
                    notes=execution_notes,
                )
                log_action(
                    request,
                    "update_asset_maintenance_execution",
                    "assets",
                    {
                        "asset_id": asset.id,
                        "base_rule_id": base_rule.id,
                        "last_execution_date": execution_date.isoformat(),
                    },
                )
                messages.success(request, "Storico manutenzione aggiornato.")
            return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id, focus_rule_id=base_rule.id))
        if action == "clear_rule_execution":
            AssetMaintenanceRuleState.objects.filter(asset=asset, base_rule=base_rule).delete()
            log_action(
                request,
                "clear_asset_maintenance_execution",
                "assets",
                {
                    "asset_id": asset.id,
                    "base_rule_id": base_rule.id,
                },
            )
            messages.success(request, "Storico manutenzione azzerato.")
            return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id, focus_rule_id=base_rule.id))

    resolved_rows = resolve_asset_maintenance_rules(asset)
    schedule_rows = build_day_based_maintenance_schedule_rows(
        asset_queryset=Asset.objects.filter(pk=asset.id).select_related("asset_category"),
    )
    schedule_by_rule_id = {row["base_rule"].id: row for row in schedule_rows}
    rule_rows: list[dict[str, object]] = []
    for row in resolved_rows:
        status_payload = _asset_maintenance_status_payload(str(row["status"]))
        override = row["override"]
        schedule_row = schedule_by_rule_id.get(row["base_rule"].id)
        row_payload = {
            **row,
            "status_label": status_payload["label"],
            "status_badge_class": status_payload["badge_class"],
            "last_execution_date": schedule_row["last_execution_date"] if schedule_row else None,
            "last_execution_notes": schedule_row["last_execution_notes"] if schedule_row else "",
            "last_execution_workorder": schedule_row["last_execution_workorder"] if schedule_row else None,
            "due_date": schedule_row["due_date"] if schedule_row else None,
            "schedule_status": schedule_row["schedule_status"] if schedule_row else "",
            "schedule_label": schedule_row["schedule_label"] if schedule_row else "Non operativo",
            "schedule_badge_class": schedule_row["schedule_badge_class"] if schedule_row else "muted",
            "effective_warning_days": row.get("effective_warning_days") or 0,
            "create_url": _asset_maintenance_rule_override_create_page_url(asset_id=asset.id, rule_id=row["base_rule"].id),
            "workorder_create_url": _workorder_create_page_url(
                asset_id=asset.id,
                rule_id=row["base_rule"].id,
                source="maintenance_rules",
            ),
            "focus_url": _asset_maintenance_rule_list_page_url(asset_id=asset.id, focus_rule_id=row["base_rule"].id),
            "edit_url": (
                _asset_maintenance_rule_override_edit_page_url(asset_id=asset.id, override_id=override.id)
                if override is not None
                else ""
            ),
            "reset_url": (
                _asset_maintenance_rule_override_reset_page_url(asset_id=asset.id, override_id=override.id)
                if override is not None
                else ""
            ),
            "is_focus": row["base_rule"].id == focus_rule_id,
        }
        rule_rows.append(row_payload)

    return render(
        request,
        "assets/pages/maintenance_asset_rule_list.html",
        {
            "page_title": f"Regole manutenzione asset {asset.asset_tag}",
            "asset": asset,
            "rule_rows": rule_rows,
            "rule_total": len(rule_rows),
            "overridden_rule_count": sum(1 for row in rule_rows if row["status"] == "overridden"),
            "disabled_rule_count": sum(1 for row in rule_rows if row["status"] == "disabled"),
            "scheduled_rule_count": sum(1 for row in rule_rows if row.get("schedule_status") not in {"", "missing"}),
            "missing_execution_count": sum(1 for row in rule_rows if row.get("schedule_status") == "missing"),
            "focus_rule_id": focus_rule_id,
            "asset_detail_url": reverse("assets:asset_view", kwargs={"id": asset.id}),
            "asset_has_category": bool(asset.asset_category_id),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_maintenance_rule_list_page_url(asset_id=asset.id),
                search_placeholder="Ricerca contestuale disponibile dal dettaglio asset",
            ),
        },
    )


@login_required
def asset_maintenance_rule_override_create(request: HttpRequest, asset_id: int, rule_id: int) -> HttpResponse:
    asset = get_object_or_404(Asset.objects.select_related("asset_category"), pk=asset_id)
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo personalizzare le regole manutenzione degli asset.")
        return redirect("assets:asset_view", id=asset.id)

    base_rule = get_object_or_404(
        MaintenanceRule.objects.select_related("asset_category", "intervention_template"),
        pk=rule_id,
    )
    if not asset.asset_category_id or asset.asset_category_id != base_rule.asset_category_id:
        messages.error(request, "La regola selezionata non appartiene alla categoria dell'asset.")
        return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id))

    existing_override = MaintenanceRuleAssetOverride.objects.filter(asset=asset, base_rule=base_rule).first()
    if existing_override is not None:
        messages.info(request, "Esiste gia una personalizzazione per questa regola. Puoi modificarla direttamente.")
        return redirect(_asset_maintenance_rule_override_edit_page_url(asset_id=asset.id, override_id=existing_override.id))

    if request.method == "POST":
        form = MaintenanceRuleAssetOverrideForm(
            request.POST,
            locked_asset=asset,
            locked_base_rule=base_rule,
        )
        if form.is_valid():
            override = form.save()
            log_action(
                request,
                "create_asset_maintenance_override",
                "assets",
                {
                    "override_id": override.id,
                    "asset_id": override.asset_id,
                    "base_rule_id": override.base_rule_id,
                    "is_disabled": override.is_disabled,
                },
            )
            messages.success(request, "Personalizzazione regola salvata.")
            return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id))
    else:
        form = MaintenanceRuleAssetOverrideForm(
            locked_asset=asset,
            locked_base_rule=base_rule,
        )

    return render(
        request,
        "assets/pages/maintenance_asset_rule_override_form.html",
        {
            "page_title": f"Personalizza regola {base_rule.intervention_template.label}",
            "asset": asset,
            "base_rule": base_rule,
            "form": form,
            "is_edit": False,
            "back_url": _asset_maintenance_rule_list_page_url(asset_id=asset.id),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_maintenance_rule_list_page_url(asset_id=asset.id),
                search_placeholder="Ricerca contestuale disponibile dal dettaglio asset",
            ),
        },
    )


@login_required
def asset_maintenance_rule_override_edit(request: HttpRequest, asset_id: int, id: int) -> HttpResponse:
    asset = get_object_or_404(Asset.objects.select_related("asset_category"), pk=asset_id)
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo personalizzare le regole manutenzione degli asset.")
        return redirect("assets:asset_view", id=asset.id)

    override = get_object_or_404(
        MaintenanceRuleAssetOverride.objects.select_related(
            "asset",
            "base_rule",
            "base_rule__asset_category",
            "base_rule__intervention_template",
            "override_intervention_template",
        ),
        pk=id,
        asset_id=asset.id,
    )

    if request.method == "POST":
        form = MaintenanceRuleAssetOverrideForm(
            request.POST,
            instance=override,
            locked_asset=asset,
            locked_base_rule=override.base_rule,
        )
        if form.is_valid():
            override = form.save()
            log_action(
                request,
                "update_asset_maintenance_override",
                "assets",
                {
                    "override_id": override.id,
                    "asset_id": override.asset_id,
                    "base_rule_id": override.base_rule_id,
                    "is_disabled": override.is_disabled,
                },
            )
            messages.success(request, "Personalizzazione regola aggiornata.")
            return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id))
    else:
        form = MaintenanceRuleAssetOverrideForm(
            instance=override,
            locked_asset=asset,
            locked_base_rule=override.base_rule,
        )

    return render(
        request,
        "assets/pages/maintenance_asset_rule_override_form.html",
        {
            "page_title": f"Modifica personalizzazione {override.base_rule.intervention_template.label}",
            "asset": asset,
            "base_rule": override.base_rule,
            "override": override,
            "form": form,
            "is_edit": True,
            "back_url": _asset_maintenance_rule_list_page_url(asset_id=asset.id),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_asset_maintenance_rule_list_page_url(asset_id=asset.id),
                search_placeholder="Ricerca contestuale disponibile dal dettaglio asset",
            ),
        },
    )


@login_required
def asset_maintenance_rule_override_reset(request: HttpRequest, asset_id: int, id: int) -> HttpResponse:
    asset = get_object_or_404(Asset.objects.only("id"), pk=asset_id)
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo ripristinare le regole manutenzione degli asset.")
        return redirect("assets:asset_view", id=asset.id)

    override = get_object_or_404(
        MaintenanceRuleAssetOverride.objects.select_related("base_rule"),
        pk=id,
        asset_id=asset.id,
    )
    if request.method != "POST":
        messages.error(request, "Usa l'azione di ripristino dalla pagina delle regole asset.")
        return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id))

    log_action(
        request,
        "reset_asset_maintenance_override",
        "assets",
        {
            "override_id": override.id,
            "asset_id": override.asset_id,
            "base_rule_id": override.base_rule_id,
        },
    )
    override.delete()
    messages.success(request, "Override rimosso. La regola e tornata ereditata.")
    return redirect(_asset_maintenance_rule_list_page_url(asset_id=asset.id))


def _maintenance_schedule_status_choices() -> list[tuple[str, str]]:
    return [
        ("attive", "Attive (escludi senza storico)"),
        ("all", "Tutte"),
        ("due", "Solo da gestire"),
        ("overdue", "Scadute"),
        ("warning", "In warning"),
        ("upcoming", "Programmate"),
        ("missing", "Senza storico"),
    ]


def _periodic_verification_schedule_status(
    *, due_date: "date | None", today: date, warning_days: int = 30
) -> dict[str, str]:
    if due_date is None:
        return {"status": "missing", "label": "Da pianificare", "badge_class": "muted"}
    delta = (due_date - today).days
    if delta < 0:
        return {"status": "overdue", "label": f"Scaduta da {abs(delta)} gg", "badge_class": "danger"}
    if delta <= max(0, int(warning_days)):
        if delta == 0:
            return {"status": "warning", "label": "Scade oggi", "badge_class": "warn"}
        return {"status": "warning", "label": f"In scadenza ({delta} gg)", "badge_class": "warn"}
    return {"status": "upcoming", "label": f"Pianificata tra {delta} gg", "badge_class": "ok"}


def _build_maintenance_lookahead_rows(
    *,
    asset_queryset,
    today: date,
    days: int = 90,
) -> list[dict[str, object]]:
    """Costruisce le righe look-ahead raggruppate per settimana ISO su un orizzonte di `days` giorni.
    Ritorna una lista di dict con chiave `week_label`, `week_start`, `week_end`, `rows`."""
    from datetime import timedelta
    horizon = today + timedelta(days=days)
    all_rows = build_day_based_maintenance_schedule_rows(asset_queryset=asset_queryset)
    week_map: dict[tuple, dict] = {}
    for row in all_rows:
        due_date = row.get("due_date")
        if not isinstance(due_date, date):
            continue
        if due_date < today or due_date > horizon:
            continue
        iso = due_date.isocalendar()
        week_key = (iso.year, iso.week)
        if week_key not in week_map:
            week_start = due_date - timedelta(days=due_date.weekday())
            week_end = week_start + timedelta(days=6)
            week_map[week_key] = {
                "week_label": f"Settimana {iso.week}/{iso.year}  ({week_start:%d/%m} - {week_end:%d/%m})",
                "week_start": week_start,
                "week_end": week_end,
                "rows": [],
            }
        row["asset_detail_url"] = reverse("assets:asset_view", kwargs={"id": row["asset"].id})
        week_map[week_key]["rows"].append(row)
    return [week_map[k] for k in sorted(week_map)]


def _maintenance_schedule_periodic_rows(
    *,
    asset_queryset,
    status_filter: str,
    q: str,
    today: date,
) -> list[dict[str, object]]:
    asset_ids = list(asset_queryset.values_list("id", flat=True))
    if not asset_ids:
        return []
    verifications = list(
        PeriodicVerification.objects.select_related("supplier")
        .prefetch_related("assets")
        .filter(is_active=True, assets__id__in=asset_ids)
        .distinct()
    )
    if not verifications:
        return []
    asset_map = {asset.id: asset for asset in asset_queryset}
    rows: list[dict[str, object]] = []
    q_lower = q.casefold() if q else ""
    for verification in verifications:
        for asset in verification.assets.all():
            if asset.id not in asset_map:
                continue
            local_asset = asset_map[asset.id]
            due_date = verification.next_verification_date if isinstance(verification.next_verification_date, date) else None
            schedule = _periodic_verification_schedule_status(due_date=due_date, today=today)
            if status_filter == "due" and schedule["status"] not in {"overdue", "warning", "missing"}:
                continue
            if status_filter == "attive" and schedule["status"] == "missing":
                continue
            if status_filter in {"overdue", "warning", "upcoming", "missing"} and schedule["status"] != status_filter:
                continue
            if q_lower:
                searchable = [
                    local_asset.asset_tag,
                    local_asset.name,
                    local_asset.reparto or "",
                    verification.name or "",
                    str(verification.supplier or ""),
                ]
                if not any(q_lower in (str(chunk) or "").casefold() for chunk in searchable):
                    continue
            rows.append(
                {
                    "asset": local_asset,
                    "asset_detail_url": reverse("assets:asset_view", kwargs={"id": local_asset.id}),
                    "verification": verification,
                    "supplier_label": str(verification.supplier) if verification.supplier_id else "",
                    "frequency_months": verification.frequency_months,
                    "last_execution_date": verification.last_verification_date,
                    "due_date": due_date,
                    "schedule_status": schedule["status"],
                    "schedule_label": schedule["label"],
                    "schedule_badge_class": schedule["badge_class"],
                    "edit_url": _periodic_verifications_page_url(
                        asset_id=local_asset.id,
                        edit_id=verification.id,
                    ),
                    "open_url": _periodic_verifications_page_url(asset_id=local_asset.id),
                    "workorder_create_url": _workorder_create_page_url(
                        asset_id=local_asset.id,
                        periodic_verification_id=verification.id,
                        source="maintenance_schedule",
                    ),
                }
            )
    rows.sort(
        key=lambda row: (
            {"overdue": 0, "warning": 1, "upcoming": 2, "missing": 3}.get(str(row["schedule_status"]), 9),
            row["due_date"] or date.max,
            (row["asset"].reparto or "").casefold(),
            (row["asset"].name or "").casefold(),
            row["asset"].id,
            row["verification"].id,
        )
    )
    return rows


@login_required
def maintenance_schedule(request: HttpRequest) -> HttpResponse:
    can_manage_outlook_calendar = _is_assets_admin(request)
    calendar_user_choices: list[tuple[str, str]] = []
    calendar_user_details: dict[str, dict[str, str]] = {}
    if can_manage_outlook_calendar:
        calendar_user_choices, calendar_user_details = _legacy_employee_options()

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action == "create_outlook_calendar_event":
            redirect_url = _maintenance_schedule_redirect_from_request(request)
            if not can_manage_outlook_calendar:
                messages.error(request, "Solo gli admin assets possono creare eventi Outlook su calendari utente.")
                return redirect(redirect_url)

            asset_id = _as_int(request.POST.get("asset_id"), default=0)
            base_rule_id = _as_int(request.POST.get("base_rule_id"), default=0)
            target_legacy_user_id = _as_int(request.POST.get("target_legacy_user_id"), default=0)
            asset = Asset.objects.select_related("asset_category").filter(pk=asset_id).first()
            target_details = calendar_user_details.get(str(target_legacy_user_id))
            if asset is None:
                messages.error(request, "Asset non trovato.")
                return redirect(redirect_url)
            if not base_rule_id:
                messages.error(request, "Regola manutenzione non valida.")
                return redirect(redirect_url)
            if target_details is None:
                messages.error(request, "Seleziona un utente valido per il calendario Outlook.")
                return redirect(redirect_url)

            schedule_row = _maintenance_schedule_row_for_asset_rule(asset=asset, base_rule_id=base_rule_id)
            if schedule_row is None:
                messages.error(request, "Scadenza manutenzione non trovata per l'asset selezionato.")
                return redirect(redirect_url)
            if not isinstance(schedule_row.get("due_date"), date):
                messages.error(request, "Questa manutenzione non ha ancora una data scadenza calendarizzabile.")
                return redirect(redirect_url)

            try:
                entry, created = _create_asset_maintenance_calendar_event(
                    request=request,
                    asset=asset,
                    schedule_row=schedule_row,
                    target_legacy_user_id=target_legacy_user_id,
                    target_details=target_details,
                )
                target_label = entry.target_display_name or entry.target_email or f"utente {entry.target_legacy_user_id}"
                if created:
                    log_action(
                        request,
                        "create_maintenance_calendar_event",
                        "assets",
                        {
                            "asset_id": asset.id,
                            "base_rule_id": schedule_row["base_rule"].id,
                            "due_date": str(schedule_row["due_date"]),
                            "target_legacy_user_id": entry.target_legacy_user_id,
                            "target_email": entry.target_email,
                            "graph_event_id": entry.graph_event_id,
                        },
                    )
                    messages.success(
                        request,
                        f"Evento Outlook creato per {target_label} sulla scadenza del {entry.due_date:%d-%m-%Y}.",
                    )
                else:
                    messages.info(
                        request,
                        f"Esiste gia un evento Outlook per {target_label} su questa stessa scadenza.",
                    )
            except Exception as exc:
                messages.error(request, f"Creazione evento Outlook fallita: {exc}")
            return redirect(redirect_url)

        if action == "record_maintenance_rule_execution":
            redirect_url = _maintenance_schedule_redirect_from_request(request)
            if not _is_assets_admin(request):
                messages.error(request, "Solo admin puo registrare esecuzioni di manutenzione.")
                return redirect(redirect_url)
            asset_id = _as_int(request.POST.get("asset_id"), default=0)
            base_rule_id = _as_int(request.POST.get("base_rule_id"), default=0)
            asset = Asset.objects.select_related("asset_category").filter(pk=asset_id).first()
            base_rule = MaintenanceRule.objects.filter(pk=base_rule_id).select_related("intervention_template").first()
            if asset is None or base_rule is None:
                messages.error(request, "Asset o regola manutenzione non trovati.")
                return redirect(redirect_url)
            today_local = timezone.localdate()
            executed_on_raw = _clean_string(request.POST.get("execution_date"))
            if executed_on_raw:
                try:
                    executed_on = date.fromisoformat(executed_on_raw)
                except ValueError:
                    messages.error(request, "Data di esecuzione non valida.")
                    return redirect(redirect_url)
            else:
                executed_on = today_local
            if executed_on > today_local:
                messages.error(request, "La data di esecuzione non puo essere futura.")
                return redirect(redirect_url)
            duration_minutes = max(0, _as_int(request.POST.get("execution_duration_minutes"), default=0))
            try:
                cost_value = _parse_execution_cost_input(request.POST.get("execution_cost_eur"))
            except ValueError:
                messages.error(request, "Costo non valido: usa un numero non negativo (es. 120.50).")
                return redirect(redirect_url)
            resolution_text = _clean_string(request.POST.get("execution_notes"))
            uploads, upload_errors = _validate_execution_attachment_uploads(request)
            if upload_errors:
                for error in upload_errors:
                    messages.error(request, error)
                return redirect(redirect_url)
            template_label = (getattr(base_rule.intervention_template, "label", "") or "Manutenzione").strip()
            attachments_total = 0
            try:
                with transaction.atomic():
                    workorder = _build_execution_workorder(
                        asset=asset,
                        title=f"Esecuzione regola: {template_label}",
                        description=resolution_text,
                        executed_on=executed_on,
                        duration_minutes=duration_minutes,
                        cost_value=cost_value,
                        resolution_text=resolution_text,
                        maintenance_rule=base_rule,
                        # Manutenzione esterna: l'OdL eredita la ditta terza della regola,
                        # così storico e costi restano attribuiti al fornitore.
                        supplier=base_rule.supplier if base_rule.is_external else None,
                    )
                    sync_workorder_maintenance_state(workorder)
                    # Checklist operativa (Fase 2): copia gli step del template e spunta
                    # quelli confermati nel form di registrazione rapida.
                    copy_template_checklist_to_workorder(workorder)
                    done_steps = [
                        int(s) for s in request.POST.getlist("checklist_done") if str(s).isdigit()
                    ]
                    if done_steps:
                        WorkOrderChecklist.objects.filter(
                            work_order=workorder, step_number__in=done_steps
                        ).update(
                            is_done=True,
                            done_at=timezone.now(),
                            done_by=request.user if request.user.is_authenticated else None,
                        )
                    if uploads:
                        attachments_total = len(
                            _save_workorder_attachments(
                                workorder=workorder,
                                uploads=uploads,
                                user=request.user,
                            )
                        )
            except ValidationError as exc:
                messages.error(request, f"Esecuzione non registrabile: {exc}")
                return redirect(redirect_url)
            except Exception as exc:
                messages.error(request, f"Errore registrazione esecuzione: {exc}")
                return redirect(redirect_url)
            log_action(
                request,
                "record_maintenance_rule_execution",
                "assets",
                {
                    "asset_id": asset.id,
                    "base_rule_id": base_rule.id,
                    "workorder_id": workorder.id,
                    "executed_on": str(executed_on),
                    "cost_eur": str(cost_value) if cost_value is not None else None,
                    "duration_minutes": duration_minutes,
                    "attachments": attachments_total,
                },
            )
            attach_suffix = f" ({attachments_total} allegati)" if attachments_total else ""
            messages.success(
                request,
                f"Esecuzione registrata su {asset.asset_tag} il {executed_on:%d-%m-%Y}{attach_suffix}.",
            )
            return redirect(redirect_url)

    selected_asset_id = _as_int(request.GET.get("asset"), default=0)
    selected_asset = None
    if selected_asset_id:
        selected_asset = Asset.objects.select_related("asset_category").filter(pk=selected_asset_id).first()
    status_filter = _clean_string(request.GET.get("status")) or "attive"
    category_id = _as_int(request.GET.get("category"), default=0)
    reparto_filter = _clean_string(request.GET.get("reparto"))
    coverage_filter = _clean_string(request.GET.get("coverage")) or "all"
    execution_filter = _clean_string(request.GET.get("execution")) or "all"
    if execution_filter not in {"all", "internal", "external"}:
        execution_filter = "all"
    q = _clean_string(request.GET.get("q"))

    asset_qs = Asset.objects.select_related("asset_category").exclude(status=Asset.STATUS_RETIRED).order_by(
        "reparto",
        "name",
        "asset_tag",
    )
    if selected_asset is not None:
        asset_qs = asset_qs.filter(pk=selected_asset.id)
    if category_id:
        asset_qs = asset_qs.filter(asset_category_id=category_id)
    if reparto_filter:
        asset_qs = asset_qs.filter(reparto__iexact=reparto_filter)
    if q:
        asset_qs = asset_qs.filter(
            Q(asset_tag__icontains=q)
            | Q(internal_number__icontains=q)
            | Q(name__icontains=q)
            | Q(reparto__icontains=q)
            | Q(asset_category__label__icontains=q)
        )

    schedule_rows = build_day_based_maintenance_schedule_rows(asset_queryset=asset_qs)
    primary_contract_by_asset_id: dict[int, AssistanceContract | None] = {}
    filtered_rows: list[dict[str, object]] = []
    # Totali per i KPI (calcolati sul sottoinsieme che passa i filtri non-stato, cosi
    # restano corretti anche quando la tabella e filtrata/snellita).
    status_totals = {"overdue": 0, "warning": 0, "upcoming": 0, "missing": 0}
    # Vista "Attive" (default): mostra solo cio che e azionabile a breve. Nasconde dalla
    # tabella le "senza storico" e le pianificate oltre l'orizzonte, restando contate.
    ATTIVE_HORIZON_DAYS = 90
    attive_hidden_total = 0
    for row in schedule_rows:
        asset = row["asset"]
        if asset.id not in primary_contract_by_asset_id:
            primary_contract_by_asset_id[asset.id] = get_primary_assistance_contract(asset)
        primary_contract = primary_contract_by_asset_id[asset.id]
        row["primary_contract"] = primary_contract
        row["contract_state"] = contract_state_payload(primary_contract) if primary_contract else None
        row["is_covered"] = primary_contract is not None
        row["asset_detail_url"] = reverse("assets:asset_view", kwargs={"id": asset.id})
        row["workorder_create_url"] = _workorder_create_page_url(
            asset_id=asset.id,
            rule_id=row["base_rule"].id,
            source="maintenance_schedule",
        )
        row["first_execution_url"] = _asset_maintenance_rule_list_page_url(
            asset_id=asset.id,
            focus_rule_id=row["base_rule"].id,
        )
        primary_action = _maintenance_row_primary_action(
            asset=asset,
            base_rule=row["base_rule"],
            schedule_status=str(row.get("schedule_status") or ""),
            source="maintenance_schedule",
        )
        row["primary_action_label"] = primary_action["label"]
        row["primary_action_url"] = primary_action["url"]
        row["contracts_url"] = _assistance_contracts_page_url(asset_id=asset.id)
        row["suggestions"] = _contextual_maintenance_suggestions(
            asset=asset,
            schedule_row=row,
            contract=primary_contract,
            source="maintenance_schedule",
        )

        # Filtri non-stato prima (copertura/ricerca), cosi il conteggio "senza storico"
        # riflette il sottoinsieme realmente pertinente.
        if coverage_filter == "covered" and not row["is_covered"]:
            continue
        if coverage_filter == "uncovered" and row["is_covered"]:
            continue
        if execution_filter == "internal" and row["base_rule"].is_external:
            continue
        if execution_filter == "external" and not row["base_rule"].is_external:
            continue
        if q:
            q_value = q.casefold()
            searchable_chunks = [
                asset.asset_tag,
                asset.name,
                asset.reparto,
                getattr(getattr(asset, "asset_category", None), "label", ""),
                row["effective_intervention_template"].label,
            ]
            if not any(q_value in _clean_string(chunk).casefold() for chunk in searchable_chunks):
                continue
        status_value = str(row["schedule_status"] or "")
        if status_value in status_totals:
            status_totals[status_value] += 1

        # Filtro stato per la tabella.
        if status_filter == "due" and status_value not in {"overdue", "warning", "missing"}:
            continue
        if status_filter in {"overdue", "warning", "upcoming", "missing"} and status_value != status_filter:
            continue
        if status_filter == "attive":
            far_upcoming = (
                status_value == "upcoming"
                and isinstance(row.get("days_until_due"), int)
                and row["days_until_due"] > ATTIVE_HORIZON_DAYS
            )
            if status_value == "missing" or far_upcoming:
                attive_hidden_total += 1
                continue
        filtered_rows.append(row)

    # Ordine di default pensato per il manutentore: prima le SCADUTE (più in ritardo
    # in cima), poi in warning, poi pianificate (più vicine prima), infine "mai eseguite".
    # È solo l'ordine iniziale: l'utente può comunque riordinare dalle intestazioni.
    _SCHEDULE_STATUS_ORDER = {"overdue": 0, "warning": 1, "upcoming": 2, "missing": 3}
    filtered_rows.sort(
        key=lambda r: (
            _SCHEDULE_STATUS_ORDER.get(str(r.get("schedule_status") or ""), 9),
            r["days_until_due"] if isinstance(r.get("days_until_due"), int) else 10 ** 9,
        )
    )

    calendar_event_map: dict[tuple[int, int, date], list[AssetCalendarEvent]] = defaultdict(list)
    if filtered_rows:
        asset_ids = set()
        base_rule_ids = set()
        due_dates = set()
        for row in filtered_rows:
            due_date = row.get("due_date")
            if not isinstance(due_date, date):
                continue
            asset_ids.add(row["asset"].id)
            base_rule_ids.add(row["base_rule"].id)
            due_dates.add(due_date)
        if asset_ids and base_rule_ids and due_dates:
            for entry in (
                AssetCalendarEvent.objects.filter(
                    event_kind=AssetCalendarEvent.KIND_MAINTENANCE,
                    asset_id__in=asset_ids,
                    maintenance_rule_id__in=base_rule_ids,
                    due_date__in=due_dates,
                )
                .order_by("target_display_name", "target_email", "created_at", "id")
            ):
                if entry.maintenance_rule_id:
                    calendar_event_map[(entry.asset_id, entry.maintenance_rule_id, entry.due_date)].append(entry)

    schedule_today = timezone.localdate()
    schedule_execution_cutoff = _periodic_execution_window_cutoff(
        PERIODIC_EXECUTION_WINDOW_DEFAULT, today=schedule_today
    )
    # Storico esecuzioni in batch: una sola query per tutte le righe visibili (evita N+1).
    exec_pairs = {(row["asset"].id, row["base_rule"].id) for row in filtered_rows}
    exec_by_pair: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    if exec_pairs:
        exec_wo_qs = WorkOrder.objects.select_related("asset", "supplier").filter(
            asset_id__in={pair[0] for pair in exec_pairs},
            maintenance_rule_id__in={pair[1] for pair in exec_pairs},
            status=WorkOrder.STATUS_DONE,
        )
        if schedule_execution_cutoff is not None:
            exec_wo_qs = exec_wo_qs.filter(closed_at__date__gte=schedule_execution_cutoff)
        for workorder in exec_wo_qs.order_by("-closed_at", "-id"):
            key = (workorder.asset_id, workorder.maintenance_rule_id)
            if key in exec_pairs and len(exec_by_pair[key]) < 5:
                exec_by_pair[key].append(_serialize_execution_workorder(workorder))

    for row in filtered_rows:
        due_date = row.get("due_date")
        if isinstance(due_date, date):
            row["calendar_event_rows"] = list(
                calendar_event_map.get((row["asset"].id, row["base_rule"].id, due_date), [])
            )
        else:
            row["calendar_event_rows"] = []
        row["default_calendar_user_id"] = _asset_calendar_default_user_id(row["asset"], calendar_user_details)
        row["execution_rows"] = exec_by_pair.get((row["asset"].id, row["base_rule"].id), [])
        row["execution_count"] = len(row["execution_rows"])

    # Checklist operativa del template per riga (Fase 2), batch anti-N+1.
    from .models import MaintenanceChecklistStep

    tmpl_ids = {
        row["effective_intervention_template"].id
        for row in filtered_rows
        if row.get("effective_intervention_template")
    }
    steps_by_tmpl: dict[int, list] = defaultdict(list)
    if tmpl_ids:
        for step in MaintenanceChecklistStep.objects.filter(
            intervention_template_id__in=tmpl_ids
        ).order_by("step_number", "id"):
            steps_by_tmpl[step.intervention_template_id].append(step)
    for row in filtered_rows:
        tmpl = row.get("effective_intervention_template")
        row["checklist_steps"] = steps_by_tmpl.get(tmpl.id, []) if tmpl else []

    reparto_options = [
        value
        for value in Asset.objects.exclude(reparto="")
        .order_by("reparto")
        .values_list("reparto", flat=True)
        .distinct()
    ]

    # --- Look-ahead 90 giorni (raggruppato per settimana) ---
    lookahead_rows = _build_maintenance_lookahead_rows(
        asset_queryset=asset_qs,
        today=schedule_today,
        days=90,
    )

    # --- Scadenze amministrative nella schedule (filtrate coerentemente) ---
    admin_deadline_qs = (
        AssetAdministrativeDeadline.objects.filter(is_active=True)
        .select_related("asset", "asset__asset_category")
        .order_by("due_date", "asset__name", "title")
    )
    if selected_asset is not None:
        admin_deadline_qs = admin_deadline_qs.filter(asset=selected_asset)
    if category_id:
        admin_deadline_qs = admin_deadline_qs.filter(asset__asset_category_id=category_id)
    if reparto_filter:
        admin_deadline_qs = admin_deadline_qs.filter(asset__reparto__iexact=reparto_filter)
    if q:
        admin_deadline_qs = admin_deadline_qs.filter(
            Q(title__icontains=q) | Q(asset__asset_tag__icontains=q) | Q(asset__name__icontains=q)
        )
    admin_deadline_rows = []
    for dl in admin_deadline_qs[:200]:
        days_left = dl.days_until_due(reference_date=schedule_today)
        if days_left is None:
            dl_status = "missing"
        elif days_left < 0:
            dl_status = "overdue"
        elif days_left <= dl.warning_days:
            dl_status = "warning"
        else:
            dl_status = "upcoming"
        if status_filter == "due" and dl_status not in {"overdue", "warning", "missing"}:
            continue
        if status_filter == "attive" and dl_status == "missing":
            continue
        if status_filter in {"overdue", "warning", "upcoming", "missing"} and dl_status != status_filter:
            continue
        admin_deadline_rows.append({
            "deadline": dl,
            "days_left": days_left,
            "schedule_status": dl_status,
            "asset_detail_url": reverse("assets:asset_view", kwargs={"id": dl.asset_id}),
        })

    periodic_schedule_rows = _maintenance_schedule_periodic_rows(
        asset_queryset=asset_qs,
        status_filter=status_filter,
        q=q,
        today=schedule_today,
    )
    periodic_overdue = sum(1 for row in periodic_schedule_rows if row["schedule_status"] == "overdue")
    periodic_warning = sum(1 for row in periodic_schedule_rows if row["schedule_status"] == "warning")
    periodic_upcoming = sum(1 for row in periodic_schedule_rows if row["schedule_status"] == "upcoming")
    periodic_missing = sum(1 for row in periodic_schedule_rows if row["schedule_status"] == "missing")

    # ── Selettore viste (Lista / Board / Per macchina) ────────────────────────
    vista = _clean_string(request.GET.get("vista")) or "lista"
    if vista not in ("lista", "board", "macchina"):
        vista = "lista"

    def _vista_url(target: str) -> str:
        params = request.GET.copy()
        params["vista"] = target
        return f"{request.path}?{params.urlencode()}"

    vista_urls = {v: _vista_url(v) for v in ("lista", "board", "macchina")}
    # Board: colonne per stato (le righe portano già tutti i dati per la card).
    board_columns = [
        {"key": "overdue", "label": "Scadute", "rows": [r for r in filtered_rows if r.get("schedule_status") == "overdue"]},
        {"key": "warning", "label": "In scadenza", "rows": [r for r in filtered_rows if r.get("schedule_status") == "warning"]},
        {"key": "upcoming", "label": "Pianificate", "rows": [r for r in filtered_rows if r.get("schedule_status") == "upcoming"]},
    ]
    # Per macchina: raggruppa le righe per asset, con lo stato peggiore in testa.
    _status_rank = {"overdue": 0, "warning": 1, "missing": 2, "upcoming": 3}
    machine_map: dict[int, dict[str, object]] = {}
    for r in filtered_rows:
        asset_obj = r["asset"]
        grp = machine_map.setdefault(
            asset_obj.id,
            {"asset": asset_obj, "rows": [], "overdue": 0, "warning": 0, "asset_detail_url": r.get("asset_detail_url")},
        )
        grp["rows"].append(r)
        if r.get("schedule_status") == "overdue":
            grp["overdue"] = int(grp["overdue"]) + 1
        elif r.get("schedule_status") == "warning":
            grp["warning"] = int(grp["warning"]) + 1
    machine_groups = sorted(
        machine_map.values(),
        key=lambda g: (
            -int(g["overdue"]), -int(g["warning"]),
            str(getattr(g["asset"], "reparto", "") or ""), str(getattr(g["asset"], "name", "") or ""),
        ),
    )

    return render(
        request,
        "assets/pages/maintenance_schedule.html",
        {
            "page_title": "Prossime manutenzioni",
            "selected_asset": selected_asset,
            "vista": vista,
            "vista_urls": vista_urls,
            "board_columns": board_columns,
            "machine_groups": machine_groups,
            "schedule_rows": filtered_rows,
            "schedule_total": len(filtered_rows),
            "periodic_schedule_rows": periodic_schedule_rows,
            "periodic_schedule_total": len(periodic_schedule_rows),
            "overdue_count": status_totals["overdue"] + periodic_overdue,
            "warning_count": status_totals["warning"] + periodic_warning,
            "upcoming_count": status_totals["upcoming"] + periodic_upcoming,
            "missing_count": status_totals["missing"] + periodic_missing,
            "attive_hidden_total": attive_hidden_total,
            "show_all_status_url": _query_url(request, status="all"),
            "covered_count": sum(1 for row in filtered_rows if row["is_covered"]),
            "uncovered_count": sum(1 for row in filtered_rows if not row["is_covered"]),
            "category_options": AssetCategory.objects.filter(is_active=True).order_by("sort_order", "label", "id"),
            "selected_category_id": category_id,
            "reparto_options": reparto_options,
            "reparto_filter": reparto_filter,
            "status_filter": status_filter,
            "status_choices": _maintenance_schedule_status_choices(),
            "coverage_filter": coverage_filter,
            "execution_filter": execution_filter,
            "execution_choices": [
                ("all", "Tutte"),
                ("internal", "Interne"),
                ("external", "Esterne (ditta terza)"),
            ],
            "q": q,
            "clear_filters_url": _maintenance_schedule_page_url(asset_id=selected_asset.id if selected_asset else 0),
            "can_manage_outlook_calendar": can_manage_outlook_calendar,
            "can_manage_assets": _is_assets_admin(request),
            "outlook_calendar_ready": _outlook_calendar_graph_ready() if can_manage_outlook_calendar else False,
            "calendar_user_choices": calendar_user_choices,
            "today_iso": schedule_today.isoformat(),
            "lookahead_rows": lookahead_rows,
            "admin_deadline_rows": admin_deadline_rows,
            "admin_deadline_total": len(admin_deadline_rows),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_maintenance_schedule_page_url(asset_id=selected_asset.id if selected_asset else 0),
                new_url=reverse("assets:asset_create"),
                new_label="+ Nuovo asset",
                search_placeholder="Ricerca manutenzioni per asset, reparto o intervento",
            ),
        },
    )


@login_required
def assistance_contract_list(request: HttpRequest) -> HttpResponse:
    from anagrafica.models import Fornitore

    today = timezone.localdate()
    can_manage_contracts = _is_assets_admin(request)
    can_manage_outlook_calendar = can_manage_contracts
    calendar_user_choices: list[tuple[str, str]] = []
    calendar_user_details: dict[str, dict[str, str]] = {}
    if can_manage_outlook_calendar:
        calendar_user_choices, calendar_user_details = _legacy_employee_options()

    selected_asset_id = _as_int(request.POST.get("asset_id") or request.GET.get("asset"), default=0)
    supplier_filter = _as_int(request.GET.get("supplier"), default=0)
    state_filter = _clean_string(request.GET.get("state")) or "all"
    scope_filter = _clean_string(request.GET.get("scope")) or "all"
    q = _clean_string(request.GET.get("q"))
    selected_asset = None
    if selected_asset_id:
        selected_asset = Asset.objects.select_related("asset_category").filter(pk=selected_asset_id).first()

    edit_id = _as_int(request.POST.get("edit_id") or request.GET.get("edit"), default=0)
    edit_contract = None
    if edit_id:
        edit_contract = (
            AssistanceContract.objects.select_related("supplier", "asset", "asset_category", "document")
            .filter(pk=edit_id)
            .first()
        )

    form = AssistanceContractForm(
        instance=edit_contract,
        locked_asset=selected_asset if selected_asset is not None and edit_contract is None else None,
    )

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action == "create_outlook_calendar_event":
            redirect_url = _assistance_contract_redirect_from_request(request)
            if not can_manage_outlook_calendar:
                messages.error(request, "Solo gli admin assets possono creare eventi Outlook su calendari utente.")
                return redirect(redirect_url)
            if selected_asset is None:
                messages.error(
                    request,
                    "Per i contratti assistenza seleziona prima un asset: l'evento Outlook viene creato sul contesto asset.",
                )
                return redirect(redirect_url)

            contract_id = _as_int(request.POST.get("contract_id"), default=0)
            target_legacy_user_id = _as_int(request.POST.get("target_legacy_user_id"), default=0)
            contract = (
                AssistanceContract.objects.select_related("supplier", "asset", "asset_category", "document")
                .filter(pk=contract_id)
                .first()
            )
            target_details = calendar_user_details.get(str(target_legacy_user_id))
            if contract is None:
                messages.error(request, "Contratto assistenza non trovato.")
                return redirect(redirect_url)
            if target_details is None:
                messages.error(request, "Seleziona un utente valido per il calendario Outlook.")
                return redirect(redirect_url)
            if not contract.applies_to_asset(selected_asset):
                messages.error(request, "Il contratto selezionato non si applica all'asset attualmente filtrato.")
                return redirect(redirect_url)
            if not isinstance(contract.end_date, date):
                messages.error(request, "Questo contratto non ha una scadenza calendarizzabile.")
                return redirect(redirect_url)

            try:
                entry, created = _create_asset_assistance_contract_calendar_event(
                    request=request,
                    asset=selected_asset,
                    contract=contract,
                    target_legacy_user_id=target_legacy_user_id,
                    target_details=target_details,
                )
                target_label = entry.target_display_name or entry.target_email or f"utente {entry.target_legacy_user_id}"
                if created:
                    log_action(
                        request,
                        "create_assistance_contract_calendar_event",
                        "assets",
                        {
                            "contract_id": contract.id,
                            "asset_id": selected_asset.id,
                            "due_date": str(contract.end_date),
                            "target_legacy_user_id": entry.target_legacy_user_id,
                            "target_email": entry.target_email,
                            "graph_event_id": entry.graph_event_id,
                        },
                    )
                    messages.success(
                        request,
                        f"Evento Outlook creato per {target_label} sulla scadenza del {entry.due_date:%d-%m-%Y}.",
                    )
                else:
                    messages.info(
                        request,
                        f"Esiste gia un evento Outlook per {target_label} su questa stessa scadenza.",
                    )
            except Exception as exc:
                messages.error(request, f"Creazione evento Outlook fallita: {exc}")
            return redirect(redirect_url)

        if action in {"create_assistance_contract", "update_assistance_contract", "delete_assistance_contract"} and not can_manage_contracts:
            messages.error(request, "Solo admin puo gestire i contratti assistenza.")
            return redirect(
                _assistance_contracts_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    supplier_filter=supplier_filter,
                    state=state_filter,
                    scope=scope_filter,
                    q=q,
                )
            )

        if action in {"create_assistance_contract", "update_assistance_contract"}:
            instance = edit_contract if action == "update_assistance_contract" else None
            form = AssistanceContractForm(
                request.POST,
                instance=instance,
                locked_asset=selected_asset if selected_asset is not None and instance is None else None,
            )
            if form.is_valid():
                contract = form.save()
                log_action(
                    request,
                    "update_assistance_contract" if instance is not None else "create_assistance_contract",
                    "assets",
                    {
                        "contract_id": contract.id,
                        "supplier_id": contract.supplier_id,
                        "asset_id": contract.asset_id,
                        "asset_category_id": contract.asset_category_id,
                    },
                )
                messages.success(
                    request,
                    "Contratto assistenza aggiornato." if instance is not None else "Contratto assistenza creato.",
                )
                return redirect(
                    _assistance_contracts_page_url(
                        asset_id=selected_asset.id if selected_asset else 0,
                        supplier_filter=supplier_filter,
                        state=state_filter,
                        scope=scope_filter,
                        q=q,
                    )
                )
        elif action == "delete_assistance_contract":
            contract = AssistanceContract.objects.filter(
                pk=_as_int(request.POST.get("contract_id"), default=0)
            ).first()
            if contract is None:
                messages.error(request, "Contratto assistenza non trovato.")
            else:
                log_action(
                    request,
                    "delete_assistance_contract",
                    "assets",
                    {
                        "contract_id": contract.id,
                        "supplier_id": contract.supplier_id,
                    },
                )
                contract.delete()
                messages.success(request, "Contratto assistenza eliminato.")
            return redirect(
                _assistance_contracts_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    supplier_filter=supplier_filter,
                    state=state_filter,
                    scope=scope_filter,
                    q=q,
                )
            )

    contract_rows: list[dict[str, object]] = []
    contract_qs = AssistanceContract.objects.select_related("supplier", "asset", "asset_category", "document").order_by(
        "-is_active",
        "end_date",
        "supplier__ragione_sociale",
        "title",
        "id",
    )
    if supplier_filter:
        contract_qs = contract_qs.filter(supplier_id=supplier_filter)
    if q:
        contract_qs = contract_qs.filter(
            Q(title__icontains=q)
            | Q(code__icontains=q)
            | Q(supplier__ragione_sociale__icontains=q)
            | Q(coverage_summary__icontains=q)
        )

    for contract in contract_qs:
        state = contract_state_payload(contract, today=today)
        scope = "asset" if contract.asset_id else "category" if contract.asset_category_id else "global"
        if selected_asset is not None and not contract.applies_to_asset(selected_asset):
            continue
        if state_filter != "all" and state["status"] != state_filter:
            continue
        if scope_filter != "all" and scope != scope_filter:
            continue
        scope_payload = _contract_scope_payload(contract)
        contract_rows.append(
            {
                "contract": contract,
                "state": state,
                "scope": scope,
                "scope_payload": scope_payload,
                "edit_url": _assistance_contracts_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    edit_id=contract.id,
                    supplier_filter=supplier_filter,
                    state=state_filter,
                    scope=scope_filter,
                    q=q,
                ),
                "supplier_url": reverse("fornitori:fornitore_detail", kwargs={"fornitore_id": contract.supplier_id}),
                "asset_url": (
                    reverse("assets:asset_view", kwargs={"id": contract.asset_id})
                    if contract.asset_id
                    else ""
                ),
            }
        )

    contract_event_map: dict[int, list[AssetCalendarEvent]] = defaultdict(list)
    if selected_asset is not None and contract_rows:
        contract_ids = [row["contract"].id for row in contract_rows if isinstance(row["contract"].end_date, date)]
        if contract_ids:
            for entry in (
                AssetCalendarEvent.objects.select_related("assistance_contract")
                .filter(
                    event_kind=AssetCalendarEvent.KIND_ASSISTANCE_CONTRACT,
                    asset_id=selected_asset.id,
                    assistance_contract_id__in=contract_ids,
                )
                .order_by("target_display_name", "target_email", "created_at", "id")
            ):
                if entry.assistance_contract_id:
                    contract_event_map[entry.assistance_contract_id].append(entry)

    default_calendar_user_id = _asset_calendar_default_user_id(selected_asset, calendar_user_details)
    for row in contract_rows:
        contract = row["contract"]
        row["can_create_calendar_event"] = bool(
            selected_asset is not None
            and isinstance(contract.end_date, date)
            and contract.applies_to_asset(selected_asset)
        )
        row["calendar_event_rows"] = list(contract_event_map.get(contract.id, []))
        row["default_calendar_user_id"] = default_calendar_user_id

    periodic_cost_total = Decimal("0")
    for row in contract_rows:
        if row["state"]["status"] in {"active", "expiring"} and row["contract"].periodic_cost_eur is not None:
            periodic_cost_total += row["contract"].periodic_cost_eur

    return render(
        request,
        "assets/pages/assistance_contract_list.html",
        {
            "page_title": "Contratti assistenza",
            "form": form,
            "contract_rows": contract_rows,
            "contract_total": len(contract_rows),
            "active_count": sum(1 for row in contract_rows if row["state"]["status"] == "active"),
            "expiring_count": sum(1 for row in contract_rows if row["state"]["status"] == "expiring"),
            "expired_count": sum(1 for row in contract_rows if row["state"]["status"] == "expired"),
            "periodic_cost_total": periodic_cost_total,
            "can_manage_contracts": can_manage_contracts,
            "is_edit": edit_contract is not None,
            "edit_contract": edit_contract,
            "selected_asset": selected_asset,
            "supplier_filter": supplier_filter,
            "state_filter": state_filter,
            "scope_filter": scope_filter,
            "q": q,
            "supplier_options": Fornitore.objects.filter(is_active=True).order_by("ragione_sociale", "id"),
            "state_choices": [
                ("all", "Tutti"),
                ("active", "Attivi"),
                ("expiring", "In scadenza"),
                ("expired", "Scaduti"),
                ("inactive", "Disattivi"),
                ("scheduled", "Non ancora attivi"),
            ],
            "scope_choices": [
                ("all", "Tutti"),
                ("global", "Generali"),
                ("category", "Per categoria"),
                ("asset", "Per asset"),
            ],
            "clear_filters_url": _assistance_contracts_page_url(asset_id=selected_asset.id if selected_asset else 0),
            "can_manage_outlook_calendar": can_manage_outlook_calendar,
            "outlook_calendar_ready": _outlook_calendar_graph_ready() if can_manage_outlook_calendar else False,
            "calendar_user_choices": calendar_user_choices,
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_assistance_contracts_page_url(asset_id=selected_asset.id if selected_asset else 0),
                search_placeholder="Ricerca contratti per fornitore, codice o copertura",
            ),
        },
        )


@login_required
def software_license_list(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    can_manage_licenses = _is_assets_admin(request)

    selected_asset_id = _as_int(request.POST.get("asset_id") or request.GET.get("asset"), default=0)
    selected_anagrafica_id = _as_int(request.POST.get("anagrafica_id") or request.GET.get("anagrafica"), default=0)
    category_filter = _clean_string(request.GET.get("category")) or "all"
    status_filter = _clean_string(request.GET.get("status")) or "all"
    assignee_filter = _clean_string(request.GET.get("assignee")) or "all"
    q = _clean_string(request.GET.get("q"))

    selected_asset = None
    if selected_asset_id:
        selected_asset = Asset.objects.select_related("asset_category").filter(pk=selected_asset_id).first()

    employee_choices, employee_details = _anagrafica_employee_options()
    selected_employee = None
    if selected_anagrafica_id:
        selected_employee = employee_details.get(str(selected_anagrafica_id))
        if selected_employee is None:
            try:
                row = (
                    AnagraficaDipendente.objects.filter(pk=selected_anagrafica_id)
                    .values(
                        "id",
                        "nome",
                        "cognome",
                        "aliasusername",
                        "reparto",
                        "email",
                        "email_notifica",
                        "utente_id",
                    )
                    .first()
                )
            except DatabaseError:
                row = None
            if row:
                display_name = " ".join(
                    [value for value in [_clean_string(row.get("cognome")), _clean_string(row.get("nome"))] if value]
                ).strip() or _clean_string(row.get("aliasusername")) or f"Dipendente #{selected_anagrafica_id}"
                selected_employee = {
                    "display_name": display_name,
                    "email": _clean_string(row.get("email")),
                    "notification_email": _clean_string(row.get("email_notifica")),
                    "reparto": _clean_string(row.get("reparto")),
                    "legacy_user_id": str(row.get("utente_id") or "").strip(),
                }
                if str(selected_anagrafica_id) not in employee_details:
                    employee_details[str(selected_anagrafica_id)] = dict(selected_employee)
                    employee_choices.append((str(selected_anagrafica_id), display_name))

    edit_id = _as_int(request.POST.get("edit_id") or request.GET.get("edit"), default=0)
    edit_license = None
    if edit_id:
        edit_license = (
            SoftwareLicense.objects.select_related("asset")
            .filter(pk=edit_id)
            .first()
        )

    form = SoftwareLicenseForm(
        instance=edit_license,
        locked_asset=selected_asset if selected_asset is not None and edit_license is None else None,
        locked_employee_id=str(selected_anagrafica_id) if selected_anagrafica_id and edit_license is None else "",
        employee_choices=employee_choices,
        employee_details=employee_details,
    )

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action in {"create_software_license", "update_software_license", "delete_software_license"} and not can_manage_licenses:
            messages.error(request, "Solo admin puo gestire le licenze software.")
            return redirect(
                _software_licenses_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    anagrafica_id=selected_anagrafica_id,
                    category=category_filter,
                    status=status_filter,
                    assignee=assignee_filter,
                    q=q,
                )
            )

        if action in {"create_software_license", "update_software_license"}:
            instance = edit_license if action == "update_software_license" else None
            form = SoftwareLicenseForm(
                request.POST,
                instance=instance,
                locked_asset=selected_asset if selected_asset is not None and instance is None else None,
                locked_employee_id=str(selected_anagrafica_id) if selected_anagrafica_id and instance is None else "",
                employee_choices=employee_choices,
                employee_details=employee_details,
            )
            if form.is_valid():
                license_row = form.save()
                log_action(
                    request,
                    "update_software_license" if instance is not None else "create_software_license",
                    "assets",
                    {
                        "license_id": license_row.id,
                        "asset_id": license_row.asset_id,
                        "assigned_anagrafica_id": license_row.assigned_anagrafica_id,
                        "assigned_legacy_user_id": license_row.assigned_legacy_user_id,
                    },
                )
                messages.success(
                    request,
                    "Licenza software aggiornata." if instance is not None else "Licenza software creata.",
                )
                return redirect(
                    _software_licenses_page_url(
                        asset_id=selected_asset.id if selected_asset else 0,
                        anagrafica_id=selected_anagrafica_id,
                        category=category_filter,
                        status=status_filter,
                        assignee=assignee_filter,
                        q=q,
                    )
                )
        elif action == "delete_software_license":
            license_row = SoftwareLicense.objects.filter(
                pk=_as_int(request.POST.get("license_id"), default=0)
            ).first()
            if license_row is None:
                messages.error(request, "Licenza software non trovata.")
            else:
                log_action(
                    request,
                    "delete_software_license",
                    "assets",
                    {
                        "license_id": license_row.id,
                        "asset_id": license_row.asset_id,
                        "assigned_anagrafica_id": license_row.assigned_anagrafica_id,
                    },
                )
                license_row.delete()
                messages.success(request, "Licenza software eliminata.")
            return redirect(
                _software_licenses_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    anagrafica_id=selected_anagrafica_id,
                    category=category_filter,
                    status=status_filter,
                    assignee=assignee_filter,
                    q=q,
                )
            )

    license_qs = SoftwareLicense.objects.select_related("asset").order_by(
        "category",
        "vendor",
        "product_name",
        "id",
    )
    if selected_asset is not None:
        license_qs = license_qs.filter(asset_id=selected_asset.id)
    if selected_anagrafica_id:
        legacy_user_id = 0
        if selected_employee is not None:
            try:
                legacy_user_id = int(selected_employee.get("legacy_user_id") or 0)
            except (TypeError, ValueError):
                legacy_user_id = 0
        if legacy_user_id:
            license_qs = license_qs.filter(
                Q(assigned_anagrafica_id=selected_anagrafica_id) | Q(assigned_legacy_user_id=legacy_user_id)
            )
        else:
            license_qs = license_qs.filter(assigned_anagrafica_id=selected_anagrafica_id)
    if category_filter != "all":
        license_qs = license_qs.filter(category=category_filter)
    if assignee_filter == "asset":
        license_qs = license_qs.filter(asset_id__isnull=False)
    elif assignee_filter == "user":
        license_qs = license_qs.filter(assigned_anagrafica_id__isnull=False)
    elif assignee_filter == "unassigned":
        license_qs = license_qs.filter(asset_id__isnull=True, assigned_anagrafica_id__isnull=True)

    if status_filter != "all":
        if status_filter == "inactive":
            license_qs = license_qs.filter(is_active=False)
        else:
            license_qs = license_qs.filter(is_active=True)
            if status_filter == "expired":
                license_qs = license_qs.filter(expiry_date__lt=today)
            elif status_filter == "expiring":
                license_qs = license_qs.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30))
            elif status_filter == "active":
                license_qs = license_qs.filter(
                    Q(expiry_date__isnull=True) | Q(expiry_date__gt=today + timedelta(days=30))
                )

    if q:
        license_qs = license_qs.filter(
            Q(product_name__icontains=q)
            | Q(vendor__icontains=q)
            | Q(edition__icontains=q)
            | Q(license_reference__icontains=q)
            | Q(account_email__icontains=q)
            | Q(assigned_to_display__icontains=q)
            | Q(asset__asset_tag__icontains=q)
            | Q(asset__name__icontains=q)
        )

    license_rows: list[dict[str, object]] = []
    for license_row in license_qs:
        state = _software_license_state_payload(license_row, today=today)
        license_rows.append(
            {
                "license": license_row,
                "state": state,
                "asset_url": (
                    reverse("assets:asset_view", kwargs={"id": license_row.asset_id})
                    if license_row.asset_id
                    else ""
                ),
                "employee_url": (
                    reverse("anagrafica:dipendente_detail", args=[license_row.assigned_anagrafica_id])
                    if license_row.assigned_anagrafica_id
                    else ""
                ),
                "edit_url": _software_licenses_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    anagrafica_id=selected_anagrafica_id,
                    edit_id=license_row.id,
                    category=category_filter,
                    status=status_filter,
                    assignee=assignee_filter,
                    q=q,
                ),
            }
        )

    return render(
        request,
        "assets/pages/software_license_list.html",
        {
            "page_title": "Licenze software",
            "form": form,
            "license_rows": license_rows,
            "license_total": len(license_rows),
            "active_count": sum(1 for row in license_rows if row["state"]["status"] == "active"),
            "expiring_count": sum(1 for row in license_rows if row["state"]["status"] == "expiring"),
            "expired_count": sum(1 for row in license_rows if row["state"]["status"] == "expired"),
            "inactive_count": sum(1 for row in license_rows if row["state"]["status"] == "inactive"),
            "unassigned_count": sum(
                1
                for row in license_rows
                if row["license"].assignment_scope == "unassigned"
            ),
            "can_manage_licenses": can_manage_licenses,
            "is_edit": edit_license is not None,
            "edit_license": edit_license,
            "selected_asset": selected_asset,
            "selected_employee": selected_employee,
            "selected_anagrafica_id": selected_anagrafica_id,
            "category_filter": category_filter,
            "status_filter": status_filter,
            "assignee_filter": assignee_filter,
            "q": q,
            "category_choices": [("all", "Tutte"), *SoftwareLicense.CATEGORY_CHOICES],
            "status_choices": [
                ("all", "Tutte"),
                ("active", "Attive"),
                ("expiring", "In scadenza"),
                ("expired", "Scadute"),
                ("inactive", "Disattive"),
            ],
            "assignee_choices": [
                ("all", "Tutte"),
                ("asset", "Assegnate a asset"),
                ("user", "Assegnate a dipendente"),
                ("unassigned", "Da assegnare"),
            ],
            "clear_filters_url": _software_licenses_page_url(
                asset_id=selected_asset.id if selected_asset else 0,
                anagrafica_id=selected_anagrafica_id,
            ),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=_software_licenses_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    anagrafica_id=selected_anagrafica_id,
                ),
                search_placeholder="Ricerca per prodotto, vendor, codice o asset",
            ),
        },
    )


@login_required
def device_list(request: HttpRequest) -> HttpResponse:
    """Lista dispositivi IT (PC, portatili, server, ecc.) con card layout."""
    form = DeviceFilterForm(request.GET or None)
    devices_qs = (
        Asset.objects
        .filter(asset_type__in=IT_DEVICE_TYPES)
        .select_related("asset_category")
        .order_by("asset_type", "reparto", "name", "asset_tag")
    )

    if form.is_valid():
        q = _clean_string(form.cleaned_data.get("q"))
        asset_type = _clean_string(form.cleaned_data.get("asset_type"))
        status = _clean_string(form.cleaned_data.get("status"))
        reparto = _clean_string(form.cleaned_data.get("reparto"))

        if q:
            devices_qs = devices_qs.filter(
                Q(asset_tag__icontains=q)
                | Q(internal_number__icontains=q)
                | Q(name__icontains=q)
                | Q(reparto__icontains=q)
                | Q(manufacturer__icontains=q)
                | Q(model__icontains=q)
                | Q(serial_number__icontains=q)
                | Q(assignment_to__icontains=q)
            )
        if asset_type:
            devices_qs = devices_qs.filter(asset_type=asset_type)
        if status:
            devices_qs = devices_qs.filter(status=status)
        if reparto:
            devices_qs = devices_qs.filter(
                Q(reparto__icontains=reparto) | Q(assignment_reparto__icontains=reparto)
            )

    device_base_qs = Asset.objects.filter(asset_type__in=IT_DEVICE_TYPES)
    total = device_base_qs.count()
    in_use_total = device_base_qs.filter(status=Asset.STATUS_IN_USE).count()
    in_stock_total = device_base_qs.filter(status=Asset.STATUS_IN_STOCK).count()
    in_repair_total = device_base_qs.filter(status=Asset.STATUS_IN_REPAIR).count()

    type_totals = list(
        device_base_qs
        .values("asset_type")
        .annotate(total=Count("id"))
        .order_by("asset_type")
    )
    type_label_map = dict(Asset.TYPE_CHOICES)
    for row in type_totals:
        row["label"] = type_label_map.get(row["asset_type"], row["asset_type"])

    visible_count = devices_qs.count()

    allowed_rows = [12, 24, 48, 96]
    rows = _as_int(request.GET.get("rows"), default=24)
    if rows not in allowed_rows:
        rows = 24
    paginator = Paginator(devices_qs, rows)
    page_number = _as_int(request.GET.get("page"), default=1)
    page_obj = paginator.get_page(page_number)
    devices = page_obj.object_list
    page_start = ((page_obj.number - 1) * rows + 1) if visible_count else 0
    page_end = (page_start + len(devices) - 1) if visible_count else 0

    rows_options = [
        {"value": v, "active": v == rows, "url": _query_url(request, rows=v, page=1)}
        for v in allowed_rows
    ]
    page_links = []
    if paginator.num_pages > 0:
        start_page = max(1, page_obj.number - 2)
        end_page = min(paginator.num_pages, page_obj.number + 2)
        for number in range(start_page, end_page + 1):
            page_links.append({
                "number": number,
                "active": number == page_obj.number,
                "url": _query_url(request, page=number, rows=rows),
            })
    prev_page_url = _query_url(request, page=page_obj.previous_page_number(), rows=rows) if page_obj.has_previous() else ""
    next_page_url = _query_url(request, page=page_obj.next_page_number(), rows=rows) if page_obj.has_next() else ""

    bulk_list_options = list(AssetListOption.objects.filter(is_active=True).order_by("field_key", "sort_order", "value", "id"))
    bulk_asset_categories = list(AssetCategory.objects.filter(is_active=True).order_by("sort_order", "label", "id"))
    maint_kpis = get_maintenance_kpis_for_types(IT_DEVICE_TYPES)
    return render(request, "assets/pages/device_list.html", {
        "page_title": "Dispositivi IT",
        "filters_form": form,
        "devices": devices,
        "visible_count": visible_count,
        "total": total,
        "in_use_total": in_use_total,
        "in_stock_total": in_stock_total,
        "in_repair_total": in_repair_total,
        "type_totals": type_totals,
        "rows": rows,
        "rows_options": rows_options,
        "page_links": page_links,
        "prev_page_url": prev_page_url,
        "next_page_url": next_page_url,
        "page_start": page_start,
        "page_end": page_end,
        "bulk_list_options": bulk_list_options,
        "bulk_asset_categories": bulk_asset_categories,
        "asset_type_choices": Asset.TYPE_CHOICES,
        "maint_kpis": maint_kpis,
        **_assets_shell_context(
            request,
            rows=rows,
            search_action=reverse("assets:device_list"),
            new_url=reverse("assets:asset_create"),
            new_label="+ Nuovo dispositivo",
            search_placeholder="Ricerca per tag, nome, utente, reparto o seriale",
        ),
    })


def _machine_availability_map(asset_ids: list[int]) -> dict[int, str]:
    """Ritorna un dict {asset_id: stato} per le macchine richieste.

    Stati: 'occupata' | 'manutenzione' | 'libera'
    """
    if not asset_ids:
        return {}
    today = timezone.localdate()
    result = {aid: "libera" for aid in asset_ids}

    # Macchine in manutenzione (OdL aperto)
    maint_ids = set(
        WorkOrder.objects
        .filter(asset_id__in=asset_ids, status=WorkOrder.STATUS_OPEN)
        .values_list("asset_id", flat=True)
    )
    for aid in maint_ids:
        result[aid] = "manutenzione"

    # Macchine occupate da task lavoro macchina attivi oggi
    try:
        from tasks.models import TaskExtraRef, TaskStatus
        occupied_ids = set(
            TaskExtraRef.objects
            .filter(
                asset_id__in=asset_ids,
                task__category__is_machine_work=True,
                task__status__in=[TaskStatus.TODO, TaskStatus.IN_PROGRESS],
                task__next_step_due__lte=today,
                task__due_date__gte=today,
            )
            .values_list("asset_id", flat=True)
        )
        for aid in occupied_ids:
            result[aid] = "occupata"
    except Exception:
        pass

    return result


@login_required
def work_machine_list(request: HttpRequest) -> HttpResponse:
    # La pagina "Asset produzione" e' confluita nell'inventario unico
    # (asset_list?group=production): stessa lista, colonne, ricerca rapida e
    # scheda. Manteniamo la rotta come redirect permanente per link, preferiti
    # e voci di navigazione esistenti, inoltrando i filtri compatibili.
    params: list[tuple[str, str]] = [("group", "production")]
    for key in ("q", "reparto", "cnc_only", "five_axes_only", "tcr_only"):
        value = _clean_string(request.GET.get(key))
        if value:
            params.append((key, value))
    asset_type_filter = _clean_string(request.GET.get("asset_type"))
    if asset_type_filter in PRODUCTION_ASSET_TYPES:
        params.append(("asset_type", asset_type_filter))
    return redirect(f"{reverse('assets:asset_list')}?{urlencode(params)}")


# ---------------------------------------------------------------------------
# EXPORT: shared helpers
# ---------------------------------------------------------------------------

def _xl_write_sheet(ws, headers: list, rows: list) -> None:
    import openpyxl.styles as xlst

    # Nomi/note/descrizioni degli asset sono testo libero: le celle passano da
    # core.excel_export (sede unica) per non finire scritte come formula viva.
    from core.excel_export import write_cell, write_row

    fill = xlst.PatternFill(fill_type="solid", fgColor="2563EB")
    hfont = xlst.Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    halign = xlst.Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        cell = write_cell(ws, 1, ci, h)
        cell.fill = fill
        cell.font = hfont
        cell.alignment = halign
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(rows, 2):
        write_row(ws, ri, row)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or "")) + 3, 10), 50)


def _apply_asset_export_filters(get_params) -> "QuerySet[Asset]":
    form = AssetFilterForm(get_params or None)
    qs = Asset.objects.select_related("asset_category").prefetch_related("endpoints")
    if form.is_valid():
        q = _clean_string(form.cleaned_data.get("q"))
        asset_type = _clean_string(form.cleaned_data.get("asset_type"))
        asset_category = form.cleaned_data.get("asset_category")
        reparto = _clean_string(form.cleaned_data.get("reparto"))
        vlan = form.cleaned_data.get("vlan")
        ip = _clean_string(form.cleaned_data.get("ip"))
        if q:
            qs = qs.filter(
                Q(asset_tag__icontains=q) | Q(name__icontains=q) | Q(serial_number__icontains=q)
                | Q(manufacturer__icontains=q) | Q(model__icontains=q)
                | Q(endpoints__endpoint_name__icontains=q) | Q(endpoints__ip__icontains=q)
            )
        if asset_type:
            qs = qs.filter(asset_type=asset_type)
        if asset_category:
            qs = qs.filter(asset_category_id__in=_category_subtree_ids(asset_category))
        if reparto:
            qs = qs.filter(reparto__icontains=reparto)
        if vlan is not None:
            qs = qs.filter(endpoints__vlan=vlan)
        if ip:
            qs = qs.filter(endpoints__ip__icontains=ip)
    return qs.distinct().order_by("name", "asset_tag")


# ---------------------------------------------------------------------------
# EXPORT: asset list
# ---------------------------------------------------------------------------

_ASSET_EXPORT_HEADERS = [
    "Tag", "Nome", "Tipo", "Categoria", "Reparto", "Stato",
    "Produttore", "Modello", "Matricola", "Assegnato a", "Posizione", "Note",
]


def _asset_export_row(a: "Asset") -> list:
    return [
        a.asset_tag or "",
        a.name or "",
        a.get_asset_type_display(),
        a.asset_category.label if a.asset_category_id else "",
        a.reparto or "",
        a.get_status_display(),
        a.manufacturer or "",
        a.model or "",
        a.serial_number or "",
        a.assignment_to or "",
        a.assignment_location or "",
        a.notes or "",
    ]


@login_required
def asset_list_export(request: HttpRequest) -> HttpResponse:
    scope = request.GET.get("scope", "filtered")
    fmt = request.GET.get("format", "xlsx")
    if scope == "full":
        assets = list(Asset.objects.select_related("asset_category").distinct().order_by("name", "asset_tag"))
    else:
        assets = list(_apply_asset_export_filters(request.GET))

    today = timezone.localdate().strftime("%Y%m%d")
    rows = [_asset_export_row(a) for a in assets]

    if fmt == "pdf":
        pdf_bytes = _report_table_pdf(title="Inventario asset", headers=_ASSET_EXPORT_HEADERS, rows=rows)
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="asset_{today}.pdf"'
        return resp

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset"
    _xl_write_sheet(ws, _ASSET_EXPORT_HEADERS, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="asset_{today}.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# EXPORT: workorder list
# ---------------------------------------------------------------------------

_WO_EXPORT_HEADERS = [
    "ID", "Asset Tag", "Asset Nome", "Reparto", "Categoria", "Titolo", "Tipo", "Origine", "Stato",
    "Assegnato a", "Eseguito da", "Fornitore", "Contratto", "Coperto da contratto",
    "Aperto il", "Chiuso il", "Durata min", "Fermo min", "Costo manodopera",
    "Costo materiali", "Costo totale", "Descrizione", "Risoluzione",
]


def _wo_export_row(wo: "WorkOrder") -> list:
    asset = wo.asset if wo.asset_id else None
    assigned = wo.assigned_to.get_full_name() or wo.assigned_to.username if wo.assigned_to_id else ""
    executed = wo.executed_by.get_full_name() or wo.executed_by.username if wo.executed_by_id else ""
    return [
        str(wo.id),
        asset.asset_tag if asset else "",
        asset.name if asset else "",
        asset.reparto if asset else "",
        asset.asset_category.label if asset and asset.asset_category_id else "",
        wo.title or "",
        wo.get_kind_display(),
        wo.get_origin_display(),
        wo.get_status_display(),
        assigned,
        executed,
        str(wo.supplier) if wo.supplier_id else "",
        wo.assistance_contract.title if wo.assistance_contract_id else "",
        "Si" if wo.covered_by_contract else "No",
        wo.opened_at.strftime("%d-%m-%Y %H:%M") if wo.opened_at else "",
        wo.closed_at.strftime("%d-%m-%Y %H:%M") if wo.closed_at else "",
        str(wo.intervention_duration_minutes or 0),
        str(wo.downtime_minutes or 0),
        str(wo.labor_cost_eur or ""),
        str(wo.materials_cost_eur or ""),
        str(wo.resolved_total_cost_eur or ""),
        wo.description or "",
        wo.resolution or "",
    ]


@login_required
def workorder_list_export(request: HttpRequest) -> HttpResponse:
    scope = request.GET.get("scope", "filtered")
    fmt = request.GET.get("format", "xlsx")
    qs = _apply_workorder_list_filters(request.GET, include_filters=(scope == "filtered"))
    if scope == "filtered" and _clean_string(request.GET.get("view")):
        qs = _apply_workorder_operational_view(
            qs,
            operational_view=_resolve_workorder_operational_view(request.GET),
            user=request.user,
        )
    workorders = list(qs.order_by("-opened_at"))
    today = timezone.localdate().strftime("%Y%m%d")
    rows = [_wo_export_row(wo) for wo in workorders]

    if fmt == "pdf":
        pdf_bytes = _report_table_pdf(title="Interventi / Work Orders", headers=_WO_EXPORT_HEADERS, rows=rows)
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="workorders_{today}.pdf"'
        return resp

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interventi"
    _xl_write_sheet(ws, _WO_EXPORT_HEADERS, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="workorders_{today}.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# EXPORT: work machine list PDF (XLSX già presente)
# ---------------------------------------------------------------------------

@login_required
def work_machine_export_pdf(request: HttpRequest) -> HttpResponse:
    scope = request.GET.get("scope", "filtered")
    if scope == "full":
        qs = Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES).select_related("work_machine")
    else:
        form = WorkMachineFilterForm(request.GET or None)
        qs = Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES).select_related("work_machine")
        if form.is_valid():
            q = _clean_string(form.cleaned_data.get("q"))
            reparto = _clean_string(form.cleaned_data.get("reparto"))
            status_f = _clean_string(form.cleaned_data.get("status"))
            cnc_only = bool(form.cleaned_data.get("cnc_only"))
            five_axes_only = bool(form.cleaned_data.get("five_axes_only"))
            tcr_only = bool(form.cleaned_data.get("tcr_only"))
            if q:
                qs = qs.filter(
                    Q(asset_tag__icontains=q) | Q(name__icontains=q) | Q(reparto__icontains=q)
                    | Q(manufacturer__icontains=q) | Q(model__icontains=q) | Q(serial_number__icontains=q)
                )
            if reparto:
                qs = qs.filter(reparto__icontains=reparto)
            if status_f:
                qs = qs.filter(status=status_f)
            if cnc_only:
                qs = qs.filter(work_machine__cnc_controlled=True)
            if five_axes_only:
                qs = qs.filter(work_machine__five_axes=True)
            if tcr_only:
                qs = qs.filter(work_machine__tcr_enabled=True)

    machines = list(qs.order_by("reparto", "name", "asset_tag"))
    headers = ["Tag", "Nome", "Reparto", "Stato", "Produttore", "Anno", "CNC", "5 assi", "TCR", "X mm", "Y mm", "Z mm"]
    rows = []
    for a in machines:
        wm = getattr(a, "work_machine", None)
        rows.append([
            a.asset_tag or "", a.name or "", a.reparto or "", a.get_status_display(),
            a.manufacturer or "",
            str(wm.year) if wm and wm.year else "",
            "Sì" if wm and wm.cnc_controlled else "No",
            "Sì" if wm and wm.five_axes else "No",
            "Sì" if wm and wm.tcr_enabled else "No",
            str(wm.x_mm) if wm and wm.x_mm else "",
            str(wm.y_mm) if wm and wm.y_mm else "",
            str(wm.z_mm) if wm and wm.z_mm else "",
        ])
    today = timezone.localdate().strftime("%Y%m%d")
    pdf_bytes = _report_table_pdf(title="Macchine di lavoro", headers=headers, rows=rows)
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="macchine_{today}.pdf"'
    return resp


@login_required
def work_machine_export_excel(request: HttpRequest) -> HttpResponse:
    """Esporta la lista macchine filtrata in formato Excel (.xlsx)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    from core.excel_export import write_cell, write_row

    form = WorkMachineFilterForm(request.GET or None)
    machines_qs = Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES).select_related("work_machine")

    if form.is_valid():
        q = _clean_string(form.cleaned_data.get("q"))
        reparto = _clean_string(form.cleaned_data.get("reparto"))
        status = _clean_string(form.cleaned_data.get("status"))
        cnc_only = bool(form.cleaned_data.get("cnc_only"))
        five_axes_only = bool(form.cleaned_data.get("five_axes_only"))
        tcr_only = bool(form.cleaned_data.get("tcr_only"))

        if q:
            machines_qs = machines_qs.filter(
                Q(asset_tag__icontains=q)
                | Q(name__icontains=q)
                | Q(reparto__icontains=q)
                | Q(manufacturer__icontains=q)
                | Q(model__icontains=q)
                | Q(serial_number__icontains=q)
            )
        if reparto:
            machines_qs = machines_qs.filter(reparto__icontains=reparto)
        if status:
            machines_qs = machines_qs.filter(status=status)
        if cnc_only:
            machines_qs = machines_qs.filter(work_machine__cnc_controlled=True)
        if five_axes_only:
            machines_qs = machines_qs.filter(work_machine__five_axes=True)
        if tcr_only:
            machines_qs = machines_qs.filter(work_machine__tcr_enabled=True)

    machines = list(machines_qs.order_by("reparto", "name", "asset_tag"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Macchine officina"

    headers = [
        "Tag", "Nome macchina", "Reparto", "Stato asset",
        "Produttore", "Modello", "N. seriale",
        "Anno", "TMC (mesi)", "Prossima manutenzione",
        "CNC", "5 assi", "TCR",
        "Cursa X (mm)", "Corsa Y (mm)", "Corsa Z (mm)",
        "Diametro (mm)", "Mandrino (mm)", "Pressione (bar)", "Accuracy",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

    for col_idx, header in enumerate(headers, 1):
        cell = write_cell(ws, 1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    for row_idx, asset in enumerate(machines, 2):
        wm = getattr(asset, "work_machine", None)
        values = [
            asset.asset_tag,
            asset.name,
            asset.reparto or "",
            asset.get_status_display(),
            asset.manufacturer or "",
            asset.model or "",
            asset.serial_number or "",
            wm.year if wm and wm.year else "",
            wm.tmc if wm and wm.tmc else "",
            wm.next_maintenance_date.strftime("%d-%m-%Y") if wm and wm.next_maintenance_date else "",
            "Sì" if wm and wm.cnc_controlled else "No",
            "Sì" if wm and wm.five_axes else "No",
            "Sì" if wm and wm.tcr_enabled else "No",
            wm.x_mm if wm and wm.x_mm else "",
            wm.y_mm if wm and wm.y_mm else "",
            wm.z_mm if wm and wm.z_mm else "",
            wm.diameter_mm if wm and wm.diameter_mm else "",
            wm.spindle_mm if wm and wm.spindle_mm else "",
            str(wm.pressure_bar) if wm and wm.pressure_bar else "",
            wm.accuracy_from if wm and wm.accuracy_from else "",
        ]
        write_row(ws, row_idx, values)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today_str = timezone.localdate().strftime("%Y%m%d")
    filename = f"macchine_officina_{today_str}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def work_machine_dashboard(request: HttpRequest) -> HttpResponse:
    now = timezone.now()
    today = timezone.localdate()
    reparto_filter = _clean_string(request.GET.get("reparto"))
    report_month_value = _clean_string(request.GET.get("month"))
    maintenance_month_dataset = _build_work_machine_maintenance_month_dataset(
        month_value=report_month_value,
        reparto_filter=reparto_filter,
        today=today,
    )

    machine_base_qs = Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES).select_related("work_machine").prefetch_related("documents")
    if reparto_filter:
        machine_base_qs = machine_base_qs.filter(reparto=reparto_filter)

    workorders_base = WorkOrder.objects.select_related("asset").filter(asset__asset_type__in=PRODUCTION_ASSET_TYPES)
    if reparto_filter:
        workorders_base = workorders_base.filter(asset__reparto=reparto_filter)

    open_workorders = workorders_base.filter(status=WorkOrder.STATUS_OPEN).order_by("opened_at", "id")
    wo_overdue_days = get_workorder_overdue_days()
    overdue_workorders = open_workorders.filter(opened_at__lt=now - timedelta(days=wo_overdue_days))
    recent_done_workorders = workorders_base.filter(
        status=WorkOrder.STATUS_DONE,
        closed_at__gte=now - timedelta(days=60),
    ).order_by("-closed_at", "-id")

    machine_rows = list(machine_base_qs.order_by("reparto", "name", "asset_tag"))
    manuals_count = 0
    specs_count = 0
    overdue_maintenance: list[dict[str, object]] = []
    warning_maintenance: list[dict[str, object]] = []
    missing_maintenance: list[dict[str, object]] = []
    for asset in machine_rows:
        extra = asset.extra_columns if isinstance(asset.extra_columns, dict) else {}
        raw_docs = extra.get("documents")
        categories = set()
        if isinstance(raw_docs, list):
            for row in raw_docs:
                if not isinstance(row, dict):
                    continue
                categories.add(_clean_string(str(row.get("category") or "SPECIFICHE")).upper())
        for uploaded in asset.documents.all():
            categories.add(uploaded.category)
        if "MANUALI" in categories:
            manuals_count += 1
        if "SPECIFICHE" in categories:
            specs_count += 1
        machine = getattr(asset, "work_machine", None)
        if not isinstance(machine, WorkMachine):
            continue
        maintenance_state = _work_machine_maintenance_state(machine, today)
        payload = {"asset": asset, "machine": machine, "state": maintenance_state}
        if maintenance_state["status"] == "overdue":
            overdue_maintenance.append(payload)
        elif maintenance_state["status"] == "warning":
            warning_maintenance.append(payload)
        elif maintenance_state["status"] == "missing":
            missing_maintenance.append(payload)

    overdue_maintenance.sort(key=lambda row: row["state"]["date"] or today)
    warning_maintenance.sort(key=lambda row: row["state"]["date"] or today)
    missing_maintenance.sort(key=lambda row: (row["asset"].reparto or "", row["asset"].name or ""))

    total_machines = len(machine_rows)
    reparto_totals = list(
        Asset.objects.filter(asset_type__in=PRODUCTION_ASSET_TYPES)
        .exclude(reparto="")
        .values("reparto")
        .annotate(total=Count("id"))
        .order_by("reparto")
    )

    _wmd_overdue = len(overdue_maintenance)
    _wmd_missing = len(missing_maintenance)
    _wmd_da_fare = _wmd_overdue + _wmd_missing
    _wmd_manutentati = max(0, total_machines - _wmd_da_fare)
    wmd_maint_kpis = {
        "coinvolti": total_machines,
        "manutentati": _wmd_manutentati,
        "da_manutentare": _wmd_da_fare,
        "percent_done": int(_wmd_manutentati / total_machines * 100) if total_machines else 0,
    }

    # P3.5 — Contatori macchine: precarica AssetMeter per le macchine con almeno un contatore
    machine_ids = [asset.id for asset in machine_rows]
    meters_by_asset_id: dict[int, list] = {}
    if machine_ids:
        for meter in AssetMeter.objects.filter(asset_id__in=machine_ids).select_related("updated_by").order_by("meter_type"):
            meters_by_asset_id.setdefault(meter.asset_id, []).append(meter)
    # Solo macchine con almeno un contatore configurato
    machines_with_meters = [
        asset for asset in machine_rows
        if asset.id in meters_by_asset_id
    ]

    return render(
        request,
        "assets/pages/work_machine_dashboard.html",
        {
            "page_title": "Dashboard officina",
            "reparto_filter": reparto_filter,
            "reparto_totals": reparto_totals,
            "total_machines": total_machines,
            "in_use_machines": sum(1 for asset in machine_rows if asset.status == Asset.STATUS_IN_USE),
            "in_repair_machines": sum(1 for asset in machine_rows if asset.status == Asset.STATUS_IN_REPAIR),
            "manuals_count": manuals_count,
            "specs_count": specs_count,
            "open_workorders": open_workorders[:10],
            "open_count": open_workorders.count(),
            "overdue_workorders": overdue_workorders[:10],
            "overdue_count": overdue_workorders.count(),
            "wo_overdue_days": wo_overdue_days,
            "overdue_workorder_ids": set(overdue_workorders.values_list("id", flat=True)),
            "recent_done_workorders": recent_done_workorders[:10],
            "recent_done_count": recent_done_workorders.count(),
            "overdue_maintenance": overdue_maintenance[:12],
            "overdue_maintenance_count": len(overdue_maintenance),
            "warning_maintenance": warning_maintenance[:12],
            "warning_maintenance_count": len(warning_maintenance),
            "missing_maintenance": missing_maintenance[:12],
            "missing_maintenance_count": len(missing_maintenance),
            "due_count": len(overdue_maintenance) + len(warning_maintenance),
            "maint_kpis": wmd_maint_kpis,
            "maintenance_month_count": maintenance_month_dataset["total_count"],
            "maintenance_month_label": maintenance_month_dataset["month_label"],
            "maintenance_month_code": maintenance_month_dataset["month_code"],
            "maintenance_month_pdf_url": _work_machine_maintenance_month_pdf_url(
                month_code=str(maintenance_month_dataset["month_code"]),
                reparto_filter=reparto_filter,
            ),
            "today": today,
            "machines_with_meters": machines_with_meters,
            "meters_by_asset_id": meters_by_asset_id,
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:work_machine_list"),
                new_url=reverse("assets:work_machine_create"),
                new_label="+ Nuova macchina",
                search_placeholder="Ricerca rapida per macchina, tag, reparto o seriale",
            ),
        },
    )


@login_required
def periodic_verification_list(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    can_manage_periodic_verifications = _is_assets_admin(request)
    can_manage_outlook_calendar = can_manage_periodic_verifications
    calendar_user_choices: list[tuple[str, str]] = []
    calendar_user_details: dict[str, dict[str, str]] = {}
    if can_manage_outlook_calendar:
        calendar_user_choices, calendar_user_details = _legacy_employee_options()

    selected_asset_id = _as_int(request.POST.get("asset_id") or request.GET.get("asset"), default=0)
    selected_asset = None
    if selected_asset_id:
        selected_asset = Asset.objects.filter(pk=selected_asset_id).only("id", "asset_tag", "name", "reparto", "asset_type").first()

    raw_scope = request.POST.get("scope") or request.GET.get("scope")
    if raw_scope:
        periodic_scope = _normalize_reports_scope(raw_scope)
    elif selected_asset is not None and selected_asset.asset_type in IT_DEVICE_TYPES:
        periodic_scope = "it"
    else:
        periodic_scope = "production"
    periodic_context = _periodic_scope_context(periodic_scope)
    periodic_asset_types = list(periodic_context["asset_types"])
    scope_asset_queryset = Asset.objects.filter(asset_type__in=periodic_asset_types)
    execution_window = _normalize_periodic_execution_window(request.GET.get("window"))
    execution_cutoff = _periodic_execution_window_cutoff(execution_window, today=today)
    periodic_view = _clean_string(request.GET.get("view")).lower()
    if periodic_view not in {"active", "attention", "planned", "archive"}:
        periodic_view = "active"
    periodic_q = _clean_string(request.GET.get("q"))

    if "scope" not in request.GET and request.method == "GET":
        query = request.GET.copy()
        query["scope"] = periodic_scope
        return redirect(f"{reverse('assets:periodic_verifications')}?{query.urlencode()}")

    if selected_asset is not None and selected_asset.asset_type not in periodic_asset_types:
        query = request.GET.copy()
        query["scope"] = "it" if selected_asset.asset_type in IT_DEVICE_TYPES else "production"
        return redirect(f"{reverse('assets:periodic_verifications')}?{query.urlencode()}")

    edit_id = _as_int(request.POST.get("edit_id") or request.GET.get("edit"), default=0)
    edit_verification = None
    if edit_id:
        edit_verification = (
            PeriodicVerification.objects.select_related("supplier", "created_by")
            .prefetch_related("assets")
            .filter(pk=edit_id)
            .first()
        )

    form = PeriodicVerificationForm(
        instance=edit_verification,
        actor=request.user,
        preselected_asset_id=selected_asset.id if selected_asset else 0,
        asset_queryset=scope_asset_queryset,
    )

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        if action == "create_outlook_calendar_event":
            redirect_url = _periodic_verification_redirect_from_request(request)
            if not can_manage_outlook_calendar:
                messages.error(request, "Solo gli admin assets possono creare eventi Outlook su calendari utente.")
                return redirect(redirect_url)
            if selected_asset is None:
                messages.error(
                    request,
                    "Per la manutenzione periodica seleziona prima un asset: l'evento Outlook viene creato sul contesto asset.",
                )
                return redirect(redirect_url)

            verification_id = _as_int(request.POST.get("verification_id"), default=0)
            target_legacy_user_id = _as_int(request.POST.get("target_legacy_user_id"), default=0)
            verification = (
                PeriodicVerification.objects.select_related("supplier", "created_by")
                .prefetch_related("assets")
                .filter(pk=verification_id)
                .first()
            )
            target_details = calendar_user_details.get(str(target_legacy_user_id))
            if verification is None:
                messages.error(request, "Piano di manutenzione periodica non trovato.")
                return redirect(redirect_url)
            if target_details is None:
                messages.error(request, "Seleziona un utente valido per il calendario Outlook.")
                return redirect(redirect_url)
            if not verification.assets.filter(pk=selected_asset.id).exists():
                messages.error(request, "La manutenzione periodica selezionata non e collegata all'asset attualmente filtrato.")
                return redirect(redirect_url)
            if not isinstance(verification.next_verification_date, date):
                messages.error(request, "Questa manutenzione periodica non ha una prossima data calendarizzabile.")
                return redirect(redirect_url)

            try:
                entry, created = _create_asset_periodic_verification_calendar_event(
                    request=request,
                    asset=selected_asset,
                    verification=verification,
                    target_legacy_user_id=target_legacy_user_id,
                    target_details=target_details,
                )
                target_label = entry.target_display_name or entry.target_email or f"utente {entry.target_legacy_user_id}"
                if created:
                    log_action(
                        request,
                        "create_periodic_verification_calendar_event",
                        "assets",
                        {
                            "verification_id": verification.id,
                            "asset_id": selected_asset.id,
                            "due_date": str(verification.next_verification_date),
                            "target_legacy_user_id": entry.target_legacy_user_id,
                            "target_email": entry.target_email,
                            "graph_event_id": entry.graph_event_id,
                        },
                    )
                    messages.success(
                        request,
                        f"Evento Outlook creato per {target_label} sulla scadenza del {entry.due_date:%d-%m-%Y}.",
                    )
                else:
                    messages.info(
                        request,
                        f"Esiste gia un evento Outlook per {target_label} su questa stessa scadenza.",
                    )
            except Exception as exc:
                messages.error(request, f"Creazione evento Outlook fallita: {exc}")
            return redirect(redirect_url)

        if action == "record_periodic_verification_execution":
            redirect_url = _periodic_verification_redirect_from_request(request)
            if not can_manage_periodic_verifications:
                messages.error(request, "Solo admin puo registrare esecuzioni di manutenzione periodica.")
                return redirect(redirect_url)
            verification_id = _as_int(request.POST.get("verification_id"), default=0)
            verification = (
                PeriodicVerification.objects.select_related("supplier")
                .prefetch_related("assets")
                .filter(pk=verification_id)
                .first()
            )
            if verification is None:
                messages.error(request, "Piano di manutenzione periodica non trovato.")
                return redirect(redirect_url)
            raw_asset_ids = request.POST.getlist("execution_asset_ids")
            if not raw_asset_ids:
                legacy_single = request.POST.get("execution_asset_id")
                if legacy_single:
                    raw_asset_ids = [legacy_single]
            target_asset_ids: list[int] = []
            for raw_id in raw_asset_ids:
                parsed = _as_int(raw_id, default=0)
                if parsed and parsed not in target_asset_ids:
                    target_asset_ids.append(parsed)
            target_assets = list(
                verification.assets.filter(pk__in=target_asset_ids)
            ) if target_asset_ids else []
            if not target_assets:
                messages.error(request, "Seleziona almeno un asset coinvolto nel piano per registrare l'esecuzione.")
                return redirect(redirect_url)
            executed_on_raw = _clean_string(request.POST.get("execution_date"))
            if executed_on_raw:
                try:
                    executed_on = date.fromisoformat(executed_on_raw)
                except ValueError:
                    messages.error(request, "Data di esecuzione non valida.")
                    return redirect(redirect_url)
            else:
                executed_on = today
            if executed_on > today:
                messages.error(request, "La data di esecuzione non puo essere futura.")
                return redirect(redirect_url)
            duration_minutes = max(0, _as_int(request.POST.get("execution_duration_minutes"), default=0))
            try:
                cost_value = _parse_execution_cost_input(request.POST.get("execution_cost_eur"))
            except ValueError:
                messages.error(request, "Costo non valido: usa un numero non negativo (es. 120.50).")
                return redirect(redirect_url)
            resolution_text = _clean_string(request.POST.get("execution_notes"))
            uploads, upload_errors = _validate_execution_attachment_uploads(request)
            if upload_errors:
                for error in upload_errors:
                    messages.error(request, error)
                return redirect(redirect_url)
            created_workorders: list[WorkOrder] = []
            attachments_total = 0
            try:
                with transaction.atomic():
                    for target_asset in target_assets:
                        workorder = _build_execution_workorder(
                            asset=target_asset,
                            title=f"Esecuzione manutenzione periodica: {verification.name}",
                            description=resolution_text,
                            executed_on=executed_on,
                            duration_minutes=duration_minutes,
                            cost_value=cost_value,
                            resolution_text=resolution_text,
                            periodic_verification=verification,
                            supplier=verification.supplier,
                        )
                        created_workorders.append(workorder)
                        if uploads:
                            attachments_total += len(
                                _save_workorder_attachments(
                                    workorder=workorder,
                                    uploads=uploads,
                                    user=request.user,
                                )
                            )
                    previous_last = verification.last_verification_date
                    if previous_last is None or executed_on >= previous_last:
                        verification.last_verification_date = executed_on
                        verification.next_verification_date = _add_months(
                            executed_on, verification.frequency_months
                        )
                        verification.save(update_fields=["last_verification_date", "next_verification_date", "updated_at"])
            except ValidationError as exc:
                messages.error(request, f"Esecuzione non registrabile: {exc}")
                return redirect(redirect_url)
            except Exception as exc:
                messages.error(request, f"Errore registrazione esecuzione: {exc}")
                return redirect(redirect_url)
            log_action(
                request,
                "record_periodic_verification_execution",
                "assets",
                {
                    "verification_id": verification.id,
                    "asset_ids": [asset.id for asset in target_assets],
                    "workorder_ids": [wo.id for wo in created_workorders],
                    "executed_on": str(executed_on),
                    "cost_eur": str(cost_value) if cost_value is not None else None,
                    "duration_minutes": duration_minutes,
                    "attachments": attachments_total,
                },
            )
            asset_label = (
                target_assets[0].asset_tag
                if len(target_assets) == 1
                else f"{len(target_assets)} asset"
            )
            attach_suffix = f" ({attachments_total} allegati)" if attachments_total else ""
            messages.success(
                request,
                f"Esecuzione registrata su {asset_label} il {executed_on:%d-%m-%Y}{attach_suffix}.",
            )
            return redirect(redirect_url)

        if action == "convert_periodic_to_rule":
            redirect_url = _periodic_verification_redirect_from_request(request)
            if not can_manage_periodic_verifications:
                messages.error(request, "Solo admin puo inglobare la manutenzione periodica nei piani ordinari.")
                return redirect(redirect_url)
            verification_id = _as_int(request.POST.get("verification_id"), default=0)
            verification = (
                PeriodicVerification.objects.prefetch_related("assets__asset_category")
                .filter(pk=verification_id)
                .first()
            )
            if verification is None:
                messages.error(request, "Piano di manutenzione periodica non trovato.")
                return redirect(redirect_url)
            if verification.is_legacy:
                messages.info(request, "Questo piano e gia gestito da una regola di manutenzione.")
                return redirect(redirect_url)
            from .services.periodic_migration import migrate_periodic_verification_to_rule

            result = migrate_periodic_verification_to_rule(verification)
            if not result.get("ok"):
                messages.error(request, result.get("message") or "Inglobamento non possibile per questo piano.")
                return redirect(redirect_url)
            log_action(
                request,
                "convert_periodic_verification_to_rule",
                "assets",
                {
                    "verification_id": verification.id,
                    "rule_id": result["rule"].id,
                    "template_id": result["template"].id,
                    "created_template": result["created_template"],
                    "created_rule": result["created_rule"],
                    "threshold_days": result["threshold_days"],
                },
            )
            messages.success(
                request,
                f"Piano inglobato in {len(result.get('rules') or [result['rule']])} piano/i ordinari "
                f"({result['threshold_days']} giorni), con asset e storico preservati.",
            )
            return redirect(redirect_url)

        if action in {"create_periodic_verification", "update_periodic_verification", "delete_periodic_verification"} and not can_manage_periodic_verifications:
            messages.error(request, "Solo admin puo gestire la manutenzione periodica.")
            return redirect(_periodic_verifications_page_url(asset_id=selected_asset.id if selected_asset else 0, scope=periodic_scope))

        if action in {"create_periodic_verification", "update_periodic_verification"}:
            instance = edit_verification if action == "update_periodic_verification" else None
            if action == "update_periodic_verification" and instance is None:
                messages.error(request, "Piano di manutenzione periodica non trovato.")
                return redirect(_periodic_verifications_page_url(asset_id=selected_asset.id if selected_asset else 0, scope=periodic_scope))
            form = PeriodicVerificationForm(
                request.POST,
                instance=instance,
                actor=request.user,
                preselected_asset_id=selected_asset.id if selected_asset else 0,
                asset_queryset=scope_asset_queryset,
            )
            if form.is_valid():
                verification = form.save()
                message = (
                    "Manutenzione periodica aggiornata."
                    if instance is not None
                    else "Manutenzione periodica creata."
                )
                messages.success(request, message)
                return redirect(_periodic_verifications_page_url(asset_id=selected_asset.id if selected_asset else 0, scope=periodic_scope))
        elif action == "delete_periodic_verification":
            verification_id = _as_int(request.POST.get("verification_id"), default=0)
            verification = PeriodicVerification.objects.filter(pk=verification_id).first()
            if verification is None:
                messages.error(request, "Piano di manutenzione periodica non trovato.")
            else:
                verification_name = verification.name
                verification.delete()
                messages.success(request, f'Manutenzione periodica "{verification_name}" eliminata.')
            return redirect(_periodic_verifications_page_url(asset_id=selected_asset.id if selected_asset else 0, scope=periodic_scope))

    verification_rows: list[dict[str, object]] = []
    verification_event_map: dict[int, list[AssetCalendarEvent]] = defaultdict(list)
    if selected_asset is not None:
        linked_ids = list(
            PeriodicVerification.objects.filter(assets__id=selected_asset.id).values_list("id", flat=True).distinct()
        )
        if linked_ids:
            for entry in (
                AssetCalendarEvent.objects.select_related("periodic_verification")
                .filter(
                    event_kind=AssetCalendarEvent.KIND_PERIODIC_VERIFICATION,
                    asset_id=selected_asset.id,
                    periodic_verification_id__in=linked_ids,
                )
                .order_by("target_display_name", "target_email", "created_at", "id")
            ):
                if entry.periodic_verification_id:
                    verification_event_map[entry.periodic_verification_id].append(entry)

    default_calendar_user_id = _asset_calendar_default_user_id(selected_asset, calendar_user_details)
    verification_queryset = (
        PeriodicVerification.objects.select_related("supplier", "created_by")
        .prefetch_related("assets")
        .filter(assets__asset_type__in=periodic_asset_types)
        .distinct()
        .order_by("-is_active", "next_verification_date", "name", "id")
    )
    if selected_asset is not None:
        verification_queryset = verification_queryset.filter(assets=selected_asset)

    for verification in verification_queryset:
        linked_assets = [asset for asset in verification.assets.all() if asset.asset_type in periodic_asset_types]
        is_selected_asset_linked = bool(selected_asset and any(asset.id == selected_asset.id for asset in linked_assets))
        execution_rows = _periodic_execution_rows_for_verification(
            verification_id=verification.id,
            asset_id=selected_asset.id if selected_asset is not None else 0,
            cutoff_date=execution_cutoff,
        )
        # Idoneità alla conversione in regola (check leggero su asset già prefetchati;
        # la validazione completa avviene nell'azione convert_periodic_to_rule).
        all_assets = list(verification.assets.all())
        all_category_ids = {asset.asset_category_id for asset in all_assets}
        can_convert_to_rule = bool(
            can_manage_periodic_verifications
            and not verification.is_legacy
            and all_assets
            and None not in all_category_ids
            and len(all_category_ids) == 1
        )
        convert_block_reason = ""
        if can_manage_periodic_verifications and not verification.is_legacy and not can_convert_to_rule:
            convert_block_reason = (
                "Nessun asset collegato" if not all_assets
                else "Asset di categorie diverse o senza categoria"
            )
        verification_rows.append(
            {
                "verification": verification,
                "state": _periodic_verification_state(verification, today=today),
                "is_legacy": verification.is_legacy,
                "can_convert_to_rule": can_convert_to_rule,
                "convert_block_reason": convert_block_reason,
                "linked_assets": linked_assets,
                "linked_assets_count": len(linked_assets),
                "is_selected_asset_linked": is_selected_asset_linked,
                "can_create_calendar_event": bool(
                    selected_asset is not None
                    and is_selected_asset_linked
                    and isinstance(verification.next_verification_date, date)
                ),
                "calendar_event_rows": list(verification_event_map.get(verification.id, [])),
                "default_calendar_user_id": default_calendar_user_id,
                "edit_url": _periodic_verifications_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    edit_id=verification.id,
                    scope=periodic_scope,
                    window=execution_window,
                    view=periodic_view,
                    q=periodic_q,
                ),
                "execution_rows": execution_rows,
                "execution_count": len(execution_rows),
                "execution_assets": linked_assets,
            }
        )

    state_priority = {"overdue": 0, "warning": 1, "missing": 2, "ok": 3, "inactive": 4}
    verification_rows.sort(
        key=lambda row: (
            bool(row["is_legacy"]),
            state_priority.get(row["state"]["status"], 9),
            row["verification"].next_verification_date or date.max,
            row["verification"].name.lower(),
        )
    )
    periodic_counts = {
        "active": sum(
            1 for row in verification_rows if row["verification"].is_active and not row["is_legacy"]
        ),
        "attention": sum(
            1
            for row in verification_rows
            if row["verification"].is_active
            and not row["is_legacy"]
            and row["state"]["status"] in {"overdue", "warning", "missing"}
        ),
        "planned": sum(
            1
            for row in verification_rows
            if row["verification"].is_active
            and not row["is_legacy"]
            and row["state"]["status"] == "ok"
        ),
        "archive": sum(
            1 for row in verification_rows if row["is_legacy"] or not row["verification"].is_active
        ),
    }
    verification_total_count = len(verification_rows)
    legacy_verification_count = sum(1 for row in verification_rows if row["is_legacy"])

    def row_matches_view(row) -> bool:
        if periodic_view == "attention":
            return bool(
                row["verification"].is_active
                and not row["is_legacy"]
                and row["state"]["status"] in {"overdue", "warning", "missing"}
            )
        if periodic_view == "planned":
            return bool(
                row["verification"].is_active
                and not row["is_legacy"]
                and row["state"]["status"] == "ok"
            )
        if periodic_view == "archive":
            return bool(row["is_legacy"] or not row["verification"].is_active)
        return bool(row["verification"].is_active and not row["is_legacy"])

    verification_rows = [row for row in verification_rows if row_matches_view(row)]
    if periodic_q:
        search_term = periodic_q.casefold()
        verification_rows = [
            row
            for row in verification_rows
            if search_term
            in " ".join(
                [
                    row["verification"].name,
                    str(row["verification"].supplier or ""),
                    row["verification"].notes or "",
                    *[
                        f"{asset.asset_tag} {asset.name} {asset.reparto}"
                        for asset in row["linked_assets"]
                    ],
                ]
            ).casefold()
        ]

    selected_asset_linked_count = (
        verification_total_count
        if selected_asset is not None
        else 0
    )

    return render(
        request,
        "assets/pages/periodic_verification_list.html",
        {
            "page_title": periodic_context["title"],
            "periodic_scope": periodic_context["scope"],
            "periodic_page_title": periodic_context["title"],
            "periodic_page_subtitle": periodic_context["subtitle"],
            "periodic_base_url": _periodic_verifications_page_url(scope=periodic_scope),
            "form": form,
            "verification_rows": verification_rows,
            "verification_total": verification_total_count,
            "active_verification_count": periodic_counts["active"],
            "due_verification_count": periodic_counts["attention"],
            "legacy_verification_count": legacy_verification_count,
            "periodic_counts": periodic_counts,
            "periodic_view": periodic_view,
            "periodic_q": periodic_q,
            "periodic_view_urls": {
                view_key: _periodic_verifications_page_url(
                    asset_id=selected_asset.id if selected_asset else 0,
                    scope=periodic_scope,
                    view=view_key,
                )
                for view_key in ("active", "attention", "planned", "archive")
            },
            "periodic_current_view_url": _periodic_verifications_page_url(
                asset_id=selected_asset.id if selected_asset else 0,
                scope=periodic_scope,
                view=periodic_view,
                q=periodic_q,
            ),
            "periodic_clear_url": _periodic_verifications_page_url(
                asset_id=selected_asset.id if selected_asset else 0,
                scope=periodic_scope,
                view=periodic_view,
            ),
            "periodic_create_url": _periodic_verifications_page_url(
                asset_id=selected_asset.id if selected_asset else 0,
                scope=periodic_scope,
                view=periodic_view,
                q=periodic_q,
                create=True,
            ),
            "show_periodic_form": bool(
                can_manage_periodic_verifications
                and (edit_verification is not None or request.GET.get("create") == "1" or form.errors)
            ),
            "selected_asset": selected_asset,
            "selected_asset_linked_count": selected_asset_linked_count,
            "execution_window": execution_window,
            "execution_window_choices": list(PERIODIC_EXECUTION_WINDOW_CHOICES),
            "today_iso": today.isoformat(),
            "can_manage_periodic_verifications": can_manage_periodic_verifications,
            "can_manage_outlook_calendar": can_manage_outlook_calendar,
            "outlook_calendar_ready": _outlook_calendar_graph_ready() if can_manage_outlook_calendar else False,
            "calendar_user_choices": calendar_user_choices,
            "is_edit": edit_verification is not None,
            "edit_verification": edit_verification,
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def plant_layout_map(request: HttpRequest) -> HttpResponse:
    focus_asset_id = _as_int(request.GET.get("asset"), default=0)
    requested_category = _clean_string(request.GET.get("category"))
    if not requested_category and focus_asset_id:
        focus_marker = (
            PlantLayoutMarker.objects.filter(asset_id=focus_asset_id, layout__is_active=True)
            .select_related("layout")
            .order_by("layout__category", "layout__name", "id")
            .first()
        )
        if focus_marker is not None:
            requested_category = focus_marker.layout.category

    active_layouts = list(_plant_layout_queryset().filter(is_active=True).order_by("category", "name", "id"))
    selected_category = _preferred_plant_layout_category(active_layouts, requested_category=requested_category)
    layout = next(
        (row for row in active_layouts if _clean_string(row.category).casefold() == _clean_string(selected_category).casefold()),
        active_layouts[0] if active_layouts else None,
    )
    payload = _plant_layout_public_payload(layout)
    category_switches = _plant_layout_category_switches(
        active_layouts=active_layouts,
        selected_category=selected_category,
        focus_asset_id=focus_asset_id,
    )

    return render(
        request,
        "assets/pages/plant_layout_map.html",
        {
            "page_title": "Planimetrie impianti",
            "layout": layout,
            "plant_layout_payload": payload,
            "focus_asset_id": focus_asset_id,
            "selected_layout_category": selected_category,
            "layout_category_switches": category_switches,
            "can_manage_map": user_can_modulo_action(request, "assets", "admin_assets"),
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:work_machine_list"),
                new_url=reverse("assets:work_machine_create"),
                new_label="+ Nuova macchina",
                search_placeholder="Ricerca rapida per macchina, tag, reparto o seriale",
            ),
        },
    )


@login_required
def plant_layout_editor(request: HttpRequest) -> HttpResponse:
    all_layouts = list(
        _plant_layout_queryset().all().order_by("category", "-is_active", "-updated_at", "name", "id")
    )
    current_category = _clean_string(
        request.POST.get("category_filter") or request.POST.get("category") or request.GET.get("category")
    )
    current_layout_id = _as_int(request.POST.get("layout_id") or request.GET.get("layout"), default=0)
    create_new = _clean_string(request.POST.get("layout_mode") or request.GET.get("new")) in {"1", "new", "true"}
    selected_layout = None
    filtered_layouts = [
        row for row in all_layouts if not current_category or _clean_string(row.category).casefold() == current_category.casefold()
    ]
    if not create_new:
        selected_layout = next((row for row in all_layouts if row.id == current_layout_id), None)
        if selected_layout is not None and not current_category:
            current_category = selected_layout.category
            filtered_layouts = [
                row for row in all_layouts if _clean_string(row.category).casefold() == current_category.casefold()
            ]
        if selected_layout is None and filtered_layouts:
            selected_layout = next((row for row in filtered_layouts if row.is_active), None) or filtered_layouts[0]
        if selected_layout is None and all_layouts:
            selected_layout = next((row for row in all_layouts if row.is_active), None) or all_layouts[0]
            current_category = current_category or getattr(selected_layout, "category", "")
            filtered_layouts = [
                row for row in all_layouts
                if not current_category or _clean_string(row.category).casefold() == current_category.casefold()
            ]

    area_rows = _plant_layout_editor_area_rows(selected_layout)
    marker_rows = _plant_layout_editor_marker_rows(selected_layout)

    if request.method == "POST":
        action = _clean_string(request.POST.get("action")) or "save_layout"
        if action == "activate_layout":
            if selected_layout is None:
                messages.error(request, "Planimetria non trovata.")
            else:
                selected_layout.is_active = True
                selected_layout.save()
                messages.success(request, f"Planimetria \"{selected_layout.name}\" pubblicata nella categoria {selected_layout.category}.")
            layout_id = selected_layout.id if selected_layout else ""
            category_qs = f"&category={quote(selected_layout.category)}" if selected_layout else ""
            return redirect(f"{reverse('assets:plant_layout_editor')}?layout={layout_id}{category_qs}")

        if action == "delete_layout":
            if selected_layout is None:
                messages.error(request, "Planimetria non trovata.")
            else:
                deleted_name = selected_layout.name
                deleted_category = selected_layout.category
                selected_layout.delete()
                messages.success(request, f"Planimetria \"{deleted_name}\" eliminata.")
            if current_category:
                return redirect(f"{reverse('assets:plant_layout_editor')}?category={quote(current_category or deleted_category)}")
            return redirect("assets:plant_layout_editor")

        form = PlantLayoutForm(
            request.POST,
            request.FILES,
            instance=None if create_new else selected_layout,
        )
        if form.is_valid():
            layout = form.save()
            messages.success(request, f"Planimetria \"{layout.name}\" aggiornata nella categoria {layout.category}.")
            return redirect(f"{reverse('assets:plant_layout_editor')}?layout={layout.id}&category={quote(layout.category)}")
        area_rows = _safe_editor_json_rows(request.POST.get("areas_payload"))
        marker_rows = _safe_editor_json_rows(request.POST.get("markers_payload"))
    else:
        form = PlantLayoutForm(
            instance=selected_layout,
            initial={"category": current_category or getattr(selected_layout, "category", "") or PlantLayout.DEFAULT_CATEGORY},
        )

    editor_payload = {
        "layout": {
            "id": getattr(selected_layout, "id", None),
            "category": getattr(selected_layout, "category", current_category or PlantLayout.DEFAULT_CATEGORY),
            "name": getattr(selected_layout, "name", ""),
            "description": getattr(selected_layout, "description", ""),
            "image_url": (
                selected_layout.image.url
                if selected_layout is not None and selected_layout.image
                else ""
            ),
            "is_active": bool(getattr(selected_layout, "is_active", False)),
        },
        "areas": area_rows,
        "markers": marker_rows,
        "machines": _plant_layout_machine_catalog(),
    }
    available_categories = []
    seen_categories: set[str] = set()
    for row in all_layouts:
        category_key = _clean_string(row.category).casefold()
        if category_key in seen_categories:
            continue
        seen_categories.add(category_key)
        available_categories.append(row.category)
    layout_choices = [
        {
            "id": row.id,
            "category": row.category,
            "name": row.name,
            "is_active": row.is_active,
            "updated_at": timezone.localtime(row.updated_at).strftime("%d-%m-%Y %H:%M"),
            "edit_url": f"{reverse('assets:plant_layout_editor')}?layout={row.id}&category={quote(row.category)}",
        }
        for row in (filtered_layouts if current_category else all_layouts)
    ]
    layout_category_filters = [
        {
            "label": "Tutte",
            "active": not current_category,
            "url": reverse("assets:plant_layout_editor"),
        }
    ]
    for category in available_categories:
        layout_category_filters.append(
            {
                "label": category,
                "active": _clean_string(category).casefold() == current_category.casefold(),
                "url": f"{reverse('assets:plant_layout_editor')}?category={quote(category)}",
            }
        )

    return render(
        request,
        "assets/pages/plant_layout_editor.html",
        {
            "page_title": "Editor planimetrie impianti",
            "form": form,
            "selected_layout": selected_layout,
            "layout_choices": layout_choices,
            "layout_category_filters": layout_category_filters,
            "current_layout_category": current_category or getattr(selected_layout, "category", PlantLayout.DEFAULT_CATEGORY),
            "create_new": create_new,
            "editor_payload": editor_payload,
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:work_machine_list"),
                new_url=(
                    f"{reverse('assets:plant_layout_editor')}?new=1&category={quote(current_category)}"
                    if current_category
                    else f"{reverse('assets:plant_layout_editor')}?new=1"
                ),
                new_label="+ Nuova planimetria",
                search_placeholder="Ricerca rapida per macchina, tag, reparto o seriale",
            ),
        },
    )


@login_required
def work_machine_create(request: HttpRequest) -> HttpResponse:
    list_suggestions = _build_asset_list_suggestions()
    assignment_kwargs = _assignment_form_kwargs()
    if request.method == "POST":
        uploads, upload_errors = _validate_asset_document_uploads(request)
        form = WorkMachineAssetForm(request.POST, list_suggestions=list_suggestions, **assignment_kwargs)
        if form.is_valid() and not upload_errors:
            asset = form.save()
            if "foto_targhetta" in request.FILES:
                asset.foto_targhetta = request.FILES["foto_targhetta"]
                asset.save(update_fields=["foto_targhetta"])
            if form.cleaned_data.get("include_in_plant_layout"):
                marker_warning = _ensure_asset_plant_layout_marker(asset)
                if marker_warning:
                    messages.warning(request, marker_warning)
            _apply_asset_document_changes(
                asset,
                uploads=uploads,
                remove_ids=set(),
                actor=request.user,
            )
            messages.success(request, "Macchina di lavoro creata correttamente.")
            return redirect("assets:asset_view", id=asset.id)
        for error in upload_errors:
            form.add_error(None, error)
    else:
        form = WorkMachineAssetForm(
            initial={"status": Asset.STATUS_IN_USE},
            list_suggestions=list_suggestions,
            **assignment_kwargs,
        )

    return render(
        request,
        "assets/pages/work_machine_form.html",
        {
            "page_title": "Nuova macchina di lavoro",
            "form": form,
            "is_edit": False,
            "list_suggestions": list_suggestions,
            "assignment_department_choices": assignment_kwargs["assignment_department_choices"],
            "assignment_employee_details": assignment_kwargs["assignment_employee_details"],
            "asset_field_names": form.asset_field_names,
            "category_field_groups": form.category_field_groups,
            "category_dynamic_field_names": form.category_dynamic_field_names,
            "machine_field_names": form.machine_field_names,
            "assignment_field_names": form.assignment_field_names,
            "verification_field_names": form.verification_field_names,
            "document_field_map": form.document_field_map,
            "uploaded_documents_by_category": _build_uploaded_documents_context(None),
            "document_upload_field_map": ASSET_DOCUMENT_UPLOAD_FIELDS,
            "asset_label_designer_url": reverse("assets:asset_label_designer") + f"?scope=asset_type&asset_type={Asset.TYPE_WORK_MACHINE}",
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:work_machine_list"),
                new_url=reverse("assets:work_machine_create"),
                new_label="+ Nuova macchina",
                search_placeholder="Ricerca rapida per macchina, tag, reparto o seriale",
            ),
        },
    )


@login_required
def work_machine_edit(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:work_machine_list")
    asset = get_object_or_404(
        Asset.objects.select_related("work_machine").prefetch_related("documents"),
        pk=id,
        asset_type__in=PRODUCTION_ASSET_TYPES,
    )
    list_suggestions = _build_asset_list_suggestions()
    assignment_kwargs = _assignment_form_kwargs(asset)

    if request.method == "POST":
        uploads, upload_errors = _validate_asset_document_uploads(request, asset)
        remove_ids = {_as_int(value, default=0) for value in request.POST.getlist("remove_document_ids")}
        remove_ids = {value for value in remove_ids if value > 0}
        form = WorkMachineAssetForm(
            request.POST,
            instance=asset,
            work_machine=getattr(asset, "work_machine", None),
            list_suggestions=list_suggestions,
            **assignment_kwargs,
        )
        if form.is_valid() and not upload_errors:
            asset = form.save()
            if form.cleaned_data.get("include_in_plant_layout"):
                marker_warning = _ensure_asset_plant_layout_marker(asset)
                if marker_warning:
                    messages.warning(request, marker_warning)
            if request.POST.get("clear_foto_targhetta") == "1":
                if asset.foto_targhetta:
                    asset.foto_targhetta.delete(save=False)
                asset.foto_targhetta = None
                asset.save(update_fields=["foto_targhetta"])
            elif "foto_targhetta" in request.FILES:
                asset.foto_targhetta = request.FILES["foto_targhetta"]
                asset.save(update_fields=["foto_targhetta"])
            wm = getattr(asset, "work_machine", None)
            if wm:
                if request.POST.get("clear_photo") == "1":
                    if wm.photo:
                        wm.photo.delete(save=False)
                    wm.photo = None
                    wm.save(update_fields=["photo"])
                elif "photo" in request.FILES:
                    wm.photo = request.FILES["photo"]
                    wm.save(update_fields=["photo"])
            _apply_asset_document_changes(
                asset,
                uploads=uploads,
                remove_ids=remove_ids,
                actor=request.user,
            )
            messages.success(request, "Macchina di lavoro aggiornata.")
            return redirect("assets:asset_view", id=asset.id)
        for error in upload_errors:
            form.add_error(None, error)
    else:
        form = WorkMachineAssetForm(
            instance=asset,
            work_machine=getattr(asset, "work_machine", None),
            list_suggestions=list_suggestions,
            **assignment_kwargs,
        )

    return render(
        request,
        "assets/pages/work_machine_form.html",
        {
            "page_title": f"Modifica {asset.asset_tag}",
            "form": form,
            "asset": asset,
            "is_edit": True,
            "list_suggestions": list_suggestions,
            "assignment_department_choices": assignment_kwargs["assignment_department_choices"],
            "assignment_employee_details": assignment_kwargs["assignment_employee_details"],
            "asset_field_names": form.asset_field_names,
            "category_field_groups": form.category_field_groups,
            "category_dynamic_field_names": form.category_dynamic_field_names,
            "machine_field_names": form.machine_field_names,
            "assignment_field_names": form.assignment_field_names,
            "verification_field_names": form.verification_field_names,
            "document_field_map": form.document_field_map,
            "uploaded_documents_by_category": _build_uploaded_documents_context(asset),
            "document_upload_field_map": ASSET_DOCUMENT_UPLOAD_FIELDS,
            "asset_label_designer_url": reverse("assets:asset_label_designer") + f"?scope=asset&asset_id={asset.id}",
            **_assets_shell_context(
                request,
                rows=_as_int(request.GET.get("rows"), default=25),
                search_action=reverse("assets:work_machine_list"),
                new_url=reverse("assets:work_machine_create"),
                new_label="+ Nuova macchina",
                search_placeholder="Ricerca rapida per macchina, tag, reparto o seriale",
            ),
        },
    )


@login_required
def assignment_set(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(Asset, pk=id)
    user_choices, user_details = _legacy_employee_options()
    list_suggestions = _build_asset_list_suggestions(employee_details=user_details)
    selected_user_id = asset.assigned_legacy_user_id if asset.assigned_legacy_user_id else ""

    if request.method == "POST":
        form = AssetAssignmentForm(
            request.POST,
            instance=asset,
            user_choices=user_choices,
            selected_user_id=selected_user_id,
            list_suggestions=list_suggestions,
        )
        if form.is_valid():
            selected_raw = _clean_string(form.cleaned_data.get("assigned_user_id"))
            saved_asset: Asset = form.save(commit=False)
            if selected_raw:
                selected_info = user_details.get(selected_raw, {})
                saved_asset.assigned_legacy_user_id = int(selected_raw)
                selected_name = _clean_string(selected_info.get("display_name"))
                if selected_name:
                    saved_asset.assignment_to = selected_name
                selected_reparto = _clean_string(selected_info.get("reparto"))
                if selected_reparto:
                    saved_asset.assignment_reparto = selected_reparto
            else:
                saved_asset.assigned_legacy_user_id = None
            saved_asset.save()
            messages.success(request, "Assegnazione aggiornata.")
            return redirect("assets:asset_view", id=asset.id)
    else:
        form = AssetAssignmentForm(
            instance=asset,
            user_choices=user_choices,
            selected_user_id=selected_user_id,
            list_suggestions=list_suggestions,
        )

    selected_user_admin_url = ""
    try:
        selected_id = int(form.initial.get("assigned_user_id") or 0)
    except (TypeError, ValueError):
        selected_id = 0
    if selected_id:
        try:
            anag_id = AnagraficaDipendente.objects.filter(
                utente_id=selected_id
            ).values_list("id", flat=True).first()
            if anag_id:
                selected_user_admin_url = reverse(
                    "anagrafica:dipendente_detail",
                    kwargs={"legacy_id": anag_id},
                )
        except (NoReverseMatch, ValueError, TypeError):
            selected_user_admin_url = ""

    return render(
        request,
        "assets/pages/asset_assignment.html",
        {
            "page_title": f"Assegna {asset.asset_tag}",
            "asset": asset,
            "form": form,
            "selected_user_admin_url": selected_user_admin_url,
            "list_suggestions": list_suggestions,
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


def _workorder_attachment_accept_attr() -> str:
    return ",".join(sorted(ASSET_DOCUMENT_ALLOWED_EXTENSIONS))


def _validate_workorder_attachment_uploads(
    request: HttpRequest, *, field_name: str = "attachments"
) -> tuple[list, list[str]]:
    uploads = []
    errors: list[str] = []
    for upload in request.FILES.getlist(field_name):
        file_name = Path(getattr(upload, "name", "") or "").name
        try:
            validate_extension_and_mime(
                upload,
                allowed_extensions=ASSET_DOCUMENT_ALLOWED_EXTENSIONS,
                allowed_mimes=ASSET_DOCUMENT_ALLOWED_MIMES,
                max_bytes=ASSET_DOCUMENT_MAX_BYTES,
                label=file_name or "Allegato",
            )
        except UploadMimeValidationError as exc:
            errors.append(str(exc))
            continue
        uploads.append(upload)
    return uploads, errors


def _validate_execution_attachment_uploads(request: HttpRequest) -> tuple[list, list[str]]:
    return _validate_workorder_attachment_uploads(request, field_name="execution_files")


def _save_workorder_attachments(
    *, workorder: WorkOrder, uploads: list, user, checklist_item: "WorkOrderChecklist | None" = None
) -> list[WorkOrderAttachment]:
    created: list[WorkOrderAttachment] = []
    for upload in uploads:
        created.append(
            WorkOrderAttachment.objects.create(
                work_order=workorder,
                checklist_item=checklist_item,
                file=upload,
                original_name=Path(getattr(upload, "name", "") or "").name[:255],
                uploaded_by=user if getattr(user, "is_authenticated", False) else None,
            )
        )
    return created


def _prepopulate_workorder_checklist_from_template(workorder: WorkOrder) -> int:
    """Copia gli step del template intervento nel WorkOrder alla creazione.
    Delega all'helper condiviso in maintenance.py (stesso comportamento usato dalla
    generazione periodica automatica). Ritorna il numero di step creati."""
    return copy_template_checklist_to_workorder(workorder)


def _add_form_validation_errors(form, exc: ValidationError) -> None:
    if hasattr(exc, "message_dict"):
        for field_name, field_errors in exc.message_dict.items():
            target_field = field_name if field_name in form.fields else None
            for error in field_errors:
                form.add_error(target_field, error)
        return
    for error in list(getattr(exc, "messages", []) or [str(exc)]):
        form.add_error(None, error)


def _workorder_priority_order_case() -> Case:
    return Case(
        When(priority=WorkOrder.PRIORITY_URGENT, then=0),
        When(priority=WorkOrder.PRIORITY_NORMAL, then=1),
        When(priority=WorkOrder.PRIORITY_LOW, then=2),
        default=1,
        output_field=IntegerField(),
    )


def _apply_workorder_list_filters(get_params, *, include_filters: bool = True):
    qs = (
        WorkOrder.objects
        .select_related(
            "asset",
            "asset__asset_category",
            "periodic_verification",
            "supplier",
            "maintenance_rule",
            "maintenance_rule__intervention_template",
            "assistance_contract",
            "assigned_to",
            "executed_by",
        )
        .all()
    )
    if not include_filters:
        return qs

    status = _clean_string(get_params.get("status"))
    kind = _clean_string(get_params.get("kind"))
    origin = _clean_string(get_params.get("origin"))
    coverage = _clean_string(get_params.get("coverage")) or "all"
    reparto = _clean_string(get_params.get("reparto"))
    category_id = _as_int(get_params.get("category"), default=0)
    assigned_id = _as_int(get_params.get("assigned"), default=0)
    open_age_days = _as_int(get_params.get("open_age"), default=0)
    asset_id = _as_int(get_params.get("asset"), default=0)
    priority = _clean_string(get_params.get("priority"))
    q = _clean_string(get_params.get("q"))

    if priority in dict(WorkOrder.PRIORITY_CHOICES):
        qs = qs.filter(priority=priority)
    if status:
        qs = qs.filter(status=status)
    if kind:
        qs = qs.filter(kind=kind)
    if origin:
        qs = qs.filter(origin=origin)
    if asset_id:
        qs = qs.filter(asset_id=asset_id)
    if coverage == "covered":
        qs = qs.filter(covered_by_contract=True)
    elif coverage == "uncovered":
        qs = qs.filter(covered_by_contract=False)
    if reparto:
        qs = qs.filter(asset__reparto__iexact=reparto)
    if category_id:
        qs = qs.filter(asset__asset_category_id=category_id)
    if assigned_id:
        qs = qs.filter(Q(assigned_to_id=assigned_id) | Q(executed_by_id=assigned_id))
    if open_age_days:
        qs = qs.filter(
            status=WorkOrder.STATUS_OPEN,
            opened_at__lte=timezone.now() - timedelta(days=max(0, open_age_days)),
        )
    if q:
        qs = qs.filter(
            Q(title__icontains=q)
            | Q(asset__asset_tag__icontains=q)
            | Q(asset__internal_number__icontains=q)
            | Q(asset__name__icontains=q)
            | Q(asset__reparto__icontains=q)
            | Q(asset__asset_category__label__icontains=q)
            | Q(description__icontains=q)
            | Q(resolution__icontains=q)
            | Q(supplier__ragione_sociale__icontains=q)
        )
    return qs


def _resolve_workorder_operational_view(get_params) -> str:
    requested_view = _clean_string(get_params.get("view")).lower()
    if requested_view in {"open", "mine", "unassigned", "closed", "all"}:
        return requested_view
    requested_status = _clean_string(get_params.get("status"))
    if requested_status in {WorkOrder.STATUS_DONE, WorkOrder.STATUS_CANCELED}:
        return "closed"
    return "open"


def _apply_workorder_operational_view(qs, *, operational_view: str, user):
    if operational_view == "mine":
        return qs.filter(status=WorkOrder.STATUS_OPEN, assigned_to=user)
    if operational_view == "unassigned":
        return qs.filter(status=WorkOrder.STATUS_OPEN, assigned_to__isnull=True)
    if operational_view == "closed":
        return qs.filter(status__in=[WorkOrder.STATUS_DONE, WorkOrder.STATUS_CANCELED])
    if operational_view == "all":
        return qs
    return qs.filter(status=WorkOrder.STATUS_OPEN)


def _workorder_list_filter_remove_url(get_params, filter_key: str) -> str:
    query = get_params.copy()
    for key in (filter_key, "page", "create", "export"):
        if key in query:
            query.pop(key)
    query_string = query.urlencode()
    base_url = reverse("assets:wo_list")
    return f"{base_url}?{query_string}" if query_string else base_url


def _workorder_list_page_url(**filters) -> str:
    query = {
        key: str(value)
        for key, value in filters.items()
        if value not in (None, "", 0, "0", False)
    }
    base_url = reverse("assets:wo_list")
    query_string = urlencode(query)
    return f"{base_url}?{query_string}" if query_string else base_url


def _workorder_list_filter_chips(
    request: HttpRequest,
    *,
    status: str,
    kind: str,
    origin: str,
    coverage: str,
    reparto: str,
    category_id: int,
    assigned_id: int,
    open_age_days: int,
    priority: str = "",
    q: str,
    operational_view: str,
    category_options,
    user_options,
    asset_filter=None,
) -> list[dict[str, str]]:
    status_labels = dict(WorkOrder.STATUS_CHOICES)
    kind_labels = dict(WorkOrder.KIND_CHOICES)
    origin_labels = dict(WorkOrder.ORIGIN_CHOICES)
    priority_labels = dict(WorkOrder.PRIORITY_CHOICES)
    coverage_labels = {
        "covered": "Con contratto",
        "uncovered": "Senza contratto",
    }
    category_by_id = {category.id: category for category in category_options}
    user_by_id = {user.id: user for user in user_options}
    chips: list[dict[str, str]] = []

    def add(filter_key: str, label: str, value: str) -> None:
        value = _clean_string(value)
        if not value:
            return
        remove_url = _workorder_list_filter_remove_url(request.GET, filter_key)
        if "view" not in request.GET:
            parsed_remove_url = urlsplit(remove_url)
            remove_query = parse_qs(parsed_remove_url.query)
            remove_query["view"] = [operational_view]
            remove_url = urlunsplit(
                (
                    parsed_remove_url.scheme,
                    parsed_remove_url.netloc,
                    parsed_remove_url.path,
                    urlencode(remove_query, doseq=True),
                    parsed_remove_url.fragment,
                )
            )
        chips.append(
            {
                "label": label,
                "value": value,
                "remove_url": remove_url,
            }
        )

    if q:
        add("q", "Ricerca", q)
    if status:
        add("status", "Stato", status_labels.get(status, status))
    if kind:
        add("kind", "Tipo", kind_labels.get(kind, kind))
    if priority:
        add("priority", "Priorità", priority_labels.get(priority, priority))
    if origin:
        add("origin", "Origine", origin_labels.get(origin, origin))
    if coverage in coverage_labels:
        add("coverage", "Copertura", coverage_labels[coverage])
    if reparto:
        add("reparto", "Reparto", reparto)
    if category_id:
        category = category_by_id.get(category_id)
        add("category", "Categoria", category.label if category else f"ID {category_id}")
    if assigned_id:
        user = user_by_id.get(assigned_id)
        user_label = (user.get_full_name() or user.username) if user else f"ID {assigned_id}"
        add("assigned", "Responsabile", user_label)
    if open_age_days:
        add("open_age", "Aperti da", f"{open_age_days} giorni")
    if asset_filter is not None:
        add("asset", "Asset", asset_filter.asset_tag or asset_filter.name)
    return chips


@login_required
def workorder_list(request: HttpRequest) -> HttpResponse:
    status = _clean_string(request.GET.get("status"))
    kind = _clean_string(request.GET.get("kind"))
    origin = _clean_string(request.GET.get("origin"))
    coverage = _clean_string(request.GET.get("coverage")) or "all"
    reparto = _clean_string(request.GET.get("reparto"))
    category_id = _as_int(request.GET.get("category"), default=0)
    assigned_id = _as_int(request.GET.get("assigned"), default=0)
    open_age_days = _as_int(request.GET.get("open_age"), default=0)
    asset_id = _as_int(request.GET.get("asset"), default=0)
    priority = _clean_string(request.GET.get("priority"))
    q = _clean_string(request.GET.get("q"))
    operational_view = _resolve_workorder_operational_view(request.GET)
    display = _clean_string(request.GET.get("display")).lower()
    display = display if display in ("list", "board") else "list"
    asset_filter = Asset.objects.filter(pk=asset_id).first() if asset_id else None

    workorders = _apply_workorder_operational_view(
        _apply_workorder_list_filters(request.GET),
        operational_view=operational_view,
        user=request.user,
    )
    if operational_view in ("open", "mine", "unassigned"):
        # Coda di lavoro: urgente prima, poi il più vecchio aperto — non l'ultimo creato.
        workorders = workorders.order_by(_workorder_priority_order_case(), "opened_at", "id")
    else:
        workorders = workorders.order_by("-opened_at", "-id")
    count_source = WorkOrder.objects.all()
    operational_counts = {
        "open": count_source.filter(status=WorkOrder.STATUS_OPEN).count(),
        "mine": count_source.filter(status=WorkOrder.STATUS_OPEN, assigned_to=request.user).count(),
        "unassigned": count_source.filter(status=WorkOrder.STATUS_OPEN, assigned_to__isnull=True).count(),
        "closed": count_source.filter(status__in=[WorkOrder.STATUS_DONE, WorkOrder.STATUS_CANCELED]).count(),
    }
    reparto_options = list(
        Asset.objects.exclude(reparto="")
        .order_by("reparto")
        .values_list("reparto", flat=True)
        .distinct()
    )
    category_options = list(AssetCategory.objects.order_by("sort_order", "label", "id"))
    from django.contrib.auth import get_user_model
    User = get_user_model()
    user_options = list(User.objects.filter(is_active=True).order_by("last_name", "first_name", "username"))
    active_filter_chips = _workorder_list_filter_chips(
        request,
        status=status,
        kind=kind,
        origin=origin,
        coverage=coverage,
        reparto=reparto,
        category_id=category_id,
        assigned_id=assigned_id,
        open_age_days=open_age_days,
        priority=priority,
        q=q,
        operational_view=operational_view,
        category_options=category_options,
        user_options=user_options,
        asset_filter=asset_filter,
    )

    board_columns = []
    if display == "board":
        # Kanban degli stati operativi (unassigned/assigned/in_progress/waiting), diversa dalla
        # board per scadenza gia' presente in maintenance_hub.html: qui la colonna e' lo stato
        # reale dell'OdL, non quanto e' vecchio.
        board_by_state = {key: [] for key in WorkOrder.OPSTATE_LABELS}
        for wo in workorders.filter(status=WorkOrder.STATUS_OPEN).select_related("asset", "assigned_to"):
            state = wo.operational_state
            if state in board_by_state:
                board_by_state[state].append(wo)
        board_columns = [
            {"key": key, "label": label, "workorders": board_by_state[key]}
            for key, label in WorkOrder.OPSTATE_LABELS.items()
        ]

    return render(
        request,
        "assets/pages/workorder_list.html",
        {
            "page_title": "Interventi",
            "workorders": workorders,
            "workorder_display": display,
            "board_columns": board_columns,
            "workorder_display_toggle_url": _workorder_list_page_url(
                view=operational_view, display="list" if display == "board" else "board"
            ),
            "status_filter": status,
            "kind_filter": kind,
            "origin_filter": origin,
            "coverage_filter": coverage,
            "reparto_filter": reparto,
            "priority_filter": priority,
            "priority_choices": WorkOrder.PRIORITY_CHOICES,
            "selected_category_id": category_id,
            "assigned_filter": assigned_id,
            "open_age_filter": open_age_days,
            "asset_filter": asset_filter,
            "q_filter": q,
            "workorder_view": operational_view,
            "workorder_view_url": _workorder_list_page_url(view=operational_view),
            "operational_counts": operational_counts,
            "workorder_view_urls": {
                view_key: _workorder_list_page_url(view=view_key)
                for view_key in ("open", "mine", "unassigned", "closed", "all")
            },
            "advanced_filters_open": bool(
                kind
                or origin
                or coverage != "all"
                or reparto
                or category_id
                or assigned_id
                or open_age_days
                or priority
                or status
            ),
            "status_choices": WorkOrder.STATUS_CHOICES,
            "kind_choices": WorkOrder.KIND_CHOICES,
            "origin_choices": WorkOrder.ORIGIN_CHOICES,
            "reparto_options": reparto_options,
            "category_options": category_options,
            "user_options": user_options,
            "active_filter_chips": active_filter_chips,
            "workorder_create_asset_options": Asset.objects.exclude(status=Asset.STATUS_RETIRED).order_by(
                "asset_tag",
                "name",
                "id",
            ),
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def il_mio_turno(request: HttpRequest) -> HttpResponse:
    """Home del manutentore: cosa fare adesso, non l'archivio di chi pianifica.
    Ogni OdL compare in una sola sezione (la prima che gli si applica, nell'ordine
    Bloccati > Emergenze > Scaduti > Oggi > In corso), cosi' non si duplica la coda."""
    user = request.user
    now = timezone.now()
    today = timezone.localdate()
    priority_order = _workorder_priority_order_case()

    mine_open = (
        WorkOrder.objects.select_related("asset", "asset__asset_category", "maintenance_rule__intervention_template")
        .filter(status=WorkOrder.STATUS_OPEN, assigned_to=user)
    )
    unassigned_open = (
        WorkOrder.objects.select_related("asset", "asset__asset_category", "maintenance_rule__intervention_template")
        .filter(status=WorkOrder.STATUS_OPEN, assigned_to__isnull=True)
        .order_by(priority_order, "opened_at", "id")
    )

    bloccati = list(mine_open.filter(is_waiting=True).order_by("waiting_since"))
    bloccati_ids = {wo.id for wo in bloccati}
    emergenze = list(
        mine_open.filter(priority=WorkOrder.PRIORITY_URGENT)
        .exclude(id__in=bloccati_ids)
        .order_by("opened_at", "id")
    )
    emergenze_ids = bloccati_ids | {wo.id for wo in emergenze}
    scaduti = list(
        mine_open.filter(due_at__lt=now)
        .exclude(id__in=emergenze_ids)
        .order_by("due_at", "id")
    )
    scaduti_ids = emergenze_ids | {wo.id for wo in scaduti}
    oggi = list(
        mine_open.filter(due_at__date=today)
        .exclude(id__in=scaduti_ids)
        .order_by("due_at", "id")
    )
    oggi_ids = scaduti_ids | {wo.id for wo in oggi}
    in_corso = list(
        mine_open.filter(started_at__isnull=False)
        .exclude(id__in=oggi_ids)
        .order_by("started_at")
    )
    in_corso_ids = oggi_ids | {wo.id for wo in in_corso}
    altri_assegnati = list(
        mine_open.exclude(id__in=in_corso_ids).order_by(priority_order, "opened_at", "id")
    )
    da_prendere = list(unassigned_open[:20])

    sections = [
        {"key": "bloccati", "title": "Bloccati / in attesa", "rows": bloccati, "empty": "Nessun intervento in attesa."},
        {"key": "emergenze", "title": "Emergenze", "rows": emergenze, "empty": "Nessuna emergenza aperta."},
        {"key": "scaduti", "title": "Scaduti", "rows": scaduti, "empty": "Nessun intervento scaduto."},
        {"key": "oggi", "title": "Oggi", "rows": oggi, "empty": "Nessuna scadenza per oggi."},
        {"key": "in_corso", "title": "In corso", "rows": in_corso, "empty": "Nessun intervento iniziato."},
        {"key": "altri", "title": "Altri assegnati a te", "rows": altri_assegnati, "empty": "Nessun altro intervento assegnato."},
        {"key": "da_prendere", "title": "Da prendere in carico", "rows": da_prendere, "empty": "Niente in coda da prendere in carico."},
    ]
    return render(
        request,
        "assets/pages/il_mio_turno.html",
        {
            "page_title": "Il mio turno",
            "sections": sections,
            "total_mine_open": mine_open.count(),
            **_assets_shell_context(request, rows=25),
            "assets_section_nav": None,
        },
    )


@login_required
def workorder_create(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        selected_asset_id = _as_int(request.GET.get("asset") or request.POST.get("asset"), default=0)
        if selected_asset_id:
            query = request.GET.copy()
            query.pop("asset", None)
            query.setdefault("source", "workorder_list")
            target_url = reverse("assets:wo_create", kwargs={"id": selected_asset_id})
            query_string = query.urlencode()
            return redirect(f"{target_url}?{query_string}" if query_string else target_url)
        # Un OdL richiede sempre un asset: un ingresso globale deve aprire il
        # selettore canonico, non arrestarsi sulla lista interventi.
        return redirect(_workorder_list_page_url(create=1))
    asset = get_object_or_404(Asset, pk=id)
    source = normalize_workorder_source(request.GET.get("source"))
    back_to_list = source == "workorder_list"
    workorder_back_url = reverse("assets:wo_list") if back_to_list else reverse("assets:asset_view", kwargs={"id": id})
    workorder_back_label = "Torna agli interventi" if back_to_list else "Torna al dettaglio asset"
    prefill_payload = build_workorder_prefill_payload(
        asset=asset,
        base_rule_id=_as_int(request.GET.get("rule"), default=0),
        source=source,
    )
    selected_periodic_verification = None
    periodic_verification_id = _as_int(request.GET.get("periodic"), default=0)
    if periodic_verification_id:
        selected_periodic_verification = get_object_or_404(
            PeriodicVerification.objects.select_related("supplier").filter(
                is_active=True,
                assets=asset,
            ),
            pk=periodic_verification_id,
        )
        prefill_payload["periodic_verification"] = selected_periodic_verification
        prefill_payload["title"] = (selected_periodic_verification.name or "Manutenzione periodica").strip()
        prefill_payload["description"] = (selected_periodic_verification.notes or "").strip()
        prefill_payload["kind"] = WorkOrder.KIND_PREVENTIVE
        prefill_payload["supplier"] = selected_periodic_verification.supplier
        suggested_contract = prefill_payload.get("contract")
        if (
            suggested_contract is not None
            and selected_periodic_verification.supplier_id
            and suggested_contract.supplier_id != selected_periodic_verification.supplier_id
        ):
            prefill_payload["contract"] = None
            prefill_payload["covered_by_contract"] = False
    selected_rule_id = getattr(prefill_payload.get("maintenance_rule"), "id", 0)
    form_kwargs = {
        "asset": asset,
        "prefill_payload": prefill_payload,
    }
    if request.method == "POST":
        form = WorkOrderForm(request.POST, **form_kwargs)
        uploads, upload_errors = _validate_workorder_attachment_uploads(request)
        form_is_valid = form.is_valid()
        for error in upload_errors:
            form.add_error(None, error)
        if form_is_valid and not upload_errors:
            workorder = form.save(commit=False)
            workorder.asset = asset
            workorder.status = WorkOrder.STATUS_OPEN
            workorder.origin = (
                WorkOrder.ORIGIN_PERIODIC
                if workorder.periodic_verification_id or workorder.maintenance_rule_id
                else WorkOrder.ORIGIN_MANUAL
            )
            if not workorder.opened_at:
                workorder.opened_at = timezone.now()
            try:
                workorder.full_clean()
            except ValidationError as exc:
                _add_form_validation_errors(form, exc)
            else:
                with transaction.atomic():
                    workorder.save()
                    _prepopulate_workorder_checklist_from_template(workorder)
                    created_attachments = _save_workorder_attachments(
                        workorder=workorder,
                        uploads=uploads,
                        user=request.user,
                    )
                    log_note = "Intervento creato."
                    if created_attachments:
                        log_note = f"{log_note} Allegati caricati: {len(created_attachments)}."
                    WorkOrderLog.objects.create(
                        work_order=workorder,
                        note=log_note,
                        author=request.user if request.user.is_authenticated else None,
                    )
                notify_workorder_assigned(
                    workorder,
                    actor=request.user if request.user.is_authenticated else None,
                )
                messages.success(request, "Intervento creato.")
                if request.POST.get("submit_action") == "close":
                    return redirect("assets:wo_close", id=workorder.id)
                return redirect("assets:wo_view", id=workorder.id)
    else:
        initial = {"status": WorkOrder.STATUS_OPEN}
        kind_param = request.GET.get("kind", "").strip().upper()
        if kind_param in {c[0] for c in WorkOrder.KIND_CHOICES}:
            initial["kind"] = kind_param
        form = WorkOrderForm(initial=initial, **form_kwargs)
    periodic_verification_supplier_map = {
        str(verification.id): {
            "supplier_id": str(verification.supplier_id or ""),
            "supplier_label": str(verification.supplier) if verification.supplier_id else "",
        }
        for verification in form.fields["periodic_verification"].queryset
    }
    maintenance_rule_suggestion_map = form.build_rule_suggestion_map()
    contract_suggestion_map = form.build_contract_suggestion_map()
    initial_rule = None
    rule_value = form["maintenance_rule"].value()
    if rule_value:
        initial_rule = form.fields["maintenance_rule"].queryset.filter(pk=_as_int(rule_value, default=0)).first()
    initial_contract = None
    contract_value = form["assistance_contract"].value()
    if contract_value:
        initial_contract = form.fields["assistance_contract"].queryset.filter(pk=_as_int(contract_value, default=0)).first()
    return render(
        request,
        "assets/pages/workorder_form.html",
        {
            "page_title": f"Nuovo intervento - {asset.asset_tag}",
            "asset": asset,
            "form": form,
            "workorder_back_url": workorder_back_url,
            "workorder_back_label": workorder_back_label,
            "selected_rule_id": selected_rule_id,
            "initial_rule": initial_rule,
            "initial_contract": initial_contract,
            "workorder_prefill": prefill_payload,
            "available_contract_count": form.fields["assistance_contract"].queryset.count(),
            "attachment_accept": _workorder_attachment_accept_attr(),
            "attachment_max_mb": int(ASSET_DOCUMENT_MAX_BYTES / (1024 * 1024)),
            "periodic_verification_supplier_map_json": json.dumps(periodic_verification_supplier_map),
            "maintenance_rule_suggestion_map_json": json.dumps(maintenance_rule_suggestion_map),
            "contract_suggestion_map_json": json.dumps(contract_suggestion_map),
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
            # Il form e' un flusso transazionale: evita tab e azioni globali
            # duplicate mentre l'utente sta compilando l'intervento.
            "assets_section_nav": None,
        },
    )


@login_required
def workorder_detail(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:wo_list")
    workorder = get_object_or_404(
        WorkOrder.objects.select_related(
            "asset",
            "periodic_verification",
            "supplier",
            "maintenance_rule",
            "maintenance_rule__intervention_template",
            "assistance_contract",
            "assistance_contract__supplier",
        ),
        pk=id,
    )

    if request.method == "POST":
        action = _clean_string(request.POST.get("action"))
        author = request.user if request.user.is_authenticated else None

        if action == "set_waiting":
            reason = _clean_string(request.POST.get("wait_reason"))
            note = _clean_string(request.POST.get("wait_note"))
            reason_choices = dict(WorkOrder.WAIT_REASON_CHOICES)
            if workorder.status != WorkOrder.STATUS_OPEN:
                messages.error(request, "Solo un intervento aperto puo essere messo in attesa.")
            elif reason not in reason_choices:
                messages.error(request, "Seleziona un motivo di attesa valido.")
            else:
                workorder.set_waiting(reason=reason, note=note)
                log_text = f"Intervento messo in attesa · motivo: {reason_choices[reason]}"
                if note:
                    log_text = f"{log_text} · {note}"
                WorkOrderLog.objects.create(work_order=workorder, note=log_text, author=author)
                messages.success(request, "Intervento messo in attesa.")
            return redirect("assets:wo_view", id=workorder.id)

        if action == "start":
            if workorder.status != WorkOrder.STATUS_OPEN:
                messages.error(request, "Solo un intervento aperto puo essere iniziato.")
            else:
                was_waiting = workorder.is_waiting
                already_started = bool(workorder.started_at)
                workorder.start()
                if not already_started:
                    log_text = "Intervento iniziato."
                    if was_waiting:
                        log_text = "Intervento ripreso e iniziato."
                    WorkOrderLog.objects.create(work_order=workorder, note=log_text, author=author)
                    messages.success(request, "Intervento iniziato.")
                elif was_waiting:
                    WorkOrderLog.objects.create(
                        work_order=workorder, note="Intervento ripreso dall'attesa.", author=author
                    )
                    messages.success(request, "Intervento ripreso.")
            return redirect("assets:wo_view", id=workorder.id)

        if action == "set_priority":
            priority_choices = dict(WorkOrder.PRIORITY_CHOICES)
            new_priority = _clean_string(request.POST.get("priority"))
            if new_priority not in priority_choices:
                messages.error(request, "Seleziona una priorità valida.")
            elif new_priority != workorder.priority:
                previous_label = workorder.get_priority_display()
                workorder.priority = new_priority
                workorder.save(update_fields=["priority"])
                WorkOrderLog.objects.create(
                    work_order=workorder,
                    note=f"Priorità aggiornata: {previous_label} → {priority_choices[new_priority]}",
                    author=author,
                )
                messages.success(request, "Priorità aggiornata.")
            return redirect("assets:wo_view", id=workorder.id)

        if action == "resume_from_waiting":
            if workorder.is_waiting:
                workorder.resume_from_waiting()
                WorkOrderLog.objects.create(
                    work_order=workorder, note="Intervento ripreso dall'attesa.", author=author
                )
                messages.success(request, "Intervento ripreso.")
            return redirect("assets:wo_view", id=workorder.id)

        if action == "reassign":
            from django.contrib.auth import get_user_model

            User = get_user_model()
            if workorder.status != WorkOrder.STATUS_OPEN:
                messages.error(request, "Solo un intervento aperto puo essere riassegnato.")
                return redirect("assets:wo_view", id=workorder.id)
            new_assignee_id = _clean_string(request.POST.get("new_assignee"))
            reason = _clean_string(request.POST.get("reassign_reason"))
            new_assignee = None
            if new_assignee_id:
                new_assignee = User.objects.filter(pk=new_assignee_id, is_active=True).first()
                if new_assignee is None:
                    messages.error(request, "Seleziona un manutentore valido.")
                    return redirect("assets:wo_view", id=workorder.id)
            previous_assignee = workorder.assigned_to
            if new_assignee_id == "" and previous_assignee is None:
                messages.info(request, "Nessuna modifica: assegnatario invariato.")
                return redirect("assets:wo_view", id=workorder.id)
            if new_assignee is not None and previous_assignee is not None and new_assignee.pk == previous_assignee.pk:
                messages.info(request, "Nessuna modifica: assegnatario invariato.")
                return redirect("assets:wo_view", id=workorder.id)
            workorder.assigned_to = new_assignee
            workorder.save(update_fields=["assigned_to"])
            previous_label = (
                previous_assignee.get_full_name() or previous_assignee.username if previous_assignee else "nessuno"
            )
            new_label = new_assignee.get_full_name() or new_assignee.username if new_assignee else "nessuno"
            actor_label = author.get_full_name() or author.username if author else "—"
            log_text = f"Riassegnato da {previous_label} a {new_label} · da {actor_label}"
            if reason:
                log_text = f"{log_text} · motivo: {reason}"
            WorkOrderLog.objects.create(work_order=workorder, note=log_text, author=author)
            notify_workorder_assigned(workorder, actor=request.user)
            notify_workorder_reassigned(workorder, previous_assignee=previous_assignee, actor=request.user)
            messages.success(request, "Intervento riassegnato.")
            return redirect("assets:wo_view", id=workorder.id)

        log_note = _clean_string(request.POST.get("log_note"))
        if log_note:
            WorkOrderLog.objects.create(
                work_order=workorder,
                note=log_note,
                author=author,
            )
            messages.success(request, "Nota aggiunta.")
            return redirect("assets:wo_view", id=workorder.id)

    from django.contrib.auth import get_user_model

    User = get_user_model()
    assignable_users = User.objects.filter(is_active=True).order_by("last_name", "first_name", "username")
    logs = workorder.logs.select_related("author").all()
    attachments = workorder.attachments.all()
    checklist_items = list(
        workorder.checklist_items.select_related("done_by", "skipped_by").prefetch_related("photos").order_by("step_number", "id")
    )
    checklist_done_count = sum(1 for it in checklist_items if it.is_complete or it.is_skipped)
    execution_days = list(workorder.execution_days.order_by("execution_date"))
    duration_hours, duration_remainder = divmod(max(0, int(workorder.intervention_duration_minutes or 0)), 60)
    downtime_hours, downtime_remainder = divmod(max(0, int(workorder.downtime_minutes or 0)), 60)
    return render(
        request,
        "assets/pages/workorder_detail.html",
        {
            "page_title": f"Intervento #{workorder.id}",
            "workorder": workorder,
            "logs": logs,
            "attachments": attachments,
            "checklist_items": checklist_items,
            "checklist_done_count": checklist_done_count,
            "checklist_total": len(checklist_items),
            "execution_days": execution_days,
            "duration_hours": duration_hours,
            "duration_remainder": duration_remainder,
            "downtime_hours": downtime_hours,
            "downtime_remainder": downtime_remainder,
            "is_open": workorder.status == WorkOrder.STATUS_OPEN,
            "workorder_asset_url": reverse("assets:asset_view", kwargs={"id": workorder.asset_id}),
            "workorder_rule_url": (
                _asset_maintenance_rule_list_page_url(
                    asset_id=workorder.asset_id,
                    focus_rule_id=workorder.maintenance_rule_id or 0,
                )
                if workorder.maintenance_rule_id
                else ""
            ),
            "workorder_contract_url": (
                _assistance_contracts_page_url(
                    asset_id=workorder.asset_id,
                    edit_id=workorder.assistance_contract_id or 0,
                )
                if workorder.assistance_contract_id
                else _assistance_contracts_page_url(asset_id=workorder.asset_id)
            ),
            "workorder_contract_label": "Apri contratto" if workorder.assistance_contract_id else "Apri contratti",
            "workorder_status": _workorder_status_payload(workorder.status),
            "workorder_kind": _workorder_kind_payload(workorder.kind),
            "workorder_coverage": _coverage_status_payload(
                is_covered=bool(workorder.covered_by_contract),
                contract=workorder.assistance_contract,
            ),
            "wait_reason_choices": WorkOrder.WAIT_REASON_CHOICES,
            "priority_choices": WorkOrder.PRIORITY_CHOICES,
            "operational_state": workorder.operational_state,
            "operational_state_label": workorder.operational_state_label,
            "is_overdue": workorder.is_overdue,
            "assignable_users": assignable_users,
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


def _render_checklist_fragment(request: HttpRequest, workorder: WorkOrder) -> HttpResponse:
    items = list(
        workorder.checklist_items.select_related("done_by", "skipped_by").prefetch_related("photos").order_by("step_number", "id")
    )
    done_count = sum(1 for it in items if it.is_complete or it.is_skipped)
    return render(
        request,
        "assets/components/workorder_checklist.html",
        {
            "workorder": workorder,
            "checklist_items": items,
            "checklist_done_count": done_count,
            "checklist_total": len(items),
            "is_open": workorder.status == WorkOrder.STATUS_OPEN,
        },
    )


@login_required
def workorder_checklist_add(request: HttpRequest, id: int) -> HttpResponse:
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    description = _clean_string(request.POST.get("description"))
    if not description:
        return _render_checklist_fragment(request, workorder)
    max_step = (
        WorkOrderChecklist.objects.filter(work_order=workorder)
        .order_by("-step_number")
        .values_list("step_number", flat=True)
        .first()
    ) or 0
    WorkOrderChecklist.objects.create(
        work_order=workorder,
        step_number=max_step + 1,
        description=description,
    )
    return _render_checklist_fragment(request, workorder)


@login_required
def workorder_checklist_toggle(request: HttpRequest, id: int, item_id: int) -> HttpResponse:
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    item = get_object_or_404(WorkOrderChecklist, pk=item_id, work_order=workorder)
    item.toggle(request.user)
    return _render_checklist_fragment(request, workorder)


@login_required
def workorder_checklist_delete(request: HttpRequest, id: int, item_id: int) -> HttpResponse:
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    WorkOrderChecklist.objects.filter(pk=item_id, work_order=workorder).delete()
    return _render_checklist_fragment(request, workorder)


@login_required
def workorder_checklist_set_value(request: HttpRequest, id: int, item_id: int) -> HttpResponse:
    """Valorizza uno step Misura o Testo (gli step Sì/No usano il toggle, le Foto l'upload)."""
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    item = get_object_or_404(WorkOrderChecklist, pk=item_id, work_order=workorder)
    if item.step_type == WorkOrderChecklist.TYPE_MEASURE:
        raw_value = _clean_string(request.POST.get("value_numeric"))
        try:
            value_numeric = Decimal(raw_value.replace(",", ".")) if raw_value else None
        except (InvalidOperation, ValueError):
            value_numeric = None
        item.set_value(value_numeric=value_numeric, user=request.user)
    elif item.step_type == WorkOrderChecklist.TYPE_TEXT:
        item.set_value(value_text=request.POST.get("value_text", ""), user=request.user)
    return _render_checklist_fragment(request, workorder)


@login_required
def workorder_checklist_photo(request: HttpRequest, id: int, item_id: int) -> HttpResponse:
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    item = get_object_or_404(WorkOrderChecklist, pk=item_id, work_order=workorder, step_type=WorkOrderChecklist.TYPE_PHOTO)
    uploads, upload_errors = _validate_workorder_attachment_uploads(request, field_name="photo")
    if uploads:
        _save_workorder_attachments(workorder=workorder, uploads=uploads[:1], user=request.user, checklist_item=item)
    elif upload_errors:
        messages.error(request, upload_errors[0])
    return redirect(f"{reverse('assets:wo_view', args=[workorder.id])}#wod-checklist-section")


@login_required
def workorder_checklist_skip(request: HttpRequest, id: int, item_id: int) -> HttpResponse:
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return HttpResponseForbidden("OdL non aperto.")
    if request.method != "POST":
        return HttpResponseForbidden("Metodo non consentito.")
    item = get_object_or_404(WorkOrderChecklist, pk=item_id, work_order=workorder)
    if item.is_skipped:
        item.unskip()
    else:
        reason = _clean_string(request.POST.get("skip_reason"))
        if reason:
            item.skip(reason=reason, user=request.user)
    return _render_checklist_fragment(request, workorder)


@login_required
def workorder_claim(request: HttpRequest, id: int) -> HttpResponse:
    """Assegna l'intervento aperto all'utente corrente ("prendi in carico").

    Azione leggera richiamabile dal Centro Manutenzione senza aprire il dettaglio.
    Non modifica lo stato dell'OdL: imposta solo ``assigned_to`` e logga la presa
    in carico. Ritorna al referer (hub) o al dettaglio.
    """
    from django.utils.http import url_has_allowed_host_and_scheme

    workorder = get_object_or_404(WorkOrder, pk=id)
    raw_next = request.POST.get("next") or ""
    if raw_next and url_has_allowed_host_and_scheme(
        raw_next, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = raw_next
    else:
        next_url = reverse("assets:wo_view", args=[workorder.id])
    if request.method != "POST":
        return redirect(next_url)
    if workorder.status != WorkOrder.STATUS_OPEN:
        messages.error(request, "L'intervento non è aperto: impossibile prenderlo in carico.")
        return redirect(next_url)
    if workorder.assigned_to_id == request.user.id:
        messages.info(request, f"Intervento #{workorder.id} già assegnato a te.")
        return redirect(next_url)

    previous_assignee = workorder.assigned_to
    workorder.assigned_to = request.user
    workorder.save(update_fields=["assigned_to"])
    actor = request.user.get_full_name() or request.user.username
    WorkOrderLog.objects.create(
        work_order=workorder,
        note=f"Preso in carico da {actor}.",
        author=request.user if request.user.is_authenticated else None,
    )
    notify_workorder_taken_over(workorder, previous_assignee=previous_assignee, actor=request.user)
    messages.success(request, f"Intervento #{workorder.id} preso in carico.")
    return redirect(next_url)


@login_required
def workorder_set_state(request: HttpRequest, id: int) -> JsonResponse:
    """Endpoint della board Kanban OdL (drag&drop): sposta un intervento aperto tra gli stati
    operativi derivati (assegna/avvia/metti in attesa). Risposta sempre JSON, mai redirect —
    e' l'unico consumatore, via fetch(), stesso pattern di tasks/_board.html."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Metodo non consentito."}, status=405)
    workorder = get_object_or_404(WorkOrder, pk=id)
    if workorder.status != WorkOrder.STATUS_OPEN:
        return JsonResponse({"ok": False, "error": "Intervento non aperto."}, status=409)

    target_state = _clean_string(request.POST.get("state"))
    actor = request.user.get_full_name() or request.user.username

    if target_state == WorkOrder.OPSTATE_ASSIGNED:
        previous_assignee = workorder.assigned_to
        workorder.assign(request.user)
        WorkOrderLog.objects.create(
            work_order=workorder,
            note=f"Preso in carico da {actor} (board).",
            author=request.user,
        )
        notify_workorder_taken_over(workorder, previous_assignee=previous_assignee, actor=request.user)
    elif target_state == WorkOrder.OPSTATE_IN_PROGRESS:
        workorder.start()
        WorkOrderLog.objects.create(
            work_order=workorder,
            note=f"Intervento avviato da {actor} (board).",
            author=request.user,
        )
    elif target_state == WorkOrder.OPSTATE_WAITING:
        workorder.set_waiting(reason=WorkOrder.WAIT_REASON_ALTRO, note="")
        WorkOrderLog.objects.create(
            work_order=workorder,
            note=f"Messo in attesa da {actor} (board). Motivo da precisare in scheda.",
            author=request.user,
        )
    else:
        return JsonResponse({"ok": False, "error": "Stato non valido."}, status=400)

    return JsonResponse({"ok": True, "operational_state": workorder.operational_state})


@login_required
def workorder_close(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:wo_list")
    workorder = get_object_or_404(
        WorkOrder.objects.select_related("asset", "maintenance_rule", "assistance_contract", "assistance_contract__supplier"),
        pk=id,
    )

    if request.method == "POST":
        form = WorkOrderCloseForm(request.POST, asset=workorder.asset, workorder=workorder)
        uploads, upload_errors = _validate_workorder_attachment_uploads(request, field_name="close_attachments")
        form_is_valid = form.is_valid()
        for error in upload_errors:
            form.add_error(None, error)
        if form_is_valid and form.cleaned_data["status"] == WorkOrder.STATUS_DONE:
            blocking_items = [item for item in workorder.checklist_items.all() if item.blocks_closure]
            if blocking_items:
                labels = ", ".join(f"#{item.step_number} {item.description}" for item in blocking_items)
                form.add_error(
                    None,
                    f"Completa o salta con motivazione gli step obbligatori della checklist prima di chiudere: {labels}.",
                )
                form_is_valid = False
        if form_is_valid and not upload_errors:
            resolved_supplier = form.cleaned_data.get("resolved_supplier")
            if resolved_supplier is not None and workorder.supplier_id is None:
                workorder.supplier = resolved_supplier
            assigned_to = form.cleaned_data.get("assigned_to")
            executed_by = form.cleaned_data.get("executed_by")
            previous_assignee_id = workorder.assigned_to_id
            previous_assignee_label = (
                workorder.assigned_to.get_full_name() or workorder.assigned_to.username
                if workorder.assigned_to_id
                else "nessuno"
            )
            if assigned_to is not None:
                workorder.assigned_to = assigned_to
            if executed_by is not None:
                workorder.executed_by = executed_by
            assignee_changed = assigned_to is not None and assigned_to.pk != previous_assignee_id
            requested_status = form.cleaned_data["status"]
            esito = form.cleaned_data.get("esito")
            # "Non risolto" non deve mai risultare chiuso, nemmeno per un istante: se il
            # WO fosse chiuso prima di riaprirlo, close()/sync_workorder_maintenance_state
            # avanzerebbero la prossima scadenza su una manutenzione che non ha risolto nulla.
            reopened_as_unresolved = (
                requested_status == WorkOrder.STATUS_DONE and esito == WorkOrder.OUTCOME_NOT_RESOLVED
            )
            effective_status = WorkOrder.STATUS_OPEN if reopened_as_unresolved else requested_status
            try:
                workorder.close(
                    status=effective_status,
                    closed_at=form.cleaned_data.get("closed_at"),
                    resolution=form.cleaned_data.get("resolution") or "",
                    intervention_duration=form.cleaned_data.get("intervention_duration_minutes"),
                    downtime=form.cleaned_data.get("downtime_minutes"),
                    labor_cost=form.cleaned_data.get("labor_cost_eur"),
                    materials_cost=form.cleaned_data.get("materials_cost_eur"),
                    cost=form.cleaned_data.get("cost_eur"),
                    covered_by_contract=form.cleaned_data.get("covered_by_contract"),
                    assistance_contract=form.cleaned_data.get("assistance_contract"),
                )
            except ValidationError as exc:
                _add_form_validation_errors(form, exc)
            else:
                failure_cause = form.cleaned_data.get("failure_cause") or ""
                if failure_cause != workorder.failure_cause:
                    workorder.failure_cause = failure_cause
                    workorder.save(update_fields=["failure_cause"])
                execution_days = form.cleaned_data.get("execution_days_list") or []
                workorder.execution_days.all().delete()
                WorkOrderExecutionDay.objects.bulk_create(
                    [
                        WorkOrderExecutionDay(work_order=workorder, execution_date=execution_day)
                        for execution_day in execution_days
                    ]
                )
                if workorder.status == WorkOrder.STATUS_DONE and workorder.maintenance_rule_id:
                    sync_workorder_maintenance_state(workorder)

                author = request.user if request.user.is_authenticated else None
                follow_up_child = None
                if reopened_as_unresolved:
                    workorder.outcome = esito
                    workorder.save(update_fields=["outcome"])
                elif workorder.status == WorkOrder.STATUS_DONE and esito:
                    outcome_fields = ["outcome"]
                    workorder.outcome = esito
                    if esito == WorkOrder.OUTCOME_RESOLVED_TEMP:
                        workorder.follow_up_date = form.cleaned_data.get("follow_up_date")
                        outcome_fields.append("follow_up_date")
                    workorder.save(update_fields=outcome_fields)
                    if esito == WorkOrder.OUTCOME_RESOLVED_TEMP:
                        follow_up_child = WorkOrder.objects.create(
                            asset=workorder.asset,
                            supplier=workorder.supplier,
                            maintenance_rule=workorder.maintenance_rule,
                            kind=workorder.kind,
                            origin=WorkOrder.ORIGIN_MANUAL,
                            assigned_to=workorder.assigned_to,
                            title=f"Follow-up: {workorder.title}",
                            description=(
                                f"Verifica di follow-up dell'intervento #{workorder.id}, "
                                f"risolto temporaneamente il {timezone.localtime(workorder.closed_at):%d/%m/%Y}."
                            ),
                            follow_up_of=workorder,
                        )
                        WorkOrderLog.objects.create(
                            work_order=follow_up_child,
                            note=f"Creato come follow-up dell'intervento #{workorder.id}.",
                            author=author,
                        )

                log_note = _clean_string(form.cleaned_data.get("log_note"))
                if workorder.status == WorkOrder.STATUS_DONE:
                    closure_note = f"Intervento chiuso · esito: {workorder.get_outcome_display()}."
                elif workorder.status == WorkOrder.STATUS_CANCELED:
                    closure_note = "Intervento annullato."
                else:
                    closure_note = f"Intervento non risolto, resta aperto · esito: {workorder.get_outcome_display()}."
                if follow_up_child is not None:
                    closure_note = (
                        f"{closure_note} Follow-up #{follow_up_child.id} creato "
                        f"(verifica entro il {workorder.follow_up_date:%d/%m/%Y})."
                    )
                if assignee_changed:
                    new_assignee_label = (
                        workorder.assigned_to.get_full_name() or workorder.assigned_to.username
                        if workorder.assigned_to_id
                        else "nessuno"
                    )
                    closure_note = (
                        f"{closure_note} Assegnatario aggiornato: {previous_assignee_label} → {new_assignee_label}."
                    )
                if execution_days:
                    closure_note = f"{closure_note} Giorni esecuzione: {', '.join(day.strftime('%d/%m/%Y') for day in execution_days)}."
                if log_note:
                    closure_note = f"{closure_note} {log_note}"
                if uploads:
                    created_attachments = _save_workorder_attachments(
                        workorder=workorder,
                        uploads=uploads,
                        user=request.user,
                    )
                    closure_note = f"{closure_note} Allegati caricati: {len(created_attachments)}."
                WorkOrderLog.objects.create(
                    work_order=workorder,
                    note=closure_note,
                    author=author,
                )
                if workorder.status == WorkOrder.STATUS_DONE:
                    success_message = "Intervento chiuso."
                elif workorder.status == WorkOrder.STATUS_CANCELED:
                    success_message = "Intervento annullato."
                else:
                    success_message = "Intervento registrato come non risolto: resta aperto."
                if follow_up_child is not None:
                    success_message = f"{success_message} Creato il follow-up #{follow_up_child.id}."
                messages.success(request, success_message)
                return redirect("assets:wo_view", id=workorder.id)
    else:
        form = WorkOrderCloseForm(
            initial={
                "status": WorkOrder.STATUS_DONE,
                "intervention_duration_minutes": workorder.intervention_duration_minutes,
                "downtime_minutes": workorder.downtime_minutes,
                "labor_cost_eur": workorder.labor_cost_eur,
                "materials_cost_eur": workorder.materials_cost_eur,
                "cost_eur": workorder.cost_eur,
                "assistance_contract": workorder.assistance_contract_id,
                "covered_by_contract": workorder.covered_by_contract,
                "assigned_to": workorder.assigned_to_id,
                # Chi chiude l'intervento è quasi sempre chi l'ha eseguito: precompilare
                # evita al manutentore di doversi ricercare in un elenco ad ogni chiusura.
                "executed_by": workorder.executed_by_id or (
                    request.user.id if request.user.is_authenticated else None
                ),
                "failure_cause": workorder.failure_cause,
            },
            asset=workorder.asset,
            workorder=workorder,
        )

    return render(
        request,
        "assets/pages/workorder_close.html",
        {
            "page_title": f"Chiudi intervento #{workorder.id}",
            "workorder": workorder,
            "form": form,
            "close_submit_label": (
                "Conferma annullamento"
                if (form["status"].value() or WorkOrder.STATUS_DONE) == WorkOrder.STATUS_CANCELED
                else "Conferma chiusura"
            ),
            "workorder_status": _workorder_status_payload(workorder.status),
            "workorder_kind": _workorder_kind_payload(workorder.kind),
            "workorder_coverage": _coverage_status_payload(
                is_covered=bool(workorder.covered_by_contract),
                contract=workorder.assistance_contract,
            ),
            "attachment_accept": _workorder_attachment_accept_attr(),
            "attachment_max_mb": int(ASSET_DOCUMENT_MAX_BYTES / (1024 * 1024)),
            "checklist_blocking_items": [item for item in workorder.checklist_items.all() if item.blocks_closure],
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
            "assets_section_nav": None,
        },
    )


@login_required
def asset_meter_update(request: HttpRequest, asset_id: int) -> HttpResponse:
    """P2.2 — Aggiornamento rapido contatore (ore/km/cicli) di un asset.

    GET: restituisce il partial HTML con i contatori correnti e il form inline.
    POST: aggiorna il valore del contatore selezionato e restituisce lo stesso partial.
    Usato via HTMX (hx-post + hx-swap='outerHTML') dalla asset_detail e dalla work_machine_dashboard.
    """
    asset = get_object_or_404(Asset, pk=asset_id)
    meters = list(AssetMeter.objects.filter(asset=asset).order_by("meter_type"))
    error = None
    success_message = None

    if request.method == "POST":
        meter_id = _as_int(request.POST.get("meter_id"), default=0)
        new_value_raw = request.POST.get("new_value", "").strip()
        try:
            new_value = Decimal(new_value_raw.replace(",", "."))
            if new_value < 0:
                raise ValueError("Il valore non può essere negativo.")
        except Exception:
            error = "Valore non valido. Inserire un numero positivo."
            new_value = None

        if new_value is not None and meter_id:
            try:
                meter = AssetMeter.objects.get(pk=meter_id, asset=asset)
                history = meter.update_value(new_value, request.user)
                # La firma è log_action(request, azione, modulo, dettaglio): passare
                # request.user come request faceva fallire l'audit in silenzio (fire-and-forget)
                # proprio sul dato che pilota la generazione degli OdL a soglia.
                log_action(
                    request,
                    "asset_meter_update",
                    "assets",
                    {
                        "asset_id": asset.id,
                        "asset_tag": asset.asset_tag,
                        "meter_id": meter.id,
                        "meter_type": meter.meter_type,
                        "old_value": str(history.old_value),
                        "new_value": str(history.new_value),
                    },
                )
                meters = list(AssetMeter.objects.filter(asset=asset).order_by("meter_type"))
                success_message = f"Contatore aggiornato a {new_value}."
            except AssetMeter.DoesNotExist:
                error = "Contatore non trovato."

    recent_history = (
        AssetMeterHistory.objects
        .filter(meter__asset=asset)
        .select_related("meter", "recorded_by")
        .order_by("-recorded_at")[:10]
    )

    ctx = {
        "asset": asset,
        "meters": meters,
        "recent_history": recent_history,
        "error": error,
        "success_message": success_message,
    }
    return render(request, "assets/components/asset_meter_panel.html", ctx)


@login_required
def asset_quick_report(request: HttpRequest) -> HttpResponse:
    """P3.2 — Form segnalazione rapida per operatori non-admin.
    Crea un Ticket MAN precompilato con asset, descrizione e priorità.
    """
    from tickets.models import TipoTicket, StatoTicket, PrioritaTicket, Ticket, get_categorie

    # Pre-selezione asset da querystring (es. da landing QR)
    preselected_asset_id = _as_int(request.GET.get("asset"), default=0)
    preselected_asset = None
    if preselected_asset_id:
        preselected_asset = Asset.objects.filter(pk=preselected_asset_id).only("id", "asset_tag", "name", "reparto", "asset_type").first()

    # Asset selezionabili: macchine e asset di produzione (no IT puro)
    asset_choices = list(
        Asset.objects
        .filter(status=Asset.STATUS_IN_USE)
        .exclude(asset_type__in=IT_DEVICE_TYPES)
        .order_by("reparto", "name", "asset_tag")
        .values("id", "asset_tag", "name", "reparto")
    )

    categorie_man = get_categorie(TipoTicket.MAN)
    priorita_choices = list(PrioritaTicket.choices)

    error: str = ""
    success_ticket_id: int | None = None

    if request.method == "POST":
        asset_id = _as_int(request.POST.get("asset_id"), default=0)
        asset_libera = _clean_string(request.POST.get("asset_descrizione_libera"))
        categoria = _clean_string(request.POST.get("categoria"))
        titolo = _clean_string(request.POST.get("titolo"))
        descrizione = _clean_string(request.POST.get("descrizione"))
        priorita = _clean_string(request.POST.get("priorita")) or PrioritaTicket.MEDIA
        incide_sicurezza = request.POST.get("incide_sicurezza") == "1"

        if not titolo or not descrizione:
            error = "Titolo e descrizione sono obbligatori."
        elif not categoria:
            error = "Seleziona una categoria."
        elif not asset_id and not asset_libera:
            error = "Seleziona un asset o descrivi il punto di intervento."
        else:
            if priorita not in dict(priorita_choices):
                priorita = PrioritaTicket.MEDIA

            asset_obj = None
            if asset_id:
                asset_obj = Asset.objects.filter(pk=asset_id).only("id").first()

            # Identità richiedente dal legacy user o dall'utente Django
            from anagrafica.models import UserExtraInfo
            legacy_user = getattr(request, "legacy_user", None)
            if legacy_user is None:
                try:
                    extra = UserExtraInfo.objects.filter(user=request.user).select_related("user").first()
                    legacy_user = extra
                except Exception:
                    legacy_user = None

            if legacy_user and hasattr(legacy_user, "nome"):
                req_name = (_clean_string(getattr(legacy_user, "nome", "")) or request.user.get_full_name() or request.user.get_username())
                req_email = (_clean_string(getattr(legacy_user, "email", "")) or request.user.email or "").lower()
                req_legacy_id = getattr(legacy_user, "id", None)
            else:
                req_name = request.user.get_full_name() or request.user.get_username()
                req_email = (request.user.email or "").lower()
                req_legacy_id = None

            ticket = Ticket(
                tipo=TipoTicket.MAN,
                titolo=titolo,
                descrizione=descrizione,
                categoria=categoria,
                priorita=priorita,
                incide_sicurezza=incide_sicurezza,
                asset=asset_obj,
                asset_descrizione_libera=asset_libera if not asset_obj else "",
                richiedente_nome=req_name,
                richiedente_email=req_email,
                richiedente_legacy_user_id=req_legacy_id,
                include_in_maintenance_register=True,
            )
            try:
                ticket.save()
                log_action(
                    request,
                    "asset_quick_report_create",
                    "assets",
                    {
                        "ticket_id": ticket.id,
                        "ticket_numero": ticket.numero_ticket,
                        "asset_id": asset_id,
                        "categoria": categoria,
                        "priorita": priorita,
                    },
                )
                success_ticket_id = ticket.id
                # Reset form su successo
                preselected_asset = None
                asset_id = 0
            except Exception as exc:
                error = f"Errore nella creazione del ticket: {exc}"

    return render(
        request,
        "assets/pages/asset_quick_report.html",
        {
            "page_title": "Segnala un problema",
            "asset_choices": asset_choices,
            "preselected_asset": preselected_asset,
            "categorie_man": categorie_man,
            "priorita_choices": priorita_choices,
            "error": error,
            "success_ticket_id": success_ticket_id,
            **_assets_shell_context(request, rows=25),
        },
    )



def _due_state(due_date, today):
    """Sorgente unica per lo stato 'giorni residui → danger/warn/ok/muted'.

    Restituisce (state, days_left, days_abs). Usato dalle tabelle dello
    scadenzario (verifiche / scadenze amm. / contratti), che prima
    ricalcolavano questo stesso blocco a mano tre volte.
    """
    if due_date is None:
        return "muted", None, None
    days_left = (due_date - today).days
    if days_left < 0:
        state = "danger"
    elif days_left <= 30:
        state = "warn"
    else:
        state = "ok"
    return state, days_left, abs(days_left)


@login_required
def maintenance_history(request: HttpRequest) -> HttpResponse:
    """Storico aziendale unico: OdL conclusi e ticket MAN, senza alterarne i flussi."""
    from tickets.models import StatoTicket, Ticket, TipoTicket

    q = _clean_string(request.GET.get("q"))
    source = _clean_string(request.GET.get("source")).lower() or "all"
    if source not in {"all", "workorders", "tickets"}:
        source = "all"

    def parsed_date(name: str) -> date | None:
        raw = _clean_string(request.GET.get(name))
        try:
            return date.fromisoformat(raw) if raw else None
        except ValueError:
            return None

    date_from = parsed_date("date_from")
    date_to = parsed_date("date_to")
    rows: list[dict[str, object]] = []

    if source in {"all", "workorders"}:
        workorders = WorkOrder.objects.filter(
            status__in=[WorkOrder.STATUS_DONE, WorkOrder.STATUS_CANCELED]
        ).select_related(
            "asset", "executed_by", "assigned_to", "supplier", "maintenance_rule__intervention_template"
        )
        if q:
            workorders = workorders.filter(
                Q(title__icontains=q)
                | Q(description__icontains=q)
                | Q(resolution__icontains=q)
                | Q(asset__asset_tag__icontains=q)
                | Q(asset__name__icontains=q)
                | Q(maintenance_rule__intervention_template__label__icontains=q)
            )
        if date_from:
            workorders = workorders.filter(closed_at__date__gte=date_from)
        if date_to:
            workorders = workorders.filter(closed_at__date__lte=date_to)
        for workorder in workorders.order_by("-closed_at", "-id")[:500]:
            technician = workorder.executed_by or workorder.assigned_to
            rows.append(
                {
                    "source": "workorder",
                    "source_label": "OdL",
                    "date": workorder.closed_at or workorder.opened_at,
                    "title": workorder.title,
                    "asset": workorder.asset,
                    "type_label": workorder.get_kind_display(),
                    "status_label": workorder.get_status_display(),
                    "technician": technician.get_full_name() or technician.username if technician else "-",
                    "duration_minutes": workorder.intervention_duration_minutes,
                    "cost": workorder.resolved_total_cost_eur,
                    "url": reverse("assets:wo_view", kwargs={"id": workorder.id}),
                }
            )

    if source in {"all", "tickets"}:
        tickets = Ticket.objects.filter(
            tipo=TipoTicket.MAN,
            include_in_maintenance_register=True,
            stato__in=[StatoTicket.RISOLTO, StatoTicket.CHIUSO, StatoTicket.ANNULLATO],
        ).select_related("asset")
        if q:
            tickets = tickets.filter(
                Q(titolo__icontains=q)
                | Q(descrizione__icontains=q)
                | Q(numero_ticket__icontains=q)
                | Q(asset__asset_tag__icontains=q)
                | Q(asset__name__icontains=q)
            )
        if date_from:
            tickets = tickets.filter(closed_at__date__gte=date_from)
        if date_to:
            tickets = tickets.filter(closed_at__date__lte=date_to)
        for ticket in tickets.order_by("-closed_at", "-id")[:500]:
            rows.append(
                {
                    "source": "ticket",
                    "source_label": "Ticket MAN",
                    "date": ticket.closed_at or ticket.created_at,
                    "title": ticket.titolo,
                    "asset": ticket.asset,
                    "type_label": ticket.label_categoria,
                    "status_label": ticket.label_stato,
                    "technician": ticket.risolto_da_nome or ticket.assegnato_a or "-",
                    "duration_minutes": 0,
                    "cost": None,
                    "url": reverse("tickets:detail", kwargs={"pk": ticket.id}),
                }
            )

    rows.sort(key=lambda row: row["date"] or timezone.now(), reverse=True)
    total_duration = sum(int(row["duration_minutes"] or 0) for row in rows)
    total_cost = sum((row["cost"] or Decimal("0")) for row in rows)

    return render(
        request,
        "assets/pages/maintenance_history.html",
        {
            **_assets_shell_context(request),
            "page_title": "Storico manutenzione aziendale",
            "history_rows": rows[:500],
            "history_total": len(rows),
            "history_workorder_count": sum(1 for row in rows if row["source"] == "workorder"),
            "history_ticket_count": sum(1 for row in rows if row["source"] == "ticket"),
            "history_duration_hours": round(total_duration / 60, 1),
            "history_total_cost": total_cost,
            "q": q,
            "source": source,
            "date_from": date_from,
            "date_to": date_to,
        },
    )


@login_required
def maintenance_hub(request: HttpRequest) -> HttpResponse:
    """Centro operativo per priorita, OdL e scadenze di manutenzione.

    Lo scadenzario canonico vive in ``maintenance_schedule``; i vecchi
    deep-link ``?tab=scadenzario`` vi confluiscono senza duplicarne la UI.
    """
    from datetime import timedelta
    from .models import AssistanceContract, WorkMachine

    today = timezone.localdate()
    horizon_7  = today + timedelta(days=7)
    horizon_14 = today + timedelta(days=14)
    horizon_30 = today + timedelta(days=30)
    wo_overdue_days = get_workorder_overdue_days()
    overdue_threshold = today - timedelta(days=wo_overdue_days)
    is_admin = _is_assets_admin(request)

    # Il Centro Manutenzione è ora solo cockpit "Da fare": lo scadenzario unico
    # (regole + verifiche + amministrative) vive in /prossime/ (maintenance_schedule).
    # Le vecchie URL/bookmark ?tab=scadenzario (e i deep-link sub=...) confluiscono lì,
    # preservando il filtro reparto, per non avere due scadenzari.
    reparto_filter = _clean_string(request.GET.get("reparto"))
    if _clean_string(request.GET.get("tab")) == "scadenzario":
        return redirect(_maintenance_schedule_page_url(reparto=reparto_filter))
    active_tab = "da_fare"

    # Filtri condivisi (tab "da fare")
    assigned_filter = _clean_string(request.GET.get("assigned"))

    # ── OdL aperti ─────────────────────────────────────────────────────────
    wo_qs = (
        WorkOrder.objects
        .filter(status=WorkOrder.STATUS_OPEN)
        .select_related("asset", "assigned_to", "executed_by", "maintenance_rule__intervention_template")
        .order_by("opened_at")
    )
    if not is_admin:
        wo_qs = wo_qs.filter(Q(assigned_to=request.user) | Q(executed_by=request.user))
    elif assigned_filter:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            assigned_user = User.objects.get(pk=int(assigned_filter))
            wo_qs = wo_qs.filter(Q(assigned_to=assigned_user) | Q(executed_by=assigned_user))
        except (ValueError, User.DoesNotExist):
            assigned_filter = ""
    if reparto_filter:
        wo_qs = wo_qs.filter(asset__reparto=reparto_filter)
    open_workorders = list(wo_qs[:50])
    for wo in open_workorders:
        wo.assigned_to_me = wo.assigned_to_id == request.user.id
    wo_overdue = [wo for wo in open_workorders if wo.opened_at and wo.opened_at.date() <= overdue_threshold]
    wo_recent  = [wo for wo in open_workorders if wo not in wo_overdue]
    wo_total   = len(open_workorders)
    wo_open_url = _workorder_list_page_url(
        status=WorkOrder.STATUS_OPEN,
        reparto=reparto_filter,
        assigned=assigned_filter,
    )
    wo_overdue_url = _workorder_list_page_url(
        status=WorkOrder.STATUS_OPEN,
        reparto=reparto_filter,
        assigned=assigned_filter,
        open_age=wo_overdue_days,
    )
    wo_done_url = reverse("assets:maintenance_history")

    # ── KPI condivisi (count) ──────────────────────────────────────────────
    upcoming_deadlines_count = AssetAdministrativeDeadline.objects.filter(
        is_active=True, due_date__gte=today, due_date__lte=horizon_30
    ).count()
    overdue_deadlines_count = AssetAdministrativeDeadline.objects.filter(
        is_active=True, due_date__lt=today
    ).count()
    # Le verifiche is_legacy=True sono ora gestite dalle MaintenanceRule (trigger temporale):
    # escluse dai conteggi/scadenzario a tempo per evitare il doppio conteggio con le regole.
    overdue_verifications_count = PeriodicVerification.objects.filter(
        is_active=True, is_legacy=False, next_verification_date__lt=today
    ).count()
    upcoming_verifications_count = PeriodicVerification.objects.filter(
        is_active=True, is_legacy=False, next_verification_date__gte=today, next_verification_date__lte=horizon_30
    ).count()
    rules_count = MaintenanceRule.objects.filter(is_active=True).count()
    contracts_count = AssistanceContract.objects.filter(is_active=True).count()
    contracts_expiring_count = AssistanceContract.objects.filter(
        is_active=True, end_date__isnull=False, end_date__gte=today, end_date__lte=horizon_30
    ).count()
    closed_recent_count = WorkOrder.objects.filter(
        status=WorkOrder.STATUS_DONE,
        closed_at__gte=today - timedelta(days=30)
    ).count()

    # ── Scadenze/verifiche urgenti e imminenti (tab da_fare) ───────────────
    _url_deadlines = reverse("assets:asset_administrative_deadline_list")
    _url_verifications = reverse("assets:periodic_verifications")

    def _deadline_items(qs, is_overdue):
        items = []
        for d in qs.select_related("asset")[:20]:
            asset_label = f"{d.asset.asset_tag} — {d.asset.name}" if d.asset else "—"
            try:
                item_url = reverse("assets:asset_administrative_deadline_edit", args=[d.pk])
            except Exception:
                item_url = _url_deadlines
            items.append({
                "title": d.title,
                "asset_label": asset_label,
                "due_date": d.due_date,
                "is_overdue": is_overdue,
                "url": item_url,
            })
        return items

    def _verification_items(qs, is_overdue):
        items = []
        for v in qs.prefetch_related("assets")[:20]:
            assets_list = list(v.assets.all()[:3])
            if assets_list:
                tags = ", ".join(a.asset_tag for a in assets_list[:2])
                total = v.assets.count()
                asset_label = f"{tags}{f' +{total - 2} altri' if total > 2 else ''}"
            else:
                asset_label = "Nessun asset"
            items.append({
                "title": v.name,
                "asset_label": asset_label,
                "due_date": v.next_verification_date,
                "is_overdue": is_overdue,
                "url": _url_verifications,
            })
        return items

    urgent_items = (
        _deadline_items(
            AssetAdministrativeDeadline.objects.filter(is_active=True, due_date__lt=today).order_by("due_date"),
            True,
        )
        + _verification_items(
            PeriodicVerification.objects.filter(is_active=True, is_legacy=False, next_verification_date__lt=today).order_by("next_verification_date"),
            True,
        )
    )
    urgent_items.sort(key=lambda x: x["due_date"])

    upcoming_items = (
        _deadline_items(
            AssetAdministrativeDeadline.objects.filter(is_active=True, due_date__gte=today, due_date__lte=horizon_30).order_by("due_date"),
            False,
        )
        + _verification_items(
            PeriodicVerification.objects.filter(is_active=True, is_legacy=False, next_verification_date__gte=today, next_verification_date__lte=horizon_30).order_by("next_verification_date"),
            False,
        )
    )
    upcoming_items.sort(key=lambda x: x["due_date"])

    # ── Macchine con manutenzione in ritardo o prossima (14gg) ─────────────
    wm_qs = (
        WorkMachine.objects
        .filter(next_maintenance_date__isnull=False, next_maintenance_date__lte=horizon_14)
        .select_related("asset")
        .order_by("next_maintenance_date")
    )
    if reparto_filter:
        wm_qs = wm_qs.filter(asset__reparto=reparto_filter)
    machines_due = list(wm_qs[:20])
    machines_overdue = [m for m in machines_due if m.next_maintenance_date <= today]
    machines_warning = [m for m in machines_due if m.next_maintenance_date > today]

    # Regole manutenzione effettive: asset con baseline mancante,
    # regole scadute o prossime, senza duplicare la logica dello scadenzario.
    maintenance_rule_counts = {"overdue": 0, "warning": 0, "upcoming": 0, "missing": 0}
    maintenance_rule_rows: list[dict[str, object]] = []
    maintenance_asset_qs = Asset.objects.select_related("asset_category").filter(asset_category__isnull=False)
    if reparto_filter:
        maintenance_asset_qs = maintenance_asset_qs.filter(reparto__iexact=reparto_filter)
    for row in build_day_based_maintenance_schedule_rows(asset_queryset=maintenance_asset_qs, today=today):
        status = str(row.get("schedule_status") or "")
        if status in maintenance_rule_counts:
            maintenance_rule_counts[status] += 1
        if status not in {"overdue", "warning", "missing"}:
            continue
        if len(maintenance_rule_rows) >= 12:
            continue
        asset = row["asset"]
        primary_action = _maintenance_row_primary_action(
            asset=asset,
            base_rule=row["base_rule"],
            schedule_status=status,
            source="maintenance_schedule",
        )
        row["asset_detail_url"] = reverse("assets:asset_view", kwargs={"id": asset.id})
        row["primary_action_label"] = primary_action["label"]
        row["primary_action_url"] = primary_action["url"]
        maintenance_rule_rows.append(row)
    maintenance_rule_critical_count = (
        maintenance_rule_counts["overdue"]
        + maintenance_rule_counts["warning"]
        + maintenance_rule_counts["missing"]
    )

    # ── Ticket MAN aperti (integrazione modulo tickets) ────────────────────
    man_tickets = []
    try:
        from assets.services.dashboard_kpi import _base_ticket_man_qs, _ticket_open_statuses
        man_qs = _base_ticket_man_qs().filter(stato__in=_ticket_open_statuses())
        if not is_admin:
            man_qs = man_qs.filter(assegnato_a=request.user)
        man_tickets = list(man_qs.order_by("data_apertura")[:20])
    except Exception:
        pass

    # ── Prossimi 7gg (colonna destra) ──────────────────────────────────────
    next7_items: list[dict] = []
    for d in AssetAdministrativeDeadline.objects.filter(is_active=True, due_date__gte=today, due_date__lte=horizon_7).order_by("due_date").select_related("asset")[:10]:
        next7_items.append({"title": d.title, "due_date": d.due_date, "kind": "scadenza", "kind_label": "Scadenza"})
    for v in PeriodicVerification.objects.filter(is_active=True, is_legacy=False, next_verification_date__gte=today, next_verification_date__lte=horizon_7).order_by("next_verification_date")[:10]:
        next7_items.append({"title": v.name, "due_date": v.next_verification_date, "kind": "verifica", "kind_label": "Verifica"})
    for c in AssistanceContract.objects.filter(is_active=True, end_date__gte=today, end_date__lte=horizon_7).order_by("end_date")[:5]:
        next7_items.append({"title": c.title, "due_date": c.end_date, "kind": "contratto", "kind_label": "Contratto"})
    next7_items.sort(key=lambda x: x["due_date"])

    # Opzioni filtri
    reparto_options = list(
        Asset.objects.exclude(reparto="")
        .order_by("reparto")
        .values_list("reparto", flat=True)
        .distinct()
    )
    user_options = []
    if is_admin:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user_options = list(User.objects.filter(is_active=True).order_by("last_name", "first_name", "username"))

    # ── Dati tab "scadenzario" ─────────────────────────────────────────────
    scope_filter = _clean_string(request.GET.get("scope")) or "all"
    verif_qs = PeriodicVerification.objects.exclude(is_legacy=True).prefetch_related("assets").select_related("supplier")
    if scope_filter == "it":
        verif_qs = verif_qs.filter(assets__asset_type__in=IT_DEVICE_TYPES).distinct()
    elif scope_filter == "production":
        verif_qs = verif_qs.exclude(assets__asset_type__in=IT_DEVICE_TYPES).distinct()
    verif_rows = []
    for v in verif_qs.order_by("next_verification_date")[:60]:
        state, days_left, days_abs = _due_state(v.next_verification_date, today)
        verif_rows.append({"v": v, "days_left": days_left, "days_abs": days_abs, "state": state})

    deadline_type_filter = _clean_string(request.GET.get("dtype")) or ""
    deadline_qs = AssetAdministrativeDeadline.objects.filter(is_active=True).select_related("asset")
    if deadline_type_filter:
        deadline_qs = deadline_qs.filter(deadline_type=deadline_type_filter)
    deadline_rows = []
    for d in deadline_qs.order_by("due_date")[:80]:
        state, days_left, days_abs = _due_state(d.due_date, today)
        deadline_rows.append({"d": d, "days_left": days_left, "days_abs": days_abs, "state": state})

    contract_qs = AssistanceContract.objects.filter(is_active=True).select_related("supplier", "asset", "asset_category")
    contract_rows = []
    for c in contract_qs.order_by("end_date")[:60]:
        state, days_left, days_abs = _due_state(c.end_date, today)
        contract_rows.append({"c": c, "days_left": days_left, "days_abs": days_abs, "state": state})

    # Sotto-tab dello scadenzario (verifiche/scadenze/contratti)
    scad_sub = _clean_string(request.GET.get("sub")) or "verifiche"
    if scad_sub not in ("verifiche", "scadenze", "contratti"):
        scad_sub = "verifiche"

    return render(
        request,
        "assets/pages/maintenance_hub.html",
        {
            **_assets_shell_context(request),
            "page_title": "Manutenzione",
            "today": today,
            "is_admin": is_admin,
            "active_tab": active_tab,
            "scad_sub": scad_sub,
            # filtri tab da_fare
            "reparto_filter": reparto_filter,
            "reparto_options": reparto_options,
            "assigned_filter": assigned_filter,
            "user_options": user_options,
            # OdL
            "wo_overdue": wo_overdue,
            "wo_recent": wo_recent,
            "wo_total": wo_total,
            # KPI
            "upcoming_deadlines_count": upcoming_deadlines_count,
            "overdue_deadlines_count": overdue_deadlines_count,
            "overdue_verifications_count": overdue_verifications_count,
            "upcoming_verifications_count": upcoming_verifications_count,
            "contracts_count": contracts_count,
            "contracts_expiring_count": contracts_expiring_count,
            "rules_count": rules_count,
            "closed_recent_count": closed_recent_count,
            # liste tab da_fare
            "urgent_items": urgent_items,
            "upcoming_items": upcoming_items,
            "machines_overdue": machines_overdue,
            "machines_warning": machines_warning,
            "maintenance_rule_rows": maintenance_rule_rows,
            "maintenance_rule_counts": maintenance_rule_counts,
            "maintenance_rule_critical_count": maintenance_rule_critical_count,
            "man_tickets": man_tickets,
            "next7_items": next7_items[:12],
            # dati tab scadenzario
            "verif_rows": verif_rows,
            "scope_filter": scope_filter,
            "deadline_rows": deadline_rows,
            "deadline_type_filter": deadline_type_filter,
            "deadline_type_choices": AssetAdministrativeDeadline.TYPE_CHOICES,
            "contract_rows": contract_rows,
            "overdue_verifications": overdue_verifications_count,
            "upcoming_verifications": upcoming_verifications_count,
            "overdue_deadlines": overdue_deadlines_count,
            "upcoming_deadlines": upcoming_deadlines_count,
            "contracts_active": contracts_count,
            "contracts_expiring": contracts_expiring_count,
            # URL
            "url_wo_list": reverse("assets:wo_list"),
            "url_wo_open": wo_open_url,
            "url_wo_overdue": wo_overdue_url,
            "url_wo_done": wo_done_url,
            "url_wo_create": _workorder_list_page_url(create=1),
            "url_hub_scadenze": _url_deadlines,
            "url_hub_verifiche": _url_verifications,
            "url_hub_contratti": reverse("assets:assistance_contract_list"),
            "url_maintenance_schedule": _maintenance_schedule_page_url(status="due", reparto=reparto_filter),
            "url_impostazioni": reverse("assets:maintenance_impostazioni"),
            "url_verifications_full": _url_verifications,
            "url_deadlines_full": _url_deadlines,
            "url_contracts_full": reverse("assets:assistance_contract_list"),
            "work_machine_list_url": reverse("assets:work_machine_list"),
        },
    )


@login_required
def maintenance_scadenzario(request: HttpRequest) -> HttpResponse:
    """Deprecata: lo scadenzario unico è ora /prossime/ (maintenance_schedule).
    Redirige lì preservando il filtro reparto."""
    reparto_filter = _clean_string(request.GET.get("reparto"))
    return redirect(_maintenance_schedule_page_url(reparto=reparto_filter), permanent=True)


def _maintenance_supplier_rows() -> list[dict[str, object]]:
    """Fornitori usati in manutenzione (regole esterne, OdL, contratti, verifiche periodiche),
    con i relativi conteggi. La gestione dei fornitori resta nel modulo dedicato /fornitori/:
    qui è una vista focalizzata con link, non un duplicato."""
    from django.db.models import Count

    from anagrafica.models import Fornitore

    from .models import AssistanceContract

    def _counts(qs) -> dict[int, int]:
        # .order_by() azzera l'ordering di default del modello: senza, su SQL Server
        # le colonne dell'ORDER BY ereditato non stanno nel GROUP BY -> errore 8127.
        return {
            row["supplier"]: row["c"]
            for row in qs.filter(supplier__isnull=False)
            .values("supplier")
            .annotate(c=Count("id"))
            .order_by()
        }

    rule_counts = _counts(MaintenanceRule.objects.filter(is_active=True))
    wo_counts = _counts(WorkOrder.objects.all())
    contract_counts = _counts(AssistanceContract.objects.filter(is_active=True))
    verif_counts = _counts(PeriodicVerification.objects.filter(is_active=True))

    # Includi anche i fornitori catalogati come "Manutenzione" anche se non ancora
    # collegati a una regola/OdL: compaiono con conteggi a 0, pronti da assegnare.
    catalog_ids = set(
        Fornitore.objects.filter(
            categoria=Fornitore.CATEGORIA_MANUTENZIONE, is_active=True
        ).values_list("id", flat=True)
    )

    supplier_ids = set(rule_counts) | set(wo_counts) | set(contract_counts) | set(verif_counts) | catalog_ids
    if not supplier_ids:
        return []

    rows: list[dict[str, object]] = []
    for supplier in Fornitore.objects.filter(pk__in=supplier_ids).order_by("ragione_sociale", "id"):
        rows.append(
            {
                "f": supplier,
                "rules": rule_counts.get(supplier.id, 0),
                "workorders": wo_counts.get(supplier.id, 0),
                "contracts": contract_counts.get(supplier.id, 0),
                "verifications": verif_counts.get(supplier.id, 0),
                "detail_url": reverse("fornitori:fornitore_detail", kwargs={"fornitore_id": supplier.id}),
            }
        )
    return rows


def _maintenance_plan_by_category_rows() -> list[dict[str, object]]:
    """Piano di manutenzione aggregato per categoria asset (filo conduttore Impostazioni → operativo).

    Per ogni categoria con almeno una regola attiva: conta regole, asset coinvolti e la
    salute della pianificazione (scadute / in scadenza / pianificate / mai eseguite),
    riusando lo scadenzario a giorni. Le regole a contatore (ore/km/cicli) sono conteggiate
    a parte: non rientrano ancora nello scadenzario giornaliero.
    """
    rules_by_cat: dict[int, list[MaintenanceRule]] = {}
    for rule in (
        MaintenanceRule.objects.filter(is_active=True)
        .select_related("asset_category", "intervention_template")
        .order_by("asset_category__sort_order", "asset_category__label", "sort_order", "id")
    ):
        if not rule.asset_category_id:
            continue
        rules_by_cat.setdefault(rule.asset_category_id, []).append(rule)

    agg: dict[int, dict[str, object]] = {}
    try:
        for row in build_day_based_maintenance_schedule_rows():
            asset = row["asset"]
            bucket = agg.setdefault(
                asset.asset_category_id,
                {"assets": set(), "overdue": 0, "warning": 0, "upcoming": 0, "missing": 0},
            )
            bucket["assets"].add(asset.id)
            status = str(row.get("schedule_status") or "")
            if status in ("overdue", "warning", "upcoming", "missing"):
                bucket[status] += 1
    except Exception:
        agg = {}

    rows: list[dict[str, object]] = []
    for cat_id, rules in rules_by_cat.items():
        category = rules[0].asset_category
        bucket = agg.get(cat_id, {"assets": set(), "overdue": 0, "warning": 0, "upcoming": 0, "missing": 0})
        meter_rules = sum(1 for r in rules if r.threshold_type != MaintenanceRule.THRESHOLD_DAYS)
        overdue = int(bucket["overdue"])
        warning = int(bucket["warning"])
        upcoming = int(bucket["upcoming"])
        missing = int(bucket["missing"])
        rows.append(
            {
                "category": category,
                "rules_count": len(rules),
                "day_rules_count": len(rules) - meter_rules,
                "meter_rules_count": meter_rules,
                "assets_count": len(bucket["assets"]),
                "overdue": overdue,
                "warning": warning,
                "upcoming": upcoming,
                "missing": missing,
                "due_total": overdue + warning + missing,
                "schedule_url": _maintenance_schedule_page_url(category_id=cat_id),
                "schedule_due_url": _maintenance_schedule_page_url(category_id=cat_id, status="due"),
                "rules_url": _maintenance_rule_list_page_url(category_id=cat_id),
            }
        )
    rows.sort(
        key=lambda r: (-int(r["overdue"]), -int(r["warning"]), -int(r["missing"]), str(r["category"].label).casefold())
    )
    return rows


@login_required
def maintenance_impostazioni(request: HttpRequest) -> HttpResponse:
    """Catalogo attivita, piani ordinari e copertura: centro di governo manutenzione."""
    # I fornitori hanno ora una pagina dedicata (nav di sezione): vecchi link ?tab=fornitori lì.
    if _clean_string(request.GET.get("tab")) == "fornitori":
        return redirect("assets:maintenance_suppliers")

    # POST: genera in blocco gli OdL delle manutenzioni in scadenza di una categoria
    # (riusa il comando periodico scoped alla categoria; dedup su OdL aperti, checklist inclusa).
    if request.method == "POST" and _clean_string(request.POST.get("action")) == "generate_workorders":
        piano_url = reverse("assets:maintenance_impostazioni") + "?tab=piano"
        if not _is_assets_admin(request):
            messages.error(request, "Solo admin può generare OdL di manutenzione.")
            return redirect(piano_url)
        cat_id = _as_int(request.POST.get("category_id"), default=0)
        category = AssetCategory.objects.filter(pk=cat_id).first() if cat_id else None
        if category is None:
            messages.error(request, "Categoria non valida.")
            return redirect(piano_url)
        import re as _re
        from io import StringIO as _StringIO

        from django.core.management import call_command as _call_command

        buf = _StringIO()
        try:
            _call_command("generate_scheduled_workorders", category=cat_id, limit=200, stdout=buf)
        except Exception as exc:  # pragma: no cover - dipende dai dati
            messages.error(request, f"Generazione OdL fallita: {exc}")
            return redirect(piano_url)
        text = buf.getvalue()
        m_created = _re.search(r"Creati=(\d+)", text)
        m_open = _re.search(r"GiaAperti=(\d+)", text)
        created = int(m_created.group(1)) if m_created else 0
        already = int(m_open.group(1)) if m_open else 0
        log_action(request, "bulk_generate_workorders", "assets", {"category_id": cat_id, "created": created})
        if created:
            messages.success(
                request,
                f"Creati {created} ordini di lavoro per «{category.label}»"
                + (f" ({already} già aperti, saltati)." if already else "."),
            )
        else:
            messages.info(
                request,
                f"Nessun nuovo OdL per «{category.label}»: le manutenzioni in scadenza hanno già un intervento aperto.",
            )
        return redirect(piano_url)

    active_tab = _clean_string(request.GET.get("tab")) or "catalogo"
    active_tab = {
        "interventi": "catalogo",
        "templates": "catalogo",
        "rules": "piani",
        "piano": "copertura",
    }.get(active_tab, active_tab)
    if active_tab not in ("catalogo", "piani", "copertura"):
        active_tab = "catalogo"
    is_admin = _is_assets_admin(request)

    from .models import MaintenanceInterventionTemplate

    # ── Filtri integrati (sostituiscono le pagine standalone "Vista avanzata") ──
    selected_category_id = _as_int(request.GET.get("category"), default=0)
    selected_category = (
        AssetCategory.objects.filter(pk=selected_category_id).first() if selected_category_id else None
    )
    active_filter = _clean_string(request.GET.get("active")).lower() or "active"
    if active_filter not in {"active", "inactive", "all"}:
        active_filter = "active"
    execution_filter = _clean_string(request.GET.get("execution")) or "all"
    if execution_filter not in {"all", "internal", "external"}:
        execution_filter = "all"
    q = _clean_string(request.GET.get("q"))

    template_qs = (
        MaintenanceInterventionTemplate.objects.select_related("asset_category")
        .prefetch_related("maintenance_rules__asset_category", "maintenance_rules__supplier")
        .order_by("asset_category__label", "sort_order", "label")
    )
    if selected_category is not None:
        template_qs = template_qs.filter(
            Q(asset_category__isnull=True) | Q(asset_category_id=selected_category.id)
        )
    if active_filter == "active":
        template_qs = template_qs.filter(is_active=True)
    elif active_filter == "inactive":
        template_qs = template_qs.filter(is_active=False)
    # Filtro esecuzione: template che hanno almeno una regola interna/esterna.
    if execution_filter == "external":
        template_qs = template_qs.filter(
            maintenance_rules__execution_mode=MaintenanceRule.MODE_EXTERNAL
        ).distinct()
    elif execution_filter == "internal":
        template_qs = template_qs.filter(
            maintenance_rules__execution_mode=MaintenanceRule.MODE_INTERNAL
        ).distinct()
    if q:
        template_qs = template_qs.filter(
            Q(code__icontains=q)
            | Q(label__icontains=q)
            | Q(description__icontains=q)
            | Q(asset_category__label__icontains=q)
        )

    template_rows = []
    for t in template_qs:
        rules = [r for r in t.maintenance_rules.all() if r.is_active]
        template_rows.append({"t": t, "rules": rules, "rules_count": len(rules)})

    has_active_filters = bool(
        selected_category is not None or q or active_filter != "active" or execution_filter != "all"
    )

    maintenance_plan_qs = (
        MaintenanceRule.objects.select_related(
            "intervention_template", "asset_category", "supplier", "assigned_to"
        )
        .prefetch_related("assets", "legacy_periodic_verifications")
        .annotate(
            history_count=Count("workorders", distinct=True),
            completed_count=Count(
                "workorders",
                filter=Q(workorders__status=WorkOrder.STATUS_DONE),
                distinct=True,
            ),
        )
        .order_by("asset_category__label", "sort_order", "intervention_template__label", "id")
    )
    if selected_category is not None:
        maintenance_plan_qs = maintenance_plan_qs.filter(asset_category=selected_category)
    if active_filter == "active":
        maintenance_plan_qs = maintenance_plan_qs.filter(is_active=True)
    elif active_filter == "inactive":
        maintenance_plan_qs = maintenance_plan_qs.filter(is_active=False)
    if execution_filter == "external":
        maintenance_plan_qs = maintenance_plan_qs.filter(execution_mode=MaintenanceRule.MODE_EXTERNAL)
    elif execution_filter == "internal":
        maintenance_plan_qs = maintenance_plan_qs.filter(execution_mode=MaintenanceRule.MODE_INTERNAL)
    if q:
        maintenance_plan_qs = maintenance_plan_qs.filter(
            Q(intervention_template__label__icontains=q)
            | Q(intervention_template__code__icontains=q)
            | Q(asset_category__label__icontains=q)
            | Q(assets__asset_tag__icontains=q)
            | Q(assets__name__icontains=q)
            | Q(supplier__ragione_sociale__icontains=q)
            | Q(assigned_to__username__icontains=q)
            | Q(assigned_to__first_name__icontains=q)
            | Q(assigned_to__last_name__icontains=q)
        ).distinct()

    category_asset_counts = dict(
        Asset.objects.filter(status=Asset.STATUS_IN_USE, asset_category_id__isnull=False)
        .values_list("asset_category_id")
        .annotate(total=Count("id"))
    )
    maintenance_plan_rows = []
    for rule in maintenance_plan_qs:
        targeted_assets = list(rule.assets.all())
        maintenance_plan_rows.append(
            {
                "rule": rule,
                "asset_count": (
                    len(targeted_assets)
                    if rule.scope_type == MaintenanceRule.SCOPE_ASSETS
                    else category_asset_counts.get(rule.asset_category_id, 0)
                ),
                "targeted_assets": targeted_assets,
                "legacy_count": len(rule.legacy_periodic_verifications.all()),
            }
        )

    plan_rows = _maintenance_plan_by_category_rows() if active_tab == "copertura" else []
    plan_totals = {
        "overdue": sum(int(r["overdue"]) for r in plan_rows),
        "warning": sum(int(r["warning"]) for r in plan_rows),
        "missing": sum(int(r["missing"]) for r in plan_rows),
    }
    # Conteggio leggero (sempre disponibile) per il badge del tab Piano
    plan_category_count = (
        MaintenanceRule.objects.filter(is_active=True, asset_category__isnull=False)
        .values("asset_category_id")
        .distinct()
        .count()
    )

    return render(
        request,
        "assets/pages/maintenance_impostazioni.html",
        {
            **_assets_shell_context(request),
            "page_title": "Catalogo e piani manutenzione",
            "active_tab": active_tab,
            "is_admin": is_admin,
            "template_rows": template_rows,
            "template_count": len(template_rows),
            "maintenance_plan_rows": maintenance_plan_rows,
            "maintenance_plan_count": len(maintenance_plan_rows),
            "ingested_periodic_count": PeriodicVerification.objects.filter(
                maintenance_rules__isnull=False
            ).distinct().count(),
            "pending_periodic_count": PeriodicVerification.objects.filter(
                maintenance_rules__isnull=True
            ).count(),
            "plan_rows": plan_rows,
            "plan_category_count": plan_category_count,
            "plan_totals": plan_totals,
            # Filtri integrati
            "category_options": list(AssetCategory.objects.order_by("sort_order", "label", "id")),
            "selected_category": selected_category,
            "active_filter": active_filter,
            "execution_filter": execution_filter,
            "q": q,
            "has_active_filters": has_active_filters,
            "clear_filters_url": reverse("assets:maintenance_impostazioni") + f"?tab={active_tab}",
            "url_suppliers": reverse("assets:maintenance_suppliers"),
            "url_hub": reverse("assets:maintenance_hub"),
            "url_scadenzario": reverse("assets:maintenance_scadenzario"),
            "url_schedule": reverse("assets:maintenance_schedule"),
            "url_template_new": reverse("assets:maintenance_template_create"),
            "url_rule_new": reverse("assets:maintenance_rule_create"),
        },
    )


@login_required
def maintenance_suppliers(request: HttpRequest) -> HttpResponse:
    """Pagina dedicata: fornitori (ditte terze) usati in manutenzione, con conteggi e link
    al modulo /fornitori/ (fonte unica di gestione/anagrafica). Raggiungibile dalla nav di
    sezione manutenzione; nessun duplicato di anagrafica."""
    rows = _maintenance_supplier_rows()
    return render(
        request,
        "assets/pages/maintenance_suppliers.html",
        {
            **_assets_shell_context(request),
            "page_title": "Fornitori manutenzione",
            "is_admin": _is_assets_admin(request),
            "supplier_rows": rows,
            "supplier_total": len(rows),
            "url_fornitori_list": reverse("fornitori:fornitori_list"),
            "url_fornitore_new": reverse("fornitori:fornitore_create"),
        },
    )


@login_required
def maintenance_worksheet(request: HttpRequest, asset_id: int, rule_id: int) -> HttpResponse:
    """Scheda intervento stampabile (A4) da portare alla macchina: dati macchina, intervento,
    checklist da spuntare a penna, ultime esecuzioni e campi per la registrazione manuale."""
    from .models import AssetMaintenanceRuleState, MaintenanceChecklistStep

    asset = get_object_or_404(Asset.objects.select_related("asset_category"), pk=asset_id)
    base_rule = get_object_or_404(
        MaintenanceRule.objects.select_related("intervention_template", "supplier", "asset_category"),
        pk=rule_id,
    )
    template = base_rule.intervention_template
    checklist_steps = list(
        MaintenanceChecklistStep.objects.filter(intervention_template=template).order_by("step_number", "id")
    )
    last_executions = list(
        WorkOrder.objects.filter(
            asset=asset, maintenance_rule=base_rule, status=WorkOrder.STATUS_DONE
        ).order_by("-closed_at", "-id")[:5]
    )
    state = AssetMaintenanceRuleState.objects.filter(asset=asset, base_rule=base_rule).first()
    return render(
        request,
        "assets/pages/maintenance_worksheet.html",
        {
            "page_title": f"Scheda intervento — {asset.asset_tag}",
            "asset": asset,
            "base_rule": base_rule,
            "template": template,
            "checklist_steps": checklist_steps,
            "last_executions": last_executions,
            "last_execution_date": state.last_execution_date if state else None,
            "today": timezone.localdate(),
        },
    )


@login_required
def maintenance_todo(request: HttpRequest) -> HttpResponse:
    """Deprecata: il 'to-do' è ora la tab 'da_fare' del Centro Manutenzione unificato."""
    params = request.GET.copy()
    params["tab"] = "da_fare"
    return redirect(f"{reverse('assets:maintenance_hub')}?{params.urlencode()}", permanent=True)


@login_required
def reports_dashboard(request: HttpRequest) -> HttpResponse:
    if "scope" not in request.GET:
        query = request.GET.copy()
        query["scope"] = "production"
        return redirect(f"{reverse('assets:reports')}?{query.urlencode()}")

    now = timezone.now()
    today = timezone.localdate()
    period_start = now - timedelta(days=90)
    reports_scope = _normalize_reports_scope(request.GET.get("scope"))
    reports_context = _reports_scope_context(reports_scope)
    report_asset_types = list(reports_context["asset_types"])
    scoped_asset_qs = Asset.objects.filter(asset_type__in=report_asset_types).select_related("asset_category")
    schedule_rows = build_day_based_maintenance_schedule_rows(asset_queryset=scoped_asset_qs, today=today)
    maintenance_report_kpis = build_maintenance_report_kpis(
        asset_queryset=scoped_asset_qs,
        schedule_rows=schedule_rows,
        today=today,
    )
    pm_kpi = maintenance_report_kpis["pm"]
    budget_kpi = maintenance_report_kpis["budget"]
    budget_rows = []
    for row in budget_kpi["rows"][:8]:
        row = dict(row)
        category = row.get("category")
        category_id = getattr(category, "id", 0)
        row["workorders_url"] = (
            _workorder_list_page_url(status=WorkOrder.STATUS_DONE, category=category_id)
            if category_id
            else ""
        )
        budget_rows.append(row)

    open_workorders = WorkOrder.objects.select_related("asset", "supplier").filter(
        status=WorkOrder.STATUS_OPEN,
        asset__asset_type__in=report_asset_types,
    )
    late_open_workorders = open_workorders.filter(opened_at__lt=now - timedelta(days=30)).order_by("opened_at")
    open_workorder_rows: list[dict[str, object]] = []
    for workorder in open_workorders.order_by("opened_at", "id")[:12]:
        opened_at = workorder.opened_at
        opened_on = timezone.localtime(opened_at).date() if timezone.is_aware(opened_at) else opened_at.date()
        open_days = max((today - opened_on).days, 0)
        open_workorder_rows.append(
            {
                "workorder": workorder,
                "open_days": open_days,
                "is_late": open_days > 30,
            }
        )
    recent_done_workorders = (
        WorkOrder.objects.select_related(
            "asset",
            "asset__asset_category",
            "supplier",
            "assistance_contract",
        )
        .filter(status=WorkOrder.STATUS_DONE, closed_at__gte=period_start)
        .filter(asset__asset_type__in=report_asset_types)
        .order_by("-closed_at", "-id")
    )
    if reports_context["maintenance_month_enabled"]:
        maintenance_month_dataset = _build_work_machine_maintenance_month_dataset(
            month_value=request.GET.get("month"),
            today=today,
        )
    else:
        month_start = _month_start_from_value(request.GET.get("month"), today=today)
        month_end = _month_end(month_start)
        maintenance_month_dataset = {
            "rows": [],
            "total_count": 0,
            "overdue_count": 0,
            "warning_count": 0,
            "ok_count": 0,
            "month_code": month_start.strftime("%Y-%m"),
            "month_label": _month_label(month_start),
            "period_label": f'{month_start.strftime("%d-%m-%Y")} - {month_end.strftime("%d-%m-%Y")}',
        }
    done_rows = list(recent_done_workorders)
    total_cost = sum((workorder.resolved_total_cost_eur or Decimal("0")) for workorder in done_rows)
    duration_values = [
        workorder.intervention_duration_minutes
        for workorder in done_rows
        if workorder.intervention_duration_minutes is not None
    ]
    avg_duration_minutes = round(sum(duration_values) / len(duration_values), 1) if duration_values else None
    avg_downtime_minutes = (
        WorkOrder.objects.filter(
            status=WorkOrder.STATUS_DONE,
            closed_at__gte=period_start,
            asset__asset_type__in=report_asset_types,
        )
        .aggregate(value=Avg("downtime_minutes"))
        .get("value")
    )
    covered_count = sum(1 for workorder in done_rows if workorder.covered_by_contract)
    uncovered_count = len(done_rows) - covered_count

    def _summary_rows(key_builder, label_builder):
        grouped: dict[tuple, dict[str, object]] = {}
        for workorder in done_rows:
            key = key_builder(workorder)
            bucket = grouped.setdefault(
                key,
                {
                    "label": label_builder(workorder),
                    "count": 0,
                    "cost": Decimal("0"),
                    "covered_count": 0,
                },
            )
            bucket["count"] += 1
            bucket["cost"] += workorder.resolved_total_cost_eur or Decimal("0")
            if workorder.covered_by_contract:
                bucket["covered_count"] += 1
        rows = list(grouped.values())
        rows.sort(key=lambda row: (-int(row["cost"] > 0), -row["cost"], -row["count"], str(row["label"]).casefold()))
        return rows

    asset_summary_rows = _summary_rows(
        lambda workorder: (workorder.asset_id,),
        lambda workorder: f"{workorder.asset.asset_tag} - {workorder.asset.name}",
    )[:8]
    category_summary_rows = _summary_rows(
        lambda workorder: (getattr(workorder.asset, "asset_category_id", 0),),
        lambda workorder: (
            getattr(getattr(workorder.asset, "asset_category", None), "label", "")
            or "Senza categoria"
        ),
    )[:8]
    supplier_summary_rows = _summary_rows(
        lambda workorder: (workorder.supplier_id or 0,),
        lambda workorder: str(workorder.supplier) if workorder.supplier_id else "Fornitore non indicato",
    )[:8]

    overdue_rows = [row for row in schedule_rows if row["schedule_status"] == "overdue"]
    warning_rows = [row for row in schedule_rows if row["schedule_status"] == "warning"]
    upcoming_rows = [row for row in schedule_rows if row["schedule_status"] == "upcoming"]
    missing_rows = [row for row in schedule_rows if row["schedule_status"] == "missing"]
    critical_count = len(overdue_rows) + len(warning_rows) + len(missing_rows)
    critical_rows: list[dict[str, object]] = []
    for row in schedule_rows:
        if row["schedule_status"] not in {"overdue", "warning", "missing"}:
            continue
        asset = row["asset"]
        critical_rows.append(
            {
                **row,
                "asset_detail_url": reverse("assets:asset_view", kwargs={"id": asset.id}),
                "workorder_create_url": _workorder_create_page_url(
                    asset_id=asset.id,
                    rule_id=row["base_rule"].id,
                    source="maintenance_reports",
                ),
                "primary_action": _maintenance_row_primary_action(
                    asset=asset,
                    base_rule=row["base_rule"],
                    schedule_status=str(row.get("schedule_status") or ""),
                    source="maintenance_reports",
                ),
            }
        )

    return render(
        request,
        "assets/pages/reports_dashboard.html",
        {
            "page_title": reports_context["title"],
            "reports_scope": reports_context["scope"],
            "reports_page_title": reports_context["title"],
            "reports_page_subtitle": reports_context["subtitle"],
            "maintenance_month_enabled": reports_context["maintenance_month_enabled"],
            "maintenance_month_empty_message": reports_context["empty_month_message"],
            "open_workorders": open_workorders.order_by("opened_at", "id")[:8],
            "open_workorder_rows": open_workorder_rows,
            "late_open_workorders": late_open_workorders[:8],
            "recent_done_workorders": done_rows[:8],
            "asset_summary_rows": asset_summary_rows,
            "category_summary_rows": category_summary_rows,
            "supplier_summary_rows": supplier_summary_rows,
            "critical_rows": critical_rows[:12],
            "overdue_rows": overdue_rows[:8],
            "warning_rows": warning_rows[:8],
            "upcoming_rows": upcoming_rows[:8],
            "open_count": open_workorders.count(),
            "late_count": late_open_workorders.count(),
            "done_recent_count": len(done_rows),
            "total_cost": total_cost,
            "avg_duration_minutes": avg_duration_minutes,
            "avg_downtime_minutes": round(avg_downtime_minutes, 1) if avg_downtime_minutes is not None else None,
            "covered_count": covered_count,
            "uncovered_count": uncovered_count,
            "critical_count": critical_count,
            "overdue_count": len(overdue_rows),
            "warning_count": len(warning_rows),
            "upcoming_count": len(upcoming_rows),
            "missing_count": len(missing_rows),
            "pm_kpi": pm_kpi,
            "budget_kpi": budget_kpi,
            "budget_rows": budget_rows,
            "open_workorders_url": _workorder_list_page_url(status=WorkOrder.STATUS_OPEN),
            "late_workorders_url": _workorder_list_page_url(status=WorkOrder.STATUS_OPEN, open_age=30),
            "done_workorders_url": _workorder_list_page_url(status=WorkOrder.STATUS_DONE),
            "maintenance_schedule_due_url": _maintenance_schedule_page_url(status="due"),
            "maintenance_schedule_missing_url": _maintenance_schedule_page_url(status="missing"),
            "maintenance_month_rows": maintenance_month_dataset["rows"][:10],
            "maintenance_month_count": maintenance_month_dataset["total_count"],
            "maintenance_month_overdue_count": maintenance_month_dataset["overdue_count"],
            "maintenance_month_warning_count": maintenance_month_dataset["warning_count"],
            "maintenance_month_ok_count": maintenance_month_dataset["ok_count"],
            "maintenance_month_label": maintenance_month_dataset["month_label"],
            "maintenance_month_period_label": maintenance_month_dataset["period_label"],
            "maintenance_month_code": maintenance_month_dataset["month_code"],
            "maintenance_month_pdf_base_url": reverse("assets:work_machine_maintenance_month_pdf"),
            "maintenance_month_pdf_url": _work_machine_maintenance_month_pdf_url(
                month_code=str(maintenance_month_dataset["month_code"])
            ),
            "maintenance_schedule_url": reverse("assets:maintenance_schedule"),
            "assistance_contracts_url": reverse("assets:assistance_contract_list"),
            "report_templates_manage_url": (
                reverse("assets:report_template_admin")
                if _is_assets_admin(request)
                else ""
            ),
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def asset_report_pdf(request: HttpRequest, id: int | None = None) -> HttpResponse:
    if id is None:
        return redirect("assets:asset_list")
    asset = get_object_or_404(
        Asset.objects.select_related("asset_category", "it_details", "work_machine")
        .prefetch_related(
            "workorders",
            "documents",
            "tickets",
            "periodic_verifications",
            "periodic_verifications__supplier",
            "asset_category__category_fields",
        ),
        pk=id,
    )
    active_template = _active_asset_report_template(AssetReportTemplate.REPORT_ASSET_DETAIL)
    snapshot = _build_asset_report_snapshot(asset)
    pdf_bytes = render_asset_report_pdf(
        asset,
        snapshot,
        template_name=active_template.name if active_template else "",
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{asset.asset_tag or "asset"}-report.pdf"'
    return response


@login_required
def asset_detail_export_xlsx(request: HttpRequest, id: int) -> HttpResponse:
    import openpyxl
    import openpyxl.styles as xlst

    from core.excel_export import write_cell

    asset = get_object_or_404(
        Asset.objects.select_related("asset_category", "it_details", "work_machine"),
        pk=id,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scheda asset"

    hfill = xlst.PatternFill(fill_type="solid", fgColor="2563EB")
    hfont = xlst.Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    sfill = xlst.PatternFill(fill_type="solid", fgColor="EFF6FF")
    sfont = xlst.Font(bold=True, name="Calibri", size=9)
    kfont = xlst.Font(bold=True, name="Calibri", size=9)
    vfont = xlst.Font(name="Calibri", size=9)

    r = 1
    for ci, h in enumerate(["Attributo", "Valore"], 1):
        cell = write_cell(ws, r, ci, h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = xlst.Alignment(horizontal="center", vertical="center")
    r += 1

    def _sec(label):
        nonlocal r
        c = write_cell(ws, r, 1, label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c.fill = sfill
        c.font = sfont
        r += 1

    def _kv(key, val):
        nonlocal r
        write_cell(ws, r, 1, key).font = kfont
        write_cell(ws, r, 2, str(val) if val is not None else "").font = vfont
        r += 1

    _sec("Anagrafica")
    _kv("Tag", asset.asset_tag)
    _kv("Nome", asset.name)
    _kv("Tipo", asset.get_asset_type_display())
    _kv("Categoria", asset.asset_category.label if asset.asset_category_id else "")
    _kv("Reparto", asset.reparto)
    _kv("Stato", asset.get_status_display())
    _kv("Produttore", asset.manufacturer)
    _kv("Modello", asset.model)
    _kv("Matricola", asset.serial_number)
    _kv("Note", asset.notes)

    _sec("Assegnazione")
    _kv("Assegnato a", asset.assignment_to)
    _kv("Reparto assegnazione", asset.assignment_reparto)
    _kv("Ubicazione", asset.assignment_location)

    it = getattr(asset, "it_details", None)
    if it:
        _sec("Dettagli IT")
        _kv("OS", it.os)
        _kv("CPU", it.cpu)
        _kv("RAM", it.ram)
        _kv("Disco", it.disco)
        _kv("Dominio", "Sì" if it.domain_joined else "No")
        _kv("EDR", "Sì" if it.edr_enabled else "No")
        _kv("2FA Office", "Sì" if it.office_2fa_enabled else "No")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tag = asset.asset_tag or str(asset.pk)
    today = timezone.localdate().strftime("%Y%m%d")
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="asset_{tag}_{today}.xlsx"'
    return resp


@login_required
def report_template_admin(request: HttpRequest) -> HttpResponse:
    if not _is_assets_admin(request):
        messages.error(request, "Solo admin puo gestire i template report.")
        return redirect("assets:reports")

    report_tables_ready = _model_table_exists(AssetReportDefinition) and _model_table_exists(AssetReportTemplate)
    _ensure_default_asset_report_definitions()

    if request.method == "POST":
        if not report_tables_ready:
            messages.error(request, "Le tabelle dei template report non sono ancora disponibili. Esegui prima le migrazioni.")
            return redirect("assets:report_template_admin")
        action = _clean_string(request.POST.get("action"))
        redirect_url = reverse("assets:report_template_admin")
        definition_map = _asset_report_definition_map()

        try:
            if action == "create_report_definition":
                code = slugify(request.POST.get("code", ""))[:80]
                label = _clean_string(request.POST.get("label"))[:120]
                description = _clean_string(request.POST.get("description"))[:255]
                sort_order = _as_int(request.POST.get("sort_order"), default=100)
                if not code or not label:
                    messages.error(request, "Inserisci codice e nome del report.")
                    return redirect(redirect_url)
                if AssetReportDefinition.objects.filter(code=code).exists():
                    messages.error(request, "Esiste gia un report con questo codice.")
                    return redirect(redirect_url)
                AssetReportDefinition.objects.create(
                    code=code,
                    label=label,
                    description=description,
                    sort_order=sort_order,
                    is_active=True,
                )
                messages.success(request, f'Report "{label}" creato.')
                return redirect(redirect_url)

            if action == "upload_report_template":
                report_code = _clean_string(request.POST.get("report_code"))
                name = _clean_string(request.POST.get("name"))[:120]
                version = _clean_string(request.POST.get("version"))[:40]
                description = _clean_string(request.POST.get("description"))[:255]
                uploaded_file = request.FILES.get("template_file")

                if report_code not in definition_map:
                    messages.error(request, "Tipo report non valido.")
                    return redirect(redirect_url)
                if not name:
                    messages.error(request, "Inserisci il nome del template.")
                    return redirect(redirect_url)
                if not uploaded_file:
                    messages.error(request, "Seleziona un file template.")
                    return redirect(redirect_url)

                suffix = Path(uploaded_file.name or "").suffix.lower()
                if suffix not in REPORT_TEMPLATE_ALLOWED_EXTENSIONS:
                    messages.error(request, "Formato template non supportato.")
                    return redirect(redirect_url)

                should_activate = bool(request.POST.get("is_active")) or not AssetReportTemplate.objects.filter(
                    report_code=report_code
                ).exists()
                AssetReportTemplate.objects.create(
                    report_code=report_code,
                    name=name,
                    version=version,
                    description=description,
                    file=uploaded_file,
                    original_name=(uploaded_file.name or "")[:255],
                    is_active=should_activate,
                    uploaded_by=request.user,
                )
                messages.success(request, f'Template report "{name}" caricato.')
                return redirect(redirect_url)

            if action == "activate_report_template":
                template_id = _as_int(request.POST.get("template_id"), default=0)
                template = AssetReportTemplate.objects.filter(pk=template_id).first()
                if not template:
                    messages.error(request, "Template report non trovato.")
                    return redirect(redirect_url)
                template.is_active = True
                template.save(update_fields=["is_active", "updated_at"])
                report_label = getattr(definition_map.get(template.report_code), "label", template.report_code)
                messages.success(request, f"Template attivo aggiornato per {report_label.lower()}.")
                return redirect(redirect_url)

            if action == "delete_report_template":
                template_id = _as_int(request.POST.get("template_id"), default=0)
                template = AssetReportTemplate.objects.filter(pk=template_id).first()
                if not template:
                    messages.error(request, "Template report non trovato.")
                    return redirect(redirect_url)
                report_label = getattr(definition_map.get(template.report_code), "label", template.report_code)
                template_name = template.name
                template.delete()
                messages.success(request, f'Template "{template_name}" eliminato ({report_label}).')
                return redirect(redirect_url)

            if action == "delete_report_definition":
                report_id = _as_int(request.POST.get("report_definition_id"), default=0)
                definition = AssetReportDefinition.objects.filter(pk=report_id).first()
                if not definition:
                    messages.error(request, "Report non trovato.")
                    return redirect(redirect_url)
                if definition.code in {
                    AssetReportTemplate.REPORT_ASSET_DETAIL,
                    AssetReportTemplate.REPORT_WORK_MACHINE_MAINTENANCE,
                }:
                    messages.error(request, "I report base non possono essere eliminati.")
                    return redirect(redirect_url)
                AssetReportTemplate.objects.filter(report_code=definition.code).delete()
                report_label = definition.label
                definition.delete()
                messages.success(request, f'Report "{report_label}" eliminato.')
                return redirect(redirect_url)
        except DatabaseError:
            messages.error(request, "Le tabelle dei template report non sono ancora disponibili. Esegui prima le migrazioni.")
            return redirect(redirect_url)

    return render(
        request,
        "assets/pages/report_template_admin.html",
        {
            "page_title": "Gestione template report",
            "report_template_groups": _report_templates_grouped(),
            "report_template_choices": [
                (row.code, row.label) for row in _ensure_default_asset_report_definitions() if row.is_active
            ],
            "report_definitions": _ensure_default_asset_report_definitions(),
            "report_template_extensions": ", ".join(sorted(REPORT_TEMPLATE_ALLOWED_EXTENSIONS)),
            **_assets_shell_context(request, rows=_as_int(request.GET.get("rows"), default=25)),
        },
    )


@login_required
def work_machine_maintenance_month_pdf(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    reparto_filter = _clean_string(request.GET.get("reparto"))
    dataset = _build_work_machine_maintenance_month_dataset(
        month_value=request.GET.get("month"),
        reparto_filter=reparto_filter,
        today=today,
    )
    generated_at = timezone.localtime()
    filename_parts = ["report", "macchine", "manutenzione", str(dataset["month_code"])]
    if reparto_filter:
        reparto_slug = slugify(reparto_filter)
        if reparto_slug:
            filename_parts.append(reparto_slug)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{"_".join(filename_parts)}.pdf"'

    buffer = io.BytesIO()
    theme = PdfTheme.from_branding()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))
    pdf.setTitle(f'Report manutenzioni macchine {dataset["month_label"]}')
    pdf.setAuthor(theme.portal_name)
    pdf.setSubject(f'Macchine con manutenzione pianificata nel periodo {dataset["period_label"]}')

    _draw_work_machine_maintenance_month_pdf(pdf, theme=theme, dataset=dataset, generated_at=generated_at)

    pdf.showPage()
    pdf.save()
    response.write(buffer.getvalue())
    return response


@legacy_admin_required
def gestione_admin(request: HttpRequest) -> HttpResponse:
    """Pagina di gestione interna Assets — accesso solo admin."""

    # --- Statistiche ---
    total_assets = Asset.objects.count()
    assets_by_status = dict(
        Asset.objects.values_list("status").annotate(n=Count("id")).order_by()
    )
    assets_by_type = list(
        Asset.objects.values("asset_type").annotate(n=Count("id")).order_by("-n")
    )
    total_wo = WorkOrder.objects.count()
    wo_by_status = dict(
        WorkOrder.objects.values_list("status").annotate(n=Count("id")).order_by()
    )

    # --- Configurazione: AssetListOption per campo ---
    list_options = {}
    for field_key, _ in AssetListOption.FIELD_CHOICES:
        list_options[field_key] = list(
            AssetListOption.objects.filter(field_key=field_key).order_by("sort_order", "value")
        )
    custom_fields = list(AssetCustomField.objects.all())
    default_label_template = _get_default_asset_label_template()
    type_template_map = {
        row.asset_type: row
        for row in AssetLabelTemplate.objects.filter(
            scope=AssetLabelTemplate.SCOPE_ASSET_TYPE,
            asset__isnull=True,
        )
    }
    asset_override_templates_qs = (
        AssetLabelTemplate.objects.filter(scope=AssetLabelTemplate.SCOPE_ASSET, asset__isnull=False)
        .select_related("asset")
        .order_by("asset__name", "asset__asset_tag")
    )
    asset_override_template_count = asset_override_templates_qs.count()
    asset_override_templates = list(asset_override_templates_qs[:30])
    asset_override_templates_limited = asset_override_template_count > len(asset_override_templates)
    asset_counts_by_type = {
        row["asset_type"]: row["n"]
        for row in Asset.objects.values("asset_type").annotate(n=Count("id")).order_by()
    }
    label_type_rows = []
    for asset_type_code, asset_type_label in Asset.TYPE_CHOICES:
        scoped_template = type_template_map.get(asset_type_code)
        effective_template = scoped_template or default_label_template
        label_type_rows.append(
            {
                "asset_type": asset_type_code,
                "asset_type_label": asset_type_label,
                "asset_count": asset_counts_by_type.get(asset_type_code, 0),
                "template": scoped_template,
                "effective_template": effective_template,
                "designer_url": reverse("assets:asset_label_designer") + f"?scope=asset_type&asset_type={asset_type_code}",
                "uses_default": scoped_template is None,
            }
        )
    sidebar_buttons = list(AssetSidebarButton.objects.select_related("parent").order_by("section", "sort_order", "label", "id"))
    sidebar_parent_choices = _sidebar_parent_choices()
    sidebar_target_suggestions, sidebar_active_match_suggestions = _sidebar_input_suggestions()
    for sidebar_item in sidebar_buttons:
        sidebar_item.label = _ui_label(sidebar_item.label)
    for parent_item in sidebar_parent_choices:
        parent_item.label = _ui_label(parent_item.label)

    # --- Azioni POST sulla configurazione ---
    if request.method == "POST":
        branding_response = handle_module_branding_post(
            request,
            module_key="assets",
            redirect_to=request.get_full_path() or f"{reverse('assets:gestione_admin')}?tab=config",
            audit_module="assets",
            legacy_logo_keys=("assets_logo_image",),
            sync_legacy_logo_keys=("assets_logo_image",),
            fallback_label="Assets",
        )
        if branding_response is not None:
            return branding_response
        action = request.POST.get("action")
        config_redirect = redirect(f"{reverse('assets:gestione_admin')}?tab=config")
        category_redirect = redirect(f"{reverse('assets:gestione_admin')}?tab=categorie")

        if action == "add_list_option":
            fk = request.POST.get("field_key", "").strip()
            val = request.POST.get("value", "").strip()
            if fk and val:
                AssetListOption.objects.get_or_create(field_key=fk, value=val)
                log_action(request, "add_list_option", "assets", {"field_key": fk, "value": val})
            return config_redirect

        if action == "delete_list_option":
            opt_id = _as_int(request.POST.get("opt_id"))
            if opt_id:
                AssetListOption.objects.filter(pk=opt_id).delete()
                log_action(request, "delete_list_option", "assets", {"opt_id": opt_id})
            return config_redirect

        if action == "add_custom_field":
            code = slugify(request.POST.get("code", ""))
            label = request.POST.get("label", "").strip()
            ftype = request.POST.get("field_type", AssetCustomField.TYPE_TEXT)
            if code and label:
                AssetCustomField.objects.get_or_create(code=code, defaults={"label": label, "field_type": ftype})
                log_action(request, "add_custom_field", "assets", {"code": code})
            return config_redirect

        if action == "delete_custom_field":
            cf_id = _as_int(request.POST.get("cf_id"))
            if cf_id:
                AssetCustomField.objects.filter(pk=cf_id).delete()
                log_action(request, "delete_custom_field", "assets", {"cf_id": cf_id})
            return config_redirect

        if action == "delete_label_template":
            template_id = _as_int(request.POST.get("template_id"))
            template = AssetLabelTemplate.objects.filter(pk=template_id).select_related("asset").first()
            if not template:
                messages.error(request, "Template etichetta non trovato.")
                return config_redirect
            if template.scope == AssetLabelTemplate.SCOPE_DEFAULT:
                messages.error(request, "Il template generale non puo essere eliminato.")
                return config_redirect
            scope_info = template.scope_display_label()
            template.delete()
            log_action(request, "delete_label_template", "assets", {"template_id": template_id, "scope": scope_info})
            messages.success(request, f"Template etichetta rimosso ({scope_info}).")
            return config_redirect

        if action in CATEGORY_ACTIONS:
            ok, text = _handle_asset_category_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return category_redirect

        if action in SIDEBAR_ACTIONS:
            ok, text = _handle_sidebar_button_request(request)
            if ok:
                messages.success(request, text)
            else:
                messages.error(request, text)
            return config_redirect

        if action == "save_assets_logo":
            logo_file = request.FILES.get("logo_file")
            logo_url = request.POST.get("logo_url", "").strip()
            if logo_file:
                _LOGO_ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".svg", ".webp"}
                _LOGO_ALLOWED_MIMES = {"image/png", "image/jpeg", "image/svg+xml", "image/webp"}
                if logo_file.size > 512 * 1024:
                    messages.error(request, "Immagine troppo grande (max 512 KB).")
                    return config_redirect
                try:
                    validate_extension_and_mime(
                        logo_file,
                        allowed_extensions=_LOGO_ALLOWED_EXTS,
                        allowed_mimes=_LOGO_ALLOWED_MIMES,
                        max_bytes=None,
                        label="Logo",
                    )
                except UploadMimeValidationError as exc:
                    messages.error(request, str(exc))
                    return config_redirect
                import os
                from django.core.files.storage import default_storage
                raw_ext = os.path.splitext(logo_file.name)[1].lower()
                ext = raw_ext if raw_ext in _LOGO_ALLOWED_EXTS else ".png"
                save_path = f"assets_logo/logo{ext}"
                if default_storage.exists(save_path):
                    default_storage.delete(save_path)
                saved = default_storage.save(save_path, logo_file)
                final_url = default_storage.url(saved)
                SiteConfig.set("assets_logo_image", final_url, "Logo personalizzato modulo Inventario Assets")
                log_action(request, "save_assets_logo", "assets", {"path": saved})
                messages.success(request, "Logo aggiornato.")
            elif logo_url:
                parsed = urlsplit(logo_url)
                if parsed.scheme not in ("http", "https"):
                    messages.error(request, "L'URL deve iniziare con http:// o https://")
                    return config_redirect
                SiteConfig.set("assets_logo_image", logo_url, "Logo personalizzato modulo Inventario Assets")
                log_action(request, "save_assets_logo_url", "assets", {"url": logo_url})
                messages.success(request, "Logo URL salvato.")
            else:
                messages.error(request, "Seleziona un file o inserisci un URL.")
            return config_redirect

        if action == "remove_assets_logo":
            SiteConfig.set("assets_logo_image", "", "Logo personalizzato modulo Inventario Assets")
            log_action(request, "remove_assets_logo", "assets", {})
            messages.success(request, "Logo personalizzato rimosso.")
            return config_redirect

        if action in ("add_categoria_ticket", "toggle_categoria_ticket", "delete_categoria_ticket"):
            from tickets.models import CategoriaTicket, Ticket as _Ticket, TipoTicket
            ticket_redirect = redirect(f"{reverse('assets:gestione_admin')}?tab=ticket")

            if action == "add_categoria_ticket":
                tipo = request.POST.get("cat_tipo", "").strip()
                codice = request.POST.get("cat_codice", "").strip().upper().replace(" ", "_")[:30]
                etichetta = request.POST.get("cat_etichetta", "").strip()[:100]
                ordine = _as_int(request.POST.get("cat_ordine"), default=0)
                if tipo in (TipoTicket.IT, TipoTicket.MAN) and codice and etichetta:
                    _, created = CategoriaTicket.objects.get_or_create(
                        codice=codice,
                        defaults={"tipo": tipo, "etichetta": etichetta, "ordine": ordine, "attivo": True},
                    )
                    if created:
                        log_action(request, "add_categoria_ticket", "tickets", {"tipo": tipo, "codice": codice})
                        messages.success(request, f"Categoria '{etichetta}' aggiunta.")
                    else:
                        messages.warning(request, f"Codice '{codice}' già esistente.")
                else:
                    messages.error(request, "Dati mancanti o non validi.")

            elif action == "toggle_categoria_ticket":
                cat_id = _as_int(request.POST.get("cat_id"))
                if cat_id:
                    cat = CategoriaTicket.objects.filter(pk=cat_id).first()
                    if cat:
                        cat.attivo = not cat.attivo
                        cat.save(update_fields=["attivo"])
                        log_action(request, "toggle_categoria_ticket", "tickets", {"id": cat_id, "attivo": cat.attivo})

            elif action == "delete_categoria_ticket":
                cat_id = _as_int(request.POST.get("cat_id"))
                if cat_id:
                    cat = CategoriaTicket.objects.filter(pk=cat_id).first()
                    if cat:
                        in_use = _Ticket.objects.filter(categoria=cat.codice).exists()
                        if in_use:
                            messages.error(request, f"Impossibile eliminare '{cat.etichetta}': è usata da ticket esistenti. Disattivala.")
                        else:
                            cat.delete()
                            log_action(request, "delete_categoria_ticket", "tickets", {"id": cat_id})
                            messages.success(request, "Categoria eliminata.")

            return ticket_redirect

    # --- Record / tab ---
    q_asset = request.GET.get("q_asset", "").strip()
    q_wo = request.GET.get("q_wo", "").strip()
    tab = request.GET.get("tab", "riepilogo")

    asset_categories = []
    asset_category_fields = []
    asset_category_rows = []
    if tab == "categorie":
        asset_categories = list(
            AssetCategory.objects.select_related("parent").prefetch_related("category_fields").order_by("sort_order", "label", "id")
        )
        asset_category_fields = list(
            AssetCategoryField.objects.select_related("category").order_by(
                "category__sort_order",
                "category__label",
                "sort_order",
                "label",
                "id",
            )
        )
        asset_category_rows = _build_asset_category_admin_rows(asset_categories)

    # --- Categorie ticket (solo tab ticket) ---
    from tickets.models import CategoriaTicket as _CategoriaTicket, TipoTicket as _TipoTicket
    categorie_ticket = (
        list(_CategoriaTicket.objects.order_by("tipo", "ordine", "etichetta"))
        if tab == "ticket" else []
    )

    assets_qs = Asset.objects.order_by("name")
    if q_asset:
        assets_qs = assets_qs.filter(Q(name__icontains=q_asset) | Q(asset_tag__icontains=q_asset))
    assets_page = Paginator(assets_qs, 50).get_page(request.GET.get("asset_page"))

    wo_qs = WorkOrder.objects.select_related("asset").order_by("-opened_at")
    if q_wo:
        wo_qs = wo_qs.filter(Q(title__icontains=q_wo) | Q(asset__name__icontains=q_wo))
    wo_page = Paginator(wo_qs, 50).get_page(request.GET.get("wo_page"))

    # --- Log ---
    audit_entries = AuditLog.objects.filter(modulo="assets").order_by("-created_at")[:100]

    return render(
        request,
        "assets/pages/gestione_admin.html",
        {
            **get_module_branding_context(
                "assets",
                fallback_label="Assets",
                legacy_logo_keys=("assets_logo_image",),
            ),
            "page_title": "Impostazioni Assets",
            "tab": tab,
            # stats
            "total_assets": total_assets,
            "assets_by_status": assets_by_status,
            "assets_by_type": assets_by_type,
            "total_wo": total_wo,
            "wo_by_status": wo_by_status,
            # config
            "list_options": list_options,
            "list_option_fields": AssetListOption.FIELD_CHOICES,
            "custom_fields": custom_fields,
            "cf_type_choices": AssetCustomField.TYPE_CHOICES,
            # records
            "assets_page": assets_page,
            "wo_page": wo_page,
            "q_asset": q_asset,
            "q_wo": q_wo,
            # log
            "audit_entries": audit_entries,
            "asset_status_labels": dict(Asset.STATUS_CHOICES),
            "asset_type_labels": dict(Asset.TYPE_CHOICES),
            "wo_status_labels": dict(WorkOrder.STATUS_CHOICES),
            "label_template": default_label_template,
            "label_template_default_designer_url": reverse("assets:asset_label_designer") + "?scope=default",
            "label_type_rows": label_type_rows,
            "asset_override_templates": asset_override_templates,
            "asset_override_templates_limited": asset_override_templates_limited,
            "label_template_override_count": asset_override_template_count,
            "sidebar_buttons": sidebar_buttons,
            "sidebar_parent_choices": sidebar_parent_choices,
            "sidebar_section_choices": _ui_choices(AssetSidebarButton.SECTION_CHOICES),
            "sidebar_target_suggestions": sidebar_target_suggestions,
            "sidebar_active_match_suggestions": sidebar_active_match_suggestions,
            "asset_categories": asset_categories,
            "asset_category_rows": asset_category_rows,
            "asset_category_fields": asset_category_fields,
            "asset_category_type_choices": _ui_choices(Asset.TYPE_CHOICES),
            "asset_category_field_type_choices": _ui_choices(AssetCategoryField.TYPE_CHOICES),
            "detail_section_choices": _ui_choices(AssetDetailField.SECTION_CHOICES),
            "detail_format_choices": _ui_choices(AssetDetailField.FORMAT_CHOICES),
            "asset_categories_active_count": sum(1 for row in asset_categories if row.is_active),
            "asset_category_fields_active_count": sum(1 for row in asset_category_fields if row.is_active),
            "categorie_ticket": categorie_ticket,
            "ticket_tipo_choices": _TipoTicket.choices,
            **_assets_shell_context(request),
        },
    )


@login_required
def asset_bulk_update(request: HttpRequest) -> JsonResponse:
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Metodo non consentito"}, status=405)
    # Scrittura di massa (stato, categoria, assegnazioni, note) su N asset: solo admin.
    # Endpoint JSON -> 403 JSON, mai redirect HTML.
    if not _is_assets_admin(request):
        log_action(
            request,
            "asset_bulk_update",
            "assets",
            {"esito": "denied", "motivo": "permission_denied"},
        )
        return JsonResponse({"ok": False, "error": "Permessi insufficienti."}, status=403)
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON non valido"}, status=400)

    ids = data.get("ids", [])
    fields = data.get("fields", {})

    if not ids:
        return JsonResponse({"ok": False, "error": "Nessun asset selezionato"}, status=400)
    if not fields:
        return JsonResponse({"ok": False, "error": "Nessun campo da aggiornare"}, status=400)

    _ALLOWED_BULK_FIELDS = {
        "status", "asset_type", "asset_category_id",
        "reparto", "manufacturer", "model",
        "assignment_to", "assignment_reparto", "assignment_location",
        "notes",
    }
    _valid_statuses = {k for k, _ in Asset.STATUS_CHOICES}
    _valid_asset_types = {k for k, _ in Asset.TYPE_CHOICES}

    update_kwargs: dict = {}
    for field, value in fields.items():
        if field not in _ALLOWED_BULK_FIELDS:
            continue
        if field == "status":
            if value not in _valid_statuses:
                return JsonResponse({"ok": False, "error": f"Stato non valido: {value}"}, status=400)
            update_kwargs[field] = str(value)
        elif field == "asset_type":
            if value not in _valid_asset_types:
                return JsonResponse({"ok": False, "error": f"Tipo asset non valido: {value}"}, status=400)
            update_kwargs[field] = str(value)
        elif field == "asset_category_id":
            str_value = str(value).strip()
            if str_value in ("", "0", "null", "None"):
                update_kwargs["asset_category_id"] = None
            else:
                try:
                    cat_id = int(str_value)
                except (ValueError, TypeError):
                    return JsonResponse({"ok": False, "error": "ID categoria non valido"}, status=400)
                if not AssetCategory.objects.filter(pk=cat_id).exists():
                    return JsonResponse({"ok": False, "error": "Categoria non trovata"}, status=400)
                update_kwargs["asset_category_id"] = cat_id
        else:
            update_kwargs[field] = str(value)

    if not update_kwargs:
        return JsonResponse({"ok": False, "error": "Nessun campo valido da aggiornare"}, status=400)

    try:
        clean_ids = [int(i) for i in ids]
    except (ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "ID asset non validi"}, status=400)

    updated = Asset.objects.filter(pk__in=clean_ids).update(**update_kwargs)
    log_action(
        request,
        "asset_bulk_update",
        "assets",
        {
            "asset_ids": clean_ids[:200],
            "asset_count": len(clean_ids),
            "campi": update_kwargs,
            "aggiornati": updated,
            "esito": "success",
        },
    )
    return JsonResponse({"ok": True, "updated": updated})


# ---------------------------------------------------------------------------
# Dashboard Assets
# ---------------------------------------------------------------------------

_WIDGET_META: dict[str, dict] = {
    "totale_asset": {"label": "Totale asset", "icon": "package", "color": "blue"},
    "asset_in_uso": {"label": "In uso", "icon": "check-circle", "color": "green"},
    "asset_in_riparazione": {"label": "In riparazione", "icon": "wrench", "color": "orange"},
    "scadenze_scadute": {"label": "Scadenze scadute", "icon": "alert-triangle", "color": "red"},
    "scadenze_30gg": {"label": "Scadenze 30 gg", "icon": "clock", "color": "yellow"},
    "scadenze_90gg": {"label": "Scadenze 90 gg", "icon": "calendar", "color": "slate"},
    "wo_aperte": {"label": "OdL aperte", "icon": "clipboard-list", "color": "blue"},
    "wo_chiuse_mese": {"label": "OdL chiuse (mese)", "icon": "clipboard-check", "color": "green"},
    "verifiche_scadute": {"label": "Verifiche scadute", "icon": "shield-alert", "color": "red"},
    "verifiche_30gg": {"label": "Verifiche 30 gg", "icon": "shield-check", "color": "yellow"},
    "asset_per_stato": {"label": "Asset per stato", "icon": "bar-chart", "color": "slate"},
    "asset_per_categoria": {"label": "Asset per categoria", "icon": "layers", "color": "slate"},
}


def _compute_dashboard_kpis(today: date) -> dict:
    """Calcola tutti i valori KPI per la dashboard assets."""
    from django.db.models import Min

    in_30 = today + timedelta(days=30)
    in_90 = today + timedelta(days=90)
    first_day_month = today.replace(day=1)

    total = Asset.objects.exclude(status=Asset.STATUS_RETIRED).count()
    in_uso = Asset.objects.filter(status=Asset.STATUS_IN_USE).count()
    in_repair = Asset.objects.filter(status=Asset.STATUS_IN_REPAIR).count()

    # Scadenze amministrative
    dl_qs = AssetAdministrativeDeadline.objects.filter(is_active=True)
    dl_scadute = dl_qs.filter(due_date__lt=today).count()
    dl_30 = dl_qs.filter(due_date__gte=today, due_date__lte=in_30).count()
    dl_90 = dl_qs.filter(due_date__gt=in_30, due_date__lte=in_90).count()

    # OdL
    wo_aperte = WorkOrder.objects.filter(status=WorkOrder.STATUS_OPEN).count()
    wo_chiuse_mese = WorkOrder.objects.filter(
        status=WorkOrder.STATUS_DONE, closed_at__date__gte=first_day_month
    ).count()

    # Verifiche periodiche
    pv_qs = PeriodicVerification.objects.filter(is_active=True, is_legacy=False, next_verification_date__isnull=False)
    pv_scadute = pv_qs.filter(next_verification_date__lt=today).count()
    pv_30 = pv_qs.filter(next_verification_date__gte=today, next_verification_date__lte=in_30).count()

    # Asset per stato
    status_map = dict(Asset.STATUS_CHOICES)
    per_stato_qs = Asset.objects.values("status").annotate(n=Count("id")).order_by("-n")
    per_stato = [
        {"status": r["status"], "label": status_map.get(r["status"], r["status"]), "count": r["n"]}
        for r in per_stato_qs
    ]

    # Asset per categoria (top 10 + "Altro")
    per_cat_qs = list(
        Asset.objects.values("asset_category__id", "asset_category__label")
        .annotate(n=Count("id"))
        .order_by("-n")[:12]
    )
    per_categoria = []
    for r in per_cat_qs:
        per_categoria.append({
            "id": r["asset_category__id"],
            "label": r["asset_category__label"] or "Senza categoria",
            "count": r["n"],
        })

    # Prossime scadenze (lista breve)
    prossime_scadenze = list(
        AssetAdministrativeDeadline.objects.filter(is_active=True, due_date__gte=today, due_date__lte=in_30)
        .select_related("asset")
        .order_by("due_date")[:8]
    )
    prossime_verifiche = list(
        PeriodicVerification.objects.filter(is_active=True, next_verification_date__gte=today, next_verification_date__lte=in_30)
        .order_by("next_verification_date")[:8]
    )
    scadenze_arretrate = list(
        AssetAdministrativeDeadline.objects.filter(is_active=True, due_date__lt=today)
        .select_related("asset")
        .order_by("due_date")[:8]
    )

    return {
        "totale_asset": total,
        "asset_in_uso": in_uso,
        "asset_in_riparazione": in_repair,
        "scadenze_scadute": dl_scadute,
        "scadenze_30gg": dl_30,
        "scadenze_90gg": dl_90,
        "wo_aperte": wo_aperte,
        "wo_chiuse_mese": wo_chiuse_mese,
        "verifiche_scadute": pv_scadute,
        "verifiche_30gg": pv_30,
        "asset_per_stato": per_stato,
        "asset_per_categoria": per_categoria,
        # dati lista per le card detail
        "prossime_scadenze": prossime_scadenze,
        "prossime_verifiche": prossime_verifiche,
        "scadenze_arretrate": scadenze_arretrate,
    }


def _legacy_asset_dashboard_list_redirect_url(request: HttpRequest) -> str:
    if request.method != "GET" or not request.GET:
        return ""
    legacy_list_query_keys = {"q", "asset_type", "reparto", "vlan", "ip", "rows", "page"}
    if not any(key in request.GET for key in legacy_list_query_keys):
        return ""
    query_string = request.GET.urlencode()
    target_url = reverse("assets:asset_list")
    return f"{target_url}?{query_string}" if query_string else target_url


@login_required
def asset_dashboard(request: HttpRequest) -> HttpResponse:
    """Dashboard principale del modulo Assets con KPI personalizzabili."""
    legacy_list_redirect = _legacy_asset_dashboard_list_redirect_url(request)
    if legacy_list_redirect:
        return redirect(legacy_list_redirect)

    today = timezone.localdate()
    selected_family_id = None
    selected_family_label = ""
    raw_family_id = request.GET.get("family")
    if raw_family_id:
        try:
            candidate_family_id = int(raw_family_id)
        except (TypeError, ValueError):
            candidate_family_id = None
        if candidate_family_id:
            selected_family = AssetCategory.objects.filter(pk=candidate_family_id, is_active=True).first()
            if selected_family:
                selected_family_id = selected_family.id
                selected_family_label = selected_family.label

    # Configurazione widget utente
    cfg, _ = AssetDashboardConfig.objects.get_or_create(user=request.user)
    enabled = cfg.get_enabled_widgets()

    kpis = _compute_dashboard_kpis(today)
    family_kpis = get_family_dashboard_kpis(selected_family_id, today=today)
    family_distribution = get_families_distribution(today=today)
    maintenance_by_family = get_maintenance_by_family(today=today)
    downtime_by_family = get_downtime_by_family(today=today)
    fire_safety_kpis = get_fire_safety_kpis(today=today)
    try:
        maintenance_perf = get_maintenance_performance_kpis(today=today)
    except Exception:
        maintenance_perf = {
            "mttr_hours": 0, "downtime_hours_month": 0, "maintenance_cost_month": 0,
            "wo_open_by_kind": {}, "wo_open_total": 0, "ticket_man_open": 0,
            "wo_closed_month": 0, "has_data": False,
        }

    # Cruscotto operativo (cose da fare + segnalazioni arrivate), condiviso con il Centro Manutenzione
    from assets.services.dashboard_kpi import get_cose_da_fare_overview, get_segnalazioni_overview
    cose_da_fare = get_cose_da_fare_overview(today=today)
    segnalazioni = get_segnalazioni_overview(today=today)

    # Categorie asset attive (solo principali, per i link in cima)
    categories = list(AssetCategory.objects.filter(is_active=True).order_by("sort_order", "label"))
    family_options = categories

    # Metadati widget arricchiti con valore
    widgets_all = []
    for code, meta in _WIDGET_META.items():
        widgets_all.append({
            "code": code,
            "label": meta["label"],
            "icon": meta["icon"],
            "color": meta["color"],
            "enabled": code in enabled,
            "order": enabled.index(code) if code in enabled else 999,
        })
    widgets_all.sort(key=lambda w: (0 if w["enabled"] else 1, w["order"]))

    branding = get_module_branding_context("assets")

    ctx = _assets_shell_context(request, search_action=reverse("assets:asset_list"))
    ctx.update({
        "today": today,
        "kpis": kpis,
        "enabled_widgets": enabled,
        "enabled_widgets_json": json.dumps(enabled),
        "widgets_all": widgets_all,
        "categories": categories,
        "selected_family_id": selected_family_id,
        "selected_family_label": selected_family_label,
        "family_options": family_options,
        "family_kpis": family_kpis,
        "family_distribution": family_distribution,
        "maintenance_by_family": maintenance_by_family,
        "downtime_by_family": downtime_by_family,
        "fire_safety_kpis": fire_safety_kpis,
        "maintenance_perf": maintenance_perf,
        "cose_da_fare": cose_da_fare,
        "segnalazioni": segnalazioni,
        "branding": branding,
        "page_title": "Dashboard Assets",
    })
    return render(request, "assets/pages/asset_dashboard.html", ctx)


@login_required
def api_asset_dashboard_save_config(request: HttpRequest) -> JsonResponse:
    """Salva la configurazione widget dashboard per l'utente corrente."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Metodo non consentito"}, status=405)
    try:
        payload = json.loads(request.body or b"{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "error": "Payload non valido"}, status=400)

    widgets = payload.get("widgets")
    if not isinstance(widgets, list):
        return JsonResponse({"ok": False, "error": "Campo widgets obbligatorio (lista)"}, status=400)

    valid = set(AssetDashboardConfig.DEFAULT_WIDGETS)
    clean = [w for w in widgets if isinstance(w, str) and w in valid]

    cfg, _ = AssetDashboardConfig.objects.get_or_create(user=request.user)
    cfg.enabled_widgets = clean
    cfg.save(update_fields=["enabled_widgets", "updated_at"])
    return JsonResponse({"ok": True, "saved": len(clean)})


# ---------------------------------------------------------------------------
# Calendario Asset (lavori macchina + manutenzioni)
# ---------------------------------------------------------------------------

def _asset_calendar_events(asset_id: int) -> list[dict]:
    """Ritorna tutti gli eventi (lavori + manutenzioni) per un singolo asset."""
    events: list[dict] = []
    today = timezone.localdate()

    # -- Lavori macchina (Task con categoria is_machine_work=True) --
    try:
        from tasks.models import TaskExtraRef, TaskStatus
        refs = (
            TaskExtraRef.objects
            .filter(asset_id=asset_id, task__category__is_machine_work=True)
            .exclude(task__status__in=[TaskStatus.DONE, TaskStatus.CANCELED])
            .select_related("task__category", "task__assigned_to", "task__project")
        )
        for ref in refs:
            t = ref.task
            start = t.next_step_due or t.due_date
            end = t.due_date or t.next_step_due
            if not start:
                continue
            assignee = ""
            if t.assigned_to_id:
                assignee = t.assigned_to.get_full_name() or t.assigned_to.get_username()
            events.append({
                "id": f"task-{t.pk}",
                "title": t.title,
                "start": str(start),
                "end": str(end) if end else str(start),
                "kind": "lavoro",
                "status": t.status,
                "assignee": assignee,
                "project": t.project.name if t.project_id else "",
                "url": f"/tasks/detail/{t.pk}/",
                "color": "#2563eb" if t.status == TaskStatus.IN_PROGRESS else "#93c5fd",
                "textColor": "#fff",
            })
    except Exception:
        pass

    # -- Manutenzioni (WorkOrder aperti) --
    try:
        wos = (
            WorkOrder.objects
            .filter(asset_id=asset_id, status=WorkOrder.STATUS_OPEN)
            .select_related("asset")
        )
        for wo in wos:
            opened = wo.opened_at.date() if wo.opened_at else today
            events.append({
                "id": f"wo-{wo.pk}",
                "title": f"OdL: {wo.title}",
                "start": str(opened),
                "end": str(opened),
                "kind": "manutenzione",
                "status": wo.status,
                "assignee": "",
                "project": "",
                "url": f"/assets/workorders/view/{wo.pk}/",
                "color": "#f59e0b",
                "textColor": "#fff",
            })
    except Exception:
        pass

    # -- Verifiche periodiche pianificate --
    try:
        verifications = (
            AssetCalendarEvent.objects
            .filter(asset_id=asset_id, due_date__gte=today)
            .order_by("due_date")
        )
        for v in verifications:
            events.append({
                "id": f"ace-{v.pk}",
                "title": v.source_object_label or v.subject or "Scadenza",
                "start": str(v.due_date),
                "end": str(v.due_date),
                "kind": "scadenza",
                "status": "",
                "assignee": v.target_display_name,
                "project": "",
                "url": "",
                "color": "#10b981",
                "textColor": "#fff",
            })
    except Exception:
        pass

    # -- Manutenzioni programmate predette (regole a giorni con prossima scadenza) --
    try:
        from .maintenance import build_maintenance_schedule_rows
        _pm_colors = {"overdue": "#dc2626", "warning": "#f59e0b", "upcoming": "#10b981", "missing": "#94a3b8"}
        for row in build_maintenance_schedule_rows(
            asset_queryset=Asset.objects.filter(pk=asset_id).select_related("asset_category")
        ):
            due = row.get("due_date")
            if not due:
                continue
            template = row.get("effective_intervention_template")
            label = getattr(template, "label", "") or "Manutenzione programmata"
            status = str(row.get("schedule_status") or "")
            base_rule = row["base_rule"]
            is_external = base_rule.is_external
            # Riempimento = urgenza (stato); bordo viola + marcatore 🏢 = manutenzione esterna (ditta terza).
            pm_event = {
                "id": f"pm-{base_rule.id}",
                "title": f"{'🏢 ' if is_external else ''}Manut.: {label}",
                "start": str(due),
                "end": str(due),
                "kind": "manutenzione_prog",
                "status": status,
                "assignee": str(base_rule.supplier) if (is_external and base_rule.supplier_id) else "",
                "project": "",
                "url": "",
                "color": _pm_colors.get(status, "#10b981"),
                "textColor": "#fff",
            }
            if is_external:
                pm_event["borderColor"] = "#6d28d9"
                pm_event["external"] = True
            events.append(pm_event)
    except Exception:
        pass

    return events


@login_required
def asset_calendar_json(request: HttpRequest, id: int) -> JsonResponse:
    """JSON eventi calendario per un singolo asset (usato da FullCalendar)."""
    asset = get_object_or_404(Asset, pk=id)
    events = _asset_calendar_events(asset.pk)
    return JsonResponse({"ok": True, "asset_id": asset.pk, "events": events})


@login_required
def calendario_asset(request: HttpRequest) -> HttpResponse:
    """Pagina globale Calendario Asset: vista a calendario o a Gantt."""
    machines = list(
        Asset.objects
        .filter(asset_type__in=[Asset.TYPE_WORK_MACHINE, Asset.TYPE_CNC])
        .order_by("reparto", "name", "asset_tag")
    )
    reparti = sorted({m.reparto for m in machines if m.reparto})
    return render(request, "assets/pages/calendario_asset.html", {
        "machines": machines,
        "reparti": reparti,
        "page_title": "Calendario Asset",
        **_assets_shell_context(request),
    })


@login_required
def calendario_asset_json(request: HttpRequest) -> JsonResponse:
    """JSON eventi per tutte le macchine (usato dalla pagina calendario globale)."""
    machines = Asset.objects.filter(
        asset_type__in=[Asset.TYPE_WORK_MACHINE, Asset.TYPE_CNC]
    ).values_list("pk", "asset_tag", "name", "reparto")

    all_events: list[dict] = []
    resources: list[dict] = []
    for pk, tag, name, reparto in machines:
        resources.append({
            "id": str(pk),
            "title": f"{tag} – {name}",
            "tag": tag,
            "name": name,
            "reparto": reparto or "",
        })
        for ev in _asset_calendar_events(pk):
            ev["resourceId"] = str(pk)
            all_events.append(ev)

    return JsonResponse({"ok": True, "resources": resources, "events": all_events})
