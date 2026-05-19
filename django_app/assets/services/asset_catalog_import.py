from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from assets.models import Asset, AssetCategory


ASSET_CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9._-]{1,31}$")

COLUMN_ALIASES = {
    "asset_id": {"asset id", "asset_id", "assetid", "codice asset", "codice", "asset tag", "asset_tag", "id asset"},
    "famiglia": {"famiglia", "family", "categoria", "categoria padre"},
    "sottocategoria": {"sottocategoria", "sotto categoria", "subcategory", "tipo asset", "tipo", "tipologia"},
    "descrizione": {"descrizione", "description", "desc"},
    "nome": {"nome", "name", "asset name", "denominazione"},
    "ubicazione": {"ubicazione", "location", "posizione", "sede", "reparto", "assignment location"},
    "matricola": {"matricola", "seriale", "serial", "serial number", "s/n", "sn"},
    "stato": {"stato", "status"},
    "manufacturer": {"produttore", "manufacturer", "marca", "costruttore", "brand"},
    "model": {"modello", "model"},
}


def normalize_spaces(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_key(value: Any) -> str:
    text = normalize_spaces(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_key(value: Any) -> str:
    return normalize_key(value).replace(" ", "")


# Regole euristiche nome categoria/sottocategoria -> tipo asset.
# Ordine significativo: vince la prima voce il cui keyword compare nel nome
# compattato (minuscolo, senza spazi ne accenti). Le voci piu specifiche
# vanno prima di quelle generiche (es. "macchinavirtuale" prima di "macchina").
# I keyword usano la radice (stem) per intercettare singolare e plurale sul
# nome compattato: es. "portatil" copre "portatile"/"portatili".
_ASSET_TYPE_KEYWORD_RULES: tuple[tuple[str, str], ...] = (
    ("carropont", Asset.TYPE_CARROPONTE),
    ("carripont", Asset.TYPE_CARROPONTE),
    ("videosorveglianz", Asset.TYPE_CCTV),
    ("telecamer", Asset.TYPE_CCTV),
    ("tvcc", Asset.TYPE_CCTV),
    ("cctv", Asset.TYPE_CCTV),
    ("firewall", Asset.TYPE_FIREWALL),
    ("switch", Asset.TYPE_FIREWALL),
    ("router", Asset.TYPE_FIREWALL),
    ("accesspoint", Asset.TYPE_FIREWALL),
    ("rete", Asset.TYPE_FIREWALL),
    ("reti", Asset.TYPE_FIREWALL),
    ("server", Asset.TYPE_SERVER),
    ("virtual", Asset.TYPE_VM),
    ("hypervisor", Asset.TYPE_VM),
    ("cnc", Asset.TYPE_CNC),
    ("utensil", Asset.TYPE_WORK_MACHINE),
    ("tornio", Asset.TYPE_WORK_MACHINE),
    ("torni", Asset.TYPE_WORK_MACHINE),
    ("fresa", Asset.TYPE_WORK_MACHINE),
    ("frese", Asset.TYPE_WORK_MACHINE),
    ("rettific", Asset.TYPE_WORK_MACHINE),
    ("macchin", Asset.TYPE_WORK_MACHINE),
    ("portatil", Asset.TYPE_NOTEBOOK),
    ("notebook", Asset.TYPE_NOTEBOOK),
    ("laptop", Asset.TYPE_NOTEBOOK),
    ("stampant", Asset.TYPE_STAMPANTE),
    ("plotter", Asset.TYPE_STAMPANTE),
    ("multifunzion", Asset.TYPE_STAMPANTE),
    ("fonia", Asset.TYPE_FONIA),
    ("telefon", Asset.TYPE_FONIA),
    ("centralin", Asset.TYPE_FONIA),
    ("voip", Asset.TYPE_FONIA),
    ("desktop", Asset.TYPE_PC),
    ("workstation", Asset.TYPE_PC),
    ("computer", Asset.TYPE_PC),
    ("pc", Asset.TYPE_PC),
)


def classify_asset_type(label: str) -> str:
    """Deduce un tipo asset dal nome categoria/sottocategoria.

    Euristica a parole chiave: ritorna ``Asset.TYPE_OTHER`` se nessuna regola
    corrisponde. Pensata per import e riallineamento, non come fonte
    autoritativa: l'esito va sempre verificato (es. con ``--dry-run``).
    """
    key = compact_key(label)
    if not key:
        return Asset.TYPE_OTHER
    for keyword, asset_type in _ASSET_TYPE_KEYWORD_RULES:
        if keyword in key:
            return asset_type
    return Asset.TYPE_OTHER


def category_code(label: str, parent: AssetCategory | None = None) -> str:
    base = slugify(normalize_key(label))[:60] or "categoria"
    if parent is not None:
        prefix = parent.code[: max(1, 79 - len(base) - 1)]
        return f"{prefix}-{base}"[:80]
    return base[:80]


def source_key_for_row(row: dict[str, str]) -> str:
    parts = [
        compact_key(row.get("famiglia")),
        compact_key(row.get("sottocategoria")),
        compact_key(row.get("asset_id")),
        compact_key(row.get("matricola")),
        compact_key(row.get("nome")),
        compact_key(row.get("descrizione")),
        compact_key(row.get("ubicazione")),
    ]
    raw = "|".join(parts)
    digest = hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()
    return f"asset-catalog:{digest}"


class AssetCodeGenerator:
    def validate(self, value: str) -> str:
        code = normalize_spaces(value).upper()
        if not code:
            raise ValueError("asset_id vuoto")
        if not ASSET_CODE_RE.match(code):
            raise ValueError("asset_id non valido: usa 2-32 caratteri A-Z, 0-9, punto, underscore o trattino")
        return code

    def generate(self, row: dict[str, str]) -> str:
        family = compact_key(row.get("famiglia")).upper()[:3] or "AST"
        subcategory = compact_key(row.get("sottocategoria")).upper()[:3] or "GEN"
        matricola = re.sub(r"[^A-Z0-9]+", "", normalize_spaces(row.get("matricola")).upper())
        if matricola:
            base = f"A{family}{subcategory}-{matricola[:18]}"
        else:
            digest = hashlib.sha1(source_key_for_row(row).encode("utf-8", errors="ignore")).hexdigest()
            fingerprint = int(digest[:10], 16) % 1_000_000
            base = f"A{family}{subcategory}-{fingerprint:06d}"
        return self._unique(base[:32])

    def _unique(self, base: str) -> str:
        candidate = self.validate(base)
        if not Asset.objects.filter(asset_tag__iexact=candidate).exists():
            return candidate
        for suffix in range(2, 1000):
            tail = f"-{suffix}"
            candidate = self.validate(f"{base[: 32 - len(tail)]}{tail}")
            if not Asset.objects.filter(asset_tag__iexact=candidate).exists():
                return candidate
        raise ValueError(f"impossibile generare un asset_id univoco da {base}")


@dataclass
class RowError:
    # Localizzatore riga gia descrittivo (es. "riga 5" oppure "foglio 'IT' riga 5").
    row_number: str
    message: str


@dataclass
class ImportStats:
    rows_seen: int = 0
    rows_valid: int = 0
    parent_categories_to_create: int = 0
    subcategories_to_create: int = 0
    assets_to_create: int = 0
    assets_to_update: int = 0
    assets_unchanged: int = 0


@dataclass
class ImportPreview:
    stats: ImportStats = field(default_factory=ImportStats)
    errors: list[RowError] = field(default_factory=list)

    @property
    def has_blocking_errors(self) -> bool:
        return bool(self.errors)


class AssetCatalogImporter:
    def __init__(self, code_generator: AssetCodeGenerator | None = None):
        self.code_generator = code_generator or AssetCodeGenerator()

    def load_rows(self, file_path: str | Path) -> list[tuple[str, dict[str, str]]]:
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._load_xlsx(path)
        raise ValueError("Formato non supportato: usa .csv o .xlsx")

    def preview(self, file_path: str | Path) -> ImportPreview:
        return self._plan(self.load_rows(file_path))

    @transaction.atomic
    def commit(self, file_path: str | Path) -> ImportPreview:
        loaded_rows = self.load_rows(file_path)
        preview = self._plan(loaded_rows)
        if preview.has_blocking_errors:
            return preview

        for _, row in loaded_rows:
            normalized = self._normalize_row(row)
            if not self._row_is_valid(normalized):
                continue
            parent = self._get_or_create_parent(normalized["famiglia"])
            subcategory = self._get_or_create_subcategory(normalized["sottocategoria"], parent)
            asset_id = normalized["asset_id"] or self.code_generator.generate(normalized)
            source_key = source_key_for_row(normalized)
            payload = self._asset_payload(normalized, subcategory, asset_id, source_key)
            asset = Asset.objects.filter(asset_tag__iexact=asset_id).first()
            if asset is None:
                asset = Asset.objects.filter(source_key=source_key).first()
            if asset is None:
                Asset.objects.create(**payload)
            else:
                self._update_asset(asset, payload)
        return preview

    def _plan(self, loaded_rows: list[tuple[str, dict[str, str]]]) -> ImportPreview:
        preview = ImportPreview()
        parent_ids_by_key = {
            normalize_key(label): category_id
            for category_id, label in AssetCategory.objects.filter(parent__isnull=True).values_list("id", "label")
        }
        parent_keys = set(parent_ids_by_key)
        child_keys = {
            (parent_id, normalize_key(label))
            for parent_id, label in AssetCategory.objects.filter(parent__isnull=False).values_list("parent_id", "label")
        }
        staged_parent_ids: dict[str, int] = {}
        staged_parent_counter = -1
        seen_asset_tags = {tag.upper() for tag in Asset.objects.values_list("asset_tag", flat=True)}
        seen_source_keys = set(Asset.objects.exclude(source_key__isnull=True).values_list("source_key", flat=True))
        staged_asset_tags: set[str] = set()
        staged_source_keys: set[str] = set()

        for row_number, raw_row in loaded_rows:
            if all(not normalize_spaces(value) for value in raw_row.values()):
                continue
            preview.stats.rows_seen += 1
            row = self._normalize_row(raw_row)
            row_errors = self._validate_row(row_number, row)
            if row_errors:
                preview.errors.extend(row_errors)
                continue
            preview.stats.rows_valid += 1

            parent_key = normalize_key(row["famiglia"])
            if parent_key not in parent_keys:
                parent_keys.add(parent_key)
                staged_parent_ids[parent_key] = staged_parent_counter
                staged_parent_counter -= 1
                preview.stats.parent_categories_to_create += 1

            parent_id = staged_parent_ids.get(parent_key, parent_ids_by_key.get(parent_key, 0))
            child_key = (int(parent_id or 0), normalize_key(row["sottocategoria"]))
            if child_key not in child_keys:
                child_keys.add(child_key)
                preview.stats.subcategories_to_create += 1

            source_key = source_key_for_row(row)
            asset_id = row["asset_id"]
            if asset_id:
                asset_id = self.code_generator.validate(asset_id)
                if asset_id in staged_asset_tags:
                    preview.errors.append(RowError(row_number, f"asset_id duplicato nel file: {asset_id}"))
                    continue
                staged_asset_tags.add(asset_id)
            if source_key in staged_source_keys and not asset_id:
                preview.errors.append(RowError(row_number, "riga duplicata nel file senza asset_id"))
                continue
            staged_source_keys.add(source_key)

            if asset_id and asset_id in seen_asset_tags:
                preview.stats.assets_to_update += 1
            elif source_key in seen_source_keys:
                preview.stats.assets_to_update += 1
            else:
                preview.stats.assets_to_create += 1
        return preview

    def _validate_row(self, row_number: int, row: dict[str, str]) -> list[RowError]:
        errors: list[RowError] = []
        if not row["famiglia"]:
            errors.append(RowError(row_number, "famiglia obbligatoria mancante"))
        if not row["sottocategoria"]:
            errors.append(RowError(row_number, "sottocategoria obbligatoria mancante"))
        if row["asset_id"]:
            try:
                row["asset_id"] = self.code_generator.validate(row["asset_id"])
            except ValueError as exc:
                errors.append(RowError(row_number, str(exc)))
        return errors

    def _row_is_valid(self, row: dict[str, str]) -> bool:
        return bool(row.get("famiglia") and row.get("sottocategoria"))

    def _normalize_row(self, raw_row: dict[str, str]) -> dict[str, Any]:
        return {
            "asset_id": normalize_spaces(raw_row.get("asset_id")).upper(),
            "famiglia": normalize_spaces(raw_row.get("famiglia")),
            "sottocategoria": normalize_spaces(raw_row.get("sottocategoria")),
            "descrizione": normalize_spaces(raw_row.get("descrizione")),
            "nome": normalize_spaces(raw_row.get("nome")),
            "ubicazione": normalize_spaces(raw_row.get("ubicazione")),
            "matricola": normalize_spaces(raw_row.get("matricola")),
            "stato": normalize_spaces(raw_row.get("stato")),
            "manufacturer": normalize_spaces(raw_row.get("manufacturer")),
            "model": normalize_spaces(raw_row.get("model")),
            "extra": {
                key[len("extra:") :]: normalize_spaces(value)
                for key, value in raw_row.items()
                if key.startswith("extra:") and normalize_spaces(value)
            },
        }

    def _asset_payload(
        self,
        row: dict[str, Any],
        subcategory: AssetCategory,
        asset_id: str,
        source_key: str,
    ) -> dict[str, Any]:
        is_generic_cn = self._is_generic_cn(asset_id)
        notes = "Asset generico CN" if is_generic_cn else ""
        extra_columns: dict[str, Any] = {"is_generic_asset": True} if is_generic_cn else {}
        # Le colonne non standard del file confluiscono in extra_columns.
        extra_columns.update(row.get("extra") or {})
        description = row["descrizione"]
        return {
            "asset_tag": asset_id,
            "name": (row["nome"] or description or asset_id)[:255],
            "asset_type": self._asset_type_for_subcategory(row["sottocategoria"]),
            "asset_category": subcategory,
            "reparto": row["ubicazione"][:120],
            "manufacturer": row["manufacturer"][:120] or None,
            "model": row["model"][:120] or None,
            "serial_number": row["matricola"][:120] or None,
            "status": self._status(row["stato"]),
            "notes": notes,
            "extra_columns": extra_columns,
            "source_key": source_key,
        }

    def _update_asset(self, asset: Asset, payload: dict[str, Any]) -> None:
        changed: list[str] = []
        for field_name in [
            "name",
            "asset_type",
            "asset_category",
            "reparto",
            "manufacturer",
            "model",
            "serial_number",
            "status",
            "source_key",
        ]:
            value = payload[field_name]
            if getattr(asset, field_name) != value:
                setattr(asset, field_name, value)
                changed.append(field_name)
        if payload["notes"] and asset.notes != payload["notes"]:
            asset.notes = payload["notes"]
            changed.append("notes")
        merged_extra = dict(asset.extra_columns or {})
        for key, value in (payload["extra_columns"] or {}).items():
            if merged_extra.get(key) != value:
                merged_extra[key] = value
                changed.append("extra_columns")
        if "extra_columns" in changed:
            asset.extra_columns = merged_extra
        if changed:
            asset.save(update_fields=sorted(set(changed + ["updated_at"])))

    def _get_or_create_parent(self, label: str) -> AssetCategory:
        normalized_label = normalize_key(label)
        for category in AssetCategory.objects.filter(parent__isnull=True):
            if normalize_key(category.label) == normalized_label:
                return category
        code = category_code(label)
        return AssetCategory.objects.get_or_create(
            code=code,
            defaults={"label": label, "parent": None, "base_asset_type": Asset.TYPE_OTHER},
        )[0]

    def _get_or_create_subcategory(self, label: str, parent: AssetCategory) -> AssetCategory:
        normalized_label = normalize_key(label)
        for category in AssetCategory.objects.filter(parent=parent):
            if normalize_key(category.label) == normalized_label:
                return category
        code = category_code(label, parent)
        return AssetCategory.objects.get_or_create(
            code=code,
            defaults={"label": label, "parent": parent, "base_asset_type": self._asset_type_for_subcategory(label)},
        )[0]

    def _asset_type_for_subcategory(self, label: str) -> str:
        return classify_asset_type(label)

    def _status(self, value: str) -> str:
        key = normalize_key(value)
        if not key or "uso" in key or "attiv" in key:
            return Asset.STATUS_IN_USE
        if "magazz" in key or "stock" in key:
            return Asset.STATUS_IN_STOCK
        if "ripar" in key:
            return Asset.STATUS_IN_REPAIR
        if "dism" in key or "ritir" in key:
            return Asset.STATUS_RETIRED
        return Asset.STATUS_IN_USE

    def _is_generic_cn(self, asset_id: str) -> bool:
        return bool(re.match(r"^CN-[A-Z0-9]+$", asset_id))

    def _load_xlsx(self, path: Path) -> list[tuple[str, dict[str, str]]]:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        loaded: list[tuple[str, dict[str, str]]] = []
        try:
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = self._map_headers(rows[0])
                if not headers:
                    # Foglio senza intestazioni riconoscibili (es. legenda): saltato.
                    continue
                for idx, values in enumerate(rows[1:], start=2):
                    row = {
                        canonical: normalize_spaces(values[col_idx]) if col_idx < len(values) else ""
                        for col_idx, canonical in headers.items()
                    }
                    loaded.append((f"foglio '{sheet_name}' riga {idx}", row))
        finally:
            workbook.close()
        return loaded

    def _load_csv(self, path: Path) -> list[tuple[str, dict[str, str]]]:
        raw = path.read_bytes()
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Encoding CSV non riconosciuto: provati UTF-8 e cp1252")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.reader(text.splitlines(), dialect)
        try:
            header_row = next(reader)
        except StopIteration:
            return []
        headers = self._map_headers(header_row)
        loaded = []
        for idx, values in enumerate(reader, start=2):
            loaded.append(
                (f"riga {idx}", {canonical: normalize_spaces(values[col_idx]) if col_idx < len(values) else "" for col_idx, canonical in headers.items()})
            )
        return loaded

    def _map_headers(self, headers: Iterable[Any]) -> dict[int, str]:
        alias_to_canonical = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                alias_to_canonical[compact_key(alias)] = canonical
        mapped: dict[int, str] = {}
        used_canonical: set[str] = set()
        for idx, header in enumerate(headers):
            label = normalize_spaces(header)
            if not label:
                continue
            canonical = alias_to_canonical.get(compact_key(header))
            if canonical:
                if canonical not in used_canonical:
                    mapped[idx] = canonical
                    used_canonical.add(canonical)
            else:
                # Colonna non standard: il valore confluisce in extra_columns.
                mapped[idx] = f"extra:{label}"
        return mapped
