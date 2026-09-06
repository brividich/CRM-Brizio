"""Import iniziale dello storico manutenzioni (specifica §34 / §62).

Il nuovo motore calcola la prossima scadenza a partire dall'ultima esecuzione. Al
primo avvio quella data non esiste da nessuna parte: senza storico ogni piano
risulterebbe "dovuto subito" e il portale aprirebbe centinaia di scadenze false il
giorno del passaggio. Questo modulo la importa da un foglio Excel/CSV.

Quattro colonne, niente di piu':

    asset | piano | ultima esecuzione | note

Per ogni riga valida crea due occorrenze:
  - una **eseguita**, datata all'ultima esecuzione dichiarata (e' lo storico);
  - la **prossima aperta**, calcolata dalla periodicita' dell'applicazione secondo
    l'ancoraggio del piano.

Il parsing e la validazione sono separati dalla scrittura: la stessa funzione
alimenta l'anteprima della pagina e il management command, cosi' cio' che l'utente
vede in anteprima e' esattamente cio' che verra' scritto.
"""

from __future__ import annotations

import csv
import io
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterable

from django.db import transaction

from assets.models import (
    Asset,
    MaintenanceInterventionTemplate,
    MaintenanceOccurrence,
)
from assets.services.maintenance_domain import build_plan_resolutions
from assets.services.recurrence import compute_next_due

# Intestazioni accettate per ogni colonna, normalizzate (minuscole, senza accenti).
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "asset": ("asset", "asset tag", "asset_tag", "matricola", "targa", "codice asset"),
    "plan": ("piano", "manutenzione", "piano di manutenzione", "intervento"),
    "last_execution": (
        "ultima esecuzione",
        "ultima_esecuzione",
        "data ultima esecuzione",
        "eseguita il",
        "data",
    ),
    "notes": ("note", "annotazioni", "commento"),
}

TEMPLATE_HEADERS = ["asset", "piano", "ultima esecuzione", "note"]

OUTCOME_CREATE = "create"
OUTCOME_DUPLICATE = "duplicate"
OUTCOME_ERROR = "error"


def _normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class ImportRow:
    """Una riga del foglio, con l'esito che avrebbe (o che ha avuto)."""

    number: int
    asset_tag: str
    plan_label: str
    raw_date: str
    notes: str
    outcome: str = OUTCOME_ERROR
    error: str = ""
    asset: Asset | None = None
    plan: MaintenanceInterventionTemplate | None = None
    assignment: Any = None
    last_execution: date | None = None
    next_due: date | None = None
    recurrence_label: str = ""

    @property
    def is_valid(self) -> bool:
        return self.outcome != OUTCOME_ERROR


@dataclass
class ImportReport:
    rows: list[ImportRow] = field(default_factory=list)
    header_error: str = ""
    created_history: int = 0
    created_next: int = 0
    kept_open: int = 0
    applied: bool = False

    @property
    def valid_rows(self) -> list[ImportRow]:
        return [row for row in self.rows if row.outcome == OUTCOME_CREATE]

    @property
    def duplicate_rows(self) -> list[ImportRow]:
        return [row for row in self.rows if row.outcome == OUTCOME_DUPLICATE]

    @property
    def error_rows(self) -> list[ImportRow]:
        return [row for row in self.rows if row.outcome == OUTCOME_ERROR]

    @property
    def can_apply(self) -> bool:
        return not self.header_error and bool(self.valid_rows)


# ---------------------------------------------------------------------------
# Lettura del file
# ---------------------------------------------------------------------------

def _read_xlsx(stream) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(stream, data_only=True, read_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _read_csv(raw: bytes) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - una delle tre decodifiche riesce sempre
        text = raw.decode("utf-8", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def read_table(uploaded_file) -> list[list[Any]]:
    """Righe grezze da un .xlsx o da un .csv, senza interpretarle."""
    name = _normalize(getattr(uploaded_file, "name", ""))
    uploaded_file.seek(0)
    if name.endswith(".csv") or name.endswith(".txt"):
        return _read_csv(uploaded_file.read())
    return _read_xlsx(uploaded_file)


def _map_columns(header: Iterable[Any]) -> dict[str, int]:
    normalized = [_normalize(cell) for cell in header]
    mapping: dict[str, int] = {}
    for key, aliases in COLUMN_ALIASES.items():
        for index, cell in enumerate(normalized):
            if cell in aliases:
                mapping[key] = index
                break
    return mapping


# ---------------------------------------------------------------------------
# Analisi
# ---------------------------------------------------------------------------

def analyze(table: list[list[Any]], *, today: date | None = None) -> ImportReport:
    """Valida il foglio riga per riga. Non scrive nulla."""
    from django.utils import timezone

    today = today or timezone.localdate()
    report = ImportReport()

    if not table:
        report.header_error = "Il file è vuoto."
        return report

    columns = _map_columns(table[0])
    missing = [key for key in ("asset", "plan", "last_execution") if key not in columns]
    if missing:
        attesi = ", ".join(TEMPLATE_HEADERS)
        report.header_error = (
            "Intestazioni non riconosciute: la prima riga deve contenere le colonne "
            f"«{attesi}». Scarica il modello e ricompilalo."
        )
        return report

    def cell(row: list[Any], key: str) -> Any:
        index = columns.get(key)
        if index is None or index >= len(row):
            return ""
        return row[index]

    # Prima passata: parsing puro, senza query.
    parsed: list[ImportRow] = []
    for number, raw_row in enumerate(table[1:], start=2):
        if not any(str(value or "").strip() for value in raw_row):
            continue
        parsed.append(
            ImportRow(
                number=number,
                asset_tag=str(cell(raw_row, "asset") or "").strip(),
                plan_label=str(cell(raw_row, "plan") or "").strip(),
                raw_date=str(cell(raw_row, "last_execution") or "").strip(),
                notes=str(cell(raw_row, "notes") or "").strip(),
                last_execution=_parse_date(cell(raw_row, "last_execution")),
            )
        )

    if not parsed:
        report.header_error = "Il file non contiene righe da importare."
        return report

    # Lookup in blocco: un file di 800 righe non deve fare 2400 query.
    tags = {_normalize(row.asset_tag) for row in parsed if row.asset_tag}
    assets = {
        _normalize(asset.asset_tag): asset
        for asset in Asset.objects.filter(asset_tag__isnull=False)
        if _normalize(asset.asset_tag) in tags
    }
    plans_by_key: dict[str, MaintenanceInterventionTemplate] = {}
    for plan in MaintenanceInterventionTemplate.objects.all():
        plans_by_key.setdefault(_normalize(plan.code), plan)
        plans_by_key.setdefault(_normalize(plan.label), plan)

    resolutions = build_plan_resolutions(
        asset_queryset=Asset.objects.filter(pk__in=[a.pk for a in assets.values()])
    )
    existing = {
        (plan_id, asset_id, due)
        for plan_id, asset_id, due in MaintenanceOccurrence.objects.filter(
            asset_id__in=[a.pk for a in assets.values()]
        ).values_list("plan_id", "asset_id", "due_date")
    }

    seen: set[tuple[int, int, date]] = set()
    for row in parsed:
        row.asset = assets.get(_normalize(row.asset_tag))
        row.plan = plans_by_key.get(_normalize(row.plan_label))

        if row.asset is None:
            row.error = f"Asset «{row.asset_tag}» non trovato."
            continue
        if row.plan is None:
            row.error = f"Piano «{row.plan_label}» non trovato."
            continue
        if row.last_execution is None:
            row.error = (
                f"Data «{row.raw_date}» non leggibile: usare il formato gg/mm/aaaa."
                if row.raw_date
                else "Manca la data dell'ultima esecuzione."
            )
            continue
        if row.last_execution > today:
            row.error = "L'ultima esecuzione è nel futuro."
            continue

        resolution = resolutions.get((row.plan.pk, row.asset.pk))
        # L'ordine conta: in conflitto e in esclusione l'applicazione risolta e' None,
        # e senza questo controllo prima l'utente leggerebbe "non si applica" davanti a
        # un piano che invece si applica fin troppo, da due gruppi diversi.
        if resolution is not None and resolution.is_conflict:
            row.error = "Periodicità in conflitto: va risolta prima, l'import non sceglie al posto tuo."
            continue
        if resolution is not None and resolution.is_excluded:
            row.error = "Questo asset è escluso dal piano."
            continue
        if resolution is None or resolution.assignment is None:
            row.error = "Il piano non si applica a questo asset: applicalo prima di importare lo storico."
            continue

        key = (row.plan.pk, row.asset.pk, row.last_execution)
        if key in existing or key in seen:
            row.outcome = OUTCOME_DUPLICATE
            row.error = "Già presente: la riga verrà saltata."
            continue
        seen.add(key)

        assignment = resolution.assignment
        row.assignment = assignment
        row.recurrence_label = resolution.recurrence_label
        row.next_due = compute_next_due(
            assignment,
            anchor=assignment.effective_schedule_anchor,
            previous_due=row.last_execution,
            completion_date=row.last_execution,
        )
        row.outcome = OUTCOME_CREATE

    report.rows = parsed
    return report


# ---------------------------------------------------------------------------
# Scrittura
# ---------------------------------------------------------------------------

@transaction.atomic
def apply_report(report: ImportReport, *, user=None, today: date | None = None) -> ImportReport:
    """Scrive le righe valide. Le righe in errore non bloccano le altre.

    Tutto dentro una transazione: un import a metà sarebbe peggio di nessun import,
    perché nessuno saprebbe quali asset hanno già lo storico.
    """
    from django.utils import timezone

    today = today or timezone.localdate()
    author = user if getattr(user, "is_authenticated", False) else None

    for row in report.valid_rows:
        # L'applicazione e' gia' stata risolta in anteprima: ci si scrive sopra la
        # stessa cosa che l'utente ha visto, senza rifare le query per riga.
        assignment = row.assignment
        if assignment is None:
            continue

        MaintenanceOccurrence.objects.create(
            plan=row.plan,
            assignment=assignment,
            asset=row.asset,
            due_date=row.last_execution,
            warning_days=assignment.warning_days,
            schedule_anchor=assignment.effective_schedule_anchor,
            status=MaintenanceOccurrence.STATUS_DONE,
            completed_on=row.last_execution,
            completed_by=author,
            completion_notes=row.notes,
            source=MaintenanceOccurrence.SOURCE_IMPORT,
        )
        report.created_history += 1

        if row.next_due is None:
            continue
        # Se una scadenza aperta esiste gia' (generata prima dell'import) non se ne
        # crea una seconda: l'occorrenza aperta e' unica per coppia piano/asset.
        already_open = MaintenanceOccurrence.objects.filter(
            plan=row.plan,
            asset=row.asset,
            status=MaintenanceOccurrence.STATUS_OPEN,
        ).exists()
        if already_open:
            # C'e' gia' una scadenza aperta per questa coppia: non la si sposta in
            # silenzio sulla data calcolata dallo storico. Chi guarda il numero deve
            # sapere quante ne sono rimaste come stavano.
            report.kept_open += 1
            continue
        _, created = MaintenanceOccurrence.objects.get_or_create(
            plan=row.plan,
            asset=row.asset,
            due_date=row.next_due,
            defaults={
                "assignment": assignment,
                "warning_days": assignment.warning_days,
                "schedule_anchor": assignment.effective_schedule_anchor,
                "previous_due_date": row.last_execution,
                "status": MaintenanceOccurrence.STATUS_OPEN,
                "source": MaintenanceOccurrence.SOURCE_IMPORT,
            },
        )
        if created:
            report.created_next += 1

    report.applied = True
    return report


def build_template_workbook():
    """Modello Excel con le intestazioni attese e una riga di esempio."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Storico manutenzioni"
    sheet.append(TEMPLATE_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.append(["TORNIO01", "Cambio olio", "15/03/2026", "eseguita dal manutentore interno"])
    for index, width in enumerate((22, 34, 20, 46), start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    return workbook
