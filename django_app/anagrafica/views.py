from __future__ import annotations

import calendar
import csv
import json
import logging
from datetime import timedelta as _timedelta
from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from io import StringIO
from pathlib import Path

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, connections, transaction
from django.db.models import Count, Max, Q, Sum
from django.http import (
    Http404, HttpResponse, HttpResponseForbidden, HttpResponseRedirect, JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from django.contrib.auth.decorators import login_required
from core.csv_export import BOM, CSV_CONTENT_TYPE, safe_csv_writer
from core.legacy_anagrafica import (
    count_anagrafica_statuses,
    ensure_anagrafica_schema,
    fetch_anagrafica_rows,
    generate_username,
    normalize_legacy_alias,
    upsert_anagrafica_dipendente,
)
from core.legacy_models import AnagraficaDipendente, Ruolo, UtenteLegacy
from core.legacy_utils import get_legacy_user, is_legacy_admin, legacy_table_columns

from .forms import (
    AnagraficaAziendaleForm,
    AnagraficaCivileForm,
    AttestatoFormazioneConfigForm,
    ElearningConfigForm,
    FiglioACaricoFormSet,
    DipendenteLegacyForm,
    TrainingCompletionRuleForm,
    TrainingCourseDependencyForm,
    TrainingCourseForm,
    TrainingCourseVersionForm,
    TrainingEnrollmentEditForm,
    TrainingInstructorForm,
    TrainingLessonAttendanceForm,
    TrainingLessonForm,
    TrainingPlanForm,
    TrainingRequirementRuleForm,
    TrainingSessionForm,
    TrainingSlideForm,
    TrainingQuizQuestionForm,
    TrainingQuizOptionForm,
    VisitaMedicaForm,
)
from .models import (
    AnagraficaHRPermission,
    AnagraficaStatPermission,
    AnagraficaVisiteMedichePermission,
    AreaAziendale,
    Reparto,
    CartellaDocumentoDipendente,
    SubnavCategoriaAnagrafica,
    SubnavLinkAnagrafica,
    DipendenteAnagraficaAziendale,
    DipendenteAnagraficaCivile,
    DipendenteCambiamentoOrganizzativo,
    DipendenteQualifica,
    DipendenteRuoloOperativo,
    DipendenteStatLayout,
    DocumentoDipendente,
    ImportazioneRetributiva,
    LivelloContrattuale,
    Mansione,
    OffboardingPratica,
    OffboardingTask,
    OnboardingOffboardingCampo,
    OnboardingPratica,
    OnboardingTask,
    RuoloAziendale,
    RuoloOperativo,
    ImportazioneCedolini,
    SaldoCedolino,
    StoricoContratto,
    TipologiaContratto,
    QualificaSessione,
    TipoQualifica,
    TipoVisitaMedica,
    VisitaMedica,
    VoceRetributiva,
    _classify_pay_item,
)
from .models_formazione import (
    AnagraficaFormazionePermission,
    AttestatoFormazioneConfig,
    ElearningConfig,
    TrainingAssignment,
    TrainingAttachment,
    TrainingCertificate,
    TrainingCompletionRule,
    TrainingCourse,
    TrainingCourseDependency,
    TrainingCourseModule,
    TrainingCourseVersion,
    TrainingDeadline,
    TrainingEmployeeRecord,
    TrainingEnrollment,
    TrainingInstructor,
    TrainingLesson,
    TrainingLessonAttendance,
    TrainingExportLog,
    TrainingPlan,
    TrainingRequirementRule,
    TrainingSession,
    TrainingSlide,
    TrainingQuizQuestion,
    TrainingQuizOption,
    TrainingElearningEnrollment,
    TrainingQuizAttempt,
)
from .services.dpi_ingresso import (
    RigaConsegnaIniziale,
    archivia_pdf_cumulativo,
    crea_consegne_iniziali,
    proposta_righe_iniziali,
)
from .services.visite import stato_visite, ultime_visite_correnti_ids, visite_storico
from .services import conformita as conformita_service
from .services import onboarding as onboarding_service
from .services import mansionario as mansionario_service

logger = logging.getLogger(__name__)


def _file_field_url(file_field) -> str:
    if not file_field:
        return ""
    try:
        return file_field.url
    except (OSError, ValueError):
        return ""


def _foto_dipendente_url(civile, legacy_id) -> str:
    """URL della view protetta che serve la foto profilo (storage privato).

    La foto NON ha più un URL pubblico /media/: ``ImageField.url`` solleva
    ``NotImplementedError`` su ``PrivateAnagraficaStorage``. Restituisce l'URL
    della view solo se il dipendente ha effettivamente una foto, altrimenti ""
    (così i template applicano il fallback iniziali).
    """
    if not civile or not getattr(civile, "foto", None):
        return ""
    return reverse("anagrafica:foto_dipendente", args=[int(legacy_id)])


def _cessati_legacy_ids() -> set[int]:
    """Insieme dei ``legacy_anagrafica_id`` con rapporto cessato (ex dipendenti).

    Un dipendente è considerato ex dipendente quando l'anagrafica aziendale ha
    una ``data_cessazione`` valorizzata. Questi nominativi restano a sistema con
    il fascicolo completo ma sono esclusi dalla lista dipendenti in forza.
    """
    return {
        int(lid)
        for lid in DipendenteAnagraficaAziendale.objects
        .filter(data_cessazione__isnull=False)
        .values_list("legacy_anagrafica_id", flat=True)
        if lid
    }


def _audit_safe(request, azione: str, modulo: str, dettaglio: dict | None = None) -> None:
    try:
        from core.audit import log_action

        log_action(request, azione, modulo, dettaglio or {})
    except Exception:
        logger.warning("Audit %s fallito", azione, exc_info=True)


def _int_or_none(value):
    try:
        if value in (None, ""):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _unique_legacy_user(qs):
    users = list(qs.order_by("id")[:2])
    if len(users) == 1:
        return users[0]
    return None


def _resolve_account_portale_dipendente(dip: dict, aziendale=None):
    stored_id = _int_or_none(getattr(aziendale, "utente_id_pre_offboarding", None)) or 0
    if stored_id:
        user = UtenteLegacy.objects.filter(id=stored_id).first()
        if user:
            return user, "utente_id_pre_offboarding"

    current_id = _int_or_none(dip.get("utente_id")) or 0
    if current_id:
        user = UtenteLegacy.objects.filter(id=current_id).first()
        if user:
            return user, "utente_id_corrente"

    email = str(dip.get("email") or "").strip()
    if email:
        user = _unique_legacy_user(UtenteLegacy.objects.filter(email__iexact=email))
        if user:
            return user, "email"

    alias = str(dip.get("aliasusername") or "").strip()
    if alias:
        if "@" in alias:
            user = _unique_legacy_user(UtenteLegacy.objects.filter(email__iexact=alias))
            if user:
                return user, "alias_email"
        else:
            user = _unique_legacy_user(UtenteLegacy.objects.filter(email__istartswith=f"{alias}@"))
            if user:
                return user, "alias_upn"

    nome = str(dip.get("nome") or "").strip()
    cognome = str(dip.get("cognome") or "").strip()
    candidate_names = [name for name in {f"{nome} {cognome}".strip(), f"{cognome} {nome}".strip()} if name]
    if candidate_names:
        name_filter = Q()
        for candidate_name in candidate_names:
            name_filter |= Q(nome__iexact=candidate_name)
        user = _unique_legacy_user(UtenteLegacy.objects.filter(name_filter))
        if user:
            return user, "nome_cognome"

    return None, ""


OFFBOARDING_RESTITUZIONI_LABELS = {
    "badge_chiavi": "Badge, chiavi, tessere",
    "device_it": "PC, telefono, SIM, token",
    "dpi_divise": "DPI, divise, attrezzature",
    "mezzi_carte": "Mezzi, carte, carburante",
    "documenti_archivi": "Documenti e archivi",
    "accessi_account": "Accessi e account da revocare",
}


OFFBOARDING_TASK_BASE = [
    {
        "codice": "hr_documenti_finali",
        "categoria": OffboardingTask.CATEGORIA_HR,
        "titolo": "Preparare documenti e chiusura HR",
        "descrizione": "Verificare comunicazioni, documenti finali e passaggi amministrativi HR.",
    },
    {
        "codice": "it_revoca_accessi",
        "categoria": OffboardingTask.CATEGORIA_IT,
        "titolo": "Revocare accessi e account",
        "descrizione": "Pianificare disattivazione account portale, AD/email, gruppi e credenziali collegate.",
    },
    {
        "codice": "responsabile_passaggio_consegne",
        "categoria": OffboardingTask.CATEGORIA_RESPONSABILE,
        "titolo": "Passaggio consegne con responsabile",
        "descrizione": "Verificare attivita aperte, documenti di reparto e consegne operative.",
    },
]

OFFBOARDING_TASK_BY_RESTITUZIONE = {
    "badge_chiavi": {
        "codice": "restituzione_badge_chiavi",
        "categoria": OffboardingTask.CATEGORIA_HR,
        "titolo": "Recuperare badge, chiavi e tessere",
        "descrizione": "Ritirare badge, chiavi, tessere e accessi fisici assegnati.",
    },
    "device_it": {
        "codice": "restituzione_device_it",
        "categoria": OffboardingTask.CATEGORIA_IT,
        "titolo": "Recuperare dotazioni IT",
        "descrizione": "Ritirare PC, telefono, SIM, token, accessori e verificare eventuali backup.",
    },
    "dpi_divise": {
        "codice": "restituzione_dpi_divise",
        "categoria": OffboardingTask.CATEGORIA_DPI,
        "titolo": "Recuperare DPI, divise e attrezzature",
        "descrizione": "Verificare DPI, divise, utensili e attrezzature assegnate.",
    },
    "mezzi_carte": {
        "codice": "restituzione_mezzi_carte",
        "categoria": OffboardingTask.CATEGORIA_AMMINISTRAZIONE,
        "titolo": "Recuperare mezzi, carte e carburante",
        "descrizione": "Ritirare veicoli, carte aziendali, carte carburante o altri strumenti amministrativi.",
    },
    "documenti_archivi": {
        "codice": "restituzione_documenti_archivi",
        "categoria": OffboardingTask.CATEGORIA_RESPONSABILE,
        "titolo": "Recuperare documenti e archivi",
        "descrizione": "Verificare documenti, archivi locali e materiali di reparto da rientrare.",
    },
    "accessi_account": {
        "codice": "restituzione_accessi_account",
        "categoria": OffboardingTask.CATEGORIA_IT,
        "titolo": "Verificare accessi applicativi da revocare",
        "descrizione": "Controllare applicativi, cartelle condivise, licenze e permessi specifici.",
    },
}


def _offboarding_task_definitions(restituzioni_codes: list[str]) -> list[dict]:
    tasks = list(OFFBOARDING_TASK_BASE)
    existing_codes = {task["codice"] for task in tasks}
    for code in restituzioni_codes:
        task = OFFBOARDING_TASK_BY_RESTITUZIONE.get(code)
        if task and task["codice"] not in existing_codes:
            tasks.append(task)
            existing_codes.add(task["codice"])
    for task in _offboarding_configured_field_tasks():
        if task["codice"] not in existing_codes:
            tasks.append(task)
            existing_codes.add(task["codice"])
    return tasks


def _offboarding_is_admin(request) -> bool:
    legacy_user = get_legacy_user(request.user)
    return bool(request.user.is_superuser or is_legacy_admin(legacy_user))


def _workflow_task_code(field_key: str) -> str:
    safe = "".join(
        ch if ch.isalnum() else "_"
        for ch in (field_key or "").strip().lower()
    ).strip("_")
    return f"campo_{safe or 'configurato'}"[:60]


def _offboarding_configured_field_tasks() -> list[dict]:
    configured = OnboardingOffboardingCampo.objects.filter(
        fase=OnboardingOffboardingCampo.FASE_OFFBOARDING,
        is_active=True,
    ).order_by("ordine", "campo_label")
    tasks: list[dict] = []
    for item in configured:
        description_parts = []
        if item.sezione:
            description_parts.append(f"Sezione + Nuovo dipendente: {item.sezione}.")
        if item.note:
            description_parts.append(item.note)
        if item.obbligatorio:
            description_parts.append("Campo marcato come obbligatorio nel workflow.")
        tasks.append({
            "codice": _workflow_task_code(item.campo_key),
            "categoria": item.categoria,
            "titolo": f"Verificare {item.campo_label}",
            "descrizione": " ".join(description_parts).strip(),
        })
    return tasks


def _dipendente_workflow_field_groups() -> list[dict]:
    forms_by_source = {
        "legacy": DipendenteLegacyForm(),
        "aziendale": AnagraficaAziendaleForm(),
        "civile": AnagraficaCivileForm(),
    }
    raw_groups = [
        ("Dati account", "legacy", [
            "nome", "cognome", "matricola", "aliasusername", "reparto",
            "mansione", "ruolo", "email", "email_notifica", "attivo",
        ]),
        ("Ruolo e organizzazione", "aziendale", ["area", "ruolo_aziendale", "badge"]),
        ("Ruoli operativi e DPI", "manual", [
            ("ruoli_operativi_ids", "Ruoli operativi"),
            ("dpi_consegna_iniziale", "DPI consegnati all'ingresso"),
        ]),
        ("Contratto e inquadramento", "manual", [
            ("contratto_data_inizio", "Data inizio contratto"),
            ("contratto_data_fine", "Data fine contratto"),
            ("contratto_tipologia_id", "Tipologia contratto"),
            ("contratto_codice_livello", "Livello CCNL"),
            ("contratto_ccnl", "CCNL applicato"),
            ("contratto_qualifica_nome", "Qualifica professionale"),
        ]),
        ("Periodo di prova e prima assunzione", "aziendale", [
            "data_prima_assunzione", "data_assunzione_ultima", "data_cessazione",
            "prova_data_inizio", "prova_data_fine",
        ]),
        ("Contatti aziendali", "aziendale", ["email_aziendale", "telefono_aziendale"]),
        ("Taglie DPI / abbigliamento", "aziendale", ["taglia_scarpe", "taglia_pantalone", "taglia_maglia"]),
        ("Privacy", "aziendale", ["consenso_privacy", "data_consenso_privacy"]),
        ("Dati personali", "civile", [
            "foto", "data_nascita", "luogo_nascita", "provincia_nascita",
            "nazionalita", "genere", "titolo_studio",
        ]),
        ("Residenza", "civile", [
            "indirizzo_residenza", "citta_residenza", "provincia_residenza",
            "cap_residenza", "nazione_residenza",
        ]),
        ("Domicilio", "civile", ["indirizzo_domicilio", "citta_domicilio", "cap_domicilio", "nazione_domicilio"]),
        ("Contatti privati e patente", "civile", ["email_privata", "telefono_privato", "patente_auto"]),
        ("Dati riservati HR", "civile", [
            "codice_fiscale", "nome_banca", "iban", "intestatario_conto",
            "categoria_protetta", "categoria_disabili", "percentuale_disabilita",
        ]),
    ]
    groups: list[dict] = []
    for section, source, fields in raw_groups:
        rows = []
        for entry in fields:
            if isinstance(entry, tuple):
                key, label = entry
                required = False
            else:
                key = entry
                field = forms_by_source.get(source).fields.get(key) if source in forms_by_source else None
                label = field.label if field else key.replace("_", " ").title()
                required = bool(field.required) if field else False
            rows.append({
                "key": key,
                "label": label,
                "section": section,
                "source": source,
                "required": required,
            })
        groups.append({"label": section, "fields": rows})
    return groups


def _dipendente_workflow_field_map() -> dict[str, dict]:
    return {
        field["key"]: field
        for group in _dipendente_workflow_field_groups()
        for field in group["fields"]
    }


def _offboarding_dipendente_nome(dip: dict) -> str:
    cognome = str(dip.get("cognome") or "").strip()
    nome = str(dip.get("nome") or "").strip()
    legacy_id = dip.get("id") or ""
    return f"{cognome} {nome}".strip() or f"#{legacy_id}".strip()


# ---------------------------------------------------------------------------
# Dashboard anagrafica
# ---------------------------------------------------------------------------

# Righe del blocco «Cose da gestire»: una per sorgente dello scadenzario, nello
# stesso ordine in cui vanno affrontate. Il gating per sorgente è già dentro
# _build_scadenzario_voci — qui non si decide chi vede cosa.
_COSE_DA_GESTIRE_KINDS = [
    ("visita",     "🩺", "Visite mediche"),
    ("qualifica",  "🎓", "Qualifiche"),
    ("formazione", "📘", "Corsi obbligatori"),
    ("contratto",  "📄", "Contratti e periodi di prova"),
]


def _build_cose_da_gestire(request, dip_map: dict) -> list[dict]:
    """Righe azionabili della dashboard HR, contate sulle voci dello scadenzario.

    Ogni riga porta allo scadenzario già filtrato su ciò che ha appena contato:
    numero e lista vengono dalla stessa funzione, quindi non possono divergere.
    Le righe a zero non compaiono — una dashboard elenca le cose da fare, non le
    cose che non ci sono.
    """
    voci = _build_scadenzario_voci(request, dip_map=dip_map)
    base = reverse("anagrafica:scadenzario")

    scadute: list[dict] = []
    in_scadenza: list[dict] = []
    for kind, icona, label in _COSE_DA_GESTIRE_KINDS:
        voci_kind = [v for v in voci if v["kind"] == kind]
        n_scadute = sum(1 for v in voci_kind if v["scaduta"])
        # Le voci non scadute che _build_scadenzario_voci restituisce senza filtro
        # stato sono già solo quelle entro 60 giorni.
        n_prossime = len(voci_kind) - n_scadute
        if n_scadute:
            scadute.append({
                "icona": icona,
                "titolo": f"{label} scadute",
                "count": n_scadute,
                "urgente": True,
                "url": f"{base}?tipo={kind}&stato=scaduta",
            })
        if n_prossime:
            in_scadenza.append({
                "icona": icona,
                "titolo": f"{label} in scadenza (60 giorni)",
                "count": n_prossime,
                "urgente": False,
                "url": f"{base}?tipo={kind}&stato=60",
            })
    return scadute + in_scadenza


@login_required
def index(request):
    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(deduplicate=True)
    # Esclude gli ex dipendenti (rapporto cessato) dai KPI "in organico":
    # restano a sistema solo per storico ma non concorrono al conteggio attivi.
    cessati_ids = _cessati_legacy_ids()
    rows_attivi = [row for row in rows if int(row.get("id") or 0) not in cessati_ids]
    n_dipendenti = len(rows_attivi)
    n_reparti = len({str(row.get("reparto") or "").strip().casefold() for row in rows_attivi if str(row.get("reparto") or "").strip()})

    # Conteggi catalogo per dashboard HR
    n_mansioni = Mansione.objects.filter(is_active=True).count()
    n_aree = AreaAziendale.objects.filter(is_active=True).count()
    n_reparti_catalog = Reparto.objects.filter(is_active=True).count()
    n_qualifiche = TipoQualifica.objects.filter(is_active=True).count()

    # Qualifiche scadute e in scadenza (prossimi 60 giorni)
    from datetime import timedelta
    from django.utils import timezone as tz
    oggi = tz.localdate()
    soglia = oggi + timedelta(days=60)
    n_qualifiche_scadute = DipendenteQualifica.objects.filter(
        data_scadenza__isnull=False, data_scadenza__lt=oggi
    ).count()
    n_qualifiche_scadenza = DipendenteQualifica.objects.filter(
        data_scadenza__isnull=False, data_scadenza__gte=oggi, data_scadenza__lte=soglia
    ).count()

    # Visite mediche scadute (dato sanitario sensibile: gating)
    can_view_visite = _can_view_visite_mediche(request)
    n_visite_scadute = 0
    if can_view_visite:
        n_visite_scadute = VisitaMedica.objects.filter(
            id__in=ultime_visite_correnti_ids(), data_scadenza__lt=oggi
        ).count()

    # Cose da gestire: righe azionabili contate sulle voci dello scadenzario.
    # `dip_map` riusa le righe legacy già lette a inizio view — nessun secondo fetch.
    dip_map = {int(row["id"]): row for row in rows if row.get("id")}
    cose_da_gestire = _build_cose_da_gestire(request, dip_map)

    return render(request, "anagrafica/pages/index.html", {
        "n_dipendenti": n_dipendenti,
        "n_reparti": n_reparti,
        "n_mansioni": n_mansioni,
        # Il catalogo ha due livelli: Reparto (contenitore) e AreaAziendale (figlia,
        # FK `reparto`). Il KPI «Reparti (catalogo)» mostrava `n_aree`, cioè le aree:
        # un numero giusto sotto l'etichetta sbagliata.
        "n_reparti_catalog": n_reparti_catalog,
        "n_aree": n_aree,
        "n_qualifiche": n_qualifiche,
        "n_qualifiche_scadute": n_qualifiche_scadute,
        "n_qualifiche_scadenza": n_qualifiche_scadenza,
        "can_view_visite": can_view_visite,
        "n_visite_scadute": n_visite_scadute,
        "cose_da_gestire": cose_da_gestire,
        # Fascia «Vai a»: i sottomoduli gated non compaiono a chi vedrebbe solo un
        # rifiuto. La nav non è un confine di sicurezza — le view restano gated.
        "can_view_formazione": _can_view_formazione(request),
        "can_view_hr": _check_hr_permission(request),
    })


# ---------------------------------------------------------------------------
# Dipendenti (sola lettura — dati da legacy SQL Server)
# ---------------------------------------------------------------------------

def _dipendenti_base_rows() -> list[dict]:
    """Righe dell'anagrafica in forza (legacy, deduplicate), pre-filtro.

    Esclude i rapporti cessati e risolve il reparto dalla **fonte unica
    canonica** (``dipendente → area_aziendale → reparto``): il reparto
    canonico ha **precedenza sul testo legacy**, così le tabelle non mostrano
    più il vecchio reparto quando l'assegnazione canonica è cambiata. Se non
    c'è canonico si tiene il testo legacy della riga; se anche quello è vuoto
    si ripiega su ``DipendenteAnagraficaAziendale.area``. Ogni riga riceve
    anche ``area_aziendale_nome`` (l'accoppiata reparto + area).
    """
    from anagrafica.services.reparto_canonico import enrich_rows_reparto_canonico

    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(deduplicate=True)
    # Gli ex dipendenti (rapporto cessato) restano a sistema ma non compaiono
    # mai in questa lista: hanno una vista dedicata `ex_dipendenti_list`.
    cessati_ids = _cessati_legacy_ids()
    rows = [row for row in rows if int(row.get("id") or 0) not in cessati_ids]

    # Reparto/area CANONICI (precedenza sul testo legacy stantio).
    enrich_rows_reparto_canonico(rows)

    # Fallback finale: testo `.area` sull'aziendale, solo per chi non ha né
    # canonico (reparto non toccato dall'enricher) né testo legacy sulla riga.
    _ids_no_reparto = [int(r.get("id") or 0) for r in rows if not str(r.get("reparto") or "").strip()]
    if _ids_no_reparto:
        _az_area_text = dict(
            DipendenteAnagraficaAziendale.objects
            .filter(legacy_anagrafica_id__in=_ids_no_reparto)
            .exclude(area="")
            .values_list("legacy_anagrafica_id", "area")
        )
        for row in rows:
            if not str(row.get("reparto") or "").strip():
                lid = int(row.get("id") or 0)
                if lid in _az_area_text:
                    row["reparto"] = _az_area_text[lid]
    return rows


def _filter_dipendenti_rows(rows: list[dict], request) -> list[dict]:
    """Filtri della lista dipendenti (q / reparto / area / tipologia_contratto)."""
    q = request.GET.get("q", "").strip()
    reparto = request.GET.get("reparto", "").strip()
    area_filter = request.GET.get("area", "").strip()
    contratto_filter = request.GET.get("tipologia_contratto", "").strip()

    if q:
        q_norm = q.casefold()
        rows = [
            row
            for row in rows
            if any(
                q_norm in value.casefold()
                for value in [
                    str(row.get("nome") or "").strip(),
                    str(row.get("cognome") or "").strip(),
                    str(row.get("aliasusername") or "").strip(),
                    str(row.get("matricola") or "").strip(),
                ]
                if value
            )
        ]
    if reparto:
        rows = [row for row in rows if str(row.get("reparto") or "").strip().casefold() == reparto.casefold()]

    # Filtri su campi Django (area, tipologia_contratto)
    if area_filter or contratto_filter:
        az_qs = DipendenteAnagraficaAziendale.objects.all()
        if area_filter:
            az_qs = az_qs.filter(area__iexact=area_filter)
        if contratto_filter:
            az_qs = az_qs.filter(tipologia_contratto=contratto_filter)
        allowed_ids = set(az_qs.values_list("legacy_anagrafica_id", flat=True))
        rows = [row for row in rows if int(row.get("id") or 0) in allowed_ids]
    return rows


def _sort_dipendenti_rows(rows: list[dict]) -> list[dict]:
    rows.sort(key=lambda row: (
        str(row.get("cognome") or "").strip().casefold(),
        str(row.get("nome") or "").strip().casefold(),
        str(row.get("aliasusername") or "").strip().casefold(),
        int(row.get("id") or 0),
    ))
    return rows


def build_dipendenti_rows(
    request,
    *,
    apply_filters: bool = True,
    base_rows: list[dict] | None = None,
) -> list[dict]:
    """Righe dell'elenco dipendenti (legacy + fallback reparto), filtrate e ordinate.

    Fonte unica condivisa tra la view ``dipendenti_list`` e l'export
    (``anagrafica.exports``): il filtro non va duplicato altrove, o le due
    superfici divergono (drift). La view passa ``base_rows`` (le righe
    pre-filtro già calcolate per ``n_totale``/``reparti``) per evitare un
    secondo fetch legacy; l'export, che non ha già una base, lascia che
    l'helper la calcoli da sé.
    """
    rows = base_rows if base_rows is not None else _dipendenti_base_rows()
    if apply_filters:
        rows = _filter_dipendenti_rows(rows, request)
    return _sort_dipendenti_rows(list(rows))


@login_required
def dipendenti_list(request):
    q = request.GET.get("q", "").strip()
    reparto = request.GET.get("reparto", "").strip()
    area_filter = request.GET.get("area", "").strip()
    contratto_filter = request.GET.get("tipologia_contratto", "").strip()

    rows_all = _dipendenti_base_rows()
    # NB: `reparti_list` e `n_totale` sono calcolati sulle righe PRE-filtro.
    reparti_list = sorted({str(row.get("reparto") or "").strip() for row in rows_all if str(row.get("reparto") or "").strip()})
    n_totale = len(rows_all)
    cessati_ids = _cessati_legacy_ids()

    # Stessa costruzione righe dell'export (`anagrafica.exports._dipendenti_rows`
    # → `build_dipendenti_rows`): passare `rows_all` come base evita un secondo
    # fetch legacy pur restando sull'unica fonte condivisa filtro/ordina.
    rows = build_dipendenti_rows(request, base_rows=rows_all)

    aree_list = sorted(
        DipendenteAnagraficaAziendale.objects.exclude(area="")
        .values_list("area", flat=True)
        .distinct()
        .order_by("area")
    )

    user_map = {
        int(user.id): user
        for user in UtenteLegacy.objects.filter(
            id__in=[int(row.get("utente_id") or 0) for row in rows if int(row.get("utente_id") or 0) > 0]
        )
    }
    for row in rows:
        raw_attivo = row.get("attivo")
        row["anagrafica_attivo"] = True if raw_attivo is None else bool(raw_attivo)
        row["matricola_legacy"] = str(row.get("matricola") or "").strip()
        row["ruolo_legacy"] = str(row.get("ruolo") or row.get("mansione") or "").strip()
        linked_user = user_map.get(int(row.get("utente_id") or 0))
        row["account_attivo"] = bool(getattr(linked_user, "attivo", False))
        row["has_account"] = bool(linked_user)
        row["timbri_operator_id"] = None
        row["timbri_count"] = 0
        row["timbri_legacy_id"] = int(row.get("id") or 0) or None

    # Pagina unica: la tabella usa ricerca/ordina/filtra per colonna lato client
    # (fm-table-enhanced.js), che opera sulle righe nel DOM. Con il volume reale
    # dell'anagrafica (~150 dipendenti) carichiamo tutto in un blocco così filtro
    # e ordinamento coprono l'intero elenco e non solo la pagina corrente.
    paginator = Paginator(rows, 500)
    page = paginator.get_page(request.GET.get("page"))

    civile_map = {
        int(obj.legacy_anagrafica_id): obj
        for obj in DipendenteAnagraficaCivile.objects.filter(
            legacy_anagrafica_id__in=[int(dip.get("id") or 0) for dip in list(page.object_list)]
        ).only("legacy_anagrafica_id", "foto")
    }
    for dip in list(page.object_list):
        civile = civile_map.get(int(dip.get("id") or 0))
        dip["foto_url"] = _foto_dipendente_url(civile, dip.get("id") or 0)

    try:
        from timbri.models import OperatoreTimbri, RegistroTimbro

        operator_map: dict[int, OperatoreTimbri] = {
            int(obj.legacy_anagrafica_id): obj
            for obj in OperatoreTimbri.objects.filter(
                legacy_anagrafica_id__in=[int(dip.get("id") or 0) for dip in list(page.object_list)]
            )
            if obj.legacy_anagrafica_id
        }

        counts = {
            int(row["operatore_id"]): int(row["n"])
            for row in RegistroTimbro.objects.filter(operatore_id__in=[op.id for op in operator_map.values()])
            .order_by()
            .values("operatore_id")
            .annotate(n=Count("id"))
        }
        for dip in list(page.object_list):
            legacy_id = int(dip.get("id") or 0)
            operatore = operator_map.get(legacy_id)
            dip["timbri_operator_id"] = getattr(operatore, "id", None)
            dip["timbri_count"] = counts.get(getattr(operatore, "id", 0), 0)
            dip["timbri_legacy_id"] = legacy_id if legacy_id > 0 else None
    except Exception:
        logger.exception("Impossibile arricchire l'elenco dipendenti con i dati timbri.")
        for dip in list(page.object_list):
            dip["timbri_operator_id"] = None
            dip["timbri_count"] = 0
            dip["timbri_legacy_id"] = None

    status_stats = count_anagrafica_statuses()
    tipologie_contratto_list = list(
        TipologiaContratto.objects.filter(is_active=True).order_by("ordine", "codice")
    )
    return render(request, "anagrafica/pages/dipendenti_list.html", {
        "page_obj": page,
        "q": q,
        "reparto": reparto,
        "area": area_filter,
        "tipologia_contratto": contratto_filter,
        "reparti": reparti_list,
        "aree": aree_list,
        "tipologie_contratto": tipologie_contratto_list,
        "n_totale": n_totale,
        "n_reparti": len(reparti_list),
        "n_attivi": status_stats["active"],
        "n_non_attivi": status_stats["inactive"],
        "n_ex": len(cessati_ids),
    })


# ---------------------------------------------------------------------------
# Ex dipendenti — vista dedicata ai rapporti cessati
# ---------------------------------------------------------------------------

@login_required
def ex_dipendenti_list(request):
    """Elenco degli ex dipendenti (rapporto di lavoro cessato).

    Vista separata dalla lista dipendenti in forza: questi nominativi restano a
    sistema con il fascicolo completo (documenti, retribuzioni, storico) ma non
    compaiono tra i dipendenti attivi. Un dipendente è qui quando la sua
    anagrafica aziendale ha una ``data_cessazione`` valorizzata.
    """
    ensure_anagrafica_schema()
    q = request.GET.get("q", "").strip()

    cessati_ids = _cessati_legacy_ids()
    rows = [
        row
        for row in fetch_anagrafica_rows(deduplicate=True)
        if int(row.get("id") or 0) in cessati_ids
    ]
    if q:
        q_norm = q.casefold()
        rows = [
            row
            for row in rows
            if any(
                q_norm in str(row.get(field) or "").strip().casefold()
                for field in ("nome", "cognome", "aliasusername", "matricola")
                if str(row.get(field) or "").strip()
            )
        ]

    az_map = {
        obj.legacy_anagrafica_id: obj
        for obj in DipendenteAnagraficaAziendale.objects.filter(
            legacy_anagrafica_id__in=cessati_ids
        )
    }
    civile_map = {
        int(obj.legacy_anagrafica_id): obj
        for obj in DipendenteAnagraficaCivile.objects.filter(
            legacy_anagrafica_id__in=cessati_ids
        ).only("legacy_anagrafica_id", "foto")
    }
    for row in rows:
        legacy_id = int(row.get("id") or 0)
        az = az_map.get(legacy_id)
        civile = civile_map.get(legacy_id)
        row["legacy_id"] = legacy_id
        row["matricola_legacy"] = str(row.get("matricola") or "").strip()
        row["foto_url"] = _foto_dipendente_url(civile, legacy_id)
        row["data_cessazione"] = getattr(az, "data_cessazione", None)
        row["data_assunzione"] = (
            getattr(az, "data_assunzione_ultima", None)
            or getattr(az, "data_prima_assunzione", None)
        )
        row["tipologia_contratto_display"] = (
            az.get_tipologia_contratto_display() if az and az.tipologia_contratto else ""
        )

    rows.sort(key=lambda row: (
        row.get("data_cessazione") or date.min,
        str(row.get("cognome") or "").strip().casefold(),
        str(row.get("nome") or "").strip().casefold(),
    ), reverse=True)

    # Pagina unica per filtro/ordinamento client-side completo (vedi dipendenti_list).
    paginator = Paginator(rows, 500)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "anagrafica/pages/ex_dipendenti_list.html", {
        "page_obj": page,
        "q": q,
        "n_ex": len(cessati_ids),
    })


# ---------------------------------------------------------------------------
# Dipendente — creazione completa (legacy + civile + aziendale)
# ---------------------------------------------------------------------------

@login_required
def dipendente_create(request):
    ensure_anagrafica_schema()
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Non hai i permessi per creare un dipendente.")
        return redirect("anagrafica:dipendenti_list")

    can_hr = _check_hr_permission(request)

    if request.method == "POST":
        legacy_form = DipendenteLegacyForm(request.POST)
        form_civile = AnagraficaCivileForm(request.POST, request.FILES)
        form_aziendale = AnagraficaAziendaleForm(request.POST)
        formset_figli = FiglioACaricoFormSet(request.POST, prefix="figli")

        # Le form civile/aziendale sono facoltative: ignora errori se tutti i campi sono vuoti
        civile_has_data = any(
            v for k, v in request.POST.items()
            if k not in ("csrfmiddlewaretoken",) and not k.startswith("az_") and not k.startswith("legacy_")
            and v and k in AnagraficaCivileForm().fields
        ) or bool(request.FILES)
        aziendale_has_data = any(
            v for k, v in request.POST.items()
            if k not in ("csrfmiddlewaretoken",) and v and k in AnagraficaAziendaleForm().fields
        )

        if legacy_form.is_valid():
            data = legacy_form.cleaned_data
            try:
                _nome = (data.get("nome") or "").strip()
                _cognome = (data.get("cognome") or "").strip()
                _alias = (data.get("aliasusername") or "").strip()
                if not _alias:
                    _existing_u = set(
                        AnagraficaDipendente.objects
                        .exclude(aliasusername="")
                        .exclude(aliasusername__isnull=True)
                        .values_list("aliasusername", flat=True)
                    )
                    _alias = generate_username(_nome, _cognome, _existing_u)
                row = upsert_anagrafica_dipendente(
                    aliasusername=_alias,
                    nome=_nome,
                    cognome=_cognome,
                    matricola=(data.get("matricola") or "").strip(),
                    reparto=(data.get("reparto") or "").strip(),
                    mansione=(data.get("mansione") or "").strip(),
                    ruolo=(data.get("ruolo") or "").strip(),
                    email=(data.get("email") or "").strip(),
                    email_notifica=(data.get("email_notifica") or "").strip(),
                    attivo=bool(data.get("attivo")),
                    utente_id=None,
                )
                new_id = int(row.get("id") or 0)

                # Salva anagrafica aziendale se il form è valido
                if new_id and form_aziendale.is_valid():
                    az = form_aziendale.save(commit=False)
                    az.legacy_anagrafica_id = new_id
                    az.updated_by = request.user
                    az.save()

                # Salva anagrafica civile se il form è valido
                if new_id and form_civile.is_valid():
                    civ = form_civile.save(commit=False)
                    civ.legacy_anagrafica_id = new_id
                    civ.updated_by = request.user
                    civ.save()
                    # Figli a carico: salva il formset e allinea il flag ai record registrati
                    if formset_figli.is_valid():
                        formset_figli.instance = civ
                        formset_figli.save()
                        ha_figli = civ.figli.exists()
                        if civ.figli_a_carico != ha_figli:
                            civ.figli_a_carico = ha_figli
                            civ.save(update_fields=["figli_a_carico"])

                # Reparto effettivo: sincronizza area aziendale e caporeparto dal catalogo
                if new_id:
                    _sync_aziendale_from_reparto(
                        new_id, (data.get("reparto") or "").strip(), saved_by=request.user
                    )

                # Crea automaticamente account portale (solo se non già collegato).
                # Solo per dipendenti ATTIVI: un dipendente inserito come inattivo
                # (storico, cessato, non ancora in forza) non deve ottenere un
                # account con password iniziale prevedibile che nessuno presidia.
                if new_id and _alias and bool(data.get("attivo")):
                    if not row.get("utente_id"):
                        try:
                            from django.db.models import Q as _Q
                            from werkzeug.security import generate_password_hash as _gph
                            _existing_utente = UtenteLegacy.objects.filter(
                                _Q(email__iexact=_alias) | _Q(email__istartswith=f"{_alias}@")
                            ).order_by("id").first()
                            if _existing_utente is not None:
                                # Utente LDAP o locale già presente: collega senza creare
                                AnagraficaDipendente.objects.filter(id=new_id).update(utente_id=_existing_utente.id)
                            else:
                                _ruolo_utente = Ruolo.objects.filter(nome__iexact="utente").first()
                                _anno_nascita = None
                                if form_civile.is_valid():
                                    _dn = form_civile.cleaned_data.get("data_nascita")
                                    if _dn:
                                        _anno_nascita = _dn.year
                                _pwd_iniziale = str(_anno_nascita) if _anno_nascita else f"Portale{date.today().year}"
                                _nome_completo = f"{data.get('cognome', '')} {data.get('nome', '')}".strip()
                                _nuovo_utente = UtenteLegacy.objects.create(
                                    nome=_nome_completo,
                                    email=_alias,
                                    password=_gph(_pwd_iniziale),
                                    ruolo="utente",
                                    ruolo_id=_ruolo_utente.id if _ruolo_utente else None,
                                    attivo=True,
                                    deve_cambiare_password=True,
                                )
                                AnagraficaDipendente.objects.filter(id=new_id).update(utente_id=_nuovo_utente.id)
                        except Exception:
                            logger.warning(
                                "Creazione automatica account portale fallita per dipendente %s", new_id,
                                exc_info=True,
                            )

                # Assegna ruoli operativi scelti nel form (multiselect "ruoli_operativi_ids")
                if new_id:
                    ruoli_raw = request.POST.getlist("ruoli_operativi_ids")
                    ruoli_ids: list[int] = []
                    for raw in ruoli_raw:
                        try:
                            ruoli_ids.append(int(raw))
                        except (TypeError, ValueError):
                            continue
                    if ruoli_ids:
                        ruoli_validi = list(
                            RuoloOperativo.objects.filter(pk__in=ruoli_ids, is_active=True)
                        )
                        for ruolo in ruoli_validi:
                            DipendenteRuoloOperativo.objects.get_or_create(
                                legacy_anagrafica_id=new_id,
                                ruolo=ruolo,
                                defaults={"assegnato_da": request.user},
                            )

                # DPI consegnati all'ingresso (proposti dalla mansione e confermati da HR)
                if new_id:
                    civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=new_id).first()
                    aziendale_obj = DipendenteAnagraficaAziendale.objects.filter(legacy_anagrafica_id=new_id).first()
                    righe_iniziali = _parse_dpi_iniziali_post(request.POST)
                    if righe_iniziali:
                        consegne = crea_consegne_iniziali(
                            civile_obj, aziendale_obj, righe_iniziali, request.user
                        )
                        if consegne:
                            archivia_pdf_cumulativo(consegne, request.user)
                            try:
                                from core.audit import log_action
                                log_action(
                                    request, "DPI_CONSEGNA_INGRESSO", "anagrafica",
                                    f"Consegna iniziale DPI a #{new_id}: {len(consegne)} articoli",
                                )
                            except Exception:
                                logger.warning("Audit DPI_CONSEGNA_INGRESSO fallito", exc_info=True)

                # Crea primo StoricoContratto se la sezione contratto è stata compilata
                if new_id:
                    from datetime import date as _date
                    data_inizio_raw = (request.POST.get("contratto_data_inizio") or "").strip()
                    if data_inizio_raw:
                        try:
                            data_inizio = _date.fromisoformat(data_inizio_raw)
                        except ValueError:
                            data_inizio = None
                        if data_inizio:
                            data_fine_raw = (request.POST.get("contratto_data_fine") or "").strip()
                            data_fine = None
                            if data_fine_raw:
                                try:
                                    data_fine = _date.fromisoformat(data_fine_raw)
                                except ValueError:
                                    data_fine = None
                            tip_id = request.POST.get("contratto_tipologia_id") or ""
                            tipologia_codice = ""
                            if tip_id.isdigit():
                                tip = TipologiaContratto.objects.filter(pk=int(tip_id)).first()
                                if tip:
                                    tipologia_codice = tip.codice
                            cod_lvl = (request.POST.get("contratto_codice_livello") or "").strip().upper()[:10]
                            descr_lvl = ""
                            if cod_lvl:
                                lvl = LivelloContrattuale.objects.filter(codice=cod_lvl).first()
                                if lvl:
                                    descr_lvl = lvl.descrizione
                            tax_code = (request.POST.get("codice_fiscale") or "").upper().strip()[:16]
                            StoricoContratto.objects.create(
                                legacy_anagrafica_id=new_id,
                                tax_code=tax_code,
                                data_inizio=data_inizio,
                                data_fine=data_fine,
                                tipologia_contratto=tipologia_codice,
                                codice_livello=cod_lvl,
                                descrizione_livello=descr_lvl,
                                ccnl=(request.POST.get("contratto_ccnl") or "").strip()[:100],
                                qualifica_nome=(request.POST.get("contratto_qualifica_nome") or "").strip()[:150],
                                importato_da=request.user,
                            )

                # Onboarding strutturato: avvio opzionale della pratica + checklist
                if new_id and request.POST.get("avvia_onboarding"):
                    try:
                        nome_onb = f"{data.get('cognome', '')} {data.get('nome', '')}".strip()
                        ruoli_onb = list(
                            DipendenteRuoloOperativo.objects
                            .filter(legacy_anagrafica_id=new_id)
                            .values_list("ruolo_id", flat=True)
                        )
                        data_ass = None
                        _ass_raw = (request.POST.get("contratto_data_inizio") or "").strip()
                        if _ass_raw:
                            try:
                                from datetime import date as _date2
                                data_ass = _date2.fromisoformat(_ass_raw)
                            except ValueError:
                                data_ass = None
                        onboarding_service.avvia_onboarding(
                            legacy_id=new_id,
                            dipendente_nome=nome_onb or f"#{new_id}",
                            reparto=(data.get("reparto") or "").strip(),
                            mansione=(data.get("mansione") or "").strip(),
                            data_assunzione=data_ass,
                            user=request.user,
                            ruolo_ids=ruoli_onb,
                        )
                    except Exception:
                        logger.warning("Avvio onboarding automatico fallito per dipendente %s", new_id, exc_info=True)

                # Formazione sicurezza pregressa dichiarata in preinserimento
                if new_id:
                    try:
                        fsic_items: list[dict] = []
                        for cid in request.POST.getlist("fsic_corso"):
                            if not str(cid).isdigit():
                                continue
                            raw = (request.POST.get(f"fsic_data_{cid}") or "").strip()
                            if not raw:
                                continue
                            try:
                                from datetime import date as _date3
                                fsic_items.append({"corso_id": int(cid), "data": _date3.fromisoformat(raw)})
                            except ValueError:
                                continue
                        if fsic_items:
                            n_fsic = onboarding_service.registra_formazione_pregressa(
                                new_id, fsic_items, user=request.user
                            )
                            if n_fsic:
                                messages.info(request, f"Registrati {n_fsic} corsi di formazione pregressa.")
                    except Exception:
                        logger.warning("Registrazione formazione pregressa fallita per %s", new_id, exc_info=True)

                nome_disp = f"{data.get('cognome', '')} {data.get('nome', '')}".strip() or "Dipendente"
                messages.success(request, f'Dipendente "{nome_disp}" creato.')
                if new_id:
                    return redirect("anagrafica:dipendente_detail", legacy_id=new_id)
                return redirect("anagrafica:dipendenti_list")
            except Exception as exc:
                logger.exception("Errore creazione dipendente")
                messages.error(request, f"Impossibile creare il dipendente: {exc}")
        else:
            messages.error(request, "Controlla i campi obbligatori nella sezione Dati account.")
    else:
        legacy_form = DipendenteLegacyForm(initial={"attivo": True})
        form_civile = AnagraficaCivileForm()
        form_aziendale = AnagraficaAziendaleForm()
        formset_figli = FiglioACaricoFormSet(prefix="figli")

    reparti_catalogo = list(
        Reparto.objects.filter(is_active=True).order_by("nome")
    )
    # Corsi di sicurezza per la dichiarazione "formazione pregressa" in preinserimento.
    from .models_formazione import TrainingCourse
    formazione_sicurezza_corsi = list(
        TrainingCourse.objects
        .filter(is_active=True, stato="ATTIVO", obbligatorio=True)
        .order_by("titolo")
    )
    return render(request, "anagrafica/pages/dipendente_create.html", {
        "legacy_form": legacy_form,
        "form_civile": form_civile,
        "form_aziendale": form_aziendale,
        "formset_figli": formset_figli,
        "can_hr": can_hr,
        "mansioni_catalogo": list(Mansione.objects.filter(is_active=True).order_by("nome")),
        "reparti_catalogo": reparti_catalogo,
        "contratto_choices": DipendenteAnagraficaAziendale.CONTRATTO_CHOICES,
        "tipologie_contratto": list(TipologiaContratto.objects.filter(is_active=True).order_by("ordine", "codice")),
        "livelli_contrattuali": list(LivelloContrattuale.objects.filter(is_active=True).order_by("ordine", "codice")),
        "ruoli_operativi_catalogo": list(RuoloOperativo.objects.filter(is_active=True).order_by("nome")),
        "formazione_sicurezza_corsi": formazione_sicurezza_corsi,
    })


def _parse_dpi_iniziali_post(post) -> list[RigaConsegnaIniziale]:
    """Estrae le righe DPI iniziali confermate (checkbox 'consegnato') dal POST.

    Convenzioni di name dei campi (per indice intero ``i``):
      - ``dpi_consegnato_<i>`` checkbox (presenza = riga selezionata)
      - ``dpi_categoria_id_<i>`` int obbligatorio
      - ``dpi_modello_id_<i>`` int opzionale
      - ``dpi_taglia_id_<i>`` int opzionale
      - ``dpi_quantita_<i>`` int (default 1)

    Gli indici disponibili sono enumerati dal campo nascosto ``dpi_indici``
    (CSV) inviato dal partial HTMX.
    """
    righe: list[RigaConsegnaIniziale] = []
    indici_raw = post.get("dpi_indici", "").strip()
    if not indici_raw:
        return righe
    for token in indici_raw.split(","):
        token = token.strip()
        if not token.isdigit():
            continue
        i = int(token)
        if not post.get(f"dpi_consegnato_{i}"):
            continue
        cat_raw = post.get(f"dpi_categoria_id_{i}", "").strip()
        if not cat_raw.isdigit():
            continue
        try:
            quantita = int(post.get(f"dpi_quantita_{i}", "1") or 1)
        except ValueError:
            quantita = 1
        def _int_or_none(value: str) -> int | None:
            value = (value or "").strip()
            return int(value) if value.isdigit() else None

        righe.append(RigaConsegnaIniziale(
            categoria_id=int(cat_raw),
            quantita=max(1, quantita),
            tipo_id=_int_or_none(post.get(f"dpi_tipo_id_{i}", "")),
            modello_id=_int_or_none(post.get(f"dpi_modello_id_{i}", "")),
            taglia_id=_int_or_none(post.get(f"dpi_taglia_id_{i}", "")),
        ))
    return righe


# ---------------------------------------------------------------------------
# NOTE: tutte le view "Fornitori" sono state spostate nel modulo dedicato
# `fornitori` (URL prefix /fornitori/). I modelli Fornitore* restano in
# `anagrafica.models` perché referenziati da `assets.models` via FK storiche
# (tabelle DB invariate).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Definizione widget statistiche dipendente
# ---------------------------------------------------------------------------

STAT_WIDGETS = [
    {
        "id": "tickets_aperti",
        "title": "Ticket aperti",
        "icon": "🎫",
        "color": "#ef4444",
        "link_url": "/tickets/",
        "link_label": "Vai ai ticket",
    },
    {
        "id": "tickets_totali",
        "title": "Ticket totali",
        "icon": "📋",
        "color": "#3b82f6",
        "link_url": "/tickets/",
        "link_label": "Vai ai ticket",
    },
    {
        "id": "anomalie",
        "title": "Anomalie",
        "icon": "⚠️",
        "color": "#f59e0b",
        "link_url": "/gestione-anomalie/",
        "link_label": "Vai alle anomalie",
    },
    {
        "id": "diario_preposto",
        "title": "Diario preposto",
        "icon": "📔",
        "color": "#8b5cf6",
        "link_url": "/diario-preposto/",
        "link_label": "Vai al diario",
    },
    {
        "id": "rilevazioni",
        "title": "Rilevazioni sicurezza",
        "icon": "🦺",
        "color": "#10b981",
        "link_url": "/rilevazione-incidenti/",
        "link_label": "Vai alle rilevazioni",
    },
    {
        "id": "assenze",
        "title": "Assenze",
        "icon": "📅",
        "color": "#06b6d4",
        "link_url": "/assenze/",
        "link_label": "Vai alle assenze",
    },
    {
        "id": "timbri",
        "title": "Timbri",
        "icon": "🕐",
        "color": "#64748b",
        "link_url": "",
        "link_label": "Vai ai timbri",
    },
    {
        "id": "dpi",
        "title": "DPI richiesti",
        "icon": "🦺",
        "color": "#16a34a",
        "link_url": "/dpi/gestione/",
        "link_label": "Vai ai DPI",
    },
]

_WIDGET_DEFAULT_ORDER = [w["id"] for w in STAT_WIDGETS]


def _can_view_stats(request) -> bool:
    """Verifica se l'utente corrente può visualizzare la sezione statistiche."""
    if request.user.is_superuser:
        return True
    perm = AnagraficaStatPermission.get_instance()
    if perm.accesso == AnagraficaStatPermission.ACCESSO_TUTTI:
        return True
    legacy_user = get_legacy_user(request.user)
    if is_legacy_admin(legacy_user):
        return True
    if perm.accesso == AnagraficaStatPermission.ACCESSO_ADMIN:
        return False
    # ACCESSO_RUOLI: controlla se il ruolo dell'utente è nella lista
    if legacy_user and legacy_user.ruolo_id is not None:
        return int(legacy_user.ruolo_id) in [int(r) for r in (perm.ruolo_ids or [])]
    return False


def _check_hr_permission(request) -> bool:
    """Verifica se l'utente può vedere i dati HR riservati (IBAN, CF, disabilità)."""
    if request.user.is_superuser:
        return True
    perm = AnagraficaHRPermission.get_instance()
    if perm.accesso == AnagraficaHRPermission.ACCESSO_TUTTI:
        return True
    legacy_user = get_legacy_user(request.user)
    if is_legacy_admin(legacy_user):
        if perm.accesso in (AnagraficaHRPermission.ACCESSO_ADMIN, AnagraficaHRPermission.ACCESSO_TUTTI):
            return True
        # ACCESSO_RUOLI: legacy admin controlla comunque la lista ruoli
    if perm.accesso == AnagraficaHRPermission.ACCESSO_ADMIN:
        return False
    if legacy_user and legacy_user.ruolo_id is not None:
        return int(legacy_user.ruolo_id) in [int(r) for r in (perm.ruolo_ids or [])]
    return False


def _check_skm_permission(request, code: str) -> bool:
    """Permesso canonico Skill Matrix MOD.187 (governabile in /admin-portale/acl-canonico/).

    Bypass per superuser/admin legacy; altrimenti richiede il grant del ruolo sul
    ``code`` canonico (``anagrafica.skillmatrix.view`` / ``.manage``). Allinea la
    guardia in-view alla decisione del middleware ACL canonico, così ciò che si
    imposta in ACL canonico governa davvero l'accesso (in dev/test il middleware è
    disattivo: qui resta l'unico controllo).
    """
    from core.acl_v2 import evaluate_permission_code_access
    try:
        legacy_user = get_legacy_user(request.user)
    except Exception:
        legacy_user = None
    return bool(evaluate_permission_code_access(
        permission_code=code, legacy_user=legacy_user, django_user=request.user,
    ).get("allowed"))


def _can_view_visite_mediche(request) -> bool:
    """Verifica se l'utente può vedere/registrare le visite mediche.

    Le visite mediche contengono dati sanitari sensibili (idoneità, prescrizioni).
    Il default del singleton è ADMIN: solo superuser + amministratori legacy.
    """
    if request.user.is_superuser:
        return True
    perm = AnagraficaVisiteMedichePermission.get_instance()
    if perm.accesso == AnagraficaVisiteMedichePermission.ACCESSO_TUTTI:
        return True
    legacy_user = get_legacy_user(request.user)
    if is_legacy_admin(legacy_user):
        if perm.accesso in (
            AnagraficaVisiteMedichePermission.ACCESSO_ADMIN,
            AnagraficaVisiteMedichePermission.ACCESSO_TUTTI,
        ):
            return True
        # ACCESSO_RUOLI: legacy admin controlla comunque la lista ruoli
    if perm.accesso == AnagraficaVisiteMedichePermission.ACCESSO_ADMIN:
        return False
    if legacy_user and legacy_user.ruolo_id is not None:
        return int(legacy_user.ruolo_id) in [int(r) for r in (perm.ruolo_ids or [])]
    return False


def _can_view_formazione(request) -> bool:
    """Verifica se l'utente può visualizzare la sezione formazione HR.
    Usa AnagraficaFormazionePermission.accesso_visualizzazione (singleton).
    """
    if request.user.is_superuser:
        return True
    perm = AnagraficaFormazionePermission.get_instance()
    if perm.accesso_visualizzazione == AnagraficaFormazionePermission.ACCESSO_TUTTI:
        return True
    legacy_user = get_legacy_user(request.user)
    if is_legacy_admin(legacy_user):
        if perm.accesso_visualizzazione in (
            AnagraficaFormazionePermission.ACCESSO_ADMIN,
            AnagraficaFormazionePermission.ACCESSO_TUTTI,
        ):
            return True
        # ACCESSO_RUOLI: legacy admin controlla comunque la lista ruoli
    if perm.accesso_visualizzazione == AnagraficaFormazionePermission.ACCESSO_ADMIN:
        return False
    if legacy_user and legacy_user.ruolo_id is not None:
        return int(legacy_user.ruolo_id) in [int(r) for r in (perm.ruoli_autorizzati_json or [])]
    return False


def _compute_widget_counts(dip: dict) -> dict[str, int]:
    """Calcola i contatori per ogni widget statistiche del dipendente."""
    legacy_id = int(dip.get("id") or 0)
    nome = str(dip.get("nome") or "").strip()
    cognome = str(dip.get("cognome") or "").strip()
    nome_completo = f"{nome} {cognome}".strip()
    counts: dict[str, int] = {w["id"]: 0 for w in STAT_WIDGETS}

    # Ticket aperti e totali
    try:
        from tickets.models import StatoTicket, Ticket
        qs = Ticket.objects.filter(richiedente_legacy_user_id=legacy_id) if legacy_id else Ticket.objects.none()
        counts["tickets_totali"] = qs.count()
        counts["tickets_aperti"] = qs.filter(stato=StatoTicket.APERTA).count()
    except Exception:
        logger.exception("Errore conteggio ticket per dipendente %s", legacy_id)

    # Anomalie (legacy SQL Server — solo se connesso)
    try:
        from django.db import connections
        with connections["default"].cursor() as cur:
            # Cerca per nome in capo_commessa (campo testo con nome)
            cur.execute(
                "SELECT COUNT(*) FROM anomalie WHERE UPPER(COALESCE(capo_commessa,'')) LIKE UPPER(%s)",
                [f"%{nome_completo}%"],
            )
            row = cur.fetchone()
            counts["anomalie"] = int(row[0]) if row else 0
    except Exception:
        logger.debug("Impossibile contare anomalie per dipendente %s (tabella legacy assente in dev)", legacy_id)

    # Diario preposto
    try:
        from diario_preposto.models import SegnalazionePreposto
        counts["diario_preposto"] = SegnalazionePreposto.objects.filter(
            Q(chi_segnala__icontains=nome_completo) | Q(preposto__icontains=nome_completo)
        ).count()
    except Exception:
        logger.exception("Errore conteggio diario preposto per dipendente %s", legacy_id)

    # Rilevazioni sicurezza
    try:
        from rilevazione_incidenti.models import RilevazioneIncidente
        counts["rilevazioni"] = RilevazioneIncidente.objects.filter(
            nominativo__icontains=nome_completo
        ).count()
    except Exception:
        logger.exception("Errore conteggio rilevazioni per dipendente %s", legacy_id)

    # Assenze (legacy SQL Server)
    try:
        from django.db import connections
        with connections["default"].cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM assenze WHERE UPPER(COALESCE(copia_nome,'')) LIKE UPPER(%s)",
                [f"%{nome_completo}%"],
            )
            row = cur.fetchone()
            counts["assenze"] = int(row[0]) if row else 0
    except Exception:
        logger.debug("Impossibile contare assenze per dipendente %s (tabella legacy assente in dev)", legacy_id)

    # Timbri
    try:
        from timbri.models import OperatoreTimbri, RegistroTimbro
        operatore = OperatoreTimbri.objects.filter(legacy_anagrafica_id=legacy_id).first()
        if operatore:
            counts["timbri"] = RegistroTimbro.objects.filter(operatore=operatore).count()
    except Exception:
        logger.exception("Errore conteggio timbri per dipendente %s", legacy_id)

    # DPI richiesti
    try:
        from dpi.models import RichiestaDPI
        counts["dpi"] = RichiestaDPI.objects.filter(richiedente_legacy_id=legacy_id).count() if legacy_id else 0
    except Exception:
        logger.exception("Errore conteggio DPI per dipendente %s", legacy_id)

    return counts


# ---------------------------------------------------------------------------
# Scheda dettaglio dipendente — helper assenze
# ---------------------------------------------------------------------------

def _query_assenze_dipendente(dip: dict) -> tuple[list[dict], bool]:
    """Read-only: assenze del dipendente (anno corrente + precedente).

    Match robusto, allineato al widget conteggio (`_compute_widget_counts`) e al
    modulo assenze (`assenze.views._load_events`):
    - `copia_nome` LIKE nome (ordini "Nome Cognome" e "Cognome Nome");
    - `utente_id` diretto su `assenze`, se la colonna esiste;
    - JOIN `dipendenti` via `dipendente_id`, solo se la colonna esiste.

    La sola JOIN su `assenze.dipendente_id` era fragile: quella colonna può
    mancare in prod (referenziarla mandava in errore l'intera query → 42S22 →
    catch → lista vuota) o restare NULL (INNER JOIN senza match), per cui il tab
    mostrava sempre "nessuna assenza" mentre il widget conteggio — che filtra per
    `copia_nome` — ne riportava di presenti.

    Ritorna `(lista, no_link)` dove `no_link` indica che il dipendente non è
    identificabile (né nome né utente_id). Chiavi dict: data_inizio, data_fine,
    tipo_assenza, moderation_status.
    """
    import datetime as _dt
    from django.utils import timezone as _tz

    nome = str(dip.get("nome") or "").strip()
    cognome = str(dip.get("cognome") or "").strip()
    utente_id = int(dip.get("utente_id") or 0) or None
    nome_cognome = f"{nome} {cognome}".strip()
    cognome_nome = f"{cognome} {nome}".strip()

    if not nome_cognome and not utente_id:
        return [], True

    assenze_cols = legacy_table_columns("assenze")
    if not assenze_cols:
        # Tabella legacy non disponibile (es. dev senza SQL Server): non è un
        # problema di linking, semplicemente non ci sono dati da mostrare.
        return [], False

    or_clauses: list[str] = []
    params: list = []

    if "copia_nome" in assenze_cols and nome_cognome:
        or_clauses.append("UPPER(COALESCE(a.copia_nome,'')) LIKE UPPER(%s)")
        params.append(f"%{nome_cognome}%")
        if cognome_nome != nome_cognome:
            or_clauses.append("UPPER(COALESCE(a.copia_nome,'')) LIKE UPPER(%s)")
            params.append(f"%{cognome_nome}%")

    if utente_id and "utente_id" in assenze_cols:
        or_clauses.append("a.utente_id = %s")
        params.append(utente_id)

    join = ""
    if utente_id and "dipendente_id" in assenze_cols and legacy_table_columns("dipendenti"):
        join = " LEFT JOIN dipendenti d ON d.id = a.dipendente_id "
        or_clauses.append("d.utente_id = %s")
        params.append(utente_id)

    if not or_clauses:
        return [], False

    data_da = _dt.date(_tz.localdate().year - 1, 1, 1)
    where = "(" + " OR ".join(or_clauses) + ") AND a.data_inizio >= %s"
    params.append(data_da)
    sql = f"""
        SELECT a.data_inizio, a.data_fine, a.tipo_assenza, a.moderation_status
        FROM assenze a
        {join}
        WHERE {where}
        ORDER BY a.data_inizio DESC
    """
    try:
        with connections["default"].cursor() as cur:
            cur.execute(sql, params)
            cols = [str(c[0]) for c in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()], False
    except Exception:
        logger.warning("_query_assenze_dipendente: query assenze fallita (dip_id=%s)", dip.get("id"))
        return [], False


def _assenza_tipo_meta(tipo) -> tuple[str, str]:
    """Ritorna ``(icona, accent)`` per un tipo di assenza, match per keyword.

    ``accent`` ∈ {ferie, malattia, permesso, congedo, altro} → mappa sulle
    classi CSS ``dp-abs-*`` usate dalla scheda dipendente (card riepilogo e
    chip della tabella storico). Tollerante a stringhe legacy eterogenee.
    """
    t = str(tipo or "").lower()
    if "feri" in t:
        return "🏖️", "ferie"
    if "malatt" in t or "infort" in t or "donaz" in t or "sangue" in t:
        return "🩺", "malattia"
    if "matern" in t or "patern" in t or "conged" in t or "allatt" in t or "104" in t:
        return "👶", "congedo"
    if "festiv" in t:
        return "🎉", "permesso"
    if "rol" in t or "permess" in t or "recuper" in t:
        return "⏱️", "permesso"
    if "sciop" in t:
        return "✊", "altro"
    return "🗓️", "altro"


# Nomi mese indicizzabili per numero (1-12). NB: più sotto nel modulo esiste un
# altro globale `_MESI_IT` (dict nome→numero, usato dall'import cedolini) che,
# essendo definito dopo, sovrascriverebbe questa tupla a livello di modulo: per
# questo qui il nome è distinto (`_MESI_IT_NOMI`) ed evitiamo la collisione.
_MESI_IT_NOMI = (
    "", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
)


def _assenze_cronologia(assenze_list: list[dict]) -> list[dict]:
    """Raggruppa le assenze (già ordinate per data_inizio DESC) in Anno → Mese.

    Ritorna ``[{anno, totale, mesi: [{mese_label, items: [...]}]}]``. Si appoggia
    all'ordinamento decrescente della query (anni e mesi contigui) per costruire
    i bucket in un solo passaggio, senza riordinare. Nomi mese in italiano via
    ``_MESI_IT_NOMI`` per non dipendere dal locale dei template.
    """
    crono: list[dict] = []
    for _a in assenze_list:
        _d_i = _a.get("data_inizio")
        if not _d_i:
            continue
        _y, _m = _d_i.year, _d_i.month
        if not crono or crono[-1]["anno"] != _y:
            crono.append({"anno": _y, "totale": 0, "mesi": []})
        anno = crono[-1]
        anno["totale"] += 1
        if not anno["mesi"] or anno["mesi"][-1]["mese_num"] != _m:
            _label = _MESI_IT_NOMI[_m] if 1 <= _m <= 12 else ""
            anno["mesi"].append({"mese_num": _m, "mese_label": _label, "items": []})
        anno["mesi"][-1]["items"].append(_a)
    return crono


# ---------------------------------------------------------------------------
# Scheda dettaglio dipendente
# ---------------------------------------------------------------------------

@login_required
def dipendente_detail(request, legacy_id: int):
    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    can_stats = _can_view_stats(request)
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    can_hr = _check_hr_permission(request)

    # Anagrafica civile e aziendale estese
    civile = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    aziendale = DipendenteAnagraficaAziendale.objects.filter(legacy_anagrafica_id=legacy_id).first()
    civile_foto_url = _foto_dipendente_url(civile, legacy_id)
    form_civile = AnagraficaCivileForm(instance=civile) if is_admin else None
    form_aziendale = AnagraficaAziendaleForm(instance=aziendale) if is_admin else None
    formset_figli = FiglioACaricoFormSet(instance=civile, prefix="figli") if is_admin else None
    figli_list = list(civile.figli.all()) if civile else []

    # Widget layout viewer
    layout_obj = DipendenteStatLayout.objects.filter(viewer_user_id=request.user.id).first()
    hidden_ids: list[str] = list(layout_obj.hidden) if layout_obj else []
    order_ids: list[str] = list(layout_obj.order) if layout_obj else []

    # Costruisce lista widget ordinata
    all_widget_ids = [w["id"] for w in STAT_WIDGETS]
    if order_ids:
        ordered = [wid for wid in order_ids if wid in all_widget_ids]
        ordered += [wid for wid in all_widget_ids if wid not in ordered]
    else:
        ordered = all_widget_ids[:]

    # Contatori
    widget_counts: dict[str, int] = {}
    if can_stats:
        widget_counts = _compute_widget_counts(dip)

    # Costruisce URL filtrati per dipendente
    from urllib.parse import urlencode
    from django.urls import reverse as _reverse

    _nome = str(dip.get("nome") or "").strip()
    _cognome = str(dip.get("cognome") or "").strip()
    _nome_completo = f"{_cognome} {_nome}".strip()
    _alias = str(dip.get("aliasusername") or "").strip()

    def _qs(**params):
        return "?" + urlencode(params) if params else ""

    try:
        _ticket_base = _reverse("tickets:gestione_list")
    except Exception:
        _ticket_base = "/tickets/gestione/"
    try:
        _diario_base = _reverse("diario_preposto:lista")
    except Exception:
        _diario_base = "/diario-preposto/"
    try:
        _rilevazioni_base = _reverse("rilevazione_incidenti:lista")
    except Exception:
        _rilevazioni_base = "/rilevazione-incidenti/"
    try:
        _timbri_base = _reverse("timbri:operatore_detail_by_legacy", args=[legacy_id])
    except Exception:
        _timbri_base = ""

    _widget_links: dict[str, str] = {
        "tickets_aperti": _ticket_base + _qs(q=_nome_completo, stato="APERTA"),
        "tickets_totali": _ticket_base + _qs(q=_nome_completo),
        "anomalie": "/gestione-anomalie/",
        "diario_preposto": _diario_base + _qs(q=_nome_completo),
        "rilevazioni": _rilevazioni_base + _qs(q=_nome_completo),
        "assenze": "/assenze/",
        "timbri": _timbri_base,
        "dpi": "/dpi/gestione/" + _qs(q=_nome_completo),
    }

    # Costruisce lista widget con dati
    widget_map = {w["id"]: w for w in STAT_WIDGETS}
    widgets_visible = []
    widgets_hidden = []
    for wid in ordered:
        w = dict(widget_map[wid])
        w["count"] = widget_counts.get(wid, 0)
        if wid in _widget_links:
            w["link_url"] = _widget_links[wid]
        if wid in hidden_ids:
            widgets_hidden.append(w)
        else:
            widgets_visible.append(w)

    # Ruoli operativi
    ruoli_assegnati = list(
        DipendenteRuoloOperativo.objects.filter(legacy_anagrafica_id=legacy_id)
        .select_related("ruolo", "assegnato_da")
        .order_by("ruolo__nome")
    )
    ruoli_disponibili = RuoloOperativo.objects.filter(is_active=True).exclude(
        id__in=[a.ruolo_id for a in ruoli_assegnati]
    )

    # Mansioni catalogo
    mansioni_catalogo = list(Mansione.objects.filter(is_active=True).order_by("nome"))

    # Qualifiche dipendente
    from django.utils import timezone as tz
    from datetime import timedelta
    oggi = tz.localdate()
    qualifiche_dip = list(
        DipendenteQualifica.objects.filter(legacy_anagrafica_id=legacy_id)
        .select_related("tipo", "record_formazione", "record_formazione__corso", "sessione")
        .prefetch_related("storico", "tipo__corsi")
        .order_by("data_scadenza", "tipo__nome")
    )
    tipi_qualifica = list(TipoQualifica.objects.filter(is_active=True).order_by("categoria", "nome"))

    # DPI dipendente
    dpi_consegnati = []
    dpi_richieste = []
    try:
        from dpi.models import RichiestaDPI
        dpi_richieste = list(
            RichiestaDPI.objects.filter(richiedente_legacy_id=legacy_id)
            .select_related("categoria", "consegna")
            .order_by("-created_at")[:20]
        )
        dpi_consegnati = [r for r in dpi_richieste if r.stato == "CONSEGNATA" and hasattr(r, "consegna")]
    except Exception:
        logger.exception("Errore caricamento DPI per dipendente %s", legacy_id)

    # PDF consegne DPI archiviati: mappa consegna_id -> documento per linkare dal tab
    dpi_consegna_doc_map: dict[int, int] = {}
    try:
        dpi_docs = DocumentoDipendente.objects.filter(
            legacy_anagrafica_id=legacy_id,
            tipo=DocumentoDipendente.Tipo.DPI_CONSEGNA,
            oggetto_riferimento_tipo="dpi.consegna",
        ).values_list("oggetto_riferimento_id", "id")
        dpi_consegna_doc_map = {int(c): int(d) for c, d in dpi_docs if c is not None}
    except Exception:
        logger.exception("Errore caricamento documenti DPI per dipendente %s", legacy_id)

    # Visite mediche (dato sanitario sensibile: gating su _can_view_visite_mediche)
    can_view_visite = _can_view_visite_mediche(request)
    visite_stato_list: list = []
    visite_storico_list: list = []
    tipi_visita_attivi: list = []
    if can_view_visite:
        try:
            visite_stato_list = stato_visite(legacy_id)
            visite_storico_list = visite_storico(legacy_id)
            tipi_visita_attivi = list(
                TipoVisitaMedica.objects.filter(is_active=True).order_by("nome")
            )
        except Exception:
            logger.exception("Errore caricamento visite mediche per dipendente %s", legacy_id)

    # Documenti dipendente (spazio privato): riga in tab "Documenti"
    documenti_dipendente = []
    cartelle_documenti = []
    try:
        qs_doc = DocumentoDipendente.objects.filter(legacy_anagrafica_id=legacy_id).select_related("cartella")
        # Nasconde i referti sanitari a chi non ha il permesso visite
        if not can_view_visite:
            qs_doc = qs_doc.exclude(tipo=DocumentoDipendente.Tipo.VISITA_MEDICA_REFERTO)
        documenti_dipendente = list(qs_doc.order_by("-created_at")[:100])
        # Scheletro cartelle: solo quelle applicabili al dipendente (targeting per
        # reparto/ruoli operativi; cartella senza targeting = universale).
        from .models import DipendenteAnagraficaAziendale as _Az, DipendenteRuoloOperativo as _RO
        _area_nome = (_Az.objects.filter(legacy_anagrafica_id=legacy_id).values_list("area", flat=True).first() or "")
        _ruoli_ids = set(_RO.objects.filter(legacy_anagrafica_id=legacy_id).values_list("ruolo_id", flat=True))
        cartelle_documenti = [
            c for c in CartellaDocumentoDipendente.objects
            .filter(attiva=True).prefetch_related("reparti", "ruoli_operativi").order_by("ordine", "nome")
            if c.si_applica(_area_nome, _ruoli_ids)
        ]
    except Exception:
        logger.exception("Errore caricamento documenti per dipendente %s", legacy_id)

    software_licenses = []
    try:
        from assets.models import SoftwareLicense

        legacy_user_id = int(dip.get("utente_id") or 0)
        license_qs = SoftwareLicense.objects.filter(
            Q(assigned_anagrafica_id=legacy_id)
            | (Q(assigned_legacy_user_id=legacy_user_id) if legacy_user_id else Q())
        ).select_related("asset").order_by("category", "vendor", "product_name", "id")
        software_licenses = list(license_qs)
    except Exception:
        logger.exception("Errore caricamento licenze software per dipendente %s", legacy_id)

    # Asset assegnati al dipendente e asset del reparto (se caporeparto)
    assets_assegnati: list = []
    assets_reparto: list = []
    reparti_capo: list = []
    try:
        from assets.models import Asset
        from core.models import RepartoCapoMapping

        utente_id_asset = int(dip.get("utente_id") or 0)
        if utente_id_asset:
            assets_assegnati = list(
                Asset.objects.filter(assigned_legacy_user_id=utente_id_asset)
                .exclude(status=Asset.STATUS_RETIRED)
                .select_related("asset_category")
                .order_by("name")
            )

        # Verifica se è caporeparto: cerca RepartoCapoMapping dove caporeparto
        # corrisponde a email, id numerico o nome dell'utente legacy
        capo_user = UtenteLegacy.objects.filter(id=utente_id_asset).first() if utente_id_asset else None
        if capo_user:
            capo_email = str(getattr(capo_user, "email", "") or "").strip().lower()
            capo_nome = str(getattr(capo_user, "nome", "") or "").strip()
            capo_filters = Q(caporeparto=str(utente_id_asset))
            if capo_email:
                capo_filters |= Q(caporeparto__iexact=capo_email)
            if capo_nome:
                capo_filters |= Q(caporeparto__iexact=capo_nome)
            reparti_capo = list(
                RepartoCapoMapping.objects.filter(capo_filters, is_active=True)
                .values_list("reparto", flat=True)
                .distinct()
            )
            if reparti_capo:
                assets_reparto = list(
                    Asset.objects.filter(assignment_reparto__in=reparti_capo)
                    .exclude(status=Asset.STATUS_RETIRED)
                    .select_related("asset_category")
                    .order_by("assignment_reparto", "name")
                )
    except Exception:
        logger.exception("Errore caricamento asset per dipendente %s", legacy_id)

    # Voci retributive (solo se l'utente ha accesso HR)
    retribuzioni_latest: list = []
    retribuzioni_timeline: list = []
    retribuzioni_importazione = None
    retribuzioni_data_competenza = None
    retribuzioni_has_changes = False
    if can_hr:
        _voci_q = Q(legacy_anagrafica_id=legacy_id)
        if civile and civile.codice_fiscale:
            _voci_q |= Q(tax_code=civile.codice_fiscale.strip().upper())
        voci_qs = VoceRetributiva.objects.filter(
            _voci_q
        ).select_related("importazione").order_by(
            "-data_competenza", "-importazione__data_importazione", "categoria", "pay_item_key"
        )
        voci_all = list(voci_qs)
        if voci_all:
            # Raggruppa per mese, separando CSV (solo più recente per mese) e manuali (sempre incluse).
            # Le voci manuali fanno override delle CSV con stesso pay_item_key.
            _by_month: dict = {}
            for v in voci_all:
                key = v.data_competenza
                if key not in _by_month:
                    _by_month[key] = {"csv_imp": None, "csv_voci": [], "manuale_voci": []}
                if v.manuale:
                    _by_month[key]["manuale_voci"].append(v)
                else:
                    if _by_month[key]["csv_imp"] is None:
                        _by_month[key]["csv_imp"] = v.importazione
                    if v.importazione_id == _by_month[key]["csv_imp"].id:
                        _by_month[key]["csv_voci"].append(v)

            def _merge(entry):
                manuale_keys = {v.pay_item_key for v in entry["manuale_voci"]}
                return [v for v in entry["csv_voci"] if v.pay_item_key not in manuale_keys] + entry["manuale_voci"]

            _months_desc = sorted(_by_month.keys(), reverse=True)
            if _months_desc:
                _latest_month = _months_desc[0]
                _latest_entry = _by_month[_latest_month]
                retribuzioni_data_competenza = _latest_month
                retribuzioni_importazione = _latest_entry["csv_imp"]  # può essere None se mese ha solo manuali
                retribuzioni_latest = _merge(_latest_entry)
                retribuzioni_has_changes = any(v.is_changed for v in retribuzioni_latest)
                retribuzioni_timeline = [(d, _merge(_by_month[d])) for d in _months_desc]

    _TOTALI_SORT = {"retribuzione di fatto": 0, "totale elementi variabili": 1, "rml": 2, "ral": 3}
    retribuzioni_fissi = [v for v in retribuzioni_latest if v.categoria == "fisso"]
    retribuzioni_variabili = [v for v in retribuzioni_latest if v.categoria == "variabile"]
    retribuzioni_totali = sorted(
        [v for v in retribuzioni_latest if v.categoria == "totale"],
        key=lambda v: _TOTALI_SORT.get(v.pay_item_key, 99),
    )
    retribuzioni_altri = [v for v in retribuzioni_latest if v.categoria == "altro"]

    # Ratei ferie/ROL/ex-festività da cedolini (solo se l'utente ha accesso HR)
    ratei_saldi: list = []
    if can_hr:
        _ratei_q = Q(legacy_anagrafica_id=legacy_id)
        if civile and civile.codice_fiscale:
            _ratei_q |= Q(tax_code=civile.codice_fiscale.strip().upper())
        ratei_saldi = list(
            SaldoCedolino.objects.filter(_ratei_q).order_by("-data_competenza")[:1]
        )

    # Assenze (read-only, ultimi 2 anni) — match per nome/utente_id, vedi helper
    assenze_list, assenze_no_link = _query_assenze_dipendente(dip)
    _summary_map: dict[str, int] = {}
    for _a in assenze_list:
        _d_i = _a.get("data_inizio")
        _d_f = _a.get("data_fine") or _d_i
        # Durata (giorni) + icona/accent per ogni riga, per la tabella storico
        _giorni_row = None
        if _d_i:
            try:
                _giorni_row = max(1, (_d_f - _d_i).days + 1) if _d_f and _d_f != _d_i else 1
            except Exception:
                _giorni_row = None
        _a["giorni"] = _giorni_row
        _ic, _acc = _assenza_tipo_meta(_a.get("tipo_assenza"))
        _a["icona"] = _ic
        _a["accent"] = _acc
        # Riepilogo: solo anno corrente e approvate
        if not _d_i:
            continue
        try:
            _year = _d_i.year if hasattr(_d_i, "year") else int(str(_d_i)[:4])
            if _year != oggi.year:
                continue
            if int(_a.get("moderation_status") or -1) != 0:
                continue
            _tipo = str(_a.get("tipo_assenza") or "Altro").strip() or "Altro"
            _summary_map[_tipo] = _summary_map.get(_tipo, 0) + (_giorni_row or 1)
        except Exception:
            pass

    # Riepilogo come lista ordinata (giorni desc) arricchita con icona/accent + totale
    assenze_summary_anno: list[dict] = []
    assenze_tot_anno = 0
    for _tipo, _giorni in sorted(_summary_map.items(), key=lambda kv: kv[1], reverse=True):
        _ic, _acc = _assenza_tipo_meta(_tipo)
        assenze_summary_anno.append({"tipo": _tipo, "giorni": _giorni, "icona": _ic, "accent": _acc})
        assenze_tot_anno += _giorni

    # Storico come cronologia raggruppata Anno → Mese
    assenze_cronologia = _assenze_cronologia(assenze_list)

    # Storico cambiamenti organizzativi (mansione, reparto, area, ruolo aziendale) — solo admin
    storico_cambiamenti: list = []
    if is_admin:
        storico_cambiamenti = list(
            DipendenteCambiamentoOrganizzativo.objects
            .filter(legacy_anagrafica_id=legacy_id)
            .select_related("created_by")
            .order_by("-data_effetto", "-created_at")[:50]
        )

    # Storico contrattuale
    storico_contratti: list = []
    tipologie_contratto: list = []
    livelli_catalogo: list = []
    livelli_json: str = "{}"
    tipi_qualifica_prof: list = []
    if can_hr:
        _contr_q = Q(legacy_anagrafica_id=legacy_id)
        tax_code_contr = civile.codice_fiscale.strip().upper() if civile and civile.codice_fiscale else None
        if tax_code_contr:
            _contr_q |= Q(tax_code=tax_code_contr)
        storico_contratti = list(
            StoricoContratto.objects.filter(_contr_q).order_by("-data_inizio", "-created_at")
        )
        tipologie_contratto = list(TipologiaContratto.objects.filter(is_active=True))
        _tipologie_map = {t.codice: t.nome for t in tipologie_contratto}
        for _c in storico_contratti:
            _c.tipologia_nome = _tipologie_map.get(_c.tipologia_contratto, _c.tipologia_contratto)
        livelli_catalogo = list(LivelloContrattuale.objects.filter(is_active=True))
        livelli_json = json.dumps({l.codice: l.descrizione for l in livelli_catalogo})
        tipi_qualifica_prof = [q for q in tipi_qualifica if q.categoria == TipoQualifica.CAT_PROFESSIONALE]

    # Formazione dipendente (gating su _can_view_formazione — accesso controllato)
    can_view_formazione_tab = _can_view_formazione(request)
    fm_storico: list = []
    fm_scadenze_urgenti: list = []
    fm_n_completati = 0
    fm_ore_totali = 0.0
    fm_n_attestati = 0
    # KPI formazione divisi per anno corrente / anni precedenti
    fm_anno_corrente = tz.localdate().year
    fm_kpi_anno_corrente = {"anno": fm_anno_corrente, "n": 0, "ore": 0.0}
    fm_kpi_precedenti_totale = {"n": 0, "ore": 0.0}
    fm_kpi_precedenti_per_anno: list = []
    if can_view_formazione_tab:
        try:
            # Carico TUTTI i completamenti (non più solo 30): lo storico mostra di
            # default i primi 30 (cap lato template) ma quando si filtra per anno
            # dai KPI il filtro deve poter agire su tutti i record senza limite.
            fm_storico = list(
                TrainingEmployeeRecord.objects.filter(legacy_anagrafica_id=legacy_id)
                .select_related("corso", "corso__piano")
                .order_by("-data_completamento")
            )
            # Aggregazione KPI sugli stessi record (nessuna seconda query)
            _per_anno: dict = {}
            for _rec in fm_storico:
                _data = _rec.data_completamento
                _anno = _data.year if _data else None
                _ore_f = float(_rec.ore_frequentate or 0)
                if _anno == fm_anno_corrente:
                    fm_kpi_anno_corrente["n"] += 1
                    fm_kpi_anno_corrente["ore"] += _ore_f
                else:
                    fm_kpi_precedenti_totale["n"] += 1
                    fm_kpi_precedenti_totale["ore"] += _ore_f
                    if _anno is not None:
                        _bucket = _per_anno.setdefault(_anno, {"anno": _anno, "n": 0, "ore": 0.0})
                        _bucket["n"] += 1
                        _bucket["ore"] += _ore_f
            fm_kpi_precedenti_per_anno = sorted(
                _per_anno.values(), key=lambda b: b["anno"], reverse=True
            )
            fm_n_completati = fm_kpi_anno_corrente["n"] + fm_kpi_precedenti_totale["n"]
            fm_ore_totali = fm_kpi_anno_corrente["ore"] + fm_kpi_precedenti_totale["ore"]
            fm_scadenze_urgenti = list(
                TrainingDeadline.objects.filter(
                    legacy_anagrafica_id=legacy_id,
                    stato_scadenza__in=["SCADUTO", "IN_SCADENZA_30", "IN_SCADENZA_90", "MAI_FREQUENTATO"],
                )
                .select_related("corso", "corso__piano")
                .order_by("data_scadenza")[:20]
            )
            fm_n_attestati = TrainingCertificate.objects.filter(
                legacy_anagrafica_id=legacy_id
            ).count()
        except Exception:
            logger.exception("Errore caricamento formazione per dipendente %s", legacy_id)

    offboarding_pratica_attiva = None
    offboarding_tasks = []
    offboarding_task_pending_count = 0
    offboarding_task_exception_count = 0
    offboarding_pratiche_recenti = []
    if is_admin:
        try:
            offboarding_pratica_attiva = (
                OffboardingPratica.objects
                .filter(
                    legacy_anagrafica_id=legacy_id,
                    stato__in=OffboardingPratica.STATI_APERTI,
                )
                .prefetch_related("tasks")
                .order_by("-created_at", "-id")
                .first()
            )
            if offboarding_pratica_attiva:
                offboarding_tasks = list(offboarding_pratica_attiva.tasks.all())
                offboarding_task_pending_count = sum(
                    1 for task in offboarding_tasks
                    if task.stato == OffboardingTask.STATO_DA_FARE
                )
                offboarding_task_exception_count = sum(
                    1 for task in offboarding_tasks
                    if task.stato == OffboardingTask.STATO_ECCEZIONE
                )
            offboarding_pratiche_recenti = list(
                OffboardingPratica.objects
                .filter(legacy_anagrafica_id=legacy_id)
                .exclude(stato__in=OffboardingPratica.STATI_APERTI)
                .order_by("-created_at", "-id")[:5]
            )
        except Exception:
            logger.exception("Errore caricamento pratiche offboarding per dipendente %s", legacy_id)

    # Onboarding strutturato (visibile a chi ha accesso HR)
    onboarding_pratica_attiva = None
    onboarding_pratiche_recenti = []
    if can_hr:
        try:
            onboarding_pratica_attiva = onboarding_service.pratica_aperta(legacy_id)
            onboarding_pratiche_recenti = list(
                OnboardingPratica.objects
                .filter(legacy_anagrafica_id=legacy_id)
                .exclude(stato__in=OnboardingPratica.STATI_APERTI)
                .order_by("-created_at", "-id")[:5]
            )
        except Exception:
            logger.exception("Errore caricamento pratiche onboarding per dipendente %s", legacy_id)

    # Catalogo reparti per dropdown + label caporeparto dall'aziendale
    reparti_catalog = list(Reparto.objects.filter(is_active=True).order_by("nome"))
    reparto_corrente = (dip.get("reparto") or "").strip()
    reparto_in_catalog = reparto_corrente and any(
        r.nome.strip().casefold() == reparto_corrente.casefold() for r in reparti_catalog
    )
    _dip_picker_map_detail = {item["id"]: item["label"] for item in _dipendenti_picker_rows()}
    caporeparto_label = _dip_picker_map_detail.get(aziendale.caporeparto_legacy_id, "") if aziendale and aziendale.caporeparto_legacy_id else ""
    reparto_autofill_json = json.dumps({
        r.nome: {
            "capo_label": _dip_picker_map_detail.get(r.caporeparto_legacy_id or 0, ""),
        }
        for r in reparti_catalog
    })

    _area_corrente_id = aziendale.area_aziendale_id if aziendale else None
    aree_by_reparto: dict[str, list[dict]] = {}
    for a in (
        AreaAziendale.objects.filter(Q(is_active=True) | Q(pk=_area_corrente_id))
        .select_related("reparto")
        .order_by("nome")
    ):
        if a.reparto_id is None:
            continue
        aree_by_reparto.setdefault(a.reparto.nome, []).append({"id": a.id, "nome": a.nome})
    aree_by_reparto_json = json.dumps(aree_by_reparto)

    # Anzianità di servizio (KPI scheda sintetica del Riepilogo). Calcolata dalla
    # prima assunzione (fallback su assunzione corrente) fino a oggi o alla
    # cessazione se il rapporto è chiuso.
    anzianita_label = ""
    _data_assunzione_anz = (aziendale.data_prima_assunzione or aziendale.data_assunzione_ultima) if aziendale else None
    if _data_assunzione_anz:
        _fine_anz = aziendale.data_cessazione if (aziendale and aziendale.data_cessazione) else oggi
        if _fine_anz >= _data_assunzione_anz:
            _tot_mesi = (_fine_anz.year - _data_assunzione_anz.year) * 12 + (_fine_anz.month - _data_assunzione_anz.month)
            if _fine_anz.day < _data_assunzione_anz.day:
                _tot_mesi -= 1
            _tot_mesi = max(0, _tot_mesi)
            _anni, _mesi = divmod(_tot_mesi, 12)
            if _anni and _mesi:
                anzianita_label = f"{_anni} ann{'o' if _anni == 1 else 'i'} e {_mesi} mes{'e' if _mesi == 1 else 'i'}"
            elif _anni:
                anzianita_label = f"{_anni} ann{'o' if _anni == 1 else 'i'}"
            elif _mesi:
                anzianita_label = f"{_mesi} mes{'e' if _mesi == 1 else 'i'}"
            else:
                anzianita_label = "< 1 mese"

    return render(request, "anagrafica/pages/dipendente_detail.html", {
        "dip": dip,
        "legacy_id": legacy_id,
        "can_stats": can_stats,
        "is_admin": is_admin,
        "can_hr": can_hr,
        "civile": civile,
        "civile_foto_url": civile_foto_url,
        "aziendale": aziendale,
        "form_civile": form_civile,
        "form_aziendale": form_aziendale,
        "formset_figli": formset_figli,
        "figli_list": figli_list,
        "reparti_catalog": reparti_catalog,
        "reparto_in_catalog": reparto_in_catalog,
        "caporeparto_label": caporeparto_label,
        "reparto_autofill_json": reparto_autofill_json,
        "aree_by_reparto_json": aree_by_reparto_json,
        "anzianita_label": anzianita_label,
        "widgets_visible": widgets_visible,
        "widgets_hidden": widgets_hidden,
        "all_widgets": STAT_WIDGETS,
        "ruoli_assegnati": ruoli_assegnati,
        "ruoli_disponibili": ruoli_disponibili,
        "mansioni_catalogo": mansioni_catalogo,
        "qualifiche_dip": qualifiche_dip,
        "tipi_qualifica": tipi_qualifica,
        "oggi": oggi,
        "oggi_plus60": oggi + timedelta(days=60),
        "dpi_richieste": dpi_richieste,
        "dpi_consegnati": dpi_consegnati,
        "dpi_consegna_doc_map": dpi_consegna_doc_map,
        "can_view_visite": can_view_visite,
        "visite_stato_list": visite_stato_list,
        "visite_storico_list": visite_storico_list,
        "tipi_visita_attivi": tipi_visita_attivi,
        "documenti_dipendente": documenti_dipendente,
        "cartelle_documenti": cartelle_documenti,
        "visita_esiti": VisitaMedica.Esito.choices,
        "software_licenses": software_licenses,
        "assets_assegnati": assets_assegnati,
        "assets_reparto": assets_reparto,
        "reparti_capo": reparti_capo,
        "retribuzioni_latest": retribuzioni_latest,
        "retribuzioni_fissi": retribuzioni_fissi,
        "retribuzioni_variabili": retribuzioni_variabili,
        "retribuzioni_totali": retribuzioni_totali,
        "retribuzioni_altri": retribuzioni_altri,
        "retribuzioni_timeline": retribuzioni_timeline,
        "retribuzioni_importazione": retribuzioni_importazione,
        "retribuzioni_data_competenza": retribuzioni_data_competenza,
        "retribuzioni_has_changes": retribuzioni_has_changes,
        "storico_cambiamenti": storico_cambiamenti,
        "storico_contratti": storico_contratti,
        "tipologie_contratto": tipologie_contratto,
        "livelli_catalogo": livelli_catalogo,
        "livelli_json": livelli_json,
        "tipi_qualifica_prof": tipi_qualifica_prof,
        "ratei_saldi": ratei_saldi,
        "assenze_list": assenze_list,
        "assenze_summary_anno": assenze_summary_anno,
        "assenze_tot_anno": assenze_tot_anno,
        "assenze_cronologia": assenze_cronologia,
        "assenze_no_link": assenze_no_link,
        # Formazione
        "can_view_formazione_tab": can_view_formazione_tab,
        "fm_storico": fm_storico,
        "fm_scadenze_urgenti": fm_scadenze_urgenti,
        "fm_n_completati": fm_n_completati,
        "fm_ore_totali": fm_ore_totali,
        "fm_n_attestati": fm_n_attestati,
        "fm_kpi_anno_corrente": fm_kpi_anno_corrente,
        "fm_kpi_precedenti_totale": fm_kpi_precedenti_totale,
        "fm_kpi_precedenti_per_anno": fm_kpi_precedenti_per_anno,
        "fm_storico_cap": 30,
        # Offboarding
        "offboarding_pratica_attiva": offboarding_pratica_attiva,
        "offboarding_tasks": offboarding_tasks,
        "offboarding_task_pending_count": offboarding_task_pending_count,
        "offboarding_task_exception_count": offboarding_task_exception_count,
        "offboarding_pratiche_recenti": offboarding_pratiche_recenti,
        "offboarding_motivo_choices": OffboardingPratica.MOTIVO_CHOICES,
        "offboarding_restituzioni_labels": OFFBOARDING_RESTITUZIONI_LABELS,
        # Onboarding
        "onboarding_pratica_attiva": onboarding_pratica_attiva,
        "onboarding_pratiche_recenti": onboarding_pratiche_recenti,
    })


# ---------------------------------------------------------------------------
# Stampa scheda dipendente — anagrafica civile + aziendale + dati bancari
# ---------------------------------------------------------------------------

@login_required
def dipendente_print(request, legacy_id: int):
    """Pagina stampa scheda dipendente completa.

    Aggrega in un'unica vista A4-friendly:
      - dati legacy base (nome, cognome, username, reparto, mansione, email);
      - anagrafica civile (residenza, domicilio, CF, titolo studio, contatti
        privati, patente);
      - anagrafica aziendale (badge, area, ruolo, contratto, livello, date
        assunzione/cessazione/prova, taglie DPI, consenso privacy);
      - dati bancari (nome banca, IBAN, intestatario) — gated da
        `_check_hr_permission`; per gli utenti senza permesso l'IBAN appare
        mascherato e il blocco è etichettato come riservato.

    La pagina ha un'action bar (non stampata) con bottone "Stampa" che invoca
    `window.print()`. CSS `@media print` rimuove sfondi e action bar.
    """
    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    civile = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    aziendale = DipendenteAnagraficaAziendale.objects.filter(legacy_anagrafica_id=legacy_id).first()
    civile_foto_url = _foto_dipendente_url(civile, legacy_id)

    can_hr = _check_hr_permission(request)

    return render(request, "anagrafica/pages/dipendente_print.html", {
        "dip": dip,
        "civile": civile,
        "aziendale": aziendale,
        "civile_foto_url": civile_foto_url,
        "can_hr": can_hr,
        "legacy_id": legacy_id,
    })


# ---------------------------------------------------------------------------
# Libretto formativo dipendente — pagina stampa per audit ISO / uscita
# ---------------------------------------------------------------------------

@login_required
def dipendente_libretto_formativo(request, legacy_id: int):
    """Curriculum formativo completo del dipendente, A4-friendly.

    Aggrega lo storico completamenti (`TrainingEmployeeRecord`, usando i campi
    snapshot per integrità storica), gli attestati (`TrainingCertificate`) e
    lo stato corrente degli obblighi (`TrainingDeadline` con is_required).
    Stesso pattern di `dipendente_print`: action bar non stampata +
    `window.print()`. La generazione è tracciata in `TrainingExportLog`
    (tipo STORICO_DIP).
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la formazione.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    # Variante PDF server-side (stessa fonte dati): ?formato=pdf
    formato = (request.GET.get("formato") or "").strip().lower()
    if formato == "pdf":
        from .services.attestato_pdf import build_libretto_pdf_bytes
        try:
            pdf = build_libretto_pdf_bytes(legacy_id)
        except Exception:
            logger.exception("Errore generazione PDF libretto %s", legacy_id)
            messages.error(request, "Errore nella generazione del PDF del libretto.")
            return redirect("anagrafica:dipendente_libretto_formativo", legacy_id=legacy_id)
        try:
            TrainingExportLog.objects.create(
                tipo="STORICO_DIP",
                filtri_json={"legacy_anagrafica_id": legacy_id, "formato": "libretto_pdf"},
                righe_esportate=0,
                generato_da=request.user,
                ip_address=request.META.get("REMOTE_ADDR") or None,
            )
        except Exception:
            logger.exception("Errore TrainingExportLog per libretto PDF")
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = f'inline; filename="libretto_formativo_{legacy_id}.pdf"'
        return resp

    record_storici = list(
        TrainingEmployeeRecord.objects
        .filter(legacy_anagrafica_id=legacy_id)
        .select_related("attestato", "corso")
        .order_by("-data_completamento", "-created_at")
    )
    ore_totali = sum((r.ore_frequentate or 0) for r in record_storici)

    from django.utils import timezone as tz
    oggi = tz.localdate()
    obblighi = list(
        TrainingDeadline.objects
        .filter(legacy_anagrafica_id=legacy_id, is_required=True)
        .select_related("corso")
        .order_by("data_scadenza")
    )

    try:
        TrainingExportLog.objects.create(
            tipo="STORICO_DIP",
            filtri_json={"legacy_anagrafica_id": legacy_id, "formato": "libretto_html"},
            righe_esportate=len(record_storici),
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore registrazione TrainingExportLog per libretto formativo")

    from .services.attestato_pdf import RIFERIMENTO_LIBRETTO
    doc_libretto = (
        DocumentoDipendente.objects
        .filter(oggetto_riferimento_tipo=RIFERIMENTO_LIBRETTO, oggetto_riferimento_id=legacy_id)
        .order_by("-id").first()
    )
    return render(request, "anagrafica/pages/dipendente_libretto.html", {
        "dip": dip,
        "legacy_id": legacy_id,
        "record_storici": record_storici,
        "ore_totali": ore_totali,
        "obblighi": obblighi,
        "oggi": oggi,
        "can_edit": _can_edit_formazione(request),
        "doc_libretto": doc_libretto,
    })


@login_required
@require_POST
def libretto_salva_box(request, legacy_id: int):
    """Genera e salva (manualmente) il libretto formativo PDF nel box del dipendente.

    Un solo libretto per dipendente: la copia viene sostituita ad ogni salvataggio
    (è una fotografia aggiornata del curriculum). Gated dal permesso di modifica.
    """
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per archiviare il libretto.")
        return redirect("anagrafica:dipendente_libretto_formativo", legacy_id=legacy_id)
    from .services.attestato_pdf import archivia_libretto
    try:
        doc = archivia_libretto(legacy_id, user=request.user)
        messages.success(request, f"Libretto formativo archiviato nel box ({doc.nome_originale}).")
    except Exception:
        logger.exception("Archiviazione libretto fallita per %s", legacy_id)
        messages.error(request, "Errore durante l'archiviazione del libretto.")
    return redirect("anagrafica:dipendente_libretto_formativo", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Attestato di formazione — foglio A4 autogenerato (layout email NOVICROM HUB)
# ---------------------------------------------------------------------------

@login_required
def attestato_formazione(request, record_id: int):
    """Attestato autogenerato per un singolo completamento corso.

    Foglio A4 stampabile nello stile delle email NOVICROM HUB (header navy con
    logo + banda arancio). Vale per corsi, qualifiche e formazione interna
    generica ("altro"): il tipo viene derivato dalla qualifica àncora del corso.
    Riporta i dati del corso e del dipendente e due blocchi firma —
    Responsabile del corso e Dipendente. Nulla viene scritto: l'attestato si
    autogenera dal record di completamento (campi snapshot per stabilità
    storica). La generazione è tracciata in `TrainingExportLog` (tipo ATTESTATO).
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la formazione.")
        return redirect("anagrafica:formazione_dashboard")

    record = get_object_or_404(
        TrainingEmployeeRecord.objects.select_related(
            "corso", "corso__piano", "corso__qualifica",
            "sessione", "sessione__docente",
        ),
        pk=record_id,
    )

    # Derivazione condivisa con il builder PDF (tipo, responsabile, nominativo,
    # sede, numero, dati anagrafici) — un'unica fonte di verità.
    from .services.attestato_pdf import build_attestato_context, _documento_esistente
    ctx = build_attestato_context(record)
    legacy_id = ctx["legacy_id"]

    # Variante "stampa": layout sobrio a basso consumo d'inchiostro (B/N),
    # mantenendo la versione a colori come default.
    stile = (request.GET.get("stile") or "").strip().lower()
    is_stampa = stile == "stampa"
    template_name = (
        "anagrafica/pages/attestato_formazione_stampa.html" if is_stampa
        else "anagrafica/pages/attestato_formazione.html"
    )

    try:
        TrainingExportLog.objects.create(
            tipo="ATTESTATO",
            filtri_json={
                "record_id": record.pk,
                "legacy_anagrafica_id": legacy_id,
                "formato": "attestato_stampa_html" if is_stampa else "attestato_html",
            },
            righe_esportate=1,
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore registrazione TrainingExportLog per attestato record %s", record_id)

    return render(request, template_name, {
        **ctx,
        "stile": stile,
        "can_edit": _can_edit_formazione(request),
        "doc_archiviato": _documento_esistente(record),
    })


@login_required
def attestato_impostazioni(request):
    """Impostazioni del template attestato di formazione (singleton).

    Gestione dei testi fissi (intestazioni, formule, etichette firma, nota
    legale, logo) e del toggle privacy dei dati personali, da Impostazioni
    Anagrafica HR. Gated dallo stesso permesso di modifica della formazione.
    """
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare le impostazioni della formazione.")
        return redirect("anagrafica:formazione_dashboard")

    from .services.attestato_pdf import RIFERIMENTO_TIPO, archivia_attestato

    cfg = AttestatoFormazioneConfig.get_instance()
    form = AttestatoFormazioneConfigForm(instance=cfg)

    if request.method == "POST":
        action = (request.POST.get("action") or "save").strip()

        # ── Azioni di gestione archivio report ───────────────────────────
        if action == "backfill":
            # Genera e archivia gli attestati mancanti per i completamenti
            # esistenti (idempotente, bounded). Fail-safe per singolo record.
            mancanti = list(
                TrainingEmployeeRecord.objects.exclude(
                    pk__in=DocumentoDipendente.objects.filter(
                        oggetto_riferimento_tipo=RIFERIMENTO_TIPO
                    ).values_list("oggetto_riferimento_id", flat=True)
                ).order_by("-data_completamento", "-id")[:500]
            )
            ok = err = 0
            for rec in mancanti:
                try:
                    archivia_attestato(rec, cfg=cfg, user=request.user)
                    ok += 1
                except Exception:
                    err += 1
                    logger.exception("Backfill attestato fallito per record %s", rec.pk)
            messages.success(
                request,
                f"Archiviazione completata: {ok} attestati generati"
                + (f", {err} errori" if err else "")
                + (". Limite di 500 per esecuzione: rilancia se restano completamenti." if len(mancanti) >= 500 else "."),
            )
            return redirect("anagrafica:attestato_impostazioni")

        if action == "purge":
            qs = DocumentoDipendente.objects.filter(oggetto_riferimento_tipo=RIFERIMENTO_TIPO)
            n = qs.count()
            for doc in qs:
                try:
                    doc.file.delete(save=False)
                except Exception:
                    logger.warning("File attestato non eliminabile (doc %s)", doc.pk, exc_info=True)
            qs.delete()
            try:
                from core.audit import log_action
                log_action(request, "ATTESTATI_ARCHIVIO_PURGE", "anagrafica", {"eliminati": n})
            except Exception:
                logger.warning("Audit ATTESTATI_ARCHIVIO_PURGE fallito", exc_info=True)
            messages.success(request, f"Archivio attestati svuotato: {n} documenti eliminati.")
            return redirect("anagrafica:attestato_impostazioni")

        # ── Salvataggio impostazioni (default) ───────────────────────────
        form = AttestatoFormazioneConfigForm(request.POST, instance=cfg)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Impostazioni attestato salvate.")
            return redirect("anagrafica:attestato_impostazioni")
        messages.error(request, "Controlla i campi evidenziati.")

    # Statistiche archivio (per la sezione "Gestione report salvati").
    from datetime import date as _date, timedelta as _timedelta
    arch_qs = DocumentoDipendente.objects.filter(oggetto_riferimento_tipo=RIFERIMENTO_TIPO)
    n_archiviati = arch_qs.count()
    ultimo = arch_qs.order_by("-created_at").values_list("created_at", flat=True).first()
    n_completamenti = TrainingEmployeeRecord.objects.count()

    # Stato conservazione GDPR (retention_until calcolata su DocumentoDipendente).
    # Il command `cleanup_expired_documents` elimina solo i doc scaduti di cessati.
    _oggi = _date.today()
    n_retention_scaduta = arch_qs.filter(retention_until__lt=_oggi).count()
    n_retention_vicina = arch_qs.filter(
        retention_until__gte=_oggi, retention_until__lte=_oggi + _timedelta(days=90)
    ).count()

    # Record di esempio per il pulsante "anteprima" (il più recente disponibile).
    sample = TrainingEmployeeRecord.objects.order_by("-data_completamento", "-id").first()
    return render(request, "anagrafica/pages/attestato_impostazioni.html", {
        "form": form,
        "cfg": cfg,
        "sample_record_id": sample.pk if sample else None,
        "n_archiviati": n_archiviati,
        "n_completamenti": n_completamenti,
        "n_mancanti": max(0, n_completamenti - n_archiviati),
        "ultimo_archiviato": ultimo,
        "n_retention_scaduta": n_retention_scaduta,
        "n_retention_vicina": n_retention_vicina,
    })


@login_required
@require_POST
def attestato_salva_box(request, record_id: int):
    """Genera e salva (manualmente) l'attestato nel box documenti del dipendente.

    Pulsante «💾 Salva nel box» dalla pagina attestato. Forza la rigenerazione
    così l'utente ottiene sempre la versione aggiornata. Gated dal permesso di
    modifica formazione (scrive nello spazio documenti del dipendente).
    """
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per archiviare gli attestati.")
        return redirect("anagrafica:attestato_formazione", record_id=record_id)

    record = get_object_or_404(TrainingEmployeeRecord, pk=record_id)
    from .services.attestato_pdf import archivia_attestato
    try:
        doc = archivia_attestato(record, user=request.user, force=True)
        messages.success(
            request,
            f"Attestato archiviato nel box del dipendente ({doc.nome_originale}).",
        )
    except Exception:
        logger.exception("Archiviazione manuale attestato fallita per record %s", record_id)
        messages.error(request, "Errore durante l'archiviazione dell'attestato.")
    return redirect("anagrafica:attestato_formazione", record_id=record_id)


@login_required
def attestato_report_export(request):
    """Esporta in CSV l'elenco degli attestati archiviati (riepilogo audit).

    Gated dal permesso di visualizzazione formazione. Nessun dato sanitario;
    contiene riferimento dipendente, corso e metadati del documento.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per esportare i report.")
        return redirect("anagrafica:attestato_impostazioni")

    import csv
    from .services.attestato_pdf import RIFERIMENTO_TIPO

    nomi_map = _build_nomi_map()
    docs = list(
        DocumentoDipendente.objects
        .filter(oggetto_riferimento_tipo=RIFERIMENTO_TIPO)
        .select_related("cartella")
        .order_by("-created_at")[:5000]
    )

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="attestati_archiviati.csv"'
    response.write("﻿")  # BOM per Excel
    writer = safe_csv_writer(response, delimiter=";")
    writer.writerow([
        "Documento ID", "Dipendente", "ID anagrafica", "Record completamento",
        "Cartella", "Nome file", "Dimensione (KB)", "Archiviato il", "Conservare fino al",
    ])
    for d in docs:
        writer.writerow([
            d.pk,
            nomi_map.get(d.legacy_anagrafica_id, f"#{d.legacy_anagrafica_id}"),
            d.legacy_anagrafica_id,
            d.oggetto_riferimento_id or "",
            d.cartella.nome if d.cartella else "",
            d.nome_originale,
            round((d.dimensione_bytes or 0) / 1024, 1),
            d.created_at.strftime("%d-%m-%Y %H:%M") if d.created_at else "",
            d.retention_until.strftime("%d-%m-%Y") if d.retention_until else "",
        ])

    try:
        TrainingExportLog.objects.create(
            tipo="ATTESTATO",
            filtri_json={"formato": "archivio_csv", "righe": len(docs)},
            righe_esportate=len(docs),
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog export archivio attestati")
    return response


@login_required
@require_POST
def formazione_sessione_attestati(request, sessione_id: int):
    """Genera e archivia nel box gli attestati di tutti i completati della sessione.

    Comodo a fine edizione: un click produce gli attestati PDF di tutti i
    `TrainingEmployeeRecord` della sessione e li salva nel box documenti
    (idempotente). Gated dal permesso di modifica formazione.
    """
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per archiviare gli attestati.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)

    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    records = list(TrainingEmployeeRecord.objects.filter(sessione=sessione))
    from .services.attestato_pdf import archivia_attestato

    ok = err = 0
    for rec in records:
        try:
            archivia_attestato(rec, user=request.user)
            ok += 1
        except Exception:
            err += 1
            logger.exception(
                "Archiviazione attestato sessione %s record %s fallita", sessione_id, rec.pk
            )

    if not records:
        messages.info(
            request,
            "Nessun completamento registrato per questa sessione: nessun attestato da archiviare.",
        )
    else:
        messages.success(
            request,
            f"Attestati archiviati nel box: {ok}" + (f", {err} errori" if err else "") + ".",
        )
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


# ---------------------------------------------------------------------------
# Allegati formazione: registro firme firmato (sessione / lezione) e materiale
# ---------------------------------------------------------------------------

@login_required
@require_POST
def formazione_allegato_upload(request, sessione_id: int):
    """Carica un allegato (registro firme firmato/materiale) di sessione o lezione.

    Storage privato fuori webroot. Se ``lezione_id`` è valorizzato l'allegato è
    legato alla singola lezione, altrimenti all'intera sessione. Gated dal
    permesso di modifica formazione. Foglio firme = dato personale: come gli
    altri documenti HR, scaricabile solo dalla view protetta.
    """
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per caricare allegati.")
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)

    lezione = None
    lezione_raw = (request.POST.get("lezione_id") or "").strip()
    if lezione_raw.isdigit():
        lezione = TrainingLesson.objects.filter(pk=int(lezione_raw), sessione=sessione).first()

    tipo = (request.POST.get("tipo") or TrainingAttachment.Tipo.REGISTRO_FIRMATO).strip()
    if tipo not in TrainingAttachment.Tipo.values:
        tipo = TrainingAttachment.Tipo.REGISTRO_FIRMATO

    def _back():
        if lezione is not None:
            return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione.pk)
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)

    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "Seleziona un file da caricare.")
        return _back()

    suffix = Path(uploaded.name or "").suffix.lower()
    if suffix not in _ALLOWED_DOC_EXTENSIONS:
        messages.error(request, f"Formato non consentito ({suffix}). Ammessi: PDF, immagini, DOC/XLS.")
        return _back()
    if uploaded.size > _MAX_DOC_SIZE:
        messages.error(request, f"File troppo grande ({uploaded.size // (1024*1024)} MB). Limite: 50 MB.")
        return _back()
    try:
        from core.upload_mime import sniff_mime
        mime = sniff_mime(uploaded)
    except Exception:
        mime = uploaded.content_type or "application/octet-stream"
    if mime not in _ALLOWED_DOC_MIMES:
        messages.error(request, "Tipo di file non consentito (contenuto non valido).")
        return _back()

    att = TrainingAttachment(
        sessione=sessione,
        lezione=lezione,
        tipo=tipo,
        nome_originale=uploaded.name[:255],
        tipo_mime=mime,
        dimensione_bytes=uploaded.size,
        descrizione=(request.POST.get("descrizione") or "").strip()[:300],
        created_by=request.user,
        created_by_display=request.user.get_full_name() or request.user.username,
    )
    att.file = uploaded
    att.save()

    try:
        from core.audit import log_action
        log_action(request, "FORMAZIONE_ALLEGATO_UPLOAD", "anagrafica", {
            "attachment_id": att.pk, "sessione_id": sessione_id,
            "lezione_id": lezione.pk if lezione else None, "tipo": tipo,
            "nome_originale": att.nome_originale,
        })
    except Exception:
        logger.warning("Audit FORMAZIONE_ALLEGATO_UPLOAD fallito", exc_info=True)

    messages.success(request, f"Allegato «{att.nome_originale}» caricato.")
    return _back()


@login_required
@require_POST
def formazione_allegato_delete(request, attachment_id: int):
    att = get_object_or_404(TrainingAttachment.objects.select_related("sessione", "lezione"), pk=attachment_id)
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per eliminare allegati.")
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=att.sessione_id)
    sessione_id = att.sessione_id
    lezione_id = att.lezione_id
    nome = att.nome_originale
    att.delete()
    try:
        from core.audit import log_action
        log_action(request, "FORMAZIONE_ALLEGATO_ELIMINATO", "anagrafica", {
            "attachment_id": attachment_id, "sessione_id": sessione_id,
            "lezione_id": lezione_id, "nome_originale": nome,
        })
    except Exception:
        logger.warning("Audit FORMAZIONE_ALLEGATO_ELIMINATO fallito", exc_info=True)
    messages.success(request, "Allegato eliminato.")
    if lezione_id:
        return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)
    return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)


@login_required
def formazione_allegato_download(request, attachment_id: int):
    """Scarica un allegato formazione dallo storage privato (ACL + audit)."""
    if not _can_view_formazione(request):
        return HttpResponse(status=403)
    att = get_object_or_404(TrainingAttachment, pk=attachment_id)
    if not att.file:
        return HttpResponse("File non disponibile.", status=404)
    try:
        from core.audit import log_action
        log_action(request, "FORMAZIONE_ALLEGATO_DOWNLOAD", "anagrafica", {
            "attachment_id": att.pk, "tipo": att.tipo,
            "sessione_id": att.sessione_id, "lezione_id": att.lezione_id,
        })
    except Exception:
        logger.warning("Audit FORMAZIONE_ALLEGATO_DOWNLOAD fallito", exc_info=True)
    from django.http import FileResponse
    try:
        fh = att.file.open("rb")
    except FileNotFoundError:
        return HttpResponse("File non trovato sul server.", status=404)
    resp = FileResponse(fh, as_attachment=True, filename=att.nome_originale or f"allegato_{att.pk}.bin")
    if att.tipo_mime:
        resp["Content-Type"] = att.tipo_mime
    return resp


# ---------------------------------------------------------------------------
# Report rapidi del corso (CSV iscritti, attestati ZIP, fogli firme PDF)
# ---------------------------------------------------------------------------

@login_required
def formazione_corso_report_iscritti_csv(request, corso_id: int):
    """CSV dei dipendenti del corso (aggregato su tutte le sessioni)."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per esportare i report.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)

    from collections import defaultdict
    agg = defaultdict(lambda: {"n_sessioni": 0, "stati": set(), "data": None, "idoneo": None, "n_compl": 0, "perc": None})
    for e in TrainingEnrollment.objects.filter(sessione__corso=corso).values(
        "legacy_anagrafica_id", "stato", "percentuale_presenza"
    ):
        a = agg[e["legacy_anagrafica_id"]]
        a["n_sessioni"] += 1
        a["stati"].add(e["stato"])
        if e["percentuale_presenza"] is not None:
            a["perc"] = e["percentuale_presenza"]
    for r in TrainingEmployeeRecord.objects.filter(corso=corso).order_by("-data_completamento").values(
        "legacy_anagrafica_id", "data_completamento", "idoneo"
    ):
        a = agg[r["legacy_anagrafica_id"]]
        a["n_compl"] += 1
        if a["data"] is None:
            a["data"] = r["data_completamento"]
            a["idoneo"] = r["idoneo"]

    nomi = _build_nomi_map()
    _prio = ["COMPLETATO", "IN_CORSO", "ISCRITTO", "NON_IDONEO", "ASSENTE", "RITIRATO"]

    import csv
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="corso_{corso.codice}_iscritti.csv"'
    response.write("﻿")
    writer = safe_csv_writer(response, delimiter=";")
    writer.writerow(["Dipendente", "ID anagrafica", "Stato", "Sessioni", "Completamenti",
                     "% presenza", "Idoneo", "Ultimo completamento"])
    for lid, a in sorted(agg.items(), key=lambda kv: nomi.get(kv[0], f"#{kv[0]}").casefold()):
        stato = next((s for s in _prio if s in a["stati"]), next(iter(a["stati"]), ""))
        writer.writerow([
            nomi.get(lid, f"#{lid}"), lid, stato, a["n_sessioni"], a["n_compl"],
            a["perc"] if a["perc"] is not None else "",
            "" if a["idoneo"] is None else ("Sì" if a["idoneo"] else "No"),
            a["data"].strftime("%d-%m-%Y") if a["data"] else "",
        ])

    try:
        TrainingExportLog.objects.create(
            tipo="CORSI", filtri_json={"corso_id": corso.pk, "formato": "iscritti_csv"},
            righe_esportate=len(agg), generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog export iscritti corso")
    return response


@login_required
def formazione_corso_attestati_zip(request, corso_id: int):
    """ZIP con gli attestati PDF di tutti i completamenti idonei del corso."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per esportare gli attestati.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)

    import io
    import zipfile
    from .services.attestato_pdf import build_attestato_pdf_bytes, build_attestato_context

    records = list(
        TrainingEmployeeRecord.objects
        .filter(corso=corso)
        .select_related("corso", "corso__piano", "corso__qualifica", "sessione", "sessione__docente")
        .order_by("-data_completamento", "-id")
    )
    if not records:
        messages.info(request, "Nessun completamento da esportare per questo corso.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)

    nomi = _build_nomi_map()
    buf = io.BytesIO()
    ok = err = 0
    used_names: set[str] = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rec in records:
            try:
                pdf = build_attestato_pdf_bytes(rec)
                try:
                    numero = build_attestato_context(rec)["numero_display"]
                except Exception:
                    numero = f"FORM-{rec.pk:05d}"
                nome_dip = nomi.get(rec.legacy_anagrafica_id, f"dip{rec.legacy_anagrafica_id}")
                base = f"{numero}_{nome_dip}".replace("/", "-").replace(" ", "_")[:120]
                fname = f"{base}.pdf"
                i = 2
                while fname in used_names:
                    fname = f"{base}_{i}.pdf"
                    i += 1
                used_names.add(fname)
                zf.writestr(fname, pdf)
                ok += 1
            except Exception:
                err += 1
                logger.exception("Attestato ZIP corso %s record %s fallito", corso_id, rec.pk)

    if not ok:
        messages.error(request, "Generazione attestati fallita per tutti i record.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)

    try:
        TrainingExportLog.objects.create(
            tipo="ATTESTATO", filtri_json={"corso_id": corso.pk, "formato": "attestati_zip", "ok": ok, "err": err},
            righe_esportate=ok, generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog export attestati ZIP")

    resp = HttpResponse(buf.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="attestati_{corso.codice}.zip"'
    return resp


@login_required
def formazione_corso_registri_pdf(request, corso_id: int):
    """PDF del foglio firme (vuoto) di tutte le lezioni del corso."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per generare i report.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    from .services.attestato_pdf import build_registri_corso_pdf_bytes
    try:
        pdf = build_registri_corso_pdf_bytes(corso)
    except Exception:
        logger.exception("Errore generazione fogli firme corso %s", corso_id)
        messages.error(request, "Errore nella generazione dei fogli firme.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    try:
        TrainingExportLog.objects.create(
            tipo="REPORT_FIRMA", filtri_json={"corso_id": corso.pk, "formato": "fogli_firme_corso_pdf"},
            righe_esportate=0, generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog fogli firme corso")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="fogli_firme_{corso.codice}.pdf"'
    return resp


# ---------------------------------------------------------------------------
# API: salva layout widget scheda dipendente
# ---------------------------------------------------------------------------

@login_required
@require_POST
def api_dipendente_widget_layout(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"ok": False, "error": "JSON non valido"}, status=400)

    hidden = [str(wid) for wid in (data.get("hidden") or []) if wid in _WIDGET_DEFAULT_ORDER]
    order = [str(wid) for wid in (data.get("order") or []) if wid in _WIDGET_DEFAULT_ORDER]

    DipendenteStatLayout.objects.update_or_create(
        viewer_user_id=request.user.id,
        defaults={"hidden": hidden, "order": order},
    )
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Ruoli operativi — assegna/rimuovi a dipendente
# ---------------------------------------------------------------------------

@login_required
@require_POST
def dipendente_ruolo_assegna(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per assegnare ruoli operativi.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    ruolo_id = int(request.POST.get("ruolo_id") or 0)
    ruolo = get_object_or_404(RuoloOperativo, pk=ruolo_id, is_active=True)
    _, created = DipendenteRuoloOperativo.objects.get_or_create(
        legacy_anagrafica_id=legacy_id,
        ruolo=ruolo,
        defaults={"assegnato_da": request.user},
    )
    if created:
        messages.success(request, f'Ruolo "{ruolo.nome}" assegnato.')
    else:
        messages.info(request, f'Ruolo "{ruolo.nome}" già assegnato.')
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_ruolo_rimuovi(request, legacy_id: int, assegnazione_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per rimuovere ruoli operativi.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    assegnazione = get_object_or_404(DipendenteRuoloOperativo, pk=assegnazione_id, legacy_anagrafica_id=legacy_id)
    nome = assegnazione.ruolo.nome
    assegnazione.delete()
    messages.success(request, f'Ruolo "{nome}" rimosso.')
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Ruoli operativi — gestione catalogo
# ---------------------------------------------------------------------------

def _parse_role_id(raw) -> int | None:
    """Interpreta l'id di un ruolo da POST (``None`` se assente/non valido)."""
    try:
        value = int(str(raw or "").strip())
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def _riporta_a_valido(ruolo_id: int | None, riporta_a_id: int | None) -> bool:
    """True se assegnare ``riporta_a`` non crea un ciclo (né un self-loop).

    Risale la catena ``riporta_a`` a partire dal candidato: se incontra
    ``ruolo_id`` la relazione chiuderebbe un ciclo → non valida.
    """
    if not riporta_a_id:
        return True
    if riporta_a_id == ruolo_id:
        return False
    seen: set[int] = set()
    current = RuoloOperativo.objects.filter(pk=riporta_a_id).select_related("riporta_a").first()
    while current is not None and current.pk not in seen:
        if current.pk == ruolo_id:
            return False
        seen.add(current.pk)
        current = current.riporta_a
    return True


@login_required
def ruoli_operativi_list(request):
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)

    ruoli = (
        RuoloOperativo.objects
        .annotate(n_assegnati=Count("assegnazioni"))
        .select_related("riporta_a")
        .order_by("nome")
    )
    # Catalogo per il dropdown «riporta a» (tutti i ruoli, incluso lo storico).
    ruoli_catalogo = list(RuoloOperativo.objects.order_by("nome").values("id", "nome"))
    ruoli_suggeriti = [
        "Preposto", "RSPP", "ASPP", "RLS",
        "Squadra antincendio", "Squadra primo soccorso",
        "Addetto emergenze", "Rappresentante sicurezza",
    ]
    return render(request, "anagrafica/pages/ruoli_operativi.html", {
        "ruoli": ruoli,
        "ruoli_catalogo": ruoli_catalogo,
        "is_admin": is_admin,
        "ruoli_suggeriti": ruoli_suggeriti,
    })


@login_required
@require_POST
def ruolo_operativo_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare ruoli operativi.")
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")

    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome del ruolo è obbligatorio.")
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")

    ruolo, created = RuoloOperativo.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "descrizione": (request.POST.get("descrizione") or "").strip(),
            "colore": (request.POST.get("colore") or "#64748b").strip()[:7],
            "icona": (request.POST.get("icona") or "").strip()[:10],
            "certificazione_competenza": (request.POST.get("certificazione_competenza") or "").strip()[:200],
        },
    )
    if created:
        riporta_a_id = _parse_role_id(request.POST.get("riporta_a"))
        if riporta_a_id and _riporta_a_valido(ruolo.pk, riporta_a_id):
            ruolo.riporta_a_id = riporta_a_id
            ruolo.save(update_fields=["riporta_a"])
        messages.success(request, f'Ruolo "{nome}" creato.')
    else:
        messages.warning(request, f'Esiste già un ruolo con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:ruoli_operativi_list")


@login_required
@require_POST
def ruolo_operativo_edit(request, ruolo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare ruoli operativi.")
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")

    ruolo = get_object_or_404(RuoloOperativo, pk=ruolo_id)
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome del ruolo è obbligatorio.")
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")

    ruolo.nome = nome
    ruolo.descrizione = (request.POST.get("descrizione") or "").strip()
    ruolo.colore = (request.POST.get("colore") or "#64748b").strip()[:7]
    ruolo.icona = (request.POST.get("icona") or "").strip()[:10]
    ruolo.certificazione_competenza = (request.POST.get("certificazione_competenza") or "").strip()[:200]
    riporta_a_id = _parse_role_id(request.POST.get("riporta_a"))
    if riporta_a_id and _riporta_a_valido(ruolo.pk, riporta_a_id):
        ruolo.riporta_a_id = riporta_a_id
    else:
        if riporta_a_id and not _riporta_a_valido(ruolo.pk, riporta_a_id):
            messages.warning(request, "Relazione «riporta a» ignorata: creerebbe un ciclo nella gerarchia.")
        ruolo.riporta_a = None
    ruolo.is_active = request.POST.get("is_active") == "1"
    ruolo.save()
    messages.success(request, f'Ruolo "{ruolo.nome}" aggiornato.')
    return _back_to_caller(request, "anagrafica:ruoli_operativi_list")


@login_required
@require_POST
def ruolo_operativo_delete(request, ruolo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare ruoli operativi.")
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")

    ruolo = get_object_or_404(RuoloOperativo, pk=ruolo_id)
    n_uso = ruolo.assegnazioni.count()
    if n_uso:
        if ruolo.is_active:
            ruolo.is_active = False
            ruolo.save(update_fields=["is_active"])
            messages.warning(request, f'"{ruolo.nome}" è in uso ({n_uso} assegnazioni): disattivato (non eliminato) per preservare lo storico. Puoi riattivarlo dalla modifica.')
        else:
            messages.error(request, f'"{ruolo.nome}" ha {n_uso} assegnazioni: non eliminabile (già disattivo).')
        return _back_to_caller(request, "anagrafica:ruoli_operativi_list")
    nome = ruolo.nome
    ruolo.delete()
    messages.success(request, f'Ruolo "{nome}" eliminato.')
    return _back_to_caller(request, "anagrafica:ruoli_operativi_list")


# ---------------------------------------------------------------------------
# Impostazioni permessi sezione statistiche (solo admin)
# ---------------------------------------------------------------------------

@login_required
def widget_permissions(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("anagrafica:index")

    from core.legacy_models import Ruolo
    perm = AnagraficaStatPermission.get_instance()

    if request.method == "POST":
        perm.accesso = request.POST.get("accesso", AnagraficaStatPermission.ACCESSO_ADMIN)
        if perm.accesso not in (
            AnagraficaStatPermission.ACCESSO_TUTTI,
            AnagraficaStatPermission.ACCESSO_ADMIN,
            AnagraficaStatPermission.ACCESSO_RUOLI,
        ):
            perm.accesso = AnagraficaStatPermission.ACCESSO_ADMIN
        ruolo_ids_raw = request.POST.getlist("ruolo_ids")
        perm.ruolo_ids = [int(r) for r in ruolo_ids_raw if str(r).isdigit()]
        perm.save()
        messages.success(request, "Impostazioni salvate.")
        return redirect("anagrafica:widget_permissions")

    try:
        ruoli_acl = list(Ruolo.objects.order_by("nome"))
    except Exception:
        ruoli_acl = []

    return render(request, "anagrafica/pages/widget_permissions.html", {
        "perm": perm,
        "ruoli_acl": ruoli_acl,
        "ACCESSO_TUTTI": AnagraficaStatPermission.ACCESSO_TUTTI,
        "ACCESSO_ADMIN": AnagraficaStatPermission.ACCESSO_ADMIN,
        "ACCESSO_RUOLI": AnagraficaStatPermission.ACCESSO_RUOLI,
    })


# ---------------------------------------------------------------------------
# Mansione dipendente — set dal catalogo (scrive su legacy DB)
# ---------------------------------------------------------------------------

def _registra_cambiamento(
    legacy_id: int,
    tipo: str,
    vecchio: str,
    nuovo: str,
    user,
    data_effetto=None,
    note: str = "",
) -> DipendenteCambiamentoOrganizzativo | None:
    """Crea una riga di storico solo se il valore è effettivamente cambiato.

    Confronto case-insensitive con strip, così micro-edit di whitespace o maiuscole
    non generano voci storiche spurie.
    """
    from django.utils import timezone as _tz
    v = (vecchio or "").strip()
    n = (nuovo or "").strip()
    if v.casefold() == n.casefold():
        return None
    return DipendenteCambiamentoOrganizzativo.objects.create(
        legacy_anagrafica_id=legacy_id,
        tipo=tipo,
        valore_precedente=v[:300],
        valore_nuovo=n[:300],
        data_effetto=data_effetto or _tz.localdate(),
        note=note,
        created_by=user if getattr(user, "is_authenticated", False) else None,
    )


def _notifica_gap_idoneita(legacy_id: int, dip: dict, mansione_nome: str, user=None) -> None:
    """All'assegnazione di una nuova mansione, ricalcola l'idoneità e notifica
    (email, fail-open) i requisiti mancanti/scaduti a caporeparto + RSPP/HR.
    Nessun dato clinico: le visite restano in forma generica."""
    try:
        stato = conformita_service.stato_conformita(legacy_id, mansione=mansione_nome)
        idn = stato.get("idoneita", {})
        gap = list(idn.get("scaduti", [])) + list(idn.get("mancanti", []))
        if idn.get("esito") not in ("warn", "ko") or not gap:
            return
        from core.email_utils import send_hub_mail
        from .services.onboarding import _caporeparto_emails
        from .services.reminders import get_reminder_recipients
        reparto = (dip.get("reparto") or "").strip()
        nome = f"{dip.get('cognome', '')} {dip.get('nome', '')}".strip() or f"#{legacy_id}"
        dest = sorted(set(
            get_reminder_recipients("idoneita_reminder_emails") + _caporeparto_emails(reparto)
        ))
        if not dest:
            return
        send_hub_mail(
            subject=f"[Idoneità] {nome} → mansione «{mansione_nome}»: requisiti da verificare",
            body_text=(
                f"{nome}{(' — reparto ' + reparto) if reparto else ''} è stato assegnato alla "
                f"mansione «{mansione_nome}».\n\nRequisiti della mansione ancora da soddisfare:\n- "
                + "\n- ".join(gap)
            ),
            recipients=dest, title="Idoneità alla mansione",
            email_type="Anagrafica HR", section_label="Idoneità", fail_silently=True,
        )
    except Exception:
        logger.warning("Notifica gap idoneità (cambio mansione) fallita per %s", legacy_id, exc_info=True)


@login_required
@require_POST
def dipendente_mansione_set(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare la mansione.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    mansione_nome = (request.POST.get("mansione_nome") or "").strip()[:200]

    # Recupera la riga esistente per non sovrascrivere gli altri campi
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]
    mansione_vecchia = (dip.get("mansione") or "").strip()

    try:
        upsert_anagrafica_dipendente(
            row_id=legacy_id,
            aliasusername=dip.get("aliasusername") or "",
            nome=dip.get("nome") or "",
            cognome=dip.get("cognome") or "",
            reparto=dip.get("reparto") or "",
            mansione=mansione_nome,
            ruolo=dip.get("ruolo") or "",
            matricola=dip.get("matricola") or "",
            email=dip.get("email") or "",
            email_notifica=dip.get("email_notifica") or "",
            attivo=bool(dip.get("attivo", True)),
        )
        _registra_cambiamento(
            legacy_id,
            DipendenteCambiamentoOrganizzativo.TIPO_MANSIONE,
            mansione_vecchia, mansione_nome,
            request.user,
        )
        messages.success(request, f'Mansione aggiornata a "{mansione_nome}".' if mansione_nome else "Mansione rimossa.")
        # E: cambio mansione → ricalcola idoneità e notifica i requisiti mancanti
        if mansione_nome and mansione_nome.casefold() != mansione_vecchia.casefold():
            _notifica_gap_idoneita(legacy_id, dip, mansione_nome, request.user)
    except Exception:
        logger.exception("Errore aggiornamento mansione dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'aggiornamento della mansione.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_reparto_set(request, legacy_id: int):
    """Modifica il reparto di un dipendente con storicizzazione e auto-fill area/caporeparto."""
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare il reparto.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    reparto_nome = (request.POST.get("reparto") or "").strip()[:200]
    area_aziendale_raw = (request.POST.get("area_aziendale") or "").strip()
    area_aziendale_id = int(area_aziendale_raw) if area_aziendale_raw.isdigit() else None

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]
    reparto_vecchio = (dip.get("reparto") or "").strip()

    try:
        upsert_anagrafica_dipendente(
            row_id=legacy_id,
            aliasusername=dip.get("aliasusername") or "",
            nome=dip.get("nome") or "",
            cognome=dip.get("cognome") or "",
            reparto=reparto_nome,
            mansione=dip.get("mansione") or "",
            ruolo=dip.get("ruolo") or "",
            matricola=dip.get("matricola") or "",
            email=dip.get("email") or "",
            email_notifica=dip.get("email_notifica") or "",
            attivo=bool(dip.get("attivo", True)),
        )
        _registra_cambiamento(
            legacy_id,
            DipendenteCambiamentoOrganizzativo.TIPO_REPARTO,
            reparto_vecchio, reparto_nome,
            request.user,
        )
        # Auto-fill area aziendale e caporeparto da catalogo
        _sync_aziendale_from_reparto(
            legacy_id, reparto_nome, area_aziendale_id=area_aziendale_id, saved_by=request.user
        )
        messages.success(request, f'Reparto aggiornato a "{reparto_nome}".' if reparto_nome else "Reparto rimosso.")
    except Exception:
        logger.exception("Errore aggiornamento reparto dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'aggiornamento del reparto.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_username_set(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare lo username.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    new_alias = (request.POST.get("aliasusername") or "").strip()[:150]

    if not new_alias:
        messages.error(request, "Lo username non può essere vuoto.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
    if " " in new_alias:
        messages.error(request, "Lo username non può contenere spazi.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    # Normalizza con la stessa regola di upsert_anagrafica_dipendente, così
    # anagrafica e account portale restano allineati sullo stesso valore.
    new_alias = normalize_legacy_alias(new_alias)
    if not new_alias:
        messages.error(request, "Lo username non è valido.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    # Unicità: nessun altro dipendente deve avere lo stesso alias
    conflict = AnagraficaDipendente.objects.filter(
        aliasusername__iexact=new_alias
    ).exclude(id=legacy_id).first()
    if conflict:
        messages.error(request, f"Lo username '{new_alias}' è già in uso da un altro dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]
    old_alias = (dip.get("aliasusername") or "").strip()

    # Account portale Django collegato: anagrafica.utente_id -> Profile.legacy_user_id -> User.
    # aliasusername è la fonte di verità: lo username Django viene tenuto allineato.
    from django.db import transaction
    from django.contrib.auth import get_user_model
    from core.models import Profile

    linked_utente_id = int(dip.get("utente_id") or 0)
    django_profile = None
    if linked_utente_id > 0:
        django_profile = (
            Profile.objects.select_related("user")
            .filter(legacy_user_id=linked_utente_id)
            .first()
        )

    # Fail-closed: blocca prima di scrivere se lo username è già preso da un altro account.
    if django_profile and django_profile.user_id:
        User = get_user_model()
        username_conflict = (
            User.objects.filter(username__iexact=new_alias)
            .exclude(id=django_profile.user_id)
            .first()
        )
        if username_conflict:
            messages.error(request, f"Lo username '{new_alias}' è già in uso da un altro account portale.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    try:
        with transaction.atomic():
            upsert_anagrafica_dipendente(
                row_id=legacy_id,
                aliasusername=new_alias,
                nome=dip.get("nome") or "",
                cognome=dip.get("cognome") or "",
                reparto=dip.get("reparto") or "",
                mansione=dip.get("mansione") or "",
                ruolo=dip.get("ruolo") or "",
                matricola=dip.get("matricola") or "",
                email=dip.get("email") or "",
                email_notifica=dip.get("email_notifica") or "",
                attivo=bool(dip.get("attivo", True)),
            )
            if (
                django_profile
                and django_profile.user_id
                and django_profile.user.username != new_alias
            ):
                django_profile.user.username = new_alias
                django_profile.user.save(update_fields=["username"])
        _registra_cambiamento(legacy_id, "USERNAME", old_alias, new_alias, request.user)
        if django_profile and django_profile.user_id:
            messages.success(request, f"Username aggiornato a '{new_alias}' (anagrafica + account portale).")
        else:
            messages.success(request, f"Username aggiornato a '{new_alias}'.")
    except Exception:
        logger.exception("Errore aggiornamento username dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'aggiornamento dello username.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_toggle_active(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare lo stato del dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]
    currently_active = dip.get("attivo") not in {None, 0, False, "0"}
    new_active = not currently_active

    try:
        upsert_anagrafica_dipendente(
            row_id=legacy_id,
            aliasusername=dip.get("aliasusername") or "",
            nome=dip.get("nome") or "",
            cognome=dip.get("cognome") or "",
            reparto=dip.get("reparto") or "",
            mansione=dip.get("mansione") or "",
            ruolo=dip.get("ruolo") or "",
            matricola=dip.get("matricola") or "",
            email=dip.get("email") or "",
            email_notifica=dip.get("email_notifica") or "",
            attivo=new_active,
            detach_account=not new_active,
        )
        if new_active:
            messages.success(request, "Dipendente riattivato.")
        else:
            messages.success(request, "Dipendente disattivato. L'account portale è stato scollegato.")
    except Exception:
        logger.exception("Errore toggle attivo dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'aggiornamento dello stato.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Qualifiche dipendente — assegna/rimuovi
# ---------------------------------------------------------------------------

@login_required
@require_POST
def dipendente_offboarding_licenziamento(request, legacy_id: int):
    if not _offboarding_is_admin(request):
        messages.error(request, "Non hai i permessi per avviare l'offboarding del dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    from django.db import transaction
    from django.utils import timezone

    oggi = timezone.localdate()
    raw_date = (request.POST.get("data_cessazione") or "").strip()
    data_cessazione = oggi
    if raw_date:
        try:
            data_cessazione = date.fromisoformat(raw_date)
        except ValueError:
            messages.error(request, "Data cessazione non valida.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
    raw_ultimo_giorno = (request.POST.get("ultimo_giorno_operativo") or "").strip()
    ultimo_giorno_operativo = data_cessazione
    if raw_ultimo_giorno:
        try:
            ultimo_giorno_operativo = date.fromisoformat(raw_ultimo_giorno)
        except ValueError:
            messages.error(request, "Ultimo giorno operativo non valido.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
    motivo = (request.POST.get("motivo") or OffboardingPratica.MOTIVO_LICENZIAMENTO).strip()
    valid_motivi = {choice[0] for choice in OffboardingPratica.MOTIVO_CHOICES}
    if motivo not in valid_motivi:
        motivo = OffboardingPratica.MOTIVO_ALTRO
    old_utente_id = _int_or_none(dip.get("utente_id")) or None
    restituzioni_codes = [
        code
        for code in request.POST.getlist("restituzioni")
        if code in OFFBOARDING_RESTITUZIONI_LABELS
    ]
    restituzioni_note = (request.POST.get("restituzioni_note") or "").strip()[:1000]
    dipendente_nome = _offboarding_dipendente_nome(dip)

    aziendale = DipendenteAnagraficaAziendale.objects.filter(legacy_anagrafica_id=legacy_id).first()
    if aziendale and aziendale.data_cessazione:
        messages.warning(request, "Dipendente gia cessato: usa 'Rimetti in forza' se devi riaprire il rapporto.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    pratica_aperta = OffboardingPratica.objects.filter(
        legacy_anagrafica_id=legacy_id,
        stato__in=OffboardingPratica.STATI_APERTI,
    ).first()
    if pratica_aperta:
        messages.warning(request, "Esiste gia una pratica offboarding aperta per questo dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    try:
        with transaction.atomic():
            pratica = OffboardingPratica.objects.create(
                legacy_anagrafica_id=legacy_id,
                dipendente_nome=dipendente_nome,
                reparto=str(dip.get("reparto") or "").strip(),
                mansione=str(dip.get("mansione") or "").strip(),
                motivo=motivo,
                data_cessazione_prevista=data_cessazione,
                ultimo_giorno_operativo=ultimo_giorno_operativo,
                note_hr=restituzioni_note,
                utente_id_pre_offboarding=old_utente_id,
                created_by=request.user,
                updated_by=request.user,
            )
            OffboardingTask.objects.bulk_create([
                OffboardingTask(
                    pratica=pratica,
                    codice=task["codice"],
                    categoria=task["categoria"],
                    titolo=task["titolo"],
                    descrizione=task["descrizione"],
                )
                for task in _offboarding_task_definitions(restituzioni_codes)
            ])

        try:
            from core.audit import log_action
            log_action(
                request,
                "DIPENDENTE_OFFBOARDING_PRATICA_APERTA",
                "anagrafica",
                {
                    "pratica_id": pratica.id,
                    "legacy_anagrafica_id": legacy_id,
                    "dipendente_nome": dipendente_nome,
                    "motivo": motivo,
                    "data_cessazione_prevista": data_cessazione.isoformat(),
                    "ultimo_giorno_operativo": ultimo_giorno_operativo.isoformat() if ultimo_giorno_operativo else "",
                    "legacy_attivo": bool(dip.get("attivo", True)),
                    "account_scollegato": False,
                    "utente_id_pre_offboarding": old_utente_id,
                    "restituzioni_richieste": restituzioni_codes,
                    "restituzioni_richieste_label": [
                        OFFBOARDING_RESTITUZIONI_LABELS[code]
                        for code in restituzioni_codes
                    ],
                    "restituzioni_note": restituzioni_note,
                },
            )
        except Exception:
            logger.warning("Audit DIPENDENTE_OFFBOARDING_PRATICA_APERTA fallito", exc_info=True)

        messages.success(
            request,
            f"Pratica offboarding avviata per il {data_cessazione:%d-%m-%Y}. "
            "Completa le restituzioni e poi conferma la chiusura del rapporto.",
        )
    except Exception:
        logger.exception("Errore avvio pratica offboarding dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'avvio della pratica offboarding.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_offboarding_task_update(request, legacy_id: int, pratica_id: int, task_id: int):
    if not _offboarding_is_admin(request):
        messages.error(request, "Non hai i permessi per aggiornare la pratica offboarding.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    task = get_object_or_404(
        OffboardingTask.objects.select_related("pratica"),
        pk=task_id,
        pratica_id=pratica_id,
        pratica__legacy_anagrafica_id=legacy_id,
        pratica__stato__in=OffboardingPratica.STATI_APERTI,
    )
    stato = (request.POST.get("stato") or "").strip()
    valid_stati = {choice[0] for choice in OffboardingTask.STATO_CHOICES}
    if stato not in valid_stati:
        messages.error(request, "Stato task offboarding non valido.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    from django.utils import timezone

    before = task.stato
    task.stato = stato
    task.note = (request.POST.get("note") or "").strip()[:1000]
    if stato in (OffboardingTask.STATO_COMPLETATO, OffboardingTask.STATO_ECCEZIONE):
        task.completed_at = timezone.now()
        task.completed_by = request.user
    else:
        task.completed_at = None
        task.completed_by = None
    task.save(update_fields=["stato", "note", "completed_at", "completed_by", "updated_at"])

    task.pratica.updated_by = request.user
    task.pratica.save(update_fields=["updated_by", "updated_at"])
    _audit_safe(request, "DIPENDENTE_OFFBOARDING_TASK_UPDATE", "anagrafica", {
        "pratica_id": pratica_id,
        "task_id": task_id,
        "legacy_anagrafica_id": legacy_id,
        "task_codice": task.codice,
        "stato_precedente": before,
        "stato_nuovo": task.stato,
        "note": task.note,
    })
    messages.success(request, f"Task offboarding aggiornato: {task.titolo}.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_offboarding_chiudi(request, legacy_id: int, pratica_id: int):
    if not _offboarding_is_admin(request):
        messages.error(request, "Non hai i permessi per chiudere la pratica offboarding.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    pratica = get_object_or_404(
        OffboardingPratica.objects.prefetch_related("tasks"),
        pk=pratica_id,
        legacy_anagrafica_id=legacy_id,
        stato__in=OffboardingPratica.STATI_APERTI,
    )
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    tasks = list(pratica.tasks.all())
    pending = [task for task in tasks if task.stato == OffboardingTask.STATO_DA_FARE]
    if pending:
        messages.warning(request, "Completa o marca come eccezione tutti i task prima di chiudere il rapporto.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    from django.db import transaction
    from django.utils import timezone

    oggi = timezone.localdate()
    data_cessazione = pratica.data_cessazione_prevista
    if data_cessazione > oggi:
        messages.warning(
            request,
            "La data cessazione prevista e futura: puoi completare i task, ma la chiusura effettiva va fatta alla data corretta.",
        )
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    old_utente_id = pratica.utente_id_pre_offboarding or (_int_or_none(dip.get("utente_id")) or None)
    try:
        with transaction.atomic():
            aziendale, _ = DipendenteAnagraficaAziendale.objects.get_or_create(
                legacy_anagrafica_id=legacy_id
            )
            old_data_cessazione = aziendale.data_cessazione
            aziendale.data_cessazione = data_cessazione
            update_fields = ["data_cessazione", "updated_by", "updated_at"]
            if old_utente_id and aziendale.utente_id_pre_offboarding != old_utente_id:
                aziendale.utente_id_pre_offboarding = old_utente_id
                update_fields.append("utente_id_pre_offboarding")
            aziendale.updated_by = request.user
            aziendale.save(update_fields=update_fields)

            upsert_anagrafica_dipendente(
                row_id=legacy_id,
                aliasusername=dip.get("aliasusername") or "",
                nome=dip.get("nome") or "",
                cognome=dip.get("cognome") or "",
                reparto=dip.get("reparto") or "",
                mansione=dip.get("mansione") or "",
                ruolo=dip.get("ruolo") or "",
                matricola=dip.get("matricola") or "",
                email=dip.get("email") or "",
                email_notifica=dip.get("email_notifica") or "",
                attivo=False,
                detach_account=True,
            )

            has_exceptions = any(task.stato == OffboardingTask.STATO_ECCEZIONE for task in tasks)
            pratica.stato = (
                OffboardingPratica.STATO_CHIUSA_CON_ECCEZIONI
                if has_exceptions else OffboardingPratica.STATO_CHIUSA
            )
            pratica.closed_at = timezone.now()
            pratica.closed_by = request.user
            pratica.updated_by = request.user
            pratica.save(update_fields=["stato", "closed_at", "closed_by", "updated_by", "updated_at"])

        _audit_safe(request, "DIPENDENTE_OFFBOARDING_CHIUSO", "anagrafica", {
            "pratica_id": pratica.id,
            "legacy_anagrafica_id": legacy_id,
            "data_cessazione": data_cessazione.isoformat(),
            "data_cessazione_precedente": old_data_cessazione.isoformat() if old_data_cessazione else "",
            "legacy_attivo": False,
            "account_scollegato": True,
            "utente_id_pre_offboarding": old_utente_id,
            "stato_pratica": pratica.stato,
            "task_totali": len(tasks),
            "task_eccezioni": sum(1 for task in tasks if task.stato == OffboardingTask.STATO_ECCEZIONE),
        })
        messages.success(
            request,
            f"Pratica offboarding chiusa: dipendente cessato dal {data_cessazione:%d-%m-%Y}, "
            "non piu in forza e account scollegato.",
        )
    except Exception:
        logger.exception("Errore chiusura pratica offboarding dipendente %s", legacy_id)
        messages.error(request, "Errore durante la chiusura della pratica offboarding.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_rimetti_in_forza(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per rimettere in forza il dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    try:
        with transaction.atomic():
            aziendale = DipendenteAnagraficaAziendale.objects.filter(
                legacy_anagrafica_id=legacy_id
            ).first()
            old_data_cessazione = aziendale.data_cessazione if aziendale else None
            account_user, account_match = _resolve_account_portale_dipendente(dip, aziendale)
            account_id = int(account_user.id) if account_user else None
            if aziendale and old_data_cessazione:
                aziendale.data_cessazione = None
                aziendale.updated_by = request.user
                update_fields = ["data_cessazione", "updated_by", "updated_at"]
                if account_id and aziendale.utente_id_pre_offboarding:
                    aziendale.utente_id_pre_offboarding = None
                    update_fields.append("utente_id_pre_offboarding")
                aziendale.save(update_fields=update_fields)

            upsert_anagrafica_dipendente(
                row_id=legacy_id,
                aliasusername=dip.get("aliasusername") or "",
                nome=dip.get("nome") or "",
                cognome=dip.get("cognome") or "",
                reparto=dip.get("reparto") or "",
                mansione=dip.get("mansione") or "",
                ruolo=dip.get("ruolo") or "",
                matricola=dip.get("matricola") or "",
                email=dip.get("email") or "",
                email_notifica=dip.get("email_notifica") or "",
                attivo=True,
                utente_id=account_id,
                detach_account=False,
            )

        _audit_safe(
            request,
            "DIPENDENTE_RIMESSO_IN_FORZA",
            "anagrafica",
            {
                "legacy_anagrafica_id": legacy_id,
                "data_cessazione_precedente": old_data_cessazione.isoformat() if old_data_cessazione else "",
                "legacy_attivo": True,
                "account_ricollegato": bool(account_id),
                "account_legacy_user_id": account_id,
                "account_match": account_match,
            },
        )

        if old_data_cessazione:
            if account_id:
                messages.success(
                    request,
                    "Dipendente rimesso in forza: data cessazione rimossa, stato legacy riattivato "
                    "e account portale ricollegato automaticamente.",
                )
            else:
                messages.warning(
                    request,
                    "Dipendente rimesso in forza: data cessazione rimossa e stato legacy riattivato. "
                    "Non ho trovato un account portale univoco da ricollegare automaticamente.",
                )
        else:
            messages.success(request, "Dipendente gia in forza: stato legacy confermato attivo.")
    except Exception:
        logger.exception("Errore rimessa in forza dipendente %s", legacy_id)
        messages.error(request, "Errore durante la rimessa in forza del dipendente.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


def _upsert_dipendente_qualifica(legacy_id, tipo, data_conseguimento, data_scadenza,
                                 *, note="", user=None, sessione=None,
                                 numero=None, livello=None, ente=None,
                                 documento=None, documento_nome=None, origine=None):
    """Crea o aggiorna la qualifica corrente di un dipendente per un tipo.

    Convenzione (come ``import_asr``): una sola ``DipendenteQualifica`` corrente
    per (dipendente, tipo), aggiornata al rinnovo — niente duplicati. Lo storico
    dei rinnovi vive nelle ``QualificaSessione`` collegate. Ritorna (obj, created).

    I campi Fase 2 (``numero``/``livello``/``ente``/``documento``) sono opzionali:
    vengono scritti solo se passati esplicitamente (None = non toccare), così gli
    altri chiamanti (import ASR, sessioni) restano invariati.
    """
    if data_scadenza is None and tipo.durata_mesi and data_conseguimento:
        from anagrafica.models import _add_months
        data_scadenza = _add_months(data_conseguimento, tipo.durata_mesi)
    obj = (
        DipendenteQualifica.objects
        .filter(legacy_anagrafica_id=legacy_id, tipo=tipo)
        .order_by("-data_conseguimento", "-id").first()
    )
    created = obj is None
    if obj is None:
        obj = DipendenteQualifica(legacy_anagrafica_id=legacy_id, tipo=tipo)
    obj.data_conseguimento = data_conseguimento
    obj.data_scadenza = data_scadenza
    if note:
        obj.note = note[:255]
    if numero is not None:
        obj.numero = numero[:100]
    if livello is not None:
        obj.livello = livello[:80]
    if ente is not None:
        obj.ente = ente[:200]
    if documento is not None:
        obj.documento = documento
        obj.documento_nome_originale = (documento_nome or getattr(documento, "name", "") or "")[:255]
        # Un nuovo documento richiede una nuova verifica HR.
        obj.verificata = False
        obj.verificata_da = None
        obj.verificata_il = None
    if sessione is not None:
        obj.sessione = sessione
    if user is not None:
        obj.assegnato_da = user
    obj.save()

    # Storico append-only (Fase 2c): una riga per rilascio/rinnovo. Dedup contro
    # l'ultima riga identica (evita doppioni su re-import idempotenti).
    from anagrafica.models import DipendenteQualificaStorico
    doc_name = getattr(obj.documento, "name", "") or ""
    snap = (obj.data_conseguimento, obj.data_scadenza, obj.numero, obj.ente, doc_name)
    last = None if created else obj.storico.order_by("-id").first()
    last_snap = None if last is None else (
        last.data_conseguimento, last.data_scadenza, last.numero, last.ente,
        getattr(last.documento, "name", "") or "",
    )
    if last_snap != snap:
        if sessione is not None:
            org = DipendenteQualificaStorico.Origine.SESSIONE
        else:
            org = origine or DipendenteQualificaStorico.Origine.MANUALE
        DipendenteQualificaStorico.objects.create(
            qualifica=obj,
            data_conseguimento=obj.data_conseguimento,
            data_scadenza=obj.data_scadenza,
            numero=obj.numero, livello=obj.livello, ente=obj.ente,
            documento=(obj.documento or None),
            documento_nome_originale=obj.documento_nome_originale,
            note=(note or "")[:255], origine=org, registrato_da=user,
        )
    return obj, created


@login_required
@require_POST
def dipendente_qualifica_add(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per aggiungere qualifiche.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    tipo_id = int(request.POST.get("tipo_id") or 0)
    tipo = get_object_or_404(TipoQualifica, pk=tipo_id, is_active=True)

    data_cons_raw = (request.POST.get("data_conseguimento") or "").strip()
    data_scad_raw = (request.POST.get("data_scadenza") or "").strip()

    from datetime import date, timedelta
    data_conseguimento = None
    data_scadenza = None

    if data_cons_raw:
        try:
            data_conseguimento = date.fromisoformat(data_cons_raw)
        except ValueError:
            pass

    if data_scad_raw:
        try:
            data_scadenza = date.fromisoformat(data_scad_raw)
        except ValueError:
            pass
    elif tipo.durata_mesi > 0 and data_conseguimento:
        # Auto-calcola scadenza
        from dateutil.relativedelta import relativedelta
        try:
            data_scadenza = data_conseguimento + relativedelta(months=tipo.durata_mesi)
        except Exception:
            data_scadenza = data_conseguimento + timedelta(days=tipo.durata_mesi * 30)

    note = (request.POST.get("note") or "").strip()[:255]
    numero = (request.POST.get("numero") or "").strip()[:100]
    livello = (request.POST.get("livello") or "").strip()[:80]
    ente = (request.POST.get("ente") or "").strip()[:200]

    # Evidenza documentale opzionale (storage privato). Validazione estensione/dimensione.
    documento = request.FILES.get("documento")
    documento_nome = None
    if documento:
        from pathlib import Path as _Path
        allowed = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".heic", ".tif", ".tiff"}
        suffix = _Path(documento.name or "").suffix.lower()
        if suffix not in allowed:
            messages.error(request, "Formato evidenza non ammesso (usa PDF o immagine).")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
        if documento.size > 50 * 1024 * 1024:
            messages.error(request, "Evidenza troppo grande (max 50 MB).")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
        documento_nome = documento.name

    _, created = _upsert_dipendente_qualifica(
        legacy_id, tipo, data_conseguimento, data_scadenza,
        note=note, user=request.user,
        numero=numero, livello=livello, ente=ente,
        documento=documento or None, documento_nome=documento_nome,
    )
    messages.success(
        request,
        f'Qualifica "{tipo.nome}" {"aggiunta" if created else "aggiornata (rinnovo)"}.',
    )
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_qualifica_delete(request, legacy_id: int, q_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per rimuovere qualifiche.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    qualifica = get_object_or_404(DipendenteQualifica, pk=q_id, legacy_anagrafica_id=legacy_id)
    nome = qualifica.tipo.nome
    qualifica.delete()
    messages.success(request, f'Qualifica "{nome}" rimossa.')
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
def dipendente_qualifica_evidenza(request, legacy_id: int, q_id: int):
    """Serve l'evidenza documentale di una qualifica da storage privato (fuori
    webroot). ACL: admin legacy / superuser / HR (può contenere dati personali)."""
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user) or _check_hr_permission(request)):
        return HttpResponse(status=403)
    q = get_object_or_404(DipendenteQualifica, pk=q_id, legacy_anagrafica_id=legacy_id)
    if not q.documento:
        return HttpResponse("Evidenza non disponibile.", status=404)
    try:
        from core.audit import log_action
        log_action(
            request, "QUALIFICA_EVIDENZA_DOWNLOAD", "anagrafica",
            {"qualifica_id": q.pk, "tipo": q.tipo.nome, "legacy_id": legacy_id},
        )
    except Exception:
        logger.warning("Audit QUALIFICA_EVIDENZA_DOWNLOAD fallito", exc_info=True)
    from django.http import FileResponse
    try:
        fh = q.documento.open("rb")
    except FileNotFoundError:
        return HttpResponse("File non trovato sul server.", status=404)
    return FileResponse(
        fh, as_attachment=True,
        filename=q.documento_nome_originale or f"qualifica_{q.pk}.bin",
    )


@login_required
def dipendente_qualifica_storico_evidenza(request, legacy_id: int, q_id: int, storico_id: int):
    """Serve l'evidenza documentale storicizzata di un rinnovo (storage privato).
    ACL: admin legacy / superuser / HR (come l'evidenza corrente)."""
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user) or _check_hr_permission(request)):
        return HttpResponse(status=403)
    from anagrafica.models import DipendenteQualificaStorico
    s = get_object_or_404(
        DipendenteQualificaStorico, pk=storico_id,
        qualifica_id=q_id, qualifica__legacy_anagrafica_id=legacy_id,
    )
    if not s.documento:
        return HttpResponse("Evidenza non disponibile.", status=404)
    try:
        from core.audit import log_action
        log_action(
            request, "QUALIFICA_STORICO_EVIDENZA_DOWNLOAD", "anagrafica",
            {"storico_id": s.pk, "qualifica_id": q_id, "legacy_id": legacy_id},
        )
    except Exception:
        logger.warning("Audit QUALIFICA_STORICO_EVIDENZA_DOWNLOAD fallito", exc_info=True)
    from django.http import FileResponse
    try:
        fh = s.documento.open("rb")
    except FileNotFoundError:
        return HttpResponse("File non trovato sul server.", status=404)
    return FileResponse(
        fh, as_attachment=True,
        filename=s.documento_nome_originale or f"qualifica_storico_{s.pk}.bin",
    )


@login_required
@require_POST
def dipendente_qualifica_verifica(request, legacy_id: int, q_id: int):
    """Toggle del flag «verificata» (controllo HR dell'evidenza). ACL: admin/HR."""
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user) or _check_hr_permission(request)):
        messages.error(request, "Permessi insufficienti per verificare la qualifica.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
    q = get_object_or_404(DipendenteQualifica, pk=q_id, legacy_anagrafica_id=legacy_id)
    if q.verificata:
        q.verificata = False
        q.verificata_da = None
        q.verificata_il = None
        msg = f'Verifica rimossa da "{q.tipo.nome}".'
    else:
        from django.utils import timezone as _tz
        q.verificata = True
        q.verificata_da = request.user
        q.verificata_il = _tz.now()
        msg = f'Qualifica "{q.tipo.nome}" contrassegnata come verificata.'
    q.save(update_fields=["verificata", "verificata_da", "verificata_il"])
    messages.success(request, msg)
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Anagrafica civile e aziendale — salvataggio
# ---------------------------------------------------------------------------

@login_required
@require_POST
def dipendente_anagrafica_civile_save(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare l'anagrafica civile.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    instance, _ = DipendenteAnagraficaCivile.objects.get_or_create(
        legacy_anagrafica_id=legacy_id
    )
    form = AnagraficaCivileForm(request.POST, request.FILES, instance=instance)
    formset = FiglioACaricoFormSet(request.POST, instance=instance, prefix="figli")
    if form.is_valid() and formset.is_valid():
        obj = form.save(commit=False)
        obj.legacy_anagrafica_id = legacy_id
        obj.updated_by = request.user
        obj.save()
        formset.instance = obj
        formset.save()
        # Allinea il flag al numero effettivo di figli registrati.
        ha_figli = obj.figli.exists()
        if obj.figli_a_carico != ha_figli:
            obj.figli_a_carico = ha_figli
            obj.save(update_fields=["figli_a_carico"])
        messages.success(request, "Anagrafica civile salvata.")
    else:
        for field, errs in form.errors.items():
            for err in errs:
                messages.error(request, f"{field}: {err}")
        for sub in formset:
            for field, errs in sub.errors.items():
                for err in errs:
                    messages.error(request, f"Figli — {field}: {err}")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_anagrafica_aziendale_save(request, legacy_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare l'anagrafica aziendale.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    instance, _ = DipendenteAnagraficaAziendale.objects.get_or_create(
        legacy_anagrafica_id=legacy_id
    )
    area_vecchia = instance.area or ""
    ruolo_az_vecchio = instance.ruolo_aziendale or ""

    form = AnagraficaAziendaleForm(request.POST, instance=instance)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.legacy_anagrafica_id = legacy_id
        obj.updated_by = request.user
        obj.save()
        _sync_aziendale_from_reparto(
            legacy_id, obj.area or "", area_aziendale_id=obj.area_aziendale_id, saved_by=request.user
        )
        _registra_cambiamento(
            legacy_id,
            DipendenteCambiamentoOrganizzativo.TIPO_AREA,
            area_vecchia, obj.area or "",
            request.user,
        )
        _registra_cambiamento(
            legacy_id,
            DipendenteCambiamentoOrganizzativo.TIPO_RUOLO_AZIENDALE,
            ruolo_az_vecchio, obj.ruolo_aziendale or "",
            request.user,
        )
        messages.success(request, "Anagrafica aziendale salvata.")
    else:
        for field, errs in form.errors.items():
            for err in errs:
                messages.error(request, f"{field}: {err}")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Voci retributive — importazione CSV dallo studio paghe
# ---------------------------------------------------------------------------

def _import_csv_retribuzioni(file_obj, user, file_nome: str = "") -> ImportazioneRetributiva:
    """Parsa il CSV paghe e crea ImportazioneRetributiva + VoceRetributiva."""
    raw = file_obj.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Impossibile decodificare il file CSV (provati utf-8, latin-1, cp1252).")

    reader = csv.DictReader(StringIO(text), delimiter=";")
    rows = [r for r in reader if any(v.strip() for v in r.values())]
    if not rows:
        raise ValueError("Il file CSV è vuoto o non contiene righe valide.")

    # Data competenza dal primo record
    first_date_str = (rows[0].get("date") or "").strip()
    try:
        data_comp_raw = datetime.strptime(first_date_str, "%d/%m/%Y").date()
        data_competenza = data_comp_raw.replace(day=1)
    except ValueError:
        raise ValueError(f"Formato data non riconosciuto: '{first_date_str}' (atteso gg/mm/aaaa).")

    importazione = ImportazioneRetributiva.objects.create(
        data_competenza=data_competenza,
        importato_da=user,
        file_nome=file_nome,
        righe_totali=len(rows),
    )

    # Mappa codice fiscale → legacy_anagrafica_id
    all_cf = {(r.get("tax_code") or "").strip().upper() for r in rows if (r.get("tax_code") or "").strip()}
    cf_to_id: dict[str, int] = dict(
        DipendenteAnagraficaCivile.objects.filter(
            codice_fiscale__in=all_cf
        ).values_list("codice_fiscale", "legacy_anagrafica_id")
    )

    # Fallback nome "COGNOME NOME" → legacy_anagrafica_id quando il CF non è ancora in DipendenteAnagraficaCivile
    nome_to_id: dict[str, int] = {}
    try:
        _legacy_all = fetch_anagrafica_rows(deduplicate=True)
        for _lr in _legacy_all:
            _cog = str(_lr.get("cognome") or "").strip().upper()
            _nom = str(_lr.get("nome") or "").strip().upper()
            _full = f"{_cog} {_nom}".strip()
            _lid = int(_lr.get("id") or 0)
            if _full and _lid:
                nome_to_id[_full] = _lid
    except Exception:
        logger.warning("Fallback nome-based per retribuzioni non disponibile (legacy DB non raggiungibile)")

    # Voci dell'ultima importazione precedente per rilevare variazioni
    prev_import = (
        ImportazioneRetributiva.objects
        .filter(data_competenza__lt=data_competenza)
        .order_by("-data_competenza")
        .first()
    )
    prev_voci: dict[tuple, Decimal] = {}
    if prev_import:
        for voce in VoceRetributiva.objects.filter(importazione=prev_import):
            prev_voci[(voce.tax_code, voce.pay_item_key)] = voce.importo

    ok = err = 0
    voci_to_create: list[VoceRetributiva] = []
    for row in rows:
        try:
            tax_code = (row.get("tax_code") or "").strip().upper()
            pay_item = (row.get("pay_item") or "").strip()
            value_str = (row.get("value") or "0").strip().replace(",", ".")
            date_str = (row.get("date") or "").strip()
            if not tax_code or not pay_item:
                err += 1
                continue
            importo = Decimal(value_str)
            data_riga = datetime.strptime(date_str, "%d/%m/%Y").date()
            pay_item_key = pay_item.lower()
            categoria = _classify_pay_item(pay_item_key)
            legacy_id = cf_to_id.get(tax_code)
            if not legacy_id:
                nome_csv = (row.get("nome") or "").strip().upper()
                legacy_id = nome_to_id.get(nome_csv)
            prev_importo = prev_voci.get((tax_code, pay_item_key))
            is_changed = prev_importo is not None and prev_importo != importo
            voci_to_create.append(VoceRetributiva(
                importazione=importazione,
                tax_code=tax_code,
                legacy_anagrafica_id=legacy_id,
                data_competenza=data_riga,
                pay_item=pay_item,
                pay_item_key=pay_item_key,
                categoria=categoria,
                importo=importo,
                is_changed=is_changed,
                importo_precedente=prev_importo if is_changed else None,
            ))
            ok += 1
        except (InvalidOperation, ValueError):
            err += 1

    VoceRetributiva.objects.bulk_create(voci_to_create)
    importazione.righe_ok = ok
    importazione.righe_errore = err
    importazione.save(update_fields=["righe_ok", "righe_errore"])
    return importazione


@login_required
def retribuzioni_import(request):
    """Pagina importazione CSV retributivo (solo admin)."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("anagrafica:dipendenti_list")

    if request.method == "POST":
        file_obj = request.FILES.get("file_csv")
        if not file_obj:
            messages.error(request, "Nessun file selezionato.")
        else:
            try:
                imp = _import_csv_retribuzioni(file_obj, request.user, file_obj.name)
                messages.success(
                    request,
                    f"Importazione completata: {imp.righe_ok} voci caricate"
                    f"{f', {imp.righe_errore} errori' if imp.righe_errore else ''}."
                )
                return redirect("anagrafica:retribuzioni_import")
            except Exception as exc:
                logger.exception("Errore importazione CSV retribuzioni")
                messages.error(request, f"Errore durante l'importazione: {exc}")

    importazioni = list(
        ImportazioneRetributiva.objects
        .filter(origine=ImportazioneRetributiva.ORIGINE_CSV)
        .select_related("importato_da")[:30]
    )
    return render(request, "anagrafica/pages/retribuzioni_import.html", {
        "importazioni": importazioni,
        "is_admin": is_admin,
    })


@login_required
def dipendente_retribuzioni(request, legacy_id: int):
    """Storico completo voci retributive per un dipendente (accesso HR)."""
    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    civile_retr = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    _retr_q = Q(legacy_anagrafica_id=legacy_id)
    if civile_retr and civile_retr.codice_fiscale:
        _retr_q |= Q(tax_code=civile_retr.codice_fiscale.strip().upper())
    voci_all = list(
        VoceRetributiva.objects.filter(_retr_q)
        .select_related("importazione")
        .order_by("-data_competenza", "-importazione__data_importazione", "categoria", "pay_item_key")
    )

    # Merge: una sola importazione CSV per mese (la più recente) + voci manuali del mese
    # (le voci manuali fanno override delle voci CSV con lo stesso pay_item_key).
    _by_month: dict = {}
    for v in voci_all:
        key = v.data_competenza
        if key not in _by_month:
            _by_month[key] = {"csv_imp": None, "csv_voci": [], "manuale_voci": []}
        if v.manuale:
            _by_month[key]["manuale_voci"].append(v)
        else:
            if _by_month[key]["csv_imp"] is None:
                _by_month[key]["csv_imp"] = v.importazione
            if v.importazione_id == _by_month[key]["csv_imp"].id:
                _by_month[key]["csv_voci"].append(v)

    # Costruisco mese -> {pay_item_key: voce effettiva (manuale prevale su CSV)}
    _mesi_dict: dict = {}
    for _date, entry in _by_month.items():
        _manuale_keys = {v.pay_item_key for v in entry["manuale_voci"]}
        _voci = [v for v in entry["csv_voci"] if v.pay_item_key not in _manuale_keys] + entry["manuale_voci"]
        _mesi_dict[_date] = {v.pay_item_key: v for v in _voci}

    # Ordinamento colonne pivot: per categoria (fisso, variabile, altro, totale)
    # e all'interno per ordine di prima apparizione cronologica (rispecchia l'ordine CSV originale).
    _CAT_ORDER = {"fisso": 0, "variabile": 1, "altro": 2, "totale": 3}
    _TOTALI_SORT = {"retribuzione di fatto": 0, "totale elementi variabili": 1, "rml": 2, "ral": 3}
    _key_meta: dict = {}
    # Iterazione cronologica ascendente per registrare prima apparizione
    _voci_for_meta = sorted(voci_all, key=lambda v: (v.data_competenza, v.categoria, v.pay_item_key))
    _seq = 0
    for v in _voci_for_meta:
        if v.pay_item_key not in _key_meta:
            _key_meta[v.pay_item_key] = {
                "key": v.pay_item_key,
                "label": v.pay_item,
                "categoria": v.categoria,
                "_seq": _seq,
            }
            _seq += 1

    def _col_sort(m):
        cat = _CAT_ORDER.get(m["categoria"], 99)
        sub = _TOTALI_SORT.get(m["key"], 50) if m["categoria"] == "totale" else m["_seq"]
        return (cat, sub)

    colonne = sorted(_key_meta.values(), key=_col_sort)

    # Costruzione righe pivot: per ogni mese cronologico ASC calcolo le celle e il flag changed
    # (confronto con il mese precedente per cui esiste un valore nella stessa colonna).
    _mesi_asc = sorted(_mesi_dict.keys())
    _prev_value_by_key: dict = {}  # pay_item_key -> ultimo importo visto
    _rows_by_date: dict = {}
    for _date in _mesi_asc:
        voci_map = _mesi_dict[_date]
        celle = []
        for col in colonne:
            v = voci_map.get(col["key"])
            importo = v.importo if v else None
            prev = _prev_value_by_key.get(col["key"])
            # Cella variata se il valore differisce dal mese precedente
            # (None vs valore = variazione; entrambi None = invariato).
            changed = (importo != prev) and not (importo is None and prev is None)
            # Aggiorno il "valore precedente" solo se in questo mese c'è un valore
            if importo is not None:
                _prev_value_by_key[col["key"]] = importo
            celle.append({
                "voce": v,
                "importo": importo,
                "changed": changed,
                "manuale": bool(v and v.manuale),
                "categoria": col["categoria"],
                "is_totale": col["categoria"] == "totale",
                "is_ral": col["key"] == "ral",
            })
        _rows_by_date[_date] = {
            "data": _date,
            "celle": celle,
            "importazione": _by_month[_date]["csv_imp"],
            "n_manuale": len(_by_month[_date]["manuale_voci"]),
            "n_changed": sum(1 for c in celle if c["changed"] and c["importo"] is not None),
        }

    # Raggruppo per anno (anno decrescente, mesi decrescenti dentro l'anno)
    _today = date.today()
    _anni_default_aperti = {_today.year, _today.year - 1}
    _by_year: dict = {}
    for _date in _mesi_asc:
        _by_year.setdefault(_date.year, []).append(_rows_by_date[_date])

    anni = []
    for anno in sorted(_by_year.keys(), reverse=True):
        mesi_desc = list(reversed(_by_year[anno]))
        anni.append({
            "anno": anno,
            "mesi": mesi_desc,
            "n_mesi": len(mesi_desc),
            "n_changed": sum(m["n_changed"] for m in mesi_desc),
            "open": anno in _anni_default_aperti,
        })

    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)

    return render(request, "anagrafica/pages/dipendente_retribuzioni.html", {
        "dip": dip,
        "legacy_id": legacy_id,
        "colonne": colonne,
        "anni": anni,
        "n_mesi": len(_mesi_asc),
        "n_colonne": len(colonne),
        "can_hr": can_hr,
        "is_admin": is_admin,
        "tax_code_dipendente": civile_retr.codice_fiscale.strip().upper() if civile_retr and civile_retr.codice_fiscale else "",
        "categorie_voce": VoceRetributiva.CATEGORIA_CHOICES,
    })


@login_required
def dipendente_retribuzioni_export_xlsx(request, legacy_id: int):
    """Esporta in Excel lo storico retributivo del dipendente come tabella pivot
    (righe = mesi cronologici desc, colonne = pay_item per categoria, celle = importi
    con highlight delle variazioni vs mese precedente)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from core.excel_export import write_cell

    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    ensure_anagrafica_schema()
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    civile = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    _q = Q(legacy_anagrafica_id=legacy_id)
    if civile and civile.codice_fiscale:
        _q |= Q(tax_code=civile.codice_fiscale.strip().upper())
    voci_all = list(
        VoceRetributiva.objects.filter(_q)
        .order_by("-data_competenza", "categoria", "pay_item_key")
    )

    # Merge manuali/CSV identico alla view principale
    _by_month: dict = {}
    for v in voci_all:
        key = v.data_competenza
        if key not in _by_month:
            _by_month[key] = {"csv_imp": None, "csv_voci": [], "manuale_voci": []}
        if v.manuale:
            _by_month[key]["manuale_voci"].append(v)
        else:
            if _by_month[key]["csv_imp"] is None:
                _by_month[key]["csv_imp"] = v.importazione
            if v.importazione_id == _by_month[key]["csv_imp"].id:
                _by_month[key]["csv_voci"].append(v)

    _mesi_dict: dict = {}
    for _date, entry in _by_month.items():
        _manuale_keys = {v.pay_item_key for v in entry["manuale_voci"]}
        _voci = [v for v in entry["csv_voci"] if v.pay_item_key not in _manuale_keys] + entry["manuale_voci"]
        _mesi_dict[_date] = {v.pay_item_key: v for v in _voci}

    _CAT_ORDER = {"fisso": 0, "variabile": 1, "altro": 2, "totale": 3}
    _TOTALI_SORT = {"retribuzione di fatto": 0, "totale elementi variabili": 1, "rml": 2, "ral": 3}
    _key_meta: dict = {}
    _seq = 0
    for v in sorted(voci_all, key=lambda v: (v.data_competenza, v.categoria, v.pay_item_key)):
        if v.pay_item_key not in _key_meta:
            _key_meta[v.pay_item_key] = {
                "key": v.pay_item_key, "label": v.pay_item,
                "categoria": v.categoria, "_seq": _seq,
            }
            _seq += 1

    def _col_sort(m):
        cat = _CAT_ORDER.get(m["categoria"], 99)
        sub = _TOTALI_SORT.get(m["key"], 50) if m["categoria"] == "totale" else m["_seq"]
        return (cat, sub)

    colonne = sorted(_key_meta.values(), key=_col_sort)
    _mesi_asc = sorted(_mesi_dict.keys())

    # Workbook + foglio
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Storico retributivo"

    # Stili
    thin = Side(border_style="thin", color="E2E8F0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_by_cat = {
        "fisso":     PatternFill(fill_type="solid", fgColor="F1F5F9"),
        "variabile": PatternFill(fill_type="solid", fgColor="FFF7ED"),
        "altro":     PatternFill(fill_type="solid", fgColor="F5F3FF"),
        "totale":    PatternFill(fill_type="solid", fgColor="ECFDF5"),
    }
    font_header = Font(bold=True, name="Calibri", size=10, color="1E293B")
    font_date = Font(bold=True, name="Calibri", size=10)
    font_normal = Font(name="Calibri", size=10)
    font_totale = Font(bold=True, name="Calibri", size=10, color="065F46")
    fill_changed = PatternFill(fill_type="solid", fgColor="DBEAFE")
    fill_changed_totale = PatternFill(fill_type="solid", fgColor="BBF7D0")
    fill_manuale_marker = PatternFill(fill_type="solid", fgColor="EDE9FE")

    # Intestazione (riga 1)
    write_cell(ws, 1, 1, "Data retribuzione").font = font_header
    ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=1).border = border_all
    for ci, col in enumerate(colonne, start=2):
        c = write_cell(ws, 1, ci, col["label"])
        c.font = font_header
        c.fill = fill_by_cat.get(col["categoria"], PatternFill(fill_type="solid", fgColor="F8FAFC"))
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        c.border = border_all

    # Righe: mesi in ordine cronologico discendente
    _prev_value_by_key: dict = {}
    # Per il calcolo "changed" itero ASC ma scrivo le righe DESC: costruisco prima la matrice
    _cell_data_by_date: dict = {}  # date -> list of (importo, changed, manuale)
    for _date in _mesi_asc:
        voci_map = _mesi_dict[_date]
        row_cells = []
        for col in colonne:
            v = voci_map.get(col["key"])
            importo = v.importo if v else None
            prev = _prev_value_by_key.get(col["key"])
            changed = (importo != prev) and not (importo is None and prev is None)
            if importo is not None:
                _prev_value_by_key[col["key"]] = importo
            row_cells.append((importo, changed, bool(v and v.manuale)))
        _cell_data_by_date[_date] = row_cells

    for ri, _date in enumerate(reversed(_mesi_asc), start=2):
        c_date = write_cell(ws, ri, 1, _date)  # date: resta tipizzata (non stringa)
        c_date.number_format = "DD/MM/YYYY"
        c_date.font = font_date
        c_date.alignment = Alignment(horizontal="left", vertical="center")
        c_date.border = border_all
        for ci, (col, (importo, changed, manuale)) in enumerate(zip(colonne, _cell_data_by_date[_date]), start=2):
            cell = ws.cell(row=ri, column=ci)
            if importo is not None:
                cell.value = float(importo)
                cell.number_format = '#,##0.00 "€"'
            cell.font = font_totale if col["categoria"] == "totale" else font_normal
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border_all
            if changed and importo is not None:
                cell.fill = fill_changed_totale if col["categoria"] == "totale" else fill_changed
            elif manuale:
                cell.fill = fill_manuale_marker

    # Freeze pane: prima riga + prima colonna
    ws.freeze_panes = "B2"

    # Auto-width approssimativo (label + padding)
    ws.column_dimensions["A"].width = 18
    for ci, col in enumerate(colonne, start=2):
        ws.column_dimensions[get_column_letter(ci)].width = max(14, min(28, len(col["label"]) + 4))

    _cognome = str(dip.get("cognome") or "").strip()
    _nome = str(dip.get("nome") or "").strip()

    # Riga riepilogo finale
    summary_row = len(_mesi_asc) + 3
    write_cell(
        ws, summary_row, 1,
        f"Dipendente: {_cognome} {_nome} — {len(_mesi_asc)} mesi · {len(colonne)} voci",
    ).font = Font(italic=True, size=9, color="64748B")

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y%m%d")
    safe_name = "".join(ch for ch in f"{_cognome}_{_nome}" if ch.isalnum() or ch in ("_", "-")).strip("_") or str(legacy_id)
    filename = f"storico_retributivo_{safe_name}_{today}.xlsx"

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ---------------------------------------------------------------------------
# Voci retributive — data-entry manuale (HR/admin)
# ---------------------------------------------------------------------------

def _get_or_create_import_manuale(data_competenza, user) -> ImportazioneRetributiva:
    """Recupera o crea l'unica `ImportazioneRetributiva` manuale per quel mese di competenza."""
    imp, created = ImportazioneRetributiva.objects.get_or_create(
        origine=ImportazioneRetributiva.ORIGINE_MANUALE,
        data_competenza=data_competenza,
        defaults={
            "importato_da": user,
            "file_nome": "(inserimento manuale)",
            "note": "Container voci retributive inserite manualmente",
        },
    )
    return imp


def _parse_data_competenza(raw: str):
    """Accetta 'YYYY-MM' o 'YYYY-MM-DD' e restituisce date(primo giorno mese)."""
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        if len(raw) == 7:  # YYYY-MM
            return datetime.strptime(raw + "-01", "%Y-%m-%d").date()
        d = datetime.strptime(raw, "%Y-%m-%d").date()
        return d.replace(day=1)
    except ValueError:
        return None


def _parse_importo(raw: str):
    raw = (raw or "").strip().replace(",", ".")
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


@login_required
@require_POST
def dipendente_retribuzione_voce_add(request, legacy_id: int):
    """Aggiunge una voce retributiva manuale (HR + admin)."""
    can_hr = _check_hr_permission(request)
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not (can_hr or is_admin):
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    data_comp = _parse_data_competenza(request.POST.get("data_competenza"))
    if not data_comp:
        messages.error(request, "Mese di competenza non valido.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    pay_item = (request.POST.get("pay_item") or "").strip()[:150]
    importo = _parse_importo(request.POST.get("importo"))
    if not pay_item or importo is None:
        messages.error(request, "Voce e importo sono obbligatori.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    categoria_raw = (request.POST.get("categoria") or "").strip().lower()
    valid_cats = {c for c, _ in VoceRetributiva.CATEGORIA_CHOICES}
    categoria = categoria_raw if categoria_raw in valid_cats else _classify_pay_item(pay_item.lower())

    civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    tax_code = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else ""

    imp = _get_or_create_import_manuale(data_comp, request.user)
    VoceRetributiva.objects.create(
        importazione=imp,
        tax_code=tax_code,
        legacy_anagrafica_id=legacy_id,
        data_competenza=data_comp,
        pay_item=pay_item,
        pay_item_key=pay_item.lower(),
        categoria=categoria,
        importo=importo,
        manuale=True,
        note=(request.POST.get("note") or "").strip()[:300],
        updated_by=request.user,
    )
    imp.righe_ok = imp.voci.count()
    imp.save(update_fields=["righe_ok"])

    messages.success(request, f'Voce "{pay_item}" aggiunta per {data_comp.strftime("%B %Y")}.')
    return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_retribuzione_voce_edit(request, legacy_id: int, voce_id: int):
    """Modifica una voce retributiva (solo voci con flag manuale=True)."""
    can_hr = _check_hr_permission(request)
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not (can_hr or is_admin):
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    voce = get_object_or_404(VoceRetributiva, pk=voce_id)
    if not voce.manuale:
        messages.error(request, "Solo le voci inserite manualmente sono modificabili. Le voci da CSV vanno reimportate.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    tc = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else ""
    if voce.legacy_anagrafica_id != legacy_id and (not tc or voce.tax_code != tc):
        messages.error(request, "Voce non appartenente a questo dipendente.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    pay_item = (request.POST.get("pay_item") or "").strip()[:150]
    importo = _parse_importo(request.POST.get("importo"))
    if not pay_item or importo is None:
        messages.error(request, "Voce e importo sono obbligatori.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    categoria_raw = (request.POST.get("categoria") or "").strip().lower()
    valid_cats = {c for c, _ in VoceRetributiva.CATEGORIA_CHOICES}
    categoria = categoria_raw if categoria_raw in valid_cats else _classify_pay_item(pay_item.lower())

    voce.pay_item = pay_item
    voce.pay_item_key = pay_item.lower()
    voce.importo = importo
    voce.categoria = categoria
    voce.note = (request.POST.get("note") or "").strip()[:300]
    voce.updated_by = request.user
    voce.save(update_fields=["pay_item", "pay_item_key", "importo", "categoria", "note", "updated_by", "updated_at"])

    messages.success(request, "Voce aggiornata.")
    return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_retribuzione_voce_delete(request, legacy_id: int, voce_id: int):
    """Elimina una voce retributiva (solo voci con flag manuale=True)."""
    can_hr = _check_hr_permission(request)
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not (can_hr or is_admin):
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    voce = get_object_or_404(VoceRetributiva, pk=voce_id)
    if not voce.manuale:
        messages.error(request, "Solo le voci inserite manualmente sono eliminabili.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    tc = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else ""
    if voce.legacy_anagrafica_id != legacy_id and (not tc or voce.tax_code != tc):
        messages.error(request, "Voce non appartenente a questo dipendente.")
        return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)

    imp = voce.importazione
    voce.delete()
    # Se l'importazione manuale è rimasta senza voci, eliminala
    if imp.origine == ImportazioneRetributiva.ORIGINE_MANUALE and not imp.voci.exists():
        imp.delete()
    else:
        imp.righe_ok = imp.voci.count()
        imp.save(update_fields=["righe_ok"])

    messages.success(request, "Voce eliminata.")
    return redirect("anagrafica:dipendente_retribuzioni", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Storico contrattuale — import CSV + CRUD manuale
# ---------------------------------------------------------------------------

_CONTRATTO_LABEL_MAP: dict[str, str] = {
    label.lower(): value
    for value, label in DipendenteAnagraficaAziendale.CONTRATTO_CHOICES
}
_CONTRATTO_LABEL_MAP.update({
    "indeterminato": DipendenteAnagraficaAziendale.CONTRATTO_INDETERMINATO,
    "determinato": DipendenteAnagraficaAziendale.CONTRATTO_DETERMINATO,
    "apprendistato": DipendenteAnagraficaAziendale.CONTRATTO_APPRENDISTATO,
    "somministrazione": DipendenteAnagraficaAziendale.CONTRATTO_SOMMINISTRAZIONE,
    "collaborazione": DipendenteAnagraficaAziendale.CONTRATTO_COLLABORAZIONE,
    "stage": DipendenteAnagraficaAziendale.CONTRATTO_STAGE,
    "tirocinio": DipendenteAnagraficaAziendale.CONTRATTO_STAGE,
    "altro": DipendenteAnagraficaAziendale.CONTRATTO_ALTRO,
})


def _normalize_contratto_choice(raw: str) -> str:
    key = raw.strip().lower()
    if key in _CONTRATTO_LABEL_MAP:
        return _CONTRATTO_LABEL_MAP[key]
    # partial match
    for k, v in _CONTRATTO_LABEL_MAP.items():
        if k in key or key in k:
            return v
    return ""


def _import_csv_contratti(file_obj, user) -> tuple[int, int, int]:
    """Parsa il CSV contrattuale e crea/aggiorna record StoricoContratto.

    Formato: Codice fiscale;Data Inizio;Data Fine;Tipo di contratto;Qualifica;Livello;CCNL;Descrizione livello
    Più righe con stesso (CF, Data Inizio, Data Fine) vengono aggregate automaticamente.
    Restituisce (created, updated, skipped).
    """
    raw = file_obj.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Impossibile decodificare il file CSV (provati utf-8, latin-1, cp1252).")

    # Mappa CF → legacy_id
    cf_to_id: dict[str, int] = {}
    for row in DipendenteAnagraficaCivile.objects.filter(
        codice_fiscale__isnull=False, legacy_anagrafica_id__isnull=False
    ).exclude(codice_fiscale="").values("codice_fiscale", "legacy_anagrafica_id"):
        cf_to_id[str(row["codice_fiscale"]).strip().upper()] = row["legacy_anagrafica_id"]

    # Fallback nome-based
    nome_to_id: dict[str, int] = {}
    try:
        _legacy_all = fetch_anagrafica_rows(deduplicate=True)
        for _lr in _legacy_all:
            _cog = str(_lr.get("cognome") or "").strip().upper()
            _nom = str(_lr.get("nome") or "").strip().upper()
            _full = f"{_cog} {_nom}".strip()
            _lid = int(_lr.get("id") or 0)
            if _full and _lid:
                nome_to_id[_full] = _lid
    except Exception:
        logger.warning("Fallback nome-based per contratti non disponibile")

    # Raggruppa righe per (CF, data_inizio, data_fine)
    groups: dict[tuple, dict] = {}
    reader = csv.DictReader(StringIO(text), delimiter=";")
    for row in reader:
        cf = (row.get("Codice fiscale") or "").strip().upper()
        if not cf:
            continue
        di_raw = (row.get("Data Inizio") or "").strip()
        df_raw = (row.get("Data Fine") or "").strip()
        if not di_raw:
            continue
        try:
            di = datetime.strptime(di_raw, "%d/%m/%Y").date()
            df = datetime.strptime(df_raw, "%d/%m/%Y").date() if df_raw else None
        except ValueError:
            continue
        key = (cf, di, df)
        if key not in groups:
            groups[key] = {
                "tax_code": cf, "data_inizio": di, "data_fine": df,
                "tipologia_contratto": "", "qualifica_nome": "",
                "codice_livello": "", "ccnl": "", "descrizione_livello": "",
            }
        g = groups[key]
        tipo = (row.get("Tipo di contratto") or "").strip()
        qualifica = (row.get("Qualifica") or "").strip()
        livello = (row.get("Livello") or "").strip()
        ccnl = (row.get("CCNL") or "").strip()
        desc = (row.get("Descrizione livello") or "").strip()
        if tipo and not g["tipologia_contratto"]:
            g["tipologia_contratto"] = _normalize_contratto_choice(tipo)
        if qualifica and not g["qualifica_nome"]:
            g["qualifica_nome"] = qualifica
        if livello and not g["codice_livello"]:
            g["codice_livello"] = livello.upper()
        if ccnl and not g["ccnl"]:
            g["ccnl"] = ccnl
        if desc and not g["descrizione_livello"]:
            g["descrizione_livello"] = desc

    created_n = updated_n = 0
    for (cf, di, df), data in groups.items():
        lid = cf_to_id.get(cf) or nome_to_id.get(cf)
        defaults = {
            "tipologia_contratto": data["tipologia_contratto"],
            "qualifica_nome": data["qualifica_nome"],
            "codice_livello": data["codice_livello"],
            "ccnl": data["ccnl"],
            "descrizione_livello": data["descrizione_livello"],
            "legacy_anagrafica_id": lid,
            "importato_da": user,
        }
        _, created = StoricoContratto.objects.update_or_create(
            tax_code=cf, data_inizio=di, data_fine=df,
            defaults=defaults,
        )
        if created:
            created_n += 1
        else:
            updated_n += 1
    return created_n, updated_n, 0


@login_required
def contratti_import(request):
    """Pagina importazione CSV contratti (solo admin)."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("anagrafica:dipendenti_list")

    stats = None
    if request.method == "POST":
        file_obj = request.FILES.get("file_csv")
        if not file_obj:
            messages.error(request, "Nessun file selezionato.")
        else:
            try:
                created, updated, _ = _import_csv_contratti(file_obj, request.user)
                messages.success(
                    request,
                    f"Importazione completata: {created} record creati, {updated} aggiornati.",
                )
                stats = {"created": created, "updated": updated}
                return redirect("anagrafica:contratti_import")
            except Exception as exc:
                logger.exception("Errore importazione CSV contratti")
                messages.error(request, f"Errore durante l'importazione: {exc}")

    recenti = list(
        StoricoContratto.objects.select_related("importato_da").order_by("-created_at")[:30]
    )
    return render(request, "anagrafica/pages/contratti_import.html", {
        "recenti": recenti,
        "is_admin": is_admin,
    })


@login_required
@require_POST
def dipendente_contratto_add(request, legacy_id: int):
    """Aggiunge manualmente un record storico contrattuale."""
    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    di_raw = request.POST.get("data_inizio", "").strip()
    df_raw = request.POST.get("data_fine", "").strip()
    try:
        data_inizio = datetime.strptime(di_raw, "%Y-%m-%d").date()
        data_fine = datetime.strptime(df_raw, "%Y-%m-%d").date() if df_raw else None
    except ValueError:
        messages.error(request, "Formato data non valido.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
    tax_code = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else ""

    # Auto-chiudi il record "in corso" se il nuovo inizia dopo
    _q_inc = Q(legacy_anagrafica_id=legacy_id)
    if tax_code:
        _q_inc |= Q(tax_code=tax_code)
    in_corso = (
        StoricoContratto.objects.filter(_q_inc, data_fine__isnull=True)
        .order_by("-data_inizio")
        .first()
    )
    if in_corso and in_corso.data_inizio < data_inizio:
        in_corso.data_fine = data_inizio
        in_corso.save(update_fields=["data_fine"])

    codice_liv = request.POST.get("codice_livello", "").strip().upper()
    # Auto-popola descrizione dal catalogo se non specificata manualmente
    desc_liv = request.POST.get("descrizione_livello", "").strip()
    if codice_liv and not desc_liv:
        _lc = LivelloContrattuale.objects.filter(codice=codice_liv).first()
        if _lc:
            desc_liv = _lc.descrizione

    StoricoContratto.objects.create(
        legacy_anagrafica_id=legacy_id,
        tax_code=tax_code,
        data_inizio=data_inizio,
        data_fine=data_fine,
        tipologia_contratto=request.POST.get("tipologia_contratto", "").strip(),
        codice_livello=codice_liv,
        descrizione_livello=desc_liv,
        qualifica_nome=request.POST.get("qualifica_nome", "").strip(),
        importato_da=request.user,
    )
    messages.success(request, "Record contrattuale aggiunto.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_contratto_delete(request, legacy_id: int, contratto_id: int):
    """Elimina un record storico contrattuale."""
    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    contratto = get_object_or_404(StoricoContratto, pk=contratto_id)
    if contratto.legacy_anagrafica_id != legacy_id:
        civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
        tc = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else None
        if contratto.tax_code != tc:
            messages.error(request, "Record non appartenente a questo dipendente.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    contratto.delete()
    messages.success(request, "Record contrattuale eliminato.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_contratto_edit(request, legacy_id: int, contratto_id: int):
    """Aggiorna un record storico contrattuale esistente."""
    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso riservato agli utenti HR.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    contratto = get_object_or_404(StoricoContratto, pk=contratto_id)
    if contratto.legacy_anagrafica_id != legacy_id:
        civile_obj = DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id=legacy_id).first()
        tc = civile_obj.codice_fiscale.strip().upper() if civile_obj and civile_obj.codice_fiscale else None
        if contratto.tax_code != tc:
            messages.error(request, "Record non appartenente a questo dipendente.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    di_raw = request.POST.get("data_inizio", "").strip()
    df_raw = request.POST.get("data_fine", "").strip()
    try:
        data_inizio = datetime.strptime(di_raw, "%Y-%m-%d").date()
        data_fine = datetime.strptime(df_raw, "%Y-%m-%d").date() if df_raw else None
    except ValueError:
        messages.error(request, "Formato data non valido.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    codice_liv = request.POST.get("codice_livello", "").strip().upper()
    desc_liv = request.POST.get("descrizione_livello", "").strip()
    if codice_liv and not desc_liv:
        _lc = LivelloContrattuale.objects.filter(codice=codice_liv).first()
        if _lc:
            desc_liv = _lc.descrizione

    contratto.data_inizio = data_inizio
    contratto.data_fine = data_fine
    contratto.tipologia_contratto = request.POST.get("tipologia_contratto", "").strip()
    contratto.codice_livello = codice_liv
    contratto.descrizione_livello = desc_liv
    contratto.qualifica_nome = request.POST.get("qualifica_nome", "").strip()
    contratto.save()
    messages.success(request, "Record contrattuale aggiornato.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Report dipendenti — filtri avanzati + export CSV
# ---------------------------------------------------------------------------

@login_required
def dipendenti_report(request):
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("anagrafica:dipendenti_list")

    q = request.GET.get("q", "").strip()
    reparto_filter = request.GET.get("reparto", "").strip()
    area_filter = request.GET.get("area", "").strip()
    contratto_filter = request.GET.get("tipologia_contratto", "").strip()
    consenso_filter = request.GET.get("consenso_privacy", "").strip()
    cat_protetta_filter = request.GET.get("categoria_protetta", "").strip()

    # Parte da tutti i dipendenti legacy
    all_rows = fetch_anagrafica_rows(deduplicate=True)

    # Filtro per testo su nome/cognome
    if q:
        q_norm = q.casefold()
        all_rows = [
            row for row in all_rows
            if any(
                q_norm in str(row.get(k) or "").casefold()
                for k in ("nome", "cognome", "aliasusername", "matricola")
            )
        ]

    # Reparto/area canonici: sovrascrive row["reparto"] col Reparto canonico
    # (fallback al testo legacy per i non ancora mappati) e valorizza
    # row["area_aziendale_nome"]. Da qui il filtro reparto è sul canonico.
    from anagrafica.services.reparto_canonico import enrich_rows_reparto_canonico
    enrich_rows_reparto_canonico(all_rows)

    # Filtro per reparto (canonico)
    if reparto_filter:
        all_rows = [
            row for row in all_rows
            if str(row.get("reparto") or "").strip().casefold() == reparto_filter.casefold()
        ]

    # Filtri su campi Django: area aziendale canonica e tipologia_contratto
    django_filter_ids: set[int] | None = None
    az_qs = DipendenteAnagraficaAziendale.objects.all()
    if area_filter:
        az_qs = az_qs.filter(area_aziendale__nome__iexact=area_filter)
    if contratto_filter:
        az_qs = az_qs.filter(tipologia_contratto=contratto_filter)
    if consenso_filter == "si":
        az_qs = az_qs.filter(consenso_privacy=True)
    elif consenso_filter == "no":
        az_qs = az_qs.filter(consenso_privacy=False)
    if area_filter or contratto_filter or consenso_filter:
        django_filter_ids = set(az_qs.values_list("legacy_anagrafica_id", flat=True))

    # Filtro categoria protetta (da DipendenteAnagraficaCivile)
    civ_filter_ids: set[int] | None = None
    if cat_protetta_filter == "si":
        civ_filter_ids = set(
            DipendenteAnagraficaCivile.objects.filter(categoria_protetta=True)
            .values_list("legacy_anagrafica_id", flat=True)
        )
    elif cat_protetta_filter == "no":
        civ_filter_ids = set(
            DipendenteAnagraficaCivile.objects.filter(categoria_protetta=False)
            .values_list("legacy_anagrafica_id", flat=True)
        )

    if django_filter_ids is not None:
        all_rows = [row for row in all_rows if int(row.get("id") or 0) in django_filter_ids]
    if civ_filter_ids is not None:
        all_rows = [row for row in all_rows if int(row.get("id") or 0) in civ_filter_ids]

    # Arricchisce ogni riga con dati Django
    legacy_ids = [int(row.get("id") or 0) for row in all_rows if int(row.get("id") or 0)]
    az_map = {
        obj.legacy_anagrafica_id: obj
        for obj in DipendenteAnagraficaAziendale.objects.filter(legacy_anagrafica_id__in=legacy_ids)
    }
    civ_map = {
        obj.legacy_anagrafica_id: obj
        for obj in DipendenteAnagraficaCivile.objects.filter(legacy_anagrafica_id__in=legacy_ids)
    }
    for row in all_rows:
        lid = int(row.get("id") or 0)
        az = az_map.get(lid)
        row["az"] = az
        row["civ"] = civ_map.get(lid)

    # Export CSV (no campi sensibili)
    fmt = request.GET.get("format", "").strip()
    if fmt == "csv":
        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type=CSV_CONTENT_TYPE)
        response["Content-Disposition"] = 'attachment; filename="dipendenti_report.csv"'
        response.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(response, delimiter=";")
        writer.writerow([
            "ID", "Cognome", "Nome", "Matricola", "Reparto",
            "Area aziendale", "Ruolo aziendale", "Tipologia contratto",
            "Livello inquadramento", "Data prima assunzione",
            "Consenso privacy", "Email aziendale", "Telefono aziendale",
        ])
        for row in all_rows:
            az = row.get("az")
            writer.writerow([
                row.get("id", ""),
                row.get("cognome", ""),
                row.get("nome", ""),
                row.get("matricola", ""),
                row.get("reparto", ""),
                row.get("area_aziendale_nome", "") or "",
                getattr(az, "ruolo_aziendale", "") or "",
                getattr(az, "get_tipologia_contratto_display", lambda: "")() if az else "",
                getattr(az, "livello_inquadramento", "") or "",
                getattr(az, "data_prima_assunzione", "") or "",
                "Sì" if getattr(az, "consenso_privacy", False) else "No",
                getattr(az, "email_aziendale", "") or "",
                getattr(az, "telefono_aziendale", "") or "",
            ])
        return response

    # Filtri dai cataloghi canonici (non più dal testo legacy).
    reparti_list = list(
        Reparto.objects.filter(is_active=True).order_by("nome").values_list("nome", flat=True)
    )
    aree_list = list(
        AreaAziendale.objects.filter(is_active=True).order_by("nome").values_list("nome", flat=True)
    )

    paginator = Paginator(all_rows, 50)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "anagrafica/pages/dipendenti_report.html", {
        "page_obj": page,
        "q": q,
        "reparto": reparto_filter,
        "area": area_filter,
        "tipologia_contratto": contratto_filter,
        "consenso_privacy": consenso_filter,
        "categoria_protetta": cat_protetta_filter,
        "reparti": reparti_list,
        "aree": aree_list,
        "contratto_choices": DipendenteAnagraficaAziendale.CONTRATTO_CHOICES,
        "n_totale": len(all_rows),
    })


# ---------------------------------------------------------------------------
# Export tabellari (PDF/Excel) — endpoint unico parametrico
# ---------------------------------------------------------------------------

@login_required
def export_view(request, key: str):
    """Endpoint unico di export delle liste di anagrafica (PDF/Excel).

    La chiave identifica la ``ExportSpec`` (vedi `anagrafica/exports.py`); il gate
    ACL della lista di origine è applicato dalla spec, l'audit dall'helper.
    """
    from anagrafica.exports import EXPORT_SPECS, build_export_response

    spec = EXPORT_SPECS.get(key)
    if spec is None:
        raise Http404("Export non disponibile.")
    if not spec.permission(request):
        return HttpResponseForbidden("Permessi insufficienti.")
    return build_export_response(
        request,
        key,
        request.GET.get("format", "xlsx"),
        request.GET.get("scope", "filtered"),
    )


# ---------------------------------------------------------------------------
# Mansioni — gestione catalogo
# ---------------------------------------------------------------------------

@login_required
def mansioni_list(request):
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    can_view_requisiti = _can_view_formazione(request)

    filtro_rischio = (request.GET.get("rischio") or "").strip().upper()
    if filtro_rischio not in dict(Mansione.LIVELLO_RISCHIO_CHOICES):
        filtro_rischio = ""
    solo_rischio = request.GET.get("solo_rischio") == "1"
    q_text = (request.GET.get("q") or "").strip()

    mansioni = list(
        Mansione.objects.all()
        .order_by("nome")
        .prefetch_related("visite_richieste", "dpi_richiesti")
    )

    # Conta dipendenti per mansione dal DB legacy
    mansione_counts: dict[str, int] = {}
    try:
        from django.db import connections
        conn_name = "legacy" if "legacy" in connections else "default"
        with connections[conn_name].cursor() as cur:
            cur.execute(
                "SELECT LOWER(mansione), COUNT(*) FROM anagrafica_dipendenti "
                "WHERE mansione IS NOT NULL AND mansione != '' GROUP BY LOWER(mansione)"
            )
            for row in cur.fetchall():
                mansione_counts[row[0]] = row[1]
    except Exception:
        pass

    for m in mansioni:
        m.n_dipendenti = mansione_counts.get(m.nome.lower(), 0)
        # Contatori requisiti (usano la cache di prefetch_related → nessuna query extra)
        m.n_visite = len(m.visite_richieste.all())
        try:
            m.n_dpi = len(m.dpi_richiesti.all())
        except Exception:
            m.n_dpi = 0
        m.livello_label = m.get_livello_rischio_display() if m.livello_rischio else ""
        m.is_rischio = bool(m.livello_rischio or m.n_dpi or m.n_visite)

    n_rischio_tot = sum(1 for m in mansioni if m.is_rischio)

    # Filtri (livello rischio, solo mansioni di rischio, ricerca nome)
    def _match(m) -> bool:
        if q_text and q_text.casefold() not in m.nome.casefold():
            return False
        if filtro_rischio and m.livello_rischio != filtro_rischio:
            return False
        if solo_rischio and not m.is_rischio:
            return False
        return True

    if filtro_rischio or solo_rischio or q_text:
        mansioni = [m for m in mansioni if _match(m)]

    # Raggruppa per categoria nell'ordine definito
    cat_order = [c for c, _ in Mansione.CATEGORIA_CHOICES]
    cat_labels = dict(Mansione.CATEGORIA_CHOICES)
    # Le mansioni senza categoria vanno in "Altro"
    grouped: list[tuple[str, str, list]] = []
    seen_cats: set[str] = set()
    for cat_code in cat_order:
        items = [m for m in mansioni if (m.categoria or Mansione.CAT_ALTRO) == cat_code]
        if items:
            grouped.append((cat_code, cat_labels[cat_code], items))
            seen_cats.add(cat_code)
    # Mansioni senza categoria classificate come "Altro"
    senza_cat = [m for m in mansioni if not m.categoria and Mansione.CAT_ALTRO not in seen_cats]
    if senza_cat:
        altro_label = cat_labels.get(Mansione.CAT_ALTRO, "Altro")
        grouped.append((Mansione.CAT_ALTRO, altro_label, senza_cat))

    mansioni_suggerite = [
        "Operaio generico", "Operaio specializzato", "Operaio qualificato",
        "Impiegato", "Impiegato tecnico", "Impiegato amministrativo",
        "Quadro", "Dirigente",
    ]

    return render(request, "anagrafica/pages/mansioni_list.html", {
        "mansioni": mansioni,
        "mansioni_grouped": grouped,
        "is_admin": is_admin,
        "can_view_requisiti": can_view_requisiti,
        "mansioni_suggerite": mansioni_suggerite,
        "CATEGORIA_CHOICES": Mansione.CATEGORIA_CHOICES,
        "LIVELLO_RISCHIO_CHOICES": Mansione.LIVELLO_RISCHIO_CHOICES,
        "filtro_rischio": filtro_rischio,
        "solo_rischio": solo_rischio,
        "q_text": q_text,
        "n_rischio_tot": n_rischio_tot,
        "n_visibili": len(mansioni),
    })


@login_required
@require_POST
def mansione_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare mansioni.")
        return _back_to_caller(request, "anagrafica:mansioni_list")

    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome della mansione è obbligatorio.")
        return _back_to_caller(request, "anagrafica:mansioni_list")

    _lr = (request.POST.get("livello_rischio") or "").strip().upper()
    _lr = _lr if _lr in dict(Mansione.LIVELLO_RISCHIO_CHOICES) else ""
    _, created = Mansione.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "categoria": (request.POST.get("categoria") or "").strip()[:20],
            "descrizione": (request.POST.get("descrizione") or "").strip(),
            "colore": (request.POST.get("colore") or "#64748b").strip()[:7],
            "livello_rischio": _lr,
        },
    )
    if created:
        messages.success(request, f'Mansione "{nome}" creata.')
    else:
        messages.warning(request, f'Esiste già una mansione con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:mansioni_list")


@login_required
@require_POST
def mansione_edit(request, mansione_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare mansioni.")
        return _back_to_caller(request, "anagrafica:mansioni_list")

    mansione = get_object_or_404(Mansione, pk=mansione_id)
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome della mansione è obbligatorio.")
        return _back_to_caller(request, "anagrafica:mansioni_list")

    _lr = (request.POST.get("livello_rischio") or "").strip().upper()
    mansione.nome = nome
    mansione.categoria = (request.POST.get("categoria") or "").strip()[:20]
    mansione.descrizione = (request.POST.get("descrizione") or "").strip()
    mansione.colore = (request.POST.get("colore") or "#64748b").strip()[:7]
    mansione.livello_rischio = _lr if _lr in dict(Mansione.LIVELLO_RISCHIO_CHOICES) else ""
    mansione.is_active = request.POST.get("is_active") == "1"
    mansione.save()
    messages.success(request, f'Mansione "{mansione.nome}" aggiornata.')
    return _back_to_caller(request, "anagrafica:mansioni_list")


@login_required
@require_POST
def mansione_delete(request, mansione_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare mansioni.")
        return _back_to_caller(request, "anagrafica:mansioni_list")

    mansione = get_object_or_404(Mansione, pk=mansione_id)
    nome = mansione.nome
    mansione.delete()
    messages.success(request, f'Mansione "{nome}" eliminata.')
    return _back_to_caller(request, "anagrafica:mansioni_list")


@login_required
def mansione_requisiti(request, mansione_id: int):
    """Profilo "mansione di rischio": DPI / visite / formazione richiesti.

    GET mostra i requisiti **diretti** (modificabili) della mansione, quelli
    **ereditati** dai fattori di rischio esposti (read-only) e il riepilogo dei
    requisiti effettivi (unione). POST salva i M2M diretti ``dpi_richiesti`` e
    ``visite_richieste``. Gate: visualizzazione/edit formazione (dominio Safety).
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per i requisiti mansione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    mansione = get_object_or_404(Mansione, pk=mansione_id)

    if request.method == "POST":
        if not is_editor:
            messages.error(request, "Permessi insufficienti per modificare i requisiti.")
            return redirect("anagrafica:mansione_requisiti", mansione_id=mansione.pk)
        _lr = (request.POST.get("livello_rischio") or "").strip().upper()
        mansione.livello_rischio = _lr if _lr in dict(Mansione.LIVELLO_RISCHIO_CHOICES) else ""
        mansione.save(update_fields=["livello_rischio"])
        visite_ids = [int(v) for v in request.POST.getlist("visite_richieste") if str(v).isdigit()]
        mansione.visite_richieste.set(
            TipoVisitaMedica.objects.filter(pk__in=visite_ids, is_active=True)
        )
        try:
            from dpi.models import CategoriaDPI
            dpi_ids = [int(v) for v in request.POST.getlist("dpi_richiesti") if str(v).isdigit()]
            mansione.dpi_richiesti.set(CategoriaDPI.objects.filter(pk__in=dpi_ids, is_active=True))
        except Exception:
            logger.warning("Salvataggio DPI mansione fallito (modulo dpi?)", exc_info=True)
        messages.success(request, f'Requisiti della mansione "{mansione.nome}" aggiornati.')
        return redirect("anagrafica:mansione_requisiti", mansione_id=mansione.pk)

    # Requisiti effettivi (unione diretti + ereditati) per il riepilogo.
    requisiti = mansionario_service.requisiti_mansione(mansione)

    # Cataloghi per i selettori dei requisiti diretti.
    visite_opts = list(TipoVisitaMedica.objects.filter(is_active=True).order_by("nome"))
    sel_visite_ids = set(mansione.visite_richieste.values_list("pk", flat=True))
    dpi_opts: list = []
    sel_dpi_ids: set[int] = set()
    try:
        from dpi.models import CategoriaDPI
        dpi_opts = list(CategoriaDPI.objects.filter(is_active=True).order_by("order_index", "nome"))
        sel_dpi_ids = set(mansione.dpi_richiesti.values_list("pk", flat=True))
    except Exception:
        dpi_opts = []

    return render(request, "anagrafica/pages/mansione_requisiti.html", {
        "mansione": mansione,
        "is_editor": is_editor,
        "requisiti": requisiti,
        "visite_opts": visite_opts,
        "sel_visite_ids": sel_visite_ids,
        "dpi_opts": dpi_opts,
        "sel_dpi_ids": sel_dpi_ids,
        "livello_choices": Mansione.LIVELLO_RISCHIO_CHOICES,
    })


# ---------------------------------------------------------------------------
# Aree aziendali + Reparti — gerarchia a due livelli
# ---------------------------------------------------------------------------

def _dipendenti_picker_rows() -> list[dict]:
    """Elenco compatto dipendenti in forza per la selezione del caporeparto."""
    rows = fetch_anagrafica_rows(deduplicate=True)
    cessati = _cessati_legacy_ids()
    items: list[dict] = []
    for row in rows:
        legacy_id = int(row.get("id") or 0)
        if legacy_id <= 0 or legacy_id in cessati:
            continue
        nome = str(row.get("nome") or "").strip()
        cognome = str(row.get("cognome") or "").strip()
        label = " ".join(part for part in [cognome, nome] if part) or str(row.get("aliasusername") or "").strip() or f"#{legacy_id}"
        items.append({"id": legacy_id, "label": label})
    items.sort(key=lambda r: r["label"].casefold())
    return items


def _resolve_caporeparto_id(raw: str | None) -> int | None:
    """Normalizza il valore POST del caporeparto in un legacy_id valido o None."""
    if not raw:
        return None
    try:
        value = int(str(raw).strip())
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def _sync_aziendale_from_reparto(
    legacy_id: int, reparto_nome: str, *, area_aziendale_id: int | None = None, saved_by
) -> None:
    """Aggiorna caporeparto_legacy_id e area_aziendale su DipendenteAnagraficaAziendale
    in base al Reparto assegnato. Chiamato ogni volta che il reparto (o l'area) cambia.

    L'Area aziendale deve appartenere al Reparto risolto, altrimenti viene azzerata
    silenziosamente (reparto cambiato altrove, reparto non trovato/disattivato, o area
    di un altro reparto) invece di bloccare il salvataggio. Non si richiede che l'area
    sia attiva: un'assegnazione a un'area nel frattempo disattivata resta valida, come
    già avviene per area/ruolo_aziendale (forms.py preserva il valore corrente anche
    se non più nel catalogo "attive").
    """
    capo_id = None
    rep = None
    if reparto_nome:
        rep = Reparto.objects.filter(nome__iexact=reparto_nome, is_active=True).first()
        if rep:
            capo_id = rep.caporeparto_legacy_id

    area_id_valido = None
    if area_aziendale_id and rep is not None:
        area = AreaAziendale.objects.filter(pk=area_aziendale_id, reparto_id=rep.id).first()
        if area is not None:
            area_id_valido = area.id
            # Responsabile effettivo: il responsabile dell'AREA aziendale vince
            # sul caporeparto del REPARTO quando differisce (fallback al capo
            # reparto se l'area non ha responsabile). Fonte unica nel service.
            from anagrafica.services.reparto_canonico import resolve_responsabile_effettivo
            capo_id = resolve_responsabile_effettivo(area=area, reparto=rep)

    az, _ = DipendenteAnagraficaAziendale.objects.get_or_create(
        legacy_anagrafica_id=legacy_id,
        defaults={"updated_by": saved_by},
    )
    az.area = reparto_nome
    az.caporeparto_legacy_id = capo_id
    az.area_aziendale_id = area_id_valido
    az.updated_by = saved_by
    az.save(update_fields=["area", "caporeparto_legacy_id", "area_aziendale", "updated_by", "updated_at"])


@login_required
def aree_list(request):
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    reparti = list(Reparto.objects.prefetch_related("aree_aziendali").order_by("nome"))
    aree_senza_reparto = list(AreaAziendale.objects.filter(reparto__isnull=True).order_by("nome"))
    dipendenti = _dipendenti_picker_rows()
    dip_map = {item["id"]: item["label"] for item in dipendenti}
    for rep in reparti:
        rep.caporeparto_label = dip_map.get(rep.caporeparto_legacy_id or 0, "")
        for area in rep.aree_aziendali.all():
            area.responsabile_label = dip_map.get(area.responsabile_legacy_id or 0, "")
    for area in aree_senza_reparto:
        area.responsabile_label = dip_map.get(area.responsabile_legacy_id or 0, "")
    return render(request, "anagrafica/pages/aree_list.html", {
        "reparti": reparti,
        "aree_senza_reparto": aree_senza_reparto,
        "is_admin": is_admin,
        "dipendenti_picker": dipendenti,
    })


# ── Area Aziendale CRUD (ora il livello FIGLIO) ─────────────────────────────

@login_required
@require_POST
def area_aziendale_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare aree aziendali.")
        return _back_to_caller(request, "anagrafica:aree_list")
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome dell'area aziendale è obbligatorio.")
        return _back_to_caller(request, "anagrafica:aree_list")
    reparto_id = request.POST.get("reparto_id") or None
    reparto = None
    if reparto_id:
        try:
            reparto = Reparto.objects.get(pk=int(reparto_id))
        except (Reparto.DoesNotExist, ValueError):
            pass
    obj, created = AreaAziendale.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "descrizione": (request.POST.get("descrizione") or "").strip(),
            "reparto": reparto,
            "responsabile_legacy_id": _resolve_caporeparto_id(request.POST.get("responsabile_legacy_id")),
        },
    )
    if created:
        messages.success(request, f'Area aziendale "{nome}" creata.')
    else:
        messages.warning(request, f'Esiste già un\'area aziendale con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:aree_list")


@login_required
@require_POST
def area_aziendale_edit(request, area_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare aree aziendali.")
        return _back_to_caller(request, "anagrafica:aree_list")
    area = get_object_or_404(AreaAziendale, pk=area_id)
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome dell'area aziendale è obbligatorio.")
        return _back_to_caller(request, "anagrafica:aree_list")
    reparto_id = request.POST.get("reparto_id") or None
    reparto = None
    if reparto_id:
        try:
            reparto = Reparto.objects.get(pk=int(reparto_id))
        except (Reparto.DoesNotExist, ValueError):
            pass
    area.nome = nome
    area.descrizione = (request.POST.get("descrizione") or "").strip()
    area.reparto = reparto
    area.responsabile_legacy_id = _resolve_caporeparto_id(request.POST.get("responsabile_legacy_id"))
    area.is_active = request.POST.get("is_active") == "1"
    area.save()
    messages.success(request, f'Area aziendale "{area.nome}" aggiornata.')
    return _back_to_caller(request, "anagrafica:aree_list")


@login_required
@require_POST
def area_aziendale_delete(request, area_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare aree aziendali.")
        return _back_to_caller(request, "anagrafica:aree_list")
    area = get_object_or_404(AreaAziendale, pk=area_id)
    nome = area.nome
    area.delete()
    messages.success(request, f'Area aziendale "{nome}" eliminata.')
    return _back_to_caller(request, "anagrafica:aree_list")


# ── Reparto CRUD (ora il livello PADRE) ─────────────────────────────────────

def _sync_reparto_capo_mapping(rep) -> None:
    """Allinea RepartoCapoMapping al valore di Reparto.caporeparto_legacy_id.

    Chiamata dopo ogni create/edit di un Reparto per mantenere la tabella
    RepartoCapoMapping (usata da assenze e automazioni) in sincronia con la
    fonte di verità in Anagrafica HR.
    """
    from core.caporeparto_utils import canonical_caporeparto_value
    from core.models import RepartoCapoMapping

    reparto_nome = (rep.nome or "").strip()
    if not reparto_nome:
        return

    RepartoCapoMapping.objects.filter(reparto__iexact=reparto_nome).delete()

    if not rep.caporeparto_legacy_id:
        return

    capo_str = canonical_caporeparto_value(legacy_user_id=rep.caporeparto_legacy_id)
    if not capo_str:
        return

    RepartoCapoMapping.objects.create(
        reparto=reparto_nome,
        caporeparto=capo_str,
        is_active=True,
    )


@login_required
@require_POST
def area_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare reparti.")
        return _back_to_caller(request, "anagrafica:aree_list")
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome del reparto è obbligatorio.")
        return _back_to_caller(request, "anagrafica:aree_list")
    capo_id = _resolve_caporeparto_id(request.POST.get("caporeparto_legacy_id"))
    obj, created = Reparto.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "descrizione": (request.POST.get("descrizione") or "").strip(),
            "colore": (request.POST.get("colore") or "#64748b").strip()[:7],
            "caporeparto_legacy_id": capo_id,
        },
    )
    if created:
        messages.success(request, f'Reparto "{nome}" creato.')
        _sync_reparto_capo_mapping(obj)
    else:
        messages.warning(request, f'Esiste già un reparto con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:aree_list")


@login_required
@require_POST
def area_edit(request, area_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare reparti.")
        return _back_to_caller(request, "anagrafica:aree_list")
    rep = get_object_or_404(Reparto, pk=area_id)
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not nome:
        messages.error(request, "Il nome del reparto è obbligatorio.")
        return _back_to_caller(request, "anagrafica:aree_list")
    rep.nome = nome
    rep.descrizione = (request.POST.get("descrizione") or "").strip()
    rep.colore = (request.POST.get("colore") or "#64748b").strip()[:7]
    rep.is_active = request.POST.get("is_active") == "1"
    rep.caporeparto_legacy_id = _resolve_caporeparto_id(request.POST.get("caporeparto_legacy_id"))
    rep.save()
    _sync_reparto_capo_mapping(rep)
    DipendenteAnagraficaAziendale.objects.filter(area__iexact=rep.nome).update(
        caporeparto_legacy_id=rep.caporeparto_legacy_id
    )
    messages.success(request, f'Reparto "{rep.nome}" aggiornato.')
    return _back_to_caller(request, "anagrafica:aree_list")


@login_required
@require_POST
def area_delete(request, area_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare reparti.")
        return _back_to_caller(request, "anagrafica:aree_list")
    rep = get_object_or_404(Reparto, pk=area_id)
    if rep.aree_aziendali.exists():
        messages.error(request, f'Impossibile eliminare: il reparto "{rep.nome}" ha aree aziendali associate. Riassegna prima le aree.')
        return _back_to_caller(request, "anagrafica:aree_list")
    nome = rep.nome
    rep.delete()
    messages.success(request, f'Reparto "{nome}" eliminato.')
    return _back_to_caller(request, "anagrafica:aree_list")


# ---------------------------------------------------------------------------
# Ruoli aziendali — catalogo dropdown
# ---------------------------------------------------------------------------

@login_required
def ruoli_aziendali_list(request):
    # Fase 2: «Ruoli aziendali» e «Ruoli operativi» sono un catalogo unico.
    # La pagina dedicata è confluita in quella unificata dei Ruoli.
    messages.info(
        request,
        "I ruoli aziendali e operativi sono ora un catalogo unico: gestiscili da qui.",
    )
    return redirect("anagrafica:ruoli_operativi_list")


@login_required
@require_POST
def ruolo_aziendale_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare ruoli aziendali.")
        return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")
    nome = (request.POST.get("nome") or "").strip()[:200]
    if not nome:
        messages.error(request, "Il nome del ruolo è obbligatorio.")
        return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")
    _, created = RuoloAziendale.objects.get_or_create(
        nome__iexact=nome,
        defaults={"nome": nome, "descrizione": (request.POST.get("descrizione") or "").strip()},
    )
    if created:
        messages.success(request, f'Ruolo aziendale "{nome}" creato.')
    else:
        messages.warning(request, f'Esiste già un ruolo aziendale con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")


@login_required
@require_POST
def ruolo_aziendale_edit(request, ruolo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare ruoli aziendali.")
        return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")
    ruolo = get_object_or_404(RuoloAziendale, pk=ruolo_id)
    nome = (request.POST.get("nome") or "").strip()[:200]
    if not nome:
        messages.error(request, "Il nome del ruolo è obbligatorio.")
        return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")
    ruolo.nome = nome
    ruolo.descrizione = (request.POST.get("descrizione") or "").strip()
    ruolo.is_active = request.POST.get("is_active") == "1"
    ruolo.save()
    messages.success(request, f'Ruolo aziendale "{ruolo.nome}" aggiornato.')
    return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")


@login_required
@require_POST
def ruolo_aziendale_delete(request, ruolo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare ruoli aziendali.")
        return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")
    ruolo = get_object_or_404(RuoloAziendale, pk=ruolo_id)
    nome = ruolo.nome
    ruolo.delete()
    messages.success(request, f'Ruolo aziendale "{nome}" eliminato.')
    return _back_to_caller(request, "anagrafica:ruoli_aziendali_list")


# ---------------------------------------------------------------------------
# Qualifiche — gestione catalogo + panoramica scadenze
# ---------------------------------------------------------------------------

@login_required
def qualifiche_list(request):
    from datetime import timedelta
    from django.utils import timezone as tz

    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)

    # Catalogo unico, viste filtrate per categoria: la pagina è la "casa" delle
    # qualifiche (modulo Formazione) ma può aprirsi già filtrata (es. Salute e
    # Sicurezza → ?categoria=SICUREZZA).
    # Pseudo-categoria virtuale per i Processi qualificati MOD.128: NON è una choice
    # del modello (riusa i ProcessoQualificato caricati più sotto), quindi vive solo
    # come voce di tab + flag di visibilità. Nessuna migrazione (stream 3).
    CAT_PROCESSI, LBL_PROCESSI = "PROCESSI", "Processi qualificati"
    valid_cats = {c for c, _ in TipoQualifica.CATEGORIA_CHOICES} | {CAT_PROCESSI}
    cat_filter = (request.GET.get("categoria") or "").strip().upper()
    if cat_filter not in valid_cats:
        cat_filter = ""

    tipi = list(
        TipoQualifica.objects.annotate(n_assegnazioni=Count("assegnazioni"))
        .prefetch_related("corsi")  # corso(i) che rilasciano la qualifica (Step 3)
        .order_by("categoria", "nome")
    )

    oggi = tz.localdate()
    soglia = oggi + timedelta(days=60)

    # Qualifiche in scadenza o scadute
    scadenze_qs = list(
        DipendenteQualifica.objects.filter(data_scadenza__isnull=False)
        .filter(data_scadenza__lte=soglia)
        .select_related("tipo")
        .order_by("data_scadenza")
    )

    # Arricchisce con nome dipendente
    legacy_ids = list({q.legacy_anagrafica_id for q in scadenze_qs})
    nome_map: dict[int, str] = {}
    if legacy_ids:
        try:
            rows = fetch_anagrafica_rows(ids=legacy_ids)
            for r in rows:
                rid = int(r.get("id") or 0)
                if rid:
                    nome_map[rid] = f"{r.get('cognome', '')} {r.get('nome', '')}".strip()
        except Exception:
            pass

    scadenze = []
    for q in scadenze_qs:
        scadenze.append({
            "id": q.id,
            "legacy_anagrafica_id": q.legacy_anagrafica_id,
            "dipendente": nome_map.get(q.legacy_anagrafica_id, f"Dip. #{q.legacy_anagrafica_id}"),
            "tipo_nome": q.tipo.nome,
            "tipo_categoria": q.tipo.categoria,
            "data_conseguimento": q.data_conseguimento,
            "data_scadenza": q.data_scadenza,
            "scaduta": q.data_scadenza < oggi,
            "in_scadenza": oggi <= q.data_scadenza <= soglia,
            "note": q.note,
        })

    tipi_suggeriti = [
        ("Patentino carrellista", "PROFESSIONALE", 60),
        ("Primo soccorso", "SICUREZZA", 36),
        ("Addetto antincendio (livello 1)", "SICUREZZA", 60),
        ("Addetto antincendio (livello 2)", "SICUREZZA", 60),
        ("Addetto antincendio (livello 3)", "SICUREZZA", 60),
        ("RSPP", "SICUREZZA", 60),
        ("RLS", "SICUREZZA", 60),
        ("Preposto sicurezza", "SICUREZZA", 60),
        ("Uso DPI anticaduta", "SICUREZZA", 60),
    ]

    # Raggruppa tipi per categoria
    cat_order_q = [c for c, _ in TipoQualifica.CATEGORIA_CHOICES]
    cat_labels_q = dict(TipoQualifica.CATEGORIA_CHOICES)
    tipi_grouped: list[tuple[str, str, list]] = []
    for cat_code in cat_order_q:
        items = [t for t in tipi if t.categoria == cat_code]
        if items:
            tipi_grouped.append((cat_code, cat_labels_q[cat_code], items))

    # Barra tab "Tutte + categorie" con conteggi (sempre su tutto il catalogo).
    tabs = [("", "Tutte", len(tipi))]
    for cat_code in cat_order_q:
        tabs.append((cat_code, cat_labels_q[cat_code],
                     sum(1 for t in tipi if t.categoria == cat_code)))

    # Vista filtrata: restringe gruppi e scadenze alla categoria selezionata.
    if cat_filter:
        tipi_grouped = [g for g in tipi_grouped if g[0] == cat_filter]
        scadenze = [s for s in scadenze if s["tipo_categoria"] == cat_filter]

    # Processi qualificati MOD.128 (modulo additivo, fail-safe): mostrati nel
    # catalogo come famiglia a sé, così si vedono nella sezione Qualifiche.
    processi_qualificati = []
    try:
        from .models_mpq import ProcessoQualificato as _PQ
        processi_qualificati = list(
            _PQ.objects.select_related("cliente")
            .annotate(n_abil=Count("abilitazioni", distinct=True))
            .order_by("cliente__nome", "nome")
        )
    except Exception:
        processi_qualificati = []

    # Chip virtuale "Processi qualificati": voce di tab (con conteggio MOD.128) +
    # flag di visibilità. `mostra_processi` regge la sezione MOD.128 (Tutte + filtro
    # PROCESSI); `solo_processi` nasconde il catalogo tipi quando si filtra su PROCESSI.
    # Con cat_filter == "PROCESSI" il blocco `if cat_filter:` ha già svuotato
    # tipi_grouped e scadenze (nessun gruppo/scadenza ha quella categoria).
    tabs.append((CAT_PROCESSI, LBL_PROCESSI, len(processi_qualificati)))
    mostra_processi = cat_filter in ("", CAT_PROCESSI)
    solo_processi = cat_filter == CAT_PROCESSI

    return render(request, "anagrafica/pages/qualifiche_list.html", {
        "tipi": tipi,
        "tipi_grouped": tipi_grouped,
        "processi_qualificati": processi_qualificati,
        "mostra_processi": mostra_processi,
        "solo_processi": solo_processi,
        "scadenze": scadenze,
        "is_admin": is_admin,
        "oggi": oggi,
        "CATEGORIA_CHOICES": TipoQualifica.CATEGORIA_CHOICES,
        "tipi_suggeriti": tipi_suggeriti,
        "tabs": tabs,
        "active_categoria": cat_filter,
        "active_categoria_label": LBL_PROCESSI if cat_filter == CAT_PROCESSI else cat_labels_q.get(cat_filter, ""),
        "is_safety_view": cat_filter == TipoQualifica.CAT_SICUREZZA,
    })


def _classifica_scadenza_qualifica(data_scadenza, oggi, soglia_30, soglia_60):
    """Stato RAG di una qualifica in base alla scadenza. Fonte unica per
    cruscotto e scadenzario dedicato (coerente con matrice/scadenzario unificato)."""
    if data_scadenza is None:
        return ("permanente", "Permanente")
    if data_scadenza < oggi:
        return ("scaduta", "Scaduta")
    if data_scadenza <= soglia_30:
        return ("s30", "In scadenza ≤30gg")
    if data_scadenza <= soglia_60:
        return ("s60", "In scadenza ≤60gg")
    return ("valida", "Valida")


@login_required
def qualifiche_dashboard(request):
    """Cruscotto Qualifiche & Certificazioni — vista trasversale di sola lettura.

    AGGREGA (non duplica) i tre modelli sorgente — ``TipoQualifica``,
    ``DipendenteQualifica``, ``QualificaSessione`` — le stesse fonti usate da
    Formazione, ``matrice_competenze``, ``conformita_report`` e dalla scheda
    dipendente: qualsiasi rilascio/rinnovo fatto altrove si riflette qui.

    Le scadenze (promemoria email) restano gestite dal modulo automazioni
    (report settimanale + pacchetto ``au12``): il cruscotto le mostra e linka
    alla loro configurazione, non le ridefinisce.
    """
    from datetime import timedelta
    from collections import OrderedDict
    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    soglia_30 = oggi + timedelta(days=30)
    soglia_60 = oggi + timedelta(days=60)
    is_admin = request.user.is_superuser or is_legacy_admin(get_legacy_user(request.user))

    tipi_attivi = TipoQualifica.objects.filter(is_active=True).count()

    quals = list(DipendenteQualifica.objects.select_related("tipo").all())
    n_valide = n_scadute = n_s30 = n_s60 = n_permanenti = 0
    n_con_evidenza = n_da_verificare = 0
    dipendenti_ids: set[int] = set()
    cat_labels = dict(TipoQualifica.CATEGORIA_CHOICES)
    dist_cat = {c: 0 for c, _ in TipoQualifica.CATEGORIA_CHOICES}

    # Timeline scadenze prossimi 12 mesi
    buckets: "OrderedDict[tuple[int, int], int]" = OrderedDict()
    yy, mm = oggi.year, oggi.month
    for _ in range(12):
        buckets[(yy, mm)] = 0
        mm += 1
        if mm > 12:
            mm, yy = 1, yy + 1
    primo_del_mese = oggi.replace(day=1)

    for q in quals:
        dipendenti_ids.add(q.legacy_anagrafica_id)
        if q.tipo.categoria in dist_cat:
            dist_cat[q.tipo.categoria] += 1
        if q.documento:
            n_con_evidenza += 1
            if not q.verificata:
                n_da_verificare += 1
        d = q.data_scadenza
        if d is None:
            n_permanenti += 1
            n_valide += 1
        elif d < oggi:
            n_scadute += 1
        elif d <= soglia_30:
            n_s30 += 1
        elif d <= soglia_60:
            n_s60 += 1
        else:
            n_valide += 1
        if d and d >= primo_del_mese and (d.year, d.month) in buckets:
            buckets[(d.year, d.month)] += 1

    mesi_abbr = ["gen", "feb", "mar", "apr", "mag", "giu", "lug", "ago", "set", "ott", "nov", "dic"]
    max_n = max(buckets.values()) if buckets else 0
    timeline = [
        {"label": f"{mesi_abbr[m - 1]} {str(y)[2:]}", "n": n,
         "pct": int(round(n / max_n * 100)) if max_n else 0}
        for (y, m), n in buckets.items()
    ]

    distribuzione = [
        {"code": c, "label": cat_labels[c], "n": dist_cat[c]}
        for c, _ in TipoQualifica.CATEGORIA_CHOICES if dist_cat[c]
    ]

    # Top scadenze urgenti (scadute + ≤60gg)
    nomi = _build_nomi_map()
    urgenti = sorted(
        (q for q in quals if q.data_scadenza is not None and q.data_scadenza <= soglia_60),
        key=lambda q: q.data_scadenza,
    )[:15]
    scadenze_urgenti = []
    for q in urgenti:
        stato_code, stato_label = _classifica_scadenza_qualifica(q.data_scadenza, oggi, soglia_30, soglia_60)
        scadenze_urgenti.append({
            "legacy_id": q.legacy_anagrafica_id,
            "dipendente": nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}"),
            "tipo_id": q.tipo_id,
            "tipo_nome": q.tipo.nome,
            "categoria": q.tipo.get_categoria_display(),
            "data_scadenza": q.data_scadenza,
            "giorni": (q.data_scadenza - oggi).days,
            "stato": stato_code,
            "stato_label": stato_label,
        })

    # Prossime sessioni di rilascio/rinnovo collettivo
    prossime_sessioni = list(
        QualificaSessione.objects.select_related("tipo")
        .filter(data_conseguimento__gte=oggi)
        .order_by("data_conseguimento")[:8]
    )

    # MOD.128 — riepilogo processi qualificati (modulo additivo, fail-safe).
    mpq = None
    try:
        from .models_mpq import ProcessoQualificato as _PQ
        _procs = list(_PQ.objects.all())
        _mpq_scaduti = _mpq_scad60 = 0
        for _p in _procs:
            _se = _p.scadenza_effettiva
            if _se is None:
                continue
            if _se < oggi:
                _mpq_scaduti += 1
            elif _se <= soglia_60:
                _mpq_scad60 += 1
        mpq = {
            "n": len(_procs),
            "attivi": sum(1 for _p in _procs if _p.stato == "ATTIVO"),
            "scaduti": _mpq_scaduti,
            "in_scadenza": _mpq_scad60,
        }
    except Exception:
        mpq = None

    return render(request, "anagrafica/pages/qualifiche_dashboard.html", {
        "oggi": oggi,
        "is_admin": is_admin,
        "mpq": mpq,
        "tipi_attivi": tipi_attivi,
        "tot_assegnazioni": len(quals),
        "dipendenti_con_qualifica": len(dipendenti_ids),
        "n_valide": n_valide,
        "n_scadute": n_scadute,
        "n_in_scadenza": n_s30 + n_s60,
        "n_s30": n_s30,
        "n_permanenti": n_permanenti,
        "n_con_evidenza": n_con_evidenza,
        "n_da_verificare": n_da_verificare,
        "distribuzione": distribuzione,
        "timeline": timeline,
        "scadenze_urgenti": scadenze_urgenti,
        "prossime_sessioni": prossime_sessioni,
    })


def _raggruppa_voci_per_tipo(voci, *, tipo_of, scaduta_of, giorni_of):
    """Raggruppa voci di scadenzario per «tipo» (vista a gruppi espandibili).

    Generico: ``tipo_of``/``scaduta_of``/``giorni_of`` estraggono da ogni voce
    il nome-tipo, il flag «scaduta» e i giorni residui. Ordina le voci coi più
    urgenti in cima e i gruppi con scadute per primi. Riusato dagli scadenzari
    qualifiche e skill-matrix (lo scadenzario anagrafica ha il proprio helper).
    """
    from collections import OrderedDict

    buckets = OrderedDict()
    for v in voci:
        buckets.setdefault(tipo_of(v) or "—", []).append(v)
    gruppi = []
    for tipo_nome, gv in buckets.items():
        gv_sorted = sorted(gv, key=lambda x: (not scaduta_of(x), giorni_of(x)))
        n_scadute = sum(1 for x in gv if scaduta_of(x))
        worst = min((giorni_of(x) for x in gv), default=99999)
        gruppi.append({
            "tipo_nome": tipo_nome,
            "voci": gv_sorted,
            "n_totale": len(gv),
            "n_scadute": n_scadute,
            "worst_giorni": worst,
            "has_scadute": n_scadute > 0,
        })
    gruppi.sort(key=lambda g: (not g["has_scadute"], g["worst_giorni"], str(g["tipo_nome"]).lower()))
    return gruppi


@login_required
def qualifiche_scadenzario(request):
    """Scadenzario dedicato alle sole qualifiche/certificazioni.

    È lo ``scadenzario`` unificato ristretto alle qualifiche, ma più ricco:
    filtri per stato / categoria / reparto / tipo ed export CSV. Legge le stesse
    ``DipendenteQualifica`` di tutto il resto: nessun dato duplicato.
    """
    from datetime import timedelta
    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    soglia_30 = oggi + timedelta(days=30)
    soglia_60 = oggi + timedelta(days=60)

    filtro_stato = (request.GET.get("stato") or "").strip()      # scaduta/30/60/valide/tutte/""
    filtro_cat = (request.GET.get("categoria") or "").strip().upper()
    filtro_reparto = (request.GET.get("reparto") or "").strip()
    filtro_tipo = (request.GET.get("tipo") or "").strip()
    export_csv = request.GET.get("format") == "csv"
    valid_cats = {c for c, _ in TipoQualifica.CATEGORIA_CHOICES}
    if filtro_cat not in valid_cats:
        filtro_cat = ""

    dip_rows = fetch_anagrafica_rows(deduplicate=True)
    dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}

    qs = DipendenteQualifica.objects.select_related("tipo")
    if filtro_cat:
        qs = qs.filter(tipo__categoria=filtro_cat)
    if filtro_tipo.isdigit():
        qs = qs.filter(tipo_id=int(filtro_tipo))

    # Filtro stato lato DB
    if filtro_stato == "scaduta":
        qs = qs.filter(data_scadenza__isnull=False, data_scadenza__lt=oggi)
    elif filtro_stato == "30":
        qs = qs.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_30)
    elif filtro_stato == "60":
        qs = qs.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_60)
    elif filtro_stato == "valide":
        qs = qs.filter(Q(data_scadenza__isnull=True) | Q(data_scadenza__gt=soglia_60))
    elif filtro_stato == "tutte":
        pass
    else:  # default: tutto ciò che richiede attenzione (scadute + ≤60gg)
        qs = qs.filter(data_scadenza__isnull=False, data_scadenza__lte=soglia_60)

    qs = qs.order_by("data_scadenza", "tipo__nome")

    nomi = _build_nomi_map()
    voci: list[dict] = []
    counts = {"scaduta": 0, "s30": 0, "s60": 0, "valida": 0, "permanente": 0}
    for q in qs:
        dip = dip_map.get(q.legacy_anagrafica_id, {})
        reparto = str(dip.get("reparto") or "").strip()
        if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
            continue
        stato_code, stato_label = _classifica_scadenza_qualifica(q.data_scadenza, oggi, soglia_30, soglia_60)
        counts[stato_code] = counts.get(stato_code, 0) + 1
        voci.append({
            "legacy_id": q.legacy_anagrafica_id,
            "dipendente": nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}"),
            "reparto": reparto,
            "tipo_nome": q.tipo.nome,
            "tipo_id": q.tipo_id,
            "categoria": q.tipo.get_categoria_display(),
            "numero": q.numero,
            "livello": q.livello,
            "ente": q.ente,
            "data_conseguimento": q.data_conseguimento,
            "data_scadenza": q.data_scadenza,
            "giorni": (q.data_scadenza - oggi).days if q.data_scadenza else None,
            "stato": stato_code,
            "stato_label": stato_label,
            "ha_evidenza": bool(q.documento),
            "verificata": q.verificata,
            "note": q.note,
        })

    if export_csv:
        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="scadenzario_qualifiche.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(resp, delimiter=";")
        writer.writerow(["Dipendente", "Reparto", "Qualifica", "Categoria", "N°", "Livello",
                         "Ente", "Conseguimento", "Scadenza", "Giorni", "Stato",
                         "Evidenza", "Verificata"])
        for v in voci:
            writer.writerow([
                v["dipendente"], v["reparto"], v["tipo_nome"], v["categoria"],
                v["numero"], v["livello"], v["ente"],
                v["data_conseguimento"].strftime("%d/%m/%Y") if v["data_conseguimento"] else "",
                v["data_scadenza"].strftime("%d/%m/%Y") if v["data_scadenza"] else "",
                v["giorni"] if v["giorni"] is not None else "",
                v["stato_label"],
                "Sì" if v["ha_evidenza"] else "No",
                "Sì" if v["verificata"] else "No",
            ])
        return resp

    reparti = sorted({str(r.get("reparto") or "").strip() for r in dip_rows if str(r.get("reparto") or "").strip()})
    tipi_opts = list(TipoQualifica.objects.order_by("categoria", "nome").values("id", "nome"))

    gruppi = _raggruppa_voci_per_tipo(
        voci,
        tipo_of=lambda v: v["tipo_nome"],
        scaduta_of=lambda v: v["stato"] == "scaduta",
        giorni_of=lambda v: v["giorni"] if v["giorni"] is not None else 99999,
    )

    return render(request, "anagrafica/pages/qualifiche_scadenzario.html", {
        "oggi": oggi,
        "voci": voci,
        "gruppi": gruppi,
        "counts": counts,
        "totale": len(voci),
        "filtro_stato": filtro_stato,
        "filtro_categoria": filtro_cat,
        "filtro_reparto": filtro_reparto,
        "filtro_tipo": filtro_tipo,
        "reparti": reparti,
        "tipi_opts": tipi_opts,
        "CATEGORIA_CHOICES": TipoQualifica.CATEGORIA_CHOICES,
    })


@login_required
def tipo_qualifica_detail(request, tipo_id: int):
    """Dettaglio di una singola qualifica/abilitazione: chi la possiede (con
    stato), corsi collegati, sessioni di rinnovo, e la formazione collegata
    (sessioni corso, lezioni, attestati/completamenti) — modello qualifica àncora.
    """
    from datetime import timedelta
    from django.utils import timezone as _tz
    from .models_formazione import TrainingSession, TrainingEmployeeRecord, TrainingLesson

    tipo = get_object_or_404(TipoQualifica, pk=tipo_id)
    is_admin = _qualifiche_can_edit(request)
    oggi = _tz.localdate()
    soglia = oggi + timedelta(days=60)

    corsi = list(tipo.corsi.select_related("piano").order_by("titolo"))
    corso_ids = [c.id for c in corsi]

    nomi = _build_nomi_map()
    holders: list[dict] = []
    n_scaduti = n_scadenza = 0
    for q in (DipendenteQualifica.objects.filter(tipo=tipo)
              .select_related("record_formazione", "record_formazione__corso", "sessione")
              .order_by("data_scadenza", "id")):
        if q.data_scadenza is None:
            stato = "valida"
        elif q.data_scadenza < oggi:
            stato = "scaduta"; n_scaduti += 1
        elif q.data_scadenza <= soglia:
            stato = "in_scadenza"; n_scadenza += 1
        else:
            stato = "valida"
        holders.append({
            "q": q, "stato": stato,
            "nome": nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}"),
        })
    _ord = {"scaduta": 0, "in_scadenza": 1, "valida": 2}
    holders.sort(key=lambda h: (_ord.get(h["stato"], 9), h["nome"].casefold()))

    sessioni = list(
        QualificaSessione.objects.filter(tipo=tipo)
        .annotate(n_part=Count("qualifiche")).order_by("-data_conseguimento", "-id")
    )

    corso_sessioni: list = []
    attestati_recenti: list = []
    n_attestati = n_lezioni = 0
    if corso_ids:
        corso_sessioni = list(
            TrainingSession.objects.filter(corso_id__in=corso_ids)
            .select_related("corso")
            .annotate(n_lezioni=Count("lezioni", distinct=True),
                      n_iscritti=Count("iscrizioni", distinct=True))
            .order_by("-data_inizio")[:30]
        )
        n_lezioni = TrainingLesson.objects.filter(sessione__corso_id__in=corso_ids).count()
        rec_qs = (TrainingEmployeeRecord.objects.filter(corso_id__in=corso_ids)
                  .select_related("corso").order_by("-data_completamento", "-id"))
        n_attestati = rec_qs.count()
        for r in rec_qs[:30]:
            attestati_recenti.append({
                "r": r, "nome": nomi.get(r.legacy_anagrafica_id, f"#{r.legacy_anagrafica_id}"),
            })

    return render(request, "anagrafica/pages/tipo_qualifica_detail.html", {
        "tipo": tipo,
        "is_admin": is_admin,
        "corsi": corsi,
        "holders": holders,
        "n_holders": len(holders),
        "n_scaduti": n_scaduti,
        "n_scadenza": n_scadenza,
        "sessioni": sessioni,
        "corso_sessioni": corso_sessioni,
        "attestati_recenti": attestati_recenti,
        "n_attestati": n_attestati,
        "n_lezioni": n_lezioni,
    })


@login_required
@require_POST
def tipo_qualifica_create(request):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per creare tipi di qualifica.")
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    nome = (request.POST.get("nome") or "").strip()[:150]
    if not nome:
        messages.error(request, "Il nome della qualifica è obbligatorio.")
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    durata_raw = request.POST.get("durata_mesi") or "0"
    try:
        durata_mesi = max(0, int(durata_raw))
    except ValueError:
        durata_mesi = 0

    _, created = TipoQualifica.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "categoria": (request.POST.get("categoria") or TipoQualifica.CAT_ALTRO).strip()[:20],
            "durata_mesi": durata_mesi,
            "descrizione": (request.POST.get("descrizione") or "").strip(),
        },
    )
    if created:
        messages.success(request, f'Tipo qualifica "{nome}" creato.')
    else:
        messages.warning(request, f'Esiste già un tipo qualifica con il nome "{nome}".')
    return _back_to_caller(request, "anagrafica:qualifiche_list")


@login_required
@require_POST
def tipo_qualifica_edit(request, tipo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per modificare tipi di qualifica.")
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    tipo = get_object_or_404(TipoQualifica, pk=tipo_id)
    nome = (request.POST.get("nome") or "").strip()[:150]
    if not nome:
        messages.error(request, "Il nome della qualifica è obbligatorio.")
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    durata_raw = request.POST.get("durata_mesi") or "0"
    try:
        durata_mesi = max(0, int(durata_raw))
    except ValueError:
        durata_mesi = 0

    tipo.nome = nome
    tipo.categoria = (request.POST.get("categoria") or TipoQualifica.CAT_ALTRO).strip()[:20]
    tipo.durata_mesi = durata_mesi
    tipo.descrizione = (request.POST.get("descrizione") or "").strip()
    tipo.is_active = request.POST.get("is_active") == "1"
    tipo.save()
    messages.success(request, f'Tipo qualifica "{tipo.nome}" aggiornato.')
    return _back_to_caller(request, "anagrafica:qualifiche_list")


@login_required
@require_POST
def tipo_qualifica_delete(request, tipo_id: int):
    legacy_user = get_legacy_user(request.user)
    if not (request.user.is_superuser or is_legacy_admin(legacy_user)):
        messages.error(request, "Non hai i permessi per eliminare tipi di qualifica.")
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    tipo = get_object_or_404(TipoQualifica, pk=tipo_id)
    n_uso = tipo.assegnazioni.count()
    if n_uso:
        if tipo.is_active:
            tipo.is_active = False
            tipo.save(update_fields=["is_active"])
            messages.warning(request, f'"{tipo.nome}" è in uso ({n_uso} assegnazioni): disattivato (non eliminato) per preservare lo storico. Puoi riattivarlo dalla modifica.')
        else:
            messages.error(request, f'"{tipo.nome}" ha {n_uso} assegnazioni: non eliminabile (già disattivo).')
        return _back_to_caller(request, "anagrafica:qualifiche_list")

    nome = tipo.nome
    tipo.delete()
    messages.success(request, f'Tipo qualifica "{nome}" eliminato.')
    return _back_to_caller(request, "anagrafica:qualifiche_list")


# ---------------------------------------------------------------------------
# Sessioni di rinnovo qualifica — rilascio/rinnovo collettivo "a sessioni"
# (speculare alle sessioni corsi; pattern batch come le visite mediche)
# ---------------------------------------------------------------------------

def _qualifiche_can_edit(request) -> bool:
    return request.user.is_superuser or is_legacy_admin(get_legacy_user(request.user))


def _build_candidati_qualifica(tipo, oggi) -> list[dict]:
    """Dipendenti (in forza) che detengono già la qualifica, con lo stato di
    rinnovo: scaduta / in scadenza (≤90gg) / valida. I nuovi rilasci si
    aggiungono dal picker. Pre-seleziona scadute e in scadenza."""
    from datetime import timedelta
    soglia = oggi + timedelta(days=90)
    ultima_per_id: dict[int, "DipendenteQualifica"] = {}
    for q in (DipendenteQualifica.objects.filter(tipo=tipo)
              .order_by("legacy_anagrafica_id", "-data_conseguimento", "-id")):
        ultima_per_id.setdefault(q.legacy_anagrafica_id, q)
    nomi = _build_nomi_map()
    cessati = _cessati_legacy_ids()
    out: list[dict] = []
    for lid, q in ultima_per_id.items():
        if lid in cessati:
            continue
        if q.data_scadenza is None:
            status = "valida"
        elif q.data_scadenza < oggi:
            status = "scaduta"
        elif q.data_scadenza <= soglia:
            status = "in_scadenza"
        else:
            status = "valida"
        out.append({
            "legacy_id": lid, "nome": nomi.get(lid, f"#{lid}"),
            "ultima": q, "status": status,
            "preselect": status in ("scaduta", "in_scadenza"),
        })
    order = {"scaduta": 0, "in_scadenza": 1, "valida": 2}
    out.sort(key=lambda c: (order.get(c["status"], 9), c["nome"].casefold()))
    return out


@login_required
def qualifica_sessioni_list(request):
    qs = QualificaSessione.objects.select_related("tipo").annotate(n_part=Count("qualifiche"))
    filtro_tipo = (request.GET.get("tipo") or "").strip()
    q_text = (request.GET.get("q") or "").strip()
    if filtro_tipo.isdigit():
        qs = qs.filter(tipo_id=int(filtro_tipo))
    if q_text:
        qs = qs.filter(Q(tipo__nome__icontains=q_text) | Q(ente__icontains=q_text))
    sessioni = list(qs.order_by("-data_conseguimento", "-id"))
    # Tipi con almeno una sessione, per il filtro a tendina.
    tipi_con_sessioni = list(
        TipoQualifica.objects.filter(sessioni__isnull=False).distinct().order_by("nome")
    )
    return render(request, "anagrafica/pages/qualifica_sessioni_list.html", {
        "sessioni": sessioni,
        "is_admin": _qualifiche_can_edit(request),
        "tipi_con_sessioni": tipi_con_sessioni,
        "filtro_tipo": filtro_tipo,
        "q_text": q_text,
    })


@login_required
def qualifica_sessione_create(request):
    if not _qualifiche_can_edit(request):
        messages.error(request, "Non hai i permessi per creare sessioni di rinnovo.")
        return redirect("anagrafica:qualifiche_list")

    from datetime import date as _date
    from django.utils import timezone as _tz
    from django.db import transaction

    oggi = _tz.localdate()
    tipi = list(TipoQualifica.objects.filter(is_active=True).order_by("categoria", "nome"))

    # ---- Step 2: salva la sessione e i record dei partecipanti ----------
    if request.method == "POST" and request.POST.get("step") == "2":
        try:
            tipo = TipoQualifica.objects.get(pk=request.POST.get("tipo_id", "").strip(), is_active=True)
        except (TipoQualifica.DoesNotExist, ValueError):
            messages.error(request, "Tipo qualifica non valido.")
            return redirect("anagrafica:qualifica_sessione_create")
        try:
            data_cons = _date.fromisoformat(request.POST.get("data_conseguimento", "").strip())
        except (ValueError, TypeError):
            messages.error(request, "Data conseguimento non valida.")
            return redirect("anagrafica:qualifica_sessione_create")
        data_scad = None
        ds_raw = (request.POST.get("data_scadenza") or "").strip()
        if ds_raw:
            try:
                data_scad = _date.fromisoformat(ds_raw)
            except (ValueError, TypeError):
                data_scad = None
        ente = (request.POST.get("ente") or "").strip()[:200]
        note = (request.POST.get("note") or "").strip()

        ids: list[int] = []
        seen: set[int] = set()
        raw = list(request.POST.getlist("dipendenti_selezionati"))
        raw += [x for x in (request.POST.get("extra_ids") or "").split(",")]
        for s in raw:
            s = str(s).strip()
            if s.isdigit():
                lid = int(s)
                if lid > 0 and lid not in seen:
                    seen.add(lid)
                    ids.append(lid)
        if not ids:
            messages.warning(request, "Nessun dipendente selezionato.")
            return redirect("anagrafica:qualifica_sessione_create")

        with transaction.atomic():
            sess = QualificaSessione.objects.create(
                tipo=tipo, data_conseguimento=data_cons, data_scadenza=data_scad,
                ente=ente, note=note, created_by=request.user,
            )
            scad_eff = sess.scadenza_effettiva
            for lid in ids:
                _upsert_dipendente_qualifica(
                    lid, tipo, data_cons, scad_eff, user=request.user, sessione=sess,
                )
        messages.success(
            request,
            f'Sessione «{tipo.nome}» del {data_cons:%d/%m/%Y}: {len(ids)} qualifiche registrate/rinnovate.',
        )
        return redirect("anagrafica:qualifica_sessione_detail", sessione_id=sess.id)

    # ---- Pagina unica (GET) — form + candidati caricati dinamicamente ----
    # Se arriva ?tipo=<id> (da scadenzario / dettaglio qualifica / matrice) la
    # tabella candidati è renderizzata subito lato server (deep-link no-JS);
    # altrimenti si popola via HTMX al cambio del tipo nel select.
    pre_tipo = None
    pre_tipo_id = (request.GET.get("tipo") or "").strip()
    if pre_tipo_id.isdigit():
        pre_tipo = TipoQualifica.objects.filter(pk=int(pre_tipo_id), is_active=True).first()
    candidati = _build_candidati_qualifica(pre_tipo, oggi) if pre_tipo else []
    n_pre = sum(1 for c in candidati if c["preselect"])
    return render(request, "anagrafica/pages/qualifica_sessione_create.html", {
        "tipi": tipi, "oggi": oggi,
        "pre_tipo": pre_tipo, "candidati": candidati, "n_pre": n_pre,
        "dipendenti_picker": _dipendenti_picker_rows(),
    })


@login_required
def qualifica_sessione_candidati(request):
    """Partial HTMX: tabella dei candidati (chi detiene già la qualifica, con
    stato di rinnovo) per il tipo selezionato. Popola dinamicamente la pagina di
    creazione sessione senza ricaricarla."""
    if not _qualifiche_can_edit(request):
        return HttpResponse(status=403)
    from django.utils import timezone as _tz
    raw = (request.GET.get("tipo") or request.GET.get("tipo_id") or "").strip()
    tipo = None
    if raw.isdigit():
        tipo = TipoQualifica.objects.filter(pk=int(raw), is_active=True).first()
    candidati = _build_candidati_qualifica(tipo, _tz.localdate()) if tipo else []
    n_pre = sum(1 for c in candidati if c["preselect"])
    return render(request, "anagrafica/pages/_qualifica_candidati.html", {
        "tipo": tipo, "candidati": candidati, "n_pre": n_pre,
    })


@login_required
def qualifica_sessione_detail(request, sessione_id: int):
    sess = get_object_or_404(QualificaSessione.objects.select_related("tipo"), pk=sessione_id)
    nomi = _build_nomi_map()
    rows = [
        {"q": q, "nome": nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}")}
        for q in sess.qualifiche.all()
    ]
    rows.sort(key=lambda r: r["nome"].casefold())
    is_admin = _qualifiche_can_edit(request)
    return render(request, "anagrafica/pages/qualifica_sessione_detail.html", {
        "sess": sess,
        "rows": rows,
        "is_admin": is_admin,
        "dipendenti_picker": _dipendenti_picker_rows() if is_admin else [],
    })


@login_required
@require_POST
def qualifica_sessione_partecipante_add(request, sessione_id: int):
    if not _qualifiche_can_edit(request):
        messages.error(request, "Permessi insufficienti.")
        return redirect("anagrafica:qualifica_sessione_detail", sessione_id=sessione_id)
    sess = get_object_or_404(QualificaSessione.objects.select_related("tipo"), pk=sessione_id)
    try:
        lid = int(request.POST.get("legacy_id") or 0)
    except (ValueError, TypeError):
        lid = 0
    if lid > 0:
        _upsert_dipendente_qualifica(
            lid, sess.tipo, sess.data_conseguimento, sess.scadenza_effettiva,
            user=request.user, sessione=sess,
        )
        messages.success(request, "Partecipante aggiunto alla sessione.")
    else:
        messages.error(request, "Dipendente non valido.")
    return redirect("anagrafica:qualifica_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def qualifica_sessione_partecipante_remove(request, sessione_id: int, q_id: int):
    if not _qualifiche_can_edit(request):
        messages.error(request, "Permessi insufficienti.")
        return redirect("anagrafica:qualifica_sessione_detail", sessione_id=sessione_id)
    q = get_object_or_404(DipendenteQualifica, pk=q_id, sessione_id=sessione_id)
    # Stacca dalla sessione (la qualifica corrente del dipendente resta).
    q.sessione = None
    q.save(update_fields=["sessione"])
    messages.success(request, "Partecipante rimosso dalla sessione (qualifica conservata).")
    return redirect("anagrafica:qualifica_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def qualifica_sessione_delete(request, sessione_id: int):
    if not _qualifiche_can_edit(request):
        messages.error(request, "Permessi insufficienti.")
        return redirect("anagrafica:qualifica_sessioni_list")
    sess = get_object_or_404(QualificaSessione, pk=sessione_id)
    # SET_NULL: le qualifiche dei dipendenti restano, perdono solo il legame.
    sess.delete()
    messages.success(request, "Sessione eliminata (qualifiche dei dipendenti conservate).")
    return redirect("anagrafica:qualifica_sessioni_list")


@login_required
def qualifica_sessione_report_csv(request, sessione_id: int):
    """Esporta in CSV i partecipanti di una sessione di qualifica/abilitazione."""
    sess = get_object_or_404(QualificaSessione.objects.select_related("tipo"), pk=sessione_id)
    nomi = _build_nomi_map()
    rows = sorted(
        sess.qualifiche.all(),
        key=lambda q: nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}").casefold(),
    )

    import csv
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    fname = f"sessione_qualifica_{sess.tipo.nome}_{sess.data_conseguimento:%Y%m%d}.csv".replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    response.write("﻿")
    writer = safe_csv_writer(response, delimiter=";")
    writer.writerow(["Qualifica", "Ente", "Conseguimento", "Scadenza sessione"])
    writer.writerow([
        sess.tipo.nome, sess.ente or "",
        sess.data_conseguimento.strftime("%d-%m-%Y") if sess.data_conseguimento else "",
        sess.scadenza_effettiva.strftime("%d-%m-%Y") if sess.scadenza_effettiva else "",
    ])
    writer.writerow([])
    writer.writerow(["Dipendente", "ID anagrafica", "Conseguimento", "Scadenza"])
    for q in rows:
        writer.writerow([
            nomi.get(q.legacy_anagrafica_id, f"#{q.legacy_anagrafica_id}"),
            q.legacy_anagrafica_id,
            q.data_conseguimento.strftime("%d-%m-%Y") if q.data_conseguimento else "",
            q.data_scadenza.strftime("%d-%m-%Y") if q.data_scadenza else "",
        ])
    return response


# ---------------------------------------------------------------------------
# Helper: redirect verso il pannello impostazioni con tab attivo
# ---------------------------------------------------------------------------

def _redirect_impostazioni(tab: str | None = None):
    url = reverse("anagrafica:impostazioni")
    if tab:
        url = f"{url}?tab={tab}#tab-{tab}"
    return HttpResponseRedirect(url)


def _back_to_caller(request, fallback_view_name: str):
    """
    Redirect intelligente per le view CRUD dei cataloghi: se il form proviene dal
    pannello impostazioni (campo nascosto `next_tab`) torna lì, altrimenti
    redirige alla pagina di lista standalone.
    """
    next_tab = (request.POST.get("next_tab") or "").strip()
    if next_tab:
        return _redirect_impostazioni(next_tab)
    return redirect(fallback_view_name)


def _impostazioni_admin_check(request, tab: str | None = None):
    """Restituisce (is_admin, redirect_response_or_None)."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Permessi insufficienti.")
        return False, _redirect_impostazioni(tab)
    return True, None


def _workflow_campo_payload(request, existing: OnboardingOffboardingCampo | None = None) -> tuple[dict | None, str | None]:
    field_map = _dipendente_workflow_field_map()
    campo_key = (request.POST.get("campo_key") or getattr(existing, "campo_key", "") or "").strip()
    meta = field_map.get(campo_key)
    if not meta:
        return None, "Campo del form nuovo dipendente non valido."

    fase = (request.POST.get("fase") or getattr(existing, "fase", "") or "").strip()
    valid_fasi = {value for value, _label in OnboardingOffboardingCampo.FASE_CHOICES}
    if fase not in valid_fasi:
        return None, "Fase onboarding/offboarding non valida."

    categoria = (request.POST.get("categoria") or getattr(existing, "categoria", "") or "").strip()
    valid_categorie = {value for value, _label in OnboardingOffboardingCampo.CATEGORIA_CHOICES}
    if categoria not in valid_categorie:
        categoria = OnboardingOffboardingCampo.CATEGORIA_HR

    label = (request.POST.get("campo_label") or meta["label"]).strip()[:160]
    if not label:
        label = meta["label"][:160]
    try:
        ordine = int((request.POST.get("ordine") or getattr(existing, "ordine", 50) or 50))
    except (TypeError, ValueError):
        ordine = 50
    ordine = max(0, min(ordine, 9999))

    return {
        "fase": fase,
        "campo_key": campo_key,
        "campo_label": label,
        "sezione": meta["section"][:120],
        "categoria": categoria,
        "obbligatorio": request.POST.get("obbligatorio") == "1",
        "is_active": request.POST.get("is_active", "1") == "1",
        "ordine": ordine,
        "note": (request.POST.get("note") or "").strip()[:1000],
        "updated_by": request.user,
    }, None


@login_required
@require_POST
def workflow_campo_create(request):
    ok, resp = _impostazioni_admin_check(request, "workflow")
    if not ok:
        return resp

    payload, error = _workflow_campo_payload(request)
    if error:
        messages.error(request, error)
        return _redirect_impostazioni("workflow")

    try:
        obj, created = OnboardingOffboardingCampo.objects.get_or_create(
            fase=payload["fase"],
            campo_key=payload["campo_key"],
            defaults=payload,
        )
        if created:
            messages.success(request, "Campo aggiunto alla lista onboarding/offboarding.")
        else:
            for key, value in payload.items():
                setattr(obj, key, value)
            obj.save()
            messages.info(request, "Associazione gia presente: aggiornata con i nuovi valori.")
    except IntegrityError:
        messages.error(request, "Esiste gia una associazione per questo campo e questa fase.")
    return _redirect_impostazioni("workflow")


@login_required
@require_POST
def workflow_campo_update(request, campo_id: int):
    ok, resp = _impostazioni_admin_check(request, "workflow")
    if not ok:
        return resp

    obj = get_object_or_404(OnboardingOffboardingCampo, pk=campo_id)
    payload, error = _workflow_campo_payload(request, existing=obj)
    if error:
        messages.error(request, error)
        return _redirect_impostazioni("workflow")

    try:
        for key, value in payload.items():
            setattr(obj, key, value)
        obj.save()
        messages.success(request, "Associazione onboarding/offboarding aggiornata.")
    except IntegrityError:
        messages.error(request, "Esiste gia una associazione per questo campo e questa fase.")
    return _redirect_impostazioni("workflow")


@login_required
@require_POST
def workflow_campo_delete(request, campo_id: int):
    ok, resp = _impostazioni_admin_check(request, "workflow")
    if not ok:
        return resp
    obj = get_object_or_404(OnboardingOffboardingCampo, pk=campo_id)
    label = obj.campo_label
    obj.delete()
    messages.success(request, f'Associazione "{label}" eliminata.')
    return _redirect_impostazioni("workflow")


# ---------------------------------------------------------------------------
# Livelli contrattuali — catalogo (CRUD)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def livello_contrattuale_create(request):
    ok, resp = _impostazioni_admin_check(request, "livelli")
    if not ok:
        return resp

    codice = (request.POST.get("codice") or "").strip().upper()[:10]
    if not codice:
        messages.error(request, "Il codice del livello è obbligatorio.")
        return _redirect_impostazioni("livelli")

    descrizione = (request.POST.get("descrizione") or "").strip()[:200]
    try:
        ordine = max(0, int(request.POST.get("ordine") or "0"))
    except ValueError:
        ordine = 0

    _, created = LivelloContrattuale.objects.get_or_create(
        codice=codice,
        defaults={"descrizione": descrizione, "ordine": ordine},
    )
    if created:
        messages.success(request, f'Livello "{codice}" creato.')
    else:
        messages.warning(request, f'Esiste già un livello con codice "{codice}".')
    return _redirect_impostazioni("livelli")


@login_required
@require_POST
def livello_contrattuale_edit(request, livello_id: int):
    ok, resp = _impostazioni_admin_check(request, "livelli")
    if not ok:
        return resp

    livello = get_object_or_404(LivelloContrattuale, pk=livello_id)
    codice = (request.POST.get("codice") or "").strip().upper()[:10]
    if not codice:
        messages.error(request, "Il codice del livello è obbligatorio.")
        return _redirect_impostazioni("livelli")

    try:
        ordine = max(0, int(request.POST.get("ordine") or "0"))
    except ValueError:
        ordine = 0

    livello.codice = codice
    livello.descrizione = (request.POST.get("descrizione") or "").strip()[:200]
    livello.ordine = ordine
    livello.is_active = request.POST.get("is_active") == "1"
    livello.save()
    messages.success(request, f'Livello "{livello.codice}" aggiornato.')
    return _redirect_impostazioni("livelli")


@login_required
@require_POST
def livello_contrattuale_delete(request, livello_id: int):
    ok, resp = _impostazioni_admin_check(request, "livelli")
    if not ok:
        return resp

    livello = get_object_or_404(LivelloContrattuale, pk=livello_id)
    codice = livello.codice
    # Controllo se esistono StoricoContratto con questo codice
    if StoricoContratto.objects.filter(codice_livello=codice).exists():
        messages.error(
            request,
            f'"{codice}" è referenziato nello storico contrattuale — non eliminabile. Disattivalo.',
        )
        return _redirect_impostazioni("livelli")

    livello.delete()
    messages.success(request, f'Livello "{codice}" eliminato.')
    return _redirect_impostazioni("livelli")


# ---------------------------------------------------------------------------
# Tipologie contratto — catalogo (CRUD)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def tipologia_contratto_create(request):
    ok, resp = _impostazioni_admin_check(request, "tipologie")
    if not ok:
        return resp

    codice = (request.POST.get("codice") or "").strip().upper()[:20]
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not codice or not nome:
        messages.error(request, "Codice e nome della tipologia sono obbligatori.")
        return _redirect_impostazioni("tipologie")

    try:
        ordine = max(0, int(request.POST.get("ordine") or "0"))
    except ValueError:
        ordine = 0

    _, created = TipologiaContratto.objects.get_or_create(
        codice=codice,
        defaults={"nome": nome, "ordine": ordine},
    )
    if created:
        messages.success(request, f'Tipologia "{codice}" creata.')
    else:
        messages.warning(request, f'Esiste già una tipologia con codice "{codice}".')
    return _redirect_impostazioni("tipologie")


@login_required
@require_POST
def tipologia_contratto_edit(request, tipologia_id: int):
    ok, resp = _impostazioni_admin_check(request, "tipologie")
    if not ok:
        return resp

    tipologia = get_object_or_404(TipologiaContratto, pk=tipologia_id)
    codice = (request.POST.get("codice") or "").strip().upper()[:20]
    nome = (request.POST.get("nome") or "").strip()[:100]
    if not codice or not nome:
        messages.error(request, "Codice e nome della tipologia sono obbligatori.")
        return _redirect_impostazioni("tipologie")

    try:
        ordine = max(0, int(request.POST.get("ordine") or "0"))
    except ValueError:
        ordine = 0

    tipologia.codice = codice
    tipologia.nome = nome
    tipologia.ordine = ordine
    tipologia.is_active = request.POST.get("is_active") == "1"
    tipologia.save()
    messages.success(request, f'Tipologia "{tipologia.codice}" aggiornata.')
    return _redirect_impostazioni("tipologie")


@login_required
@require_POST
def tipologia_contratto_delete(request, tipologia_id: int):
    ok, resp = _impostazioni_admin_check(request, "tipologie")
    if not ok:
        return resp

    tipologia = get_object_or_404(TipologiaContratto, pk=tipologia_id)
    codice = tipologia.codice
    if StoricoContratto.objects.filter(tipologia_contratto=codice).exists():
        messages.error(
            request,
            f'"{codice}" è referenziata nello storico contrattuale — non eliminabile. Disattivala.',
        )
        return _redirect_impostazioni("tipologie")

    tipologia.delete()
    messages.success(request, f'Tipologia "{codice}" eliminata.')
    return _redirect_impostazioni("tipologie")


# ---------------------------------------------------------------------------
# Pannello impostazioni anagrafica — vista aggregata con tabs
# ---------------------------------------------------------------------------
# Scadenzario unificato qualifiche + visite mediche
# ---------------------------------------------------------------------------

def _build_scadenzario_voci(
    request,
    *,
    filtro_tipo: str = "",
    filtro_stato: str = "",
    filtro_reparto: str = "",
    dip_map: dict | None = None,
) -> list[dict]:
    """Voci dello scadenzario unificato, ordinate per urgenza.

    Fonte unica per la pagina scadenzario e per il blocco «Cose da gestire» della
    dashboard: se i due contassero per conto proprio, il numero sulla dashboard e
    la lista che quel numero apre potrebbero contraddirsi.

    Ogni sorgente entra solo se il permesso lo consente (visite mediche = dato
    sanitario, formazione e contratti = dato HR), quindi il chiamante non deve
    rifare il gating. ``dip_map`` evita un secondo fetch legacy quando il
    chiamante ha già le righe anagrafica.
    """
    from django.utils import timezone as tz
    oggi = tz.localdate()
    soglia_30 = oggi + _timedelta(days=30)
    soglia_60 = oggi + _timedelta(days=60)

    # Gli ex dipendenti (rapporto cessato) non devono comparire nello scadenzario
    # in NESSUNA sorgente: il filtro va applicato a monte, non solo ai contratti.
    cessati = _cessati_legacy_ids()

    can_view_visite = _can_view_visite_mediche(request)
    can_view_formazione = _can_view_formazione(request)
    can_view_contratti = _check_hr_permission(request)

    if dip_map is None:
        dip_rows = fetch_anagrafica_rows(deduplicate=True)
        dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}

    voci: list[dict] = []

    # ── Qualifiche ──────────────────────────────────────────────────────────
    if filtro_tipo in ("", "qualifica"):
        qs_q = DipendenteQualifica.objects.select_related("tipo").filter(
            data_scadenza__isnull=False
        ).exclude(legacy_anagrafica_id__in=cessati)
        if filtro_stato == "scaduta":
            qs_q = qs_q.filter(data_scadenza__lt=oggi)
        elif filtro_stato == "30":
            qs_q = qs_q.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_30)
        elif filtro_stato == "60":
            qs_q = qs_q.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_60)
        else:
            qs_q = qs_q.filter(data_scadenza__lte=soglia_60)

        for q in qs_q:
            dip = dip_map.get(q.legacy_anagrafica_id, {})
            reparto = str(dip.get("reparto") or "").strip()
            if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
                continue
            delta = (q.data_scadenza - oggi).days
            voci.append({
                "kind":         "qualifica",
                "kind_label":   "Qualifica",
                "legacy_id":    q.legacy_anagrafica_id,
                "cognome":      str(dip.get("cognome") or f"ID {q.legacy_anagrafica_id}").strip(),
                "nome":         str(dip.get("nome") or "").strip(),
                "reparto":      reparto,
                "tipo_nome":    q.tipo.nome,
                "tipo_id":      q.tipo_id,
                "categoria":    q.tipo.get_categoria_display(),
                "data_scadenza": q.data_scadenza,
                "giorni":       delta,
                "scaduta":      delta < 0,
            })

    # ── Visite mediche (gated) ───────────────────────────────────────────────
    if can_view_visite and filtro_tipo in ("", "visita"):
        qs_v = VisitaMedica.objects.select_related("tipo").filter(
            id__in=ultime_visite_correnti_ids(), data_scadenza__isnull=False
        ).exclude(legacy_anagrafica_id__in=cessati)
        if filtro_stato == "scaduta":
            qs_v = qs_v.filter(data_scadenza__lt=oggi)
        elif filtro_stato == "30":
            qs_v = qs_v.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_30)
        elif filtro_stato == "60":
            qs_v = qs_v.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_60)
        else:
            qs_v = qs_v.filter(data_scadenza__lte=soglia_60)

        for v in qs_v:
            dip = dip_map.get(v.legacy_anagrafica_id, {})
            reparto = str(dip.get("reparto") or "").strip()
            if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
                continue
            delta = (v.data_scadenza - oggi).days
            voci.append({
                "kind":         "visita",
                "kind_label":   "Visita medica",
                "legacy_id":    v.legacy_anagrafica_id,
                "cognome":      str(dip.get("cognome") or f"ID {v.legacy_anagrafica_id}").strip(),
                "nome":         str(dip.get("nome") or "").strip(),
                "reparto":      reparto,
                "tipo_nome":    v.tipo.nome,
                "tipo_id":      v.tipo_id,
                "categoria":    "Visita medica",
                "data_scadenza": v.data_scadenza,
                "giorni":       delta,
                "scaduta":      delta < 0,
            })

    # ── Formazione: corsi obbligatori da TrainingDeadline (gated) ───────────
    if can_view_formazione and filtro_tipo in ("", "formazione"):
        qs_f = (
            TrainingDeadline.objects.select_related("corso")
            .filter(is_required=True, data_scadenza__isnull=False)
            .exclude(legacy_anagrafica_id__in=cessati)
        )
        if filtro_stato == "scaduta":
            qs_f = qs_f.filter(data_scadenza__lt=oggi)
        elif filtro_stato == "30":
            qs_f = qs_f.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_30)
        elif filtro_stato == "60":
            qs_f = qs_f.filter(data_scadenza__gte=oggi, data_scadenza__lte=soglia_60)
        else:
            qs_f = qs_f.filter(data_scadenza__lte=soglia_60)

        for d in qs_f:
            dip = dip_map.get(d.legacy_anagrafica_id, {})
            reparto = str(dip.get("reparto") or "").strip()
            if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
                continue
            delta = (d.data_scadenza - oggi).days
            voci.append({
                "kind":         "formazione",
                "kind_label":   "Formazione",
                "legacy_id":    d.legacy_anagrafica_id,
                "cognome":      str(dip.get("cognome") or f"ID {d.legacy_anagrafica_id}").strip(),
                "nome":         str(dip.get("nome") or "").strip(),
                "reparto":      reparto,
                "tipo_nome":    d.corso.titolo,
                "corso_id":     d.corso_id,
                "categoria":    "Corso obbligatorio",
                "data_scadenza": d.data_scadenza,
                "giorni":       delta,
                "scaduta":      delta < 0,
            })

    # ── Contratti a termine e periodi di prova (gated: dato HR) ─────────────
    if can_view_contratti and filtro_tipo in ("", "contratto"):
        # `cessati` è già calcolato a monte (usato da tutte le sorgenti).
        # Solo l'ultimo contratto per dipendente: i precedenti sono storia chiusa.
        ultimo_contratto: dict[int, StoricoContratto] = {}
        for c in (
            StoricoContratto.objects
            .filter(legacy_anagrafica_id__isnull=False)
            .order_by("legacy_anagrafica_id", "-data_inizio", "-created_at")
        ):
            ultimo_contratto.setdefault(c.legacy_anagrafica_id, c)

        scadenze_contratti: list[tuple[int, date, str]] = []
        for legacy_id, c in ultimo_contratto.items():
            if legacy_id in cessati or c.data_fine is None:
                continue
            tip = c.tipologia_contratto or "a termine"
            scadenze_contratti.append((legacy_id, c.data_fine, f"Contratto {tip}"))
        # Periodi di prova: solo futuri — una prova già conclusa non è un'azione da fare.
        prova_rows = DipendenteAnagraficaAziendale.objects.filter(
            data_cessazione__isnull=True,
            prova_data_fine__isnull=False,
            prova_data_fine__gte=oggi,
        ).values_list("legacy_anagrafica_id", "prova_data_fine")
        for legacy_id, fine_prova in prova_rows:
            scadenze_contratti.append((legacy_id, fine_prova, "Fine periodo di prova"))

        for legacy_id, data_fine, descrizione in scadenze_contratti:
            if filtro_stato == "scaduta":
                if data_fine >= oggi:
                    continue
            elif filtro_stato == "30":
                if not (oggi <= data_fine <= soglia_30):
                    continue
            elif filtro_stato == "60":
                if not (oggi <= data_fine <= soglia_60):
                    continue
            elif data_fine > soglia_60:
                continue
            dip = dip_map.get(legacy_id, {})
            reparto = str(dip.get("reparto") or "").strip()
            if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
                continue
            delta = (data_fine - oggi).days
            voci.append({
                "kind":         "contratto",
                "kind_label":   "Contratto",
                "legacy_id":    legacy_id,
                "cognome":      str(dip.get("cognome") or f"ID {legacy_id}").strip(),
                "nome":         str(dip.get("nome") or "").strip(),
                "reparto":      reparto,
                "tipo_nome":    descrizione,
                "categoria":    "Contratto",
                "data_scadenza": data_fine,
                "giorni":       delta,
                "scaduta":      delta < 0,
            })

    # Ordina per urgenza: prima le più scadute (giorni più negativi), poi le più vicine
    voci.sort(key=lambda x: x["giorni"])
    return voci


def _raggruppa_scadenze_per_tipo(voci: list[dict]) -> list[dict]:
    """Raggruppa le voci scadenzario per (kind, tipo_nome) — per la vista a
    gruppi espandibili. Ogni gruppo espone i conteggi e le voci ordinate coi
    più urgenti in cima; i gruppi sono ordinati con gli scaduti/urgenti prima.
    """
    from collections import OrderedDict

    buckets: "OrderedDict[tuple, list[dict]]" = OrderedDict()
    for v in voci:
        buckets.setdefault((v.get("kind"), v.get("tipo_nome") or "—"), []).append(v)

    gruppi: list[dict] = []
    for (kind, tipo_nome), gv in buckets.items():
        gv_sorted = sorted(gv, key=lambda x: (not x.get("scaduta"), x.get("giorni", 9999)))
        n_scadute = sum(1 for x in gv if x.get("scaduta"))
        worst = min((x.get("giorni", 9999) for x in gv), default=9999)
        gruppi.append({
            "kind": kind,
            "kind_label": gv[0].get("kind_label", ""),
            "tipo_nome": tipo_nome,
            "tipo_id": gv[0].get("tipo_id"),
            "voci": gv_sorted,
            "n_totale": len(gv),
            "n_scadute": n_scadute,
            "worst_giorni": worst,
            "has_scadute": n_scadute > 0,
        })
    gruppi.sort(key=lambda g: (not g["has_scadute"], g["worst_giorni"], g["tipo_nome"].lower()))
    return gruppi


@login_required
def scadenzario(request):
    """Scadenzario unificato: qualifiche, visite mediche, formazione obbligatoria
    e contratti/periodi di prova in scadenza o scaduti.

    Accesso: login obbligatorio. Le voci — e con esse il gating per sorgente —
    arrivano da ``_build_scadenzario_voci``, condiviso con la dashboard.
    """
    from django.utils import timezone as tz
    oggi = tz.localdate()

    can_view_visite = _can_view_visite_mediche(request)
    can_view_formazione = _can_view_formazione(request)
    can_view_contratti = _check_hr_permission(request)

    filtro_tipo    = request.GET.get("tipo", "")         # "qualifica" / "visita" / "formazione" / "contratto" / ""
    filtro_stato   = request.GET.get("stato", "")        # "scaduta" / "30" / "60" / ""
    filtro_reparto = request.GET.get("reparto", "").strip()
    export_csv     = request.GET.get("format") == "csv"

    voci = _build_scadenzario_voci(
        request,
        filtro_tipo=filtro_tipo,
        filtro_stato=filtro_stato,
        filtro_reparto=filtro_reparto,
    )

    # KPI
    n_scadute = sum(1 for v in voci if v["scaduta"])
    n_30gg    = sum(1 for v in voci if not v["scaduta"] and v["giorni"] <= 30)
    n_60gg    = sum(1 for v in voci if not v["scaduta"] and 30 < v["giorni"] <= 60)

    # Raggruppamento per TIPO (con espansione dei dipendenti): un gruppo per
    # (kind, tipo_nome), ordinato con i più urgenti (scaduti / meno giorni) in alto.
    gruppi = _raggruppa_scadenze_per_tipo(voci)

    reparti = sorted({v["reparto"] for v in voci if v["reparto"]})

    # Export CSV
    if export_csv:
        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="scadenzario_anagrafica.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(resp, delimiter=";")
        writer.writerow(["Dipendente", "Reparto", "Tipo entità", "Descrizione", "Scadenza", "Stato"])
        for v in voci:
            stato = "Scaduta" if v["scaduta"] else f"Scade in {v['giorni']} giorni"
            writer.writerow([
                f"{v['cognome']} {v['nome']}".strip(),
                v["reparto"],
                v["kind_label"],
                v["tipo_nome"],
                v["data_scadenza"].strftime("%d-%m-%Y") if v["data_scadenza"] else "",
                stato,
            ])
        return resp

    # Paginazione
    paginator  = Paginator(voci, 50)
    page_obj   = paginator.get_page(request.GET.get("page"))

    # Formazione: riepilogo scadenze urgenti da mostrare come sezione aggiuntiva
    fm_n_scaduti = 0
    fm_n_30gg    = 0
    fm_n_90gg    = 0
    fm_is_cache_empty = True
    if can_view_formazione:
        try:
            fm_n_scaduti = TrainingDeadline.objects.filter(stato_scadenza="SCADUTO").count()
            fm_n_30gg    = TrainingDeadline.objects.filter(stato_scadenza="IN_SCADENZA_30").count()
            fm_n_90gg    = TrainingDeadline.objects.filter(stato_scadenza="IN_SCADENZA_90").count()
            fm_is_cache_empty = not TrainingDeadline.objects.exists()
        except Exception:
            logger.exception("Errore caricamento KPI formazione per scadenzario")

    # Toggle di vista: gruppi (default), calendario (griglia mensile),
    # affiancata (due colonne Visite | Formazione). Valore ignoto → gruppi.
    layout = request.GET.get("layout", "gruppi")
    if layout not in ("gruppi", "calendario", "affiancata"):
        layout = "gruppi"

    voci_visite = [v for v in voci if v["kind"] == "visita"]
    voci_formazione = [v for v in voci if v["kind"] == "formazione"]

    cal_settimane = cal_label = cal_prev = cal_next = None
    if layout == "calendario":
        try:
            cal_anno = int(request.GET.get("anno") or oggi.year)
            cal_mese = int(request.GET.get("mese") or oggi.month)
        except (TypeError, ValueError):
            cal_anno, cal_mese = oggi.year, oggi.month
        if not 1 <= cal_mese <= 12:
            cal_mese = oggi.month
        primo = date(cal_anno, cal_mese, 1)
        per_giorno: dict = {}
        for v in voci:
            ds = v["data_scadenza"]
            if ds and ds.year == cal_anno and ds.month == cal_mese:
                per_giorno.setdefault(ds.day, []).append(v)
        settimane = []
        for week in calendar.Calendar(firstweekday=0).monthdatescalendar(cal_anno, cal_mese):
            giorni = []
            for gg in week:
                in_mese = (gg.month == cal_mese)
                giorni.append({
                    "data": gg,
                    "in_mese": in_mese,
                    "voci": per_giorno.get(gg.day, []) if in_mese else [],
                    "is_oggi": gg == oggi,
                })
            settimane.append(giorni)
        cal_settimane = settimane
        cal_label = primo.strftime("%B %Y").capitalize()
        cal_prev = _add_months(primo, -1)
        cal_next = _add_months(primo, 1)

    return render(request, "anagrafica/pages/scadenzario.html", {
        "page_obj":      page_obj,
        "n_scadute":     n_scadute,
        "n_30gg":        n_30gg,
        "n_60gg":        n_60gg,
        "oggi":          oggi,
        "filtro_tipo":   filtro_tipo,
        "filtro_stato":  filtro_stato,
        "filtro_reparto": filtro_reparto,
        "reparti":       reparti,
        "gruppi":        gruppi,
        "can_view_visite": can_view_visite,
        "totale":        len(voci),
        "can_view_formazione": can_view_formazione,
        "can_view_contratti": can_view_contratti,
        "is_qual_admin": _qualifiche_can_edit(request),
        "fm_n_scaduti":  fm_n_scaduti,
        "fm_n_30gg":     fm_n_30gg,
        "fm_n_90gg":     fm_n_90gg,
        "fm_is_cache_empty": fm_is_cache_empty,
        "layout":        layout,
        "voci_visite":   voci_visite,
        "voci_formazione": voci_formazione,
        "cal_settimane": cal_settimane,
        "cal_label":     cal_label,
        "cal_prev":      cal_prev,
        "cal_next":      cal_next,
    })


# Cedolini — import XLSX via pagina web (replica logica del management command)
# ---------------------------------------------------------------------------

_MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}

_COL_CF       = 0
_COL_MESE     = 3
_COL_ANNO     = 4
_COL_DATA_PER = 5
_COL_ANZ_ANNI = 8
_COL_ANZ_MESI = 9
_COL_FERIE_AP  = 16
_COL_FERIE_MAT = 17
_COL_FERIE_GOD = 18
_COL_FERIE_RES = 19
_COL_PERM_AP   = 20
_COL_PERM_MAT  = 21
_COL_PERM_GOD  = 22
_COL_PERM_RES  = 23
_COL_ROL_AP    = 24
_COL_ROL_MAT   = 25
_COL_ROL_GOD   = 26
_COL_ROL_RES   = 27
_COL_EXFEST_AP  = 28
_COL_EXFEST_MAT = 29
_COL_EXFEST_GOD = 30
_COL_EXFEST_RES = 31


def _ced_dec(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _ced_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _ced_data_competenza(row) -> "date | None":
    dt = row[_COL_DATA_PER]
    if isinstance(dt, datetime):
        return dt.date()
    if isinstance(dt, date):
        return dt
    mese_str = str(row[_COL_MESE] or "").strip().lower()
    anno = _ced_int(row[_COL_ANNO])
    mese_num = _MESI_IT.get(mese_str)
    if mese_num and anno:
        ultimo = calendar.monthrange(anno, mese_num)[1]
        return date(anno, mese_num, ultimo)
    return None


def _import_xlsx_cedolini(file_obj, user, file_nome: str) -> ImportazioneCedolini:
    """Importa saldi cedolini da file XLSX e crea record su DB."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("openpyxl non installato sul server.") from exc

    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    foglio = "Dati"
    if foglio not in wb.sheetnames:
        raise ValueError(
            f"Foglio '{foglio}' non trovato nel file. Fogli disponibili: {', '.join(wb.sheetnames)}"
        )
    ws = wb[foglio]

    cf_to_lid = dict(
        DipendenteAnagraficaCivile.objects
        .exclude(codice_fiscale="")
        .values_list("codice_fiscale", "legacy_anagrafica_id")
    )

    periodi: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cf = str(row[_COL_CF] or "").strip().upper()
        if not cf:
            continue
        dt = _ced_data_competenza(row)
        if dt is None:
            continue
        periodi.setdefault(dt, []).append((cf, row))

    if not periodi:
        raise ValueError("Il file non contiene righe valide (nessun codice fiscale o data competenza leggibili).")

    totale_ok = totale_err = totale_nf = 0
    ultima_imp = None

    with transaction.atomic():
        for dt in sorted(periodi):
            righe = periodi[dt]
            imp = ImportazioneCedolini.objects.create(
                data_competenza=dt,
                origine=ImportazioneCedolini.ORIGINE_XLSX,
                importato_da=user,
                file_nome=file_nome,
                righe_totali=len(righe),
            )
            ok = err = nf = 0
            for cf, row in righe:
                try:
                    lid = cf_to_lid.get(cf)
                    if lid is None:
                        nf += 1
                    SaldoCedolino.objects.update_or_create(
                        tax_code=cf,
                        data_competenza=dt,
                        defaults=dict(
                            importazione=imp,
                            legacy_anagrafica_id=lid,
                            anzianita_anni=_ced_int(row[_COL_ANZ_ANNI]),
                            anzianita_mesi=_ced_int(row[_COL_ANZ_MESI]),
                            ferie_anni_prec=_ced_dec(row[_COL_FERIE_AP]),
                            ferie_maturati=_ced_dec(row[_COL_FERIE_MAT]),
                            ferie_goduti=_ced_dec(row[_COL_FERIE_GOD]),
                            ferie_residui=_ced_dec(row[_COL_FERIE_RES]),
                            permessi_anni_prec=_ced_dec(row[_COL_PERM_AP]),
                            permessi_maturati=_ced_dec(row[_COL_PERM_MAT]),
                            permessi_goduti=_ced_dec(row[_COL_PERM_GOD]),
                            permessi_residui=_ced_dec(row[_COL_PERM_RES]),
                            rol_anni_prec=_ced_dec(row[_COL_ROL_AP]),
                            rol_maturati=_ced_dec(row[_COL_ROL_MAT]),
                            rol_goduti=_ced_dec(row[_COL_ROL_GOD]),
                            rol_residui=_ced_dec(row[_COL_ROL_RES]),
                            ex_fest_anni_prec=_ced_dec(row[_COL_EXFEST_AP]),
                            ex_fest_maturati=_ced_dec(row[_COL_EXFEST_MAT]),
                            ex_fest_goduti=_ced_dec(row[_COL_EXFEST_GOD]),
                            ex_fest_residui=_ced_dec(row[_COL_EXFEST_RES]),
                        ),
                    )
                    ok += 1
                except Exception:
                    err += 1
                    logger.exception("Errore import cedolino CF=%s mese=%s", cf, dt)

            imp.righe_ok = ok
            imp.righe_errore = err
            imp.righe_non_trovate = nf
            imp.save(update_fields=["righe_ok", "righe_errore", "righe_non_trovate"])
            totale_ok += ok
            totale_err += err
            totale_nf += nf
            ultima_imp = imp

    return ultima_imp


@login_required
def cedolini_import(request):
    """Pagina importazione XLSX cedolini (solo admin)."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not is_admin:
        messages.error(request, "Accesso riservato agli amministratori.")
        return redirect("anagrafica:dipendenti_list")

    if request.method == "POST":
        file_obj = request.FILES.get("file_xlsx")
        if not file_obj:
            messages.error(request, "Nessun file selezionato.")
        else:
            try:
                imp = _import_xlsx_cedolini(file_obj, request.user, file_obj.name)
                periodi_n = ImportazioneCedolini.objects.filter(file_nome=file_obj.name).count()
                messages.success(
                    request,
                    f"Importazione completata: {imp.righe_ok} saldi salvati"
                    f"{f', {imp.righe_errore} errori' if imp.righe_errore else ''}"
                    f"{f', {imp.righe_non_trovate} CF non in anagrafica' if imp.righe_non_trovate else ''}.",
                )
                return redirect("anagrafica:cedolini_import")
            except Exception as exc:
                logger.exception("Errore importazione XLSX cedolini")
                messages.error(request, f"Errore durante l'importazione: {exc}")

    importazioni = list(
        ImportazioneCedolini.objects
        .select_related("importato_da")
        .order_by("-data_competenza", "-data_importazione")[:50]
    )
    return render(request, "anagrafica/pages/cedolini_import.html", {
        "importazioni": importazioni,
        "is_admin": is_admin,
    })


# Ratei ferie/ROL/ex-festività — vista aggregata con filtro mese + dipendente
# ---------------------------------------------------------------------------

@login_required
def ratei_list(request):
    """Lista aggregata saldi cedolini con filtro per periodo e dipendente (solo HR)."""
    from core.legacy_models import UtenteLegacy
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    can_hr = _check_hr_permission(request)

    if not can_hr:
        messages.error(request, "Accesso non autorizzato ai dati HR.")
        return redirect("anagrafica:index")

    # Periodi disponibili
    periodi_qs = (
        SaldoCedolino.objects.values("data_competenza")
        .distinct()
        .order_by("-data_competenza")
    )
    periodi = [p["data_competenza"] for p in periodi_qs]

    # Build opzioni multi-select dipendente: tax_code → "Cognome Nome"
    all_cf_legacy = list(
        SaldoCedolino.objects.values("tax_code", "legacy_anagrafica_id")
        .distinct()
        .order_by("tax_code")
    )
    # Fallback CF → legacy_id via DipendenteAnagraficaCivile (per cedolini senza legacy_id)
    cf_civile_legacy: dict = {}
    for c in DipendenteAnagraficaCivile.objects.exclude(codice_fiscale="").values("codice_fiscale", "legacy_anagrafica_id"):
        cf_u = (c["codice_fiscale"] or "").strip().upper()
        if cf_u:
            cf_civile_legacy[cf_u] = c["legacy_anagrafica_id"]

    cf_to_legacy_rl: dict = {}
    for row in all_cf_legacy:
        cf = (row["tax_code"] or "").strip()
        if not cf or cf in cf_to_legacy_rl:
            continue
        cf_to_legacy_rl[cf] = row["legacy_anagrafica_id"] or cf_civile_legacy.get(cf.upper())

    legacy_ids = sorted({lid for lid in cf_to_legacy_rl.values() if lid})
    dip_qs = list(AnagraficaDipendente.objects.filter(id__in=legacy_ids).values("id", "cognome", "nome", "reparto"))
    id_to_nome: dict = {
        d["id"]: f'{(d["cognome"] or "").strip()} {(d["nome"] or "").strip()}'.strip()
        for d in dip_qs
    }
    # Fallback reparto: AnagraficaDipendente.reparto → DipendenteAnagraficaAziendale.area
    id_to_az_reparto_rl: dict = dict(
        DipendenteAnagraficaAziendale.objects
        .filter(legacy_anagrafica_id__in=legacy_ids)
        .exclude(area="")
        .values_list("legacy_anagrafica_id", "area")
    )
    id_to_reparto: dict = {
        d["id"]: (d["reparto"] or id_to_az_reparto_rl.get(d["id"], "")).strip()
        for d in dip_qs
    }

    seen_cf: set = set()
    dipendenti_options: list = []
    for cf, lid in cf_to_legacy_rl.items():
        if cf in seen_cf:
            continue
        seen_cf.add(cf)
        nome = id_to_nome.get(lid) if lid else None
        reparto = id_to_reparto.get(lid, "") if lid else ""
        dipendenti_options.append({"cf": cf, "nome": nome or cf, "reparto": reparto})
    dipendenti_options.sort(key=lambda x: x["nome"])
    cf_to_nome: dict = {d["cf"]: d["nome"] for d in dipendenti_options}

    reparti_options: list = sorted({d["reparto"] for d in dipendenti_options if d["reparto"]})

    # Filtri GET — backward compat: ?cf=CF_CODE (link da dipendente_detail legacy)
    filtro_periodo = request.GET.get("periodo", "")
    filtro_dipendenti: list = request.GET.getlist("dipendente")
    if not filtro_dipendenti:
        cf_compat = request.GET.get("cf", "").strip().upper()
        if cf_compat:
            filtro_dipendenti = [cf_compat]
    filtro_reparti: list = request.GET.getlist("reparto")
    filtro_allerta: bool = request.GET.get("allerta", "") == "1"

    qs = SaldoCedolino.objects.all().order_by("-data_competenza", "tax_code")

    if filtro_periodo:
        try:
            from datetime import date as _date
            anno, mese, giorno = filtro_periodo.split("-")
            qs = qs.filter(data_competenza=_date(int(anno), int(mese), int(giorno)))
        except (ValueError, AttributeError):
            pass

    if filtro_reparti:
        ids_in_reparto = [lid for lid, rep in id_to_reparto.items() if rep in filtro_reparti]
        qs = qs.filter(legacy_anagrafica_id__in=ids_in_reparto)

    if filtro_dipendenti:
        qs = qs.filter(tax_code__in=filtro_dipendenti)

    # AB1-D — semaforo/alert residuo ferie (solo HR/amministrazione)
    from .ratei_alert import filtro_allerta_q, soglie_ratei, valuta_residuo_ferie
    soglie = soglie_ratei()
    # KPI sul set filtrato (periodo/dipendente/reparto), prima del toggle allerta
    n_negativi = qs.filter(ferie_residui__lt=0).count()
    n_accumulo = qs.filter(ferie_residui__gte=soglie["ore_max"]).count()
    n_allerta = qs.filter(filtro_allerta_q(soglie)).count()
    if filtro_allerta:
        qs = qs.filter(filtro_allerta_q(soglie))

    totale = qs.count()
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    for s in page_obj.object_list:
        s.nome_display = cf_to_nome.get(s.tax_code.upper(), s.tax_code)
        s.semaforo = valuta_residuo_ferie(s.ferie_residui, soglie)

    return render(request, "anagrafica/pages/ratei_list.html", {
        "page_obj": page_obj,
        "periodi": periodi,
        "filtro_periodo": filtro_periodo,
        "filtro_dipendenti": filtro_dipendenti,
        "filtro_reparti": filtro_reparti,
        "filtro_allerta": filtro_allerta,
        "dipendenti_options": dipendenti_options,
        "reparti_options": reparti_options,
        "can_hr": can_hr,
        "is_admin": is_admin,
        "totale": totale,
        "soglie": soglie,
        "n_negativi": n_negativi,
        "n_accumulo": n_accumulo,
        "n_allerta": n_allerta,
    })


# ---------------------------------------------------------------------------
# Ratei ferie — export XLSX (stessa logica filtri di ratei_list)
# ---------------------------------------------------------------------------

@login_required
def ratei_export(request):
    """Scarica XLSX dei saldi cedolini con i filtri correnti (solo HR)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    from core.excel_export import append_row

    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso non autorizzato ai dati HR.")
        return redirect("anagrafica:index")

    # Mappa legacy_id → nome/reparto
    all_cf_legacy = list(
        SaldoCedolino.objects.values("tax_code", "legacy_anagrafica_id")
        .distinct()
        .order_by("tax_code")
    )
    cf_civile_legacy_exp: dict = {}
    for c in DipendenteAnagraficaCivile.objects.exclude(codice_fiscale="").values("codice_fiscale", "legacy_anagrafica_id"):
        cf_u = (c["codice_fiscale"] or "").strip().upper()
        if cf_u:
            cf_civile_legacy_exp[cf_u] = c["legacy_anagrafica_id"]

    cf_to_legacy_exp: dict = {}
    for row in all_cf_legacy:
        cf = (row["tax_code"] or "").strip()
        if not cf or cf in cf_to_legacy_exp:
            continue
        cf_to_legacy_exp[cf] = row["legacy_anagrafica_id"] or cf_civile_legacy_exp.get(cf.upper())

    legacy_ids = sorted({lid for lid in cf_to_legacy_exp.values() if lid})
    dip_qs = list(AnagraficaDipendente.objects.filter(id__in=legacy_ids).values("id", "cognome", "nome", "reparto"))
    id_to_nome: dict = {
        d["id"]: f'{(d["cognome"] or "").strip()} {(d["nome"] or "").strip()}'.strip()
        for d in dip_qs
    }
    id_to_az_reparto_exp: dict = dict(
        DipendenteAnagraficaAziendale.objects
        .filter(legacy_anagrafica_id__in=legacy_ids)
        .exclude(area="")
        .values_list("legacy_anagrafica_id", "area")
    )
    id_to_reparto: dict = {
        d["id"]: (d["reparto"] or id_to_az_reparto_exp.get(d["id"], "")).strip()
        for d in dip_qs
    }
    cf_to_nome: dict = {}
    cf_to_reparto: dict = {}
    for cf, lid in cf_to_legacy_exp.items():
        cf_to_nome[cf] = id_to_nome.get(lid, cf) if lid else cf
        cf_to_reparto[cf] = id_to_reparto.get(lid, "") if lid else ""

    # Filtri GET (identici a ratei_list)
    filtro_periodo = request.GET.get("periodo", "")
    filtro_dipendenti: list = request.GET.getlist("dipendente")
    if not filtro_dipendenti:
        cf_compat = request.GET.get("cf", "").strip().upper()
        if cf_compat:
            filtro_dipendenti = [cf_compat]
    filtro_reparti: list = request.GET.getlist("reparto")
    filtro_allerta: bool = request.GET.get("allerta", "") == "1"

    qs = SaldoCedolino.objects.all().order_by("-data_competenza", "tax_code")
    if filtro_periodo:
        try:
            anno, mese, giorno = filtro_periodo.split("-")
            qs = qs.filter(data_competenza=date(int(anno), int(mese), int(giorno)))
        except (ValueError, AttributeError):
            pass
    if filtro_reparti:
        ids_in_reparto = [lid for lid, rep in id_to_reparto.items() if rep in filtro_reparti]
        qs = qs.filter(legacy_anagrafica_id__in=ids_in_reparto)
    if filtro_dipendenti:
        qs = qs.filter(tax_code__in=filtro_dipendenti)
    if filtro_allerta:
        from .ratei_alert import filtro_allerta_q, soglie_ratei
        qs = qs.filter(filtro_allerta_q(soglie_ratei()))

    # --- Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratei Ferie"

    fill_hdr  = PatternFill("solid", fgColor="F1F5F9")
    fill_fer  = PatternFill("solid", fgColor="FEF9C3")
    fill_rol  = PatternFill("solid", fgColor="DBEAFE")
    fill_exf  = PatternFill("solid", fgColor="DCFCE7")
    font_b    = Font(bold=True)
    font_fer  = Font(bold=True, color="854D0E")
    font_rol  = Font(bold=True, color="1E40AF")
    font_exf  = Font(bold=True, color="166534")
    c_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_right   = Alignment(horizontal="right")

    # Riga 1 — gruppi
    groups = [
        ("A1:A2", "Dipendente",   fill_hdr, font_b),
        ("B1:B2", "Reparto",      fill_hdr, font_b),
        ("C1:C2", "Periodo",      fill_hdr, font_b),
        ("D1:D2", "Anzianità",    fill_hdr, font_b),
        ("E1:H1", "Ferie",        fill_fer, font_fer),
        ("I1:L1", "ROL",          fill_rol, font_rol),
        ("M1:P1", "Ex-Festività", fill_exf, font_exf),
    ]
    for rng, label, fill, fnt in groups:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = label; c.fill = fill; c.font = fnt; c.alignment = c_center

    # Riga 2 — sotto-intestazioni
    sub = [
        ("E", "Anni Prec.", fill_fer, font_fer), ("F", "Maturate",  fill_fer, font_fer),
        ("G", "Godute",     fill_fer, font_fer), ("H", "Residue",   fill_fer, font_fer),
        ("I", "Anni Prec.", fill_rol, font_rol), ("J", "Maturati",  fill_rol, font_rol),
        ("K", "Goduti",     fill_rol, font_rol), ("L", "Residui",   fill_rol, font_rol),
        ("M", "Anni Prec.", fill_exf, font_exf), ("N", "Maturate",  fill_exf, font_exf),
        ("O", "Godute",     fill_exf, font_exf), ("P", "Residue",   fill_exf, font_exf),
    ]
    for col, label, fill, fnt in sub:
        c = ws[f"{col}2"]
        c.value = label; c.fill = fill; c.font = fnt; c.alignment = c_center

    ws.freeze_panes = "A3"

    # Dati
    for s in qs.iterator():
        cf_upper = s.tax_code.upper()
        nome = cf_to_nome.get(cf_upper, s.tax_code)
        reparto = cf_to_reparto.get(cf_upper, "")
        periodo = s.data_competenza.strftime("%b %Y") if s.data_competenza else ""
        if s.anzianita_anni is not None:
            anz = f"{s.anzianita_anni}a"
            if s.anzianita_mesi:
                anz += f" {s.anzianita_mesi}m"
        else:
            anz = ""
        append_row(ws, [
            nome, reparto, periodo, anz,
            s.ferie_anni_prec, s.ferie_maturati, s.ferie_goduti, s.ferie_residui,
            s.rol_anni_prec, s.rol_maturati, s.rol_goduti, s.rol_residui,
            s.ex_fest_anni_prec, s.ex_fest_maturati, s.ex_fest_goduti, s.ex_fest_residui,
        ])

    # Larghezze colonne
    for i, w in enumerate([28, 20, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Nome file
    parts = []
    if filtro_reparti:
        parts.append("_".join(r.replace(" ", "-") for r in filtro_reparti[:2]))
    if filtro_periodo:
        parts.append(filtro_periodo[:7])
    fname = "ratei_ferie" + (f"_{'_'.join(parts)}" if parts else "") + ".xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Retribuzioni — vista globale pivot (dipendente+mese × pay_item) con filtri
# ---------------------------------------------------------------------------

_GENERE_LABEL = dict(DipendenteAnagraficaCivile.GENERE_CHOICES)


def _retribuzioni_globale_context(request) -> dict:
    """Costruisce dataset e filtri condivisi dalla vista globale retribuzioni e
    dal relativo export. Restituisce mappe display, colonne pivot, righe-gruppo
    (chiavi tax_code+mese ordinate) e i valori dei filtri correnti."""
    from django.db.models import Min

    # Coppie tax_code/legacy presenti nelle voci retributive
    all_pairs = list(
        VoceRetributiva.objects.values("tax_code", "legacy_anagrafica_id").distinct()
    )

    # Anagrafica civile: la voce retributiva non sempre ha `legacy_anagrafica_id`
    # valorizzato, quindi il collegamento al dipendente legacy (nome, reparto)
    # passa dal codice fiscale tramite l'anagrafica civile.
    cf_civile_legacy: dict = {}
    cf_to_sesso: dict = {}
    for c in (
        DipendenteAnagraficaCivile.objects
        .exclude(codice_fiscale="")
        .values("codice_fiscale", "legacy_anagrafica_id", "genere")
    ):
        cf_u = (c["codice_fiscale"] or "").strip().upper()
        if not cf_u:
            continue
        cf_civile_legacy[cf_u] = c["legacy_anagrafica_id"]
        cf_to_sesso[cf_u] = c["genere"] or ""

    # Mappa tax_code -> legacy_id: prima il valore della voce, poi l'anagrafica civile
    cf_to_legacy: dict = {}
    for p in all_pairs:
        cf = (p["tax_code"] or "").strip()
        if not cf or cf in cf_to_legacy:
            continue
        cf_to_legacy[cf] = p["legacy_anagrafica_id"] or cf_civile_legacy.get(cf.upper())

    legacy_ids = sorted({lid for lid in cf_to_legacy.values() if lid})

    dip_qs = list(
        AnagraficaDipendente.objects.filter(id__in=legacy_ids)
        .values("id", "cognome", "nome", "reparto")
    )
    id_to_nome = {
        d["id"]: f'{(d["cognome"] or "").strip()} {(d["nome"] or "").strip()}'.strip()
        for d in dip_qs
    }
    # Fallback reparto: AnagraficaDipendente.reparto → DipendenteAnagraficaAziendale.area
    _id_to_az_reparto: dict = dict(
        DipendenteAnagraficaAziendale.objects
        .filter(legacy_anagrafica_id__in=legacy_ids)
        .exclude(area="")
        .values_list("legacy_anagrafica_id", "area")
    )
    id_to_reparto = {
        d["id"]: (d["reparto"] or _id_to_az_reparto.get(d["id"], "")).strip()
        for d in dip_qs
    }

    # Livello contrattuale corrente: contratto piu' recente per dipendente
    id_to_livello: dict = {}
    cf_to_livello: dict = {}
    for sc in (
        StoricoContratto.objects
        .order_by("-data_inizio", "-created_at")
        .values("legacy_anagrafica_id", "tax_code", "codice_livello")
    ):
        liv = (sc["codice_livello"] or "").strip()
        if not liv:
            continue
        lid = sc["legacy_anagrafica_id"]
        cf = (sc["tax_code"] or "").strip().upper()
        if lid and lid not in id_to_livello:
            id_to_livello[lid] = liv
        if cf and cf not in cf_to_livello:
            cf_to_livello[cf] = liv

    def _nome(cf):
        lid = cf_to_legacy.get(cf)
        return (id_to_nome.get(lid) if lid else None) or cf

    def _reparto(cf):
        lid = cf_to_legacy.get(cf)
        return id_to_reparto.get(lid, "") if lid else ""

    def _sesso(cf):
        return cf_to_sesso.get(cf.upper(), "")

    def _livello(cf):
        lid = cf_to_legacy.get(cf)
        return (id_to_livello.get(lid) if lid else None) or cf_to_livello.get(cf.upper(), "")

    # Opzioni filtri
    dipendenti_options = sorted(
        ({"cf": cf, "nome": _nome(cf), "reparto": _reparto(cf)} for cf in cf_to_legacy),
        key=lambda x: x["nome"],
    )
    reparti_options = sorted({r for r in id_to_reparto.values() if r})
    livelli_db = list(
        LivelloContrattuale.objects.filter(is_active=True)
        .order_by("ordine", "codice")
        .values_list("codice", flat=True)
    )
    livelli_usati = {_livello(cf) for cf in cf_to_legacy if _livello(cf)}
    livelli_options = list(dict.fromkeys(list(livelli_db) + sorted(livelli_usati)))
    periodi = list(
        VoceRetributiva.objects.values_list("data_competenza", flat=True)
        .distinct().order_by("-data_competenza")
    )

    # Filtri GET
    filtro_dipendenti = [c for c in request.GET.getlist("dipendente") if c.strip()]
    filtro_reparti = request.GET.getlist("reparto")
    filtro_sesso = request.GET.get("sesso", "").strip()
    filtro_livelli = [v.strip() for v in request.GET.getlist("livello") if v.strip()]
    filtro_periodo = request.GET.get("periodo", "")

    # Insieme tax_code ammessi dopo i filtri dipendente/reparto/sesso/livello
    allowed = set(cf_to_legacy.keys())
    if filtro_dipendenti:
        allowed &= set(filtro_dipendenti)
    if filtro_reparti:
        allowed = {cf for cf in allowed if _reparto(cf) in filtro_reparti}
    if filtro_sesso:
        allowed = {cf for cf in allowed if _sesso(cf) == filtro_sesso}
    if filtro_livelli:
        allowed = {cf for cf in allowed if _livello(cf) in filtro_livelli}

    has_filters = bool(filtro_dipendenti or filtro_reparti or filtro_sesso or filtro_livelli)
    qs = VoceRetributiva.objects.all()
    if has_filters:
        qs = qs.filter(tax_code__in=allowed)
    if filtro_periodo:
        try:
            anno, mese, giorno = filtro_periodo.split("-")
            qs = qs.filter(data_competenza=date(int(anno), int(mese), int(giorno)))
        except (ValueError, AttributeError):
            pass

    # Colonne pivot: tutti i pay_item del set filtrato, ordine di prima apparizione
    col_rows = list(
        qs.values("pay_item_key", "pay_item", "categoria")
        .annotate(_first=Min("data_competenza"))
        .order_by("_first")
    )
    _CAT_ORDER = {"fisso": 0, "variabile": 1, "altro": 2, "totale": 3}
    _TOTALI_SORT = {"retribuzione di fatto": 0, "totale elementi variabili": 1, "rml": 2, "ral": 3}
    key_meta: dict = {}
    seq = 0
    for r in col_rows:
        k = r["pay_item_key"]
        if k not in key_meta:
            key_meta[k] = {
                "key": k, "label": r["pay_item"], "categoria": r["categoria"], "_seq": seq,
            }
            seq += 1

    def _col_sort(m):
        cat = _CAT_ORDER.get(m["categoria"], 99)
        sub = _TOTALI_SORT.get(m["key"], 50) if m["categoria"] == "totale" else m["_seq"]
        return (cat, sub)

    colonne = sorted(key_meta.values(), key=_col_sort)

    _CAT_LABEL = dict(VoceRetributiva.CATEGORIA_CHOICES)
    gruppi: list = []
    for c in colonne:
        if gruppi and gruppi[-1]["categoria"] == c["categoria"]:
            gruppi[-1]["n"] += 1
        else:
            gruppi.append({
                "categoria": c["categoria"],
                "label": _CAT_LABEL.get(c["categoria"], c["categoria"]),
                "n": 1,
            })

    # Righe-gruppo: una per ogni combinazione dipendente+mese
    group_rows = list(
        qs.values("tax_code", "data_competenza").distinct()
        .order_by("-data_competenza", "tax_code")
    )

    return {
        "qs": qs,
        "colonne": colonne,
        "gruppi": gruppi,
        "group_rows": group_rows,
        "periodi": periodi,
        "dipendenti_options": dipendenti_options,
        "reparti_options": reparti_options,
        "livelli_options": livelli_options,
        "sesso_options": DipendenteAnagraficaCivile.GENERE_CHOICES,
        "filtro_dipendenti": filtro_dipendenti,
        "filtro_reparti": filtro_reparti,
        "filtro_sesso": filtro_sesso,
        "filtro_livelli": filtro_livelli,
        "filtro_periodo": filtro_periodo,
        "_nome": _nome,
        "_reparto": _reparto,
        "_sesso": _sesso,
        "_livello": _livello,
    }


def _retribuzioni_globale_rows(ctx: dict, groups: list) -> list:
    """Per un sottoinsieme di righe-gruppo costruisce le righe pivot effettive.

    Applica la stessa regola di merge della scheda dipendente: per ogni mese vale
    l'importazione CSV piu' recente, le voci manuali fanno override sullo stesso
    pay_item_key."""
    colonne = ctx["colonne"]
    qs = ctx["qs"]
    tax_set = {g["tax_code"] for g in groups}
    date_set = {g["data_competenza"] for g in groups}
    key_set = {(g["tax_code"], g["data_competenza"]) for g in groups}

    bucket: dict = {}
    for v in (
        qs.filter(tax_code__in=tax_set, data_competenza__in=date_set)
        .select_related("importazione")
    ):
        kk = (v.tax_code, v.data_competenza)
        if kk in key_set:
            bucket.setdefault(kk, []).append(v)

    rows: list = []
    for g in groups:
        cf = g["tax_code"]
        kk = (cf, g["data_competenza"])
        gv = bucket.get(kk, [])
        manuali = [v for v in gv if v.manuale]
        csv_voci = [v for v in gv if not v.manuale]
        if csv_voci:
            latest = max((v.importazione for v in csv_voci), key=lambda i: i.data_importazione)
            csv_voci = [v for v in csv_voci if v.importazione_id == latest.id]
        manuale_keys = {v.pay_item_key for v in manuali}
        eff = {v.pay_item_key: v for v in csv_voci if v.pay_item_key not in manuale_keys}
        eff.update({v.pay_item_key: v for v in manuali})

        celle = []
        for col in colonne:
            v = eff.get(col["key"])
            celle.append({
                "importo": v.importo if v else None,
                "manuale": bool(v and v.manuale),
                "is_totale": col["categoria"] == "totale",
            })
        sesso = ctx["_sesso"](cf)
        rows.append({
            "tax_code": cf,
            "legacy_anagrafica_id": gv[0].legacy_anagrafica_id if gv else None,
            "nome": ctx["_nome"](cf),
            "reparto": ctx["_reparto"](cf),
            "livello": ctx["_livello"](cf),
            "sesso": _GENERE_LABEL.get(sesso, sesso),
            "data_competenza": g["data_competenza"],
            "celle": celle,
        })
    return rows


@login_required
def retribuzioni_globale(request):
    """Vista globale voci retributive: una riga per dipendente+mese, colonne =
    pay_item raggruppate per categoria. Filtri per dipendente (multi), reparto,
    sesso e livello contrattuale (solo HR)."""
    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso non autorizzato ai dati HR.")
        return redirect("anagrafica:index")

    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)

    ctx = _retribuzioni_globale_context(request)

    paginator = Paginator(ctx["group_rows"], 40)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    rows = _retribuzioni_globale_rows(ctx, list(page_obj.object_list))

    return render(request, "anagrafica/pages/retribuzioni_globale.html", {
        "page_obj": page_obj,
        "rows": rows,
        "colonne": ctx["colonne"],
        "gruppi": ctx["gruppi"],
        "n_colonne": len(ctx["colonne"]),
        "periodi": ctx["periodi"],
        "dipendenti_options": ctx["dipendenti_options"],
        "reparti_options": ctx["reparti_options"],
        "livelli_options": ctx["livelli_options"],
        "sesso_options": ctx["sesso_options"],
        "filtro_dipendenti": ctx["filtro_dipendenti"],
        "filtro_reparti": ctx["filtro_reparti"],
        "filtro_sesso": ctx["filtro_sesso"],
        "filtro_livelli": ctx["filtro_livelli"],
        "filtro_periodo": ctx["filtro_periodo"],
        "totale": len(ctx["group_rows"]),
        "can_hr": can_hr,
        "is_admin": is_admin,
    })


@login_required
def retribuzioni_globale_export(request):
    """Esporta in XLSX la vista globale retribuzioni con i filtri correnti (solo HR)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    from core.excel_export import append_row, write_cell

    can_hr = _check_hr_permission(request)
    if not can_hr:
        messages.error(request, "Accesso non autorizzato ai dati HR.")
        return redirect("anagrafica:index")

    ctx = _retribuzioni_globale_context(request)
    colonne = ctx["colonne"]
    rows = _retribuzioni_globale_rows(ctx, ctx["group_rows"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Retribuzioni"

    _CAT_FILL = {
        "fisso": PatternFill("solid", fgColor="EFF6FF"),
        "variabile": PatternFill("solid", fgColor="FEF9C3"),
        "totale": PatternFill("solid", fgColor="DCFCE7"),
        "altro": PatternFill("solid", fgColor="F1F5F9"),
    }
    fill_hdr = PatternFill("solid", fgColor="F1F5F9")
    font_b = Font(bold=True)
    c_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fixed = ["Dipendente", "Periodo", "Reparto", "Livello", "Sesso"]
    n_fixed = len(fixed)

    # Riga 1 — gruppi categoria
    for i, label in enumerate(fixed, 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        c = write_cell(ws, 1, i, label)
        c.fill = fill_hdr
        c.font = font_b
        c.alignment = c_center
    col = n_fixed + 1
    for g in ctx["gruppi"]:
        start, end = col, col + g["n"] - 1
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = write_cell(ws, 1, start, g["label"])
        c.fill = _CAT_FILL.get(g["categoria"], fill_hdr)
        c.font = font_b
        c.alignment = c_center
        col = end + 1

    # Riga 2 — pay_item
    for j, meta in enumerate(colonne):
        c = write_cell(ws, 2, n_fixed + 1 + j, meta["label"])
        c.fill = _CAT_FILL.get(meta["categoria"], fill_hdr)
        c.font = font_b
        c.alignment = c_center
    ws.freeze_panes = "A3"

    for r in rows:
        periodo = r["data_competenza"].strftime("%m-%Y") if r["data_competenza"] else ""
        line = [r["nome"], periodo, r["reparto"], r["livello"], r["sesso"]]
        for cell in r["celle"]:
            line.append(float(cell["importo"]) if cell["importo"] is not None else None)
        append_row(ws, line)

    ws.column_dimensions["A"].width = 28
    for i in range(2, n_fixed + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    for j in range(len(colonne)):
        ws.column_dimensions[get_column_letter(n_fixed + 1 + j)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "retribuzioni_globale"
    if ctx["filtro_periodo"]:
        fname += f"_{ctx['filtro_periodo'][:7]}"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}.xlsx"'
    return response


# ---------------------------------------------------------------------------

@login_required
def impostazioni(request):
    """Pannello unico di gestione dei cataloghi/configurazioni del modulo anagrafica."""
    from datetime import timedelta
    from django.utils import timezone as tz

    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)

    active_tab = (request.GET.get("tab") or "mansioni").strip().lower()

    # --- Mansioni ---
    mansioni = list(Mansione.objects.all().order_by("nome"))
    mansione_counts: dict[str, int] = {}
    try:
        from django.db import connections
        conn_name = "legacy" if "legacy" in connections else "default"
        with connections[conn_name].cursor() as cur:
            cur.execute(
                "SELECT LOWER(mansione), COUNT(*) FROM anagrafica_dipendenti "
                "WHERE mansione IS NOT NULL AND mansione != '' GROUP BY LOWER(mansione)"
            )
            for row in cur.fetchall():
                mansione_counts[row[0]] = row[1]
    except Exception:
        pass
    for m in mansioni:
        m.n_dipendenti = mansione_counts.get(m.nome.lower(), 0)

    cat_order = [c for c, _ in Mansione.CATEGORIA_CHOICES]
    cat_labels = dict(Mansione.CATEGORIA_CHOICES)
    mansioni_grouped: list[tuple[str, str, list]] = []
    for cat_code in cat_order:
        items = [m for m in mansioni if (m.categoria or Mansione.CAT_ALTRO) == cat_code]
        if items:
            mansioni_grouped.append((cat_code, cat_labels[cat_code], items))

    # --- Reparti e Aree aziendali ---
    aree = list(Reparto.objects.prefetch_related("aree_aziendali").order_by("nome"))
    aree_aziendali = list(AreaAziendale.objects.order_by("nome"))
    dipendenti_picker = _dipendenti_picker_rows()
    _dip_picker_map = {item["id"]: item["label"] for item in dipendenti_picker}
    for a in aree:
        a.caporeparto_label = _dip_picker_map.get(a.caporeparto_legacy_id or 0, "")
    for az in aree_aziendali:
        az.responsabile_label = _dip_picker_map.get(az.responsabile_legacy_id or 0, "")

    # --- Ruoli aziendali ---
    ruoli_aziendali = list(RuoloAziendale.objects.all().order_by("nome"))

    # --- Ruoli operativi sicurezza ---
    # select_related("riporta_a") + catalogo/suggeriti: il pannello "Ruoli" inline
    # di impostazioni.html riusa il partial _ruoli_operativi_body.html e ha bisogno
    # dello stesso context della pagina autonoma ruoli_operativi_list.
    ruoli_operativi = (
        RuoloOperativo.objects
        .annotate(n_assegnati=Count("assegnazioni"))
        .select_related("riporta_a")
        .order_by("nome")
    )
    ruoli_catalogo = list(RuoloOperativo.objects.order_by("nome").values("id", "nome"))
    ruoli_suggeriti = [
        "Preposto", "RSPP", "ASPP", "RLS",
        "Squadra antincendio", "Squadra primo soccorso",
        "Addetto emergenze", "Rappresentante sicurezza",
    ]

    # --- Qualifiche professionali ---
    tipi_qualifica = list(
        TipoQualifica.objects.annotate(n_assegnazioni=Count("assegnazioni"))
        .order_by("categoria", "nome")
    )
    oggi = tz.localdate()
    soglia_q = oggi + timedelta(days=60)
    scadenze_q_count = DipendenteQualifica.objects.filter(
        data_scadenza__isnull=False, data_scadenza__lte=soglia_q
    ).count()

    # --- Livelli contrattuali ---
    livelli = list(LivelloContrattuale.objects.all().order_by("ordine", "codice"))

    # --- Tipologie contratto ---
    tipologie = list(TipologiaContratto.objects.all().order_by("ordine", "codice"))

    # --- Permessi HR / widget statistiche (singleton) ---
    stat_perm = AnagraficaStatPermission.get_instance()
    hr_perm = AnagraficaHRPermission.get_instance()
    visite_perm = AnagraficaVisiteMedichePermission.get_instance()
    try:
        from core.legacy_models import Ruolo
        ruoli_acl = list(Ruolo.objects.order_by("nome"))
    except Exception:
        ruoli_acl = []

    # --- Tipi visita medica ---
    tipi_visita = list(
        TipoVisitaMedica.objects
        .annotate(n_visite=Count("visite"))
        .prefetch_related("ruoli_operativi")
        .order_by("nome")
    )
    scadenze_vm_count = VisitaMedica.objects.filter(
        data_scadenza__isnull=False, data_scadenza__lte=soglia_q
    ).count()

    # --- DPI catalogo (gestione diretta dal pannello) ---
    dpi_categorie = []
    dpi_tipi = []
    dpi_modelli = []
    dpi_taglie = []
    try:
        from dpi.models import CategoriaDPI, TipoDPI, ModelloDPI, TagliaDPI as _TagliaDPI
        dpi_categorie = list(CategoriaDPI.objects.order_by("order_index", "nome"))
        dpi_tipi = list(TipoDPI.objects.select_related("categoria").prefetch_related("modelli").order_by("categoria__order_index", "categoria__nome", "ordine", "nome"))
        dpi_modelli = list(ModelloDPI.objects.select_related("tipo", "tipo__categoria").order_by("tipo__categoria__order_index", "tipo__ordine", "codice", "nome"))
        dpi_taglie = list(_TagliaDPI.objects.select_related("modello", "modello__tipo", "modello__tipo__categoria").order_by("modello__tipo__categoria__order_index", "modello__tipo__ordine", "ordine", "valore"))
    except Exception:
        pass

    # --- Cartelle documenti (ordinate ad albero: parent → figlie, con livello) ---
    _cartelle_all = list(
        CartellaDocumentoDipendente.objects
        .annotate(n_documenti=Count("documenti"))
        .prefetch_related("reparti", "ruoli_operativi")
        .order_by("ordine", "nome")
    )
    _cartelle_by_parent: dict = {}
    for _c in _cartelle_all:
        # Targeting: id selezionati (per i multiselect del form) + flag "mirata"
        _c.reparti_ids = {r.id for r in _c.reparti.all()}
        _c.ruoli_ids = {r.id for r in _c.ruoli_operativi.all()}
        _c.is_mirata = bool(_c.reparti_ids or _c.ruoli_ids)
        _cartelle_by_parent.setdefault(_c.parent_id, []).append(_c)
    cartelle_documenti: list = []

    def _walk_cartelle(parent_id, livello):
        for _c in _cartelle_by_parent.get(parent_id, []):
            _c.livello = livello
            cartelle_documenti.append(_c)
            _walk_cartelle(_c.id, livello + 1)

    _walk_cartelle(None, 0)

    # --- Subnav navigazione ---
    subnav_categorie = list(SubnavCategoriaAnagrafica.objects.order_by("ordine", "nome"))
    subnav_links = list(
        SubnavLinkAnagrafica.objects
        .select_related("categoria")
        .order_by("ordine", "etichetta")
    )
    # Rotte del modulo selezionabili nei link subnav (pagine GET senza parametri)
    subnav_route_choices = [
        ("anagrafica:index", "Dashboard modulo (home)"),
        ("anagrafica:scadenzario", "Scadenzario unificato (filtrabile)"),
        # Persone
        ("anagrafica:dipendenti_list", "Persone — elenco dipendenti"),
        ("anagrafica:dipendente_create", "Persone — nuovo dipendente"),
        ("anagrafica:ex_dipendenti_list", "Persone — ex dipendenti"),
        ("anagrafica:organigramma", "Persone — organigramma"),
        ("anagrafica:onboarding_list", "Persone — onboarding (pratiche)"),
        ("anagrafica:documenti_list", "Persone — documenti"),
        ("anagrafica:dipendenti_report", "Persone — report dipendenti"),
        # Competenze (formazione + qualifiche)
        ("anagrafica:formazione_dashboard", "Competenze — dashboard formazione"),
        ("anagrafica:formazione_piani_list", "Competenze — piani formativi"),
        ("anagrafica:formazione_corsi_list", "Competenze — corsi"),
        ("anagrafica:formazione_sessioni_list", "Competenze — sessioni"),
        ("anagrafica:formazione_istruttori_list", "Competenze — istruttori"),
        ("anagrafica:formazione_elearning_hub", "Competenze — e-learning (hub)"),
        ("anagrafica:formazione_online_catalog", "Competenze — corsi online"),
        ("anagrafica:formazione_copertura", "Competenze — copertura / gap"),
        ("anagrafica:qualifiche_dashboard", "Competenze — cruscotto qualifiche"),
        ("anagrafica:qualifiche_list", "Competenze — catalogo qualifiche"),
        ("anagrafica:qualifica_sessioni_list", "Competenze — sessioni di rinnovo"),
        ("anagrafica:matrice_competenze", "Competenze — matrice competenze"),
        # Compliance (salute & sicurezza)
        ("anagrafica:sicurezza_hub", "Compliance — hub sicurezza"),
        ("anagrafica:visite_mediche_dashboard", "Compliance — visite mediche"),
        ("anagrafica:visite_mediche_nuova_sessione", "Compliance — visite, nuova sessione"),
        ("anagrafica:conformita_report", "Compliance — conformità alla mansione"),
        # Amministrazione (paghe & contratti)
        ("anagrafica:retribuzioni_globale", "Amministrazione — analisi retribuzioni"),
        ("anagrafica:retribuzioni_import", "Amministrazione — import retribuzioni"),
        ("anagrafica:contratti_import", "Amministrazione — contratti"),
        ("anagrafica:cedolini_import", "Amministrazione — cedolini"),
        ("anagrafica:ratei_list", "Amministrazione — ratei ferie / ROL"),
        # Cataloghi struttura / impostazioni
        ("anagrafica:ruoli_operativi_list", "Catalogo — ruoli operativi"),
        ("anagrafica:mansioni_list", "Catalogo — mansioni"),
        ("anagrafica:aree_list", "Catalogo — reparti"),
        ("anagrafica:ruoli_aziendali_list", "Catalogo — ruoli aziendali"),
        ("anagrafica:widget_permissions", "Impostazioni — permessi widget"),
        ("anagrafica:impostazioni", "Impostazioni anagrafica"),
    ]

    # --- Workflow onboarding/offboarding ---
    workflow_field_groups = _dipendente_workflow_field_groups()
    workflow_field_map = {
        field["key"]: field
        for group in workflow_field_groups
        for field in group["fields"]
    }
    workflow_campi = list(
        OnboardingOffboardingCampo.objects.order_by("fase", "ordine", "campo_label")
    )
    for item in workflow_campi:
        meta = workflow_field_map.get(item.campo_key)
        item.catalog_missing = meta is None
        item.catalog_label = meta["label"] if meta else item.campo_label
        item.catalog_section = meta["section"] if meta else item.sezione

    return render(request, "anagrafica/pages/impostazioni.html", {
        "is_admin": is_admin,
        "active_tab": active_tab,
        # Mansioni
        "mansioni": mansioni,
        "mansioni_grouped": mansioni_grouped,
        "CATEGORIA_CHOICES": Mansione.CATEGORIA_CHOICES,
        # Reparti
        "aree": aree,
        "aree_aziendali": aree_aziendali,
        "dipendenti_picker": dipendenti_picker,
        # Ruoli aziendali
        "ruoli_aziendali": ruoli_aziendali,
        # Ruoli operativi
        "ruoli_operativi": ruoli_operativi,
        "ruoli_catalogo": ruoli_catalogo,
        "ruoli_suggeriti": ruoli_suggeriti,
        # Qualifiche
        "tipi_qualifica": tipi_qualifica,
        "QUAL_CATEGORIA_CHOICES": TipoQualifica.CATEGORIA_CHOICES,
        "scadenze_q_count": scadenze_q_count,
        # Livelli
        "livelli": livelli,
        # Tipologie
        "tipologie": tipologie,
        # Visite mediche
        "tipi_visita": tipi_visita,
        "scadenze_vm_count": scadenze_vm_count,
        # DPI
        "dpi_categorie": dpi_categorie,
        "dpi_tipi": dpi_tipi,
        "dpi_modelli": dpi_modelli,
        "dpi_taglie": dpi_taglie,
        # Permessi
        "stat_perm": stat_perm,
        "hr_perm": hr_perm,
        "visite_perm": visite_perm,
        "ruoli_acl": ruoli_acl,
        "ACCESSO_TUTTI": AnagraficaStatPermission.ACCESSO_TUTTI,
        "ACCESSO_ADMIN": AnagraficaStatPermission.ACCESSO_ADMIN,
        "ACCESSO_RUOLI": AnagraficaStatPermission.ACCESSO_RUOLI,
        # Cartelle documenti
        "cartelle_documenti": cartelle_documenti,
        # Subnav navigazione
        "subnav_categorie": subnav_categorie,
        "subnav_links": subnav_links,
        "subnav_route_choices": subnav_route_choices,
        # Workflow onboarding/offboarding
        "workflow_campi": workflow_campi,
        "workflow_field_groups": workflow_field_groups,
        "workflow_fase_choices": OnboardingOffboardingCampo.FASE_CHOICES,
        "workflow_categoria_choices": OnboardingOffboardingCampo.CATEGORIA_CHOICES,
    })


@login_required
@require_POST
def impostazioni_permessi_save(request):
    """Salvataggio combinato dei permessi statistiche e dati HR riservati."""
    ok, resp = _impostazioni_admin_check(request, "permessi")
    if not ok:
        return resp

    def _parse_accesso(prefix: str) -> str:
        val = (request.POST.get(f"{prefix}_accesso") or "").strip()
        if val in (
            AnagraficaStatPermission.ACCESSO_TUTTI,
            AnagraficaStatPermission.ACCESSO_ADMIN,
            AnagraficaStatPermission.ACCESSO_RUOLI,
        ):
            return val
        return AnagraficaStatPermission.ACCESSO_ADMIN

    def _parse_ruoli(prefix: str) -> list[int]:
        raw = request.POST.getlist(f"{prefix}_ruolo_ids")
        return [int(r) for r in raw if str(r).isdigit()]

    stat_perm = AnagraficaStatPermission.get_instance()
    stat_perm.accesso = _parse_accesso("stat")
    stat_perm.ruolo_ids = _parse_ruoli("stat")
    stat_perm.save()

    hr_perm = AnagraficaHRPermission.get_instance()
    hr_perm.accesso = _parse_accesso("hr")
    hr_perm.ruolo_ids = _parse_ruoli("hr")
    hr_perm.save()

    visite_perm = AnagraficaVisiteMedichePermission.get_instance()
    visite_perm.accesso = _parse_accesso("visite")
    visite_perm.ruolo_ids = _parse_ruoli("visite")
    visite_perm.save()

    messages.success(request, "Permessi salvati.")
    return _redirect_impostazioni("permessi")


@login_required
@require_POST
def tipo_visita_medica_create(request):
    ok, resp = _impostazioni_admin_check(request, "visite-mediche")
    if not ok:
        return resp
    nome = (request.POST.get("nome") or "").strip()[:150]
    if not nome:
        messages.error(request, "Il nome della visita è obbligatorio.")
        return _redirect_impostazioni("visite-mediche")
    durata_raw = request.POST.get("durata_mesi") or "12"
    try:
        durata_mesi = max(0, int(durata_raw))
    except ValueError:
        durata_mesi = 12
    tv, created = TipoVisitaMedica.objects.get_or_create(
        nome__iexact=nome,
        defaults={
            "nome": nome,
            "durata_mesi": durata_mesi,
            "obbligatoria": request.POST.get("obbligatoria") == "1",
            "descrizione": (request.POST.get("descrizione") or "").strip(),
        },
    )
    if created:
        ruolo_ids = [int(x) for x in request.POST.getlist("ruolo_ids") if str(x).isdigit()]
        if ruolo_ids:
            tv.ruoli_operativi.set(ruolo_ids)
        messages.success(request, f'Tipo visita "{nome}" creato.')
    else:
        messages.warning(request, f'Esiste già un tipo visita con il nome "{nome}".')
    return _redirect_impostazioni("visite-mediche")


@login_required
@require_POST
def tipo_visita_medica_edit(request, tipo_id: int):
    ok, resp = _impostazioni_admin_check(request, "visite-mediche")
    if not ok:
        return resp
    tipo = get_object_or_404(TipoVisitaMedica, pk=tipo_id)
    nome = (request.POST.get("nome") or "").strip()[:150]
    if not nome:
        messages.error(request, "Il nome della visita è obbligatorio.")
        return _redirect_impostazioni("visite-mediche")
    durata_raw = request.POST.get("durata_mesi") or "12"
    try:
        durata_mesi = max(0, int(durata_raw))
    except ValueError:
        durata_mesi = 12
    tipo.nome = nome
    tipo.durata_mesi = durata_mesi
    tipo.obbligatoria = request.POST.get("obbligatoria") == "1"
    tipo.descrizione = (request.POST.get("descrizione") or "").strip()
    tipo.is_active = request.POST.get("is_active") == "1"
    tipo.save()
    ruolo_ids = [int(x) for x in request.POST.getlist("ruolo_ids") if str(x).isdigit()]
    tipo.ruoli_operativi.set(ruolo_ids)
    messages.success(request, f'Tipo visita "{tipo.nome}" aggiornato.')
    return _redirect_impostazioni("visite-mediche")


@login_required
@require_POST
def tipo_visita_medica_delete(request, tipo_id: int):
    ok, resp = _impostazioni_admin_check(request, "visite-mediche")
    if not ok:
        return resp
    tipo = get_object_or_404(TipoVisitaMedica, pk=tipo_id)
    n_uso = tipo.visite.count()
    if n_uso:
        if tipo.is_active:
            tipo.is_active = False
            tipo.save(update_fields=["is_active"])
            messages.warning(request, f'"{tipo.nome}" è in uso ({n_uso} visite registrate): disattivato (non eliminato) per preservare lo storico. Puoi riattivarlo dalla modifica.')
        else:
            messages.error(request, f'"{tipo.nome}" ha {n_uso} visite registrate: non eliminabile (già disattivo).')
        return _redirect_impostazioni("visite-mediche")
    nome = tipo.nome
    tipo.delete()
    messages.success(request, f'Tipo visita "{nome}" eliminato.')
    return _redirect_impostazioni("visite-mediche")


# ---------------------------------------------------------------------------
# DPI catalogo — CRUD inline da anagrafica/impostazioni
# ---------------------------------------------------------------------------

def _dpi_parse_int(value) -> int | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        parsed = int(raw)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _dpi_validate_image(uploaded) -> str | None:
    if uploaded is None:
        return None
    try:
        from core.upload_mime import UploadMimeValidationError, validate_extension_and_mime
        validate_extension_and_mime(
            uploaded,
            allowed_extensions={".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"},
            allowed_mimes={"image/jpeg", "image/png", "image/webp", "image/gif", "image/bmp", "image/x-ms-bmp"},
            label=getattr(uploaded, "name", "immagine"),
        )
    except Exception as exc:
        return str(exc)
    return None


@login_required
@require_POST
def dpi_categoria_create(request):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import CategoriaDPI
    nome = (request.POST.get("nome") or "").strip()
    if not nome:
        messages.error(request, "Il nome della categoria DPI è obbligatorio.")
        return _redirect_impostazioni("dpi")
    cat = CategoriaDPI(
        nome=nome,
        descrizione=(request.POST.get("descrizione") or "").strip(),
        icona_emoji=(request.POST.get("icona_emoji") or "🦺").strip() or "🦺",
        vita_utile_giorni=_dpi_parse_int(request.POST.get("vita_utile_giorni")),
        obbligatoria_mansionario=request.POST.get("obbligatoria_mansionario") == "1",
        unita_misura=(request.POST.get("unita_misura") or "pz").strip() or "pz",
        scorta_minima=_dpi_parse_int(request.POST.get("scorta_minima")) or 0,
        is_active=True,
        order_index=_dpi_parse_int(request.POST.get("order_index")) or 0,
    )
    uploaded = request.FILES.get("immagine")
    if uploaded:
        err = _dpi_validate_image(uploaded)
        if err:
            messages.error(request, err)
            return _redirect_impostazioni("dpi")
        cat.immagine = uploaded
    cat.save()
    messages.success(request, f'Categoria DPI "{nome}" creata.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_categoria_edit(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import CategoriaDPI
    cat = get_object_or_404(CategoriaDPI, pk=pk)
    nome = (request.POST.get("nome") or "").strip()
    if not nome:
        messages.error(request, "Il nome della categoria DPI è obbligatorio.")
        return _redirect_impostazioni("dpi")
    cat.nome = nome
    cat.descrizione = (request.POST.get("descrizione") or "").strip()
    cat.icona_emoji = (request.POST.get("icona_emoji") or "🦺").strip() or "🦺"
    cat.vita_utile_giorni = _dpi_parse_int(request.POST.get("vita_utile_giorni"))
    cat.obbligatoria_mansionario = request.POST.get("obbligatoria_mansionario") == "1"
    cat.unita_misura = (request.POST.get("unita_misura") or "pz").strip() or "pz"
    cat.scorta_minima = _dpi_parse_int(request.POST.get("scorta_minima")) or 0
    cat.is_active = request.POST.get("is_active") == "1"
    cat.order_index = _dpi_parse_int(request.POST.get("order_index")) or 0
    uploaded = request.FILES.get("immagine")
    if uploaded:
        err = _dpi_validate_image(uploaded)
        if err:
            messages.error(request, err)
            return _redirect_impostazioni("dpi")
        cat.immagine = uploaded
    cat.save()
    messages.success(request, f'Categoria DPI "{cat.nome}" aggiornata.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_categoria_delete(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import CategoriaDPI
    cat = get_object_or_404(CategoriaDPI, pk=pk)
    if cat.richieste.exists():
        messages.error(request, f'"{cat.nome}" ha richieste associate — disattivala invece di eliminarla.')
        return _redirect_impostazioni("dpi")
    nome = cat.nome
    cat.delete()
    messages.success(request, f'Categoria DPI "{nome}" eliminata.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_tipo_create(request):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import CategoriaDPI, TipoDPI
    categoria_id = (request.POST.get("categoria_id") or "").strip()
    nome = (request.POST.get("nome") or "").strip()
    if not categoria_id or not nome:
        messages.error(request, "Categoria e nome del tipo DPI sono obbligatori.")
        return _redirect_impostazioni("dpi")
    try:
        categoria = CategoriaDPI.objects.get(pk=int(categoria_id))
    except (CategoriaDPI.DoesNotExist, ValueError):
        messages.error(request, "Categoria DPI non valida.")
        return _redirect_impostazioni("dpi")
    if TipoDPI.objects.filter(categoria=categoria, nome=nome).exists():
        messages.error(request, f'Esiste già un tipo "{nome}" in questa categoria.')
        return _redirect_impostazioni("dpi")
    TipoDPI.objects.create(
        categoria=categoria,
        nome=nome,
        descrizione=(request.POST.get("descrizione") or "").strip(),
        ordine=_dpi_parse_int(request.POST.get("ordine")) or 0,
        is_active=True,
    )
    messages.success(request, f'Tipo DPI "{nome}" creato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_tipo_edit(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import TipoDPI
    tipo = get_object_or_404(TipoDPI, pk=pk)
    nome = (request.POST.get("nome") or "").strip()
    if not nome:
        messages.error(request, "Il nome del tipo DPI è obbligatorio.")
        return _redirect_impostazioni("dpi")
    tipo.nome = nome
    tipo.descrizione = (request.POST.get("descrizione") or "").strip()
    tipo.ordine = _dpi_parse_int(request.POST.get("ordine")) or 0
    tipo.is_active = request.POST.get("is_active") == "1"
    tipo.save()
    messages.success(request, f'Tipo DPI "{tipo.nome}" aggiornato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_tipo_delete(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import TipoDPI
    tipo = get_object_or_404(TipoDPI, pk=pk)
    if tipo.modelli.exists():
        messages.error(request, f'"{tipo.nome}" ha modelli associati — eliminali prima.')
        return _redirect_impostazioni("dpi")
    nome = tipo.nome
    tipo.delete()
    messages.success(request, f'Tipo DPI "{nome}" eliminato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_modello_create(request):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import TipoDPI, ModelloDPI
    tipo_id = (request.POST.get("tipo_id") or "").strip()
    codice = (request.POST.get("codice") or "").strip()
    nome = (request.POST.get("nome") or "").strip()
    if not tipo_id or not codice or not nome:
        messages.error(request, "Tipo, codice e nome del modello DPI sono obbligatori.")
        return _redirect_impostazioni("dpi")
    try:
        tipo = TipoDPI.objects.select_related("categoria").get(pk=int(tipo_id))
    except (TipoDPI.DoesNotExist, ValueError):
        messages.error(request, "Tipo DPI non valido.")
        return _redirect_impostazioni("dpi")
    if ModelloDPI.objects.filter(codice=codice).exists():
        messages.error(request, f'Esiste già un modello con codice "{codice}".')
        return _redirect_impostazioni("dpi")
    modello = ModelloDPI(
        tipo=tipo, codice=codice, nome=nome,
        produttore=(request.POST.get("produttore") or "").strip(),
        descrizione=(request.POST.get("descrizione") or "").strip(),
        vita_utile_giorni=_dpi_parse_int(request.POST.get("vita_utile_giorni")),
        is_active=True,
    )
    uploaded = request.FILES.get("immagine")
    if uploaded:
        err = _dpi_validate_image(uploaded)
        if err:
            messages.error(request, err)
            return _redirect_impostazioni("dpi")
        modello.immagine = uploaded
    modello.save()
    messages.success(request, f'Modello DPI "{codice}" creato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_modello_edit(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import ModelloDPI
    modello = get_object_or_404(ModelloDPI, pk=pk)
    nome = (request.POST.get("nome") or "").strip()
    if not nome:
        messages.error(request, "Il nome del modello DPI è obbligatorio.")
        return _redirect_impostazioni("dpi")
    modello.nome = nome
    modello.produttore = (request.POST.get("produttore") or "").strip()
    modello.descrizione = (request.POST.get("descrizione") or "").strip()
    modello.vita_utile_giorni = _dpi_parse_int(request.POST.get("vita_utile_giorni"))
    modello.is_active = request.POST.get("is_active") == "1"
    uploaded = request.FILES.get("immagine")
    if uploaded:
        err = _dpi_validate_image(uploaded)
        if err:
            messages.error(request, err)
            return _redirect_impostazioni("dpi")
        modello.immagine = uploaded
    modello.save()
    messages.success(request, f'Modello DPI "{modello.codice}" aggiornato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_modello_delete(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import ModelloDPI
    modello = get_object_or_404(ModelloDPI, pk=pk)
    if modello.taglie.exists():
        messages.error(request, f'"{modello.codice}" ha taglie associate — eliminale prima.')
        return _redirect_impostazioni("dpi")
    codice = modello.codice
    modello.delete()
    messages.success(request, f'Modello DPI "{codice}" eliminato.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_taglia_create(request):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import ModelloDPI, TagliaDPI as _TagliaDPI
    modello_id = (request.POST.get("modello_id") or "").strip()
    valore = (request.POST.get("valore") or "").strip()
    if not modello_id or not valore:
        messages.error(request, "Modello e valore della taglia DPI sono obbligatori.")
        return _redirect_impostazioni("dpi")
    try:
        modello = ModelloDPI.objects.get(pk=int(modello_id))
    except (ModelloDPI.DoesNotExist, ValueError):
        messages.error(request, "Modello DPI non valido.")
        return _redirect_impostazioni("dpi")
    if _TagliaDPI.objects.filter(modello=modello, valore=valore).exists():
        messages.error(request, f'Esiste già la taglia "{valore}" per questo modello.')
        return _redirect_impostazioni("dpi")
    _TagliaDPI.objects.create(
        modello=modello,
        valore=valore,
        ordine=_dpi_parse_int(request.POST.get("ordine")) or 0,
        is_active=True,
    )
    messages.success(request, f'Taglia DPI "{valore}" creata.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_taglia_edit(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import TagliaDPI as _TagliaDPI
    taglia = get_object_or_404(_TagliaDPI, pk=pk)
    valore = (request.POST.get("valore") or "").strip()
    if not valore:
        messages.error(request, "Il valore della taglia DPI è obbligatorio.")
        return _redirect_impostazioni("dpi")
    taglia.valore = valore
    taglia.ordine = _dpi_parse_int(request.POST.get("ordine")) or 0
    taglia.is_active = request.POST.get("is_active") == "1"
    taglia.save()
    messages.success(request, f'Taglia DPI "{taglia.valore}" aggiornata.')
    return _redirect_impostazioni("dpi")


@login_required
@require_POST
def dpi_taglia_delete(request, pk: int):
    ok, resp = _impostazioni_admin_check(request, "dpi")
    if not ok:
        return resp
    from dpi.models import TagliaDPI as _TagliaDPI
    taglia = get_object_or_404(_TagliaDPI, pk=pk)
    valore = taglia.valore
    taglia.delete()
    messages.success(request, f'Taglia DPI "{valore}" eliminata.')
    return _redirect_impostazioni("dpi")


# ===========================================================================
# Visite mediche / Documenti dipendente / HTMX DPI iniziali
# ===========================================================================

def _ensure_admin(request):
    """Ritorna (legacy_user, is_admin). Riusa la convenzione del modulo."""
    legacy_user = get_legacy_user(request.user)
    return legacy_user, (request.user.is_superuser or is_legacy_admin(legacy_user))


def _salva_referto_visita(request, visita: VisitaMedica, referto_file) -> DocumentoDipendente:
    """Crea il ``DocumentoDipendente`` VISITA_MEDICA_REFERTO (storage privato)
    e lo aggancia a ``visita.referto_documento``. Percorso unico per form
    singolo e sessione batch."""
    doc = DocumentoDipendente(
        legacy_anagrafica_id=visita.legacy_anagrafica_id,
        tipo=DocumentoDipendente.Tipo.VISITA_MEDICA_REFERTO,
        nome_originale=getattr(referto_file, "name", "") or "referto",
        tipo_mime=getattr(referto_file, "content_type", "") or "",
        dimensione_bytes=getattr(referto_file, "size", 0) or 0,
        descrizione=f"Referto visita {visita.tipo.nome} del {visita.data_svolgimento}",
        oggetto_riferimento_tipo="anagrafica.visitamedica",
        oggetto_riferimento_id=visita.pk,
        created_by=request.user,
        created_by_display=request.user.get_full_name() or request.user.username,
    )
    doc.file.save(referto_file.name, referto_file, save=True)
    visita.referto_documento = doc
    visita.save(update_fields=["referto_documento", "updated_at"])
    return doc


@login_required
@require_POST
def dipendente_visita_add(request, legacy_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per registrare visite mediche.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    form = VisitaMedicaForm(request.POST, request.FILES)
    if not form.is_valid():
        for err in form.errors.values():
            messages.error(request, "; ".join(err))
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    visita = form.save(commit=False)
    visita.legacy_anagrafica_id = legacy_id
    visita.created_by = request.user
    visita.updated_by = request.user
    visita.save()

    referto_file = form.cleaned_data.get("referto_file")
    if referto_file:
        _salva_referto_visita(request, visita, referto_file)

    try:
        from core.audit import log_action
        log_action(
            request, "VISITA_MEDICA_CREATA", "anagrafica",
            f"Nuova visita {visita.tipo.nome} per #{legacy_id} ({visita.data_svolgimento})",
        )
    except Exception:
        logger.warning("Audit VISITA_MEDICA_CREATA fallito", exc_info=True)

    messages.success(request, f"Visita medica registrata. Scadenza: {visita.data_scadenza or '—'}")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_visita_edit(request, legacy_id: int, v_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per modificare le visite mediche.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    visita = get_object_or_404(VisitaMedica, pk=v_id, legacy_anagrafica_id=legacy_id)
    form = VisitaMedicaForm(request.POST, request.FILES, instance=visita)
    if not form.is_valid():
        for err in form.errors.values():
            messages.error(request, "; ".join(err))
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    visita = form.save(commit=False)
    visita.updated_by = request.user
    visita.save()

    referto_file = form.cleaned_data.get("referto_file")
    if referto_file:
        # Se esisteva un referto, lo sostituiamo con uno nuovo
        if visita.referto_documento_id:
            try:
                visita.referto_documento.delete()
            except Exception:
                logger.warning(
                    "Impossibile rimuovere referto precedente di visita %s", v_id, exc_info=True,
                )
            visita.referto_documento = None
        _salva_referto_visita(request, visita, referto_file)

    try:
        from core.audit import log_action
        log_action(
            request, "VISITA_MEDICA_MODIFICATA", "anagrafica",
            f"Modificata visita {visita.tipo.nome} per #{legacy_id} (v_id={v_id})",
        )
    except Exception:
        logger.warning("Audit VISITA_MEDICA_MODIFICATA fallito", exc_info=True)

    messages.success(request, "Visita medica aggiornata.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
@require_POST
def dipendente_visita_delete(request, legacy_id: int, v_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per eliminare visite mediche.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)
    _, is_admin = _ensure_admin(request)
    if not is_admin:
        messages.error(request, "Solo gli amministratori possono eliminare una visita medica.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    visita = get_object_or_404(VisitaMedica, pk=v_id, legacy_anagrafica_id=legacy_id)
    tipo_nome = visita.tipo.nome
    data = visita.data_svolgimento
    visita.delete()
    try:
        from core.audit import log_action
        log_action(
            request, "VISITA_MEDICA_ELIMINATA", "anagrafica",
            f"Eliminata visita {tipo_nome} per #{legacy_id} ({data})",
        )
    except Exception:
        logger.warning("Audit VISITA_MEDICA_ELIMINATA fallito", exc_info=True)
    messages.success(request, "Visita medica eliminata.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


@login_required
def documento_dipendente_download(request, doc_id: int):
    doc = get_object_or_404(DocumentoDipendente, pk=doc_id)

    # ACL: referti visite mediche → permesso visite. Altri → admin/HR.
    if doc.tipo == DocumentoDipendente.Tipo.VISITA_MEDICA_REFERTO:
        if not _can_view_visite_mediche(request):
            return HttpResponse(status=403)
    else:
        legacy_user = get_legacy_user(request.user)
        if not (request.user.is_superuser or is_legacy_admin(legacy_user) or _check_hr_permission(request)):
            return HttpResponse(status=403)

    if not doc.file:
        return HttpResponse("File non disponibile.", status=404)

    try:
        from core.audit import log_action
        log_action(
            request, "DOCUMENTO_DIPENDENTE_DOWNLOAD", "anagrafica",
            f"Download documento #{doc.pk} ({doc.tipo}) di dipendente #{doc.legacy_anagrafica_id}",
        )
    except Exception:
        logger.warning("Audit DOCUMENTO_DIPENDENTE_DOWNLOAD fallito", exc_info=True)

    from django.http import FileResponse
    try:
        fh = doc.file.open("rb")
    except FileNotFoundError:
        return HttpResponse("File non trovato sul server.", status=404)
    response = FileResponse(fh, as_attachment=True, filename=doc.nome_originale or f"documento_{doc.pk}.bin")
    if doc.tipo_mime:
        response["Content-Type"] = doc.tipo_mime
    return response


@login_required
def foto_dipendente(request, legacy_id: int):
    """Serve la foto profilo del dipendente da storage privato (fuori webroot).

    La foto è un dato personale e NON è esposta su URL pubblico /media/: passa
    da questa view, che richiede autenticazione. Compare nelle liste/rubriche
    interne, quindi è visibile a ogni utente loggato (no ACL HR specifica).
    Lo storage decifra in automatico i file cifrati at-rest.
    """
    civile = DipendenteAnagraficaCivile.objects.filter(
        legacy_anagrafica_id=legacy_id
    ).only("legacy_anagrafica_id", "foto").first()
    if not civile or not civile.foto:
        return HttpResponse(status=404)

    from django.http import FileResponse
    try:
        fh = civile.foto.open("rb")
    except FileNotFoundError:
        return HttpResponse(status=404)

    import mimetypes
    content_type = mimetypes.guess_type(civile.foto.name)[0] or "application/octet-stream"
    response = FileResponse(fh, content_type=content_type)
    # Cache breve lato browser: la foto cambia di rado, ma resta dietro auth.
    response["Cache-Control"] = "private, max-age=3600"
    return response


@login_required
@require_POST
def documento_dipendente_delete(request, doc_id: int):
    _, is_admin = _ensure_admin(request)
    if not is_admin:
        messages.error(request, "Solo gli amministratori possono eliminare un documento.")
        return redirect("anagrafica:dipendenti_list")
    doc = get_object_or_404(DocumentoDipendente, pk=doc_id)
    legacy_id = doc.legacy_anagrafica_id
    tipo = doc.tipo
    doc.delete()
    try:
        from core.audit import log_action
        log_action(
            request, "DOCUMENTO_DIPENDENTE_ELIMINATO", "anagrafica",
            f"Eliminato documento #{doc_id} ({tipo}) di dipendente #{legacy_id}",
        )
    except Exception:
        logger.warning("Audit DOCUMENTO_DIPENDENTE_ELIMINATO fallito", exc_info=True)
    messages.success(request, "Documento eliminato.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Upload manuale documento dipendente
# ---------------------------------------------------------------------------

_ALLOWED_DOC_MIMES = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "image/jpeg",
    "image/png",
    "image/webp",
}
_ALLOWED_DOC_EXTENSIONS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".jpg", ".jpeg", ".png", ".webp"}
_MAX_DOC_SIZE = 50 * 1024 * 1024  # 50 MB


@login_required
@require_POST
def documento_dipendente_upload(request, legacy_id: int):
    """Carica un documento manuale nella cartella virtuale del dipendente."""
    _, is_admin = _ensure_admin(request)
    if not is_admin:
        messages.error(request, "Solo gli amministratori possono caricare documenti.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    cartella_id = request.POST.get("cartella_id") or None
    descrizione = (request.POST.get("descrizione") or "").strip()[:300]
    uploaded = request.FILES.get("file")

    if not uploaded:
        messages.error(request, "Seleziona un file da caricare.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    suffix = Path(uploaded.name or "").suffix.lower()
    if suffix not in _ALLOWED_DOC_EXTENSIONS:
        messages.error(request, f"Formato non consentito ({suffix}). Formati ammessi: PDF, DOC, DOCX, XLS, XLSX, JPG, PNG.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    if uploaded.size > _MAX_DOC_SIZE:
        messages.error(request, f"File troppo grande ({uploaded.size // (1024*1024)} MB). Limite: 50 MB.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    try:
        from core.upload_mime import sniff_mime
        mime = sniff_mime(uploaded)
    except Exception:
        mime = uploaded.content_type or "application/octet-stream"

    if mime not in _ALLOWED_DOC_MIMES:
        messages.error(request, "Tipo di file non consentito (contenuto non valido).")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    cartella = None
    if cartella_id:
        cartella = CartellaDocumentoDipendente.objects.filter(pk=cartella_id, attiva=True).first()

    doc = DocumentoDipendente(
        legacy_anagrafica_id=legacy_id,
        tipo=DocumentoDipendente.Tipo.MANUALE,
        cartella=cartella,
        nome_originale=uploaded.name[:255],
        tipo_mime=mime,
        dimensione_bytes=uploaded.size,
        descrizione=descrizione,
        created_by=request.user,
        created_by_display=request.user.get_full_name() or request.user.username,
    )
    doc.file = uploaded
    doc.save()

    try:
        from core.audit import log_action
        log_action(
            request, "DOCUMENTO_DIPENDENTE_UPLOAD", "anagrafica",
            {
                "documento_id": doc.pk,
                "nome_originale": doc.nome_originale,
                "cartella_id": cartella.pk if cartella else None,
                "cartella_nome": cartella.nome if cartella else "",
                "legacy_anagrafica_id": legacy_id,
            },
        )
    except Exception:
        logger.warning("Audit DOCUMENTO_DIPENDENTE_UPLOAD fallito", exc_info=True)

    messages.success(request, f"Documento '{doc.nome_originale}' caricato correttamente.")
    return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)


# ---------------------------------------------------------------------------
# Lista globale documenti manuali (tab generale)
# ---------------------------------------------------------------------------

@login_required
def documenti_list(request):
    """Lista di tutti i documenti caricati manualmente, filtrabile per dipendente/cartella."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    hr_ok = _check_hr_permission(request)

    if not (is_admin or hr_ok):
        messages.error(request, "Accesso non autorizzato.")
        return redirect("anagrafica:index")

    cartelle = list(CartellaDocumentoDipendente.objects.all())
    qs = DocumentoDipendente.objects.filter(tipo=DocumentoDipendente.Tipo.MANUALE).select_related("cartella", "created_by").order_by("-created_at")

    # Cartelle riservate (solo_admin): visibili nell'archivio solo ai super-amministratori.
    if not request.user.is_superuser:
        cartelle = [c for c in cartelle if not c.solo_admin]
        qs = qs.exclude(cartella__solo_admin=True)

    filtro_cartella = request.GET.get("cartella", "").strip()
    filtro_cerca = request.GET.get("q", "").strip()
    filtro_anno = request.GET.get("anno", "").strip()

    if filtro_cartella:
        if filtro_cartella == "__nessuna__":
            qs = qs.filter(cartella__isnull=True)
        else:
            try:
                qs = qs.filter(cartella_id=int(filtro_cartella))
            except (ValueError, TypeError):
                pass

    if filtro_anno:
        try:
            qs = qs.filter(created_at__year=int(filtro_anno))
        except (ValueError, TypeError):
            pass

    nomi_map = _build_nomi_map()
    documenti = list(qs[:500])
    if filtro_cerca:
        q_low = filtro_cerca.lower()
        documenti = [
            d for d in documenti
            if q_low in nomi_map.get(d.legacy_anagrafica_id, "").lower()
            or q_low in (d.nome_originale or "").lower()
            or q_low in (d.descrizione or "").lower()
        ]

    for d in documenti:
        d.nome_dipendente = nomi_map.get(d.legacy_anagrafica_id, f"#{d.legacy_anagrafica_id}")

    _oggi = __import__("datetime").date.today()
    return render(request, "anagrafica/pages/documenti_list.html", {
        "is_admin": is_admin,
        "documenti": documenti,
        "cartelle": cartelle,
        "filtro_cartella": filtro_cartella,
        "filtro_cerca": filtro_cerca,
        "filtro_anno": filtro_anno,
        "anni_disponibili": list(range(2020, _oggi.year + 1)),
        "today": _oggi,
        "today_plus_1y": _oggi.replace(year=_oggi.year + 1),
    })


@login_required
@require_POST
def documento_sposta_cartella(request, doc_id: int):
    """Sposta un documento manuale in un'altra cartella (o «senza cartella»).

    Gestione del container: stessa autorizzazione dell'archivio (HR/admin)."""
    legacy_user = get_legacy_user(request.user)
    is_admin = request.user.is_superuser or is_legacy_admin(legacy_user)
    if not (is_admin or _check_hr_permission(request)):
        messages.error(request, "Accesso non autorizzato.")
        return redirect("anagrafica:documenti_list")
    doc = get_object_or_404(
        DocumentoDipendente, pk=doc_id, tipo=DocumentoDipendente.Tipo.MANUALE
    )
    target = (request.POST.get("cartella") or "").strip()
    if target in ("", "__nessuna__"):
        doc.cartella = None
    else:
        try:
            doc.cartella = CartellaDocumentoDipendente.objects.get(pk=int(target))
        except (CartellaDocumentoDipendente.DoesNotExist, ValueError, TypeError):
            messages.error(request, "Cartella di destinazione non valida.")
            return redirect("anagrafica:documenti_list")
    doc.save(update_fields=["cartella"])
    messages.success(request, "Documento spostato.")
    nxt = request.POST.get("next") or ""
    if nxt.startswith("/") and not nxt.startswith("//"):
        return redirect(nxt)
    return redirect("anagrafica:documenti_list")


# ---------------------------------------------------------------------------
# CRUD cartelle documenti (da impostazioni)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def cartella_documento_create(request):
    ok, resp = _impostazioni_admin_check(request, "documenti")
    if not ok:
        return resp
    nome = (request.POST.get("nome") or "").strip()[:100]
    descrizione = (request.POST.get("descrizione") or "").strip()[:300]
    ordine_raw = (request.POST.get("ordine") or "0").strip()
    if not nome:
        messages.error(request, "Il nome della cartella è obbligatorio.")
        return _redirect_impostazioni("documenti")
    try:
        ordine = int(ordine_raw)
    except (ValueError, TypeError):
        ordine = 0
    # Cartella superiore (sottocartella) opzionale
    parent = None
    parent_raw = (request.POST.get("parent") or "").strip()
    if parent_raw and parent_raw not in ("0", "__root__"):
        try:
            parent = CartellaDocumentoDipendente.objects.get(pk=int(parent_raw))
        except (CartellaDocumentoDipendente.DoesNotExist, ValueError, TypeError):
            parent = None
    if CartellaDocumentoDipendente.objects.filter(nome__iexact=nome, parent=parent).exists():
        dove = f" in «{parent.nome}»" if parent else " di primo livello"
        messages.error(request, f"Esiste già una cartella «{nome}»{dove}.")
        return _redirect_impostazioni("documenti")
    try:
        retention = max(1, min(99, int((request.POST.get("retention_anni") or "10").strip())))
    except (ValueError, TypeError):
        retention = 10
    solo_admin = request.POST.get("solo_admin") == "1"
    CartellaDocumentoDipendente.objects.create(
        nome=nome, parent=parent, descrizione=descrizione, ordine=ordine,
        retention_anni=retention, solo_admin=solo_admin,
    )
    dove = f" in «{parent.nome}»" if parent else ""
    messages.success(request, f"Cartella «{nome}»{dove} creata.")
    return _redirect_impostazioni("documenti")


@login_required
@require_POST
def cartella_documento_edit(request, cartella_id: int):
    ok, resp = _impostazioni_admin_check(request, "documenti")
    if not ok:
        return resp
    cartella = get_object_or_404(CartellaDocumentoDipendente, pk=cartella_id)
    nome = (request.POST.get("nome") or "").strip()[:100]
    descrizione = (request.POST.get("descrizione") or "").strip()[:300]
    ordine_raw = (request.POST.get("ordine") or "0").strip()
    attiva = request.POST.get("attiva") == "1"
    if not nome:
        messages.error(request, "Il nome della cartella è obbligatorio.")
        return _redirect_impostazioni("documenti")
    if CartellaDocumentoDipendente.objects.filter(nome__iexact=nome, parent=cartella.parent).exclude(pk=cartella_id).exists():
        messages.error(request, f"Esiste già un'altra cartella «{nome}» nello stesso livello.")
        return _redirect_impostazioni("documenti")
    try:
        ordine = int(ordine_raw)
    except (ValueError, TypeError):
        ordine = cartella.ordine
    cartella.nome = nome
    cartella.descrizione = descrizione
    cartella.ordine = ordine
    cartella.attiva = attiva
    cartella.solo_admin = request.POST.get("solo_admin") == "1"
    try:
        cartella.retention_anni = max(1, min(99, int((request.POST.get("retention_anni") or "10").strip())))
    except (ValueError, TypeError):
        pass
    cartella.save()
    # Targeting (visibilità mirata): reparti + ruoli operativi. Vuoto = universale.
    def _ids(name):
        out = []
        for raw in request.POST.getlist(name):
            if str(raw).strip().isdigit():
                out.append(int(raw))
        return out
    cartella.reparti.set(_ids("reparti_ids"))
    cartella.ruoli_operativi.set(_ids("ruoli_ids"))
    messages.success(request, f"Cartella «{nome}» aggiornata.")
    return _redirect_impostazioni("documenti")


@login_required
@require_POST
def cartella_documento_delete(request, cartella_id: int):
    ok, resp = _impostazioni_admin_check(request, "documenti")
    if not ok:
        return resp
    cartella = get_object_or_404(CartellaDocumentoDipendente, pk=cartella_id)
    n_figlie = cartella.figlie.count()
    if n_figlie > 0:
        messages.error(request, f"La cartella «{cartella.nome}» contiene {n_figlie} sottocartella/e: eliminale o spostale prima.")
        return _redirect_impostazioni("documenti")
    n_docs = cartella.documenti.count()
    if n_docs > 0:
        if cartella.attiva:
            cartella.attiva = False
            cartella.save(update_fields=["attiva"])
            messages.warning(request, f"La cartella «{cartella.nome}» contiene {n_docs} documento/i: disattivata (non eliminata) per preservare lo storico. Riattivala dalla modifica.")
        else:
            messages.error(request, f"La cartella «{cartella.nome}» contiene {n_docs} documento/i: non eliminabile (già disattiva).")
        return _redirect_impostazioni("documenti")
    nome = cartella.nome
    cartella.delete()
    messages.success(request, f"Cartella «{nome}» eliminata.")
    return _redirect_impostazioni("documenti")


# ---------------------------------------------------------------------------
# CRUD subnav — categorie e link (da impostazioni)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def subnav_categoria_create(request):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    nome = (request.POST.get("nome") or "").strip()[:80]
    icona = (request.POST.get("icona") or "").strip()[:20]
    ordine_raw = (request.POST.get("ordine") or "0").strip()
    if not nome:
        messages.error(request, "Il nome della categoria è obbligatorio.")
        return _redirect_impostazioni("navigazione")
    try:
        ordine = int(ordine_raw)
    except (ValueError, TypeError):
        ordine = 0
    landing_type = request.POST.get("landing_url_type") or "named"
    if landing_type not in ("named", "raw"):
        landing_type = "named"
    landing_value = (request.POST.get("landing_url_value") or "").strip()[:255]
    SubnavCategoriaAnagrafica.objects.create(
        nome=nome, icona=icona, ordine=ordine,
        landing_url_type=landing_type, landing_url_value=landing_value,
    )
    messages.success(request, f"Categoria «{nome}» creata.")
    return _redirect_impostazioni("navigazione")


@login_required
@require_POST
def subnav_categoria_edit(request, cat_id: int):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    cat = get_object_or_404(SubnavCategoriaAnagrafica, pk=cat_id)
    cat.nome = (request.POST.get("nome") or "").strip()[:80] or cat.nome
    cat.icona = (request.POST.get("icona") or "").strip()[:20]
    cat.is_active = request.POST.get("is_active") == "1"
    try:
        cat.ordine = int(request.POST.get("ordine") or cat.ordine)
    except (ValueError, TypeError):
        pass
    landing_type = request.POST.get("landing_url_type") or cat.landing_url_type
    if landing_type in ("named", "raw"):
        cat.landing_url_type = landing_type
    cat.landing_url_value = (request.POST.get("landing_url_value") or "").strip()[:255]
    cat.save()
    messages.success(request, f"Categoria «{cat.nome}» aggiornata.")
    return _redirect_impostazioni("navigazione")


@login_required
@require_POST
def subnav_categoria_delete(request, cat_id: int):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    cat = get_object_or_404(SubnavCategoriaAnagrafica, pk=cat_id)
    n_links = cat.links.count()
    if n_links > 0:
        messages.error(request, f"Impossibile eliminare: la categoria contiene {n_links} link/s. Spostarli o eliminarli prima.")
        return _redirect_impostazioni("navigazione")
    nome = cat.nome
    cat.delete()
    messages.success(request, f"Categoria «{nome}» eliminata.")
    return _redirect_impostazioni("navigazione")


@login_required
@require_POST
def subnav_link_create(request):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    etichetta = (request.POST.get("etichetta") or "").strip()[:80]
    icona = (request.POST.get("icona") or "").strip()[:20]
    gruppo = (request.POST.get("gruppo") or "").strip()[:40]
    url_type = request.POST.get("url_type") or "raw"
    if url_type not in ("named", "raw"):
        url_type = "raw"
    url_value = (request.POST.get("url_value") or "").strip()[:255]
    cat_id_raw = (request.POST.get("categoria_id") or "").strip()
    apri_nuova_tab = request.POST.get("apri_nuova_tab") == "1"
    try:
        ordine = int(request.POST.get("ordine") or 0)
    except (ValueError, TypeError):
        ordine = 0
    if not etichetta or not url_value:
        messages.error(request, "Etichetta e URL sono obbligatori.")
        return _redirect_impostazioni("navigazione")
    cat = None
    if cat_id_raw:
        cat = SubnavCategoriaAnagrafica.objects.filter(pk=cat_id_raw).first()
    SubnavLinkAnagrafica.objects.create(
        etichetta=etichetta, icona=icona, gruppo=gruppo, url_type=url_type, url_value=url_value,
        categoria=cat, apri_nuova_tab=apri_nuova_tab, ordine=ordine, is_sistema=False,
    )
    messages.success(request, f"Link «{etichetta}» aggiunto.")
    return _redirect_impostazioni("navigazione")


@login_required
@require_POST
def subnav_link_edit(request, link_id: int):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    link = get_object_or_404(SubnavLinkAnagrafica, pk=link_id)
    link.etichetta = (request.POST.get("etichetta") or "").strip()[:80] or link.etichetta
    link.icona = (request.POST.get("icona") or "").strip()[:20]
    link.gruppo = (request.POST.get("gruppo") or "").strip()[:40]
    link.is_active = request.POST.get("is_active") == "1"
    link.apri_nuova_tab = request.POST.get("apri_nuova_tab") == "1"
    try:
        link.ordine = int(request.POST.get("ordine") or link.ordine)
    except (ValueError, TypeError):
        pass
    cat_id_raw = (request.POST.get("categoria_id") or "").strip()
    if cat_id_raw:
        link.categoria = SubnavCategoriaAnagrafica.objects.filter(pk=cat_id_raw).first()
    else:
        link.categoria = None
    # Per link non di sistema permetti anche di cambiare URL
    if not link.is_sistema:
        url_type = request.POST.get("url_type") or link.url_type
        if url_type in ("named", "raw"):
            link.url_type = url_type
        url_value = (request.POST.get("url_value") or "").strip()[:255]
        if url_value:
            link.url_value = url_value
    link.save()
    messages.success(request, f"Link «{link.etichetta}» aggiornato.")
    return _redirect_impostazioni("navigazione")


@login_required
@require_POST
def subnav_link_delete(request, link_id: int):
    ok, resp = _impostazioni_admin_check(request, "navigazione")
    if not ok:
        return resp
    link = get_object_or_404(SubnavLinkAnagrafica, pk=link_id)
    if link.is_sistema:
        messages.error(request, "I link di sistema non possono essere eliminati. Puoi disattivarli.")
        return _redirect_impostazioni("navigazione")
    nome = link.etichetta
    link.delete()
    messages.success(request, f"Link «{nome}» eliminato.")
    return _redirect_impostazioni("navigazione")


@login_required
def dipendente_dpi_iniziali_proposti(request):
    """Partial HTMX: lista righe DPI iniziali in base ai ruoli operativi scelti.

    Accetta ``ruoli_operativi_ids`` (multi-value) via GET. Ritorna l'HTML
    parziale ``_dpi_iniziali_righe.html`` con il formset pre-compilato.
    """
    ruoli_raw = request.GET.getlist("ruoli_operativi_ids")
    ruoli_ids: list[int] = []
    for raw in ruoli_raw:
        try:
            ruoli_ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    righe = proposta_righe_iniziali(ruoli_ids)
    return render(
        request,
        "anagrafica/partials/_dpi_iniziali_righe.html",
        {"righe": righe, "indici": list(range(len(righe)))},
    )


@login_required
def visite_mediche_dashboard(request):
    """Dashboard globale visite mediche.

    Mostra (gating su ``_can_view_visite_mediche``):
    - contatori: scadute / in scadenza / valide / mancanti aggregati sull'azienda
    - tabella "scadute o in scadenza" con dettaglio dipendente
    - tabella "ultime visite registrate"
    - elenco tipologie attive con counter di dipendenti coperti vs richiesti
    """
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per visualizzare le visite mediche.")
        return redirect("anagrafica:index")

    import calendar as _calendar
    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    soglia_avviso = oggi + _timedelta(days=60)

    # Filtro mese dalla query string (?scad=mese_corrente | prossimo_mese | tutti)
    filtro_scad = request.GET.get("scad", "").strip()

    # Map legacy_id -> "Cognome Nome" da AnagraficaDipendente (lookup unico)
    nomi_map = _build_nomi_map()

    # Fonte unica "visite correnti" (l'ultima per dipendente+tipo): le righe
    # storiche superate non contano più come scadute/in scadenza.
    correnti_qs = VisitaMedica.objects.filter(id__in=ultime_visite_correnti_ids())

    # KPI globali (sulle visite correnti dell'intero dataset, non filtrato per mese)
    kpi_scadute = correnti_qs.filter(
        data_scadenza__isnull=False, data_scadenza__lt=oggi
    ).count()
    kpi_in_scad = correnti_qs.filter(
        data_scadenza__isnull=False, data_scadenza__range=[oggi, soglia_avviso]
    ).count()
    kpi_visite_totali = VisitaMedica.objects.count()

    _MESI_ITA = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
                 "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    # Ultime visite registrate (globale, 30 più recenti)
    ultime_visite = list(
        VisitaMedica.objects
        .select_related("tipo")
        .order_by("-data_svolgimento", "-id")[:30]
    )
    for v in ultime_visite:
        v.dipendente_nome = nomi_map.get(v.legacy_anagrafica_id, f"#{v.legacy_anagrafica_id}")
        if v.data_svolgimento:
            v.mese_label = f"{_MESI_ITA[v.data_svolgimento.month]} {v.data_svolgimento.year}"
        else:
            v.mese_label = "Data non disponibile"

    # Scadute o in scadenza — con filtro mese opzionale
    if filtro_scad == "mese_corrente":
        _, last_day_n = _calendar.monthrange(oggi.year, oggi.month)
        _range_start = oggi.replace(day=1)
        _range_end = oggi.replace(day=last_day_n)
        scad_qs = correnti_qs.filter(
            data_scadenza__isnull=False,
            data_scadenza__range=[_range_start, _range_end],
        )
    elif filtro_scad == "prossimo_mese":
        pm_y = oggi.year + 1 if oggi.month == 12 else oggi.year
        pm_m = 1 if oggi.month == 12 else oggi.month + 1
        _, last_day_n = _calendar.monthrange(pm_y, pm_m)
        _range_start = date(pm_y, pm_m, 1)
        _range_end = date(pm_y, pm_m, last_day_n)
        scad_qs = correnti_qs.filter(
            data_scadenza__isnull=False,
            data_scadenza__range=[_range_start, _range_end],
        )
    else:
        filtro_scad = "tutti"
        scad_qs = correnti_qs.filter(
            data_scadenza__isnull=False, data_scadenza__lte=soglia_avviso
        )

    scad_o_in_scad = list(scad_qs.select_related("tipo").order_by("data_scadenza"))
    for v in scad_o_in_scad:
        v.dipendente_nome = nomi_map.get(v.legacy_anagrafica_id, f"#{v.legacy_anagrafica_id}")
        v.giorni_a_scadenza = (v.data_scadenza - oggi).days if v.data_scadenza else None
        if v.data_scadenza:
            v.mese_label = f"{_MESI_ITA[v.data_scadenza.month]} {v.data_scadenza.year}"
        else:
            v.mese_label = "Senza scadenza"

    # Pre-aggrega conteggi DB per tipo (evita N query nel loop) — solo visite correnti
    _valide_per_tipo = {
        row["tipo_id"]: row["n"]
        for row in correnti_qs
        .filter(data_scadenza__gte=oggi)
        .order_by()
        .values("tipo_id")
        .annotate(n=Count("legacy_anagrafica_id", distinct=True))
    }
    _scadute_per_tipo = {
        row["tipo_id"]: row["n"]
        for row in correnti_qs
        .filter(data_scadenza__lt=oggi)
        .order_by()
        .values("tipo_id")
        .annotate(n=Count("legacy_anagrafica_id", distinct=True))
    }

    # Tipologie con copertura ruoli + conteggi reali DB
    tipi_attivi_qs = TipoVisitaMedica.objects.filter(is_active=True).prefetch_related("ruoli_operativi")
    tipologie_stats = []
    for t in tipi_attivi_qs:
        ruoli_ids = list(t.ruoli_operativi.values_list("id", flat=True))
        if ruoli_ids:
            legacy_ids_richiesti = set(
                DipendenteRuoloOperativo.objects
                .filter(ruolo_id__in=ruoli_ids)
                .values_list("legacy_anagrafica_id", flat=True)
            )
        else:
            legacy_ids_richiesti = set()
        legacy_ids_coperti = set(
            correnti_qs
            .filter(tipo=t, data_scadenza__gte=oggi)
            .values_list("legacy_anagrafica_id", flat=True)
        )
        mancanti = legacy_ids_richiesti - legacy_ids_coperti
        tipologie_stats.append({
            "tipo": t,
            "ha_ruoli": bool(ruoli_ids),
            "richiesti": len(legacy_ids_richiesti),
            "coperti": len(legacy_ids_richiesti & legacy_ids_coperti),
            "mancanti": len(mancanti),
            "valide_db": _valide_per_tipo.get(t.pk, 0),
            "scadute_db": _scadute_per_tipo.get(t.pk, 0),
        })

    kpi_tipi_attivi = tipi_attivi_qs.count()

    # Sessioni per tipo (latest per dipendente): usate nel pannello dettaglio interattivo
    from collections import defaultdict as _defaultdict
    _all_tipo_ids = [row["tipo"].pk for row in tipologie_stats]
    if _all_tipo_ids:
        _latest_sessions_bulk = list(
            correnti_qs
            .filter(tipo_id__in=_all_tipo_ids)
            .select_related("tipo")
            .order_by("tipo_id", "legacy_anagrafica_id")
        )
        _soglia_avviso = oggi + _timedelta(days=60)
        _sessions_by_tipo: dict = _defaultdict(list)
        for _v in _latest_sessions_bulk:
            _v.dipendente_nome = nomi_map.get(_v.legacy_anagrafica_id, f"#{_v.legacy_anagrafica_id}")
            _v.giorni_a_scadenza = (_v.data_scadenza - oggi).days if _v.data_scadenza else None
            _sessions_by_tipo[_v.tipo_id].append(_v)

        def _session_sort_key(_v):
            if not _v.data_scadenza:
                return (3, "")
            if _v.data_scadenza < oggi:
                return (0, str(_v.data_scadenza))
            if _v.data_scadenza <= _soglia_avviso:
                return (1, str(_v.data_scadenza))
            return (2, str(_v.data_scadenza))

        for row in tipologie_stats:
            sessioni = sorted(_sessions_by_tipo.get(row["tipo"].pk, []), key=_session_sort_key)
            row["sessioni"] = sessioni
            row["n_scadute"] = sum(1 for s in sessioni if s.data_scadenza and s.giorni_a_scadenza is not None and s.giorni_a_scadenza < 0)
            row["n_in_scadenza"] = sum(1 for s in sessioni if s.data_scadenza and s.giorni_a_scadenza is not None and 0 <= s.giorni_a_scadenza <= 60)
            row["n_valide"] = sum(1 for s in sessioni if s.data_scadenza and s.giorni_a_scadenza is not None and s.giorni_a_scadenza > 60)
    else:
        for row in tipologie_stats:
            row["sessioni"] = []
            row["n_scadute"] = 0
            row["n_in_scadenza"] = 0
            row["n_valide"] = 0

    return render(request, "anagrafica/pages/visite_mediche_dashboard.html", {
        "oggi": oggi,
        "soglia_avviso": oggi + _timedelta(days=60),
        "kpi_scadute": kpi_scadute,
        "kpi_in_scad": kpi_in_scad,
        "kpi_totali": kpi_visite_totali,
        "kpi_tipi_attivi": kpi_tipi_attivi,
        "ultime_visite": ultime_visite,
        "scad_o_in_scad": scad_o_in_scad,
        "tipologie_stats": tipologie_stats,
        "can_manage": _can_view_visite_mediche(request),
        "filtro_scad": filtro_scad,
    })


# ---------------------------------------------------------------------------
# Helpers: registrazione sessione batch visite mediche
# ---------------------------------------------------------------------------

def _build_nomi_map() -> dict[int, str]:
    """Ritorna dict {legacy_anagrafica_id: 'Cognome Nome'} per tutti i dipendenti."""
    nomi: dict[int, str] = {}
    try:
        for r in AnagraficaDipendente.objects.values("id", "cognome", "nome"):
            try:
                lid = int(r.get("id") or 0)
            except (TypeError, ValueError):
                continue
            cog = (r.get("cognome") or "").strip()
            nom = (r.get("nome") or "").strip()
            nomi[lid] = f"{cog} {nom}".strip() or f"#{lid}"
    except Exception:
        logger.exception("Errore lookup nomi dipendenti per sessione visita")
    return nomi


def _cessati_legacy_ids() -> set[int]:
    """Id legacy dei dipendenti cessati (``data_cessazione`` valorizzata)."""
    return set(
        DipendenteAnagraficaAziendale.objects
        .filter(data_cessazione__isnull=False)
        .values_list("legacy_anagrafica_id", flat=True)
    )


def _requisiti_tipo_visita(tipo: TipoVisitaMedica) -> dict:
    """Chi è tenuto alla visita ``tipo`` e da dove nasce l'obbligo.

    Ritorna ``{"da_ruoli": set, "da_processi": set, "ha_vincoli": bool}``.
    ``ha_vincoli`` = il tipo ha ruoli operativi o processi MOD.128 COLLEGATI
    in configurazione (anche se nessuna persona li possiede): governa il
    fallback storico e la valutazione di pertinenza. Cessati NON filtrati qui.
    """
    ruolo_ids = list(tipo.ruoli_operativi.values_list("id", flat=True))
    da_ruoli: set[int] = set()
    if ruolo_ids:
        da_ruoli = set(
            DipendenteRuoloOperativo.objects
            .filter(ruolo_id__in=ruolo_ids)
            .values_list("legacy_anagrafica_id", flat=True)
        )
    da_processi: set[int] = set()
    ha_processi = False
    try:
        from .models_mpq import AbilitazioneProcesso
        ha_processi = tipo.processi_richiedenti.exists()
        if ha_processi:
            da_processi = set(
                AbilitazioneProcesso.objects
                .filter(
                    stato=AbilitazioneProcesso.STATO_ATTIVA,
                    processo__visite_richieste=tipo,
                )
                .exclude(legacy_anagrafica_id=0)
                .values_list("legacy_anagrafica_id", flat=True)
            )
    except Exception:
        logger.warning(
            "Lookup requisiti MOD.128 per tipo visita %s fallito", tipo.pk, exc_info=True,
        )
    return {
        "da_ruoli": da_ruoli,
        "da_processi": da_processi,
        "ha_vincoli": bool(ruolo_ids) or ha_processi,
    }


def _build_candidati_sessione(tipo: TipoVisitaMedica, oggi) -> list[dict]:
    """Dipendenti candidati per una sessione di visita del tipo dato.

    Il tipo è "consono" quando è richiesto dai ruoli operativi del dipendente
    o da un processo MOD.128 a cui è abilitato; se il tipo non ha vincoli
    configurati si propone chi ha quel tipo nello storico (stato calcolato
    sull'ultima visita). Cessati sempre esclusi.

    Candidato = ultima visita del tipo scaduta, in scadenza entro 90 giorni,
    oppure mai effettuata (solo pool ruoli/processi). Un'ultima visita senza
    scadenza (durata 0) è valida per sempre: non viene riproposta.
    """
    soglia = oggi + _timedelta(days=90)
    cessati = _cessati_legacy_ids()
    req = _requisiti_tipo_visita(tipo)
    da_ruoli, da_processi = req["da_ruoli"], req["da_processi"]

    if req["ha_vincoli"]:
        pool_ids = (da_ruoli | da_processi) - cessati
    else:
        # Tipo non collegato a ruoli/processi: si propone chi ha quel tipo
        # nello storico, non tutta l'azienda.
        pool_ids = set(
            VisitaMedica.objects
            .filter(tipo=tipo)
            .values_list("legacy_anagrafica_id", flat=True)
        ) - cessati

    if not pool_ids:
        return []

    # Ultima visita del tipo per ogni candidato (spareggio: pk più alto).
    ultima_per_id: dict[int, VisitaMedica] = {}
    for v in (
        VisitaMedica.objects
        .filter(tipo=tipo, legacy_anagrafica_id__in=pool_ids)
        .select_related("tipo")
        .order_by("legacy_anagrafica_id", "-data_svolgimento", "-pk")
    ):
        if v.legacy_anagrafica_id not in ultima_per_id:
            ultima_per_id[v.legacy_anagrafica_id] = v

    nomi_map = _build_nomi_map()

    candidati = []
    for lid in pool_ids:
        ultima = ultima_per_id.get(lid)
        if ultima is None:
            status = "mai_effettuata"
            giorni = None
        elif ultima.data_scadenza is None:
            continue  # valida senza scadenza: non riproporre
        elif ultima.data_scadenza < oggi:
            status = "scaduta"
            giorni = (ultima.data_scadenza - oggi).days
        elif ultima.data_scadenza <= soglia:
            status = "in_scadenza"
            giorni = (ultima.data_scadenza - oggi).days
        else:
            continue  # ancora valida oltre la soglia dei 90 giorni

        if lid in da_ruoli:
            origine = "ruolo"
        elif lid in da_processi:
            origine = "processo"
        else:
            origine = "storico"

        candidati.append({
            "legacy_id": lid,
            "nome": nomi_map.get(lid, f"#{lid}"),
            "ultima_visita": ultima,
            "data_scadenza": ultima.data_scadenza if ultima else None,
            "status": status,
            "giorni_a_scadenza": giorni,
            "origine": origine,
        })

    # Ordine: in_scadenza → scaduta → mai_effettuata; poi alfabetico per nome
    _status_order = {"in_scadenza": 0, "scaduta": 1, "mai_effettuata": 2}
    candidati.sort(key=lambda c: (_status_order.get(c["status"], 9), c["nome"]))
    return candidati


def _build_candidati_giornata(oggi, tipo_id=None) -> list[dict]:
    """Righe candidate per una "giornata visite" multi-tipo: per ogni tipo di
    visita attivo (o solo ``tipo_id`` se dato) prende i candidati "consoni" da
    ``_build_candidati_sessione`` e li appiattisce in righe ``(persona, tipo)``.
    Chi ha più tipi dovuti compare in più righe (= più visite nel giorno)."""
    if tipo_id:
        tipi = list(TipoVisitaMedica.objects.filter(pk=tipo_id, is_active=True))
    else:
        tipi = list(TipoVisitaMedica.objects.filter(is_active=True).order_by("nome"))

    righe: list[dict] = []
    for tipo in tipi:
        for c in _build_candidati_sessione(tipo, oggi):
            righe.append({
                **c,
                "tipo": tipo,
                "preselect": c["status"] in ("scaduta", "in_scadenza"),
            })
    _status_order = {"in_scadenza": 0, "scaduta": 1, "mai_effettuata": 2}
    righe.sort(key=lambda r: (_status_order.get(r["status"], 9), r["nome"].casefold(), r["tipo"].nome))
    return righe


# ---------------------------------------------------------------------------
# View: registrazione nuova sessione batch visite mediche
# ---------------------------------------------------------------------------

@login_required
def visite_mediche_nuova_sessione(request):
    """Registrazione batch di visite mediche per un tipo e una data.

    Step 1 (GET / POST step=1): selezione tipo visita, data, medico competente.
    Step 2 (POST step=2): tabella dipendenti candidati con esito per ognuno → salva.
    """
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per registrare le visite mediche.")
        return redirect("anagrafica:visite_mediche_dashboard")

    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    tipi_attivi = list(TipoVisitaMedica.objects.filter(is_active=True).order_by("nome"))

    # ---- POST: salva la giornata (sessione + visite multi-tipo) -----------
    if request.method == "POST":
        data_str = request.POST.get("data_svolgimento", "").strip()
        medico = request.POST.get("medico_competente", "").strip()
        luogo = request.POST.get("luogo", "").strip()
        note_sess = request.POST.get("note", "").strip()

        try:
            data_svolgimento = date.fromisoformat(data_str)
        except (ValueError, TypeError):
            messages.error(request, "Data non valida.")
            return redirect("anagrafica:visite_mediche_nuova_sessione")
        if data_svolgimento > oggi:
            messages.error(request, "La data di svolgimento non può essere nel futuro.")
            return redirect("anagrafica:visite_mediche_nuova_sessione")

        # Righe selezionate: campo sel_<legacy>_<tipo> presente.
        selezioni: list[tuple[int, int]] = []
        for key in request.POST.keys():
            if not key.startswith("sel_"):
                continue
            parts = key.split("_")
            if len(parts) != 3:
                continue
            try:
                selezioni.append((int(parts[1]), int(parts[2])))
            except (ValueError, TypeError):
                continue
        if not selezioni:
            messages.warning(request, "Nessuna visita selezionata.")
            return redirect("anagrafica:visite_mediche_nuova_sessione")

        from .models import VisitaSessione
        sess = VisitaSessione.objects.create(
            data_svolgimento=data_svolgimento, medico_competente=medico,
            luogo=luogo, note=note_sess, created_by=request.user,
        )

        tipi_cache: dict[int, TipoVisitaMedica] = {}
        creati = 0
        doppioni = 0
        errori = []
        for legacy_id, tipo_id in selezioni:
            tipo = tipi_cache.get(tipo_id)
            if tipo is None:
                tipo = TipoVisitaMedica.objects.filter(pk=tipo_id, is_active=True).first()
                tipi_cache[tipo_id] = tipo
            if tipo is None:
                continue
            esito = request.POST.get(f"esito_{legacy_id}_{tipo_id}", VisitaMedica.Esito.IDONEO)
            if esito not in VisitaMedica.Esito.values:
                esito = VisitaMedica.Esito.IDONEO
            prescrizioni = request.POST.get(f"prescrizioni_{legacy_id}_{tipo_id}", "").strip()
            note = request.POST.get(f"note_{legacy_id}_{tipo_id}", "").strip()
            try:
                if VisitaMedica.objects.filter(
                    legacy_anagrafica_id=legacy_id, tipo=tipo,
                    data_svolgimento=data_svolgimento,
                ).exists():
                    doppioni += 1
                    continue
                visita = VisitaMedica.objects.create(
                    legacy_anagrafica_id=legacy_id, tipo=tipo,
                    data_svolgimento=data_svolgimento, esito=esito,
                    prescrizioni=prescrizioni, note=note, medico_competente=medico,
                    sessione=sess, created_by=request.user, updated_by=request.user,
                )
                referto_file = request.FILES.get(f"referto_{legacy_id}_{tipo_id}")
                if referto_file:
                    _salva_referto_visita(request, visita, referto_file)
                creati += 1
            except Exception:
                logger.exception("Errore creazione VisitaMedica giornata legacy=%s tipo=%s", legacy_id, tipo_id)
                errori.append(f"{legacy_id}/{tipo_id}")

        if creati == 0:
            sess.delete()  # nessuna visita creata: non lasciare sessioni vuote

        try:
            from core.audit import log_action
            log_action(
                request, "VISITA_MEDICA_BATCH_CREATA", "anagrafica",
                f"Giornata visite del {data_svolgimento} (sessione {sess.pk if creati else '—'}): "
                f"{creati} visite registrate, {doppioni} doppioni saltati.",
            )
        except Exception:
            logger.warning("Audit VISITA_MEDICA_BATCH_CREATA fallito", exc_info=True)

        if errori:
            messages.warning(request, f"{creati} visite registrate. Errori: {', '.join(errori)}.")
        elif creati == 0:
            messages.info(request, "Nessuna visita registrata (tutte già presenti in pari data).")
            return redirect("anagrafica:visite_mediche_nuova_sessione")
        else:
            msg = f"Giornata del {data_svolgimento.strftime('%d-%m-%Y')}: {creati} visite registrate."
            if doppioni:
                msg += f" {doppioni} già presenti in pari data: saltate."
            messages.success(request, msg)
        return redirect("anagrafica:visite_mediche_sessione_detail", sessione_id=sess.pk)

    # ---- GET: pagina Giornata visite (reattiva) --------------------------
    pre_tipo_id = (request.GET.get("tipo") or "").strip()
    tipo_pre = None
    if pre_tipo_id.isdigit():
        tipo_pre = TipoVisitaMedica.objects.filter(pk=int(pre_tipo_id), is_active=True).first()
    righe = (
        _build_candidati_giornata(oggi, tipo_id=(tipo_pre.pk if tipo_pre else None))
        if (tipo_pre or request.GET.get("all")) else []
    )
    medici_precedenti = list(
        VisitaMedica.objects.exclude(medico_competente="")
        .order_by("medico_competente").values_list("medico_competente", flat=True).distinct()[:20]
    )
    return render(request, "anagrafica/pages/visite_mediche_nuova_sessione.html", {
        "tipi_attivi": tipi_attivi,
        "tipo_pre": tipo_pre,
        "righe": righe,
        "n_pre": sum(1 for r in righe if r["preselect"]),
        "oggi": oggi,
        "esiti": VisitaMedica.Esito.choices,
        "esito_default": VisitaMedica.Esito.IDONEO,
        "medici_precedenti": medici_precedenti,
    })


# ---------------------------------------------------------------------------
# API: ricerca live dipendente per aggiunta manuale in sessione
# ---------------------------------------------------------------------------

@login_required
def visite_mediche_api_cerca_dipendente(request):
    """Ricerca live dipendenti per il popup '+Aggiungi dipendente' nella sessione.

    GET ?q=QUERY&exclude=ID1,ID2,...&tipo_id=N
    Ritorna JSON {results: [{legacy_id, nome, pertinente}, ...]}.

    ``pertinente`` = il tipo è richiesto al dipendente da ruoli operativi o
    processi MOD.128; se il tipo non ha vincoli configurati la pertinenza non
    è valutabile e vale sempre ``true``. I cessati non compaiono mai.
    """
    if not _can_view_visite_mediche(request):
        return JsonResponse({"error": "Forbidden"}, status=403)

    q = request.GET.get("q", "").strip()
    if len(q) < 2:
        return JsonResponse({"results": []})

    exclude_ids: set[int] = set()
    for s in request.GET.get("exclude", "").split(","):
        try:
            exclude_ids.add(int(s.strip()))
        except (ValueError, TypeError):
            pass
    exclude_ids |= _cessati_legacy_ids()

    pertinenti: set[int] | None = None
    tipo_id_raw = request.GET.get("tipo_id", "").strip()
    if tipo_id_raw:
        try:
            tipo = TipoVisitaMedica.objects.get(pk=int(tipo_id_raw))
        except (TipoVisitaMedica.DoesNotExist, ValueError, TypeError):
            tipo = None
        if tipo is not None:
            req = _requisiti_tipo_visita(tipo)
            if req["ha_vincoli"]:
                pertinenti = req["da_ruoli"] | req["da_processi"]

    qs = (
        AnagraficaDipendente.objects
        .filter(Q(cognome__icontains=q) | Q(nome__icontains=q))
        .exclude(id__in=exclude_ids)
        .order_by("cognome", "nome")[:25]
    )
    results = []
    for d in qs:
        cog = (getattr(d, "cognome", "") or "").strip()
        nom = (getattr(d, "nome", "") or "").strip()
        results.append({
            "legacy_id": d.id,
            "nome": f"{cog} {nom}".strip() or f"#{d.id}",
            "pertinente": True if pertinenti is None else (d.id in pertinenti),
        })
    return JsonResponse({"results": results})


@login_required
def visite_mediche_candidati(request):
    """Partial HTMX: righe candidate della giornata per il tipo selezionato
    (o tutti i tipi se ``tipo`` assente). Popola la tabella senza reload."""
    if not _can_view_visite_mediche(request):
        return HttpResponse(status=403)
    from django.utils import timezone as _tz
    raw = (request.GET.get("tipo") or "").strip()
    tipo_id = int(raw) if raw.isdigit() else None
    righe = _build_candidati_giornata(_tz.localdate(), tipo_id=tipo_id)
    return render(request, "anagrafica/partials/_visite_candidati.html", {
        "righe": righe,
        "esiti": VisitaMedica.Esito.choices,
        "esito_default": VisitaMedica.Esito.IDONEO,
        "n_pre": sum(1 for r in righe if r["preselect"]),
    })


# ---------------------------------------------------------------------------
# Sessioni visite (VisitaSessione): hub proposte, dettaglio, aggiungi, elimina
# ---------------------------------------------------------------------------

@login_required
def visite_mediche_sessioni(request):
    """Hub: proposte di rinnovo per tipo (da rinnovare = ultima visita corrente
    scaduta/in scadenza) + elenco delle giornate salvate."""
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per le visite mediche.")
        return redirect("anagrafica:index")
    from django.utils import timezone as _tz
    from .models import VisitaSessione
    oggi = _tz.localdate()
    soglia = oggi + _timedelta(days=60)
    correnti = VisitaMedica.objects.filter(id__in=ultime_visite_correnti_ids())
    proposte = []
    for t in TipoVisitaMedica.objects.filter(is_active=True).order_by("nome"):
        n_scad = correnti.filter(tipo=t, data_scadenza__lt=oggi).count()
        n_insc = correnti.filter(tipo=t, data_scadenza__gte=oggi, data_scadenza__lte=soglia).count()
        if n_scad or n_insc:
            proposte.append({"tipo": t, "n_scadute": n_scad, "n_in_scadenza": n_insc})
    sessioni = list(
        VisitaSessione.objects.annotate(n_visite=Count("visite")).order_by("-data_svolgimento", "-id")[:50]
    )
    return render(request, "anagrafica/pages/visite_mediche_sessioni.html", {
        "proposte": proposte, "sessioni": sessioni, "oggi": oggi,
        "tot_da_rinnovare": sum(p["n_scadute"] + p["n_in_scadenza"] for p in proposte),
    })


@login_required
def visite_mediche_sessione_detail(request, sessione_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Non hai i permessi per le visite mediche.")
        return redirect("anagrafica:index")
    from .models import VisitaSessione
    sess = get_object_or_404(VisitaSessione, pk=sessione_id)
    nomi = _build_nomi_map()
    visite = list(sess.visite.select_related("tipo").order_by("tipo__nome"))
    for v in visite:
        v.dipendente_nome = nomi.get(v.legacy_anagrafica_id, f"#{v.legacy_anagrafica_id}")
    _, is_admin = _ensure_admin(request)
    return render(request, "anagrafica/pages/visite_mediche_sessione_detail.html", {
        "sess": sess, "visite": visite, "is_admin": is_admin,
        "tipi_attivi": list(TipoVisitaMedica.objects.filter(is_active=True).order_by("nome")),
        "esiti": VisitaMedica.Esito.choices, "esito_default": VisitaMedica.Esito.IDONEO,
    })


@login_required
@require_POST
def visite_mediche_sessione_partecipante_add(request, sessione_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Permessi insufficienti.")
        return redirect("anagrafica:visite_mediche_sessione_detail", sessione_id=sessione_id)
    from .models import VisitaSessione
    sess = get_object_or_404(VisitaSessione, pk=sessione_id)
    try:
        legacy_id = int(request.POST.get("legacy_id") or 0)
        tipo = TipoVisitaMedica.objects.get(pk=request.POST.get("tipo_id"), is_active=True)
    except (ValueError, TypeError, TipoVisitaMedica.DoesNotExist):
        messages.error(request, "Dipendente o tipo non validi.")
        return redirect("anagrafica:visite_mediche_sessione_detail", sessione_id=sessione_id)
    esito = request.POST.get("esito", VisitaMedica.Esito.IDONEO)
    if esito not in VisitaMedica.Esito.values:
        esito = VisitaMedica.Esito.IDONEO
    if VisitaMedica.objects.filter(
        legacy_anagrafica_id=legacy_id, tipo=tipo, data_svolgimento=sess.data_svolgimento
    ).exists():
        messages.warning(request, "Visita già presente in pari data: non aggiunta.")
    else:
        VisitaMedica.objects.create(
            legacy_anagrafica_id=legacy_id, tipo=tipo, data_svolgimento=sess.data_svolgimento,
            esito=esito, medico_competente=sess.medico_competente, sessione=sess,
            created_by=request.user, updated_by=request.user,
        )
        messages.success(request, "Partecipante aggiunto alla giornata.")
    return redirect("anagrafica:visite_mediche_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def visite_mediche_sessione_delete(request, sessione_id: int):
    if not _can_view_visite_mediche(request):
        messages.error(request, "Permessi insufficienti.")
        return redirect("anagrafica:visite_mediche_sessioni")
    _, is_admin = _ensure_admin(request)
    if not is_admin:
        messages.error(request, "Solo gli amministratori possono eliminare una giornata.")
        return redirect("anagrafica:visite_mediche_sessione_detail", sessione_id=sessione_id)
    from .models import VisitaSessione
    sess = get_object_or_404(VisitaSessione, pk=sessione_id)
    sess.delete()  # SET_NULL: le visite restano
    messages.success(request, "Giornata eliminata (le visite registrate sono conservate).")
    return redirect("anagrafica:visite_mediche_sessioni")


# ---------------------------------------------------------------------------
# Export Excel: copertura per tipologia e visite in scadenza
# ---------------------------------------------------------------------------

@login_required
def visite_mediche_export_scadenze(request):
    """Scarica le visite scadute/in scadenza come file .xlsx.

    Accetta ?scad=mese_corrente | prossimo_mese (default: tutti ≤60g).
    """
    if not _can_view_visite_mediche(request):
        return HttpResponse(status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl non installato.", status=500)

    from core.excel_export import append_row

    import calendar as _cal
    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    filtro = request.GET.get("scad", "tutti").strip()
    correnti_ids = ultime_visite_correnti_ids()

    if filtro == "mese_corrente":
        _, ld = _cal.monthrange(oggi.year, oggi.month)
        qs = VisitaMedica.objects.filter(
            id__in=correnti_ids,
            data_scadenza__isnull=False,
            data_scadenza__range=[oggi.replace(day=1), oggi.replace(day=ld)],
        )
        label = f"mese_corrente_{oggi.strftime('%Y%m')}"
    elif filtro == "prossimo_mese":
        pm_y = oggi.year + 1 if oggi.month == 12 else oggi.year
        pm_m = 1 if oggi.month == 12 else oggi.month + 1
        _, ld = _cal.monthrange(pm_y, pm_m)
        qs = VisitaMedica.objects.filter(
            id__in=correnti_ids,
            data_scadenza__isnull=False,
            data_scadenza__range=[date(pm_y, pm_m, 1), date(pm_y, pm_m, ld)],
        )
        label = f"prossimo_mese_{pm_y}{pm_m:02d}"
    else:
        qs = VisitaMedica.objects.filter(
            id__in=correnti_ids,
            data_scadenza__isnull=False, data_scadenza__lte=oggi + _timedelta(days=60)
        )
        label = "scadenze_60gg"

    qs = qs.select_related("tipo").order_by("data_scadenza")
    nomi_map = _build_nomi_map()

    # CF lookup
    cf_map: dict[int, str] = dict(
        DipendenteAnagraficaCivile.objects
        .exclude(codice_fiscale="")
        .values_list("legacy_anagrafica_id", "codice_fiscale")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scadenze"

    header = ["Dipendente", "Codice Fiscale", "Tipo visita", "Data svolgimento",
              "Data scadenza", "Giorni a scadenza", "Esito", "Note"]
    append_row(ws, header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for v in qs:
        nome = nomi_map.get(v.legacy_anagrafica_id, f"#{v.legacy_anagrafica_id}")
        cf = cf_map.get(v.legacy_anagrafica_id, "")
        giorni = (v.data_scadenza - oggi).days if v.data_scadenza else ""
        append_row(ws, [
            nome, cf, v.tipo.nome,
            v.data_svolgimento.strftime("%d-%m-%Y") if v.data_svolgimento else "",
            v.data_scadenza.strftime("%d-%m-%Y") if v.data_scadenza else "",
            giorni,
            v.get_esito_display(),
            v.prescrizioni or "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="visite_scadenze_{label}.xlsx"'
    wb.save(response)
    return response


@login_required
def visite_mediche_export_copertura(request):
    """Scarica la tabella copertura per tipologia come .xlsx."""
    if not _can_view_visite_mediche(request):
        return HttpResponse(status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl non installato.", status=500)

    from core.excel_export import append_row

    from django.utils import timezone as _tz

    oggi = _tz.localdate()
    correnti_qs = VisitaMedica.objects.filter(id__in=ultime_visite_correnti_ids())

    _valide_per_tipo = {
        row["tipo_id"]: row["n"]
        for row in correnti_qs
        .filter(data_scadenza__gte=oggi)
        .order_by()
        .values("tipo_id")
        .annotate(n=Count("legacy_anagrafica_id", distinct=True))
    }
    _scadute_per_tipo = {
        row["tipo_id"]: row["n"]
        for row in correnti_qs
        .filter(data_scadenza__lt=oggi)
        .order_by()
        .values("tipo_id")
        .annotate(n=Count("legacy_anagrafica_id", distinct=True))
    }

    tipi = TipoVisitaMedica.objects.filter(is_active=True).prefetch_related("ruoli_operativi")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Copertura tipologie"

    header = ["Tipologia", "Periodicità (mesi)", "Obbligatoria",
              "Ruoli collegati", "Richiesti (ruoli)", "Coperti (ruoli)", "Mancanti (ruoli)",
              "Valide (DB)", "Scadute (DB)"]
    append_row(ws, header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for t in tipi:
        ruoli_ids = list(t.ruoli_operativi.values_list("id", flat=True))
        if ruoli_ids:
            legacy_ids_richiesti = set(
                DipendenteRuoloOperativo.objects
                .filter(ruolo_id__in=ruoli_ids)
                .values_list("legacy_anagrafica_id", flat=True)
            )
        else:
            legacy_ids_richiesti = set()
        legacy_ids_coperti = set(
            correnti_qs
            .filter(tipo=t, data_scadenza__gte=oggi)
            .values_list("legacy_anagrafica_id", flat=True)
        )
        mancanti = legacy_ids_richiesti - legacy_ids_coperti
        ruoli_nomi = ", ".join(t.ruoli_operativi.values_list("nome", flat=True)) or "—"

        append_row(ws, [
            t.nome, t.durata_mesi, "Sì" if t.obbligatoria else "No",
            ruoli_nomi,
            len(legacy_ids_richiesti),
            len(legacy_ids_richiesti & legacy_ids_coperti),
            len(mancanti),
            _valide_per_tipo.get(t.pk, 0),
            _scadute_per_tipo.get(t.pk, 0),
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(c.value or "")) for c in col
        ) + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="visite_copertura_tipologie.xlsx"'
    wb.save(response)
    return response


# ============================================================================
# FORMAZIONE HR — Dashboard (PATCH-02)
# ============================================================================

@login_required
def formazione_dashboard(request):
    """Dashboard globale formazione HR.

    KPI letti da TrainingDeadline (cache ricalcolabile) e TrainingEmployeeRecord.
    Se TrainingDeadline è vuota (non ancora ricalcolata) i KPI scadenze sono 0
    e viene mostrato un banner informativo.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    from django.utils import timezone as _tz

    oggi = _tz.localdate()

    # KPI da TrainingEmployeeRecord (non dipende dalla cache)
    kpi_dipendenti_formazione = (
        TrainingEmployeeRecord.objects
        .values("legacy_anagrafica_id").distinct().count()
    )

    # KPI da TrainingDeadline (cache — può essere 0 se non ancora ricalcolata)
    kpi_scaduti         = TrainingDeadline.objects.filter(stato_scadenza="SCADUTO").count()
    kpi_in_scadenza_30  = TrainingDeadline.objects.filter(stato_scadenza="IN_SCADENZA_30").count()
    kpi_obbligatori_mancanti = TrainingDeadline.objects.filter(
        stato_scadenza__in=["MAI_FREQUENTATO", "SCADUTO"],
        is_required=True,
    ).count()

    # KPI catalogo
    kpi_piani_attivi = TrainingPlan.objects.filter(stato="ATTIVO", is_active=True).count()
    kpi_corsi_attivi = TrainingCourse.objects.filter(stato="ATTIVO", is_active=True).count()

    # Scadenzario urgente TOP 20 (ordinato per urgenza crescente)
    scadenze_raw = list(
        TrainingDeadline.objects
        .filter(stato_scadenza__in=["SCADUTO", "IN_SCADENZA_30", "IN_SCADENZA_90"])
        .select_related("corso", "corso__piano")
        .order_by("data_scadenza", "legacy_anagrafica_id")[:20]
    )
    nomi_map = _build_nomi_map()
    for d in scadenze_raw:
        d.dipendente_nome = nomi_map.get(d.legacy_anagrafica_id, f"#{d.legacy_anagrafica_id}")
        d.giorni = (d.data_scadenza - oggi).days if d.data_scadenza else None

    deadline_cache_empty = not TrainingDeadline.objects.exists()

    return render(request, "anagrafica/pages/formazione_dashboard.html", {
        "oggi": oggi,
        "can_edit_formazione": _can_edit_formazione(request),
        "kpi_dipendenti_formazione": kpi_dipendenti_formazione,
        "kpi_scaduti": kpi_scaduti,
        "kpi_in_scadenza_30": kpi_in_scadenza_30,
        "kpi_obbligatori_mancanti": kpi_obbligatori_mancanti,
        "kpi_piani_attivi": kpi_piani_attivi,
        "kpi_corsi_attivi": kpi_corsi_attivi,
        "scadenze_urgenti": scadenze_raw,
        "deadline_cache_empty": deadline_cache_empty,
    })


@login_required
def formazione_ricerca(request):
    """Ricerca globale formazione: corsi, sessioni, piani, qualifiche,
    dipendenti (libretto) e attestati. Un'unica casella dalla dashboard.

    Best-effort, limitata in righe: serve a "saltare" velocemente all'entità,
    non è un motore full-text. Gated dal permesso di visualizzazione formazione.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    q = (request.GET.get("q") or "").strip()
    risultati: dict[str, list] = {
        "corsi": [], "sessioni": [], "piani": [], "qualifiche": [],
        "dipendenti": [], "attestati": [],
    }
    totale = 0
    if q:
        ql = q.lower()

        from django.db.models import Count as _Count, Prefetch as _Prefetch
        # Edizioni + lezioni precaricate per il sottoelenco espandibile,
        # sia in pagina risultati sia nella tendina (no query N+1).
        risultati["corsi"] = list(
            TrainingCourse.objects.select_related("piano")
            .filter(Q(titolo__icontains=q) | Q(codice__icontains=q))
            .annotate(n_sessioni=_Count("sessioni", distinct=True))
            .prefetch_related(
                _Prefetch(
                    "sessioni",
                    queryset=TrainingSession.objects.prefetch_related("lezioni").order_by("-data_inizio"),
                )
            )
            .order_by("titolo")[:25]
        )
        risultati["piani"] = list(
            TrainingPlan.objects.filter(Q(nome__icontains=q) | Q(codice__icontains=q))
            .annotate(n_corsi=_Count("corsi", distinct=True))
            .order_by("nome")[:25]
        )
        risultati["sessioni"] = list(
            TrainingSession.objects.select_related("corso")
            .filter(Q(codice_sessione__icontains=q) | Q(corso__titolo__icontains=q) | Q(sede__icontains=q))
            .annotate(n_iscritti=_Count("iscrizioni", distinct=True))
            .order_by("-data_inizio")[:25]
        )
        risultati["qualifiche"] = list(
            TipoQualifica.objects.filter(Q(nome__icontains=q) | Q(descrizione__icontains=q))
            .annotate(n_assegnazioni=_Count("assegnazioni", distinct=True))
            .order_by("nome")[:25]
        )

        # Dipendenti per nome → link al libretto formativo.
        nomi = _build_nomi_map()
        dip_match = [
            {"legacy_id": lid, "nome": nome}
            for lid, nome in nomi.items() if ql in nome.lower()
        ]
        dip_match.sort(key=lambda d: d["nome"].casefold())
        risultati["dipendenti"] = dip_match[:25]

        # Attestati: per nome dipendente o titolo corso (snapshot).
        matched_ids = [d["legacy_id"] for d in dip_match]
        rec_qs = TrainingEmployeeRecord.objects.filter(
            Q(course_title_snapshot__icontains=q)
            | Q(course_code_snapshot__icontains=q)
            | Q(legacy_anagrafica_id__in=matched_ids)
        ).order_by("-data_completamento")[:25]
        att = []
        for r in rec_qs:
            att.append({
                "record": r,
                "nome": nomi.get(r.legacy_anagrafica_id, f"#{r.legacy_anagrafica_id}"),
            })
        risultati["attestati"] = att

        totale = sum(len(v) for v in risultati.values())

    # Ricerca live: su richiesta HTMX rende solo il frammento dei risultati;
    # con ?suggest=1 (casella nella dashboard) rende la tendina compatta.
    if request.headers.get("HX-Request"):
        template = (
            "anagrafica/partials/_formazione_search_suggest.html"
            if request.GET.get("suggest")
            else "anagrafica/partials/_formazione_search_results.html"
        )
    else:
        template = "anagrafica/pages/formazione_ricerca.html"
    return render(request, template, {
        "q": q,
        "risultati": risultati,
        "totale": totale,
    })


def _can_edit_formazione(request) -> bool:
    """Verifica permesso di modifica sezione formazione (modifica catalogo piani/corsi/istruttori)."""
    if request.user.is_superuser:
        return True
    perm = AnagraficaFormazionePermission.get_instance()
    if perm.accesso_modifica == AnagraficaFormazionePermission.ACCESSO_TUTTI:
        return True
    legacy_user = get_legacy_user(request.user)
    if is_legacy_admin(legacy_user):
        return perm.accesso_modifica in (
            AnagraficaFormazionePermission.ACCESSO_ADMIN,
            AnagraficaFormazionePermission.ACCESSO_TUTTI,
        )
    if perm.accesso_modifica == AnagraficaFormazionePermission.ACCESSO_ADMIN:
        return False
    if legacy_user and legacy_user.ruolo_id is not None:
        return int(legacy_user.ruolo_id) in [int(r) for r in (perm.ruoli_autorizzati_json or [])]
    return False


# ============================================================================
# FORMAZIONE HR — Piani formativi (PATCH-03)
# ============================================================================

@login_required
def formazione_piani_list(request):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    filtro_stato = request.GET.get("stato", "")
    filtro_cat   = request.GET.get("categoria", "")
    qs = TrainingPlan.objects.all()
    if filtro_stato:
        qs = qs.filter(stato=filtro_stato)
    if filtro_cat:
        qs = qs.filter(categoria=filtro_cat)
    piani = list(qs.order_by("nome").annotate(n_corsi=Count("corsi")))

    form = TrainingPlanForm()
    return render(request, "anagrafica/pages/formazione_piani.html", {
        "piani": piani,
        "form": form,
        "is_editor": is_editor,
        "filtro_stato": filtro_stato,
        "filtro_cat": filtro_cat,
        "STATO_CHOICES": TrainingPlan.STATO_CHOICES,
        "CATEGORIA_CHOICES": TrainingPlan.CATEGORIA_CHOICES,
    })


@login_required
@require_POST
def formazione_piano_create(request):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per creare piani formativi.")
        return redirect("anagrafica:formazione_piani_list")
    form = TrainingPlanForm(request.POST)
    if form.is_valid():
        piano = form.save(commit=False)
        piano.created_by = request.user
        piano.save()
        messages.success(request, f'Piano "{piano.nome}" creato.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_piani_list")


@login_required
def formazione_piano_detail(request, piano_id: int):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    piano = get_object_or_404(TrainingPlan, pk=piano_id)
    corsi = list(
        TrainingCourse.objects.filter(piano=piano)
        .order_by("titolo")
        .annotate(n_sessioni=Count("sessioni", distinct=True))
    )
    regole = list(
        TrainingRequirementRule.objects.filter(piano=piano, is_active=True)
        .select_related("mansione", "area", "ruolo_operativo")
        .order_by("-priority")
    )
    nomi_singoli: dict[int, str] = {}
    ids_singoli = [r.legacy_anagrafica_id for r in regole if r.legacy_anagrafica_id]
    if ids_singoli:
        nomi_singoli = _build_nomi_map()
    for r in regole:
        r.dipendente_nome = nomi_singoli.get(r.legacy_anagrafica_id, "") if r.legacy_anagrafica_id else ""

    edit_form = TrainingPlanForm(instance=piano)
    return render(request, "anagrafica/pages/formazione_piano_detail.html", {
        "piano": piano,
        "corsi": corsi,
        "regole": regole,
        "edit_form": edit_form,
        "is_editor": is_editor,
    })


@login_required
@require_POST
def formazione_piano_edit(request, piano_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare piani formativi.")
        return redirect("anagrafica:formazione_piani_list")
    piano = get_object_or_404(TrainingPlan, pk=piano_id)
    form = TrainingPlanForm(request.POST, instance=piano)
    if form.is_valid():
        form.save()
        messages.success(request, f'Piano "{piano.nome}" aggiornato.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_piano_detail", piano_id=piano_id)


@login_required
@require_POST
def formazione_piano_delete(request, piano_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per eliminare piani formativi.")
        return redirect("anagrafica:formazione_piani_list")
    piano = get_object_or_404(TrainingPlan, pk=piano_id)
    if piano.corsi.exists():
        messages.error(request, f'Il piano "{piano.nome}" ha corsi associati. Archivia o rimuovi prima i corsi.')
        return redirect("anagrafica:formazione_piano_detail", piano_id=piano_id)
    nome = piano.nome
    piano.delete()
    messages.success(request, f'Piano "{nome}" eliminato.')
    return redirect("anagrafica:formazione_piani_list")


# ============================================================================
# FORMAZIONE HR — Corsi (PATCH-03)
# ============================================================================

@login_required
def formazione_corsi_list(request):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    filtro_piano = request.GET.get("piano", "")
    filtro_stato = request.GET.get("stato", "")
    filtro_obbligatorio = request.GET.get("obbligatorio", "")
    q_search = (request.GET.get("q") or "").strip()

    qs = TrainingCourse.objects.select_related("piano").all()
    if filtro_piano:
        qs = qs.filter(piano_id=filtro_piano)
    if filtro_stato:
        qs = qs.filter(stato=filtro_stato)
    if filtro_obbligatorio == "1":
        qs = qs.filter(obbligatorio=True)
    elif filtro_obbligatorio == "0":
        qs = qs.filter(obbligatorio=False)
    if q_search:
        qs = qs.filter(Q(titolo__icontains=q_search) | Q(codice__icontains=q_search))

    paginator = Paginator(qs.order_by("piano__nome", "titolo"), 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    piani = list(TrainingPlan.objects.filter(is_active=True).order_by("nome"))

    return render(request, "anagrafica/pages/formazione_corsi.html", {
        "page_obj": page_obj,
        "is_editor": is_editor,
        "filtro_piano": filtro_piano,
        "filtro_stato": filtro_stato,
        "filtro_obbligatorio": filtro_obbligatorio,
        "q_search": q_search,
        "piani": piani,
        "STATO_CHOICES": TrainingCourse.STATO_CHOICES,
    })


@login_required
def formazione_corso_create(request):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per creare corsi formativi.")
        return redirect("anagrafica:formazione_corsi_list")
    if request.method == "POST":
        form = TrainingCourseForm(request.POST)
        if form.is_valid():
            corso = form.save(commit=False)
            corso.created_by = request.user
            corso.save()
            form.salva_processi(corso)
            messages.success(request, f'Corso "{corso.titolo}" creato.')
            return redirect("anagrafica:formazione_corso_detail", corso_id=corso.pk)
    else:
        initial = {}
        if request.GET.get("piano"):
            initial["piano"] = request.GET.get("piano")
        # Preset "nuovo corso e-learning" dal hub di gestione e-learning: applica i
        # default configurati nelle Impostazioni e-learning.
        if request.GET.get("elearning") in ("1", "true", "on"):
            cfg = ElearningConfig.get_instance()
            initial["is_elearning"] = True
            initial["quiz_punteggio_minimo"] = cfg.quiz_punteggio_minimo_default
            if cfg.validita_mesi_default:
                initial["validita_mesi"] = cfg.validita_mesi_default
        form = TrainingCourseForm(initial=initial)
    return render(request, "anagrafica/pages/formazione_corso_form.html", {
        "form": form,
        "modo": "crea",
    })


# ── Quick-add inline (JSON) delle entità collegate dei form formazione ───────
# Creano l'entità minimale e ritornano {ok, id, label} così la UI la appende al
# <select> e la seleziona senza ricaricare. Gated dal permesso di modifica.

@login_required
@require_POST
def formazione_quickadd_piano(request):
    if not _can_edit_formazione(request):
        return JsonResponse({"ok": False, "error": "Permesso negato."}, status=403)
    codice = (request.POST.get("codice") or "").strip().upper()[:20]
    nome = (request.POST.get("nome") or "").strip()[:200]
    if not codice or not nome:
        return JsonResponse({"ok": False, "error": "Codice e nome sono obbligatori."}, status=400)
    if TrainingPlan.objects.filter(codice=codice).exists():
        return JsonResponse({"ok": False, "error": f"Esiste già un piano con codice {codice}."}, status=400)
    p = TrainingPlan.objects.create(codice=codice, nome=nome, created_by=request.user)
    return JsonResponse({"ok": True, "id": p.pk, "label": str(p)})


@login_required
@require_POST
def formazione_quickadd_categoria(request):
    if not _can_edit_formazione(request):
        return JsonResponse({"ok": False, "error": "Permesso negato."}, status=403)
    from .models_rischi import CategoriaCorso
    codice = (request.POST.get("codice") or "").strip().upper()[:20]
    nome = (request.POST.get("nome") or "").strip()[:200]
    if not codice or not nome:
        return JsonResponse({"ok": False, "error": "Codice e nome sono obbligatori."}, status=400)
    if CategoriaCorso.objects.filter(codice=codice).exists():
        return JsonResponse({"ok": False, "error": f"Esiste già una categoria con codice {codice}."}, status=400)
    c = CategoriaCorso.objects.create(codice=codice, nome=nome)
    return JsonResponse({"ok": True, "id": c.pk, "label": c.nome})


@login_required
@require_POST
def formazione_quickadd_qualifica(request):
    if not _can_edit_formazione(request):
        return JsonResponse({"ok": False, "error": "Permesso negato."}, status=403)
    from .models import TipoQualifica
    nome = (request.POST.get("nome") or "").strip()[:150]
    if not nome:
        return JsonResponse({"ok": False, "error": "Il nome è obbligatorio."}, status=400)
    try:
        durata = max(0, int(request.POST.get("durata_mesi") or 0))
    except (TypeError, ValueError):
        durata = 0
    obj, _created = TipoQualifica.objects.get_or_create(nome=nome, defaults={"durata_mesi": durata})
    return JsonResponse({"ok": True, "id": obj.pk, "label": obj.nome})


@login_required
@require_POST
def formazione_quickadd_docente(request):
    if not _can_edit_formazione(request):
        return JsonResponse({"ok": False, "error": "Permesso negato."}, status=403)
    nome = (request.POST.get("nome") or "").strip()[:200]
    if not nome:
        return JsonResponse({"ok": False, "error": "Il nome è obbligatorio."}, status=400)
    tipo = (request.POST.get("tipo") or "ESTERNO").strip().upper()
    if tipo not in ("INTERNO", "ESTERNO"):
        tipo = "ESTERNO"
    d = TrainingInstructor.objects.create(nome=nome, tipo=tipo)
    return JsonResponse({"ok": True, "id": d.pk, "label": d.nome})


@login_required
def formazione_corso_codice_suggest(request):
    """Suggerisce un codice corso UNIVOCO a partire dal titolo (JSON).

    Base: alfanumerico maiuscolo del titolo (iniziali delle parole se lungo);
    se già usato accoda un progressivo. Usato dal form corso per precompilare il
    codice quando l'utente non lo digita."""
    if not _can_edit_formazione(request):
        return JsonResponse({"ok": False}, status=403)
    import re
    titolo = (request.GET.get("titolo") or "").strip()
    parole = re.findall(r"[A-Za-z0-9]+", titolo)
    if not parole:
        base = "CORSO"
    elif len("".join(parole)) <= 8:
        base = "".join(parole).upper()
    else:
        base = ("".join(p[0] for p in parole).upper() or parole[0].upper())
    base = (base or "CORSO")[:12]
    codice, i = base, 1
    while TrainingCourse.objects.filter(codice=codice).exists():
        i += 1
        codice = f"{base}-{i}"[:30]
    return JsonResponse({"ok": True, "codice": codice})


@login_required
def formazione_qualifica_durata(request):
    """Durata (mesi) di una TipoQualifica (JSON): il form corso preimposta la
    validità del corso da quella della qualifica àncora selezionata."""
    if not _can_view_formazione(request):
        return JsonResponse({"ok": False}, status=403)
    from .models import TipoQualifica
    try:
        q = TipoQualifica.objects.get(pk=int(request.GET.get("id") or 0))
    except (TipoQualifica.DoesNotExist, TypeError, ValueError):
        return JsonResponse({"ok": False}, status=404)
    return JsonResponse({"ok": True, "durata_mesi": q.durata_mesi or 0})


@login_required
def formazione_corso_detail(request, corso_id: int):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    corso = get_object_or_404(TrainingCourse.objects.select_related("piano"), pk=corso_id)

    prerequisiti = list(
        TrainingCourseDependency.objects.filter(corso_principale=corso)
        .select_related("prerequisito")
    )
    moduli = list(
        TrainingCourseModule.objects.filter(corso_padre=corso)
        .select_related("corso_modulo")
        .order_by("ordine")
    )
    versioni = list(
        TrainingCourseVersion.objects.filter(corso=corso).order_by("-created_at")
    )
    regole = list(
        TrainingRequirementRule.objects.filter(corso=corso, is_active=True)
        .select_related("mansione", "area", "ruolo_operativo")
        .order_by("-priority")
    )
    nomi_singoli: dict[int, str] = {}
    ids_singoli = [r.legacy_anagrafica_id for r in regole if r.legacy_anagrafica_id]
    if ids_singoli:
        nomi_singoli = _build_nomi_map()
    for r in regole:
        r.dipendente_nome = nomi_singoli.get(r.legacy_anagrafica_id, "") if r.legacy_anagrafica_id else ""

    try:
        completion_rule = corso.regola_superamento
    except TrainingCompletionRule.DoesNotExist:
        completion_rule = None

    n_sessioni = corso.sessioni.count() if hasattr(corso, "sessioni") else 0
    n_completamenti = TrainingEmployeeRecord.objects.filter(corso=corso).count()

    # Lista sessioni del corso con conteggi lezioni/iscritti per la sezione
    # "Sessioni del corso" nel dettaglio. Limita a 100 per evitare pagine enormi
    # su corsi storici molto frequentati (ordine: piu recenti prima).
    sessioni_list = list(
        corso.sessioni
        .select_related("docente")
        .annotate(n_lezioni=Count("lezioni", distinct=True),
                  n_iscritti=Count("iscrizioni", distinct=True))
        .order_by("-data_inizio")[:100]
    )

    # Aggregazione dipendenti iscritti distinti (across all sessioni del corso).
    # Per ciascuno: n. sessioni, stato sintetico, ultimo completamento, idoneo.
    from collections import defaultdict
    agg_dip = defaultdict(lambda: {
        "n_sessioni": 0,
        "stati": set(),
        "data_completamento": None,
        "idoneo": None,
        "n_completati": 0,
    })
    for e in (TrainingEnrollment.objects
              .filter(sessione__corso=corso)
              .values("legacy_anagrafica_id", "stato")):
        a = agg_dip[e["legacy_anagrafica_id"]]
        a["n_sessioni"] += 1
        a["stati"].add(e["stato"])
    for r in (TrainingEmployeeRecord.objects
              .filter(corso=corso)
              .order_by("-data_completamento")
              .values("legacy_anagrafica_id", "data_completamento", "idoneo")):
        a = agg_dip[r["legacy_anagrafica_id"]]
        a["n_completati"] += 1
        if a["data_completamento"] is None:
            a["data_completamento"] = r["data_completamento"]
            a["idoneo"] = r["idoneo"]

    nomi_dipendenti = _build_nomi_map() if agg_dip else {}
    # Ordine sintetico stati per scegliere il più rappresentativo
    _stato_priority = ["COMPLETATO", "IN_CORSO", "ISCRITTO", "NON_IDONEO", "ASSENTE", "RITIRATO"]
    dipendenti_iscritti = []
    for lid, a in agg_dip.items():
        stato_sint = next((s for s in _stato_priority if s in a["stati"]), next(iter(a["stati"]), ""))
        dipendenti_iscritti.append({
            "legacy_id": lid,
            "nome": nomi_dipendenti.get(lid, f"#{lid}"),
            "n_sessioni": a["n_sessioni"],
            "n_completati": a["n_completati"],
            "stato_sint": stato_sint,
            "data_completamento": a["data_completamento"],
            "idoneo": a["idoneo"],
        })
    dipendenti_iscritti.sort(key=lambda x: x["nome"].lower())
    n_dipendenti_iscritti = len(dipendenti_iscritti)

    edit_form = TrainingCourseForm(instance=corso)
    dep_form = TrainingCourseDependencyForm(corso_principale=corso)
    completion_form = TrainingCompletionRuleForm(instance=completion_rule)
    version_form = TrainingCourseVersionForm()
    req_rule_form = TrainingRequirementRuleForm(initial={"corso": corso})

    # Candidati all'assegnazione al corso (primo anello corso→sessione): idonei non
    # ancora assegnati. Solo per editor (evita il calcolo in sola lettura).
    candidati_assegnazione: list = []
    assegnazione_pool_filtrato = False
    if is_editor:
        from .services.training_eligibility import candidati_corso
        _res = candidati_corso(corso)
        assegnazione_pool_filtrato = _res["pool_filtrato"]
        _gia_assegnati = set(
            TrainingAssignment.objects.filter(corso=corso)
            .values_list("legacy_anagrafica_id", flat=True)
        )
        candidati_assegnazione = [
            c for c in _res["idonei"] if c["legacy_id"] not in _gia_assegnati
        ]

    return render(request, "anagrafica/pages/formazione_corso_detail.html", {
        "corso": corso,
        "prerequisiti": prerequisiti,
        "moduli": moduli,
        "versioni": versioni,
        "regole": regole,
        "completion_rule": completion_rule,
        "n_sessioni": n_sessioni,
        "n_completamenti": n_completamenti,
        "sessioni_list": sessioni_list,
        "dipendenti_iscritti": dipendenti_iscritti,
        "n_dipendenti_iscritti": n_dipendenti_iscritti,
        "edit_form": edit_form,
        "dep_form": dep_form,
        "completion_form": completion_form,
        "version_form": version_form,
        "req_rule_form": req_rule_form,
        "is_editor": is_editor,
        "tutti_corsi": TrainingCourse.objects.exclude(pk=corso_id).filter(is_active=True).order_by("titolo"),
        "candidati_assegnazione": candidati_assegnazione,
        "assegnazione_pool_filtrato": assegnazione_pool_filtrato,
    })


@login_required
@require_POST
def formazione_corso_edit(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corsi_list")
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    form = TrainingCourseForm(request.POST, instance=corso)
    if form.is_valid():
        corso = form.save()
        form.salva_processi(corso)
        messages.success(request, f'Corso "{corso.titolo}" aggiornato.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


@login_required
@require_POST
def formazione_corso_delete(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per eliminare corsi formativi.")
        return redirect("anagrafica:formazione_corsi_list")
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    if corso.sessioni.exists():
        messages.error(request, f'Il corso "{corso.titolo}" ha sessioni associate. Archivia o rimuovi prima le sessioni.')
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    piano_id = corso.piano_id
    nome = corso.titolo
    corso.delete()
    messages.success(request, f'Corso "{nome}" eliminato.')
    return redirect("anagrafica:formazione_piano_detail", piano_id=piano_id)


# ── Prerequisiti ────────────────────────────────────────────────────────────

@login_required
@require_POST
def formazione_corso_dep_add(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    form = TrainingCourseDependencyForm(request.POST, corso_principale=corso)
    if form.is_valid():
        prereq_id = form.cleaned_data["prerequisito_id"]
        prereq = get_object_or_404(TrainingCourse, pk=prereq_id)
        if prereq == corso:
            messages.error(request, "Un corso non può essere prerequisito di sé stesso.")
        else:
            _, created = TrainingCourseDependency.objects.get_or_create(
                corso_principale=corso,
                prerequisito=prereq,
                defaults={"obbligatorio": form.cleaned_data.get("obbligatorio", True)},
            )
            if created:
                messages.success(request, f'Prerequisito "{prereq.codice}" aggiunto.')
            else:
                messages.warning(request, "Prerequisito già presente.")
    else:
        messages.error(request, "Dati non validi per il prerequisito.")
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


@login_required
@require_POST
def formazione_corso_dep_delete(request, corso_id: int, dep_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    dep = get_object_or_404(TrainingCourseDependency, pk=dep_id, corso_principale_id=corso_id)
    dep.delete()
    messages.success(request, "Prerequisito rimosso.")
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


# ── Versioni corso ───────────────────────────────────────────────────────────

@login_required
@require_POST
def formazione_corso_version_add(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    form = TrainingCourseVersionForm(request.POST)
    if form.is_valid():
        ver = form.save(commit=False)
        ver.corso = corso
        ver.revised_by = request.user
        ver.save()
        messages.success(request, f'Versione "{ver.version_label}" aggiunta.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


@login_required
@require_POST
def formazione_corso_version_delete(request, corso_id: int, ver_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    ver = get_object_or_404(TrainingCourseVersion, pk=ver_id, corso_id=corso_id)
    ver.delete()
    messages.success(request, "Versione rimossa.")
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


# ── Regola superamento ───────────────────────────────────────────────────────

@login_required
@require_POST
def formazione_corso_completion_rule_save(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    try:
        instance = corso.regola_superamento
    except TrainingCompletionRule.DoesNotExist:
        instance = None
    form = TrainingCompletionRuleForm(request.POST, instance=instance)
    if form.is_valid():
        rule = form.save(commit=False)
        rule.corso = corso
        rule.save()
        messages.success(request, "Regola di superamento salvata.")
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


# ── Regole obbligatorietà sul corso ─────────────────────────────────────────

@login_required
@require_POST
def formazione_corso_req_rule_add(request, corso_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    form = TrainingRequirementRuleForm(request.POST)
    if form.is_valid():
        rule = form.save(commit=False)
        rule.corso = corso
        rule.piano = None
        rule.created_by = request.user
        rule.save()
        messages.success(request, "Regola di obbligatorietà aggiunta.")
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


@login_required
@require_POST
def formazione_corso_req_rule_delete(request, corso_id: int, rule_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare corsi formativi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    rule = get_object_or_404(TrainingRequirementRule, pk=rule_id, corso_id=corso_id)
    rule.delete()
    messages.success(request, "Regola rimossa.")
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


@login_required
@require_POST
def formazione_corso_assegna(request, corso_id: int):
    """Assegna in blocco dei dipendenti al corso (TrainingAssignment, stato ASSEGNATO).

    Primo anello del ciclo corso→sessione: chi è assegnato al corso viene poi proposto
    in cima ai candidati delle sue edizioni. Idempotente (unique_together corso×dip)."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per assegnare corsi.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse.objects.select_related("piano"), pk=corso_id)

    ids: list[int] = []
    seen: set[int] = set()
    for raw in request.POST.getlist("dipendenti_selezionati"):
        s = str(raw).strip()
        if s.isdigit():
            lid = int(s)
            if lid > 0 and lid not in seen:
                seen.add(lid)
                ids.append(lid)
    if not ids:
        messages.warning(request, "Nessun dipendente selezionato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)

    n_new = 0
    for lid in ids:
        _, created = TrainingAssignment.objects.get_or_create(
            corso=corso, legacy_anagrafica_id=lid,
            defaults={
                "stato": "ASSEGNATO",
                "piano": corso.piano,
                "assigned_by": request.user,
            },
        )
        if created:
            n_new += 1
    if n_new:
        messages.success(request, f"{n_new} dipendenti assegnati al corso.")
    else:
        messages.info(request, "I dipendenti selezionati erano già assegnati al corso.")
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


# ============================================================================
# FORMAZIONE HR — Istruttori (PATCH-03)
# ============================================================================

@login_required
def formazione_istruttori_list(request):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    filtro_tipo = request.GET.get("tipo", "")
    q_search = (request.GET.get("q") or "").strip()
    qs = TrainingInstructor.objects.all()
    if filtro_tipo:
        qs = qs.filter(tipo=filtro_tipo)
    if q_search:
        qs = qs.filter(Q(nome__icontains=q_search) | Q(ragione_sociale__icontains=q_search))

    istruttori = list(qs.order_by("nome"))
    form = TrainingInstructorForm()
    return render(request, "anagrafica/pages/formazione_istruttori.html", {
        "istruttori": istruttori,
        "form": form,
        "is_editor": is_editor,
        "filtro_tipo": filtro_tipo,
        "q_search": q_search,
        "TIPO_CHOICES": TrainingInstructor.TIPO_CHOICES,
    })


@login_required
@require_POST
def formazione_istruttore_create(request):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per creare istruttori.")
        return redirect("anagrafica:formazione_istruttori_list")
    form = TrainingInstructorForm(request.POST)
    if form.is_valid():
        istr = form.save()
        messages.success(request, f'Istruttore "{istr.nome}" creato.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_istruttori_list")


@login_required
@require_POST
def formazione_istruttore_edit(request, istruttore_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare istruttori.")
        return redirect("anagrafica:formazione_istruttori_list")
    istr = get_object_or_404(TrainingInstructor, pk=istruttore_id)
    form = TrainingInstructorForm(request.POST, instance=istr)
    if form.is_valid():
        form.save()
        messages.success(request, f'Istruttore "{istr.nome}" aggiornato.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_istruttori_list")


@login_required
@require_POST
def formazione_istruttore_delete(request, istruttore_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per eliminare istruttori.")
        return redirect("anagrafica:formazione_istruttori_list")
    istr = get_object_or_404(TrainingInstructor, pk=istruttore_id)
    if istr.sessioni.exists() or istr.lezioni.exists():
        istr.is_active = False
        istr.save()
        messages.warning(request, f'Istruttore "{istr.nome}" disattivato (ha sessioni/lezioni associate).')
    else:
        nome = istr.nome
        istr.delete()
        messages.success(request, f'Istruttore "{nome}" eliminato.')
    return redirect("anagrafica:formazione_istruttori_list")


# ============================================================================
# FORMAZIONE HR — Sessioni formative (PATCH-04)
# ============================================================================

@login_required
def formazione_sessioni_list(request):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    filtro_corso  = request.GET.get("corso", "")
    filtro_stato  = request.GET.get("stato", "")
    filtro_anno   = request.GET.get("anno", "")
    q_search      = (request.GET.get("q") or "").strip()

    from django.utils import timezone as _tz
    qs = TrainingSession.objects.select_related("corso", "corso__piano", "docente").all()
    if filtro_corso:
        qs = qs.filter(corso_id=filtro_corso)
    if filtro_stato:
        qs = qs.filter(stato=filtro_stato)
    if filtro_anno:
        qs = qs.filter(data_inizio__year=filtro_anno)
    if q_search:
        qs = qs.filter(
            Q(codice_sessione__icontains=q_search) |
            Q(corso__titolo__icontains=q_search) |
            Q(sede__icontains=q_search)
        )

    paginator = Paginator(qs.order_by("-data_inizio"), 50)
    page_obj  = paginator.get_page(request.GET.get("page"))

    corsi_attivi = list(TrainingCourse.objects.filter(is_active=True).order_by("titolo"))
    anni = list(
        TrainingSession.objects
        .values_list("data_inizio__year", flat=True)
        .distinct()
        .order_by("-data_inizio__year")
    )

    return render(request, "anagrafica/pages/formazione_sessioni.html", {
        "page_obj": page_obj,
        "is_editor": is_editor,
        "filtro_corso": filtro_corso,
        "filtro_stato": filtro_stato,
        "filtro_anno": filtro_anno,
        "q_search": q_search,
        "corsi_attivi": corsi_attivi,
        "anni": anni,
        "STATO_CHOICES": TrainingSession.STATO_CHOICES,
    })


@login_required
@require_POST
def formazione_rinnovo_da_scadenzario(request):
    """Punto d'ingresso «seleziona dipendenti → sessione di rinnovo»: raccoglie i
    dipendenti selezionati per un corso e li porta nel FLUSSO STANDARD di creazione
    sessione (``formazione_sessione_create``). Gli id restano in ``request.session``
    e vengono iscritti in blocco al salvataggio della sessione."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per creare sessioni di rinnovo.")
        return redirect("anagrafica:formazione_scadenzario")
    try:
        corso = TrainingCourse.objects.get(pk=request.POST.get("corso_id"), is_active=True)
    except (TrainingCourse.DoesNotExist, ValueError, TypeError):
        messages.error(request, "Corso non valido.")
        return redirect("anagrafica:formazione_scadenzario")
    ids, seen = [], set()
    for raw in request.POST.getlist("dipendenti_selezionati"):
        s = str(raw).strip()
        if s.isdigit():
            lid = int(s)
            if lid > 0 and lid not in seen:
                seen.add(lid)
                ids.append(lid)
    if not ids:
        messages.warning(request, "Nessun dipendente selezionato.")
        return redirect(request.POST.get("back") or "anagrafica:scadenzario")
    request.session["rinnovo_preselect"] = {"corso": corso.pk, "ids": ids}
    messages.info(request, f"{len(ids)} dipendenti pronti per il rinnovo: compila la sessione.")
    return redirect(f"{reverse('anagrafica:formazione_sessione_create')}?corso={corso.pk}")


@login_required
def formazione_sessione_create(request):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per creare sessioni formative.")
        return redirect("anagrafica:formazione_sessioni_list")
    if request.method == "POST":
        form = TrainingSessionForm(request.POST)
        if form.is_valid():
            sessione = form.save(commit=False)
            sessione.created_by = request.user
            sessione.save()
            # Rinnovo dallo scadenzario: se ci sono dipendenti pre-selezionati per
            # QUESTO corso, iscrivili in blocco (idempotente) e vai agli iscritti.
            pre = request.session.get("rinnovo_preselect")
            if pre and pre.get("corso") == sessione.corso_id and pre.get("ids"):
                n_new = 0
                for lid in pre["ids"]:
                    _, created = TrainingEnrollment.objects.get_or_create(
                        sessione=sessione, legacy_anagrafica_id=lid,
                        defaults={"stato": "ISCRITTO", "iscritto_da": request.user},
                    )
                    if created:
                        n_new += 1
                request.session.pop("rinnovo_preselect", None)
                messages.success(request, f'Sessione "{sessione.codice_sessione}" creata; {n_new} dipendenti iscritti per il rinnovo.')
                return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione.pk)
            messages.success(request, f'Sessione "{sessione.codice_sessione}" creata.')
            return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione.pk)
    else:
        initial = {}
        if request.GET.get("corso"):
            initial["corso"] = request.GET.get("corso")
        form = TrainingSessionForm(initial=initial)
    return render(request, "anagrafica/pages/formazione_sessione_form.html", {
        "form": form,
        "modo": "crea",
    })


@login_required
def formazione_sessione_detail(request, sessione_id: int):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    sessione = get_object_or_404(
        TrainingSession.objects.select_related("corso", "corso__piano", "docente"),
        pk=sessione_id,
    )
    lezioni  = list(sessione.lezioni.select_related("docente").order_by("data", "ora_inizio"))
    n_iscritti = sessione.iscrizioni.count()

    # Presenze registrate per lezione (1 query aggregata).
    presenze_per_lezione = {
        row["lezione_id"]: row["n"]
        for row in (
            TrainingLessonAttendance.objects
            .filter(lezione__sessione=sessione)
            .values("lezione_id")
            .annotate(n=Count("id"))
        )
    }
    # Allegati (registro firmato / materiale): livello sessione + per-lezione.
    allegati = list(sessione.allegati.select_related("lezione").order_by("-created_at"))
    allegati_sessione = [a for a in allegati if a.lezione_id is None]
    allegati_per_lezione: dict[int, list] = {}
    for a in allegati:
        if a.lezione_id is not None:
            allegati_per_lezione.setdefault(a.lezione_id, []).append(a)
    for lz in lezioni:
        lz.n_presenze = presenze_per_lezione.get(lz.pk, 0)
        lz.allegati_list = allegati_per_lezione.get(lz.pk, [])

    edit_form   = TrainingSessionForm(instance=sessione)
    lezione_form = TrainingLessonForm(sessione=sessione)

    return render(request, "anagrafica/pages/formazione_sessione_detail.html", {
        "sessione":     sessione,
        "lezioni":      lezioni,
        "n_iscritti":   n_iscritti,
        "edit_form":    edit_form,
        "lezione_form": lezione_form,
        "is_editor":    is_editor,
        "allegati_sessione": allegati_sessione,
        "ATTACH_TIPI":  TrainingAttachment.Tipo.choices,
    })


@login_required
@require_POST
def formazione_sessione_edit(request, sessione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare sessioni formative.")
        return redirect("anagrafica:formazione_sessioni_list")
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    form = TrainingSessionForm(request.POST, instance=sessione)
    if form.is_valid():
        form.save()
        messages.success(request, f'Sessione "{sessione.codice_sessione}" aggiornata.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_sessione_delete(request, sessione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per eliminare sessioni formative.")
        return redirect("anagrafica:formazione_sessioni_list")
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    if sessione.iscrizioni.exists():
        messages.error(
            request,
            f'La sessione "{sessione.codice_sessione}" ha iscrizioni. '
            "Rimuovi prima gli iscritti o annulla la sessione.",
        )
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)
    corso_id = sessione.corso_id
    codice   = sessione.codice_sessione
    sessione.delete()
    messages.success(request, f'Sessione "{codice}" eliminata.')
    return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)


# ── Lezioni ──────────────────────────────────────────────────────────────────

@login_required
@require_POST
def formazione_lezione_add(request, sessione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare sessioni formative.")
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    form = TrainingLessonForm(request.POST, sessione=sessione)
    if form.is_valid():
        lezione = form.save(commit=False)
        lezione.sessione = sessione
        lezione.updated_by = request.user
        lezione.save()
        messages.success(request, f'Lezione {lezione.numero} aggiunta.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_lezione_edit(request, sessione_id: int, lezione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare sessioni formative.")
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    lezione  = get_object_or_404(TrainingLesson, pk=lezione_id, sessione=sessione)
    form = TrainingLessonForm(request.POST, instance=lezione, sessione=sessione)
    if form.is_valid():
        lz = form.save(commit=False)
        lz.updated_by = request.user
        lz.save()
        messages.success(request, f'Lezione {lezione.numero} aggiornata.')
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_lezione_delete(request, sessione_id: int, lezione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare sessioni formative.")
        return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)
    lezione = get_object_or_404(TrainingLesson, pk=lezione_id, sessione_id=sessione_id)
    num = lezione.numero
    lezione.delete()
    messages.success(request, f'Lezione {num} eliminata.')
    return redirect("anagrafica:formazione_sessione_detail", sessione_id=sessione_id)


# ============================================================================
# FORMAZIONE HR — Iscritti e Presenze (PATCH-05)
# ============================================================================

def _add_months(dt, months: int):
    """Aggiunge N mesi a una data, gestendo correttamente fine mese."""
    import calendar
    month = dt.month - 1 + months
    year  = dt.year + month // 12
    month = month % 12 + 1
    day   = min(dt.day, calendar.monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)


def _calcola_percentuale_presenza(enrollment: TrainingEnrollment) -> float | None:
    """Calcola la percentuale di presenza sull'intero monte ore della sessione."""
    lezioni = list(enrollment.sessione.lezioni.all())
    if not lezioni:
        return None
    ore_totali = sum(lz.durata_ore for lz in lezioni)
    if ore_totali <= 0:
        return None
    presenze = {
        p.lezione_id: p
        for p in TrainingLessonAttendance.objects.filter(
            lezione__in=lezioni,
            legacy_anagrafica_id=enrollment.legacy_anagrafica_id,
        )
    }
    ore_effettive = 0.0
    for lz in lezioni:
        p = presenze.get(lz.pk)
        if p and p.stato_presenza == "PRESENTE":
            ore_effettive += float(p.ore_effettive) if p.ore_effettive else lz.durata_ore
        elif p and p.stato_presenza == "PARZIALE":
            ore_effettive += float(p.ore_effettive) if p.ore_effettive else (lz.durata_ore * 0.5)
    return round(ore_effettive / ore_totali * 100, 2)


def _ricalcola_presenza_enrollment(enrollment: TrainingEnrollment) -> None:
    """Ricalcola e salva ore_frequentate e percentuale_presenza dell'iscrizione dalle
    presenze registrate su tutte le lezioni della sessione."""
    perc = _calcola_percentuale_presenza(enrollment)
    ore = sum(
        float(pr.ore_effettive) if pr.ore_effettive else lz.durata_ore
        for lz, pr in [
            (lz, TrainingLessonAttendance.objects.filter(
                lezione=lz, legacy_anagrafica_id=enrollment.legacy_anagrafica_id
            ).first())
            for lz in enrollment.sessione.lezioni.all()
        ]
        if pr and pr.stato_presenza in ("PRESENTE", "PARZIALE")
    )
    enrollment.ore_frequentate = round(ore, 2)
    enrollment.percentuale_presenza = perc
    enrollment.save(update_fields=["ore_frequentate", "percentuale_presenza"])


def _motivi_blocco_completamento(iscrizione: TrainingEnrollment) -> list[str]:
    """Motivi che impediscono un completamento *conforme* (Accordo Stato-Regioni 2025),
    secondo la regola di superamento del corso: verifica finale non superata e/o
    frequenza sotto la soglia minima. Lista vuota = nessun blocco (o nessuna regola)."""
    corso = iscrizione.sessione.corso
    try:
        regola = corso.regola_superamento
    except TrainingCompletionRule.DoesNotExist:
        return []
    motivi: list[str] = []
    if regola.richiede_esame_finale and iscrizione.verifica_superata is not True:
        motivi.append("verifica finale di apprendimento non superata o non registrata")
    soglia = regola.presenza_minima_percentuale or 0
    if soglia:
        perc = iscrizione.percentuale_presenza
        if perc is None:
            perc = _calcola_percentuale_presenza(iscrizione)
        if perc is not None and float(perc) < float(soglia):
            motivi.append(f"frequenza {float(perc):.0f}% inferiore al minimo richiesto ({soglia}%)")
    return motivi


def _crea_employee_record(enrollment: TrainingEnrollment, created_by) -> TrainingEmployeeRecord | None:
    """Crea TrainingEmployeeRecord con tutti i campi snapshot al completamento.

    Ritorna None se il record esiste già (idempotente).
    Segna TrainingDeadline.needs_refresh=True per invalidare la cache scadenze.
    """
    if hasattr(enrollment, "record_completamento") and enrollment.record_completamento_id:
        return None

    corso    = enrollment.sessione.corso
    sessione = enrollment.sessione
    from django.utils import timezone as _tz
    oggi = _tz.localdate()

    data_completamento = enrollment.data_completamento or oggi
    data_scadenza = None
    if corso.validita_mesi:
        data_scadenza = _add_months(data_completamento, corso.validita_mesi)

    try:
        completion_rule = corso.regola_superamento
        rule_json = {
            "ore_minime": completion_rule.ore_minime_percentuale,
            "presenza_minima": completion_rule.presenza_minima_percentuale,
            "esame": completion_rule.richiede_esame_finale,
        }
    except TrainingCompletionRule.DoesNotExist:
        rule_json = {}

    record = TrainingEmployeeRecord.objects.create(
        corso=corso,
        sessione=sessione,
        enrollment=enrollment,
        legacy_anagrafica_id=enrollment.legacy_anagrafica_id,
        data_completamento=data_completamento,
        ore_frequentate=enrollment.ore_frequentate or 0,
        percentuale_presenza=enrollment.percentuale_presenza,
        idoneo=enrollment.idoneo if enrollment.idoneo is not None else True,
        data_scadenza=data_scadenza,
        # Snapshot storici
        course_code_snapshot=corso.codice,
        course_title_snapshot=corso.titolo,
        course_version_snapshot=corso.versione,
        plan_code_snapshot=corso.piano.codice if corso.piano_id else "",
        plan_name_snapshot=corso.piano.nome if corso.piano_id else "",
        duration_hours_snapshot=corso.durata_ore_teorica,
        validity_months_snapshot=corso.validita_mesi,
        completion_rule_snapshot_json=rule_json,
        session_code_snapshot=sessione.codice_sessione,
        teacher_name_snapshot=sessione.docente_nome or "",
        completion_calculation_snapshot_json={
            "ore_frequentate": str(enrollment.ore_frequentate or 0),
            "percentuale": str(enrollment.percentuale_presenza or ""),
        },
    )

    # Invalida cache scadenze per questo dipendente × corso
    TrainingDeadline.objects.filter(
        corso=corso,
        legacy_anagrafica_id=enrollment.legacy_anagrafica_id,
    ).update(needs_refresh=True)

    # Tenta ricalcolo immediato (stub — NON lancia eccezione se non ancora implementato)
    try:
        from .services.training_deadline_service import refresh_deadlines
        refresh_deadlines(legacy_id=enrollment.legacy_anagrafica_id, corso_id=corso.pk)
    except NotImplementedError:
        pass  # PATCH-06 implementerà il ricalcolo completo

    # Allineamento qualifica (competency management): se il corso rilascia/rinnova una
    # qualifica e il completamento è idoneo, crea/aggiorna la DipendenteQualifica corrente
    # collegandola al record (la prova formativa). Riusa la convenzione "una qualifica
    # corrente per (dip, tipo), niente duplicati". Fail-safe: non deve bloccare il
    # completamento.
    try:
        if corso.qualifica_id and record.idoneo:
            qual, _ = _upsert_dipendente_qualifica(
                record.legacy_anagrafica_id, corso.qualifica,
                record.data_completamento, record.data_scadenza,
                user=created_by,
            )
            if qual.record_formazione_id != record.pk:
                qual.record_formazione = record
                qual.save(update_fields=["record_formazione"])
    except Exception:
        logger.exception("Allineamento qualifica fallito per record %s", record.pk)

    # Archiviazione automatica dell'attestato nel box documenti del dipendente
    # (se abilitata da Impostazioni → Template attestato). Fail-safe: un errore
    # qui NON deve impedire la registrazione del completamento.
    try:
        from .models_formazione import AttestatoFormazioneConfig
        cfg = AttestatoFormazioneConfig.get_instance()
        if cfg.auto_salva_attestato:
            from .services.attestato_pdf import archivia_attestato
            archivia_attestato(record, cfg=cfg, user=created_by)
    except Exception:
        logger.exception("Archiviazione automatica attestato fallita per record %s", record.pk)

    return record


def _candidati_rinnovo_corso(corso, sessione=None) -> list[dict]:
    """Dipendenti da (ri)formare per un corso, dalla cache scadenze
    (`TrainingDeadline`): scaduti / in scadenza / mai frequentati, esclusi i
    cessati e chi è già iscritto alla sessione. Pre-seleziona scaduti e in
    scadenza (il "mai frequentato" è incluso ma non pre-spuntato: è un primo
    rilascio, non un rinnovo). Specchio di `_build_candidati_qualifica`."""
    rilevanti = ["SCADUTO", "IN_SCADENZA_30", "IN_SCADENZA_90", "MAI_FREQUENTATO"]
    deadlines = list(
        TrainingDeadline.objects.filter(corso=corso, stato_scadenza__in=rilevanti)
        .order_by("data_scadenza", "legacy_anagrafica_id")
    )
    iscritti_ids: set[int] = set()
    if sessione is not None:
        iscritti_ids = set(
            TrainingEnrollment.objects.filter(sessione=sessione)
            .values_list("legacy_anagrafica_id", flat=True)
        )
    nomi = _build_nomi_map()
    cessati = _cessati_legacy_ids()
    out: list[dict] = []
    for d in deadlines:
        lid = d.legacy_anagrafica_id
        if lid in cessati or lid in iscritti_ids:
            continue
        out.append({
            "legacy_id": lid,
            "nome": nomi.get(lid, f"#{lid}"),
            "stato": d.stato_scadenza,
            "stato_label": d.get_stato_scadenza_display(),
            "data_scadenza": d.data_scadenza,
            "preselect": d.stato_scadenza in ("SCADUTO", "IN_SCADENZA_30", "IN_SCADENZA_90"),
        })
    order = {"SCADUTO": 0, "IN_SCADENZA_30": 1, "IN_SCADENZA_90": 2, "MAI_FREQUENTATO": 3}
    out.sort(key=lambda c: (order.get(c["stato"], 9), c["nome"].casefold()))
    return out


@login_required
def formazione_sessione_iscritti(request, sessione_id: int):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    sessione = get_object_or_404(
        TrainingSession.objects.select_related("corso", "corso__piano"),
        pk=sessione_id,
    )
    iscrizioni = list(
        TrainingEnrollment.objects.filter(sessione=sessione)
        .order_by("legacy_anagrafica_id")
    )
    lezioni = list(sessione.lezioni.order_by("data", "ora_inizio"))

    # Mappa presenze: {(legacy_id, lezione_id): stato_presenza}
    ids = [i.legacy_anagrafica_id for i in iscrizioni]
    presenze_qs = TrainingLessonAttendance.objects.filter(
        lezione__in=lezioni,
        legacy_anagrafica_id__in=ids,
    ).values("lezione_id", "legacy_anagrafica_id", "stato_presenza")
    presenze_map: dict[tuple, str] = {
        (p["legacy_anagrafica_id"], p["lezione_id"]): p["stato_presenza"]
        for p in presenze_qs
    }

    # Turni: a quali lezioni è assegnato ciascun iscritto (vuoto = tutte le lezioni).
    from .services.training_turni import mappa_turni_sessione
    turni_map = mappa_turni_sessione(sessione) if lezioni else {}

    nomi_map = _build_nomi_map()
    for i in iscrizioni:
        i.nome_dip = nomi_map.get(i.legacy_anagrafica_id, f"#{i.legacy_anagrafica_id}")
        # Griglia presenze per riga (lista parallela alle lezioni)
        presenze_griglia = [
            presenze_map.get((i.legacy_anagrafica_id, lz.pk), "")
            for lz in lezioni
        ]
        i.lezioni_presenze = list(zip(lezioni, presenze_griglia))
        # Turni assegnati: set di lezione_id (vuoto = tutte). Lista parallela per la UI.
        i.turni_ids = turni_map.get(i.pk, set())
        i.turni_espliciti = bool(i.turni_ids)
        i.turni_griglia = [(lz, lz.pk in i.turni_ids) for lz in lezioni]
        i.turni_label = (
            ", ".join(f"L{lz.numero}" for lz in lezioni if lz.pk in i.turni_ids)
            if i.turni_espliciti else "Tutte"
        )

    # Lista dipendenti attivi per il form di iscrizione manuale
    dipendenti_attivi = _build_nomi_map()

    # Candidati all'iscrizione: motore di idoneità (pertinenza + scadenze + prerequisiti).
    # Restituisce idonei (proponibili, pre-spuntati i rinnovi) e non idonei (prerequisiti
    # mancanti, in coda e disabilitati). Solo per editor.
    candidati = {"idonei": [], "non_idonei": [], "pool_filtrato": False, "n_preselect": 0}
    if is_editor:
        from .services.training_eligibility import candidati_corso
        candidati = candidati_corso(sessione.corso, sessione=sessione)
    candidati_rinnovo = candidati["idonei"] + candidati["non_idonei"]

    return render(request, "anagrafica/pages/formazione_iscritti.html", {
        "sessione":          sessione,
        "iscrizioni":        iscrizioni,
        "lezioni":           lezioni,
        "is_editor":         is_editor,
        "dipendenti_attivi": sorted(dipendenti_attivi.items(), key=lambda x: x[1]),
        "STATO_CHOICES":     TrainingEnrollment.STATO_CHOICES,
        "candidati":         candidati,
        "candidati_rinnovo": candidati_rinnovo,
        "n_rinnovo_pre":     candidati["n_preselect"],
    })


@login_required
@require_POST
def formazione_iscrizione_add(request, sessione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le iscrizioni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)

    try:
        legacy_id = int(request.POST.get("legacy_anagrafica_id") or 0)
    except (TypeError, ValueError):
        legacy_id = 0
    if not legacy_id:
        messages.error(request, "Selezionare un dipendente valido.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)

    # Enforcement *soft* dei prerequisiti: se mancano e non c'è forzatura esplicita,
    # blocca con avviso; con force=1 procede ma traccia la deroga in audit.
    from .services.training_eligibility import prerequisiti_mancanti
    mancanti = prerequisiti_mancanti(sessione.corso, legacy_id)
    force = request.POST.get("force") == "1"
    if mancanti and not force:
        nome = _build_nomi_map().get(legacy_id, f"#{legacy_id}")
        messages.error(
            request,
            f'"{nome}" non soddisfa i prerequisiti del corso '
            f'(manca: {", ".join(mancanti)}). Per iscriverlo comunque, conferma la forzatura.',
        )
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    if mancanti and force:
        _audit_safe(request, "formazione_iscrizione_forzata", "formazione", {
            "sessione_id": sessione.pk, "legacy_id": legacy_id, "prerequisiti_mancanti": mancanti,
        })

    # Collega l'eventuale assegnazione a livello corso (TrainingAssignment).
    assignment = TrainingAssignment.objects.filter(
        corso=sessione.corso, legacy_anagrafica_id=legacy_id
    ).first()

    _, created = TrainingEnrollment.objects.get_or_create(
        sessione=sessione,
        legacy_anagrafica_id=legacy_id,
        defaults={
            "stato": "ISCRITTO",
            "iscritto_da": request.user,
            "assignment": assignment,
            "note": (request.POST.get("note") or "").strip(),
        },
    )
    if created:
        nomi = _build_nomi_map()
        nome = nomi.get(legacy_id, f"#{legacy_id}")
        messages.success(request, f'Dipendente "{nome}" iscritto alla sessione.')
    else:
        messages.warning(request, "Dipendente già iscritto a questa sessione.")
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_iscrizione_edit(request, sessione_id: int, iscrizione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le iscrizioni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    iscrizione = get_object_or_404(TrainingEnrollment, pk=iscrizione_id, sessione_id=sessione_id)
    stato_precedente = iscrizione.stato

    form = TrainingEnrollmentEditForm(request.POST, instance=iscrizione)
    if form.is_valid():
        iscrizione = form.save()
        # Se diventa COMPLETATO per la prima volta: gating conformità (Accordo SR 2025).
        if iscrizione.stato == "COMPLETATO" and stato_precedente != "COMPLETATO":
            force = request.POST.get("force") == "1"
            motivi = _motivi_blocco_completamento(iscrizione)
            if motivi and not force:
                # Conserva i dati inseriti ma non avanza lo stato.
                iscrizione.stato = stato_precedente
                iscrizione.save(update_fields=["stato"])
                messages.error(
                    request,
                    "Completamento bloccato: " + "; ".join(motivi)
                    + ". Conferma la forzatura per registrarlo comunque.",
                )
                return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
            if motivi and force:
                _audit_safe(request, "formazione_completamento_forzato", "formazione", {
                    "sessione_id": sessione_id, "iscrizione_id": iscrizione.pk, "motivi": motivi,
                })
            _crea_employee_record(iscrizione, request.user)
        messages.success(request, "Iscrizione aggiornata.")
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_iscrizione_attestato_upload(request, sessione_id: int, iscrizione_id: int):
    """Carica l'attestato dell'organizzatore esterno e **chiude il corso** per l'iscritto:
    porta l'iscrizione a COMPLETATO, crea il record storico, archivia il file nel box
    documenti (CERTIFICATO_FORMAZIONE) e registra il TrainingCertificate. Rispetta il
    gating conformità (verifica/frequenza), forzabile con force=1 (tracciato)."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le iscrizioni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    iscrizione = get_object_or_404(
        TrainingEnrollment.objects.select_related("sessione", "sessione__corso"),
        pk=iscrizione_id, sessione_id=sessione_id,
    )

    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "Seleziona l'attestato da caricare.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    suffix = Path(uploaded.name or "").suffix.lower()
    if suffix not in _ALLOWED_DOC_EXTENSIONS:
        messages.error(request, f"Formato non consentito ({suffix}). Ammessi: PDF, immagini, DOC/XLS.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    if uploaded.size > _MAX_DOC_SIZE:
        messages.error(request, f"File troppo grande ({uploaded.size // (1024*1024)} MB). Limite: 50 MB.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    try:
        from core.upload_mime import sniff_mime
        mime = sniff_mime(uploaded)
    except Exception:
        mime = uploaded.content_type or "application/octet-stream"
    if mime not in _ALLOWED_DOC_MIMES:
        messages.error(request, "Tipo di file non consentito (contenuto non valido).")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)

    force = request.POST.get("force") == "1"
    # Gating conformità (se non già completata)
    if iscrizione.stato != "COMPLETATO":
        motivi = _motivi_blocco_completamento(iscrizione)
        if motivi and not force:
            messages.error(
                request,
                "Chiusura bloccata: " + "; ".join(motivi)
                + ". Conferma la forzatura per chiudere comunque.",
            )
            return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
        if motivi and force:
            _audit_safe(request, "formazione_completamento_forzato", "formazione", {
                "sessione_id": sessione_id, "iscrizione_id": iscrizione.pk,
                "motivi": motivi, "via": "attestato_caricato",
            })
        from django.utils import timezone as _tz
        if not iscrizione.data_completamento:
            iscrizione.data_completamento = _tz.localdate()
        if iscrizione.idoneo is None:
            iscrizione.idoneo = True
        iscrizione.stato = "COMPLETATO"
        iscrizione.save(update_fields=["stato", "idoneo", "data_completamento"])

    record = _crea_employee_record(iscrizione, request.user)
    if record is None:
        record = getattr(iscrizione, "record_completamento", None)
    if record is None:
        messages.error(request, "Impossibile registrare il completamento.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)

    rilasciato_da = (request.POST.get("rilasciato_da") or "").strip()[:200]
    numero = (request.POST.get("numero_attestato") or "").strip()[:100]
    data_rilascio = None
    raw = (request.POST.get("data_rilascio") or "").strip()
    if raw:
        try:
            data_rilascio = date.fromisoformat(raw)
        except ValueError:
            data_rilascio = None
    if data_rilascio is None:
        data_rilascio = record.data_completamento

    try:
        from .services.attestato_pdf import archivia_attestato, archivia_attestato_caricato
        # Copia interna NOVICROM: assicurane la presenza, così affianca l'esterno
        # (storico completo nel box). Fail-safe: non deve bloccare l'archiviazione esterna.
        try:
            archivia_attestato(record, user=request.user)
        except Exception:
            logger.exception("Copia interna attestato non generata (record %s)", record.pk)
        # Attestato esterno = principale (slot dedicato), collegato al certificato.
        descr = "Attestato organizzatore esterno (principale)" + (f" — {rilasciato_da}" if rilasciato_da else "")
        doc = archivia_attestato_caricato(
            record, uploaded, user=request.user, force=True, mime=mime, descrizione=descr,
        )
        TrainingCertificate.objects.update_or_create(
            record=record,
            defaults={
                "legacy_anagrafica_id": record.legacy_anagrafica_id,
                "numero_attestato": numero,
                "data_rilascio": data_rilascio,
                "rilasciato_da": rilasciato_da,
                "file_attestato": doc,
                "created_by": request.user,
            },
        )
        _audit_safe(request, "formazione_attestato_caricato", "formazione", {
            "sessione_id": sessione_id, "iscrizione_id": iscrizione.pk,
            "record_id": record.pk, "documento_id": doc.pk,
        })
        messages.success(request, "Attestato caricato e corso chiuso: completamento registrato e archiviato.")
    except Exception:
        logger.exception("Upload attestato organizzatore fallito (iscrizione %s)", iscrizione.pk)
        messages.error(request, "Errore durante l'archiviazione dell'attestato.")
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_iscrizione_delete(request, sessione_id: int, iscrizione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le iscrizioni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    iscrizione = get_object_or_404(TrainingEnrollment, pk=iscrizione_id, sessione_id=sessione_id)
    iscrizione.delete()
    messages.success(request, "Iscrizione rimossa.")
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_iscrizione_bulk(request, sessione_id: int):
    """Iscrive in blocco i dipendenti selezionati (candidati al rinnovo del
    corso) a questa edizione. Idempotente: chi è già iscritto viene saltato."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le iscrizioni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)

    ids: list[int] = []
    seen: set[int] = set()
    for raw in request.POST.getlist("dipendenti_selezionati"):
        s = str(raw).strip()
        if s.isdigit():
            lid = int(s)
            if lid > 0 and lid not in seen:
                seen.add(lid)
                ids.append(lid)
    if not ids:
        messages.warning(request, "Nessun dipendente selezionato.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)

    # Collega alle assegnazioni a livello corso (TrainingAssignment), se presenti.
    assegnazioni = {
        a.legacy_anagrafica_id: a
        for a in TrainingAssignment.objects.filter(
            corso=sessione.corso, legacy_anagrafica_id__in=ids
        )
    }
    n_new = 0
    for lid in ids:
        _, created = TrainingEnrollment.objects.get_or_create(
            sessione=sessione, legacy_anagrafica_id=lid,
            defaults={
                "stato": "ISCRITTO",
                "iscritto_da": request.user,
                "assignment": assegnazioni.get(lid),
            },
        )
        if created:
            n_new += 1
    if n_new:
        messages.success(request, f"{n_new} dipendenti iscritti all'edizione di rinnovo.")
    else:
        messages.info(request, "I dipendenti selezionati erano già iscritti.")
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
@require_POST
def formazione_iscrizione_turni(request, sessione_id: int, iscrizione_id: int):
    """Imposta i turni (lezioni) di un'iscrizione = le lezioni selezionate.

    Nessuna selezione ⇒ l'iscritto torna "su tutte le lezioni" (default storico).
    Vedi :mod:`services.training_turni`."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire i turni.")
        return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)
    iscrizione = get_object_or_404(
        TrainingEnrollment, pk=iscrizione_id, sessione_id=sessione_id
    )
    lezione_ids: list[int] = []
    for raw in request.POST.getlist("lezioni_turno"):
        s = str(raw).strip()
        if s.isdigit():
            lezione_ids.append(int(s))

    from .services.training_turni import set_turni
    set_turni(iscrizione, lezione_ids, user=request.user)
    if lezione_ids:
        messages.success(request, "Turni dell'iscritto aggiornati.")
    else:
        messages.success(request, "Turni rimossi: l'iscritto frequenta tutte le lezioni.")
    return redirect("anagrafica:formazione_sessione_iscritti", sessione_id=sessione_id)


@login_required
def formazione_lezione_presenze(request, sessione_id: int, lezione_id: int):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)

    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    lezione  = get_object_or_404(TrainingLesson, pk=lezione_id, sessione=sessione)

    # Turni: mostra solo gli iscritti *attesi* a questa lezione (chi non ha turni
    # espliciti vale per tutte le lezioni — fallback storico).
    from .services.training_turni import iscritti_attesi_lezione
    tutte_iscrizioni = list(
        TrainingEnrollment.objects.filter(sessione=sessione)
        .order_by("legacy_anagrafica_id")
    )
    iscrizioni = iscritti_attesi_lezione(sessione, lezione, tutte_iscrizioni)
    turno_filtrato = len(iscrizioni) != len(tutte_iscrizioni)
    ids = [i.legacy_anagrafica_id for i in iscrizioni]
    presenze_map = {
        p.legacy_anagrafica_id: p
        for p in TrainingLessonAttendance.objects.filter(
            lezione=lezione, legacy_anagrafica_id__in=ids
        )
    }
    nomi_map = _build_nomi_map()
    righe = []
    for i in iscrizioni:
        righe.append({
            "enrollment": i,
            "nome": nomi_map.get(i.legacy_anagrafica_id, f"#{i.legacy_anagrafica_id}"),
            "presenza": presenze_map.get(i.legacy_anagrafica_id),
        })

    n_presenti = sum(1 for r in righe if r["presenza"] is not None)

    allegati_lezione = list(lezione.allegati.order_by("-created_at"))

    return render(request, "anagrafica/pages/formazione_presenze.html", {
        "sessione":   sessione,
        "lezione":    lezione,
        "righe":      righe,
        "is_editor":  is_editor,
        "n_presenti": n_presenti,
        "turno_filtrato": turno_filtrato,
        "n_totale_iscritti": len(tutte_iscrizioni),
        "STATO_PRESENZA_CHOICES": TrainingLessonAttendance.STATO_PRESENZA_CHOICES,
        "allegati_lezione": allegati_lezione,
        "ATTACH_TIPI": TrainingAttachment.Tipo.choices,
    })


@login_required
def formazione_lezione_registro(request, sessione_id: int, lezione_id: int):
    """Foglio firme (registro presenze) di una lezione, in **PDF**.

    Veste UNIFICATA del portale (reportlab via `core/pdf`), identica ai fogli firme di
    corso e del fascicolo: intestazione corso/sessione/lezione, tabella nominativi +
    colonne firma ingresso/uscita, righe vuote di scorta e firma docente. Turno-aware
    (elenca gli iscritti attesi alla lezione). Tracciato in `TrainingExportLog`.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    sessione = get_object_or_404(TrainingSession.objects.select_related("corso", "corso__piano"), pk=sessione_id)
    lezione  = get_object_or_404(TrainingLesson, pk=lezione_id, sessione=sessione)

    # Foglio firme PDF nella veste UNICA del portale (la stessa dei fogli firme di
    # corso e del fascicolo): riusa il builder reportlab condiviso, turno-aware
    # (elenca gli iscritti attesi a questa lezione, fallback tutti).
    from .services.attestato_pdf import build_registro_lezione_pdf_bytes
    pdf = build_registro_lezione_pdf_bytes(lezione)
    try:
        TrainingExportLog.objects.create(
            tipo="REPORT_FIRMA",
            filtri_json={
                "sessione_id": sessione.pk,
                "lezione_id": lezione.pk,
                "formato": "foglio_firme_pdf",
            },
            righe_esportate=sessione.iscrizioni.count(),
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog per foglio firme lezione %s", lezione_id)

    fname = f"Foglio_firme_{sessione.codice_sessione}_L{lezione.numero}.pdf".replace("/", "-").replace(" ", "_")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{fname}"'
    return resp


@login_required
@require_POST
def formazione_presenza_set(request, sessione_id: int, lezione_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le presenze.")
        return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)
    lezione = get_object_or_404(TrainingLesson, pk=lezione_id, sessione_id=sessione_id)

    try:
        legacy_id = int(request.POST.get("legacy_anagrafica_id") or 0)
    except (TypeError, ValueError):
        legacy_id = 0
    if not legacy_id:
        messages.error(request, "ID dipendente non valido.")
        return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)

    presenza, _ = TrainingLessonAttendance.objects.get_or_create(
        lezione=lezione,
        legacy_anagrafica_id=legacy_id,
        defaults={"registrato_da": request.user},
    )
    form = TrainingLessonAttendanceForm(request.POST, instance=presenza)
    if form.is_valid():
        p = form.save(commit=False)
        p.registrato_da = request.user
        p.save()

        # Aggiorna ore_frequentate e percentuale nell'iscrizione
        try:
            enrollment = TrainingEnrollment.objects.get(
                sessione_id=sessione_id, legacy_anagrafica_id=legacy_id
            )
            _ricalcola_presenza_enrollment(enrollment)
        except TrainingEnrollment.DoesNotExist:
            pass

        messages.success(request, "Presenza registrata.")
    else:
        for field_errors in form.errors.values():
            for err in field_errors:
                messages.error(request, err)
    return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)


@login_required
@require_POST
def formazione_registro_autocompila(request, sessione_id: int, lezione_id: int):
    """Autocompila le presenze di una lezione dal **registro firme** firmato.

    Per ogni iscritto *atteso* alla lezione (rispetta i turni) imposta firma
    ingresso/uscita in base ai campi spuntati nel modulo (= "a seconda dei campi
    firmati"), lo stato presenza (PRESENTE se entrambe, PARZIALE se una sola),
    ``signature_status=FIRMATO`` / ``signature_method=UPLOAD`` / ``signed_at``, e
    ricalcola ore/percentuale dell'iscrizione. Gli iscritti senza alcuna firma
    spuntata non vengono toccati.
    """
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per gestire le presenze.")
        return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)
    sessione = get_object_or_404(TrainingSession, pk=sessione_id)
    lezione = get_object_or_404(TrainingLesson, pk=lezione_id, sessione=sessione)

    from django.utils import timezone as _tz
    from .services.training_turni import iscritti_attesi_lezione

    attesi = iscritti_attesi_lezione(sessione, lezione)
    now = _tz.now()
    n = 0
    toccati: list[TrainingEnrollment] = []
    for e in attesi:
        lid = e.legacy_anagrafica_id
        ing = request.POST.get(f"ingresso_{lid}") == "1"
        usc = request.POST.get(f"uscita_{lid}") == "1"
        if not ing and not usc:
            continue  # nessuna firma: non sovrascrive una presenza eventualmente già registrata
        att, _ = TrainingLessonAttendance.objects.get_or_create(
            lezione=lezione, legacy_anagrafica_id=lid,
            defaults={"registrato_da": request.user},
        )
        att.firma_ingresso = ing
        att.firma_uscita = usc
        att.signature_status = "FIRMATO"
        att.signature_method = "UPLOAD"
        att.signed_at = now
        att.stato_presenza = "PRESENTE" if (ing and usc) else "PARZIALE"
        att.registrato_da = request.user
        att.save()
        toccati.append(e)
        n += 1

    for e in toccati:
        _ricalcola_presenza_enrollment(e)

    if n:
        _audit_safe(request, "formazione_registro_autocompila", "formazione", {
            "sessione_id": sessione_id, "lezione_id": lezione_id, "iscritti": n,
        })
        messages.success(request, f"Presenze autocompilate dal registro per {n} iscritti.")
    else:
        messages.info(request, "Nessuna firma selezionata: nessuna presenza modificata.")
    return redirect("anagrafica:formazione_lezione_presenze", sessione_id=sessione_id, lezione_id=lezione_id)


@login_required
def formazione_sessione_fascicolo(request, sessione_id: int):
    """Fascicolo formativo dell'edizione in PDF (progettazione + programma + partecipanti
    ed esiti + relazione): documento unico per la tracciabilità Accordo SR 2025."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")
    sessione = get_object_or_404(
        TrainingSession.objects.select_related("corso", "corso__piano", "docente"),
        pk=sessione_id,
    )
    from .services.attestato_pdf import build_fascicolo_sessione_pdf_bytes
    pdf = build_fascicolo_sessione_pdf_bytes(sessione)
    try:
        TrainingExportLog.objects.create(
            tipo="REPORT_FIRMA",
            filtri_json={"sessione_id": sessione.pk, "formato": "fascicolo_edizione"},
            righe_esportate=sessione.iscrizioni.count(),
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR") or None,
        )
    except Exception:
        logger.exception("Errore TrainingExportLog fascicolo sessione %s", sessione_id)
    fname = f"Fascicolo_{sessione.codice_sessione}.pdf".replace("/", "-").replace(" ", "_")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{fname}"'
    return resp


@login_required
def formazione_scadenzario(request):
    """Scadenzario formazione: corsi scaduti, in scadenza e mai frequentati."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    filtro_stato = request.GET.get("stato", "").strip()
    filtro_corso = request.GET.get("corso", "").strip()
    filtro_q     = request.GET.get("q", "").strip()

    qs = (
        TrainingDeadline.objects
        .select_related("corso", "corso__piano")
        .order_by("stato_scadenza", "data_scadenza", "legacy_anagrafica_id")
    )

    if filtro_stato:
        qs = qs.filter(stato_scadenza=filtro_stato)
    else:
        qs = qs.filter(
            stato_scadenza__in=["SCADUTO", "IN_SCADENZA_30", "IN_SCADENZA_90", "MAI_FREQUENTATO"]
        )

    if filtro_corso:
        try:
            qs = qs.filter(corso_id=int(filtro_corso))
        except (ValueError, TypeError):
            pass

    nomi_map = _build_nomi_map()

    if filtro_q:
        q_lower = filtro_q.lower()
        matched_ids = [lid for lid, nome in nomi_map.items() if q_lower in nome.lower()]
        qs = qs.filter(legacy_anagrafica_id__in=matched_ids)

    scadenze = list(qs)
    for s in scadenze:
        s.nome_dip = nomi_map.get(s.legacy_anagrafica_id, f"#{s.legacy_anagrafica_id}")

    # KPI globali (su tutti, non filtrati)
    all_qs = TrainingDeadline.objects
    n_scaduti  = all_qs.filter(stato_scadenza="SCADUTO").count()
    n_30gg     = all_qs.filter(stato_scadenza="IN_SCADENZA_30").count()
    n_90gg     = all_qs.filter(stato_scadenza="IN_SCADENZA_90").count()
    n_mai      = all_qs.filter(stato_scadenza="MAI_FREQUENTATO", is_required=True).count()

    is_cache_empty = not all_qs.exists()

    paginator = Paginator(scadenze, 50)
    page_obj  = paginator.get_page(request.GET.get("page"))

    corsi_list = TrainingCourse.objects.filter(is_active=True).order_by("codice")

    return render(request, "anagrafica/pages/formazione_scadenzario.html", {
        "page_obj":      page_obj,
        "filtro_stato":  filtro_stato,
        "filtro_corso":  filtro_corso,
        "filtro_q":      filtro_q,
        "corsi_list":    corsi_list,
        "n_scaduti":     n_scaduti,
        "n_30gg":        n_30gg,
        "n_90gg":        n_90gg,
        "n_mai":         n_mai,
        "is_cache_empty": is_cache_empty,
        "totale":        len(scadenze),
        "STATO_SCADENZA_CHOICES": TrainingDeadline.STATO_SCADENZA_CHOICES,
        "is_editor":     _can_edit_formazione(request),
    })


@login_required
def formazione_copertura(request):
    """Report «copertura / gap formativo»: per i corsi *obbligatori* (bersaglio di una
    regola di obbligo attiva) elenca i dipendenti **non in regola** (corso scaduto, in
    scadenza o mai frequentato), con reparto e mansione. Riusa il motore di idoneità
    (`candidati_corso`, che restituisce proprio i candidati = chi non è già a posto).
    Filtri GET: reparto, corso, stato."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    from .services.training_eligibility import candidati_corso

    filtro_reparto = (request.GET.get("reparto") or "").strip()
    filtro_corso   = (request.GET.get("corso") or "").strip()
    filtro_stato   = (request.GET.get("stato") or "").strip()

    # Corsi obbligatori: bersaglio di una regola attiva e obbligatoria (corso o piano).
    corso_ids = set(
        TrainingRequirementRule.objects
        .filter(is_active=True, is_mandatory=True, corso__isnull=False)
        .values_list("corso_id", flat=True)
    )
    piano_ids = set(
        TrainingRequirementRule.objects
        .filter(is_active=True, is_mandatory=True, piano__isnull=False)
        .values_list("piano_id", flat=True)
    )
    if piano_ids:
        corso_ids |= set(
            TrainingCourse.objects.filter(piano_id__in=piano_ids, is_active=True)
            .values_list("id", flat=True)
        )
    corsi_all = list(
        TrainingCourse.objects.filter(id__in=corso_ids, is_active=True).order_by("titolo")
    )
    corsi_iter = (
        [c for c in corsi_all if str(c.pk) == filtro_corso] if filtro_corso.isdigit() else corsi_all
    )

    # Mappa reparto/mansione per dipendente (accessor legacy canonico).
    rep_map: dict[int, dict] = {}
    try:
        for r in fetch_anagrafica_rows():
            try:
                lid = int(r.get("id") or 0)
            except (TypeError, ValueError):
                continue
            rep_map[lid] = {
                "reparto": (r.get("reparto") or "").strip(),
                "mansione": (r.get("mansione") or "").strip(),
            }
    except Exception:
        logger.exception("Errore lookup reparti per copertura formativa")

    _ord = {"SCADUTO": 0, "IN_SCADENZA_30": 1, "IN_SCADENZA_90": 2, "MAI_FREQUENTATO": 3}
    righe = []
    for corso in corsi_iter:
        res = candidati_corso(corso)
        for c in res["idonei"] + res["non_idonei"]:
            if filtro_stato and c["stato"] != filtro_stato:
                continue
            info = rep_map.get(c["legacy_id"], {})
            reparto = info.get("reparto", "")
            if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
                continue
            righe.append({
                "legacy_id": c["legacy_id"],
                "nome": c["nome"],
                "reparto": reparto or "—",
                "mansione": info.get("mansione", "") or "—",
                "corso": corso,
                "stato": c["stato"],
                "stato_label": c["stato_label"],
                "data_scadenza": c["data_scadenza"],
            })
    righe.sort(key=lambda r: (r["reparto"].casefold(), r["nome"].casefold(), _ord.get(r["stato"], 9)))

    reparti = sorted({v["reparto"] for v in rep_map.values() if v.get("reparto")})

    return render(request, "anagrafica/pages/formazione_copertura.html", {
        "righe": righe,
        "corsi": corsi_all,
        "reparti": reparti,
        "n_dip": len({r["legacy_id"] for r in righe}),
        "n_gap": len(righe),
        "n_scaduti": sum(1 for r in righe if r["stato"] == "SCADUTO"),
        "n_mai": sum(1 for r in righe if r["stato"] == "MAI_FREQUENTATO"),
        "filtro_reparto": filtro_reparto,
        "filtro_corso": filtro_corso,
        "filtro_stato": filtro_stato,
        "is_editor": _can_edit_formazione(request),
    })


# ─────────────────────────────────────────────────────────────────────────────
# PLAN — calendario mese per mese di corsi, scadenze, DPI, assenze
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def formazione_plan(request, legacy_id: int | None = None):
    """Vista PLAN: eventi formazione/safety in 3 modalità.

    Senza `legacy_id` mostra il plan aggregato (tutto il personale).
    Con `legacy_id` filtra sui soli eventi del dipendente.

    Modalità (querystring `?view=`):
      - `mese`      (default) — card per mese con elenco eventi
      - `calendario` — griglia mensile classica (settimane × giorni)
      - `matrice`   — X giorni × Y dipendenti (planning resource view).
                      Disponibile solo nella vista globale (non per-dipendente).

    Eventi aggregati:
      - Sessioni formative (`TrainingSession`) per `data_inizio`, con elenco
        dipendenti iscritti (top 5 nomi + count totale).
      - Scadenze formazione (`TrainingDeadline`) per `data_scadenza`.
      - DPI scadenze — TODO PATCH successiva.
      - Assenze programmate — TODO PATCH successiva.

    Range default: mese corrente ± 3 mesi (7 mesi totali). Override via
    querystring `?from=YYYY-MM-DD&to=YYYY-MM-DD`. Per la vista `calendario` si
    può anche specificare `?anno=YYYY&mese=MM` per centrare su un singolo mese.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare la sezione formazione.")
        return redirect("anagrafica:index")

    from datetime import timedelta
    import calendar as _calendar
    from collections import OrderedDict, defaultdict

    today = date.today()
    view_mode = (request.GET.get("view") or "mese").lower()
    if view_mode not in {"mese", "calendario", "matrice"}:
        view_mode = "mese"

    def _parse_iso(s: str) -> date | None:
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None

    # ── Calcolo range a seconda della modalità ──────────────────────────────
    if view_mode == "calendario":
        try:
            cal_anno = int(request.GET.get("anno") or today.year)
            cal_mese = int(request.GET.get("mese") or today.month)
        except (TypeError, ValueError):
            cal_anno, cal_mese = today.year, today.month
        if cal_mese < 1 or cal_mese > 12:
            cal_mese = today.month
        cal_first = date(cal_anno, cal_mese, 1)
        cal_last = (_add_months(cal_first, 1) - timedelta(days=1))
        range_from, range_to = cal_first, cal_last
    else:
        range_from = _parse_iso(request.GET.get("from") or "") or _add_months(today.replace(day=1), -3)
        range_to_input = _parse_iso(request.GET.get("to") or "")
        if range_to_input is None:
            plus3 = _add_months(today.replace(day=1), 4)
            range_to = plus3 - timedelta(days=1)
        else:
            range_to = range_to_input
        if range_to < range_from:
            range_from, range_to = range_to, range_from
        cal_anno, cal_mese = today.year, today.month

    # ── Mappa nomi dipendenti ──────────────────────────────────────────────
    nomi_map = _build_nomi_map()

    # ── Raccolta eventi ────────────────────────────────────────────────────
    eventi: list[dict] = []

    # 1) Sessioni formative
    qs_sessioni = (
        TrainingSession.objects
        .select_related("corso", "corso__piano", "docente")
        .filter(data_inizio__gte=range_from, data_inizio__lte=range_to)
    )
    if legacy_id is not None:
        sess_ids_dip = list(
            TrainingEnrollment.objects
            .filter(legacy_anagrafica_id=legacy_id,
                    sessione__data_inizio__gte=range_from,
                    sessione__data_inizio__lte=range_to)
            .values_list("sessione_id", flat=True)
        )
        qs_sessioni = qs_sessioni.filter(pk__in=sess_ids_dip)

    # Pre-fetch iscritti per sessione (per mostrare nomi)
    sess_ids = [s.pk for s in qs_sessioni]
    iscritti_per_sess: dict[int, list[int]] = defaultdict(list)
    if sess_ids:
        for sess_id, lid in (
            TrainingEnrollment.objects
            .filter(sessione_id__in=sess_ids)
            .values_list("sessione_id", "legacy_anagrafica_id")
        ):
            iscritti_per_sess[sess_id].append(lid)

    for s in qs_sessioni:
        legacy_ids = iscritti_per_sess.get(s.pk, [])
        nomi_top = [nomi_map.get(lid, f"#{lid}") for lid in legacy_ids[:5]]
        eventi.append({
            "data":   s.data_inizio,
            "data_fine": s.data_fine,
            "tipo":   "CORSO",
            "tipo_label": "Corso",
            "titolo": f"{s.corso.codice} — {s.corso.titolo[:60]}",
            "sub":    f"Sessione {s.codice_sessione} · {s.get_stato_display()}",
            "url":    f"/anagrafica/formazione/sessioni/{s.pk}/",
            "stato":  s.stato,
            "badge":  s.stato.lower(),
            "iscritti_count": len(legacy_ids),
            "iscritti_nomi":  nomi_top,
            "iscritti_extra": max(0, len(legacy_ids) - 5),
            "iscritti_legacy_ids": legacy_ids,
        })

    # 2) Scadenze formazione
    qs_scad = (
        TrainingDeadline.objects
        .select_related("corso", "corso__piano")
        .filter(data_scadenza__gte=range_from, data_scadenza__lte=range_to)
        .exclude(stato_scadenza="UNA_TANTUM")
    )
    if legacy_id is not None:
        qs_scad = qs_scad.filter(legacy_anagrafica_id=legacy_id)

    for d in qs_scad:
        nome_dip = nomi_map.get(d.legacy_anagrafica_id, "") if legacy_id is None else ""
        sub_parts = [d.get_stato_scadenza_display()]
        if nome_dip:
            sub_parts.insert(0, nome_dip)
        eventi.append({
            "data":   d.data_scadenza,
            "data_fine": d.data_scadenza,
            "tipo":   "SCADENZA",
            "tipo_label": "Scadenza",
            "titolo": f"{d.corso.codice} — {d.corso.titolo[:60]}",
            "sub":    " · ".join(sub_parts),
            "url":    f"/anagrafica/formazione/corsi/{d.corso_id}/",
            "stato":  d.stato_scadenza,
            "badge":  d.stato_scadenza.lower(),
            "iscritti_count": 1,
            "iscritti_nomi":  [nome_dip] if nome_dip else [],
            "iscritti_extra": 0,
            "iscritti_legacy_ids": [d.legacy_anagrafica_id],
        })

    # ── 3) Raggruppamento per mese (modalità "mese", per i KPI riepilogo) ──
    mesi: "OrderedDict[tuple[int,int], dict]" = OrderedDict()
    cursor = range_from.replace(day=1)
    end_month = range_to.replace(day=1)
    while cursor <= end_month:
        mesi[(cursor.year, cursor.month)] = {
            "anno": cursor.year, "mese": cursor.month,
            "label_mese": cursor.strftime("%B %Y").capitalize(),
            "is_corrente": (cursor.year, cursor.month) == (today.year, today.month),
            "is_passato":  cursor < today.replace(day=1),
            "is_futuro":   cursor > today.replace(day=1),
            "eventi": [],
        }
        cursor = _add_months(cursor, 1)
    for ev in eventi:
        key = (ev["data"].year, ev["data"].month)
        if key in mesi:
            mesi[key]["eventi"].append(ev)
    for m in mesi.values():
        m["eventi"].sort(key=lambda e: (e["data"], e["tipo"]))
        m["n_eventi"] = len(m["eventi"])
        m["n_corsi"]    = sum(1 for e in m["eventi"] if e["tipo"] == "CORSO")
        m["n_scadenze"] = sum(1 for e in m["eventi"] if e["tipo"] == "SCADENZA")

    # ── 4) Griglia calendario mensile (modalità "calendario") ──────────────
    calendar_weeks: list[list[dict]] = []
    if view_mode == "calendario":
        # Index eventi per giorno
        eventi_per_giorno: dict[date, list[dict]] = defaultdict(list)
        for ev in eventi:
            eventi_per_giorno[ev["data"]].append(ev)
        cal = _calendar.Calendar(firstweekday=0)  # 0 = lunedì
        for week in cal.monthdatescalendar(cal_anno, cal_mese):
            row = []
            for d in week:
                row.append({
                    "data": d,
                    "is_other_month": d.month != cal_mese,
                    "is_today": d == today,
                    "eventi": sorted(eventi_per_giorno.get(d, []), key=lambda e: e["tipo"]),
                })
            calendar_weeks.append(row)

    # ── 5) Matrice giorni × dipendenti (modalità "matrice") ────────────────
    matrice = None
    filtro_reparto  = (request.GET.get("reparto") or "").strip()
    filtro_mansione = (request.GET.get("mansione") or "").strip()
    filtro_area     = (request.GET.get("area") or "").strip()
    filtro_ruolo    = (request.GET.get("ruolo") or "").strip()
    if view_mode == "matrice" and legacy_id is None:
        # Cap dimensione: max 60 giorni.
        n_days = (range_to - range_from).days + 1
        if n_days > 60:
            range_to = range_from + timedelta(days=59)
            n_days = 60
        giorni = [range_from + timedelta(days=i) for i in range(n_days)]

        # Aggrega per (legacy_id, giorno) → list di eventi
        cell_index: dict[tuple[int, date], list[dict]] = defaultdict(list)
        legacy_ids_coinvolti: set[int] = set()
        for ev in eventi:
            for lid in (ev.get("iscritti_legacy_ids") or []):
                cell_index[(lid, ev["data"])].append(ev)
                legacy_ids_coinvolti.add(lid)

        # Carica attributi dipendente (mansione/reparto da legacy; area da DipendenteAnagraficaAziendale)
        from core.legacy_models import AnagraficaDipendente as _LegAna
        dip_attr: dict[int, dict] = {}
        if legacy_ids_coinvolti:
            for r in _LegAna.objects.filter(id__in=list(legacy_ids_coinvolti)).values("id", "mansione", "reparto"):
                dip_attr[int(r["id"])] = {
                    "mansione": (r.get("mansione") or "").strip(),
                    "reparto":  (r.get("reparto") or "").strip(),
                    "area":     "",
                }
            for r in DipendenteAnagraficaAziendale.objects.filter(
                legacy_anagrafica_id__in=list(legacy_ids_coinvolti)
            ).values("legacy_anagrafica_id", "area", "ruolo_aziendale"):
                lid = int(r["legacy_anagrafica_id"])
                attrs = dip_attr.setdefault(lid, {"mansione": "", "reparto": "", "area": "", "ruolo": ""})
                attrs["area"]  = (r.get("area") or "").strip()
                attrs["ruolo"] = (r.get("ruolo_aziendale") or "").strip()

        # Valori distinti per i menu filtro (calcolati PRIMA di filtrare, su tutti i coinvolti)
        opzioni_reparto  = sorted({a.get("reparto", "")  for a in dip_attr.values() if a.get("reparto")})
        opzioni_mansione = sorted({a.get("mansione", "") for a in dip_attr.values() if a.get("mansione")})
        opzioni_area     = sorted({a.get("area", "")     for a in dip_attr.values() if a.get("area")})
        opzioni_ruolo    = sorted({a.get("ruolo", "")    for a in dip_attr.values() if a.get("ruolo")})

        # Applica filtri sulle righe
        def _row_passes(lid: int) -> bool:
            a = dip_attr.get(lid) or {}
            if filtro_reparto  and a.get("reparto", "") != filtro_reparto:  return False
            if filtro_mansione and a.get("mansione", "") != filtro_mansione: return False
            if filtro_area     and a.get("area", "") != filtro_area:         return False
            if filtro_ruolo    and a.get("ruolo", "") != filtro_ruolo:       return False
            return True

        rows_legacy = sorted(
            (lid for lid in legacy_ids_coinvolti if _row_passes(lid)),
            key=lambda lid: nomi_map.get(lid, f"#{lid}").lower(),
        )
        rows = []
        for lid in rows_legacy:
            a = dip_attr.get(lid) or {}
            cells = []
            for g in giorni:
                evs = cell_index.get((lid, g), [])
                cells.append({"data": g, "is_today": g == today, "eventi": evs})
            rows.append({
                "legacy_id": lid,
                "nome":      nomi_map.get(lid, f"#{lid}"),
                "reparto":   a.get("reparto", ""),
                "mansione":  a.get("mansione", ""),
                "area":      a.get("area", ""),
                "ruolo":     a.get("ruolo", ""),
                "cells":     cells,
                "n_eventi":  sum(len(c["eventi"]) for c in cells),
            })

        matrice = {
            "giorni":  giorni,
            "rows":    rows,
            "n_rows":  len(rows),
            "n_totale_dipendenti": len(legacy_ids_coinvolti),
            "n_days":  n_days,
            "opzioni_reparto":  opzioni_reparto,
            "opzioni_mansione": opzioni_mansione,
            "opzioni_area":     opzioni_area,
            "opzioni_ruolo":    opzioni_ruolo,
            "filtro_reparto":   filtro_reparto,
            "filtro_mansione":  filtro_mansione,
            "filtro_area":      filtro_area,
            "filtro_ruolo":     filtro_ruolo,
            "any_filter":       bool(filtro_reparto or filtro_mansione or filtro_area or filtro_ruolo),
        }

    # ── Contesto dipendente (per vista per-dipendente) ─────────────────────
    contesto_dipendente = None
    if legacy_id is not None:
        from core.legacy_models import AnagraficaDipendente
        dip = AnagraficaDipendente.objects.filter(id=legacy_id).first()
        if dip:
            contesto_dipendente = {
                "legacy_id": legacy_id,
                "nome":      f"{(dip.cognome or '').strip()} {(dip.nome or '').strip()}".strip() or f"#{legacy_id}",
            }

    # Mese precedente/successivo per nav calendario
    cal_prev = _add_months(date(cal_anno, cal_mese, 1), -1)
    cal_next = _add_months(date(cal_anno, cal_mese, 1), 1)

    return render(request, "anagrafica/pages/formazione_plan.html", {
        "view_mode":       view_mode,
        "mesi":            list(mesi.values()),
        "calendar_weeks":  calendar_weeks,
        "cal_anno":        cal_anno,
        "cal_mese":        cal_mese,
        "cal_mese_label":  date(cal_anno, cal_mese, 1).strftime("%B %Y").capitalize(),
        "cal_prev_anno":   cal_prev.year, "cal_prev_mese": cal_prev.month,
        "cal_next_anno":   cal_next.year, "cal_next_mese": cal_next.month,
        "matrice":         matrice,
        "range_from":      range_from,
        "range_to":        range_to,
        "totale_eventi":   sum(m["n_eventi"] for m in mesi.values()),
        "totale_corsi":    sum(m["n_corsi"] for m in mesi.values()),
        "totale_scadenze": sum(m["n_scadenze"] for m in mesi.values()),
        "dip":             contesto_dipendente,
        "is_editor":       _can_edit_formazione(request),
        "today":           today,
    })


# ─────────────────────────────────────────────────────────────────────────────
# RISCHI / CATEGORIE / ESPOSIZIONI — CRUD (PATCH-RISK-02)
# ─────────────────────────────────────────────────────────────────────────────

from .models_rischi import CategoriaCorso, EsposizioneRischio, FattoreRischio  # noqa: E402
from .forms import (  # noqa: E402
    CategoriaCorsoForm,
    EsposizioneRischioForm,
    FattoreRischioForm,
)


# ── Fattori di Rischio ──────────────────────────────────────────────────────

@login_required
def fattori_rischio_list(request, form=None, open_create=False, status=200):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare i fattori di rischio.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    fattori = list(
        FattoreRischio.objects
        .annotate(n_categorie=Count("categorie_corso", distinct=True),
                  n_esposizioni=Count("esposizioni", distinct=True))
        .prefetch_related("tipi_visita", "categorie_dpi")
        .order_by("categoria", "nome")
    )
    # Id selezionati per il modale di modifica JS (requisiti generati dal fattore).
    for f in fattori:
        f.sel_visite_ids = [t.pk for t in f.tipi_visita.all()]
        f.sel_dpi_ids = [c.pk for c in f.categorie_dpi.all()]

    visite_opts = list(
        TipoVisitaMedica.objects.filter(is_active=True).order_by("nome").values("id", "nome")
    )
    dpi_opts: list[dict] = []
    try:
        from dpi.models import CategoriaDPI
        dpi_opts = list(
            CategoriaDPI.objects.filter(is_active=True)
            .order_by("order_index", "nome").values("id", "nome")
        )
    except Exception:
        dpi_opts = []

    if form is None:
        form = FattoreRischioForm()
    return render(request, "anagrafica/pages/rischi_fattori_list.html", {
        "fattori":   fattori,
        "form":      form,
        "is_editor": is_editor,
        "CATEGORIA_CHOICES": FattoreRischio.CATEGORIA_CHOICES,
        "visite_opts": visite_opts,
        "dpi_opts":    dpi_opts,
        "open_create": open_create,
    }, status=status)


@login_required
@require_POST
def fattore_rischio_create(request):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    form = FattoreRischioForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Fattore di rischio creato.")
        return redirect("anagrafica:fattori_rischio_list")
    messages.error(request, "Correggi gli errori evidenziati nel form.")
    return fattori_rischio_list(request, form=form, open_create=True, status=400)


@login_required
@require_POST
def fattore_rischio_edit(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(FattoreRischio, pk=pk)
    form = FattoreRischioForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, f"Fattore «{obj.nome}» aggiornato.")
    else:
        messages.error(request, "Errore aggiornamento: " + "; ".join(f"{k}: {v}" for k, v in form.errors.items()))
    return redirect("anagrafica:fattori_rischio_list")


@login_required
@require_POST
def fattore_rischio_delete(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(FattoreRischio, pk=pk)
    if obj.categorie_corso.exists() or obj.esposizioni.exists():
        obj.is_active = False
        obj.save(update_fields=["is_active", "updated_at"])
        messages.warning(request, f"Fattore «{obj.nome}» disattivato (è collegato a categorie/esposizioni).")
    else:
        obj.delete()
        messages.success(request, "Fattore di rischio eliminato.")
    return redirect("anagrafica:fattori_rischio_list")


# ── Categorie Corso ──────────────────────────────────────────────────────────

@login_required
def categorie_corso_list(request, form=None, open_create=False, status=200):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare le categorie corso.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    categorie = list(
        CategoriaCorso.objects
        .prefetch_related("fattori_rischio")
        .annotate(n_corsi=Count("corsi", distinct=True))
        .order_by("nome")
    )
    if form is None:
        form = CategoriaCorsoForm()
    return render(request, "anagrafica/pages/rischi_categorie_list.html", {
        "categorie": categorie,
        "form":      form,
        "is_editor": is_editor,
        "open_create": open_create,
    }, status=status)


@login_required
@require_POST
def categoria_corso_create(request):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    form = CategoriaCorsoForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Categoria corso creata.")
        return redirect("anagrafica:categorie_corso_list")
    messages.error(request, "Correggi gli errori evidenziati nel form.")
    return categorie_corso_list(request, form=form, open_create=True, status=400)


@login_required
@require_POST
def categoria_corso_edit(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(CategoriaCorso, pk=pk)
    form = CategoriaCorsoForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, f"Categoria «{obj.nome}» aggiornata.")
    else:
        messages.error(request, "Errore aggiornamento: " + "; ".join(f"{k}: {v}" for k, v in form.errors.items()))
    return redirect("anagrafica:categorie_corso_list")


@login_required
@require_POST
def categoria_corso_delete(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(CategoriaCorso, pk=pk)
    if obj.corsi.exists():
        obj.is_active = False
        obj.save(update_fields=["is_active", "updated_at"])
        messages.warning(request, f"Categoria «{obj.nome}» disattivata (collegata a corsi).")
    else:
        obj.delete()
        messages.success(request, "Categoria eliminata.")
    return redirect("anagrafica:categorie_corso_list")


# ── Esposizioni Rischio ──────────────────────────────────────────────────────

@login_required
def esposizioni_rischio_list(request, form=None, open_create=False, status=200):
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per visualizzare le esposizioni.")
        return redirect("anagrafica:index")
    is_editor = _can_edit_formazione(request)
    esposizioni = list(
        EsposizioneRischio.objects
        .select_related("fattore", "mansione", "area")
        .order_by("fattore__categoria", "fattore__nome", "mansione__nome", "area__nome")
    )
    if form is None:
        form = EsposizioneRischioForm()
    return render(request, "anagrafica/pages/rischi_esposizioni_list.html", {
        "esposizioni": esposizioni,
        "form":        form,
        "is_editor":   is_editor,
        "open_create": open_create,
    }, status=status)


@login_required
@require_POST
def esposizione_rischio_create(request):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    form = EsposizioneRischioForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, "Esposizione creata.")
        return redirect("anagrafica:esposizioni_rischio_list")
    messages.error(request, "Correggi gli errori evidenziati nel form.")
    return esposizioni_rischio_list(request, form=form, open_create=True, status=400)


@login_required
@require_POST
def esposizione_rischio_edit(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(EsposizioneRischio, pk=pk)
    form = EsposizioneRischioForm(request.POST, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, "Esposizione aggiornata.")
    else:
        messages.error(request, "Errore aggiornamento: " + "; ".join(f"{k}: {v}" for k, v in form.errors.items()))
    return redirect("anagrafica:esposizioni_rischio_list")


@login_required
@require_POST
def esposizione_rischio_delete(request, pk: int):
    if not _can_edit_formazione(request):
        return _forbid_json_or_redirect(request, "Permessi insufficienti.")
    obj = get_object_or_404(EsposizioneRischio, pk=pk)
    obj.delete()
    messages.success(request, "Esposizione eliminata.")
    return redirect("anagrafica:esposizioni_rischio_list")


def _forbid_json_or_redirect(request, msg: str):
    """Helper: 403 JSON per AJAX, altrimenti messaggio flash + redirect."""
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({"ok": False, "error": msg}, status=403)
    messages.error(request, msg)
    return redirect("anagrafica:index")


# ---------------------------------------------------------------------------
# Organigramma visuale — Reparto → Aree aziendali → caporeparto → dipendenti
# ---------------------------------------------------------------------------

@login_required
def organigramma(request):
    """Organigramma navigabile SSR: reparti, aree aziendali, capi e membri.

    Il reparto di ogni dipendente è risolto dalla **fonte unica canonica**
    (`dipendente → area_aziendale → reparto`, via
    :func:`anagrafica.services.reparto_canonico.build_reparto_canonico_map`),
    con fallback al testo legacy finché la copertura canonica non è completa.
    I dipendenti che non si risolvono per nessuna via finiscono nel bucket
    "Non mappati", che resta come spia dei residui da bonificare.
    """
    from anagrafica.services.reparto_canonico import (
        build_area_canonica_map,
        build_reparto_canonico_map,
        resolve_reparto_for_row,
        resolve_responsabile_effettivo,
    )

    ensure_anagrafica_schema()
    filtro_reparto = (request.GET.get("reparto") or "").strip()

    dip_rows = [r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("attivo")]
    dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}

    reparti = list(
        Reparto.objects.filter(is_active=True)
        .prefetch_related("aree_aziendali")
        .order_by("nome")
    )
    reparto_by_name = {r.nome.strip().casefold(): r for r in reparti}

    legacy_ids = list(dip_map.keys())
    canonico_map = build_reparto_canonico_map(legacy_ids)
    area_map = build_area_canonica_map(legacy_ids)

    membri_per_reparto: dict[int, list[dict]] = {}
    non_mappati: list[dict] = []
    for row in dip_rows:
        rep = resolve_reparto_for_row(row, canonico_map=canonico_map, reparto_by_name=reparto_by_name)
        # Area aziendale canonica del membro (per la vista); None se non assegnata.
        area = area_map.get(int(row.get("id") or 0))
        row["area_aziendale_nome"] = area.nome if area is not None else ""
        if rep is None:
            non_mappati.append(row)
        else:
            membri_per_reparto.setdefault(rep.id, []).append(row)

    def _sort_key(row: dict):
        return (str(row.get("cognome") or "").casefold(), str(row.get("nome") or "").casefold())

    def _blocco_reparto(rep: Reparto) -> dict:
        capo = dip_map.get(rep.caporeparto_legacy_id or 0)
        membri = sorted(membri_per_reparto.get(rep.id, []), key=_sort_key)
        if capo:
            membri = [m for m in membri if int(m.get("id") or 0) != int(capo.get("id") or 0)]
        aree = list(rep.aree_aziendali.filter(is_active=True).order_by("nome"))
        # Responsabile effettivo per AREA: il responsabile dell'area vince sul
        # caporeparto del reparto quando differisce (dominio: "caporeparto
        # dall'area aziendale se differisce"). Il capo del blocco resta il
        # caporeparto del reparto come fallback complessivo.
        for area in aree:
            rid = resolve_responsabile_effettivo(area=area, reparto=rep)
            area.responsabile_effettivo_id = rid
            resp_row = dip_map.get(rid or 0)
            area.responsabile_effettivo_label = (
                f"{resp_row.get('cognome', '')} {resp_row.get('nome', '')}".strip()
                if resp_row else ""
            )
            area.responsabile_distinto = bool(
                area.responsabile_legacy_id
                and area.responsabile_legacy_id != (rep.caporeparto_legacy_id or 0)
            )
        return {
            "reparto": rep,
            "capo": capo,
            "membri": membri,
            "aree_aziendali": aree,
            "n_totale": len(membri) + (1 if capo else 0),
        }

    blocchi = [_blocco_reparto(r) for r in reparti]

    nomi_reparti = [r.nome for r in reparti]
    if filtro_reparto:
        blocchi = [b for b in blocchi if b["reparto"].nome.casefold() == filtro_reparto.casefold()]

    non_mappati.sort(key=_sort_key)
    n_dipendenti = len(dip_rows)

    return render(request, "anagrafica/pages/organigramma.html", {
        "blocchi": blocchi,
        "non_mappati": non_mappati,
        "nomi_reparti": nomi_reparti,
        "filtro_reparto": filtro_reparto,
        "n_dipendenti": n_dipendenti,
        "n_reparti": len(reparti),
        "n_non_mappati": len(non_mappati),
    })


@login_required
def organigramma_albero(request):
    """Organigramma ad albero: gerarchia dei RUOLI (RuoloOperativo.riporta_a),
    persone come foglie titolari. Con ?certificazione=<TipoQualifica.pk> mostra
    la copertura ad albero della certificazione (chi la possiede). La gerarchia
    è SEMPRE tra ruoli, mai tra persone."""
    from anagrafica.services.organigramma_albero import (
        build_ruolo_albero, build_certificazione_copertura,
    )
    raw = (request.GET.get("certificazione") or "").strip()
    cert_id = int(raw) if raw.isdigit() else None
    albero = build_certificazione_copertura(cert_id) if cert_id else build_ruolo_albero()
    return render(request, "anagrafica/pages/organigramma_albero.html", {
        "albero": albero,
        "cert_id": cert_id,
        "certificazioni": TipoQualifica.objects.filter(is_active=True).order_by("nome"),
    })


# ---------------------------------------------------------------------------
# Fascicolo conformità — "il dipendente è in regola con la sua mansione?"
# (H2: semaforo aggregato formazione + visite + qualifiche + DPI)
# ---------------------------------------------------------------------------

@login_required
def dipendente_conformita_panel(request, legacy_id: int):
    """Pannello semaforo conformità per la scheda dipendente (HTMX lazy-load).

    Accesso: login. Il semaforo visite mostra solo l'esito valido/scaduto; il
    dettaglio (nomi tipologia) è incluso solo se ``_can_view_visite_mediche``.
    Mai esiti/prescrizioni. La logica privacy è applicata dal service via
    ``include_visite_dettaglio``.
    """
    can_view_visite = _can_view_visite_mediche(request)
    mansione_nome = (
        AnagraficaDipendente.objects
        .filter(id=legacy_id).values_list("mansione", flat=True).first()
    ) or ""
    stato = conformita_service.stato_conformita(
        legacy_id,
        include_visite_dettaglio=can_view_visite,
        mansione=mansione_nome,
    )
    mansione_obj = (
        Mansione.objects.filter(nome__iexact=mansione_nome.strip()).only("id").first()
        if mansione_nome else None
    )
    return render(request, "anagrafica/partials/conformita_panel.html", {
        "legacy_id": legacy_id,
        "stato": stato,
        "mansione_nome": mansione_nome,
        "mansione_id": mansione_obj.id if mansione_obj else None,
        "can_view_visite": can_view_visite,
    })


@login_required
def dipendente_verbale_dpi(request, legacy_id: int):
    """Verbale di consegna DPI (MOD.155) precompilato con i DPI richiesti dalla
    mansione del dipendente. Pagina stampabile (``window.print()``) da firmare;
    riusa i requisiti del resolver mansionario, niente record di consegna creati."""
    from django.utils import timezone
    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]
    mansione_nome = str(dip.get("mansione") or "").strip()
    requisiti = (
        mansionario_service.requisiti_per_nome_mansione(mansione_nome)
        if mansione_nome else mansionario_service.requisiti_vuoti()
    )
    nome = f"{str(dip.get('cognome') or '').strip()} {str(dip.get('nome') or '').strip()}".strip()
    return render(request, "anagrafica/pages/verbale_dpi.html", {
        "dip": dip,
        "nome": nome,
        "mansione": mansione_nome,
        "reparto": str(dip.get("reparto") or "").strip(),
        "dpi": requisiti["dpi"],
        "oggi": timezone.localdate(),
    })


@login_required
def sicurezza_hub(request):
    """Cruscotto "Sicurezza & Idoneità": numeri chiave e collegamenti a tutte le
    sezioni correlate (fattori, esposizioni, mansioni di rischio, conformità,
    catalogo DPI), così le pagine non risultano isolate. Le metriche di idoneità
    (aggregate, non nominative) sono mostrate solo con permesso HR.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per il cruscotto sicurezza.")
        return redirect("anagrafica:index")
    can_hr = _check_hr_permission(request)

    n_mansioni = Mansione.objects.filter(is_active=True).count()
    n_mansioni_rischio = (
        Mansione.objects.filter(is_active=True)
        .filter(Q(dpi_richiesti__isnull=False) | Q(visite_richieste__isnull=False)
                | ~Q(livello_rischio=""))
        .distinct().count()
    )

    idoneita_kpi = None
    if can_hr:
        dip_rows = [r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("attivo")]
        dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}
        mansioni_per_legacy = {
            lid: str(d.get("mansione") or "").strip()
            for lid, d in dip_map.items() if str(d.get("mansione") or "").strip()
        }
        stati = conformita_service.stato_conformita_batch(
            list(dip_map), mansioni_per_legacy=mansioni_per_legacy
        )
        c = {"ok": 0, "warn": 0, "ko": 0, "na": 0}
        for s in stati.values():
            e = s.get("idoneita", {}).get("esito", "na")
            c[e] = c.get(e, 0) + 1
        idoneita_kpi = {
            "idonei": c["ok"], "riserve": c["warn"], "non_idonei": c["ko"],
            "valutati": c["ok"] + c["warn"] + c["ko"],
        }

    return render(request, "anagrafica/pages/sicurezza_hub.html", {
        "can_hr": can_hr,
        "n_mansioni": n_mansioni,
        "n_mansioni_rischio": n_mansioni_rischio,
        "idoneita_kpi": idoneita_kpi,
    })


@login_required
def sicurezza_ricerca(request):
    """Ricerca dell'area Sicurezza & Idoneità: mansioni di rischio, dipendenti
    (idoneità), qualifiche di sicurezza e fattori di rischio. Speculare alla
    ricerca globale formazione, ma con le categorie pertinenti alla safety.
    Gated dallo stesso permesso del cruscotto sicurezza.
    """
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per il cruscotto sicurezza.")
        return redirect("anagrafica:index")

    from .models_rischi import FattoreRischio

    q = (request.GET.get("q") or "").strip()
    risultati: dict[str, list] = {
        "mansioni": [], "dipendenti": [], "qualifiche": [], "fattori": [],
    }
    totale = 0
    if q:
        ql = q.lower()

        from django.db.models import Count as _Count
        risultati["mansioni"] = list(
            Mansione.objects.filter(is_active=True)
            .filter(Q(nome__icontains=q) | Q(descrizione__icontains=q))
            .annotate(n_esposizioni=_Count("esposizioni_rischio", distinct=True))
            .order_by("nome")[:25]
        )
        risultati["qualifiche"] = list(
            TipoQualifica.objects.filter(categoria=TipoQualifica.CAT_SICUREZZA)
            .filter(Q(nome__icontains=q) | Q(descrizione__icontains=q))
            .annotate(n_assegnazioni=_Count("assegnazioni", distinct=True))
            .order_by("nome")[:25]
        )
        risultati["fattori"] = list(
            FattoreRischio.objects.filter(
                Q(nome__icontains=q) | Q(codice__icontains=q) | Q(descrizione__icontains=q)
            ).annotate(n_esposizioni=_Count("esposizioni", distinct=True))
            .order_by("nome")[:25]
        )

        nomi = _build_nomi_map()
        dip_match = [
            {"legacy_id": lid, "nome": nome}
            for lid, nome in nomi.items() if ql in nome.lower()
        ]
        dip_match.sort(key=lambda d: d["nome"].casefold())
        risultati["dipendenti"] = dip_match[:25]

        totale = sum(len(v) for v in risultati.values())

    # Ricerca live: su richiesta HTMX rende solo il frammento dei risultati;
    # con ?suggest=1 (casella nella dashboard) rende la tendina compatta.
    if request.headers.get("HX-Request"):
        template = (
            "anagrafica/partials/_safety_search_suggest.html"
            if request.GET.get("suggest")
            else "anagrafica/partials/_safety_search_results.html"
        )
    else:
        template = "anagrafica/pages/sicurezza_ricerca.html"
    return render(request, template, {
        "q": q,
        "risultati": risultati,
        "totale": totale,
    })


@login_required
def sicurezza_wizard(request):
    """Configurazione guidata della "mansione di rischio": passi in ordine, con
    stato (fatto/da fare) calcolato dai dati reali e CTA verso ogni sezione.
    Rende esplicito il flusso e collega le sezioni tra loro."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per la guida sicurezza.")
        return redirect("anagrafica:index")

    n_mansioni_rischio = (
        Mansione.objects.filter(is_active=True)
        .filter(Q(dpi_richiesti__isnull=False) | Q(visite_richieste__isnull=False)
                | ~Q(livello_rischio=""))
        .distinct().count()
    )
    try:
        from dpi.models import CategoriaDPI
        n_dpi = CategoriaDPI.objects.filter(is_active=True).count()
    except Exception:
        n_dpi = 0
    n_visite = TipoVisitaMedica.objects.filter(is_active=True).count()

    steps = [
        {
            "n": 1, "titolo": "Cataloghi di base",
            "desc": "Verifica che esistano le categorie DPI e le tipologie di visita medica da assegnare alle mansioni.",
            "stato": "ok" if (n_dpi or n_visite) else "todo",
            "info": f"{n_dpi} categorie DPI · {n_visite} tipi visita",
            "url": reverse("dpi:impostazioni"), "cta": "Catalogo DPI",
            "url2": reverse("anagrafica:impostazioni") + "?tab=visite", "cta2": "Tipi visita",
        },
        {
            "n": 2, "titolo": "Requisiti delle mansioni",
            "desc": "Su ogni mansione imposta il livello di rischio e i requisiti (DPI, visite, formazione). Il rischio è dedotto dalla mansione e dai DPI associati.",
            "stato": "ok" if n_mansioni_rischio else "todo",
            "info": f"{n_mansioni_rischio} mansioni di rischio configurate",
            "url": reverse("anagrafica:mansioni_list"), "cta": "Mansioni di rischio",
        },
        {
            "n": 3, "titolo": "Verifica idoneità",
            "desc": "Controlla il semaforo di idoneità di ogni dipendente: requisiti mancanti (avviso) o scaduti (non idoneo). Vedi anche la matrice competenze.",
            "stato": "info",
            "info": "Report trasversale su tutto il personale",
            "url": reverse("anagrafica:conformita_report"), "cta": "Apri conformità",
        },
    ]
    return render(request, "anagrafica/pages/sicurezza_wizard.html", {"steps": steps})


@login_required
def matrice_competenze(request):
    """Matrice **dipendenti × competenze/abilitazioni** (qualifiche) per audit ISO 45001.

    Celle: valido / in scadenza (≤60gg) / scaduto / mancante. Colonne = i
    ``TipoQualifica`` con almeno un'assegnazione (competenze realmente in uso,
    incl. quelle importate dall'ASR). Filtro reparto + export CSV.
    Accesso: ``_check_hr_permission`` (vista trasversale sul personale).
    """
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per la matrice competenze.")
        return redirect("anagrafica:index")

    from datetime import timedelta
    from django.utils import timezone
    ensure_anagrafica_schema()
    filtro_reparto = (request.GET.get("reparto") or "").strip()
    export_csv = request.GET.get("format") == "csv"
    valid_cats = {c for c, _ in TipoQualifica.CATEGORIA_CHOICES}
    cat_filter = (request.GET.get("categoria") or "").strip().upper()
    if cat_filter not in valid_cats:
        cat_filter = ""
    oggi = timezone.localdate()
    soglia = oggi + timedelta(days=60)

    dip_rows = [r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("attivo")]
    dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}
    legacy_ids = list(dip_map.keys())

    # Colonne = TipoQualifica con almeno un'assegnazione; tab per categoria.
    cat_labels = dict(TipoQualifica.CATEGORIA_CHOICES)
    tipi_all = list(
        TipoQualifica.objects.annotate(_n=Count("assegnazioni"))
        .filter(_n__gt=0).order_by("categoria", "nome")
    )
    tabs = [("", "Tutte", len(tipi_all))]
    for cat_code, cat_lbl in TipoQualifica.CATEGORIA_CHOICES:
        n = sum(1 for t in tipi_all if t.categoria == cat_code)
        if n:
            tabs.append((cat_code, cat_lbl, n))
    tipi = [t for t in tipi_all if not cat_filter or t.categoria == cat_filter]
    tipo_ids = [t.id for t in tipi]
    q_map: dict[tuple[int, int], DipendenteQualifica] = {}
    for q in DipendenteQualifica.objects.filter(
        legacy_anagrafica_id__in=legacy_ids, tipo_id__in=tipo_ids
    ):
        q_map[(q.legacy_anagrafica_id, q.tipo_id)] = q

    def _stato(q):
        if q is None:
            return "mancante"
        if q.data_scadenza is None:
            return "valido"
        if q.data_scadenza < oggi:
            return "scaduto"
        if q.data_scadenza <= soglia:
            return "in_scadenza"
        return "valido"

    righe: list[dict] = []
    for lid, dip in dip_map.items():
        reparto = str(dip.get("reparto") or "").strip()
        if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
            continue
        celle = []
        for t in tipi:
            q = q_map.get((lid, t.id))
            celle.append({"stato": _stato(q), "data": q.data_scadenza if q else None})
        righe.append({
            "legacy_id": lid,
            "cognome": str(dip.get("cognome") or f"ID {lid}").strip(),
            "nome": str(dip.get("nome") or "").strip(),
            "reparto": reparto,
            "celle": celle,
        })
    righe.sort(key=lambda r: (r["cognome"].casefold(), r["nome"].casefold()))
    reparti = sorted({r["reparto"] for r in righe if r["reparto"]})

    if export_csv:
        _LAB = {"valido": "OK", "in_scadenza": "In scadenza", "scaduto": "SCADUTO", "mancante": "—"}
        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="matrice_competenze.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(resp, delimiter=";")
        writer.writerow(["Dipendente", "Reparto"] + [t.nome for t in tipi])
        for r in righe:
            cells = []
            for c in r["celle"]:
                lab = _LAB.get(c["stato"], c["stato"])
                if c["data"]:
                    lab = f"{lab} {c['data']:%d/%m/%Y}"
                cells.append(lab)
            writer.writerow([f"{r['cognome']} {r['nome']}".strip(), r["reparto"]] + cells)
        return resp

    return render(request, "anagrafica/pages/matrice_competenze.html", {
        "tipi": tipi,
        "righe": righe,
        "reparti": reparti,
        "filtro_reparto": filtro_reparto,
        "totale": len(righe),
        "tabs": tabs,
        "active_categoria": cat_filter,
        "active_categoria_label": cat_labels.get(cat_filter, ""),
    })


# ---------------------------------------------------------------------------
# Skill Matrix MOD.187 — validazione abbinamento competenza→asset (F2a, UI)
# ---------------------------------------------------------------------------

@login_required
def skm_match_validazione(request):
    """Specchietto in-portal per abbinare le competenze-macchina MOD.187 agli
    asset **reali** dell'ambiente (dev o prod) e **confermare** il match — è il
    gate F2a, propedeutico all'import baseline.

    Lavora sugli asset live e salva nel catalogo ``CompetenzaSkm``: NON scrive
    baseline (nessuna ``AbilitazioneMacchina``). Le conferme manuali sono
    preservate dalle ri-sincronizzazioni.
    Accesso: ``_check_hr_permission`` (ACL formale in F7).
    """
    from .acl_bootstrap import PERM_SKM_MANAGE
    if not _check_skm_permission(request, PERM_SKM_MANAGE):
        messages.error(request, "Non hai i permessi per la validazione Skill Matrix.")
        return redirect("anagrafica:index")

    from .models import CompetenzaSkm
    from .services.skillmatrix_seed import sincronizza_catalogo

    if request.method == "POST":
        azione = request.POST.get("azione", "")
        if azione == "sincronizza":
            stats = sincronizza_catalogo()
            messages.success(
                request,
                f"Catalogo sincronizzato dagli asset live: {stats['macchine']} macchine "
                f"({stats['esatti']} esatti, {stats['parziali']} parziali, "
                f"{stats['assenti']} assenti; {stats['confermati']} confermati). "
                f"Processi {stats['processi']} (collegati a qualifica {stats['processi_collegati']}).",
            )
            return redirect("anagrafica:skm_match_validazione")

        if azione == "salva":
            from assets.models import Asset
            macchine = list(CompetenzaSkm.objects.filter(tipo=CompetenzaSkm.TIPO_MACCHINA))
            tag_richiesti = {
                (request.POST.get(f"asset_{c.id}") or "").strip()
                for c in macchine
            }
            tag_richiesti.discard("")
            asset_per_tag = {
                a.asset_tag: a for a in Asset.objects.filter(asset_tag__in=tag_richiesti)
            } if tag_richiesti else {}
            n_conf = n_escl = n_aperti = n_err = 0
            for c in macchine:
                dec = request.POST.get(f"decisione_{c.id}", "da_validare")
                tag = (request.POST.get(f"asset_{c.id}") or "").strip()
                if dec == "conferma":
                    asset = asset_per_tag.get(tag)
                    if asset is None:
                        n_err += 1
                        continue
                    c.asset = asset
                    c.match_confermato = True
                    n_conf += 1
                elif dec == "escludi":
                    c.asset = None
                    c.match_confermato = True
                    n_escl += 1
                else:  # da_validare: aggiorna l'asset proposto ma non confermare
                    asset = asset_per_tag.get(tag) if tag else None
                    if asset is not None:
                        c.asset = asset
                    c.match_confermato = False
                    n_aperti += 1
                c.save()
            msg = f"Salvato: {n_conf} confermati, {n_escl} esclusi, {n_aperti} da validare."
            if n_err:
                messages.warning(request, msg + f" {n_err} righe non salvate (asset_tag non valido).")
            else:
                messages.success(request, msg)
            return redirect("anagrafica:skm_match_validazione")

    macchine = list(
        CompetenzaSkm.objects.filter(tipo=CompetenzaSkm.TIPO_MACCHINA)
        .select_related("asset").order_by("competenza_key")
    )

    def _stato(c):
        if c.match_confermato:
            return "confermato" if c.asset_id else "escluso"
        return "da_validare"

    ordine = {"da_validare": 0, "confermato": 1, "escluso": 2}
    righe = sorted(
        ({"c": c, "stato": _stato(c)} for c in macchine),
        key=lambda r: (ordine.get(r["stato"], 9), r["c"].competenza_key),
    )
    n_conf = sum(1 for r in righe if r["stato"] == "confermato")
    n_escl = sum(1 for r in righe if r["stato"] == "escluso")
    n_val = sum(1 for r in righe if r["stato"] == "da_validare")

    # Datalist asset: esclude i tipi palesemente non-macchina per ridurre rumore.
    from assets.models import Asset
    nonmacc = {"PC", "NOTEBOOK", "SERVER", "VM", "FIREWALL", "STAMPANTE", "FONIA", "CCTV"}
    asset_opts = list(
        Asset.objects.exclude(asset_type__in=nonmacc)
        .only("id", "asset_tag", "name").order_by("asset_tag")
    )

    return render(request, "anagrafica/pages/skm_match_validazione.html", {
        "righe": righe,
        "n_tot": len(righe),
        "n_conf": n_conf,
        "n_escl": n_escl,
        "n_val": n_val,
        "catalogo_vuoto": len(righe) == 0,
        "asset_opts": asset_opts,
    })


# ---------------------------------------------------------------------------
# Skill Matrix MOD.187 — matrice persone × macchine (F4)
# ---------------------------------------------------------------------------

@login_required
def skill_matrix_macchina(request):
    """Matrice **persone × macchine** con livelli I/L/U/O (MOD.187).

    Celle = livello (etichetta configurabile); marker ``▲`` sotto livello
    richiesto; tratteggio = rivalutazione arretrata (NON bloccante); barra blu =
    multivoce; punto = continuità monitorata. KPI in testa (prontezza squadra,
    macchine scoperte, rischio uomo-solo, continuità persa) dal resolver F3.
    Tab gemella "Processi qualificati" rimanda alla matrice qualifiche esistente.

    Sola lettura. Accesso: ``_check_hr_permission`` (ACL formale in F7).
    Finché la baseline (F2b) non è importata, la matrice è vuota: la pagina
    mostra comunque struttura e KPI, con rimando alla validazione match (F2a).
    """
    from .acl_bootstrap import PERM_SKM_VIEW
    if not _check_skm_permission(request, PERM_SKM_VIEW):
        messages.error(request, "Non hai i permessi per la Skill Matrix macchine.")
        return redirect("anagrafica:index")

    from django.utils import timezone
    from .models import (
        AbilitazioneMacchina, CompetenzaSkm, ContinuitaOperativa, SkillMatrixConfig,
    )
    from .services import skillmatrix_resolver as resolver

    config = SkillMatrixConfig.get_instance()
    filtro_reparto = (request.GET.get("reparto") or "").strip()
    export_csv = request.GET.get("format") == "csv"
    oggi = timezone.localdate()

    # Colonne = macchine MOD.187 con asset risolto (catalogo F2a).
    comp_all = list(
        CompetenzaSkm.objects
        .filter(tipo=CompetenzaSkm.TIPO_MACCHINA, asset__isnull=False)
        .select_related("asset").order_by("display", "competenza_key")
    )
    reparti = sorted({
        (c.asset.reparto or "").strip() for c in comp_all if (c.asset.reparto or "").strip()
    })
    comp_macchine = [
        c for c in comp_all
        if not filtro_reparto or (c.asset.reparto or "").strip().casefold() == filtro_reparto.casefold()
    ]
    asset_ids = [c.asset_id for c in comp_macchine]

    # Abilitazioni su quelle macchine (vuoto finché non c'è la baseline F2b).
    ab_map: dict[tuple[int, int], AbilitazioneMacchina] = {}
    legacy_ids: set[int] = set()
    if asset_ids:
        for ab in AbilitazioneMacchina.objects.filter(asset_id__in=asset_ids).prefetch_related("voci"):
            ab_map[(ab.legacy_anagrafica_id, ab.asset_id)] = ab
            legacy_ids.add(ab.legacy_anagrafica_id)

    # Nomi dall'anagrafica legacy: solo se servono, fail-safe (la sorgente può
    # non essere disponibile, es. in test) → fallback "ID <n>".
    dip_map: dict[int, dict] = {}
    if legacy_ids:
        try:
            dip_map = {int(r["id"]): r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("id")}
        except Exception:
            logger.warning("Skill Matrix: anagrafica legacy non disponibile, nomi non risolti", exc_info=True)
    soglia_ord = config.soglia_operativa_ordinale

    # Disponibilità per una data scelta (Skill Matrix × assenze): chi è in
    # ferie/malattia/permesso quel giorno. Read-only, fail-safe: se la sorgente
    # assenze non risponde la matrice resta identica (nessun marker).
    import datetime as _dt
    data_str = (request.GET.get("data") or "").strip()
    try:
        data_sel = _dt.date.fromisoformat(data_str) if data_str else oggi
    except ValueError:
        data_sel = oggi
    disp_map: dict[int, list[dict]] = {}
    if legacy_ids:
        try:
            from assenze.availability import disponibilita_per_anagrafica
            disp_map = disponibilita_per_anagrafica(sorted(legacy_ids), data_sel, data_sel)
        except Exception:
            logger.warning("Skill Matrix: disponibilità assenze non risolta", exc_info=True)

    righe: list[dict] = []
    for lid in legacy_ids:
        dip = dip_map.get(lid, {})
        celle = []
        for c in comp_macchine:
            ab = ab_map.get((lid, c.asset_id))
            if ab is None:
                celle.append({"vuota": True})
                continue
            celle.append({
                "vuota": False,
                "livello": ab.livello,
                "etichetta": config.etichetta(ab.livello),
                "operativa": ab.is_operativa(soglia_ord),
                "sotto": ab.sotto_livello_richiesto,
                "arretrata": bool(ab.prossima_revisione and ab.prossima_revisione < oggi),
                "multivoce": bool(ab.voci.all()),
                "sospesa": ab.stato == AbilitazioneMacchina.STATO_SOSPESA,
                "in_lista": ab.in_lista,
            })
        assenze_lid = disp_map.get(lid, [])
        if any(a["stato"] == "confermata" for a in assenze_lid):
            disp_stato = "assente"
        elif any(a["stato"] == "pendente" for a in assenze_lid):
            disp_stato = "da_confermare"
        else:
            disp_stato = "presente"
        righe.append({
            "legacy_id": lid,
            "cognome": str(dip.get("cognome") or f"ID {lid}").strip(),
            "nome": str(dip.get("nome") or "").strip(),
            "reparto": str(dip.get("reparto") or "").strip(),
            "celle": celle,
            "disp_stato": disp_stato,
            "disp_assenze": assenze_lid,
        })
    righe.sort(key=lambda r: (r["cognome"].casefold(), r["nome"].casefold()))

    if export_csv:
        def _disp_csv(r):
            if r["disp_stato"] == "assente":
                tipi = []
                for a in r["disp_assenze"]:
                    if a["stato"] != "confermata":
                        continue
                    t = a["tipo"] + (" (parziale)" if a.get("parziale") else "")
                    if t not in tipi:
                        tipi.append(t)
                return "assente (%s)" % ", ".join(tipi) if tipi else "assente"
            if r["disp_stato"] == "da_confermare":
                return "da confermare"
            return ""

        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="skill_matrix_macchina.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(resp, delimiter=";")
        disp_col = f"Disponibilità {data_sel.strftime('%d/%m/%Y')}"
        writer.writerow(
            ["Dipendente", "Reparto", disp_col] + [c.display or c.competenza_key for c in comp_macchine]
        )
        for r in righe:
            cells = []
            for cell in r["celle"]:
                cells.append("" if cell["vuota"] else cell["livello"])
            writer.writerow(
                [f"{r['cognome']} {r['nome']}".strip(), r["reparto"], _disp_csv(r)] + cells
            )
        return resp

    # KPI dal resolver (rispettano il filtro reparto).
    rep = filtro_reparto or None
    prontezza = resolver.prontezza_squadra(rep, config=config)
    n_scoperte = len(resolver.macchine_scoperte(rep))
    n_uomo_solo = sum(
        1 for c in comp_macchine if resolver.kpi_uomo_solo(c.asset, config=config)["a_rischio"]
    )
    n_cont_persa = sum(
        1 for co in ContinuitaOperativa.objects.all()
        if co.stato() == ContinuitaOperativa.STATO_PERSA
    )

    from .acl_bootstrap import PERM_SKM_MANAGE
    n_assenti = sum(1 for r in righe if r["disp_stato"] == "assente")
    n_da_confermare = sum(1 for r in righe if r["disp_stato"] == "da_confermare")
    return render(request, "anagrafica/pages/skill_matrix_macchina.html", {
        "macchine": comp_macchine,
        "righe": righe,
        "reparti": reparti,
        "filtro_reparto": filtro_reparto,
        "totale_persone": len(righe),
        "config": config,
        "kpi": {
            "prontezza": prontezza,
            "scoperte": n_scoperte,
            "uomo_solo": n_uomo_solo,
            "continuita_persa": n_cont_persa,
        },
        "matrice_vuota": len(righe) == 0,
        "n_macchine": len(comp_macchine),
        "data_sel": data_sel,
        "n_assenti": n_assenti,
        "n_da_confermare": n_da_confermare,
        "can_manage": _check_skm_permission(request, PERM_SKM_MANAGE),
    })


# ---------------------------------------------------------------------------
# Skill Matrix MOD.187 — refresh semestrale CAR (F6)
# ---------------------------------------------------------------------------

@login_required
def skm_refresh(request):
    """Schermata CAR di **refresh semestrale** delle abilitazioni macchina.

    ① Rivaluta le abilitazioni in lista del reparto (conferma invariato / modifica
    livello / rimuovi); ② aggiunge nuove abilitazioni. Ogni azione scrive uno
    scatto in ``AbilitazioneMacchinaStorico`` (fonte ``refresh``) e sposta in avanti
    ``prossima_revisione``. Arretrato visibile, non bloccante.

    Accesso: ``_check_hr_permission`` (lo scoping per CAR sul proprio reparto è una
    rifinitura successiva). Merito = CAR; la campagna è solo l'innesco.
    """
    from .acl_bootstrap import PERM_SKM_MANAGE
    if not _check_skm_permission(request, PERM_SKM_MANAGE):
        messages.error(request, "Non hai i permessi per il refresh Skill Matrix.")
        return redirect("anagrafica:index")

    from django.utils import timezone
    from .models import CampagnaRefresh, CompetenzaSkm, LivelloSkm
    from .services import skillmatrix_refresh as refresh

    reparto = (request.GET.get("reparto") or request.POST.get("reparto") or "").strip()

    comp = list(
        CompetenzaSkm.objects
        .filter(tipo=CompetenzaSkm.TIPO_MACCHINA, asset__isnull=False)
        .select_related("asset")
    )
    reparti = sorted({(c.asset.reparto or "").strip() for c in comp if (c.asset.reparto or "").strip()})
    macchine_reparto = [
        {"asset_id": c.asset_id, "label": f"{c.competenza_key} — {c.asset.name}"}
        for c in comp if reparto and (c.asset.reparto or "").strip() == reparto
    ]

    if request.method == "POST" and reparto:
        camp = refresh.apri_campagna(reparto, avviatore_ruolo=(request.POST.get("avviatore_ruolo") or "").strip())
        decisioni = {}
        for k, v in request.POST.items():
            if k.startswith("azione_"):
                ab_id = k[len("azione_"):]
                decisioni[ab_id] = {"azione": v, "livello": request.POST.get(f"livello_{ab_id}", "")}
        stats = refresh.applica_refresh(reparto=reparto, decisioni=decisioni, campagna=camp)
        # ② aggiunta manuale (facoltativa).
        nl = (request.POST.get("nuovo_legacy") or "").strip()
        na = (request.POST.get("nuovo_asset") or "").strip()
        nlv = (request.POST.get("nuovo_livello") or "").strip()
        agg = 0
        if nl and na and nlv:
            try:
                refresh.aggiungi_abilitazione(
                    legacy_anagrafica_id=int(nl), asset_id=int(na), livello=nlv)
                agg = 1
            except (ValueError, TypeError):
                messages.warning(request, "Aggiunta manuale non valida (legacy id / asset / livello).")
        messages.success(
            request,
            f"Refresh salvato: {stats['confermate']} confermate, {stats['modificate']} modificate, "
            f"{stats['rimosse']} rimosse" + (f", {agg} aggiunta." if agg else "."),
        )
        return redirect(f"{reverse('anagrafica:skm_refresh')}?reparto={reparto}")

    oggi = timezone.localdate()
    abil = list(refresh.abilitazioni_reparto(reparto)) if reparto else []
    dip_map: dict[int, dict] = {}
    if abil:
        try:
            dip_map = {int(r["id"]): r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("id")}
        except Exception:
            logger.warning("Skill Matrix refresh: anagrafica legacy non disponibile", exc_info=True)

    def _nome(lid):
        d = dip_map.get(lid, {})
        return f"{str(d.get('cognome') or '').strip()} {str(d.get('nome') or '').strip()}".strip() or f"ID {lid}"

    righe = [{
        "ab": a, "nome": _nome(a.legacy_anagrafica_id),
        "arretrata": bool(a.prossima_revisione and a.prossima_revisione < oggi),
    } for a in abil]
    campagna = (
        CampagnaRefresh.objects.filter(reparto=reparto, stato=CampagnaRefresh.STATO_APERTA).first()
        if reparto else None
    )

    return render(request, "anagrafica/pages/skm_refresh.html", {
        "reparto": reparto,
        "reparti": reparti,
        "righe": righe,
        "macchine_reparto": macchine_reparto,
        "campagna": campagna,
        "arretrati": sum(1 for r in righe if r["arretrata"]),
        "livelli": LivelloSkm.choices,
        "totale": len(righe),
    })


@login_required
def skm_scadenzario(request):
    """Scadenzario abilitazioni macchina per reparto (MOD.187).

    Mostra, per reparto, la prossima revisione, gli arretrati (non bloccanti) e lo
    stato campagna. HR "dà il via" al refresh (apre la campagna e avvisa il CAR: in-app
    + email); il merito della rivalutazione resta al CAR (pagina Refresh).
    Accesso: ``anagrafica.skillmatrix.manage``.
    """
    from .acl_bootstrap import PERM_SKM_MANAGE
    if not _check_skm_permission(request, PERM_SKM_MANAGE):
        messages.error(request, "Non hai i permessi per lo scadenzario Skill Matrix.")
        return redirect("anagrafica:index")

    from django.utils import timezone
    from .services import skillmatrix_refresh as refresh

    if request.method == "POST" and request.POST.get("azione") == "avvia":
        reparto = (request.POST.get("reparto") or "").strip()
        if reparto:
            legacy_user = get_legacy_user(request.user)
            _, created = refresh.avvia_refresh(
                reparto=reparto, avviatore_ruolo="HR",
                avviatore_legacy_id=int(legacy_user.id) if legacy_user else None)
            if created:
                messages.success(request, f"Refresh avviato per «{reparto}»: il CAR è stato avvisato.")
            else:
                messages.info(request, f"Il reparto «{reparto}» ha già una campagna di refresh aperta.")
        return redirect("anagrafica:skm_scadenzario")

    oggi = timezone.localdate()
    tutte = refresh.scadenzario_reparti(oggi=oggi)
    kpi = {
        "reparti_scaduti": sum(1 for r in tutte if r["stato"] == "scaduto"),
        "abil_scadute": sum(r["n_scadute"] for r in tutte),
        "campagne_aperte": sum(1 for r in tutte if r["campagna_aperta"]),
    }

    filtro_stato = (request.GET.get("stato") or "").strip()
    righe = [r for r in tutte if r["stato"] == filtro_stato] if filtro_stato in ("scaduto", "in_arrivo") else tutte

    if request.GET.get("format") == "csv":
        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="scadenzario_abilitazioni.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        w = safe_csv_writer(resp, delimiter=";")
        w.writerow(["Reparto", "Prossima revisione", "Totali", "Scadute", "In arrivo", "Stato", "Campagna aperta"])
        for r in righe:
            w.writerow([
                r["reparto"],
                r["prossima_revisione"].strftime("%d/%m/%Y") if r["prossima_revisione"] else "",
                r["n_totali"], r["n_scadute"], r["n_in_arrivo"], r["stato"],
                "Sì" if r["campagna_aperta"] else "No",
            ])
        return resp

    return render(request, "anagrafica/pages/skm_scadenzario.html", {
        "oggi": oggi, "righe": righe, "kpi": kpi, "filtro_stato": filtro_stato,
        "totale": len(righe),
    })


@login_required
def skm_impostazioni(request):
    """Gestione del singleton ``SkillMatrixConfig`` (parametri Skill Matrix MOD.187).

    Governa "operativo" (soglia livello, CAR come riserva), il rischio uomo-solo,
    le cadenze continuità/refresh e le etichette della scala. Sola configurazione,
    nessun dato personale. Accesso: ``anagrafica.skillmatrix.manage``.
    """
    from .acl_bootstrap import PERM_SKM_MANAGE
    if not _check_skm_permission(request, PERM_SKM_MANAGE):
        messages.error(request, "Non hai i permessi per gestire la configurazione Skill Matrix.")
        return redirect("anagrafica:index")

    from .forms import SkillMatrixConfigForm
    from .models import SkillMatrixConfig

    config = SkillMatrixConfig.get_instance()
    if request.method == "POST":
        form = SkillMatrixConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configurazione Skill Matrix aggiornata.")
            return redirect("anagrafica:skm_impostazioni")
        messages.error(request, "Controlla i campi: la configurazione non è stata salvata.")
    else:
        form = SkillMatrixConfigForm(instance=config)

    return render(request, "anagrafica/pages/skill_matrix_impostazioni.html", {
        "form": form,
        "config": config,
    })


@login_required
def conformita_report(request):
    """Elenco conformità di tutti i dipendenti attivi (semaforo per dominio).

    Accesso: ``_check_hr_permission`` (vista trasversale su tutto il personale).
    Filtri reparto/esito complessivo, ordinamento peggiori-prima, export CSV.
    Performance: un'unica chiamata batch (numero di query costante).
    """
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per il report conformità.")
        return redirect("anagrafica:index")

    ensure_anagrafica_schema()
    can_view_visite = _can_view_visite_mediche(request)

    filtro_reparto = (request.GET.get("reparto") or "").strip()
    filtro_esito = (request.GET.get("esito") or "").strip()
    filtro_idoneita = (request.GET.get("idoneita") or "").strip()
    filtro_mansione = (request.GET.get("mansione") or "").strip()
    export_csv = request.GET.get("format") == "csv"

    dip_rows = [r for r in fetch_anagrafica_rows(deduplicate=True) if r.get("attivo")]
    dip_map = {int(r["id"]): r for r in dip_rows if r.get("id")}
    legacy_ids = list(dip_map.keys())

    # Mappa nome mansione → id (per collegare ogni riga alla sua pagina Requisiti).
    mansioni_map = {
        m.nome.casefold(): m.id
        for m in Mansione.objects.filter(is_active=True).only("id", "nome")
    }

    mansioni_per_legacy = {
        lid: str(dip.get("mansione") or "").strip()
        for lid, dip in dip_map.items()
        if str(dip.get("mansione") or "").strip()
    }

    stati = conformita_service.stato_conformita_batch(
        legacy_ids,
        include_visite_dettaglio=can_view_visite,
        mansioni_per_legacy=mansioni_per_legacy,
    )

    _ORDINE_ESITO = {
        conformita_service.ESITO_KO: 0,
        conformita_service.ESITO_WARN: 1,
        conformita_service.ESITO_OK: 2,
        conformita_service.ESITO_NA: 3,
    }
    na = {"esito": conformita_service.ESITO_NA, "dettagli": []}

    righe: list[dict] = []
    for legacy_id, dip in dip_map.items():
        reparto = str(dip.get("reparto") or "").strip()
        if filtro_reparto and reparto.casefold() != filtro_reparto.casefold():
            continue
        mansione_nome = str(dip.get("mansione") or "").strip()
        if filtro_mansione and mansione_nome.casefold() != filtro_mansione.casefold():
            continue
        stato = stati.get(legacy_id, {"complessivo": conformita_service.ESITO_NA})
        complessivo = stato.get("complessivo", conformita_service.ESITO_NA)
        if filtro_esito and complessivo != filtro_esito:
            continue
        idoneita = stato.get("idoneita", {"esito": conformita_service.ESITO_NA})
        if filtro_idoneita and idoneita.get("esito") != filtro_idoneita:
            continue
        righe.append({
            "legacy_id": legacy_id,
            "mansione_id": mansioni_map.get(mansione_nome.casefold()),
            "cognome": str(dip.get("cognome") or f"ID {legacy_id}").strip(),
            "nome": str(dip.get("nome") or "").strip(),
            "reparto": reparto,
            "mansione": str(dip.get("mansione") or "").strip(),
            "complessivo": complessivo,
            "idoneita": idoneita,
            "formazione": stato.get("formazione", na),
            "visite": stato.get("visite", na),
            "qualifiche": stato.get("qualifiche", na),
            "dpi": stato.get("dpi", na),
        })

    righe.sort(key=lambda r: (
        _ORDINE_ESITO.get(r["complessivo"], 9),
        r["cognome"].casefold(),
        r["nome"].casefold(),
    ))

    # KPI
    n_ko = sum(1 for r in righe if r["complessivo"] == conformita_service.ESITO_KO)
    n_warn = sum(1 for r in righe if r["complessivo"] == conformita_service.ESITO_WARN)
    n_ok = sum(1 for r in righe if r["complessivo"] == conformita_service.ESITO_OK)
    n_na = sum(1 for r in righe if r["complessivo"] == conformita_service.ESITO_NA)

    reparti = sorted({r["reparto"] for r in righe if r["reparto"]})

    if export_csv:
        _LABEL = {
            conformita_service.ESITO_OK: "In regola",
            conformita_service.ESITO_WARN: "In scadenza",
            conformita_service.ESITO_KO: "Non conforme",
            conformita_service.ESITO_NA: "Nessun requisito",
        }
        _LABEL_IDN = {
            conformita_service.ESITO_OK: "Idoneo",
            conformita_service.ESITO_WARN: "Idoneo con riserve",
            conformita_service.ESITO_KO: "Non idoneo",
            conformita_service.ESITO_NA: "Non valutabile",
        }
        resp = HttpResponse(content_type=CSV_CONTENT_TYPE)
        resp["Content-Disposition"] = 'attachment; filename="conformita_anagrafica.csv"'
        resp.write(BOM)  # una volta sola: Excel riconosce l'UTF-8
        writer = safe_csv_writer(resp, delimiter=";")
        writer.writerow([
            "Dipendente", "Reparto", "Mansione", "Conformità", "Idoneità mansione",
            "Requisiti da soddisfare", "Formazione", "Visite mediche", "Qualifiche", "DPI",
        ])
        for r in righe:
            idn = r["idoneita"]
            da_soddisfare = "; ".join(list(idn.get("scaduti", [])) + list(idn.get("mancanti", [])))
            writer.writerow([
                f"{r['cognome']} {r['nome']}".strip(),
                r["reparto"],
                r["mansione"],
                _LABEL.get(r["complessivo"], r["complessivo"]),
                _LABEL_IDN.get(idn.get("esito"), idn.get("esito")),
                da_soddisfare,
                _LABEL.get(r["formazione"]["esito"], r["formazione"]["esito"]),
                _LABEL.get(r["visite"]["esito"], r["visite"]["esito"]),
                _LABEL.get(r["qualifiche"]["esito"], r["qualifiche"]["esito"]),
                _LABEL.get(r["dpi"]["esito"], r["dpi"]["esito"]),
            ])
        return resp

    paginator = Paginator(righe, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "anagrafica/pages/conformita_report.html", {
        "page_obj": page_obj,
        "totale": len(righe),
        "n_ko": n_ko,
        "n_warn": n_warn,
        "n_ok": n_ok,
        "n_na": n_na,
        "reparti": reparti,
        "filtro_reparto": filtro_reparto,
        "filtro_esito": filtro_esito,
        "filtro_idoneita": filtro_idoneita,
        "filtro_mansione": filtro_mansione,
        "can_view_visite": can_view_visite,
    })


# ---------------------------------------------------------------------------
# Onboarding strutturato — pratica + checklist (H1, speculare a offboarding)
# ---------------------------------------------------------------------------

def _onboarding_counts(pratica) -> dict[str, int]:
    tasks = list(pratica.tasks.all())
    return {
        "totale": len(tasks),
        "da_fare": sum(1 for t in tasks if t.stato == OnboardingTask.STATO_DA_FARE),
        "completati": sum(1 for t in tasks if t.stato == OnboardingTask.STATO_COMPLETATO),
        "eccezioni": sum(1 for t in tasks if t.stato == OnboardingTask.STATO_ECCEZIONE),
    }


@login_required
def onboarding_list(request):
    """Elenco pratiche onboarding con filtro stato e KPI. Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per le pratiche di onboarding.")
        return redirect("anagrafica:index")

    filtro_stato = (request.GET.get("stato") or "").strip()
    qs = OnboardingPratica.objects.prefetch_related("tasks").all()
    valid_stati = {choice[0] for choice in OnboardingPratica.STATO_CHOICES}
    if filtro_stato in valid_stati:
        qs = qs.filter(stato=filtro_stato)

    pratiche = list(qs)
    for pratica in pratiche:
        pratica.counts = _onboarding_counts(pratica)

    n_in_corso = OnboardingPratica.objects.filter(
        stato__in=OnboardingPratica.STATI_APERTI
    ).count()
    n_chiuse = OnboardingPratica.objects.filter(
        stato=OnboardingPratica.STATO_CHIUSA
    ).count()
    n_eccezioni = OnboardingPratica.objects.filter(
        stato=OnboardingPratica.STATO_CHIUSA_CON_ECCEZIONI
    ).count()

    paginator = Paginator(pratiche, 30)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "anagrafica/pages/onboarding_list.html", {
        "page_obj": page_obj,
        "totale": len(pratiche),
        "filtro_stato": filtro_stato,
        "stato_choices": OnboardingPratica.STATO_CHOICES,
        "n_in_corso": n_in_corso,
        "n_chiuse": n_chiuse,
        "n_eccezioni": n_eccezioni,
    })


@login_required
def onboarding_detail(request, pratica_id: int):
    """Dettaglio pratica onboarding con checklist task. Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per le pratiche di onboarding.")
        return redirect("anagrafica:index")

    pratica = get_object_or_404(
        OnboardingPratica.objects.prefetch_related("tasks"), pk=pratica_id
    )
    tasks = list(pratica.tasks.all())
    counts = _onboarding_counts(pratica)

    return render(request, "anagrafica/pages/onboarding_detail.html", {
        "pratica": pratica,
        "tasks": tasks,
        "counts": counts,
        "task_stato_choices": OnboardingTask.STATO_CHOICES,
        "is_admin": _offboarding_is_admin(request),
    })


@login_required
@require_POST
def onboarding_avvia(request, legacy_id: int):
    """Avvia una pratica onboarding per il dipendente. Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per avviare l'onboarding.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    rows = fetch_anagrafica_rows(ids=[legacy_id])
    if not rows:
        messages.error(request, "Dipendente non trovato.")
        return redirect("anagrafica:dipendenti_list")
    dip = rows[0]

    if onboarding_service.pratica_aperta(legacy_id):
        messages.warning(request, "Esiste gia una pratica onboarding aperta per questo dipendente.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    data_assunzione = None
    raw_date = (request.POST.get("data_assunzione") or "").strip()
    if raw_date:
        try:
            data_assunzione = date.fromisoformat(raw_date)
        except ValueError:
            messages.error(request, "Data assunzione non valida.")
            return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    ruolo_ids = list(
        DipendenteRuoloOperativo.objects
        .filter(legacy_anagrafica_id=legacy_id)
        .values_list("ruolo_id", flat=True)
    )
    nome = f"{str(dip.get('cognome') or '').strip()} {str(dip.get('nome') or '').strip()}".strip()
    try:
        pratica = onboarding_service.avvia_onboarding(
            legacy_id=legacy_id,
            dipendente_nome=nome or f"#{legacy_id}",
            reparto=str(dip.get("reparto") or "").strip(),
            mansione=str(dip.get("mansione") or "").strip(),
            data_assunzione=data_assunzione,
            note_hr=(request.POST.get("note_hr") or "").strip()[:1000],
            user=request.user,
            ruolo_ids=ruolo_ids,
        )
    except Exception:
        logger.exception("Errore avvio pratica onboarding dipendente %s", legacy_id)
        messages.error(request, "Errore durante l'avvio della pratica onboarding.")
        return redirect("anagrafica:dipendente_detail", legacy_id=legacy_id)

    _audit_safe(request, "DIPENDENTE_ONBOARDING_PRATICA_APERTA", "anagrafica", {
        "pratica_id": pratica.id,
        "legacy_anagrafica_id": legacy_id,
        "dipendente_nome": nome,
    })
    messages.success(request, "Pratica onboarding avviata. Completa la checklist di inserimento.")
    return redirect("anagrafica:onboarding_detail", pratica_id=pratica.id)


@login_required
@require_POST
def onboarding_task_update(request, pratica_id: int, task_id: int):
    """Aggiorna lo stato/nota di un task onboarding. Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per aggiornare la pratica onboarding.")
        return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)

    task = get_object_or_404(
        OnboardingTask.objects.select_related("pratica"),
        pk=task_id,
        pratica_id=pratica_id,
        pratica__stato__in=OnboardingPratica.STATI_APERTI,
    )
    stato = (request.POST.get("stato") or "").strip()
    valid_stati = {choice[0] for choice in OnboardingTask.STATO_CHOICES}
    if stato not in valid_stati:
        messages.error(request, "Stato task onboarding non valido.")
        return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)

    from django.utils import timezone

    before = task.stato
    task.stato = stato
    task.note = (request.POST.get("note") or "").strip()[:1000]
    if stato in (OnboardingTask.STATO_COMPLETATO, OnboardingTask.STATO_ECCEZIONE):
        task.completed_at = timezone.now()
        task.completed_by = request.user
    else:
        task.completed_at = None
        task.completed_by = None
    task.save(update_fields=["stato", "note", "completed_at", "completed_by", "updated_at"])

    task.pratica.updated_by = request.user
    task.pratica.save(update_fields=["updated_by", "updated_at"])
    _audit_safe(request, "DIPENDENTE_ONBOARDING_TASK_UPDATE", "anagrafica", {
        "pratica_id": pratica_id,
        "task_id": task_id,
        "task_codice": task.codice,
        "stato_precedente": before,
        "stato_nuovo": task.stato,
    })
    messages.success(request, f"Task onboarding aggiornato: {task.titolo}.")
    return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)


@login_required
@require_POST
def onboarding_chiudi(request, pratica_id: int):
    """Chiude la pratica onboarding (CHIUSA o CHIUSA_CON_ECCEZIONI). Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per chiudere la pratica onboarding.")
        return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)

    pratica = get_object_or_404(
        OnboardingPratica.objects.prefetch_related("tasks"),
        pk=pratica_id,
        stato__in=OnboardingPratica.STATI_APERTI,
    )
    stato_finale = onboarding_service.chiudi_pratica(pratica, user=request.user)
    _audit_safe(request, "DIPENDENTE_ONBOARDING_PRATICA_CHIUSA", "anagrafica", {
        "pratica_id": pratica_id,
        "legacy_anagrafica_id": pratica.legacy_anagrafica_id,
        "stato_finale": stato_finale,
    })
    if stato_finale == OnboardingPratica.STATO_CHIUSA_CON_ECCEZIONI:
        messages.warning(request, "Pratica onboarding chiusa con eccezioni (task non completati).")
    else:
        messages.success(request, "Pratica onboarding chiusa: checklist completata.")
    return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)


@login_required
@require_POST
def onboarding_annulla(request, pratica_id: int):
    """Annulla una pratica onboarding aperta. Gated HR."""
    if not _check_hr_permission(request):
        messages.error(request, "Non hai i permessi per annullare la pratica onboarding.")
        return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)

    pratica = get_object_or_404(
        OnboardingPratica,
        pk=pratica_id,
        stato__in=OnboardingPratica.STATI_APERTI,
    )
    onboarding_service.annulla_pratica(pratica, user=request.user)
    _audit_safe(request, "DIPENDENTE_ONBOARDING_PRATICA_ANNULLATA", "anagrafica", {
        "pratica_id": pratica_id,
        "legacy_anagrafica_id": pratica.legacy_anagrafica_id,
    })
    messages.success(request, "Pratica onboarding annullata.")
    return redirect("anagrafica:onboarding_detail", pratica_id=pratica_id)


# ============================================================================
# FORMAZIONE HR — E-LEARNING: MICRO-CORSI INTERNI (slide + quiz)
# ============================================================================
# Layer self-service sopra TrainingCourse (is_elearning=True). Distinzione ruoli
# via ACL formazione esistente:
#   - DISCENTE  -> @login_required (fruisce i corsi pubblicati, traccia solo i propri dati)
#   - AUTORE    -> _can_edit_formazione (crea/modifica slide e quiz)
#   - HR/ADMIN  -> _can_view_formazione (report completamenti gia esistenti)
# Identita discente = legacy_anagrafica_id (convenzione del modulo).


def _current_legacy_anagrafica_id(request) -> int | None:
    """Risolve il legacy_anagrafica_id del dipendente collegato all'utente loggato.

    Catena: utente Django -> Profile.legacy_user_id (UtenteLegacy) -> AnagraficaDipendente
    (utente_id). Ritorna None se l'utente non e collegato a un'anagrafica."""
    legacy_user = get_legacy_user(request.user)
    if not legacy_user:
        return None
    try:
        ana = AnagraficaDipendente.objects.filter(utente_id=legacy_user.id).only("id").first()
    except Exception:
        logger.debug("Impossibile risolvere anagrafica per utente legacy %s", getattr(legacy_user, "id", None))
        return None
    return int(ana.id) if ana else None


def _crea_record_completamento_elearning(corso, legacy_id, attempt, created_by):
    """Crea un TrainingEmployeeRecord storicizzato per il superamento di un micro-corso
    e-learning, riusando la tabella audit esistente (niente duplicazione).

    Allinea la qualifica e invalida/ricalcola la cache scadenze, come il flusso d'aula
    (`_crea_employee_record`)."""
    from django.utils import timezone as _tz
    oggi = _tz.localdate()
    data_scadenza = _add_months(oggi, corso.validita_mesi) if corso.validita_mesi else None

    record = TrainingEmployeeRecord.objects.create(
        corso=corso,
        sessione=None,
        enrollment=None,
        legacy_anagrafica_id=legacy_id,
        data_completamento=oggi,
        ore_frequentate=corso.durata_ore_teorica or 0,
        percentuale_presenza=None,
        idoneo=True,
        data_scadenza=data_scadenza,
        # Snapshot storici (immutabili)
        course_code_snapshot=corso.codice,
        course_title_snapshot=corso.titolo,
        course_version_snapshot=corso.versione,
        plan_code_snapshot=corso.piano.codice if corso.piano_id else "",
        plan_name_snapshot=corso.piano.nome if corso.piano_id else "",
        duration_hours_snapshot=corso.durata_ore_teorica,
        validity_months_snapshot=corso.validita_mesi,
        completion_rule_snapshot_json={"modalita": "ELEARNING", "quiz_minimo_pct": corso.quiz_punteggio_minimo},
        session_code_snapshot="",
        teacher_name_snapshot="",
        completion_calculation_snapshot_json={
            "modalita": "ELEARNING",
            "quiz_punteggio_pct": str(attempt.punteggio_pct),
            "quiz_corrette": attempt.n_corrette,
            "quiz_totali": attempt.n_totali,
            "quiz_minimo_pct": corso.quiz_punteggio_minimo,
        },
    )

    # Invalida cache scadenze e tenta ricalcolo immediato (fail-safe come il flusso d'aula)
    TrainingDeadline.objects.filter(corso=corso, legacy_anagrafica_id=legacy_id).update(needs_refresh=True)
    try:
        from .services.training_deadline_service import refresh_deadlines
        refresh_deadlines(legacy_id=legacy_id, corso_id=corso.pk)
    except NotImplementedError:
        pass
    except Exception:
        logger.exception("Refresh scadenze e-learning fallito per dip=%s corso=%s", legacy_id, corso.pk)

    # Allineamento qualifica (competency management), fail-safe
    try:
        if corso.qualifica_id and record.idoneo:
            qual, _ = _upsert_dipendente_qualifica(
                legacy_id, corso.qualifica, record.data_completamento, record.data_scadenza, user=created_by,
            )
            if qual.record_formazione_id != record.pk:
                qual.record_formazione = record
                qual.save(update_fields=["record_formazione"])
    except Exception:
        logger.exception("Allineamento qualifica e-learning fallito per record %s", record.pk)

    # Archiviazione automatica dell'attestato nella cartella documenti del dipendente
    # (stesso flusso dei corsi d'aula: cartella «Attestati formazione» o quella scelta
    # in Impostazioni → Template attestato). Fail-safe: non deve bloccare il completamento.
    try:
        cfg = AttestatoFormazioneConfig.get_instance()
        if cfg.auto_salva_attestato:
            from .services.attestato_pdf import archivia_attestato
            archivia_attestato(record, cfg=cfg, user=created_by)
    except Exception:
        logger.exception("Archiviazione automatica attestato e-learning fallita per record %s", record.pk)

    return record


# -- GESTIONE E-LEARNING: hub autori/HR --------------------------------------

def _elearning_salute(n_slide: int, n_domande: int, n_invalid: int) -> tuple[str, str]:
    """Classifica lo stato di salute di un micro-corso e-learning (codice, etichetta)."""
    if n_slide == 0:
        return ("CRIT", "Senza slide")
    if n_domande == 0:
        return ("INFO", "Senza quiz")
    if n_invalid:
        return ("WARN", "Quiz incompleto")
    return ("OK", "Pronto")


def _elearning_iscritti_rows(corso):
    """Righe «iscritti & esiti» di un micro-corso: nome dipendente + avanzamento + esito."""
    enrollments = list(TrainingElearningEnrollment.objects.filter(corso=corso))
    nomi = _build_nomi_map() if enrollments else {}
    rows = []
    for e in enrollments:
        rows.append({
            "legacy_id": e.legacy_anagrafica_id,
            "nome": nomi.get(e.legacy_anagrafica_id, f"#{e.legacy_anagrafica_id}"),
            "stato": e.stato,
            "stato_disp": e.get_stato_display(),
            "avanzamento": f"{e.ultima_slide_ordine}/{e.n_slide_totali}" if e.n_slide_totali else "—",
            "best": e.best_punteggio_pct,
            "n_tentativi": e.n_tentativi,
            "data_completamento": e.data_completamento,
        })
    rows.sort(key=lambda x: x["nome"].lower())
    return rows


@login_required
def formazione_elearning_hub(request):
    """Hub di gestione dei micro-corsi e-learning (autori/HR): elenco corsi con stato di
    salute (slide/quiz), iscritti e completati, e azioni rapide."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per la sezione formazione.")
        return redirect("anagrafica:index")
    from django.db.models import Count, Q
    corsi = list(
        TrainingCourse.objects.filter(is_elearning=True)
        .select_related("piano")
        .prefetch_related("quiz_domande__opzioni")
        .annotate(
            n_slide=Count("slides", filter=Q(slides__is_active=True), distinct=True),
            n_domande=Count("quiz_domande", filter=Q(quiz_domande__is_active=True), distinct=True),
            n_iscritti=Count("iscrizioni_elearning", distinct=True),
            n_completati=Count("iscrizioni_elearning", filter=Q(iscrizioni_elearning__stato="COMPLETATO"), distinct=True),
        )
        .order_by("-is_active", "stato", "titolo")
    )
    tot = {"corsi": len(corsi), "pubblicati": 0, "iscritti": 0, "completati": 0, "problemi": 0}
    rows = []
    for c in corsi:
        n_invalid = sum(
            1 for d in c.quiz_domande.all()
            if d.is_active and not any(o.corretta for o in d.opzioni.all())
        )
        salute = _elearning_salute(c.n_slide, c.n_domande, n_invalid)
        if c.stato == "ATTIVO" and c.is_active:
            tot["pubblicati"] += 1
        tot["iscritti"] += c.n_iscritti
        tot["completati"] += c.n_completati
        if salute[0] in ("CRIT", "WARN"):
            tot["problemi"] += 1
        rows.append({"corso": c, "n_invalid": n_invalid, "salute": salute})
    return render(request, "anagrafica/pages/formazione_elearning_hub.html", {
        "rows": rows,
        "tot": tot,
        "is_editor": _can_edit_formazione(request),
    })


@login_required
def formazione_elearning_manage(request, corso_id: int):
    """Cabina di regia di un singolo micro-corso e-learning: contenuti, iscritti & esiti,
    stato di pubblicazione ed export — tutto in un'unica pagina."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per la sezione formazione.")
        return redirect("anagrafica:index")
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    slides = list(corso.slides.filter(is_active=True))
    domande = list(corso.quiz_domande.filter(is_active=True).prefetch_related("opzioni"))
    n_invalid = sum(1 for d in domande if not any(o.corretta for o in d.opzioni.all()))
    salute = _elearning_salute(len(slides), len(domande), n_invalid)

    iscritti = _elearning_iscritti_rows(corso)
    counts = {"iscritti": len(iscritti), "in_corso": 0, "completati": 0, "non_superato": 0}
    for r in iscritti:
        if r["stato"] == "COMPLETATO":
            counts["completati"] += 1
        elif r["stato"] == "NON_SUPERATO":
            counts["non_superato"] += 1
        elif r["stato"] == "IN_CORSO":
            counts["in_corso"] += 1

    is_editor = _can_edit_formazione(request)

    # Assegnazioni (obbligo) del corso + pool dipendenti assegnabili per il picker.
    assegnazioni = list(TrainingAssignment.objects.filter(corso=corso))
    nomi = _build_nomi_map() if (assegnazioni or is_editor) else {}
    for a in assegnazioni:
        a.nome = nomi.get(a.legacy_anagrafica_id, f"#{a.legacy_anagrafica_id}")
    assegnazioni.sort(key=lambda a: (a.nome or "").lower())
    assegnabili = []
    if is_editor:
        assegnati_ids = {a.legacy_anagrafica_id for a in assegnazioni}
        attivi_ids = set(
            DipendenteAnagraficaAziendale.objects
            .filter(data_cessazione__isnull=True)
            .values_list("legacy_anagrafica_id", flat=True)
        )
        for lid in sorted(attivi_ids):
            if lid not in assegnati_ids and lid in nomi:
                assegnabili.append({"legacy_id": lid, "nome": nomi[lid]})
        assegnabili.sort(key=lambda x: x["nome"].lower())

    return render(request, "anagrafica/pages/formazione_elearning_manage.html", {
        "corso": corso,
        "is_editor": is_editor,
        "n_slide": len(slides),
        "n_domande": len(domande),
        "n_invalid": n_invalid,
        "salute": salute,
        "iscritti": iscritti,
        "counts": counts,
        "pubblicato": corso.stato == "ATTIVO" and corso.is_active,
        "assegnazioni": assegnazioni,
        "assegnabili": assegnabili,
    })


@login_required
@require_POST
def formazione_elearning_publish_toggle(request, corso_id: int):
    """Pubblica/ritira un micro-corso. La pubblicazione è bloccata se mancano le slide o
    il quiz ha domande senza risposta corretta (controllo qualità centralizzato)."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_elearning_hub")
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    if corso.stato == "ATTIVO":
        corso.stato = "BOZZA"
        corso.save(update_fields=["stato", "updated_at"])
        messages.info(request, "Corso ritirato (bozza): non più visibile ai discenti.")
    else:
        n_slide = corso.slides.filter(is_active=True).count()
        domande = list(corso.quiz_domande.filter(is_active=True).prefetch_related("opzioni"))
        n_invalid = sum(1 for d in domande if not any(o.corretta for o in d.opzioni.all()))
        if n_slide == 0:
            messages.error(request, "Impossibile pubblicare: aggiungi almeno una slide.")
        elif domande and n_invalid:
            messages.error(request, "Impossibile pubblicare: il quiz ha domande senza risposta corretta. Completa le domande o disattivale.")
        else:
            corso.stato = "ATTIVO"
            corso.is_active = True
            corso.save(update_fields=["stato", "is_active", "updated_at"])
            messages.success(request, "Corso pubblicato: ora visibile in «Corsi online».")
    # Ritorna alla pagina di provenienza interna, altrimenti alla cabina di regia
    nxt = request.POST.get("next") or ""
    if nxt.startswith("/") and not nxt.startswith("//"):
        return redirect(nxt)
    return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)


@login_required
def formazione_elearning_iscritti_csv(request, corso_id: int):
    """Export CSV di iscritti ed esiti del micro-corso (audit). Tracciato in TrainingExportLog."""
    if not _can_view_formazione(request):
        messages.error(request, "Non hai i permessi per esportare i report.")
        return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    rows = _elearning_iscritti_rows(corso)

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="elearning_{corso.codice}_iscritti.csv"'
    resp.write("﻿")  # BOM per apertura corretta in Excel
    writer = safe_csv_writer(resp, delimiter=";")
    writer.writerow(["Dipendente", "ID", "Stato", "Avanzamento slide", "Miglior punteggio %", "Tentativi", "Data completamento"])
    for r in rows:
        writer.writerow([
            r["nome"], r["legacy_id"], r["stato_disp"], r["avanzamento"],
            r["best"] if r["best"] is not None else "", r["n_tentativi"],
            r["data_completamento"] or "",
        ])
    try:
        TrainingExportLog.objects.create(
            tipo="ISCRITTI",
            filtri_json={"modalita": "ELEARNING", "corso": corso.codice},
            righe_esportate=len(rows),
            generato_da=request.user,
            ip_address=request.META.get("REMOTE_ADDR"),
        )
    except Exception:
        logger.warning("Audit export e-learning iscritti fallito", exc_info=True)
    return resp


@login_required
@require_POST
def formazione_elearning_assign(request, corso_id: int):
    """Assegna (obbligo) il micro-corso a uno o più dipendenti: crea TrainingAssignment
    e richiama l'hook notifica (predisposto, invio non attivo). Idempotente."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per assegnare corsi.")
        return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse.objects.select_related("piano"), pk=corso_id, is_elearning=True)

    ids, seen = [], set()
    for raw in request.POST.getlist("dipendenti_selezionati"):
        s = str(raw).strip()
        if s.isdigit():
            lid = int(s)
            if lid > 0 and lid not in seen:
                seen.add(lid)
                ids.append(lid)
    if not ids:
        messages.warning(request, "Nessun dipendente selezionato.")
        return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)

    due = None
    raw_due = (request.POST.get("due_date") or "").strip()
    if raw_due:
        from datetime import datetime as _dt
        try:
            due = _dt.strptime(raw_due, "%Y-%m-%d").date()
        except ValueError:
            due = None

    from .services.elearning_notifications import notify_corso_assegnato
    n_new = 0
    for lid in ids:
        obj, created = TrainingAssignment.objects.get_or_create(
            corso=corso, legacy_anagrafica_id=lid,
            defaults={"stato": "ASSEGNATO", "piano": corso.piano, "due_date": due, "assigned_by": request.user},
        )
        if created:
            n_new += 1
            try:
                notify_corso_assegnato(corso.pk, lid)
            except Exception:
                logger.debug("Hook notifica assegnazione e-learning fallito", exc_info=True)
    if n_new:
        messages.success(request, f"{n_new} dipendente/i assegnato/i al corso.")
    else:
        messages.info(request, "I dipendenti selezionati erano già assegnati.")
    return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)


@login_required
@require_POST
def formazione_elearning_unassign(request, corso_id: int, assignment_id: int):
    """Rimuove un'assegnazione non ancora completata."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)
    a = get_object_or_404(TrainingAssignment, pk=assignment_id, corso_id=corso_id)
    if a.stato == "COMPLETATO":
        messages.error(request, "Non puoi rimuovere un'assegnazione già completata.")
    else:
        a.delete()
        messages.success(request, "Assegnazione rimossa.")
    return redirect("anagrafica:formazione_elearning_manage", corso_id=corso_id)


@login_required
def formazione_elearning_settings(request):
    """Impostazioni e-learning (singleton): default dei nuovi micro-corsi e percorso
    LibreOffice per l'import PowerPoint. Gated dal permesso di modifica formazione."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare le impostazioni e-learning.")
        return redirect("anagrafica:formazione_elearning_hub")
    cfg = ElearningConfig.get_instance()
    if request.method == "POST":
        form = ElearningConfigForm(request.POST, instance=cfg)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Impostazioni e-learning salvate.")
            return redirect("anagrafica:formazione_elearning_settings")
        messages.error(request, "Controlla i campi: " + form.errors.as_text())
    else:
        form = ElearningConfigForm(instance=cfg)
    # Diagnostica LibreOffice (per l'import PowerPoint)
    from .services.elearning_import import find_libreoffice
    return render(request, "anagrafica/pages/formazione_elearning_settings.html", {
        "form": form,
        "cfg": cfg,
        "libreoffice_trovato": find_libreoffice(),
    })


# -- AUTORE: gestione contenuti (slide + quiz) -------------------------------

@login_required
def formazione_corso_elearning(request, corso_id: int):
    """Pagina autore: gestione slide e quiz di un micro-corso e-learning."""
    if not _can_edit_formazione(request):
        messages.error(request, "Non hai i permessi per modificare i contenuti e-learning.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    slides = list(corso.slides.all())
    domande = list(corso.quiz_domande.prefetch_related("opzioni").all())
    # Segnala le domande attive senza alcuna opzione corretta: sarebbero impossibili da
    # superare e vengono escluse dal quiz del discente finché non si completa l'autoring.
    n_domande_incomplete = 0
    for d in domande:
        d.senza_corretta = d.is_active and not any(o.corretta for o in d.opzioni.all())
        if d.senza_corretta:
            n_domande_incomplete += 1
    return render(request, "anagrafica/pages/formazione_corso_elearning.html", {
        "corso": corso,
        "slides": slides,
        "domande": domande,
        "n_domande_incomplete": n_domande_incomplete,
        "slide_form": TrainingSlideForm(initial={"ordine": (slides[-1].ordine + 1) if slides else 1}),
        "question_form": TrainingQuizQuestionForm(initial={"ordine": (domande[-1].ordine + 1) if domande else 1}),
    })


@login_required
@require_POST
def formazione_slide_save(request, corso_id: int):
    """Crea o aggiorna una slide. slide_id vuoto = creazione."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    slide_id = request.POST.get("slide_id")
    instance = get_object_or_404(TrainingSlide, pk=slide_id, corso=corso) if slide_id else None
    form = TrainingSlideForm(request.POST, instance=instance)
    if form.is_valid():
        slide = form.save(commit=False)
        slide.corso = corso
        if not slide.pk:
            slide.created_by = request.user
        slide.save()
        messages.success(request, f'Slide "{slide.titolo}" salvata.')
    else:
        messages.error(request, "Errore nel salvataggio della slide: " + form.errors.as_text())
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
@require_POST
def formazione_slide_delete(request, corso_id: int, slide_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    slide = get_object_or_404(TrainingSlide, pk=slide_id, corso_id=corso_id)
    slide.delete()
    messages.success(request, "Slide eliminata.")
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
@require_POST
def formazione_slide_import(request, corso_id: int):
    """Importa slide da un file PowerPoint/PDF: una slide-immagine per pagina."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Nessun file selezionato.")
        return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)
    # Limite dimensione (50 MB) per evitare upload abnormi
    if f.size and f.size > 50 * 1024 * 1024:
        messages.error(request, "File troppo grande (massimo 50 MB).")
        return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)
    from .services.elearning_import import importa_slides_da_file, ImportError_
    try:
        n = importa_slides_da_file(corso, f, user=request.user)
        messages.success(request, f"Importate {n} slide da «{f.name}».")
    except ImportError_ as e:
        messages.error(request, str(e))
    except Exception:
        logger.exception("Import slide e-learning fallito per corso %s", corso_id)
        messages.error(request, "Import non riuscito: errore imprevisto nella conversione.")
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
def formazione_slide_image(request, slide_id: int):
    """Serve inline l'immagine di una slide dallo storage privato.

    Accesso: editor formazione, oppure qualsiasi utente autenticato se il corso è un
    e-learning pubblicato (così il discente vede le slide-immagine nel player)."""
    slide = get_object_or_404(TrainingSlide.objects.select_related("corso"), pk=slide_id)
    corso = slide.corso
    pubblicato = corso.is_elearning and corso.is_active and corso.stato == "ATTIVO"
    if not (_can_edit_formazione(request) or pubblicato):
        return HttpResponse(status=403)
    if not slide.immagine:
        return HttpResponse("Immagine non disponibile.", status=404)
    from django.http import FileResponse
    try:
        fh = slide.immagine.open("rb")
    except FileNotFoundError:
        return HttpResponse("Immagine non trovata sul server.", status=404)
    resp = FileResponse(fh, content_type="image/png")
    resp["Content-Disposition"] = f'inline; filename="slide_{slide.pk}.png"'
    resp["Cache-Control"] = "private, max-age=300"
    return resp


@login_required
@require_POST
def formazione_question_save(request, corso_id: int):
    """Crea o aggiorna una domanda del quiz."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    corso = get_object_or_404(TrainingCourse, pk=corso_id)
    q_id = request.POST.get("question_id")
    instance = get_object_or_404(TrainingQuizQuestion, pk=q_id, corso=corso) if q_id else None
    form = TrainingQuizQuestionForm(request.POST, instance=instance)
    if form.is_valid():
        q = form.save(commit=False)
        q.corso = corso
        q.save()
        messages.success(request, "Domanda salvata.")
    else:
        messages.error(request, "Errore nella domanda: " + form.errors.as_text())
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
@require_POST
def formazione_question_delete(request, corso_id: int, question_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    q = get_object_or_404(TrainingQuizQuestion, pk=question_id, corso_id=corso_id)
    q.delete()
    messages.success(request, "Domanda eliminata.")
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
@require_POST
def formazione_option_save(request, corso_id: int, question_id: int):
    """Aggiunge/aggiorna un'opzione di risposta a una domanda."""
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    domanda = get_object_or_404(TrainingQuizQuestion, pk=question_id, corso_id=corso_id)
    opt_id = request.POST.get("option_id")
    instance = get_object_or_404(TrainingQuizOption, pk=opt_id, domanda=domanda) if opt_id else None
    form = TrainingQuizOptionForm(request.POST, instance=instance)
    if form.is_valid():
        opt = form.save(commit=False)
        opt.domanda = domanda
        opt.save()
        messages.success(request, "Opzione salvata.")
    else:
        messages.error(request, "Errore nell'opzione: " + form.errors.as_text())
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


@login_required
@require_POST
def formazione_option_delete(request, corso_id: int, option_id: int):
    if not _can_edit_formazione(request):
        messages.error(request, "Permesso negato.")
        return redirect("anagrafica:formazione_corso_detail", corso_id=corso_id)
    opt = get_object_or_404(TrainingQuizOption, pk=option_id, domanda__corso_id=corso_id)
    opt.delete()
    messages.success(request, "Opzione eliminata.")
    return redirect("anagrafica:formazione_corso_elearning", corso_id=corso_id)


# -- DISCENTE: catalogo, player slide (HTMX), quiz ---------------------------

@login_required
def formazione_online_catalog(request):
    """Catalogo dei micro-corsi e-learning pubblicati, con il mio stato per ciascuno."""
    legacy_id = _current_legacy_anagrafica_id(request)
    corsi = list(
        TrainingCourse.objects.filter(is_elearning=True, is_active=True, stato="ATTIVO")
        .select_related("piano", "categoria")
        .order_by("titolo")
    )
    iscrizioni = {}
    assegnazioni = {}
    if legacy_id:
        iscrizioni = {
            e.corso_id: e
            for e in TrainingElearningEnrollment.objects.filter(
                legacy_anagrafica_id=legacy_id, corso__in=corsi
            )
        }
        assegnazioni = {
            a.corso_id: a
            for a in TrainingAssignment.objects.filter(
                legacy_anagrafica_id=legacy_id, corso__in=corsi
            )
        }
    cards = []
    for c in corsi:
        e = iscrizioni.get(c.pk)
        a = assegnazioni.get(c.pk)
        # «Da fare» = assegnato (obbligo) e non ancora completato
        assegnato_da_fare = bool(a) and a.stato not in ("COMPLETATO", "ESONERATO") and (e is None or e.stato != "COMPLETATO")
        cards.append({
            "corso": c,
            "stato": e.stato if e else None,
            "best_pct": e.best_punteggio_pct if e else None,
            "n_slide": c.slides.filter(is_active=True).count(),
            "n_domande": c.quiz_domande.filter(is_active=True).count(),
            "assegnato": bool(a),
            "assegnato_da_fare": assegnato_da_fare,
            "due_date": a.due_date if a else None,
        })
    # Ordina: prima i corsi assegnati da completare, poi gli altri (per titolo)
    cards.sort(key=lambda x: (not x["assegnato_da_fare"], x["corso"].titolo.lower()))
    return render(request, "anagrafica/pages/formazione_online_catalog.html", {
        "cards": cards,
        "no_anagrafica": legacy_id is None,
        "is_editor": _can_edit_formazione(request),
    })


def _enrollment_corrente(corso, legacy_id):
    """Ritorna (creandola se serve) l'iscrizione e-learning del discente al corso."""
    n_slide = corso.slides.filter(is_active=True).count()
    enr, _created = TrainingElearningEnrollment.objects.get_or_create(
        corso=corso, legacy_anagrafica_id=legacy_id,
        defaults={"stato": "ISCRITTO", "n_slide_totali": n_slide},
    )
    if enr.n_slide_totali != n_slide:
        enr.n_slide_totali = n_slide
        enr.save(update_fields=["n_slide_totali"])
    return enr


@login_required
def formazione_online_player(request, corso_id: int):
    """Apre il player del micro-corso: iscrive il discente (se non gia) e mostra la
    prima slide (o quella piu avanti gia vista)."""
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    if not (corso.is_active and corso.stato == "ATTIVO") and not _can_edit_formazione(request):
        messages.error(request, "Corso non disponibile.")
        return redirect("anagrafica:formazione_online_catalog")
    legacy_id = _current_legacy_anagrafica_id(request)
    slides = list(corso.slides.filter(is_active=True))
    enr = None
    slide_iniziale = slides[0].ordine if slides else 1
    if legacy_id and slides:
        enr = _enrollment_corrente(corso, legacy_id)
        ordini = [s.ordine for s in slides]
        if enr.ultima_slide_ordine in ordini:
            slide_iniziale = enr.ultima_slide_ordine
    return render(request, "anagrafica/pages/formazione_online_player.html", {
        "corso": corso,
        "slides": slides,
        "n_slide": len(slides),
        "slide_iniziale": slide_iniziale,
        "enrollment": enr,
        "no_anagrafica": legacy_id is None,
        "n_domande": corso.quiz_domande.filter(is_active=True).count(),
    })


@login_required
def formazione_online_slide(request, corso_id: int, ordine: int):
    """Partial HTMX: rende la slide <ordine> del corso e aggiorna l'avanzamento.

    Ritorna il partial quando chiamata via HTMX, altrimenti reindirizza al player."""
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    slides = list(corso.slides.filter(is_active=True))
    if not slides:
        if request.headers.get("HX-Request"):
            return HttpResponse('<div class="fmd-empty"><span class="fmd-et">Nessuna slide disponibile</span></div>')
        return redirect("anagrafica:formazione_online_catalog")

    ordini = [s.ordine for s in slides]
    pos = ordini.index(ordine) if ordine in ordini else 0
    slide = slides[pos]

    from .services.elearning_markdown import render_markdown
    contenuto_html = render_markdown(slide.contenuto)

    # Avanzamento (solo se il discente e tracciabile)
    legacy_id = _current_legacy_anagrafica_id(request)
    if legacy_id:
        enr = _enrollment_corrente(corso, legacy_id)
        campi = []
        if slide.ordine > enr.ultima_slide_ordine:
            enr.ultima_slide_ordine = slide.ordine
            campi.append("ultima_slide_ordine")
        if enr.stato == "ISCRITTO":
            enr.stato = "IN_CORSO"
            campi.append("stato")
        if campi:
            enr.save(update_fields=campi + ["updated_at"])

    ctx = {
        "corso": corso,
        "slide": slide,
        "contenuto_html": contenuto_html,
        "indice": pos + 1,
        "n_slide": len(slides),
        "ordine_prec": ordini[pos - 1] if pos > 0 else None,
        "ordine_succ": ordini[pos + 1] if pos < len(slides) - 1 else None,
        "is_ultima": pos == len(slides) - 1,
        "n_domande": corso.quiz_domande.filter(is_active=True).count(),
        "progress_pct": round((pos + 1) / len(slides) * 100),
    }
    if request.headers.get("HX-Request"):
        return render(request, "anagrafica/partials/_formazione_online_slide.html", ctx)
    return redirect("anagrafica:formazione_online_player", corso_id=corso_id)


@login_required
def formazione_online_quiz(request, corso_id: int):
    """Quiz finale del micro-corso: GET mostra le domande, POST corregge, scrive il
    tentativo (audit) e — al superamento — il record di completamento storicizzato."""
    corso = get_object_or_404(TrainingCourse, pk=corso_id, is_elearning=True)
    domande = list(corso.quiz_domande.filter(is_active=True).prefetch_related("opzioni"))
    legacy_id = _current_legacy_anagrafica_id(request)

    if not domande:
        messages.info(request, "Questo corso non ha ancora un quiz finale.")
        return redirect("anagrafica:formazione_online_player", corso_id=corso_id)

    # Solo domande "valide" (almeno un'opzione corretta): una domanda senza risposta
    # corretta sarebbe impossibile da indovinare e bloccherebbe il superamento, quindi
    # viene esclusa dal quiz finché l'autore non la completa (segnalata in pagina autore).
    domande = [d for d in domande if any(o.corretta for o in d.opzioni.all())]
    if not domande:
        messages.info(request, "Il quiz non è ancora pronto: nessuna domanda ha una risposta corretta configurata.")
        return redirect("anagrafica:formazione_online_player", corso_id=corso_id)

    if request.method != "POST":
        return render(request, "anagrafica/pages/formazione_online_quiz.html", {
            "corso": corso,
            "domande": domande,
            "no_anagrafica": legacy_id is None,
            "esito": None,
        })

    if legacy_id is None:
        messages.error(request, "Il tuo profilo non e collegato all'anagrafica: il completamento non puo essere registrato. Contatta HR.")
        return redirect("anagrafica:formazione_online_player", corso_id=corso_id)

    # Limite tentativi (Impostazioni e-learning): blocca un nuovo invio se esaurito.
    cfg_el = ElearningConfig.get_instance()
    if cfg_el.max_tentativi_quiz:
        _enr = TrainingElearningEnrollment.objects.filter(corso=corso, legacy_anagrafica_id=legacy_id).first()
        if _enr and _enr.stato != "COMPLETATO" and (_enr.n_tentativi or 0) >= cfg_el.max_tentativi_quiz:
            messages.error(request, f"Hai esaurito i tentativi disponibili per questo quiz ({cfg_el.max_tentativi_quiz}).")
            return redirect("anagrafica:formazione_online_player", corso_id=corso_id)

    # -- Correzione -----------------------------------------------------------
    n_totali = len(domande)
    n_corrette = 0
    risposte_snapshot = []
    for d in domande:
        scelte = set(int(x) for x in request.POST.getlist(f"q_{d.pk}") if str(x).isdigit())
        corrette = set(o.pk for o in d.opzioni.all() if o.corretta)
        giusta = bool(corrette) and scelte == corrette
        if giusta:
            n_corrette += 1
        risposte_snapshot.append({
            "domanda_id": d.pk,
            "domanda": d.testo,
            "scelte": sorted(scelte),
            "corrette": sorted(corrette),
            "giusta": giusta,
        })

    from decimal import Decimal
    punteggio = Decimal(str(round(n_corrette / n_totali * 100, 2))) if n_totali else Decimal("0")
    superato = punteggio >= corso.quiz_punteggio_minimo

    with transaction.atomic():
        enr = _enrollment_corrente(corso, legacy_id)
        attempt = TrainingQuizAttempt.objects.create(
            corso=corso,
            enrollment=enr,
            legacy_anagrafica_id=legacy_id,
            punteggio_pct=punteggio,
            n_corrette=n_corrette,
            n_totali=n_totali,
            superato=superato,
            risposte_json={"risposte": risposte_snapshot},
            utente=request.user,
        )
        enr.n_tentativi = (enr.n_tentativi or 0) + 1
        if enr.best_punteggio_pct is None or punteggio > enr.best_punteggio_pct:
            enr.best_punteggio_pct = punteggio
        campi = ["n_tentativi", "best_punteggio_pct", "updated_at"]
        if superato and enr.stato != "COMPLETATO":
            from django.utils import timezone as _tz
            enr.stato = "COMPLETATO"
            enr.data_completamento = _tz.localdate()
            campi += ["stato", "data_completamento"]
            if not enr.record_completamento_id:
                record = _crea_record_completamento_elearning(corso, legacy_id, attempt, request.user)
                enr.record_completamento = record
                attempt.record = record
                attempt.save(update_fields=["record"])
                campi.append("record_completamento")
            # Chiude eventuali assegnazioni (obbligo) aperte per questo corso/dipendente
            TrainingAssignment.objects.filter(
                corso=corso, legacy_anagrafica_id=legacy_id,
            ).exclude(stato="COMPLETATO").update(stato="COMPLETATO")
        elif not superato and enr.stato != "COMPLETATO":
            enr.stato = "NON_SUPERATO"
            campi.append("stato")
        enr.save(update_fields=list(dict.fromkeys(campi)))

    return render(request, "anagrafica/pages/formazione_online_quiz.html", {
        "corso": corso,
        "domande": domande,
        "no_anagrafica": False,
        "esito": {
            "superato": superato,
            "punteggio": punteggio,
            "n_corrette": n_corrette,
            "n_totali": n_totali,
            "minimo": corso.quiz_punteggio_minimo,
            "risposte": risposte_snapshot,
        },
    })
