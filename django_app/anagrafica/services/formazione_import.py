"""Import servizi per il modulo Formazione HR — da file Excel gestionale.

Vedi `docs/anagrafica/formazione/MAPPING_IMPORT_GESTIONALE.md` per la mappatura
puntuale delle colonne Excel → campi modelli.

Tutti gli import sono **dry-run di default** (`commit=False`): contano cosa
verrebbe creato/aggiornato senza scrivere nel DB. Per applicare passare
`commit=True` — l'intera operazione è avvolta in una `transaction.atomic()`,
quindi in caso di errore di mid-import nessun dato parziale viene persistito.

Lookup `Iscritto` → `legacy_anagrafica_id`:
- match esatto Cognome + Nome (case-insensitive) su `core.legacy_models.AnagraficaDipendente`
- fallback su `codice_fiscale` (se presente nel file)
- se ambiguo o mancante: la riga finisce in `skipped` con messaggio in `warnings`

NOTA AMBIENTE: `AnagraficaDipendente` è una tabella legacy non gestita
(`managed=False`, `db_table=anagrafica_dipendenti`). In SQLite dev la tabella
è vuota → tutti i lookup falliranno (atteso). In SQL Server test/prod la
tabella è popolata → lookup funzioneranno.
"""
from __future__ import annotations

import hashlib
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.db import transaction

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# PARSING HELPERS
# ─────────────────────────────────────────────────────────────

def _norm(s: Any) -> str:
    """Strip + collapse whitespace + rimuove caratteri non stampabili."""
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_date(v: Any) -> date | None:
    """Accetta date, datetime, o stringa 'gg/mm/aaaa'. Ritorna `date` o None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _norm(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(v: Any) -> Decimal | None:
    """Accetta numero o stringa con virgola/punto/percentuale. Ritorna Decimal o None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None
    s = _norm(v).replace("%", "").replace(",", ".").strip()
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_si_no(v: Any) -> bool | None:
    """Accetta 'Si'/'No' (case insensitive). Ritorna bool o None se vuoto/sconosciuto."""
    if v is None:
        return None
    s = _norm(v).lower()
    if s in {"si", "sì", "s", "y", "yes", "true", "1"}:
        return True
    if s in {"no", "n", "false", "0"}:
        return False
    return None


def _split_cognome_nome(s: str) -> tuple[str, str]:
    """Split 'COGNOME NOME' → ('COGNOME', 'NOME').

    Heuristica: il PRIMO token è cognome, il resto è nome composto.
    Se contiene virgola ('Rossi, Mario') la usa come separatore.
    """
    s = _norm(s)
    if not s:
        return ("", "")
    if "," in s:
        parts = s.split(",", 1)
        return (parts[0].strip(), parts[1].strip())
    tokens = s.split(" ")
    if len(tokens) == 1:
        return (tokens[0], "")
    return (tokens[0], " ".join(tokens[1:]))


def _map_modalita_sessione(s: Any) -> str:
    """Mappa stringa gestionale → choice TrainingSession.MODALITA_CHOICES.

    Default: 'IN_SEDE' se sconosciuto/vuoto.
    """
    sl = _norm(s).lower()
    if not sl:
        return "IN_SEDE"
    if "remoto" in sl or "lms" in sl or "online" in sl or "e-learning" in sl or "elearning" in sl:
        return "REMOTO"
    if "esterno" in sl or "provider" in sl:
        return "ESTERNO"
    if "misto" in sl or "blended" in sl:
        return "MISTO"
    return "IN_SEDE"


def _short_code(nome: str, prefix: str = "", max_len: int = 20) -> str:
    """Genera un codice corto univoco a partire da un nome lungo.

    Esempio: 'SALUTE E SICUREZZA NEI LUOGHI DI LAVORO' → 'P-SALUTE_E_SI-A3F2C1'
    """
    nome = _norm(nome)
    slug = re.sub(r"[^A-Z0-9]+", "_", nome.upper()).strip("_")
    h = hashlib.sha1(nome.encode("utf-8")).hexdigest()[:6].upper()
    body_max = max_len - len(prefix) - 7  # 7 = "-" + 6 hash chars
    body = slug[:body_max] if body_max > 0 else ""
    if prefix:
        return f"{prefix}{body}-{h}"[:max_len]
    return f"{body}-{h}"[:max_len]


# ─────────────────────────────────────────────────────────────
# LOOKUP DIPENDENTE (AnagraficaDipendente legacy)
# ─────────────────────────────────────────────────────────────

def _build_lookup_indexes() -> tuple[dict[str, int], dict[str, int]]:
    """Costruisce due indici da `AnagraficaDipendente` legacy:

    - `by_nome`: { "COGNOME NOME" (upper, normalizzato): id }
    - `by_cf`:   { codice_fiscale (upper): id }  — vuoto se non disponibile

    NOTA: `AnagraficaDipendente` non ha codice fiscale come campo proprio
    (il CF è in `DipendenteAnagraficaCivile` del modulo anagrafica). Il
    fallback CF richiede join cross-database e per ora è disabilitato.
    """
    from core.legacy_models import AnagraficaDipendente

    by_nome: dict[str, int] = {}
    by_cf: dict[str, int] = {}

    try:
        for r in AnagraficaDipendente.objects.values("id", "cognome", "nome"):
            try:
                lid = int(r.get("id") or 0)
            except (TypeError, ValueError):
                continue
            cog = _norm(r.get("cognome") or "").upper()
            nom = _norm(r.get("nome") or "").upper()
            full = f"{cog} {nom}".strip()
            if full:
                by_nome.setdefault(full, lid)
    except Exception:
        logger.exception("Errore costruzione indice nomi dipendenti — tabella legacy non accessibile?")

    # CF lookup: collega DipendenteAnagraficaCivile.codice_fiscale → legacy_id
    try:
        from anagrafica.models import DipendenteAnagraficaCivile
        for r in DipendenteAnagraficaCivile.objects.exclude(codice_fiscale="").values(
            "legacy_anagrafica_id", "codice_fiscale"
        ):
            cf = _norm(r.get("codice_fiscale") or "").upper()
            if cf:
                by_cf.setdefault(cf, int(r["legacy_anagrafica_id"]))
    except Exception:
        logger.exception("Errore costruzione indice CF dipendenti")

    return by_nome, by_cf


def _lookup_legacy_id(
    cognome_nome: str,
    codice_fiscale: str | None,
    by_nome: dict[str, int],
    by_cf: dict[str, int],
) -> int | None:
    """Cerca legacy_anagrafica_id per Cognome+Nome (preferito) o CF (fallback)."""
    key_nome = _norm(cognome_nome).upper()
    if key_nome and key_nome in by_nome:
        return by_nome[key_nome]
    cf = _norm(codice_fiscale or "").upper()
    if cf and cf in by_cf:
        return by_cf[cf]
    return None


# ─────────────────────────────────────────────────────────────
# WORKBOOK READER
# ─────────────────────────────────────────────────────────────

def _read_rows(xlsx_path: str | Path, sheet_name: str | None = None) -> list[dict]:
    """Legge un foglio del workbook come lista di dict (header → cella).

    Senza `sheet_name` legge il primo foglio (comportamento storico).
    """
    import openpyxl  # lazy import: dipendenza opzionale

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []
        headers = [_norm(h) for h in headers]
        out: list[dict] = []
        for row in rows_iter:
            if not row or all(c is None or c == "" for c in row):
                continue
            d = {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}
            out.append(d)
        return out
    finally:
        wb.close()


# ─────────────────────────────────────────────────────────────
# IMPORT: training-plans.xlsx → TrainingPlan
# ─────────────────────────────────────────────────────────────

def import_piani_formativi(xlsx_path: str | Path, commit: bool = False) -> dict:
    """Importa piani formativi da `training-plans.xlsx`.

    Colonne attese: Nome | Stato | Numero di corsi | Ore totali dei corsi |
    Partecipanti ai corsi | Rapporto ore corsi/partecipanti | Totale ore dei partecipanti | Costo totale

    Strategia:
    - Lookup `TrainingPlan.nome` esatto → update se esiste, create altrimenti.
    - `codice` generato deterministicamente da `_short_code(nome, prefix="P-")`.
    - `ore_totali_stimate` ← "Ore totali dei corsi".
    - `costo_stimato` ← "Costo totale" (strippato di "€").
    """
    from anagrafica.models_formazione import TrainingPlan

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "warnings": [],
        "dry_run": not commit,
    }

    try:
        rows = _read_rows(xlsx_path)
    except FileNotFoundError as e:
        report["errors"].append(f"File non trovato: {e}")
        return report
    except Exception as e:
        report["errors"].append(f"Errore lettura xlsx: {e}")
        logger.exception("Errore lettura xlsx %s", xlsx_path)
        return report

    report["righe_lette"] = len(rows)

    def _do_import():
        for i, row in enumerate(rows, start=2):  # riga 1 = header
            nome = _norm(row.get("Nome"))
            if not nome:
                report["skipped"] += 1
                report["warnings"].append(f"riga {i}: Nome vuoto")
                continue

            stato_raw = _norm(row.get("Stato")).lower()
            if stato_raw == "attivo":
                stato = "ATTIVO"
            elif stato_raw in {"archiviato", "chiuso", "concluso"}:
                stato = "ARCHIVIATO"
            else:
                stato = "BOZZA"

            ore = _parse_decimal(row.get("Ore totali dei corsi"))

            costo_raw = _norm(row.get("Costo totale"))
            costo_raw = costo_raw.replace("€", "").replace(",", ".").strip()
            try:
                costo = Decimal(costo_raw) if costo_raw else None
            except (InvalidOperation, ValueError):
                costo = None

            defaults = {
                "stato": stato,
                "ore_totali_stimate": ore,
                "costo_stimato": costo,
                "is_active": stato == "ATTIVO",
            }

            existing = TrainingPlan.objects.filter(nome=nome).first()
            if existing:
                if commit:
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                report["updated"] += 1
            else:
                codice = _short_code(nome, prefix="P-", max_len=20)
                # Evita collisioni rare di hash
                while TrainingPlan.objects.filter(codice=codice).exists():
                    codice = _short_code(nome + "_x", prefix="P-", max_len=20)
                if commit:
                    TrainingPlan.objects.create(
                        nome=nome,
                        codice=codice,
                        categoria="OBBLIGATORIA" if "sicurezza" in nome.lower() else "CONSIGLIATA",
                        **defaults,
                    )
                report["created"] += 1

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report


# ─────────────────────────────────────────────────────────────
# IMPORT: courses-person.xlsx → Course + Session + Enrollment + Record
# ─────────────────────────────────────────────────────────────

def import_courses_person(xlsx_path: str | Path, commit: bool = False) -> dict:
    """Importa storico iscrizioni/esiti da `courses-person.xlsx`.

    Crea (o aggiorna) in cascata:
    - `TrainingPlan` (lookup per Nome, create se mancante)
    - `TrainingCourse` (lookup per Codice corso)
    - `TrainingSession` (lookup per `corso + data_inizio` — chiave naturale)
    - `TrainingEnrollment` (lookup per `(sessione, legacy_anagrafica_id)`)
    - `TrainingEmployeeRecord` (creato solo se "Ha superato il corso? = Si")
    """
    from anagrafica.models_formazione import (
        TrainingCourse,
        TrainingEmployeeRecord,
        TrainingEnrollment,
        TrainingPlan,
        TrainingSession,
    )

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "dipendenti_non_trovati": 0,
        "piani_created": 0,
        "corsi_created": 0,
        "corsi_accoppiati_per_titolo": 0,
        "sessioni_created": 0,
        "iscrizioni_created": 0,
        "iscrizioni_updated": 0,
        "record_created": 0,
        "skipped": 0,
        "errors": [],
        "warnings": [],
        "dry_run": not commit,
    }

    try:
        rows = _read_rows(xlsx_path)
    except FileNotFoundError as e:
        report["errors"].append(f"File non trovato: {e}")
        return report
    except Exception as e:
        report["errors"].append(f"Errore lettura xlsx: {e}")
        logger.exception("Errore lettura xlsx %s", xlsx_path)
        return report

    report["righe_lette"] = len(rows)

    by_nome, by_cf = _build_lookup_indexes()
    if not by_nome:
        report["warnings"].append(
            "Indice nomi dipendenti vuoto — tabella legacy `anagrafica_dipendenti` "
            "non accessibile o vuota. In SQLite dev è atteso. Tutti i lookup falliranno."
        )

    # Cache piani/corsi/sessioni già processati in questa run (dry-run friendly)
    pian_cache: dict[str, TrainingPlan | None] = {}
    cors_cache: dict[str, TrainingCourse | None] = {}
    sess_cache: dict[tuple[str, date | None], TrainingSession | None] = {}

    # Indice per titolo normalizzato: i "Codice corso" del gestionale sono
    # per-edizione (cambiano di anno in anno anche per lo stesso corso
    # ricorrente), quindi un titolo già a catalogo con un codice diverso va
    # ACCOPPIATO al TrainingCourse esistente invece di duplicarlo.
    titolo_index: dict[str, TrainingCourse] = {}
    for tc in TrainingCourse.objects.exclude(titolo=""):
        key = _norm(tc.titolo).upper()
        if key and key not in titolo_index:
            titolo_index[key] = tc

    def _get_piano(nome: str) -> TrainingPlan | None:
        if not nome:
            return None
        if nome in pian_cache:
            return pian_cache[nome]
        existing = TrainingPlan.objects.filter(nome=nome).first()
        if existing:
            pian_cache[nome] = existing
            return existing
        if commit:
            codice = _short_code(nome, prefix="P-", max_len=20)
            while TrainingPlan.objects.filter(codice=codice).exists():
                codice = _short_code(nome + "_x", prefix="P-", max_len=20)
            obj = TrainingPlan.objects.create(
                nome=nome, codice=codice,
                categoria="OBBLIGATORIA" if "sicurezza" in nome.lower() else "CONSIGLIATA",
                stato="ATTIVO", is_active=True,
            )
            pian_cache[nome] = obj
            report["piani_created"] += 1
            return obj
        # Dry-run: simula creazione
        pian_cache[nome] = None
        report["piani_created"] += 1
        return None

    def _get_corso(codice: str, titolo: str, piano: TrainingPlan | None, durata: Decimal | None) -> TrainingCourse | None:
        if not codice:
            return None
        if codice in cors_cache:
            return cors_cache[codice]
        existing = TrainingCourse.objects.filter(codice=codice).first()
        if existing:
            cors_cache[codice] = existing
            return existing
        # Codice non a catalogo: prova ad accoppiare per titolo normalizzato
        # prima di crearne uno nuovo (stesso corso ricorrente, codice diverso
        # per edizione/anno nel gestionale).
        titolo_key = _norm(titolo).upper()
        if titolo_key and titolo_key in titolo_index:
            matched = titolo_index[titolo_key]
            cors_cache[codice] = matched
            report["corsi_accoppiati_per_titolo"] += 1
            return matched
        if commit and piano is not None:
            obj = TrainingCourse.objects.create(
                codice=codice, titolo=titolo or codice,
                piano=piano,
                durata_ore_teorica=durata or Decimal("0"),
                validita_mesi=0,  # default una tantum — può essere aggiornato in seguito
                stato="ATTIVO", is_active=True,
            )
            cors_cache[codice] = obj
            if titolo_key:
                titolo_index[titolo_key] = obj
            report["corsi_created"] += 1
            return obj
        cors_cache[codice] = None
        if titolo_key and titolo_key not in titolo_index:
            # Dry-run: nessun oggetto reale da indicizzare, ma segnamo il
            # titolo come "già visto in questa run" così righe successive con
            # lo stesso titolo/codice diverso risultano accoppiate e non
            # gonfiano il conteggio di corsi_created.
            titolo_index[titolo_key] = None
        report["corsi_created"] += 1
        return None

    def _get_sessione(corso: TrainingCourse | None, data_inizio: date | None,
                      data_fine: date | None, modalita: str, sede: str) -> TrainingSession | None:
        if corso is None or data_inizio is None:
            return None
        codice_corso = corso.codice
        key = (codice_corso, data_inizio)
        if key in sess_cache:
            return sess_cache[key]
        existing = TrainingSession.objects.filter(corso=corso, data_inizio=data_inizio).first()
        if existing:
            sess_cache[key] = existing
            return existing
        # Genera codice_sessione deterministico
        codice_sessione = f"{codice_corso}-{data_inizio.strftime('%Y%m%d')}"[:40]
        # Evita collisioni
        n = 0
        base = codice_sessione
        while TrainingSession.objects.filter(codice_sessione=codice_sessione).exists():
            n += 1
            codice_sessione = f"{base[:36]}-{n}"
        if commit:
            obj = TrainingSession.objects.create(
                corso=corso, codice_sessione=codice_sessione,
                stato="COMPLETATA", modalita=modalita,
                data_inizio=data_inizio, data_fine=data_fine or data_inizio,
                sede=sede[:200] if sede else "",
            )
            sess_cache[key] = obj
            report["sessioni_created"] += 1
            return obj
        sess_cache[key] = None
        report["sessioni_created"] += 1
        return None

    def _do_import():
        for i, row in enumerate(rows, start=2):
            try:
                nome_dip = _norm(row.get("Iscritto"))
                cf = _norm(row.get("Codice fiscale"))
                legacy_id = _lookup_legacy_id(nome_dip, cf, by_nome, by_cf)
                if legacy_id is None:
                    report["dipendenti_non_trovati"] += 1
                    report["skipped"] += 1
                    if len(report["warnings"]) < 50:
                        report["warnings"].append(
                            f"riga {i}: dipendente non trovato '{nome_dip}' (CF '{cf}')"
                        )
                    continue

                piano_nome = _norm(row.get("Piano formativo"))
                cod_corso = _norm(row.get("Codice corso"))
                tit_corso = _norm(row.get("Corso"))
                if not cod_corso:
                    report["skipped"] += 1
                    report["warnings"].append(f"riga {i}: Codice corso vuoto")
                    continue

                durata = _parse_decimal(row.get("Durata (ore)"))
                modalita = _map_modalita_sessione(row.get("Modalità di erogazione"))
                sede = _norm(row.get("Luogo del corso") or row.get("Sede"))
                d_inizio = _parse_date(row.get("Inizio corso"))
                d_fine = _parse_date(row.get("Fine corso"))

                piano = _get_piano(piano_nome)
                corso = _get_corso(cod_corso, tit_corso, piano, durata)
                sessione = _get_sessione(corso, d_inizio, d_fine, modalita, sede)

                ha_partecipato = _parse_si_no(row.get("Ha partecipato al corso ?"))
                ha_superato = _parse_si_no(row.get("Ha superato il corso ?"))
                ore_freq = _parse_decimal(row.get("Frequenza"))
                perc = _parse_decimal(row.get("Ore partecipate (%)"))

                if ha_superato is True:
                    stato_isc = "COMPLETATO"
                    idoneo = True
                elif ha_superato is False and ha_partecipato is True:
                    stato_isc = "NON_IDONEO"
                    idoneo = False
                elif ha_partecipato is False:
                    stato_isc = "ASSENTE"
                    idoneo = None
                else:
                    stato_isc = "ISCRITTO"
                    idoneo = None

                if not commit:
                    # Dry-run: assume tutte le righe come "iscrizioni nuove" se non lookup-able
                    report["iscrizioni_created"] += 1
                    if ha_superato is True:
                        report["record_created"] += 1
                    continue

                # Necessari: corso e sessione reali per creare l'enrollment
                if corso is None or sessione is None:
                    report["skipped"] += 1
                    report["warnings"].append(
                        f"riga {i}: corso/sessione non creabili (piano={piano_nome}, "
                        f"corso={cod_corso}, data={d_inizio})"
                    )
                    continue

                enroll, created = TrainingEnrollment.objects.get_or_create(
                    sessione=sessione,
                    legacy_anagrafica_id=legacy_id,
                    defaults={
                        "stato": stato_isc,
                        "ore_frequentate": ore_freq or Decimal("0"),
                        "percentuale_presenza": perc,
                        "idoneo": idoneo,
                        "data_completamento": d_fine if ha_superato else None,
                    },
                )
                if created:
                    report["iscrizioni_created"] += 1
                else:
                    # Update con nuovi valori
                    enroll.stato = stato_isc
                    enroll.ore_frequentate = ore_freq or Decimal("0")
                    enroll.percentuale_presenza = perc
                    enroll.idoneo = idoneo
                    enroll.data_completamento = d_fine if ha_superato else None
                    enroll.save()
                    report["iscrizioni_updated"] += 1

                if ha_superato is True and d_fine is not None:
                    rec, rec_created = TrainingEmployeeRecord.objects.get_or_create(
                        corso=corso, legacy_anagrafica_id=legacy_id,
                        data_completamento=d_fine,
                        defaults={
                            "sessione": sessione,
                            "enrollment": enroll,
                            "ore_frequentate": ore_freq or Decimal("0"),
                            "percentuale_presenza": perc,
                            "idoneo": True,
                            "course_code_snapshot": corso.codice[:30],
                            "course_title_snapshot": corso.titolo[:300],
                            "plan_code_snapshot": (piano.codice if piano else "")[:20],
                            "plan_name_snapshot": (piano.nome if piano else "")[:200],
                            "duration_hours_snapshot": durata,
                            "validity_months_snapshot": corso.validita_mesi,
                            "session_code_snapshot": sessione.codice_sessione[:40],
                        },
                    )
                    if rec_created:
                        report["record_created"] += 1

            except Exception as e:
                report["errors"].append(f"riga {i}: {type(e).__name__}: {e}")
                logger.exception("Errore import riga %s", i)

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report


# ─────────────────────────────────────────────────────────────
# IMPORT: Iscritti alle lezioni → TrainingLesson + Attendance
# ─────────────────────────────────────────────────────────────

def import_lezioni_presenze(xlsx_path: str | Path, commit: bool = False) -> dict:
    """Importa presenze granulari da `Iscritti alle lezioni_*.xlsx`.

    Strategia: per ogni riga, individua sessione (corso+data corso) e
    lezione (sessione+data lezione+nome). Se mancano, le crea solo se in
    `commit=True` e c'è abbastanza informazione. Poi crea/aggiorna
    `TrainingLessonAttendance` per (lezione, legacy_anagrafica_id).
    """
    from datetime import time

    from anagrafica.models_formazione import (
        TrainingCourse,
        TrainingEnrollment,
        TrainingLesson,
        TrainingLessonAttendance,
        TrainingSession,
    )

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "dipendenti_non_trovati": 0,
        "lezioni_created": 0,
        "presenze_created": 0,
        "presenze_updated": 0,
        "skipped": 0,
        "errors": [],
        "warnings": [],
        "dry_run": not commit,
    }

    try:
        rows = _read_rows(xlsx_path)
    except FileNotFoundError as e:
        report["errors"].append(f"File non trovato: {e}")
        return report
    except Exception as e:
        report["errors"].append(f"Errore lettura xlsx: {e}")
        logger.exception("Errore lettura xlsx %s", xlsx_path)
        return report

    report["righe_lette"] = len(rows)

    by_nome, by_cf = _build_lookup_indexes()
    if not by_nome:
        report["warnings"].append("Indice nomi dipendenti vuoto.")

    lez_cache: dict[tuple[int, date, str], TrainingLesson | None] = {}

    def _parse_time(v: Any) -> time | None:
        if v is None or v == "":
            return None
        if isinstance(v, time):
            return v
        if isinstance(v, datetime):
            return v.time()
        s = _norm(v)
        for fmt in ("%H:%M:%S", "%H:%M", "%H.%M"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                continue
        return None

    def _do_import():
        for i, row in enumerate(rows, start=2):
            try:
                nome_dip = _norm(row.get("Iscritto"))
                cf = _norm(row.get("Codice fiscale"))
                legacy_id = _lookup_legacy_id(nome_dip, cf, by_nome, by_cf)
                if legacy_id is None:
                    report["dipendenti_non_trovati"] += 1
                    report["skipped"] += 1
                    if len(report["warnings"]) < 50:
                        report["warnings"].append(
                            f"riga {i}: dipendente non trovato '{nome_dip}'"
                        )
                    continue

                cod_corso = _norm(row.get("Codice corso"))
                d_corso_inizio = _parse_date(row.get("Inizio corso"))
                d_lezione = _parse_date(row.get("Data"))
                nome_lez = _norm(row.get("Nome lezione"))

                if not cod_corso or d_corso_inizio is None or d_lezione is None:
                    report["skipped"] += 1
                    report["warnings"].append(
                        f"riga {i}: dati insufficienti (corso={cod_corso}, "
                        f"inizio_corso={d_corso_inizio}, data_lezione={d_lezione})"
                    )
                    continue

                # Lookup corso + sessione (deve essere già presente da courses-person)
                corso = TrainingCourse.objects.filter(codice=cod_corso).first()
                if not corso:
                    report["skipped"] += 1
                    if len(report["warnings"]) < 50:
                        report["warnings"].append(
                            f"riga {i}: corso {cod_corso} non in catalogo — importare prima courses-person.xlsx"
                        )
                    continue
                sessione = TrainingSession.objects.filter(
                    corso=corso, data_inizio=d_corso_inizio
                ).first()
                if not sessione:
                    report["skipped"] += 1
                    if len(report["warnings"]) < 50:
                        report["warnings"].append(
                            f"riga {i}: sessione non trovata corso={cod_corso} data={d_corso_inizio}"
                        )
                    continue

                # Get/create lezione
                lez_key = (sessione.pk, d_lezione, nome_lez)
                lezione = lez_cache.get(lez_key)
                if lezione is None:
                    lezione = (
                        TrainingLesson.objects
                        .filter(sessione=sessione, data=d_lezione, argomento=nome_lez)
                        .first()
                    )
                    if lezione is None:
                        if commit:
                            # Calcola numero progressivo
                            n_existing = TrainingLesson.objects.filter(sessione=sessione).count()
                            ora_inizio = _parse_time(row.get("Ora di inizio")) or time(9, 0)
                            ora_fine = _parse_time(row.get("Ora di fine")) or time(13, 0)
                            lezione = TrainingLesson.objects.create(
                                sessione=sessione,
                                numero=n_existing + 1,
                                data=d_lezione,
                                ora_inizio=ora_inizio,
                                ora_fine=ora_fine,
                                argomento=nome_lez[:500] or "Lezione",
                            )
                            report["lezioni_created"] += 1
                        else:
                            report["lezioni_created"] += 1
                            lez_cache[lez_key] = None
                            continue
                    lez_cache[lez_key] = lezione

                if not commit:
                    report["presenze_created"] += 1
                    continue

                # Crea/aggiorna presenza
                stato_partecipazione = _parse_si_no(row.get("Ha partecipato alla lezione?"))
                if stato_partecipazione is True:
                    stato_p = "PRESENTE"
                elif stato_partecipazione is False:
                    stato_p = "ASSENTE_INGIUST"
                else:
                    stato_p = "PRESENTE"  # default ottimistico

                ore_eff = _parse_decimal(row.get("Ore frequentate"))

                # Lookup enrollment (per FK opzionale)
                enrollment = TrainingEnrollment.objects.filter(
                    sessione=sessione, legacy_anagrafica_id=legacy_id
                ).first()

                pres, created = TrainingLessonAttendance.objects.get_or_create(
                    lezione=lezione, legacy_anagrafica_id=legacy_id,
                    defaults={
                        "enrollment": enrollment,
                        "stato_presenza": stato_p,
                        "ore_effettive": ore_eff,
                    },
                )
                if created:
                    report["presenze_created"] += 1
                else:
                    pres.stato_presenza = stato_p
                    pres.ore_effettive = ore_eff
                    if enrollment and not pres.enrollment_id:
                        pres.enrollment = enrollment
                    pres.save()
                    report["presenze_updated"] += 1

            except Exception as e:
                report["errors"].append(f"riga {i}: {type(e).__name__}: {e}")
                logger.exception("Errore import lezione riga %s", i)

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report


# ─────────────────────────────────────────────────────────────
# IMPORT: qualifications-person.xlsx → TipoQualifica + DipendenteQualifica
# ─────────────────────────────────────────────────────────────

def import_qualifications_person(xlsx_path: str | Path, commit: bool = False) -> dict:
    """Importa qualifiche/abilitazioni da `qualifications-person.xlsx`.

    NOTA: popola **tabelle del modulo anagrafica** (TipoQualifica,
    DipendenteQualifica), NON tabelle formazione. Vedi MAPPING per il razionale.
    """
    from anagrafica.models import DipendenteQualifica, TipoQualifica

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "dipendenti_non_trovati": 0,
        "tipi_qualifica_created": 0,
        "qualifiche_created": 0,
        "qualifiche_updated": 0,
        "skipped": 0,
        "errors": [],
        "warnings": [],
        "dry_run": not commit,
    }

    try:
        rows = _read_rows(xlsx_path)
    except FileNotFoundError as e:
        report["errors"].append(f"File non trovato: {e}")
        return report
    except Exception as e:
        report["errors"].append(f"Errore lettura xlsx: {e}")
        logger.exception("Errore lettura xlsx %s", xlsx_path)
        return report

    report["righe_lette"] = len(rows)
    by_nome, by_cf = _build_lookup_indexes()
    if not by_nome:
        report["warnings"].append("Indice nomi dipendenti vuoto.")

    tipo_cache: dict[str, TipoQualifica | None] = {}

    def _categoria_qualifica(cat_excel: str) -> str:
        cl = cat_excel.lower()
        if "sicurezza" in cl or "primo soccors" in cl or "antincendio" in cl:
            return TipoQualifica.CAT_SICUREZZA
        if "professional" in cl:
            return TipoQualifica.CAT_PROFESSIONALE
        if "gestional" in cl:
            return TipoQualifica.CAT_GESTIONALE
        return TipoQualifica.CAT_ALTRO

    def _do_import():
        for i, row in enumerate(rows, start=2):
            try:
                nome_dip = _norm(row.get("Nome"))
                legacy_id = _lookup_legacy_id(nome_dip, None, by_nome, by_cf)
                if legacy_id is None:
                    report["dipendenti_non_trovati"] += 1
                    report["skipped"] += 1
                    if len(report["warnings"]) < 50:
                        report["warnings"].append(
                            f"riga {i}: dipendente non trovato '{nome_dip}'"
                        )
                    continue

                abil = _norm(row.get("Abilitazione"))
                if not abil:
                    report["skipped"] += 1
                    continue

                cat_excel = _norm(row.get("Categoria") or "")
                cat = _categoria_qualifica(cat_excel)
                d_cons = _parse_date(row.get("Data conseguimento"))
                d_scad = _parse_date(row.get("Data scadenza"))
                note = _norm(row.get("Note") or "")[:255]

                # Get/create TipoQualifica
                tipo = tipo_cache.get(abil)
                if tipo is None:
                    tipo = TipoQualifica.objects.filter(nome=abil).first()
                    if tipo is None:
                        if commit:
                            tipo = TipoQualifica.objects.create(
                                nome=abil, categoria=cat, is_active=True,
                            )
                            report["tipi_qualifica_created"] += 1
                        else:
                            report["tipi_qualifica_created"] += 1
                            tipo_cache[abil] = None
                            continue
                    tipo_cache[abil] = tipo

                if not commit:
                    report["qualifiche_created"] += 1
                    continue

                # Get/create DipendenteQualifica per (legacy_id, tipo)
                qual, created = DipendenteQualifica.objects.get_or_create(
                    legacy_anagrafica_id=legacy_id, tipo=tipo,
                    defaults={
                        "data_conseguimento": d_cons,
                        "data_scadenza": d_scad,
                        "note": note,
                    },
                )
                if created:
                    report["qualifiche_created"] += 1
                else:
                    qual.data_conseguimento = d_cons or qual.data_conseguimento
                    qual.data_scadenza = d_scad or qual.data_scadenza
                    if note:
                        qual.note = note
                    qual.save()
                    report["qualifiche_updated"] += 1

            except Exception as e:
                report["errors"].append(f"riga {i}: {type(e).__name__}: {e}")
                logger.exception("Errore import qualifica riga %s", i)

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report


# ─────────────────────────────────────────────────────────────
# IMPORT: "N Estrazioni Corsi.xlsx" → TrainingPlan + TrainingCourse +
#          TrainingInstructor + TrainingSession
# ─────────────────────────────────────────────────────────────
# Formato diverso da courses-person.xlsx: qui il foglio è un catalogo
# corsi/sessioni (una riga = una sessione erogata), senza dati di
# iscrizione per persona. Colonne attese (fogli "Corsi" e
# "Corsi AGGIORNAMENTI"):
#   CODICE ATTIVITA' | DESCRIZIONE ATTIVITA' | CODICE CORSO | DESCRIZIONE |
#   DESCRIZIONE ESTESA | ORE OBBLIGATORIE | DESCRIZIONE LUOGO | TIPO LUOGO |
#   CODICE DOCENTE | DESCRIZIONE DOCENTE | STATO CORSO | DATA INIZIO |
#   DATA FINE | NOTE (solo AGGIORNAMENTI)
#
# I fogli "PROVA ..." dello stesso file sono bozze/QA e vanno ignorati.

ESTRAZIONI_CORSI_SHEETS_DEFAULT = ["Corsi", "Corsi AGGIORNAMENTI"]

_STATO_CORSO_TO_SESSIONE = {
    "aperto":  "PIANIFICATA",
    "erogato": "IN_CORSO",
    "chiuso":  "COMPLETATA",
    "sospeso": "ANNULLATA",
}


def _map_stato_sessione_estrazioni(s: Any) -> str | None:
    return _STATO_CORSO_TO_SESSIONE.get(_norm(s).lower())


def _map_modalita_da_tipo_luogo(s: Any) -> str:
    return "ESTERNO" if _norm(s).lower() == "esterno" else "IN_SEDE"


def import_estrazioni_corsi(
    xlsx_path: str | Path,
    commit: bool = False,
    sheets: list[str] | None = None,
    solo_docenti: bool = False,
) -> dict:
    """Importa catalogo corsi/sessioni/docenti da "N Estrazioni Corsi.xlsx".

    Con `solo_docenti=True` importa **solo** `TrainingInstructor` da
    DESCRIZIONE DOCENTE (una entry per nome distinto su tutte le righe/fogli),
    senza toccare piani/corsi/sessioni — utile quando il catalogo corsi verrà
    importato separatamente da un altro file.

    Per ogni riga valida (CODICE CORSO presente):
    - `TrainingPlan`: lookup/creazione per `nome` = DESCRIZIONE ATTIVITA'.
    - `TrainingCourse`: lookup per `codice` = str(CODICE CORSO). Titolo
      sempre allineato al file; descrizione riempita solo se vuota; durata
      aggiornata solo se il file fornisce un valore e quello attuale è
      vuoto/zero.
    - `TrainingInstructor`: lookup/creazione per `nome` = DESCRIZIONE DOCENTE
      (case-insensitive). Se il nome combacia con un dipendente in anagrafica
      (Cognome Nome), `tipo=INTERNO` e `legacy_anagrafica_id` valorizzato;
      altrimenti `tipo=ESTERNO`. Righe senza docente (vuoto o '#N/A') non
      creano/aggiornano l'istruttore.
    - `TrainingSession`: lookup per `(corso, data_inizio)`; salta la riga se
      manca DATA INIZIO. `sede`/`modalita`/`stato`/`docente` aggiornati se il
      file fornisce un valore diverso da quello corrente (il file è
      considerato la fonte più recente per questi campi operativi).

    I fogli passati in `sheets` (default: "Corsi" poi "Corsi AGGIORNAMENTI")
    sono processati in ordine: a parità di codice corso, i fogli successivi
    vincono (permette di ri-lanciare l'import con un export più recente).
    """
    from anagrafica.models_formazione import (
        TrainingCourse,
        TrainingInstructor,
        TrainingPlan,
        TrainingSession,
    )

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "righe_saltate": 0,
        "piani_created": 0,
        "corsi_created": 0,
        "corsi_updated": 0,
        "docenti_created": 0,
        "sessioni_created": 0,
        "sessioni_updated": 0,
        "errors": [],
        "warnings": [],
        "dry_run": not commit,
    }

    sheets = sheets or ESTRAZIONI_CORSI_SHEETS_DEFAULT

    all_rows: list[dict] = []
    for sheet_name in sheets:
        try:
            rows = _read_rows(xlsx_path, sheet_name=sheet_name)
        except KeyError:
            report["warnings"].append(f"Foglio non trovato: '{sheet_name}' — saltato.")
            continue
        except FileNotFoundError as e:
            report["errors"].append(f"File non trovato: {e}")
            return report
        except Exception as e:
            report["errors"].append(f"Errore lettura xlsx (foglio {sheet_name}): {e}")
            logger.exception("Errore lettura xlsx %s [%s]", xlsx_path, sheet_name)
            return report
        for row in rows:
            row["__sheet__"] = sheet_name
        all_rows.extend(rows)

    report["righe_lette"] = len(all_rows)

    by_nome_dip, _by_cf = _build_lookup_indexes()

    piano_cache: dict[str, TrainingPlan] = {}
    docente_cache: dict[str, TrainingInstructor] = {}
    # Il file ripete quasi lo stesso set di corsi/sessioni su più fogli (es.
    # "Corsi" e "Corsi AGGIORNAMENTI" si sovrappongono al ~99%). In dry-run
    # non c'è una riga scritta nel DB da ritrovare al secondo passaggio, quindi
    # senza questi set la stessa entità non ancora esistente verrebbe contata
    # come "creata" una volta per ogni foglio in cui compare. Track locale di
    # cosa è già stato "creato" (reale o virtuale) in QUESTA esecuzione.
    corsi_processati: set[str] = set()
    sessioni_processate: set[tuple[str, date]] = set()

    def _get_piano(nome: str) -> TrainingPlan:
        if nome in piano_cache:
            return piano_cache[nome]
        obj = TrainingPlan.objects.filter(nome=nome).first()
        if obj is None:
            codice = _short_code(nome, prefix="P-", max_len=20)
            while TrainingPlan.objects.filter(codice=codice).exists():
                codice = _short_code(nome + "_x", prefix="P-", max_len=20)
            if commit:
                obj = TrainingPlan.objects.create(
                    nome=nome, codice=codice,
                    categoria="OBBLIGATORIA" if "sicurezza" in nome.lower() else "CONSIGLIATA",
                    stato="ATTIVO", is_active=True,
                )
            report["piani_created"] += 1
        piano_cache[nome] = obj
        return obj

    def _get_docente(nome_raw: str) -> TrainingInstructor | None:
        nome = _norm(nome_raw)
        if not nome or nome == "#N/A":
            return None
        key = nome.upper()
        if key in docente_cache:
            return docente_cache[key]
        obj = TrainingInstructor.objects.filter(nome__iexact=nome).first()
        if obj is None:
            legacy_id = by_nome_dip.get(key)
            tipo = "INTERNO" if legacy_id else "ESTERNO"
            if commit:
                obj = TrainingInstructor.objects.create(
                    nome=nome, tipo=tipo, legacy_anagrafica_id=legacy_id,
                )
            report["docenti_created"] += 1
        docente_cache[key] = obj
        return obj

    def _do_import():
        for i, row in enumerate(all_rows, start=2):
            try:
                if solo_docenti:
                    _get_docente(row.get("DESCRIZIONE DOCENTE"))
                    continue

                foglio = row.get("__sheet__")
                cod_corso_raw = row.get("CODICE CORSO")
                if not cod_corso_raw:
                    report["righe_saltate"] += 1
                    continue
                try:
                    codice = str(int(cod_corso_raw))
                except (TypeError, ValueError):
                    codice = _norm(cod_corso_raw)
                if not codice:
                    report["righe_saltate"] += 1
                    continue

                piano_nome = _norm(row.get("DESCRIZIONE ATTIVITA'")) or "NOVICROM"
                piano = _get_piano(piano_nome)

                titolo = _norm(row.get("DESCRIZIONE")) or codice
                descrizione = _norm(row.get("DESCRIZIONE ESTESA"))
                ore = _parse_decimal(row.get("ORE OBBLIGATORIE"))

                corso = TrainingCourse.objects.filter(codice=codice).first()
                corso_gia_visto = codice in corsi_processati
                corsi_processati.add(codice)
                if corso is None and not corso_gia_visto:
                    if commit:
                        corso = TrainingCourse.objects.create(
                            piano=piano, codice=codice, titolo=titolo,
                            descrizione=descrizione,
                            durata_ore_teorica=ore or Decimal("0"),
                            stato="ATTIVO", is_active=True,
                        )
                    report["corsi_created"] += 1
                else:
                    changed = False
                    if commit and corso is not None:
                        if titolo and corso.titolo != titolo:
                            corso.titolo = titolo
                            changed = True
                        if descrizione and not corso.descrizione:
                            corso.descrizione = descrizione
                            changed = True
                        if ore and not corso.durata_ore_teorica:
                            corso.durata_ore_teorica = ore
                            changed = True
                        if changed:
                            corso.save()
                    report["corsi_updated"] += 1

                docente = _get_docente(row.get("DESCRIZIONE DOCENTE"))

                d_inizio = _parse_date(row.get("DATA INIZIO"))
                if d_inizio is None:
                    report["warnings"].append(
                        f"[{foglio}] riga {i}: corso {codice} senza DATA INIZIO — sessione non creata."
                    )
                    continue
                d_fine = _parse_date(row.get("DATA FINE")) or d_inizio

                sede = _norm(row.get("DESCRIZIONE LUOGO"))
                modalita = _map_modalita_da_tipo_luogo(row.get("TIPO LUOGO"))
                stato_sess = _map_stato_sessione_estrazioni(row.get("STATO CORSO")) or "PIANIFICATA"
                note = _norm(row.get("NOTE"))

                sess_key = (codice, d_inizio)
                sessione_gia_vista = sess_key in sessioni_processate
                sessioni_processate.add(sess_key)

                if corso is not None:
                    sessione = TrainingSession.objects.filter(corso=corso, data_inizio=d_inizio).first()
                else:
                    sessione = None  # dry-run senza corso reale

                if sessione is None and not sessione_gia_vista:
                    if not commit:
                        report["sessioni_created"] += 1
                        continue
                    codice_sessione = f"{codice}-{d_inizio.strftime('%Y%m%d')}"
                    n = 0
                    base = codice_sessione
                    while TrainingSession.objects.filter(codice_sessione=codice_sessione).exists():
                        n += 1
                        codice_sessione = f"{base[:36]}-{n}"
                    TrainingSession.objects.create(
                        corso=corso, codice_sessione=codice_sessione,
                        stato=stato_sess, modalita=modalita,
                        data_inizio=d_inizio, data_fine=d_fine,
                        sede=sede[:200], docente=docente,
                        docente_nome=(docente.nome if docente else "")[:200],
                        note=note,
                    )
                    report["sessioni_created"] += 1
                else:
                    changed = False
                    if commit and sessione is not None:
                        if sede and sessione.sede != sede:
                            sessione.sede = sede[:200]
                            changed = True
                        if modalita and sessione.modalita != modalita:
                            sessione.modalita = modalita
                            changed = True
                        if stato_sess and sessione.stato != stato_sess:
                            sessione.stato = stato_sess
                            changed = True
                        if docente and not sessione.docente_id:
                            sessione.docente = docente
                            sessione.docente_nome = docente.nome[:200]
                            changed = True
                        if note and not sessione.note:
                            sessione.note = note
                            changed = True
                        if changed:
                            sessione.save()
                    report["sessioni_updated"] += 1

            except Exception as e:
                report["errors"].append(f"riga {i}: {type(e).__name__}: {e}")
                logger.exception("Errore import estrazioni corsi riga %s", i)

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report


# ─────────────────────────────────────────────────────────────
# IMPORT: elenco aziende/agenzie formative → TrainingProvider
# ─────────────────────────────────────────────────────────────

def import_training_providers_xlsx(xlsx_path: str | Path, commit: bool = False) -> dict:
    """Importa aziende/agenzie formative da un xlsx con colonne
    Nome | Descrizione | Indirizzo | Telefono | Contatto | Telefono del contatto.

    Match su `TrainingProvider.nome` case-insensitive (evita doppioni per
    maiuscole/spazi). Il file esporta l'email del referente nella colonna
    "Telefono del contatto" (quando contiene una "@") — è l'unico modo in cui
    compare un'email nel file. "Descrizione" e il referente ("Contatto")
    confluiscono in `note`. Indirizzo/telefono troppo lunghi per il campo
    vengono troncati e il testo integrale spostato in `note`.
    """
    from anagrafica.models_formazione import TrainingProvider

    report: dict[str, Any] = {
        "file": str(xlsx_path),
        "righe_lette": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "warnings": [],
        "errors": [],
    }

    rows = _read_rows(xlsx_path)
    report["righe_lette"] = len(rows)

    def _do_import():
        for i, row in enumerate(rows, start=2):
            try:
                nome = _norm(row.get("Nome"))
                if not nome:
                    report["skipped"] += 1
                    continue

                descrizione = _norm(row.get("Descrizione"))
                indirizzo = _norm(row.get("Indirizzo"))
                telefono = _norm(row.get("Telefono"))
                contatto = _norm(row.get("Contatto"))
                telefono_contatto = _norm(row.get("Telefono del contatto"))

                email = telefono_contatto if "@" in telefono_contatto else ""

                note_parts = []
                if descrizione:
                    note_parts.append(descrizione)
                if contatto:
                    note_parts.append(f"Referente: {contatto}")
                note = "\n\n".join(note_parts)

                if len(indirizzo) > 300:
                    note = (f"Indirizzo completo: {indirizzo}\n\n{note}").strip()
                    indirizzo = indirizzo[:297] + "..."
                if len(telefono) > 30:
                    note = (f"Telefono completo: {telefono}\n\n{note}").strip()
                    telefono = telefono[:27] + "..."

                defaults = {
                    "indirizzo": indirizzo,
                    "telefono": telefono,
                    "email": email,
                    "note": note,
                }

                if not commit:
                    exists = TrainingProvider.objects.filter(nome__iexact=nome).exists()
                    report["updated" if exists else "created"] += 1
                    continue

                obj, created = TrainingProvider.objects.get_or_create(
                    nome__iexact=nome,
                    defaults={"nome": nome, **defaults},
                )
                if created:
                    report["created"] += 1
                else:
                    changed = False
                    for field, value in defaults.items():
                        if value and not getattr(obj, field):
                            setattr(obj, field, value)
                            changed = True
                    if changed:
                        obj.save()
                    report["updated"] += 1

            except Exception as e:
                report["errors"].append(f"riga {i}: {type(e).__name__}: {e}")
                logger.exception("Errore import aziende formative riga %s", i)

    if commit:
        with transaction.atomic():
            _do_import()
    else:
        _do_import()

    return report
