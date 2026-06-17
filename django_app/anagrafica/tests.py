from __future__ import annotations

import os
import tempfile
from datetime import date
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.legacy_anagrafica import cleanup_duplicate_anagrafica_rows, ensure_anagrafica_schema
from core.legacy_models import UtenteLegacy
from assets.models import Asset, AssetCategory, SoftwareLicense
from .models import (
    DipendenteAnagraficaAziendale,
    DipendenteAnagraficaCivile,
    OffboardingPratica,
    OffboardingTask,
    OnboardingOffboardingCampo,
    SaldoCedolino,
)

User = get_user_model()

VALID_GIF_1X1 = (
    b"GIF87a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00"
    b"\xff\xff\xff,\x00\x00\x00\x00\x01\x00\x01\x00"
    b"\x00\x02\x02D\x01\x00;"
)


def _ensure_anagrafica_table() -> None:
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute("DROP TABLE IF EXISTS anagrafica_dipendenti")
            cursor.execute(
                """
                CREATE TABLE anagrafica_dipendenti (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    aliasusername VARCHAR(200) NULL,
                    nome VARCHAR(200) NULL,
                    cognome VARCHAR(200) NULL,
                    mansione VARCHAR(200) NULL,
                    reparto VARCHAR(200) NULL,
                    ruolo VARCHAR(200) NULL,
                    matricola VARCHAR(100) NULL,
                    attivo INTEGER NULL,
                    email VARCHAR(200) NULL,
                    email_notifica VARCHAR(200) NULL,
                    utente_id INTEGER NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('anagrafica_dipendenti', 'U') IS NOT NULL
                    DROP TABLE anagrafica_dipendenti
                CREATE TABLE anagrafica_dipendenti (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    aliasusername NVARCHAR(200) NULL,
                    nome NVARCHAR(200) NULL,
                    cognome NVARCHAR(200) NULL,
                    mansione NVARCHAR(200) NULL,
                    reparto NVARCHAR(200) NULL,
                    ruolo NVARCHAR(200) NULL,
                    matricola NVARCHAR(100) NULL,
                    attivo BIT NULL,
                    email NVARCHAR(200) NULL,
                    email_notifica NVARCHAR(200) NULL,
                    utente_id INT NULL
                )
                """
            )
    ensure_anagrafica_schema()


def _ensure_utenti_table() -> None:
    vendor = connection.vendor
    with connection.cursor() as cursor:
        if vendor == "sqlite":
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS utenti (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome VARCHAR(200) NOT NULL,
                    email VARCHAR(200) NULL,
                    password VARCHAR(500) NOT NULL,
                    ruolo VARCHAR(100) NULL,
                    attivo INTEGER NOT NULL DEFAULT 1,
                    deve_cambiare_password INTEGER NOT NULL DEFAULT 0,
                    ruolo_id INTEGER NULL
                )
                """
            )
        else:
            cursor.execute(
                """
                IF OBJECT_ID('utenti', 'U') IS NULL
                CREATE TABLE utenti (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    nome NVARCHAR(200) NOT NULL,
                    email NVARCHAR(200) NULL,
                    password NVARCHAR(500) NOT NULL,
                    ruolo NVARCHAR(100) NULL,
                    attivo BIT NOT NULL DEFAULT 1,
                    deve_cambiare_password BIT NOT NULL DEFAULT 0,
                    ruolo_id INT NULL
                )
                """
            )


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AnagraficaRateiExportTests(TestCase):
    def setUp(self):
        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["m.rossi", "Mario", "Rossi", "AMMINISTRAZIONE", 1],
            )
            cursor.execute(
                """
                SELECT id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["m.rossi"],
            )
            self.legacy_id = int(cursor.fetchone()[0])

        self.user = User.objects.create_superuser(
            username="ratei-export",
            email="ratei-export@example.com",
            password="pass12345",
        )

    def test_ratei_export_with_reparto_filter_returns_valid_xlsx(self):
        SaldoCedolino.objects.create(
            tax_code="RSSMRA80A01H501U",
            legacy_anagrafica_id=self.legacy_id,
            data_competenza=date(2026, 5, 31),
            anzianita_anni=3,
            anzianita_mesi=4,
            ferie_anni_prec="1.50",
            ferie_maturati="8.00",
            ferie_goduti="2.00",
            ferie_residui="7.50",
            rol_anni_prec="0.00",
            rol_maturati="4.00",
            rol_goduti="1.00",
            rol_residui="3.00",
            ex_fest_anni_prec="0.00",
            ex_fest_maturati="2.00",
            ex_fest_goduti="0.00",
            ex_fest_residui="2.00",
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("anagrafica:ratei_export"),
            {"periodo": "2026-05-31", "reparto": "AMMINISTRAZIONE"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb["Ratei Ferie"]
        self.assertEqual(ws["A1"].value, "Dipendente")
        self.assertEqual(ws["E1"].value, "Ferie")
        self.assertEqual(ws["E2"].value, "Anni Prec.")
        self.assertEqual(ws["A3"].value, "Rossi Mario")
        self.assertEqual(ws["B3"].value, "AMMINISTRAZIONE")


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class AnagraficaDipendentiViewTests(TestCase):
    def setUp(self):
        _ensure_anagrafica_table()
        _ensure_utenti_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            cursor.execute("DELETE FROM utenti")
        self.user = User.objects.create_superuser(
            username="anagrafica-view",
            email="anagrafica-view@example.com",
            password="pass12345",
        )

    def test_can_create_inactive_employee_without_account(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("anagrafica:dipendente_create"),
            {
                "nome": "Mario",
                "cognome": "Rossi",
                "aliasusername": "m.rossi",
                "matricola": "MR001",
                "reparto": "Produzione",
                "mansione": "Saldatore",
                "ruolo": "Operaio",
                "email_notifica": "m.rossi@example.com",
            },
        )

        self.assertEqual(response.status_code, 302)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT nome, cognome, attivo, utente_id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["m.rossi"],
            )
            row = cursor.fetchone()

        self.assertIsNotNone(row)
        self.assertEqual(row[0], "Mario")
        self.assertEqual(row[1], "Rossi")
        self.assertEqual(int(row[2] or 0), 0)
        self.assertIsNone(row[3])

    def test_create_employee_defaults_notification_email_to_account_email(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("anagrafica:dipendente_create"),
            {
                "nome": "Federico",
                "cognome": "Bernini",
                "aliasusername": "f.bernini",
                "email": "federicobernini3@gmail.com",
                "attivo": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT email, email_notifica
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["f.bernini"],
            )
            row = cursor.fetchone()

        self.assertIsNotNone(row)
        self.assertEqual(row[0], "federicobernini3@gmail.com")
        self.assertEqual(row[1], "federicobernini3@gmail.com")

    def test_detail_falls_back_to_account_email_when_notification_email_is_missing(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, email, email_notifica, attivo)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
                """,
                ["f.bernini", "Federico", "Bernini", "federicobernini3@gmail.com", "", 1],
            )
            cursor.execute(
                """
                SELECT id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["f.bernini"],
            )
            row = cursor.fetchone()

        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:dipendente_detail", args=[int(row[0])]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["dip"]["email_notifica"], "federicobernini3@gmail.com")
        self.assertContains(response, "federicobernini3@gmail.com")

    def test_detail_shows_assigned_software_licenses(self):
        asset_category = AssetCategory.objects.create(
            code="anagrafica-license",
            label="Categoria Licenze",
            base_asset_type=Asset.TYPE_PC,
        )
        asset = Asset.objects.create(
            name="Laptop marketing",
            asset_type=Asset.TYPE_PC,
            asset_category=asset_category,
            reparto="MKT",
            source_key="anagrafica-asset-license",
        )
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, email, email_notifica, attivo)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
                """,
                ["l.gallo", "Lara", "Gallo", "l.gallo@example.com", "l.gallo@example.com", 1],
            )
            cursor.execute(
                """
                SELECT id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["l.gallo"],
            )
            row = cursor.fetchone()

        SoftwareLicense.objects.create(
            category=SoftwareLicense.CATEGORY_OFFICE,
            vendor="Microsoft",
            product_name="Office 365",
            assigned_anagrafica_id=int(row[0]),
            assigned_to_display="Gallo Lara",
            asset=asset,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:dipendente_detail", args=[int(row[0])]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Office 365")

    def test_list_deduplicates_uppercase_legacy_rows(self):
        with connection.cursor() as cursor:
            vendor = connection.vendor
            if vendor == "sqlite":
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, matricola, ruolo, email)
                    VALUES
                        (NULL, 'DERYA', 'AKSOY', 'INT010', 'Operatore Aggiustaggio', 'legacy@test.local')
                    """
                )
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, email)
                    VALUES
                        ('d.aksoy', 'Derya', 'Aksoy', 'd.aksoy@example.local')
                    """
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, matricola, ruolo, email)
                    VALUES
                        (NULL, 'DERYA', 'AKSOY', 'INT010', 'Operatore Aggiustaggio', 'legacy@test.local')
                    """
                )
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, email)
                    VALUES
                        ('d.aksoy', 'Derya', 'Aksoy', 'd.aksoy@example.local')
                    """
                )

        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:dipendenti_list"))
        html = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(html.count(">Aksoy Derya<"), 1)
        self.assertNotIn(">AKSOY DERYA<", html)

    def test_ex_dipendenti_are_separated_from_active_list(self):
        """Un dipendente con data_cessazione non compare nella lista in forza
        ma compare nella vista dedicata agli ex dipendenti."""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti (aliasusername, nome, cognome, attivo)
                VALUES (%s, %s, %s, %s)
                """,
                ["m.attivo", "Marco", "Attivo", 1],
            )
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti (aliasusername, nome, cognome, attivo)
                VALUES (%s, %s, %s, %s)
                """,
                ["e.cessato", "Elia", "Cessato", 0],
            )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s",
                ["e.cessato"],
            )
            ex_legacy_id = int(cursor.fetchone()[0])

        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=ex_legacy_id,
            data_cessazione=date(2025, 12, 31),
        )

        self.client.force_login(self.user)

        list_html = self.client.get(reverse("anagrafica:dipendenti_list")).content.decode()
        self.assertIn(">Attivo Marco<", list_html)
        self.assertNotIn(">Cessato Elia<", list_html)

        ex_response = self.client.get(reverse("anagrafica:ex_dipendenti_list"))
        ex_html = ex_response.content.decode()
        self.assertEqual(ex_response.status_code, 200)
        self.assertIn(">Cessato Elia<", ex_html)
        self.assertNotIn(">Attivo Marco<", ex_html)
        self.assertEqual(ex_response.context["page_obj"].paginator.count, 1)

    def test_offboarding_licenziamento_creates_pratica_then_closes_employee(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, attivo, utente_id)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["l.licenziamento", "Luca", "Licenziamento", 1, 123],
            )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s",
                ["l.licenziamento"],
            )
            legacy_id = int(cursor.fetchone()[0])

        self.client.force_login(self.user)
        detail_response = self.client.get(reverse("anagrafica:dipendente_detail", args=[legacy_id]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Avvia uscita dipendente")
        self.assertContains(detail_response, "Data uscita")
        self.assertContains(detail_response, "Restituzioni")
        self.assertNotContains(detail_response, 'name="ultimo_giorno_operativo"')
        self.assertNotContains(detail_response, "Onboarding / Offboarding")
        index_response = self.client.get(reverse("anagrafica:index"))
        self.assertContains(index_response, "Nuovo dipendente")
        self.assertNotContains(index_response, "Onboarding / Offboarding")

        data_cessazione = date.today()
        with patch("core.audit.log_action") as log_action:
            response = self.client.post(
                reverse("anagrafica:dipendente_offboarding_licenziamento", args=[legacy_id]),
                {
                    "motivo": OffboardingPratica.MOTIVO_LICENZIAMENTO,
                    "data_cessazione": data_cessazione.isoformat(),
                    "restituzioni": ["badge_chiavi", "device_it"],
                    "restituzioni_note": "Recuperare anche il telecomando cancello.",
                },
            )

        self.assertEqual(response.status_code, 302)
        audit_detail = log_action.call_args.args[3]
        self.assertEqual(audit_detail["data_cessazione_prevista"], data_cessazione.isoformat())
        self.assertEqual(audit_detail["ultimo_giorno_operativo"], data_cessazione.isoformat())
        self.assertFalse(audit_detail["account_scollegato"])
        self.assertEqual(audit_detail["restituzioni_richieste"], ["badge_chiavi", "device_it"])
        self.assertEqual(
            audit_detail["restituzioni_richieste_label"],
            ["Badge, chiavi, tessere", "PC, telefono, SIM, token"],
        )
        self.assertEqual(audit_detail["restituzioni_note"], "Recuperare anche il telecomando cancello.")

        pratica = OffboardingPratica.objects.get(legacy_anagrafica_id=legacy_id)
        self.assertEqual(pratica.stato, OffboardingPratica.STATO_IN_CORSO)
        self.assertEqual(pratica.data_cessazione_prevista, data_cessazione)
        self.assertEqual(pratica.ultimo_giorno_operativo, data_cessazione)
        self.assertEqual(pratica.utente_id_pre_offboarding, 123)
        self.assertEqual(pratica.tasks.count(), 5)
        self.assertTrue(pratica.tasks.filter(codice="hr_documenti_finali").exists())
        self.assertTrue(pratica.tasks.filter(codice="it_revoca_accessi").exists())
        self.assertTrue(pratica.tasks.filter(codice="responsabile_passaggio_consegne").exists())
        self.assertTrue(pratica.tasks.filter(codice="restituzione_badge_chiavi").exists())
        self.assertTrue(pratica.tasks.filter(codice="restituzione_device_it").exists())

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT attivo, utente_id FROM anagrafica_dipendenti WHERE id = %s",
                [legacy_id],
            )
            row = cursor.fetchone()
        self.assertEqual(int(row[0] or 0), 1)
        self.assertEqual(int(row[1] or 0), 123)

        list_html = self.client.get(reverse("anagrafica:dipendenti_list")).content.decode()
        ex_html = self.client.get(reverse("anagrafica:ex_dipendenti_list")).content.decode()
        self.assertIn(">Licenziamento Luca<", list_html)
        self.assertNotIn(">Licenziamento Luca<", ex_html)

        detail_response = self.client.get(reverse("anagrafica:dipendente_detail", args=[legacy_id]))
        self.assertContains(detail_response, "Pratica offboarding in corso")
        self.assertContains(detail_response, "Conferma chiusura rapporto")

        response = self.client.post(
            reverse("anagrafica:dipendente_offboarding_chiudi", args=[legacy_id, pratica.id])
        )
        self.assertEqual(response.status_code, 302)
        pratica.refresh_from_db()
        self.assertEqual(pratica.stato, OffboardingPratica.STATO_IN_CORSO)

        for task in pratica.tasks.all():
            response = self.client.post(
                reverse(
                    "anagrafica:dipendente_offboarding_task_update",
                    args=[legacy_id, pratica.id, task.id],
                ),
                {"stato": OffboardingTask.STATO_COMPLETATO, "note": "OK"},
            )
            self.assertEqual(response.status_code, 302)

        response = self.client.post(
            reverse("anagrafica:dipendente_offboarding_chiudi", args=[legacy_id, pratica.id])
        )
        self.assertEqual(response.status_code, 302)
        pratica.refresh_from_db()
        self.assertEqual(pratica.stato, OffboardingPratica.STATO_CHIUSA)
        self.assertIsNotNone(pratica.closed_at)

        aziendale = DipendenteAnagraficaAziendale.objects.get(legacy_anagrafica_id=legacy_id)
        self.assertEqual(aziendale.data_cessazione, data_cessazione)
        self.assertEqual(aziendale.utente_id_pre_offboarding, 123)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT attivo, utente_id FROM anagrafica_dipendenti WHERE id = %s",
                [legacy_id],
            )
            row = cursor.fetchone()
        self.assertEqual(int(row[0] or 0), 0)
        self.assertIsNone(row[1])

        list_html = self.client.get(reverse("anagrafica:dipendenti_list")).content.decode()
        ex_html = self.client.get(reverse("anagrafica:ex_dipendenti_list")).content.decode()
        self.assertNotIn(">Licenziamento Luca<", list_html)
        self.assertIn(">Licenziamento Luca<", ex_html)

    def test_offboarding_creates_tasks_from_configured_fields(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, attivo, utente_id)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["c.config", "Carlo", "Config", 1, 321],
            )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s",
                ["c.config"],
            )
            legacy_id = int(cursor.fetchone()[0])

        OnboardingOffboardingCampo.objects.create(
            fase=OnboardingOffboardingCampo.FASE_OFFBOARDING,
            campo_key="badge",
            campo_label="Badge",
            sezione="Ruolo e organizzazione",
            categoria=OnboardingOffboardingCampo.CATEGORIA_HR,
            obbligatorio=True,
            ordine=5,
            note="Recuperare badge fisico.",
        )

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("anagrafica:dipendente_offboarding_licenziamento", args=[legacy_id]),
            {"data_cessazione": date.today().isoformat()},
        )
        self.assertEqual(response.status_code, 302)

        pratica = OffboardingPratica.objects.get(legacy_anagrafica_id=legacy_id)
        task = pratica.tasks.get(codice="campo_badge")
        self.assertEqual(task.titolo, "Verificare Badge")
        self.assertEqual(task.categoria, OffboardingTask.CATEGORIA_HR)
        self.assertIn("Recuperare badge fisico.", task.descrizione)

    def test_rimetti_in_forza_relinks_saved_pre_offboarding_account(self):
        legacy_user = UtenteLegacy.objects.create(
            nome="Account Storico",
            email="portal.storico@example.com",
            password="x",
            ruolo="Dipendente",
            attivo=True,
        )
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, email, attivo, utente_id)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
                """,
                ["s.storico", "Sara", "Storico", "legacy.storico@example.com", 0, None],
            )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s",
                ["s.storico"],
            )
            legacy_id = int(cursor.fetchone()[0])

        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=legacy_id,
            data_cessazione=date(2026, 2, 15),
            utente_id_pre_offboarding=legacy_user.id,
        )

        self.client.force_login(self.user)
        response = self.client.post(reverse("anagrafica:dipendente_rimetti_in_forza", args=[legacy_id]))
        self.assertEqual(response.status_code, 302)

        aziendale = DipendenteAnagraficaAziendale.objects.get(legacy_anagrafica_id=legacy_id)
        self.assertIsNone(aziendale.data_cessazione)
        self.assertIsNone(aziendale.utente_id_pre_offboarding)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT attivo, utente_id FROM anagrafica_dipendenti WHERE id = %s",
                [legacy_id],
            )
            row = cursor.fetchone()
        self.assertEqual(int(row[0] or 0), 1)
        self.assertEqual(int(row[1] or 0), legacy_user.id)

    def test_rimetti_in_forza_clears_cessazione_and_restores_active_status(self):
        legacy_user = UtenteLegacy.objects.create(
            nome="Rita Forza",
            email="r.forza@example.com",
            password="x",
            ruolo="Dipendente",
            attivo=True,
        )
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, email, attivo, utente_id)
                VALUES
                    (%s, %s, %s, %s, %s, %s)
                """,
                ["r.forza", "Rita", "Forza", "r.forza@example.com", 0, None],
            )
            cursor.execute(
                "SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s",
                ["r.forza"],
            )
            legacy_id = int(cursor.fetchone()[0])

        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=legacy_id,
            data_cessazione=date(2026, 1, 31),
        )

        self.client.force_login(self.user)
        detail_response = self.client.get(reverse("anagrafica:dipendente_detail", args=[legacy_id]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "Rimetti in forza")

        response = self.client.post(reverse("anagrafica:dipendente_rimetti_in_forza", args=[legacy_id]))
        self.assertEqual(response.status_code, 302)

        aziendale = DipendenteAnagraficaAziendale.objects.get(legacy_anagrafica_id=legacy_id)
        self.assertIsNone(aziendale.data_cessazione)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT attivo, utente_id FROM anagrafica_dipendenti WHERE id = %s",
                [legacy_id],
            )
            row = cursor.fetchone()
        self.assertEqual(int(row[0] or 0), 1)
        self.assertEqual(int(row[1] or 0), legacy_user.id)

        list_html = self.client.get(reverse("anagrafica:dipendenti_list")).content.decode()
        ex_html = self.client.get(reverse("anagrafica:ex_dipendenti_list")).content.decode()
        self.assertIn(">Forza Rita<", list_html)
        self.assertNotIn(">Forza Rita<", ex_html)

    def test_list_orders_employees_by_surname_and_name_on_first_load(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["z.zeta", "Zoe", "Zeta", "Produzione", 1],
            )
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["a.alfa", "Anna", "Alfa", "Produzione", 1],
            )

        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:dipendenti_list"))

        html = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertLess(html.index("Alfa Anna"), html.index("Zeta Zoe"))

    def test_list_uses_employee_photo_or_gray_avatar_fallback(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["a.alfa", "Anna", "Alfa", "Produzione", 1],
            )
            cursor.execute("SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["a.alfa"])
            photo_legacy_id = int(cursor.fetchone()[0])
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["b.beta", "Bruno", "Beta", "Produzione", 1],
            )

        with tempfile.TemporaryDirectory() as private_root, override_settings(
            ANAGRAFICA_PRIVATE_ROOT=private_root
        ):
            civile = DipendenteAnagraficaCivile.objects.create(legacy_anagrafica_id=photo_legacy_id)
            civile.foto.save("profilo.gif", ContentFile(VALID_GIF_1X1), save=True)

            self.client.force_login(self.user)
            response = self.client.get(reverse("anagrafica:dipendenti_list"))

            html = response.content.decode()
            self.assertEqual(response.status_code, 200)
            # La foto è un dato personale: NON più su /media/ pubblico, ma servita
            # dalla view protetta anagrafica:foto_dipendente.
            self.assertIn(
                reverse("anagrafica:foto_dipendente", args=[photo_legacy_id]),
                html,
            )
            self.assertNotIn(f"/media/anagrafica/dipendenti/{photo_legacy_id}/foto/", html)
            self.assertIn("ana-avatar-img", html)
            self.assertIn("ana-avatar-fallback", html)

    def test_civil_form_can_upload_employee_photo(self):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["a.alfa", "Anna", "Alfa", "Produzione", 1],
            )
            cursor.execute("SELECT id FROM anagrafica_dipendenti WHERE aliasusername = %s", ["a.alfa"])
            legacy_id = int(cursor.fetchone()[0])

        with tempfile.TemporaryDirectory() as private_root, override_settings(
            ANAGRAFICA_PRIVATE_ROOT=private_root
        ):
            self.client.force_login(self.user)
            response = self.client.post(
                reverse("anagrafica:dipendente_civile_save", args=[legacy_id]),
                {
                    "foto": SimpleUploadedFile(
                        "profilo.gif",
                        VALID_GIF_1X1,
                        content_type="image/gif",
                    )
                },
            )

            self.assertEqual(response.status_code, 302)
            civile = DipendenteAnagraficaCivile.objects.get(legacy_anagrafica_id=legacy_id)
            self.assertTrue(civile.foto.name.startswith(f"anagrafica/dipendenti/{legacy_id}/foto/"))
            # Il file deve essere salvato nello storage privato (fuori webroot),
            # non più sotto MEDIA_ROOT pubblico.
            self.assertTrue(
                os.path.exists(os.path.join(private_root, civile.foto.name.replace("/", os.sep)))
            )

    def test_cleanup_duplicate_rows_merges_and_deletes_duplicates(self):
        with connection.cursor() as cursor:
            if connection.vendor == "sqlite":
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, matricola, ruolo, email)
                    VALUES
                        (NULL, 'DERYA', 'AKSOY', 'INT010', 'Operatore Aggiustaggio', 'legacy@test.local')
                    """
                )
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, email)
                    VALUES
                        ('d.aksoy', 'Derya', 'Aksoy', 'd.aksoy@example.local')
                    """
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, matricola, ruolo, email)
                    VALUES
                        (NULL, 'DERYA', 'AKSOY', 'INT010', 'Operatore Aggiustaggio', 'legacy@test.local')
                    """
                )
                cursor.execute(
                    """
                    INSERT INTO anagrafica_dipendenti
                        (aliasusername, nome, cognome, email)
                    VALUES
                        ('d.aksoy', 'Derya', 'Aksoy', 'd.aksoy@example.local')
                    """
                )

        summary = cleanup_duplicate_anagrafica_rows()
        self.assertEqual(summary["groups"], 1)
        self.assertEqual(summary["rows_deleted"], 1)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT nome, cognome, aliasusername, matricola, ruolo
                FROM anagrafica_dipendenti
                """
            )
            rows = cursor.fetchall()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "Derya")
        self.assertEqual(rows[0][1], "Aksoy")
        self.assertEqual(rows[0][2], "d.aksoy")
        self.assertEqual(rows[0][3], "INT010")
        self.assertEqual(rows[0][4], "Operatore Aggiustaggio")


# ===========================================================================
# Visite mediche / Documenti dipendente / Servizi DPI ingresso
# ===========================================================================

from datetime import timedelta

from .models import (
    AnagraficaVisiteMedichePermission,
    DipendenteRuoloOperativo,
    DocumentoDipendente,
    RuoloOperativo,
    TipoVisitaMedica,
    VisitaMedica,
)
from .services.visite import (
    STATO_IN_SCADENZA,
    STATO_MANCANTE,
    STATO_SCADUTA,
    STATO_VALIDA,
    stato_visite,
    tipi_visita_richiesti_per_dipendente,
)


class VisitaMedicaScadenzaTests(TestCase):
    def setUp(self):
        self.tipo_12 = TipoVisitaMedica.objects.create(nome="Tipo 12 mesi", durata_mesi=12)
        self.tipo_24 = TipoVisitaMedica.objects.create(nome="Tipo 24 mesi", durata_mesi=24)
        self.tipo_0 = TipoVisitaMedica.objects.create(nome="Senza scadenza", durata_mesi=0)

    def test_scadenza_calcolata_12_mesi(self):
        v = VisitaMedica.objects.create(
            legacy_anagrafica_id=1,
            tipo=self.tipo_12,
            data_svolgimento=date(2026, 1, 15),
        )
        self.assertEqual(v.data_scadenza, date(2027, 1, 15))

    def test_scadenza_calcolata_24_mesi(self):
        v = VisitaMedica.objects.create(
            legacy_anagrafica_id=1,
            tipo=self.tipo_24,
            data_svolgimento=date(2026, 5, 1),
        )
        self.assertEqual(v.data_scadenza, date(2028, 5, 1))

    def test_scadenza_durata_zero_non_imposta_data(self):
        v = VisitaMedica.objects.create(
            legacy_anagrafica_id=1,
            tipo=self.tipo_0,
            data_svolgimento=date(2026, 5, 1),
        )
        self.assertIsNone(v.data_scadenza)

    def test_is_scaduta_e_in_scadenza(self):
        from django.utils import timezone
        oggi = timezone.localdate()
        v_scaduta = VisitaMedica.objects.create(
            legacy_anagrafica_id=1,
            tipo=self.tipo_12,
            data_svolgimento=oggi - timedelta(days=400),
        )
        self.assertTrue(v_scaduta.is_scaduta)

        v_in_scad = VisitaMedica.objects.create(
            legacy_anagrafica_id=2,
            tipo=self.tipo_12,
            data_svolgimento=oggi - timedelta(days=340),  # scadenza tra ~25g
        )
        self.assertFalse(v_in_scad.is_scaduta)
        self.assertTrue(v_in_scad.in_scadenza)


class StatoVisiteServiceTests(TestCase):
    def setUp(self):
        self.ruolo = RuoloOperativo.objects.create(nome="Saldatore")
        self.tipo = TipoVisitaMedica.objects.create(
            nome="Visita saldatori", durata_mesi=12, obbligatoria=True,
        )
        self.tipo.ruoli_operativi.add(self.ruolo)
        self.tipo_non_obb = TipoVisitaMedica.objects.create(
            nome="Visita facoltativa", durata_mesi=12, obbligatoria=False,
        )
        self.tipo_non_obb.ruoli_operativi.add(self.ruolo)

    def test_dipendente_senza_ruoli_non_ha_visite_richieste(self):
        self.assertEqual(tipi_visita_richiesti_per_dipendente(999), [])

    def test_dipendente_con_ruolo_vede_solo_obbligatorie(self):
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=10, ruolo=self.ruolo)
        tipi = tipi_visita_richiesti_per_dipendente(10)
        nomi = {t.nome for t in tipi}
        self.assertIn("Visita saldatori", nomi)
        self.assertNotIn("Visita facoltativa", nomi)

    def test_stato_visite_mancante(self):
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=11, ruolo=self.ruolo)
        out = stato_visite(11)
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0]["stato"], STATO_MANCANTE)

    def test_stato_visite_valida_e_scaduta(self):
        from django.utils import timezone
        oggi = timezone.localdate()
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=12, ruolo=self.ruolo)
        VisitaMedica.objects.create(
            legacy_anagrafica_id=12, tipo=self.tipo, data_svolgimento=oggi - timedelta(days=30),
        )
        out = stato_visite(12)
        self.assertEqual(out[0]["stato"], STATO_VALIDA)

        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=13, ruolo=self.ruolo)
        VisitaMedica.objects.create(
            legacy_anagrafica_id=13, tipo=self.tipo, data_svolgimento=oggi - timedelta(days=400),
        )
        out = stato_visite(13)
        self.assertEqual(out[0]["stato"], STATO_SCADUTA)


class DPIPDFRenderTests(TestCase):
    def test_render_modulo_consegna_dpi_produce_pdf_bytes(self):
        from dpi.models import (
            CategoriaDPI, ConsegnaDPI, RichiestaDPI, StatoRichiesta,
        )
        from dpi.pdf import render_modulo_consegna_dpi

        cat = CategoriaDPI.objects.create(nome="Elmetto", vita_utile_giorni=365)
        richiesta = RichiestaDPI.objects.create(
            categoria=cat,
            quantita=1,
            stato=StatoRichiesta.APPROVATA,
            richiedente_legacy_id=99,
            richiedente_nome="Test Dipendente",
        )
        consegna = ConsegnaDPI.objects.create(
            richiesta=richiesta,
            data_consegna=date(2026, 5, 21),
            consegnato_da_nome="HR",
            firma_immagine="",
        )
        pdf_bytes = render_modulo_consegna_dpi(consegna)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertGreater(len(pdf_bytes), 1000)


class VisiteMedichePermissionTests(TestCase):
    def setUp(self):
        self.user_super = User.objects.create_superuser(
            username="su", email="su@test.local", password="x"
        )
        self.user_plain = User.objects.create_user(
            username="plain", email="plain@test.local", password="x"
        )

    def test_superuser_sempre_autorizzato(self):
        from .views import _can_view_visite_mediche
        class FakeReq:
            user = self.user_super
        # Lo helper accetta `request`-like con `.user`. Funziona anche con questo stub.
        self.assertTrue(_can_view_visite_mediche(FakeReq()))

    def test_default_admin_blocca_utente_normale(self):
        from .views import _can_view_visite_mediche
        AnagraficaVisiteMedichePermission.objects.all().delete()
        # default get_instance crea ACCESSO_ADMIN
        class FakeReq:
            user = self.user_plain
        self.assertFalse(_can_view_visite_mediche(FakeReq()))


class DocumentoDipendenteDownloadACLTests(TestCase):
    def setUp(self):
        self.user_super = User.objects.create_superuser(
            username="su2", email="su2@test.local", password="x"
        )
        self.user_plain = User.objects.create_user(
            username="plain2", email="plain2@test.local", password="x"
        )
        # Forziamo ACCESSO_ADMIN per le visite (default)
        perm = AnagraficaVisiteMedichePermission.get_instance()
        perm.accesso = AnagraficaVisiteMedichePermission.ACCESSO_ADMIN
        perm.save()

        self.doc_referto = DocumentoDipendente.objects.create(
            legacy_anagrafica_id=1,
            tipo=DocumentoDipendente.Tipo.VISITA_MEDICA_REFERTO,
            nome_originale="referto.pdf",
        )
        # Salviamo un file vuoto per evitare 404 prima del 403
        self.doc_referto.file.save("referto.pdf", ContentFile(b"%PDF-1.4 stub"), save=True)

    def test_referto_403_per_utente_normale(self):
        # Bypassa eventuali middleware globali (SetupWizard, NotizieMandatory)
        # chiamando direttamente la view: il test interessa la sola ACL.
        from django.test import RequestFactory
        from .views import documento_dipendente_download

        rf = RequestFactory()
        request = rf.get(f"/anagrafica/documenti/{self.doc_referto.id}/download")
        request.user = self.user_plain
        resp = documento_dipendente_download(request, self.doc_referto.id)
        self.assertEqual(resp.status_code, 403)

    def test_referto_autorizzato_per_superuser(self):
        from django.test import RequestFactory
        from .views import documento_dipendente_download

        rf = RequestFactory()
        request = rf.get(f"/anagrafica/documenti/{self.doc_referto.id}/download")
        request.user = self.user_super
        resp = documento_dipendente_download(request, self.doc_referto.id)
        # FileResponse o HttpResponse 200/404 (in test il file dovrebbe esserci)
        self.assertNotEqual(resp.status_code, 403)


class DocumentoDipendenteListTests(TestCase):
    def setUp(self):
        _ensure_anagrafica_table()
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM anagrafica_dipendenti")
            cursor.execute(
                """
                INSERT INTO anagrafica_dipendenti
                    (aliasusername, nome, cognome, reparto, attivo)
                VALUES
                    (%s, %s, %s, %s, %s)
                """,
                ["l.bianchi", "Luca", "Bianchi", "HR", 1],
            )
            cursor.execute(
                """
                SELECT id
                FROM anagrafica_dipendenti
                WHERE aliasusername = %s
                """,
                ["l.bianchi"],
            )
            self.legacy_id = int(cursor.fetchone()[0])

        self.user_super = User.objects.create_superuser(
            username="docs-admin", email="docs-admin@test.local", password="x"
        )
        self.doc = DocumentoDipendente.objects.create(
            legacy_anagrafica_id=self.legacy_id,
            tipo=DocumentoDipendente.Tipo.MANUALE,
            nome_originale="contratto.pdf",
            descrizione="Documento contratto",
            created_by=self.user_super,
            created_by_display="Admin Docs",
        )

    def test_documenti_list_renderizza_nome_dipendente(self):
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.backends.signed_cookies import SessionStore
        from django.test import RequestFactory
        from .views import documenti_list

        rf = RequestFactory()
        request = rf.get("/anagrafica/documenti/")
        request.user = self.user_super
        request.session = SessionStore()
        request._messages = FallbackStorage(request)

        resp = documenti_list(request)

        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8", errors="ignore")
        self.assertIn("Bianchi Luca", body)
        self.assertIn("contratto.pdf", body)


class DocumentoDipendenteUploadTests(TestCase):
    def setUp(self):
        self.user_super = User.objects.create_superuser(
            username="docs-upload-admin",
            email="docs-upload-admin@test.local",
            password="x",
        )

    def test_documento_upload_salva_pdf_manuale(self):
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.backends.signed_cookies import SessionStore
        from django.test import RequestFactory
        from .views import documento_dipendente_upload

        with tempfile.TemporaryDirectory() as private_root, override_settings(
            ANAGRAFICA_PRIVATE_ROOT=private_root
        ):
            uploaded = SimpleUploadedFile(
                "contratto.pdf",
                b"%PDF-1.4\n% test pdf\n",
                content_type="application/pdf",
            )
            request = RequestFactory().post(
                "/anagrafica/dipendenti/277/documenti/upload",
                data={"file": uploaded, "descrizione": "Contratto firmato"},
            )
            request.user = self.user_super
            request.session = SessionStore()
            request._messages = FallbackStorage(request)

            resp = documento_dipendente_upload(request, 277)

            self.assertEqual(resp.status_code, 302)
            doc = DocumentoDipendente.objects.get(legacy_anagrafica_id=277)
            self.assertEqual(doc.tipo, DocumentoDipendente.Tipo.MANUALE)
            self.assertEqual(doc.nome_originale, "contratto.pdf")
            self.assertEqual(doc.tipo_mime, "application/pdf")
            self.assertEqual(doc.descrizione, "Contratto firmato")
            self.assertTrue(doc.file.name.endswith(".pdf"))


class ImpostazioniRedirectTests(TestCase):
    def setUp(self):
        self.user_super = User.objects.create_superuser(
            username="settings-admin", email="settings-admin@test.local", password="x"
        )

    def _post_request(self, path: str, data: dict):
        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.backends.signed_cookies import SessionStore
        from django.test import RequestFactory

        request = RequestFactory().post(path, data=data)
        request.user = self.user_super
        request.session = SessionStore()
        request._messages = FallbackStorage(request)
        return request

    def test_cartella_documento_create_redirects_to_documenti_tab(self):
        from .models import CartellaDocumentoDipendente
        from .views import cartella_documento_create

        request = self._post_request(
            "/anagrafica/cartelle-documenti/nuovo",
            {"nome": "Contratti", "descrizione": "Archivio contratti", "ordine": "10"},
        )

        resp = cartella_documento_create(request)

        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp["Location"], f"{reverse('anagrafica:impostazioni')}?tab=documenti#tab-documenti")
        self.assertTrue(CartellaDocumentoDipendente.objects.filter(nome="Contratti").exists())

    def test_subnav_categoria_create_redirects_to_navigazione_tab(self):
        from .models import SubnavCategoriaAnagrafica
        from .views import subnav_categoria_create

        request = self._post_request(
            "/anagrafica/subnav/categoria/nuovo",
            {"nome": "Custom", "icona": "", "ordine": "5"},
        )

        resp = subnav_categoria_create(request)

        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp["Location"], f"{reverse('anagrafica:impostazioni')}?tab=navigazione#tab-navigazione")
        self.assertTrue(SubnavCategoriaAnagrafica.objects.filter(nome="Custom").exists())

    def test_workflow_settings_maps_new_employee_field(self):
        self.client.force_login(self.user_super)
        response = self.client.get(reverse("anagrafica:impostazioni"), {"tab": "workflow"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Campi onboarding / offboarding")
        self.assertContains(response, "Data inizio contratto")

        response = self.client.post(
            reverse("anagrafica:workflow_campo_create"),
            {
                "campo_key": "contratto_data_inizio",
                "fase": OnboardingOffboardingCampo.FASE_ONBOARDING,
                "categoria": OnboardingOffboardingCampo.CATEGORIA_HR,
                "obbligatorio": "1",
                "is_active": "1",
                "ordine": "10",
                "note": "Da compilare prima dell'ingresso.",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(resp_url := response["Location"], f"{reverse('anagrafica:impostazioni')}?tab=workflow#tab-workflow")
        self.assertIn("tab=workflow", resp_url)

        mapping = OnboardingOffboardingCampo.objects.get(campo_key="contratto_data_inizio")
        self.assertEqual(mapping.campo_label, "Data inizio contratto")
        self.assertEqual(mapping.sezione, "Contratto e inquadramento")
        self.assertTrue(mapping.obbligatorio)

        response = self.client.post(
            reverse("anagrafica:workflow_campo_update", args=[mapping.id]),
            {
                "campo_key": "contratto_data_inizio",
                "campo_label": "Data inizio rapporto",
                "fase": OnboardingOffboardingCampo.FASE_OFFBOARDING,
                "categoria": OnboardingOffboardingCampo.CATEGORIA_AMMINISTRAZIONE,
                "is_active": "1",
                "ordine": "20",
                "note": "Verifica documentale finale.",
            },
        )
        self.assertEqual(response.status_code, 302)
        mapping.refresh_from_db()
        self.assertEqual(mapping.fase, OnboardingOffboardingCampo.FASE_OFFBOARDING)
        self.assertEqual(mapping.campo_label, "Data inizio rapporto")
        self.assertFalse(mapping.obbligatorio)


class VisiteMedicheDashboardTests(TestCase):
    def setUp(self):
        self.user_super = User.objects.create_superuser(
            username="su3", email="su3@test.local", password="x"
        )
        self.user_plain = User.objects.create_user(
            username="plain3", email="plain3@test.local", password="x"
        )

    def test_dashboard_403_per_utente_normale(self):
        from django.test import RequestFactory
        from .views import visite_mediche_dashboard

        rf = RequestFactory()
        request = rf.get("/anagrafica/visite-mediche/")
        request.user = self.user_plain
        # _check_session_setup richiede session+messages middleware: usiamo middleware factories
        from django.contrib.messages.storage.fallback import FallbackStorage
        request.session = {}
        request._messages = FallbackStorage(request)
        resp = visite_mediche_dashboard(request)
        # Redirect verso anagrafica:index (302)
        self.assertEqual(resp.status_code, 302)

    def test_dashboard_render_superuser(self):
        # Smoke test: la view ritorna 200 e prepara il context senza errori.
        # Usiamo RequestFactory + chiamiamo la view; per evitare l'accesso
        # ai context processor (es. legacy_nav che richiede session), il
        # template usa solo variabili nostre. La risposta NON renderizza
        # context processors quando si chiama direttamente la view: il
        # rendering avviene comunque, quindi forniamo una session vuota.
        ruolo = RuoloOperativo.objects.create(nome="Operatore")
        tipo = TipoVisitaMedica.objects.create(
            nome="Visita test", durata_mesi=12, obbligatoria=True
        )
        tipo.ruoli_operativi.add(ruolo)
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=77, ruolo=ruolo)

        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.sessions.backends.signed_cookies import SessionStore
        from django.test import RequestFactory
        from .views import visite_mediche_dashboard

        rf = RequestFactory()
        request = rf.get("/anagrafica/visite-mediche/")
        request.user = self.user_super
        request.session = SessionStore()
        request._messages = FallbackStorage(request)
        resp = visite_mediche_dashboard(request)
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8", errors="ignore")
        self.assertIn("Visite mediche", body)
        self.assertIn("Copertura per tipologia", body)


class DPIIngressoServiceTests(TestCase):
    def test_crea_consegne_iniziali_crea_richiesta_e_consegna(self):
        from dpi.models import CategoriaDPI, ConsegnaDPI, RichiestaDPI, StatoRichiesta
        from .services.dpi_ingresso import RigaConsegnaIniziale, crea_consegne_iniziali

        cat = CategoriaDPI.objects.create(
            nome="Guanti", vita_utile_giorni=180, obbligatoria_mansionario=True,
        )
        civile = DipendenteAnagraficaCivile.objects.create(legacy_anagrafica_id=42)
        user = User.objects.create_user(username="hr", email="hr@x.local", password="x")

        consegne = crea_consegne_iniziali(
            civile, None, [RigaConsegnaIniziale(categoria_id=cat.id, quantita=2)], user,
        )
        self.assertEqual(len(consegne), 1)
        consegna = consegne[0]
        self.assertIsInstance(consegna, ConsegnaDPI)
        self.assertEqual(consegna.richiesta.stato, StatoRichiesta.CONSEGNATA)
        self.assertEqual(consegna.richiesta.quantita, 2)
        self.assertEqual(consegna.richiesta.richiedente_legacy_id, 42)
        # scadenza stimata calcolata dalla vita utile della categoria
        from django.utils import timezone
        self.assertEqual(
            consegna.data_scadenza_stimata,
            timezone.localdate() + timedelta(days=180),
        )


# ---------------------------------------------------------------------------
# H3 — send_contratti_expiry_reminders
# ---------------------------------------------------------------------------

class ContrattiExpiryRemindersCommandTests(TestCase):
    """Digest contratti a termine / periodi di prova (management command)."""

    def _run(self, *args):
        from io import StringIO
        from django.core.management import call_command
        out = StringIO()
        call_command("send_contratti_expiry_reminders", *args, stdout=out)
        return out.getvalue()

    def _make_attivo(self, legacy_id: int, **kwargs) -> DipendenteAnagraficaAziendale:
        return DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=legacy_id, **kwargs
        )

    def test_contratto_determinato_in_scadenza_incluso(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        self._make_attivo(101)
        StoricoContratto.objects.create(
            legacy_anagrafica_id=101,
            data_inizio=oggi - timedelta(days=300),
            data_fine=oggi + timedelta(days=30),
            tipologia_contratto="DETERMINATO",
        )
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertIn("CONTRATTI IN SCADENZA", out)
        self.assertIn("#101", out)

    def test_dipendente_cessato_escluso(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        self._make_attivo(102, data_cessazione=oggi - timedelta(days=10))
        StoricoContratto.objects.create(
            legacy_anagrafica_id=102,
            data_inizio=oggi - timedelta(days=300),
            data_fine=oggi + timedelta(days=20),
            tipologia_contratto="DETERMINATO",
        )
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertNotIn("#102", out)

    def test_solo_ultimo_contratto_considerato(self):
        """Un contratto vecchio chiuso non genera alert se l'ultimo è indeterminato."""
        from datetime import timedelta
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        self._make_attivo(103)
        StoricoContratto.objects.create(
            legacy_anagrafica_id=103,
            data_inizio=oggi - timedelta(days=700),
            data_fine=oggi - timedelta(days=400),
            tipologia_contratto="DETERMINATO",
        )
        StoricoContratto.objects.create(
            legacy_anagrafica_id=103,
            data_inizio=oggi - timedelta(days=399),
            data_fine=None,
            tipologia_contratto="INDETERMINATO",
        )
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertNotIn("#103", out)

    def test_periodo_prova_in_sezione_separata(self):
        from datetime import timedelta
        from django.utils import timezone
        oggi = timezone.localdate()
        self._make_attivo(104, prova_data_fine=oggi + timedelta(days=7))
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertIn("PERIODI DI PROVA", out)
        self.assertIn("#104", out)

    def test_fallback_a_termine_senza_storico(self):
        self._make_attivo(105, tipologia_contratto="DETERMINATO")
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertIn("SENZA STORICO IMPORTATO", out)
        self.assertIn("#105", out)

    def test_dry_run_non_invia_email(self):
        from datetime import timedelta
        from django.core import mail
        from django.utils import timezone
        oggi = timezone.localdate()
        self._make_attivo(106, prova_data_fine=oggi + timedelta(days=3))
        self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertEqual(len(mail.outbox), 0)

    def test_invio_reale_con_recipients(self):
        from datetime import timedelta
        from django.core import mail
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        self._make_attivo(107)
        StoricoContratto.objects.create(
            legacy_anagrafica_id=107,
            data_inizio=oggi - timedelta(days=100),
            data_fine=oggi + timedelta(days=10),
            tipologia_contratto="DETERMINATO",
        )
        self._run("--recipients", "hr@example.local")
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("[CONTRATTI]", mail.outbox[0].subject)

    def test_nessuna_voce_output_pulito(self):
        out = self._run("--dry-run", "--recipients", "hr@example.local")
        self.assertIn("Nessun contratto", out)


# ---------------------------------------------------------------------------
# H4 — scadenzario unificato esteso (formazione + contratti/prova)
# ---------------------------------------------------------------------------

class ScadenzarioEstesoTests(TestCase):
    """Le nuove sorgenti (formazione obbligatoria, contratti, prova) entrano
    nello scadenzario unificato e rispettano i gate permessi."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="sc_admin", email="sc_admin@x.local", password="x"
        )

    def setUp(self):
        self.client.force_login(self.admin)

    def _make_deadline(self, legacy_id: int, giorni: int, titolo: str = "Corso Sicurezza Base"):
        from datetime import timedelta
        from django.utils import timezone
        from .models_formazione import TrainingCourse, TrainingDeadline, TrainingPlan
        piano = TrainingPlan.objects.create(codice=f"P{legacy_id}", nome="Piano test")
        corso = TrainingCourse.objects.create(
            piano=piano, codice=f"C{legacy_id}", titolo=titolo,
            durata_ore_teorica=4,
        )
        return TrainingDeadline.objects.create(
            corso=corso,
            legacy_anagrafica_id=legacy_id,
            data_scadenza=timezone.localdate() + timedelta(days=giorni),
            stato_scadenza="IN_SCADENZA_30" if giorni >= 0 else "SCADUTO",
            giorni_alla_scadenza=giorni,
            is_required=True,
        )

    def test_formazione_obbligatoria_in_scadenzario(self):
        self._make_deadline(201, 10, titolo="Antincendio rischio alto")
        resp = self.client.get(reverse("anagrafica:scadenzario"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Antincendio rischio alto")
        self.assertContains(resp, "Formazione")

    def test_formazione_non_obbligatoria_esclusa(self):
        deadline = self._make_deadline(202, 10, titolo="Corso facoltativo X")
        deadline.is_required = False
        deadline.save(update_fields=["is_required"])
        resp = self.client.get(reverse("anagrafica:scadenzario"))
        self.assertNotContains(resp, "Corso facoltativo X")

    def test_qualifica_scaduta_mostra_azione_rinnova(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import TipoQualifica, DipendenteQualifica
        tipo = TipoQualifica.objects.create(nome="Carrellista SC", categoria=TipoQualifica.CAT_SICUREZZA)
        DipendenteQualifica.objects.create(
            legacy_anagrafica_id=205, tipo=tipo,
            data_scadenza=timezone.localdate() - timedelta(days=5),
        )
        resp = self.client.get(reverse("anagrafica:scadenzario"), {"tipo": "qualifica"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, reverse("anagrafica:qualifica_sessione_create") + "?tipo=" + str(tipo.id))

    def test_contratto_in_scadenza_in_scadenzario(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        DipendenteAnagraficaAziendale.objects.create(legacy_anagrafica_id=203)
        StoricoContratto.objects.create(
            legacy_anagrafica_id=203,
            data_inizio=oggi - timedelta(days=100),
            data_fine=oggi + timedelta(days=20),
            tipologia_contratto="DETERMINATO",
        )
        resp = self.client.get(reverse("anagrafica:scadenzario"), {"tipo": "contratto"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Contratto DETERMINATO")

    def test_prova_futura_inclusa_prova_passata_esclusa(self):
        from datetime import timedelta
        from django.utils import timezone
        oggi = timezone.localdate()
        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=204, prova_data_fine=oggi + timedelta(days=5)
        )
        DipendenteAnagraficaAziendale.objects.create(
            legacy_anagrafica_id=205, prova_data_fine=oggi - timedelta(days=30)
        )
        resp = self.client.get(reverse("anagrafica:scadenzario"), {"tipo": "contratto"})
        content = resp.content.decode()
        self.assertIn("Fine periodo di prova", content)
        self.assertEqual(content.count("Fine periodo di prova"), 1)

    def test_contratti_nascosti_senza_permesso_hr(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import StoricoContratto
        oggi = timezone.localdate()
        DipendenteAnagraficaAziendale.objects.create(legacy_anagrafica_id=206)
        StoricoContratto.objects.create(
            legacy_anagrafica_id=206,
            data_inizio=oggi - timedelta(days=100),
            data_fine=oggi + timedelta(days=20),
            tipologia_contratto="DETERMINATO",
        )
        with patch("anagrafica.views._check_hr_permission", return_value=False):
            resp = self.client.get(reverse("anagrafica:scadenzario"))
        self.assertNotContains(resp, "Contratto DETERMINATO")

    def test_csv_include_nuove_voci(self):
        self._make_deadline(207, 10, titolo="Primo soccorso aggiornamento")
        resp = self.client.get(reverse("anagrafica:scadenzario"), {"format": "csv"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Primo soccorso aggiornamento", resp.content.decode("utf-8-sig"))


# ---------------------------------------------------------------------------
# H5 — libretto formativo dipendente
# ---------------------------------------------------------------------------

class LibrettoFormativoTests(TestCase):
    """Pagina stampa libretto formativo: snapshot storici, gate permessi, audit."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="lb_admin", email="lb_admin@x.local", password="x"
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) "
                "VALUES (301, 'Mario', 'Rossi', 'Operatore', 'PROD', 1)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def _make_record(self, titolo_snapshot: str):
        from django.utils import timezone
        from .models_formazione import TrainingCourse, TrainingEmployeeRecord, TrainingPlan
        piano = TrainingPlan.objects.create(codice="PLB1", nome="Piano sicurezza")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CLB1", titolo="Titolo attuale corso",
            durata_ore_teorica=8,
        )
        return TrainingEmployeeRecord.objects.create(
            corso=corso,
            legacy_anagrafica_id=301,
            data_completamento=timezone.localdate(),
            ore_frequentate=8,
            idoneo=True,
            course_title_snapshot=titolo_snapshot,
            course_code_snapshot="CLB1-v1",
        )

    def test_libretto_mostra_snapshot_non_titolo_attuale(self):
        self._make_record("Titolo storico al completamento")
        resp = self.client.get(
            reverse("anagrafica:dipendente_libretto_formativo", args=[301])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Titolo storico al completamento")
        self.assertNotContains(resp, "Titolo attuale corso")

    def test_libretto_dipendente_senza_record(self):
        resp = self.client.get(
            reverse("anagrafica:dipendente_libretto_formativo", args=[301])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Nessun corso completato registrato")

    def test_libretto_negato_senza_permesso_formazione(self):
        with patch("anagrafica.views._can_view_formazione", return_value=False):
            resp = self.client.get(
                reverse("anagrafica:dipendente_libretto_formativo", args=[301])
            )
        self.assertEqual(resp.status_code, 302)

    def test_generazione_tracciata_in_export_log(self):
        from .models_formazione import TrainingExportLog
        self._make_record("Corso X")
        self.client.get(reverse("anagrafica:dipendente_libretto_formativo", args=[301]))
        log = TrainingExportLog.objects.filter(tipo="STORICO_DIP").last()
        self.assertIsNotNone(log)
        self.assertEqual(log.righe_esportate, 1)
        self.assertEqual(log.filtri_json.get("legacy_anagrafica_id"), 301)


# ---------------------------------------------------------------------------
# H5b — attestato di formazione autogenerato
# ---------------------------------------------------------------------------

class AttestatoFormazioneTests(TestCase):
    """Attestato A4 autogenerato per singolo completamento: tipo derivato,
    responsabile, snapshot storici, gate permessi, audit."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="att_admin", email="att_admin@x.local", password="x"
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) "
                "VALUES (401, 'Anna', 'Bianchi', 'Saldatrice', 'PROD', 1)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def _make_record(self, *, qualifica=None, obbligatorio=False, teacher="Ing. Verdi"):
        from django.utils import timezone
        from .models_formazione import TrainingCourse, TrainingEmployeeRecord, TrainingPlan
        piano = TrainingPlan.objects.create(codice="PATT", nome="Piano sicurezza")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CATT", titolo="Titolo attuale",
            durata_ore_teorica=8, obbligatorio=obbligatorio, qualifica=qualifica,
        )
        return TrainingEmployeeRecord.objects.create(
            corso=corso,
            legacy_anagrafica_id=401,
            data_completamento=timezone.localdate(),
            ore_frequentate=8,
            idoneo=True,
            course_title_snapshot="Titolo storico corso",
            teacher_name_snapshot=teacher,
        )

    def test_attestato_render_snapshot_e_firme(self):
        rec = self._make_record()
        resp = self.client.get(
            reverse("anagrafica:attestato_formazione", args=[rec.pk])
        )
        self.assertEqual(resp.status_code, 200)
        # Usa lo snapshot storico, non il titolo attuale del corso
        self.assertContains(resp, "Titolo storico corso")
        self.assertNotContains(resp, "Titolo attuale")
        # Nominativo dall'anagrafica + entrambi i blocchi firma
        self.assertContains(resp, "Bianchi Anna")
        self.assertContains(resp, "Il Responsabile del corso")
        self.assertContains(resp, "Il Dipendente")
        self.assertContains(resp, "Ing. Verdi")

    def test_tipo_partecipazione_default(self):
        rec = self._make_record()
        resp = self.client.get(reverse("anagrafica:attestato_formazione", args=[rec.pk]))
        self.assertContains(resp, "Attestato di partecipazione")

    def test_tipo_qualifica_quando_corso_ancorato(self):
        from .models import TipoQualifica
        qual = TipoQualifica.objects.create(nome="Saldatore certificato")
        rec = self._make_record(qualifica=qual)
        resp = self.client.get(reverse("anagrafica:attestato_formazione", args=[rec.pk]))
        self.assertContains(resp, "Attestato di qualifica")
        self.assertContains(resp, "Saldatore certificato")

    def test_attestato_negato_senza_permesso_formazione(self):
        rec = self._make_record()
        with patch("anagrafica.views._can_view_formazione", return_value=False):
            resp = self.client.get(
                reverse("anagrafica:attestato_formazione", args=[rec.pk])
            )
        self.assertEqual(resp.status_code, 302)

    def test_generazione_tracciata_in_export_log(self):
        from .models_formazione import TrainingExportLog
        rec = self._make_record()
        self.client.get(reverse("anagrafica:attestato_formazione", args=[rec.pk]))
        log = TrainingExportLog.objects.filter(tipo="ATTESTATO").last()
        self.assertIsNotNone(log)
        self.assertEqual(log.righe_esportate, 1)
        self.assertEqual(log.filtri_json.get("record_id"), rec.pk)

    def test_attestato_usa_titolo_configurato(self):
        from .models_formazione import AttestatoFormazioneConfig
        cfg = AttestatoFormazioneConfig.get_instance()
        cfg.titolo_partecipazione = "Certificato interno NOVICROM"
        cfg.firma_dipendente_label = "Il Partecipante"
        cfg.save()
        rec = self._make_record()
        resp = self.client.get(reverse("anagrafica:attestato_formazione", args=[rec.pk]))
        self.assertContains(resp, "Certificato interno NOVICROM")
        self.assertContains(resp, "Il Partecipante")

    def test_impostazioni_get_e_salva(self):
        from .models_formazione import AttestatoFormazioneConfig
        url = reverse("anagrafica:attestato_impostazioni")
        self.assertEqual(self.client.get(url).status_code, 200)
        data = {
            "intestazione_eyebrow": "Formazione interna",
            "sezione_label": "NOVICROM HUB · Attestazione formativa",
            "titolo_partecipazione": "Certificato interno NOVICROM",
            "titolo_frequenza": "Attestato di frequenza",
            "titolo_qualifica": "Attestato di qualifica",
            "formula_attestazione": "Si attesta che",
            "firma_responsabile_label": "Il Responsabile del corso",
            "firma_dipendente_label": "Il Dipendente",
            "responsabile_default": "",
            "mostra_dati_personali": "on",
            "nota_legale": "Nota di prova.",
            "logo_url": "",
            "pie_organizzazione": "NOVICROM HUB",
        }
        resp = self.client.post(url, data)
        self.assertEqual(resp.status_code, 302)
        cfg = AttestatoFormazioneConfig.get_instance()
        self.assertEqual(cfg.titolo_partecipazione, "Certificato interno NOVICROM")
        self.assertEqual(cfg.updated_by_id, self.admin.id)

    def test_impostazioni_negato_senza_permesso_modifica(self):
        with patch("anagrafica.views._can_edit_formazione", return_value=False):
            resp = self.client.get(reverse("anagrafica:attestato_impostazioni"))
        self.assertEqual(resp.status_code, 302)

    def test_variante_stampa_render(self):
        rec = self._make_record()
        resp = self.client.get(
            reverse("anagrafica:attestato_formazione", args=[rec.pk]) + "?stile=stampa"
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, "anagrafica/pages/attestato_formazione_stampa.html")
        self.assertContains(resp, "Bianchi Anna")
        self.assertContains(resp, "Versione a colori")

    def test_default_usa_variante_a_colori(self):
        rec = self._make_record()
        resp = self.client.get(reverse("anagrafica:attestato_formazione", args=[rec.pk]))
        self.assertTemplateUsed(resp, "anagrafica/pages/attestato_formazione.html")


# ---------------------------------------------------------------------------
# H5c — registro presenze lezione (foglio firme stampabile)
# ---------------------------------------------------------------------------

class RegistroPresenzeLezioneTests(TestCase):
    """Foglio firme A4 di una lezione: nominativi iscritti + righe vuote + gate."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="reg_admin", email="reg_admin@x.local", password="x"
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, attivo) "
                "VALUES (501, 'Luca', 'Conti', 1)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def _make_lezione(self):
        from datetime import date, time
        from .models_formazione import (
            TrainingCourse, TrainingEnrollment, TrainingLesson,
            TrainingPlan, TrainingSession,
        )
        piano = TrainingPlan.objects.create(codice="PREG", nome="Piano")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CREG", titolo="Sicurezza base", durata_ore_teorica=4,
        )
        sess = TrainingSession.objects.create(
            corso=corso, codice_sessione="SREG1",
            data_inizio=date.today(), data_fine=date.today(), sede="Aula A",
        )
        lez = TrainingLesson.objects.create(
            sessione=sess, numero=1, data=date.today(),
            ora_inizio=time(9, 0), ora_fine=time(13, 0), argomento="Rischi generali",
        )
        TrainingEnrollment.objects.create(sessione=sess, legacy_anagrafica_id=501)
        return sess, lez

    def test_registro_render_con_iscritti_e_righe_vuote(self):
        sess, lez = self._make_lezione()
        resp = self.client.get(
            reverse("anagrafica:formazione_lezione_registro", args=[sess.pk, lez.pk])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Conti Luca")
        self.assertContains(resp, "Firma ingresso")
        self.assertContains(resp, "Firma uscita")
        self.assertContains(resp, "Rischi generali")

    def test_registro_tracciato_in_export_log(self):
        from .models_formazione import TrainingExportLog
        sess, lez = self._make_lezione()
        self.client.get(reverse("anagrafica:formazione_lezione_registro", args=[sess.pk, lez.pk]))
        log = TrainingExportLog.objects.filter(tipo="REPORT_FIRMA").last()
        self.assertIsNotNone(log)
        self.assertEqual(log.filtri_json.get("lezione_id"), lez.pk)

    def test_registro_negato_senza_permesso(self):
        sess, lez = self._make_lezione()
        with patch("anagrafica.views._can_view_formazione", return_value=False):
            resp = self.client.get(
                reverse("anagrafica:formazione_lezione_registro", args=[sess.pk, lez.pk])
            )
        self.assertEqual(resp.status_code, 302)


# ---------------------------------------------------------------------------
# H5d — archiviazione attestato PDF nel box documenti del dipendente
# ---------------------------------------------------------------------------

class AttestatoArchivioTests(TestCase):
    """PDF attestato + archiviazione nel box: auto-save a fine corso, idempotenza,
    salvataggio manuale, backfill/purge/export, cartella predefinita, gate."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="arch_admin", email="arch_admin@x.local", password="x"
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, attivo) "
                "VALUES (601, 'Marco', 'Rossi', 1)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def _make_record(self):
        from django.utils import timezone
        from .models_formazione import TrainingCourse, TrainingEmployeeRecord, TrainingPlan
        piano = TrainingPlan.objects.create(codice="PARC", nome="Piano archivio")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CARC", titolo="Corso archivio", durata_ore_teorica=6,
        )
        return TrainingEmployeeRecord.objects.create(
            corso=corso, legacy_anagrafica_id=601,
            data_completamento=timezone.localdate(),
            ore_frequentate=6, idoneo=True,
            course_title_snapshot="Corso archivio",
            teacher_name_snapshot="Ing. Bianchi",
        )

    def _make_enrollment(self):
        from datetime import date
        from .models_formazione import (
            TrainingCourse, TrainingEnrollment, TrainingPlan, TrainingSession,
        )
        piano = TrainingPlan.objects.create(codice="PAUT", nome="Piano auto")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CAUT", titolo="Corso auto", durata_ore_teorica=4,
        )
        sess = TrainingSession.objects.create(
            corso=corso, codice_sessione="SAUT",
            data_inizio=date.today(), data_fine=date.today(),
        )
        return TrainingEnrollment.objects.create(
            sessione=sess, legacy_anagrafica_id=601, ore_frequentate=4, idoneo=True,
        )

    def test_build_pdf_bytes(self):
        from .services.attestato_pdf import build_attestato_pdf_bytes
        rec = self._make_record()
        pdf = build_attestato_pdf_bytes(rec)
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 800)

    def test_archivia_crea_documento_idempotente_e_cartella(self):
        from .services.attestato_pdf import RIFERIMENTO_TIPO, archivia_attestato
        rec = self._make_record()
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            doc1 = archivia_attestato(rec, user=self.admin)
            doc2 = archivia_attestato(rec, user=self.admin)  # idempotente: stesso doc
        self.assertEqual(doc1.pk, doc2.pk)
        self.assertEqual(doc1.tipo, DocumentoDipendente.Tipo.CERTIFICATO_FORMAZIONE)
        self.assertEqual(doc1.cartella.nome, "Attestati formazione")
        self.assertEqual(
            DocumentoDipendente.objects.filter(
                oggetto_riferimento_tipo=RIFERIMENTO_TIPO, oggetto_riferimento_id=rec.pk
            ).count(),
            1,
        )

    def test_auto_save_a_fine_corso(self):
        from .models_formazione import AttestatoFormazioneConfig
        from .views import _crea_employee_record
        cfg = AttestatoFormazioneConfig.get_instance()
        cfg.auto_salva_attestato = True
        cfg.save()
        enr = self._make_enrollment()
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            rec = _crea_employee_record(enr, self.admin)
            self.assertIsNotNone(rec)
            self.assertEqual(
                DocumentoDipendente.objects.filter(
                    tipo=DocumentoDipendente.Tipo.CERTIFICATO_FORMAZIONE,
                    oggetto_riferimento_id=rec.pk,
                ).count(),
                1,
            )

    def test_no_auto_save_se_disattivato(self):
        from .views import _crea_employee_record
        # default cfg.auto_salva_attestato = False
        enr = self._make_enrollment()
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            rec = _crea_employee_record(enr, self.admin)
        self.assertEqual(
            DocumentoDipendente.objects.filter(oggetto_riferimento_id=rec.pk).count(), 0
        )

    def test_salva_box_manuale_view(self):
        rec = self._make_record()
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            resp = self.client.post(reverse("anagrafica:attestato_salva_box", args=[rec.pk]))
            self.assertEqual(resp.status_code, 302)
            self.assertEqual(
                DocumentoDipendente.objects.filter(oggetto_riferimento_id=rec.pk).count(), 1
            )

    def test_salva_box_negato_senza_permesso(self):
        rec = self._make_record()
        with patch("anagrafica.views._can_edit_formazione", return_value=False):
            resp = self.client.post(reverse("anagrafica:attestato_salva_box", args=[rec.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(
            DocumentoDipendente.objects.filter(oggetto_riferimento_id=rec.pk).count(), 0
        )

    def test_backfill_e_purge(self):
        rec = self._make_record()
        url = reverse("anagrafica:attestato_impostazioni")
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            self.client.post(url, {"action": "backfill"})
            self.assertEqual(
                DocumentoDipendente.objects.filter(oggetto_riferimento_id=rec.pk).count(), 1
            )
            self.client.post(url, {"action": "purge"})
            self.assertEqual(
                DocumentoDipendente.objects.filter(oggetto_riferimento_id=rec.pk).count(), 0
            )

    def test_export_csv(self):
        from .services.attestato_pdf import archivia_attestato
        rec = self._make_record()
        with tempfile.TemporaryDirectory() as root, override_settings(ANAGRAFICA_PRIVATE_ROOT=root):
            archivia_attestato(rec, user=self.admin)
            resp = self.client.get(reverse("anagrafica:attestato_report_export"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])
        self.assertIn(b"Rossi Marco", resp.content)


# ---------------------------------------------------------------------------
# H6 — organigramma visuale
# ---------------------------------------------------------------------------

class OrganigrammaTests(TestCase):
    """Albero area → reparto → capo → membri, con bucket disallineamenti."""

    @classmethod
    def setUpTestData(cls):
        from .models import AreaAziendale, Reparto
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="org_admin", email="org_admin@x.local", password="x"
        )
        cls.area = AreaAziendale.objects.create(nome="Produzione", colore="#1d4ed8")
        cls.rep_prod = Reparto.objects.create(
            nome="PROD", area_aziendale=cls.area, caporeparto_legacy_id=401
        )
        cls.rep_orfano = Reparto.objects.create(nome="MAG")  # senza area, senza capo
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) VALUES "
                "(401, 'Capo', 'Reparto', 'Caporeparto', 'PROD', 1), "
                "(402, 'Mario', 'Verdi', 'Operatore', 'PROD', 1), "
                "(403, 'Luigi', 'Bianchi', 'Operatore', 'INESISTENTE', 1), "
                "(404, 'Anna', 'Cessata', 'Operatore', 'PROD', 0)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def test_albero_aree_reparti_capo(self):
        resp = self.client.get(reverse("anagrafica:organigramma"))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertIn("Produzione", content)
        self.assertIn("PROD", content)
        self.assertIn("Reparto Capo", content)          # capo evidenziato
        self.assertIn("Verdi Mario", content)            # membro (capo escluso dai membri)
        self.assertIn("Senza area", content)              # reparto MAG senza area
        self.assertIn("Caporeparto non assegnato", content)

    def test_non_mappati_visibili(self):
        resp = self.client.get(reverse("anagrafica:organigramma"))
        content = resp.content.decode()
        self.assertIn("Bianchi Luigi", content)
        self.assertIn("non a catalogo", content)

    def test_cessati_esclusi(self):
        resp = self.client.get(reverse("anagrafica:organigramma"))
        self.assertNotContains(resp, "Cessata")

    def test_filtro_area(self):
        resp = self.client.get(reverse("anagrafica:organigramma"), {"area": "Produzione"})
        content = resp.content.decode()
        self.assertIn("PROD", content)
        # il gruppo "Senza area" sparisce col filtro
        self.assertNotIn("Senza area</span>", content)


# ---------------------------------------------------------------------------
# H2 — fascicolo conformità "idoneità alla mansione"
# ---------------------------------------------------------------------------

class ConformitaServiceTests(TestCase):
    """Semaforo aggregato: esiti per dominio e privacy dettaglio visite."""

    @classmethod
    def setUpTestData(cls):
        from .models import TipoQualifica
        _ensure_anagrafica_table()
        cls.tipo_q = TipoQualifica.objects.create(nome="Patentino muletto", categoria="PROFESSIONALE")

    def _qualifica(self, legacy_id: int, giorni):
        from datetime import timedelta
        from django.utils import timezone
        from .models import DipendenteQualifica
        scad = None if giorni is None else timezone.localdate() + timedelta(days=giorni)
        return DipendenteQualifica.objects.create(
            legacy_anagrafica_id=legacy_id, tipo=self.tipo_q, data_scadenza=scad
        )

    def test_qualifica_valida_ok(self):
        from .services import conformita
        self._qualifica(601, 365)
        stato = conformita.stato_conformita(601)
        self.assertEqual(stato["qualifiche"]["esito"], conformita.ESITO_OK)
        self.assertEqual(stato["complessivo"], conformita.ESITO_OK)

    def test_qualifica_scaduta_ko(self):
        from .services import conformita
        self._qualifica(602, -5)
        stato = conformita.stato_conformita(602)
        self.assertEqual(stato["qualifiche"]["esito"], conformita.ESITO_KO)
        self.assertEqual(stato["complessivo"], conformita.ESITO_KO)

    def test_qualifica_in_scadenza_warn(self):
        from .services import conformita
        self._qualifica(603, 30)
        stato = conformita.stato_conformita(603)
        self.assertEqual(stato["qualifiche"]["esito"], conformita.ESITO_WARN)

    def test_nessun_requisito_na(self):
        from .services import conformita
        stato = conformita.stato_conformita(699)
        self.assertEqual(stato["complessivo"], conformita.ESITO_NA)

    def test_complessivo_e_il_peggiore(self):
        from .services import conformita
        # qualifica valida (OK) ma visita obbligatoria realmente scaduta (KO) →
        # complessivo KO (il peggiore fra i domini applicabili).
        self._qualifica(604, 365)
        self._visita_obbligatoria_scaduta(604)
        stato = conformita.stato_conformita(604)
        self.assertEqual(stato["qualifiche"]["esito"], conformita.ESITO_OK)
        self.assertEqual(stato["visite"]["esito"], conformita.ESITO_KO)
        self.assertEqual(stato["complessivo"], conformita.ESITO_KO)

    def test_visita_mancante_e_neutra(self):
        from .services import conformita
        # Dipendente già in forza: visita obbligatoria mai registrata = dato
        # mancante → neutro (NA), NON "scaduto"/KO; non trascina il complessivo.
        self._visita_obbligatoria_mancante(606)
        stato = conformita.stato_conformita(606)
        self.assertEqual(stato["visite"]["esito"], conformita.ESITO_NA)
        self.assertEqual(stato["complessivo"], conformita.ESITO_NA)

    def _visita_obbligatoria_mancante(self, legacy_id: int):
        from .models import (
            DipendenteRuoloOperativo, RuoloOperativo, TipoVisitaMedica,
        )
        ruolo = RuoloOperativo.objects.create(nome=f"Saldatore {legacy_id}")
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=legacy_id, ruolo=ruolo)
        tipo = TipoVisitaMedica.objects.create(
            nome=f"Sorveglianza fumi saldatura {legacy_id}", obbligatoria=True, is_active=True
        )
        tipo.ruoli_operativi.add(ruolo)

    def _visita_obbligatoria_scaduta(self, legacy_id: int):
        """Come ``_visita_obbligatoria_mancante`` ma con una visita registrata e
        scaduta (data_svolgimento ~2 anni fa, durata 12 mesi → scaduta)."""
        from datetime import timedelta
        from django.utils import timezone
        from .models import (
            DipendenteRuoloOperativo, RuoloOperativo, TipoVisitaMedica, VisitaMedica,
        )
        ruolo = RuoloOperativo.objects.create(nome=f"Saldatore {legacy_id}")
        DipendenteRuoloOperativo.objects.create(legacy_anagrafica_id=legacy_id, ruolo=ruolo)
        tipo = TipoVisitaMedica.objects.create(
            nome=f"Sorveglianza fumi saldatura {legacy_id}", obbligatoria=True,
            is_active=True, durata_mesi=12,
        )
        tipo.ruoli_operativi.add(ruolo)
        VisitaMedica.objects.create(
            legacy_anagrafica_id=legacy_id, tipo=tipo,
            data_svolgimento=timezone.localdate() - timedelta(days=730),
        )

    def test_dettaglio_visite_gated_da_privacy(self):
        from .services import conformita
        self._visita_obbligatoria_scaduta(605)
        # con dettaglio: appare il nome della tipologia
        con = conformita.stato_conformita(605, include_visite_dettaglio=True)
        self.assertTrue(any("Sorveglianza fumi" in d for d in con["visite"]["dettagli"]))
        # senza dettaglio: etichetta generica, nessun nome tipologia
        senza = conformita.stato_conformita(605, include_visite_dettaglio=False)
        self.assertEqual(senza["visite"]["esito"], conformita.ESITO_KO)
        self.assertFalse(any("Sorveglianza fumi" in d for d in senza["visite"]["dettagli"]))
        self.assertTrue(any("Visita richiesta" in d for d in senza["visite"]["dettagli"]))


class SicurezzaHubTests(TestCase):
    """Smoke test del cruscotto Sicurezza & Idoneità (rendering + KPI)."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="sh_admin", email="sh_admin@x.local", password="x"
        )

    def test_hub_render_ok(self):
        from .models import Mansione
        m = Mansione.objects.create(nome="Saldatore", livello_rischio=Mansione.RISCHIO_ALTO)
        from dpi.models import CategoriaDPI
        m.dpi_richiesti.add(CategoriaDPI.objects.create(nome="Guanti", is_active=True))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:sicurezza_hub"))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("Sicurezza", body)
        self.assertIn("Mansioni di rischio", body)
        # la mansione con requisiti conta come "di rischio"
        self.assertEqual(resp.context["n_mansioni_rischio"], 1)

    def test_wizard_render_ok(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:sicurezza_wizard"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Configurazione guidata", resp.content.decode())
        self.assertEqual(len(resp.context["steps"]), 3)

    def test_mansioni_list_mostra_requisiti_e_contatori_dpi_visite(self):
        """La lista mansioni espone il pulsante Requisiti e i contatori DPI/visite,
        così DPI e visite sono visibili e raggiungibili dal "setup" mansione."""
        from .models import Mansione, TipoVisitaMedica
        from dpi.models import CategoriaDPI
        m = Mansione.objects.create(nome="Saldatore", livello_rischio=Mansione.RISCHIO_ALTO)
        m.dpi_richiesti.add(CategoriaDPI.objects.create(nome="Guanti", is_active=True))
        m.visite_richieste.add(TipoVisitaMedica.objects.create(nome="Visita audiometrica"))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:mansioni_list"))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        # pulsante requisiti presente per la mansione
        self.assertIn(reverse("anagrafica:mansione_requisiti", args=[m.id]), body)
        # contatori visibili sulla card
        self.assertIn("1 DPI", body)
        self.assertIn("1 visite", body)
        # i contatori sono calcolati sulla view
        mansione_ctx = next(x for x in resp.context["mansioni"] if x.id == m.id)
        self.assertEqual(mansione_ctx.n_dpi, 1)
        self.assertEqual(mansione_ctx.n_visite, 1)

    def test_mansioni_list_filtri_rischio(self):
        """Filtri lista mansioni: livello di rischio e 'solo mansioni di rischio'."""
        from .models import Mansione
        Mansione.objects.create(nome="Verniciatore X", livello_rischio=Mansione.RISCHIO_ALTO)
        Mansione.objects.create(nome="Impiegato XYZ")  # nessun requisito → non di rischio
        self.client.force_login(self.admin)
        r1 = self.client.get(reverse("anagrafica:mansioni_list") + "?rischio=A")
        self.assertContains(r1, "Verniciatore X")
        self.assertNotContains(r1, "Impiegato XYZ")
        r2 = self.client.get(reverse("anagrafica:mansioni_list") + "?solo_rischio=1")
        self.assertContains(r2, "Verniciatore X")
        self.assertNotContains(r2, "Impiegato XYZ")


class QualificheCatalogoTests(TestCase):
    """Catalogo qualifiche unico con viste filtrate per categoria + navigazione."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="ql_admin", email="ql_admin@x.local", password="x"
        )

    def setUp(self):
        from .models import TipoQualifica
        TipoQualifica.objects.create(nome="Carrellista ASR", categoria=TipoQualifica.CAT_SICUREZZA)
        TipoQualifica.objects.create(nome="Saldatore certificato", categoria=TipoQualifica.CAT_PROFESSIONALE)

    def test_catalogo_completo_mostra_tutte_le_categorie_e_tab(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:qualifiche_list"))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("Carrellista ASR", body)
        self.assertIn("Saldatore certificato", body)
        self.assertEqual(resp.context["active_categoria"], "")
        self.assertEqual(len(resp.context["tabs"]), 5)  # Tutte + 4 categorie

    def test_vista_filtrata_sicurezza_esclude_altre_categorie(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:qualifiche_list") + "?categoria=SICUREZZA")
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("Carrellista ASR", body)
        self.assertNotIn("Saldatore certificato", body)
        self.assertEqual(resp.context["active_categoria"], "SICUREZZA")
        self.assertTrue(resp.context["is_safety_view"])
        self.assertEqual(len(resp.context["tipi_grouped"]), 1)

    def test_categoria_non_valida_ignorata(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:qualifiche_list") + "?categoria=PIPPO")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["active_categoria"], "")

    def test_catalogo_mostra_corso_collegato(self):
        """Step 3: il catalogo qualifiche mostra il corso collegato (qualifica àncora)."""
        from decimal import Decimal
        from .models import TipoQualifica
        from .models_formazione import TrainingPlan, TrainingCourse
        tipo = TipoQualifica.objects.get(nome="Carrellista ASR")
        plan = TrainingPlan.objects.create(nome="Piano cat", codice="PCAT")
        TrainingCourse.objects.create(
            piano=plan, codice="CC1", titolo="Corso Carrellisti",
            durata_ore_teorica=Decimal("0"), qualifica=tipo,
        )
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:qualifiche_list"))
        self.assertContains(resp, "Corso Carrellisti")

    def test_dettaglio_qualifica_render(self):
        """Step 3: dal catalogo si apre il dettaglio della singola qualifica
        con sezioni dipendenti/corsi/sessioni/attestati."""
        from .models import TipoQualifica
        tipo = TipoQualifica.objects.get(nome="Carrellista ASR")
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:tipo_qualifica_detail", args=[tipo.id]))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("Carrellista ASR", body)
        self.assertIn("Dipendenti che la possiedono", body)
        self.assertIn("Corsi collegati", body)

    def test_subnav_binding_qualifiche_su_formazione_non_impostazioni(self):
        """Dopo la 0045 l'highlight top-nav delle qualifiche è su Formazione e
        non più su Impostazioni (niente doppio-highlight, come per la 0044)."""
        from .models import SubnavLinkAnagrafica
        form = SubnavLinkAnagrafica.objects.filter(url_value="anagrafica:formazione_dashboard").first()
        imp = SubnavLinkAnagrafica.objects.filter(url_value="anagrafica:impostazioni").first()
        if form:
            self.assertIn("anagrafica:qualifiche_list", form.active_view_names)
        if imp:
            self.assertNotIn("anagrafica:qualifiche_list", imp.active_view_names)


class QualificaSessioneTests(TestCase):
    """Rinnovi qualifiche: singolo (update-or-create) + sessioni persistenti."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="qsess_admin", email="qsess@x.local", password="x"
        )

    def setUp(self):
        from .models import TipoQualifica
        self.tipo = TipoQualifica.objects.create(
            nome="Carrellista test", categoria=TipoQualifica.CAT_SICUREZZA, durata_mesi=60,
        )

    def test_pagine_sessioni_render_ok(self):
        from .models import QualificaSessione
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("anagrafica:qualifica_sessioni_list")).status_code, 200)
        # Pagina unica di creazione (GET) + deep-link ?tipo= con candidati inline
        self.assertEqual(self.client.get(reverse("anagrafica:qualifica_sessione_create")).status_code, 200)
        r_pre = self.client.get(reverse("anagrafica:qualifica_sessione_create") + f"?tipo={self.tipo.id}")
        self.assertEqual(r_pre.status_code, 200)
        self.assertEqual(r_pre.context["pre_tipo"].id, self.tipo.id)
        # Partial HTMX dei candidati per il tipo selezionato
        r_cand = self.client.get(reverse("anagrafica:qualifica_sessione_candidati") + f"?tipo={self.tipo.id}")
        self.assertEqual(r_cand.status_code, 200)
        self.assertContains(r_cand, self.tipo.nome)
        # dettaglio
        sess = QualificaSessione.objects.create(tipo=self.tipo, data_conseguimento=date(2026, 6, 14))
        self.assertEqual(
            self.client.get(reverse("anagrafica:qualifica_sessione_detail", args=[sess.id])).status_code, 200
        )

    def test_single_add_aggiorna_non_duplica(self):
        from .models import DipendenteQualifica
        self.client.force_login(self.admin)
        url = reverse("anagrafica:dipendente_qualifica_add", args=[5001])
        self.client.post(url, {"tipo_id": self.tipo.id, "data_conseguimento": "2024-01-10"})
        self.client.post(url, {"tipo_id": self.tipo.id, "data_conseguimento": "2026-01-10"})
        qs = DipendenteQualifica.objects.filter(legacy_anagrafica_id=5001, tipo=self.tipo)
        self.assertEqual(qs.count(), 1)  # rinnovo = aggiornamento, non duplicato
        q = qs.first()
        self.assertEqual(q.data_conseguimento, date(2026, 1, 10))
        self.assertEqual(q.data_scadenza, date(2031, 1, 10))  # +60 mesi (ASR)

    def test_sessione_create_upsert_e_scadenza_auto(self):
        from .models import DipendenteQualifica, QualificaSessione
        self.client.force_login(self.admin)
        resp = self.client.post(reverse("anagrafica:qualifica_sessione_create"), {
            "step": "2", "tipo_id": self.tipo.id, "data_conseguimento": "2026-06-14",
            "ente": "Ente X", "dipendenti_selezionati": ["5002", "5003"],
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(QualificaSessione.objects.count(), 1)
        sess = QualificaSessione.objects.first()
        self.assertEqual(DipendenteQualifica.objects.filter(sessione=sess).count(), 2)
        q = DipendenteQualifica.objects.get(legacy_anagrafica_id=5002, tipo=self.tipo)
        self.assertEqual(q.data_scadenza, date(2031, 6, 14))  # scadenza calcolata dal tipo

    def test_remove_partecipante_stacca_ma_conserva_qualifica(self):
        from .models import DipendenteQualifica, QualificaSessione
        self.client.force_login(self.admin)
        self.client.post(reverse("anagrafica:qualifica_sessione_create"), {
            "step": "2", "tipo_id": self.tipo.id, "data_conseguimento": "2026-06-14",
            "dipendenti_selezionati": ["5004"],
        })
        sess = QualificaSessione.objects.first()
        q = DipendenteQualifica.objects.get(legacy_anagrafica_id=5004)
        self.client.post(reverse("anagrafica:qualifica_sessione_partecipante_remove", args=[sess.id, q.id]))
        q.refresh_from_db()
        self.assertIsNone(q.sessione_id)
        self.assertTrue(DipendenteQualifica.objects.filter(pk=q.pk).exists())

    def test_delete_sessione_conserva_qualifiche(self):
        from .models import DipendenteQualifica, QualificaSessione
        self.client.force_login(self.admin)
        self.client.post(reverse("anagrafica:qualifica_sessione_create"), {
            "step": "2", "tipo_id": self.tipo.id, "data_conseguimento": "2026-06-14",
            "dipendenti_selezionati": ["5005"],
        })
        sess = QualificaSessione.objects.first()
        self.client.post(reverse("anagrafica:qualifica_sessione_delete", args=[sess.id]))
        self.assertEqual(QualificaSessione.objects.count(), 0)
        q = DipendenteQualifica.objects.filter(legacy_anagrafica_id=5005).first()
        self.assertIsNotNone(q)            # qualifica conservata (SET_NULL)
        self.assertIsNone(q.sessione_id)


class LinkQualificheCorsiTests(TestCase):
    """Back-fill dei legami qualifica↔corso↔completamento (competency management)."""

    def test_backfill_collega_corso_e_record(self):
        from decimal import Decimal
        from django.core.management import call_command
        from .models import TipoQualifica, DipendenteQualifica
        from .models_formazione import TrainingPlan, TrainingCourse, TrainingEmployeeRecord

        plan = TrainingPlan.objects.create(nome="Piano test", codice="PTEST")
        tipo = TipoQualifica.objects.create(
            nome="Carrellista X", categoria=TipoQualifica.CAT_SICUREZZA, durata_mesi=60,
        )
        corso = TrainingCourse.objects.create(
            piano=plan, codice="CTEST", titolo="Carrellista X",
            durata_ore_teorica=Decimal("0"), validita_mesi=60, stato="ATTIVO",
        )
        rec = TrainingEmployeeRecord.objects.create(
            corso=corso, legacy_anagrafica_id=6001, data_completamento=date(2024, 1, 10),
        )
        dq = DipendenteQualifica.objects.create(
            legacy_anagrafica_id=6001, tipo=tipo, data_conseguimento=date(2024, 1, 10),
        )
        # corso che non corrisponde ad alcuna qualifica → deve restare scollegato
        altro = TrainingCourse.objects.create(
            piano=plan, codice="CX", titolo="Excel base", durata_ore_teorica=Decimal("0"),
        )

        call_command("link_qualifiche_corsi", "--commit", verbosity=0)

        corso.refresh_from_db(); dq.refresh_from_db(); altro.refresh_from_db()
        self.assertEqual(corso.qualifica_id, tipo.id)
        self.assertEqual(dq.record_formazione_id, rec.id)
        self.assertIsNone(altro.qualifica_id)  # nessun falso legame

    def test_backfill_dry_run_non_scrive(self):
        from decimal import Decimal
        from django.core.management import call_command
        from .models import TipoQualifica
        from .models_formazione import TrainingPlan, TrainingCourse

        plan = TrainingPlan.objects.create(nome="Piano test2", codice="PT2")
        TipoQualifica.objects.create(nome="BLSD test", categoria=TipoQualifica.CAT_SICUREZZA)
        corso = TrainingCourse.objects.create(
            piano=plan, codice="CB", titolo="BLSD test", durata_ore_teorica=Decimal("0"),
        )
        call_command("link_qualifiche_corsi", verbosity=0)  # dry-run
        corso.refresh_from_db()
        self.assertIsNone(corso.qualifica_id)


class MatriceCompetenzeTests(TestCase):
    """Matrice competenze: tab per categoria filtrano le colonne."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="mc_admin", email="mc@x.local", password="x"
        )

    def setUp(self):
        from .models import TipoQualifica, DipendenteQualifica
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (7001, 'Mario', 'Rossi', 'Saldatore', 1)"
            )
        tsic = TipoQualifica.objects.create(nome="Carrellista MC", categoria=TipoQualifica.CAT_SICUREZZA)
        tprof = TipoQualifica.objects.create(nome="Saldatore MC", categoria=TipoQualifica.CAT_PROFESSIONALE)
        DipendenteQualifica.objects.create(legacy_anagrafica_id=7001, tipo=tsic)
        DipendenteQualifica.objects.create(legacy_anagrafica_id=7001, tipo=tprof)

    def test_matrice_tutte_le_colonne(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:matrice_competenze"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["tipi"]), 2)
        self.assertEqual(resp.context["active_categoria"], "")

    def test_matrice_tab_categoria_filtra_colonne(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:matrice_competenze") + "?categoria=SICUREZZA")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["active_categoria"], "SICUREZZA")
        self.assertEqual(len(resp.context["tipi"]), 1)
        self.assertEqual(resp.context["tipi"][0].nome, "Carrellista MC")


class VerbaleDpiTests(TestCase):
    """Verbale consegna DPI (MOD.155) precompilato dai requisiti della mansione."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="vd_admin", email="vd@x.local", password="x"
        )

    def test_verbale_render_con_dpi_mansione(self):
        from .models import Mansione
        from dpi.models import CategoriaDPI
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (8001, 'Luca', 'Verdi', 'Verniciatore VD', 1)"
            )
        m = Mansione.objects.create(nome="Verniciatore VD")
        m.dpi_richiesti.add(CategoriaDPI.objects.create(nome="Maschera vapori", is_active=True))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:dipendente_verbale_dpi", args=[8001]))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("MOD.155", body)
        self.assertIn("Maschera vapori", body)
        self.assertIn("Taglia / Modello", body)


class SafetyIntegrationsTests(TestCase):
    """A/B/D/E: digest idoneità, verbale DPI, matrice competenze, gap al cambio mansione."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="si2_admin", email="si2_admin@x.local", password="x"
        )

    def _ins(self, lid, mansione="", reparto="", attivo=1):
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) "
                "VALUES (%s, 'Mario', 'Rossi', %s, %s, %s)", [lid, mansione, reparto, attivo],
            )

    def _cat_dpi(self, nome):
        from dpi.models import CategoriaDPI
        return CategoriaDPI.objects.create(nome=nome, is_active=True)

    # B — verbale DPI
    def test_verbale_dpi_render(self):
        from .models import Mansione
        self._ins(951, mansione="Saldatore")
        m = Mansione.objects.create(nome="Saldatore")
        m.dpi_richiesti.add(self._cat_dpi("Maschera saldatura"))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:dipendente_verbale_dpi", args=[951]))
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("VERBALE DI CONSEGNA", body)
        self.assertIn("Maschera saldatura", body)

    # D — matrice competenze
    def test_matrice_competenze_render_e_csv(self):
        from .models import DipendenteQualifica, TipoQualifica
        self._ins(952, reparto="PROD")
        tipo = TipoQualifica.objects.create(nome="Patentino muletto", categoria="PROFESSIONALE")
        DipendenteQualifica.objects.create(legacy_anagrafica_id=952, tipo=tipo)
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("anagrafica:matrice_competenze"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("Patentino muletto", resp.content.decode())
        csv_resp = self.client.get(reverse("anagrafica:matrice_competenze"), {"format": "csv"})
        self.assertEqual(csv_resp.status_code, 200)
        self.assertIn("text/csv", csv_resp["Content-Type"])

    # A — digest idoneità
    def test_idoneita_digest(self):
        from django.core import mail
        from django.core.management import call_command
        from core.models import SiteConfig
        from .models import Mansione
        SiteConfig.set("idoneita_reminder_emails", "rspp@x.local")
        self._ins(953, mansione="Saldatore")
        m = Mansione.objects.create(nome="Saldatore")
        m.dpi_richiesti.add(self._cat_dpi("Guanti"))  # mancante → idoneo con riserve
        mail.outbox = []
        call_command("send_idoneita_digest", verbosity=0)
        destinatari = [a for msg in mail.outbox for a in msg.to]
        self.assertIn("rspp@x.local", destinatari)

    # E — gap idoneità al cambio mansione
    def test_notifica_gap_cambio_mansione(self):
        from django.core import mail
        from core.models import SiteConfig
        from . import views
        from .models import Mansione
        SiteConfig.set("idoneita_reminder_emails", "rspp@x.local")
        m = Mansione.objects.create(nome="Carrellista")
        m.dpi_richiesti.add(self._cat_dpi("Gilet AV"))
        mail.outbox = []
        views._notifica_gap_idoneita(954, {"cognome": "Verdi", "nome": "Ada", "reparto": ""}, "Carrellista")
        destinatari = [a for msg in mail.outbox for a in msg.to]
        self.assertIn("rspp@x.local", destinatari)


class MansionarioIdoneitaTests(TestCase):
    """Resolver requisiti "mansione di rischio" (diretti + ereditati) e idoneità."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()

    def _cat_dpi(self, nome, vita=None):
        from dpi.models import CategoriaDPI
        return CategoriaDPI.objects.create(nome=nome, is_active=True, vita_utile_giorni=vita)

    def _tipo_visita(self, nome, durata=12):
        from .models import TipoVisitaMedica
        return TipoVisitaMedica.objects.create(
            nome=nome, obbligatoria=True, is_active=True, durata_mesi=durata
        )

    def test_resolver_unione_diretti_ed_ereditati(self):
        from .models import Mansione
        from .models_rischi import EsposizioneRischio, FattoreRischio
        from .services import mansionario
        m = Mansione.objects.create(nome="Saldatore")
        m.dpi_richiesti.add(self._cat_dpi("Maschera saldatura"))
        m.visite_richieste.add(self._tipo_visita("Visita base"))
        fattore = FattoreRischio.objects.create(codice="R-FUMI", nome="Fumi saldatura")
        fattore.categorie_dpi.add(self._cat_dpi("Filtrante FFP3"))
        fattore.tipi_visita.add(self._tipo_visita("Spirometria"))
        EsposizioneRischio.objects.create(fattore=fattore, mansione=m)

        req = mansionario.requisiti_mansione(m)
        self.assertEqual({c.nome for c in req["dpi"]}, {"Maschera saldatura", "Filtrante FFP3"})
        self.assertEqual({v.nome for v in req["visite"]}, {"Visita base", "Spirometria"})
        self.assertIn(fattore, req["fattori"])

    def test_resolver_esposizione_disattivata_non_eredita(self):
        from .models import Mansione
        from .models_rischi import EsposizioneRischio, FattoreRischio
        from .services import mansionario
        m = Mansione.objects.create(nome="Magazziniere")
        fattore = FattoreRischio.objects.create(codice="R-MMC", nome="Movimentazione carichi")
        fattore.tipi_visita.add(self._tipo_visita("Visita MMC"))
        EsposizioneRischio.objects.create(fattore=fattore, mansione=m, is_active=False)
        req = mansionario.requisiti_mansione(m)
        self.assertEqual(req["visite"], [])

    def test_resolver_match_per_nome_case_insensitive(self):
        from .models import Mansione
        from .services import mansionario
        m = Mansione.objects.create(nome="Carrellista")
        m.visite_richieste.add(self._tipo_visita("Visita carrellista"))
        req = mansionario.requisiti_per_nome_mansione("  carrellista ")
        self.assertEqual({v.nome for v in req["visite"]}, {"Visita carrellista"})
        self.assertEqual(mansionario.requisiti_per_nome_mansione("inesistente")["visite"], [])

    def test_idoneita_dpi_mancante_warn(self):
        from .models import Mansione
        from .services import conformita
        m = Mansione.objects.create(nome="Operaio A")
        m.dpi_richiesti.add(self._cat_dpi("Guanti antitaglio"))
        idn = conformita.stato_conformita(701, mansione="Operaio A")["idoneita"]
        self.assertEqual(idn["esito"], conformita.ESITO_WARN)
        self.assertTrue(any("Guanti antitaglio" in d for d in idn["mancanti"]))

    def test_idoneita_visita_scaduta_ko(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import Mansione, VisitaMedica
        from .services import conformita
        m = Mansione.objects.create(nome="Operaio B")
        tipo = self._tipo_visita("Sorveglianza rumore")
        m.visite_richieste.add(tipo)
        VisitaMedica.objects.create(
            legacy_anagrafica_id=702, tipo=tipo,
            data_svolgimento=timezone.localdate() - timedelta(days=730),
        )
        idn = conformita.stato_conformita(
            702, mansione="Operaio B", include_visite_dettaglio=True
        )["idoneita"]
        self.assertEqual(idn["esito"], conformita.ESITO_KO)
        self.assertTrue(any("Sorveglianza rumore" in d for d in idn["scaduti"]))

    def test_idoneita_completa_ok(self):
        from datetime import timedelta
        from django.utils import timezone
        from dpi.models import ConsegnaDPI, RichiestaDPI, StatoRichiesta
        from .models import Mansione, VisitaMedica
        from .services import conformita
        m = Mansione.objects.create(nome="Operaio C")
        cat = self._cat_dpi("Occhiali", vita=365)
        tipo = self._tipo_visita("Visita C")
        m.dpi_richiesti.add(cat)
        m.visite_richieste.add(tipo)
        rich = RichiestaDPI.objects.create(
            categoria=cat, quantita=1, stato=StatoRichiesta.CONSEGNATA,
            richiedente_legacy_id=703, richiedente_nome="X",
        )
        ConsegnaDPI.objects.create(
            richiesta=rich, data_consegna=timezone.localdate(),
            data_scadenza_stimata=timezone.localdate() + timedelta(days=365),
        )
        VisitaMedica.objects.create(
            legacy_anagrafica_id=703, tipo=tipo, data_svolgimento=timezone.localdate(),
        )
        self.assertEqual(
            conformita.stato_conformita(703, mansione="Operaio C")["idoneita"]["esito"],
            conformita.ESITO_OK,
        )

    def test_idoneita_senza_mansione_na(self):
        from .services import conformita
        self.assertEqual(conformita.stato_conformita(704)["idoneita"]["esito"], conformita.ESITO_NA)

    def test_idoneita_privacy_visite(self):
        from .models import Mansione
        from .services import conformita
        m = Mansione.objects.create(nome="Operaio D")
        m.visite_richieste.add(self._tipo_visita("Sorveglianza speciale"))
        senza = conformita.stato_conformita(
            705, mansione="Operaio D", include_visite_dettaglio=False
        )["idoneita"]
        self.assertTrue(any("Visita richiesta" in d for d in senza["mancanti"]))
        self.assertFalse(any("Sorveglianza speciale" in d for d in senza["mancanti"]))
        con = conformita.stato_conformita(
            705, mansione="Operaio D", include_visite_dettaglio=True
        )["idoneita"]
        self.assertTrue(any("Sorveglianza speciale" in d for d in con["mancanti"]))


class ConformitaPanelTests(TestCase):
    """Pannello HTMX nella scheda dipendente: semaforo + gate dettaglio visite."""

    @classmethod
    def setUpTestData(cls):
        from .models import TipoQualifica
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="cf_admin", email="cf_admin@x.local", password="x"
        )
        cls.tipo_q = TipoQualifica.objects.create(nome="Abilitazione PLE", categoria="PROFESSIONALE")

    def setUp(self):
        self.client.force_login(self.admin)

    def test_panel_mostra_ko_per_qualifica_scaduta(self):
        from datetime import timedelta
        from django.utils import timezone
        from .models import DipendenteQualifica
        DipendenteQualifica.objects.create(
            legacy_anagrafica_id=611, tipo=self.tipo_q,
            data_scadenza=timezone.localdate() - timedelta(days=3),
        )
        resp = self.client.get(reverse("anagrafica:dipendente_conformita_panel", args=[611]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Non conforme")
        self.assertContains(resp, "Abilitazione PLE")

    def test_panel_richiede_login(self):
        self.client.logout()
        resp = self.client.get(reverse("anagrafica:dipendente_conformita_panel", args=[611]))
        self.assertEqual(resp.status_code, 302)


class ConformitaReportTests(TestCase):
    """Elenco conformità: gate HR, filtri, esclusione cessati, export CSV."""

    @classmethod
    def setUpTestData(cls):
        from datetime import timedelta
        from django.utils import timezone
        from .models import DipendenteQualifica, TipoQualifica
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="cr_admin", email="cr_admin@x.local", password="x"
        )
        tipo = TipoQualifica.objects.create(nome="Carrellista", categoria="PROFESSIONALE")
        oggi = timezone.localdate()
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) VALUES "
                "(701, 'Mario', 'Rossi', 'Operatore', 'PROD', 1), "
                "(702, 'Luigi', 'Verdi', 'Operatore', 'PROD', 1), "
                "(703, 'Anna', 'Cessata', 'Operatore', 'PROD', 0)"
            )
        # 701 non conforme (scaduta), 702 in regola (valida)
        DipendenteQualifica.objects.create(
            legacy_anagrafica_id=701, tipo=tipo, data_scadenza=oggi - timedelta(days=2)
        )
        DipendenteQualifica.objects.create(
            legacy_anagrafica_id=702, tipo=tipo, data_scadenza=oggi + timedelta(days=365)
        )

    def setUp(self):
        self.client.force_login(self.admin)

    def test_report_richiede_permesso_hr(self):
        with patch("anagrafica.views._check_hr_permission", return_value=False):
            resp = self.client.get(reverse("anagrafica:conformita_report"))
        self.assertEqual(resp.status_code, 302)

    def test_report_elenca_attivi_ed_esclude_cessati(self):
        resp = self.client.get(reverse("anagrafica:conformita_report"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Rossi Mario")
        self.assertContains(resp, "Verdi Luigi")
        self.assertNotContains(resp, "Cessata")

    def test_report_filtro_esito_ko(self):
        resp = self.client.get(reverse("anagrafica:conformita_report"), {"esito": "ko"})
        content = resp.content.decode()
        self.assertIn("Rossi Mario", content)       # non conforme
        self.assertNotIn("Verdi Luigi", content)    # in regola: escluso dal filtro KO

    def test_report_csv_export(self):
        resp = self.client.get(reverse("anagrafica:conformita_report"), {"format": "csv"})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv; charset=utf-8-sig")
        body = resp.content.decode("utf-8-sig")
        self.assertIn("Conformità", body)
        self.assertIn("Rossi Mario", body)
        self.assertIn("Non conforme", body)


# ---------------------------------------------------------------------------
# H1 — onboarding strutturato (pratica + checklist)
# ---------------------------------------------------------------------------

class OnboardingServiceTests(TestCase):
    """Generazione checklist (base + config + corsi obbligatori) e chiusura."""

    def test_avvia_genera_task_base(self):
        from .models import OnboardingTask
        from .services import onboarding
        pratica = onboarding.avvia_onboarding(
            legacy_id=801, dipendente_nome="Rossi Mario", reparto="PROD", mansione="Operatore"
        )
        codici = set(pratica.tasks.values_list("codice", flat=True))
        for atteso in (
            "it_account_ad", "hr_badge_accessi", "dpi_consegna_iniziale",
            "formazione_corsi_obbligatori", "visita_preassuntiva",
        ):
            self.assertIn(atteso, codici)
        self.assertTrue(
            all(t.stato == OnboardingTask.STATO_DA_FARE for t in pratica.tasks.all())
        )

    def test_task_da_configurazione_workflow(self):
        from .models import OnboardingOffboardingCampo
        from .services import onboarding
        OnboardingOffboardingCampo.objects.create(
            fase=OnboardingOffboardingCampo.FASE_ONBOARDING,
            campo_key="consegna_portatile",
            campo_label="Consegna portatile aziendale",
            categoria=OnboardingOffboardingCampo.CATEGORIA_IT,
            is_active=True,
        )
        pratica = onboarding.avvia_onboarding(legacy_id=802, dipendente_nome="Verdi Anna")
        titoli = list(pratica.tasks.values_list("titolo", flat=True))
        self.assertIn("Verificare Consegna portatile aziendale", titoli)

    def test_corso_obbligatorio_in_descrizione(self):
        from .models import Mansione
        from .models_formazione import TrainingCourse, TrainingPlan, TrainingRequirementRule
        from .services import onboarding
        mans = Mansione.objects.create(nome="Saldatore")
        piano = TrainingPlan.objects.create(codice="PSAL", nome="Piano saldatori")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CSAL", titolo="Sicurezza saldatura", durata_ore_teorica=8
        )
        TrainingRequirementRule.objects.create(
            corso=corso, mansione=mans, is_active=True, is_mandatory=True
        )
        pratica = onboarding.avvia_onboarding(
            legacy_id=803, dipendente_nome="Bianchi Luca", mansione="Saldatore"
        )
        task_form = pratica.tasks.get(codice="formazione_corsi_obbligatori")
        self.assertIn("Sicurezza saldatura", task_form.descrizione)

    def test_chiusura_tutti_completati(self):
        from .models import OnboardingPratica, OnboardingTask
        from .services import onboarding
        pratica = onboarding.avvia_onboarding(legacy_id=804, dipendente_nome="X")
        pratica.tasks.update(stato=OnboardingTask.STATO_COMPLETATO)
        stato = onboarding.chiudi_pratica(pratica)
        self.assertEqual(stato, OnboardingPratica.STATO_CHIUSA)

    def test_chiusura_con_task_aperti_da_eccezioni(self):
        from .models import OnboardingPratica, OnboardingTask
        from .services import onboarding
        pratica = onboarding.avvia_onboarding(legacy_id=805, dipendente_nome="Y")
        # lascia tutti i task DA_FARE
        stato = onboarding.chiudi_pratica(pratica)
        self.assertEqual(stato, OnboardingPratica.STATO_CHIUSA_CON_ECCEZIONI)


class OnboardingMansioneRischioTests(TestCase):
    """Task derivati dalla mansione di rischio, notifiche AMM/CAR, formazione pregressa."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()

    def _cat_dpi(self, nome):
        from dpi.models import CategoriaDPI
        return CategoriaDPI.objects.create(nome=nome, is_active=True)

    def _tipo_visita(self, nome):
        from .models import TipoVisitaMedica
        return TipoVisitaMedica.objects.create(nome=nome, obbligatoria=True, is_active=True)

    def test_task_descrizioni_da_mansione_rischio(self):
        from .models import Mansione
        from .services import onboarding
        m = Mansione.objects.create(nome="Verniciatore")
        m.dpi_richiesti.add(self._cat_dpi("Maschera vapori"))
        m.visite_richieste.add(self._tipo_visita("Sorveglianza chimica"))
        pratica = onboarding.avvia_onboarding(
            legacy_id=820, dipendente_nome="Neri Ugo", mansione="Verniciatore", notifica_dpi=False
        )
        self.assertIn("Maschera vapori", pratica.tasks.get(codice="dpi_consegna_iniziale").descrizione)
        self.assertIn("Sorveglianza chimica", pratica.tasks.get(codice="visita_preassuntiva").descrizione)

    def test_notifica_amm_e_car(self):
        from django.core import mail
        from core.models import SiteConfig
        from .models import Mansione, Reparto
        from .services import onboarding
        SiteConfig.set("dpi_amm_emails", "amm@x.local")
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, attivo, email_notifica) "
                "VALUES (821, 'Capo', 'Reparto', 1, 'capo@x.local')"
            )
        Reparto.objects.create(nome="VERNICIATURA", caporeparto_legacy_id=821)
        m = Mansione.objects.create(nome="Verniciatore2")
        m.dpi_richiesti.add(self._cat_dpi("Guanti chimici"))
        mail.outbox = []
        onboarding.avvia_onboarding(
            legacy_id=822, dipendente_nome="Gialli Tom",
            mansione="Verniciatore2", reparto="VERNICIATURA",
        )
        destinatari = [addr for msg in mail.outbox for addr in msg.to]
        self.assertIn("amm@x.local", destinatari)
        self.assertIn("capo@x.local", destinatari)

    def test_nessuna_notifica_senza_dpi(self):
        from django.core import mail
        from core.models import SiteConfig
        from .models import Mansione
        from .services import onboarding
        SiteConfig.set("dpi_amm_emails", "amm@x.local")
        Mansione.objects.create(nome="Impiegato ufficio")
        mail.outbox = []
        onboarding.avvia_onboarding(
            legacy_id=823, dipendente_nome="Blu Ada", mansione="Impiegato ufficio"
        )
        self.assertEqual(len(mail.outbox), 0)

    def test_registra_formazione_pregressa(self):
        from datetime import date
        from .models_formazione import TrainingCourse, TrainingEmployeeRecord, TrainingPlan
        from .services import onboarding
        piano = TrainingPlan.objects.create(codice="PSEC", nome="Sicurezza")
        corso = TrainingCourse.objects.create(
            piano=piano, codice="CSEC1", titolo="Antincendio", durata_ore_teorica=8, validita_mesi=36
        )
        n = onboarding.registra_formazione_pregressa(
            830, [{"corso_id": corso.pk, "data": date(2025, 1, 10)}]
        )
        self.assertEqual(n, 1)
        rec = TrainingEmployeeRecord.objects.get(legacy_anagrafica_id=830, corso=corso)
        self.assertEqual(rec.course_title_snapshot, "Antincendio")
        self.assertIsNotNone(rec.data_scadenza)
        # idempotente: non duplica
        self.assertEqual(
            onboarding.registra_formazione_pregressa(830, [{"corso_id": corso.pk, "data": date(2025, 1, 10)}]),
            0,
        )


class OnboardingViewTests(TestCase):
    """View lista/dettaglio/avvio/aggiornamento/chiusura + gate HR."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()
        cls.admin = User.objects.create_superuser(
            username="onb_admin", email="onb_admin@x.local", password="x"
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, reparto, attivo) "
                "VALUES (811, 'Mario', 'Rossi', 'Operatore', 'PROD', 1)"
            )

    def setUp(self):
        self.client.force_login(self.admin)

    def test_list_richiede_permesso_hr(self):
        with patch("anagrafica.views._check_hr_permission", return_value=False):
            resp = self.client.get(reverse("anagrafica:onboarding_list"))
        self.assertEqual(resp.status_code, 302)

    def test_avvia_crea_pratica_con_task(self):
        from .models import OnboardingPratica
        resp = self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        self.assertEqual(resp.status_code, 302)
        pratica = OnboardingPratica.objects.get(legacy_anagrafica_id=811)
        self.assertEqual(pratica.stato, OnboardingPratica.STATO_IN_CORSO)
        self.assertGreater(pratica.tasks.count(), 0)
        self.assertIn(f"/onboarding/{pratica.id}/", resp["Location"])

    def test_avvia_blocca_doppia_pratica(self):
        from .models import OnboardingPratica
        self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        self.assertEqual(
            OnboardingPratica.objects.filter(legacy_anagrafica_id=811).count(), 1
        )

    def test_task_update_completa(self):
        from .models import OnboardingPratica, OnboardingTask
        self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        pratica = OnboardingPratica.objects.get(legacy_anagrafica_id=811)
        task = pratica.tasks.first()
        self.client.post(
            reverse("anagrafica:onboarding_task_update", args=[pratica.id, task.id]),
            {"stato": OnboardingTask.STATO_COMPLETATO, "note": "fatto"},
        )
        task.refresh_from_db()
        self.assertEqual(task.stato, OnboardingTask.STATO_COMPLETATO)
        self.assertIsNotNone(task.completed_at)

    def test_chiudi_pratica_con_eccezioni(self):
        from .models import OnboardingPratica
        self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        pratica = OnboardingPratica.objects.get(legacy_anagrafica_id=811)
        self.client.post(reverse("anagrafica:onboarding_chiudi", args=[pratica.id]))
        pratica.refresh_from_db()
        self.assertEqual(pratica.stato, OnboardingPratica.STATO_CHIUSA_CON_ECCEZIONI)

    def test_detail_render(self):
        from .models import OnboardingPratica
        self.client.post(reverse("anagrafica:onboarding_avvia", args=[811]))
        pratica = OnboardingPratica.objects.get(legacy_anagrafica_id=811)
        resp = self.client.get(reverse("anagrafica:onboarding_detail", args=[pratica.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Checklist di inserimento")


class QueryAssenzeDipendenteTests(TestCase):
    """Match assenze nel tab scheda dipendente: deve restare allineato al widget
    conteggio (copia_nome) e non dipendere dalla sola colonna `dipendente_id`,
    che può mancare in prod."""

    def _fake_cursor(self, rows, description):
        cur = MagicMock()
        cur.fetchall.return_value = rows
        cur.description = description
        cm = MagicMock()
        cm.__enter__.return_value = cur
        cm.__exit__.return_value = False
        return cur, cm

    _DESC = [("data_inizio",), ("data_fine",), ("tipo_assenza",), ("moderation_status",)]

    @patch("anagrafica.views.legacy_table_columns")
    @patch("anagrafica.views.connections")
    def test_match_per_copia_nome_senza_dipendente_id(self, mock_conn, mock_cols):
        """Scenario prod: `assenze` ha copia_nome ma non dipendente_id."""
        from anagrafica import views

        mock_cols.side_effect = lambda table: (
            {"id", "copia_nome", "data_inizio", "data_fine", "tipo_assenza", "moderation_status"}
            if table == "assenze"
            else set()
        )
        cur, cm = self._fake_cursor(
            rows=[(date(2026, 1, 10), date(2026, 1, 12), "Ferie", 0)],
            description=self._DESC,
        )
        mock_conn.__getitem__.return_value.cursor.return_value = cm

        lista, no_link = views._query_assenze_dipendente(
            {"id": 1, "nome": "Mario", "cognome": "Rossi", "utente_id": 77}
        )

        self.assertFalse(no_link)
        self.assertEqual(len(lista), 1)
        sql, params = cur.execute.call_args.args
        self.assertIn("copia_nome", sql)
        self.assertNotIn("dipendente_id", sql)
        self.assertNotIn("JOIN", sql.upper())
        # match in entrambi gli ordini del nome
        self.assertIn("%Mario Rossi%", params)
        self.assertIn("%Rossi Mario%", params)

    @patch("anagrafica.views.legacy_table_columns")
    @patch("anagrafica.views.connections")
    def test_join_dipendenti_quando_colonna_presente(self, mock_conn, mock_cols):
        from anagrafica import views

        def cols(table):
            if table == "assenze":
                return {"id", "copia_nome", "dipendente_id", "utente_id",
                        "data_inizio", "data_fine", "tipo_assenza", "moderation_status"}
            if table == "dipendenti":
                return {"id", "utente_id"}
            return set()

        mock_cols.side_effect = cols
        cur, cm = self._fake_cursor(rows=[], description=self._DESC)
        mock_conn.__getitem__.return_value.cursor.return_value = cm

        lista, no_link = views._query_assenze_dipendente(
            {"id": 1, "nome": "Mario", "cognome": "Rossi", "utente_id": 77}
        )

        self.assertFalse(no_link)
        self.assertEqual(lista, [])
        sql, _params = cur.execute.call_args.args
        self.assertIn("LEFT JOIN dipendenti", sql)
        self.assertIn("a.utente_id = %s", sql)
        self.assertIn("d.utente_id = %s", sql)

    @patch("anagrafica.views.legacy_table_columns")
    def test_no_link_senza_identita(self, mock_cols):
        from anagrafica import views

        lista, no_link = views._query_assenze_dipendente(
            {"id": 1, "nome": "", "cognome": "", "utente_id": None}
        )
        self.assertTrue(no_link)
        self.assertEqual(lista, [])
        mock_cols.assert_not_called()

    @patch("anagrafica.views.legacy_table_columns", return_value=set())
    def test_tabella_assente_non_e_no_link(self, _mock_cols):
        """Senza tabella legacy (dev) non è un problema di linking: lista vuota,
        no_link False (mostra 'Nessuna assenza' anziché 'non collegato')."""
        from anagrafica import views

        lista, no_link = views._query_assenze_dipendente(
            {"id": 1, "nome": "Mario", "cognome": "Rossi", "utente_id": 77}
        )
        self.assertFalse(no_link)
        self.assertEqual(lista, [])


class RateiAlertLogicTests(TestCase):
    """AB1-D — logica semaforo residuo ferie (SaldoCedolino)."""

    def test_valuta_residuo_ferie_toni(self):
        from anagrafica.ratei_alert import valuta_residuo_ferie
        soglie = {"ore_max": 200.0, "ore_warn": 160.0}
        self.assertEqual(valuta_residuo_ferie(-3, soglie)["tono"], "rosso")
        self.assertEqual(valuta_residuo_ferie(210, soglie)["tono"], "rosso")
        self.assertEqual(valuta_residuo_ferie(170, soglie)["tono"], "giallo")
        self.assertEqual(valuta_residuo_ferie(50, soglie)["tono"], "verde")
        self.assertEqual(valuta_residuo_ferie(None, soglie)["tono"], "verde")

    def test_soglie_ratei_da_siteconfig(self):
        from core.models import SiteConfig
        from anagrafica.ratei_alert import soglie_ratei
        SiteConfig.set("ratei_ferie_alert_ore_max", "120")
        SiteConfig.set("ratei_ferie_alert_ore_warn", "100")
        soglie = soglie_ratei()
        self.assertEqual(soglie["ore_max"], 120.0)
        self.assertEqual(soglie["ore_warn"], 100.0)

    def test_soglie_warn_clamped_a_max(self):
        from core.models import SiteConfig
        from anagrafica.ratei_alert import soglie_ratei
        SiteConfig.set("ratei_ferie_alert_ore_max", "100")
        SiteConfig.set("ratei_ferie_alert_ore_warn", "250")
        self.assertEqual(soglie_ratei()["ore_warn"], 100.0)

    def test_filtro_allerta_q(self):
        from anagrafica.ratei_alert import filtro_allerta_q
        q = filtro_allerta_q({"ore_max": 200.0, "ore_warn": 160.0})
        for tax, residui in [("AAA", -5), ("BBB", 210), ("CCC", 170), ("DDD", 50)]:
            SaldoCedolino.objects.create(
                tax_code=tax, data_competenza=date(2026, 5, 31), ferie_residui=residui,
            )
        codici = set(SaldoCedolino.objects.filter(q).values_list("tax_code", flat=True))
        self.assertEqual(codici, {"AAA", "BBB", "CCC"})


@override_settings(LEGACY_AUTH_ENABLED=False, SECURE_SSL_REDIRECT=False)
class RateiAlertViewTests(TestCase):
    """AB1-D — vista ratei_list con semaforo e filtro allerta (HR)."""

    def setUp(self):
        self.user = User.objects.create_superuser(
            username="ratei-alert", email="ratei-alert@example.com", password="pass12345",
        )
        for tax, residui in [("AAA", -5), ("BBB", 210), ("CCC", 170), ("DDD", 50)]:
            SaldoCedolino.objects.create(
                tax_code=tax, legacy_anagrafica_id=None,
                data_competenza=date(2026, 5, 31), ferie_residui=residui,
            )

    def test_list_renders_with_kpi(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:ratei_list"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["n_negativi"], 1)
        self.assertEqual(response.context["n_accumulo"], 1)
        self.assertEqual(response.context["n_allerta"], 3)
        self.assertEqual(response.context["totale"], 4)

    def test_filtro_solo_allerta(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("anagrafica:ratei_list"), {"allerta": "1"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["totale"], 3)
        codici = {s.tax_code for s in response.context["page_obj"].object_list}
        self.assertEqual(codici, {"AAA", "BBB", "CCC"})


class ImportASRTests(TestCase):
    """Importer matrice ASR: match per CF, qualifiche con date, livello rischio."""

    @classmethod
    def setUpTestData(cls):
        _ensure_anagrafica_table()

    def _build_xlsx(self, path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Salute e Sicurezza"
        ws1.append(["Cognome Nome", "CODICE FISCALE", "Rischio",
                    "Ultimo Corso Lavoratori", "Anno Scadenza (5 anni)"])
        ws1.append(["ROSSI MARIO", "RSSMRA80A01H501U", "A", date(2024, 1, 10), 2029])
        ws2 = wb.create_sheet("Cartel1.XLS")
        ws2.append(["nome", "BLSD"])
        ws2.append(["ROSSI MARIO", date(2024, 2, 1)])
        wb.save(path)

    def test_import_commit_popola_qualifiche_e_rischio(self):
        import tempfile, os
        from django.core.management import call_command
        from .models import (
            DipendenteAnagraficaCivile, DipendenteQualifica, Mansione,
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (901, 'Mario', 'Rossi', 'Saldatore', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=901, codice_fiscale="RSSMRA80A01H501U"
        )
        mans = Mansione.objects.create(nome="Saldatore")

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self._build_xlsx(path)
            call_command("import_asr", path, "--commit", verbosity=0)
        finally:
            os.remove(path)

        # abilitazione BLSD con data + scadenza calcolata (36 mesi default)
        q = DipendenteQualifica.objects.get(legacy_anagrafica_id=901, tipo__nome="BLSD")
        self.assertEqual(q.data_conseguimento, date(2024, 2, 1))
        self.assertEqual(q.data_scadenza, date(2027, 2, 1))
        # corso lavoratori con scadenza all'anno indicato
        corso = DipendenteQualifica.objects.get(
            legacy_anagrafica_id=901, tipo__nome="Formazione lavoratori (ASR)"
        )
        self.assertEqual(corso.data_scadenza.year, 2029)
        # livello rischio impostato sulla mansione del dipendente
        mans.refresh_from_db()
        self.assertEqual(mans.livello_rischio, Mansione.RISCHIO_ALTO)

    def test_import_dry_run_non_scrive(self):
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile, DipendenteQualifica
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (902, 'Anna', 'Verdi', 'Operaia', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=902, codice_fiscale="VRDNNA85A41H501T"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            # CF diverso nel file → uso lo stesso schema ma con la dipendente 902
            import openpyxl
            wb = openpyxl.Workbook()
            ws1 = wb.active; ws1.title = "Salute e Sicurezza"
            ws1.append(["Cognome Nome", "CODICE FISCALE", "Rischio",
                        "Ultimo Corso Lavoratori", "Anno Scadenza (5 anni)"])
            ws1.append(["VERDI ANNA", "VRDNNA85A41H501T", "B", date(2024, 3, 1), 2029])
            ws2 = wb.create_sheet("Cartel1.XLS")
            ws2.append(["nome", "Carrellisti"]); ws2.append(["VERDI ANNA", date(2024, 3, 1)])
            wb.save(path)
            call_command("import_asr", path, verbosity=0)  # dry-run di default
        finally:
            os.remove(path)
        self.assertEqual(DipendenteQualifica.objects.filter(legacy_anagrafica_id=902).count(), 0)

    def test_import_match_nome_con_annotazioni(self):
        """Il foglio date può avere il nominativo "sporcato" da annotazioni
        ("(PREPOSTO)", "- PENSIONAMENTO"): la normalizzazione le rimuove e il
        match col foglio principale (nome pulito) avviene comunque."""
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile, DipendenteQualifica
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (904, 'Iacopo', 'Simoncini', 'Saldatore', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=904, codice_fiscale="SMNCPI80A01H501U"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws1 = wb.active; ws1.title = "Salute e Sicurezza"
            ws1.append(["Cognome Nome", "CODICE FISCALE", "Rischio",
                        "Ultimo Corso Lavoratori", "Anno Scadenza (5 anni)"])
            ws1.append(["SIMONCINI IACOPO", "SMNCPI80A01H501U", "A", date(2024, 1, 10), 2029])
            ws2 = wb.create_sheet("Cartel1.XLS")
            ws2.append(["nome", "BLSD"])
            # nominativo annotato nel foglio date: deve comunque matchare
            ws2.append(["SIMONCINI IACOPO (PREPOSTO)", date(2024, 2, 1)])
            wb.save(path)
            call_command("import_asr", path, "--commit", verbosity=0)
        finally:
            os.remove(path)
        self.assertTrue(
            DipendenteQualifica.objects.filter(
                legacy_anagrafica_id=904, tipo__nome="BLSD"
            ).exists()
        )

    def test_norm_nome_preserva_cognomi_composti(self):
        """La pulizia annotazioni non deve spezzare i cognomi con trattino senza
        spazi (es. "FERRARI-ROSSI MARIO")."""
        from anagrafica.management.commands.import_asr import _norm_nome
        self.assertEqual(_norm_nome("FERRARI-ROSSI MARIO"), "FERRARI-ROSSI MARIO")
        self.assertEqual(_norm_nome("Simoncini Iacopo (Preposto)"), "SIMONCINI IACOPO")
        self.assertEqual(_norm_nome("Pasqualetti Marco - Pensionamento"), "PASQUALETTI MARCO")

    def test_import_collega_corso_qualifica_e_record(self):
        """Step 2: l'import collega alla fonte corso→qualifica e
        DipendenteQualifica→completamento (modello qualifica àncora)."""
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile, DipendenteQualifica, TipoQualifica
        from .models_formazione import TrainingCourse, TrainingEmployeeRecord
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (905, 'Sara', 'Neri', 'Saldatore', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=905, codice_fiscale="NRESRA80A41H501T"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self._build_xlsx_dip(path, "NERI SARA", "NRESRA80A41H501T", abil="BLSD")
            call_command("import_asr", path, "--commit", verbosity=0)
        finally:
            os.remove(path)
        tipo = TipoQualifica.objects.get(nome="BLSD")
        corso = TrainingCourse.objects.get(titolo="BLSD")
        self.assertEqual(corso.qualifica_id, tipo.id)  # corso -> qualifica
        dq = DipendenteQualifica.objects.get(legacy_anagrafica_id=905, tipo=tipo)
        rec = TrainingEmployeeRecord.objects.get(legacy_anagrafica_id=905, corso=corso)
        self.assertEqual(dq.record_formazione_id, rec.id)  # qualifica dip -> evidenza

    def _build_xlsx_dip(self, path, cognome_nome, cf, abil="BLSD", data_abil=date(2024, 2, 1)):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "Salute e Sicurezza"
        ws1.append(["Cognome Nome", "CODICE FISCALE", "Rischio",
                    "Ultimo Corso Lavoratori", "Anno Scadenza (5 anni)"])
        ws1.append([cognome_nome, cf, "A", date(2024, 1, 10), 2029])
        ws2 = wb.create_sheet("Cartel1.XLS")
        ws2.append(["nome", abil]); ws2.append([cognome_nome, data_abil])
        wb.save(path)

    def test_import_commit_popola_formazione_e_idempotente(self):
        """Fase formazione: corso creato, sessione partecipata + record completamento;
        doppio run non genera doppioni (match per titolo)."""
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile
        from .models_formazione import (
            TrainingCourse, TrainingEmployeeRecord, TrainingSession,
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (903, 'Luca', 'Bianchi', 'Saldatore', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=903, codice_fiscale="BNCLCU80A01H501U"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self._build_xlsx_dip(path, "BIANCHI LUCA", "BNCLCU80A01H501U")
            call_command("import_asr", path, "--commit", verbosity=0)
            call_command("import_asr", path, "--commit", verbosity=0)  # idempotenza
        finally:
            os.remove(path)
        # corso BLSD creato nel modulo formazione, una sola volta
        self.assertEqual(TrainingCourse.objects.filter(titolo__icontains="BLSD").count(), 1)
        # 2 completamenti (BLSD + corso lavoratori), senza doppioni dopo due run
        recs = TrainingEmployeeRecord.objects.filter(legacy_anagrafica_id=903)
        self.assertEqual(recs.count(), 2)
        blsd = recs.get(corso__titolo__icontains="BLSD")
        self.assertEqual(blsd.data_completamento, date(2024, 2, 1))
        self.assertEqual(blsd.data_scadenza, date(2027, 2, 1))  # BLSD 36 mesi
        self.assertTrue(TrainingSession.objects.filter(corso__titolo__icontains="BLSD").exists())

    def test_import_riusa_corso_esistente_senza_duplicare(self):
        """Se un corso a catalogo matcha (alias per titolo), viene riusato."""
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile
        from .models_formazione import TrainingCourse, TrainingPlan
        piano = TrainingPlan.objects.create(
            nome="Piano X", codice="P-X", categoria="OBBLIGATORIA", stato="ATTIVO"
        )
        esistente = TrainingCourse.objects.create(
            piano=piano, codice="C-BLSD-OLD", titolo="BLS-D Corso Operatore",
            durata_ore_teorica=5, stato="ATTIVO",
        )
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (904, 'Sara', 'Neri', 'Operaia', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=904, codice_fiscale="NRESRA85A41H501T"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self._build_xlsx_dip(path, "NERI SARA", "NRESRA85A41H501T")
            call_command("import_asr", path, "--commit", "--no-qualifiche", verbosity=0)
        finally:
            os.remove(path)
        # nessun corso BLSD nuovo: riusato quello esistente
        self.assertEqual(TrainingCourse.objects.filter(titolo__icontains="bls").count(), 1)
        self.assertEqual(
            esistente.record_completamenti.filter(legacy_anagrafica_id=904).count(), 1
        )

    def test_no_corsi_salta_formazione(self):
        import tempfile, os
        from django.core.management import call_command
        from .models import DipendenteAnagraficaCivile
        from .models_formazione import TrainingEmployeeRecord
        with connection.cursor() as cur:
            cur.execute(
                "INSERT INTO anagrafica_dipendenti (id, nome, cognome, mansione, attivo) "
                "VALUES (905, 'Ivo', 'Galli', 'Operaio', 1)"
            )
        DipendenteAnagraficaCivile.objects.create(
            legacy_anagrafica_id=905, codice_fiscale="GLLIVO80A01H501U"
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            self._build_xlsx_dip(path, "GALLI IVO", "GLLIVO80A01H501U")
            call_command("import_asr", path, "--commit", "--no-corsi", verbosity=0)
        finally:
            os.remove(path)
        self.assertEqual(
            TrainingEmployeeRecord.objects.filter(legacy_anagrafica_id=905).count(), 0
        )
