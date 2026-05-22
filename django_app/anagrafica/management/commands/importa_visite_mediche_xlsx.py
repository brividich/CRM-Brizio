"""Import storico visite mediche da file Excel (.xlsx).

Formato atteso: foglio con header in prima riga, colonne come da export
``medical-examinations-people.xlsx`` (Codice fiscale, Visita, Data visita,
Data di richiamo, Esito, Note, ...).

Esempi:
    python manage.py importa_visite_mediche_xlsx django_app/medical-examinations-people.xlsx --dry-run
    python manage.py importa_visite_mediche_xlsx django_app/medical-examinations-people.xlsx
    python manage.py importa_visite_mediche_xlsx django_app/medical-examinations-people.xlsx --skip-existing
    python manage.py importa_visite_mediche_xlsx django_app/medical-examinations-people.xlsx --sheet Worksheet --limit 10
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from anagrafica.models import (
    DipendenteAnagraficaCivile,
    TipoVisitaMedica,
    VisitaMedica,
)


# ---------------------------------------------------------------------------
# Mapping esito: stringa Excel → VisitaMedica.Esito
# ---------------------------------------------------------------------------

def _norm_str(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


ESITO_MAP: dict[str, str] = {
    "idoneo": VisitaMedica.Esito.IDONEO,
    "idoneo mansione specifica": VisitaMedica.Esito.IDONEO_MANSIONE,
    "idoneo con prescrizioni": VisitaMedica.Esito.IDONEO_CON_PRESCRIZIONI,
    "idoneo con limitazioni": VisitaMedica.Esito.IDONEO_CON_LIMITAZIONI,
    "idoneo con limitazioni e prescrizioni": VisitaMedica.Esito.IDONEO_LIM_PRESCR,
    "non idoneo": VisitaMedica.Esito.NON_IDONEO_TEMPORANEO,
    "non idoneo temporaneo": VisitaMedica.Esito.NON_IDONEO_TEMPORANEO,
    "non idoneo (temporaneo)": VisitaMedica.Esito.NON_IDONEO_TEMPORANEO,
    "non idoneo definitivo": VisitaMedica.Esito.NON_IDONEO_DEFINITIVO,
    "non idoneo (definitivo)": VisitaMedica.Esito.NON_IDONEO_DEFINITIVO,
}


def _map_esito(raw) -> str:
    key = _norm_str(raw)
    if key in ESITO_MAP:
        return ESITO_MAP[key]
    # Fuzzy fallback
    if "limitazioni" in key and "prescriz" in key:
        return VisitaMedica.Esito.IDONEO_LIM_PRESCR
    if "limitazioni" in key:
        return VisitaMedica.Esito.IDONEO_CON_LIMITAZIONI
    if "prescriz" in key:
        return VisitaMedica.Esito.IDONEO_CON_PRESCRIZIONI
    if "mansione" in key:
        return VisitaMedica.Esito.IDONEO_MANSIONE
    if "non idoneo" in key:
        return VisitaMedica.Esito.NON_IDONEO_TEMPORANEO
    if "idoneo" in key:
        return VisitaMedica.Esito.IDONEO
    return VisitaMedica.Esito.IDONEO


# ---------------------------------------------------------------------------
# Durata mesi inferita dal nome del tipo visita
# ---------------------------------------------------------------------------

def _infer_durata_mesi(nome: str) -> int:
    n = nome.lower()
    if "quinquennale" in n or "5 anni" in n:
        return 60
    if "triennale" in n or "3 anni" in n:
        return 36
    if "biennale" in n or "2 anni" in n:
        return 24
    return 12  # Default: annuale


# ---------------------------------------------------------------------------
# Header mapping
# ---------------------------------------------------------------------------

def _norm_header(raw) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(raw or "").lower())


HEADER_ALIAS: dict[str, str] = {
    "codicefiscale": "codice_fiscale",
    "persona": "persona",
    "visita": "tipo_visita",
    "datavisita": "data_svolgimento",
    "datasvolgimento": "data_svolgimento",
    "datadirichiamo": "data_richiamo",
    "datarichiamo": "data_richiamo",
    "esito": "esito",
    "note": "note",
    # Colonne ignorate (sempre vuote o non rilevanti)
    "valoremonetario": "_skip",
    "categoria": "_skip",
    "azienda": "_skip",
    "sede": "_skip",
    "ruolo": "_skip",
}


def _build_header_index(header_row) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, raw in enumerate(header_row):
        if raw is None:
            continue
        key = _norm_header(str(raw))
        mapped = HEADER_ALIAS.get(key)
        if mapped and mapped != "_skip":
            out[mapped] = i
    return out


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa storico visite mediche da file Excel (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Percorso al file .xlsx")
        parser.add_argument("--sheet", type=str, default="", help="Nome foglio (default: primo)")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla, mostra solo il report")
        parser.add_argument("--skip-existing", action="store_true", default=True,
                            help="Salta se esiste già (legacy_id, tipo, data_svolgimento) [default: True]")
        parser.add_argument("--overwrite-existing", action="store_true",
                            help="Aggiorna i record già esistenti invece di saltarli")
        parser.add_argument("--limit", type=int, default=0, help="Limita a N righe (0 = tutte)")
        parser.add_argument("--verbose", action="store_true", help="Mostra ogni riga processata")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl non installato.") from exc

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet_name = opts.get("sheet") or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' non trovato. Disponibili: {wb.sheetnames}")
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise CommandError("Foglio vuoto.") from exc

        idx = _build_header_index(header_row)
        required = ["codice_fiscale", "tipo_visita", "data_svolgimento", "esito"]
        missing = [k for k in required if k not in idx]
        if missing:
            raise CommandError(
                f"Colonne obbligatorie mancanti: {missing}\nHeader letto: {list(header_row)}"
            )

        dry_run = bool(opts.get("dry_run"))
        overwrite = bool(opts.get("overwrite_existing"))
        # skip_existing è il default: True a meno che non si passi --overwrite-existing
        skip_existing = not overwrite
        limit = int(opts.get("limit") or 0)
        verbose = bool(opts.get("verbose"))

        self.stdout.write(
            self.style.NOTICE(
                f"Import visite mediche da {path.name}"
                f" [sheet={sheet_name}, dry_run={dry_run},"
                f" skip_existing={skip_existing}, limit={limit or 'all'}]"
            )
        )

        # Pre-carica mappa CF → legacy_anagrafica_id
        cf_to_legacy: dict[str, int] = {}
        for cf, lid in (
            DipendenteAnagraficaCivile.objects
            .exclude(codice_fiscale="")
            .values_list("codice_fiscale", "legacy_anagrafica_id")
        ):
            cf_to_legacy[cf.strip().upper()] = lid

        self.stdout.write(f"  Dipendenti con CF registrato: {len(cf_to_legacy)}")

        # Pre-carica TipoVisitaMedica esistenti (nome normalizzato → oggetto)
        tipo_cache: dict[str, TipoVisitaMedica] = {
            t.nome.strip().lower(): t
            for t in TipoVisitaMedica.objects.all()
        }

        stats = {
            "lette": 0,
            "create": 0,
            "aggiornate": 0,
            "saltate_esistenti": 0,
            "saltate_cf_non_trovato": 0,
            "saltate_dati_mancanti": 0,
            "errori": 0,
            "tipi_creati": 0,
        }
        errors: list[str] = []
        tipi_creati_nomi: list[str] = []

        class _DryRunAbort(Exception):
            pass

        def cell(row, key):
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        try:
            with transaction.atomic():
                for n, raw_row in enumerate(rows_iter, start=2):
                    if limit and stats["lette"] >= limit:
                        break
                    if raw_row is None or all(c in (None, "") for c in raw_row):
                        continue
                    stats["lette"] += 1

                    try:
                        with transaction.atomic():
                            result = self._process_row(
                                raw_row=raw_row,
                                cell=cell,
                                cf_to_legacy=cf_to_legacy,
                                tipo_cache=tipo_cache,
                                skip_existing=skip_existing,
                                overwrite=overwrite,
                                stats=stats,
                                errors=errors,
                                tipi_creati_nomi=tipi_creati_nomi,
                                line=n,
                                verbose=verbose,
                                dry_run=dry_run,
                            )
                            if verbose and result:
                                self.stdout.write(f"  riga {n}: {result}")
                    except Exception as exc:
                        stats["errori"] += 1
                        errors.append(f"riga {n}: {exc.__class__.__name__}: {exc}")

                if dry_run:
                    raise _DryRunAbort()
        except _DryRunAbort:
            pass

        self._print_report(stats, errors, tipi_creati_nomi, dry_run=dry_run)

    # ------------------------------------------------------------------

    def _process_row(
        self, *, raw_row, cell, cf_to_legacy, tipo_cache,
        skip_existing, overwrite, stats, errors, tipi_creati_nomi,
        line, verbose, dry_run,
    ) -> str:
        cf_raw = cell(raw_row, "codice_fiscale")
        tipo_nome_raw = cell(raw_row, "tipo_visita")
        data_raw = cell(raw_row, "data_svolgimento")
        esito_raw = cell(raw_row, "esito")

        cf = re.sub(r"\s+", "", str(cf_raw or "")).upper()
        tipo_nome = str(tipo_nome_raw or "").strip()
        data_svolgimento = _parse_date(data_raw)
        data_richiamo = _parse_date(cell(raw_row, "data_richiamo"))
        note = str(cell(raw_row, "note") or "").strip()

        if not cf or not tipo_nome or not data_svolgimento:
            stats["saltate_dati_mancanti"] += 1
            errors.append(f"riga {line}: dati mancanti (CF={cf!r}, tipo={tipo_nome!r}, data={data_raw!r})")
            return ""

        # Risolvi legacy_id
        legacy_id = cf_to_legacy.get(cf)
        if legacy_id is None:
            stats["saltate_cf_non_trovato"] += 1
            errors.append(f"riga {line}: CF '{cf}' non trovato in anagrafica")
            return ""

        # Risolvi / crea TipoVisitaMedica
        tipo_key = tipo_nome.strip().lower()
        if tipo_key not in tipo_cache:
            durata = _infer_durata_mesi(tipo_nome)
            tipo_obj = TipoVisitaMedica(nome=tipo_nome, durata_mesi=durata, obbligatoria=True)
            tipo_obj.save()
            tipo_cache[tipo_key] = tipo_obj
            stats["tipi_creati"] += 1
            tipi_creati_nomi.append(tipo_nome)
        else:
            tipo_obj = tipo_cache[tipo_key]

        esito = _map_esito(esito_raw)

        # Cerca duplicato (legacy_id + tipo + data_svolgimento)
        existing = VisitaMedica.objects.filter(
            legacy_anagrafica_id=legacy_id,
            tipo=tipo_obj,
            data_svolgimento=data_svolgimento,
        ).first()

        if existing:
            if skip_existing and not overwrite:
                stats["saltate_esistenti"] += 1
                return f"SALTATA (già presente id={existing.pk})"
            if overwrite:
                existing.esito = esito
                existing.prescrizioni = note
                if data_richiamo:
                    # Forza la data scadenza storica bypassando save()
                    existing.save(update_fields=["esito", "prescrizioni", "updated_at"])
                    if data_richiamo:
                        VisitaMedica.objects.filter(pk=existing.pk).update(data_scadenza=data_richiamo)
                else:
                    existing.save(update_fields=["esito", "prescrizioni", "updated_at"])
                stats["aggiornate"] += 1
                return f"AGGIORNATA id={existing.pk}"
        else:
            visita = VisitaMedica(
                legacy_anagrafica_id=legacy_id,
                tipo=tipo_obj,
                data_svolgimento=data_svolgimento,
                esito=esito,
                prescrizioni=note,
            )
            visita.save()  # auto-calcola data_scadenza dal tipo
            # Se l'Excel riporta una data richiamo diversa, la usiamo (dato storico)
            if data_richiamo and data_richiamo != visita.data_scadenza:
                VisitaMedica.objects.filter(pk=visita.pk).update(data_scadenza=data_richiamo)
            stats["create"] += 1
            return f"CREATA id={visita.pk} — {tipo_nome} [{esito}]"

        return ""

    # ------------------------------------------------------------------

    def _print_report(self, stats, errors, tipi_creati_nomi, *, dry_run):
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== Report import visite mediche ==="))
        labels = {
            "lette": "Righe lette",
            "create": "Visite create",
            "aggiornate": "Visite aggiornate",
            "saltate_esistenti": "Saltate (già presenti)",
            "saltate_cf_non_trovato": "Saltate (CF non in anagrafica)",
            "saltate_dati_mancanti": "Saltate (dati mancanti)",
            "tipi_creati": "Tipi visita creati",
            "errori": "Errori",
        }
        for key, label in labels.items():
            val = stats.get(key, 0)
            style = self.style.WARNING if (key == "errori" and val) else self.style.SUCCESS if val else str
            styled = self.style.SUCCESS(str(val)) if val and key in ("create", "aggiornate", "tipi_creati") else str(val)
            self.stdout.write(f"  {label:40s} {styled}")

        if tipi_creati_nomi:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Tipi visita creati automaticamente (controlla durata_mesi):"))
            for nome in tipi_creati_nomi:
                self.stdout.write(f"  • {nome}")

        if errors:
            self.stdout.write("")
            n_cf = stats.get("saltate_cf_non_trovato", 0)
            label = f"=== Dettaglio problemi ({len(errors)}"
            if n_cf:
                label += f" — di cui {n_cf} CF non trovati in anagrafica"
            label += ") ==="
            self.stdout.write(self.style.WARNING(label))
            for e in errors[:60]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 60:
                self.stdout.write(f"  ... e altri {len(errors) - 60}")

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.NOTICE("DRY-RUN: nessuna scrittura effettuata."))
        else:
            self.stdout.write(self.style.SUCCESS("Import completato."))
