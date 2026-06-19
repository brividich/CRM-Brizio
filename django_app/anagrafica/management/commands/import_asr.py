"""Import della matrice formazione/abilitazioni "ASR" (Excel) nelle qualifiche.

Sorgenti tipiche del file ``Programmazione ... ASR.xlsx``:
- foglio principale (default "Salute e Sicurezza"): ha **Codice Fiscale**, Reparto,
  livello **Rischio** (A=alto/16h, B=basso/8h, M=medio/12h) e la data dell'ultimo
  **Corso Lavoratori** con relativa scadenza quinquennale;
- foglio "date" (default "Cartel1.XLS"): stesse colonne ruolo/abilitazione ma con
  le **date** di conseguimento (per nominativo, senza CF).

L'import incrocia i due fogli per **nominativo normalizzato**, risale al
``legacy_anagrafica_id`` via CF, e popola **due punti di gestione** della stessa
competenza (nessun doppione: stessi dati, viste diverse):

1. **Qualifiche** (lato Salute e Sicurezza / patentini) — fase ``--no-qualifiche``
   per saltarla:
   - ``DipendenteQualifica`` per ogni abilitazione con data (DPI III, Carrellisti,
     Funi/Carroponti, PLE, PES/PAV/PEI, BLSD, Primo soccorso, Antincendio, Preposto,
     RSPP, RLS, ASPP) — ``TipoQualifica`` creato a catalogo se assente;
   - una qualifica "Formazione lavoratori (ASR)" con la scadenza a 5 anni;
   - ``Mansione.livello_rischio`` della mansione del dipendente (best-effort).

2. **Formazione** (lato modulo Corsi) — fase ``--no-corsi`` per saltarla:
   - ``TrainingCourse`` per ogni corso/abilitazione, **riusando** un corso già a
     catalogo che corrisponde per titolo (alias-match) → niente doppioni; i nuovi
     vanno nel piano "Sicurezza" con categoria "Sicurezza ASR" (così sono
     gestibili anche dal modulo Salute e Sicurezza);
   - le **sessioni partecipate**: ``TrainingSession`` (raggruppata per corso+data),
     ``TrainingEnrollment`` (completato) e ``TrainingEmployeeRecord`` (completamento
     con scadenza) per ogni dipendente con una data. Idempotente.

DRY-RUN di default: nessuna scrittura senza ``--commit``. Privacy: l'output non
stampa nominativi/CF (solo conteggi); usare ``--verbose`` per le iniziali.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal
from functools import reduce
from operator import or_

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Count, Q

# Per ogni colonna ASR: (alias per riconoscere un TipoQualifica GIÀ a catalogo,
# nome canonico da creare se nessun alias matcha, durata_mesi rinnovo).
# La durata è usata per CALCOLARE la scadenza della singola assegnazione
# (anche se il tipo a catalogo ha durata 0). Gli alias evitano i doppioni con i
# nomi reali già presenti (es. "CONDUZIONE CARRELLI…", "PREPOSTI", "BLS-D…").
ABILITAZIONI = {
    "Preposto":           (["prepost"], "Formazione Preposto", 24),
    "RSPP":               (["rspp"], "RSPP", 60),
    "RLS":                (["rls"], "RLS", 12),
    "ASPP":               (["aspp"], "ASPP", 60),
    "DPI III CAT":        (["dpi iii", "dpi 3", "iii cat", "iii categoria"], "Addestramento DPI III categoria", 0),
    "I° Soccorso":        (["primo soccorso", "soccorso"], "Primo soccorso", 36),
    "I° Soccorso ":       (["primo soccorso", "soccorso"], "Primo soccorso", 36),
    "Antincendio":        (["antincendio"], "Addetto antincendio", 60),
    "Carrellisti":        (["carrell", "conduzione carrelli", "muletto"], "Carrellista / carrello elevatore", 60),
    "Funi - Carroponti":  (["funi", "carropont"], "Funi e carroponti", 60),
    "Piattaforme Aeree":  (["piattaform", "elevabil", "ple"], "Piattaforme aeree (PLE)", 60),
    "PES PAV":            (["operatori elettric", "pes/pav", "pes pav", "pes ", "pav"], "PES/PAV/PEI", 60),
    "PES PEI":            (["operatori elettric", "pes/pav", "pes pav", "pes ", "pei"], "PES/PAV/PEI", 60),
    "BLSD":               (["blsd", "bls-d", "bls d"], "BLSD", 36),
}

CORSO_LAVORATORI_ALIAS = ["formazione lavoratori", "corso lavoratori", "lavoratori (asr)"]
CORSO_LAVORATORI_TIPO = "Formazione lavoratori (ASR)"
CORSO_LAVORATORI_DURATA = 60  # ASR: aggiornamento ogni 5 anni


# Annotazioni che nel foglio date sporcano il nominativo e impediscono il match
# col foglio principale: "(PREPOSTO)", "(AGGIORNAMENTO)" e suffissi dopo un
# trattino con spazi ("- PENSIONAMENTO"). Lo spazio attorno al trattino evita di
# spezzare i cognomi composti ("FERRARI-ROSSI", senza spazi).
_ANNOT_RE = re.compile(r"\([^)]*\)|\s[-–]\s.*$")


def _norm_nome(value) -> str:
    s = _ANNOT_RE.sub(" ", str(value or "").upper())
    return " ".join(s.split())


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _add_months(d: date, months: int) -> date:
    from anagrafica.models import _add_months as add
    return add(d, months)


class Command(BaseCommand):
    help = "Importa la matrice formazione/abilitazioni ASR (Excel) nelle qualifiche dipendente."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Percorso del file .xlsx")
        parser.add_argument("--sheet-main", default="Salute e Sicurezza",
                            help="Foglio con CF/Reparto/Rischio/Corso Lavoratori")
        parser.add_argument("--sheet-dates", default="Cartel1.XLS",
                            help="Foglio con le date delle abilitazioni (per nominativo)")
        parser.add_argument("--commit", action="store_true",
                            help="Scrive su DB (di default è dry-run)")
        parser.add_argument("--user-id", type=int, default=None,
                            help="ID utente Django che risulta come assegnatore")
        parser.add_argument("--no-qualifiche", action="store_true",
                            help="Salta la fase qualifiche (DipendenteQualifica)")
        parser.add_argument("--no-corsi", action="store_true",
                            help="Salta la fase formazione (TrainingCourse/sessioni)")

    def handle(self, *args, **opt):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl non disponibile.")
        from anagrafica.models import (
            DipendenteAnagraficaCivile, DipendenteQualifica, Mansione, TipoQualifica,
        )
        from anagrafica.models_formazione import (
            TrainingCourse, TrainingEmployeeRecord, TrainingEnrollment,
            TrainingPlan, TrainingSession,
        )
        from anagrafica.models_rischi import CategoriaCorso
        from anagrafica.services.formazione_import import _short_code
        from core.legacy_models import AnagraficaDipendente

        dry = not opt["commit"]
        verbose = opt["verbosity"] >= 2
        skip_q = opt["no_qualifiche"]
        skip_c = opt["no_corsi"]

        wb = openpyxl.load_workbook(opt["file"], data_only=True, read_only=True)
        for sh in (opt["sheet_main"], opt["sheet_dates"]):
            if sh not in wb.sheetnames:
                raise CommandError(f"Foglio '{sh}' non trovato. Disponibili: {wb.sheetnames}")

        # CF → legacy_id
        cf_to_lid = {
            (cf or "").strip().upper(): lid
            for cf, lid in DipendenteAnagraficaCivile.objects
            .exclude(codice_fiscale="").values_list("codice_fiscale", "legacy_anagrafica_id")
        }
        # legacy_id → mansione (per livello rischio)
        lid_to_mansione = dict(
            AnagraficaDipendente.objects.exclude(mansione="").exclude(mansione__isnull=True)
            .values_list("id", "mansione")
        )

        # 1) Foglio principale: nome → dati anagrafici/rischio/corso
        main = wb[opt["sheet_main"]]
        mrows = list(main.iter_rows(values_only=True))
        mh = mrows[0]
        def mi(name):
            for i, h in enumerate(mh):
                if h and name.lower() in str(h).lower():
                    return i
            return None
        c_nome, c_cf, c_ris = mi("Cognome Nome"), mi("Codice Fiscale"), mi("Rischio")
        c_corso, c_scad = mi("Ultimo Corso Lavoratori"), mi("Anno Scadenza")
        if c_nome is None or c_cf is None:
            raise CommandError("Colonne 'Cognome Nome' / 'Codice Fiscale' non trovate nel foglio principale.")

        info: dict[str, dict] = {}
        for r in mrows[1:]:
            nome = _norm_nome(r[c_nome])
            if not nome:
                continue
            cf = (str(r[c_cf]).strip().upper() if r[c_cf] else "")
            info[nome] = {
                "cf": cf,
                "lid": cf_to_lid.get(cf),
                "rischio": (str(r[c_ris]).strip().upper() if c_ris is not None and r[c_ris] else ""),
                "corso": _parse_date(r[c_corso]) if c_corso is not None else None,
                "anno_scad": r[c_scad] if c_scad is not None else None,
            }

        # 2) Foglio date: header colonne ruolo/abilitazione
        dsheet = wb[opt["sheet_dates"]]
        drows = list(dsheet.iter_rows(values_only=True))
        wb.close()  # read_only tiene il file aperto: chiudere per liberare l'handle
        dh = drows[0]
        d_nome = next((i for i, h in enumerate(dh) if h and "nome" in str(h).lower()
                       and "fiscale" not in str(h).lower()), None)
        if d_nome is None:
            raise CommandError("Colonna nominativo non trovata nel foglio date.")
        abil_cols = {i: str(h).strip() for i, h in enumerate(dh)
                     if h and str(h).strip() in ABILITAZIONI}

        counters = {"matched": 0, "unmatched_name": 0, "unmatched_cf": 0,
                    "qualifiche": 0, "corso_lavoratori": 0, "mansioni_rischio": 0,
                    "corsi_nuovi": 0, "corsi_match": 0, "sessioni": 0, "completamenti": 0,
                    "link_corso_qualifica": 0}
        tipi_cache: dict[str, TipoQualifica] = {}
        corso_cache: dict[str, TrainingCourse | None] = {}
        sess_cache: dict[tuple, TrainingSession | None] = {}
        seen_sess: set[tuple] = set()       # (canonical, data) → sessioni logiche
        seen_compl: set[tuple] = set()      # (lid, canonical, data) → completamenti logici
        mansioni_seen: dict[str, str] = {}  # nome_mansione → rischio (per riepilogo)
        assegnatore = None
        if opt["user_id"]:
            from django.contrib.auth import get_user_model
            assegnatore = get_user_model().objects.filter(pk=opt["user_id"]).first()

        # Piano + categoria condivisi per i corsi ASR nuovi (lazy, una volta sola).
        formazione_ctx: dict[str, object | None] = {"piano": None, "categoria": None, "ready": False}

        def _ensure_formazione_ctx():
            if formazione_ctx["ready"] or dry:
                formazione_ctx["ready"] = True
                return
            piano = TrainingPlan.objects.filter(nome__iexact="Sicurezza").first()
            if piano is None:
                piano = TrainingPlan.objects.create(
                    nome="Sicurezza", codice=_short_code("Sicurezza", prefix="P-", max_len=20),
                    categoria="OBBLIGATORIA", stato="ATTIVO", is_active=True, created_by=assegnatore,
                )
            cat = CategoriaCorso.objects.filter(nome__iexact="Sicurezza ASR").first()
            if cat is None:
                cat = CategoriaCorso.objects.create(
                    nome="Sicurezza ASR", codice=_short_code("Sicurezza ASR", max_len=20),
                )
            formazione_ctx["piano"], formazione_ctx["categoria"] = piano, cat
            formazione_ctx["ready"] = True

        def get_tipo(aliases, canonical, durata):
            """Riusa un TipoQualifica esistente che matcha un alias (il più usato);
            altrimenti crea il tipo canonico. Evita i doppioni col catalogo reale."""
            if canonical in tipi_cache:
                return tipi_cache[canonical]
            q = reduce(or_, (Q(nome__icontains=a) for a in aliases))
            t = (TipoQualifica.objects.filter(q)
                 .annotate(_n=Count("assegnazioni")).order_by("-_n", "id").first())
            if t is None and not dry:
                t = TipoQualifica.objects.create(
                    nome=canonical, categoria=TipoQualifica.CAT_SICUREZZA, durata_mesi=durata,
                )
            tipi_cache[canonical] = t
            return t

        def get_corso(aliases, canonical, validita, tipo=None):
            """Riusa un TrainingCourse esistente che matcha un alias (il più
            usato per n. completamenti); altrimenti crea il corso canonico nel
            piano "Sicurezza". Evita i doppioni con il catalogo formazione.

            Se ``tipo`` è dato, collega il corso alla qualifica (la qualifica è
            l'àncora): ``TrainingCourse.qualifica`` viene impostato una volta sola.
            """
            if canonical in corso_cache:
                c = corso_cache[canonical]
            else:
                q = reduce(or_, (Q(titolo__icontains=a) for a in aliases))
                c = (TrainingCourse.objects.filter(q)
                     .annotate(_n=Count("record_completamenti")).order_by("-_n", "id").first())
                if c is not None:
                    counters["corsi_match"] += 1
                else:
                    counters["corsi_nuovi"] += 1
                    if dry:
                        c = None
                    else:
                        _ensure_formazione_ctx()
                        codice = _short_code(canonical, prefix="C-", max_len=30)
                        while TrainingCourse.objects.filter(codice=codice).exists():
                            codice = _short_code(canonical + "_x", prefix="C-", max_len=30)
                        c = TrainingCourse.objects.create(
                            piano=formazione_ctx["piano"], categoria=formazione_ctx["categoria"],
                            codice=codice, titolo=canonical, durata_ore_teorica=Decimal("0"),
                            validita_mesi=validita or 0, obbligatorio=True,
                            stato="ATTIVO", is_active=True, created_by=assegnatore,
                        )
                corso_cache[canonical] = c
            # Collega corso → qualifica (una volta, idempotente).
            if c is not None and tipo is not None and not dry and c.qualifica_id is None:
                c.qualifica = tipo
                c.save(update_fields=["qualifica"])
                counters["link_corso_qualifica"] += 1
            return c

        def get_sessione(corso, data):
            if corso is None or data is None:
                return None
            key = (corso.pk, data)
            if key in sess_cache:
                return sess_cache[key]
            s = TrainingSession.objects.filter(corso=corso, data_inizio=data).first()
            if s is None and not dry:
                code = f"{corso.codice}-{data.strftime('%Y%m%d')}"[:40]
                base, n = code, 0
                while TrainingSession.objects.filter(codice_sessione=code).exists():
                    n += 1
                    code = f"{base[:36]}-{n}"
                s = TrainingSession.objects.create(
                    corso=corso, codice_sessione=code, stato="COMPLETATA",
                    modalita="ESTERNO", data_inizio=data, data_fine=data,
                )
            sess_cache[key] = s
            return s

        def register_completamento(lid, aliases, canonical, validita, data, ore=Decimal("0"), tipo=None):
            """Crea/riusa corso + sessione partecipata + iscrizione + record di
            completamento per (dipendente, corso, data). Idempotente.

            Collega il corso alla ``tipo`` qualifica (via ``get_corso``) e
            ritorna il ``TrainingEmployeeRecord`` (o ``None`` in dry-run / skip),
            così il chiamante può agganciarlo a ``DipendenteQualifica``.
            """
            if not data:
                return None
            corso = get_corso(aliases, canonical, validita, tipo=tipo)
            skey = (canonical, data)
            if skey not in seen_sess:
                seen_sess.add(skey)
                counters["sessioni"] += 1
            ckey = (lid, canonical, data)
            if ckey in seen_compl:
                return None
            seen_compl.add(ckey)
            counters["completamenti"] += 1
            if dry or corso is None:
                return None
            scad = _add_months(data, validita) if validita else None
            sess = get_sessione(corso, data)
            enroll, _ = TrainingEnrollment.objects.get_or_create(
                sessione=sess, legacy_anagrafica_id=lid,
                defaults={"stato": "COMPLETATO", "idoneo": True,
                          "ore_frequentate": ore, "data_completamento": data},
            )
            record, _ = TrainingEmployeeRecord.objects.get_or_create(
                corso=corso, legacy_anagrafica_id=lid, data_completamento=data,
                defaults={
                    "sessione": sess, "enrollment": enroll, "ore_frequentate": ore,
                    "idoneo": True, "data_scadenza": scad,
                    "validato_da": assegnatore, "validato_il": date.today(),
                    "course_code_snapshot": corso.codice[:30],
                    "course_title_snapshot": corso.titolo[:300],
                    "plan_code_snapshot": (corso.piano.codice if corso.piano_id else "")[:20],
                    "plan_name_snapshot": (corso.piano.nome if corso.piano_id else "")[:200],
                    "duration_hours_snapshot": corso.durata_ore_teorica,
                    "validity_months_snapshot": corso.validita_mesi,
                },
            )
            return record

        @transaction.atomic
        def run():
            # Abilitazioni dal foglio date (per nominativo → lid via CF)
            for r in drows[1:]:
                nome = _norm_nome(r[d_nome])
                if not nome:
                    continue
                rec = info.get(nome)
                if rec is None:
                    counters["unmatched_name"] += 1
                    continue
                lid = rec["lid"]
                if not lid:
                    counters["unmatched_cf"] += 1
                    continue
                counters["matched"] += 1
                for col, header in abil_cols.items():
                    data_cons = _parse_date(r[col])
                    if data_cons is None:
                        continue
                    aliases, canonical, durata = ABILITAZIONI[header]
                    # La qualifica è l'àncora: ottieni prima il tipo, poi il
                    # completamento (che collega corso→qualifica e ritorna il
                    # record), infine aggancia la qualifica del dipendente al record.
                    tipo = get_tipo(aliases, canonical, durata) if not skip_q else None
                    record = None
                    if not skip_c:
                        record = register_completamento(
                            lid, aliases, canonical, durata, data_cons, tipo=tipo,
                        )
                    if not skip_q:
                        counters["qualifiche"] += 1
                        if not dry and tipo is not None:
                            scad = _add_months(data_cons, durata) if durata else None
                            defaults = {"data_conseguimento": data_cons,
                                        "data_scadenza": scad, "assegnato_da": assegnatore}
                            if record is not None:
                                defaults["record_formazione"] = record
                            DipendenteQualifica.objects.update_or_create(
                                legacy_anagrafica_id=lid, tipo=tipo, defaults=defaults,
                            )

            # Corso lavoratori + livello rischio mansione (dal foglio principale)
            for nome, rec in info.items():
                lid = rec["lid"]
                if not lid:
                    continue
                if rec["corso"]:
                    ore_liv = Decimal({"A": 16, "M": 12, "B": 8}.get(rec["rischio"], 0))
                    tipo = (get_tipo(CORSO_LAVORATORI_ALIAS, CORSO_LAVORATORI_TIPO, CORSO_LAVORATORI_DURATA)
                            if not skip_q else None)
                    record = None
                    # Fase 2 — corso lavoratori come formazione + sessione partecipata
                    if not skip_c:
                        record = register_completamento(
                            lid, CORSO_LAVORATORI_ALIAS, CORSO_LAVORATORI_TIPO,
                            CORSO_LAVORATORI_DURATA, rec["corso"], ore=ore_liv, tipo=tipo,
                        )
                    # Fase 1 — qualifica "Formazione lavoratori (ASR)" agganciata al record
                    if not skip_q:
                        scad = None
                        try:
                            if rec["anno_scad"]:
                                scad = date(int(str(rec["anno_scad"])[:4]),
                                            rec["corso"].month, rec["corso"].day)
                        except (ValueError, TypeError):
                            scad = _add_months(rec["corso"], CORSO_LAVORATORI_DURATA)
                        if scad is None:
                            scad = _add_months(rec["corso"], CORSO_LAVORATORI_DURATA)
                        counters["corso_lavoratori"] += 1
                        if not dry and tipo is not None:
                            defaults = {"data_conseguimento": rec["corso"],
                                        "data_scadenza": scad, "assegnato_da": assegnatore}
                            if record is not None:
                                defaults["record_formazione"] = record
                            DipendenteQualifica.objects.update_or_create(
                                legacy_anagrafica_id=lid, tipo=tipo, defaults=defaults,
                            )
                # livello rischio sulla mansione del dipendente
                liv = {"A": Mansione.RISCHIO_ALTO, "B": Mansione.RISCHIO_BASSO,
                       "M": Mansione.RISCHIO_MEDIO}.get(rec["rischio"])
                man_nome = (lid_to_mansione.get(lid) or "").strip()
                if liv and man_nome:
                    mansioni_seen[man_nome.casefold()] = man_nome
                    if not dry:
                        man = Mansione.objects.filter(nome__iexact=man_nome).first()
                        if man and not man.livello_rischio:
                            man.livello_rischio = liv
                            man.save(update_fields=["livello_rischio"])
                            counters["mansioni_rischio"] += 1
            if dry:
                transaction.set_rollback(True)

        run()

        mode = "DRY-RUN (nessuna scrittura)" if dry else "COMMIT"
        self.stdout.write(self.style.WARNING(f"== Import ASR — {mode} =="))
        self.stdout.write(f"  Lavoratori nel foglio principale: {len(info)}")
        self.stdout.write(f"  Match (nome+CF): {counters['matched']} · "
                          f"nome non trovato: {counters['unmatched_name']} · "
                          f"CF non in anagrafica: {counters['unmatched_cf']}")
        if not skip_q:
            self.stdout.write(self.style.MIGRATE_HEADING("  -- Qualifiche (lato Salute e Sicurezza) --"))
            self.stdout.write(f"    Qualifiche/abilitazioni: {counters['qualifiche']}")
            self.stdout.write(f"    Corso lavoratori: {counters['corso_lavoratori']}")
        if not skip_c:
            self.stdout.write(self.style.MIGRATE_HEADING("  -- Formazione (modulo Corsi) --"))
            self.stdout.write(f"    Corsi riusati (match): {counters['corsi_match']} · "
                              f"nuovi: {counters['corsi_nuovi']}")
            self.stdout.write(f"    Sessioni partecipate: {counters['sessioni']}")
            self.stdout.write(f"    Completamenti (record): {counters['completamenti']}")
        self.stdout.write(f"  Legami corso -> qualifica creati: {counters['link_corso_qualifica']}")
        self.stdout.write(f"  Mansioni con livello rischio impostato: {counters['mansioni_rischio']}")
        if verbose:
            self.stdout.write(f"  Mansioni coinvolte: {sorted(mansioni_seen.values())}")
        if dry:
            self.stdout.write(self.style.NOTICE("  Rilancia con --commit per scrivere su DB."))
