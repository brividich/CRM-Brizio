"""Import del catalogo ruoli e delle assegnazioni dal gestionale storico.

Due file, due mestieri:

- ``roles.xlsx`` — il **catalogo**: una riga per ruolo (`Nome`, più
  `Descrizione` e `Scopo` dove ci sono). È l'elenco completo, anche dei ruoli
  che oggi nessuno ricopre.
- ``people-roles-summary.xlsx`` — le **assegnazioni**: una riga per
  persona×ruolo, con codice fiscale, `Tipologia associazione`
  (Principale/Secondario/Ad interim), `Data inizio`, `Data fine` e il
  `Ruolo responsabile`, da cui si ricava la gerarchia fra ruoli.

Scelte che vale la pena conoscere prima di lanciarlo:

**Le persone si riconoscono dal codice fiscale**, mai dal nome e mai dall'id:
gli id anagrafica non coincidono fra sviluppo e produzione, e due omonimi
esistono. Chi non ha un CF corrispondente in anagrafica viene elencato e
saltato — l'import non crea dipendenti.

**Il ruolo principale è uno solo.** La riga `Principale` diventa anche il
«Ruolo aziendale» della scheda; secondari e ad interim restano assegnazioni.
Se il gestionale dà a una persona due principali (succede: il file lo segnala
da sé) si tiene la più recente per data di inizio e si riporta l'anomalia.

**La gerarchia fra ruoli** si deduce da `Ruolo responsabile` sulle assegnazioni
in corso — dove «in corso» significa senza data di fine *oppure con una fine
futura*, perché il gestionale registra le uscite in anticipo — e si scrive
**solo dove è univoca**: un ruolo che secondo il file
riporta a due ruoli diversi resta senza `riporta_a` e finisce nel report. Un
ruolo che risulta riportare a sé stesso viene ignorato.

**Le assegnazioni già presenti non si toccano** se non per completarle: date e
tipologia si scrivono su quelle esistenti, nessuna assegnazione viene mai
cancellata. Rilanciare l'import non duplica nulla.

Uso:
    python manage.py import_ruoli_gestionale --roles roles.xlsx --people people-roles-summary.xlsx
    python manage.py import_ruoli_gestionale --roles ... --people ... --apply
    python manage.py import_ruoli_gestionale --roles ... --apply --solo-catalogo
"""
from __future__ import annotations

import datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from anagrafica.models import (
    DipendenteAnagraficaAziendale,
    DipendenteAnagraficaCivile,
    DipendenteRuoloOperativo,
    RuoloOperativo,
)

TIPOLOGIE = {
    "principale": DipendenteRuoloOperativo.TIPOLOGIA_PRINCIPALE,
    "secondario": DipendenteRuoloOperativo.TIPOLOGIA_SECONDARIO,
    "secondaria": DipendenteRuoloOperativo.TIPOLOGIA_SECONDARIO,
    "ad interim": DipendenteRuoloOperativo.TIPOLOGIA_AD_INTERIM,
}


def _testo(value) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _data(value) -> datetime.date | None:
    """Le date arrivano come datetime, come ``gg/mm/aaaa`` o come ISO."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    testo = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(testo, fmt).date()
        except ValueError:
            continue
    return None


def _in_corso(data_fine: datetime.date | None) -> bool:
    """Una data di fine futura non conclude nulla: il ruolo è ancora in essere.

    Il gestionale registra le uscite in anticipo — «Operatore CN5 fino al
    31/08» è una persona che oggi quel ruolo lo ricopre eccome.
    """
    return data_fine is None or data_fine >= timezone.localdate()


def _leggi(path: Path, foglio: str | None = None) -> tuple[list[str], list[tuple]]:
    from openpyxl import load_workbook

    if not path.exists():
        raise CommandError(f"File non trovato: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[foglio] if foglio else wb.worksheets[0]
    righe = list(ws.iter_rows(values_only=True))
    wb.close()
    if not righe:
        return [], []
    header = [_testo(c).casefold() for c in righe[0]]
    dati = [r for r in righe[1:] if any(c not in (None, "") for c in r)]
    return header, dati


class Command(BaseCommand):
    help = (
        "Importa il catalogo ruoli (roles.xlsx) e le assegnazioni persona×ruolo "
        "(people-roles-summary.xlsx) dal gestionale storico. Dry-run di default."
    )

    def add_arguments(self, parser):
        parser.add_argument("--roles", default=None, help="Percorso di roles.xlsx (catalogo).")
        parser.add_argument("--people", default=None, help="Percorso di people-roles-summary.xlsx.")
        parser.add_argument(
            "--apply", action="store_true",
            help="Scrive. Senza questo flag il comando non tocca nulla.",
        )
        parser.add_argument(
            "--solo-catalogo", action="store_true",
            help="Importa solo i ruoli, senza assegnazioni né gerarchia.",
        )
        parser.add_argument(
            "--rimuovi-estranee", action="store_true",
            help=(
                "Con --apply, rimuove le assegnazioni presenti nell'HUB che il file "
                "del gestionale non conferma (ruoli nati da prove o da assegnazioni "
                "manuali). Senza il flag vengono soltanto elencate."
            ),
        )
        parser.add_argument(
            "--no-gerarchia", action="store_true",
            help="Non deduce «riporta a» dal file: la gerarchia resta quella già impostata.",
        )

    # ------------------------------------------------------------------ setup
    def handle(self, *args, **options):
        self.apply = bool(options["apply"])
        roles_path = options["roles"]
        people_path = options["people"]
        if not roles_path and not people_path:
            raise CommandError("Serve almeno --roles o --people.")

        self.catalogo_nuovi: list[str] = []
        self.catalogo_aggiornati: list[str] = []

        with transaction.atomic():
            if roles_path:
                self._importa_catalogo(Path(roles_path))
            if people_path and not options["solo_catalogo"]:
                self._importa_assegnazioni(
                    Path(people_path),
                    gerarchia=not options["no_gerarchia"],
                    rimuovi_estranee=bool(options["rimuovi_estranee"]),
                )
            if not self.apply:
                transaction.set_rollback(True)

        self.stdout.write("")
        if self.apply:
            self.stdout.write(self.style.SUCCESS("Import applicato."))
        else:
            self.stdout.write(self.style.WARNING(
                "Nessuna modifica salvata: rilancia con --apply per scrivere."
            ))

    # --------------------------------------------------------------- catalogo
    def _importa_catalogo(self, path: Path) -> None:
        header, dati = _leggi(path)
        try:
            i_nome = header.index("nome")
        except ValueError:
            raise CommandError(f"{path.name}: manca la colonna «Nome».")
        i_desc = header.index("descrizione") if "descrizione" in header else None
        i_scopo = header.index("scopo") if "scopo" in header else None

        esistenti = {r.nome.strip().casefold(): r for r in RuoloOperativo.objects.all()}
        troppo_lunghi: list[str] = []

        for riga in dati:
            nome = _testo(riga[i_nome])
            if not nome:
                continue
            if len(nome) > 100:
                troppo_lunghi.append(nome)
                nome = nome[:100]
            descrizione = _testo(riga[i_desc]) if i_desc is not None else ""
            scopo = _testo(riga[i_scopo]) if i_scopo is not None else ""
            testo = "\n\n".join(p for p in [descrizione, f"Scopo: {scopo}" if scopo else ""] if p)

            ruolo = esistenti.get(nome.casefold())
            if ruolo is None:
                ruolo = RuoloOperativo(nome=nome, descrizione=testo)
                ruolo.save()
                esistenti[nome.casefold()] = ruolo
                self.catalogo_nuovi.append(nome)
            elif testo and not (ruolo.descrizione or "").strip():
                # La descrizione già scritta nel portale vince: il gestionale
                # ne ha appena quattro, non deve sovrascrivere il lavoro fatto.
                ruolo.descrizione = testo
                ruolo.save(update_fields=["descrizione"])
                self.catalogo_aggiornati.append(nome)

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(f"Catalogo ruoli — {path.name}"))
        self.stdout.write(f"  righe lette          : {len(dati)}")
        self.stdout.write(f"  ruoli nuovi          : {len(self.catalogo_nuovi)}")
        self.stdout.write(f"  descrizioni aggiunte : {len(self.catalogo_aggiornati)}")
        self.stdout.write(f"  già in catalogo      : {len(dati) - len(self.catalogo_nuovi)}")
        if troppo_lunghi:
            self.stdout.write(self.style.WARNING(
                f"  nomi troncati a 100 caratteri: {len(troppo_lunghi)}"
            ))

    # ----------------------------------------------------------- assegnazioni
    def _importa_assegnazioni(self, path: Path, *, gerarchia: bool, rimuovi_estranee: bool = False) -> None:
        header, dati = _leggi(path)
        def col(nome: str, obbligatoria: bool = True) -> int | None:
            if nome in header:
                return header.index(nome)
            if obbligatoria:
                raise CommandError(f"{path.name}: manca la colonna «{nome}».")
            return None

        i_cf = col("codice fiscale")
        i_ruolo = col("ruolo")
        i_tipo = col("tipologia associazione", False)
        i_inizio = col("data inizio", False)
        i_fine = col("data fine", False)
        i_ruolo_resp = col("ruolo responsabile", False)
        i_dip = col("dipendente", False)

        cf_map = {
            (c.codice_fiscale or "").strip().upper(): c.legacy_anagrafica_id
            for c in DipendenteAnagraficaCivile.objects.exclude(codice_fiscale="")
        }
        catalogo = {r.nome.strip().casefold(): r for r in RuoloOperativo.objects.all()}

        creati = completati = 0
        viste: set[tuple[int, int]] = set()
        persone_ignote: set[str] = set()
        ruoli_ignoti: set[str] = set()
        principali: dict[int, tuple[str, datetime.date | None]] = {}
        doppi_principali: set[int] = set()
        supervisori: dict[str, set[str]] = {}

        for riga in dati:
            cf = _testo(riga[i_cf]).upper()
            nome_ruolo = _testo(riga[i_ruolo])
            if not cf or not nome_ruolo:
                continue

            ruolo = catalogo.get(nome_ruolo.casefold())
            if ruolo is None:
                ruoli_ignoti.add(nome_ruolo)
                continue

            data_inizio = _data(riga[i_inizio]) if i_inizio is not None else None
            data_fine = _data(riga[i_fine]) if i_fine is not None else None
            tipologia = ""
            if i_tipo is not None:
                tipologia = TIPOLOGIE.get(_testo(riga[i_tipo]).casefold(), "")

            # La gerarchia guarda solo i ruoli in essere: un riporto chiuso nel
            # 2019 non descrive l'organizzazione di oggi.
            if gerarchia and i_ruolo_resp is not None and _in_corso(data_fine):
                capo = _testo(riga[i_ruolo_resp])
                if capo and capo.casefold() != nome_ruolo.casefold():
                    supervisori.setdefault(nome_ruolo.casefold(), set()).add(capo)

            legacy_id = cf_map.get(cf)
            if legacy_id is not None:
                viste.add((legacy_id, ruolo.pk))
            if legacy_id is None:
                persone_ignote.add(_testo(riga[i_dip]) if i_dip is not None else cf)
                continue

            assegnazione, creata = DipendenteRuoloOperativo.objects.get_or_create(
                legacy_anagrafica_id=legacy_id, ruolo=ruolo,
            )
            campi = []
            for attr, valore in (
                ("tipologia", tipologia),
                ("data_inizio", data_inizio),
                ("data_fine", data_fine),
            ):
                if valore and getattr(assegnazione, attr) != valore:
                    setattr(assegnazione, attr, valore)
                    campi.append(attr)
            if campi:
                assegnazione.save(update_fields=campi)
            creati += int(creata)
            completati += int(bool(campi) and not creata)

            if tipologia == DipendenteRuoloOperativo.TIPOLOGIA_PRINCIPALE and _in_corso(data_fine):
                precedente = principali.get(legacy_id)
                if precedente is None:
                    principali[legacy_id] = (ruolo.nome, data_inizio)
                else:
                    doppi_principali.add(legacy_id)
                    # A parità di anomalia vince l'inizio più recente.
                    if data_inizio and (precedente[1] is None or data_inizio > precedente[1]):
                        principali[legacy_id] = (ruolo.nome, data_inizio)

        # Ruolo principale → «Ruolo aziendale» della scheda.
        scritti_principali = 0
        for legacy_id, (nome, _inizio) in principali.items():
            az, _ = DipendenteAnagraficaAziendale.objects.get_or_create(
                legacy_anagrafica_id=legacy_id,
            )
            if (az.ruolo_aziendale or "").strip().casefold() != nome.casefold():
                az.ruolo_aziendale = nome[:200]
                az.save(update_fields=["ruolo_aziendale", "updated_at"])
                scritti_principali += 1

        # Assegnazioni che l'HUB ha e il gestionale no: nate da prove o da
        # inserimenti a mano. Non si cancellano di default — potrebbero essere
        # legittime e più aggiornate del file.
        persone_del_file = {lid for lid, _ in viste}
        estranee = [
            a for a in DipendenteRuoloOperativo.objects
            .filter(legacy_anagrafica_id__in=persone_del_file)
            .select_related("ruolo")
            if (a.legacy_anagrafica_id, a.ruolo_id) not in viste
        ]
        rimosse = 0
        if estranee and rimuovi_estranee:
            rimosse = len(estranee)
            DipendenteRuoloOperativo.objects.filter(
                pk__in=[a.pk for a in estranee]
            ).delete()

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(f"Assegnazioni — {path.name}"))
        self.stdout.write(f"  righe lette              : {len(dati)}")
        self.stdout.write(f"  assegnazioni create      : {creati}")
        self.stdout.write(f"  assegnazioni completate  : {completati}  (date/tipologia su righe già esistenti)")
        self.stdout.write(f"  ruoli principali scritti : {scritti_principali}")
        if doppi_principali:
            self.stdout.write(self.style.WARNING(
                f"  persone con PIÙ ruoli principali: {len(doppi_principali)} "
                "(tenuto il più recente per data di inizio)"
            ))
        if estranee:
            testo = "rimosse" if rimuovi_estranee else "NON confermate dal file (solo elencate)"
            self.stdout.write(self.style.WARNING(
                f"  assegnazioni nell'HUB {testo}: {len(estranee)}"
            ))
            for a in estranee:
                self.stdout.write(f"    - [{a.legacy_anagrafica_id}] {a.ruolo.nome}")
            if not rimuovi_estranee:
                self.stdout.write("    (per toglierle: --rimuovi-estranee)")
        if persone_ignote:
            self.stdout.write(self.style.WARNING(
                f"  persone senza corrispondenza in anagrafica (saltate): {len(persone_ignote)}"
            ))
            for p in sorted(persone_ignote):
                self.stdout.write(f"    - {p}")
        if ruoli_ignoti:
            self.stdout.write(self.style.WARNING(
                f"  ruoli non in catalogo (saltati): {len(ruoli_ignoti)}"
            ))
            for r in sorted(ruoli_ignoti):
                self.stdout.write(f"    - {r}")

        if gerarchia:
            self._scrivi_gerarchia(supervisori, catalogo)

    # ---------------------------------------------------------------- riporti
    def _scrivi_gerarchia(self, supervisori: dict[str, set[str]], catalogo: dict) -> None:
        from anagrafica.views import _riporta_a_valido

        scritti = 0
        conflitti: list[tuple[str, list[str]]] = []
        capi_ignoti: set[str] = set()
        cicli: list[str] = []

        for nome_ruolo, capi in supervisori.items():
            ruolo = catalogo.get(nome_ruolo)
            if ruolo is None:
                continue
            if len(capi) > 1:
                conflitti.append((ruolo.nome, sorted(capi)))
                continue
            capo_nome = next(iter(capi))
            capo = catalogo.get(capo_nome.casefold())
            if capo is None:
                capi_ignoti.add(capo_nome)
                continue
            if ruolo.riporta_a_id == capo.pk:
                continue
            if not _riporta_a_valido(ruolo.pk, capo.pk):
                cicli.append(f"{ruolo.nome} → {capo.nome}")
                continue
            ruolo.riporta_a = capo
            ruolo.save(update_fields=["riporta_a"])
            scritti += 1

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Gerarchia fra ruoli («riporta a»)"))
        self.stdout.write(f"  riporti scritti : {scritti}")
        if conflitti:
            self.stdout.write(self.style.WARNING(
                f"  ruoli con più responsabili nel file (lasciati vuoti): {len(conflitti)}"
            ))
            for nome, capi in sorted(conflitti):
                self.stdout.write(f"    - {nome} → {', '.join(capi)}")
        if capi_ignoti:
            self.stdout.write(self.style.WARNING(
                f"  ruoli-responsabile non in catalogo: {len(capi_ignoti)}"
            ))
        if cicli:
            self.stdout.write(self.style.WARNING(
                f"  riporti che avrebbero creato un ciclo (ignorati): {', '.join(cicli)}"
            ))
