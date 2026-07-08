"""Import anagrafica dipendenti da file Excel (`.xlsx`).

Formato atteso: il foglio principale ha l'header sulla prima riga con le colonne
documentate nel modello-tipo `people.xlsx` (Cognome, Nome, E-mail, Codice fiscale,
Matricola, Badge, Data di nascita, residenza, IBAN, taglie, consenso privacy,
grado di istruzione, ecc.).

Le righe con `Data cessazione` valorizzata vengono importate come ex dipendenti:
restano a sistema con il fascicolo completo ma vengono marcate non attive e
l'eventuale account portale viene scollegato. Non compaiono nella lista
dipendenti in forza (vedi `anagrafica:ex_dipendenti_list`).

Esempi:
    python manage.py import_dipendenti_xlsx doc/people.xlsx --dry-run
    python manage.py import_dipendenti_xlsx doc/people.xlsx --update-existing
    python manage.py import_dipendenti_xlsx doc/people.xlsx --sheet Worksheet --limit 5
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

from anagrafica.models import (
    DipendenteAnagraficaAziendale,
    DipendenteAnagraficaCivile,
    Reparto,
)
from core.legacy_anagrafica import (
    fetch_anagrafica_rows,
    generate_username,
    upsert_anagrafica_dipendente,
)
from werkzeug.security import generate_password_hash
from core.legacy_models import AnagraficaDipendente, Ruolo, UtenteLegacy


# ---------------------------------------------------------------------------
# Mapping costanti
# ---------------------------------------------------------------------------

GENERE_MAP = {
    "maschio": DipendenteAnagraficaCivile.GENERE_M,
    "femmina": DipendenteAnagraficaCivile.GENERE_F,
    "m": DipendenteAnagraficaCivile.GENERE_M,
    "f": DipendenteAnagraficaCivile.GENERE_F,
}

TITOLO_MAP = {
    "istruzione-di-primo-grado": DipendenteAnagraficaCivile.TITOLO_PRIMO_GRADO,
    "istruzione-di-secondo-grado": DipendenteAnagraficaCivile.TITOLO_SECONDO_GRADO,
    "laurea-triennale": DipendenteAnagraficaCivile.TITOLO_LAUREA_TRIENNALE,
    "laurea-magistrale": DipendenteAnagraficaCivile.TITOLO_LAUREA_MAGISTRALE,
    "laurea-ciclo-unico": DipendenteAnagraficaCivile.TITOLO_LAUREA_CICLO_UNICO,
    "laurea-magistrale-ciclo-unico": DipendenteAnagraficaCivile.TITOLO_LAUREA_CICLO_UNICO,
}

RAPPORTO_MAP = {
    "dipendente": DipendenteAnagraficaAziendale.CONTRATTO_INDETERMINATO,
    "somministrato": DipendenteAnagraficaAziendale.CONTRATTO_SOMMINISTRAZIONE,
    "apprendista": DipendenteAnagraficaAziendale.CONTRATTO_APPRENDISTATO,
    "stage": DipendenteAnagraficaAziendale.CONTRATTO_STAGE,
    "tirocinio": DipendenteAnagraficaAziendale.CONTRATTO_STAGE,
    "collaboratore": DipendenteAnagraficaAziendale.CONTRATTO_COLLABORAZIONE,
    "determinato": DipendenteAnagraficaAziendale.CONTRATTO_DETERMINATO,
    "indeterminato": DipendenteAnagraficaAziendale.CONTRATTO_INDETERMINATO,
}

CONSENSO_TRUE = {"si", "sì", "yes", "y", "true", "1"}
CONSENSO_FALSE = {"no", "n", "false", "0"}


# ---------------------------------------------------------------------------
# Normalizzatori
# ---------------------------------------------------------------------------

def _txt(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_badge(value) -> str:
    v = _txt(value)
    if v:
        stripped = v.lstrip("0")
        return stripped if stripped else "0"
    return v


def _norm_telefono(value) -> str:
    raw = _txt(value)
    if not raw:
        return ""
    raw = re.sub(r"\s+", "", raw)
    raw = re.sub(r"^\++", "+", raw)
    return raw


def _norm_iban(value) -> str:
    raw = _txt(value)
    if not raw:
        return ""
    return re.sub(r"\s+", "", raw).upper()


def _norm_cf(value) -> str:
    raw = _txt(value)
    return re.sub(r"\s+", "", raw).upper()


def _norm_genere(value) -> str:
    raw = _txt(value).lower()
    return GENERE_MAP.get(raw, "")


def _norm_titolo(value) -> str:
    raw = _txt(value).lower()
    return TITOLO_MAP.get(raw, "")


def _norm_consenso(value) -> bool:
    raw = _txt(value).lower()
    if raw in CONSENSO_TRUE:
        return True
    if raw in CONSENSO_FALSE:
        return False
    return False


def _norm_bool_categoria(value) -> bool:
    raw = _txt(value).lower()
    if not raw:
        return False
    return raw in CONSENSO_TRUE


def _norm_tipologia(value) -> str:
    raw = _txt(value).lower()
    return RAPPORTO_MAP.get(raw, "")


def _norm_matricola(value) -> str:
    raw = _txt(value)
    return raw.upper()


def _norm_taglia_maglia(value) -> str:
    raw = _txt(value).upper()
    valid = {choice[0] for choice in DipendenteAnagraficaAziendale.TAGLIA_MAGLIA_CHOICES}
    return raw if raw in valid else ""


def _norm_intestatario(value) -> str:
    raw = _txt(value)
    return re.sub(r"\s+", " ", raw)


def _parse_date(value, *, formats=("%d/%m/%Y", "%Y-%m-%d")) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _compose_indirizzo(indirizzo, civico) -> str:
    base = _txt(indirizzo)
    civ = _txt(civico)
    if base and civ:
        return f"{base}, {civ}"
    return base or civ


def _compose_luogo_nascita(luogo, provincia) -> str:
    base = _txt(luogo)
    prov = _txt(provincia)
    if base and prov and prov.upper() not in base.upper():
        return f"{base} ({prov})"
    return base


# ---------------------------------------------------------------------------
# Lettura header → indice
# ---------------------------------------------------------------------------

def _normalize_header(raw: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(raw or "").lower())


HEADER_ALIAS = {
    "cognome": "cognome",
    "nome": "nome",
    "email": "email",
    "emailpersonale": "email_personale",
    "numeroditelefono": "telefono",
    "telefono": "telefono",
    "codicefiscale": "codice_fiscale",
    "matricola": "matricola",
    "badge": "badge",
    "datadinascita": "data_nascita",
    "paesedinascita": "luogo_nascita",
    "comunedinascita": "luogo_nascita",
    "luogodinascita": "luogo_nascita",
    "provinciadinascita": "provincia_nascita",
    "genere": "genere",
    "nazionalit": "nazionalita",
    "nazionalita": "nazionalita",
    "indirizzodiresidenza": "indirizzo_res",
    "civicodiresidenza": "civico_res",
    "capdiresidenza": "cap_res",
    "comunediresidenza": "comune_res",
    "provinciadiresidenza": "provincia_res",
    "nazionediresidenza": "nazione_res",
    "datadiassunzione": "data_assunzione",
    "dataprimaassunzione": "data_prima_assunzione",
    "datacessazione": "data_cessazione",
    "rapportodilavoro": "rapporto_lavoro",
    "numerodiscarpe": "scarpe",
    "tagliamaglia": "taglia_maglia",
    "tagliapantalone": "taglia_pantalone",
    "iban": "iban",
    "banca": "banca",
    "titolarecontocorente": "intestatario",
    "titolarecontocorrente": "intestatario",
    "consensoprivacy": "consenso",
    # nota: gli header reali tipo "data-del-rilascio-o-del-divieto" sono
    # intercettati dal prefisso "datadelrilascio" in _build_header_index.
    "gradodiistruzione": "titolo_studio",
    "categoriaprotetta": "categoria_protetta",
    "categoriadisabili": "categoria_disabili",
    "percentualedisabilita": "percentuale_disabilita",
    "reparto": "reparto",
    "repartodiappartenenza": "reparto",
    "unitaorganizzativa": "reparto",
}


def _build_header_index(header_row) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        key = _normalize_header(str(raw))
        # match diretto
        if key in HEADER_ALIAS:
            out[HEADER_ALIAS[key]] = idx
            continue
        # match "starts-with" per la colonna data-del-rilascio-...
        if key.startswith("datadelrilascio") or key.startswith("datarilascio"):
            out["data_consenso"] = idx
            continue
    return out


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa anagrafica dipendenti da file Excel (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Percorso al file .xlsx")
        parser.add_argument("--sheet", type=str, default="", help="Nome foglio (default: primo)")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla, mostra solo il report")
        parser.add_argument("--update-existing", action="store_true", help="Aggiorna i record già presenti (default: solo create)")
        parser.add_argument("--limit", type=int, default=0, help="Limita N righe (0 = tutte)")
        parser.add_argument("--verbose-errors", action="store_true", help="Stampa traceback completo sugli errori per riga")

    def handle(self, *args, **opts):
        try:
            import openpyxl  # noqa: WPS433
        except ImportError as exc:
            raise CommandError("openpyxl non installato. pip install openpyxl") from exc

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet_name = opts.get("sheet") or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' non presente. Disponibili: {wb.sheetnames}")
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise CommandError("Foglio vuoto") from exc

        idx = _build_header_index(header_row)
        missing_required = [k for k in ("cognome", "nome", "codice_fiscale") if k not in idx]
        if missing_required:
            raise CommandError(
                f"Colonne obbligatorie mancanti: {missing_required}. Header letto: {list(header_row)}"
            )

        dry_run = bool(opts.get("dry_run"))
        update_existing = bool(opts.get("update_existing"))
        limit = int(opts.get("limit") or 0)
        verbose_errors = bool(opts.get("verbose_errors"))

        self.stdout.write(
            self.style.NOTICE(
                f"Import da {path.name} (sheet={sheet_name}, dry_run={dry_run}, "
                f"update_existing={update_existing}, limit={limit or 'all'})"
            )
        )

        stats = {
            "letti": 0,
            "creati_civile": 0,
            "creati_aziendale": 0,
            "aggiornati_civile": 0,
            "aggiornati_aziendale": 0,
            "match_per_cf": 0,
            "match_per_email": 0,
            "match_per_nome": 0,
            "cessati": 0,
            "saltati": 0,
            "errori": 0,
            "creati_utenti_portale": 0,
            "aggiornati_reparto": 0,
        }
        errors: list[str] = []

        # CF → legacy_anagrafica_id (via DipendenteAnagraficaCivile)
        cf_to_legacy: dict[str, int] = {
            cf: lid
            for cf, lid in DipendenteAnagraficaCivile.objects
            .exclude(codice_fiscale="")
            .values_list("codice_fiscale", "legacy_anagrafica_id")
        }

        # email → legacy_anagrafica_id (via tabella legacy)
        email_to_legacy: dict[str, int] = {}
        for row in fetch_anagrafica_rows():
            em = _txt(row.get("email")).lower()
            if em and em not in email_to_legacy:
                email_to_legacy[em] = int(row.get("id") or 0)

        # username già presenti (per generazione senza collisioni)
        self._existing_usernames: set = set(
            AnagraficaDipendente.objects
            .exclude(aliasusername="")
            .exclude(aliasusername__isnull=True)
            .values_list("aliasusername", flat=True)
        )
        # id → aliasusername già assegnato
        self._legacy_id_to_alias: dict = dict(
            AnagraficaDipendente.objects
            .exclude(aliasusername="")
            .values_list("id", "aliasusername")
        )

        def cell(row, key):
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        class _DryRunAbort(Exception):
            pass

        try:
            with transaction.atomic():
                self._run_rows(
                    rows_iter=rows_iter,
                    cell=cell,
                    cf_to_legacy=cf_to_legacy,
                    email_to_legacy=email_to_legacy,
                    limit=limit,
                    update_existing=update_existing,
                    stats=stats,
                    errors=errors,
                    verbose_errors=verbose_errors,
                )
                if dry_run:
                    raise _DryRunAbort()
        except _DryRunAbort:
            pass

        # --- Report ---
        self._print_report(stats, errors, dry_run=dry_run)

    # ------------------------------------------------------------------
    # Loop principale
    # ------------------------------------------------------------------

    def _run_rows(
        self,
        *,
        rows_iter,
        cell,
        cf_to_legacy,
        email_to_legacy,
        limit,
        update_existing,
        stats,
        errors,
        verbose_errors,
    ):
        for n, raw_row in enumerate(rows_iter, start=2):
            if limit and stats["letti"] >= limit:
                break
            if raw_row is None or all(c in (None, "") for c in raw_row):
                continue
            stats["letti"] += 1

            try:
                with transaction.atomic():
                    self._process_row(
                        raw_row=raw_row,
                        cell=cell,
                        cf_to_legacy=cf_to_legacy,
                        email_to_legacy=email_to_legacy,
                        update_existing=update_existing,
                        stats=stats,
                        errors=errors,
                        line=n,
                    )
            except Exception as exc:  # noqa: BLE001
                stats["errori"] += 1
                msg = f"riga {n}: {exc.__class__.__name__}: {exc}"
                if verbose_errors:
                    import traceback
                    msg += "\n" + traceback.format_exc()
                errors.append(msg)

    def _process_row(
        self,
        *,
        raw_row,
        cell,
        cf_to_legacy,
        email_to_legacy,
        update_existing,
        stats,
        errors,
        line,
    ):
        cognome = _txt(cell(raw_row, "cognome"))
        nome = _txt(cell(raw_row, "nome"))
        cf = _norm_cf(cell(raw_row, "codice_fiscale"))
        email = _txt(cell(raw_row, "email")).lower()

        if not (cognome and nome and cf):
            stats["saltati"] += 1
            errors.append(f"riga {line}: campi obbligatori mancanti (cognome/nome/CF)")
            return

        # 1) Trova / crea record legacy
        legacy_id = cf_to_legacy.get(cf) or email_to_legacy.get(email)
        matched_by = ""
        if legacy_id:
            matched_by = "cf" if cf_to_legacy.get(cf) else "email"

        # Genera username solo se il record esistente non ne ha già uno
        current_alias = self._legacy_id_to_alias.get(legacy_id, "") if legacy_id else ""
        if current_alias:
            alias_to_use = current_alias
        else:
            alias_to_use = generate_username(nome, cognome, self._existing_usernames)
            if alias_to_use:
                self._existing_usernames.add(alias_to_use)

        # Se il dipendente ha una data di cessazione è un ex dipendente:
        # va marcato non attivo e l'account portale va scollegato (utente_id
        # azzerato), coerentemente con `dipendente_toggle_active`.
        is_cessato = _parse_date(cell(raw_row, "data_cessazione")) is not None
        legacy_row = upsert_anagrafica_dipendente(
            row_id=legacy_id,
            nome=nome,
            cognome=cognome,
            email=email,
            aliasusername=alias_to_use,
            matricola=_norm_matricola(cell(raw_row, "matricola")),
            attivo=not is_cessato,
            detach_account=is_cessato,
        )
        new_legacy_id = int(legacy_row["id"])
        if is_cessato:
            stats["cessati"] += 1
        if not matched_by:
            matched_by = "nome"
        if matched_by == "cf":
            stats["match_per_cf"] += 1
        elif matched_by == "email":
            stats["match_per_email"] += 1
        else:
            stats["match_per_nome"] += 1

        if cf:
            cf_to_legacy[cf] = new_legacy_id
        if email:
            email_to_legacy[email] = new_legacy_id

        # 2) Anagrafica civile
        civile, created_civ = DipendenteAnagraficaCivile.objects.get_or_create(
            legacy_anagrafica_id=new_legacy_id
        )
        civ_fields = self._build_civile_payload(raw_row, cell, cf)
        if created_civ:
            for k, v in civ_fields.items():
                setattr(civile, k, v)
            civile.save()
            stats["creati_civile"] += 1
        elif update_existing:
            changed = False
            for k, v in civ_fields.items():
                if v in (None, "") and getattr(civile, k):
                    continue
                if getattr(civile, k) != v:
                    setattr(civile, k, v)
                    changed = True
            if changed:
                civile.save()
                stats["aggiornati_civile"] += 1

        # 3) Anagrafica aziendale
        aziendale, created_az = DipendenteAnagraficaAziendale.objects.get_or_create(
            legacy_anagrafica_id=new_legacy_id
        )
        az_fields = self._build_aziendale_payload(raw_row, cell, email)
        if created_az:
            for k, v in az_fields.items():
                setattr(aziendale, k, v)
            aziendale.save()
            stats["creati_aziendale"] += 1
        elif update_existing:
            changed = False
            for k, v in az_fields.items():
                if v in (None, "") and getattr(aziendale, k):
                    continue
                if getattr(aziendale, k) != v:
                    setattr(aziendale, k, v)
                    changed = True
            if changed:
                aziendale.save()
                stats["aggiornati_aziendale"] += 1

        # 3b) Sync reparto → caporeparto (l'area aziendale non si autopopola
        # più: con l'inversione della gerarchia un Reparto può avere più
        # Aree aziendali figlie, quindi non è più derivabile da un singolo
        # Reparto — vedi Fase 2 nello spec).
        reparto_nome = _txt(cell(raw_row, "reparto"))
        if reparto_nome and (created_az or update_existing):
            rep = Reparto.objects.filter(nome__iexact=reparto_nome, is_active=True).first()
            capo_id = rep.caporeparto_legacy_id if rep else None
            update_fields = []
            if aziendale.area != reparto_nome:
                aziendale.area = reparto_nome
                update_fields.append("area")
            if aziendale.caporeparto_legacy_id != capo_id:
                aziendale.caporeparto_legacy_id = capo_id
                update_fields.append("caporeparto_legacy_id")
            if update_fields:
                aziendale.save(update_fields=update_fields)
                stats["aggiornati_reparto"] += 1

        # 4) Account portale: collega esistente o crea locale (solo dipendenti attivi non già collegati)
        if alias_to_use and not is_cessato and not legacy_row.get("utente_id"):
            existing_utente = UtenteLegacy.objects.filter(
                Q(email__iexact=alias_to_use) | Q(email__istartswith=f"{alias_to_use}@")
            ).order_by("id").first()
            if existing_utente is not None:
                # Utente LDAP o locale già esistente: collega senza creare
                AnagraficaDipendente.objects.filter(id=new_legacy_id).update(utente_id=existing_utente.id)
            else:
                ruolo_utente = Ruolo.objects.filter(nome__iexact="utente").first()
                data_nascita = _parse_date(cell(raw_row, "data_nascita"))
                anno_nascita = data_nascita.year if data_nascita else None
                pwd_iniziale = str(anno_nascita) if anno_nascita else f"Portale{date.today().year}"
                nome_completo = f"{cognome} {nome}".strip()
                nuovo_utente = UtenteLegacy.objects.create(
                    nome=nome_completo,
                    email=alias_to_use,
                    password=generate_password_hash(pwd_iniziale),
                    ruolo="utente",
                    ruolo_id=ruolo_utente.id if ruolo_utente else None,
                    attivo=True,
                    deve_cambiare_password=True,
                )
                AnagraficaDipendente.objects.filter(id=new_legacy_id).update(utente_id=nuovo_utente.id)
                stats["creati_utenti_portale"] += 1

    # ------------------------------------------------------------------
    # Report
    # ------------------------------------------------------------------

    def _print_report(self, stats, errors, *, dry_run):
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== Report import ==="))
        for key, val in stats.items():
            self.stdout.write(f"  {key:25s} {val}")

        if errors:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING(f"=== Errori / scartati ({len(errors)}) ==="))
            for e in errors[:50]:
                self.stdout.write(f"  - {e}")
            if len(errors) > 50:
                self.stdout.write(f"  ... e altri {len(errors) - 50}")

        if dry_run:
            self.stdout.write("")
            self.stdout.write(self.style.NOTICE("DRY-RUN: nessuna scrittura effettuata."))

    # ------------------------------------------------------------------
    # Payload builders
    # ------------------------------------------------------------------

    def _build_civile_payload(self, raw_row, cell, cf) -> dict:
        return {
            "codice_fiscale": cf,
            "data_nascita": _parse_date(cell(raw_row, "data_nascita")),
            "luogo_nascita": _compose_luogo_nascita(
                cell(raw_row, "luogo_nascita"),
                cell(raw_row, "provincia_nascita"),
            ),
            "provincia_nascita": _txt(cell(raw_row, "provincia_nascita")).upper(),
            "nazionalita": _txt(cell(raw_row, "nazionalita")).title(),
            "genere": _norm_genere(cell(raw_row, "genere")),
            "indirizzo_residenza": _compose_indirizzo(
                cell(raw_row, "indirizzo_res"),
                cell(raw_row, "civico_res"),
            ),
            "citta_residenza": _txt(cell(raw_row, "comune_res")),
            "provincia_residenza": _txt(cell(raw_row, "provincia_res")).upper()[:5],
            "nazione_residenza": _txt(cell(raw_row, "nazione_res")).title() or "Italia",
            "cap_residenza": _txt(cell(raw_row, "cap_res")),
            "email_privata": _txt(cell(raw_row, "email_personale")).lower(),
            "telefono_privato": _norm_telefono(cell(raw_row, "telefono")),
            "titolo_studio": _norm_titolo(cell(raw_row, "titolo_studio")),
            "nome_banca": _txt(cell(raw_row, "banca")),
            "iban": _norm_iban(cell(raw_row, "iban")),
            "intestatario_conto": _norm_intestatario(cell(raw_row, "intestatario")),
            "categoria_protetta": _norm_bool_categoria(cell(raw_row, "categoria_protetta")),
            "categoria_disabili": _norm_bool_categoria(cell(raw_row, "categoria_disabili")),
            "percentuale_disabilita": _parse_decimal(cell(raw_row, "percentuale_disabilita")),
        }

    def _build_aziendale_payload(self, raw_row, cell, email) -> dict:
        return {
            "badge": _norm_badge(cell(raw_row, "badge")),
            "taglia_scarpe": _txt(cell(raw_row, "scarpe")),
            "taglia_pantalone": _txt(cell(raw_row, "taglia_pantalone")),
            "taglia_maglia": _norm_taglia_maglia(cell(raw_row, "taglia_maglia")),
            "consenso_privacy": _norm_consenso(cell(raw_row, "consenso")),
            "data_consenso_privacy": _parse_date(cell(raw_row, "data_consenso")),
            "data_prima_assunzione": _parse_date(cell(raw_row, "data_prima_assunzione")),
            "data_assunzione_ultima": _parse_date(cell(raw_row, "data_assunzione")),
            "data_cessazione": _parse_date(cell(raw_row, "data_cessazione")),
            "tipologia_contratto": _norm_tipologia(cell(raw_row, "rapporto_lavoro")),
            "email_aziendale": email if email.endswith("costruzioninovicrom.it") else "",
        }
