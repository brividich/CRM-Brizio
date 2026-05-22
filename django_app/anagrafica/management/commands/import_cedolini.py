"""Import mensile saldi ore (ferie/ROL/ex-festività) da file cedolini XLSX.

Legge il foglio "Dati" del file cedolini e popola ImportazioneCedolini +
SaldoCedolino con upsert su (tax_code, data_competenza).

Esempi:
    python manage.py import_cedolini cedolini_maggio2026.xlsx
    python manage.py import_cedolini cedolini_maggio2026.xlsx --dry-run
    python manage.py import_cedolini cedolini_maggio2026.xlsx --foglio Dati --verbose
"""

from __future__ import annotations

import calendar
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from anagrafica.models import (
    DipendenteAnagraficaCivile,
    ImportazioneCedolini,
    SaldoCedolino,
)

User = get_user_model()

# Mappa nome mese italiano → numero mese
MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}

# Indici colonna (0-based) del foglio "Dati"
COL_CF        = 0   # Codice Fiscale
COL_MESE      = 3   # Mese (stringa italiana)
COL_ANNO      = 4   # Anno (intero)
COL_DATA_PER  = 5   # Data Periodo (datetime — ultimo gg mese)
COL_ANZ_ANNI  = 8   # Anzianità Anni
COL_ANZ_MESI  = 9   # Anzianità Mesi
COL_FERIE_AP  = 16  # FERIE - Anni Prec.
COL_FERIE_MAT = 17  # FERIE - Maturati
COL_FERIE_GOD = 18  # FERIE - Goduti
COL_FERIE_RES = 19  # FERIE - Residui
COL_PERM_AP   = 20  # PERMESSI - Anni Prec.
COL_PERM_MAT  = 21  # PERMESSI - Maturati
COL_PERM_GOD  = 22  # PERMESSI - Goduti
COL_PERM_RES  = 23  # PERMESSI - Residui
COL_ROL_AP    = 24  # ROL - Anni Prec.
COL_ROL_MAT   = 25  # ROL - Maturati
COL_ROL_GOD   = 26  # ROL - Goduti
COL_ROL_RES   = 27  # ROL - Residui
COL_EXFEST_AP  = 28  # EX FESTIVITA - Anni Prec.
COL_EXFEST_MAT = 29  # EX FESTIVITA - Maturati
COL_EXFEST_GOD = 30  # EX FESTIVITA - Goduti
COL_EXFEST_RES = 31  # EX FESTIVITA - Residui


def _dec(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _int_or_none(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _data_competenza(row) -> date | None:
    """Ricava data_competenza dall'ultimo giorno del mese."""
    # Preferisce la colonna Data Periodo se valorizzata
    dt = row[COL_DATA_PER]
    if isinstance(dt, datetime):
        return dt.date()
    if isinstance(dt, date):
        return dt
    # Fallback: costruisce da Mese + Anno
    mese_str = str(row[COL_MESE] or "").strip().lower()
    anno = _int_or_none(row[COL_ANNO])
    mese_num = MESI_IT.get(mese_str)
    if mese_num and anno:
        ultimo = calendar.monthrange(anno, mese_num)[1]
        return date(anno, mese_num, ultimo)
    return None


class Command(BaseCommand):
    help = "Import mensile saldi ferie/ROL/ex-festività da file cedolini XLSX"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Percorso del file XLSX cedolini")
        parser.add_argument("--foglio", default="Dati", help="Nome foglio (default: Dati)")
        parser.add_argument("--dry-run", action="store_true", help="Simula senza scrivere su DB")
        parser.add_argument("--user-id", type=int, default=None, help="ID utente Django da associare all'importazione")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl non installato. Esegui: pip install openpyxl")

        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        foglio = options["foglio"]
        dry_run = options["dry_run"]
        verboso = options["verbosity"] >= 2

        self.stdout.write(f"Lettura {path.name} (foglio: {foglio}) ...")
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        if foglio not in wb.sheetnames:
            raise CommandError(f"Foglio '{foglio}' non trovato. Fogli disponibili: {wb.sheetnames}")
        ws = wb[foglio]

        # Pre-carica la mappa CF → legacy_anagrafica_id
        cf_to_lid = dict(
            DipendenteAnagraficaCivile.objects
            .exclude(codice_fiscale="")
            .values_list("codice_fiscale", "legacy_anagrafica_id")
        )

        importatore = None
        if options["user_id"]:
            try:
                importatore = User.objects.get(pk=options["user_id"])
            except User.DoesNotExist:
                self.stderr.write(f"Utente ID {options['user_id']} non trovato, importazione anonima.")

        # Raggruppa per periodo per creare un ImportazioneCedolini per mese
        periodi: dict[date, list] = {}
        skipped = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            cf = str(row[COL_CF] or "").strip().upper()
            if not cf:
                skipped += 1
                continue
            dt = _data_competenza(row)
            if dt is None:
                self.stderr.write(f"  Riga {idx+2}: impossibile determinare data_competenza per CF {cf}, saltata.")
                skipped += 1
                continue
            periodi.setdefault(dt, []).append((cf, row))

        totale_righe = sum(len(v) for v in periodi.values())
        self.stdout.write(f"  {totale_righe} righe valide, {len(periodi)} periodi, {skipped} saltate.")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna scrittura su DB."))
            for dt in sorted(periodi):
                self.stdout.write(f"  Periodo {dt.strftime('%m/%Y')}: {len(periodi[dt])} dipendenti")
            return

        totale_ok = totale_err = totale_nf = 0

        with transaction.atomic():
            for dt in sorted(periodi):
                righe = periodi[dt]
                imp = ImportazioneCedolini.objects.create(
                    data_competenza=dt,
                    origine=ImportazioneCedolini.ORIGINE_XLSX,
                    importato_da=importatore,
                    file_nome=path.name,
                    righe_totali=len(righe),
                )
                ok = err = nf = 0
                for cf, row in righe:
                    try:
                        lid = cf_to_lid.get(cf)
                        if lid is None:
                            nf += 1
                            if verboso:
                                self.stdout.write(f"    CF {cf}: non trovato in anagrafica")

                        SaldoCedolino.objects.update_or_create(
                            tax_code=cf,
                            data_competenza=dt,
                            defaults=dict(
                                importazione=imp,
                                legacy_anagrafica_id=lid,
                                anzianita_anni=_int_or_none(row[COL_ANZ_ANNI]),
                                anzianita_mesi=_int_or_none(row[COL_ANZ_MESI]),
                                ferie_anni_prec=_dec(row[COL_FERIE_AP]),
                                ferie_maturati=_dec(row[COL_FERIE_MAT]),
                                ferie_goduti=_dec(row[COL_FERIE_GOD]),
                                ferie_residui=_dec(row[COL_FERIE_RES]),
                                permessi_anni_prec=_dec(row[COL_PERM_AP]),
                                permessi_maturati=_dec(row[COL_PERM_MAT]),
                                permessi_goduti=_dec(row[COL_PERM_GOD]),
                                permessi_residui=_dec(row[COL_PERM_RES]),
                                rol_anni_prec=_dec(row[COL_ROL_AP]),
                                rol_maturati=_dec(row[COL_ROL_MAT]),
                                rol_goduti=_dec(row[COL_ROL_GOD]),
                                rol_residui=_dec(row[COL_ROL_RES]),
                                ex_fest_anni_prec=_dec(row[COL_EXFEST_AP]),
                                ex_fest_maturati=_dec(row[COL_EXFEST_MAT]),
                                ex_fest_goduti=_dec(row[COL_EXFEST_GOD]),
                                ex_fest_residui=_dec(row[COL_EXFEST_RES]),
                            ),
                        )
                        ok += 1
                    except Exception as exc:
                        err += 1
                        self.stderr.write(f"    Errore CF {cf} {dt}: {exc}")

                imp.righe_ok = ok
                imp.righe_errore = err
                imp.righe_non_trovate = nf
                imp.save(update_fields=["righe_ok", "righe_errore", "righe_non_trovate"])
                totale_ok += ok
                totale_err += err
                totale_nf += nf
                self.stdout.write(
                    f"  {dt.strftime('%m/%Y')}: {ok} OK, {err} errori, {nf} CF non in anagrafica"
                )

        self.stdout.write(self.style.SUCCESS(
            f"\nImport completato: {totale_ok} saldi salvati, {totale_err} errori, {totale_nf} CF sconosciuti."
        ))
