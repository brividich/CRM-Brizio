"""Import storico retributivo (voci paga) da file XLSX dello studio paghe.

Il file deve avere un foglio con le colonne:
    tax_code | nome | pay_item | value | date

Crea una `ImportazioneRetributiva` per ogni mese di competenza presente nel
file e popola le relative `VoceRetributiva`, rilevando le variazioni rispetto
all'ultima importazione precedente.

Esempi:
    python manage.py import_retribuzioni Storico_retributivo.xlsx --dry-run
    python manage.py import_retribuzioni Storico_retributivo.xlsx --foglio Tutti
    python manage.py import_retribuzioni Storico_retributivo.xlsx --user-id 1 --verbose
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from anagrafica.models import (
    DipendenteAnagraficaCivile,
    ImportazioneRetributiva,
    VoceRetributiva,
    _classify_pay_item,
)

User = get_user_model()

# Intestazioni colonna attese (case-insensitive)
COLONNE = ("tax_code", "nome", "pay_item", "value", "date")


def _parse_data(val) -> date | None:
    """Converte la cella `date` (datetime o stringa gg/mm/aaaa) in date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _dec(val) -> Decimal | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return Decimal(str(val).strip().replace(",", ".")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Import storico retributivo (voci paga) da file XLSX studio paghe"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Percorso del file XLSX")
        parser.add_argument("--foglio", default=None, help="Nome foglio (default: primo foglio)")
        parser.add_argument("--dry-run", action="store_true", help="Simula senza scrivere su DB")
        parser.add_argument("--user-id", type=int, default=None, help="ID utente Django importatore")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl non installato. Esegui: pip install openpyxl")

        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"File non trovato: {path}")

        dry_run = options["dry_run"]
        verboso = options["verbosity"] >= 2

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        foglio = options["foglio"] or wb.sheetnames[0]
        if foglio not in wb.sheetnames:
            raise CommandError(f"Foglio '{foglio}' non trovato. Fogli disponibili: {wb.sheetnames}")
        ws = wb[foglio]
        self.stdout.write(f"Lettura {path.name} (foglio: {foglio}) ...")

        # Mappa intestazione → indice colonna
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise CommandError("Foglio vuoto.")
        col_idx = {str(c or "").strip().lower(): i for i, c in enumerate(header)}
        mancanti = [c for c in COLONNE if c not in col_idx]
        if mancanti:
            raise CommandError(f"Colonne mancanti nel foglio: {mancanti}. Trovate: {list(col_idx)}")

        importatore = None
        if options["user_id"]:
            try:
                importatore = User.objects.get(pk=options["user_id"])
            except User.DoesNotExist:
                self.stderr.write(f"Utente ID {options['user_id']} non trovato, importazione anonima.")

        # Lettura righe → raggruppamento per mese di competenza (primo giorno mese)
        periodi: dict[date, list] = {}
        skipped = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            tax_code = str(row[col_idx["tax_code"]] or "").strip().upper()
            pay_item = str(row[col_idx["pay_item"]] or "").strip()
            data_riga = _parse_data(row[col_idx["date"]])
            importo = _dec(row[col_idx["value"]])
            if not tax_code or not pay_item or data_riga is None or importo is None:
                skipped += 1
                if verboso:
                    self.stderr.write(f"  Riga {idx + 2}: dati incompleti, saltata.")
                continue
            nome = str(row[col_idx["nome"]] or "").strip().upper()
            periodo = data_riga.replace(day=1)
            periodi.setdefault(periodo, []).append((tax_code, nome, pay_item, importo, data_riga))

        totale = sum(len(v) for v in periodi.values())
        self.stdout.write(f"  {totale} voci valide, {len(periodi)} periodi, {skipped} saltate.")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN: nessuna scrittura su DB."))
            for periodo in sorted(periodi):
                righe = periodi[periodo]
                dipendenti = len({r[0] for r in righe})
                self.stdout.write(
                    f"  Periodo {periodo.strftime('%m-%Y')}: "
                    f"{len(righe)} voci, {dipendenti} dipendenti"
                )
            return

        # Mappe CF → legacy_anagrafica_id e nome → id (fallback)
        cf_to_id = dict(
            DipendenteAnagraficaCivile.objects
            .exclude(codice_fiscale="")
            .values_list("codice_fiscale", "legacy_anagrafica_id")
        )
        nome_to_id: dict[str, int] = {}
        try:
            from core.legacy_anagrafica import fetch_anagrafica_rows
            for lr in fetch_anagrafica_rows(deduplicate=True):
                cog = str(lr.get("cognome") or "").strip().upper()
                nom = str(lr.get("nome") or "").strip().upper()
                full = f"{cog} {nom}".strip()
                lid = int(lr.get("id") or 0)
                if full and lid:
                    nome_to_id[full] = lid
        except Exception:
            self.stderr.write("Fallback nome-based non disponibile (legacy DB non raggiungibile).")

        totale_ok = totale_err = 0

        with transaction.atomic():
            for periodo in sorted(periodi):
                righe = periodi[periodo]

                # Voci dell'ultima importazione precedente per rilevare variazioni
                prev_import = (
                    ImportazioneRetributiva.objects
                    .filter(data_competenza__lt=periodo)
                    .order_by("-data_competenza")
                    .first()
                )
                prev_voci: dict[tuple, Decimal] = {}
                if prev_import:
                    for voce in VoceRetributiva.objects.filter(importazione=prev_import):
                        prev_voci[(voce.tax_code, voce.pay_item_key)] = voce.importo

                imp = ImportazioneRetributiva.objects.create(
                    data_competenza=periodo,
                    origine=ImportazioneRetributiva.ORIGINE_CSV,
                    importato_da=importatore,
                    file_nome=path.name,
                    righe_totali=len(righe),
                )

                voci_to_create: list[VoceRetributiva] = []
                ok = err = 0
                for tax_code, nome, pay_item, importo, data_riga in righe:
                    try:
                        pay_item_key = pay_item.lower()
                        legacy_id = cf_to_id.get(tax_code) or nome_to_id.get(nome)
                        prev_importo = prev_voci.get((tax_code, pay_item_key))
                        is_changed = prev_importo is not None and prev_importo != importo
                        voci_to_create.append(VoceRetributiva(
                            importazione=imp,
                            tax_code=tax_code,
                            legacy_anagrafica_id=legacy_id,
                            data_competenza=data_riga,
                            pay_item=pay_item,
                            pay_item_key=pay_item_key,
                            categoria=_classify_pay_item(pay_item_key),
                            importo=importo,
                            is_changed=is_changed,
                            importo_precedente=prev_importo if is_changed else None,
                        ))
                        ok += 1
                    except Exception as exc:
                        err += 1
                        self.stderr.write(f"    Errore {tax_code} {pay_item}: {exc}")

                VoceRetributiva.objects.bulk_create(voci_to_create)
                imp.righe_ok = ok
                imp.righe_errore = err
                imp.save(update_fields=["righe_ok", "righe_errore"])
                totale_ok += ok
                totale_err += err
                self.stdout.write(f"  {periodo.strftime('%m-%Y')}: {ok} voci OK, {err} errori")

        self.stdout.write(self.style.SUCCESS(
            f"\nImport completato: {totale_ok} voci salvate, {totale_err} errori."
        ))
