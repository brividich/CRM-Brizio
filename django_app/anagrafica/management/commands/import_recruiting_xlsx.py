"""Import dello storico candidati dal Mod. 05-01 «Valutazione Selezione Risorse».

Il foglio Excel di HR non ha intestazioni standardizzate (cambiano tra revisioni
e tra fogli), quindi le colonne si riconoscono per **sinonimi normalizzati**
invece che per posizione fissa. ``--dry-run`` stampa la mappatura riconosciuta e
le colonne rimaste fuori: si verifica quella prima di scrivere qualcosa.

I criteri di valutazione non sono hardcoded: le colonne di punteggio si agganciano
ai ``RecruitingCriterio`` presenti a DB, per codice o per label. Un criterio
aggiunto o rinominato da HR viene riconosciuto senza toccare questo file.

Privacy: il comando non stampa mai valori di cella con dati personali — solo
numeri di riga, conteggi e nomi di colonna. Il file sorgente non va committato.

Esempi:
    python manage.py import_recruiting_xlsx "C:/tmp/MOD 05-01.xlsx" --dry-run
    python manage.py import_recruiting_xlsx "C:/tmp/MOD 05-01.xlsx" --sheet "SELEZIONE Operatore"
    python manage.py import_recruiting_xlsx "C:/tmp/MOD 05-01.xlsx" --limit 20 --dry-run
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from anagrafica.models_recruiting import (
    Candidato,
    CandidatoPunteggio,
    RecruitingCriterio,
)
from anagrafica.services import recruiting as recruiting_service


# ---------------------------------------------------------------------------
# Normalizzazione e riconoscimento delle intestazioni
# ---------------------------------------------------------------------------

def _norm(value) -> str:
    """Minuscolo, senza accenti, senza punteggiatura, spazi compattati."""
    if value is None:
        return ""
    testo = unicodedata.normalize("NFKD", str(value))
    testo = "".join(ch for ch in testo if not unicodedata.combining(ch))
    testo = re.sub(r"[^a-z0-9]+", " ", testo.lower())
    return re.sub(r"\s+", " ", testo).strip()


# campo modello -> sinonimi di intestazione (già normalizzati).
# Il primo che matcha vince; il match è esatto oppure "l'intestazione inizia con".
COLONNE = {
    "codice_riferimento": ["riferimento", "progressivo", "rif", "n candidato", "n"],
    "data_primo_colloquio": ["data 1 colloquio", "data primo colloquio", "data colloquio", "data"],
    "canale_provenienza": ["provenienza", "canale", "canale cv", "provenienza cv", "fonte"],
    "canale_dettaglio": ["agenzia", "dettaglio provenienza", "segnalato da"],
    "cognome": ["cognome"],
    "nome": ["nome"],
    "cellulare": ["cellulare", "telefono", "cell", "tel"],
    "email": ["email", "e mail", "mail"],
    "localita": ["localita", "citta", "comune", "residenza"],
    "provincia": ["provincia", "prov"],
    "mansione_cercata": ["mansione primaria cercata", "mansione cercata", "mansione richiesta", "posizione"],
    "azienda_attuale": ["azienda attuale", "azienda"],
    "mansione_attuale": ["mansione attuale"],
    "livello_contratto_attuale": ["livello contratto attuale", "livello contratto", "livello", "contratto attuale"],
    "occupato_attualmente": ["occupazione attuale", "occupato", "occupato attualmente"],
    "eta": ["eta", "anni"],
    "titolo_studio": ["titolo di studio", "titolo studio", "studi"],
    "cittadinanza": ["cittadinanza", "nazionalita"],
    "cv_esito": ["c v", "cv", "esito cv"],
    "colloquio_effettuato": ["colloquio effettuato", "colloquio"],
    "lingua_inglese_livello": ["lingua inglese", "inglese"],
    "idoneita_tirocinio": ["idoneita tirocinio", "tirocinio"],
    "idoneita_apprendistato": ["idoneita apprendistato", "apprendistato"],
    "disponibilita": ["disponibilita"],
    "motivo_cambio_lavoro": ["motivo del cambio lavoro", "motivo cambio lavoro", "motivo cambio"],
    "note": ["note", "note libere", "annotazioni"],
    "rischio_abbandono": ["rischio di abbandono", "rischio abbandono", "rischio"],
    "giudizio_finale": ["giudizio finale", "giudizio", "esito finale"],
    "data_secondo_colloquio": ["data 2 colloquio", "data secondo colloquio"],
    "note_secondo_colloquio": ["note 2 colloquio", "note secondo colloquio"],
    "comunicazione_esito": ["comunicazione esito", "comunicazione", "esito comunicato"],
    "data_assunzione": ["data assunzione", "assunzione"],
}

# Intestazioni da ignorare senza segnalarle come "non riconosciute": sono i
# totali/medie di fondo pagina, ricalcolati dal portale.
IGNORA = [
    "punteggio medio totale", "punteggio medio", "media ponderata", "media",
    "totale", "conteggio", "n", "progressivo",
]

CANALE_MAP = {
    "autocandidatura": Candidato.CANALE_AUTOCANDIDATURA,
    "agenzia": Candidato.CANALE_AGENZIA,
    "agenzia interinale": Candidato.CANALE_AGENZIA,
    "interinale": Candidato.CANALE_AGENZIA,
    "collocamento mirato": Candidato.CANALE_COLLOCAMENTO_MIRATO,
    "collocamento": Candidato.CANALE_COLLOCAMENTO_MIRATO,
    "segnalazione": Candidato.CANALE_SEGNALAZIONE,
    "annuncio": Candidato.CANALE_ANNUNCIO,
    "scuola": Candidato.CANALE_SCUOLA,
    "universita": Candidato.CANALE_SCUOLA,
}

COMUNICAZIONE_MAP = {
    "si": Candidato.COMUNICAZIONE_SI,
    "no": Candidato.COMUNICAZIONE_NO,
    "academy": Candidato.COMUNICAZIONE_ACADEMY,
    "rinuncia": Candidato.COMUNICAZIONE_RINUNCIA,
}

VERO = {"si", "s", "sl", "x", "true", "vero", "1", "ok", "yes", "y"}
FALSO = {"no", "n", "false", "falso", "0", "ko"}


def _match_colonna(intestazione: str) -> str | None:
    """Campo del modello per questa intestazione, o None se non riconosciuta."""
    testa = _norm(intestazione)
    if not testa:
        return None
    for campo, sinonimi in COLONNE.items():
        for sinonimo in sinonimi:
            if testa == sinonimo:
                return campo
    # Secondo giro, più permissivo: prefisso (es. "note 2° colloquio (sintesi)").
    for campo, sinonimi in COLONNE.items():
        for sinonimo in sinonimi:
            if testa.startswith(sinonimo + " ") or testa.startswith(sinonimo):
                return campo
    return None


def _is_ignorabile(intestazione: str) -> bool:
    testa = _norm(intestazione)
    return any(testa == v or testa.startswith(v) for v in IGNORA)


# ---------------------------------------------------------------------------
# Conversioni di valore
# ---------------------------------------------------------------------------

def _testo(value, limite: int = 250) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())[:limite]


def _data(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    testo = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(testo, fmt).date()
        except ValueError:
            continue
    return None


def _intero(value, minimo: int | None = None, massimo: int | None = None) -> int | None:
    if value is None or value == "":
        return None
    try:
        numero = int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return None
    if minimo is not None and numero < minimo:
        return None
    if massimo is not None and numero > massimo:
        return None
    return numero


def _booleano(value) -> bool | None:
    testa = _norm(value)
    if not testa:
        return None
    if testa in VERO:
        return True
    if testa in FALSO:
        return False
    return None


def _decimale(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _canale(value) -> tuple[str, str]:
    """(codice canale, dettaglio). Un canale non riconosciuto finisce in ALTRO
    conservando il testo originale nel dettaglio: nessuna informazione persa."""
    testa = _norm(value)
    if not testa:
        return Candidato.CANALE_AUTOCANDIDATURA, ""
    if testa in CANALE_MAP:
        return CANALE_MAP[testa], ""
    for chiave, codice in CANALE_MAP.items():
        if chiave in testa:
            return codice, _testo(value, 160)
    return Candidato.CANALE_ALTRO, _testo(value, 160)


def _giudizio(value) -> str:
    testa = _norm(value)
    if testa.startswith("pos"):
        return Candidato.GIUDIZIO_POSITIVO
    if testa.startswith("neg"):
        return Candidato.GIUDIZIO_NEGATIVO
    return ""


def _cv_esito(value) -> str:
    testa = _norm(value)
    if not testa:
        return ""
    if testa in {"ok", "si", "buono", "positivo"}:
        return Candidato.CV_OK
    if testa in {"0", "no", "ko", "negativo"}:
        return Candidato.CV_KO
    return Candidato.CV_OK if "ok" in testa else Candidato.CV_KO


def _comunicazione(value) -> str:
    testa = _norm(value)
    if not testa:
        return ""
    if testa in COMUNICAZIONE_MAP:
        return COMUNICAZIONE_MAP[testa]
    for chiave, codice in COMUNICAZIONE_MAP.items():
        if chiave in testa:
            return codice
    return ""


def _stato_iniziale(dati: dict) -> str:
    """Stato dell'iter dedotto dai dati importati, senza inventare transizioni."""
    if dati.get("data_assunzione"):
        return Candidato.STATO_ASSUNTO
    if _norm(dati.get("comunicazione_esito")) == _norm(Candidato.COMUNICAZIONE_RINUNCIA):
        return Candidato.STATO_RINUNCIA
    if dati.get("giudizio_finale") == Candidato.GIUDIZIO_NEGATIVO:
        return Candidato.STATO_SCARTATO
    if dati.get("data_secondo_colloquio"):
        return Candidato.STATO_COLLOQUIO_2
    if dati.get("data_primo_colloquio") or dati.get("colloquio_effettuato"):
        return Candidato.STATO_COLLOQUIO_1
    if dati.get("cv_esito"):
        return Candidato.STATO_CV_VALUTATO
    return Candidato.STATO_NUOVO


class Command(BaseCommand):
    help = "Importa lo storico candidati dal Mod. 05-01 (Excel)."

    def add_arguments(self, parser):
        parser.add_argument("percorso", help="Percorso del file .xlsx")
        parser.add_argument("--sheet", default=None, help="Nome del foglio (default: il primo)")
        parser.add_argument(
            "--header-row", type=int, default=None,
            help="Riga delle intestazioni (1-based). Se omesso, viene rilevata automaticamente "
                 "(il Mod. 05-01 ha una testata sopra le intestazioni).",
        )
        parser.add_argument("--limit", type=int, default=0, help="Importa solo le prime N righe dati")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla: mostra mappatura ed esito")
        parser.add_argument(
            "--scan", action="store_true",
            help="Diagnostica: per le prime righe mostra quante colonne verrebbero riconosciute, "
                 "così si individua la riga delle intestazioni. Non importa nulla.",
        )
        parser.add_argument(
            "--allow-duplicates", action="store_true",
            help="Non saltare i candidati già presenti (cognome+nome+data 1° colloquio)",
        )

    # Quante righe in cima esaminare per trovare le intestazioni.
    RIGHE_SCANSIONE = 25

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dipendenza di progetto
            raise CommandError("openpyxl non disponibile: pip install openpyxl") from exc

        percorso = Path(opts["percorso"])
        if not percorso.exists():
            raise CommandError(f"File non trovato: {percorso}")

        wb = load_workbook(percorso, data_only=True, read_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb[wb.sheetnames[0]]
        self.stdout.write(f"Foglio: {ws.title}")

        righe = list(ws.iter_rows(values_only=True))
        if not righe:
            raise CommandError("Foglio vuoto.")

        # Modalità diagnostica: mostra dove sono le intestazioni e basta.
        if opts["scan"]:
            self._scan_intestazioni(righe)
            wb.close()
            return

        if opts["header_row"] is not None:
            indice_header = max(0, opts["header_row"] - 1)
            if indice_header >= len(righe):
                raise CommandError("Riga di intestazione oltre la fine del foglio.")
        else:
            # Auto-detect: il Mod. 05-01 ha una testata (logo/codice/revisione)
            # sopra le intestazioni di colonna, che quindi NON sono sulla riga 1.
            indice_header, riconosciute = self._rileva_riga_header(righe)
            if indice_header is None:
                raise CommandError(
                    "Nessuna riga di intestazioni riconosciuta nelle prime "
                    f"{self.RIGHE_SCANSIONE} righe. Controlla --sheet, oppure indica la riga "
                    "a mano con --header-row N (usa --scan per individuarla)."
                )
            self.stdout.write(self.style.SUCCESS(
                f"Intestazioni rilevate alla riga {indice_header + 1} "
                f"({riconosciute} colonne riconosciute)."
            ))

        intestazioni = list(righe[indice_header])
        mappa, mappa_criteri, ignorate, sconosciute = self._analizza(intestazioni)
        self._stampa_mappatura(mappa, mappa_criteri, ignorate, sconosciute)

        anonimo = not mappa.get("cognome") and not mappa.get("nome")
        if anonimo:
            # Il Mod. 05-01 di HR arriva spesso senza nominativi (rimossi alla
            # fonte per privacy): non è un errore, è un import da completare a
            # mano. Serve però qualcosa che distingua una riga vera da una vuota.
            if not mappa.get("data_primo_colloquio") and not mappa.get("mansione_cercata") \
                    and not mappa_criteri:
                raise CommandError(
                    "Nessuna colonna nome/cognome né dati riconoscibili (data colloquio, "
                    "mansione, punteggi): controlla --header-row e --sheet."
                )
            self.stdout.write(self.style.WARNING(
                "Import ANONIMO: nessuna colonna nominativo. Le schede saranno create "
                "«Da completare» — inserisci nome e cognome dal portale. Riconciliazione "
                "tramite la colonna di riferimento se presente, altrimenti il numero di riga."
            ))

        dati_righe = righe[indice_header + 1:]
        if opts["limit"]:
            dati_righe = dati_righe[: opts["limit"]]

        creati = saltati = vuote = 0
        for offset, riga in enumerate(dati_righe, start=indice_header + 2):
            valori = self._leggi(riga, mappa, numero_riga=offset)
            if not self._riga_significativa(valori, riga, mappa_criteri):
                vuote += 1
                continue

            if not opts["allow_duplicates"] and self._esiste(valori):
                saltati += 1
                continue

            voti = self._leggi_voti(riga, mappa_criteri)
            if opts["dry_run"]:
                creati += 1
                continue

            try:
                with transaction.atomic():
                    candidato = Candidato.objects.create(**valori)
                    for criterio, voto in voti:
                        CandidatoPunteggio.objects.create(
                            candidato=candidato, criterio=criterio, valore=voto,
                            peso_snapshot=criterio.peso_percentuale,
                        )
                    recruiting_service.ricalcola_punteggio(candidato)
                creati += 1
            except Exception as exc:  # riga malformata: si prosegue
                saltati += 1
                self.stderr.write(f"  riga {offset}: scartata ({exc.__class__.__name__})")

        wb.close()
        prefisso = "[dry-run] " if opts["dry_run"] else ""
        coda = " (da completare col nominativo)" if anonimo else ""
        self.stdout.write(self.style.SUCCESS(
            f"{prefisso}candidati importati: {creati}{coda} · già presenti: {saltati} · righe vuote: {vuote}"
        ))
        if opts["dry_run"]:
            self.stdout.write("Nessuna scrittura eseguita. Rilancia senza --dry-run per importare.")

    # -- rilevamento della riga di intestazioni ------------------------------

    def _riconoscibili(self, intestazioni) -> int:
        """Quante colonne (campi + criteri) verrebbero riconosciute da questa riga."""
        mappa, mappa_criteri, _ignorate, _sconosciute = self._analizza(intestazioni)
        return len(mappa) + len(mappa_criteri)

    def _rileva_riga_header(self, righe):
        """(indice_riga_migliore, n_colonne) tra le prime righe; (None, 0) se nessuna.

        Sceglie la riga che massimizza le colonne riconosciute — così una testata
        di documento (logo/codice/revisione) sopra le intestazioni non inganna.
        A parità di conteggio vince la più in alto (la vera riga di intestazioni
        precede i dati).
        """
        migliore_idx, migliore_n = None, 0
        for idx in range(min(self.RIGHE_SCANSIONE, len(righe))):
            n = self._riconoscibili(list(righe[idx]))
            if n > migliore_n:
                migliore_idx, migliore_n = idx, n
        return migliore_idx, migliore_n

    def _scan_intestazioni(self, righe):
        """Diagnostica --scan: per ogni riga in cima, quante colonne riconoscerebbe."""
        self.stdout.write(f"Scansione delle prime {self.RIGHE_SCANSIONE} righe:\n")
        migliore_idx, migliore_n = None, 0
        for idx in range(min(self.RIGHE_SCANSIONE, len(righe))):
            n = self._riconoscibili(list(righe[idx]))
            marca = ""
            if n > migliore_n:
                migliore_idx, migliore_n = idx, n
            if n > 0:
                marca = f"  <-- {n} colonne riconosciute"
            self.stdout.write(f"  riga {idx + 1:>2}: {n} colonne{marca}")
        self.stdout.write("")
        if migliore_idx is None:
            self.stdout.write(self.style.WARNING(
                "Nessuna riga riconosciuta: le intestazioni potrebbero usare nomi molto diversi "
                "dai sinonimi attesi, oppure stare oltre la finestra di scansione."
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"Riga di intestazioni più probabile: {migliore_idx + 1} "
                f"({migliore_n} colonne). Importa con:  --header-row {migliore_idx + 1}  "
                f"(oppure senza --header-row: viene rilevata da sola)."
            ))

    # -- analisi intestazioni ------------------------------------------------

    def _analizza(self, intestazioni):
        """(mappa campo->colonna, [(criterio, colonna)], ignorate, sconosciute)."""
        criteri = list(RecruitingCriterio.objects.all())
        per_codice = {_norm(c.codice): c for c in criteri}
        per_label = {_norm(c.label): c for c in criteri}

        mappa: dict[str, int] = {}
        mappa_criteri: list[tuple[RecruitingCriterio, int]] = []
        ignorate: list[str] = []
        sconosciute: list[str] = []

        for indice, intestazione in enumerate(intestazioni):
            testa = _norm(intestazione)
            if not testa:
                continue

            criterio = per_codice.get(testa) or per_label.get(testa)
            if criterio is None:
                for chiave, cand in per_label.items():
                    if chiave and (testa.startswith(chiave) or chiave.startswith(testa)):
                        criterio = cand
                        break
            if criterio is not None:
                mappa_criteri.append((criterio, indice))
                continue

            campo = _match_colonna(intestazione)
            if campo and campo not in mappa:
                mappa[campo] = indice
                continue

            if _is_ignorabile(intestazione):
                ignorate.append(str(intestazione))
            else:
                sconosciute.append(str(intestazione))

        return mappa, mappa_criteri, ignorate, sconosciute

    def _stampa_mappatura(self, mappa, mappa_criteri, ignorate, sconosciute):
        self.stdout.write("\nColonne riconosciute:")
        for campo, indice in sorted(mappa.items(), key=lambda kv: kv[1]):
            self.stdout.write(f"  col {indice + 1:>3}  ->  {campo}")
        for criterio, indice in mappa_criteri:
            self.stdout.write(f"  col {indice + 1:>3}  ->  punteggio «{criterio.label}» ({criterio.peso_percentuale}%)")
        if not mappa_criteri:
            self.stdout.write(self.style.WARNING(
                "  nessuna colonna di punteggio riconosciuta: i candidati saranno importati "
                "senza valutazione (i criteri si agganciano per codice o label)."
            ))
        if ignorate:
            self.stdout.write(f"\nColonne ignorate (aggregati ricalcolati dal portale): {', '.join(ignorate)}")
        if sconosciute:
            self.stdout.write(self.style.WARNING(
                f"\nColonne NON riconosciute (saranno perse): {', '.join(sconosciute)}"
            ))
        self.stdout.write("")

    # -- lettura riga --------------------------------------------------------

    def _leggi(self, riga, mappa, *, numero_riga: int) -> dict:
        def cella(campo):
            indice = mappa.get(campo)
            if indice is None or indice >= len(riga):
                return None
            return riga[indice]

        # Riferimento: la colonna del foglio se c'è, altrimenti il numero di riga
        # del file — così una scheda anonima resta agganciabile alla sua origine.
        riferimento = _testo(cella("codice_riferimento"), 60)
        if not riferimento:
            riferimento = f"riga {numero_riga}"

        canale, dettaglio = _canale(cella("canale_provenienza"))
        dati = {
            "cognome": _testo(cella("cognome"), 120),
            "nome": _testo(cella("nome"), 120),
            "codice_riferimento": riferimento,
            "cellulare": _testo(cella("cellulare"), 40),
            "email": _testo(cella("email"), 254),
            "localita": _testo(cella("localita"), 120),
            "provincia": _testo(cella("provincia"), 4).upper(),
            "canale_provenienza": canale,
            "canale_dettaglio": _testo(cella("canale_dettaglio"), 160) or dettaglio,
            "mansione_cercata": _testo(cella("mansione_cercata"), 160),
            "azienda_attuale": _testo(cella("azienda_attuale"), 160),
            "mansione_attuale": _testo(cella("mansione_attuale"), 160),
            "livello_contratto_attuale": _testo(cella("livello_contratto_attuale"), 120),
            "occupato_attualmente": _booleano(cella("occupato_attualmente")),
            "eta": _intero(cella("eta"), 14, 99),
            "titolo_studio": _testo(cella("titolo_studio"), 160),
            "cittadinanza": _testo(cella("cittadinanza"), 120),
            "data_primo_colloquio": _data(cella("data_primo_colloquio")),
            "cv_esito": _cv_esito(cella("cv_esito")),
            "colloquio_effettuato": bool(_booleano(cella("colloquio_effettuato"))),
            "lingua_inglese_livello": _testo(cella("lingua_inglese_livello"), 60),
            "idoneita_tirocinio": _booleano(cella("idoneita_tirocinio")),
            "idoneita_apprendistato": _booleano(cella("idoneita_apprendistato")),
            "disponibilita": _testo(cella("disponibilita"), 200),
            "motivo_cambio_lavoro": _testo(cella("motivo_cambio_lavoro"), 2000),
            "note": _testo(cella("note"), 2000),
            "rischio_abbandono": _intero(cella("rischio_abbandono"), 1, 10),
            "giudizio_finale": _giudizio(cella("giudizio_finale")),
            "data_secondo_colloquio": _data(cella("data_secondo_colloquio")),
            "note_secondo_colloquio": _testo(cella("note_secondo_colloquio"), 2000),
            "comunicazione_esito": _comunicazione(cella("comunicazione_esito")),
            "data_assunzione": _data(cella("data_assunzione")),
        }
        dati["stato"] = _stato_iniziale(dati)
        return dati

    def _leggi_voti(self, riga, mappa_criteri):
        voti = []
        for criterio, indice in mappa_criteri:
            if indice >= len(riga):
                continue
            voto = _intero(riga[indice], 1, 5)
            if voto is not None:
                voti.append((criterio, voto))
        return voti

    def _riga_significativa(self, valori, riga, mappa_criteri) -> bool:
        """Distingue una riga vera da una vuota o di separazione.

        Con nominativo basta quello; per gli import anonimi serve almeno un dato
        di sostanza (data colloquio, mansione, età, o un voto), così le righe
        vuote/di totale non diventano schede fantasma.
        """
        if valori.get("cognome") or valori.get("nome"):
            return True
        if valori.get("data_primo_colloquio") or valori.get("mansione_cercata") \
                or valori.get("eta") or valori.get("giudizio_finale"):
            return True
        return bool(self._leggi_voti(riga, mappa_criteri))

    def _esiste(self, valori) -> bool:
        """Duplicato: per le schede con nominativo, nome + data 1° colloquio;
        per quelle anonime, il codice di riferimento (colonna del foglio se c'era).

        Meglio saltare una riga in più che creare doppioni in un database
        consultato per le ricerche future. Il fallback ``riga N`` NON deduplica
        (cambia se il file cambia): al reimport di un file anonimo senza colonna
        di riferimento, usare ``--limit``/una tantum per non duplicare.
        """
        cognome = valori.get("cognome") or ""
        nome = valori.get("nome") or ""
        if cognome or nome:
            query = Candidato.objects.filter(cognome__iexact=cognome, nome__iexact=nome)
            if valori.get("data_primo_colloquio"):
                query = query.filter(data_primo_colloquio=valori["data_primo_colloquio"])
            return query.exists()

        riferimento = valori.get("codice_riferimento") or ""
        if riferimento and not riferimento.startswith("riga "):
            return Candidato.objects.filter(
                cognome="", nome="", codice_riferimento=riferimento,
            ).exists()
        return False
