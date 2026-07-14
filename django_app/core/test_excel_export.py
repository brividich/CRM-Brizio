from __future__ import annotations

import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from django.test import TestCase

from openpyxl import Workbook, load_workbook

from core.excel_export import _brand_header, build_xlsx_bytes


class ExcelExportUtilTests(TestCase):
    # TestCase (non SimpleTestCase): con `title` valorizzato build_xlsx_bytes
    # legge il branding via PdfTheme.from_branding() (query DB SiteConfig).
    def test_build_xlsx_bytes_valid(self):
        from openpyxl import load_workbook

        from core.excel_export import build_xlsx_bytes

        data = build_xlsx_bytes(
            columns=["A", "B"], rows=[["x", 1], ["y", 2]], title="Titolo", sheet_title="Foglio"
        )
        self.assertEqual(data[:2], b"PK")  # firma xlsx (zip)

        ws = load_workbook(BytesIO(data)).active
        self.assertEqual(ws.title, "Foglio")
        # Con `title` valorizzato il blocco intestazione HUB sposta titolo/header:
        # titolo in A3, header tabella in riga 7 (vedi XlsxDocumentHeaderTests).
        self.assertEqual(ws.cell(row=3, column=1).value, "Titolo")
        self.assertEqual(ws.cell(row=7, column=1).value, "A")
        self.assertEqual(ws.cell(row=8, column=2).value, 1)

    def test_make_xlsx_response(self):
        from core.excel_export import make_xlsx_response

        resp = make_xlsx_response(filename='c"on.xlsx', columns=["A"], rows=[["x"]])
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn('filename="con.xlsx"', resp["Content-Disposition"])  # virgolette interne ripulite
        self.assertEqual(resp.content[:2], b"PK")


class XlsxDocumentHeaderTests(TestCase):
    def test_header_block_contains_title_subtitle_and_filters(self):
        data = build_xlsx_bytes(
            columns=["Nominativo", "Reparto"],
            rows=[["Mario Bianchi", "Officina"]],
            title="Elenco dipendenti",
            subtitle="Generato il 12-07-2026 da Mario Rossi",
            filters_label="Filtri: Reparto = Officina · 1 righe",
        )
        ws = load_workbook(BytesIO(data)).active
        self.assertEqual(ws["A3"].value, "Elenco dipendenti")
        self.assertEqual(ws["A4"].value, "Generato il 12-07-2026 da Mario Rossi")
        self.assertEqual(ws["A5"].value, "Filtri: Reparto = Officina · 1 righe")
        self.assertEqual(ws["A7"].value, "Nominativo")
        self.assertEqual(ws["A8"].value, "Mario Bianchi")
        self.assertEqual(ws.freeze_panes, "A8")

    def test_without_title_header_stays_on_first_row(self):
        data = build_xlsx_bytes(columns=["Nominativo"], rows=[["Mario Bianchi"]])
        ws = load_workbook(BytesIO(data)).active
        self.assertEqual(ws["A1"].value, "Nominativo")
        self.assertEqual(ws["A2"].value, "Mario Bianchi")

    def test_header_uses_print_title_rows_for_repeated_header(self):
        # ws.print_title_rows definisce l'area di stampa con riga header ripetuta:
        # requisito di progetto, non va rimosso (vedi consegna del task).
        data = build_xlsx_bytes(
            columns=["Nominativo"],
            rows=[["Mario Bianchi"]],
            title="Elenco dipendenti",
        )
        ws = load_workbook(BytesIO(data)).active
        self.assertEqual(ws.print_title_rows, "$7:$7")


class XlsxFormulaInjectionTests(TestCase):
    """SICUREZZA: nessuna cella del file prodotto deve essere una formula viva.

    openpyxl scrive come formula (``data_type == 'f'``) qualunque stringa che
    inizia con ``=``: i testi degli export arrivano dal DB e — per l'etichetta
    filtri — perfino dalla querystring, quindi un link malevolo potrebbe far
    scaricare a un utente HR un .xlsx con dentro una formula attiva.
    """

    PAYLOAD = '=HYPERLINK("http://evil.invalid","apri")'

    def _cells(self, data):
        ws = load_workbook(BytesIO(data)).active
        return ws

    def test_data_cell_starting_with_equals_is_not_a_formula(self):
        data = build_xlsx_bytes(columns=["Descrizione"], rows=[[self.PAYLOAD]])
        cell = self._cells(data)["A2"]
        self.assertNotEqual(cell.data_type, "f")
        self.assertEqual(cell.data_type, "s")
        self.assertEqual(cell.value, self.PAYLOAD)  # valore preservato, come testo
        self.assertNotIn(b"<f>", self._sheet_xml(data))

    def _sheet_xml(self, data) -> bytes:
        import zipfile

        with zipfile.ZipFile(BytesIO(data)) as zf:
            return zf.read("xl/worksheets/sheet1.xml")

    def test_header_block_texts_are_not_formulas(self):
        data = build_xlsx_bytes(
            columns=["Nominativo"],
            rows=[["Mario Bianchi"]],
            title=f"Titolo {self.PAYLOAD}",
            subtitle=self.PAYLOAD,
            filters_label=self.PAYLOAD,  # riflessa dalla querystring (?q=…)
        )
        ws = self._cells(data)
        for ref in ("A3", "A4", "A5"):
            with self.subTest(ref=ref):
                self.assertNotEqual(ws[ref].data_type, "f")
        self.assertEqual(ws["A4"].value, self.PAYLOAD)
        self.assertEqual(ws["A5"].value, self.PAYLOAD)
        self.assertNotIn(b"<f>", self._sheet_xml(data))

    def test_all_dangerous_prefixes_are_neutralised(self):
        payloads = [
            '=cmd|"/c calc"!A1',
            "+1+1",
            "-1+1",
            "@SUM(A1)",
            "\t=SUM(A1)",
            "\r=SUM(A1)",
        ]
        data = build_xlsx_bytes(columns=["X"], rows=[[p] for p in payloads])
        ws = self._cells(data)
        for i, payload in enumerate(payloads, start=2):
            with self.subTest(payload=payload):
                self.assertNotEqual(ws.cell(row=i, column=1).data_type, "f")
        self.assertNotIn(b"<f>", self._sheet_xml(data))

    def test_column_header_starting_with_equals_is_not_a_formula(self):
        data = build_xlsx_bytes(columns=[self.PAYLOAD], rows=[["x"]])
        self.assertNotEqual(self._cells(data)["A1"].data_type, "f")

    def test_legitimate_values_are_not_altered(self):
        from datetime import date
        from decimal import Decimal

        data = build_xlsx_bytes(
            columns=["Testo", "Intero", "Decimale", "Data", "Negativo testo"],
            rows=[["Cabina di verniciatura", 12, Decimal("10.5"), date(2026, 3, 10), "-12,50"]],
        )
        ws = self._cells(data)
        self.assertEqual(ws["A2"].value, "Cabina di verniciatura")
        # I NON stringa restano tipizzati (numeri/date, non testo).
        self.assertEqual(ws["B2"].value, 12)
        self.assertEqual(ws["B2"].data_type, "n")
        self.assertEqual(ws["C2"].value, 10.5)
        self.assertEqual(ws["C2"].data_type, "n")
        # openpyxl rilegge le date come datetime: resta un valore-data, non testo.
        self.assertEqual(ws["D2"].value.date(), date(2026, 3, 10))
        self.assertEqual(ws["D2"].data_type, "d")
        # Numero negativo passato come STRINGA: resta la stringa originale,
        # senza apice visibile né altre deturpazioni.
        self.assertEqual(ws["E2"].value, "-12,50")


class BrandHeaderLogoResizeTests(TestCase):
    """Copre il bug: img.height veniva sovrascritto PRIMA di calcolare il
    rapporto d'aspetto per img.width, quindi il logo usciva distorto
    (larghezza nativa in pixel invece che proporzionale a 36px di altezza).
    """

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="excel_export_logo_")
        self.addCleanup(shutil.rmtree, self._tmpdir, ignore_errors=True)

    def _make_logo(self, width, height):
        from PIL import Image as PILImage

        logo_path = Path(self._tmpdir) / "logo.png"
        PILImage.new("RGB", (width, height), color="white").save(logo_path)
        return str(logo_path)

    def test_logo_is_resized_proportionally_not_distorted(self):
        logo_path = self._make_logo(400, 100)
        fake_theme = SimpleNamespace(portal_name="NOVICROM HUB", logo_path=logo_path)

        wb = Workbook()
        ws = wb.active

        with mock.patch("core.pdf.PdfTheme.from_branding", return_value=fake_theme):
            _brand_header(ws, title="T", subtitle=None, filters_label=None, logo=True)

        self.assertEqual(len(ws._images), 1)
        img = ws._images[0]
        # Altezza forzata a 36px; larghezza proporzionale (400 * 36/100 = 144),
        # non la larghezza nativa (400) come nel bug originale.
        self.assertEqual(img.height, 36)
        self.assertEqual(img.width, 144)

    def test_logo_resize_with_different_aspect_ratio(self):
        # Logo quasi quadrato: 120x100 -> larghezza attesa 43 (120 * 36/100 = 43.2 -> int 43)
        logo_path = self._make_logo(120, 100)
        fake_theme = SimpleNamespace(portal_name="NOVICROM HUB", logo_path=logo_path)

        wb = Workbook()
        ws = wb.active

        with mock.patch("core.pdf.PdfTheme.from_branding", return_value=fake_theme):
            _brand_header(ws, title="T", subtitle=None, filters_label=None, logo=True)

        img = ws._images[0]
        self.assertEqual(img.height, 36)
        self.assertEqual(img.width, 43)
