"""Util riusabile per esportare un "report a tabella" in .xlsx (openpyxl).

Stile NOVICROM HUB: intestazione navy con testo bianco, larghezze auto, riga
intestazione bloccata e autofiltro. Da usare nelle view per restituire un download
Excel coerente, riusando i dati gia' calcolati (ACL/filtri a carico della view).

Esempio::

    from core.excel_export import make_xlsx_response
    return make_xlsx_response(
        filename="conformita-dpi.xlsx",
        columns=["Categoria", "Stato"],
        rows=[["Guanti", "OK"], ["Scarpe", "Mancante"]],
        title="Conformità DPI — Mario Rossi",
    )
"""
from __future__ import annotations

from io import BytesIO

from django.http import HttpResponse

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Prefissi che Excel/Calc interpretano come inizio di formula (formula injection):
# oltre a "=" anche "+", "-", "@", TAB e CR.
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def write_cell(ws, row, column, value):
    """Scrive una cella NEUTRALIZZANDO la formula injection, e la ritorna.

    SICUREZZA — openpyxl deduce il tipo dal valore: qualunque *stringa* che inizia
    con ``=`` viene scritta come formula viva (``data_type == 'f'``). Nei nostri
    export le stringhe arrivano dal DB (nomi, descrizioni, nomi file) e perfino
    dalla querystring (l'etichetta filtri riflette ``?q=``): un link malevolo
    basterebbe a far scaricare a un utente HR un file con una formula attiva
    (es. ``=HYPERLINK(...)`` che esfiltra i dati della riga accanto).

    Rimedio: le stringhe sono SEMPRE scritte come testo (``data_type = "s"``),
    senza alterarne il contenuto (nessun apice aggiunto: ``-12,50`` resta
    ``-12,50``); sui prefissi pericolosi si aggiunge il *quote prefix* di Excel
    ("tratta come testo"), che non fa parte del valore. I valori NON stringa
    (int, float, Decimal, date, bool, None) restano tipizzati.
    """
    cell = ws.cell(row=row, column=column)
    cell.value = value  # assegnazione esplicita: con None ripulisce la cella
    if isinstance(value, str):
        cell.data_type = "s"  # mai 'f': nessuna formula viva nel file prodotto
        if value[:1] in FORMULA_PREFIXES:
            cell.quotePrefix = True
    return cell


def write_cell_at(ws, coordinate, value):
    """Come `write_cell` ma con coordinata Excel ("B3"). Ritorna la cella.

    Serve dove si scriveva `ws["B3"] = valore` (assegnazione che passa dal type
    guessing di openpyxl, quindi vulnerabile alla formula injection).
    """
    from openpyxl.utils.cell import coordinate_to_tuple

    row, column = coordinate_to_tuple(coordinate)
    return write_cell(ws, row, column, value)


def write_row(ws, row, values, start_col: int = 1):
    """Scrive una riga di celle sanificate. Ritorna la lista delle celle scritte
    (cosi' il chiamante puo' continuare ad applicare font/fill/border)."""
    return [
        write_cell(ws, row, col, value)
        for col, value in enumerate(values, start=start_col)
    ]


def append_row(ws, values=(), start_col: int = 1):
    """Equivalente sanificato di `ws.append([...])`. Ritorna le celle scritte.

    Replica la semantica di openpyxl (`Worksheet.append` scrive sulla riga
    `_current_row + 1` e la avanza anche quando la riga e' vuota, es. `ws.append([])`
    usato come spaziatore).
    """
    row = getattr(ws, "_current_row", ws.max_row) + 1
    cells = write_row(ws, row, values, start_col=start_col)
    if getattr(ws, "_current_row", 0) < row:
        ws._current_row = row  # riga vuota: nessuna cella scritta, ma la riga si consuma
    return cells


def _brand_header(ws, *, title, subtitle, filters_label, logo):
    """Scrive il blocco intestazione documento (righe 1..5) e ritorna la riga di header tabella.

    Riusa `core.pdf.PdfTheme.from_branding()` cosi' Excel e PDF condividono nome
    portale/logo. Il logo e' opzionale (best-effort): se non disegnabile si
    prosegue senza sollevare eccezioni verso l'utente.
    """
    from openpyxl.styles import Font

    from core.pdf import PdfTheme

    theme = PdfTheme.from_branding()
    navy = "0C2545"

    portal_name = getattr(theme, "portal_name", "") or "NOVICROM HUB"
    write_cell(ws, 1, 2, portal_name).font = Font(bold=True, size=11, color=navy)

    logo_path = getattr(theme, "logo_path", None)
    if logo and logo_path:
        try:
            from openpyxl.drawing.image import Image as XlImage

            img = XlImage(logo_path)
            native_w, native_h = img.width, img.height
            target_h = 36
            img.height = target_h
            img.width = int(native_w * (target_h / native_h)) if native_h else 90
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 30
        except Exception:  # logo non disegnabile: si prosegue senza (mai eccezione all'utente)
            pass

    write_cell(ws, 3, 1, str(title)).font = Font(bold=True, size=14, color=navy)
    if subtitle:
        write_cell(ws, 4, 1, str(subtitle)).font = Font(size=10, color="5B6B7C")
    if filters_label:
        write_cell(ws, 5, 1, str(filters_label)).font = Font(size=10, color="5B6B7C")
    return 7


def build_xlsx_bytes(
    *,
    columns,
    rows,
    sheet_title: str = "Dati",
    title: str | None = None,
    subtitle: str | None = None,
    filters_label: str | None = None,
    logo: bool = True,
) -> bytes:
    """Costruisce il file .xlsx (bytes). `columns`: intestazioni; `rows`: iterabile di righe.

    Con `title` valorizzato scrive il blocco intestazione HUB (logo, titolo,
    sottotitolo, filtri) e la tabella parte dalla riga 7; senza `title` l'header
    della tabella resta in riga 1 (retrocompatibilita' con i consumatori esistenti).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = list(columns)
    rows = [list(r) for r in rows]

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Dati")[:31]

    header_row = 1
    if title:
        header_row = _brand_header(
            ws, title=title, subtitle=subtitle, filters_label=filters_label, logo=logo
        )

    header_fill = PatternFill("solid", fgColor="0C2545")
    header_font = Font(bold=True, color="FFFFFF")
    for c, col in enumerate(columns, start=1):
        cell = write_cell(ws, header_row, c, str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    r = header_row
    for row in rows:
        r += 1
        for c, val in enumerate(row, start=1):
            write_cell(ws, r, c, val)

    # Larghezze auto (cap 60) sul contenuto.
    for c, col in enumerate(columns, start=1):
        maxlen = len(str(col))
        for row in rows:
            if c - 1 < len(row) and row[c - 1] is not None:
                maxlen = max(maxlen, len(str(row[c - 1])))
        ws.column_dimensions[get_column_letter(c)].width = min(60, max(10, maxlen + 2))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if rows and columns:
        last = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A{header_row}:{last}{header_row + len(rows)}"
        ws.print_title_rows = f"{header_row}:{header_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_xlsx_response(
    *,
    filename: str,
    columns,
    rows,
    sheet_title: str = "Dati",
    title: str | None = None,
    subtitle: str | None = None,
    filters_label: str | None = None,
    logo: bool = True,
) -> HttpResponse:
    """Risposta HTTP di download .xlsx. La view resta responsabile di ACL/filtri."""
    data = build_xlsx_bytes(
        columns=columns,
        rows=rows,
        sheet_title=sheet_title,
        title=title,
        subtitle=subtitle,
        filters_label=filters_label,
        logo=logo,
    )
    safe_name = (filename or "export.xlsx").replace('"', "")
    response = HttpResponse(data, content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    response["Content-Length"] = str(len(data))
    return response
