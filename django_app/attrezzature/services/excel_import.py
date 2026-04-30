from __future__ import annotations

import hashlib
import io
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal

from django.db import transaction
from django.utils import timezone

from attrezzature.models import (
    Attrezzatura,
    AttrezzaturaAvanzamento,
    AttrezzaturaImportBatch,
    AttrezzaturaImportRow,
    AttrezzaturaPartNumber,
    AttrezzaturaStato,
    DisponibilitaStato,
    ImportBatchStatus,
    ImportRowAzione,
)


HEADER_ALIASES = {
    "codice": {"codice", "cod", "codice attrezzo", "cod attrezzo"},
    "part_number": {"particolare", "part number", "p/n", "pn", "part_number"},
    "numero_pezzi": {"n pezzi", "n. pezzi", "numero pezzi", "pezzi"},
    "avanzamento_percentuale": {"avanzamento", "avanzamento %", "percentuale", "% avanzamento"},
    "note_consegna": {"note e consegna", "note consegna", "consegna note"},
    "data_consegna_prevista": {"data consegna", "consegna", "data prevista consegna"},
    "descrizione": {"descrizione", "description", "desc"},
    "note_rocco": {"note rocco", "rocco"},
    "ordine": {"ordine", "ord"},
}

EXPECTED_HEADER_HINTS = {alias for aliases in HEADER_ALIASES.values() for alias in aliases}


def normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[^\w%/]+", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_header(value) -> str:
    normalized = normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return normalized.replace(" ", "_").replace("/", "_")


def _jsonable(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def compute_row_hash(payload: dict) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def normalize_part_number(value) -> str:
    return " ".join(_clean(value).upper().split())


def normalize_int(value):
    text = _clean(value)
    if not text:
        return None, []
    try:
        return int(float(text.replace(",", "."))), []
    except (TypeError, ValueError):
        return None, [f"Numero non valido: {text}"]


def normalize_progress(value):
    warnings = []
    text = _clean(value)
    if not text or text == ".":
        return None, warnings
    try:
        if text.endswith("%"):
            number = float(text[:-1].strip().replace(",", "."))
        else:
            number = float(text.replace(",", "."))
            if 0 < number <= 1:
                number *= 100
        percent = int(round(number))
        if percent < 0 or percent > 100:
            warnings.append(f"Avanzamento fuori range: {text}")
            return None, warnings
        return percent, warnings
    except (TypeError, ValueError):
        warnings.append(f"Avanzamento non valido: {text}")
        return None, warnings


def normalize_date(value):
    if value in (None, ""):
        return None, []
    if isinstance(value, datetime):
        return value.date(), []
    if isinstance(value, date):
        return value, []
    text = _clean(value)
    if not text:
        return None, []
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date(), []
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date(), []
        except ValueError:
            pass
    return None, [f"Data consegna non valida: {text}"]


def interpret_ordine(value, codice_present: bool):
    warnings = []
    raw = _clean(value).lower()
    if raw == "o":
        return {"skip": True, "stato": "", "ordine_visuale": None, "warnings": warnings}
    if raw == "x":
        return {"skip": False, "stato": AttrezzaturaStato.FINITA, "ordine_visuale": None, "warnings": warnings}
    if raw == "z":
        return {"skip": False, "stato": AttrezzaturaStato.ECCEZIONE, "ordine_visuale": None, "warnings": warnings}
    if raw == "":
        return {
            "skip": False,
            "stato": AttrezzaturaStato.DA_CLASSIFICARE if codice_present else "",
            "ordine_visuale": None,
            "warnings": warnings,
        }
    try:
        return {"skip": False, "stato": AttrezzaturaStato.IN_CORSO, "ordine_visuale": int(float(raw.replace(",", "."))), "warnings": warnings}
    except ValueError:
        warnings.append(f"Ordine non riconosciuto: {value}")
        return {"skip": False, "stato": AttrezzaturaStato.DA_CLASSIFICARE, "ordine_visuale": None, "warnings": warnings}


def default_disponibilita(stato: str) -> str:
    if stato in {AttrezzaturaStato.FINITA, AttrezzaturaStato.PRONTA_PRODUZIONE}:
        return DisponibilitaStato.DISPONIBILE
    return DisponibilitaStato.DA_VERIFICARE


@dataclass
class SheetRows:
    name: str
    rows: list[list]


def _read_xlsx(data: bytes) -> SheetRows:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_name = "ATTREZZI" if "ATTREZZI" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    return SheetRows(sheet_name, [list(row) for row in ws.iter_rows(values_only=True)])


def _read_xls(data: bytes, filename: str) -> SheetRows:
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("Import .xls non disponibile: installare pandas/xlrd o convertire in .xlsx.") from exc
    sheets = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None)
    sheet_name = "ATTREZZI" if "ATTREZZI" in sheets else next(iter(sheets.keys()))
    frame = sheets[sheet_name]
    return SheetRows(str(sheet_name), frame.where(frame.notna(), None).values.tolist())


def read_excel_rows(file_obj, filename: str = "") -> SheetRows:
    data = file_obj.read() if hasattr(file_obj, "read") else bytes(file_obj)
    name = filename or getattr(file_obj, "name", "") or ""
    ext = os.path.splitext(name.lower())[1]
    if ext == ".xls":
        return _read_xls(data, name)
    return _read_xlsx(data)


def detect_header_row(rows: list[list]) -> tuple[int, list[str]]:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:12]):
        normalized = [normalize_header(cell) for cell in row]
        score = sum(1 for cell in normalized if cell in EXPECTED_HEADER_HINTS)
        if score > best_score:
            best_idx = idx
            best_score = score
    headers = []
    seen: dict[str, int] = {}
    for pos, raw in enumerate(rows[best_idx]):
        base = canonical_header(raw) or f"colonna_{pos + 1}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return best_idx, headers


def _row_payload(headers: list[str], row: list) -> dict:
    payload = {}
    for idx, header in enumerate(headers):
        value = row[idx] if idx < len(row) else None
        if value not in (None, ""):
            payload[header] = _jsonable(value)
    return payload


def _first(payload: dict, key: str):
    if key in payload:
        return payload.get(key)
    for payload_key, value in payload.items():
        if payload_key.startswith(key + "_"):
            return value
    return None


def _normalized_values(payload: dict) -> tuple[dict, list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    codice = _clean(_first(payload, "codice"))
    part_number = normalize_part_number(_first(payload, "part_number"))
    descrizione = _clean(_first(payload, "descrizione"))
    numero_pezzi, w = normalize_int(_first(payload, "numero_pezzi"))
    warnings.extend(w)
    avanzamento, w = normalize_progress(_first(payload, "avanzamento_percentuale"))
    warnings.extend(w)
    data_consegna, w = normalize_date(_first(payload, "data_consegna_prevista"))
    warnings.extend(w)
    ordine_info = interpret_ordine(_first(payload, "ordine"), bool(codice))
    warnings.extend(ordine_info["warnings"])
    stato = ordine_info["stato"] or AttrezzaturaStato.DA_CLASSIFICARE
    values = {
        "codice": codice,
        "part_number": part_number,
        "descrizione": descrizione,
        "numero_pezzi": numero_pezzi,
        "avanzamento_percentuale": avanzamento,
        "note_consegna": _clean(_first(payload, "note_consegna")),
        "data_consegna_prevista": data_consegna,
        "note_rocco": _clean(_first(payload, "note_rocco")),
        "stato": stato,
        "disponibilita_stato": default_disponibilita(stato),
        "ordine_visuale": ordine_info["ordine_visuale"],
        "is_placeholder": bool(ordine_info["skip"]),
    }
    return values, warnings, errors


def _find_match(values: dict) -> tuple[Attrezzatura | None, list[str]]:
    warnings: list[str] = []
    codice = values["codice"]
    part_number = values["part_number"]
    descrizione = values["descrizione"]
    if not codice:
        return None, warnings

    strategies = []
    if part_number and descrizione:
        strategies.append({"codice__iexact": codice, "part_number__iexact": part_number, "descrizione__iexact": descrizione})
    if part_number:
        strategies.append({"codice__iexact": codice, "part_number__iexact": part_number})
    if not part_number and descrizione:
        strategies.append({"codice__iexact": codice, "descrizione__iexact": descrizione})

    for filters in strategies:
        matches = list(Attrezzatura.objects.filter(**filters).order_by("id")[:3])
        if len(matches) == 1:
            return matches[0], warnings
        if len(matches) > 1:
            warnings.append("Match ambiguo: piu attrezzature corrispondono alla riga.")
            return None, warnings
    return None, warnings


def _action_for(values: dict, warnings: list[str], errors: list[str]) -> tuple[str, Attrezzatura | None]:
    if values["is_placeholder"]:
        return ImportRowAzione.SKIPPED, None
    if not values["codice"] and not values["part_number"] and not values["descrizione"]:
        return ImportRowAzione.SKIPPED, None
    if errors:
        return ImportRowAzione.ERROR, None
    if not values["codice"]:
        warnings.append("Riga significativa senza codice attrezzo: import manuale richiesto.")
        return ImportRowAzione.WARNING, None
    match, match_warnings = _find_match(values)
    warnings.extend(match_warnings)
    if match_warnings:
        return ImportRowAzione.WARNING, None
    if warnings:
        return ImportRowAzione.WARNING, match
    if match is None:
        return ImportRowAzione.CREATED, None
    comparable = {
        "descrizione": values["descrizione"],
        "part_number": values["part_number"],
        "numero_pezzi": values["numero_pezzi"],
        "avanzamento_percentuale": values["avanzamento_percentuale"],
        "stato": values["stato"],
        "disponibilita_stato": values["disponibilita_stato"],
        "ordine_visuale": values["ordine_visuale"],
        "data_consegna_prevista": values["data_consegna_prevista"],
        "note_consegna": values["note_consegna"],
        "note_rocco": values["note_rocco"],
    }
    unchanged = all(getattr(match, field) == value for field, value in comparable.items())
    return (ImportRowAzione.UNCHANGED if unchanged else ImportRowAzione.UPDATED), match


def build_import_preview(file_obj, *, filename: str = "", user=None, dry_run: bool = True) -> dict:
    sheet = read_excel_rows(file_obj, filename)
    header_idx, headers = detect_header_row(sheet.rows)
    previews = []
    summary = {key: 0 for key in ["created", "updated", "unchanged", "skipped", "warning", "error"]}
    for offset, row in enumerate(sheet.rows[header_idx + 1 :], start=header_idx + 2):
        payload = _row_payload(headers, row)
        if not payload:
            continue
        values, warnings, errors = _normalized_values(payload)
        action, match = _action_for(values, warnings, errors)
        summary[action] += 1
        previews.append(
            {
                "excel_sheet": sheet.name,
                "excel_row_number": offset,
                "payload_originale_json": payload,
                "row_hash": compute_row_hash(payload),
                "normalized": values,
                "warnings": warnings,
                "errors": errors,
                "proposed_action": action,
                "attrezzatura_id": match.id if match else None,
            }
        )
    summary["total_rows"] = len(previews)
    summary["warning_count"] = summary["warning"] + sum(1 for row in previews if row["warnings"] and row["proposed_action"] != ImportRowAzione.WARNING)
    summary["error_count"] = summary["error"]
    return {
        "dry_run": dry_run,
        "file_name": filename or getattr(file_obj, "name", ""),
        "sheet": sheet.name,
        "header_row_number": header_idx + 1,
        "headers": headers,
        "summary": summary,
        "row_previews": previews,
    }


def _create_or_update_attrezzatura(row: dict, batch: AttrezzaturaImportBatch, user=None) -> Attrezzatura | None:
    values = row["normalized"]
    action = row["proposed_action"]
    if action not in {ImportRowAzione.CREATED, ImportRowAzione.UPDATED}:
        return Attrezzatura.objects.filter(pk=row.get("attrezzatura_id")).first() if row.get("attrezzatura_id") else None
    match = Attrezzatura.objects.filter(pk=row.get("attrezzatura_id")).first() if row.get("attrezzatura_id") else None
    if action == ImportRowAzione.CREATED:
        match = Attrezzatura.objects.create(
            codice=values["codice"],
            descrizione=values["descrizione"],
            part_number=values["part_number"],
            numero_pezzi=values["numero_pezzi"],
            avanzamento_percentuale=values["avanzamento_percentuale"],
            stato=values["stato"],
            disponibilita_stato=values["disponibilita_stato"],
            ordine_visuale=values["ordine_visuale"],
            data_consegna_prevista=values["data_consegna_prevista"],
            note_consegna=values["note_consegna"],
            note_rocco=values["note_rocco"],
            origine_import=batch.file_name,
            created_by=user,
            updated_by=user,
        )
        if values["part_number"]:
            AttrezzaturaPartNumber.objects.get_or_create(
                attrezzatura=match,
                part_number=values["part_number"],
                defaults={"fonte": "import_excel"},
            )
        AttrezzaturaAvanzamento.objects.create(
            attrezzatura=match,
            stato_nuovo=match.stato,
            avanzamento_nuovo=match.avanzamento_percentuale,
            nota="Creata da import Excel legacy",
            origine="import_excel",
            created_by=user,
        )
        return match
    if match is None:
        return None

    old_stato = match.stato
    old_progress = match.avanzamento_percentuale
    for field in [
        "descrizione",
        "part_number",
        "numero_pezzi",
        "avanzamento_percentuale",
        "stato",
        "disponibilita_stato",
        "ordine_visuale",
        "data_consegna_prevista",
        "note_consegna",
        "note_rocco",
    ]:
        setattr(match, field, values[field])
    match.origine_import = batch.file_name
    match.updated_by = user
    match.save()
    if values["part_number"]:
        AttrezzaturaPartNumber.objects.get_or_create(
            attrezzatura=match,
            part_number=values["part_number"],
            defaults={"fonte": "import_excel"},
        )
    if old_stato != match.stato or old_progress != match.avanzamento_percentuale:
        AttrezzaturaAvanzamento.objects.create(
            attrezzatura=match,
            stato_precedente=old_stato,
            stato_nuovo=match.stato,
            avanzamento_precedente=old_progress,
            avanzamento_nuovo=match.avanzamento_percentuale,
            nota="Aggiornata da import Excel legacy",
            origine="import_excel",
            created_by=user,
        )
    return match


@transaction.atomic
def confirm_import(file_obj, *, filename: str = "", user=None) -> dict:
    preview = build_import_preview(file_obj, filename=filename, user=user, dry_run=False)
    batch = AttrezzaturaImportBatch.objects.create(
        file_name=preview["file_name"] or "import_attrezzature.xlsx",
        uploaded_by=user,
        total_rows=preview["summary"]["total_rows"],
        status=ImportBatchStatus.IMPORTED,
        details={
            "sheet": preview["sheet"],
            "header_row_number": preview["header_row_number"],
            "imported_at": timezone.now().isoformat(),
        },
    )
    counters = {key: 0 for key in ["created", "updated", "unchanged", "skipped", "warning", "error"]}
    for row in preview["row_previews"]:
        attrezzatura = _create_or_update_attrezzatura(row, batch, user=user)
        action = row["proposed_action"]
        counters[action] += 1
        AttrezzaturaImportRow.objects.create(
            batch=batch,
            excel_sheet=row["excel_sheet"],
            excel_row_number=row["excel_row_number"],
            payload_originale_json=row["payload_originale_json"],
            row_hash=row["row_hash"],
            codice_letto=row["normalized"]["codice"],
            part_number_letto=row["normalized"]["part_number"],
            azione=action,
            warnings=row["warnings"],
            errors=row["errors"],
            attrezzatura=attrezzatura,
        )
    batch.created_count = counters["created"]
    batch.updated_count = counters["updated"]
    batch.unchanged_count = counters["unchanged"]
    batch.skipped_count = counters["skipped"]
    batch.warning_count = counters["warning"] + sum(1 for row in preview["row_previews"] if row["warnings"] and row["proposed_action"] != ImportRowAzione.WARNING)
    batch.error_count = counters["error"]
    batch.save()
    preview["batch"] = batch
    preview["summary"].update(counters)
    preview["summary"]["warning_count"] = batch.warning_count
    preview["summary"]["error_count"] = batch.error_count
    return preview
